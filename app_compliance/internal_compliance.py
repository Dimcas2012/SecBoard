from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.utils.translation import gettext_lazy as _, get_language
from django.views.decorators.http import require_http_methods, require_POST
from io import BytesIO
from app_conf.models import Company, Country, CompanyType
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date
from django.db import transaction
from django.core.exceptions import ValidationError
import PyPDF2
import re

from .models import (
    ComplianceFramework, ControlCategory, Control,
    Evidence, ControlAssignment, ComplianceAuditLog, ControlMapping,
    AccessCompliance, AccessLocalCompliance, RegulatorType, LocalComplianceRegulator,
    LocalComplianceRequirement, LocalComplianceControl, LocalRequirementCategory,
    RequirementType, RequirementStatus, RequirementPriority,
    RequirementTypeTranslation, RequirementStatusTranslation,
    RequirementPriorityTranslation, EvidenceType, EvidenceTypeTranslation,
    LocalControlEvidence,
    LocalControlAssignment, LocalControlNote, LocalControlMapping,
    AccessInternalCompliance, InternalComplianceSource, InternalComplianceRequirement,
    InternalComplianceControl, InternalRequirementCategory, InternalControlEvidence,
    InternalControlAssignment, InternalControlNote, InternalControlMapping,
    AccessControlMapping, FrameworkInstanceNote, FrameworkDomain,
    InternalComplianceGuide, InternalComplianceGuideTranslation
)

from .utils import *

# ========================
# Internal Compliance Dashboard
# ========================


def _has_internal_compliance_access(user):
    """Check if user has access to Internal Compliance module."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return AccessInternalCompliance.objects.filter(
        group__in=user.groups.all(),
        has_access=True
    ).exists()


@login_required
@require_http_methods(["GET"])
def internal_compliance_guide(request):
    """Return JSON { content: html } for the Internal Compliance guide (localized)."""
    if not _has_internal_compliance_access(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = InternalComplianceGuide.objects.first()
    if guide:
        if country:
            trans = InternalComplianceGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = InternalComplianceGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def internal_compliance_guide_translate(request):
    """API for AI translation of Internal Compliance guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@internal_compliance_access_required
def internal_compliance_dashboard(request):
    """
    Dashboard for Internal Compliance (internal company requirements)
    """
    from datetime import date, timedelta
    from django.db.models import Count, Q
    
    # Get user's accessible companies for Internal Compliance
    accessible_companies = get_user_accessible_companies_internal(request.user)
    
    # Get selected companies from GET parameters
    selected_company_ids = request.GET.getlist('company')
    
    if selected_company_ids:
        # Filter only accessible companies
        valid_selected_ids = [
            int(cid) for cid in selected_company_ids 
            if int(cid) in list(accessible_companies.values_list('id', flat=True))
        ]
        if valid_selected_ids:
            selected_companies = accessible_companies.filter(id__in=valid_selected_ids)
        else:
            selected_companies = accessible_companies
    else:
        selected_companies = accessible_companies
    
    # Filter by source type
    selected_source_type = request.GET.get('source_type', '')
    
    # Get all active sources
    sources_qs = InternalComplianceSource.objects.filter(is_active=True)
    
    if selected_source_type:
        sources_qs = sources_qs.filter(source_type=selected_source_type)
    
    # Count requirements for selected companies and annotate sources
    sources = sources_qs.annotate(
        total_requirements=Count('requirements', filter=Q(
            requirements__company__in=selected_companies
        )),
        active_requirements=Count('requirements', filter=Q(
            requirements__status='active',
            requirements__company__in=selected_companies
        ))
    ).order_by('name')
    
    # Get requirements for selected companies (including templates with company assigned)
    # Internal Requirements can be templates but still assigned to specific companies
    requirements_qs = InternalComplianceRequirement.objects.filter(
        company__in=selected_companies
    ).select_related('source', 'company')
    
    # Filter by status
    status_filter = request.GET.get('status', 'active')
    if status_filter and status_filter != 'all':
        requirements_qs = requirements_qs.filter(status=status_filter)
    
    # Filter by source if sources exist
    if sources.exists():
        requirements_qs = requirements_qs.filter(source__in=sources)
    
    # Get controls for requirements
    # If requirement is a template, controls have company=NULL
    # If requirement is an instance, controls have company assigned
    controls = InternalComplianceControl.objects.filter(
        requirement__in=requirements_qs
    ).filter(
        Q(company__in=selected_companies) | Q(company__isnull=True)
    ).select_related('requirement', 'requirement__company', 'company', 'responsible')
    
    # Overall statistics
    total_sources = sources.count()
    total_requirements = requirements_qs.count()
    total_controls = controls.count()
    
    completed_controls = controls.filter(status='completed').count()
    in_progress_controls = controls.filter(status='in_progress').count()
    not_started_controls = controls.filter(status='not_started').count()
    
    overall_completion = 0
    if total_controls > 0:
        overall_completion = round((completed_controls / total_controls) * 100, 1)
    
    # Priority statistics
    critical_controls = controls.filter(priority='critical').count()
    critical_completed = controls.filter(priority='critical', status='completed').count()
    
    high_controls = controls.filter(priority='high').count()
    high_completed = controls.filter(priority='high', status='completed').count()
    
    # Overdue controls
    today = date.today()
    overdue_controls = controls.filter(
        target_completion_date__lt=today,
        status__in=['not_started', 'in_progress']
    ).order_by('target_completion_date')[:10]
    
    # Controls becoming due in next 30 days
    upcoming_due = controls.filter(
        target_completion_date__gte=today,
        target_completion_date__lte=today + timedelta(days=30),
        status__in=['not_started', 'in_progress']
    ).order_by('target_completion_date')[:10]
    
    # Overdue requirements (by deadline)
    overdue_requirements = requirements_qs.filter(
        deadline_date__lt=today
    ).order_by('deadline_date')[:10]
    
    # Upcoming deadlines for requirements
    upcoming_requirements = requirements_qs.filter(
        deadline_date__gte=today,
        deadline_date__lte=today + timedelta(days=60)
    ).order_by('deadline_date')[:10]
    
    # Statistics by source
    source_stats = []
    for source in sources[:10]:  # Top 10 sources
        source_requirements = requirements_qs.filter(source=source)
        source_controls = controls.filter(requirement__source=source)
        
        total = source_controls.count()
        completed = source_controls.filter(status='completed').count()
        completion = round((completed / total * 100), 1) if total > 0 else 0
        
        source_stats.append({
            'source': source,
            'requirements_count': source_requirements.count(),
            'total_controls': total,
            'completed_controls': completed,
            'completion': completion,
        })
    
    # Statistics by company
    company_stats = []
    for company in selected_companies:
        company_controls = controls.filter(company=company)
        total = company_controls.count()
        completed = company_controls.filter(status='completed').count()
        completion = round((completed / total * 100), 1) if total > 0 else 0
        
        company_stats.append({
            'company': company,
            'total_controls': total,
            'completed': completed,
            'in_progress': company_controls.filter(status='in_progress').count(),
            'not_started': company_controls.filter(status='not_started').count(),
            'completion': completion,
        })
    
    # Priority Matrix
    priority_matrix = {}
    for priority_value, priority_label in Control.PRIORITY_CHOICES:
        priority_matrix[priority_value] = {
            'label': priority_label,
            'completed': controls.filter(priority=priority_value, status='completed').count(),
            'in_progress': controls.filter(priority=priority_value, status='in_progress').count(),
            'not_started': controls.filter(priority=priority_value, status='not_started').count(),
            'total': controls.filter(priority=priority_value).count(),
        }
    
    # Team Workload
    from django.contrib.auth.models import User
    
    team_workload = []
    users_with_controls = User.objects.filter(
        internal_compliance_controls__company__in=selected_companies
    ).distinct()
    
    for user in users_with_controls:
        user_controls = controls.filter(responsible=user)
        total_assigned = user_controls.count()
        
        if total_assigned > 0:
            completed = user_controls.filter(status='completed').count()
            completion_pct = round((completed / total_assigned) * 100, 1)
            
            team_workload.append({
                'user': user,
                'total': total_assigned,
                'completed': completed,
                'in_progress': user_controls.filter(status='in_progress').count(),
                'not_started': user_controls.filter(status='not_started').count(),
                'completion': completion_pct,
            })
    
    # Unassigned controls
    unassigned_count = controls.filter(responsible__isnull=True).count()
    if unassigned_count > 0:
        team_workload.append({
            'user': None,
            'total': unassigned_count,
            'completed': controls.filter(responsible__isnull=True, status='completed').count(),
            'in_progress': controls.filter(responsible__isnull=True, status='in_progress').count(),
            'not_started': controls.filter(responsible__isnull=True, status='not_started').count(),
            'completion': 0,
        })
    
    # Sort by completion (ascending) to show users needing attention first
    team_workload.sort(key=lambda x: x['completion'])
    
    # Get user permissions
    permissions = get_user_internal_compliance_permissions(request.user)
    
    # Get available source types for filter
    source_type_choices = InternalComplianceSource.SOURCE_TYPE_CHOICES
    
    # Prepare selected company IDs for template
    final_selected_ids = [c.id for c in selected_companies]
    
    context = {
        'accessible_companies': accessible_companies,
        'available_companies': accessible_companies,  # Alias for template compatibility
        'selected_companies': selected_companies,
        'selected_company_ids': final_selected_ids,
        'sources': sources,
        'source_type_choices': source_type_choices,
        'selected_source_type': selected_source_type,
        'total_sources': total_sources,
        'total_requirements': total_requirements,
        'total_controls': total_controls,
        'completed_controls': completed_controls,
        'in_progress_controls': in_progress_controls,
        'not_started_controls': not_started_controls,
        'overall_completion': overall_completion,
        'critical_controls': critical_controls,
        'critical_completed': critical_completed,
        'high_controls': high_controls,
        'high_completed': high_completed,
        'overdue_controls': overdue_controls,
        'upcoming_due': upcoming_due,
        'overdue_requirements': overdue_requirements,
        'upcoming_requirements': upcoming_requirements,
        'source_stats': source_stats,
        'company_stats': company_stats,
        'priority_matrix': priority_matrix,
        'team_workload': team_workload[:10],
        'permissions': permissions,
    }
    
    return render(request, 'app_compliance/internal_compliance_dashboard.html', context)


@login_required
@internal_compliance_access_required
def internal_requirements_library(request):
    """
    Internal Requirements Templates
    """
    # Check permissions
    if not check_user_internal_compliance_permission(request.user, 'can_view_requirements'):
        messages.error(request, _('You do not have permission to view requirements templates'))
        return redirect('compliance:internal_compliance')
    
    # Show only requirements
    requirements = InternalComplianceRequirement.objects.filter(is_template=True)
    
    requirements = requirements.select_related('created_by', 'company').prefetch_related('instances__company').annotate(
        controls_count=Count('controls', filter=Q(controls__company__isnull=True)),
        instances_count=Count('instances')
    )
    
    # Search
    search = request.GET.get('search', '')
    if search:
        requirements = requirements.filter(
            Q(name__icontains=search) |
            Q(name_local__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Filter by type
    requirement_type = request.GET.get('type', '')
    if requirement_type:
        requirements = requirements.filter(requirement_type=requirement_type)
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        requirements = requirements.filter(status=status)
    
    # Pagination
    paginator = Paginator(requirements, 25)
    page = request.GET.get('page')
    
    try:
        requirements_page = paginator.page(page)
    except PageNotAnInteger:
        requirements_page = paginator.page(1)
    except EmptyPage:
        requirements_page = paginator.page(paginator.num_pages)
    
    # Get user permissions
    permissions = get_user_internal_compliance_permissions(request.user)
    
    # Get requirement types using RequirementType model
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    requirement_types = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(InternalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    )
    status_choices = get_dictionary_options(
        RequirementStatus,
        RequirementStatusTranslation,
        'requirement_status',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(InternalComplianceRequirement.STATUS_CHOICES)
    )
    
    # Get accessible companies for import modal
    accessible_companies = get_user_accessible_companies_internal(request.user)
    
    # Get active countries for AI import
    from app_conf.models import Country
    active_countries = Country.objects.filter(is_active=True).order_by('display_order', 'name')
    
    # Check access to Control Mapping
    has_access_to_control_mapping = check_user_control_mapping_access(request.user)
    
    context = {
        'requirements': requirements_page,
        'search': search,
        'selected_type': requirement_type,
        'selected_status': status,
        'requirement_types': requirement_types,
        'status_choices': status_choices,
        'permissions': permissions,
        'accessible_companies': accessible_companies,
        'active_countries': active_countries,
        'has_access_to_control_mapping': has_access_to_control_mapping,
    }
    
    return render(request, 'app_compliance/internal_requirements_library.html', context)


@login_required
@internal_compliance_access_required
def internal_requirements_export_excel(request):
    """Export filtered Internal Requirements to Excel"""
    if not check_user_internal_compliance_permission(request.user, 'can_view_requirements'):
        messages.error(request, _('You do not have permission to export requirements'))
        return redirect('compliance:internal_requirements_library')

    requirements = InternalComplianceRequirement.objects.filter(is_template=True)
    requirements = requirements.select_related('created_by').prefetch_related('instances__company').annotate(
        controls_count=Count('controls', filter=Q(controls__company__isnull=True)),
        instances_count=Count('instances')
    )

    search = request.GET.get('search', '')
    if search:
        requirements = requirements.filter(
            Q(name__icontains=search) |
            Q(name_local__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )

    requirement_type = request.GET.get('type', '')
    if requirement_type:
        requirements = requirements.filter(requirement_type=requirement_type)

    status = request.GET.get('status', '')
    if status:
        requirements = requirements.filter(status=status)

    requirements = requirements.order_by('-created_date')

    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Internal Requirements'))

    headers = [
        str(_('Code')),
        str(_('Name')),
        str(_('Type')),
        str(_('Status')),
        str(_('Controls')),
        str(_('Applied to Companies')),
        str(_('Effective Date')),
        str(_('Deadline'))
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='28a745', end_color='1e7e34', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    status_fill_map = {
        'active': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'draft': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'archived': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
    }

    accessible_companies = get_user_accessible_companies_internal(request.user)
    accessible_company_ids = set(accessible_companies.values_list('id', flat=True))
    is_staff_or_superuser = request.user.is_superuser or request.user.is_staff

    for idx, req in enumerate(requirements, start=2):
        instances = list(req.instances.filter(company__isnull=False))
        if not is_staff_or_superuser:
            instances = [inst for inst in instances if inst.company_id in accessible_company_ids]

        applied_companies = []
        for instance in instances:
            label = instance.company.name if instance.company else str(_('(No company)'))
            if instance.is_mandatory:
                label = f"{label} ({str(_('Mandatory'))})"
            applied_companies.append(label)

        row = [
            req.code,
            req.name,
            req.get_requirement_type_display(),
            req.get_status_display(),
            req.controls_count,
            ', '.join(applied_companies) if applied_companies else str(_('Not applied')),
            req.effective_date.strftime('%d.%m.%Y') if req.effective_date else '',
            req.deadline_date.strftime('%d.%m.%Y') if req.deadline_date else '',
        ]
        ws.append(row)

        status_cell = ws.cell(row=idx, column=5)
        status_fill = status_fill_map.get(req.status)
        if status_fill:
            status_cell.fill = status_fill

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col_idx).border = thin_border

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"internal_requirements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@internal_compliance_access_required
def internal_requirements_excel_template(request):
    """Download Excel template for Internal Requirements import"""
    if not check_user_internal_compliance_permission(request.user, 'can_add_requirements'):
        messages.error(request, _('You do not have permission to download the template'))
        return redirect('compliance:internal_requirements_library')

    # Get actual codes from models
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    
    requirement_types = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(InternalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    )
    requirement_type_codes = [opt['code'] for opt in requirement_types]
    requirement_type_notes = str(_('Options: ')) + ', '.join(requirement_type_codes) if requirement_type_codes else str(_('Options: policy, standard, procedure, guideline, directive, rule, other'))
    
    status_choices = get_dictionary_options(
        RequirementStatus,
        RequirementStatusTranslation,
        'requirement_status',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(InternalComplianceRequirement.STATUS_CHOICES)
    )
    status_codes = [opt['code'] for opt in status_choices]
    status_notes = str(_('Options: ')) + ', '.join(status_codes) if status_codes else str(_('Options: draft, active, suspended, archived'))
    
    priority_choices = get_dictionary_options(
        RequirementPriority,
        RequirementPriorityTranslation,
        'requirement_priority',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(Control.PRIORITY_CHOICES)
    )
    priority_codes = [opt['code'] for opt in priority_choices]
    priority_notes = str(_('Options: ')) + ', '.join(priority_codes) if priority_codes else str(_('Options: critical, high, medium, low'))
    
    # Get default values (first active option or fallback)
    default_requirement_type = requirement_type_codes[0] if requirement_type_codes else 'policy'
    default_status = status_codes[0] if status_codes else 'draft'
    default_priority = priority_codes[0] if priority_codes else 'medium'
    
    wb = Workbook()
    
    # Requirement sheet - use "Requirement" and "Controls" (import expects exact names)
    ws_req = wb.active
    ws_req.title = "Requirement"

    req_headers = ["Field", "Value", "Notes"]
    ws_req.append(req_headers)

    req_data = [
        ["Code", "", str(_('Required: Unique requirement code (e.g., POL-001)'))],
        ["Name", "", str(_('Required: Full requirement name'))],
        ["Requirement Type", default_requirement_type, requirement_type_notes],
        ["Status", default_status, status_notes],
        ["Priority", default_priority, priority_notes],
        ["Mandatory", "yes", str(_('yes or no'))],
        ["Name Local", "", str(_('Optional: Localized name'))],
        ["Description", "", str(_('Optional: Detailed description'))],
        ["Applicable To", "", str(_('Optional: Who this applies to'))],
        ["Publication Date", "", str(_('Format: DD.MM.YYYY'))],
        ["Effective Date", "", str(_('Format: DD.MM.YYYY'))],
        ["Deadline Date", "", str(_('Format: DD.MM.YYYY'))],
    ]

    for row in req_data:
        ws_req.append(row)

    # Controls sheet - import expects exact sheet name "Controls"
    ws_ctrl = wb.create_sheet("Controls")
    ctrl_headers = [
        "Category Code",
        "Category Name",
        "Category Description",
        "Control Code",
        "Control Name",
        "Priority",
        "Target Date",
        "Periodicity (days)",
        "Implementation Notes",
        "Evidence Notes",
        "Description",
    ]
    ws_ctrl.append(ctrl_headers)
    
    # Style headers
    header_fill = PatternFill(start_color='28a745', end_color='1e7e34', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for ws in [ws_req, ws_ctrl]:
        for col_idx in range(1, len(ws[1]) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-width columns
    for ws in [ws_req, ws_ctrl]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 4, 60)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="internal_requirements_template.xlsx"'
    return response


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirements_import_excel(request):
    """Import Internal Requirement with controls from Excel"""
    if not check_user_internal_compliance_permission(request.user, 'can_add_requirements'):
        messages.error(request, _('You do not have permission to import requirements'))
        return redirect('compliance:internal_requirements_library')

    excel_file = request.FILES.get('file')
    if not excel_file:
        messages.error(request, _('No file uploaded'))
        return redirect('compliance:internal_requirements_library')

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as exc:
        messages.error(request, _('Unable to read Excel file: %(error)s') % {'error': str(exc)})
        return redirect('compliance:internal_requirements_library')

    if "Requirement" not in wb.sheetnames or "Controls" not in wb.sheetnames:
        messages.error(
            request,
            _('Invalid template structure. The file must contain sheets named exactly "Requirement" and "Controls". '
              'Please use the template from the "Download Template" button.')
        )
        return redirect('compliance:internal_requirements_library')

    ws_req = wb["Requirement"]
    requirement_data = {}
    for row in ws_req.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        field = str(row[0]).strip().lower()
        value = row[1] if len(row) > 1 else ''
        requirement_data[field] = value

    def get_required(field):
        value = str(requirement_data.get(field, '') or '').strip()
        if not value:
            raise ValueError(_('Field "%(field)s" is required') % {'field': field})
        return value

    def normalize_choice(value, choices, label):
        if not value:
            raise ValueError(_('Field "%(field)s" is required') % {'field': label})
        value_str = str(value).strip().lower()
        for code, display in choices:
            if value_str == code.lower() or value_str == str(display).lower():
                return code
        raise ValueError(_('Invalid %(field)s value: %(value)s') % {'field': label, 'value': value})

    def normalize_from_options(value, options, label, default=None):
        """Normalize value from get_dictionary_options result (list of dicts with 'code' and 'label')"""
        if not value and default is not None:
            return default
        if not value:
            raise ValueError(_('Field "%(field)s" is required') % {'field': label})
        value_str = str(value).strip().lower()
        for opt in options:
            code = opt.get('code', '')
            label_text = opt.get('label', '')
            if value_str == code.lower() or value_str == str(label_text).lower():
                return code
        raise ValueError(_('Invalid %(field)s value: %(value)s. Valid options: %(options)s') % {
            'field': label,
            'value': value,
            'options': ', '.join([opt.get('code', '') for opt in options])
        })

    def is_empty_value(value):
        if not value:
            return True
        value_str = str(value).strip()
        if not value_str:
            return True
        value_lower = value_str.lower()
        if value_lower in ('none', 'null', '', 'n/a', '-', '—'):
            return True
        return False

    def parse_date(value):
        if is_empty_value(value):
            return None
        try:
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            return parse_local_requirement_date(value)
        except (ValueError, TypeError):
            return None

    try:
        # Get company (required)
        company_id = request.POST.get('company', '')
        if not company_id:
            messages.error(request, _('Company is required'))
            return redirect('compliance:internal_requirements_library')
        
        # Verify user has access to this company
        accessible_companies = get_user_accessible_companies_internal(request.user)
        try:
            company = accessible_companies.get(id=int(company_id))
        except (ValueError, Company.DoesNotExist, TypeError):
            messages.error(request, _('Invalid company selected'))
            return redirect('compliance:internal_requirements_library')
        
        # Get actual options from models
        language_code, country_for_language, use_localized_labels = get_language_preferences(request)
        
        requirement_types = get_dictionary_options(
            RequirementType,
            RequirementTypeTranslation,
            'requirement_type',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(InternalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
        )
        
        status_choices = get_dictionary_options(
            RequirementStatus,
            RequirementStatusTranslation,
            'requirement_status',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(InternalComplianceRequirement.STATUS_CHOICES)
        )
        
        priority_choices = get_dictionary_options(
            RequirementPriority,
            RequirementPriorityTranslation,
            'requirement_priority',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(Control.PRIORITY_CHOICES)
        )
        
        with transaction.atomic():
            code = get_required('code')
            name = get_required('name')

            requirement_type_code = normalize_from_options(
                requirement_data.get('requirement type', 'policy'),
                requirement_types,
                _('Requirement Type'),
                default='policy'
            )
            status_code = normalize_from_options(
                requirement_data.get('status', 'active'),
                status_choices,
                _('Status'),
                default='active'
            )
            requirement_priority = normalize_from_options(
                requirement_data.get('priority', 'medium'),
                priority_choices,
                _('Priority'),
                default='medium'
            )
            mandatory_value = str(requirement_data.get('mandatory', 'yes') or '').strip().lower()
            is_mandatory = mandatory_value in ('yes', 'true', '1', 'y', 'так', 'да')

            requirement = InternalComplianceRequirement.objects.create(
                code=code,
                name=name,
                company=company,
                name_local=str(requirement_data.get('name local', '') or '').strip(),
                requirement_type=requirement_type_code,
                description=str(requirement_data.get('description', '') or '').strip(),
                status=status_code,
                applicable_to=str(requirement_data.get('applicable to', '') or '').strip(),
                publication_date=parse_date(requirement_data.get('publication date')),
                effective_date=parse_date(requirement_data.get('effective date')),
                deadline_date=parse_date(requirement_data.get('deadline date')),
                is_mandatory=is_mandatory,
                priority=requirement_priority,
                is_template=True,
                created_by=request.user
            )

            ws_controls = wb["Controls"]
            categories_cache = {}
            category_order = 0
            last_category = None
            control_count = 0

            priority_map = {code.lower(): code for code, _ in Control.PRIORITY_CHOICES}
            priority_map.update({str(label).lower(): code for code, label in Control.PRIORITY_CHOICES})

            for row in ws_controls.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                category_code = str(row[0]).strip() if row[0] else ''
                category_name = str(row[1]).strip() if row[1] else ''
                category_description = str(row[2]).strip() if row[2] else ''
                control_code = str(row[3]).strip() if row[3] else ''
                control_name = str(row[4]).strip() if row[4] else ''
                control_priority_raw = str(row[5]).strip() if row[5] else ''
                target_date_raw = row[6]
                periodicity_raw = row[7] if len(row) > 7 else None
                implementation_notes = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                evidence_notes = str(row[9]).strip() if len(row) > 9 and row[9] else ''
                control_description = str(row[10]).strip() if len(row) > 10 and row[10] else ''

                if category_code or category_name:
                    if not category_code:
                        category_code = f"CAT-{len(categories_cache) + 1}"
                    if category_code not in categories_cache:
                        category_order += 1
                        categories_cache[category_code] = InternalRequirementCategory.objects.create(
                            requirement=requirement,
                            code=category_code,
                            name=category_name or category_code,
                            description=category_description,
                            order=category_order
                        )
                    last_category = categories_cache[category_code]

                if not control_code or not control_name:
                    continue

                control_priority = control_priority_raw.lower() if control_priority_raw else 'medium'
                control_priority = priority_map.get(control_priority, 'medium')
                target_date = parse_date(target_date_raw)
                
                # Parse periodicity
                periodicity = None
                if periodicity_raw is not None:
                    try:
                        periodicity = int(float(str(periodicity_raw).strip()))
                        if periodicity < 0:
                            periodicity = None
                    except (ValueError, TypeError):
                        periodicity = None

                InternalComplianceControl.objects.create(
                    requirement=requirement,
                    company=None,
                    category=last_category,
                    code=control_code,
                    name=control_name,
                    description=control_description,
                    priority=control_priority,
                    target_completion_date=target_date,
                    periodicity=periodicity,
                    implementation_notes=implementation_notes,
                    evidence_notes=evidence_notes,
                    status='not_started',
                    created_by=request.user
                )
                control_count += 1

            log_compliance_action(
                request.user, 'create', 'internal_requirement', requirement,
                changes={'action': 'imported_from_excel', 'controls_count': control_count, 'categories_count': len(categories_cache)},
                request=request
            )

            messages.success(
                request,
                _('Requirement "%(code)s" imported successfully with %(controls)d controls') % {
                    'code': requirement.code,
                    'controls': control_count
                }
            )
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

    except ValueError as exc:
        messages.error(request, str(exc))
    except Exception as exc:
        messages.error(request, _('Error importing requirement: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_requirements_library')


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirements_import_ai(request):
    """Import Internal Requirement with AI parsing from PDF"""
    from .ai_import_helpers import extract_text_from_pdf, parse_pdf_with_ai, validate_parsed_data
    from app_conf.models import Country
    
    if not check_user_internal_compliance_permission(request.user, 'can_add_requirements'):
        messages.error(request, _('You do not have permission to import requirements'))
        return redirect('compliance:internal_requirements_library')

    pdf_file = request.FILES.get('pdf_file')
    if not pdf_file:
        messages.error(request, _('No PDF file uploaded'))
        return redirect('compliance:internal_requirements_library')

    ai_model = request.POST.get('ai_model', '')
    if not ai_model:
        messages.error(request, _('AI model is required'))
        return redirect('compliance:internal_requirements_library')

    company_id = request.POST.get('company', '')
    if not company_id:
        messages.error(request, _('Company is required'))
        return redirect('compliance:internal_requirements_library')
    
    country_id = request.POST.get('country', '')
    if not country_id:
        messages.error(request, _('Country is required'))
        return redirect('compliance:internal_requirements_library')
    
    # Verify user has access to this company
    accessible_companies = get_user_accessible_companies_internal(request.user)
    try:
        company = accessible_companies.get(id=int(company_id))
    except (ValueError, Company.DoesNotExist, TypeError):
        messages.error(request, _('Invalid company selected'))
        return redirect('compliance:internal_requirements_library')
    
    # Get country for language context
    try:
        country = Country.objects.get(id=int(country_id), is_active=True)
    except (ValueError, Country.DoesNotExist, TypeError):
        messages.error(request, _('Invalid country selected'))
        return redirect('compliance:internal_requirements_library')

    try:
        # Extract text from PDF
        try:
            pdf_text = extract_text_from_pdf(pdf_file)
        except Exception as e:
            messages.error(request, _('Error reading PDF file: %(error)s') % {'error': str(e)})
            return redirect('compliance:internal_requirements_library')
        
        if not pdf_text or len(pdf_text.strip()) < 100:
            messages.error(request, _('PDF file appears to be empty or contains insufficient text. Extracted %(chars)d characters.') % {'chars': len(pdf_text.strip())})
            return redirect('compliance:internal_requirements_library')
        
        # Parse with AI (pass country for language context)
        try:
            parsed_data, usage_info = parse_pdf_with_ai(pdf_text, ai_model, country)
        except ValueError as e:
            error_msg = str(e)
            # Check if it's a rate limit or quota error
            if any(keyword in error_msg.lower() for keyword in ['429', 'rate limit', 'quota', 'exhausted', 'resource']):
                messages.error(
                    request,
                    _('AI service is currently unavailable (rate limit exceeded). Please try again later or select a different AI model.')
                )
            else:
                messages.error(request, _('AI parsing error: %(error)s') % {'error': error_msg})
            return redirect('compliance:internal_requirements_library')
        except Exception as e:
            messages.error(request, _('Unexpected AI error: %(error)s') % {'error': str(e)})
            return redirect('compliance:internal_requirements_library')
        
        # Validate and clean data
        validated_data = validate_parsed_data(parsed_data)
        
        # Store in session for preview
        request.session['ai_parsed_data'] = validated_data
        request.session['ai_company_id'] = company.id
        request.session['ai_country_id'] = country.id
        request.session['ai_model_used'] = ai_model
        
        # Redirect to preview page
        return render(request, 'app_compliance/internal_requirements_ai_preview.html', {
            'requirement_data': validated_data['requirement'],
            'controls_data': validated_data['controls'],
            'company': company,
            'country': country,
            'ai_model': ai_model,
            'usage_info': usage_info
        })
        
    except Exception as e:
        import traceback
        print(f"Unexpected error in AI import: {traceback.format_exc()}")
        messages.error(request, _('Unexpected error processing PDF: %(error)s') % {'error': str(e)})

    return redirect('compliance:internal_requirements_library')


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirements_save_ai(request):
    """Save AI-parsed requirement after preview and editing"""
    if not check_user_internal_compliance_permission(request.user, 'can_add_requirements'):
        messages.error(request, _('You do not have permission to create requirements'))
        return redirect('compliance:internal_requirements_library')

    company_id = request.POST.get('company_id', '')
    if not company_id:
        messages.error(request, _('Company is required'))
        return redirect('compliance:internal_requirements_library')
    
    # Verify user has access to this company
    accessible_companies = get_user_accessible_companies_internal(request.user)
    try:
        company = accessible_companies.get(id=int(company_id))
    except (ValueError, Company.DoesNotExist, TypeError):
        messages.error(request, _('Invalid company selected'))
        return redirect('compliance:internal_requirements_library')

    try:
        with transaction.atomic():
            # Parse dates
            def parse_date(date_str):
                if date_str:
                    try:
                        return datetime.strptime(date_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        return None
                return None
            
            # Create requirement
            requirement = InternalComplianceRequirement.objects.create(
                company=company,
                code=request.POST.get('requirement_code', ''),
                name=request.POST.get('requirement_name', ''),
                name_local=request.POST.get('requirement_name_local', ''),
                requirement_type=request.POST.get('requirement_type', 'policy'),
                description=request.POST.get('requirement_description', ''),
                status=request.POST.get('requirement_status', 'draft'),
                applicable_to=request.POST.get('requirement_applicable_to', ''),
                publication_date=parse_date(request.POST.get('publication_date')),
                effective_date=parse_date(request.POST.get('effective_date')),
                deadline_date=parse_date(request.POST.get('deadline_date')),
                priority=request.POST.get('requirement_priority', 'medium'),
                is_mandatory=request.POST.get('requirement_is_mandatory', 'false') == 'true',
                is_template=True,
                created_by=request.user
            )

            # Get control data from arrays
            control_codes = request.POST.getlist('control_code[]')
            control_names = request.POST.getlist('control_name[]')
            control_descriptions = request.POST.getlist('control_description[]')
            control_priorities = request.POST.getlist('control_priority[]')
            control_categories = request.POST.getlist('control_category[]')
            control_category_codes = request.POST.getlist('control_category_code[]')
            control_category_descriptions = request.POST.getlist('control_category_description[]')
            control_target_dates = request.POST.getlist('control_target_date[]')
            control_periodicities = request.POST.getlist('control_periodicity[]')
            control_implementation_notes = request.POST.getlist('control_implementation_notes[]')
            control_evidence_notes = request.POST.getlist('control_evidence_notes[]')

            control_count = 0
            categories_cache = {}

            for idx, code in enumerate(control_codes):
                if not code:
                    continue

                # Get or create category
                category_name = control_categories[idx] if idx < len(control_categories) else ''
                category_code = control_category_codes[idx] if idx < len(control_category_codes) else ''
                category_description = control_category_descriptions[idx] if idx < len(control_category_descriptions) else ''
                category = None
                
                if category_name and category_name.strip():
                    cache_key = f"{category_code}_{category_name}"
                    if cache_key not in categories_cache:
                        category, created = InternalRequirementCategory.objects.get_or_create(
                            requirement=requirement,
                            company=None,
                            code=category_code[:50] if category_code else category_name[:50],
                            defaults={
                                'name': category_name,
                                'description': category_description
                            }
                        )
                        if not created and category_description:
                            category.description = category_description
                            category.save()
                        categories_cache[cache_key] = category
                    else:
                        category = categories_cache[cache_key]

                # Parse target date and periodicity
                target_date = None
                if idx < len(control_target_dates) and control_target_dates[idx]:
                    target_date = parse_date(control_target_dates[idx])
                
                periodicity = None
                if idx < len(control_periodicities) and control_periodicities[idx]:
                    try:
                        periodicity = int(control_periodicities[idx])
                    except (ValueError, TypeError):
                        periodicity = None

                # Create control
                InternalComplianceControl.objects.create(
                    requirement=requirement,
                    company=None,
                    category=category,
                    code=code,
                    name=control_names[idx] if idx < len(control_names) else '',
                    description=control_descriptions[idx] if idx < len(control_descriptions) else '',
                    priority=control_priorities[idx] if idx < len(control_priorities) else 'medium',
                    target_completion_date=target_date,
                    periodicity=periodicity,
                    implementation_notes=control_implementation_notes[idx] if idx < len(control_implementation_notes) else '',
                    evidence_notes=control_evidence_notes[idx] if idx < len(control_evidence_notes) else '',
                    status='not_started',
                    created_by=request.user
                )
                control_count += 1

            log_compliance_action(
                request.user, 'create', 'internal_requirement', requirement,
                changes={'action': 'imported_from_ai', 'controls_count': control_count, 'categories_count': len(categories_cache)},
                request=request
            )

            messages.success(
                request,
                _('Requirement "%(code)s" created successfully with %(controls)d controls using AI') % {
                    'code': requirement.code,
                    'controls': control_count
                }
            )
            
            # Clear session data
            if 'ai_parsed_data' in request.session:
                del request.session['ai_parsed_data']
            if 'ai_company_id' in request.session:
                del request.session['ai_company_id']
            if 'ai_model_used' in request.session:
                del request.session['ai_model_used']
            
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

    except Exception as e:
        messages.error(request, _('Error saving requirement: %(error)s') % {'error': str(e)})

    return redirect('compliance:internal_requirements_library')


@login_required
@internal_compliance_access_required
def internal_requirement_template_create(request):
    """GET: форма створення, POST: створити requirement"""
    if not check_user_internal_compliance_permission(request.user, 'can_add_requirements'):
        messages.error(request, _('You do not have permission to create requirements'))
        return redirect('compliance:internal_requirements_library')
    
    if request.method == 'GET':
        language_code, country_for_language, use_localized_labels = get_language_preferences(request)

        # Get accessible companies for the user
        accessible_companies = get_user_accessible_companies_internal(request.user)

        requirement_types = get_dictionary_options(
            RequirementType,
            RequirementTypeTranslation,
            'requirement_type',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(InternalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
        )

        status_choices = get_dictionary_options(
            RequirementStatus,
            RequirementStatusTranslation,
            'requirement_status',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(InternalComplianceRequirement.STATUS_CHOICES)
        )

        priority_choices = get_dictionary_options(
            RequirementPriority,
            RequirementPriorityTranslation,
            'requirement_priority',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(Control.PRIORITY_CHOICES)
        )
        
        context = {
            'requirement_types': requirement_types,
            'status_choices': status_choices,
            'priority_choices': priority_choices,
            'accessible_companies': accessible_companies,
            'form_action': request.path,
            'page_title': _('Create Internal Requirement'),
            'page_subtitle': _('Create a new internal requirement'),
            'submit_label': _('Create Requirement'),
            'requirement': None,
            'default_status': 'draft',
            'default_priority': 'medium',
            'default_is_mandatory': True,
        }
        return render(request, 'app_compliance/internal_requirement_template_create.html', context)
    
    # POST: створення requirement
    try:
        # Get company (required)
        company_id = request.POST.get('company', '')
        if not company_id:
            messages.error(request, _('Company is required'))
            return redirect('compliance:internal_requirements_library')
        
        # Verify user has access to this company
        accessible_companies = get_user_accessible_companies_internal(request.user)
        try:
            company = accessible_companies.get(id=int(company_id))
        except (ValueError, Company.DoesNotExist, TypeError):
            messages.error(request, _('Invalid company selected'))
            return redirect('compliance:internal_requirements_library')
        
        # Get document if selected
        document_id = request.POST.get('document', '')
        document = None
        if document_id:
            from app_doc.models import RegisterDocs
            try:
                # Verify document belongs to selected company
                document = RegisterDocs.objects.get(id=int(document_id), company=company, is_active=True)
            except (ValueError, RegisterDocs.DoesNotExist, TypeError):
                pass  # Document not found or doesn't belong to company
        
        requirement = InternalComplianceRequirement.objects.create(
            code=request.POST.get('code'),
            name=request.POST.get('name'),
            name_local=request.POST.get('name_local', ''),
            requirement_type=request.POST.get('requirement_type', 'policy'),
            description=request.POST.get('description', ''),
            status=request.POST.get('status', 'draft'),
            applicable_to=request.POST.get('applicable_to', ''),
            is_mandatory=request.POST.get('is_mandatory') == '1',
            priority=request.POST.get('priority', 'medium'),
            company=company,
            document=document,
            is_template=True,
            created_by=request.user
        )
        
        # Handle dates
        if request.POST.get('publication_date'):
            requirement.publication_date = parse_local_requirement_date(request.POST.get('publication_date'))
        if request.POST.get('effective_date'):
            requirement.effective_date = parse_local_requirement_date(request.POST.get('effective_date'))
        if request.POST.get('deadline_date'):
            requirement.deadline_date = parse_local_requirement_date(request.POST.get('deadline_date'))
        
        requirement.save()
        
        log_compliance_action(
            request.user, 'create', 'internal_requirement', requirement,
            request=request
        )
        
        messages.success(request, _('Requirement created successfully'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)
        
    except Exception as e:
        messages.error(request, _('Error creating requirement: %(error)s') % {'error': str(e)})
        return redirect('compliance:internal_requirements_library')


@login_required
@internal_compliance_access_required
def internal_requirement_template_edit(request, requirement_id):
    """Edit existing Internal Requirement"""
    requirement = get_object_or_404(
        InternalComplianceRequirement.objects.select_related('source'),
        id=requirement_id,
        is_template=True
    )

    if not check_user_internal_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to edit requirements'))
        return redirect('compliance:internal_requirements_library')

    if request.method == 'GET':
        language_code, country_for_language, use_localized_labels = get_language_preferences(request)

        # Get accessible companies for the user
        accessible_companies = get_user_accessible_companies_internal(request.user)

        requirement_types = get_dictionary_options(
            RequirementType,
            RequirementTypeTranslation,
            'requirement_type',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(InternalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
        )

        status_choices = get_dictionary_options(
            RequirementStatus,
            RequirementStatusTranslation,
            'requirement_status',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(InternalComplianceRequirement.STATUS_CHOICES)
        )

        priority_choices = get_dictionary_options(
            RequirementPriority,
            RequirementPriorityTranslation,
            'requirement_priority',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(Control.PRIORITY_CHOICES)
        )
        
        context = {
            'requirement_types': requirement_types,
            'status_choices': status_choices,
            'priority_choices': priority_choices,
            'accessible_companies': accessible_companies,
            'form_action': request.path,
            'page_title': _('Edit Internal Requirement'),
            'page_subtitle': _('Update requirement details or metadata'),
            'submit_label': _('Save Changes'),
            'requirement': requirement,
            'default_status': requirement.status,
            'default_priority': requirement.priority,
            'default_is_mandatory': requirement.is_mandatory,
        }
        return render(request, 'app_compliance/internal_requirement_template_create.html', context)

    # POST - update requirement
    try:
        # Get company (required)
        company_id = request.POST.get('company', '')
        if not company_id:
            messages.error(request, _('Company is required'))
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)
        
        # Verify user has access to this company
        accessible_companies = get_user_accessible_companies_internal(request.user)
        try:
            requirement.company = accessible_companies.get(id=int(company_id))
        except (ValueError, Company.DoesNotExist, TypeError):
            messages.error(request, _('Invalid company selected'))
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)
        
        requirement.code = request.POST.get('code', requirement.code)
        requirement.name = request.POST.get('name', requirement.name)
        requirement.name_local = request.POST.get('name_local', '')
        requirement.requirement_type = request.POST.get('requirement_type', requirement.requirement_type)
        requirement.description = request.POST.get('description', '')
        requirement.status = request.POST.get('status', requirement.status)
        requirement.applicable_to = request.POST.get('applicable_to', '')
        requirement.is_mandatory = request.POST.get('is_mandatory') == '1'
        requirement.priority = request.POST.get('priority', requirement.priority)
        
        # Get document if selected
        document_id = request.POST.get('document', '')
        if document_id:
            from app_doc.models import RegisterDocs
            try:
                # Verify document belongs to selected company
                requirement.document = RegisterDocs.objects.get(id=int(document_id), company=requirement.company, is_active=True)
            except (ValueError, RegisterDocs.DoesNotExist, TypeError):
                requirement.document = None
        else:
            requirement.document = None

        publication_date = request.POST.get('publication_date')
        effective_date = request.POST.get('effective_date')
        deadline_date = request.POST.get('deadline_date')

        requirement.publication_date = parse_local_requirement_date(publication_date) if publication_date else None
        requirement.effective_date = parse_local_requirement_date(effective_date) if effective_date else None
        requirement.deadline_date = parse_local_requirement_date(deadline_date) if deadline_date else None

        requirement.save()

        log_compliance_action(
            request.user,
            'update',
            'internal_requirement',
            requirement,
            request=request
        )

        messages.success(request, _('Requirement updated successfully'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

    except Exception as e:
        messages.error(request, _('Error updating requirement: %(error)s') % {'error': str(e)})
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_template_delete(request, requirement_id):
    """Delete Internal Requirement template"""
    if not check_user_internal_compliance_permission(request.user, 'can_delete_requirements'):
        messages.error(request, _('You do not have permission to delete requirements'))
        return redirect('compliance:internal_requirements_library')
    
    requirement = get_object_or_404(
        InternalComplianceRequirement,
        id=requirement_id,
        is_template=True
    )
    
    try:
        # Check if requirement has instances
        instances_count = requirement.instances.count()
        if instances_count > 0:
            messages.error(
                request,
                _('Cannot delete requirement: %(count)s instance(s) exist. Delete instances first.') % {'count': instances_count}
            )
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)
        
        # Log the deletion
        log_compliance_action(
            request.user,
            'delete',
            'internal_requirement',
            requirement,
            changes=f'Deleted requirement: {requirement.code} - {requirement.name}',
            request=request
        )
        
        requirement_code = requirement.code
        requirement.delete()
        
        messages.success(request, _('Requirement "%(code)s" deleted successfully') % {'code': requirement_code})
        
    except Exception as e:
        messages.error(request, _('Error deleting requirement: %(error)s') % {'error': str(e)})
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)
    
    return redirect('compliance:internal_requirements_library')


@login_required
@internal_compliance_access_required
def get_internal_company_documents(request):
    """Get RegisterDocs for a specific company (AJAX endpoint for Internal Compliance)"""
    try:
        company_id = request.GET.get('company_id')
        if not company_id:
            return JsonResponse({
                'success': False,
                'error': _('Company ID is required')
            }, status=400)

        # Check if user has access to this company using Internal Compliance access
        accessible_companies = get_user_accessible_companies_internal(request.user)
        try:
            company = accessible_companies.get(id=int(company_id))
        except (ValueError, Company.DoesNotExist, TypeError):
            return JsonResponse({
                'success': False,
                'error': _('Access denied to this company')
            }, status=403)

        # Get user's groups for access control
        user_groups = request.user.groups.all()
        
        # Filter documents by company and user's group access
        from app_doc.models import RegisterDocs
        documents = RegisterDocs.objects.filter(
            company=company,
            groups__in=user_groups,
            is_active=True
        ).distinct().select_related('company', 'type_doc', 'status_doc').order_by('name_doc')

        documents_data = [{
            'id': doc.id,
            'name_doc': doc.name_doc,
            'vers_doc': doc.vers_doc or '',
            'date_doc': doc.date_doc.strftime('%Y-%m-%d') if doc.date_doc else '',
            'type_doc': doc.type_doc.get_name_by_language() if doc.type_doc else '',
            'status_doc': doc.status_doc.get_name_by_language() if doc.status_doc else '',
            'company_name': doc.company.name if doc.company else ''
        } for doc in documents]

        return JsonResponse({
            'success': True,
            'documents': documents_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@internal_compliance_access_required
def internal_requirement_template_detail(request, requirement_id):
    """Details of requirement template with controls"""
    requirement = get_object_or_404(
        InternalComplianceRequirement.objects.select_related('created_by', 'company'),
        id=requirement_id,
        is_template=True
    )

    if not check_user_internal_compliance_permission(request.user, 'can_view_requirements'):
        messages.error(request, _('You do not have permission to view requirements'))
        return redirect('compliance:internal_requirements_library')

    # Get filter parameters
    search = request.GET.get('search', '')
    status_filter = request.GET.get('control_status', '')
    priority_filter = request.GET.get('priority', '')

    # Get requirement controls (without company)
    template_controls_qs = requirement.controls.filter(company__isnull=True).select_related(
        'category',
        'responsible'
    )
    
    # Apply filters
    if search:
        template_controls_qs = template_controls_qs.filter(
            Q(code__icontains=search) |
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    if status_filter:
        template_controls_qs = template_controls_qs.filter(status=status_filter)
    if priority_filter:
        template_controls_qs = template_controls_qs.filter(priority=priority_filter)
    
    template_controls_qs = template_controls_qs.order_by('code')
    
    # Get categories with prefetched controls
    categories = list(
        requirement.categories.order_by('order', 'code').prefetch_related(
            Prefetch(
                'controls',
                queryset=template_controls_qs,
                to_attr='template_controls'
            )
        )
    )
    
    # Get uncategorized controls
    uncategorized_controls = list(template_controls_qs.filter(category__isnull=True))
    
    # Get instances
    instances = requirement.instances.all().select_related('company')
    
    permissions = get_user_internal_compliance_permissions(request.user)
    
    # Statistics
    completion = requirement.get_completion_percentage()
    stats = requirement.get_controls_by_status()
    
    # Get filter choices
    all_controls = requirement.controls.filter(company__isnull=True)
    existing_statuses = all_controls.values_list('status', flat=True).distinct()
    status_choices = [(value, label) for value, label in Control.STATUS_CHOICES if value in existing_statuses]
    
    existing_priorities = all_controls.values_list('priority', flat=True).distinct()
    priority_choices = [(value, label) for value, label in Control.PRIORITY_CHOICES if value in existing_priorities]
    
    # Get notes for this requirement
    from .models import InternalRequirementNote
    requirement_notes = InternalRequirementNote.objects.filter(
        requirement=requirement,
        is_active=True
    ).select_related('created_by').order_by('-created_date')
    
    context = {
        'requirement': requirement,
        'template_controls': template_controls_qs,
        'categories': categories,
        'uncategorized_controls': uncategorized_controls,
        'instances': instances,
        'permissions': permissions,
        'completion': completion,
        'stats': stats,
        'search': search,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'status_choices': status_choices,
        'priority_choices': priority_choices,
        'requirement_notes': requirement_notes,
    }
    
    return render(request, 'app_compliance/internal_requirement_template_detail.html', context)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_note_create(request, requirement_id):
    """Create note for internal requirement"""
    requirement = get_object_or_404(InternalComplianceRequirement, id=requirement_id, is_template=True)

    if not check_user_internal_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to add notes'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)

        from .models import InternalRequirementNote, InternalRequirementNoteAttachment
        note = InternalRequirementNote.objects.create(
            requirement=requirement,
            note=note_text,
            attachment=request.FILES.get('attachment'),
            created_by=request.user,
        )

        # Multiple attachments
        for f in request.FILES.getlist('attachments'):
            InternalRequirementNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'create',
            'internal_requirement_note',
            note,
            request=request,
        )

        messages.success(request, _('Note added successfully'))

    except Exception as exc:
        messages.error(request, _('Error adding note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_note_update(request, note_id):
    """Update internal requirement note"""
    from .models import InternalRequirementNote, InternalRequirementNoteAttachment
    note = get_object_or_404(InternalRequirementNote, id=note_id)
    requirement = note.requirement

    if not check_user_internal_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to edit notes'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

    try:
        note_text = (request.POST.get('note') or '').strip()
        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

        note.note = note_text

        # Clear attachment if requested
        if request.POST.get('clear_attachment') == 'on':
            note.attachment = None

        # New single attachment
        if 'attachment' in request.FILES:
            note.attachment = request.FILES['attachment']

        note.save()

        # Multiple new attachments
        for f in request.FILES.getlist('attachments'):
            InternalRequirementNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'update',
            'internal_requirement_note',
            note,
            request=request,
        )

        messages.success(request, _('Note updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_note_delete(request, note_id):
    """Delete internal requirement note"""
    from .models import InternalRequirementNote
    note = get_object_or_404(InternalRequirementNote, id=note_id)
    requirement = note.requirement

    if not check_user_internal_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to delete notes'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

    try:
        log_compliance_action(
            request.user,
            'delete',
            'internal_requirement_note',
            note,
            changes=f'Deleted note: {note.note[:50]}...',
            request=request,
        )

        note.delete()
        messages.success(request, _('Note deleted successfully'))

    except Exception as exc:
        messages.error(request, _('Error deleting note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_note_attachment_delete(request, attachment_id):
    """Delete an attachment from internal requirement note"""
    from .models import InternalRequirementNoteAttachment
    attachment = get_object_or_404(InternalRequirementNoteAttachment, id=attachment_id)
    note = attachment.note
    requirement = note.requirement

    if not check_user_internal_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to delete attachments'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)

    try:
        attachment.delete()
        messages.success(request, _('Attachment deleted successfully'))
    except Exception as exc:
        messages.error(request, _('Error deleting attachment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement.id)


@login_required
@internal_compliance_access_required
def internal_control_detail(request, control_id):
    """Деталі внутрішнього контролю з повним контекстом"""
    control = get_object_or_404(
        InternalComplianceControl.objects.select_related(
            'requirement__company',
            'responsible__cabinet__position',
            'responsible__cabinet__department',
            'verified_by',
            'created_by',
            'category'
        ),
        id=control_id
    )

    is_instance = bool(control.company)
    # Для instance controls використовуємо can_view_requirement_instances, для template - can_view_controls
    permission_key = 'can_view_requirement_instances' if is_instance else 'can_view_controls'

    if not check_user_internal_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have access to this control'))
        return redirect('compliance:internal_compliance')

    if control.company:
        accessible_companies = get_user_accessible_companies_internal(request.user)
        if control.company not in accessible_companies:
            messages.error(request, _('You do not have access to this company'))
            return redirect('compliance:internal_compliance')

    logs = ComplianceAuditLog.objects.filter(
        object_type='internal_control',
        object_id=control.id
    ).select_related('user').order_by('-timestamp')[:50]

    evidences = control.evidences.filter(is_active=True).select_related('uploaded_by', 'reviewed_by')
    assignments = control.assignments.filter(is_active=True).select_related(
        'user__cabinet__position',
        'user__cabinet__department',
        'assigned_by'
    )
    notes = control.notes.filter(is_active=True).select_related('created_by')
    mappings = InternalControlMapping.objects.filter(
        internal_control=control
    ).select_related(
        'target_internal_control__requirement__company',
        'target_local_control__requirement',
        'target_framework_control__category__framework',
        'created_by'
    )

    from app_cabinet.models import CabinetUser

    if control.company:
        cabinet_users = CabinetUser.objects.filter(
            company=control.company,
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')
    else:
        cabinet_users = CabinetUser.objects.filter(
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')

    users = [{
        'id': cu.user.id,
        'username': cu.user.username,
        'email': cu.user.email,
        'full_name': cu.user.get_full_name(),
        'position': str(cu.position) if cu.position else None,
        'department': str(cu.department) if cu.department else None,
    } for cu in cabinet_users]

    # Get framework controls - only from frameworks applied to this company
    # Determine the company: use control.company if exists, otherwise use control.requirement.company
    target_company = control.company or control.requirement.company
    
    if target_company:
        # Filter only controls from framework instances applied to this company
        # Must be: company matches AND is_template=False (instance, not template)
        company_id = target_company.id
        framework_controls = Control.objects.filter(
            category__framework__company_id=company_id,
            category__framework__is_template=False
        ).exclude(
            id=control.related_framework_control_id
        ).select_related(
            'category__framework'
        ).distinct().order_by(
            'category__framework__name',
            'category__code',
            'code'
        )[:150]
    else:
        # For template controls without company, show all framework controls
        framework_controls = Control.objects.exclude(
            id=control.related_framework_control_id
        ).select_related(
            'category__framework'
        ).order_by(
            'category__framework__name',
            'category__code',
            'code'
        )[:150]

    # Use target_company for filtering other controls too
    internal_controls_for_mapping = InternalComplianceControl.objects.filter(
        requirement__company=target_company
    ).exclude(id=control.id).order_by('code')[:150] if target_company else InternalComplianceControl.objects.none()
    
    # Get local controls for the same company
    local_controls_for_mapping = []
    if target_company:
        from .models import LocalComplianceControl
        local_controls_for_mapping = LocalComplianceControl.objects.filter(
            company=target_company
        ).select_related('requirement').order_by('code')[:150]

    permissions = get_user_internal_compliance_permissions(request.user)
    # Для редагування контролів завжди використовуємо can_edit_controls
    can_edit = permissions['can_edit_controls']

    # Get available mandatory processes for the company
    from app_compliance.models import MandatoryProcess
    from app_doc.models import RegisterDocs, RelatedDocs
    mandatory_processes = MandatoryProcess.objects.filter(
        is_active=True
    ).select_related('company', 'source_document').order_by('process_name')
    
    # Filter by company if control has one
    if target_company:
        mandatory_processes = mandatory_processes.filter(
            Q(company=target_company) | Q(company__isnull=True)
        )
    
    # Get register documents
    register_docs = RegisterDocs.objects.filter(
        is_active=True
    ).exclude(
        file_doc=''
    ).select_related('status_doc', 'company').order_by('name_doc')
    if target_company:
        register_docs = register_docs.filter(
            Q(company=target_company) | Q(company__isnull=True)
        )
    
    # Get related documents
    related_docs = RelatedDocs.objects.exclude(
        file_rel_doc=''
    ).select_related('status_rel_doc', 'company').order_by('name_rel_doc')
    if target_company:
        related_docs = related_docs.filter(
            Q(company=target_company) | Q(company__isnull=True)
        )

    # Get language preferences for localization
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    
    # Get active evidence types with localized names
    evidence_types_qs = EvidenceType.objects.filter(is_active=True).order_by('display_order', 'name')
    
    # Build localized evidence types list
    evidence_types = []
    if country_for_language and use_localized_labels:
        # Get translations for the country
        translations = {
            et.evidence_type.code: et.name_local
            for et in EvidenceTypeTranslation.objects.filter(
                country=country_for_language,
                evidence_type__in=evidence_types_qs
            ).select_related('evidence_type')
        }
        for et in evidence_types_qs:
            localized_name = translations.get(et.code) or (et.name_local if language_code == 'uk' else et.name)
            evidence_types.append({
                'id': et.id,
                'name': localized_name,
                'code': et.code,
                'color': et.color
            })
    else:
        # Use English names
        for et in evidence_types_qs:
            evidence_types.append({
                'id': et.id,
                'name': et.name,
                'code': et.code,
                'color': et.color
            })

    context = {
        'control': control,
        'requirement': control.requirement,
        'evidences': evidences,
        'assignments': assignments,
        'notes': notes,
        'logs': logs,
        'mappings': mappings,
        'users': users,
        'framework_controls': framework_controls,
        'internal_controls_for_mapping': internal_controls_for_mapping,
        'local_controls_for_mapping': local_controls_for_mapping,
        'mandatory_processes': mandatory_processes,
        'register_docs': register_docs,
        'related_docs': related_docs,
        'permissions': permissions,
        'can_edit': can_edit,
        'is_instance': is_instance,
        'has_sufficient_evidence': control.has_sufficient_evidence(),
        'evidence_count': control.get_evidence_count(),
        'approved_evidence_count': control.get_approved_evidence_count(),
        'Control': Control,
        'InternalControlEvidence': InternalControlEvidence,
        'InternalControlAssignment': InternalControlAssignment,
        'InternalControlMapping': InternalControlMapping,
        'evidence_types': evidence_types,
    }

    return render(request, 'app_compliance/internal_control_detail.html', context)


# ========================
# Internal Requirement Categories CRUD
# ========================

@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_category_create(request, requirement_id):
    """Create category for requirement template"""
    if not check_user_internal_compliance_permission(request.user, 'can_add_controls'):
        messages.error(request, _('You do not have permission to add categories'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)

    requirement = get_object_or_404(InternalComplianceRequirement, id=requirement_id, is_template=True)

    try:
        order_value = request.POST.get('order') or 0
        order_value = int(order_value)
    except ValueError:
        order_value = 0

    try:
        category = InternalRequirementCategory.objects.create(
            requirement=requirement,
            code=request.POST.get('code', '').strip() or _('CAT'),
            name=request.POST.get('name', '').strip() or _('New Category'),
            description=request.POST.get('description', '').strip(),
            order=order_value
        )

        log_compliance_action(
            request.user, 'create', 'internal_requirement_category', category,
            request=request
        )

        messages.success(request, _('Category created successfully'))
    except Exception as e:
        messages.error(request, _('Error creating category: %(error)s') % {'error': str(e)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_category_update(request, category_id):
    """Update category"""
    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit categories'))
        category = get_object_or_404(InternalRequirementCategory, id=category_id)
        return redirect('compliance:internal_requirement_template_detail', requirement_id=category.requirement.id)

    category = get_object_or_404(InternalRequirementCategory, id=category_id, requirement__is_template=True)
    requirement_id = category.requirement.id

    try:
        old_values = {
            'code': category.code,
            'name': category.name,
            'description': category.description,
            'order': category.order,
        }

        category.code = request.POST.get('code', category.code).strip()
        category.name = request.POST.get('name', category.name).strip()
        category.description = request.POST.get('description', category.description).strip()
        
        try:
            order_value = request.POST.get('order') or category.order
            category.order = int(order_value)
        except ValueError:
            pass

        category.save()

        new_values = {
            'code': category.code,
            'name': category.name,
            'description': category.description,
            'order': category.order,
        }

        log_compliance_action(
            request.user, 'update', 'internal_requirement_category', category,
            changes={'old': old_values, 'new': new_values},
            request=request
        )

        messages.success(request, _('Category updated successfully'))

    except Exception as e:
        messages.error(request, _('Error updating category: %(error)s') % {'error': str(e)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_category_delete(request, category_id):
    """Delete category"""
    if not check_user_internal_compliance_permission(request.user, 'can_delete_controls'):
        messages.error(request, _('You do not have permission to delete categories'))
        return redirect('compliance:internal_requirements_library')

    category = get_object_or_404(InternalRequirementCategory, id=category_id, requirement__is_template=True)
    requirement_id = category.requirement_id

    try:
        log_compliance_action(
            request.user, 'delete', 'internal_requirement_category', category,
            request=request
        )
        category.delete()
        messages.success(request, _('Category deleted successfully'))
    except Exception as e:
        messages.error(request, _('Error deleting category: %(error)s') % {'error': str(e)})

    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


# ========================
# Internal Requirement Controls CRUD
# ========================

@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_control_create(request, requirement_id):
    """Create control for requirement template"""
    if not check_user_internal_compliance_permission(request.user, 'can_add_controls'):
        messages.error(request, _('You do not have permission to add controls'))
        return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)
    
    requirement = get_object_or_404(InternalComplianceRequirement, id=requirement_id, is_template=True)
    
    try:
        category = None
        category_id = request.POST.get('category')
        if category_id:
            category = InternalRequirementCategory.objects.filter(
                id=category_id,
                requirement=requirement
            ).first()

        def parse_required_count(value):
            try:
                return max(0, int(value))
            except (TypeError, ValueError):
                return 1
        
        def parse_periodicity(value):
            try:
                periodicity_val = int(value)
                return periodicity_val if periodicity_val > 0 else None
            except (TypeError, ValueError):
                return None

        required_evidence_count = parse_required_count(request.POST.get('required_evidence_count', 1))
        periodicity = parse_periodicity(request.POST.get('periodicity'))

        control = InternalComplianceControl.objects.create(
            requirement=requirement,
            company=None,  # Template control without company
            code=request.POST.get('code'),
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            status='not_started',
            priority=request.POST.get('priority', 'medium'),
            implementation_notes=request.POST.get('implementation_notes', ''),
            evidence_notes=request.POST.get('evidence_notes', ''),
            required_evidence_count=required_evidence_count,
            periodicity=periodicity,
            category=category,
            created_by=request.user
        )
        
        # Handle target_completion_date
        if request.POST.get('target_completion_date'):
            try:
                control.target_completion_date = parse_local_requirement_date(
                    request.POST.get('target_completion_date')
                )
                control.save()
            except ValueError:
                pass
        
        log_compliance_action(
            request.user, 'create', 'internal_control', control,
            request=request
        )
        
        messages.success(request, _('Control created successfully'))
        
    except Exception as e:
        messages.error(request, f'Error creating control: {str(e)}')
    
    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_control_update(request, control_id):
    """Update template or instance control"""
    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit controls'))
        control = get_object_or_404(InternalComplianceControl, id=control_id)
        if control.company:
            return redirect('compliance:internal_control_detail', control_id=control_id)
        return redirect('compliance:internal_requirement_template_detail', requirement_id=control.requirement.id)
    
    control = get_object_or_404(InternalComplianceControl, id=control_id)
    requirement_id = control.requirement.id
    is_instance = bool(control.company)
    
    try:
        old_values = {
            'code': control.code,
            'name': control.name,
            'description': control.description,
            'priority': control.priority,
            'category': control.category.id if control.category else None,
        }
        
        control.code = request.POST.get('code', control.code).strip()
        control.name = request.POST.get('name', control.name).strip()
        control.description = request.POST.get('description', control.description).strip()
        control.priority = request.POST.get('priority', control.priority)
        control.implementation_notes = request.POST.get('implementation_notes', control.implementation_notes).strip()
        control.evidence_notes = request.POST.get('evidence_notes', control.evidence_notes).strip()
        
        try:
            control.required_evidence_count = max(0, int(request.POST.get('required_evidence_count', control.required_evidence_count)))
        except (TypeError, ValueError):
            pass
        
        # Handle periodicity
        periodicity_val = request.POST.get('periodicity')
        if periodicity_val:
            try:
                periodicity_int = int(periodicity_val)
                control.periodicity = periodicity_int if periodicity_int > 0 else None
            except (TypeError, ValueError):
                pass
        elif request.POST.get('periodicity') == '':
            control.periodicity = None
        
        # Handle category
        category_id = request.POST.get('category')
        if category_id and category_id != '0':
            try:
                category_id_int = int(category_id)
                if category_id_int > 0:
                    category = InternalRequirementCategory.objects.filter(
                        id=category_id_int,
                        requirement=control.requirement
                    ).first()
                    control.category = category
            except (ValueError, TypeError):
                pass
        else:
            control.category = None
        
        # Handle target_completion_date
        target_date = request.POST.get('target_completion_date')
        if target_date:
            try:
                control.target_completion_date = parse_local_requirement_date(target_date)
            except ValueError:
                pass
        elif request.POST.get('target_completion_date') == '':
            control.target_completion_date = None
        
        control.save()
        
        new_values = {
            'code': control.code,
            'name': control.name,
            'description': control.description,
            'priority': control.priority,
            'category': control.category.id if control.category else None,
        }
        
        log_compliance_action(
            request.user, 'update', 'internal_control', control,
            changes={'old': old_values, 'new': new_values},
            request=request
        )
        
        messages.success(request, _('Control updated successfully'))
        
    except Exception as e:
        messages.error(request, f'Error updating control: {str(e)}')
    
    # Redirect based on control type
    if is_instance:
        return redirect('compliance:internal_control_detail', control_id=control_id)
    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_requirement_control_delete(request, control_id):
    """Delete template control"""
    if not check_user_internal_compliance_permission(request.user, 'can_delete_controls'):
        messages.error(request, _('You do not have permission to delete controls'))
        return redirect('compliance:internal_requirements_library')
    
    control = get_object_or_404(InternalComplianceControl, id=control_id, company__isnull=True)
    requirement_id = control.requirement.id
    
    try:
        log_compliance_action(
            request.user, 'delete', 'internal_control', control,
            request=request
        )
        
        control.delete()
        messages.success(request, _('Control deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting control: {str(e)}')
    
    return redirect('compliance:internal_requirement_template_detail', requirement_id=requirement_id)


# ========================
# Internal Control Operations (Assign, Evidence, Notes, Mapping)
# ========================

@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_assign(request, control_id):
    """Assign user to internal control"""
    control = get_object_or_404(InternalComplianceControl, id=control_id)

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to assign users'))
        return redirect('compliance:internal_control_detail', control_id=control_id)

    user_id = request.POST.get('user_id')
    assignment_type = request.POST.get('assignment_type', 'owner')
    notes = request.POST.get('notes', '')

    if not user_id:
        messages.error(request, _('Please select a user to assign'))
        return redirect('compliance:internal_control_detail', control_id=control_id)

    try:
        assignment, created = InternalControlAssignment.objects.get_or_create(
            control=control,
            user_id=user_id,
            assignment_type=assignment_type,
            defaults={
                'assigned_by': request.user,
                'notes': notes,
                'is_active': True,
            }
        )

        if not created:
            assignment.is_active = True
            assignment.notes = notes or assignment.notes
            assignment.assigned_by = request.user
            assignment.save()

        log_compliance_action(
            request.user,
            'create' if created else 'update',
            'internal_control_assignment',
            assignment,
            request=request
        )

        messages.success(request, _('User assigned successfully'))

    except Exception as exc:
        messages.error(request, _('Error assigning user: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_set_responsible(request, control_id):
    """Set responsible person for internal control"""
    control = get_object_or_404(InternalComplianceControl, id=control_id)
    
    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to set responsible'))
        return redirect('compliance:internal_control_detail', control_id=control.id)
    
    try:
        old_responsible = control.responsible.username if control.responsible else None
        
        responsible_id = request.POST.get('responsible_id')
        if responsible_id:
            from django.contrib.auth.models import User
            control.responsible = User.objects.get(id=responsible_id)
            new_responsible = control.responsible.username
        else:
            control.responsible = None
            new_responsible = None
        
        control.save()
        
        log_compliance_action(
            request.user,
            'update',
            'internal_control',
            control,
            changes={'old': {'responsible': old_responsible}, 'new': {'responsible': new_responsible}},
            request=request
        )
        
        if new_responsible:
            messages.success(request, _('Responsible person set successfully'))
        else:
            messages.success(request, _('Responsible person cleared'))
            
    except Exception as e:
        messages.error(request, _('Error setting responsible: %(error)s') % {'error': str(e)})
    
    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_update_status(request, control_id):
    """Update status for internal control (AJAX endpoint)"""
    import json
    from django.http import JsonResponse
    from django.utils import timezone
    
    control = get_object_or_404(InternalComplianceControl, id=control_id)
    
    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        return JsonResponse({
            'success': False,
            'error': str(_('You do not have permission to edit controls'))
        }, status=403)
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if not new_status:
            return JsonResponse({
                'success': False,
                'error': str(_('Status is required'))
            }, status=400)
        
        # Validate status
        valid_statuses = [choice[0] for choice in Control.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({
                'success': False,
                'error': str(_('Invalid status'))
            }, status=400)
        
        old_status = control.status
        control.status = new_status
        control.status_changed_date = timezone.now()
        
        # If status is completed, set actual completion date
        if new_status == 'completed' and not control.actual_completion_date:
            control.actual_completion_date = timezone.now().date()
        
        control.save(update_fields=['status', 'status_changed_date', 'actual_completion_date'])
        
        # Log the change
        log_compliance_action(
            request.user,
            'update',
            'internal_control',
            control,
            changes=f'Status changed from {old_status} to {new_status}',
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': str(_('Status updated successfully')),
            'status': new_status,
            'status_display': control.get_status_display()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': str(_('Invalid JSON data'))
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_assignment_delete(request, assignment_id):
    """Delete (deactivate) assignment"""
    assignment = get_object_or_404(InternalControlAssignment, id=assignment_id)
    control = assignment.control

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to remove assignments'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        assignment.is_active = False
        assignment.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'internal_control_assignment',
            assignment,
            request=request
        )

        messages.success(request, _('Assignment removed'))

    except Exception as exc:
        messages.error(request, _('Error removing assignment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_evidence_create(request, control_id):
    """Add evidence to internal control"""
    control = get_object_or_404(InternalComplianceControl, id=control_id)

    if not check_user_internal_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to add evidence'))
        return redirect('compliance:internal_control_detail', control_id=control_id)

    try:
        mandatory_process_id = request.POST.get('mandatory_process_id')
        mandatory_process = None
        if mandatory_process_id:
            from app_compliance.models import MandatoryProcess
            mandatory_process = MandatoryProcess.objects.filter(id=mandatory_process_id).first()
        
        # Get EvidenceType by ID or code, fallback to 'document' code
        evidence_type_id = request.POST.get('evidence_type', '').strip()
        evidence_type = None
        if evidence_type_id:
            try:
                # Try to get by ID first (convert to int if possible)
                try:
                    evidence_type_id_int = int(evidence_type_id)
                    evidence_type = EvidenceType.objects.get(id=evidence_type_id_int)
                except (ValueError, TypeError):
                    # If not a number, try by code
                    evidence_type = EvidenceType.objects.get(code=evidence_type_id)
            except EvidenceType.DoesNotExist:
                # Fallback to default 'document'
                evidence_type = EvidenceType.objects.filter(code='document').first()
        else:
            # Fallback to default 'document'
            evidence_type = EvidenceType.objects.filter(code='document').first()
        
        if not evidence_type:
            messages.error(request, _('Invalid evidence type selected'))
            return redirect('compliance:internal_control_detail', control_id=control_id)
        
        evidence = InternalControlEvidence.objects.create(
            control=control,
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            evidence_type=evidence_type,
            text_evidence=request.POST.get('text_evidence', ''),
            external_link=request.POST.get('external_link', ''),
            mandatory_process=mandatory_process,
            uploaded_by=request.user
        )

        # Handle file: priority: uploaded file > register_doc > related_doc
        if 'file' in request.FILES:
            evidence.file = request.FILES['file']
        else:
            # Check for register document
            register_doc_id = request.POST.get('register_doc_id')
            if register_doc_id:
                from app_doc.models import RegisterDocs
                register_doc = RegisterDocs.objects.filter(id=register_doc_id).first()
                if register_doc and register_doc.file_doc:
                    evidence.file = register_doc.file_doc
                    evidence.register_document = register_doc
            else:
                # Check for related document
                related_doc_id = request.POST.get('related_doc_id')
                if related_doc_id:
                    from app_doc.models import RelatedDocs
                    related_doc = RelatedDocs.objects.filter(id=related_doc_id).first()
                    if related_doc and related_doc.file_rel_doc:
                        evidence.file = related_doc.file_rel_doc
                        evidence.related_document = related_doc
        
        evidence.save()

        log_compliance_action(
            request.user,
            'create',
            'internal_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence uploaded successfully'))

    except Exception as exc:
        messages.error(request, _('Error uploading evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control_id)


@login_required
@internal_compliance_access_required
def internal_control_evidence_edit(request, evidence_id):
    """Get evidence data for editing or update it"""
    evidence = get_object_or_404(InternalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_internal_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to edit evidence'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    if request.method == 'GET':
        evidence_data = {
            'id': evidence.id,
            'title': evidence.title,
            'description': evidence.description,
            'evidence_type': evidence.evidence_type_id if evidence.evidence_type_id else None,
            'text_evidence': evidence.text_evidence,
            'external_link': evidence.external_link,
            'file_name': evidence.file.name if evidence.file else '',
            'mandatory_process_id': evidence.mandatory_process_id if evidence.mandatory_process else None,
        }
        return JsonResponse(evidence_data)

    return internal_control_evidence_update(request, evidence_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_evidence_update(request, evidence_id):
    """Update evidence"""
    evidence = get_object_or_404(InternalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_internal_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to edit evidence'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        evidence.title = request.POST.get('title', evidence.title)
        evidence.description = request.POST.get('description', evidence.description)
        evidence.evidence_type = request.POST.get('evidence_type', evidence.evidence_type)
        evidence.text_evidence = request.POST.get('text_evidence', evidence.text_evidence)
        evidence.external_link = request.POST.get('external_link', evidence.external_link)
        
        # Handle mandatory process update
        mandatory_process_id = request.POST.get('mandatory_process_id')
        if mandatory_process_id:
            from app_compliance.models import MandatoryProcess
            evidence.mandatory_process = MandatoryProcess.objects.filter(id=mandatory_process_id).first()
        else:
            evidence.mandatory_process = None

        if 'file' in request.FILES:
            evidence.file = request.FILES['file']

        evidence.save()

        log_compliance_action(
            request.user,
            'update',
            'internal_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_evidence_delete(request, evidence_id):
    """Delete evidence"""
    evidence = get_object_or_404(InternalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_internal_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to delete evidence'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        evidence.is_active = False
        evidence.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'internal_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_evidence_approve(request, evidence_id):
    """Approve evidence"""
    evidence = get_object_or_404(InternalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_internal_compliance_permission(request.user, 'can_approve_evidence'):
        messages.error(request, _('You do not have permission to approve evidence'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        evidence.approval_status = 'approved'
        evidence.reviewed_by = request.user
        evidence.reviewed_date = timezone.now()
        evidence.save()

        log_compliance_action(
            request.user,
            'approve',
            'internal_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence approved'))

    except Exception as exc:
        messages.error(request, _('Error approving evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_evidence_reject(request, evidence_id):
    """Reject evidence"""
    evidence = get_object_or_404(InternalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_internal_compliance_permission(request.user, 'can_approve_evidence'):
        messages.error(request, _('You do not have permission to reject evidence'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        evidence.approval_status = 'rejected'
        evidence.reviewed_by = request.user
        evidence.reviewed_date = timezone.now()
        evidence.review_comments = request.POST.get('review_comments', '')
        evidence.save()

        log_compliance_action(
            request.user,
            'reject',
            'internal_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence rejected'))

    except Exception as exc:
        messages.error(request, _('Error rejecting evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_note_create(request, control_id):
    """Create note for internal control"""
    control = get_object_or_404(InternalComplianceControl, id=control_id)

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to add notes'))
        return redirect('compliance:internal_control_detail', control_id=control_id)

    try:
        note = InternalControlNote.objects.create(
            control=control,
            note=request.POST.get('note'),
            created_by=request.user
        )
        
        # Multiple attachments
        from .models import InternalControlNoteAttachment
        for f in request.FILES.getlist('attachments'):
            InternalControlNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'create',
            'internal_control_note',
            note,
            request=request
        )

        messages.success(request, _('Note added successfully'))

    except Exception as exc:
        messages.error(request, _('Error adding note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control_id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_note_update(request, note_id):
    """Update note for internal control"""
    note = get_object_or_404(InternalControlNote, id=note_id, is_active=True)
    control = note.control

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit notes'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:internal_control_detail', control_id=control.id)

        note.note = note_text

        # Clear existing single attachment if requested
        if request.POST.get('clear_attachment') == '1' and note.attachment:
            note.attachment = None

        note.save()

        # Append new attachments if provided
        from .models import InternalControlNoteAttachment
        for f in request.FILES.getlist('attachments'):
            InternalControlNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'update',
            'internal_control_note',
            note,
            request=request
        )

        messages.success(request, _('Note updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_note_delete(request, note_id):
    """Delete note"""
    note = get_object_or_404(InternalControlNote, id=note_id)
    control = note.control

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to delete notes'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        note.is_active = False
        note.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'internal_control_note',
            note,
            request=request
        )

        messages.success(request, _('Note deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_mapping_create(request):
    """Create mapping for internal control"""
    def clean_id(value):
        """Clean and validate ID value"""
        if not value:
            return None
        cleaned = ''.join(c for c in str(value).strip() if c.isdigit())
        if not cleaned:
            return None
        try:
            return int(cleaned)
        except (ValueError, TypeError):
            return None

    control_id = clean_id(request.POST.get('internal_control_id'))
    target_internal_id = clean_id(request.POST.get('target_internal_control_id'))
    target_local_id = clean_id(request.POST.get('target_local_control_id'))
    target_framework_id = clean_id(request.POST.get('target_framework_control_id'))

    if not control_id:
        messages.error(request, _('Internal control is required'))
        return redirect('compliance:internal_compliance')

    control = get_object_or_404(InternalComplianceControl, id=control_id)

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to create mappings'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    if not target_internal_id and not target_local_id and not target_framework_id:
        messages.error(request, _('Please select a target control'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        mapping_type = request.POST.get('mapping_type', 'related')
        notes = request.POST.get('notes', '')

        mapping = InternalControlMapping.objects.create(
            internal_control=control,
            target_internal_control_id=target_internal_id,
            target_local_control_id=target_local_id,
            target_framework_control_id=target_framework_id,
            mapping_type=mapping_type,
            notes=notes,
            created_by=request.user
        )

        log_compliance_action(
            request.user,
            'create',
            'internal_control_mapping',
            mapping,
            request=request
        )

        messages.success(request, _('Mapping created successfully'))

    except Exception as exc:
        messages.error(request, _('Error creating mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@internal_compliance_access_required
@require_http_methods(["POST"])
def internal_control_mapping_delete(request, mapping_id):
    """Delete mapping"""
    mapping = get_object_or_404(InternalControlMapping, id=mapping_id)
    control = mapping.internal_control

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to delete mappings'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        log_compliance_action(
            request.user,
            'delete',
            'internal_control_mapping',
            mapping,
            request=request
        )

        mapping.delete()
        messages.success(request, _('Mapping deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


# ========================
# Framework Control Mapping Operations
# ========================

@login_required
@compliance_access_required
@require_http_methods(["POST"])
def control_mapping_create(request):
    """Create mapping for framework control to internal/local/framework controls"""
    def clean_id(value):
        """Clean and validate ID value"""
        if not value:
            return None
        cleaned = ''.join(c for c in str(value).strip() if c.isdigit())
        if not cleaned:
            return None
        try:
            return int(cleaned)
        except (ValueError, TypeError):
            return None

    source_control_id = clean_id(request.POST.get('source_control_id'))
    target_framework_id = clean_id(request.POST.get('target_framework_control_id'))
    target_internal_id = clean_id(request.POST.get('target_internal_control_id'))
    target_local_id = clean_id(request.POST.get('target_local_control_id'))

    if not source_control_id:
        messages.error(request, _('Source control is required'))
        return redirect('compliance:dashboard')

    source_control = get_object_or_404(Control, id=source_control_id)

    if not check_user_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to create mappings'))
        return redirect('compliance:control_detail', control_id=source_control.id)

    if not target_framework_id and not target_internal_id and not target_local_id:
        messages.error(request, _('Please select a target control'))
        return redirect('compliance:control_detail', control_id=source_control.id)

    try:
        mapping_type = request.POST.get('mapping_type', 'related')
        notes = request.POST.get('notes', '')

        mapping = ControlMapping.objects.create(
            source_control=source_control,
            target_control_id=target_framework_id,
            target_internal_control_id=target_internal_id,
            target_local_control_id=target_local_id,
            mapping_type=mapping_type,
            notes=notes,
            created_by=request.user
        )

        log_compliance_action(
            request.user,
            'create',
            'control_mapping',
            mapping,
            request=request
        )

        messages.success(request, _('Mapping created successfully'))

    except Exception as exc:
        messages.error(request, _('Error creating mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:control_detail', control_id=source_control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def control_mapping_delete(request, mapping_id):
    """Delete control mapping"""
    mapping = get_object_or_404(ControlMapping, id=mapping_id)
    source_control = mapping.source_control

    if not check_user_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to delete mappings'))
        return redirect('compliance:control_detail', control_id=source_control.id)

    try:
        log_compliance_action(
            request.user,
            'delete',
            'control_mapping',
            mapping,
            request=request
        )

        mapping.delete()
        messages.success(request, _('Mapping deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:control_detail', control_id=source_control.id)
