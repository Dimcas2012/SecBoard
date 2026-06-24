from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.utils.translation import gettext_lazy as _, get_language
from django.views.decorators.http import require_http_methods, require_POST
from io import BytesIO
from app_conf.models import Company, Country, CompanyType
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date
from django.db import transaction
from django.core.exceptions import ValidationError
import PyPDF2
import re

from .models import (
    ComplianceFramework, ControlCategory, Control,
    Evidence, ControlAssignment, ComplianceAuditLog, ControlMapping,
    AccessCompliance, AccessLocalCompliance, RegulatorType, LocalComplianceRegulator,
    LocalComplianceRequirement, LocalComplianceControl, LocalRequirementCategory,
    RequirementType, RequirementStatus, RequirementPriority,
    RequirementTypeTranslation, RequirementStatusTranslation,
    RequirementPriorityTranslation, EvidenceType, EvidenceTypeTranslation,
    LocalControlEvidence,
    LocalControlAssignment, LocalControlNote, LocalControlMapping,
    AccessInternalCompliance, InternalComplianceSource, InternalComplianceRequirement,
    InternalComplianceControl, InternalRequirementCategory, InternalControlEvidence,
    InternalControlAssignment, InternalControlNote, InternalControlMapping,
    AccessControlMapping, FrameworkInstanceNote, FrameworkDomain,
    LocalComplianceGuide, LocalComplianceGuideTranslation
)

from .utils import *

# ========================
# Local Compliance Dashboard
# ========================


def _has_local_compliance_access(user):
    """Check if user has access to Local Compliance module."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return AccessLocalCompliance.objects.filter(
        group__in=user.groups.all(),
        has_access=True
    ).exists()


@login_required
@require_http_methods(["GET"])
def local_compliance_guide(request):
    """Return JSON { content: html } for the Local Compliance guide (localized)."""
    if not _has_local_compliance_access(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = LocalComplianceGuide.objects.first()
    if guide:
        if country:
            trans = LocalComplianceGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = LocalComplianceGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def local_compliance_guide_translate(request):
    """API for AI translation of Local Compliance guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@local_compliance_access_required
def local_compliance_dashboard(request):
    """
    Dashboard для Local Compliance (вимоги місцевих регуляторів)
    """
    from datetime import date, timedelta
    from django.db.models import Count, Q
    from django.core.paginator import Paginator
    
    # Get user's accessible companies for Local Compliance
    accessible_companies = get_user_accessible_companies_local(request.user)
    
    # Отримуємо вибрані компанії з GET параметрів
    selected_company_ids = request.GET.getlist('company')
    
    if selected_company_ids:
        # Фільтруємо тільки доступні компанії
        valid_selected_ids = [
            int(cid) for cid in selected_company_ids 
            if int(cid) in list(accessible_companies.values_list('id', flat=True))
        ]
        if valid_selected_ids:
            selected_companies = accessible_companies.filter(id__in=valid_selected_ids)
        else:
            selected_companies = accessible_companies
    else:
        selected_companies = accessible_companies
    
    # Фільтр по країні
    selected_country_id = request.GET.get('country', '')
    selected_country = None
    if selected_country_id:
        try:
            selected_country = Country.objects.get(id=int(selected_country_id))
        except (Country.DoesNotExist, ValueError, TypeError):
            selected_country_id = ''
            selected_country = None
    
    # Фільтр по типу регулятора
    selected_regulator_type_id = request.GET.get('regulator_type', '')
    selected_regulator_type = None
    if selected_regulator_type_id:
        try:
            selected_regulator_type = RegulatorType.objects.get(id=int(selected_regulator_type_id))
        except (RegulatorType.DoesNotExist, ValueError):
            selected_regulator_type_id = ''
    
    # Отримуємо всі активні регулятори
    regulators_qs = LocalComplianceRegulator.objects.filter(is_active=True)
    
    if selected_country_id:
        regulators_qs = regulators_qs.filter(country_id=selected_country_id)
    
    if selected_regulator_type_id:
        regulators_qs = regulators_qs.filter(regulator_type_id=selected_regulator_type_id)
    
    regulators = regulators_qs.annotate(
        total_requirements=Count('requirements'),
        active_requirements=Count('requirements', filter=Q(requirements__status='active'))
    ).order_by('country', 'name')
    
    # Отримуємо вимоги для вибраних компаній
    requirements_qs = LocalComplianceRequirement.objects.filter(
        regulator__in=regulators,
        status='active'
    ).select_related('regulator')
    
    # Отримуємо контролі для вибраних компаній
    controls = LocalComplianceControl.objects.filter(
        company__in=selected_companies,
        requirement__in=requirements_qs
    ).select_related('requirement', 'company', 'responsible')
    
    # Загальна статистика
    total_regulators = regulators.count()
    total_requirements = requirements_qs.count()
    total_controls = controls.count()
    
    completed_controls = controls.filter(status='completed').count()
    in_progress_controls = controls.filter(status='in_progress').count()
    not_started_controls = controls.filter(status='not_started').count()
    
    overall_completion = 0
    if total_controls > 0:
        overall_completion = round((completed_controls / total_controls) * 100, 1)
    
    # Статистика по пріоритетах
    critical_controls = controls.filter(priority='critical').count()
    critical_completed = controls.filter(priority='critical', status='completed').count()
    
    high_controls = controls.filter(priority='high').count()
    high_completed = controls.filter(priority='high', status='completed').count()
    
    # Overdue контролі
    today = date.today()
    overdue_controls = controls.filter(
        target_completion_date__lt=today,
        status__in=['not_started', 'in_progress']
    ).order_by('target_completion_date')[:10]
    
    # Контролі що стають due в найближчі 30 днів
    upcoming_due = controls.filter(
        target_completion_date__gte=today,
        target_completion_date__lte=today + timedelta(days=30),
        status__in=['not_started', 'in_progress']
    ).order_by('target_completion_date')[:10]
    
    # Overdue requirements (по deadline)
    overdue_requirements = requirements_qs.filter(
        deadline_date__lt=today
    ).order_by('deadline_date')[:10]
    
    # Upcoming deadlines для requirements
    upcoming_requirements = requirements_qs.filter(
        deadline_date__gte=today,
        deadline_date__lte=today + timedelta(days=60)
    ).order_by('deadline_date')[:10]
    
    # Статистика по регуляторах
    regulator_stats = []
    for regulator in regulators[:10]:  # Top 10 регуляторів
        reg_requirements = requirements_qs.filter(regulator=regulator)
        reg_controls = controls.filter(requirement__regulator=regulator)
        
        total = reg_controls.count()
        completed = reg_controls.filter(status='completed').count()
        completion = round((completed / total * 100), 1) if total > 0 else 0
        
        regulator_stats.append({
            'regulator': regulator,
            'requirements_count': reg_requirements.count(),
            'total_controls': total,
            'completed_controls': completed,
            'completion': completion,
        })
    
    # Статистика по компаніях
    company_stats = []
    for company in selected_companies:
        company_controls = controls.filter(company=company)
        total = company_controls.count()
        completed = company_controls.filter(status='completed').count()
        completion = round((completed / total * 100), 1) if total > 0 else 0
        
        company_stats.append({
            'company': company,
            'total_controls': total,
            'completed': completed,
            'in_progress': company_controls.filter(status='in_progress').count(),
            'not_started': company_controls.filter(status='not_started').count(),
            'completion': completion,
        })
    
    # Priority Matrix
    priority_matrix = {}
    for priority_value, priority_label in Control.PRIORITY_CHOICES:
        priority_matrix[priority_value] = {
            'label': priority_label,
            'completed': controls.filter(priority=priority_value, status='completed').count(),
            'in_progress': controls.filter(priority=priority_value, status='in_progress').count(),
            'not_started': controls.filter(priority=priority_value, status='not_started').count(),
            'total': controls.filter(priority=priority_value).count(),
        }
    
    # Team Workload
    from django.contrib.auth.models import User
    
    team_workload = []
    users_with_controls = User.objects.filter(
        local_compliance_controls__company__in=selected_companies
    ).distinct()
    
    for user in users_with_controls:
        user_controls = controls.filter(responsible=user)
        total_assigned = user_controls.count()
        
        if total_assigned > 0:
            completed = user_controls.filter(status='completed').count()
            completion_pct = round((completed / total_assigned) * 100, 1)
            
            team_workload.append({
                'user': user,
                'total': total_assigned,
                'completed': completed,
                'in_progress': user_controls.filter(status='in_progress').count(),
                'not_started': user_controls.filter(status='not_started').count(),
                'completion': completion_pct,
            })
    
    # Unassigned controls
    unassigned_count = controls.filter(responsible__isnull=True).count()
    if unassigned_count > 0:
        team_workload.append({
            'user': None,
            'total': unassigned_count,
            'completed': controls.filter(responsible__isnull=True, status='completed').count(),
            'in_progress': controls.filter(responsible__isnull=True, status='in_progress').count(),
            'not_started': controls.filter(responsible__isnull=True, status='not_started').count(),
            'completion': 0,
        })
    
    team_workload = sorted(team_workload, key=lambda x: x['completion'])
    
    # Recent Activity
    recent_logs = ComplianceAuditLog.objects.filter(
        Q(object_type='local_requirement') | Q(object_type='local_control')
    ).select_related('user').order_by('-timestamp')[:20]
    
    # Отримуємо доступні країни для фільтрів
    # Get unique country IDs from active regulators
    country_ids = LocalComplianceRegulator.objects.filter(
        is_active=True
    ).values_list('country_id', flat=True).distinct()
    available_countries = Country.objects.filter(
        id__in=country_ids,
        is_active=True
    ).order_by('display_order', 'name')
    
    # Отримуємо доступні типи регуляторів
    regulator_type_ids = LocalComplianceRegulator.objects.filter(
        is_active=True
    ).values_list('regulator_type_id', flat=True).distinct()
    available_regulator_types = RegulatorType.objects.filter(
        id__in=regulator_type_ids,
        is_active=True
    ).order_by('display_order', 'name')
    
    # Permissions for Local Compliance
    permissions = get_user_local_compliance_permissions(request.user)
    
    context = {
        'accessible_companies': accessible_companies,
        'available_companies': accessible_companies,  # Alias for template consistency
        'selected_companies': selected_companies,
        'selected_company_ids': [int(id) for id in selected_company_ids] if selected_company_ids else [],
        'selected_country': selected_country,
        'selected_country_id': selected_country_id,
        'selected_regulator_type': selected_regulator_type,
        'selected_regulator_type_id': selected_regulator_type_id,
        
        # Statistics
        'total_regulators': total_regulators,
        'total_requirements': total_requirements,
        'total_controls': total_controls,
        'completed_controls': completed_controls,
        'in_progress_controls': in_progress_controls,
        'not_started_controls': not_started_controls,
        'overall_completion': overall_completion,
        
        # Priority stats
        'critical_controls': critical_controls,
        'critical_completed': critical_completed,
        'high_controls': high_controls,
        'high_completed': high_completed,
        
        # Lists
        'regulators': regulators,
        'regulator_stats': regulator_stats,
        'company_stats': company_stats,
        'overdue_controls': overdue_controls,
        'upcoming_due': upcoming_due,
        'overdue_requirements': overdue_requirements,
        'upcoming_requirements': upcoming_requirements,
        'priority_matrix': priority_matrix,
        'team_workload': team_workload[:10],
        'recent_logs': recent_logs,
        
        # Filters
        'available_countries': available_countries,
        'available_regulator_types': available_regulator_types,
        
        # Permissions
        'permissions': permissions,
    }
    
    return render(request, 'app_compliance/local_compliance_dashboard.html', context)


# ========================
# Local Requirements Instances
# ========================

@login_required
@local_compliance_access_required
def local_requirement_instances_list(request):
    """List Local Requirement instances applied to companies"""
    if not check_user_local_compliance_permission(request.user, 'can_view_requirement_instances'):
        messages.error(request, _('You do not have permission to view requirement instances'))
        return redirect('compliance:local_compliance')

    accessible_companies = get_user_accessible_companies_local(request.user)
    accessible_company_ids = list(accessible_companies.values_list('id', flat=True))

    instances = LocalComplianceRequirement.objects.filter(
        is_template=False,
        company_id__in=accessible_company_ids
    ).select_related(
        'company',
        'template',
        'regulator__country',
        'created_by'
    ).prefetch_related('company_types')

    instances = instances.annotate(
        controls_total=Count('controls', distinct=True),
        controls_completed=Count('controls', filter=Q(controls__status='completed'), distinct=True),
        controls_in_progress=Count('controls', filter=Q(controls__status='in_progress'), distinct=True),
        controls_not_started=Count('controls', filter=Q(controls__status='not_started'), distinct=True)
    )

    search = request.GET.get('search', '')
    if search:
        instances = instances.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(company__name__icontains=search) |
            Q(regulator__name__icontains=search) |
            Q(description__icontains=search)
        )

    company_id = request.GET.get('company', '')
    if company_id and company_id.isdigit() and int(company_id) in accessible_company_ids:
        instances = instances.filter(company_id=company_id)

    template_id = request.GET.get('template', '')
    if template_id and template_id.isdigit():
        instances = instances.filter(template_id=template_id)

    status = request.GET.get('status', '')
    if status:
        instances = instances.filter(status=status)

    sort_by = request.GET.get('sort', '-created_date')
    instances = instances.order_by(sort_by)

    paginator = Paginator(instances, 20)
    page = request.GET.get('page')
    try:
        instances_page = paginator.page(page)
    except PageNotAnInteger:
        instances_page = paginator.page(1)
    except EmptyPage:
        instances_page = paginator.page(paginator.num_pages)

    for instance in instances_page:
        total = instance.controls_total or 0
        completed = instance.controls_completed or 0
        instance.completion = round((completed / total) * 100, 1) if total else 0
        instance.controls_summary = {
            'total': total,
            'completed': completed,
            'in_progress': instance.controls_in_progress or 0,
            'not_started': instance.controls_not_started or 0,
        }

    companies = accessible_companies.order_by('name')
    if request.user.is_superuser or request.user.is_staff:
        templates = LocalComplianceRequirement.objects.filter(is_template=True).order_by('name')
    else:
        templates = LocalComplianceRequirement.objects.filter(
            is_template=True,
            instances__company_id__in=accessible_company_ids
        ).distinct().order_by('name')

    permissions = get_user_local_compliance_permissions(request.user)
    has_access_to_control_mapping = check_user_control_mapping_access(request.user)

    context = {
        'instances': instances_page,
        'companies': companies,
        'templates': templates,
        'search': search,
        'selected_company': company_id,
        'selected_template': template_id,
        'selected_status': status,
        'sort_by': sort_by,
        'status_choices': LocalComplianceRequirement.STATUS_CHOICES,
        'permissions': permissions,
        'has_access_to_control_mapping': has_access_to_control_mapping,
    }

    return render(request, 'app_compliance/local_requirement_instances_list.html', context)


# ========================
# Local Requirements Templates Management
# ========================

@login_required
@local_compliance_access_required
def local_requirements_library(request):
    """
    Бібліотека Local Requirements Templates
    """
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_view_requirements'):
        messages.error(request, _('You do not have permission to view requirements templates'))
        return redirect('compliance:local_compliance')
    
    # Show only templates
    requirements = LocalComplianceRequirement.objects.filter(is_template=True)
    
    requirements = requirements.select_related('regulator__country', 'regulator__regulator_type', 'created_by').prefetch_related('company_types', 'instances__company').annotate(
        controls_count=Count('controls', filter=Q(controls__company__isnull=True)),
        instances_count=Count('instances')
    )
    
    # Режим відображення (table/cards)
    view_mode = request.GET.get('view', 'table')
    if view_mode not in ('table', 'cards'):
        view_mode = 'table'

    # Пошук
    search = request.GET.get('search', '')
    if search:
        requirements = requirements.filter(
            Q(name__icontains=search) |
            Q(name_local__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Фільтр по типу
    requirement_type = request.GET.get('type', '')
    if requirement_type:
        requirements = requirements.filter(requirement_type=requirement_type)
    
    # Фільтр по статусу
    status = request.GET.get('status', '')
    if status:
        requirements = requirements.filter(status=status)
    
    # Зберігаємо queryset після застосування пошуку/типу/статусу для формування списку фільтрів
    filter_options_qs = requirements

    # Фільтр по регулятору
    regulator_id = request.GET.get('regulator', '')
    if regulator_id:
        requirements = requirements.filter(regulator_id=regulator_id)

    # Сортування
    sort_by = request.GET.get('sort', '-created_date')
    requirements = requirements.order_by(sort_by)
    
    # Пагінація
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(requirements, 20)
    page = request.GET.get('page')
    
    try:
        requirements_page = paginator.page(page)
    except PageNotAnInteger:
        requirements_page = paginator.page(1)
    except EmptyPage:
        requirements_page = paginator.page(paginator.num_pages)
    
    # Отримуємо permissions для користувача (Local Compliance)
    permissions = get_user_local_compliance_permissions(request.user)
    
    # Отримуємо доступні компанії для bulk operations (Local Compliance)
    accessible_companies = get_user_accessible_companies_local(request.user)
    accessible_company_ids = set(accessible_companies.values_list('id', flat=True))
    companies = accessible_companies.order_by('name')
    
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)

    # Отримуємо регуляторів для фільтру
    regulator_ids = filter_options_qs.values_list('regulator_id', flat=True).distinct()
    regulators = LocalComplianceRegulator.objects.filter(
        id__in=regulator_ids,
        is_active=True
    ).order_by('name')

    # Отримуємо типи вимог для фільтру
    type_codes = list(filter_options_qs.values_list('requirement_type', flat=True).distinct())
    choice_labels = dict(LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    requirement_types = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        codes=type_codes,
        fallback_map=choice_labels
    )

    # Отримуємо статуси вимог для фільтру
    status_codes = list(filter_options_qs.values_list('status', flat=True).distinct())
    status_choice_labels = dict(LocalComplianceRequirement.STATUS_CHOICES)
    status_choices = get_dictionary_options(
        RequirementStatus,
        RequirementStatusTranslation,
        'requirement_status',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(LocalComplianceRequirement.STATUS_CHOICES)
    )

    type_labels_map = {opt['code']: opt['label'] for opt in requirement_types}
    status_labels_map = {opt['code']: opt['label'] for opt in status_choices}

    # Build company type map for localized display
    company_type_map = {}
    requirement_ids = [req.id for req in requirements_page.object_list]
    requirement_company_type_ids = set(
        LocalComplianceRequirement.company_types.through.objects.filter(
            localcompliancerequirement_id__in=requirement_ids
        ).values_list('companytype_id', flat=True)
    )

    available_company_types = CompanyType.objects.filter(
        Q(is_active=True) | Q(id__in=requirement_company_type_ids)
    ).order_by('display_order', 'name')

    for company_type in available_company_types:
        if use_localized_labels:
            label = company_type.get_local_name(country_for_language) if country_for_language else (company_type.name_local or company_type.name)
        else:
            label = company_type.name
        company_type_map[company_type.id] = {
            'label': label,
            'color': getattr(company_type, 'color', '#6c757d')
        }

    is_staff_or_superuser = request.user.is_superuser or request.user.is_staff

    for req in requirements_page:
        req.localized_requirement_type = type_labels_map.get(
            req.requirement_type,
            req.get_requirement_type_display()
        )
        req.localized_status = status_labels_map.get(
            req.status,
            req.get_status_display()
        )
        localized_company_types = []
        for company_type_id in req.company_types.values_list('id', flat=True):
            info = company_type_map.get(company_type_id)
            if info:
                localized_company_types.append(info)
        req.localized_company_types = localized_company_types

        instances = list(req.instances.all())
        if not is_staff_or_superuser:
            instances = [
                instance for instance in instances
                if instance.company_id in accessible_company_ids
            ]
        req.instances_data = [
            {
                'id': instance.id,
                'company__name': instance.company.name if instance.company else _('(No company)'),
                'is_mandatory': instance.is_mandatory,
            }
            for instance in instances
        ]
    
    # Базовий query string без параметрів пагінації та режиму відображення
    base_query_dict = request.GET.copy()
    for param in ['page', 'view']:
        if param in base_query_dict:
            del base_query_dict[param]
    base_query = base_query_dict.urlencode()

    context = {
        'requirements': requirements_page,
        'search': search,
        'selected_regulator': regulator_id,
        'selected_type': requirement_type,
        'selected_status': status,
        'sort_by': sort_by,
        'regulators': regulators,
        'requirement_types': requirement_types,
        'status_choices': status_choices,
        'type_labels_map': type_labels_map,
        'status_labels_map': status_labels_map,
        'permissions': permissions,
        'companies': companies,
        'view_mode': view_mode,
        'base_query': base_query,
        'current_query': request.GET.urlencode(),
    }
    
    return render(request, 'app_compliance/local_requirements_library.html', context)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_template_delete(request, requirement_id):
    """Видалити Local Requirement Template"""
    if not check_user_local_compliance_permission(request.user, 'can_delete_requirements'):
        messages.error(request, _('You do not have permission to delete requirement templates'))
        return redirect('compliance:local_requirements_library')

    requirement = get_object_or_404(LocalComplianceRequirement, id=requirement_id, is_template=True)

    try:
        log_compliance_action(
            request.user, 'delete', 'local_requirement', requirement,
            request=request
        )
        requirement.delete()
        messages.success(request, _('Requirement template deleted successfully'))
    except Exception as e:
        messages.error(request, _('Error deleting requirement template: %(error)s') % {'error': str(e)})

    return redirect('compliance:local_requirements_library')


@login_required
@local_compliance_access_required
@require_http_methods(["GET"])
def local_requirements_export_excel(request):
    """Export filtered Local Requirements Library to Excel"""
    if not check_user_local_compliance_permission(request.user, 'can_view_requirements'):
        messages.error(request, _('You do not have permission to export requirement templates'))
        return redirect('compliance:local_requirements_library')

    requirements = LocalComplianceRequirement.objects.filter(is_template=True)
    requirements = requirements.select_related(
        'regulator__country',
        'regulator__regulator_type',
        'created_by'
    ).prefetch_related(
        'company_types',
        'instances__company'
    ).annotate(
        controls_count=Count('controls', filter=Q(controls__company__isnull=True)),
        instances_count=Count('instances')
    )

    search = request.GET.get('search', '')
    if search:
        requirements = requirements.filter(
            Q(name__icontains=search) |
            Q(name_local__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )

    requirement_type = request.GET.get('type', '')
    if requirement_type:
        requirements = requirements.filter(requirement_type=requirement_type)

    status = request.GET.get('status', '')
    if status:
        requirements = requirements.filter(status=status)

    regulator_id = request.GET.get('regulator', '')
    if regulator_id:
        requirements = requirements.filter(regulator_id=regulator_id)

    sort_by = request.GET.get('sort', '-created_date')
    requirements = requirements.order_by(sort_by)

    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    type_codes = list(requirements.values_list('requirement_type', flat=True).distinct())
    status_codes = list(requirements.values_list('status', flat=True).distinct())

    requirement_types = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        codes=type_codes,
        fallback_map=dict(LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    )
    status_choices = get_dictionary_options(
        RequirementStatus,
        RequirementStatusTranslation,
        'requirement_status',
        language_code,
        country_for_language,
        use_localized_labels,
        codes=status_codes,
        fallback_map=dict(LocalComplianceRequirement.STATUS_CHOICES)
    )
    type_labels_map = {opt['code']: opt['label'] for opt in requirement_types}
    status_labels_map = {opt['code']: opt['label'] for opt in status_choices}

    requirement_ids = list(requirements.values_list('id', flat=True))
    company_type_ids = set(
        LocalComplianceRequirement.company_types.through.objects.filter(
            localcompliancerequirement_id__in=requirement_ids
        ).values_list('companytype_id', flat=True)
    )
    available_company_types = CompanyType.objects.filter(
        Q(is_active=True) | Q(id__in=company_type_ids)
    ).order_by('display_order', 'name')

    company_type_map = {}
    for company_type in available_company_types:
        if use_localized_labels:
            label = company_type.get_local_name(country_for_language) if country_for_language else (company_type.name_local or company_type.name)
        else:
            label = company_type.name
        company_type_map[company_type.id] = {
            'label': label,
            'color': getattr(company_type, 'color', '#6c757d')
        }

    accessible_companies = get_user_accessible_companies(request.user)
    accessible_company_ids = set(accessible_companies.values_list('id', flat=True))
    is_staff_or_superuser = request.user.is_superuser or request.user.is_staff

    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Local Requirements'))

    headers = [
        str(_('Code')),
        str(_('Name')),
        str(_('Regulator')),
        str(_('Type')),
        str(_('Status')),
        str(_('Company Types')),
        str(_('Controls')),
        str(_('Applied to Companies')),
        str(_('Effective Date')),
        str(_('Deadline'))
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    status_fill_map = {
        'active': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'draft': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'archived': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
    }

    for idx, req in enumerate(requirements, start=2):
        localized_company_types = []
        for company_type_id in req.company_types.values_list('id', flat=True):
            info = company_type_map.get(company_type_id)
            if info:
                localized_company_types.append(info['label'])

        instances = list(req.instances.filter(company__isnull=False))
        if not is_staff_or_superuser:
            instances = [inst for inst in instances if inst.company_id in accessible_company_ids]

        applied_companies = []
        for instance in instances:
            label = instance.company.name if instance.company else str(_('(No company)'))
            if instance.is_mandatory:
                label = f"{label} ({str(_('Mandatory'))})"
            applied_companies.append(label)

        row = [
            req.code,
            req.name,
            req.regulator.acronym or req.regulator.name if req.regulator else '',
            type_labels_map.get(req.requirement_type, req.get_requirement_type_display()),
            status_labels_map.get(req.status, req.get_status_display()),
            ', '.join(localized_company_types) if localized_company_types else '—',
            req.controls_count,
            ', '.join(applied_companies) if applied_companies else str(_('Not applied')),
            req.effective_date.strftime('%d.%m.%Y') if req.effective_date else '',
            req.deadline_date.strftime('%d.%m.%Y') if req.deadline_date else '',
        ]
        ws.append(row)

        status_cell = ws.cell(row=idx, column=5)
        status_fill = status_fill_map.get(req.status)
        if status_fill:
            status_cell.fill = status_fill

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col_idx).border = thin_border

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"local_requirements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


@login_required
@compliance_access_required
def local_requirements_excel_template(request):
    """Download Excel template for Local Requirements import"""
    if not check_user_compliance_permission(request.user, 'can_add_frameworks'):
        messages.error(request, _('You do not have permission to download the template'))
        return redirect('compliance:local_requirements_library')

    # Get language preferences for localized names
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    use_local_names = bool(language_code) and not language_code.lower().startswith('en')

    # Get all active regulators with their IDs, names, and countries
    regulators_qs = LocalComplianceRegulator.objects.filter(
        is_active=True
    ).select_related('country').order_by('country__name', 'name')
    
    regulators_list = []
    for regulator in regulators_qs:
        display_name = regulator.name_local if use_local_names and regulator.name_local else regulator.name
        country_name = regulator.country.name_local if use_local_names and regulator.country.name_local else regulator.country.name
        regulator_label = f"{regulator.acronym} - {display_name}" if regulator.acronym else display_name
        regulators_list.append(f"{regulator.id}: {regulator_label} ({country_name})")
    regulator_notes = str(_("Use the numeric ID from Regulators list. Available regulators: ")) + ", ".join(regulators_list)

    # Get all active countries with their IDs and names
    countries = Country.objects.filter(is_active=True).order_by('display_order', 'name')
    countries_list = []
    for country in countries:
        display_name = country.name_local if use_local_names and country.name_local else country.name
        countries_list.append(f"{country.id}: {display_name}")
    country_notes = str(_("Use the numeric ID from Countries list. Available countries: ")) + ", ".join(countries_list)

    # Get all active company types with their IDs and names (localized)
    company_types = CompanyType.objects.filter(is_active=True).order_by('display_order', 'name')
    company_types_list = []
    company_type_labels = []
    for ct in company_types:
        if use_localized_labels:
            label = ct.get_local_name(country_for_language) if country_for_language else (ct.name_local or ct.name)
        else:
            label = ct.name
        company_types_list.append(f"{ct.id}: {label}")
        company_type_labels.append(label)
    company_types_notes = str(_("Comma separated CompanyType IDs. Available types: ")) + ", ".join(company_types_list)

    # Applicable To examples (use current company type labels)
    applicable_to_examples = company_type_labels or [
        str(_("Banks")),
        str(_("Payment Systems")),
        str(_("Insurance Companies")),
        str(_("Investment Firms"))
    ]
    applicable_to_notes = str(_("Examples: ")) + ", ".join(applicable_to_examples)

    wb = Workbook()
    ws_req = wb.active
    ws_req.title = "Requirement"
    ws_req.append(["Field", "Value", "Notes"])
    ws_req.append(["Code", "REQ-2025-01", str(_("Required. Unique code in the system."))])
    ws_req.append(["Name", str(_("NBU Resolution No. 95")), str(_("Full requirement name."))])
    ws_req.append(["Name Local", str(_("Назва українською")), str(_("Optional local name."))])
    ws_req.append(["Country", "1", country_notes])
    ws_req.append(["Regulator ID", "1", regulator_notes])
    ws_req.append(["Requirement Type", "regulation", str(_("Allowed: regulation, law, decree, guidance, other"))])
    ws_req.append(["Status", "active", str(_("Allowed: active, draft, archived"))])
    ws_req.append(["Priority", "high", str(_("Allowed: low, medium, high, critical"))])
    ws_req.append(["Mandatory", "Yes", str(_("Yes/No"))])
    ws_req.append(["Applicable To", str(_("Banks, payment systems")), applicable_to_notes])
    ws_req.append(["Official Link", "https://...", ""])
    ws_req.append([
        "Publication Date",
        "",
        str(_("Set before import in the modal (required). Format: DD.MM.YYYY"))
    ])
    ws_req.append([
        "Effective Date",
        "",
        str(_("Set before import in the modal (required). Format: DD.MM.YYYY"))
    ])
    ws_req.append([
        "Deadline Date",
        "",
        str(_("Set before import in the modal (required). Format: DD.MM.YYYY"))
    ])
    ws_req.append(["Company Type IDs", "1,3", company_types_notes])
    ws_req.append(["Description", str(_("Short description of requirement")), ""])

    for cell in ws_req[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_req.column_dimensions['A'].width = 25
    ws_req.column_dimensions['B'].width = 40
    ws_req.column_dimensions['C'].width = 80
    
    # Enable text wrapping for Notes column (column C)
    for row in ws_req.iter_rows(min_row=2, max_row=ws_req.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_ctrl = wb.create_sheet("Controls")
    control_headers = [
        "Category Code",
        "Category Name",
        "Category Description",
        "Control Code",
        "Control Name",
        "Priority",
        "Target Date",
        "Implementation Notes",
        "Evidence Notes",
        "Description",
    ]
    ws_ctrl.append(control_headers)
    ws_ctrl.append([
        "CAT-1",
        str(_("Governance")),
        str(_("Management responsibilities")),
        "CAT-1-01",
        str(_("Create IS governance body")),
        "High",
        datetime(2025, 6, 1),
        str(_("Define members and charter")),
        str(_("Meeting minutes")),
        str(_("Establish body responsible for IS decisions")),
    ])
    ws_ctrl.cell(row=2, column=7).number_format = 'DD.MM.YYYY'
    for cell in ws_ctrl[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="10793F", end_color="10793F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, width in zip("ABCDEFGHIJ", [15, 25, 30, 18, 40, 12, 15, 40, 40, 50]):
        ws_ctrl.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="local_requirements_template.xlsx"'
    return response


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_requirements_import_excel(request):
    """Import Local Requirement template with controls from Excel"""
    if not check_user_compliance_permission(request.user, 'can_add_frameworks'):
        messages.error(request, _('You do not have permission to import requirements'))
        return redirect('compliance:local_requirements_library')

    excel_file = request.FILES.get('file')
    if not excel_file:
        messages.error(request, _('No file uploaded'))
        return redirect('compliance:local_requirements_library')

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as exc:
        messages.error(request, _('Unable to read Excel file: %(error)s') % {'error': str(exc)})
        return redirect('compliance:local_requirements_library')

    if "Requirement" not in wb.sheetnames or "Controls" not in wb.sheetnames:
        messages.error(
            request,
            _('Invalid template structure. The file must contain sheets named exactly "Requirement" and "Controls". '
              'Please use the template from the "Download Template" button, or export a requirement from its detail page.')
        )
        return redirect('compliance:local_requirements_library')

    ws_req = wb["Requirement"]
    requirement_data = {}
    for row in ws_req.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        field = str(row[0]).strip().lower()
        value = row[1] if len(row) > 1 else ''
        requirement_data[field] = value

    def get_required(field):
        value = str(requirement_data.get(field, '') or '').strip()
        if not value:
            raise ValueError(_('Field "%(field)s" is required') % {'field': field})
        return value

    def normalize_choice(value, choices, label):
        if not value:
            raise ValueError(_('Field "%(field)s" is required') % {'field': label})
        value_str = str(value).strip().lower()
        for code, display in choices:
            if value_str == code.lower() or value_str == str(display).lower():
                return code
        raise ValueError(_('Invalid %(field)s value: %(value)s') % {'field': label, 'value': value})

    def is_empty_value(value):
        """Check if value is effectively empty"""
        if not value:
            return True
        value_str = str(value).strip()
        if not value_str:
            return True
        value_lower = value_str.lower()
        if value_lower in ('none', 'null', '', 'n/a', '-', '—'):
            return True
        return False

    def parse_date(value):
        """Safely parse date value, returning None for empty/invalid"""
        if is_empty_value(value):
            return None
        try:
            return parse_local_requirement_date(value)
        except (ValueError, TypeError):
            return None

    override_inputs = {
        'publication date': request.POST.get('override_publication_date', '').strip(),
        'effective date': request.POST.get('override_effective_date', '').strip(),
        'deadline date': request.POST.get('override_deadline_date', '').strip(),
    }
    override_dates = {}
    for field_name, override_value in override_inputs.items():
        parsed_override = None
        if override_value and not is_empty_value(override_value):
            try:
                parsed_override = parse_local_requirement_date(override_value)
            except ValueError:
                messages.error(
                    request,
                    _('Invalid %(field)s override. Use DD.MM.YYYY') % {'field': field_name.title()}
                )
                return redirect('compliance:local_requirements_library')
        override_dates[field_name] = parsed_override

    def resolve_date(field_name):
        """Resolve date from Excel or override, never raises exception"""
        value = requirement_data.get(field_name, '')
        if not is_empty_value(value):
            parsed = parse_date(value)
            if parsed:
                return parsed
        return override_dates.get(field_name)

    try:
        with transaction.atomic():
            code = get_required('code')
            name = get_required('name')
            regulator_id = get_required('regulator id')
            try:
                regulator = LocalComplianceRegulator.objects.get(id=int(regulator_id))
            except (ValueError, LocalComplianceRegulator.DoesNotExist):
                raise ValueError(_('Invalid Regulator ID: %(value)s') % {'value': regulator_id})

            requirement_type_code = normalize_choice(
                requirement_data.get('requirement type', ''),
                LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES,
                _('Requirement Type')
            )
            status_code = normalize_choice(
                requirement_data.get('status', ''),
                LocalComplianceRequirement.STATUS_CHOICES,
                _('Status')
            )
            requirement_priority = normalize_choice(
                requirement_data.get('priority', 'medium'),
                Control.PRIORITY_CHOICES,
                _('Priority')
            )
            mandatory_value = str(requirement_data.get('mandatory', 'yes') or '').strip().lower()
            is_mandatory = mandatory_value in ('yes', 'true', '1', 'y', 'так', 'да')

            requirement = LocalComplianceRequirement.objects.create(
                regulator=regulator,
                code=code,
                name=name,
                name_local=str(requirement_data.get('name local', '') or '').strip(),
                requirement_type=requirement_type_code,
                description=str(requirement_data.get('description', '') or '').strip(),
                status=status_code,
                applicable_to=str(requirement_data.get('applicable to', '') or '').strip(),
                publication_date=resolve_date('publication date'),
                effective_date=resolve_date('effective date'),
                deadline_date=resolve_date('deadline date'),
                official_link=str(requirement_data.get('official link', '') or '').strip(),
                is_mandatory=is_mandatory,
                priority=requirement_priority,
                is_template=True,
                created_by=request.user
            )

            company_type_ids_raw = str(requirement_data.get('company type ids', '') or '').strip()
            if company_type_ids_raw:
                ids = [int(x.strip()) for x in company_type_ids_raw.split(',') if x.strip()]
                company_types = CompanyType.objects.filter(id__in=ids)
                if company_types.count() != len(ids):
                    raise ValueError(_('One or more Company Type IDs are invalid'))
                requirement.company_types.set(company_types)

            ws_controls = wb["Controls"]
            categories_cache = {}
            category_order = 0
            last_category = None
            control_count = 0

            priority_map = {code.lower(): code for code, _ in Control.PRIORITY_CHOICES}
            priority_map.update({str(label).lower(): code for code, label in Control.PRIORITY_CHOICES})

            for row in ws_controls.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                category_code = str(row[0]).strip() if row[0] else ''
                category_name = str(row[1]).strip() if row[1] else ''
                category_description = str(row[2]).strip() if row[2] else ''
                control_code = str(row[3]).strip() if row[3] else ''
                control_name = str(row[4]).strip() if row[4] else ''
                control_priority_raw = str(row[5]).strip() if row[5] else ''
                target_date_raw = row[6]
                implementation_notes = str(row[7]).strip() if row[7] else ''
                evidence_notes = str(row[8]).strip() if row[8] else ''
                control_description = str(row[9]).strip() if len(row) > 9 and row[9] else ''

                if category_code or category_name:
                    if not category_code:
                        category_code = f"CAT-{len(categories_cache) + 1}"
                    if category_code not in categories_cache:
                        category_order += 1
                        categories_cache[category_code] = LocalRequirementCategory.objects.create(
                            requirement=requirement,
                            code=category_code,
                            name=category_name or category_code,
                            description=category_description,
                            order=category_order
                        )
                    last_category = categories_cache[category_code]

                if not control_code or not control_name:
                    continue

                control_priority = control_priority_raw.lower() if control_priority_raw else 'medium'
                control_priority = priority_map.get(control_priority, 'medium')
                target_date = parse_date(target_date_raw)

                LocalComplianceControl.objects.create(
                    requirement=requirement,
                    company=None,
                    category=last_category,
                    code=control_code,
                    name=control_name,
                    description=control_description,
                    priority=control_priority,
                    target_completion_date=target_date,
                    implementation_notes=implementation_notes,
                    evidence_notes=evidence_notes,
                    status='not_started',
                    created_by=request.user
                )
                control_count += 1

            log_compliance_action(
                request.user, 'create', 'local_requirement', requirement,
                changes={'action': 'imported_from_excel', 'controls_count': control_count, 'categories_count': len(categories_cache)},
                request=request
            )

            messages.success(
                request,
                _('Requirement "%(code)s" imported successfully with %(controls)d controls') % {
                    'code': requirement.code,
                    'controls': control_count
                }
            )
            return redirect('compliance:local_requirement_template_detail', requirement_id=requirement.id)

    except ValueError as exc:
        messages.error(request, str(exc))
    except Exception as exc:
        messages.error(request, _('Error importing requirement: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_requirements_library')


@login_required
@local_compliance_access_required
def local_requirement_template_create(request):
    from app_conf.models import CompanyType
    """GET: форма створення, POST: створити template"""
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_add_requirements'):
        messages.error(request, _('You do not have permission to create requirement templates'))
        return redirect('compliance:local_requirements_library')
    
    if request.method == 'GET':
        from app_conf.models import Country, CompanyType

        language_code, country_for_language, use_localized_labels = get_language_preferences(request)
        use_local_names = bool(language_code) and not language_code.lower().startswith('en')

        # Get active countries
        countries_qs = Country.objects.filter(is_active=True).order_by('display_order', 'name')
        countries = []
        for country in countries_qs:
            display_name = country.name_local if use_local_names and country.name_local else country.name
            setattr(country, 'display_name', display_name)
            countries.append(country)

        # Get all active regulators
        regulators_qs = LocalComplianceRegulator.objects.filter(
            is_active=True
        ).select_related('country').order_by('country__name', 'name')

        regulators = []
        for regulator in regulators_qs:
            display_name = regulator.name_local if use_local_names and regulator.name_local else regulator.name
            display_label = f"{regulator.acronym} - {display_name}" if regulator.acronym else display_name
            setattr(regulator, 'display_name', display_name)
            setattr(regulator, 'display_label', display_label)
            regulators.append(regulator)

        company_types_qs = CompanyType.objects.filter(is_active=True).order_by('display_order', 'name')
        company_type_choices = []
        for company_type in company_types_qs:
            if use_localized_labels:
                label = company_type.get_local_name(country_for_language) if country_for_language else (company_type.name_local or company_type.name)
            else:
                label = company_type.name
            company_type_choices.append({'id': company_type.id, 'label': label})

        requirement_types = get_dictionary_options(
            RequirementType,
            RequirementTypeTranslation,
            'requirement_type',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
        )

        status_choices = get_dictionary_options(
            RequirementStatus,
            RequirementStatusTranslation,
            'requirement_status',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(LocalComplianceRequirement.STATUS_CHOICES)
        )

        priority_choices = get_dictionary_options(
            RequirementPriority,
            RequirementPriorityTranslation,
            'requirement_priority',
            language_code,
            country_for_language,
            use_localized_labels,
            fallback_map=dict(Control.PRIORITY_CHOICES)
        )

        context = {
            'countries': countries,
            'regulators': regulators,
            'requirement_types': requirement_types,
            'status_choices': status_choices,
            'priority_choices': priority_choices,
            'company_type_choices': company_type_choices,
            'form_action': request.path,
            'page_title': _('Create Local Requirement Template'),
            'page_subtitle': _('Create a template for regulatory requirement that can be applied to multiple companies'),
            'submit_label': _('Create Template'),
            'requirement': None,
            'selected_country_id': '',
            'selected_regulator_id': '',
            'selected_company_type_ids': [],
            'default_status': 'draft',
            'default_priority': 'medium',
            'default_is_mandatory': True,
        }
        return render(request, 'app_compliance/local_requirement_template_create.html', context)
    
    # POST: створення requirement
    try:
        requirement = LocalComplianceRequirement.objects.create(
            regulator_id=request.POST.get('regulator'),
            code=request.POST.get('code'),
            name=request.POST.get('name'),
            name_local=request.POST.get('name_local', ''),
            requirement_type=request.POST.get('requirement_type', 'regulation'),
            description=request.POST.get('description', ''),
            status=request.POST.get('status', 'draft'),
            applicable_to=request.POST.get('applicable_to', ''),
            official_link=request.POST.get('official_link', ''),
            is_mandatory=request.POST.get('is_mandatory') == '1',
            priority=request.POST.get('priority', 'medium'),
            is_template=True,  # Create as template
            created_by=request.user
        )

        company_type_ids = request.POST.getlist('company_types')
        if company_type_ids:
            types_qs = CompanyType.objects.filter(id__in=company_type_ids)
            requirement.company_types.set(types_qs)
        
        # Handle dates
        if request.POST.get('publication_date'):
            requirement.publication_date = parse_local_requirement_date(request.POST.get('publication_date'))
        if request.POST.get('effective_date'):
            requirement.effective_date = parse_local_requirement_date(request.POST.get('effective_date'))
        if request.POST.get('deadline_date'):
            requirement.deadline_date = parse_local_requirement_date(request.POST.get('deadline_date'))
        
        requirement.save()
        
        log_compliance_action(
            request.user, 'create', 'local_requirement', requirement,
            request=request
        )
        
        messages.success(request, _('Requirement template created successfully'))
        return redirect('compliance:local_requirement_template_detail', requirement_id=requirement.id)
        
    except Exception as e:
        messages.error(request, f'Error creating requirement: {str(e)}')
        return redirect('compliance:local_requirements_library')


@login_required
@local_compliance_access_required
def local_requirement_template_edit(request, requirement_id):
    """Edit existing Local Requirement Template"""
    requirement = get_object_or_404(
        LocalComplianceRequirement.objects.select_related('regulator__country').prefetch_related('company_types'),
        id=requirement_id,
        is_template=True
    )

    if not check_user_local_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to edit requirement templates'))
        return redirect('compliance:local_requirements_library')

    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    use_local_names = bool(language_code) and not language_code.lower().startswith('en')

    countries_qs = Country.objects.filter(is_active=True).order_by('display_order', 'name')
    countries = []
    for country in countries_qs:
        display_name = country.name_local if use_local_names and country.name_local else country.name
        setattr(country, 'display_name', display_name)
        countries.append(country)

    regulators_qs = LocalComplianceRegulator.objects.filter(
        is_active=True
    ).select_related('country').order_by('country__name', 'name')

    regulators = []
    for regulator in regulators_qs:
        display_name = regulator.name_local if use_local_names and regulator.name_local else regulator.name
        display_label = f"{regulator.acronym} - {display_name}" if regulator.acronym else display_name
        setattr(regulator, 'display_name', display_name)
        setattr(regulator, 'display_label', display_label)
        regulators.append(regulator)

    requirement_types = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    )

    status_choices = get_dictionary_options(
        RequirementStatus,
        RequirementStatusTranslation,
        'requirement_status',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(LocalComplianceRequirement.STATUS_CHOICES)
    )

    priority_choices = get_dictionary_options(
        RequirementPriority,
        RequirementPriorityTranslation,
        'requirement_priority',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(Control.PRIORITY_CHOICES)
    )

    company_types_qs = CompanyType.objects.filter(is_active=True).order_by('display_order', 'name')
    company_type_choices = []
    for company_type in company_types_qs:
        if use_localized_labels:
            label = company_type.get_local_name(country_for_language) if country_for_language else (company_type.name_local or company_type.name)
        else:
            label = company_type.name
        company_type_choices.append({'id': company_type.id, 'label': label})

    if request.method == 'GET':
        # Get instances for "Applied to Companies" block
        instances = requirement.instances.select_related('regulator', 'company').annotate(
            controls_count=Count('controls')
        ).order_by('-created_date')
        
        # Get user permissions
        permissions = get_user_local_compliance_permissions(request.user)
        
        # Get accessible companies
        accessible_companies = get_user_accessible_companies_local(request.user)
        companies = accessible_companies.order_by('name')
        
        # Get applied companies info
        applied_companies = []
        applied_companies_mandatory = []
        for instance in instances:
            if instance.company_id:
                applied_companies.append(instance.company_id)
                if instance.is_mandatory:
                    applied_companies_mandatory.append(instance.company_id)
        applied_companies = list(set(applied_companies))
        applied_companies_mandatory = list(set(applied_companies_mandatory))
        
        context = {
            'countries': countries,
            'regulators': regulators,
            'requirement_types': requirement_types,
            'status_choices': status_choices,
            'priority_choices': priority_choices,
            'company_type_choices': company_type_choices,
            'form_action': request.path,
            'page_title': _('Edit Local Requirement Template'),
            'page_subtitle': _('Update template details or metadata'),
            'submit_label': _('Save Changes'),
            'requirement': requirement,
            'selected_country_id': str(requirement.regulator.country_id),
            'selected_regulator_id': str(requirement.regulator_id),
            'selected_company_type_ids': list(requirement.company_types.values_list('id', flat=True)),
            'default_status': requirement.status,
            'default_priority': requirement.priority,
            'default_is_mandatory': requirement.is_mandatory,
            'instances': instances,
            'permissions': permissions,
            'companies': companies,
            'applied_companies': applied_companies,
            'applied_companies_mandatory': applied_companies_mandatory,
        }
        return render(request, 'app_compliance/local_requirement_template_create.html', context)

    # POST - update requirement
    try:
        requirement.regulator_id = request.POST.get('regulator') or requirement.regulator_id
        requirement.code = request.POST.get('code', requirement.code)
        requirement.name = request.POST.get('name', requirement.name)
        requirement.name_local = request.POST.get('name_local', '')
        requirement.requirement_type = request.POST.get('requirement_type', requirement.requirement_type)
        requirement.description = request.POST.get('description', '')
        requirement.status = request.POST.get('status', requirement.status)
        requirement.applicable_to = request.POST.get('applicable_to', '')
        requirement.official_link = request.POST.get('official_link', '')
        requirement.is_mandatory = request.POST.get('is_mandatory') == '1'
        requirement.priority = request.POST.get('priority', requirement.priority)

        publication_date = request.POST.get('publication_date')
        effective_date = request.POST.get('effective_date')
        deadline_date = request.POST.get('deadline_date')

        requirement.publication_date = parse_local_requirement_date(publication_date) if publication_date else None
        requirement.effective_date = parse_local_requirement_date(effective_date) if effective_date else None
        requirement.deadline_date = parse_local_requirement_date(deadline_date) if deadline_date else None

        requirement.save()

        company_type_ids = request.POST.getlist('company_types')
        if company_type_ids:
            requirement.company_types.set(CompanyType.objects.filter(id__in=company_type_ids))
        else:
            requirement.company_types.clear()

        log_compliance_action(
            request.user,
            'update',
            'local_requirement',
            requirement,
            request=request
        )

        messages.success(request, _('Requirement template updated successfully'))
        return redirect('compliance:local_requirement_template_detail', requirement_id=requirement.id)

    except Exception as e:
        messages.error(request, _('Error updating requirement: %(error)s') % {'error': str(e)})
        return redirect('compliance:local_requirement_template_detail', requirement_id=requirement.id)


@login_required
@local_compliance_access_required
def local_requirement_template_detail(request, requirement_id):
    """Деталі requirement template з контролями"""
    requirement = get_object_or_404(
        LocalComplianceRequirement.objects.select_related('regulator__country', 'regulator__regulator_type').prefetch_related('company_types'),
        id=requirement_id,
        is_template=True
    )
    
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_view_requirements'):
        messages.error(request, _('You do not have access to this requirement'))
        return redirect('compliance:local_requirements_library')
    
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)

    company_type_labels = []
    for company_type in requirement.company_types.all():
        if use_localized_labels:
            label = company_type.get_local_name(country_for_language) if country_for_language else (company_type.name_local or company_type.name)
        else:
            label = company_type.name
        company_type_labels.append({
            'label': label,
            'color': getattr(company_type, 'color', '#6c757d')
        })

    requirement_type_options = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    )
    requirement_type_label_map = {opt['code']: opt['label'] for opt in requirement_type_options}
    localized_requirement_type = requirement_type_label_map.get(
        requirement.requirement_type,
        requirement.get_requirement_type_display()
    )

    # Отримуємо контролі template (без company)
    template_controls_qs = requirement.controls.filter(company__isnull=True).select_related(
        'responsible',
        'category'
    ).order_by('code')
    controls = template_controls_qs
    
    categories = list(
        requirement.categories.order_by('order', 'code').prefetch_related(
        Prefetch(
            'controls',
            queryset=template_controls_qs,
            to_attr='template_controls'
        )
    ))
    uncategorized_controls = list(template_controls_qs.filter(category__isnull=True))
    
    # Отримуємо instances
    instances = requirement.instances.select_related('regulator', 'company').annotate(
        controls_count=Count('controls')
    ).order_by('-created_date')
    
    # Отримуємо permissions
    permissions = get_user_local_compliance_permissions(request.user)
    
    # Отримуємо доступні компанії
    accessible_companies = get_user_accessible_companies_local(request.user)
    companies = accessible_companies.order_by('name')
    
    # Get applied companies info
    applied_companies = []
    applied_companies_mandatory = []
    for instance in instances:
        if instance.company_id:
            applied_companies.append(instance.company_id)
            if instance.is_mandatory:
                applied_companies_mandatory.append(instance.company_id)
    applied_companies = list(set(applied_companies))
    applied_companies_mandatory = list(set(applied_companies_mandatory))
    
    context = {
        'requirement': requirement,
        'controls': controls,
        'categories': categories,
        'uncategorized_controls': uncategorized_controls,
        'instances': instances,
        'permissions': permissions,
        'companies': companies,
        'applied_companies': applied_companies,
        'applied_companies_mandatory': applied_companies_mandatory,
        'company_type_labels': company_type_labels,
        'localized_requirement_type': localized_requirement_type,
    }
    
    return render(request, 'app_compliance/local_requirement_template_detail.html', context)


@login_required
@local_compliance_access_required
def local_requirement_instance_detail(request, requirement_id):
    """Деталі Local Requirement Instance з категоріями та контролями"""
    requirement = get_object_or_404(
        LocalComplianceRequirement.objects.select_related('company', 'regulator__country', 'regulator__regulator_type', 'template').prefetch_related('company_types'),
        id=requirement_id,
        is_template=False
    )
    
    # Перевірка прав доступу
    if not check_user_local_compliance_permission(request.user, 'can_view_requirement_instances'):
        messages.error(request, _('You do not have access to this requirement'))
        return redirect('compliance:local_compliance')
    
    # Перевірка доступу до компанії
    if requirement.company:
        accessible_companies = get_user_accessible_companies_local(request.user)
        if requirement.company not in accessible_companies:
            messages.error(request, _('You do not have access to this requirement'))
            return redirect('compliance:local_compliance')
    
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    
    # Отримуємо параметри фільтрів
    status_filter = request.GET.get('control_status', '')
    priority_filter = request.GET.get('priority', '')
    owner_filter = request.GET.get('owner', '')
    search = request.GET.get('search', '')
    
    # Будую базовий queryset для controls з урахуванням фільтрів
    controls_queryset = LocalComplianceControl.objects.filter(
        requirement=requirement,
        company=requirement.company
    ).select_related('responsible', 'category').annotate(
        evidence_count=Count('evidences', filter=Q(evidences__is_active=True))
    )
    
    # Застосовуємо фільтри
    if status_filter:
        controls_queryset = controls_queryset.filter(status=status_filter)
    if priority_filter:
        controls_queryset = controls_queryset.filter(priority=priority_filter)
    if owner_filter:
        controls_queryset = controls_queryset.filter(responsible_id=owner_filter)
    if search:
        controls_queryset = controls_queryset.filter(
            Q(code__icontains=search) |
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Додаємо сортування
    controls_queryset = controls_queryset.order_by('code')
    
    # Отримуємо категорії з фільтрованими контролями
    categories = requirement.categories.prefetch_related(
        Prefetch('controls', queryset=controls_queryset)
    ).annotate(
        total_controls=Count('controls')
    ).order_by('order', 'code')
    
    # Отримуємо контролі без категорії
    uncategorized_controls = list(controls_queryset.filter(category__isnull=True))
    
    # Статистика (використовуємо всі controls без фільтрів)
    all_controls = LocalComplianceControl.objects.filter(
        requirement=requirement,
        company=requirement.company
    )
    
    # Обчислюємо completion
    total_controls = all_controls.count()
    completed_controls = all_controls.filter(status='completed').count()
    completion = round((completed_controls / total_controls * 100), 1) if total_controls > 0 else 0
    
    # Статистика по статусах
    stats = {
        'total': total_controls,
        'completed': all_controls.filter(status='completed').count(),
        'in_progress': all_controls.filter(status='in_progress').count(),
        'not_started': all_controls.filter(status='not_started').count(),
        'ready_for_review': all_controls.filter(status='ready_for_review').count(),
        'failed': all_controls.filter(status='failed').count(),
        'not_applicable': all_controls.filter(status='not_applicable').count(),
    }
    
    # Отримуємо унікальні статуси та пріоритети
    existing_statuses = all_controls.values_list('status', flat=True).distinct()
    status_choices = [(value, label) for value, label in Control.STATUS_CHOICES if value in existing_statuses]
    
    existing_priorities = all_controls.values_list('priority', flat=True).distinct()
    priority_choices = [(value, label) for value, label in Control.PRIORITY_CHOICES if value in existing_priorities]
    
    # Історія змін
    recent_logs = ComplianceAuditLog.objects.filter(
        Q(object_type='local_requirement', object_id=requirement.id) |
        Q(object_type='local_requirement_category', object_id__in=requirement.categories.values_list('id', flat=True)) |
        Q(object_type='local_control', object_id__in=all_controls.values_list('id', flat=True))
    ).select_related('user').order_by('-timestamp')[:20]
    
    # Отримуємо permissions
    permissions = get_user_local_compliance_permissions(request.user)
    
    # Отримуємо користувачів для owner filter
    from app_cabinet.models import CabinetUser
    from datetime import date
    
    requirement_company = requirement.company
    if requirement_company:
        cabinet_users = CabinetUser.objects.filter(
            company=requirement_company,
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')
    else:
        cabinet_users = CabinetUser.objects.filter(
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')
    
    users_data = []
    for cu in cabinet_users:
        users_data.append({
            'id': cu.user.id,
            'username': cu.user.username,
            'email': cu.user.email,
            'full_name': cu.user.get_full_name(),
            'position': str(cu.position) if cu.position else None,
            'department': str(cu.department) if cu.department else None,
        })
    
    # Localized labels
    company_type_labels = []
    for company_type in requirement.company_types.all():
        if use_localized_labels:
            label = company_type.get_local_name(country_for_language) if country_for_language else (company_type.name_local or company_type.name)
        else:
            label = company_type.name
        company_type_labels.append({
            'label': label,
            'color': getattr(company_type, 'color', '#6c757d')
        })
    
    requirement_type_options = get_dictionary_options(
        RequirementType,
        RequirementTypeTranslation,
        'requirement_type',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(LocalComplianceRequirement.REQUIREMENT_TYPE_CHOICES)
    )
    requirement_type_label_map = {opt['code']: opt['label'] for opt in requirement_type_options}
    localized_requirement_type = requirement_type_label_map.get(
        requirement.requirement_type,
        requirement.get_requirement_type_display()
    )
    
    status_options = get_dictionary_options(
        RequirementStatus,
        RequirementStatusTranslation,
        'requirement_status',
        language_code,
        country_for_language,
        use_localized_labels,
        fallback_map=dict(LocalComplianceRequirement.STATUS_CHOICES)
    )
    status_label_map = {opt['code']: opt['label'] for opt in status_options}
    localized_status = status_label_map.get(
        requirement.status,
        requirement.get_status_display()
    )
    
    # Get notes for this requirement
    from .models import LocalRequirementNote
    requirement_notes = LocalRequirementNote.objects.filter(
        requirement=requirement,
        is_active=True
    ).select_related('created_by').order_by('-created_date')
    
    context = {
        'requirement': requirement,
        'categories': categories,
        'uncategorized_controls': uncategorized_controls,
        'completion': completion,
        'stats': stats,
        'recent_logs': recent_logs,
        'search': search,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'owner_filter': owner_filter,
        'status_choices': status_choices,
        'priority_choices': priority_choices,
        'permissions': permissions,
        'users': users_data,
        'today': date.today(),
        'company_type_labels': company_type_labels,
        'localized_requirement_type': localized_requirement_type,
        'localized_status': localized_status,
        'requirement_notes': requirement_notes,
    }
    
    return render(request, 'app_compliance/local_requirement_instance_detail.html', context)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_note_create(request, requirement_id):
    """Create note for local requirement instance"""
    requirement = get_object_or_404(LocalComplianceRequirement, id=requirement_id, is_template=False)

    if not check_user_local_compliance_permission(request.user, 'can_edit_requirement_instances'):
        messages.error(request, _('You do not have permission to add notes'))
        return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement_id)

    # Check company access
    if requirement.company:
        accessible_companies = get_user_accessible_companies_local(request.user)
        if requirement.company not in accessible_companies:
            messages.error(request, _('You do not have access to this company'))
            return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement_id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement_id)

        from .models import LocalRequirementNote, LocalRequirementNoteAttachment
        note = LocalRequirementNote.objects.create(
            requirement=requirement,
            note=note_text,
            attachment=request.FILES.get('attachment'),
            created_by=request.user,
        )

        # Multiple attachments
        for f in request.FILES.getlist('attachments'):
            LocalRequirementNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'create',
            'local_requirement_note',
            note,
            request=request,
        )

        messages.success(request, _('Note added successfully'))

    except Exception as exc:
        messages.error(request, _('Error adding note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_note_update(request, note_id):
    """Update local requirement note"""
    from .models import LocalRequirementNote, LocalRequirementNoteAttachment
    note = get_object_or_404(LocalRequirementNote, id=note_id)
    requirement = note.requirement

    if not check_user_local_compliance_permission(request.user, 'can_edit_requirement_instances'):
        messages.error(request, _('You do not have permission to edit notes'))
        return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

    # Check company access
    if requirement.company:
        accessible_companies = get_user_accessible_companies_local(request.user)
        if requirement.company not in accessible_companies:
            messages.error(request, _('You do not have access to this company'))
            return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

    try:
        note_text = (request.POST.get('note') or '').strip()
        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

        note.note = note_text

        # Clear attachment if requested
        if request.POST.get('clear_attachment') == 'on':
            note.attachment = None

        # New single attachment
        if 'attachment' in request.FILES:
            note.attachment = request.FILES['attachment']

        note.save()

        # Multiple new attachments
        for f in request.FILES.getlist('attachments'):
            LocalRequirementNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'update',
            'local_requirement_note',
            note,
            request=request,
        )

        messages.success(request, _('Note updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_note_delete(request, note_id):
    """Delete local requirement note"""
    from .models import LocalRequirementNote
    note = get_object_or_404(LocalRequirementNote, id=note_id)
    requirement = note.requirement

    if not check_user_local_compliance_permission(request.user, 'can_edit_requirement_instances'):
        messages.error(request, _('You do not have permission to delete notes'))
        return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

    # Check company access
    if requirement.company:
        accessible_companies = get_user_accessible_companies_local(request.user)
        if requirement.company not in accessible_companies:
            messages.error(request, _('You do not have access to this company'))
            return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

    try:
        log_compliance_action(
            request.user,
            'delete',
            'local_requirement_note',
            note,
            changes=f'Deleted note: {note.note[:50]}...',
            request=request,
        )

        note.delete()
        messages.success(request, _('Note deleted successfully'))

    except Exception as exc:
        messages.error(request, _('Error deleting note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_note_attachment_delete(request, attachment_id):
    """Delete an attachment from local requirement note"""
    from .models import LocalRequirementNoteAttachment
    attachment = get_object_or_404(LocalRequirementNoteAttachment, id=attachment_id)
    note = attachment.note
    requirement = note.requirement

    if not check_user_local_compliance_permission(request.user, 'can_edit_requirement_instances'):
        messages.error(request, _('You do not have permission to delete attachments'))
        return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

    # Check company access
    if requirement.company:
        accessible_companies = get_user_accessible_companies_local(request.user)
        if requirement.company not in accessible_companies:
            messages.error(request, _('You do not have access to this company'))
            return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)

    try:
        attachment.delete()
        messages.success(request, _('Attachment deleted successfully'))
    except Exception as exc:
        messages.error(request, _('Error deleting attachment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_requirement_instance_detail', requirement_id=requirement.id)


@login_required
@local_compliance_access_required
def local_requirement_template_export_excel(request, requirement_id):
    """Export single Local Requirement Template with its categories & controls"""
    requirement = get_object_or_404(
        LocalComplianceRequirement.objects.select_related(
            'regulator__country',
            'regulator__regulator_type',
            'created_by'
        ).prefetch_related('company_types'),
        id=requirement_id,
        is_template=True
    )

    if not check_user_compliance_permission(request.user, 'can_view_frameworks'):
        messages.error(request, _('You do not have permission to export this requirement'))
        return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)

    template_controls_qs = requirement.controls.filter(company__isnull=True).select_related('category').order_by('code')
    categories = list(
        requirement.categories.order_by('order', 'code').prefetch_related(
            Prefetch(
                'controls',
                queryset=template_controls_qs,
                to_attr='template_controls'
            )
        )
    )
    uncategorized_controls = list(template_controls_qs.filter(category__isnull=True))

    # Use import-compatible sheet names and structure: "Requirement" and "Controls"
    wb = Workbook()
    ws_req = wb.active
    ws_req.title = "Requirement"

    company_type_ids = ','.join(str(ct.id) for ct in requirement.company_types.all())

    # Requirement sheet: same format as Template/Import expects (Field, Value, Notes)
    # Use English field names so import can parse regardless of locale
    req_rows = [
        ["Field", "Value", "Notes"],
        ["Code", requirement.code, str(_("Required. Unique code in the system."))],
        ["Name", requirement.name, str(_("Full requirement name."))],
        ["Name Local", requirement.name_local or '', str(_("Optional local name."))],
        ["Country", str(requirement.regulator.country_id) if requirement.regulator else '', ''],
        ["Regulator ID", str(requirement.regulator_id) if requirement.regulator else '', ''],
        ["Requirement Type", requirement.requirement_type, str(_("Allowed: regulation, law, decree, guidance, other"))],
        ["Status", requirement.status, str(_("Allowed: active, draft, archived"))],
        ["Priority", requirement.priority, str(_("Allowed: low, medium, high, critical"))],
        ["Mandatory", "Yes" if requirement.is_mandatory else "No", str(_("Yes/No"))],
        ["Applicable To", requirement.applicable_to or '', ''],
        ["Official Link", requirement.official_link or '', ''],
        ["Publication Date", requirement.publication_date.strftime('%d.%m.%Y') if requirement.publication_date else '', ''],
        ["Effective Date", requirement.effective_date.strftime('%d.%m.%Y') if requirement.effective_date else '', ''],
        ["Deadline Date", requirement.deadline_date.strftime('%d.%m.%Y') if requirement.deadline_date else '', ''],
        ["Company Type IDs", company_type_ids, ''],
        ["Description", requirement.description or '', ''],
    ]
    for row_idx, row_data in enumerate(req_rows, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_req.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    ws_req.column_dimensions['A'].width = 25
    ws_req.column_dimensions['B'].width = 40
    ws_req.column_dimensions['C'].width = 80

    ws_ctrl = wb.create_sheet("Controls")
    # Same structure as Template/Import: Category Code, Category Name, Category Description,
    # Control Code, Control Name, Priority, Target Date, Implementation Notes, Evidence Notes, Description
    control_headers = [
        "Category Code",
        "Category Name",
        "Category Description",
        "Control Code",
        "Control Name",
        "Priority",
        "Target Date",
        "Implementation Notes",
        "Evidence Notes",
        "Description",
    ]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    ws_ctrl.append(control_headers)
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx in range(1, len(control_headers) + 1):
        cell = ws_ctrl.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    category_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    current_row = 2
    for category in categories:
        ws_ctrl.append([
            category.code,
            category.name,
            category.description or '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
        ])
        for col in range(1, len(control_headers) + 1):
            cell = ws_ctrl.cell(row=current_row, column=col)
            cell.fill = category_fill
            cell.font = Font(bold=True)
            cell.border = thin_border
        current_row += 1

        controls = getattr(category, 'template_controls', [])
        for control in controls:
            ws_ctrl.append([
                '',
                '',
                '',
                control.code,
                control.name,
                control.priority,
                control.target_completion_date.strftime('%d.%m.%Y') if control.target_completion_date else '',
                control.implementation_notes or '',
                control.evidence_notes or '',
                control.description or '',
            ])
            for col in range(1, len(control_headers) + 1):
                cell = ws_ctrl.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top')
            current_row += 1

    if uncategorized_controls:
        ws_ctrl.append([
            'Uncategorized',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
        ])
        for col in range(1, len(control_headers) + 1):
            cell = ws_ctrl.cell(row=current_row, column=col)
            cell.fill = category_fill
            cell.font = Font(bold=True)
            cell.border = thin_border
        current_row += 1

        for control in uncategorized_controls:
            ws_ctrl.append([
                '',
                '',
                '',
                control.code,
                control.name,
                control.priority,
                control.target_completion_date.strftime('%d.%m.%Y') if control.target_completion_date else '',
                control.implementation_notes or '',
                control.evidence_notes or '',
                control.description or '',
            ])
            for col in range(1, len(control_headers) + 1):
                cell = ws_ctrl.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top')
            current_row += 1

    for column in ws_ctrl.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_ctrl.column_dimensions[column_letter].width = min(max_length + 4, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{requirement.code}_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    return response


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_template_apply(request, requirement_id):
    """Застосувати requirement template до компаній"""
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_edit_requirements'):
        messages.error(request, _('You do not have permission to apply requirements'))
        return redirect('compliance:local_requirements_library')
    
    requirement = get_object_or_404(LocalComplianceRequirement, id=requirement_id, is_template=True)
    
    try:
        raw_selected = request.POST.getlist('companies')
        selected_companies = {
            int(cid) for cid in raw_selected if cid and cid.isdigit()
        }
        
        accessible_companies = get_user_accessible_companies_local(request.user)
        accessible_company_ids = set(accessible_companies.values_list('id', flat=True))
        valid_selected_ids = {cid for cid in selected_companies if cid in accessible_company_ids}
        
        existing_instances_qs = requirement.instances.filter(
            company_id__in=accessible_company_ids,
            company__isnull=False
        )
        existing_instances = {instance.company_id: instance for instance in existing_instances_qs if instance.company_id}
        
        created_count = 0
        updated_mandatory = 0
        removed_count = 0
        
        with transaction.atomic():
            new_company_ids = valid_selected_ids - set(existing_instances.keys())
            for company_id in new_company_ids:
                company = Company.objects.get(id=company_id)
                instance = requirement.apply_to_company(company, created_by=request.user)
                is_mandatory = request.POST.get(f'mandatory_{company_id}') == '1'
                if instance.is_mandatory != is_mandatory:
                    instance.is_mandatory = is_mandatory
                    instance.save(update_fields=['is_mandatory'])
                created_count += 1
            
            # Update mandatory flags for existing instances
            for company_id, instance in existing_instances.items():
                if company_id in valid_selected_ids:
                    is_mandatory = request.POST.get(f'mandatory_{company_id}') == '1'
                    if instance.is_mandatory != is_mandatory:
                        instance.is_mandatory = is_mandatory
                        instance.save(update_fields=['is_mandatory'])
                        updated_mandatory += 1
            
            # Remove deselected companies
            removed_ids = set(existing_instances.keys()) - valid_selected_ids
            if removed_ids:
                removed_count = LocalComplianceRequirement.objects.filter(
                    template=requirement,
                    company_id__in=removed_ids
                ).delete()[0]
        
        if created_count:
            messages.success(request, _('Successfully applied to %(count)d companies') % {'count': created_count})
        if updated_mandatory:
            messages.info(request, _('Updated mandatory status for %(count)d companies') % {'count': updated_mandatory})
        if removed_count:
            messages.warning(request, _('Removed %(count)d company applications') % {'count': removed_count})
        if not (created_count or updated_mandatory or removed_count):
            messages.info(request, _('No changes were made'))
        
    except Exception as e:
        messages.error(request, f'Error applying requirement: {str(e)}')
    
    # Check if there's a next parameter for redirect
    next_url = request.POST.get('next')
    if next_url:
        return redirect(next_url)
    
    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_category_create(request, requirement_id):
    """Створення категорії для requirement template"""
    if not check_user_local_compliance_permission(request.user, 'can_add_controls'):
        messages.error(request, _('You do not have permission to add categories'))
        return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)

    requirement = get_object_or_404(LocalComplianceRequirement, id=requirement_id, is_template=True)

    try:
        order_value = request.POST.get('order') or 0
        order_value = int(order_value)
    except ValueError:
        order_value = 0

    try:
        category = LocalRequirementCategory.objects.create(
            requirement=requirement,
            code=request.POST.get('code', '').strip() or _('CAT'),
            name=request.POST.get('name', '').strip() or _('New Category'),
            description=request.POST.get('description', '').strip(),
            order=order_value
        )

        log_compliance_action(
            request.user, 'create', 'local_requirement_category', category,
            request=request
        )

        messages.success(request, _('Category created successfully'))
    except Exception as e:
        messages.error(request, _('Error creating category: %(error)s') % {'error': str(e)})

    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_category_update(request, category_id):
    """Оновлення категорії"""
    if not check_user_local_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit categories'))
        category = get_object_or_404(LocalRequirementCategory, id=category_id)
        return redirect('compliance:local_requirement_template_detail', requirement_id=category.requirement.id)

    category = get_object_or_404(LocalRequirementCategory, id=category_id, requirement__is_template=True)
    requirement_id = category.requirement.id

    try:
        old_values = {
            'code': category.code,
            'name': category.name,
            'description': category.description,
            'order': category.order,
        }

        category.code = request.POST.get('code', category.code).strip()
        category.name = request.POST.get('name', category.name).strip()
        category.description = request.POST.get('description', category.description).strip()
        
        try:
            order_value = request.POST.get('order') or category.order
            category.order = int(order_value)
        except ValueError:
            pass

        category.save()

        new_values = {
            'code': category.code,
            'name': category.name,
            'description': category.description,
            'order': category.order,
        }

        log_compliance_action(
            request.user, 'update', 'local_requirement_category', category,
            changes={'old': old_values, 'new': new_values},
            request=request
        )

        messages.success(request, _('Category updated successfully'))

    except Exception as e:
        messages.error(request, _('Error updating category: %(error)s') % {'error': str(e)})

    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_category_delete(request, category_id):
    """Видалення категорії"""
    if not check_user_local_compliance_permission(request.user, 'can_delete_controls'):
        messages.error(request, _('You do not have permission to delete categories'))
        return redirect('compliance:local_requirements_library')

    category = get_object_or_404(LocalRequirementCategory, id=category_id, requirement__is_template=True)
    requirement_id = category.requirement_id

    try:
        log_compliance_action(
            request.user, 'delete', 'local_requirement_category', category,
            request=request
        )
        category.delete()
        messages.success(request, _('Category deleted successfully'))
    except Exception as e:
        messages.error(request, _('Error deleting category: %(error)s') % {'error': str(e)})

    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_control_create(request, requirement_id):
    """Створення контролю для requirement template"""
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_add_controls'):
        messages.error(request, _('You do not have permission to add controls'))
        return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)
    
    requirement = get_object_or_404(LocalComplianceRequirement, id=requirement_id, is_template=True)
    
    try:
        category = None
        category_id = request.POST.get('category')
        if category_id:
            category = LocalRequirementCategory.objects.filter(
                id=category_id,
                requirement=requirement
            ).first()

        def parse_required_count(value):
            try:
                return max(0, int(value))
            except (TypeError, ValueError):
                return 1

        required_evidence_count = parse_required_count(request.POST.get('required_evidence_count', 1))

        control = LocalComplianceControl.objects.create(
            requirement=requirement,
            company=None,  # Template control без company
            code=request.POST.get('code'),
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            status='not_started',
            priority=request.POST.get('priority', 'medium'),
            implementation_notes=request.POST.get('implementation_notes', ''),
            evidence_notes=request.POST.get('evidence_notes', ''),
            required_evidence_count=required_evidence_count,
            category=category,
            created_by=request.user
        )
        
        # Handle target_completion_date
        if request.POST.get('target_completion_date'):
            control.target_completion_date = parse_local_requirement_date(
                request.POST.get('target_completion_date')
            )
            control.save()
        
        log_compliance_action(
            request.user, 'create', 'local_control', control,
            request=request
        )
        
        messages.success(request, _('Control created successfully'))
        
    except Exception as e:
        messages.error(request, f'Error creating control: {str(e)}')
    
    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_control_update(request, control_id):
    """Оновлення контролю template"""
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit controls'))
        control = get_object_or_404(LocalComplianceControl, id=control_id)
        return redirect('compliance:local_requirement_template_detail', requirement_id=control.requirement.id)
    
    control = get_object_or_404(LocalComplianceControl, id=control_id, company__isnull=True)
    requirement_id = control.requirement.id
    
    try:
        old_values = {
            'code': control.code,
            'name': control.name,
            'description': control.description,
            'priority': control.priority,
            'category': control.category.id if control.category else None,
        }
        
        control.code = request.POST.get('code', control.code).strip()
        control.name = request.POST.get('name', control.name).strip()
        control.description = request.POST.get('description', control.description).strip()
        control.priority = request.POST.get('priority', control.priority)
        control.implementation_notes = request.POST.get('implementation_notes', control.implementation_notes).strip()
        control.evidence_notes = request.POST.get('evidence_notes', control.evidence_notes).strip()
        try:
            control.required_evidence_count = max(0, int(request.POST.get('required_evidence_count', control.required_evidence_count)))
        except (TypeError, ValueError):
            pass
        
        # Handle category
        category_id = request.POST.get('category')
        if category_id:
            category = LocalRequirementCategory.objects.filter(
                id=category_id,
                requirement=control.requirement
            ).first()
            control.category = category
        else:
            control.category = None
        
        # Handle target_completion_date
        target_date = request.POST.get('target_completion_date')
        if target_date:
            control.target_completion_date = parse_local_requirement_date(target_date)
        elif request.POST.get('target_completion_date') == '':
            control.target_completion_date = None
        
        control.save()
        
        new_values = {
            'code': control.code,
            'name': control.name,
            'description': control.description,
            'priority': control.priority,
            'category': control.category.id if control.category else None,
        }
        
        log_compliance_action(
            request.user, 'update', 'local_control', control,
            changes={'old': old_values, 'new': new_values},
            request=request
        )
        
        messages.success(request, _('Control updated successfully'))
        
    except Exception as e:
        messages.error(request, f'Error updating control: {str(e)}')
    
    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_requirement_control_delete(request, control_id):
    """Видалення контролю template"""
    # Перевірка прав
    if not check_user_local_compliance_permission(request.user, 'can_delete_controls'):
        messages.error(request, _('You do not have permission to delete controls'))
        return redirect('compliance:local_requirements_library')
    
    control = get_object_or_404(LocalComplianceControl, id=control_id, company__isnull=True)
    requirement_id = control.requirement.id
    
    try:
        log_compliance_action(
            request.user, 'delete', 'local_control', control,
            request=request
        )
        
        control.delete()
        messages.success(request, _('Control deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting control: {str(e)}')
    
    return redirect('compliance:local_requirement_template_detail', requirement_id=requirement_id)


# ========================
# Local Control Detail & Operations
# ========================


@login_required
@local_compliance_access_required
def local_control_detail(request, control_id):
    """Деталі локального контролю з повним контекстом"""
    control = get_object_or_404(
        LocalComplianceControl.objects.select_related(
            'requirement__company',
            'requirement__regulator__country',
            'responsible__cabinet__position',
            'responsible__cabinet__department',
            'verified_by',
            'created_by',
            'category'
        ),
        id=control_id
    )

    is_instance = bool(control.company)
    # Для instance controls використовуємо can_view_requirement_instances, для template - can_view_controls
    permission_key = 'can_view_requirement_instances' if is_instance else 'can_view_controls'

    if not check_user_local_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have access to this control'))
        return redirect('compliance:local_compliance')

    if control.company:
        accessible_companies = get_user_accessible_companies_local(request.user)
        if control.company not in accessible_companies:
            messages.error(request, _('You do not have access to this company'))
            return redirect('compliance:local_compliance')

    logs = ComplianceAuditLog.objects.filter(
        object_type='local_control',
        object_id=control.id
    ).select_related('user').order_by('-timestamp')[:50]

    evidences = control.evidences.filter(is_active=True).select_related('uploaded_by', 'reviewed_by')
    assignments = control.assignments.filter(is_active=True).select_related(
        'user__cabinet__position',
        'user__cabinet__department',
        'assigned_by'
    )
    notes = control.notes.filter(is_active=True).select_related('created_by')
    mappings = LocalControlMapping.objects.filter(
        local_control=control
    ).select_related(
        'target_local_control__requirement',
        'target_internal_control__requirement',
        'target_framework_control__category__framework',
        'created_by'
    )

    from app_cabinet.models import CabinetUser

    if control.company:
        cabinet_users = CabinetUser.objects.filter(
            company=control.company,
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')
    else:
        cabinet_users = CabinetUser.objects.filter(
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')

    users = [{
        'id': cu.user.id,
        'username': cu.user.username,
        'email': cu.user.email,
        'full_name': cu.user.get_full_name(),
        'position': str(cu.position) if cu.position else None,
        'department': str(cu.department) if cu.department else None,
    } for cu in cabinet_users]

    # Filter framework controls - only from instances for this company
    if control.company:
        framework_controls = Control.objects.filter(
            category__framework__company=control.company,
            category__framework__is_template=False
        ).exclude(
            id=control.related_framework_control_id
        ).select_related('category__framework').order_by(
            'category__framework__name',
            'category__code',
            'code'
        )[:150]
    else:
        framework_controls = Control.objects.exclude(
            id=control.related_framework_control_id
        ).select_related('category__framework').order_by(
            'category__framework__name',
            'category__code',
            'code'
        )[:150]

    local_controls_for_mapping = LocalComplianceControl.objects.filter(
        requirement__company=control.company
    ).exclude(id=control.id).order_by('code')[:150]
    
    # Get internal controls for mapping
    internal_controls_for_mapping = []
    if control.company:
        internal_controls_for_mapping = InternalComplianceControl.objects.filter(
            Q(company=control.company) | Q(requirement__company=control.company)
        ).select_related('requirement').order_by('code')[:150]

    permissions = get_user_local_compliance_permissions(request.user)
    # Для редагування контролів завжди використовуємо can_edit_controls
    can_edit = permissions['can_edit_controls']

    # Get available mandatory processes and documents for the company
    from app_compliance.models import MandatoryProcess
    from app_doc.models import RegisterDocs, RelatedDocs
    mandatory_processes = MandatoryProcess.objects.filter(
        is_active=True
    ).select_related('company', 'source_document').order_by('process_name')
    
    register_docs = RegisterDocs.objects.filter(
        is_active=True
    ).exclude(
        file_doc=''
    ).select_related('company', 'status_doc').order_by('name_doc')
    
    related_docs = RelatedDocs.objects.exclude(
        file_rel_doc=''
    ).select_related('company', 'status_rel_doc').order_by('name_rel_doc')
    
    # Filter by company if control has one
    if control.company:
        mandatory_processes = mandatory_processes.filter(
            Q(company=control.company) | Q(company__isnull=True)
        )
        register_docs = register_docs.filter(
            Q(company=control.company) | Q(company__isnull=True)
        )
        related_docs = related_docs.filter(
            Q(company=control.company) | Q(company__isnull=True)
        )

    # Get language preferences for localization
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    
    # Get active evidence types with localized names
    evidence_types_qs = EvidenceType.objects.filter(is_active=True).order_by('display_order', 'name')
    
    # Build localized evidence types list
    evidence_types = []
    if country_for_language and use_localized_labels:
        # Get translations for the country
        translations = {
            et.evidence_type.code: et.name_local
            for et in EvidenceTypeTranslation.objects.filter(
                country=country_for_language,
                evidence_type__in=evidence_types_qs
            ).select_related('evidence_type')
        }
        for et in evidence_types_qs:
            localized_name = translations.get(et.code) or (et.name_local if language_code == 'uk' else et.name)
            evidence_types.append({
                'id': et.id,
                'name': localized_name,
                'code': et.code,
                'color': et.color
            })
    else:
        # Use English names
        for et in evidence_types_qs:
            evidence_types.append({
                'id': et.id,
                'name': et.name,
                'code': et.code,
                'color': et.color
            })

    context = {
        'control': control,
        'requirement': control.requirement,
        'evidences': evidences,
        'assignments': assignments,
        'notes': notes,
        'logs': logs,
        'mappings': mappings,
        'users': users,
        'framework_controls': framework_controls,
        'local_controls_for_mapping': local_controls_for_mapping,
        'internal_controls_for_mapping': internal_controls_for_mapping,
        'mandatory_processes': mandatory_processes,
        'register_docs': register_docs,
        'related_docs': related_docs,
        'permissions': permissions,
        'can_edit': can_edit,
        'is_instance': is_instance,
        'has_sufficient_evidence': control.has_sufficient_evidence(),
        'evidence_count': control.get_evidence_count(),
        'approved_evidence_count': control.get_approved_evidence_count(),
        'Control': Control,
        'LocalControlEvidence': LocalControlEvidence,
        'LocalControlAssignment': LocalControlAssignment,
        'LocalControlMapping': LocalControlMapping,
        'evidence_types': evidence_types,
    }

    return render(request, 'app_compliance/local_control_detail.html', context)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_control_update(request, control_id):
    """Оновлення детальної інформації про локальний контроль"""
    control = get_object_or_404(LocalComplianceControl, id=control_id)

    # Для редагування контролів завжди використовуємо can_edit_controls
    if not check_user_local_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to update this control'))
        return redirect('compliance:local_control_detail', control_id=control_id)

    try:
        old_values = {
            'status': control.status,
            'priority': control.priority,
            'responsible': control.responsible.username if control.responsible else None,
            'target_completion_date': control.target_completion_date.isoformat() if control.target_completion_date else None,
        }

        control.code = request.POST.get('code', control.code)
        control.name = request.POST.get('name', control.name)
        control.description = request.POST.get('description', control.description)
        control.implementation_notes = request.POST.get('implementation_notes', control.implementation_notes)
        control.evidence_notes = request.POST.get('evidence_notes', control.evidence_notes)
        control.priority = request.POST.get('priority', control.priority)

        try:
            control.required_evidence_count = max(0, int(request.POST.get('required_evidence_count', control.required_evidence_count)))
        except (TypeError, ValueError):
            pass

        new_status = request.POST.get('status', control.status)
        if new_status != control.status:
            control.status = new_status
            control.status_changed_date = timezone.now()
            if new_status == 'completed' and not control.actual_completion_date:
                control.actual_completion_date = timezone.now().date()

        target_date = request.POST.get('target_completion_date')
        if target_date:
            control.target_completion_date = parse_local_requirement_date(target_date)
        elif target_date == '':
            control.target_completion_date = None

        actual_date = request.POST.get('actual_completion_date')
        if actual_date:
            control.actual_completion_date = parse_local_requirement_date(actual_date)
        elif actual_date == '':
            control.actual_completion_date = None

        responsible_id = request.POST.get('responsible')
        if responsible_id == '' or responsible_id == '0':
            control.responsible = None
        elif responsible_id:
            try:
                responsible_id_int = int(responsible_id)
                if responsible_id_int > 0:
                    control.responsible_id = responsible_id_int
            except (ValueError, TypeError):
                pass

        control.save()

        new_values = {
            'status': control.status,
            'priority': control.priority,
            'responsible': control.responsible.username if control.responsible else None,
            'target_completion_date': control.target_completion_date.isoformat() if control.target_completion_date else None,
        }

        log_compliance_action(
            request.user,
            'update',
            'local_control',
            control,
            changes={'old': old_values, 'new': new_values},
            request=request
        )

        if old_values['status'] != control.status:
            log_compliance_action(
                request.user,
                'update',
                'local_control',
                control,
                changes={
                    'field': 'status',
                    'old_status': old_values['status'],
                    'new_status': control.status,
                    'changed_date': control.status_changed_date.isoformat() if control.status_changed_date else None
                },
                notes=_('Status changed'),
                request=request
            )

        messages.success(request, _('Control updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating control: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control_id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_assign(request, control_id):
    """Призначити користувача до локального контролю"""
    control = get_object_or_404(LocalComplianceControl, id=control_id)

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to assign users'))
        return redirect('compliance:local_control_detail', control_id=control_id)

    user_id = request.POST.get('user_id')
    assignment_type = request.POST.get('assignment_type', 'owner')
    notes = request.POST.get('notes', '')

    if not user_id:
        messages.error(request, _('Please select a user to assign'))
        return redirect('compliance:local_control_detail', control_id=control_id)

    try:
        assignment, created = LocalControlAssignment.objects.get_or_create(
            control=control,
            user_id=user_id,
            assignment_type=assignment_type,
            defaults={
                'assigned_by': request.user,
                'notes': notes,
                'is_active': True,
            }
        )

        if not created:
            assignment.is_active = True
            assignment.notes = notes or assignment.notes
            assignment.assigned_by = request.user
            assignment.save()

        log_compliance_action(
            request.user,
            'create' if created else 'update',
            'local_control_assignment',
            assignment,
            request=request
        )

        messages.success(request, _('User assigned successfully'))

    except Exception as exc:
        messages.error(request, _('Error assigning user: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control_id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_set_responsible(request, control_id):
    """Встановлення відповідальної особи для локального контролю"""
    control = get_object_or_404(
        LocalComplianceControl.objects.select_related('requirement__company'),
        id=control_id
    )
    
    is_instance = bool(control.requirement.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to set responsible'))
        return redirect('compliance:local_control_detail', control_id=control.id)
    
    try:
        old_responsible = control.responsible.username if control.responsible else None
        
        responsible_id = request.POST.get('responsible_id')
        if responsible_id:
            from django.contrib.auth.models import User
            control.responsible = User.objects.get(id=responsible_id)
            new_responsible = control.responsible.username
        else:
            control.responsible = None
            new_responsible = None
        
        control.save()
        
        log_compliance_action(
            request.user,
            'update',
            'local_control',
            control,
            changes={'old': {'responsible': old_responsible}, 'new': {'responsible': new_responsible}},
            request=request
        )
        
        if new_responsible:
            messages.success(request, _('Responsible person set successfully'))
        else:
            messages.success(request, _('Responsible person cleared'))
            
    except Exception as e:
        messages.error(request, _('Error setting responsible: %(error)s') % {'error': str(e)})
    
    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_assignment_delete(request, assignment_id):
    """Видалити (деактивувати) призначення"""
    assignment = get_object_or_404(LocalControlAssignment, id=assignment_id)
    control = assignment.control

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to remove assignments'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        assignment.is_active = False
        assignment.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'local_control_assignment',
            assignment,
            request=request
        )

        messages.success(request, _('Assignment removed'))

    except Exception as exc:
        messages.error(request, _('Error removing assignment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_control_evidence_create(request, control_id):
    """Додати доказ до локального контролю"""
    control = get_object_or_404(LocalComplianceControl, id=control_id)

    if not check_user_local_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to add evidence'))
        return redirect('compliance:local_control_detail', control_id=control_id)

    try:
        mandatory_process_id = request.POST.get('mandatory_process_id')
        mandatory_process = None
        if mandatory_process_id:
            from app_compliance.models import MandatoryProcess
            mandatory_process = MandatoryProcess.objects.filter(id=mandatory_process_id).first()

        register_doc_id = request.POST.get('register_doc_id')
        related_doc_id = request.POST.get('related_doc_id')
        register_doc = None
        related_doc = None
        
        # Get EvidenceType by ID or code, fallback to 'document' code
        evidence_type_id = request.POST.get('evidence_type', '').strip()
        evidence_type = None
        if evidence_type_id:
            try:
                # Try to get by ID first (convert to int if possible)
                try:
                    evidence_type_id_int = int(evidence_type_id)
                    evidence_type = EvidenceType.objects.get(id=evidence_type_id_int)
                except (ValueError, TypeError):
                    # If not a number, try by code
                    evidence_type = EvidenceType.objects.get(code=evidence_type_id)
            except EvidenceType.DoesNotExist:
                # Fallback to default 'document'
                evidence_type = EvidenceType.objects.filter(code='document').first()
        else:
            # Fallback to default 'document'
            evidence_type = EvidenceType.objects.filter(code='document').first()
        
        if not evidence_type:
            messages.error(request, _('Invalid evidence type selected'))
            return redirect('compliance:local_control_detail', control_id=control_id)
        
        evidence = LocalControlEvidence.objects.create(
            control=control,
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            evidence_type=evidence_type,
            text_evidence=request.POST.get('text_evidence', ''),
            external_link=request.POST.get('external_link', ''),
            mandatory_process=mandatory_process,
            register_document=None,
            related_document=None,
            uploaded_by=request.user
        )

        # Priority: uploaded file > register document > related document
        if 'file' in request.FILES:
            evidence.file = request.FILES['file']
            evidence.save()
        elif register_doc_id:
            from app_doc.models import RegisterDocs
            register_doc = RegisterDocs.objects.filter(id=register_doc_id, file_doc__isnull=False).first()
            if register_doc and register_doc.file_doc:
                evidence.file = register_doc.file_doc
                evidence.register_document = register_doc
                evidence.save()
        elif related_doc_id:
            from app_doc.models import RelatedDocs
            related_doc = RelatedDocs.objects.filter(id=related_doc_id, file_rel_doc__isnull=False).first()
            if related_doc and related_doc.file_rel_doc:
                evidence.file = related_doc.file_rel_doc
                evidence.related_document = related_doc
                evidence.save()

        log_compliance_action(
            request.user,
            'create',
            'local_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence uploaded successfully'))

    except Exception as exc:
        messages.error(request, _('Error uploading evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control_id)


@login_required
@local_compliance_access_required
def local_control_evidence_edit(request, evidence_id):
    """Повернути дані доказу для редагування або оновити їх"""
    evidence = get_object_or_404(LocalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_local_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to edit evidence'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    if request.method == 'GET':
        evidence_data = {
            'id': evidence.id,
            'title': evidence.title,
            'description': evidence.description,
            'evidence_type': evidence.evidence_type_id if evidence.evidence_type_id else None,
            'text_evidence': evidence.text_evidence,
            'external_link': evidence.external_link,
            'file_name': evidence.file.name if evidence.file else '',
            'mandatory_process_id': evidence.mandatory_process_id if evidence.mandatory_process else None,
        }
        return JsonResponse(evidence_data)

    return local_control_evidence_update(request, evidence_id)


@login_required
@local_compliance_access_required
@require_http_methods(["POST"])
def local_control_evidence_update(request, evidence_id):
    """Оновлення доказу"""
    evidence = get_object_or_404(LocalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_local_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to edit evidence'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        evidence.title = request.POST.get('title', evidence.title)
        evidence.description = request.POST.get('description', evidence.description)
        evidence.evidence_type = request.POST.get('evidence_type', evidence.evidence_type)
        evidence.text_evidence = request.POST.get('text_evidence', evidence.text_evidence)
        evidence.external_link = request.POST.get('external_link', evidence.external_link)
        
        # Handle mandatory process update
        mandatory_process_id = request.POST.get('mandatory_process_id')
        if mandatory_process_id:
            from app_compliance.models import MandatoryProcess
            evidence.mandatory_process = MandatoryProcess.objects.filter(id=mandatory_process_id).first()
        else:
            evidence.mandatory_process = None

        if 'file' in request.FILES:
            evidence.file = request.FILES['file']

        evidence.save()

        log_compliance_action(
            request.user,
            'update',
            'local_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_evidence_delete(request, evidence_id):
    """Видалити (деактивувати) доказ"""
    evidence = get_object_or_404(LocalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_local_compliance_permission(request.user, 'can_manage_evidence'):
        messages.error(request, _('You do not have permission to delete evidence'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        evidence.is_active = False
        evidence.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'local_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence deleted successfully'))

    except Exception as exc:
        messages.error(request, _('Error deleting evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_evidence_approve(request, evidence_id):
    """Схвалити доказ"""
    evidence = get_object_or_404(LocalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_local_compliance_permission(request.user, 'can_approve_evidence'):
        messages.error(request, _('You do not have permission to approve evidence'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        evidence.approval_status = 'approved'
        evidence.reviewed_by = request.user
        evidence.reviewed_date = timezone.now()
        evidence.review_comments = request.POST.get('review_comments', '')
        evidence.save()

        log_compliance_action(
            request.user,
            'approve',
            'local_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence approved'))

    except Exception as exc:
        messages.error(request, _('Error approving evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_evidence_reject(request, evidence_id):
    """Відхилити доказ"""
    evidence = get_object_or_404(LocalControlEvidence, id=evidence_id)
    control = evidence.control

    if not check_user_local_compliance_permission(request.user, 'can_approve_evidence'):
        messages.error(request, _('You do not have permission to reject evidence'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        evidence.approval_status = 'rejected'
        evidence.reviewed_by = request.user
        evidence.reviewed_date = timezone.now()
        evidence.review_comments = request.POST.get('review_comments', '')
        evidence.save()

        log_compliance_action(
            request.user,
            'reject',
            'local_control_evidence',
            evidence,
            request=request
        )

        messages.success(request, _('Evidence rejected'))

    except Exception as exc:
        messages.error(request, _('Error rejecting evidence: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_note_create(request, control_id):
    """Створити примітку до локального контролю"""
    control = get_object_or_404(LocalComplianceControl, id=control_id)

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to add notes'))
        return redirect('compliance:local_control_detail', control_id=control_id)

    try:
        note = LocalControlNote.objects.create(
            control=control,
            note=request.POST.get('note'),
            created_by=request.user
        )
        
        # Multiple attachments
        from .models import LocalControlNoteAttachment
        for f in request.FILES.getlist('attachments'):
            LocalControlNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'create',
            'local_control_note',
            note,
            request=request
        )

        messages.success(request, _('Note added successfully'))

    except Exception as exc:
        messages.error(request, _('Error adding note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control_id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_note_update(request, note_id):
    """Update note for local control"""
    note = get_object_or_404(LocalControlNote, id=note_id, is_active=True)
    control = note.control

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to edit notes'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:local_control_detail', control_id=control.id)

        note.note = note_text

        # Clear existing single attachment if requested
        if request.POST.get('clear_attachment') == '1' and note.attachment:
            note.attachment = None

        note.save()

        # Append new attachments if provided
        from .models import LocalControlNoteAttachment
        for f in request.FILES.getlist('attachments'):
            LocalControlNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'update',
            'local_control_note',
            note,
            request=request
        )

        messages.success(request, _('Note updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_note_delete(request, note_id):
    """Видалити примітку"""
    note = get_object_or_404(LocalControlNote, id=note_id)
    control = note.control

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to delete notes'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        note.is_active = False
        note.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'local_control_note',
            note,
            request=request
        )

        messages.success(request, _('Note deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_mapping_create(request):
    """Створення мапінгу для локального контролю"""
    def clean_id(value):
        """Очистити та валідувати ID значення"""
        if not value:
            return None
        # Видалити всі нечислові символи (включаючи non-breaking space \xa0)
        cleaned = ''.join(c for c in str(value).strip() if c.isdigit())
        if not cleaned:
            return None
        try:
            return int(cleaned)
        except (ValueError, TypeError):
            return None

    control_id = clean_id(request.POST.get('local_control_id'))
    target_local_id = clean_id(request.POST.get('target_local_control_id'))
    target_internal_id = clean_id(request.POST.get('target_internal_control_id'))
    target_framework_id = clean_id(request.POST.get('target_framework_control_id'))

    if not control_id:
        messages.error(request, _('Local control is required'))
        return redirect('compliance:local_compliance')

    control = get_object_or_404(LocalComplianceControl, id=control_id)

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to create mappings'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    if not target_local_id and not target_internal_id and not target_framework_id:
        messages.error(request, _('Select target control to map'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        mapping_type = request.POST.get('mapping_type', 'related')
        notes = request.POST.get('notes', '')

        mapping = LocalControlMapping.objects.create(
            local_control=control,
            target_local_control_id=target_local_id,
            target_internal_control_id=target_internal_id,
            target_framework_control_id=target_framework_id,
            mapping_type=mapping_type,
            notes=notes,
            created_by=request.user
        )

        log_compliance_action(
            request.user,
            'create',
            'local_control_mapping',
            mapping,
            request=request
        )

        messages.success(request, _('Mapping created successfully'))

    except ValidationError as exc:
        messages.error(request, _('Invalid mapping: %(error)s') % {'error': exc.messages[0]})
    except Exception as exc:
        messages.error(request, _('Error creating mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def local_control_mapping_delete(request, mapping_id):
    """Видалення мапінгу локального контролю"""
    mapping = get_object_or_404(LocalControlMapping, id=mapping_id)
    control = mapping.local_control

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to delete mappings'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        mapping.delete()
        messages.success(request, _('Mapping deleted successfully'))
    except Exception as exc:
        messages.error(request, _('Error deleting mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


# ========================
# AJAX Helper Views
# ========================

@login_required
@require_http_methods(["GET"])
def get_country_companies(request, country_id):
    """
    AJAX endpoint для отримання компаній країни
    Використовується для динамічної фільтрації в admin
    """
    try:
        country = Country.objects.get(id=country_id)
        companies = country.companies.filter(is_active=True).values('id', 'name').order_by('name')
        
        return JsonResponse({
            'companies': list(companies),
            'country_name': country.name,
            'country_code': country.code,
            'country_flag': country.flag_emoji
        })
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def translate_to_country_language(request):
    """
    AJAX endpoint для автоматичного перекладу тексту на мову країни
    Використовується в admin для перекладу Local Name
    """
    try:
        data = json.loads(request.body)
        text = data.get('text', '').strip()
        country_id = data.get('country_id')
        
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        
        # Отримуємо країну
        try:
            country = Country.objects.get(id=country_id)
        except Country.DoesNotExist:
            return JsonResponse({'error': 'Country not found'}, status=404)
        
        # Визначаємо мову країни (використовуємо code країни як мову)
        # Для більшості країн ISO code = мовний код (ua=uk для української, pl=pl для польської і т.д.)
        target_language = country.language_code if hasattr(country, 'language_code') and country.language_code else country.code.lower()
        
        # Спеціальні випадки та мапінг ISO кодів країн на мовні коди Google Translate
        language_map = {
            # Україна та сусіди
            'ua': 'uk',  # Україна -> українська
            'kz': 'kk',  # Казахстан -> казахська (або 'ru' для російської)
            'by': 'be',  # Білорусь -> білоруська
            'md': 'ro',  # Молдова -> румунська
            'ge': 'ka',  # Грузія -> грузинська
            'am': 'hy',  # Вірменія -> вірменська
            'az': 'az',  # Азербайджан -> азербайджанська
            
            # Західна Європа
            'en': 'en',  # English
            'gb': 'en',  # Great Britain -> English
            'us': 'en',  # USA -> English
            'uk': 'en',  # United Kingdom -> English
            'ch': 'de',  # Швейцарія -> німецька (також fr, it, але беремо de)
            'at': 'de',  # Австрія -> німецька
            'be': 'nl',  # Бельгія -> нідерландська (також fr)
            
            # Північна Європа
            'dk': 'da',  # Данія -> данська
            'no': 'no',  # Норвегія -> норвезька
            'se': 'sv',  # Швеція -> шведська
            'fi': 'fi',  # Фінляндія -> фінська
            'is': 'is',  # Ісландія -> ісландська
            
            # Балтія
            'ee': 'et',  # Естонія -> естонська
            'lv': 'lv',  # Латвія -> латиська
            'lt': 'lt',  # Литва -> литовська
            
            # Східна Європа
            'cz': 'cs',  # Чехія -> чеська
            'sk': 'sk',  # Словаччина -> словацька
            'hu': 'hu',  # Угорщина -> угорська
            'ro': 'ro',  # Румунія -> румунська
            'bg': 'bg',  # Болгарія -> болгарська
            
            # Балкани
            'rs': 'sr',  # Сербія -> сербська
            'hr': 'hr',  # Хорватія -> хорватська
            'si': 'sl',  # Словенія -> словенська
            'ba': 'bs',  # Боснія і Герцеговина -> боснійська
            'mk': 'mk',  # Македонія -> македонська
            'al': 'sq',  # Албанія -> албанська
            'gr': 'el',  # Греція -> грецька
            
            # Азія
            'cn': 'zh-CN',  # Китай -> китайська (спрощена)
            'tw': 'zh-TW',  # Тайвань -> китайська (традиційна)
            'jp': 'ja',  # Японія -> японська
            'kr': 'ko',  # Корея -> корейська
            'in': 'hi',  # Індія -> хінді
            'th': 'th',  # Таїланд -> тайська
            'vn': 'vi',  # В'єтнам -> в'єтнамська
            'id': 'id',  # Індонезія -> індонезійська
            
            # Близький Схід
            'il': 'iw',  # Ізраїль -> іврит
            'tr': 'tr',  # Туреччина -> турецька
            'sa': 'ar',  # Саудівська Аравія -> арабська
            'ae': 'ar',  # ОАЕ -> арабська
            'ir': 'fa',  # Іран -> перська
            
            # Латинська Америка
            'mx': 'es',  # Мексика -> іспанська
            'ar': 'es',  # Аргентина -> іспанська
            'br': 'pt',  # Бразилія -> португальська
            
            # Інші
            'za': 'af',  # ПАР -> африкаанс
        }
        target_language = language_map.get(target_language, target_language)
        
        # Перекладаємо
        from deep_translator import GoogleTranslator
        
        try:
            # Визначаємо вихідну мову (auto-detect або English)
            translator = GoogleTranslator(source='auto', target=target_language)
            translated_text = translator.translate(text)
            
            return JsonResponse({
                'success': True,
                'translated_text': translated_text,
                'target_language': target_language,
                'country_name': country.name
            })
            
        except Exception as translate_error:
            # Якщо мова не підтримується, спробуємо fallback на англійську
            error_msg = str(translate_error)
            
            if 'No support for the provided language' in error_msg:
                # Log the error for debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Language {target_language} not supported for country {country.name} ({country.code}). Returning original text.')
                
                # Повертаємо оригінальний текст якщо мова не підтримується
                return JsonResponse({
                    'success': True,
                    'translated_text': text,  # Повертаємо оригінал
                    'target_language': target_language,
                    'country_name': country.name,
                    'warning': f'Language "{target_language}" not supported, returned original text'
                })
            else:
                # Інша помилка
                raise
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


