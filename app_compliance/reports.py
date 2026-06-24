# SecBoard\SecBoard\app_compliance\reports.py

from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext as _
from io import BytesIO
import json

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_framework_to_excel(framework):
    """
    Експорт фреймворку в Excel
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Install it with: pip install openpyxl")
    
    # Створюємо workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Видаляємо дефолтний лист
    
    # Стилі
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Лист 1: Загальна інформація
    ws_info = wb.create_sheet("Framework Info")
    
    info_data = [
        ["Framework Name", framework.name],
        ["Version", framework.version],
        ["Type", framework.get_framework_type_display()],
        ["Status", framework.get_status_display()],
        ["Owner", framework.owner.username if framework.owner else "-"],
        ["Company", framework.company.name if framework.company else "-"],
        ["Created Date", framework.created_date.strftime("%Y-%m-%d")],
        ["Completion", f"{framework.get_completion_percentage()}%"],
        ["Is Mandatory", "Yes" if framework.is_mandatory else "No"],
    ]
    
    if framework.description:
        info_data.append(["Description", framework.description])
    
    for row_idx, (label, value) in enumerate(info_data, start=1):
        cell_label = ws_info.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_value = ws_info.cell(row=row_idx, column=2, value=value)
        cell_value.border = border
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 50
    
    # Лист 2: Статистика
    ws_stats = wb.create_sheet("Statistics")
    stats = framework.get_controls_by_status()
    
    ws_stats.cell(row=1, column=1, value="Control Status").font = header_font
    ws_stats.cell(row=1, column=2, value="Count").font = header_font
    ws_stats.cell(row=1, column=1).fill = header_fill
    ws_stats.cell(row=1, column=2).fill = header_fill
    
    stats_data = [
        ["Total Controls", stats['total']],
        ["Not Started", stats['not_started']],
        ["In Progress", stats['in_progress']],
        ["Ready for Review", stats['ready_for_review']],
        ["Completed", stats['completed']],
        ["Failed", stats['failed']],
    ]
    
    for row_idx, (status, count) in enumerate(stats_data, start=2):
        ws_stats.cell(row=row_idx, column=1, value=status).border = border
        ws_stats.cell(row=row_idx, column=2, value=count).border = border
    
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 15
    
    # Лист 3: Контролі по категоріям
    ws_controls = wb.create_sheet("Controls")
    
    headers = ["Category", "Code", "Name", "Domain Code", "Status", "Priority", "Owner", "Evidence", "Due Date"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_controls.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row_idx = 2
    categories = framework.categories.prefetch_related('controls').all()
    
    for category in categories:
        controls = category.controls.select_related('domain').all()
        
        for control in controls:
            ws_controls.cell(row=row_idx, column=1, value=category.name).border = border
            ws_controls.cell(row=row_idx, column=2, value=control.code).border = border
            ws_controls.cell(row=row_idx, column=3, value=control.name).border = border
            ws_controls.cell(row=row_idx, column=4, value=control.domain.code if control.domain else "-").border = border
            ws_controls.cell(row=row_idx, column=5, value=control.get_status_display()).border = border
            ws_controls.cell(row=row_idx, column=6, value=control.get_priority_display()).border = border
            ws_controls.cell(row=row_idx, column=7, value=control.owner.username if control.owner else "-").border = border
            ws_controls.cell(row=row_idx, column=8, value=f"{control.get_evidence_count()}/{control.required_evidence_count}").border = border
            ws_controls.cell(row=row_idx, column=9, value=control.target_completion_date.strftime("%Y-%m-d") if control.target_completion_date else "-").border = border
            
            row_idx += 1
    
    # Автоподбор ширины колонок
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_controls.column_dimensions[col].width = 20
    
    ws_controls.column_dimensions['C'].width = 40
    
    # Лист 4: Докази (Evidence)
    ws_evidence = wb.create_sheet("Evidence")
    
    evidence_headers = ["Control Code", "Control Name", "Evidence Title", "Type", "Status", "Uploaded By", "Upload Date"]
    for col_idx, header in enumerate(evidence_headers, start=1):
        cell = ws_evidence.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row_idx = 2
    from .models import Evidence, Control
    
    controls = Control.objects.filter(category__framework=framework).prefetch_related('evidences')
    
    for control in controls:
        for evidence in control.evidences.filter(is_active=True):
            ws_evidence.cell(row=row_idx, column=1, value=control.code).border = border
            ws_evidence.cell(row=row_idx, column=2, value=control.name).border = border
            ws_evidence.cell(row=row_idx, column=3, value=evidence.title).border = border
            ws_evidence.cell(row=row_idx, column=4, value=evidence.get_evidence_type_display()).border = border
            ws_evidence.cell(row=row_idx, column=5, value=evidence.get_approval_status_display()).border = border
            ws_evidence.cell(row=row_idx, column=6, value=evidence.uploaded_by.username if evidence.uploaded_by else "-").border = border
            ws_evidence.cell(row=row_idx, column=7, value=evidence.uploaded_date.strftime("%Y-%m-%d")).border = border
            row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_evidence.column_dimensions[col].width = 20
    
    ws_evidence.column_dimensions['B'].width = 40
    ws_evidence.column_dimensions['C'].width = 40
    
    # Зберігаємо в пам'ять
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_framework_to_pdf(framework):
    """
    Експорт фреймворку в PDF
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed. Install it with: pip install reportlab")
    
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Заголовок
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
    )
    
    title = Paragraph(f"{framework.name} {framework.version}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Загальна інформація
    heading_style = styles['Heading2']
    
    elements.append(Paragraph("Framework Information", heading_style))
    elements.append(Spacer(1, 12))
    
    info_data = [
        ["Field", "Value"],
        ["Framework Name", framework.name],
        ["Version", framework.version],
        ["Type", framework.get_framework_type_display()],
        ["Status", framework.get_status_display()],
        ["Owner", framework.owner.username if framework.owner else "-"],
        ["Company", framework.company.name if framework.company else "-"],
        ["Created Date", framework.created_date.strftime("%Y-%m-%d")],
        ["Completion", f"{framework.get_completion_percentage()}%"],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Статистика
    elements.append(Paragraph("Control Statistics", heading_style))
    elements.append(Spacer(1, 12))
    
    stats = framework.get_controls_by_status()
    stats_data = [
        ["Status", "Count"],
        ["Total Controls", str(stats['total'])],
        ["Not Started", str(stats['not_started'])],
        ["In Progress", str(stats['in_progress'])],
        ["Ready for Review", str(stats['ready_for_review'])],
        ["Completed", str(stats['completed'])],
        ["Failed", str(stats['failed'])],
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(stats_table)
    elements.append(PageBreak())
    
    # Контролі
    elements.append(Paragraph("Controls Details", heading_style))
    elements.append(Spacer(1, 12))
    
    categories = framework.categories.prefetch_related('controls').all()
    
    for category in categories:
        elements.append(Paragraph(f"{category.code} - {category.name}", styles['Heading3']))
        elements.append(Spacer(1, 8))
        
        controls = category.controls.all()
        
        if controls:
            controls_data = [["Code", "Name", "Status", "Owner"]]
            
            for control in controls:
                controls_data.append([
                    control.code,
                    control.name[:50] + "..." if len(control.name) > 50 else control.name,
                    control.get_status_display(),
                    control.owner.username if control.owner else "-"
                ])
            
            controls_table = Table(controls_data, colWidths=[1*inch, 3*inch, 1.2*inch, 1*inch])
            controls_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
            ]))
            
            elements.append(controls_table)
        else:
            elements.append(Paragraph("No controls in this category", styles['Normal']))
        
        elements.append(Spacer(1, 16))
    
    # Генеруємо PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def export_control_details_to_pdf(control):
    """
    Детальний експорт одного контролю в PDF
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Заголовок
    title = Paragraph(f"Control: {control.code} - {control.name}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Інформація про контроль
    info_data = [
        ["Field", "Value"],
        ["Code", control.code],
        ["Name", control.name],
        ["Category", f"{control.category.code} - {control.category.name}"],
        ["Framework", f"{control.category.framework.name} {control.category.framework.version}"],
        ["Status", control.get_status_display()],
        ["Priority", control.get_priority_display()],
        ["Owner", control.owner.username if control.owner else "-"],
        ["Evidence", f"{control.get_evidence_count()}/{control.required_evidence_count}"],
        ["Verified", "Yes" if control.is_verified else "No"],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Опис
    if control.description:
        elements.append(Paragraph("<b>Description:</b>", styles['Heading3']))
        elements.append(Paragraph(control.description, styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Докази
    elements.append(Paragraph("<b>Evidence:</b>", styles['Heading3']))
    evidences = control.evidences.filter(is_active=True)
    
    if evidences:
        evidence_data = [["Title", "Type", "Status", "Uploaded By"]]
        for evidence in evidences:
            evidence_data.append([
                evidence.title,
                evidence.get_evidence_type_display(),
                evidence.get_approval_status_display(),
                evidence.uploaded_by.username if evidence.uploaded_by else "-"
            ])
        
        evidence_table = Table(evidence_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        evidence_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(evidence_table)
    else:
        elements.append(Paragraph("No evidence uploaded", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

