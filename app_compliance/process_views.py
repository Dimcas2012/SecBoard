# Views for Mandatory Process Management
import json
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.db import transaction
from django.db.models import Q
from django.db import models
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.translation import gettext as _, get_language
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_http_methods, require_POST
import logging

from .models import MandatoryProcess, ProcessAttachment, ProcessExecution, ProcessEvidenceFile, ProcessFrequency, ProcessStatus, MandatoryProcessesGuide, MandatoryProcessesGuideTranslation, ComplianceAuditLog
from .utils import log_compliance_action
from app_doc.views import user_can_edit, get_mandatory_user_access_level, check_user_mandatory_edit_access, get_user_allowed_companies
from app_doc.models import RegisterDocs
from app_conf.models import Company, Country
from app_cabinet.models import CabinetUser, CabinetGroup
from django.contrib.auth.models import Group, User

logger = logging.getLogger(__name__)


def _format_user_with_info(user):
    """Format user display name with department and position info"""
    full_name = user.get_full_name()
    
    if not hasattr(user, 'cabinet') or not user.cabinet:
        return full_name
        
    cabinet = user.cabinet
    info_parts = []
    
    if cabinet.department:
        info_parts.append(cabinet.department.get_name())
    if cabinet.position:
        info_parts.append(cabinet.position.get_name())
        
    if info_parts:
        return f"{full_name} ({', '.join(info_parts)})"
    return full_name


def _format_user_name_only(user):
    """Format user display name only (without department/position)"""
    return user.get_full_name() or user.username


def _format_user_tooltip(user):
    """Format user department and position info for tooltip"""
    if not hasattr(user, 'cabinet') or not user.cabinet:
        return ''
        
    cabinet = user.cabinet
    info_parts = []
    
    if cabinet.department:
        info_parts.append(cabinet.department.get_name())
    if cabinet.position:
        info_parts.append(cabinet.position.get_name())
        
    if info_parts:
        return f"Department: {', '.join(info_parts)}"
    return ''


@login_required
def mandatory_processes(request):
    """Main view for mandatory processes page"""
    try:
        user_access = get_mandatory_user_access_level(request.user)
        if not user_access['has_access']:
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)

        user_groups = request.user.groups.all()
        
        # Handle AJAX request for DataTables
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                start = int(request.GET.get('start', 0))
                length = int(request.GET.get('length', 10))
                draw = int(request.GET.get('draw', 1))

                # Get all processes first with related data
                processes = MandatoryProcess.objects.select_related(
                    'company', 'source_document'
                ).prefetch_related(
                    'responsible_person__cabinet__department',
                    'responsible_person__cabinet__position',
                    'additional_person__cabinet__department', 
                    'additional_person__cabinet__position',
                    'attachments'
                )
                
                # Apply filters from request
                company_filter = request.GET.get('company')
                if company_filter:
                    processes = processes.filter(company_id=company_filter)
                
                priority_filter = request.GET.get('priority')
                if priority_filter:
                    processes = processes.filter(priority=priority_filter)
                
                status_filter = request.GET.get('status')
                if status_filter:
                    if status_filter == 'overdue':
                        processes = processes.filter(next_due_date__lt=timezone.now().date())
                    elif status_filter == 'upcoming':
                        processes = processes.filter(next_due_date__gte=timezone.now().date())
                    elif status_filter == 'completed':
                        processes = processes.filter(last_completed_date__isnull=False)
                    elif status_filter == 'in_progress':
                        processes = processes.filter(last_completed_date__isnull=True, next_due_date__gte=timezone.now().date())
                
                frequency_filter = request.GET.get('frequency')
                if frequency_filter:
                    processes = processes.filter(frequency=frequency_filter)
                
                responsible_filter = request.GET.get('responsible')
                if responsible_filter:
                    processes = processes.filter(responsible_person_id=responsible_filter)
                
                # Handle special actions for getting filter data
                action = request.GET.get('action')
                if action == 'get_responsible_persons':
                    responsible_persons = []
                    for process in processes:
                        for person in process.responsible_person.all():
                            if not any(p['id'] == person.id for p in responsible_persons):
                                responsible_persons.append({
                                    'id': person.id,
                                    'name': person.get_full_name() or person.username
                                })
                    
                    return JsonResponse({
                        'success': True,
                        'responsible_persons': responsible_persons
                    })
                
                elif action == 'get_companies':
                    companies = []
                    for process in processes:
                        if process.company and not any(c['id'] == process.company.id for c in companies):
                            companies.append({
                                'id': process.company.id,
                                'name': process.company.name
                            })
                    
                    return JsonResponse({
                        'success': True,
                        'companies': companies
                    })
                
                # Filter by allowed companies based on AccessMandatory settings
                allowed_companies = get_user_allowed_companies(request.user)
                if allowed_companies:
                    # If specific companies are allowed, filter by them
                    if isinstance(allowed_companies, list):
                        company_ids = [company.id for company in allowed_companies]
                        processes = processes.filter(
                            models.Q(company_id__in=company_ids) | models.Q(company__isnull=True)
                        )
                    else:
                        # If it's a QuerySet (all companies), no additional filtering needed
                        pass
                else:
                    # If no companies are allowed, return empty queryset
                    processes = MandatoryProcess.objects.none()
                
                # Filter processes based on user access
                if not request.user.is_superuser:
                    accessible_processes = []
                    for process in processes:
                        if process.has_access(request.user):
                            accessible_processes.append(process.id)
                    
                    if accessible_processes:
                        processes = processes.filter(id__in=accessible_processes)
                    else:
                        processes = MandatoryProcess.objects.none()
                
                # Get total record count
                total = processes.count()

                # Handle search
                search_value = request.GET.get('search[value]', '')
                if search_value:
                    processes = processes.filter(
                        Q(process_name__icontains=search_value) |
                        Q(description__icontains=search_value) |
                        Q(responsible_person__username__icontains=search_value) |
                        Q(responsible_person__first_name__icontains=search_value) |
                        Q(responsible_person__last_name__icontains=search_value)
                    )

                # Get filtered count
                filtered_total = processes.count()

                # Order
                order_column = int(request.GET.get('order[0][column]', 0))
                order_dir = request.GET.get('order[0][dir]', 'asc')

                # Map column index to field name (matching DataTables column order incl. checkbox)
                order_columns = [
                    None,                     # 0: Checkbox
                    'process_name',           # 1: Process Name
                    'company__name',          # 2: Company
                    'description',            # 3: Description
                    'responsible_person__username',  # 4: Responsible Person
                    'additional_person__username',   # 5: Additional Person
                    'last_completed_date',    # 6: Last Completed
                    'frequency',              # 7: Frequency
                    'next_due_date',          # 8: Next Due Date
                    'priority',               # 9: Priority
                    'next_due_date',          # 10: Status (use next_due_date for sorting)
                    'source_document__name_doc',  # 11: Source Document
                    'created_at',             # 12: Attachment (use created_at for sorting)
                    None,                     # 13: Actions
                ]
                if order_column < len(order_columns) and order_columns[order_column]:
                    order_field = order_columns[order_column]
                    if order_dir == 'desc':
                        order_field = f'-{order_field}'
                    processes = processes.order_by(order_field)

                # Apply pagination
                processes = processes[start:start + length]

                # Format data
                processes_data = []
                for process in processes:
                    # Debug logging for priority (temporarily disabled)
                    # logger.info(f"Process {process.id} - priority: '{process.priority}', display: '{process.get_priority_display()}'")
                    
                    # Format responsible persons data
                    responsible_persons = []
                    responsible_tooltips = []
                    for user in process.responsible_person.all():
                        responsible_persons.append(_format_user_name_only(user))
                        tooltip = _format_user_tooltip(user)
                        if tooltip:
                            responsible_tooltips.append(f"{_format_user_name_only(user)}: {tooltip}")
                        else:
                            responsible_tooltips.append(_format_user_name_only(user))
                    
                    # Format additional persons data
                    additional_persons = []
                    additional_tooltips = []
                    for user in process.additional_person.all():
                        additional_persons.append(_format_user_name_only(user))
                        tooltip = _format_user_tooltip(user)
                        if tooltip:
                            additional_tooltips.append(f"{_format_user_name_only(user)}: {tooltip}")
                        else:
                            additional_tooltips.append(_format_user_name_only(user))
                    
                    can_edit_result = process.can_edit(request.user)
                    
                    processes_data.append({
                        'id': process.id,
                        'process_name': process.process_name,
                        'description': process.description,
                        'frequency': process.get_frequency_display(),
                        'responsible_person': '<br>'.join(responsible_persons) or _('Not assigned'),
                        'responsible_person_tooltip': ' | '.join(responsible_tooltips) if responsible_tooltips else _('Not assigned'),
                        'additional_person': '<br>'.join(additional_persons) or '',
                        'additional_person_tooltip': ' | '.join(additional_tooltips) if additional_tooltips else '',
                        'company': {'id': process.company.id, 'name': process.company.name} if process.company else None,
                        'next_due_date': process.next_due_date.strftime('%Y-%m-%d') if process.next_due_date else '',
                        'last_completed_date': process.last_completed_date.strftime('%Y-%m-%d') if process.last_completed_date else '',
                        'priority': process.get_priority_display(),
                        'status': process.status,
                        'days_until_due': process.days_until_due,
                        'source_document': process.source_document.name_doc if process.source_document else '',
                        'source_document_id': process.source_document_id,
                        'attachments_count': process.attachments.count(),
                        'attachments': [
                            {
                                'filename': att.filename,
                                'url': att.file.url,
                                'size': att.file_size_formatted
                            }
                            for att in process.attachments.all()[:3]  # Show only first 3 in table
                        ],
                        'is_active': process.is_active,
                        'can_edit': can_edit_result
                    })

                return JsonResponse({
                    'draw': draw,
                    'recordsTotal': total,
                    'recordsFiltered': filtered_total,
                    'data': processes_data
                })

            except Exception as e:
                logger.error(f"Error processing DataTables request: {str(e)}", exc_info=True)
                return JsonResponse({
                    'error': str(e)
                }, status=500)

        # Regular page load context
        cabinet_groups = CabinetGroup.objects.select_related('group').all()
        # Source documents for Add/Edit are loaded via get_company_documents (active docs for selected company only)
        documents = RegisterDocs.objects.none()
        
        # Get allowed companies for user
        allowed_companies = get_user_allowed_companies(request.user)
        
        # Summary counts (same access logic as table: allowed companies + has_access)
        from datetime import timedelta
        from django.db.models import F
        summary_processes = MandatoryProcess.objects.filter(is_active=True).select_related(
            'company'
        ).prefetch_related(
            'responsible_person', 'additional_person', 'groups', 'access_users'
        )
        if not allowed_companies:
            summary_processes = MandatoryProcess.objects.none()
        else:
            if isinstance(allowed_companies, list):
                company_ids = [company.id for company in allowed_companies]
                summary_processes = summary_processes.filter(
                    models.Q(company_id__in=company_ids) | models.Q(company__isnull=True)
                )
        if not request.user.is_superuser:
            accessible_summary_ids = [p.id for p in summary_processes if p.has_access(request.user)]
            if accessible_summary_ids:
                summary_processes = summary_processes.filter(id__in=accessible_summary_ids)
            else:
                summary_processes = MandatoryProcess.objects.none()
        today = timezone.now().date()
        end_of_week = today + timedelta(days=(6 - today.weekday()))  # Sunday
        count_overdue = summary_processes.filter(next_due_date__lt=today).count()
        count_due_this_week = summary_processes.filter(
            next_due_date__gte=today,
            next_due_date__lte=end_of_week
        ).count()
        count_in_progress = summary_processes.filter(
            next_due_date__gte=today
        ).filter(
            models.Q(last_completed_date__isnull=True) | models.Q(last_completed_date__lt=F('next_due_date'))
        ).count()

        # Filter companies list based on allowed companies
        if allowed_companies:
            if isinstance(allowed_companies, list):
                company_ids = [company.id for company in allowed_companies]
                companies_list = Company.objects.filter(id__in=company_ids)
                # Filter cabinet groups by allowed companies
                cabinet_groups = cabinet_groups.filter(company_id__in=company_ids)
                # Get cabinet users from allowed companies: active user and currently active employee only
                _today = timezone.now().date()
                _active_emp = (Q(start_date__isnull=True) | Q(start_date__date__lte=_today)) & (Q(end_date__isnull=True) | Q(end_date__date__gte=_today))
                all_cabinet_users = CabinetUser.objects.filter(
                    user__is_active=True,
                    company_id__in=company_ids
                ).filter(_active_emp).select_related('user', 'position', 'department')
            else:
                # All companies allowed
                companies_list = Company.objects.all()
                # No additional filtering needed for documents, groups, users - only active employees
                _today = timezone.now().date()
                _active_emp = (Q(start_date__isnull=True) | Q(start_date__date__lte=_today)) & (Q(end_date__isnull=True) | Q(end_date__date__gte=_today))
                all_cabinet_users = CabinetUser.objects.filter(
                    user__is_active=True
                ).filter(_active_emp).select_related('user', 'position', 'department')
        else:
            # No companies allowed
            companies_list = Company.objects.none()
            cabinet_groups = CabinetGroup.objects.none()
            all_cabinet_users = CabinetUser.objects.none()

        context = {
            'companies': companies_list,
            'groups': cabinet_groups,
            'documents': documents,
            'user_access': user_access,
            'cabinet_users': all_cabinet_users,
            'process_frequencies': ProcessFrequency.choices,
            'process_statuses': ProcessStatus.choices,
            'summary_overdue': count_overdue,
            'summary_due_this_week': count_due_this_week,
            'summary_in_progress': count_in_progress,
        }
        return render(request, 'app_compliance/mandatory_processes.html', context)

    except Exception as e:
        logger.error(f"Error in mandatory_processes view: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def mandatory_processes_guide(request):
    """Return JSON { content: html } for the Mandatory Processes guide (localized)."""
    user_access = get_mandatory_user_access_level(request.user)
    if not user_access.get('has_access'):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = MandatoryProcessesGuide.objects.first()
    if guide:
        if country:
            trans = MandatoryProcessesGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = MandatoryProcessesGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def mandatory_processes_guide_translate(request):
    """API for AI translation of Mandatory Processes guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': _('Text is required')}, status=400)
        if not country_id:
            return JsonResponse({'error': _('Country ID is required')}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': _('Country not found')}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
@user_can_edit
def add_mandatory_process(request):
    """View for adding new mandatory process"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': _('Invalid request method')
        }, status=405)

    try:
        # Debug logging for last_completed_date
        logger.info(f"Add process - last_completed_date received: '{request.POST.get('last_completed_date')}'")
        
        with transaction.atomic():
            # Create new MandatoryProcess instance
            process = MandatoryProcess(
                process_name=request.POST.get('process_name'),
                description=request.POST.get('description'),
                frequency=request.POST.get('frequency'),
                next_due_date=request.POST.get('next_due_date'),
                last_completed_date=request.POST.get('last_completed_date') or None,
                priority=request.POST.get('priority', 'medium'),
                reminder_days=int(request.POST.get('reminder_days', 7)),
                created_by=request.user,
                updated_by=request.user
            )
            
            # Handle company
            company_id = request.POST.get('company')
            if company_id:
                process.company_id = company_id
            
            # Handle source document
            source_document_id = request.POST.get('source_document')
            if source_document_id:
                process.source_document_id = source_document_id
            
            # Save the process first before adding M2M relationships
            process.save()
            
            # Handle responsible persons (M2M)
            responsible_person_ids = request.POST.getlist('responsible_person[]')
            if responsible_person_ids:
                process.responsible_person.set(responsible_person_ids)
                
            # Handle additional persons (M2M)
            additional_person_ids = request.POST.getlist('additional_person[]')
            if additional_person_ids:
                process.additional_person.set(additional_person_ids)
            
            # Save again after M2M changes
            process.save()
                
            # Handle source document section
            source_document_section = request.POST.get('source_document_section')
            if source_document_section:
                process.source_document_section = source_document_section

            # Add groups and access users (both optional; if neither set, access is not regulated)
            groups = request.POST.getlist('groups[]')
            process.groups.set(groups)
            access_users = request.POST.getlist('access_users[]')
            process.access_users.set(access_users)
            
            # Handle multiple attachment files
            if 'attachment_files' in request.FILES:
                files = request.FILES.getlist('attachment_files')
                for file in files:
                    ProcessAttachment.objects.create(
                        process=process,
                        file=file,
                        filename=file.name,
                        uploaded_by=request.user,
                        file_size=file.size
                    )

            logger.info(f"Mandatory process {process.id} created successfully by user {request.user.id}")
            log_compliance_action(
                request.user, 'create', 'mandatory_process', process,
                request=request
            )
            return JsonResponse({
                'success': True,
                'message': _('Process added successfully'),
                'id': process.id
            })

    except ValidationError as e:
        logger.warning(f"Validation error adding process: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    except Exception as e:
        logger.error(f"Error adding process: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_can_edit
def edit_mandatory_process(request, process_id):
    """View for editing existing mandatory process"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': _('Invalid request method')
        }, status=405)

    try:
        with transaction.atomic():
            # Get the existing process
            process = MandatoryProcess.objects.get(id=process_id)
            
            # Check if user can edit
            if not process.can_edit(request.user):
                return JsonResponse({
                    'success': False,
                    'error': _('Access denied')
                }, status=403)

            # Update basic fields
            process.process_name = request.POST.get('process_name')
            process.description = request.POST.get('description')
            process.frequency = request.POST.get('frequency')
            process.next_due_date = request.POST.get('next_due_date')
            process.priority = request.POST.get('priority', 'medium')
            process.reminder_days = int(request.POST.get('reminder_days', 7))
            process.source_document_section = request.POST.get('source_document_section', '')
            process.updated_by = request.user
            
            # Handle is_active checkbox
            process.is_active = request.POST.get('is_active') == 'on'
            
            # Handle last_completed_date
            last_completed_date = request.POST.get('last_completed_date')
            if last_completed_date:
                process.last_completed_date = last_completed_date
            else:
                process.last_completed_date = None

            # Handle company
            company_id = request.POST.get('company')
            if company_id:
                process.company_id = company_id
            else:
                process.company = None
            
            # Handle source document
            source_document_id = request.POST.get('source_document')
            if source_document_id:
                process.source_document_id = source_document_id
            else:
                process.source_document = None
            
            # Save the process first before updating M2M relationships
            process.save()
            
            # Handle responsible persons (M2M)
            responsible_person_ids = request.POST.getlist('responsible_person[]')
            process.responsible_person.set(responsible_person_ids)
                
            # Handle additional persons (M2M)
            additional_person_ids = request.POST.getlist('additional_person[]')
            process.additional_person.set(additional_person_ids)

            # Handle groups and access users (both optional)
            groups = request.POST.getlist('groups[]')
            process.groups.set(groups)
            access_users = request.POST.getlist('access_users[]')
            process.access_users.set(access_users)
            
            # Handle new attachment files
            if 'attachment_files' in request.FILES:
                files = request.FILES.getlist('attachment_files')
                for file in files:
                    ProcessAttachment.objects.create(
                        process=process,
                        file=file,
                        filename=file.name,
                        uploaded_by=request.user,
                        file_size=file.size
                    )

            logger.info(f"Mandatory process {process.id} updated successfully by user {request.user.id}")
            log_compliance_action(
                request.user, 'update', 'mandatory_process', process,
                request=request, notes=_('Process updated')
            )
            return JsonResponse({
                'success': True,
                'message': _('Process updated successfully'),
                'id': process.id
            })

    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except ValidationError as e:
        logger.warning(f"Validation error editing process {process_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    except Exception as e:
        logger.error(f"Error editing process {process_id}: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_mandatory_process(request, process_id):
    """Get mandatory process data"""
    try:
        process = MandatoryProcess.objects.get(id=process_id)
        
        # Check access
        if not process.has_access(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)
            
        data = {
            'id': process.id,
            'process_name': process.process_name,
            'description': process.description,
            'company_id': process.company.id if process.company else None,
            'source_document_id': process.source_document.id if process.source_document else None,
            'source_document_section': process.source_document_section,
            'frequency': process.frequency,
            'responsible_person_ids': list(process.responsible_person.values_list('id', flat=True)),
            'additional_person_ids': list(process.additional_person.values_list('id', flat=True)),
            'responsible_persons': [{'id': u.id, 'name': u.get_full_name() or u.username} for u in process.responsible_person.all()],
            'additional_persons': [{'id': u.id, 'name': u.get_full_name() or u.username} for u in process.additional_person.all()],
            'next_due_date': process.next_due_date.strftime('%Y-%m-%d') if process.next_due_date else '',
            'last_completed_date': process.last_completed_date.strftime('%Y-%m-%d') if process.last_completed_date else None,
            'priority': process.priority,
            'reminder_days': process.reminder_days,
            'is_active': process.is_active,
            'attachments': [
                {
                    'id': att.id,
                    'filename': att.filename,
                    'url': att.file.url,
                    'size': att.file_size_formatted,
                    'uploaded_at': att.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                    'uploaded_by': att.uploaded_by.get_full_name() if att.uploaded_by else _('Unknown')
                }
                for att in process.attachments.all()
            ],
            'groups': list(process.groups.values_list('id', flat=True)),
            'groups_list': [{'id': g.id, 'name': g.name} for g in process.groups.all()],
            'access_user_ids': list(process.access_users.values_list('id', flat=True)),
            'access_users_list': [{'id': u.id, 'name': u.get_full_name() or u.username} for u in process.access_users.all()]
        }

        return JsonResponse({
            'success': True,
            'data': data
        })

    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error getting process: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




@login_required
@user_can_edit
def delete_mandatory_process(request, process_id):
    """View for deleting mandatory process"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': _('Invalid request method')
        }, status=405)

    try:
        with transaction.atomic():
            process = MandatoryProcess.objects.get(id=process_id)
            
            # Check if user can edit
            if not process.can_edit(request.user):
                return JsonResponse({
                    'success': False,
                    'error': _('Access denied')
                }, status=403)

            process_name = process.process_name
            log_compliance_action(
                request.user, 'delete', 'mandatory_process', process,
                request=request
            )
            process.delete()

            logger.info(f"Mandatory process {process_id} ({process_name}) deleted successfully by user {request.user.id}")

            return JsonResponse({
                'success': True,
                'message': _('Process deleted successfully')
            })

    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error deleting process {process_id}: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_can_edit
def mark_process_completed(request, process_id):
    """Mark process as completed and create execution record"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': _('Invalid request method')
        }, status=405)

    try:
        with transaction.atomic():
            process = MandatoryProcess.objects.get(id=process_id)
            
            # Check if user can edit
            if not process.can_edit(request.user):
                return JsonResponse({
                    'success': False,
                    'error': _('Access denied')
                }, status=403)

            # Parse execution date properly
            execution_date_str = request.POST.get('execution_date')
            if execution_date_str:
                from django.utils.dateparse import parse_date
                execution_date = parse_date(execution_date_str)
                if execution_date:
                    # Convert date to timezone-aware datetime at start of day
                    execution_date = timezone.make_aware(
                        timezone.datetime.combine(execution_date, timezone.datetime.min.time())
                    )
                else:
                    execution_date = timezone.now()
            else:
                execution_date = timezone.now()

            # Create execution record
            execution = ProcessExecution(
                process=process,
                execution_date=execution_date,
                executed_by=request.user,
                status='completed',
                notes=request.POST.get('notes', '')
            )
            
            # Handle evidence files (multiple files support)
            # Старий спосіб (для зворотної сумісності)
            if 'evidence_file' in request.FILES:
                execution.evidence_file = request.FILES['evidence_file']
            
            execution.save()
            
            # Новий спосіб - множинні файли через ProcessEvidenceFile
            if 'evidence_files' in request.FILES:
                files = request.FILES.getlist('evidence_files')
                for file in files:
                    ProcessEvidenceFile.objects.create(
                        execution=execution,
                        file=file,
                        file_name=file.name,
                        uploaded_by=request.user if request.user.is_authenticated else None
                    )

            # Update process last completed date
            if hasattr(execution.execution_date, 'date'):
                # If it's a datetime, get just the date part
                process.last_completed_date = execution.execution_date.date()
            else:
                # If it's already a date
                process.last_completed_date = execution.execution_date
            
            # Calculate next due date based on frequency
            from datetime import timedelta
            from dateutil.relativedelta import relativedelta
            base_date = process.last_completed_date
            
            if process.frequency == ProcessFrequency.DAILY:
                process.next_due_date = base_date + timedelta(days=1)
            elif process.frequency == ProcessFrequency.WEEKLY:
                process.next_due_date = base_date + timedelta(weeks=1)
            elif process.frequency == ProcessFrequency.MONTHLY:
                process.next_due_date = base_date + relativedelta(months=1)
            elif process.frequency == ProcessFrequency.QUARTERLY:
                process.next_due_date = base_date + relativedelta(months=3)
            elif process.frequency == ProcessFrequency.SEMI_ANNUALLY:
                process.next_due_date = base_date + relativedelta(months=6)
            elif process.frequency == ProcessFrequency.ANNUALLY:
                process.next_due_date = base_date + relativedelta(years=1)
            
            process.save()

            logger.info(f"Process {process.id} marked as completed by user {request.user.id}")
            _exec_date = execution.execution_date.strftime('%Y-%m-%d') if hasattr(execution.execution_date, 'strftime') else str(execution.execution_date)
            log_compliance_action(
                request.user, 'complete', 'mandatory_process', process,
                request=request,
                notes=_('Execution date: {date}').format(date=_exec_date)
            )

            # Format execution details for the success message
            execution_details = []
            
            # Format execution date properly
            if hasattr(execution.execution_date, 'strftime'):
                execution_date_str = execution.execution_date.strftime('%Y-%m-%d')
            elif hasattr(execution.execution_date, 'date'):
                execution_date_str = execution.execution_date.date().strftime('%Y-%m-%d')
            else:
                execution_date_str = str(execution.execution_date)
            
            execution_details.append(f"📅 {_('Execution Date')}: {execution_date_str}")
            execution_details.append(f"👤 {_('Executed By')}: {execution.executed_by.get_full_name() or execution.executed_by.username}")
            
            if execution.notes:
                execution_details.append(f"📝 {_('Notes')}: {execution.notes[:100]}{'...' if len(execution.notes) > 100 else ''}")
            
            # Старий спосіб (для зворотної сумісності)
            if execution.evidence_file:
                execution_details.append(f"📎 {_('Evidence File')}: {execution.evidence_file.name}")
            
            # Новий спосіб - множинні файли
            evidence_files = execution.evidence_files.all()
            if evidence_files.exists():
                file_names = [f.file_name for f in evidence_files]
                execution_details.append(f"📎 {_('Evidence Files')} ({len(file_names)}): {', '.join(file_names[:3])}{'...' if len(file_names) > 3 else ''}")
            
            execution_details.append(f"🔄 {_('Next Due Date')}: {process.next_due_date.strftime('%Y-%m-%d')}")
            
            detailed_message = f"{_('Process marked as completed successfully')}\n\n{_('Saved Information')}:\n" + "\n".join(execution_details)

            # Формуємо список файлів для JSON відповіді
            evidence_files_list = []
            if execution.evidence_file:
                evidence_files_list.append({
                    'name': execution.evidence_file.name,
                    'url': execution.evidence_file.url,
                    'is_legacy': True
                })
            for ef in evidence_files:
                evidence_files_list.append({
                    'name': ef.file_name,
                    'url': ef.file.url,
                    'size': ef.file_size,
                    'type': ef.file_type,
                    'is_archive': ef.is_archive(),
                    'is_legacy': False
                })

            return JsonResponse({
                'success': True,
                'message': detailed_message,
                'execution_details': {
                    'execution_date': execution_date_str,
                    'executed_by': execution.executed_by.get_full_name() or execution.executed_by.username,
                    'notes': execution.notes,
                    'has_evidence': bool(execution.evidence_file) or evidence_files.exists(),
                    'evidence_filename': execution.evidence_file.name if execution.evidence_file else None,
                    'evidence_files': evidence_files_list,
                    'evidence_files_count': len(evidence_files_list),
                    'next_due_date': process.next_due_date.strftime('%Y-%m-%d')
                }
            })

    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error marking process as completed: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_process_last_execution(request, process_id):
    """Get last execution data for a process to pre-fill the completion modal"""
    try:
        process = MandatoryProcess.objects.get(id=process_id)
        
        # Check access
        if not process.has_access(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)

        # Get the last execution
        last_execution = ProcessExecution.objects.filter(
            process=process
        ).select_related('executed_by').prefetch_related('evidence_files').first()
        
        if last_execution:
            # Збираємо файли з останнього виконання
            evidence_files_list = []
            if last_execution.evidence_file:
                evidence_files_list.append({
                    'name': last_execution.evidence_file.name.split('/')[-1],
                    'url': last_execution.evidence_file.url,
                    'is_legacy': True
                })
            for ef in last_execution.evidence_files.all():
                evidence_files_list.append({
                    'name': ef.file_name,
                    'url': ef.file.url,
                    'size': ef.file_size,
                    'type': ef.file_type,
                    'is_archive': ef.is_archive(),
                    'is_legacy': False
                })
            
            return JsonResponse({
                'success': True,
                'has_execution': True,
                'execution_data': {
                    'execution_date': last_execution.execution_date.strftime('%Y-%m-%d'),
                    'notes': last_execution.notes,
                    'evidence_filename': last_execution.evidence_file.name if last_execution.evidence_file else None,
                    'evidence_files': evidence_files_list,
                    'evidence_files_count': len(evidence_files_list),
                    'executed_by': last_execution.executed_by.get_full_name() if last_execution.executed_by else _('Unknown')
                }
            })
        else:
            return JsonResponse({
                'success': True,
                'has_execution': False,
                'execution_data': None
            })

    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error getting last execution: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def process_execution_history(request, process_id):
    """Get process execution history"""
    try:
        process = MandatoryProcess.objects.get(id=process_id)
        
        # Check access
        if not process.has_access(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)

        executions = ProcessExecution.objects.filter(process=process).select_related('executed_by').prefetch_related('evidence_files')
        
        executions_data = []
        for execution in executions:
            # Збираємо всі файли доказів
            evidence_files_list = []
            if execution.evidence_file:
                evidence_files_list.append({
                    'id': None,
                    'name': execution.evidence_file.name.split('/')[-1],
                    'url': execution.evidence_file.url,
                    'is_legacy': True
                })
            for ef in execution.evidence_files.all():
                evidence_files_list.append({
                    'id': ef.id,
                    'name': ef.file_name,
                    'url': ef.file.url,
                    'size': ef.file_size,
                    'type': ef.file_type,
                    'is_archive': ef.is_archive(),
                    'is_legacy': False,
                })
            
            executions_data.append({
                'id': execution.id,
                'execution_date': execution.execution_date.strftime('%Y-%m-%d'),
                'executed_by': execution.executed_by.get_full_name() if execution.executed_by else _('Unknown'),
                'status': execution.get_status_display(),
                'notes': execution.notes,
                'has_evidence': bool(execution.evidence_file) or execution.evidence_files.exists(),
                'evidence_url': execution.evidence_file.url if execution.evidence_file else None,
                'evidence_files': evidence_files_list,
                'evidence_files_count': len(evidence_files_list),
                'created_at': execution.created_at.strftime('%Y-%m-%d %H:%M')
            })

        return JsonResponse({
            'success': True,
            'executions': executions_data,
            'process_name': process.process_name,
            'can_edit': process.can_edit(request.user),
            'process_id': process.id,
        })

    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error getting execution history: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_can_edit
@require_POST
def update_execution_evidence(request, execution_id):
    """Replace or add evidence files for a process execution record."""
    try:
        execution = ProcessExecution.objects.select_related('process').get(id=execution_id)
        process = execution.process

        if not process.has_access(request.user) or not process.can_edit(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)

        replaced = []
        added = []

        replacement_file = request.FILES.get('replacement_file')
        if replacement_file:
            is_legacy = request.POST.get('is_legacy') == '1'
            evidence_file_id = request.POST.get('evidence_file_id')

            if is_legacy:
                if execution.evidence_file:
                    try:
                        execution.evidence_file.delete(save=False)
                    except Exception:
                        pass
                execution.evidence_file = replacement_file
                execution.save(update_fields=['evidence_file'])
                replaced.append(replacement_file.name)
            elif evidence_file_id:
                evidence = ProcessEvidenceFile.objects.get(id=evidence_file_id, execution=execution)
                if evidence.file:
                    try:
                        evidence.file.delete(save=False)
                    except Exception:
                        pass
                evidence.file = replacement_file
                evidence.file_name = replacement_file.name
                evidence.file_size = replacement_file.size
                evidence.uploaded_by = request.user
                evidence.save()
                replaced.append(replacement_file.name)
            else:
                return JsonResponse({
                    'success': False,
                    'error': _('Evidence file target not specified')
                }, status=400)

        if 'evidence_files' in request.FILES:
            for file in request.FILES.getlist('evidence_files'):
                ProcessEvidenceFile.objects.create(
                    execution=execution,
                    file=file,
                    file_name=file.name,
                    uploaded_by=request.user
                )
                added.append(file.name)

        if not replaced and not added:
            return JsonResponse({
                'success': False,
                'error': _('No files provided')
            }, status=400)

        notes_parts = []
        if replaced:
            notes_parts.append(_('Replaced: {files}').format(files=', '.join(replaced)))
        if added:
            notes_parts.append(_('Added: {files}').format(files=', '.join(added)))

        log_compliance_action(
            request.user, 'update', 'mandatory_process', process,
            request=request,
            notes=_('Execution evidence updated ({date}): {details}').format(
                date=execution.execution_date.strftime('%Y-%m-%d'),
                details='; '.join(notes_parts)
            )
        )

        return JsonResponse({
            'success': True,
            'message': _('Evidence updated successfully'),
            'replaced': replaced,
            'added': added,
        })

    except ProcessExecution.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Execution record not found')
        }, status=404)
    except ProcessEvidenceFile.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Evidence file not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error updating execution evidence: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_can_edit
def delete_process_attachment(request, attachment_id):
    """Delete process attachment"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': _('Invalid request method')
        }, status=405)

    try:
        attachment = ProcessAttachment.objects.get(id=attachment_id)
        
        # Check if user can edit the process
        if not attachment.process.can_edit(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)

        # Delete the file from storage
        if attachment.file:
            try:
                attachment.file.delete(save=False)
            except:
                pass  # Ignore if file doesn't exist

        # Delete the attachment record
        attachment.delete()

        logger.info(f"Process attachment {attachment_id} deleted by user {request.user.id}")

        return JsonResponse({
            'success': True,
            'message': _('Attachment deleted successfully')
        })

    except ProcessAttachment.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Attachment not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error deleting attachment: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_process_reminder_recipients(request, process_id):
    """Get recipients for process reminder emails"""
    try:
        # Get the process
        process = MandatoryProcess.objects.get(id=process_id)
        
        # Check if user has access to this process
        if not process.has_access(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)
        
        recipients = []
        
        # Add only responsible persons
        for user in process.responsible_person.all():
            if user.email:
                recipients.append({
                    'name': user.get_full_name() or user.username,
                    'email': user.email,
                    'role': _('Responsible Person')
                })
        
        
        # Remove duplicates based on email
        unique_recipients = []
        seen_emails = set()
        for recipient in recipients:
            if recipient['email'] not in seen_emails:
                unique_recipients.append(recipient)
                seen_emails.add(recipient['email'])
        
        return JsonResponse({
            'success': True,
            'recipients': unique_recipients
        })
        
    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error getting reminder recipients: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_can_edit
def send_process_reminder(request):
    """Send reminder email for mandatory process"""
    try:
        if request.method != 'POST':
            return JsonResponse({
                'success': False,
                'error': _('Invalid request method')
            }, status=405)
        
        process_id = request.POST.get('process_id')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        include_process_details = request.POST.get('include_process_details') == 'on'
        
        if not process_id or not subject or not message:
            return JsonResponse({
                'success': False,
                'error': _('Missing required fields')
            }, status=400)
        
        # Get the process
        process = MandatoryProcess.objects.get(id=process_id)
        
        # Check if user has access to this process
        if not process.has_access(request.user):
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)
        
        # Get recipients
        recipients = []
        
        # Add only responsible persons
        for user in process.responsible_person.all():
            if user.email:
                recipients.append({
                    'name': user.get_full_name() or user.username,
                    'email': user.email,
                    'role': _('Responsible Person')
                })
        
        
        # Remove duplicates based on email
        unique_recipients = []
        seen_emails = set()
        for recipient in recipients:
            if recipient['email'] not in seen_emails:
                unique_recipients.append(recipient)
                seen_emails.add(recipient['email'])
        
        if not unique_recipients:
            return JsonResponse({
                'success': False,
                'error': _('No recipients found for this process')
            }, status=400)
        
        # Send emails
        from app_doc.email_utils import send_mandatory_process_reminder
        success_count = send_mandatory_process_reminder(
            process=process,
            recipients=unique_recipients,
            subject=subject,
            message=message,
            include_process_details=include_process_details,
            sent_by=request.user
        )
        
        if success_count > 0:
            log_compliance_action(
                request.user, 'remind', 'mandatory_process', process,
                request=request,
                notes=_('Reminder sent to {count} recipients').format(count=success_count)
            )
            return JsonResponse({
                'success': True,
                'message': _('Reminder sent successfully to {count} recipients').format(count=success_count)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': _('Failed to send reminder to any recipients')
            }, status=500)
        
    except MandatoryProcess.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Process not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error sending process reminder: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def export_mandatory_processes(request):
    """Export mandatory processes to Excel (with optional attachments in ZIP)"""
    try:
        user_access = get_mandatory_user_access_level(request.user)
        if not user_access['has_access']:
            return JsonResponse({
                'success': False,
                'error': _('Access denied')
            }, status=403)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from io import BytesIO
        import zipfile
        import os
        
        # Get process IDs from request
        process_ids_json = request.POST.get('process_ids')
        export_all = request.POST.get('export_all') == 'true'
        include_attachments = request.POST.get('include_attachments') == 'true'
        
        # Get user's accessible groups and companies
        user_groups = request.user.groups.all()
        allowed_companies = get_user_allowed_companies(request.user)
        
        # Build queryset
        if export_all:
            # Apply filters if any
            processes = MandatoryProcess.objects.filter(
                is_active=True
            ).select_related(
                'company', 'source_document', 'created_by', 'updated_by'
            ).prefetch_related(
                'responsible_person', 'additional_person', 'groups', 'access_users', 'attachments'
            )
            
            # Filter by allowed companies (same as list view)
            if not allowed_companies:
                processes = MandatoryProcess.objects.none()
            else:
                company_filter = request.POST.get('company')
                if company_filter:
                    processes = processes.filter(company_id=company_filter)
                elif isinstance(allowed_companies, list):
                    company_ids = [company.id for company in allowed_companies]
                    processes = processes.filter(
                        models.Q(company_id__in=company_ids) | models.Q(company__isnull=True)
                    )
                # Apply other filters
                priority_filter = request.POST.get('priority')
                if priority_filter:
                    processes = processes.filter(priority=priority_filter)
                frequency_filter = request.POST.get('frequency')
                if frequency_filter:
                    processes = processes.filter(frequency=frequency_filter)
                responsible_filter = request.POST.get('responsible')
                if responsible_filter:
                    processes = processes.filter(responsible_person__id=responsible_filter)
            
            # Filter by user access (has_access), same as list view
            if not request.user.is_superuser:
                accessible_ids = [p.id for p in processes if p.has_access(request.user)]
                if accessible_ids:
                    processes = processes.filter(id__in=accessible_ids)
                else:
                    processes = MandatoryProcess.objects.none()
            
        else:
            # Export selected processes
            if not process_ids_json:
                return JsonResponse({
                    'success': False,
                    'error': _('No processes selected for export')
                }, status=400)
            
            process_ids = json.loads(process_ids_json)
            processes = MandatoryProcess.objects.filter(
                id__in=process_ids
            ).select_related(
                'company', 'source_document', 'created_by', 'updated_by'
            ).prefetch_related(
                'responsible_person', 'additional_person', 'groups', 'access_users', 'attachments'
            )
            
            # Restrict to allowed companies (security: only export what user can access)
            if not allowed_companies:
                processes = MandatoryProcess.objects.none()
            elif isinstance(allowed_companies, list):
                company_ids = [company.id for company in allowed_companies]
                processes = processes.filter(
                    models.Q(company_id__in=company_ids) | models.Q(company__isnull=True)
                )
            
            # Restrict to processes user has access to
            if not request.user.is_superuser:
                accessible_ids = [p.id for p in processes if p.has_access(request.user)]
                if accessible_ids:
                    processes = processes.filter(id__in=accessible_ids)
                else:
                    processes = MandatoryProcess.objects.none()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _("Mandatory Processes")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        thin_border = Side(border_style="thin", color="000000")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        data_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Add headers
        headers = [
            _('Process Name'),
            _('Company'),
            _('Description'),
            _('Responsible Person'),
            _('Additional Person'),
            _('Frequency'),
            _('Last Completed Date'),
            _('Next Due Date'),
            _('Priority'),
            _('Status'),
            _('Source Document'),
            _('Source Document Section'),
            _('Reminder Days'),
            _('Attachments Count'),
            _('Access Groups'),
            _('Access Users'),
            _('Created By'),
            _('Created At'),
            _('Updated By'),
            _('Updated At'),
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data
        for row_num, process in enumerate(processes, 2):
            # Process Name
            cell = ws.cell(row=row_num, column=1, value=process.process_name)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Company
            cell = ws.cell(row=row_num, column=2, value=process.company.name if process.company else _('All Companies'))
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Description
            cell = ws.cell(row=row_num, column=3, value=process.description)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Responsible Person
            responsible_persons = ', '.join([
                user.get_full_name() or user.username 
                for user in process.responsible_person.all()
            ])
            cell = ws.cell(row=row_num, column=4, value=responsible_persons)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Additional Person
            additional_persons = ', '.join([
                user.get_full_name() or user.username 
                for user in process.additional_person.all()
            ])
            cell = ws.cell(row=row_num, column=5, value=additional_persons)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Frequency
            cell = ws.cell(row=row_num, column=6, value=process.get_frequency_display())
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Last Completed Date
            cell = ws.cell(row=row_num, column=7, value=process.last_completed_date.strftime('%Y-%m-%d') if process.last_completed_date else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Next Due Date
            cell = ws.cell(row=row_num, column=8, value=process.next_due_date.strftime('%Y-%m-%d') if process.next_due_date else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Priority
            cell = ws.cell(row=row_num, column=9, value=process.get_priority_display())
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Status
            cell = ws.cell(row=row_num, column=10, value=process.status)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Source Document
            cell = ws.cell(row=row_num, column=11, value=process.source_document.name_doc if process.source_document else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Source Document Section
            cell = ws.cell(row=row_num, column=12, value=process.source_document_section)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Reminder Days
            cell = ws.cell(row=row_num, column=13, value=process.reminder_days)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Attachments Count
            cell = ws.cell(row=row_num, column=14, value=process.attachments.count())
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Access Groups
            groups = ', '.join([
                group.name for group in process.groups.all()
            ])
            cell = ws.cell(row=row_num, column=15, value=groups)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Access Users
            access_users = ', '.join([
                user.get_full_name() or user.username
                for user in process.access_users.all()
            ])
            cell = ws.cell(row=row_num, column=16, value=access_users)
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Created By
            cell = ws.cell(row=row_num, column=17, value=process.created_by.get_full_name() if process.created_by else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Created At
            cell = ws.cell(row=row_num, column=18, value=process.created_at.strftime('%Y-%m-%d %H:%M') if process.created_at else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Updated By
            cell = ws.cell(row=row_num, column=19, value=process.updated_by.get_full_name() if process.updated_by else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
            
            # Updated At
            cell = ws.cell(row=row_num, column=20, value=process.updated_at.strftime('%Y-%m-%d %H:%M') if process.updated_at else '')
            cell.font = data_font
            cell.alignment = cell_alignment
            cell.border = border
        
        # Adjust column widths
        column_widths = {
            1: 30,  # Process Name
            2: 20,  # Company
            3: 40,  # Description
            4: 25,  # Responsible Person
            5: 25,  # Additional Person
            6: 15,  # Frequency
            7: 15,  # Last Completed Date
            8: 15,  # Next Due Date
            9: 12,  # Priority
            10: 12, # Status
            11: 30, # Source Document
            12: 30, # Source Document Section
            13: 12, # Reminder Days
            14: 15, # Attachments Count
            15: 30, # Access Groups
            16: 30, # Access Users
            17: 20, # Created By
            18: 18, # Created At
            19: 20, # Updated By
            20: 18, # Updated At
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Save Excel to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # If attachments not included, return just Excel
        if not include_attachments:
            _export_count = processes.count()
            ComplianceAuditLog.objects.create(
                user=request.user,
                action='export',
                object_type='mandatory_process',
                object_id=0,
                object_repr=_('Mandatory processes export ({count} processes)').format(count=_export_count)[:500],
                notes=_('Exported {count} processes to Excel').format(count=_export_count),
                ip_address=request.META.get('REMOTE_ADDR') or None,
                user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:500],
            )
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"Mandatory_Processes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        # Create ZIP archive with Excel and attachments
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel file to ZIP
            excel_filename = f"Mandatory_Processes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            zip_file.writestr(excel_filename, excel_file.getvalue())
            
            # Add attachments for each process
            attachments_added = 0
            for process in processes:
                process_attachments = process.attachments.all()
                
                if process_attachments.exists():
                    # Create folder name for this process
                    # Sanitize process name for folder name
                    safe_process_name = "".join(
                        c for c in process.process_name 
                        if c.isalnum() or c in (' ', '-', '_')
                    ).strip()[:50]  # Limit length
                    
                    process_folder = f"Attachments/{safe_process_name}_({process.id})/"
                    
                    for attachment in process_attachments:
                        try:
                            # Check if file exists
                            if attachment.file and os.path.exists(attachment.file.path):
                                # Get original filename
                                original_filename = os.path.basename(attachment.file.name)
                                
                                # Add file to ZIP in process folder
                                with open(attachment.file.path, 'rb') as f:
                                    zip_file.writestr(
                                        process_folder + original_filename,
                                        f.read()
                                    )
                                attachments_added += 1
                        except Exception as e:
                            # Log error but continue with other files
                            logger.warning(
                                f"Failed to add attachment {attachment.id} to ZIP: {str(e)}"
                            )
                            continue
            
            # Add summary file
            summary = f"""Mandatory Processes Export Summary
==========================================

Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}
Total Processes: {processes.count()}
Total Attachments: {attachments_added}

Files Included:
- {excel_filename}: Excel spreadsheet with all process details
- Attachments/: Folder containing process attachments organized by process name

Notes:
- Each process with attachments has its own subfolder
- Folder names are sanitized for filesystem compatibility
- Missing or inaccessible files are skipped

Generated by: {request.user.get_full_name() or request.user.username}
"""
            zip_file.writestr('README.txt', summary.encode('utf-8'))
        
        _export_count = processes.count()
        ComplianceAuditLog.objects.create(
            user=request.user,
            action='export',
            object_type='mandatory_process',
            object_id=0,
            object_repr=_('Mandatory processes export ({count} processes)').format(count=_export_count)[:500],
            notes=_('Exported {count} processes to ZIP with attachments').format(count=_export_count),
            ip_address=request.META.get('REMOTE_ADDR') or None,
            user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:500],
        )
        zip_buffer.seek(0)
        
        # Create response with ZIP
        response = HttpResponse(
            zip_buffer.read(),
            content_type='application/zip'
        )
        
        zip_filename = f"Mandatory_Processes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.zip"
        response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting mandatory processes: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
