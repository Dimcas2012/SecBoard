from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils import timezone
from django.utils.translation import gettext_lazy as _, get_language
from django.views.decorators.http import require_http_methods, require_POST
from io import BytesIO
from app_conf.models import Company, Country, CompanyType
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date
from django.db import transaction
from django.core.exceptions import ValidationError
import PyPDF2
import re
import threading
import time
from deep_translator import GoogleTranslator

from .models import (
    ComplianceFramework, ControlCategory, Control,
    Evidence, ControlAssignment, ComplianceAuditLog, ControlMapping,
    AccessCompliance, AccessLocalCompliance, RegulatorType, LocalComplianceRegulator,
    LocalComplianceRequirement, LocalComplianceControl, LocalRequirementCategory,
    RequirementType, RequirementStatus, RequirementPriority,
    RequirementTypeTranslation, RequirementStatusTranslation,
    RequirementPriorityTranslation, EvidenceType, EvidenceTypeTranslation,
    LocalControlEvidence,
    LocalControlAssignment, LocalControlNote, LocalControlMapping,
    AccessInternalCompliance, InternalComplianceSource, InternalComplianceRequirement,
    InternalComplianceControl, InternalRequirementCategory, InternalControlEvidence,
    InternalControlAssignment, InternalControlNote, InternalControlMapping,
    AccessControlMapping, FrameworkInstanceNote, FrameworkDomain,
    FrameworkComplianceGuide, FrameworkComplianceGuideTranslation
)

from .utils import *

# Global variables for framework translation
framework_translation_in_progress = False
framework_translation_progress = {'total': 0, 'processed': 0, 'percent': 0, 'log': []}
framework_translation_stop_requested = False

# ========================
# Framework Views
# ========================


def _has_framework_compliance_access(user):
    """Check if user has access to Framework Compliance module."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return AccessCompliance.objects.filter(
        group__in=user.groups.all(),
        has_access=True
    ).exists()


@login_required
@require_http_methods(["GET"])
def framework_compliance_guide(request):
    """Return JSON { content: html } for the Framework Compliance guide (localized)."""
    if not _has_framework_compliance_access(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = FrameworkComplianceGuide.objects.first()
    if guide:
        if country:
            trans = FrameworkComplianceGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = FrameworkComplianceGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def framework_compliance_guide_translate(request):
    """API for AI translation of Framework Compliance guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@compliance_access_required
def framework_list(request):
    """Список template frameworks"""
    # Get accessible companies for the user
    accessible_companies = get_user_accessible_companies(request.user)
    accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
    
    # Show only templates that have instances in accessible companies
    # OR if user is superuser/staff (has access to all companies)
    if request.user.is_superuser or request.user.is_staff:
        # Superusers see all templates
        frameworks = ComplianceFramework.objects.filter(is_template=True)
    else:
        # Regular users see only templates that have instances in their accessible companies
        # OR templates without any instances yet (so they can apply them)
        frameworks = ComplianceFramework.objects.filter(
            Q(is_template=True) & (
                Q(instances__company_id__in=accessible_company_ids) |
                Q(instances__isnull=True)
            )
        ).distinct()
    
    frameworks = frameworks.select_related('created_by').annotate(
        total_categories=Count('categories', distinct=True),
        total_controls=Count('categories__controls', distinct=True),
        total_instances=Count('instances', distinct=True)
    )
    
    # Пошук
    search = request.GET.get('search', '')
    if search:
        frameworks = frameworks.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(version__icontains=search)
        )
    
    # Фільтр по типу
    framework_type = request.GET.get('type', '')
    if framework_type:
        frameworks = frameworks.filter(framework_type=framework_type)
    
    # Фільтр по статусу
    status = request.GET.get('status', '')
    if status:
        frameworks = frameworks.filter(status=status)
    
    # Сортування
    sort_by = request.GET.get('sort', '-created_date')
    frameworks = frameworks.order_by(sort_by)
    
    # Пагінація
    paginator = Paginator(frameworks, 20)
    page = request.GET.get('page')
    
    try:
        frameworks_page = paginator.page(page)
    except PageNotAnInteger:
        frameworks_page = paginator.page(1)
    except EmptyPage:
        frameworks_page = paginator.page(paginator.num_pages)
    
    # Додаємо статистику для кожного фреймворку
    for fw in frameworks_page:
        fw.completion = fw.get_completion_percentage()
        fw.stats = fw.get_controls_by_status()
        
        # Get instances with company and mandatory info (only for accessible companies)
        if request.user.is_superuser or request.user.is_staff:
            fw.instances_data = fw.instances.select_related('company').values(
                'id', 'company__name', 'is_mandatory'
            )
        else:
            fw.instances_data = fw.instances.filter(
                company_id__in=accessible_company_ids
            ).select_related('company').values(
                'id', 'company__name', 'is_mandatory'
            )
    
    # Отримуємо тільки ті значення фільтрів, які реально існують
    all_frameworks = ComplianceFramework.objects.filter(is_template=True)
    
    # Отримуємо унікальні типи фреймворків
    existing_types = all_frameworks.values_list('framework_type', flat=True).distinct()
    framework_types = [(value, label) for value, label in ComplianceFramework.FRAMEWORK_TYPE_CHOICES if value in existing_types]
    
    # Отримуємо унікальні статуси
    existing_statuses = all_frameworks.values_list('status', flat=True).distinct()
    status_choices = [(value, label) for value, label in ComplianceFramework.STATUS_CHOICES if value in existing_statuses]
    
    # Отримуємо permissions для користувача
    permissions = get_user_compliance_permissions(request.user)
    
    # Отримуємо доступні компанії для bulk operations
    accessible_companies = get_user_accessible_companies(request.user)
    companies = list(accessible_companies.order_by('name'))
    
    # Отримуємо доступні AI моделі для перекладу
    available_models = [('google_translate', 'Google Translator (Free, No API Key)')]
    try:
        from app_ai.models import APISettingsClaude, APISettingsGoogle, APISettingsGroq, APISettingsOllama, APISettingsDeepSeek
        
        # Перевіряємо кожну AI модель
        claude_settings = APISettingsClaude.objects.first()
        if claude_settings and claude_settings.model_name:
            available_models.append(('claude', f'Claude - {claude_settings.model_name.model_name}'))
        
        google_settings = APISettingsGoogle.objects.first()
        if google_settings and google_settings.model_name:
            available_models.append(('google', f'Google Gemini - {google_settings.model_name.model_id}'))
        
        groq_settings = APISettingsGroq.objects.first()
        if groq_settings and groq_settings.model_name:
            available_models.append(('groq', f'Groq - {groq_settings.model_name.model_name}'))
        
        ollama_settings = APISettingsOllama.objects.first()
        if ollama_settings and ollama_settings.model_name:
            available_models.append(('ollama', f'Ollama - {ollama_settings.model_name.model_id}'))
        
        deepseek_settings = APISettingsDeepSeek.objects.first()
        if deepseek_settings and deepseek_settings.model_name:
            available_models.append(('deepseek', f'DeepSeek - {deepseek_settings.model_name.model_name}'))
    except Exception as e:
        # Якщо app_ai не доступний, використовуємо тільки Google Translator
        pass
    
    context = {
        'frameworks': frameworks_page,
        'search': search,
        'selected_type': framework_type,
        'selected_status': status,
        'sort_by': sort_by,
        'framework_types': framework_types,
        'status_choices': status_choices,
        'permissions': permissions,
        'companies': companies,
        'available_models': available_models,  # Додано AI моделі
    }
    
    return render(request, 'app_compliance/framework_list.html', context)


@login_required
@compliance_access_required
def framework_detail(request, framework_id):
    """Деталі фреймворку з категоріями та контролями"""
    framework = get_object_or_404(
        ComplianceFramework.objects.select_related('company'),
        id=framework_id
    )
    
    # Перевірка прав доступу залежно від типу framework
    if framework.is_template:
        # Для template frameworks потрібен can_view_frameworks
        if not check_user_compliance_permission(request.user, 'can_view_frameworks'):
            messages.error(request, _('You do not have access to this framework'))
            return redirect('compliance:dashboard')
        
        # Для не-superuser перевіряємо чи template має instances в доступних компаніях
        if not (request.user.is_superuser or request.user.is_staff):
            accessible_companies = get_user_accessible_companies(request.user)
            accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
            
            # Перевіряємо чи є instances в доступних компаніях
            has_accessible_instances = framework.instances.filter(
                company_id__in=accessible_company_ids
            ).exists()
            
            # Дозволяємо доступ якщо є instances в доступних компаніях або template без instances
            if not has_accessible_instances and framework.instances.exists():
                messages.error(request, _('You do not have access to this framework'))
                return redirect('compliance:dashboard')
    else:
        # Для instance frameworks потрібен can_view_instance_controls
        if not check_user_compliance_permission(request.user, 'can_view_instance_controls'):
            messages.error(request, _('You do not have access to this framework'))
            return redirect('compliance:dashboard')
        
        # Також перевірка доступу до компанії
        if framework.company:
            accessible_companies = get_user_accessible_companies(request.user)
            if framework.company not in accessible_companies:
                messages.error(request, _('You do not have access to this framework'))
                return redirect('compliance:dashboard')
    
    # Отримуємо параметри фільтрів ЗАРАЗ
    status_filter = request.GET.get('control_status', '')
    priority_filter = request.GET.get('priority', '')
    owner_filter = request.GET.get('owner', '')
    domain_filter = request.GET.get('domain', '')  # ← ДОДАНО Domain filter
    search = request.GET.get('search', '')
    
    # Будую базовий queryset для controls з урахуванням фільтрів
    controls_queryset = Control.objects.select_related(
        'responsible__cabinet__position',
        'responsible__cabinet__department'
    ).prefetch_related(
        Prefetch(
            'sub_controls',
            queryset=Control.objects.select_related(
                'responsible__cabinet__position',
                'responsible__cabinet__department'
            )
        )
    ).annotate(
        evidence_count=Count('evidences', filter=Q(evidences__is_active=True))
    )
    
    # Застосовуємо фільтри до controls_queryset
    if status_filter:
        controls_queryset = controls_queryset.filter(status=status_filter)
    if priority_filter:
        controls_queryset = controls_queryset.filter(priority=priority_filter)
    if owner_filter:
        controls_queryset = controls_queryset.filter(responsible_id=owner_filter)
    if domain_filter:
        controls_queryset = controls_queryset.filter(domain_id=domain_filter)  # ← ДОДАНО Domain filter
    if search:
        controls_queryset = controls_queryset.filter(
            Q(code__icontains=search) |
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Додаємо сортування
    controls_queryset = controls_queryset.order_by('title', 'code')
    
    # Отримуємо категорії з фільтрованими контролями
    categories = framework.categories.prefetch_related(
        Prefetch('controls', queryset=controls_queryset)
    ).annotate(
        total_controls=Count('controls')
    ).order_by('order', 'code')  # Sort categories by order (Display Order), then code
    
    # Статистика (використовуємо всі controls без фільтрів)
    controls_qs = Control.objects.filter(category__framework=framework)
    
    # Статистика
    completion = framework.get_completion_percentage()
    stats = framework.get_controls_by_status()
    
    # Отримуємо тільки ті значення фільтрів, які реально існують у контролях
    all_controls = Control.objects.filter(category__framework=framework)
    
    # Отримуємо унікальні статуси
    existing_statuses = all_controls.values_list('status', flat=True).distinct()
    status_choices = [(value, label) for value, label in Control.STATUS_CHOICES if value in existing_statuses]
    
    # Отримуємо унікальні пріоритети
    existing_priorities = all_controls.values_list('priority', flat=True).distinct()
    priority_choices = [(value, label) for value, label in Control.PRIORITY_CHOICES if value in existing_priorities]
    
    # Історія змін
    recent_logs = ComplianceAuditLog.objects.filter(
        Q(object_type='framework', object_id=framework.id) |
        Q(object_type='category', object_id__in=framework.categories.values_list('id', flat=True)) |
        Q(object_type='control', object_id__in=controls_qs.values_list('id', flat=True))
    ).select_related('user').order_by('-timestamp')[:20]
    
    # Отримуємо permissions для користувача
    permissions = get_user_compliance_permissions(request.user)
    
    # Отримуємо користувачів для review owner
    from app_cabinet.models import CabinetUser
    from datetime import date
    
    framework_company = framework.company
    if framework_company:
        cabinet_users = CabinetUser.objects.filter(
            company=framework_company,
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')
    else:
        cabinet_users = CabinetUser.objects.filter(
            user__is_active=True
        ).select_related('user', 'position', 'department').order_by('user__username')
    
    users_data = []
    for cu in cabinet_users:
        users_data.append({
            'id': cu.user.id,
            'username': cu.user.username,
            'email': cu.user.email,
            'full_name': cu.user.get_full_name(),
            'position': str(cu.position) if cu.position else None,
            'department': str(cu.department) if cu.department else None,
        })
    
    # Framework-level notes (only for company instances)
    framework_notes = None
    if not framework.is_template:
        framework_notes = FrameworkInstanceNote.objects.filter(
            framework=framework,
            is_active=True
        ).select_related('created_by').prefetch_related('attachments')

    domains = FrameworkDomain.objects.filter(is_active=True).order_by('display_order', 'name')

    context = {
        'framework': framework,
        'categories': categories,
        'completion': completion,
        'stats': stats,
        'recent_logs': recent_logs,
        'search': search,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'owner_filter': owner_filter,
        'domain_filter': domain_filter,  # ← ДОДАНО Domain filter
        'status_choices': status_choices,
        'priority_choices': priority_choices,
        'permissions': permissions,  # ← ДОДАНО Permissions
        'users': users_data,  # ← ДОДАНО для lifecycle
        'today': date.today(),  # ← ДОДАНО для lifecycle
        'framework_notes': framework_notes,
        'domains': domains,
    }
    
    return render(request, 'app_compliance/framework_detail.html', context)


@login_required
def framework_create_form(request):
    """GET: показати форму створення, POST: створити фреймворк"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_add_frameworks'):
        messages.error(request, _('You do not have permission to create frameworks'))
        return redirect('compliance:framework_list')
    
    if request.method == 'GET':
        user_company = get_user_company(request.user)
        # Тільки доступні компанії
        companies = get_user_accessible_companies(request.user).order_by('name')
        
        context = {
            'framework_types': ComplianceFramework.FRAMEWORK_TYPE_CHOICES,
            'status_choices': ComplianceFramework.STATUS_CHOICES,
            'companies': companies,
            'user_company': user_company,
        }
        return render(request, 'app_compliance/framework_create_form.html', context)
    
    # POST handled by framework_create
    return framework_create(request)


@login_required
@require_http_methods(["POST"])
def framework_note_create(request, framework_id):
    """Create note for framework company instance"""
    framework = get_object_or_404(ComplianceFramework, id=framework_id)

    if framework.is_template:
        messages.error(request, _('Notes are only available for company framework instances'))
        return redirect('compliance:framework_detail', framework_id=framework_id)

    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to add notes'))
        return redirect('compliance:framework_detail', framework_id=framework_id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:framework_detail', framework_id=framework_id)

        note = FrameworkInstanceNote.objects.create(
            framework=framework,
            note=note_text,
            attachment=request.FILES.get('attachment'),
            created_by=request.user,
        )

        # Multiple attachments
        from .models import FrameworkInstanceNoteAttachment
        for f in request.FILES.getlist('attachments'):
            FrameworkInstanceNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'create',
            'framework_instance_note',
            note,
            request=request,
        )

        messages.success(request, _('Note added successfully'))

    except Exception as exc:
        messages.error(request, _('Error adding note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
@require_http_methods(["POST"])
def framework_note_update(request, note_id):
    """Update framework instance note"""
    note = get_object_or_404(FrameworkInstanceNote, id=note_id, is_active=True)
    framework = note.framework

    if framework.is_template:
        messages.error(request, _('Notes are only available for company framework instances'))
        return redirect('compliance:framework_detail', framework_id=framework.id)

    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to edit notes'))
        return redirect('compliance:framework_detail', framework_id=framework.id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:framework_detail', framework_id=framework.id)

        note.note = note_text

        # Clear existing single attachment if requested
        if request.POST.get('clear_attachment') == '1' and note.attachment:
            note.attachment = None

        # Optional single attachment (kept for compatibility)
        if 'attachment' in request.FILES:
            note.attachment = request.FILES['attachment']

        note.save()

        # Append new attachments if provided
        from .models import FrameworkInstanceNoteAttachment
        for f in request.FILES.getlist('attachments'):
            FrameworkInstanceNoteAttachment.objects.create(note=note, file=f)

        log_compliance_action(
            request.user,
            'update',
            'framework_instance_note',
            note,
            request=request,
        )

        messages.success(request, _('Note updated successfully'))

    except Exception as exc:
        messages.error(request, _('Error updating note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:framework_detail', framework_id=framework.id)


@login_required
@require_http_methods(["POST"])
def framework_note_delete(request, note_id):
    """Soft delete framework instance note"""
    note = get_object_or_404(FrameworkInstanceNote, id=note_id)
    framework = note.framework

    if framework.is_template:
        messages.error(request, _('Notes are only available for company framework instances'))
        return redirect('compliance:framework_detail', framework_id=framework.id)

    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to delete notes'))
        return redirect('compliance:framework_detail', framework_id=framework.id)

    try:
        note.is_active = False
        note.save(update_fields=['is_active'])

        log_compliance_action(
            request.user,
            'delete',
            'framework_instance_note',
            note,
            request=request,
        )

        messages.success(request, _('Note deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting note: %(error)s') % {'error': str(exc)})

    return redirect('compliance:framework_detail', framework_id=framework.id)


@login_required
@require_http_methods(["POST"])
def local_control_note_attachment_delete(request, attachment_id):
    """Delete single file attachment from local control note"""
    from .models import LocalControlNoteAttachment

    attachment = get_object_or_404(LocalControlNoteAttachment, id=attachment_id)
    note = attachment.note
    control = note.control

    is_instance = bool(control.company)
    permission_key = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, permission_key):
        messages.error(request, _('You do not have permission to delete attachments'))
        return redirect('compliance:local_control_detail', control_id=control.id)

    try:
        attachment.delete()
        messages.success(request, _('Attachment deleted'))
    except Exception as exc:
        messages.error(request, _('Error deleting attachment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:local_control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def internal_control_note_attachment_delete(request, attachment_id):
    """Delete single file attachment from internal control note"""
    from .models import InternalControlNoteAttachment

    attachment = get_object_or_404(InternalControlNoteAttachment, id=attachment_id)
    note = attachment.note
    control = note.control

    if not check_user_internal_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to delete attachments'))
        return redirect('compliance:internal_control_detail', control_id=control.id)

    try:
        attachment.delete()
        messages.success(request, _('Attachment deleted'))
    except Exception as exc:
        messages.error(request, _('Error deleting attachment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:internal_control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def framework_note_attachment_delete(request, attachment_id):
    """Delete single file attachment from framework instance note"""
    from .models import FrameworkInstanceNoteAttachment

    attachment = get_object_or_404(FrameworkInstanceNoteAttachment, id=attachment_id)
    note = attachment.note
    framework = note.framework

    if framework.is_template:
        messages.error(request, _('Notes are only available for company framework instances'))
        return redirect('compliance:framework_detail', framework_id=framework.id)

    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to delete attachments'))
        return redirect('compliance:framework_detail', framework_id=framework.id)

    try:
        attachment.delete()
        messages.success(request, _('Attachment deleted'))
    except Exception as exc:
        messages.error(request, _('Error deleting attachment: %(error)s') % {'error': str(exc)})

    return redirect('compliance:framework_detail', framework_id=framework.id)


@login_required
@require_http_methods(["POST"])
def framework_create(request):
    """Створення нового фреймворку (як template)"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_add_frameworks'):
        messages.error(request, _('You do not have permission to create frameworks'))
        return redirect('compliance:framework_list')
    
    try:
        # Get is_mandatory from checkbox
        is_mandatory = request.POST.get('is_mandatory') == '1'
        
        # Create as template by default
        framework = ComplianceFramework.objects.create(
            name=request.POST.get('name'),
            framework_type=request.POST.get('framework_type', 'custom'),
            version=request.POST.get('version', '1.0'),
            description=request.POST.get('description', ''),
            status=request.POST.get('status', 'draft'),
            is_mandatory=is_mandatory,
            is_template=True,  # Create as template
            company=None,  # Templates don't have company
            created_by=request.user
        )
        
        # Apply to selected companies if any (тільки до доступних)
        selected_companies = request.POST.getlist('companies')
        accessible_companies = get_user_accessible_companies(request.user)
        accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
        instances_created = 0
        
        for company_id in selected_companies:
            if company_id and int(company_id) in accessible_company_ids:
                company = Company.objects.get(id=company_id)
                framework.apply_to_company(company, created_by=request.user)
                instances_created += 1
        
        log_compliance_action(
            request.user, 'create', 'framework', framework,
            request=request
        )
        
        messages.success(request, _('Framework template created successfully'))
        if instances_created > 0:
            messages.success(request, f'Applied to {instances_created} companies')
        
        return redirect('compliance:framework_detail', framework_id=framework.id)
        
    except Exception as e:
        messages.error(request, f'Error creating framework: {str(e)}')
        return redirect('compliance:framework_list')


@login_required
def framework_edit_form(request, framework_id):
    """GET: показати форму редагування, POST: оновити фреймворк"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to edit frameworks'))
        return redirect('compliance:framework_list')
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    if request.method == 'GET':
        # Тільки доступні компанії
        companies = get_user_accessible_companies(request.user).order_by('name')
        
        # Get existing instances for this template
        if framework.is_template:
            instances = ComplianceFramework.objects.filter(template=framework)
            applied_companies = list(instances.values_list('company_id', flat=True))
            # Get which companies have mandatory=True
            applied_companies_mandatory = list(
                instances.filter(is_mandatory=True).values_list('company_id', flat=True)
            )
        else:
            applied_companies = []
            applied_companies_mandatory = []
        
        context = {
            'framework': framework,
            'companies': companies,
            'applied_companies': applied_companies,
            'applied_companies_mandatory': applied_companies_mandatory,
        }
        return render(request, 'app_compliance/framework_edit_form.html', context)
    
    # POST handled by framework_update
    return framework_update(request, framework_id)


@login_required
@require_http_methods(["POST"])
def framework_update(request, framework_id):
    """Оновлення фреймворку"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to edit frameworks'))
        return redirect('compliance:framework_list')
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        old_values = {
            'name': framework.name,
            'version': framework.version,
            'status': framework.status,
            'is_mandatory': framework.is_mandatory,
        }
        
        # Update basic fields
        framework.name = request.POST.get('name', framework.name)
        framework.framework_type = request.POST.get('framework_type', framework.framework_type)
        framework.version = request.POST.get('version', framework.version)
        framework.description = request.POST.get('description', framework.description)
        framework.status = request.POST.get('status', framework.status)
        framework.is_mandatory = request.POST.get('is_mandatory') == '1'
        
        # Handle template/instance logic
        if framework.is_template:
            # Template: handle company selection for instances (тільки доступні компанії)
            selected_companies = request.POST.getlist('companies')
            accessible_companies = get_user_accessible_companies(request.user)
            accessible_company_ids = set(accessible_companies.values_list('id', flat=True))
            
            # Get existing instances
            existing_instances = ComplianceFramework.objects.filter(template=framework)
            existing_company_ids = set(existing_instances.values_list('company_id', flat=True))
            # Фільтруємо тільки доступні компанії
            selected_company_ids = set(
                int(cid) for cid in selected_companies 
                if cid and int(cid) in accessible_company_ids
            )
            
            # Create new instances
            new_company_ids = selected_company_ids - existing_company_ids
            for company_id in new_company_ids:
                company = Company.objects.get(id=company_id)
                # Check if mandatory for this company
                is_mandatory_for_company = request.POST.get(f'mandatory_{company_id}') == '1'
                
                instance = framework.apply_to_company(company, created_by=request.user)
                instance.is_mandatory = is_mandatory_for_company
                instance.save()
                
                mandatory_text = " (Mandatory)" if is_mandatory_for_company else " (Optional)"
                messages.success(request, f'Applied to {company.name}{mandatory_text}')
            
            # Update mandatory status for existing instances
            for company_id in selected_company_ids & existing_company_ids:
                instance = existing_instances.get(company_id=company_id)
                is_mandatory_for_company = request.POST.get(f'mandatory_{company_id}') == '1'
                
                if instance.is_mandatory != is_mandatory_for_company:
                    instance.is_mandatory = is_mandatory_for_company
                    instance.save()
                    mandatory_text = "Mandatory" if is_mandatory_for_company else "Optional"
                    messages.info(request, f'{instance.company.name}: changed to {mandatory_text}')
            
            # Delete removed instances
            removed_company_ids = existing_company_ids - selected_company_ids
            if removed_company_ids:
                ComplianceFramework.objects.filter(
                    template=framework,
                    company_id__in=removed_company_ids
                ).delete()
                messages.info(request, f'Removed {len(removed_company_ids)} instances')
        else:
            # Instance: can change company
            company_id = request.POST.get('company')
            if company_id:
                framework.company = get_object_or_404(Company, id=company_id)
        
        framework.save()
        
        new_values = {
            'name': framework.name,
            'version': framework.version,
            'status': framework.status,
            'is_mandatory': framework.is_mandatory,
        }
        
        log_compliance_action(
            request.user, 'update', 'framework', framework,
            changes={'old': old_values, 'new': new_values},
            request=request
        )
        
        messages.success(request, _('Framework updated successfully'))
        
    except Exception as e:
        messages.error(request, f'Error updating framework: {str(e)}')
    
    return redirect('compliance:framework_detail', framework_id=framework.id)


@login_required
@require_http_methods(["POST"])
def framework_delete(request, framework_id):
    """Видалення фреймворку"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_delete_frameworks'):
        messages.error(request, _('You do not have permission to delete frameworks'))
        return redirect('compliance:framework_list')
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        log_compliance_action(
            request.user, 'delete', 'framework', framework,
            request=request
        )
        
        framework.delete()
        messages.success(request, _('Framework deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting framework: {str(e)}')
    
    return redirect('compliance:framework_list')


# ========================
# Category Views
# ========================

@login_required
@require_http_methods(["POST"])
def category_create(request, framework_id):
    """Створення категорії"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_add_controls'):
        messages.error(request, _('You do not have permission to add categories'))
        return redirect('compliance:framework_detail', framework_id=framework_id)
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        category = ControlCategory.objects.create(
            framework=framework,
            code=request.POST.get('code'),
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            order=request.POST.get('order', 0)
        )
        
        log_compliance_action(
            request.user, 'create', 'category', category,
            request=request
        )
        
        messages.success(request, _('Category created successfully'))
        
    except Exception as e:
        messages.error(request, f'Error creating category: {str(e)}')
    
    return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
@require_http_methods(["POST"])
def category_update(request, category_id):
    """Оновлення категорії"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit categories'))
        category = get_object_or_404(ControlCategory, id=category_id)
        return redirect('compliance:framework_detail', framework_id=category.framework.id)
    
    category = get_object_or_404(ControlCategory, id=category_id)
    
    try:
        old_values = {
            'code': category.code,
            'name': category.name,
        }
        
        category.code = request.POST.get('code', category.code)
        category.name = request.POST.get('name', category.name)
        category.description = request.POST.get('description', category.description)
        category.order = request.POST.get('order', category.order)
        category.save()
        
        new_values = {
            'code': category.code,
            'name': category.name,
        }
        
        log_compliance_action(
            request.user, 'update', 'category', category,
            changes={'old': old_values, 'new': new_values},
            request=request
        )
        
        messages.success(request, _('Category updated successfully'))
        
    except Exception as e:
        messages.error(request, f'Error updating category: {str(e)}')
    
    return redirect('compliance:framework_detail', framework_id=category.framework.id)


@login_required
@require_http_methods(["POST"])
def category_delete(request, category_id):
    """Видалення категорії"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_delete_controls'):
        messages.error(request, _('You do not have permission to delete categories'))
        category = get_object_or_404(ControlCategory, id=category_id)
        return redirect('compliance:framework_detail', framework_id=category.framework.id)
    
    category = get_object_or_404(ControlCategory, id=category_id)
    framework_id = category.framework.id
    
    try:
        log_compliance_action(
            request.user, 'delete', 'category', category,
            request=request
        )
        
        category.delete()
        messages.success(request, _('Category deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting category: {str(e)}')
    
    return redirect('compliance:framework_detail', framework_id=framework_id)


# ========================
# Control Views
# ========================

@login_required
def control_detail(request, control_id):
    """Деталі контролю або форма редагування (для Template)"""
    control = get_object_or_404(
        Control.objects.select_related(
            'category__framework', 
            'responsible__cabinet__position', 
            'responsible__cabinet__department',
            'verified_by', 
            'created_by', 
            'parent_control'
        ).prefetch_related(
            'evidences', 
            'sub_controls',
            Prefetch(
                'assignments',
                queryset=ControlAssignment.objects.select_related(
                    'user__cabinet__position',
                    'user__cabinet__department',
                    'assigned_by'
                )
            )
        ),
        id=control_id
    )
    
    # Перевірка прав доступу залежно від типу framework
    if control.category.framework.is_template:
        # Для template controls потрібен can_view_controls
        if not check_user_compliance_permission(request.user, 'can_view_controls'):
            messages.error(request, _('You do not have access to this control'))
            return redirect('compliance:dashboard')
        
        # Якщо це Template framework → показати форму редагування
        from .models import FrameworkDomain
        framework_domains = FrameworkDomain.objects.filter(is_active=True).order_by('display_order', 'name')
        return render(request, 'app_compliance/control_edit_form.html', {
            'control': control,
            'framework_domains': framework_domains,
        })
    else:
        # Для instance controls потрібен can_view_instance_controls
        if not check_user_compliance_permission(request.user, 'can_view_instance_controls'):
            messages.error(request, _('You do not have access to this control'))
            return redirect('compliance:dashboard')
    
    # Історія змін контролю
    logs = ComplianceAuditLog.objects.filter(
        object_type='control',
        object_id=control.id
    ).select_related('user').order_by('-timestamp')[:50]
    
    # Мапінги з іншими контролями
    mappings_to = ControlMapping.objects.filter(
        source_control=control
    ).select_related(
        'target_control__category__framework',
        'target_internal_control__requirement',
        'target_local_control__requirement'
    )
    
    mappings_from = ControlMapping.objects.filter(
        target_control=control
    ).select_related('source_control__category__framework')
    
    # Список користувачів для призначення (тільки з компанії framework)
    from django.contrib.auth.models import User
    from app_cabinet.models import CabinetUser
    
    framework_company = control.category.framework.company
    
    if framework_company:
        # Отримуємо CabinetUser з компанії
        cabinet_users = CabinetUser.objects.filter(
            company=framework_company,
            user__is_active=True
        ).select_related('user', 'position', 'department', 'company').order_by('user__username')
        
        # Формуємо список словників з повною інформацією про користувачів
        users_data = []
        for cu in cabinet_users:
            users_data.append({
                'id': cu.user.id,
                'username': cu.user.username,
                'email': cu.user.email,
                'full_name': cu.user.get_full_name(),
                'position': str(cu.position) if cu.position else None,
                'department': str(cu.department) if cu.department else None,
            })
        
        users = users_data
    else:
        # Якщо немає компанії (template), показуємо всіх активних з CabinetUser
        cabinet_users = CabinetUser.objects.filter(
            user__is_active=True
        ).select_related('user', 'position', 'department', 'company').order_by('user__username')
        
        # Формуємо список словників
        users_data = []
        for cu in cabinet_users:
            users_data.append({
                'id': cu.user.id,
                'username': cu.user.username,
                'email': cu.user.email,
                'full_name': cu.user.get_full_name(),
                'position': str(cu.position) if cu.position else None,
                'department': str(cu.department) if cu.department else None,
            })
        
        users = users_data
    
    # Список контролів для мапінгу (виключаючи поточний)
    # Отримуємо framework контролі для мапи - тільки з instances для цієї компанії
    if framework_company:
        # Filter only controls from framework instances applied to this company
        all_controls = Control.objects.filter(
            category__framework__company=framework_company,
            category__framework__is_template=False
        ).exclude(
            id=control.id
        ).select_related(
            'category__framework'
        ).order_by('category__framework__name', 'category__code', 'code')[:150]
    else:
        # For template frameworks, show all controls
        all_controls = Control.objects.exclude(id=control.id).select_related(
            'category__framework'
        ).order_by('category__framework__name', 'category__code', 'code')[:150]
    
    # Get internal controls for mapping
    internal_controls_for_mapping = []
    if framework_company:
        internal_controls_for_mapping = InternalComplianceControl.objects.filter(
            Q(company=framework_company) | Q(requirement__company=framework_company)
        ).select_related('requirement').order_by('code')[:150]
    
    # Get local controls for mapping
    local_controls_for_mapping = []
    if framework_company:
        from .models import LocalComplianceControl
        local_controls_for_mapping = LocalComplianceControl.objects.filter(
            company=framework_company
        ).select_related('requirement').order_by('code')[:150]
    
    # Примітки до контролю
    from .models import ControlNote
    notes = ControlNote.objects.filter(
        control=control,
        is_active=True
    ).select_related('created_by').order_by('-created_date')
    
    # Get user permissions
    permissions = get_user_compliance_permissions(request.user)
    
    # Determine which permission to check based on framework type
    is_instance = not control.category.framework.is_template
    can_edit = permissions['can_edit_instance_controls'] if is_instance else permissions['can_edit_controls']
    
    # Import model classes for template
    from .models import ControlMapping as ControlMappingModel
    
    # Get available mandatory processes for the company
    from app_compliance.models import MandatoryProcess
    from app_doc.models import RegisterDocs, RelatedDocs
    mandatory_processes = MandatoryProcess.objects.filter(
        is_active=True
    ).select_related('company', 'source_document').order_by('process_name')
    
    # Filter by company if framework has one
    if framework_company:
        mandatory_processes = mandatory_processes.filter(
            Q(company=framework_company) | Q(company__isnull=True)
        )
    
    # Get register documents
    register_docs = RegisterDocs.objects.filter(
        is_active=True
    ).exclude(
        file_doc=''
    ).select_related('status_doc', 'company').order_by('name_doc')
    if framework_company:
        register_docs = register_docs.filter(
            Q(company=framework_company) | Q(company__isnull=True)
        )
    
    # Get related documents
    related_docs = RelatedDocs.objects.exclude(
        file_rel_doc=''
    ).select_related('status_rel_doc', 'company').order_by('name_rel_doc')
    if framework_company:
        related_docs = related_docs.filter(
            Q(company=framework_company) | Q(company__isnull=True)
        )
    
    # Get language preferences for localization
    language_code, country_for_language, use_localized_labels = get_language_preferences(request)
    
    # Get active evidence types with localized names
    evidence_types_qs = EvidenceType.objects.filter(is_active=True).order_by('display_order', 'name')
    
    # Build localized evidence types list
    evidence_types = []
    if country_for_language and use_localized_labels:
        # Get translations for the country
        translations = {
            et.evidence_type.code: et.name_local
            for et in EvidenceTypeTranslation.objects.filter(
                country=country_for_language,
                evidence_type__in=evidence_types_qs
            ).select_related('evidence_type')
        }
        for et in evidence_types_qs:
            localized_name = translations.get(et.code) or (et.name_local if language_code == 'uk' else et.name)
            evidence_types.append({
                'id': et.id,
                'name': localized_name,
                'code': et.code,
                'color': et.color
            })
    else:
        # Use English names
        for et in evidence_types_qs:
            evidence_types.append({
                'id': et.id,
                'name': et.name,
                'code': et.code,
                'color': et.color
            })
    
    context = {
        'control': control,
        'evidences': control.evidences.filter(is_active=True).order_by('-uploaded_date'),
        'logs': logs,
        'mappings_to': mappings_to,
        'mappings_from': mappings_from,
        'notes': notes,
        'has_sufficient_evidence': control.has_sufficient_evidence(),
        'evidence_count': control.get_evidence_count(),
        'approved_evidence_count': control.get_approved_evidence_count(),
        'users': users,
        'all_controls': all_controls,
        'internal_controls_for_mapping': internal_controls_for_mapping,
        'local_controls_for_mapping': local_controls_for_mapping,
        'mandatory_processes': mandatory_processes,
        'register_docs': register_docs,
        'related_docs': related_docs,
        'permissions': permissions,
        'can_edit': can_edit,
        'is_instance': is_instance,
        'ControlMapping': ControlMappingModel,
        'evidence_types': evidence_types,
    }
    
    return render(request, 'app_compliance/control_detail.html', context)


@login_required
def control_create(request, category_id):
    """GET: показати форму створення, POST: створити контроль"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_add_controls'):
        messages.error(request, _('You do not have permission to add controls'))
        return redirect('compliance:dashboard')
    
    category = get_object_or_404(ControlCategory, id=category_id)
    
    if request.method == 'GET':
        # Отримуємо активні домени з FrameworkDomain
        domains = FrameworkDomain.objects.filter(is_active=True).order_by('display_order', 'name')
        context = {
            'category': category,
            'framework': category.framework,
            'domains': domains,
        }
        return render(request, 'app_compliance/control_create_form.html', context)
    
    try:
        parent_control_id = request.POST.get('parent_control')
        parent_control = None
        if parent_control_id:
            parent_control = Control.objects.get(id=parent_control_id)
        
        # Для template завжди встановлюємо status "not_started"
        if category.framework.is_template:
            status = 'not_started'
        else:
            status = request.POST.get('status', 'not_started')
        
        domain_id = request.POST.get('domain')
        domain_id = int(domain_id) if domain_id else None
        
        control = Control.objects.create(
            category=category,
            parent_control=parent_control,
            code=request.POST.get('code'),
            title=request.POST.get('title', ''),
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            domain_id=domain_id,
            status=status,
            priority=request.POST.get('priority', 'medium'),
            required_evidence_count=int(request.POST.get('required_evidence_count', 0)),
            evidence_description=request.POST.get('evidence_description', ''),
            implementation_guidance=request.POST.get('implementation_guidance', ''),
            testing_procedure=request.POST.get('testing_procedure', ''),
            order=int(request.POST.get('order', 0)),
            created_by=request.user
        )
        
        # Призначення відповідального
        responsible_id = request.POST.get('responsible')
        if responsible_id:
            control.responsible_id = responsible_id
            control.save()
        
        log_compliance_action(
            request.user, 'create', 'control', control,
            request=request
        )
        
        messages.success(request, _('Control created successfully'))
        next_url = request.POST.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('compliance:control_detail', control_id=control.id)
        
    except Exception as e:
        messages.error(request, f'Error creating control: {str(e)}')
        return redirect('compliance:framework_detail', framework_id=category.framework.id)


@login_required
@require_http_methods(["POST"])
def control_update(request, control_id):
    """Оновлення контролю"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to edit controls'))
        control = get_object_or_404(Control, id=control_id)
        return redirect('compliance:control_detail', control_id=control.id)
    
    control = get_object_or_404(Control, id=control_id)
    
    try:
        old_values = {
            'status': control.status,
            'priority': control.priority,
            'responsible': control.responsible.username if control.responsible else None,
        }
        
        control.code = request.POST.get('code', control.code)
        control.title = request.POST.get('title', control.title)  # ← ДОДАНО Title
        control.name = request.POST.get('name', control.name)
        control.description = request.POST.get('description', control.description)
        domain_id = request.POST.get('domain')
        if domain_id and domain_id != '' and domain_id != '0':
            from .models import FrameworkDomain
            try:
                domain_id_int = int(domain_id)
                if domain_id_int > 0:
                    control.domain = FrameworkDomain.objects.get(id=domain_id_int)
                else:
                    control.domain = None
            except (FrameworkDomain.DoesNotExist, ValueError, TypeError):
                control.domain = None
        else:
            control.domain = None  # ← ДОДАНО Domain
        
        # Оновлення статусу
        new_status = request.POST.get('status', control.status)
        if new_status != control.status:
            control.status = new_status
            control.status_changed_date = timezone.now()
        
        control.priority = request.POST.get('priority', control.priority)
        control.required_evidence_count = int(request.POST.get('required_evidence_count', control.required_evidence_count))
        control.evidence_description = request.POST.get('evidence_description', control.evidence_description)
        control.implementation_guidance = request.POST.get('implementation_guidance', control.implementation_guidance)
        control.testing_procedure = request.POST.get('testing_procedure', control.testing_procedure)
        control.order = int(request.POST.get('order', control.order))  # ← ДОДАНО Order
        
        # Оновлення відповідального
        responsible_id = request.POST.get('responsible')
        if responsible_id and responsible_id != '' and responsible_id != '0':
            try:
                responsible_id_int = int(responsible_id)
                if responsible_id_int > 0:
                    control.responsible_id = responsible_id_int
                else:
                    control.responsible_id = None
            except (ValueError, TypeError):
                control.responsible_id = None
        else:
            control.responsible_id = None
        
        # Оновлення parent_control
        parent_control_id = request.POST.get('parent_control')
        if parent_control_id and parent_control_id != '' and parent_control_id != '0':
            try:
                parent_control_id_int = int(parent_control_id)
                if parent_control_id_int > 0:
                    from .models import Control as ControlModel
                    try:
                        parent_control = ControlModel.objects.get(id=parent_control_id_int)
                        control.parent_control = parent_control
                    except ControlModel.DoesNotExist:
                        control.parent_control = None
                else:
                    control.parent_control = None
            except (ValueError, TypeError):
                control.parent_control = None
        else:
            control.parent_control = None
        
        # Якщо статус змінився на completed
        if control.status == 'completed' and old_values['status'] != 'completed':
            control.actual_completion_date = timezone.now().date()
        
        control.save()
        
        new_values = {
            'status': control.status,
            'priority': control.priority,
            'responsible': control.responsible.username if control.responsible else None,
        }
        
        # Логуємо загальне оновлення
        log_compliance_action(
            request.user, 'update', 'control', control,
            changes={'old': old_values, 'new': new_values},
            request=request
        )
        
        # Окреме логування зміни статусу
        if old_values['status'] != control.status:
            log_compliance_action(
                request.user, 'update', 'control', control,
                changes={
                    'field': 'status',
                    'old_status': old_values['status'],
                    'new_status': control.status,
                    'changed_date': control.status_changed_date.strftime('%Y-%m-%d %H:%M:%S') if control.status_changed_date else None
                },
                notes=f'Status changed from {dict(Control.STATUS_CHOICES).get(old_values["status"])} to {dict(Control.STATUS_CHOICES).get(control.status)}',
                request=request
            )
        
        messages.success(request, _('Control updated successfully'))
        
    except Exception as e:
        messages.error(request, f'Error updating control: {str(e)}')
    
    # Для Template redirect до framework, для Instance до control
    if control.category.framework.is_template:
        return redirect('compliance:framework_detail', framework_id=control.category.framework.id)
    else:
        return redirect('compliance:control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def control_delete(request, control_id):
    """Видалення контролю"""
    # Перевірка прав
    if not check_user_compliance_permission(request.user, 'can_delete_controls'):
        messages.error(request, _('You do not have permission to delete controls'))
        return redirect('compliance:dashboard')
    
    control = get_object_or_404(Control, id=control_id)
    framework_id = control.category.framework.id
    
    try:
        log_compliance_action(
            request.user, 'delete', 'control', control,
            request=request
        )
        
        control.delete()
        messages.success(request, _('Control deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting control: {str(e)}')
    
    return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
@require_http_methods(["POST"])
def control_verify(request, control_id):
    """Верифікація контролю"""
    control = get_object_or_404(Control, id=control_id)
    
    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to verify controls'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        control.is_verified = True
        control.verified_by = request.user
        control.verified_date = timezone.now()
        control.verification_notes = request.POST.get('verification_notes', '')
        
        # Оновлюємо статус і дату зміни статусу
        if control.status != 'completed':
            control.status = 'completed'
            control.status_changed_date = timezone.now()
        
        # Встановлюємо дату завершення
        if not control.actual_completion_date:
            control.actual_completion_date = timezone.now().date()
        
        control.save()
        
        log_compliance_action(
            request.user, 'verify', 'control', control,
            changes={'verified_date': str(control.verified_date)},
            request=request
        )
        
        # Логуємо зміну статусу при верифікації
        log_compliance_action(
            request.user, 'update', 'control', control,
            changes={
                'field': 'status',
                'old_status': 'ready_for_review',  # Припускаємо, що верифікуємо з ready_for_review
                'new_status': 'completed',
                'changed_date': control.status_changed_date.strftime('%Y-%m-%d %H:%M:%S') if control.status_changed_date else None
            },
            notes=f'Status changed to Completed via verification',
            request=request
        )
        
        messages.success(request, _('Control verified successfully'))
        
    except Exception as e:
        messages.error(request, f'Error verifying control: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def control_assign(request, control_id):
    """Призначення контролю користувачу"""
    control = get_object_or_404(Control, id=control_id)
    
    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to assign users'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        from django.contrib.auth.models import User
        
        user_id = request.POST.get('user_id')
        assignment_type = request.POST.get('assignment_type', 'owner')
        notes = request.POST.get('notes', '')
        
        user = User.objects.get(id=user_id)
        
        # Переконаємося що assignment_type валідний
        if assignment_type not in ['owner', 'reviewer', 'collaborator']:
            assignment_type = 'owner'
        
        assignment, created = ControlAssignment.objects.get_or_create(
            control=control,
            user=user,
            assignment_type=assignment_type,  # Має бути в lookup, не в defaults!
            defaults={
                'assigned_by': request.user,
                'notes': notes,
            }
        )
        
        if not created:
            assignment.notes = notes
            assignment.is_active = True
            assignment.save()
        
        # Якщо assignment_type == 'owner', оновлюємо поле responsible у Control
        if assignment_type == 'owner':
            control.responsible = user
            control.save()
            messages.success(request, _('User assigned as owner (responsible) for this control'))
        else:
            messages.success(request, _('User assigned as collaborator/reviewer'))
        
        log_compliance_action(
            request.user, 'assign', 'assignment', assignment,
            changes={'user': user.username, 'type': assignment_type},
            request=request
        )
        
    except Exception as e:
        messages.error(request, f'Error assigning control: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def control_set_responsible(request, control_id):
    """Встановлення відповідальної особи для контролю"""
    control = get_object_or_404(
        Control.objects.select_related('category__framework'),
        id=control_id
    )
    
    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to set responsible'))
        return redirect('compliance:control_detail', control_id=control.id)
    
    try:
        old_responsible = control.responsible.username if control.responsible else None
        
        responsible_id = request.POST.get('responsible_id')
        if responsible_id:
            from django.contrib.auth.models import User
            control.responsible = User.objects.get(id=responsible_id)
            new_responsible = control.responsible.username
        else:
            control.responsible = None
            new_responsible = None
        
        control.save()
        
        log_compliance_action(
            request.user,
            'update',
            'control',
            control,
            changes={'old': {'responsible': old_responsible}, 'new': {'responsible': new_responsible}},
            request=request
        )
        
        if new_responsible:
            messages.success(request, _('Responsible person set successfully'))
        else:
            messages.success(request, _('Responsible person cleared'))
            
    except Exception as e:
        messages.error(request, _('Error setting responsible: %(error)s') % {'error': str(e)})
    
    return redirect('compliance:control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def assignment_delete(request, assignment_id):
    """Видалення призначення"""
    assignment = get_object_or_404(ControlAssignment, id=assignment_id)
    control_id = assignment.control.id
    
    # Check permissions based on framework type
    is_instance = not assignment.control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to remove assignments'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        # Якщо це був owner, очищаємо поле responsible у Control
        if assignment.assignment_type == 'owner':
            control = assignment.control
            if control.responsible == assignment.user:
                control.responsible = None
                control.save()
        
        assignment.delete()
        messages.success(request, _('Assignment removed successfully'))
        
    except Exception as e:
        messages.error(request, f'Error removing assignment: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control_id)


# ========================
# Evidence Views
# ========================

@login_required
def evidence_list(request, control_id):
    """Список доказів для контролю"""
    control = get_object_or_404(Control, id=control_id)
    
    evidences = control.evidences.filter(is_active=True).select_related(
        'uploaded_by', 'reviewed_by'
    ).order_by('-uploaded_date')
    
    # Фільтр по типу
    evidence_type = request.GET.get('type', '')
    if evidence_type:
        evidences = evidences.filter(evidence_type=evidence_type)
    
    # Фільтр по статусу схвалення
    approval_status = request.GET.get('approval_status', '')
    if approval_status:
        evidences = evidences.filter(approval_status=approval_status)
    
    context = {
        'control': control,
        'evidences': evidences,
        'evidence_type': evidence_type,
        'approval_status': approval_status,
        'evidence_types': Evidence.EVIDENCE_TYPE_CHOICES,
        'approval_statuses': Evidence.APPROVAL_STATUS_CHOICES,
    }
    
    return render(request, 'app_compliance/evidence_list.html', context)


@login_required
@require_http_methods(["POST"])
def evidence_create(request, control_id):
    """Створення доказу"""
    control = get_object_or_404(Control, id=control_id)
    
    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_manage_evidence'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to add evidence'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        mandatory_process_id = request.POST.get('mandatory_process_id')
        mandatory_process = None
        if mandatory_process_id:
            from app_compliance.models import MandatoryProcess
            mandatory_process = MandatoryProcess.objects.filter(id=mandatory_process_id).first()
        
        # Get EvidenceType by ID or code, fallback to 'document' code
        evidence_type_id = request.POST.get('evidence_type', '').strip()
        evidence_type = None
        if evidence_type_id:
            try:
                # Try to get by ID first (convert to int if possible)
                try:
                    evidence_type_id_int = int(evidence_type_id)
                    evidence_type = EvidenceType.objects.get(id=evidence_type_id_int)
                except (ValueError, TypeError):
                    # If not a number, try by code
                    evidence_type = EvidenceType.objects.get(code=evidence_type_id)
            except EvidenceType.DoesNotExist:
                # Fallback to default 'document'
                evidence_type = EvidenceType.objects.filter(code='document').first()
        else:
            # Fallback to default 'document'
            evidence_type = EvidenceType.objects.filter(code='document').first()
        
        if not evidence_type:
            messages.error(request, _('Invalid evidence type selected'))
            return redirect('compliance:control_detail', control_id=control_id)
        
        evidence = Evidence.objects.create(
            control=control,
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            evidence_type=evidence_type,
            text_evidence=request.POST.get('text_evidence', ''),
            external_link=request.POST.get('external_link', ''),
            mandatory_process=mandatory_process,
            uploaded_by=request.user
        )
        
        # Handle file: priority: uploaded file > register_doc > related_doc
        if 'file' in request.FILES:
            evidence.file = request.FILES['file']
        else:
            # Check for register document
            register_doc_id = request.POST.get('register_doc_id')
            if register_doc_id:
                from app_doc.models import RegisterDocs
                register_doc = RegisterDocs.objects.filter(id=register_doc_id).first()
                if register_doc and register_doc.file_doc:
                    evidence.file = register_doc.file_doc
                    evidence.register_document = register_doc
            
            # Check for related document (not else - could have both, but priority to register)
            if not evidence.file:
                related_doc_id = request.POST.get('related_doc_id')
                if related_doc_id:
                    from app_doc.models import RelatedDocs
                    related_doc = RelatedDocs.objects.filter(id=related_doc_id).first()
                    if related_doc and related_doc.file_rel_doc:
                        evidence.file = related_doc.file_rel_doc
                        evidence.related_document = related_doc
        
        evidence.save()
        
        log_compliance_action(
            request.user, 'create', 'evidence', evidence,
            request=request
        )
        
        messages.success(request, _('Evidence uploaded successfully'))
        
    except Exception as e:
        messages.error(request, f'Error uploading evidence: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control_id)


@login_required
def evidence_edit(request, evidence_id):
    """GET: показати форму редагування evidence, POST: оновити evidence"""
    evidence = get_object_or_404(Evidence, id=evidence_id)
    control = evidence.control
    
    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_manage_evidence'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to edit evidence'))
        return redirect('compliance:control_detail', control_id=control.id)
    
    if request.method == 'GET':
        # Return evidence data as JSON for modal
        import json
        evidence_data = {
            'id': evidence.id,
            'title': evidence.title,
            'description': evidence.description,
            'evidence_type': evidence.evidence_type_id if evidence.evidence_type_id else None,
            'text_evidence': evidence.text_evidence,
            'external_link': evidence.external_link,
            'file_name': evidence.file.name if evidence.file else '',
            'mandatory_process_id': evidence.mandatory_process_id if evidence.mandatory_process else None,
        }
        return JsonResponse(evidence_data)
    
    # POST - update evidence
    return evidence_update(request, evidence_id)


@login_required
@require_http_methods(["POST"])
def evidence_update(request, evidence_id):
    """Оновлення доказу"""
    evidence = get_object_or_404(Evidence, id=evidence_id)
    
    # Check permissions based on framework type
    is_instance = not evidence.control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_manage_evidence'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to edit evidence'))
        return redirect('compliance:control_detail', control_id=evidence.control.id)
    
    try:
        evidence.title = request.POST.get('title', evidence.title)
        evidence.description = request.POST.get('description', evidence.description)
        evidence.evidence_type = request.POST.get('evidence_type', evidence.evidence_type)
        evidence.text_evidence = request.POST.get('text_evidence', evidence.text_evidence)
        evidence.external_link = request.POST.get('external_link', evidence.external_link)
        
        # Handle mandatory process update
        mandatory_process_id = request.POST.get('mandatory_process_id')
        if mandatory_process_id:
            from app_compliance.models import MandatoryProcess
            evidence.mandatory_process = MandatoryProcess.objects.filter(id=mandatory_process_id).first()
        else:
            evidence.mandatory_process = None
        
        # Оновлення файлу
        if 'file' in request.FILES:
            evidence.file = request.FILES['file']
        
        evidence.save()
        
        log_compliance_action(
            request.user, 'update', 'evidence', evidence,
            request=request
        )
        
        messages.success(request, _('Evidence updated successfully'))
        
    except Exception as e:
        messages.error(request, f'Error updating evidence: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=evidence.control.id)


@login_required
@require_http_methods(["POST"])
def evidence_delete(request, evidence_id):
    """Видалення доказу"""
    evidence = get_object_or_404(Evidence, id=evidence_id)
    control_id = evidence.control.id
    
    # Check permissions based on framework type
    is_instance = not evidence.control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_manage_evidence'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to delete evidence'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        evidence.is_active = False
        evidence.save()
        
        log_compliance_action(
            request.user, 'delete', 'evidence', evidence,
            request=request
        )
        
        messages.success(request, _('Evidence deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting evidence: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control_id)


@login_required
@require_http_methods(["POST"])
def evidence_approve(request, evidence_id):
    """Схвалення доказу"""
    evidence = get_object_or_404(Evidence, id=evidence_id)
    
    try:
        evidence.approval_status = 'approved'
        evidence.reviewed_by = request.user
        evidence.reviewed_date = timezone.now()
        evidence.review_comments = request.POST.get('review_comments', '')
        evidence.save()
        
        log_compliance_action(
            request.user, 'approve', 'evidence', evidence,
            changes={'approved_date': str(evidence.reviewed_date)},
            request=request
        )
        
        messages.success(request, _('Evidence approved successfully'))
        
    except Exception as e:
        messages.error(request, f'Error approving evidence: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=evidence.control.id)


@login_required
@require_http_methods(["POST"])
def evidence_reject(request, evidence_id):
    """Відхилення доказу"""
    evidence = get_object_or_404(Evidence, id=evidence_id)
    
    try:
        evidence.approval_status = 'rejected'
        evidence.reviewed_by = request.user
        evidence.reviewed_date = timezone.now()
        evidence.review_comments = request.POST.get('review_comments', '')
        evidence.save()
        
        log_compliance_action(
            request.user, 'reject', 'evidence', evidence,
            changes={'rejected_date': str(evidence.reviewed_date), 'reason': evidence.review_comments},
            request=request
        )
        
        messages.success(request, _('Evidence rejected'))
        
    except Exception as e:
        messages.error(request, f'Error rejecting evidence: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=evidence.control.id)


# ========================
# Statistics & Reports
# ========================

@login_required
@compliance_access_required
def framework_instances_list(request):
    """Список framework instances (для адміністрування)"""
    # Отримуємо доступні компанії для користувача
    accessible_companies = get_user_accessible_companies(request.user)
    accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
    
    # Show only instances для доступних компаній
    instances = ComplianceFramework.objects.filter(
        is_template=False,
        company_id__in=accessible_company_ids
    )
    
    instances = instances.select_related('company', 'template', 'created_by').annotate(
        total_categories=Count('categories', distinct=True),
        total_controls=Count('categories__controls', distinct=True)
    )
    
    # Пошук
    search = request.GET.get('search', '')
    if search:
        instances = instances.filter(
            Q(name__icontains=search) |
            Q(company__name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Фільтр по компанії (тільки з доступних)
    company_id = request.GET.get('company', '')
    if company_id and int(company_id) in accessible_company_ids:
        instances = instances.filter(company_id=company_id)
    
    # Фільтр по template
    template_id = request.GET.get('template', '')
    if template_id:
        instances = instances.filter(template_id=template_id)
    
    # Фільтр по статусу
    status = request.GET.get('status', '')
    if status:
        instances = instances.filter(status=status)
    
    # Сортування
    sort_by = request.GET.get('sort', '-created_date')
    instances = instances.order_by(sort_by)
    
    # Пагінація
    paginator = Paginator(instances, 20)
    page = request.GET.get('page')
    
    try:
        instances_page = paginator.page(page)
    except PageNotAnInteger:
        instances_page = paginator.page(1)
    except EmptyPage:
        instances_page = paginator.page(paginator.num_pages)
    
    # Додаємо статистику
    for instance in instances_page:
        instance.completion = instance.get_completion_percentage()
        instance.stats = instance.get_controls_by_status()
    
    # Список компаній (тільки доступних) та templates для фільтрів
    companies = accessible_companies.order_by('name')
    
    # Show only templates that have instances in accessible companies (or all for superuser/staff)
    if request.user.is_superuser or request.user.is_staff:
        templates = ComplianceFramework.objects.filter(is_template=True)
    else:
        # Only templates that have instances in accessible companies
        templates = ComplianceFramework.objects.filter(
            is_template=True,
            instances__company_id__in=accessible_company_ids
        ).distinct()
    
    # Отримуємо permissions для користувача
    permissions = get_user_compliance_permissions(request.user)
    
    context = {
        'instances': instances_page,
        'search': search,
        'companies': companies,
        'templates': templates,
        'selected_company': company_id,
        'selected_template': template_id,
        'selected_status': status,
        'sort_by': sort_by,
        'status_choices': ComplianceFramework.STATUS_CHOICES,
        'permissions': permissions,
    }
    
    return render(request, 'app_compliance/framework_instances_list.html', context)


@login_required
@compliance_access_required
def compliance_dashboard(request):
    """Dashboard зі статистикою"""
    from django.db.models import Q
    from datetime import date, timedelta
    from app_cabinet.models import Company
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    user_company = get_user_company(request.user)
    
    # Отримуємо вибрані компанії з GET параметрів
    selected_company_ids = request.GET.getlist('company')
    
    # Визначаємо доступні компанії для користувача через AccessCompliance
    available_companies = get_user_accessible_companies(request.user).order_by('name')
    
    # Фільтруємо frameworks за обраними компаніями
    # Спочатку обмежуємо доступними компаніями
    accessible_company_ids = list(available_companies.values_list('id', flat=True))
    
    if selected_company_ids:
        # Якщо обрано конкретні компанії, перевіряємо що вони в доступних
        valid_selected_ids = [
            int(cid) for cid in selected_company_ids 
            if int(cid) in accessible_company_ids
        ]
        if valid_selected_ids:
            frameworks = ComplianceFramework.objects.filter(
                company_id__in=valid_selected_ids, 
                status='active'
            )
            all_frameworks = ComplianceFramework.objects.filter(
                company_id__in=valid_selected_ids
            )
            selected_companies = available_companies.filter(id__in=valid_selected_ids)
        else:
            # Немає валідних обраних компаній
            frameworks = ComplianceFramework.objects.none()
            all_frameworks = ComplianceFramework.objects.none()
            selected_companies = Company.objects.none()
    else:
        # За замовчуванням показуємо всі доступні компанії
        if accessible_company_ids:
            frameworks = ComplianceFramework.objects.filter(
                company_id__in=accessible_company_ids, 
                status='active'
            )
            all_frameworks = ComplianceFramework.objects.filter(
                company_id__in=accessible_company_ids
            )
            selected_companies = available_companies
        else:
            frameworks = ComplianceFramework.objects.none()
            all_frameworks = ComplianceFramework.objects.none()
            selected_companies = Company.objects.none()
    
    # Загальна статистика
    total_frameworks = frameworks.count()
    total_controls = Control.objects.filter(category__framework__in=frameworks).count()
    completed_controls = Control.objects.filter(
        category__framework__in=frameworks,
        status='completed'
    ).count()
    in_progress_controls = Control.objects.filter(
        category__framework__in=frameworks,
        status='in_progress'
    ).count()
    not_started_controls = Control.objects.filter(
        category__framework__in=frameworks,
        status='not_started'
    ).count()
    
    # Відсоток виконання
    overall_completion = 0
    if total_controls > 0:
        overall_completion = round((completed_controls / total_controls) * 100, 2)
    
    # Статистика по пріоритетах
    critical_controls = Control.objects.filter(
        category__framework__in=frameworks,
        priority='critical'
    ).count()
    critical_completed = Control.objects.filter(
        category__framework__in=frameworks,
        priority='critical',
        status='completed'
    ).count()
    
    high_controls = Control.objects.filter(
        category__framework__in=frameworks,
        priority='high'
    ).count()
    high_completed = Control.objects.filter(
        category__framework__in=frameworks,
        priority='high',
        status='completed'
    ).count()
    
    # Overdue контролі (пройшов target_completion_date і не completed)
    today = date.today()
    overdue_controls = Control.objects.filter(
        category__framework__in=frameworks,
        target_completion_date__lt=today,
        status__in=['not_started', 'in_progress', 'ready_for_review']
    ).select_related('category__framework', 'responsible').order_by('target_completion_date')[:10]
    
    # Контролі що стають due в найближчі 7 днів
    upcoming_due = Control.objects.filter(
        category__framework__in=frameworks,
        target_completion_date__gte=today,
        target_completion_date__lte=today + timedelta(days=7),
        status__in=['not_started', 'in_progress']
    ).select_related('category__framework', 'responsible').order_by('target_completion_date')[:10]
    
    # Статистика по фреймворках
    framework_stats = []
    for fw in frameworks:
        framework_stats.append({
            'framework': fw,
            'completion': fw.get_completion_percentage(),
            'stats': fw.get_controls_by_status(),
            'company': fw.company.name if fw.company else None,
        })
    
    # Контролі, що потребують уваги
    controls_attention = Control.objects.filter(
        category__framework__in=frameworks,
        status__in=['not_started', 'in_progress']
    ).select_related('category__framework', 'responsible').order_by('target_completion_date')[:10]
    
    # Нещодавня активність (з фільтром по обраним компаніям)
    if frameworks.exists():
        # Отримуємо ID frameworks для фільтрації логів
        framework_ids = list(frameworks.values_list('id', flat=True))
        
        # Отримуємо ID контролів для фільтрації логів
        control_ids = list(Control.objects.filter(category__framework__in=frameworks).values_list('id', flat=True))
        
        # Отримуємо ID доказів для фільтрації логів
        evidence_ids = list(Evidence.objects.filter(control__category__framework__in=frameworks).values_list('id', flat=True))
        
        # Фільтруємо логи по обраним frameworks, controls та evidences
        recent_logs_queryset = ComplianceAuditLog.objects.filter(
            Q(object_type='framework', object_id__in=framework_ids) |
            Q(object_type='control', object_id__in=control_ids) |
            Q(object_type='evidence', object_id__in=evidence_ids) |
            Q(object_type='category')  # Категорії також можуть бути релевантними
        ).select_related('user').order_by('-timestamp')
    else:
        # Якщо немає frameworks, показуємо порожній список
        recent_logs_queryset = ComplianceAuditLog.objects.none()
    
    # Пагінація для Recent Activity (10 записів на сторінку)
    page = request.GET.get('page', 1)
    paginator = Paginator(recent_logs_queryset, 10)  # 10 записів на сторінку
    
    try:
        recent_logs = paginator.page(page)
    except PageNotAnInteger:
        recent_logs = paginator.page(1)
    except EmptyPage:
        recent_logs = paginator.page(paginator.num_pages)
    
    # Докази що очікують перегляду
    pending_evidences = Evidence.objects.filter(
        control__category__framework__in=frameworks,
        approval_status='pending',
        is_active=True
    ).select_related('control', 'uploaded_by').order_by('-uploaded_date')[:10]
    
    # Priority Matrix - статистика по пріоритетах та статусах
    priority_matrix = {}
    for priority_value, priority_label in Control.PRIORITY_CHOICES:
        priority_matrix[priority_value] = {
            'label': priority_label,
            'completed': Control.objects.filter(
                category__framework__in=frameworks,
                priority=priority_value,
                status='completed'
            ).count(),
            'in_progress': Control.objects.filter(
                category__framework__in=frameworks,
                priority=priority_value,
                status='in_progress'
            ).count(),
            'not_started': Control.objects.filter(
                category__framework__in=frameworks,
                priority=priority_value,
                status='not_started'
            ).count(),
            'failed': Control.objects.filter(
                category__framework__in=frameworks,
                priority=priority_value,
                status='failed'
            ).count(),
        }
        # Обчислюємо total для кожного priority
        priority_matrix[priority_value]['total'] = (
            priority_matrix[priority_value]['completed'] +
            priority_matrix[priority_value]['in_progress'] +
            priority_matrix[priority_value]['not_started'] +
            priority_matrix[priority_value]['failed']
        )
    
    # Compliance Score Card - зважена оцінка з урахуванням пріоритетів
    compliance_score = 0
    score_breakdown = {}
    
    if total_controls > 0:
        # Вага для кожного пріоритету (critical важливіше)
        priority_weights = {
            'critical': 4.0,
            'high': 3.0,
            'medium': 2.0,
            'low': 1.0,
        }
        
        total_weighted_score = 0
        total_weighted_controls = 0
        
        for priority_value, weight in priority_weights.items():
            if priority_value in priority_matrix:
                pm = priority_matrix[priority_value]
                if pm['total'] > 0:
                    # Completion для цього пріоритету
                    priority_completion = (pm['completed'] / pm['total']) * 100
                    score_breakdown[priority_value] = {
                        'completion': round(priority_completion, 1),
                        'weight': weight,
                        'total': pm['total']
                    }
                    # Додаємо до загального score
                    total_weighted_score += (pm['completed'] * weight)
                    total_weighted_controls += (pm['total'] * weight)
        
        if total_weighted_controls > 0:
            compliance_score = round((total_weighted_score / total_weighted_controls) * 100, 1)
    
    # Статус готовності до аудиту
    audit_readiness = 'not_ready'
    audit_readiness_label = _('Not Ready')
    
    if compliance_score >= 95:
        audit_readiness = 'excellent'
        audit_readiness_label = _('Excellent - Audit Ready')
    elif compliance_score >= 85:
        audit_readiness = 'good'
        audit_readiness_label = _('Good - Minor improvements needed')
    elif compliance_score >= 70:
        audit_readiness = 'fair'
        audit_readiness_label = _('Fair - Needs improvement')
    elif compliance_score >= 50:
        audit_readiness = 'poor'
        audit_readiness_label = _('Poor - Significant work required')
    else:
        audit_readiness = 'not_ready'
        audit_readiness_label = _('Not Ready - Major gaps')
    
    # Кількість доказів
    total_evidence = Evidence.objects.filter(
        control__category__framework__in=frameworks,
        is_active=True
    ).count()
    
    approved_evidence = Evidence.objects.filter(
        control__category__framework__in=frameworks,
        is_active=True,
        approval_status='approved'
    ).count()
    
    evidence_completion = round((approved_evidence / total_evidence * 100), 1) if total_evidence > 0 else 0
    
    # Team Workload Distribution - навантаження по користувачам
    from django.contrib.auth.models import User
    
    team_workload = []
    
    # Отримуємо всіх користувачів з призначеними контролями
    users_with_controls = User.objects.filter(
        assigned_controls__category__framework__in=frameworks
    ).distinct()
    
    for user in users_with_controls:
        user_controls = Control.objects.filter(
            category__framework__in=frameworks,
            responsible=user
        )
        
        total_assigned = user_controls.count()
        if total_assigned > 0:
            completed = user_controls.filter(status='completed').count()
            in_progress = user_controls.filter(status='in_progress').count()
            not_started = user_controls.filter(status='not_started').count()
            failed = user_controls.filter(status='failed').count()
            
            completion_pct = round((completed / total_assigned) * 100, 1)
            
            team_workload.append({
                'user': user,
                'total': total_assigned,
                'completed': completed,
                'in_progress': in_progress,
                'not_started': not_started,
                'failed': failed,
                'completion': completion_pct,
            })
    
    # Додаємо непризначені контролі
    unassigned_controls = Control.objects.filter(
        category__framework__in=frameworks,
        responsible__isnull=True
    )
    unassigned_count = unassigned_controls.count()
    
    if unassigned_count > 0:
        team_workload.append({
            'user': None,
            'total': unassigned_count,
            'completed': unassigned_controls.filter(status='completed').count(),
            'in_progress': unassigned_controls.filter(status='in_progress').count(),
            'not_started': unassigned_controls.filter(status='not_started').count(),
            'failed': unassigned_controls.filter(status='failed').count(),
            'completion': round((unassigned_controls.filter(status='completed').count() / unassigned_count) * 100, 1),
        })
    
    # Сортуємо по completion (найнижчі першими - потребують уваги)
    team_workload = sorted(team_workload, key=lambda x: x['completion'])
    
    # Статистика по доменам (для PCI DSS)
    from .models import ControlNote
    from django.db.models import Count, Q
    
    domain_stats = []
    if frameworks.exists():
        # Отримуємо унікальні домени з агрегованою статистикою
        domains_data = Control.objects.filter(
            category__framework__in=frameworks
        ).exclude(domain__isnull=True).values('domain__code', 'domain__name').annotate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed'))
        ).order_by('domain__code')
        
        for domain_data in domains_data:
            domain_total = domain_data['total']
            domain_completed = domain_data['completed']
            domain_completion = round((domain_completed / domain_total * 100), 1) if domain_total > 0 else 0
            
            domain_stats.append({
                'name': domain_data['domain__name'],  # Use domain__name from values()
                'domain': domain_data['domain__code'],  # Use domain__code from values()
                'total': domain_total,
                'completed': domain_completed,
                'completion': domain_completion
            })
    
    # Сортуємо domain_stats за completion (найнижчі першими)
    domain_stats = sorted(domain_stats, key=lambda x: x['completion'])
    
    # Якщо нічого не вибрано, за замовчуванням всі доступні компанії вибрані
    if not selected_company_ids:
        final_selected_ids = accessible_company_ids
    else:
        final_selected_ids = [int(id) for id in selected_company_ids]
    
    context = {
        'user_company': user_company,
        'available_companies': available_companies,
        'selected_companies': selected_companies,
        'selected_company_ids': final_selected_ids,
        'total_frameworks': total_frameworks,
        'total_controls': total_controls,
        'completed_controls': completed_controls,
        'in_progress_controls': in_progress_controls,
        'not_started_controls': not_started_controls,
        'overall_completion': overall_completion,
        'critical_controls': critical_controls,
        'critical_completed': critical_completed,
        'high_controls': high_controls,
        'high_completed': high_completed,
        'overdue_controls': overdue_controls,
        'upcoming_due': upcoming_due,
        'framework_stats': framework_stats,
        'controls_attention': controls_attention,
        'recent_logs': recent_logs,
        'pending_evidences': pending_evidences,
        'domain_stats': domain_stats[:5],  # Top 5 domains що потребують уваги
        'priority_matrix': priority_matrix,  # Priority Matrix для Dashboard
        'compliance_score': compliance_score,  # Загальна зважена оцінка
        'score_breakdown': score_breakdown,  # Деталізація оцінки по пріоритетах
        'audit_readiness': audit_readiness,  # Статус готовності
        'audit_readiness_label': audit_readiness_label,  # Текст статусу
        'total_evidence': total_evidence,  # Всього доказів
        'approved_evidence': approved_evidence,  # Затверджених доказів
        'evidence_completion': evidence_completion,  # % затверджених доказів
        'team_workload': team_workload[:10],  # Top 10 користувачів з найнижчим completion
    }
    
    return render(request, 'app_compliance/framework_dashboard.html', context)


@login_required
@control_mapping_access_required
def control_mapping_view(request):
    """Перегляд мапінгу контролів з фільтрацією по Company"""
    # Get selected company from query params
    company_id = request.GET.get('company')
    selected_company = None
    
    # Get accessible companies based on AccessControlMapping
    accessible_companies = get_user_accessible_companies_for_control_mapping(request.user)
    
    if company_id:
        try:
            selected_company = accessible_companies.get(id=int(company_id))
        except (Company.DoesNotExist, ValueError):
            selected_company = None
    
    # Build query for all types of mappings
    framework_mappings = []
    internal_mappings = []
    local_mappings = []
    
    if selected_company:
        # Framework Control mappings для цієї компанії
        framework_mappings = ControlMapping.objects.filter(
            source_control__category__framework__company=selected_company
        ).select_related(
            'source_control__category__framework',
            'target_control__category__framework',
            'target_internal_control__requirement',
            'target_local_control__requirement',
            'created_by'
        ).order_by('-created_date')
        
        # Internal Control mappings для цієї компанії
        internal_mappings = InternalControlMapping.objects.filter(
            Q(internal_control__company=selected_company) | 
            Q(internal_control__requirement__company=selected_company)
        ).select_related(
            'internal_control__requirement',
            'target_internal_control__requirement',
            'target_local_control__requirement',
            'target_framework_control__category__framework',
            'created_by'
        ).order_by('-created_date')
        
        # Local Control mappings для цієї компанії
        local_mappings = LocalControlMapping.objects.filter(
            local_control__company=selected_company
        ).select_related(
            'local_control__requirement',
            'target_local_control__requirement',
            'target_internal_control__requirement',
            'target_framework_control__category__framework',
            'created_by'
        ).order_by('-created_date')
    
    # Get user permissions
    permissions = get_user_compliance_permissions(request.user)
    can_create_mapping = permissions['can_edit_instance_controls'] or permissions['can_edit_controls']
    
    context = {
        'framework_mappings': framework_mappings,
        'internal_mappings': internal_mappings,
        'local_mappings': local_mappings,
        'companies': accessible_companies.order_by('name'),
        'selected_company': selected_company,
        'permissions': permissions,
        'can_create_mapping': can_create_mapping,
    }
    
    return render(request, 'app_compliance/control_mapping.html', context)


@login_required
@require_http_methods(["POST"])
def control_mapping_create(request):
    """Створення мапінгу між контролями"""
    try:
        source_control_id = request.POST.get('source_control_id')
        target_control_id = request.POST.get('target_control_id')
        mapping_type = request.POST.get('mapping_type', 'related')
        notes = request.POST.get('notes', '')
        
        source_control = Control.objects.get(id=source_control_id)
        target_control = Control.objects.get(id=target_control_id)
        
        # Check permissions for both controls
        permissions = get_user_compliance_permissions(request.user)
        source_is_instance = not source_control.category.framework.is_template
        target_is_instance = not target_control.category.framework.is_template
        
        # Check source control permission
        source_required = 'can_edit_instance_controls' if source_is_instance else 'can_edit_controls'
        if not check_user_compliance_permission(request.user, source_required):
            messages.error(request, _('You do not have permission to create mappings for source control'))
            return redirect('compliance:control_detail', control_id=source_control_id)
        
        # Check target control permission
        target_required = 'can_edit_instance_controls' if target_is_instance else 'can_edit_controls'
        if not check_user_compliance_permission(request.user, target_required):
            messages.error(request, _('You do not have permission to create mappings for target control'))
            return redirect('compliance:control_detail', control_id=source_control_id)
        
        mapping, created = ControlMapping.objects.get_or_create(
            source_control=source_control,
            target_control=target_control,
            defaults={
                'mapping_type': mapping_type,
                'notes': notes,
                'created_by': request.user,
            }
        )
        
        if created:
            messages.success(request, _('Control mapping created successfully'))
        else:
            messages.info(request, _('This mapping already exists'))
        
        # Redirect back to source control detail
        return redirect('compliance:control_detail', control_id=source_control_id)
        
    except Exception as e:
        messages.error(request, f'Error creating mapping: {str(e)}')
        # If we have source_control_id, redirect there; otherwise to control_mapping
        if source_control_id:
            return redirect('compliance:control_detail', control_id=source_control_id)
        return redirect('compliance:control_mapping')


# ========================
# AJAX Views
# ========================

@login_required
def get_framework_stats(request, framework_id):
    """AJAX: отримання статистики фреймворку"""
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    data = {
        'completion': framework.get_completion_percentage(),
        'stats': framework.get_controls_by_status(),
    }
    
    return JsonResponse(data)


@login_required
def get_control_evidences(request, control_id):
    """AJAX: отримання списку доказів контролю"""
    control = get_object_or_404(Control, id=control_id)
    
    evidences = control.evidences.filter(is_active=True).values(
        'id', 'title', 'evidence_type', 'approval_status',
        'uploaded_date', 'uploaded_by__username'
    )
    
    return JsonResponse(list(evidences), safe=False)


@login_required
def search_controls(request):
    """AJAX: пошук контролів"""
    query = request.GET.get('q', '')
    framework_id = request.GET.get('framework_id', '')
    
    controls = Control.objects.select_related('category__framework')
    
    if framework_id:
        controls = controls.filter(category__framework_id=framework_id)
    
    if query:
        controls = controls.filter(
            Q(code__icontains=query) |
            Q(name__icontains=query) |
            Q(description__icontains=query)
        )
    
    controls = controls[:20]
    
    data = [{
        'id': c.id,
        'code': c.code,
        'name': c.name,
        'framework': c.category.framework.name,
        'category': c.category.name,
        'status': c.status,
    } for c in controls]
    
    return JsonResponse(data, safe=False)


# ========================
# Export Views
# ========================

@login_required
def export_framework_excel(request, framework_id):
    """Експорт фреймворку в Excel"""
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        from .reports import export_framework_to_excel
        
        output = export_framework_to_excel(framework)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{framework.name}_{framework.version}_export_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError as e:
        messages.error(request, f'Export failed: {str(e)}. Please install required libraries.')
        return redirect('compliance:framework_detail', framework_id=framework_id)
    except Exception as e:
        messages.error(request, f'Error exporting to Excel: {str(e)}')
        return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
def export_framework_pdf(request, framework_id):
    """Експорт фреймворку в PDF"""
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        from .reports import export_framework_to_pdf
        
        buffer = export_framework_to_pdf(framework)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f"{framework.name}_{framework.version}_report_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError as e:
        messages.error(request, f'Export failed: {str(e)}. Please install required libraries.')
        return redirect('compliance:framework_detail', framework_id=framework_id)
    except Exception as e:
        messages.error(request, f'Error exporting to PDF: {str(e)}')
        return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
def export_control_pdf(request, control_id):
    """Експорт контролю в PDF"""
    control = get_object_or_404(Control, id=control_id)
    
    try:
        from .reports import export_control_details_to_pdf
        
        buffer = export_control_details_to_pdf(control)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        filename = f"Control_{control.code}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError as e:
        messages.error(request, f'Export failed: {str(e)}')
        return redirect('compliance:control_detail', control_id=control_id)
    except Exception as e:
        messages.error(request, f'Error exporting to PDF: {str(e)}')
        return redirect('compliance:control_detail', control_id=control_id)


# ========================
# Control Notes Views
# ========================

@login_required
@require_http_methods(["POST"])
def note_create(request, control_id):
    """Створення примітки до контролю"""
    from .models import ControlNote
    
    control = get_object_or_404(Control, id=control_id)
    
    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to add notes'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        note_text = request.POST.get('note', '').strip()
        
        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:control_detail', control_id=control.id)
        
        note = ControlNote.objects.create(
            control=control,
            note=note_text,
            attachment=request.FILES.get('attachment'),
            created_by=request.user
        )

        # Multiple attachments
        from .models import ControlNoteAttachment
        for f in request.FILES.getlist('attachments'):
            ControlNoteAttachment.objects.create(note=note, file=f)
        
        # Логування
        log_compliance_action(
            request.user, 'create', 'note', note,
            changes={'control_code': control.code, 'action': 'note_added'},
            request=request
        )
        
        messages.success(request, _('Note added successfully'))
        
    except Exception as e:
        messages.error(request, f'Error creating note: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def note_delete(request, note_id):
    """Видалення примітки"""
    from .models import ControlNote
    
    note = get_object_or_404(ControlNote, id=note_id)
    control_id = note.control.id
    
    # Check permissions based on framework type
    is_instance = not note.control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'
    
    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to delete notes'))
        return redirect('compliance:control_detail', control_id=control_id)
    
    try:
        # Soft delete
        note.is_active = False
        note.save()
        
        # Логування
        log_compliance_action(
            request.user, 'delete', 'note', note,
            notes=f'Deleted note from control {note.control.code}',
            request=request
        )
        
        messages.success(request, _('Note deleted successfully'))
        
    except Exception as e:
        messages.error(request, f'Error deleting note: {str(e)}')
    
    return redirect('compliance:control_detail', control_id=control_id)


@login_required
@require_http_methods(["POST"])
def note_update(request, note_id):
    """Редагування примітки до контролю"""
    from .models import ControlNote, ControlNoteAttachment

    note = get_object_or_404(ControlNote, id=note_id, is_active=True)
    control = note.control

    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to edit notes'))
        return redirect('compliance:control_detail', control_id=control.id)

    try:
        note_text = (request.POST.get('note') or '').strip()

        if not note_text:
            messages.error(request, _('Note text is required'))
            return redirect('compliance:control_detail', control_id=control.id)

        note.note = note_text

        # Clear existing single attachment if requested
        if request.POST.get('clear_attachment') == '1' and note.attachment:
            note.attachment = None

        # Optional single attachment (kept for compatibility)
        if 'attachment' in request.FILES:
            note.attachment = request.FILES['attachment']

        note.save()

        # Append new attachments if provided
        for f in request.FILES.getlist('attachments'):
            ControlNoteAttachment.objects.create(note=note, file=f)

        # Logging
        log_compliance_action(
            request.user, 'update', 'note', note,
            changes={'control_code': control.code, 'action': 'note_updated'},
            request=request
        )

        messages.success(request, _('Note updated successfully'))

    except Exception as e:
        messages.error(request, f'Error updating note: {str(e)}')

    return redirect('compliance:control_detail', control_id=control.id)


@login_required
@require_http_methods(["POST"])
def control_note_attachment_delete(request, attachment_id):
    """Delete single file attachment from control note"""
    from .models import ControlNoteAttachment

    attachment = get_object_or_404(ControlNoteAttachment, id=attachment_id)
    note = attachment.note
    control = note.control

    # Check permissions based on framework type
    is_instance = not control.category.framework.is_template
    required_permission = 'can_edit_instance_controls' if is_instance else 'can_edit_controls'

    if not check_user_compliance_permission(request.user, required_permission):
        messages.error(request, _('You do not have permission to delete attachments'))
        return redirect('compliance:control_detail', control_id=control.id)

    try:
        attachment.delete()
        messages.success(request, _('Attachment deleted'))
    except Exception as e:
        messages.error(request, f'Error deleting attachment: {str(e)}')

    return redirect('compliance:control_detail', control_id=control.id)


# ========================
# Framework Export/Import
# ========================

@login_required
@compliance_access_required
def framework_export_excel(request, framework_id):
    """Export framework to Excel with styles and colors"""
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    # Check permissions
    if not check_user_compliance_permission(request.user, 'can_export'):
        messages.error(request, _('You do not have permission to export frameworks'))
        if framework.is_template:
            return redirect('compliance:framework_list')
        else:
            return redirect('compliance:framework_instances_list')
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    category_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    category_font = Font(bold=True, size=10)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Status colors
    status_colors = {
        'completed': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
        'in_progress': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
        'not_started': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light red
        'ready_for_review': PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),  # Light blue
        'failed': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),  # Red
        'not_applicable': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # Gray
    }
    
    priority_colors = {
        'critical': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
        'high': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # Orange
        'medium': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow
        'low': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
    }
    
    # Sheet 1: Framework Overview
    ws_overview = wb.create_sheet("Framework Overview")
    
    ws_overview['A1'] = "Framework Export"
    ws_overview['A1'].font = Font(bold=True, size=14, color="0066CC")
    ws_overview.merge_cells('A1:D1')
    
    # Framework details
    overview_data = [
        ["Framework Name:", framework.name],
        ["Version:", framework.version],
        ["Type:", framework.get_framework_type_display()],
        ["Status:", framework.get_status_display()],
        ["Company:", framework.company.name if framework.company else "Template"],
        ["Is Template:", "Yes" if framework.is_template else "No"],
        ["Is Mandatory:", "Yes" if framework.is_mandatory else "No"],
        ["Description:", framework.description],
        ["Created Date:", framework.created_date.strftime("%d.%m.%Y")],
        ["Total Categories:", framework.categories.count()],
        ["Total Controls:", Control.objects.filter(category__framework=framework).count()],
        ["Export Date:", datetime.now().strftime("%d.%m.%Y %H:%M:%S")],
        ["Exported By:", request.user.get_full_name() or request.user.username],
    ]
    
    for idx, (label, value) in enumerate(overview_data, start=3):
        ws_overview[f'A{idx}'] = label
        ws_overview[f'A{idx}'].font = Font(bold=True)
        ws_overview[f'B{idx}'] = str(value)
        ws_overview.merge_cells(f'B{idx}:D{idx}')
    
    # Adjust column widths
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 50
    
    # Sheet 2: Controls List
    ws_controls = wb.create_sheet("Controls")
    
    # Headers
    headers = [
        "Category Code", "Category Name", "Control Code", "Control Name",
        "Domain Code", "Status", "Priority", "Responsible", "Target Date", "Actual Date",
        "Description", "Implementation Guidance", "Testing Procedure",
        "Evidence Count", "Required Evidence"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_controls.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border
    
    # Get all controls
    categories = framework.categories.prefetch_related('controls').order_by('order', 'code')
    
    row_idx = 2
    for category in categories:
        controls = category.controls.select_related('responsible', 'domain').order_by('order', 'code')
        
        for control in controls:
            # Data
            ws_controls.cell(row=row_idx, column=1, value=category.code)
            ws_controls.cell(row=row_idx, column=2, value=category.name)
            ws_controls.cell(row=row_idx, column=3, value=control.code)
            ws_controls.cell(row=row_idx, column=4, value=control.name)
            ws_controls.cell(row=row_idx, column=5, value=control.domain.code if control.domain else "")
            ws_controls.cell(row=row_idx, column=6, value=control.get_status_display())
            ws_controls.cell(row=row_idx, column=7, value=control.get_priority_display())
            ws_controls.cell(row=row_idx, column=8, value=control.responsible.get_full_name() if control.responsible else "")
            ws_controls.cell(row=row_idx, column=9, value=control.target_completion_date.strftime("%d.%m.%Y") if control.target_completion_date else "")
            ws_controls.cell(row=row_idx, column=10, value=control.actual_completion_date.strftime("%d.%m.%Y") if control.actual_completion_date else "")
            ws_controls.cell(row=row_idx, column=11, value=control.description)
            ws_controls.cell(row=row_idx, column=12, value=control.implementation_guidance)
            ws_controls.cell(row=row_idx, column=13, value=control.testing_procedure)
            ws_controls.cell(row=row_idx, column=14, value=control.get_evidence_count())
            ws_controls.cell(row=row_idx, column=15, value=control.required_evidence_count)
            
            # Apply status color
            if control.status in status_colors:
                for col in range(1, 16):
                    ws_controls.cell(row=row_idx, column=col).fill = status_colors[control.status]
            
            # Apply priority color to priority column
            if control.priority in priority_colors:
                ws_controls.cell(row=row_idx, column=7).fill = priority_colors[control.priority]
                ws_controls.cell(row=row_idx, column=7).font = Font(bold=True, color="FFFFFF")
            
            # Apply borders and alignment
            for col in range(1, 16):
                ws_controls.cell(row=row_idx, column=col).border = border
                if col <= 10:
                    ws_controls.cell(row=row_idx, column=col).alignment = alignment_center
                else:
                    ws_controls.cell(row=row_idx, column=col).alignment = alignment_left
            
            row_idx += 1
    
    # Adjust column widths
    column_widths = [15, 30, 15, 40, 20, 15, 12, 20, 12, 12, 50, 50, 50, 12, 12]
    for idx, width in enumerate(column_widths, start=1):
        ws_controls.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze first row
    ws_controls.freeze_panes = 'A2'
    
    # Sheet 3: Statistics
    ws_stats = wb.create_sheet("Statistics")
    
    ws_stats['A1'] = "Framework Statistics"
    ws_stats['A1'].font = Font(bold=True, size=14, color="0066CC")
    ws_stats.merge_cells('A1:C1')
    
    stats = framework.get_controls_by_status()
    
    stats_data = [
        ["", "Count", "Percentage"],
        ["Total Controls", stats['total'], "100%"],
        ["Completed", stats['completed'], f"{round(stats['completed']/stats['total']*100, 1) if stats['total'] > 0 else 0}%"],
        ["In Progress", stats['in_progress'], f"{round(stats['in_progress']/stats['total']*100, 1) if stats['total'] > 0 else 0}%"],
        ["Not Started", stats['not_started'], f"{round(stats['not_started']/stats['total']*100, 1) if stats['total'] > 0 else 0}%"],
        ["Failed", stats['failed'], f"{round(stats['failed']/stats['total']*100, 1) if stats['total'] > 0 else 0}%"],
    ]
    
    for idx, (label, count, pct) in enumerate(stats_data, start=3):
        ws_stats[f'A{idx}'] = label
        ws_stats[f'B{idx}'] = count
        ws_stats[f'C{idx}'] = pct
        
        if idx == 3:  # Header row
            for col in ['A', 'B', 'C']:
                ws_stats[f'{col}{idx}'].fill = header_fill
                ws_stats[f'{col}{idx}'].font = header_font
                ws_stats[f'{col}{idx}'].alignment = alignment_center
    
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 15
    ws_stats.column_dimensions['C'].width = 15
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"Framework_{framework.name.replace(' ', '_')}_{framework.version}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    # Log export action
    log_compliance_action(
        request.user, 'update', 'framework', framework,
        changes={'action': 'exported_to_excel'},
        request=request
    )
    
    return response


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def framework_import_excel(request):
    """Import framework or instance from Excel file.
    If POST contains as_instance=1 and company id, creates an instance framework for that company.
    Otherwise creates a Template framework (draft)."""
    if not check_user_compliance_permission(request.user, 'can_add_frameworks'):
        messages.error(request, _('You do not have permission to import frameworks'))
        # Redirect back to appropriate list
        return redirect(request.META.get('HTTP_REFERER') or 'compliance:framework_list')
    
    if 'file' not in request.FILES:
        messages.error(request, _('No file uploaded'))
        return redirect(request.META.get('HTTP_REFERER') or 'compliance:framework_list')
    
    excel_file = request.FILES['file']
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        # Read from "Controls" sheet
        if "Controls" not in wb.sheetnames:
            messages.error(request, _('Invalid Excel file: "Controls" sheet not found'))
            return redirect('compliance:framework_list')
        
        ws_controls = wb["Controls"]
        
        # Parse framework info from Overview sheet if exists
        framework_name = "Imported Framework"
        framework_version = "1.0"
        framework_type = "other"
        
        if "Framework Overview" in wb.sheetnames:
            ws_overview = wb["Framework Overview"]
            for row in ws_overview.iter_rows(min_row=3, max_row=10, values_only=True):
                if row[0] == "Framework Name:":
                    framework_name = row[1]
                elif row[0] == "Version:":
                    framework_version = row[1]
        
        with transaction.atomic():
            # Determine target type (template vs instance)
            as_instance = request.POST.get('as_instance') == '1'
            company_id = request.POST.get('company')

            framework_kwargs = dict(
                name=framework_name,
                version=framework_version,
                framework_type=framework_type,
                status='draft',
                created_by=request.user
            )

            if as_instance:
                try:
                    company = Company.objects.get(id=company_id)
                except Company.DoesNotExist:
                    messages.error(request, _('Selected company not found'))
                    return redirect('compliance:framework_instances_list')
                framework_kwargs.update(dict(is_template=False, company=company))
            else:
                framework_kwargs.update(dict(is_template=True))

            # Create framework (template or instance)
            framework = ComplianceFramework.objects.create(**framework_kwargs)
            
            # Parse controls
            categories_dict = {}
            imported_count = 0
            
            for row in ws_controls.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                category_code = row[0]
                category_name = row[1]
                control_code = row[2]
                control_name = row[3]
                domain_code = row[4] if len(row) > 4 and row[4] else None
                status = row[5] if len(row) > 5 else row[4]  # Support old format without domain
                priority = row[6] if len(row) > 6 else row[5]  # Support old format without domain
                description = row[10] if len(row) > 10 else (row[9] if len(row) > 9 else "")
                implementation_guidance = row[11] if len(row) > 11 else (row[10] if len(row) > 10 else "")
                testing_procedure = row[12] if len(row) > 12 else (row[11] if len(row) > 11 else "")
                
                # Create or get category
                if category_code not in categories_dict:
                    category = ControlCategory.objects.create(
                        framework=framework,
                        code=category_code,
                        name=category_name,
                        order=len(categories_dict)
                    )
                    categories_dict[category_code] = category
                else:
                    category = categories_dict[category_code]
                
                # Map status and priority
                status_mapping = {
                    'Not Started': 'not_started',
                    'In Progress': 'in_progress',
                    'Completed': 'completed',
                    'Ready for Review': 'ready_for_review',
                    'Failed': 'failed',
                    'Not Applicable': 'not_applicable',
                }
                
                priority_mapping = {
                    'Low': 'low',
                    'Medium': 'medium',
                    'High': 'high',
                    'Critical': 'critical',
                }
                
                control_status = status_mapping.get(status, 'not_started')
                control_priority = priority_mapping.get(priority, 'medium')
                
                # Find domain by code if provided
                domain = None
                if domain_code:
                    try:
                        domain = FrameworkDomain.objects.get(code=domain_code, is_active=True)
                    except FrameworkDomain.DoesNotExist:
                        pass  # Domain not found, leave as None
                
                # Create control
                Control.objects.create(
                    category=category,
                    code=control_code,
                    name=control_name,
                    domain=domain,
                    status=control_status,
                    priority=control_priority,
                    description=description or "",
                    implementation_guidance=implementation_guidance or "",
                    testing_procedure=testing_procedure or "",
                    created_by=request.user
                )
                
                imported_count += 1
            
            # Log import
            log_compliance_action(
                request.user, 'create', 'framework', framework,
                changes={'action': 'imported_from_excel', 'as_instance': as_instance, 'controls_count': imported_count, 'categories_count': len(categories_dict)},
                request=request
            )

            if as_instance:
                messages.success(
                    request,
                    _(f'Framework instance "{framework.name}" for {framework.company.name} imported successfully with {imported_count} controls in {len(categories_dict)} categories')
                )
                return redirect('compliance:framework_instances_list')
            else:
                messages.success(
                    request,
                    _(f'Template framework "{framework.name}" imported successfully with {imported_count} controls in {len(categories_dict)} categories')
                )
                return redirect('compliance:framework_detail', framework_id=framework.id)
    
    except Exception as e:
        messages.error(request, _(f'Error importing framework: {str(e)}'))
        return redirect(request.META.get('HTTP_REFERER') or 'compliance:framework_list')


@login_required
@compliance_access_required
def framework_excel_template(request):
    """Download Excel template for framework import"""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Framework Overview sheet
    ws_overview = wb.create_sheet("Framework Overview")
    ws_overview.append(["Framework Package Template"])
    ws_overview.append([])
    ws_overview.append(["Framework Name:", "Example Framework"])
    ws_overview.append(["Version:", "1.0"])
    ws_overview.append(["Description:", "Framework description"])
    ws_overview.append([])
    ws_overview.append(["Instructions:"])
    ws_overview.append(["1. Fill in the 'Controls' sheet with your framework data"])
    ws_overview.append(["2. Use the provided status and priority values"])
    ws_overview.append(["3. Save and import through the Import Excel function"])
    
    # Style overview sheet
    ws_overview['A1'].font = Font(size=16, bold=True, color="0066CC")
    for row in range(3, 6):
        ws_overview[f'A{row}'].font = Font(bold=True)
    ws_overview['A7'].font = Font(bold=True, size=12)
    
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 50
    
    # Create Controls sheet with headers and example
    ws_controls = wb.create_sheet("Controls")
    
    headers = [
        "Category Code", "Category Name", "Control Code", "Control Name",
        "Domain Code", "Status", "Priority", "Responsible", "Target Date", "Actual Date",
        "Description", "Implementation Guidance", "Testing Procedure",
        "Evidence Count", "Required Evidence"
    ]
    
    ws_controls.append(headers)
    
    # Add example row
    example_row = [
        "C.5", "Organizational Controls", "C.5.1", "Policies for information security",
        "SECURITY_PRIVACY_GOVERNANCE", "Not Started", "High", "", "",
        "", "Define and document information security policies",
        "Ensure policies cover all aspects of information security",
        "Review policy documents and approval records",
        "0", ""
    ]
    ws_controls.append(example_row)
    
    # Add second example
    example_row2 = [
        "C.5", "Organizational Controls", "C.5.2", "Information security roles",
        "SECURITY_PRIVACY_GOVERNANCE", "In Progress", "Critical", "", "",
        "", "Define roles and responsibilities for information security",
        "Document roles in organizational structure",
        "Review role definitions and assignments",
        "0", ""
    ]
    ws_controls.append(example_row2)
    
    # Style headers
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_controls.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border
    
    # Style example rows
    status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Not Started
    ws_controls.cell(row=2, column=5).fill = status_fill
    
    priority_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # High
    ws_controls.cell(row=2, column=6).fill = priority_fill
    ws_controls.cell(row=2, column=6).font = Font(color="FFFFFF")
    
    status_fill2 = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # In Progress
    ws_controls.cell(row=3, column=5).fill = status_fill2
    
    priority_fill2 = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Critical
    ws_controls.cell(row=3, column=6).fill = priority_fill2
    ws_controls.cell(row=3, column=6).font = Font(color="FFFFFF")
    
    # Add borders to example rows
    for row in range(2, 4):
        for col in range(1, len(headers) + 1):
            ws_controls.cell(row=row, column=col).border = border
    
    # Set column widths
    column_widths = [15, 30, 15, 40, 15, 12, 20, 12, 12, 50, 50, 50, 12, 12]
    for idx, width in enumerate(column_widths, 1):
        ws_controls.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Freeze header row
    ws_controls.freeze_panes = 'A2'
    
    # Create Statistics sheet with reference info
    ws_stats = wb.create_sheet("Statistics")
    ws_stats.append(["Reference Information"])
    ws_stats.append([])
    ws_stats.append(["Valid Status Values:"])
    ws_stats.append(["Not Started", "Control not yet started"])
    ws_stats.append(["In Progress", "Control implementation in progress"])
    ws_stats.append(["Completed", "Control fully implemented"])
    ws_stats.append(["Ready for Review", "Control ready for review"])
    ws_stats.append(["Failed", "Control implementation failed"])
    ws_stats.append(["Not Applicable", "Control not applicable"])
    ws_stats.append([])
    ws_stats.append(["Valid Priority Values:"])
    ws_stats.append(["Low", "Low priority"])
    ws_stats.append(["Medium", "Medium priority"])
    ws_stats.append(["High", "High priority"])
    ws_stats.append(["Critical", "Critical priority"])
    
    ws_stats['A1'].font = Font(size=14, bold=True, color="0066CC")
    ws_stats['A3'].font = Font(size=12, bold=True)
    ws_stats['A11'].font = Font(size=12, bold=True)
    
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 40
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Framework_Import_Template.xlsx"'
    
    wb.save(response)
    return response


# ========================
# Bulk Operations Views
# ========================

@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_apply_frameworks(request):
    """Застосування frameworks до кількох компаній"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to apply frameworks'))
        return redirect('compliance:framework_list')
    
    framework_ids_str = request.POST.get('framework_ids', '')
    company_ids = request.POST.getlist('company_ids')
    is_mandatory = request.POST.get('is_mandatory') == 'on'
    
    if not framework_ids_str or not company_ids:
        messages.error(request, _('Please select frameworks and companies'))
        return redirect('compliance:framework_list')
    
    framework_ids = [int(id) for id in framework_ids_str.split(',')]
    
    try:
        from app_cabinet.models import Company
        
        frameworks = ComplianceFramework.objects.filter(id__in=framework_ids, is_template=True)
        companies = Company.objects.filter(id__in=company_ids)
        
        created_count = 0
        
        with transaction.atomic():
            for framework in frameworks:
                for company in companies:
                    # Check if instance already exists
                    existing = ComplianceFramework.objects.filter(
                        template=framework,
                        company=company
                    ).first()
                    
                    if existing:
                        continue
                    
                    # Create instance
                    instance = ComplianceFramework.objects.create(
                        name=framework.name,
                        version=framework.version,
                        description=framework.description,
                        framework_type=framework.framework_type,
                        status='draft',
                        is_template=False,
                        is_mandatory=is_mandatory,
                        company=company,
                        template=framework,
                        created_by=request.user
                    )
                    
                    # Copy categories and controls
                    for category in framework.categories.all():
                        new_category = ControlCategory.objects.create(
                            framework=instance,
                            code=category.code,
                            name=category.name,
                            description=category.description,
                            order=category.order
                        )
                        
                        for control in category.controls.all():
                            Control.objects.create(
                                category=new_category,
                                code=control.code,
                                title=control.title,
                                name=control.name,
                                description=control.description,
                                domain=control.domain,
                                status='not_started',
                                priority=control.priority,
                                required_evidence_count=control.required_evidence_count,
                                evidence_description=control.evidence_description,
                                implementation_guidance=control.implementation_guidance,
                                testing_procedure=control.testing_procedure,
                                order=control.order,
                                parent_control=control.parent_control,
                                created_by=request.user
                            )
                    
                    created_count += 1
        
        messages.success(request, _(f'Successfully created {created_count} framework instances'))
        
    except Exception as e:
        messages.error(request, _(f'Error applying frameworks: {str(e)}'))
    
    return redirect('compliance:framework_list')


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_change_status(request):
    """Зміна статусу для кількох frameworks"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to change framework status'))
        return redirect('compliance:framework_list')
    
    framework_ids_str = request.POST.get('framework_ids', '')
    new_status = request.POST.get('status')
    
    if not framework_ids_str or not new_status:
        messages.error(request, _('Invalid parameters'))
        return redirect('compliance:framework_list')
    
    framework_ids = [int(id) for id in framework_ids_str.split(',')]
    
    try:
        frameworks = ComplianceFramework.objects.filter(id__in=framework_ids, is_template=True)
        updated_count = frameworks.update(status=new_status)
        
        messages.success(request, _(f'Successfully updated {updated_count} frameworks'))
        
    except Exception as e:
        messages.error(request, _(f'Error changing status: {str(e)}'))
    
    return redirect('compliance:framework_list')


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_export_frameworks(request):
    """Експорт кількох frameworks в один Excel файл"""
    # Check export permission
    if not check_user_compliance_permission(request.user, 'can_export'):
        messages.error(request, _('You do not have permission to export frameworks'))
        return redirect('compliance:framework_list')
    
    framework_ids_str = request.POST.get('framework_ids', '')
    
    if not framework_ids_str:
        messages.error(request, _('No frameworks selected'))
        return redirect('compliance:framework_list')
    
    framework_ids = [int(id) for id in framework_ids_str.split(',')]
    frameworks = ComplianceFramework.objects.filter(id__in=framework_ids, is_template=True)
    
    if not frameworks.exists():
        messages.error(request, _('Selected frameworks not found'))
        return redirect('compliance:framework_list')
    
    # Create ZIP file with multiple Excel files
    import io
    import zipfile
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for framework in frameworks:
            # Create Excel for each framework
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Framework Overview sheet
            ws_overview = wb.create_sheet("Framework Overview")
            ws_overview.append(["Framework Name:", framework.name])
            ws_overview.append(["Version:", framework.version])
            ws_overview.append(["Type:", framework.get_framework_type_display()])
            ws_overview.append(["Status:", framework.get_status_display()])
            ws_overview.append(["Description:", framework.description])
            
            # Controls sheet
            ws_controls = wb.create_sheet("Controls")
            headers = ["Category Code", "Category Name", "Control Code", "Control Name", "Domain Code", "Status", "Priority"]
            ws_controls.append(headers)
            
            for category in framework.categories.all():
                for control in category.controls.select_related('domain').all():
                    ws_controls.append([
                        category.code,
                        category.name,
                        control.code,
                        control.name,
                        control.domain.code if control.domain else "",
                        control.get_status_display(),
                        control.get_priority_display()
                    ])
            
            # Save to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Add to ZIP
            filename = f"{framework.name}_{framework.version}.xlsx".replace(' ', '_')
            zip_file.writestr(filename, excel_buffer.read())
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="Frameworks_Export_{datetime.now().strftime("%Y%m%d")}.zip"'
    
    return response


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_duplicate_frameworks(request):
    """Дублювання кількох frameworks"""
    if not check_user_compliance_permission(request.user, 'can_add_frameworks'):
        messages.error(request, _('You do not have permission to duplicate frameworks'))
        return redirect('compliance:framework_list')
    
    framework_ids_str = request.POST.get('framework_ids', '')
    
    if not framework_ids_str:
        messages.error(request, _('No frameworks selected'))
        return redirect('compliance:framework_list')
    
    framework_ids = [int(id) for id in framework_ids_str.split(',')]
    
    try:
        frameworks = ComplianceFramework.objects.filter(id__in=framework_ids, is_template=True)
        created_count = 0
        
        with transaction.atomic():
            for framework in frameworks:
                # Create copy
                new_framework = ComplianceFramework.objects.create(
                    name=f"{framework.name} (Copy)",
                    version=framework.version,
                    description=framework.description,
                    framework_type=framework.framework_type,
                    status='draft',
                    is_template=True,
                    created_by=request.user
                )
                
                # Copy categories and controls
                for category in framework.categories.all():
                    new_category = ControlCategory.objects.create(
                        framework=new_framework,
                        code=category.code,
                        name=category.name,
                        description=category.description,
                        order=category.order
                    )
                    
                    for control in category.controls.all():
                        Control.objects.create(
                            category=new_category,
                            code=control.code,
                            title=control.title,
                            name=control.name,
                            description=control.description,
                            domain=control.domain,
                            status=control.status,
                            priority=control.priority,
                            required_evidence_count=control.required_evidence_count,
                            evidence_description=control.evidence_description,
                            implementation_guidance=control.implementation_guidance,
                            testing_procedure=control.testing_procedure,
                            order=control.order,
                            created_by=request.user
                        )
                
                created_count += 1
        
        messages.success(request, _(f'Successfully duplicated {created_count} frameworks'))
        
    except Exception as e:
        messages.error(request, _(f'Error duplicating frameworks: {str(e)}'))
    
    return redirect('compliance:framework_list')


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_delete_frameworks(request):
    """Видалення кількох frameworks"""
    if not check_user_compliance_permission(request.user, 'can_delete_frameworks'):
        messages.error(request, _('You do not have permission to delete frameworks'))
        return redirect('compliance:framework_list')
    
    framework_ids_str = request.POST.get('framework_ids', '')
    
    if not framework_ids_str:
        messages.error(request, _('No frameworks selected'))
        return redirect('compliance:framework_list')
    
    framework_ids = [int(id) for id in framework_ids_str.split(',')]
    
    try:
        frameworks = ComplianceFramework.objects.filter(id__in=framework_ids, is_template=True)
        deleted_count = frameworks.count()
        frameworks.delete()
        
        messages.success(request, _(f'Successfully deleted {deleted_count} frameworks'))
        
    except Exception as e:
        messages.error(request, _(f'Error deleting frameworks: {str(e)}'))
    
    return redirect('compliance:framework_list')


# ========================
# Bulk Operations for Instances
# ========================

@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_change_instance_status(request):
    """Зміна статусу для кількох instances"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to change instance status'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids_str = request.POST.get('instance_ids', '')
    new_status = request.POST.get('status')
    
    if not instance_ids_str or not new_status:
        messages.error(request, _('Invalid parameters'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids = [int(id) for id in instance_ids_str.split(',')]
    
    try:
        # Get accessible companies for user
        accessible_companies = get_user_accessible_companies(request.user)
        accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
        
        instances = ComplianceFramework.objects.filter(
            id__in=instance_ids,
            is_template=False,
            company_id__in=accessible_company_ids
        )
        updated_count = instances.update(status=new_status)
        
        messages.success(request, _(f'Successfully updated {updated_count} instances'))
        
    except Exception as e:
        messages.error(request, _(f'Error changing status: {str(e)}'))
    
    return redirect('compliance:framework_instances_list')


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_toggle_mandatory(request):
    """Зміна mandatory флагу для кількох instances"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to change mandatory flag'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids_str = request.POST.get('instance_ids', '')
    is_mandatory = request.POST.get('is_mandatory') == 'true'
    
    if not instance_ids_str:
        messages.error(request, _('Invalid parameters'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids = [int(id) for id in instance_ids_str.split(',')]
    
    try:
        # Get accessible companies for user
        accessible_companies = get_user_accessible_companies(request.user)
        accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
        
        instances = ComplianceFramework.objects.filter(
            id__in=instance_ids,
            is_template=False,
            company_id__in=accessible_company_ids
        )
        updated_count = instances.update(is_mandatory=is_mandatory)
        
        mandatory_text = "mandatory" if is_mandatory else "optional"
        messages.success(request, _(f'Successfully marked {updated_count} instances as {mandatory_text}'))
        
    except Exception as e:
        messages.error(request, _(f'Error updating mandatory flag: {str(e)}'))
    
    return redirect('compliance:framework_instances_list')


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_export_instances(request):
    """Експорт кількох instances в один Excel файл"""
    # Check export permission
    if not check_user_compliance_permission(request.user, 'can_export'):
        messages.error(request, _('You do not have permission to export frameworks'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids_str = request.POST.get('instance_ids', '')
    
    if not instance_ids_str:
        messages.error(request, _('No instances selected'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids = [int(id) for id in instance_ids_str.split(',')]
    
    # Get accessible companies for user
    accessible_companies = get_user_accessible_companies(request.user)
    accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
    
    instances = ComplianceFramework.objects.filter(
        id__in=instance_ids,
        is_template=False,
        company_id__in=accessible_company_ids
    )
    
    if not instances.exists():
        messages.error(request, _('Selected instances not found'))
        return redirect('compliance:framework_instances_list')
    
    # Create ZIP file with multiple Excel files
    import io
    import zipfile
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for instance in instances:
            # Create Excel for each instance
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Instance Overview sheet
            ws_overview = wb.create_sheet("Instance Overview")
            ws_overview.append(["Framework Name:", instance.name])
            ws_overview.append(["Version:", instance.version])
            ws_overview.append(["Company:", instance.company.name if instance.company else "N/A"])
            ws_overview.append(["Type:", instance.get_framework_type_display()])
            ws_overview.append(["Status:", instance.get_status_display()])
            ws_overview.append(["Mandatory:", "Yes" if instance.is_mandatory else "No"])
            ws_overview.append(["Description:", instance.description])
            
            # Controls sheet
            ws_controls = wb.create_sheet("Controls")
            headers = ["Category Code", "Category Name", "Control Code", "Control Name", "Domain Code", "Status", "Priority"]
            ws_controls.append(headers)
            
            for category in instance.categories.all():
                for control in category.controls.select_related('domain').all():
                    ws_controls.append([
                        category.code,
                        category.name,
                        control.code,
                        control.name,
                        control.domain.code if control.domain else "",
                        control.get_status_display(),
                        control.get_priority_display()
                    ])
            
            # Save to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Add to ZIP
            company_name = instance.company.name.replace(' ', '_') if instance.company else 'NoCompany'
            filename = f"{instance.name}_{company_name}_{instance.version}.xlsx".replace(' ', '_')
            zip_file.writestr(filename, excel_buffer.read())
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="Instances_Export_{datetime.now().strftime("%Y%m%d")}.zip"'
    
    return response


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def bulk_delete_instances(request):
    """Видалення кількох instances"""
    if not check_user_compliance_permission(request.user, 'can_delete_frameworks'):
        messages.error(request, _('You do not have permission to delete instances'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids_str = request.POST.get('instance_ids', '')
    
    if not instance_ids_str:
        messages.error(request, _('No instances selected'))
        return redirect('compliance:framework_instances_list')
    
    instance_ids = [int(id) for id in instance_ids_str.split(',')]
    
    try:
        # Get accessible companies for user
        accessible_companies = get_user_accessible_companies(request.user)
        accessible_company_ids = list(accessible_companies.values_list('id', flat=True))
        
        instances = ComplianceFramework.objects.filter(
            id__in=instance_ids,
            is_template=False,
            company_id__in=accessible_company_ids
        )
        deleted_count = instances.count()
        instances.delete()
        
        messages.success(request, _(f'Successfully deleted {deleted_count} instances'))
        
    except Exception as e:
        messages.error(request, _(f'Error deleting instances: {str(e)}'))
    
    return redirect('compliance:framework_instances_list')


# ========================
# Framework Lifecycle Views
# ========================

@login_required
@compliance_access_required
@require_http_methods(["POST"])
def schedule_review(request, framework_id):
    """Планування review для framework"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to schedule reviews'))
        return redirect('compliance:framework_detail', framework_id=framework_id)
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        from datetime import datetime
        
        next_review_date_str = request.POST.get('next_review_date')
        review_frequency = request.POST.get('review_frequency')
        review_owner_id = request.POST.get('review_owner_id')
        
        if next_review_date_str:
            framework.next_review_date = parse_local_requirement_date(next_review_date_str)
        
        if review_frequency:
            framework.review_frequency = review_frequency
        
        if review_owner_id:
            framework.review_owner_id = review_owner_id
        else:
            framework.review_owner = None
        
        framework.save()
        
        log_compliance_action(
            request.user, 'update', 'framework', framework,
            changes={
                'action': 'review_scheduled',
                'next_review_date': str(framework.next_review_date),
                'review_frequency': framework.review_frequency,
                'review_owner': framework.review_owner.username if framework.review_owner else None
            },
            request=request
        )
        
        messages.success(request, _('Review scheduled successfully'))
        
    except Exception as e:
        messages.error(request, _(f'Error scheduling review: {str(e)}'))
    
    return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def mark_reviewed(request, framework_id):
    """Позначити framework як reviewed і автоматично запланувати наступний review"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to mark as reviewed'))
        return redirect('compliance:framework_detail', framework_id=framework_id)
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        
        review_date_str = request.POST.get('review_date')
        review_notes = request.POST.get('review_notes', '')
        
        if review_date_str:
            review_date = parse_local_requirement_date(review_date_str)
            framework.last_review_date = review_date
            
            # Auto-calculate next review based on frequency
            if framework.review_frequency == 'quarterly':
                framework.next_review_date = review_date + relativedelta(months=3)
            elif framework.review_frequency == 'semi_annual':
                framework.next_review_date = review_date + relativedelta(months=6)
            elif framework.review_frequency == 'annual':
                framework.next_review_date = review_date + relativedelta(years=1)
            elif framework.review_frequency == 'biennial':
                framework.next_review_date = review_date + relativedelta(years=2)
            
            framework.save()
            
            log_compliance_action(
                request.user, 'update', 'framework', framework,
                changes={
                    'action': 'marked_reviewed',
                    'review_date': str(review_date),
                    'next_review_date': str(framework.next_review_date),
                    'notes': review_notes
                },
                request=request
            )
            
            messages.success(request, _(f'Framework marked as reviewed. Next review scheduled for {framework.next_review_date}'))
        
    except Exception as e:
        messages.error(request, _(f'Error marking as reviewed: {str(e)}'))
    
    return redirect('compliance:framework_detail', framework_id=framework_id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def archive_framework(request, framework_id):
    """Архівування framework"""
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        messages.error(request, _('You do not have permission to archive frameworks'))
        return redirect('compliance:framework_detail', framework_id=framework_id)
    
    framework = get_object_or_404(ComplianceFramework, id=framework_id)
    
    try:
        archive_reason = request.POST.get('archive_reason', '')
        
        old_status = framework.status
        framework.status = 'archived'
        framework.save()
        
        log_compliance_action(
            request.user, 'update', 'framework', framework,
            changes={
                'action': 'archived',
                'old_status': old_status,
                'new_status': 'archived',
                'reason': archive_reason
            },
            request=request
        )
        
        messages.success(request, _('Framework archived successfully'))
        
    except Exception as e:
        messages.error(request, _(f'Error archiving framework: {str(e)}'))
    
    return redirect('compliance:framework_detail', framework_id=framework_id)


# ===== FRAMEWORK TRANSLATION FUNCTIONS =====

import threading
import time
from deep_translator import GoogleTranslator

# Global variables to track translation progress
framework_translation_in_progress = False
framework_translation_progress = {
    'total': 0,
    'processed': 0,
    'percent': 0,
    'log': []
}
framework_translation_stop_requested = False


def log_framework_translation(message):
    """Add a message to the framework translation log"""
    global framework_translation_progress
    framework_translation_progress['log'].append({
        'time': time.strftime('%H:%M:%S'),
        'message': message
    })
    print(f"[Framework Translation] {message}")


def translate_text(text, target_lang, model='google_translate'):
    """
    Translate text using selected model
    
    Args:
        text: Text to translate
        target_lang: Target language code (uk, ru, en, etc.)
        model: Translation model to use (google_translate, claude, etc.)
    
    Returns:
        Translated text or original text if translation fails
    """
    if not text or not text.strip():
        return text
    
    try:
        if model == 'google_translate':
            # Use Google Translator (free, no API key required)
            translator = GoogleTranslator(source='auto', target=target_lang)
            translated = translator.translate(text)
            return translated
            
        elif model == 'claude':
            import anthropic
            from app_ai.models import APISettingsClaude
            
            settings = APISettingsClaude.objects.first()
            if not settings or not settings.model_name:
                log_framework_translation("Claude settings not configured")
                return text
                
            client = anthropic.Anthropic(api_key=settings.api_key)
            
            # Визначення повної назви мови
            lang_names = {
                'uk': 'Ukrainian', 'ru': 'Russian', 'en': 'English', 'de': 'German',
                'fr': 'French', 'es': 'Spanish', 'it': 'Italian', 'pl': 'Polish',
                'pt': 'Portuguese', 'nl': 'Dutch', 'cs': 'Czech', 'sk': 'Slovak',
                'ro': 'Romanian', 'bg': 'Bulgarian', 'hr': 'Croatian', 'sr': 'Serbian',
                'tr': 'Turkish', 'ar': 'Arabic', 'zh': 'Chinese', 'ja': 'Japanese'
            }
            target_lang_name = lang_names.get(target_lang, target_lang)
            
            prompt = f"Translate the following text to {target_lang_name}. Provide only the translation without any comments or explanations:\n\n{text}"
            
            response = client.messages.create(
                model=settings.model_name.model_id,
                max_tokens=settings.max_tokens,
                temperature=settings.temperature,
                messages=[{"role": "user", "content": prompt}]
            )
            
            return response.content[0].text.strip()
            
        elif model == 'google':
            import google.generativeai as genai
            from app_ai.models import APISettingsGoogle
            
            settings = APISettingsGoogle.objects.first()
            if not settings or not settings.model_name:
                log_framework_translation("Google settings not configured")
                return text
                
            genai.configure(api_key=settings.api_key)
            model_instance = genai.GenerativeModel(settings.model_name.model_id)
            
            lang_names = {
                'uk': 'Ukrainian', 'ru': 'Russian', 'en': 'English', 'de': 'German',
                'fr': 'French', 'es': 'Spanish', 'it': 'Italian', 'pl': 'Polish',
                'pt': 'Portuguese', 'nl': 'Dutch', 'cs': 'Czech', 'sk': 'Slovak',
                'ro': 'Romanian', 'bg': 'Bulgarian', 'hr': 'Croatian', 'sr': 'Serbian',
                'tr': 'Turkish', 'ar': 'Arabic', 'zh': 'Chinese', 'ja': 'Japanese'
            }
            target_lang_name = lang_names.get(target_lang, target_lang)
            
            prompt = f"Translate to {target_lang_name}. Only provide the translation:\n\n{text}"
            response = model_instance.generate_content(prompt)
            return response.text.strip()
            
        elif model == 'groq':
            from groq import Groq
            from app_ai.models import APISettingsGroq
            
            settings = APISettingsGroq.objects.first()
            if not settings or not settings.model_name:
                log_framework_translation("Groq settings not configured")
                return text
                
            client = Groq(api_key=settings.api_key)
            
            lang_names = {
                'uk': 'Ukrainian', 'ru': 'Russian', 'en': 'English', 'de': 'German',
                'fr': 'French', 'es': 'Spanish', 'it': 'Italian', 'pl': 'Polish',
                'pt': 'Portuguese', 'nl': 'Dutch', 'cs': 'Czech', 'sk': 'Slovak',
                'ro': 'Romanian', 'bg': 'Bulgarian', 'hr': 'Croatian', 'sr': 'Serbian',
                'tr': 'Turkish', 'ar': 'Arabic', 'zh': 'Chinese', 'ja': 'Japanese'
            }
            target_lang_name = lang_names.get(target_lang, target_lang)
            
            response = client.chat.completions.create(
                messages=[
                    {"role": "system", "content": "You are a translation assistant. Provide only the translation."},
                    {"role": "user", "content": f"Translate to {target_lang_name}:\n\n{text}"}
                ],
                model=settings.model_name.model_id,
            )
            
            return response.choices[0].message.content.strip()
            
        elif model == 'ollama':
            from ollama import Client
            from app_ai.models import APISettingsOllama
            
            settings = APISettingsOllama.objects.first()
            if not settings or not settings.model_name:
                log_framework_translation("Ollama settings not configured")
                return text
                
            client = Client(host=settings.api_url)
            
            lang_names = {
                'uk': 'Ukrainian', 'ru': 'Russian', 'en': 'English', 'de': 'German',
                'fr': 'French', 'es': 'Spanish', 'it': 'Italian', 'pl': 'Polish',
                'pt': 'Portuguese', 'nl': 'Dutch', 'cs': 'Czech', 'sk': 'Slovak',
                'ro': 'Romanian', 'bg': 'Bulgarian', 'hr': 'Croatian', 'sr': 'Serbian',
                'tr': 'Turkish', 'ar': 'Arabic', 'zh': 'Chinese', 'ja': 'Japanese'
            }
            target_lang_name = lang_names.get(target_lang, target_lang)
            
            response = client.chat(
                model=settings.model_name.model_id,
                messages=[
                    {"role": "system", "content": "You are a translation assistant. Provide only the translation."},
                    {"role": "user", "content": f"Translate to {target_lang_name}:\n\n{text}"}
                ]
            )
            
            return response.message.content.strip()
            
        elif model == 'deepseek':
            from openai import OpenAI
            from app_ai.models import APISettingsDeepSeek
            
            settings = APISettingsDeepSeek.objects.first()
            if not settings or not settings.model_name:
                log_framework_translation("DeepSeek settings not configured")
                return text
                
            client = OpenAI(api_key=settings.api_key, base_url="https://api.deepseek.com/v1")
            
            lang_names = {
                'uk': 'Ukrainian', 'ru': 'Russian', 'en': 'English', 'de': 'German',
                'fr': 'French', 'es': 'Spanish', 'it': 'Italian', 'pl': 'Polish',
                'pt': 'Portuguese', 'nl': 'Dutch', 'cs': 'Czech', 'sk': 'Slovak',
                'ro': 'Romanian', 'bg': 'Bulgarian', 'hr': 'Croatian', 'sr': 'Serbian',
                'tr': 'Turkish', 'ar': 'Arabic', 'zh': 'Chinese', 'ja': 'Japanese'
            }
            target_lang_name = lang_names.get(target_lang, target_lang)
            
            response = client.chat.completions.create(
                model=settings.model_name.model_id,
                messages=[
                    {"role": "system", "content": "You are a translation assistant. Provide only the translation."},
                    {"role": "user", "content": f"Translate to {target_lang_name}:\n\n{text}"}
                ],
                max_tokens=settings.max_tokens,
                temperature=settings.temperature
            )
            
            return response.choices[0].message.content.strip()
            
        else:
            log_framework_translation(f"Translation model '{model}' not supported")
            return text
            
    except Exception as e:
        log_framework_translation(f"Translation error with {model}: {str(e)}")
        return text


def translate_framework_background(framework_ids, target_language, translation_model, options):
    """
    Background task to translate frameworks
    
    Args:
        framework_ids: List of framework IDs to translate
        target_language: Target language code
        translation_model: Translation model to use
        options: Dict with translate_name, translate_description, translate_categories, translate_controls
    """
    global framework_translation_in_progress, framework_translation_progress, framework_translation_stop_requested
    
    try:
        framework_translation_stop_requested = False
        
        log_framework_translation(f"Starting translation of {len(framework_ids)} frameworks to {target_language}")
        log_framework_translation(f"Using translation model: {translation_model}")
        
        # Get frameworks
        frameworks = ComplianceFramework.objects.filter(id__in=framework_ids)
        
        # Calculate total items to translate
        total_items = 0
        for framework in frameworks:
            if options['translate_name']:
                total_items += 1
            if options['translate_description'] and framework.description:
                total_items += 1
            if options['translate_categories']:
                categories = framework.categories.all()
                total_items += len(categories) * 2  # name + description
            if options['translate_controls']:
                controls = Control.objects.filter(category__framework=framework)
                total_items += len(controls) * 2  # name + description
        
        framework_translation_progress['total'] = total_items
        framework_translation_progress['processed'] = 0
        
        log_framework_translation(f"Total items to translate: {total_items}")
        
        # Translate each framework
        for framework in frameworks:
            if framework_translation_stop_requested:
                log_framework_translation("Translation stopped by user")
                break
            
            log_framework_translation(f"=" * 60)
            log_framework_translation(f"Translating Framework: {framework.name}")
            log_framework_translation(f"=" * 60)
            
            # Translate framework name
            if options['translate_name']:
                if framework_translation_stop_requested:
                    break
                    
                original_name = framework.name
                log_framework_translation(f"Translating framework name: {original_name[:50]}...")
                translated_name = translate_text(original_name, target_language, translation_model)
                framework.name = translated_name
                framework_translation_progress['processed'] += 1
                framework_translation_progress['percent'] = (framework_translation_progress['processed'] / total_items) * 100
                log_framework_translation(f"✓ Framework name translated: {translated_name[:50]}...")
                time.sleep(0.5)  # Delay to avoid rate limiting
            
            # Translate framework description
            if options['translate_description'] and framework.description:
                if framework_translation_stop_requested:
                    break
                    
                original_desc = framework.description
                log_framework_translation(f"Translating framework description ({len(original_desc)} chars)...")
                translated_desc = translate_text(original_desc, target_language, translation_model)
                framework.description = translated_desc
                framework_translation_progress['processed'] += 1
                framework_translation_progress['percent'] = (framework_translation_progress['processed'] / total_items) * 100
                log_framework_translation(f"✓ Framework description translated")
                time.sleep(0.5)
            
            # Save framework
            framework.save()
            log_framework_translation(f"✓ Framework saved")
            
            # Translate categories
            if options['translate_categories']:
                categories = framework.categories.all()
                log_framework_translation(f"Translating {len(categories)} categories...")
                
                for category in categories:
                    if framework_translation_stop_requested:
                        break
                    
                    # Translate category name
                    original_cat_name = category.name
                    log_framework_translation(f"  - Category: {original_cat_name[:40]}...")
                    translated_cat_name = translate_text(original_cat_name, target_language, translation_model)
                    category.name = translated_cat_name
                    framework_translation_progress['processed'] += 1
                    framework_translation_progress['percent'] = (framework_translation_progress['processed'] / total_items) * 100
                    time.sleep(0.5)
                    
                    # Translate category description
                    if category.description:
                        original_cat_desc = category.description
                        translated_cat_desc = translate_text(original_cat_desc, target_language, translation_model)
                        category.description = translated_cat_desc
                    framework_translation_progress['processed'] += 1
                    framework_translation_progress['percent'] = (framework_translation_progress['processed'] / total_items) * 100
                    
                    category.save()
                    log_framework_translation(f"    ✓ Category translated: {translated_cat_name[:40]}...")
                    time.sleep(0.5)
            
            # Translate controls
            if options['translate_controls']:
                controls = Control.objects.filter(category__framework=framework)
                log_framework_translation(f"Translating {len(controls)} controls...")
                
                for control in controls:
                    if framework_translation_stop_requested:
                        break
                    
                    # Translate control name
                    original_ctrl_name = control.name
                    log_framework_translation(f"  - Control: {original_ctrl_name[:40]}...")
                    translated_ctrl_name = translate_text(original_ctrl_name, target_language, translation_model)
                    control.name = translated_ctrl_name
                    framework_translation_progress['processed'] += 1
                    framework_translation_progress['percent'] = (framework_translation_progress['processed'] / total_items) * 100
                    time.sleep(0.5)
                    
                    # Translate control description
                    if control.description:
                        original_ctrl_desc = control.description
                        translated_ctrl_desc = translate_text(original_ctrl_desc, target_language, translation_model)
                        control.description = translated_ctrl_desc
                    framework_translation_progress['processed'] += 1
                    framework_translation_progress['percent'] = (framework_translation_progress['processed'] / total_items) * 100
                    
                    control.save()
                    log_framework_translation(f"    ✓ Control translated: {translated_ctrl_name[:40]}...")
                    time.sleep(0.5)
        
        log_framework_translation("=" * 60)
        log_framework_translation(f"Translation completed! Processed {framework_translation_progress['processed']}/{total_items} items")
        log_framework_translation("=" * 60)
        
    except Exception as e:
        log_framework_translation(f"Error during translation: {str(e)}")
        import traceback
        log_framework_translation(traceback.format_exc())
    
    finally:
        framework_translation_in_progress = False
        framework_translation_stop_requested = False


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def start_framework_translation(request):
    """Start framework translation process"""
    global framework_translation_in_progress, framework_translation_progress
    
    if not check_user_compliance_permission(request.user, 'can_edit_frameworks'):
        return JsonResponse({
            'success': False,
            'message': _('You do not have permission to translate frameworks')
        })
    
    if framework_translation_in_progress:
        return JsonResponse({
            'success': False,
            'message': _('Translation is already in progress')
        })
    
    try:
        data = json.loads(request.body)
        framework_ids = data.get('framework_ids', [])
        target_language = data.get('target_language', 'uk')
        translation_model = data.get('translation_model', 'google_translate')
        
        options = {
            'translate_name': data.get('translate_name', True),
            'translate_description': data.get('translate_description', True),
            'translate_categories': data.get('translate_categories', True),
            'translate_controls': data.get('translate_controls', True)
        }
        
        if not framework_ids:
            return JsonResponse({
                'success': False,
                'message': _('No frameworks selected')
            })
        
        # Reset progress
        framework_translation_progress = {
            'total': 0,
            'processed': 0,
            'percent': 0,
            'log': []
        }
        
        # Start translation in background thread
        framework_translation_in_progress = True
        thread = threading.Thread(
            target=translate_framework_background,
            args=(framework_ids, target_language, translation_model, options)
        )
        thread.daemon = True
        thread.start()
        
        return JsonResponse({
            'success': True,
            'message': _('Translation started')
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })


@login_required
@compliance_access_required
@require_http_methods(["GET"])
def get_framework_translation_progress(request):
    """Get current framework translation progress"""
    global framework_translation_in_progress, framework_translation_progress
    
    return JsonResponse({
        'in_progress': framework_translation_in_progress,
        'progress': framework_translation_progress
    })


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def stop_framework_translation(request):
    """Stop framework translation process"""
    global framework_translation_stop_requested
    
    if not framework_translation_in_progress:
        return JsonResponse({
            'success': False,
            'message': _('No translation is currently in progress')
        })
    
    framework_translation_stop_requested = True
    log_framework_translation("Stop requested by user - completing current item...")
    
    return JsonResponse({
        'success': True,
        'message': _('Stop requested')
    })


# ========================
# Framework Control Mapping Operations
# ========================

@login_required
@compliance_access_required
@require_http_methods(["POST"])
def control_mapping_create(request):
    """Create mapping for framework control to internal/local/framework controls"""
    def clean_id(value):
        """Clean and validate ID value"""
        if not value:
            return None
        cleaned = ''.join(c for c in str(value).strip() if c.isdigit())
        if not cleaned:
            return None
        try:
            return int(cleaned)
        except (ValueError, TypeError):
            return None

    source_control_id = clean_id(request.POST.get('source_control_id'))
    target_framework_id = clean_id(request.POST.get('target_framework_control_id'))
    target_internal_id = clean_id(request.POST.get('target_internal_control_id'))
    target_local_id = clean_id(request.POST.get('target_local_control_id'))

    if not source_control_id:
        messages.error(request, _('Source control is required'))
        return redirect('compliance:dashboard')

    source_control = get_object_or_404(Control, id=source_control_id)

    if not check_user_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to create mappings'))
        return redirect('compliance:control_detail', control_id=source_control.id)

    if not target_framework_id and not target_internal_id and not target_local_id:
        messages.error(request, _('Please select a target control'))
        return redirect('compliance:control_detail', control_id=source_control.id)

    try:
        mapping_type = request.POST.get('mapping_type', 'related')
        notes = request.POST.get('notes', '')

        mapping = ControlMapping.objects.create(
            source_control=source_control,
            target_control_id=target_framework_id,
            target_internal_control_id=target_internal_id,
            target_local_control_id=target_local_id,
            mapping_type=mapping_type,
            notes=notes,
            created_by=request.user
        )

        log_compliance_action(
            request.user,
            'create',
            'control_mapping',
            mapping,
            request=request
        )

        messages.success(request, _('Mapping created successfully'))

    except Exception as exc:
        messages.error(request, _('Error creating mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:control_detail', control_id=source_control.id)


@login_required
@compliance_access_required
@require_http_methods(["POST"])
def control_mapping_delete(request, mapping_id):
    """Delete control mapping"""
    mapping = get_object_or_404(ControlMapping, id=mapping_id)
    source_control = mapping.source_control

    if not check_user_compliance_permission(request.user, 'can_edit_controls'):
        messages.error(request, _('You do not have permission to delete mappings'))
        return redirect('compliance:control_detail', control_id=source_control.id)

    try:
        log_compliance_action(
            request.user,
            'delete',
            'control_mapping',
            mapping,
            request=request
        )

        mapping.delete()
        messages.success(request, _('Mapping deleted'))

    except Exception as exc:
        messages.error(request, _('Error deleting mapping: %(error)s') % {'error': str(exc)})

    return redirect('compliance:control_detail', control_id=source_control.id)

