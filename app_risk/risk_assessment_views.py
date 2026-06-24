# SecBoard/app_risk/risk_assessment_views.py
import csv
import json

import pytz
from django.http import JsonResponse, HttpResponse
from django.utils.encoding import smart_str
from django.views.decorators.http import require_http_methods, require_POST
from django.core.paginator import Paginator
from datetime import datetime
from rest_framework.decorators import api_view

from .models import Threat, Vulnerability, AssetVulnerability, SoftwareVulnerability, ExternalMediaVulnerability, RiskLevel, RiskTreatment, RiskTreatmentAttachment, AccessRisk, Treatment_type, Treatment_status, TreatmentPriority, TreatmentEffectiveness, MonitoringFrequency, FinancialImpact, OperationalImpact, ReputationalImpact, AcceptableRisk, AllowedSoftware, RiskAssessmentConfigGuide, RiskAssessmentConfigGuideTranslation, RiskAssessmentGuide, RiskAssessmentGuideTranslation
from app_asset.models import AssetGroup, CriticalityLevel, AssetType, SoftwareRegister, ExternalMediaRegister
from app_conf.models import Company, Country
from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import render, redirect, get_object_or_404
import logging
from django.utils.translation import gettext as _
from .logging_utils import (
    RiskAssessmentLogger, 
    log_risk_action, 
    log_data_access_decorator
)
from .forms import ThreatForm
from .pagination_utils import normalize_risk_table_length
from django.utils.translation import get_language
from django.db.models import F, Q, Value, CharField, IntegerField, Case, When, Count, Sum
from django.contrib import messages
from django.db.models.functions import Greatest, Cast, Coalesce
from decimal import Decimal, ROUND_DOWN
import decimal
from django.db import transaction
from rest_framework.response import Response
from django.utils import timezone
from app_asset.models import InformationAsset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth import get_user_model
from app_cabinet.models import CabinetUser

# Import helper functions from utils
from .risk_assessment_utils import (
    get_acceptable_risk_for_asset,
    get_localized_criticality,
    calculate_risk_level,
    get_risk_level_name,
    calculate_value_of_risk,
    calculate_threat_impact_value,
    calculate_threat_impact_level,
    get_risk_level,
    get_status_color,
    get_risk_level_from_name,
    get_user_risk_assessment_permissions,
    format_time,
    get_company_risk_levels_queryset
)

logger = logging.getLogger(__name__)


def _get_permitted_risk_company_ids(user):
    perms = get_user_risk_assessment_permissions(user)
    return set(perms.get('companies', []))


def has_risk_assessment_access(user):
    return AccessRisk.objects.filter(group__in=user.groups.all(), has_access_assessment=True).exists()

def has_risk_assessment_config_access(user):
    from .access_utils import has_risk_config_access
    return has_risk_config_access(user)

def can_edit_risk_assessment(user):
    return AccessRisk.objects.filter(group__in=user.groups.all(), can_edit_assessment=True).exists()


@user_passes_test(has_risk_assessment_access)
@log_risk_action("RISK_ASSESSMENT_VIEW")
def risk_assessment(request):
    try:
        user_permissions = get_user_risk_assessment_permissions(request.user)

        assets = InformationAsset.objects.filter(
            company__id__in=user_permissions['companies']
        ).select_related('company', 'asset_type')

        vulnerabilities = Vulnerability.objects.filter(
            is_active=True,
            asset_type__in=assets.values('asset_type')
        ).select_related('asset_type')

        # Get treatment types and statuses for the modal
        treatment_types = Treatment_type.objects.filter(is_active=True).order_by('name', 'code')
        treatment_statuses = Treatment_status.objects.filter(is_active=True).order_by('name', 'code')

        assets_paginator = Paginator(assets, 10)  # 10 items per page
        page_number = request.GET.get('page')
        assets_page = assets_paginator.get_page(page_number)

        context = {
            'assets_page': assets_page,
            'vulnerabilities': vulnerabilities,
            'can_edit': user_permissions['can_edit'],
            'show_link': user_permissions['show_link'],
            'total_assets': assets.count(),
            'total_vulnerabilities': vulnerabilities.count(),
            'user_companies': Company.objects.filter(id__in=user_permissions['companies']),
            'treatment_types': treatment_types,
            'treatment_statuses': treatment_statuses,
        }

        # logger.info(f"User {request.user.username} accessed risk assessment")
        return render(request, 'app_risk/risk_assessment.html', context)

    except Exception as e:
        logger.error(f"Error in risk assessment view: {str(e)}", exc_info=True)
        messages.error(request, "An error occurred while loading the risk assessment page.")
        return render(request, 'app_risk/error.html', {'error_message': "An error occurred while loading the risk assessment page."})



@user_passes_test(has_risk_assessment_config_access)
@require_http_methods(["GET", "POST"])
def risk_assessment_config(request):
    from .vulnerability_utils import get_supported_risk_language
    current_language = get_supported_risk_language(get_language())

    threats = Threat.objects.order_by('-impact')

    if request.method == 'POST':
        post_data = request.POST.copy()
        # Map default English fields (Add/Edit threat modal sends only name_en, description_en, risks_en)
        post_data['name'] = (post_data.get('name_en') or post_data.get('name_uk') or post_data.get('name_ru') or '').strip()
        post_data['description'] = (post_data.get('description_en') or post_data.get('description_uk') or post_data.get('description_ru') or '').strip()
        post_data['risks'] = (post_data.get('risks_en') or post_data.get('risks_uk') or post_data.get('risks_ru') or '').strip()
        form = ThreatForm(post_data)
        if form.is_valid():
            threat = form.save()
            # Save extra language fields (de, fr, etc.) from POST
            from .vulnerability_utils import get_vulnerability_form_languages
            from .models import Threat as ThreatModel
            for lang_code, _ in get_vulnerability_form_languages():
                if lang_code not in ('uk', 'ru', 'en'):
                    for field in ThreatModel.TRANSLATABLE_FIELDS:
                        key = f'{field}_{lang_code}'
                        if key in request.POST:
                            threat.set_translated_value(field, lang_code, request.POST.get(key, '') or '')
            threat.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': 'Threat saved successfully',
                    'id': threat.id
                })
            else:
                return redirect('risk_assessment_config')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Error saving threat',
                    'errors': form.errors
                }, status=400)
    else:
        form = ThreatForm()

    vulnerabilities = Vulnerability.objects.all()

    # Get user permissions for Risk Configuration
    from .access_utils import get_user_risk_permissions
    user_permissions = get_user_risk_permissions(request.user)

    context = {
        'threats': threats,
        'vulnerabilities': vulnerabilities,
        'form': form,
        'current_language': get_language(),
        'LANGUAGE_CODE': get_language(),
        'vulnerabilities_context': 'Config',
        'user_permissions': user_permissions,
    }

    return render(request, 'app_risk/risk_assessment_config.html', context)


@user_passes_test(has_risk_assessment_config_access)
@require_http_methods(["GET"])
def risk_assessment_config_guide(request):
    """Return JSON { content: html } for the Risk Assessment Config guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = RiskAssessmentConfigGuide.objects.first()
    if guide:
        if country:
            trans = RiskAssessmentConfigGuideTranslation.objects.filter(config_guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = RiskAssessmentConfigGuideTranslation.objects.filter(config_guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def risk_assessment_config_guide_translate(request):
    """API for AI translation of Risk Assessment Config guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@user_passes_test(has_risk_assessment_access)
@require_http_methods(["GET"])
def risk_assessment_guide(request):
    """Return JSON { content: html } for the Risk Assessment guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = RiskAssessmentGuide.objects.first()
    if guide:
        if country:
            trans = RiskAssessmentGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = RiskAssessmentGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def risk_assessment_guide_translate(request):
    """API for AI translation of Risk Assessment guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
@user_passes_test(has_risk_assessment_access)
@log_data_access_decorator("ASSET_RISKS")
def get_asset_risks(request):
    from .vulnerability_utils import get_supported_risk_language
    asset_id = request.GET.get('asset_id')
    language = get_supported_risk_language(request.GET.get('language', get_language()))

    # logger.info(f"Received request for asset risks. Asset ID: {asset_id}, Language: {language}")

    if not asset_id:
        return JsonResponse({'error': 'Asset ID is required'}, status=400)

    try:
        asset = InformationAsset.objects.get(id=asset_id)
        # logger.info(f"Found asset: {asset.name} (ID: {asset.asset_id})")
    except InformationAsset.DoesNotExist:
        logger.error(f"Asset not found with ID: {asset_id}")
        return JsonResponse({'error': 'Asset not found'}, status=404)

    # Get criticality with localization based on request language
    def get_localized_criticality(asset, language):
        levels = [
            (asset.confidentiality, asset.confidentiality.cost if asset.confidentiality else 0),
            (asset.integrity, asset.integrity.cost if asset.integrity else 0),
            (asset.availability, asset.availability.cost if asset.availability else 0)
        ]
        max_level = max(levels, key=lambda x: x[1])
        if max_level[0]:
            return {
                'name': max_level[0].get_name() if max_level[0] else '',
                'cost': max_level[0].cost,
                'color': max_level[0].color
            }
        return {'name': _("Undefined"), 'cost': 0, 'color': "#000000"}
    
    criticality = get_localized_criticality(asset, language)
    # logger.info(f"Asset criticality: {criticality}")

    # Get acceptable risk level for this asset
    acceptable_risk = get_acceptable_risk_for_asset(asset, language)

    # Get manual risk level overrides for this asset
    from .models import ManualRiskLevelOverride
    manual_overrides = []
    override_lookup = {}
    
    try:
        manual_overrides = ManualRiskLevelOverride.objects.filter(asset=asset).select_related(
            'vulnerability', 'threat', 'manual_risk_level'
        )
        
        # Create a lookup dictionary for manual overrides
        for override in manual_overrides:
            key = (override.vulnerability.id, override.threat.id if override.threat else None)
            override_lookup[key] = override
    except Exception as e:
        # If the table doesn't exist, continue without manual overrides
        logger.warning(f"ManualRiskLevelOverride table not available: {e}")
        manual_overrides = []
        override_lookup = {}

    risks = []
    # Get all vulnerabilities with status 'Yes' and 'No' (exclude orphaned records)
    asset_vulnerabilities = AssetVulnerability.objects.filter(
        asset=asset,
        status__in=['Yes', 'No'],
        vulnerability__isnull=False
    ).select_related(
        'vulnerability',
        'asset'
    )

    # logger.info(f"Found {asset_vulnerabilities.count()} vulnerabilities with status 'Yes' or 'No'")
    
    # Log each vulnerability found
    for av in asset_vulnerabilities:
        # logger.info(f"Processing AssetVulnerability ID: {av.id}, Vulnerability ID: {av.vulnerability.id}")
        # logger.info(f"Vulnerability details: {av.vulnerability.vulnerability_uk}")
        
        # Get all threats associated with this vulnerability
        threats = av.vulnerability.threats.all()
        # logger.info(f"Found {threats.count()} threats for vulnerability {av.vulnerability.id}")
        
        if not threats.exists():
            # If no threats, create a default risk entry with minimal values for 'No' status
            if av.status == 'No':
                probability = Decimal('0.001')
                impact = Decimal('0.001')
                # Вплив загрози = Ймовірність (L) × Загальний вплив (E) × 100 для відсоткового формату
                threat_impact_value = probability * impact * 100
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                calculated_risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)
                
                # Check for manual override
                override_key = (av.vulnerability.id, None)  # No threat for 'No' status
                manual_override = override_lookup.get(override_key)
                
                if manual_override:
                    risk_level = manual_override.manual_risk_level
                    is_manual_override = True
                    justification = manual_override.justification
                else:
                    risk_level = calculated_risk_level
                    is_manual_override = False
                    justification = None
                
                # Check if risk level exceeds acceptable risk
                exceeds_acceptable = False
                acceptable_risk_info = None
                if acceptable_risk and risk_level:
                    if isinstance(risk_level, RiskLevel) and risk_level.max_value > acceptable_risk['max_value']:
                        exceeds_acceptable = True
                        acceptable_risk_info = acceptable_risk
                
                risk_data = {
                    'asset_id': asset.asset_id,
                    'asset_name': asset.name,
                    'company': asset.company.name if asset.company else 'N/A',
                    'criticality': {
                        'name': criticality['name'],
                        'cost': criticality['cost'],
                        'color': criticality['color']
                    },
                    'vulnerability': av.vulnerability.get_name(language),
                    'vulnerability_id': av.vulnerability.id,
                    'vulnerability_description': av.vulnerability.get_translated_value('description', language) or av.vulnerability.description or '',
                    'vulnerability_status': av.status,
                    'threat': _('Not Applicable'),
                    'threat_id': None,
                    'probability_impact': f"{probability:.4f} / {impact:.2f}",
                    'threat_impact_value': f"{threat_impact_value:.4f}%",
                    'threat_impact_level': threat_impact_level,
                    'value_of_risk': float(value_of_risk),
                    'risk_level': {
                        'text': get_risk_level_name(risk_level, language),
                        'color': risk_level.color if isinstance(risk_level, RiskLevel) else "#000000",
                        'exceeds_acceptable': exceeds_acceptable,
                        'acceptable_risk_info': acceptable_risk_info,
                        'is_manual_override': is_manual_override,
                        'justification': justification
                    },
                    'risk_mitigation_controls': av.vulnerability.get_translated_value('risk_mitigation_controls', language) or av.vulnerability.risk_mitigation_controls or ''
                }
            else:
                # For other statuses, keep the original behavior
                # Check if risk level exceeds acceptable risk (for undefined status)
                exceeds_acceptable = False
                acceptable_risk_info = None
                if acceptable_risk:
                    # For undefined status, we consider it as potentially exceeding if acceptable risk is very low
                    if acceptable_risk['max_value'] < 1:  # Very low acceptable risk
                        exceeds_acceptable = True
                        acceptable_risk_info = acceptable_risk
                
                risk_data = {
                    'asset_id': asset.asset_id,
                    'asset_name': asset.name,
                    'company': asset.company.name if asset.company else 'N/A',
                    'criticality': {
                        'name': criticality['name'],
                        'cost': criticality['cost'],
                        'color': criticality['color']
                    },
                    'vulnerability': av.vulnerability.get_name(language),
                    'vulnerability_id': av.vulnerability.id,
                    'vulnerability_description': av.vulnerability.get_translated_value('description', language) or av.vulnerability.description or '',
                    'vulnerability_status': av.status,
                    'threat': 'N/A',
                    'threat_id': None,
                    'probability_impact': 'N/A',
                    'threat_impact_value': '0%',
                    'threat_impact_level': 0,
                    'value_of_risk': 0,
                    'risk_level': {
                        'text': _('Undefined'),
                        'color': "#808080",
                        'exceeds_acceptable': exceeds_acceptable,
                        'acceptable_risk_info': acceptable_risk_info,
                        'is_manual_override': False
                    },
                    'risk_mitigation_controls': av.vulnerability.get_translated_value('risk_mitigation_controls', language) or av.vulnerability.risk_mitigation_controls or ''
                }
            risks.append(risk_data)
            continue
        
        for threat in threats:
            # logger.info(f"Processing threat: {threat.get_name()}")
            
            try:
                # For 'No' status vulnerabilities, use minimal values for probability and impact
                if av.status == 'No':
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                else:
                    probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    # Use new impact calculation methodology if available
                    if hasattr(threat, 'calculate_overall_impact'):
                        impact = threat.calculate_overall_impact()
                    else:
                        # Fallback to original impact field
                        impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN) / 100
                
                # Вплив загрози = Ймовірність (L) × Загальний вплив (E) × 100 для відсоткового формату
                threat_impact_value = probability * impact * 100
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                calculated_risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)

                # logger.info(f"Calculated values - Probability: {probability}, Impact: {impact}, "
                #           f"Threat Impact Value: {threat_impact_value}, Level: {threat_impact_level}, "
                #           f"Risk Value: {value_of_risk}, Risk Level: {risk_level}")

                # Check for manual override
                override_key = (av.vulnerability.id, threat.id)
                manual_override = override_lookup.get(override_key)
                
                if manual_override:
                    risk_level = manual_override.manual_risk_level
                    is_manual_override = True
                    justification = manual_override.justification
                else:
                    risk_level = calculated_risk_level
                    is_manual_override = False
                    justification = None

                # Check if risk level exceeds acceptable risk
                exceeds_acceptable = False
                acceptable_risk_info = None
                if acceptable_risk and risk_level:
                    if isinstance(risk_level, RiskLevel) and risk_level.max_value > acceptable_risk['max_value']:
                        exceeds_acceptable = True
                        acceptable_risk_info = acceptable_risk

                risk_data = {
                    'asset_id': asset.asset_id,
                    'asset_name': asset.name,
                    'company': asset.company.name if asset.company else 'N/A',
                    'criticality': {
                        'name': criticality['name'],
                        'cost': criticality['cost'],
                        'color': criticality['color']
                    },
                    'vulnerability': av.vulnerability.get_name(language),
                    'vulnerability_id': av.vulnerability.id,
                    'vulnerability_description': av.vulnerability.get_translated_value('description', language) or av.vulnerability.description or '',
                    'vulnerability_status': av.status,
                    'threat': threat.get_name(language),
                    'threat_id': threat.id,
                    'probability_impact': f"{probability:.4f} / {impact:.2f}",
                    'threat_impact_value': f"{threat_impact_value:.4f}%",
                    'threat_impact_level': threat_impact_level,
                    'value_of_risk': float(value_of_risk),
                    'risk_level': {
                        'text': get_risk_level_name(risk_level, language),
                        'color': risk_level.color if isinstance(risk_level, RiskLevel) else "#000000",
                        'exceeds_acceptable': exceeds_acceptable,
                        'acceptable_risk_info': acceptable_risk_info,
                        'is_manual_override': is_manual_override,
                        'justification': justification
                    },
                    'risk_mitigation_controls': av.vulnerability.get_translated_value('risk_mitigation_controls', language) or av.vulnerability.risk_mitigation_controls or ''
                }
                risks.append(risk_data)
            except Exception as e:
                logger.error(f"Error calculating risk for threat {threat.id}: {str(e)}")
                continue

    # logger.info(f"Returning {len(risks)} risks for asset {asset_id}")
    
    # Include asset information in the response
    asset_info = {
        'asset_id': asset.asset_id,
        'name': asset.name,
        'company': asset.company.name if asset.company else 'N/A'
    }
    
    response_data = {
        'asset': asset_info,
        'risks': risks
    }
    # logger.info(f"Response data: {response_data}")
    
    return JsonResponse(response_data)


@user_passes_test(has_risk_assessment_access)
def risk_calculation_data(request):
    from .vulnerability_utils import get_supported_risk_language
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    language = get_supported_risk_language(request.GET.get('language', get_language()))

    order_column_index = int(request.GET.get('order[0][column]', 1))  # Змінено на 1, щоб за замовчуванням сортувати за asset_id
    order_direction = request.GET.get('order[0][dir]', 'desc')  # Змінено на 'desc'

    columns = [
        'asset__asset_id', 'asset__name', 'asset__company__name', 'max_cost',
        f'vulnerability__vulnerability_{language}', f'vulnerability__threats__name_{language}',
        'vulnerability__threats__probability', 'vulnerability__threats__impact',
        'max_cost', 'max_cost', 'max_cost', f'vulnerability__risk_mitigation_controls_{language}'
    ]

    user_permissions = get_user_risk_assessment_permissions(request.user)

    show_deleted = request.GET.get('showDeleted', 'false').lower() == 'true'

    queryset = AssetVulnerability.objects.filter(
        asset__company__id__in=user_permissions['companies']
    ).select_related(
        'asset', 'asset__company', 'vulnerability',
        'asset__confidentiality', 'asset__integrity', 'asset__availability'
    ).prefetch_related(
        'vulnerability__threats'
    ).annotate(
        max_cost=Greatest(
            F('asset__confidentiality__cost'),
            F('asset__integrity__cost'),
            F('asset__availability__cost')
        ),
        quantitative_criticality=F('max_cost'),
        qualitative_criticality=Case(
            When(max_cost=F('asset__confidentiality__cost'), then=F('asset__confidentiality__name')),
            When(max_cost=F('asset__integrity__cost'), then=F('asset__integrity__name')),
            When(max_cost=F('asset__availability__cost'), then=F('asset__availability__name')),
            default=Value('Невизначено'),
            output_field=CharField()
        ),
        asset_id_as_int=Cast('asset__asset_id', IntegerField())  # Додано цю анотацію
    )

    if not show_deleted:
        queryset = queryset.filter(asset__deletion_date__isnull=True, asset__is_active=True)


    if search_value:
        queryset = queryset.filter(
            Q(asset__asset_id__icontains=search_value) |
            Q(asset__name__icontains=search_value) |
            Q(asset__company__name__icontains=search_value) |
            Q(**{f'vulnerability__vulnerability_{language}__icontains': search_value}) |
            Q(**{f'vulnerability__threats__name_{language}__icontains': search_value})
        )

    total_records = queryset.count()

    # Сортування
    order_column = columns[order_column_index]
    if 'asset__asset_id' in order_column:
        queryset = queryset.order_by(f'{"-" if order_direction == "desc" else ""}asset_id_as_int')
    else:
        if order_direction == 'desc':
            order_column = f'-{order_column}'
        queryset = queryset.order_by(order_column)

    asset_vulnerabilities = queryset[start:start + length]

    grouped_data = {}
    for av in asset_vulnerabilities:
        criticality = av.asset.get_criticality()
        asset_id = av.asset.asset_id

        if asset_id not in grouped_data:
            grouped_data[asset_id] = {
                'asset_id': asset_id,
                'asset_name': av.asset.name,
                'company': av.asset.company.name,
                'criticality': {
                    'text': f"{criticality['name']} / {criticality['cost']}",
                    'color': criticality['color']
                },
                'risks': []
            }

        for threat in av.vulnerability.threats.all():
            probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
            impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
            threat_impact_value = probability * impact
            threat_impact_level = calculate_threat_impact_level(threat_impact_value)
            value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
            risk_level = get_risk_level(value_of_risk, company_id=av.asset.company_id)

            risk_data = {
                "vulnerability": av.vulnerability.get_name(language),
                "threat": threat.get_name(language),
                "probability_impact": f"{probability:.4f} / {impact:.2f}",
                "threat_impact_value": f"{threat_impact_value:.4f}%",
                "threat_impact_level": threat_impact_level,
                "value_of_risk": float(value_of_risk),
                "risk_level": {
                    'text': risk_level.name if risk_level else _("Undefined"),
                    'color': risk_level.color if risk_level else "#000000"
                },
                "risk_mitigation_controls": getattr(av.vulnerability, f'risk_mitigation_controls_{language}',
                                                    av.vulnerability.risk_mitigation_controls_uk)
            }
            grouped_data[asset_id]['risks'].append(risk_data)

    # Сортування ризиків для кожного активу
    for asset_data in grouped_data.values():
        asset_data['risks'].sort(key=lambda x: x['value_of_risk'], reverse=True)
        asset_data['top_risks'] = asset_data['risks'][:3]

    data = list(grouped_data.values())

    # Сортування даних на стороні Python, якщо потрібно
    if order_column_index in [6, 7, 8, 9]:  # Індекси колонок, які не можна сортувати на рівні БД
        reverse = order_direction == 'desc'
        if order_column_index == 6:
            data.sort(key=lambda x: float(x['top_risks'][0]['threat_impact_value'].rstrip('%')), reverse=reverse)
        elif order_column_index == 7:
            data.sort(key=lambda x: x['top_risks'][0]['threat_impact_level'], reverse=reverse)
        elif order_column_index == 8:
            data.sort(key=lambda x: x['top_risks'][0]['value_of_risk'], reverse=reverse)
        elif order_column_index == 9:
            data.sort(key=lambda x: x['top_risks'][0]['risk_level']['text'], reverse=reverse)

    response = {
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data
    }

    return JsonResponse(response)


@login_required
@user_passes_test(has_risk_assessment_access)
@log_data_access_decorator("RISK_TREATMENT_DATA")
def get_risk_treatment_data(request):
    asset_id = request.GET.get('asset_id')
    current_language = get_language()[:2]
    print(f"Received request for asset_id: {asset_id}, language: {current_language}")

    if not asset_id:
        return JsonResponse({'error': 'Asset ID is required'}, status=400)

    try:
        entity_type = 'asset'
        entity = None
        entity_filter = {}
        if str(asset_id).startswith('S'):
            entity_type = 'software'
            sw_id = str(asset_id)[1:]
            if sw_id.isdigit():
                entity = SoftwareRegister.objects.filter(id=int(sw_id)).first()
            if not entity:
                return JsonResponse({'error': 'Software entry not found'}, status=404)
            entity_filter = {'software_register': entity}
        elif str(asset_id).startswith('M'):
            entity_type = 'external_media'
            em_id = str(asset_id)[1:]
            if em_id.isdigit():
                entity = ExternalMediaRegister.objects.filter(id=int(em_id)).first()
            if not entity:
                return JsonResponse({'error': 'External media entry not found'}, status=404)
            entity_filter = {'external_media_register': entity}
        else:
            entity = InformationAsset.objects.filter(Q(asset_id=asset_id) | Q(id=asset_id)).first()
            if not entity:
                return JsonResponse({'error': 'Asset not found'}, status=404)
            entity_filter = {'asset': entity}

        # Get acceptable risk level for this asset
        acceptable_risk = get_acceptable_risk_for_asset(entity, current_language) if entity_type == 'asset' else None

        # Get manual risk level overrides for this asset
        from .models import ManualRiskLevelOverride
        manual_overrides = []
        override_lookup = {}
        
        try:
            manual_overrides = ManualRiskLevelOverride.objects.filter(asset=entity).select_related(
                'vulnerability', 'threat', 'manual_risk_level'
            ) if entity_type == 'asset' else []
            
            # Create a lookup dictionary for manual overrides
            for override in manual_overrides:
                override_key = (override.vulnerability.id, override.threat.id if override.threat else None)
                override_lookup[override_key] = override
        except Exception as e:
            # If the table doesn't exist, continue without manual overrides
            logger.warning(f"ManualRiskLevelOverride table not available: {e}")
            manual_overrides = []
            override_lookup = {}

        if entity_type == 'software':
            asset_vulnerabilities = SoftwareVulnerability.objects.filter(software_register=entity, status__in=['Yes', 'No']).select_related('vulnerability')
        elif entity_type == 'external_media':
            asset_vulnerabilities = ExternalMediaVulnerability.objects.filter(external_media_register=entity, status__in=['Yes', 'No']).select_related('vulnerability')
        else:
            asset_vulnerabilities = AssetVulnerability.objects.filter(asset=entity, status__in=['Yes', 'No']).select_related('vulnerability')
        existing_treatments = RiskTreatment.objects.filter(**entity_filter).select_related(
            'treatment_type', 'status', 'residual_risk_level', 'effectiveness', 'priority', 'monitoring_frequency'
        ).prefetch_related('monitoring_responsible', 'dependencies', 'affected_assets', 'attachments')
        existing_treatment_vulns = set(existing_treatments.values_list('vulnerability_id', flat=True))

        # Get all treatment types and statuses (active only for dropdowns)
        all_treatment_types = Treatment_type.objects.filter(is_active=True).order_by('name', 'code')
        all_treatment_statuses = Treatment_status.objects.filter(is_active=True).order_by('name', 'code')

        # Create dictionaries for treatment types and statuses
        treatment_types_dict = {
            treatment_type.code: {
                'name': treatment_type.get_name(current_language),
                'color': treatment_type.color
            }
            for treatment_type in all_treatment_types
        }

        treatment_statuses_dict = {
            status.code: {
                'name': status.get_name(current_language),
                'color': status.color
            }
            for status in all_treatment_statuses
        }

        # Helper function for localized criticality
        def get_localized_criticality(asset_obj, language):
            levels = [
                (asset_obj.confidentiality, asset_obj.confidentiality.cost if asset_obj.confidentiality else 0),
                (asset_obj.integrity, asset_obj.integrity.cost if asset_obj.integrity else 0),
                (asset_obj.availability, asset_obj.availability.cost if asset_obj.availability else 0)
            ]
            max_level = max(levels, key=lambda x: x[1])
            if max_level[0]:
                return {
                    'name': max_level[0].get_name() if max_level[0] else '',
                    'cost': max_level[0].cost,
                    'color': max_level[0].color
                }
            return {'name': _("Undefined"), 'cost': 0, 'color': "#000000"}

        data = []
        for av in asset_vulnerabilities:
            criticality = get_localized_criticality(entity, current_language)
            # The criticality dictionary structure returns {'name': str, 'cost': int, 'color': str}

            threats_and_risk_levels = []
            highest_risk_level = None

            for threat in av.vulnerability.threats.all():
                # Check for manual override first
                override_key = (av.vulnerability.id, threat.id)
                manual_override = override_lookup.get(override_key)
                
                if manual_override:
                    # Use manual override risk level
                    risk_level = manual_override.manual_risk_level
                    is_manual_override = True
                else:
                    # Calculate risk level normally
                    # For 'No' status vulnerabilities, use minimal values for probability and impact
                    if av.status == 'No':
                        probability = Decimal('0.001')
                        impact = Decimal('0.001')
                    else:
                        probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                        impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                    
                    threat_impact_value = probability * impact
                    threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                    value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                    risk_level = get_risk_level(value_of_risk, company_id=entity.company_id)
                    is_manual_override = False

                # Get localized threat name with proper fallback
                threat_name = (
                        threat.get_name(current_language) or
                        _('Undefined')
                )

                # Get localized risk level name with proper fallback
                risk_level_name = (
                    risk_level.get_name(current_language) if risk_level else _('Undefined')
                )

                # Add indicator for manual override
                if is_manual_override:
                    risk_level_name += " (Manual)"

                threats_and_risk_levels.append(f"{threat_name}({risk_level_name})")

                if highest_risk_level is None or (risk_level and risk_level.max_value > highest_risk_level.max_value):
                    highest_risk_level = risk_level

            # Get default treatment type and status objects
            default_treatment_type = Treatment_type.objects.filter(code='Undefined').first()
            default_treatment_status = Treatment_status.objects.filter(code='Undefined').first()

            # Update existing treatment or create new one with proper localization
            treatment = None
            if av.vulnerability.id in existing_treatment_vulns:
                treatment = next((t for t in existing_treatments if t.vulnerability_id == av.vulnerability.id), None)
                if not treatment:
                    treatment = RiskTreatment.objects.filter(vulnerability_id=av.vulnerability.id, **entity_filter).first()
                if treatment:
                    treatment.threats = ', '.join(threats_and_risk_levels)
                    treatment.save()
            if not treatment:
                treatment = RiskTreatment.objects.create(
                    vulnerability=av.vulnerability,
                    threats=', '.join(threats_and_risk_levels),
                    risk_mitigation_controls=av.vulnerability,
                    treatment_type=default_treatment_type,
                    description='',
                    responsible='',
                    deadline=None,
                    status=default_treatment_status,
                    highest_risk_level=highest_risk_level,
                    last_modified_by=request.user,
                    **entity_filter,
                )

            # Check if risk levels exceed acceptable risk
            highest_risk_exceeds_acceptable = False
            residual_risk_exceeds_acceptable = False
            
            if acceptable_risk and highest_risk_level:
                if isinstance(highest_risk_level, RiskLevel) and highest_risk_level.max_value > acceptable_risk['max_value']:
                    highest_risk_exceeds_acceptable = True
            
            if acceptable_risk and treatment.residual_risk_level:
                if isinstance(treatment.residual_risk_level, RiskLevel) and treatment.residual_risk_level.max_value > acceptable_risk['max_value']:
                    residual_risk_exceeds_acceptable = True

            treatment_data = {
                'id': treatment.id,
                'asset_id': (
                    f"S{entity.id:06d}" if entity_type == 'software'
                    else (f"M{entity.id:05d}" if entity_type == 'external_media' else entity.asset_id)
                ),
                'asset_name': entity.name,
                'company': entity.company.name if entity.company else '',
                'criticality': {
                    'name': criticality['name'],  # Using the correct key 'name' instead of 'name_uk'
                    'cost': criticality['cost'],
                    'color': criticality['color']
                },
                'vulnerability': av.vulnerability.get_name(current_language),
                'vulnerability_status': av.status,
                'threats_and_risk_levels': treatment.threats,
                'highest_risk_level': {
                    'name': highest_risk_level.get_name(current_language) if highest_risk_level else _('Undefined'),
                    'color': highest_risk_level.color if highest_risk_level else '#000000',
                    'exceeds_acceptable': highest_risk_exceeds_acceptable,
                    'acceptable_risk_info': acceptable_risk if highest_risk_exceeds_acceptable else None,
                    'is_manual_override': any(override_lookup.get((av.vulnerability.id, threat.id)) for threat in av.vulnerability.threats.all() if threat.id)
                },
                'risk_mitigation_controls': (av.vulnerability.get_translated_value('risk_mitigation_controls', current_language) or '') if hasattr(av.vulnerability, 'get_translated_value') else (getattr(av.vulnerability, 'risk_mitigation_controls', None) or ''),
                'treatment_type': {
                    'name': treatment.treatment_type.get_name(current_language) if treatment.treatment_type else _('Undefined'),
                    'code': treatment.treatment_type.code if treatment.treatment_type else 'Undefined',
                    'color': treatment.treatment_type.color if treatment.treatment_type else '#808080'
                },
                'description': treatment.description,
                'responsible': treatment.responsible,
                'deadline': treatment.deadline.isoformat() if treatment.deadline else '',
                'status': {
                    'name': treatment.status.get_name(current_language) if treatment.status else _('Undefined'),
                    'code': treatment.status.code if treatment.status else 'Undefined',
                    'color': treatment.status.color if treatment.status else '#808080'
                },
                'last_modified': {
                    'datetime': format_time(treatment.last_modified),
                    'user': treatment.last_modified_by.get_full_name() if treatment.last_modified_by else ''
                },
                # Treatment Details data
                'residual_risk_level': {
                    'name': treatment.residual_risk_level.get_name(current_language) if treatment.residual_risk_level else '',
                    'color': treatment.residual_risk_level.color if treatment.residual_risk_level else '#6c757d',
                    'exceeds_acceptable': residual_risk_exceeds_acceptable,
                    'acceptable_risk_info': acceptable_risk if residual_risk_exceeds_acceptable else None
                } if treatment.residual_risk_level else None,
                'residual_risk_justification': treatment.residual_risk_justification or '',
                'effectiveness': {
                    'name': treatment.effectiveness.get_name(current_language) if treatment.effectiveness else '',
                    'color': treatment.effectiveness.color if treatment.effectiveness else '#6c757d'
                } if treatment.effectiveness else None,
                'effectiveness_metrics': treatment.effectiveness_metrics or '',
                'effectiveness_evaluation_date': treatment.effectiveness_evaluation_date.strftime('%Y-%m-%d') if treatment.effectiveness_evaluation_date else '',
                'priority': {
                    'name': treatment.priority.get_name(current_language) if treatment.priority else '',
                    'color': treatment.priority.color if treatment.priority else '#6c757d'
                } if treatment.priority else None,
                'priority_justification': treatment.priority_justification or '',
                'monitoring_frequency': {
                    'name': treatment.monitoring_frequency.get_name(current_language) if treatment.monitoring_frequency else ''
                } if treatment.monitoring_frequency else None,
                'next_review_date': treatment.next_review_date.strftime('%Y-%m-%d') if treatment.next_review_date else '',
                'last_review_date': treatment.last_review_date.strftime('%Y-%m-%d') if treatment.last_review_date else '',
                'monitoring_responsible': ', '.join([f"{user.first_name} {user.last_name}".strip() or user.username for user in treatment.monitoring_responsible.all()]),
                'review_notes': treatment.review_notes or '',
                'implementation_cost': str(treatment.implementation_cost) if treatment.implementation_cost else '',
                'annual_maintenance_cost': str(treatment.annual_maintenance_cost) if treatment.annual_maintenance_cost else '',
                'roi_assessment': treatment.roi_assessment or '',
                'prerequisites': treatment.prerequisites or '',
                'dependencies': ', '.join([str(dep) for dep in treatment.dependencies.all()]),
                'affected_assets': ', '.join([asset.name for asset in treatment.affected_assets.all()]),
                # Attachments data
                'attachments': [
                    {
                        'id': attachment.id,
                        'filename': attachment.filename,
                        'file_size': attachment.get_file_size_display(),
                        'file_type': attachment.file_type,
                        'uploaded_by': attachment.uploaded_by.get_full_name() if attachment.uploaded_by else '',
                        'uploaded_at': format_time(attachment.uploaded_at),
                        'description': attachment.description or '',
                        'file_url': attachment.file.url if attachment.file else ''
                    }
                    for attachment in treatment.attachments.all()
                ],
                'attachments_count': treatment.attachments.count()
            }
            data.append(treatment_data)

        response_data = {
            'treatments': data,
            'treatment_types': treatment_types_dict,
            'treatment_statuses': treatment_statuses_dict
        }

        return JsonResponse(response_data)

    except Exception as e:
        print(f"Error processing risk treatment data: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
@log_risk_action("SAVE_RISK_TREATMENTS")
def save_all_risk_treatments(request):
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')
        treatments = data.get('treatments', [])

        # Детальне логування початкових даних
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="SAVE_RISK_TREATMENTS_START",
            details={
                'asset_id': asset_id,
                'treatments_count': len(treatments),
                'treatments_ids': [t.get('id') for t in treatments if t.get('id')]
            },
            asset_id=asset_id,
            request_path=request.path
        )

        # logger.info(f"Received data for asset_id: {asset_id}, treatments: {treatments}")

        entity_type = 'asset'
        entity = None
        treatment_filter = {}
        if str(asset_id).startswith('S'):
            entity_type = 'software'
            sw_id = str(asset_id)[1:]
            if sw_id.isdigit():
                entity = SoftwareRegister.objects.filter(id=int(sw_id)).first()
            treatment_filter = {'software_register': entity}
        elif str(asset_id).startswith('M'):
            entity_type = 'external_media'
            em_id = str(asset_id)[1:]
            if em_id.isdigit():
                entity = ExternalMediaRegister.objects.filter(id=int(em_id)).first()
            treatment_filter = {'external_media_register': entity}
        else:
            entity = InformationAsset.objects.filter(Q(asset_id=asset_id) | Q(id=asset_id)).first()
            treatment_filter = {'asset': entity}

        if not entity:
            logger.error(f"Entity not found for id: {asset_id}")
            RiskAssessmentLogger.log_error(
                user=request.user,
                error_type="ENTITY_NOT_FOUND",
                error_message=f"Entity not found for id: {asset_id}",
                request_path=request.path,
                additional_context={'asset_id': asset_id}
            )
            return JsonResponse({'success': False, 'message': f'Entity not found for id: {asset_id}'}, status=404)

        # logger.info(f"Found asset: {asset}")
        
        # Логування знайденого активу
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="ASSET_FOUND",
            details={
                'asset_name': entity.name,
                'asset_company': entity.company.name if entity.company else None,
                'asset_type': entity.asset_type.get_name() if getattr(entity, 'asset_type', None) else None
            },
            asset_id=asset_id,
            request_path=request.path
        )

        updated_treatments = []

        with transaction.atomic():
            for treatment_data in treatments:
                # logger.info(f"Processing treatment data: {treatment_data}")
                if not treatment_data.get('id'):
                    logger.warning("Skipping treatment without id")
                    continue

                treatment = RiskTreatment.objects.filter(id=treatment_data['id'], **treatment_filter).first()
                if not treatment:
                    logger.warning(f"Treatment not found for id: {treatment_data['id']}")
                    continue

                # logger.info(f"Found treatment: {treatment}")

                # Get treatment_type and status objects from their codes
                treatment_type_code = treatment_data.get('treatment_type')
                status_code = treatment_data.get('status')
                
                treatment_type = None
                if treatment_type_code:
                    treatment_type = Treatment_type.objects.filter(code=treatment_type_code).first()
                    if not treatment_type:
                        treatment_type = Treatment_type.get_by_display_name(treatment_type_code)
                    if not treatment_type:
                        treatment_type = Treatment_type.objects.filter(code='Undefined').first()
                
                status = None
                if status_code:
                    status = Treatment_status.objects.filter(code=status_code).first()
                    if not status:
                        status = Treatment_status.get_by_display_name(status_code)
                    if not status:
                        status = Treatment_status.objects.filter(code='Undefined').first()

                # Check if any fields have changed
                has_changes = False
                
                # Check treatment_type change
                if treatment_type and treatment.treatment_type != treatment_type:
                    has_changes = True
                    treatment.treatment_type = treatment_type
                
                # Check status change
                if status and treatment.status != status:
                    has_changes = True
                    treatment.status = status
                
                # Check other fields
                text_fields = ['description', 'responsible']
                for field in text_fields:
                    if field in treatment_data and getattr(treatment, field) != treatment_data.get(field):
                        has_changes = True
                        setattr(treatment, field, treatment_data.get(field))

                # Special handling for deadline
                if treatment_data.get('deadline'):
                    new_deadline = datetime.strptime(treatment_data['deadline'], '%Y-%m-%d').date()
                    if treatment.deadline != new_deadline:
                        has_changes = True
                        treatment.deadline = new_deadline
                elif treatment.deadline is not None:
                    has_changes = True
                    treatment.deadline = None

                if has_changes:
                    # Логування змін перед збереженням
                    old_data = {
                        'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                        'status': treatment.status.code if treatment.status else None,
                        'description': treatment.description,
                        'responsible': treatment.responsible,
                        'deadline': treatment.deadline.isoformat() if treatment.deadline else None
                    }
                    
                    treatment.last_modified = timezone.now()
                    treatment.last_modified_by = request.user
                    treatment.save(user=request.user)

                    # Логування після збереження
                    new_data = {
                        'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                        'status': treatment.status.code if treatment.status else None,
                        'description': treatment.description,
                        'responsible': treatment.responsible,
                        'deadline': treatment.deadline.isoformat() if treatment.deadline else None
                    }
                    
                    # Розширене логування з додатковими деталями
                    RiskAssessmentLogger.log_data_modification(
                        user=request.user,
                        operation="UPDATE",
                        data_type="RISK_TREATMENT",
                        data_before=old_data,
                        data_after=new_data,
                        asset_id=asset_id,
                        request_path=request.path
                    )
                    
                    # Додаткове логування в базу даних з деталями
                    try:
                        from .models import RiskAssessmentAuditLog
                        
                        with transaction.atomic():
                            RiskAssessmentAuditLog.objects.create(
                                user=request.user,
                                action_type='UPDATE',
                                action_name='RISK_TREATMENT_UPDATE',
                                asset=entity if entity_type == 'asset' else None,
                                object_type='RiskTreatment',
                                object_id=str(treatment.id),
                                ip_address=request.META.get('REMOTE_ADDR'),
                                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                                request_path=request.path,
                                request_method=request.method,
                                data_before=old_data,
                                data_after=new_data,
                                additional_data={
                                    'vulnerability_id': treatment.vulnerability.id,
                                    'vulnerability_name': treatment.vulnerability.get_name(),
                                    'changes_made': list(old_data.keys())
                                },
                                severity='MEDIUM',
                                success=True
                            )
                    except Exception as audit_error:
                        logger.error(f"Failed to create audit log: {str(audit_error)}")

                    # logger.info(f"Updated treatment: {treatment}")

                    updated_treatments.append({
                        'id': treatment.id,
                        'last_modified': format_time(treatment.last_modified),
                        'last_modified_by': treatment.last_modified_by.get_full_name() if treatment.last_modified_by else ''
                    })

            return JsonResponse({
                'success': True,
                'message': f'Successfully updated {len(updated_treatments)} treatments',
                'updated_treatments': updated_treatments
            })

    except Exception as e:
        logger.error(f"Error updating risk treatments: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@user_passes_test(can_edit_risk_assessment)
@require_POST
def update_risk_treatment_data(request):
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')
        new_vulnerabilities = data.get('vulnerabilities', [])
        is_cleared = data.get('is_cleared', False)  # Add this line to get the value
        print(f"Received data: {data}")

        asset = InformationAsset.objects.get(id=asset_id)

        # Отримання всіх поточних загроз, пов'язаних з вразливостями активу
        current_threats = set()
        for av in asset.assetvulnerability_set.all():
            for threat in av.vulnerability.threats.all():
                current_threats.add(threat)

        # Додавання нових вразливостей до таблиці risk treatment
        for vuln_data in new_vulnerabilities:
            vuln_id = vuln_data['id']
            print(f"Processing vulnerability: {vuln_id}")

            vulnerability = Vulnerability.objects.get(id=vuln_id)
            av, created = AssetVulnerability.objects.get_or_create(
                asset=asset,
                vulnerability=vulnerability,
                defaults={'status': 'Yes'}
            )

            # Оновлення загроз
            threats_with_risk_levels = vuln_data.get('threats', [])
            threats = []
            for threat_info in threats_with_risk_levels:
                if '(' in threat_info and ')' in threat_info:
                    threat_name, risk_level_name = threat_info.rsplit('(', 1)
                    risk_level_name = risk_level_name.rstrip(')').strip()

                    threat = Threat.get_by_display_name(threat_name)
                    threat_created = False
                    if not threat:
                        threat = Threat.objects.create(name=threat_name, probability=Decimal('0.5'), impact=Decimal('50'))
                        threat_created = True
                        print(f"New threat created: {threat}")

                    risk_level = RiskLevel.get_by_display_name(risk_level_name, company_id=asset.company_id)
                    if not risk_level:
                        print(f"Risk level '{risk_level_name}' not found, skipping...")
                        continue

                    threats.append(threat)

                av.threats = ", ".join([f"{threat.get_name()}({risk_level_name})" for threat in threats])
                av.save()

                # Оновлення risk treatment
                if is_cleared:  # Now using the is_cleared variable
                    deleted_count = RiskTreatment.objects.filter(asset=asset, vulnerability=vulnerability).delete()[0]
                    print(f"Cleared {deleted_count} existing treatments for asset and vulnerability")

                for threat in threats:
                    print(f"Processing threat: {threat}")
                    if threat in current_threats:
                        print(f"Threat {threat.get_name()} already exists in the database")
                        current_threats.remove(threat)
                    else:
                        print(f"New threat found: {threat.get_name()}")
                        risk_level = RiskLevel.get_by_display_name('Undefined', company_id=asset.company_id) or RiskLevel.get_by_display_name('Невизначено', company_id=asset.company_id)
                        if not risk_level:
                            risk_level = get_company_risk_levels_queryset(asset.company_id).first()
                        treatment, created = RiskTreatment.objects.create(
                            asset=asset,
                            vulnerability=vulnerability,
                            threat=threat,
                            risk_level=risk_level,
                            risk_mitigation_controls=vulnerability.risk_mitigation_controls_uk,
                            treatment_type='Undefined',
                            description='',
                            responsible='',
                            deadline=None,
                            status='Undefined'
                        )
                        print(f"{'Created' if created else 'Updated'} treatment for threat {threat.get_name()}: {treatment}")
                        print("Treatment details:")
                        print(f"  - Risk mitigation controls: {treatment.risk_mitigation_controls}")
                        print(f"  - Treatment type: {treatment.treatment_type}")
                        print(f"  - Description: {treatment.description}")
                        print(f"  - Responsible: {treatment.responsible}")
                        print(f"  - Deadline: {treatment.deadline}")
                        print(f"  - Status: {treatment.status}")

        print(f"Updated risk treatments for asset: {asset_id}")
        return JsonResponse({'success': True, 'message': 'Risk treatment data updated successfully'})

    except Exception as e:
        print(f"Error updating risk treatment data: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@user_passes_test(can_edit_risk_assessment)
@require_POST
def clear_risk_treatments(request):
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')

        if not asset_id:
            return JsonResponse({'success': False, 'message': 'Asset ID is required'}, status=400)

        if str(asset_id).startswith('S'):
            sw_id = str(asset_id)[1:]
            software = SoftwareRegister.objects.filter(id=int(sw_id)).first() if sw_id.isdigit() else None
            if not software:
                return JsonResponse({'success': False, 'message': 'Software entry not found'}, status=404)
            deleted_count, _ = RiskTreatment.objects.filter(software_register=software).delete()
        elif str(asset_id).startswith('M'):
            em_id = str(asset_id)[1:]
            media = ExternalMediaRegister.objects.filter(id=int(em_id)).first() if em_id.isdigit() else None
            if not media:
                return JsonResponse({'success': False, 'message': 'External media entry not found'}, status=404)
            deleted_count, _ = RiskTreatment.objects.filter(external_media_register=media).delete()
        else:
            asset = InformationAsset.objects.filter(Q(asset_id=asset_id) | Q(id=asset_id)).first()
            if not asset:
                return JsonResponse({'success': False, 'message': 'Asset not found'}, status=404)
            deleted_count, _ = RiskTreatment.objects.filter(asset=asset).delete()

        return JsonResponse({
            'success': True,
            'message': f'{deleted_count} risk treatments have been cleared for this asset',
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@require_POST
@log_risk_action("EXPORT_RISK_TREATMENTS")
def export_risk_treatments(request):
    data = json.loads(request.body)
    
    # Логування початку експорту
    RiskAssessmentLogger.log_export_action(
        user=request.user,
        export_type="RISK_TREATMENTS",
        filters={
            'asset_id': data.get('asset_id'),
            'asset_name': data.get('asset_name'),
            'company': data.get('company')
        },
        record_count=len(data.get('treatments', [])),
        request_path=request.path
    )
    treatments = data.get('treatments', [])
    treatment_ids = data.get('treatment_ids', [])
    include_details = data.get('include_details', False)
    current_language = get_language()[:2]
    
    # Create a new workbook and set active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = _("Risk Treatments")
    
    # Helper function to convert hex colors to ARGB format for Excel
    def convert_to_excel_color(hex_color):
        # Remove # if present and ensure a default if None or empty
        if not hex_color or hex_color == 'None':
            return 'FF808080'  # Default gray with full opacity
        
        hex_color = hex_color.lstrip('#')
        
        # Ensure it's a 6-character hex without alpha
        if len(hex_color) == 6:
            # Add full opacity (FF) prefix for Excel ARGB format
            return 'FF' + hex_color
        
        # If somehow we have an 8-character ARGB already, return as is
        if len(hex_color) == 8:
            return hex_color
            
        # For any other case, return default gray
        return 'FF808080'
    
    # Pre-load treatment types and statuses for proper language display
    treatment_types = {t.code: t for t in Treatment_type.objects.all()}
    treatment_statuses = {s.code: s for s in Treatment_status.objects.all()}
    
    # Load detailed treatment data if requested
    treatment_details = {}
    if include_details and treatment_ids:
        from .models import RiskTreatment
        detailed_treatments = RiskTreatment.objects.filter(id__in=treatment_ids).select_related(
            'residual_risk_level', 'effectiveness', 'priority', 'monitoring_frequency'
        ).prefetch_related('monitoring_responsible', 'dependencies', 'affected_assets')
        
        for treatment in detailed_treatments:
            treatment_details[treatment.id] = {
                'residual_risk_level': treatment.residual_risk_level.get_name(current_language) if treatment.residual_risk_level else '',
                'residual_risk_justification': treatment.residual_risk_justification or '',
                'effectiveness': treatment.effectiveness.get_name(current_language) if treatment.effectiveness else '',
                'effectiveness_metrics': treatment.effectiveness_metrics or '',
                'effectiveness_evaluation_date': treatment.effectiveness_evaluation_date.strftime('%Y-%m-%d') if treatment.effectiveness_evaluation_date else '',
                'priority': treatment.priority.get_name(current_language) if treatment.priority else '',
                'priority_justification': treatment.priority_justification or '',
                'monitoring_frequency': treatment.monitoring_frequency.get_name(current_language) if treatment.monitoring_frequency else '',
                'next_review_date': treatment.next_review_date.strftime('%Y-%m-%d') if treatment.next_review_date else '',
                'last_review_date': treatment.last_review_date.strftime('%Y-%m-%d') if treatment.last_review_date else '',
                'monitoring_responsible': ', '.join([f"{user.first_name} {user.last_name}".strip() or user.username for user in treatment.monitoring_responsible.all()]),
                'review_notes': treatment.review_notes or '',
                'implementation_cost': str(treatment.implementation_cost) if treatment.implementation_cost else '',
                'annual_maintenance_cost': str(treatment.annual_maintenance_cost) if treatment.annual_maintenance_cost else '',
                'roi_assessment': treatment.roi_assessment or '',
                'prerequisites': treatment.prerequisites or '',
                'dependencies': ', '.join([str(dep) for dep in treatment.dependencies.all()]),
                'affected_assets': ', '.join([asset.name for asset in treatment.affected_assets.all()]),
                'residual_risk_color': treatment.residual_risk_level.color if treatment.residual_risk_level else '#808080',
                'effectiveness_color': treatment.effectiveness.color if treatment.effectiveness else '#808080',
                'priority_color': treatment.priority.color if treatment.priority else '#808080'
            }
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF0366D6', end_color='FF0366D6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='E1E1E1'),
        right=Side(style='thin', color='E1E1E1'),
        top=Side(style='thin', color='E1E1E1'),
        bottom=Side(style='thin', color='E1E1E1')
    )
    
    row_fills = [
        PatternFill(start_color='FFF8F9FA', end_color='FFF8F9FA', fill_type='solid'),
        PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    ]
    
    # Status colors (now in ARGB format)
    status_colors = {
        'Planned': 'FFFFC107',       # Yellow
        'In Progress': 'FF17A2B8',   # Blue
        'Completed': 'FF28A745',     # Green
        'Undefined': 'FF6C757D'      # Gray
    }
    
    # Define headers
    if include_details:
        headers = [
            _('Asset ID'), _('Asset Name'), _('Company'), _('Criticality'), 
            _('Vulnerability'), _('Threats and Risk Levels'), _('Risk Level'), 
            _('Risk Mitigation Controls'), _('Treatment Type'), _('Description & Responsible'), 
            _('Deadline'), _('Status'), _('Treatment Details'),
            # Treatment Details columns
            _('Residual Risk Level'), _('Residual Risk Justification'),
            _('Effectiveness'), _('Effectiveness Metrics'), _('Effectiveness Evaluation Date'),
            _('Priority'), _('Priority Justification'),
            _('Monitoring Frequency'), _('Next Review Date'), _('Last Review Date'),
            _('Monitoring Responsible'), _('Review Notes'),
            _('Implementation Cost'), _('Annual Maintenance Cost'), _('ROI Assessment'),
            _('Prerequisites'), _('Dependencies'), _('Affected Assets')
        ]
    else:
        headers = [
            _('Asset ID'), _('Asset Name'), _('Company'), _('Criticality'), 
            _('Vulnerability'), _('Threats and Risk Levels'), _('Risk Level'), 
            _('Risk Mitigation Controls'), _('Treatment Type'), _('Description & Responsible'), 
            _('Deadline'), _('Status'), _('Treatment Details')
        ]
    
    # Set headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Auto-filter
    if include_details:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    else:
        ws.auto_filter.ref = f"A1:M1"
    
    # Fill data
    for row_idx, treatment in enumerate(treatments, 2):
        fill = row_fills[(row_idx - 2) % 2]
        
        # Get localized treatment type and status names
        treatment_type_code = treatment['treatment_type'].get('code') if isinstance(treatment['treatment_type'], dict) else treatment['treatment_type']
        status_code = treatment['status'].get('code') if isinstance(treatment['status'], dict) else treatment['status']
        
        treatment_type_name = treatment['treatment_type'].get('name', '')
        status_name = treatment['status'].get('name', '')
        
        # If we have codes, try to get localized names from our models
        if treatment_type_code and treatment_type_code in treatment_types:
            treatment_type_name = treatment_types[treatment_type_code].get_name(current_language)
        
        if status_code and status_code in treatment_statuses:
            status_name = treatment_statuses[status_code].get_name(current_language)
        
        # Base row data
        row_data = [
            treatment['asset_id'],
            treatment['asset_name'],
            treatment['company'],
            treatment.get('criticality', {}).get('name', ''),  # Handle missing criticality gracefully
            treatment['vulnerability'],
            treatment['threats_and_risk_levels'],
            treatment['highest_risk_level']['name'],
            treatment['risk_mitigation_controls'],
            treatment_type_name,
            (treatment['description'] or '') + ('\n\nResponsible: ' + treatment['responsible'] if treatment['responsible'] else ''),
            treatment['deadline'],
            status_name,
            f"Treatment Details (Last Modified: {treatment['last_modified']['datetime']} by {treatment['last_modified']['user']})" if treatment.get('last_modified', {}).get('datetime') else 'Treatment Details'
        ]
        
        # Add treatment details if requested
        if include_details:
            treatment_id = treatment['id']
            details = treatment_details.get(treatment_id, {})
            
            row_data.extend([
                details.get('residual_risk_level', ''),
                details.get('residual_risk_justification', ''),
                details.get('effectiveness', ''),
                details.get('effectiveness_metrics', ''),
                details.get('effectiveness_evaluation_date', ''),
                details.get('priority', ''),
                details.get('priority_justification', ''),
                details.get('monitoring_frequency', ''),
                details.get('next_review_date', ''),
                details.get('last_review_date', ''),
                details.get('monitoring_responsible', ''),
                details.get('review_notes', ''),
                details.get('implementation_cost', ''),
                details.get('annual_maintenance_cost', ''),
                details.get('roi_assessment', ''),
                details.get('prerequisites', ''),
                details.get('dependencies', ''),
                details.get('affected_assets', '')
            ])
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.fill = fill
            
            # Apply special formatting
            if col_idx == 4:  # Criticality column
                criticality_data = treatment.get('criticality', {})
                if 'color' in criticality_data and criticality_data['color']:
                    color = convert_to_excel_color(criticality_data['color'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if criticality_data['color'].startswith('#'):
                        hex_color = criticality_data['color'].lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            elif col_idx == 7:  # Risk Level column
                if 'color' in treatment['highest_risk_level']:
                    color = convert_to_excel_color(treatment['highest_risk_level']['color'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if treatment['highest_risk_level']['color'] and treatment['highest_risk_level']['color'].startswith('#'):
                        hex_color = treatment['highest_risk_level']['color'].lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            elif col_idx == 9:  # Treatment Type column
                # Get color from model if possible
                treatment_color = None
                if treatment_type_code and treatment_type_code in treatment_types:
                    treatment_color = treatment_types[treatment_type_code].color
                elif 'color' in treatment['treatment_type']:
                    treatment_color = treatment['treatment_type']['color']
                
                if treatment_color:
                    color = convert_to_excel_color(treatment_color)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if treatment_color.startswith('#'):
                        hex_color = treatment_color.lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            elif col_idx == 14:  # Last Modified column
                # Format the last modified cell with better readability
                if value and '|' in str(value):
                    datetime_part, user_part = str(value).split('|', 1)
                    cell.value = f"{datetime_part.strip()}\n{user_part.strip()}"
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            elif col_idx == 13:  # Status column
                # Get color from model if possible
                status_color = None
                if status_code and status_code in treatment_statuses:
                    status_color = treatment_statuses[status_code].color
                elif 'color' in treatment['status']:
                    status_color = treatment['status']['color']
                
                if status_color:
                    color = convert_to_excel_color(status_color)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if status_color.startswith('#'):
                        hex_color = status_color.lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            # Treatment Details formatting (if include_details is True)
            elif include_details:
                treatment_id = treatment['id']
                details = treatment_details.get(treatment_id, {})
                
                if col_idx == 15:  # Residual Risk Level column
                    risk_color = details.get('residual_risk_color')
                    if risk_color:
                        color = convert_to_excel_color(risk_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        if risk_color.startswith('#'):
                            hex_color = risk_color.lstrip('#')
                            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                
                elif col_idx == 17:  # Effectiveness column
                    eff_color = details.get('effectiveness_color')
                    if eff_color:
                        color = convert_to_excel_color(eff_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        if eff_color.startswith('#'):
                            hex_color = eff_color.lstrip('#')
                            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                
                elif col_idx == 20:  # Priority column
                    priority_color = details.get('priority_color')
                    if priority_color:
                        color = convert_to_excel_color(priority_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        if priority_color.startswith('#'):
                            hex_color = priority_color.lstrip('#')
                            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
    
    # Adjust column widths
    if include_details:
        column_widths = [
            12, 20, 20, 15, 30, 25, 15, 30, 15, 30, 20, 15, 15, 25,  # Base columns
            # Treatment Details columns
            15, 30,  # Residual Risk Level, Justification
            15, 30, 15,  # Effectiveness, Metrics, Evaluation Date
            15, 30,  # Priority, Justification
            15, 15, 15,  # Monitoring Frequency, Next Review, Last Review
            25, 30,  # Monitoring Responsible, Review Notes
            15, 15, 30,  # Implementation Cost, Annual Cost, ROI Assessment
            30, 25, 25   # Prerequisites, Dependencies, Affected Assets
        ]
    else:
        column_widths = [12, 20, 20, 15, 30, 25, 15, 30, 15, 30, 20, 15, 15, 25]
    
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="risk_treatments.xlsx"'
    
    # Save the workbook to response
    wb.save(response)
    
    return response


@login_required
@require_POST
@log_risk_action("SAVE_RISK_TREATMENTS")
def save_all_risk_treatments(request):
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')
        treatments = data.get('treatments', [])

        # Детальне логування початкових даних
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="SAVE_RISK_TREATMENTS_START",
            details={
                'asset_id': asset_id,
                'treatments_count': len(treatments),
                'treatments_ids': [t.get('id') for t in treatments if t.get('id')]
            },
            asset_id=asset_id,
            request_path=request.path
        )

        # logger.info(f"Received data for asset_id: {asset_id}, treatments: {treatments}")

        entity_type = 'asset'
        entity = None
        treatment_filter = {}
        if str(asset_id).startswith('S'):
            entity_type = 'software'
            sw_id = str(asset_id)[1:]
            if sw_id.isdigit():
                entity = SoftwareRegister.objects.filter(id=int(sw_id)).first()
            treatment_filter = {'software_register': entity}
        elif str(asset_id).startswith('M'):
            entity_type = 'external_media'
            em_id = str(asset_id)[1:]
            if em_id.isdigit():
                entity = ExternalMediaRegister.objects.filter(id=int(em_id)).first()
            treatment_filter = {'external_media_register': entity}
        else:
            entity = InformationAsset.objects.filter(Q(asset_id=asset_id) | Q(id=asset_id)).first()
            treatment_filter = {'asset': entity}
        if not entity:
            logger.error(f"Entity not found for id: {asset_id}")
            RiskAssessmentLogger.log_error(
                user=request.user,
                error_type="ENTITY_NOT_FOUND",
                error_message=f"Entity not found for id: {asset_id}",
                request_path=request.path,
                additional_context={'asset_id': asset_id}
            )
            return JsonResponse({'success': False, 'message': f'Entity not found for id: {asset_id}'}, status=404)

        # logger.info(f"Found asset: {asset}")
        
        # Логування знайденого активу
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="ASSET_FOUND",
            details={
                'asset_name': entity.name,
                'asset_company': entity.company.name if entity.company else None,
                'asset_type': entity.asset_type.get_name() if getattr(entity, 'asset_type', None) else None
            },
            asset_id=asset_id,
            request_path=request.path
        )

        updated_treatments = []

        with transaction.atomic():
            for treatment_data in treatments:
                # logger.info(f"Processing treatment data: {treatment_data}")
                if not treatment_data.get('id'):
                    logger.warning("Skipping treatment without id")
                    continue

                treatment = RiskTreatment.objects.filter(id=treatment_data['id'], **treatment_filter).first()
                if not treatment:
                    logger.warning(f"Treatment not found for id: {treatment_data['id']}")
                    continue

                # logger.info(f"Found treatment: {treatment}")

                # Get treatment_type and status objects from their codes
                treatment_type_code = treatment_data.get('treatment_type')
                status_code = treatment_data.get('status')
                
                treatment_type = None
                if treatment_type_code:
                    treatment_type = Treatment_type.objects.filter(code=treatment_type_code).first()
                    if not treatment_type:
                        treatment_type = Treatment_type.get_by_display_name(treatment_type_code)
                    if not treatment_type:
                        treatment_type = Treatment_type.objects.filter(code='Undefined').first()
                
                status = None
                if status_code:
                    status = Treatment_status.objects.filter(code=status_code).first()
                    if not status:
                        status = Treatment_status.get_by_display_name(status_code)
                    if not status:
                        status = Treatment_status.objects.filter(code='Undefined').first()

                # Check if any fields have changed
                has_changes = False
                
                # Check treatment_type change
                if treatment_type and treatment.treatment_type != treatment_type:
                    has_changes = True
                    treatment.treatment_type = treatment_type
                
                # Check status change
                if status and treatment.status != status:
                    has_changes = True
                    treatment.status = status
                
                # Check other fields
                text_fields = ['description', 'responsible']
                for field in text_fields:
                    if field in treatment_data and getattr(treatment, field) != treatment_data.get(field):
                        has_changes = True
                        setattr(treatment, field, treatment_data.get(field))

                # Special handling for deadline
                if treatment_data.get('deadline'):
                    new_deadline = datetime.strptime(treatment_data['deadline'], '%Y-%m-%d').date()
                    if treatment.deadline != new_deadline:
                        has_changes = True
                        treatment.deadline = new_deadline
                elif treatment.deadline is not None:
                    has_changes = True
                    treatment.deadline = None

                if has_changes:
                    # Логування змін перед збереженням
                    old_data = {
                        'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                        'status': treatment.status.code if treatment.status else None,
                        'description': treatment.description,
                        'responsible': treatment.responsible,
                        'deadline': treatment.deadline.isoformat() if treatment.deadline else None
                    }
                    
                    treatment.last_modified = timezone.now()
                    treatment.last_modified_by = request.user
                    treatment.save(user=request.user)

                    # Логування після збереження
                    new_data = {
                        'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                        'status': treatment.status.code if treatment.status else None,
                        'description': treatment.description,
                        'responsible': treatment.responsible,
                        'deadline': treatment.deadline.isoformat() if treatment.deadline else None
                    }
                    
                    # Розширене логування з додатковими деталями
                    RiskAssessmentLogger.log_data_modification(
                        user=request.user,
                        operation="UPDATE",
                        data_type="RISK_TREATMENT",
                        data_before=old_data,
                        data_after=new_data,
                        asset_id=asset_id,
                        request_path=request.path
                    )
                    
                    # Додаткове логування в базу даних з деталями
                    try:
                        from .models import RiskAssessmentAuditLog
                        
                        with transaction.atomic():
                            RiskAssessmentAuditLog.objects.create(
                                user=request.user,
                                action_type='UPDATE',
                                action_name='RISK_TREATMENT_UPDATE',
                                asset=entity if entity_type == 'asset' else None,
                                object_type='RiskTreatment',
                                object_id=str(treatment.id),
                                ip_address=request.META.get('REMOTE_ADDR'),
                                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                                request_path=request.path,
                                request_method=request.method,
                                data_before=old_data,
                                data_after=new_data,
                                additional_data={
                                    'vulnerability_id': treatment.vulnerability.id,
                                    'vulnerability_name': treatment.vulnerability.get_name(),
                                    'changes_made': list(old_data.keys())
                                },
                                severity='MEDIUM',
                                success=True
                            )
                    except Exception as audit_error:
                        logger.error(f"Failed to create audit log: {str(audit_error)}")

                    # logger.info(f"Updated treatment: {treatment}")

                    updated_treatments.append({
                        'id': treatment.id,
                        'last_modified': format_time(treatment.last_modified),
                        'last_modified_by': treatment.last_modified_by.get_full_name() if treatment.last_modified_by else ''
                    })

            return JsonResponse({
                'success': True,
                'message': f'Successfully updated {len(updated_treatments)} treatments',
                'updated_treatments': updated_treatments
            })

    except Exception as e:
        logger.error(f"Error updating risk treatments: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required
@user_passes_test(can_edit_risk_assessment)
@require_POST
def clear_risk_treatments(request):
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')

        if not asset_id:
            return JsonResponse({'success': False, 'message': 'Asset ID is required'}, status=400)

        if str(asset_id).startswith('S'):
            sw_id = str(asset_id)[1:]
            software = SoftwareRegister.objects.filter(id=int(sw_id)).first() if sw_id.isdigit() else None
            if not software:
                return JsonResponse({'success': False, 'message': 'Software entry not found'}, status=404)
            deleted_count, _ = RiskTreatment.objects.filter(software_register=software).delete()
        elif str(asset_id).startswith('M'):
            em_id = str(asset_id)[1:]
            media = ExternalMediaRegister.objects.filter(id=int(em_id)).first() if em_id.isdigit() else None
            if not media:
                return JsonResponse({'success': False, 'message': 'External media entry not found'}, status=404)
            deleted_count, _ = RiskTreatment.objects.filter(external_media_register=media).delete()
        else:
            asset = InformationAsset.objects.filter(Q(asset_id=asset_id) | Q(id=asset_id)).first()
            if not asset:
                return JsonResponse({'success': False, 'message': 'Asset not found'}, status=404)
            deleted_count, _ = RiskTreatment.objects.filter(asset=asset).delete()

        return JsonResponse({
            'success': True,
            'message': f'{deleted_count} risk treatments have been cleared for this asset',
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@require_POST
@log_risk_action("EXPORT_RISK_TREATMENTS")
def export_risk_treatments(request):
    data = json.loads(request.body)
    
    # Логування початку експорту
    RiskAssessmentLogger.log_export_action(
        user=request.user,
        export_type="RISK_TREATMENTS",
        filters={
            'asset_id': data.get('asset_id'),
            'asset_name': data.get('asset_name'),
            'company': data.get('company')
        },
        record_count=len(data.get('treatments', [])),
        request_path=request.path
    )
    treatments = data.get('treatments', [])
    treatment_ids = data.get('treatment_ids', [])
    include_details = data.get('include_details', False)
    current_language = get_language()[:2]
    
    # Create a new workbook and set active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = _("Risk Treatments")
    
    # Helper function to convert hex colors to ARGB format for Excel
    def convert_to_excel_color(hex_color):
        # Remove # if present and ensure a default if None or empty
        if not hex_color or hex_color == 'None':
            return 'FF808080'  # Default gray with full opacity
        
        hex_color = hex_color.lstrip('#')
        
        # Ensure it's a 6-character hex without alpha
        if len(hex_color) == 6:
            # Add full opacity (FF) prefix for Excel ARGB format
            return 'FF' + hex_color
        
        # If somehow we have an 8-character ARGB already, return as is
        if len(hex_color) == 8:
            return hex_color
            
        # For any other case, return default gray
        return 'FF808080'
    
    # Pre-load treatment types and statuses for proper language display
    treatment_types = {t.code: t for t in Treatment_type.objects.all()}
    treatment_statuses = {s.code: s for s in Treatment_status.objects.all()}
    
    # Load detailed treatment data if requested
    treatment_details = {}
    if include_details and treatment_ids:
        from .models import RiskTreatment
        detailed_treatments = RiskTreatment.objects.filter(id__in=treatment_ids).select_related(
            'residual_risk_level', 'effectiveness', 'priority', 'monitoring_frequency'
        ).prefetch_related('monitoring_responsible', 'dependencies', 'affected_assets')
        
        for treatment in detailed_treatments:
            treatment_details[treatment.id] = {
                'residual_risk_level': treatment.residual_risk_level.get_name(current_language) if treatment.residual_risk_level else '',
                'residual_risk_justification': treatment.residual_risk_justification or '',
                'effectiveness': treatment.effectiveness.get_name(current_language) if treatment.effectiveness else '',
                'effectiveness_metrics': treatment.effectiveness_metrics or '',
                'effectiveness_evaluation_date': treatment.effectiveness_evaluation_date.strftime('%Y-%m-%d') if treatment.effectiveness_evaluation_date else '',
                'priority': treatment.priority.get_name(current_language) if treatment.priority else '',
                'priority_justification': treatment.priority_justification or '',
                'monitoring_frequency': treatment.monitoring_frequency.get_name(current_language) if treatment.monitoring_frequency else '',
                'next_review_date': treatment.next_review_date.strftime('%Y-%m-%d') if treatment.next_review_date else '',
                'last_review_date': treatment.last_review_date.strftime('%Y-%m-%d') if treatment.last_review_date else '',
                'monitoring_responsible': ', '.join([f"{user.first_name} {user.last_name}".strip() or user.username for user in treatment.monitoring_responsible.all()]),
                'review_notes': treatment.review_notes or '',
                'implementation_cost': str(treatment.implementation_cost) if treatment.implementation_cost else '',
                'annual_maintenance_cost': str(treatment.annual_maintenance_cost) if treatment.annual_maintenance_cost else '',
                'roi_assessment': treatment.roi_assessment or '',
                'prerequisites': treatment.prerequisites or '',
                'dependencies': ', '.join([str(dep) for dep in treatment.dependencies.all()]),
                'affected_assets': ', '.join([asset.name for asset in treatment.affected_assets.all()]),
                'residual_risk_color': treatment.residual_risk_level.color if treatment.residual_risk_level else '#808080',
                'effectiveness_color': treatment.effectiveness.color if treatment.effectiveness else '#808080',
                'priority_color': treatment.priority.color if treatment.priority else '#808080'
            }
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF0366D6', end_color='FF0366D6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='E1E1E1'),
        right=Side(style='thin', color='E1E1E1'),
        top=Side(style='thin', color='E1E1E1'),
        bottom=Side(style='thin', color='E1E1E1')
    )
    
    row_fills = [
        PatternFill(start_color='FFF8F9FA', end_color='FFF8F9FA', fill_type='solid'),
        PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    ]
    
    # Status colors (now in ARGB format)
    status_colors = {
        'Planned': 'FFFFC107',       # Yellow
        'In Progress': 'FF17A2B8',   # Blue
        'Completed': 'FF28A745',     # Green
        'Undefined': 'FF6C757D'      # Gray
    }
    
    # Define headers
    if include_details:
        headers = [
            _('Asset ID'), _('Asset Name'), _('Company'), _('Criticality'), 
            _('Vulnerability'), _('Threats and Risk Levels'), _('Risk Level'), 
            _('Risk Mitigation Controls'), _('Treatment Type'), _('Description & Responsible'), 
            _('Deadline'), _('Status'), _('Treatment Details'),
            # Treatment Details columns
            _('Residual Risk Level'), _('Residual Risk Justification'),
            _('Effectiveness'), _('Effectiveness Metrics'), _('Effectiveness Evaluation Date'),
            _('Priority'), _('Priority Justification'),
            _('Monitoring Frequency'), _('Next Review Date'), _('Last Review Date'),
            _('Monitoring Responsible'), _('Review Notes'),
            _('Implementation Cost'), _('Annual Maintenance Cost'), _('ROI Assessment'),
            _('Prerequisites'), _('Dependencies'), _('Affected Assets')
        ]
    else:
        headers = [
            _('Asset ID'), _('Asset Name'), _('Company'), _('Criticality'), 
            _('Vulnerability'), _('Threats and Risk Levels'), _('Risk Level'), 
            _('Risk Mitigation Controls'), _('Treatment Type'), _('Description & Responsible'), 
            _('Deadline'), _('Status'), _('Treatment Details')
        ]
    
    # Set headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Auto-filter
    if include_details:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    else:
        ws.auto_filter.ref = f"A1:M1"
    
    # Fill data
    for row_idx, treatment in enumerate(treatments, 2):
        fill = row_fills[(row_idx - 2) % 2]
        
        # Get localized treatment type and status names
        treatment_type_code = treatment['treatment_type'].get('code') if isinstance(treatment['treatment_type'], dict) else treatment['treatment_type']
        status_code = treatment['status'].get('code') if isinstance(treatment['status'], dict) else treatment['status']
        
        treatment_type_name = treatment['treatment_type'].get('name', '')
        status_name = treatment['status'].get('name', '')
        
        # If we have codes, try to get localized names from our models
        if treatment_type_code and treatment_type_code in treatment_types:
            treatment_type_name = treatment_types[treatment_type_code].get_name(current_language)
        
        if status_code and status_code in treatment_statuses:
            status_name = treatment_statuses[status_code].get_name(current_language)
        
        # Base row data
        row_data = [
            treatment['asset_id'],
            treatment['asset_name'],
            treatment['company'],
            treatment.get('criticality', {}).get('name', ''),  # Handle missing criticality gracefully
            treatment['vulnerability'],
            treatment['threats_and_risk_levels'],
            treatment['highest_risk_level']['name'],
            treatment['risk_mitigation_controls'],
            treatment_type_name,
            (treatment['description'] or '') + ('\n\nResponsible: ' + treatment['responsible'] if treatment['responsible'] else ''),
            treatment['deadline'],
            status_name,
            f"Treatment Details (Last Modified: {treatment['last_modified']['datetime']} by {treatment['last_modified']['user']})" if treatment.get('last_modified', {}).get('datetime') else 'Treatment Details'
        ]
        
        # Add treatment details if requested
        if include_details:
            treatment_id = treatment['id']
            details = treatment_details.get(treatment_id, {})
            
            row_data.extend([
                details.get('residual_risk_level', ''),
                details.get('residual_risk_justification', ''),
                details.get('effectiveness', ''),
                details.get('effectiveness_metrics', ''),
                details.get('effectiveness_evaluation_date', ''),
                details.get('priority', ''),
                details.get('priority_justification', ''),
                details.get('monitoring_frequency', ''),
                details.get('next_review_date', ''),
                details.get('last_review_date', ''),
                details.get('monitoring_responsible', ''),
                details.get('review_notes', ''),
                details.get('implementation_cost', ''),
                details.get('annual_maintenance_cost', ''),
                details.get('roi_assessment', ''),
                details.get('prerequisites', ''),
                details.get('dependencies', ''),
                details.get('affected_assets', '')
            ])
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.fill = fill
            
            # Apply special formatting
            if col_idx == 4:  # Criticality column
                criticality_data = treatment.get('criticality', {})
                if 'color' in criticality_data and criticality_data['color']:
                    color = convert_to_excel_color(criticality_data['color'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if criticality_data['color'].startswith('#'):
                        hex_color = criticality_data['color'].lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            elif col_idx == 7:  # Risk Level column
                if 'color' in treatment['highest_risk_level']:
                    color = convert_to_excel_color(treatment['highest_risk_level']['color'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if treatment['highest_risk_level']['color'] and treatment['highest_risk_level']['color'].startswith('#'):
                        hex_color = treatment['highest_risk_level']['color'].lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            elif col_idx == 9:  # Treatment Type column
                # Get color from model if possible
                treatment_color = None
                if treatment_type_code and treatment_type_code in treatment_types:
                    treatment_color = treatment_types[treatment_type_code].color
                elif 'color' in treatment['treatment_type']:
                    treatment_color = treatment['treatment_type']['color']
                
                if treatment_color:
                    color = convert_to_excel_color(treatment_color)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if treatment_color.startswith('#'):
                        hex_color = treatment_color.lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            elif col_idx == 14:  # Last Modified column
                # Format the last modified cell with better readability
                if value and '|' in str(value):
                    datetime_part, user_part = str(value).split('|', 1)
                    cell.value = f"{datetime_part.strip()}\n{user_part.strip()}"
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            elif col_idx == 13:  # Status column
                # Get color from model if possible
                status_color = None
                if status_code and status_code in treatment_statuses:
                    status_color = treatment_statuses[status_code].color
                elif 'color' in treatment['status']:
                    status_color = treatment['status']['color']
                
                if status_color:
                    color = convert_to_excel_color(status_color)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # Set font color based on background brightness
                    if status_color.startswith('#'):
                        hex_color = status_color.lstrip('#')
                        r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    else:
                        cell.font = Font(color='FFFFFF', bold=True)
            
            # Treatment Details formatting (if include_details is True)
            elif include_details:
                treatment_id = treatment['id']
                details = treatment_details.get(treatment_id, {})
                
                if col_idx == 15:  # Residual Risk Level column
                    risk_color = details.get('residual_risk_color')
                    if risk_color:
                        color = convert_to_excel_color(risk_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        if risk_color.startswith('#'):
                            hex_color = risk_color.lstrip('#')
                            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                
                elif col_idx == 17:  # Effectiveness column
                    eff_color = details.get('effectiveness_color')
                    if eff_color:
                        color = convert_to_excel_color(eff_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        if eff_color.startswith('#'):
                            hex_color = eff_color.lstrip('#')
                            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                
                elif col_idx == 20:  # Priority column
                    priority_color = details.get('priority_color')
                    if priority_color:
                        color = convert_to_excel_color(priority_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        if priority_color.startswith('#'):
                            hex_color = priority_color.lstrip('#')
                            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
    
    # Adjust column widths
    if include_details:
        column_widths = [
            12, 20, 20, 15, 30, 25, 15, 30, 15, 30, 20, 15, 15, 25,  # Base columns
            # Treatment Details columns
            15, 30,  # Residual Risk Level, Justification
            15, 30, 15,  # Effectiveness, Metrics, Evaluation Date
            15, 30,  # Priority, Justification
            15, 15, 15,  # Monitoring Frequency, Next Review, Last Review
            25, 30,  # Monitoring Responsible, Review Notes
            15, 15, 30,  # Implementation Cost, Annual Cost, ROI Assessment
            30, 25, 25   # Prerequisites, Dependencies, Affected Assets
        ]
    else:
        column_widths = [12, 20, 20, 15, 30, 25, 15, 30, 15, 30, 20, 15, 15, 25]
    
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="risk_treatments.xlsx"'
    
    # Save the workbook to response
    wb.save(response)
    
    return response


@user_passes_test(has_risk_assessment_access)
@require_http_methods(["GET", "POST"])
def risk_assessment_dashboard(request):
    try:
        # Asset Vulnerability Overview
        # logger.info("Starting risk assessment dashboard data collection")

        # logger.info("Collecting vulnerability data")
        vulnerability_data = AssetVulnerability.objects.aggregate(
            vulnerable=Count(Case(When(status='Yes', then=1))),
            not_vulnerable=Count(Case(When(status='No', then=1))),
            undefined=Count(Case(When(status='Undefined', then=1)))
        )

        if vulnerability_data is None:
            vulnerability_data = {'vulnerable': 0, 'not_vulnerable': 0, 'undefined': 0}


        # logger.info(f"Vulnerability data: {vulnerability_data}")

        # logger.info("Collecting criticality data")
        # Criticality Chart - CriticalityLevel has name and name_local (no critical_name_uk/en/ru)
        current_language = get_language()[:2]
        default_name = 'Невизначено' if current_language == 'uk' else 'Undefined' if current_language == 'en' else 'Неопределено'

        criticality_data = InformationAsset.objects.annotate(
            criticality_name=Case(
                When(confidentiality__cost__gt=F('integrity__cost'), then=Coalesce(F('confidentiality__name_local'), F('confidentiality__name'))),
                When(confidentiality__cost__gt=F('availability__cost'), then=Coalesce(F('confidentiality__name_local'), F('confidentiality__name'))),
                When(integrity__cost__gt=F('confidentiality__cost'), then=Coalesce(F('integrity__name_local'), F('integrity__name'))),
                When(integrity__cost__gt=F('availability__cost'), then=Coalesce(F('integrity__name_local'), F('integrity__name'))),
                When(availability__cost__gt=F('confidentiality__cost'), then=Coalesce(F('availability__name_local'), F('availability__name'))),
                When(availability__cost__gt=F('integrity__cost'), then=Coalesce(F('availability__name_local'), F('availability__name'))),
                default=Value(default_name),
                output_field=CharField()
            ),
            color=Case(
                When(confidentiality__cost__gt=F('integrity__cost'), then=F('confidentiality__color')),
                When(confidentiality__cost__gt=F('availability__cost'), then=F('confidentiality__color')),
                When(integrity__cost__gt=F('confidentiality__cost'), then=F('integrity__color')),
                When(integrity__cost__gt=F('availability__cost'), then=F('integrity__color')),
                When(availability__cost__gt=F('confidentiality__cost'), then=F('availability__color')),
                When(availability__cost__gt=F('integrity__cost'), then=F('availability__color')),
                default=Value('#808080'),  # Default color for undefined
                output_field=CharField()
            )
        ).values('criticality_name', 'color').annotate(
            count=Count('id')
        ).order_by('criticality_name')
        # logger.info(f"Criticality data: {list(criticality_data)}")
        risk_matrix_data = []

        # 1. Get AssetVulnerability Data
        asset_vulnerabilities = AssetVulnerability.objects.filter(status='Yes').select_related(
            'asset__confidentiality', 'asset__integrity', 'asset__availability'
        ).prefetch_related(
            'vulnerability__threats'
        )

        # 2. Calculate Risk Scores and get all defined risk levels
        risk_levels_counts = {}
        current_language = get_language()[:2]
        
        # Calculate actual risk counts from data
        for av in asset_vulnerabilities:
            value_of_risk = calculate_value_of_risk(av.asset, av.vulnerability)
            risk_level = calculate_risk_level(value_of_risk, company_id=av.asset.company_id)
            if risk_level:
                risk_level_name = risk_level.get_name_by_language(current_language) or risk_level.name or risk_level.name_local or ''
                risk_levels_counts[risk_level_name] = risk_levels_counts.get(risk_level_name, 0) + 1

        # Get all defined risk levels from admin panel
        all_risk_levels = get_company_risk_levels_queryset(None)
        
        # Create risk matrix data with all defined risk levels
        risk_matrix_data = []
        for risk_level in all_risk_levels:
            risk_level_name = risk_level.get_name_by_language(current_language) or risk_level.name or risk_level.name_local or ''
            count = risk_levels_counts.get(risk_level_name, 0)
            risk_matrix_data.append({
                'probability': risk_level_name,
                'impact': risk_level_name,
                'count': count,
                'color': risk_level.color,
                'min_value': risk_level.min_value,
                'max_value': risk_level.max_value
            })
        # print('risk_matrix_data =', risk_matrix_data)
        
        # Get all available risk levels for reference
        all_risk_levels_data = []
        for risk_level in all_risk_levels:
            all_risk_levels_data.append({
                'id': risk_level.id,
                'name': risk_level.get_name_by_language(current_language) or risk_level.name or risk_level.name_local or '',
                'color': risk_level.color,
                'min_value': risk_level.min_value,
                'max_value': risk_level.max_value
            })
        
        # Risk Treatment Progress Tracker
        # Modified to work with the new model relationships
        treatment_data = []
        current_language = get_language()[:2]
        
        # Get all treatment types (active only)
        treatment_types = Treatment_type.objects.filter(is_active=True).order_by('name', 'code')
        
        for treatment_type in treatment_types:
            treatment_count = {
                'treatment_type': treatment_type.get_name(current_language),
                'treatment_type_code': treatment_type.code,
                'planned': RiskTreatment.objects.filter(
                    treatment_type=treatment_type,
                    status__code='Planned'
                ).count(),
                'in_progress': RiskTreatment.objects.filter(
                    treatment_type=treatment_type,
                    status__code='In Progress'
                ).count(),
                'completed': RiskTreatment.objects.filter(
                    treatment_type=treatment_type,
                    status__code='Completed'
                ).count()
            }
            treatment_data.append(treatment_count)
            
        # logger.info("Dashboard data collection completed successfully")
        
        # Treatment Details data collection
        # logger.info("Collecting Treatment Details data")
        
        # Effectiveness data
        effectiveness_data = []
        effectiveness_levels = TreatmentEffectiveness.objects.all()
        for effectiveness in effectiveness_levels:
            count = RiskTreatment.objects.filter(effectiveness=effectiveness).count()
            if count > 0:
                effectiveness_data.append({
                    'name': effectiveness.get_name(current_language),
                    'value': effectiveness.value,
                    'count': count,
                    'color': effectiveness.color
                })
        
        # Priority data
        priority_data = []
        priority_levels = TreatmentPriority.objects.all()
        for priority in priority_levels:
            count = RiskTreatment.objects.filter(priority=priority).count()
            if count > 0:
                priority_data.append({
                    'name': priority.get_name(current_language),
                    'value': priority.value,
                    'count': count,
                    'color': priority.color
                })
        
        # Residual Risk data
        residual_risk_data = []
        residual_risk_levels = get_company_risk_levels_queryset(None)
        for residual_risk in residual_risk_levels:
            count = RiskTreatment.objects.filter(residual_risk_level=residual_risk).count()
            if count > 0:
                residual_risk_data.append({
                    'name': residual_risk.get_name(current_language),
                    'value': residual_risk.max_value,
                    'count': count,
                    'color': residual_risk.color
                })
        
        # Monitoring Frequency data
        monitoring_data = []
        monitoring_frequencies = MonitoringFrequency.objects.all()
        for frequency in monitoring_frequencies:
            count = RiskTreatment.objects.filter(monitoring_frequency=frequency).count()
            if count > 0:
                monitoring_data.append({
                    'name': frequency.get_name(current_language),
                    'count': count
                })
        
        # Cost analysis data
        cost_data = {
            'total_implementation_cost': 0,
            'total_maintenance_cost': 0,
            'treatments_with_cost': 0,
            'cost_by_priority': []
        }
        
        treatments_with_costs = RiskTreatment.objects.filter(
            implementation_cost__isnull=False
        ).exclude(implementation_cost=0)
        
        if treatments_with_costs.exists():
            cost_data['total_implementation_cost'] = treatments_with_costs.aggregate(
                total=Sum('implementation_cost')
            )['total'] or 0
            
            maintenance_costs = RiskTreatment.objects.filter(
                annual_maintenance_cost__isnull=False
            ).exclude(annual_maintenance_cost=0)
            
            if maintenance_costs.exists():
                cost_data['total_maintenance_cost'] = maintenance_costs.aggregate(
                    total=Sum('annual_maintenance_cost')
                )['total'] or 0
            
            cost_data['treatments_with_cost'] = treatments_with_costs.count()
            
            # Cost by priority
            for priority in priority_levels:
                priority_cost = treatments_with_costs.filter(priority=priority).aggregate(
                    total=Sum('implementation_cost')
                )['total'] or 0
                if priority_cost > 0:
                    cost_data['cost_by_priority'].append({
                        'priority': priority.get_name(current_language),
                        'cost': float(priority_cost),
                        'color': priority.color
                    })
        
        # logger.info(f"Treatment Details data collected - Effectiveness: {len(effectiveness_data)}, Priority: {len(priority_data)}, Residual Risk: {len(residual_risk_data)}, Monitoring: {len(monitoring_data)}")
        
        # Get translations for all status and treatment type values
        status_translations = {}
        treatment_type_translations = {}
        
        # Get all status values and their translations
        for status in Treatment_status.objects.all():
            status_translations[status.code] = status.get_name(current_language)
            
        # Get all treatment type values and their translations
        for ttype in Treatment_type.objects.all():
            treatment_type_translations[ttype.code] = ttype.get_name(current_language)
        
        translations = {
            'planned': _('Planned'),
            'in_progress': _('In Progress'),
            'completed': _('Completed'),
            'risk_treatment_progress': _('Risk Treatment Progress'),
            'status': status_translations,
            'treatment_type': treatment_type_translations
        }
        
        # Vulnerability Analytics Data
        # 1. Vulnerability Types Distribution (by Asset Type)
        vulnerability_types_data = []
        asset_types_with_vulns = AssetType.objects.filter(
            vulnerability__isnull=False
        ).distinct().prefetch_related('group')
        
        for asset_type in asset_types_with_vulns:
            vuln_count = Vulnerability.objects.filter(asset_type=asset_type).count()
            if vuln_count > 0:
                vulnerability_types_data.append({
                    'type': asset_type.name_local or asset_type.name,
                    'group': asset_type.group.name_local or asset_type.group.name,
                    'count': vuln_count,
                    'color': asset_type.color or '#007bff'
                })
        
        # 2. Most Common Vulnerabilities
        most_common_vulns = []
        vuln_counts = AssetVulnerability.objects.filter(
            status='Yes'
        ).values('vulnerability').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        for vuln_count in vuln_counts:
            vuln = Vulnerability.objects.get(id=vuln_count['vulnerability'])
            vuln_name = vuln.get_name(current_language)
            most_common_vulns.append({
                'name': vuln_name[:50] + '...' if len(vuln_name) > 50 else vuln_name,
                'count': vuln_count['count'],
                'full_name': vuln_name
            })
        
        # 3. Vulnerability Status Distribution
        vuln_status_data = []
        status_counts = AssetVulnerability.objects.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        status_colors = {
            'Yes': '#dc3545',    # Red for vulnerable
            'No': '#28a745',     # Green for not vulnerable
            'Undefined': '#ffc107'  # Yellow for undefined
        }
        
        for status_count in status_counts:
            vuln_status_data.append({
                'status': status_count['status'],
                'count': status_count['count'],
                'color': status_colors.get(status_count['status'], '#6c757d')
            })
        
        # 4. Vulnerabilities by Priority (based on risk levels)
        vuln_priority_data = []
        priority_counts = {}
        
        # Get all asset vulnerabilities with risk calculations
        asset_vulns_with_risk = AssetVulnerability.objects.filter(
            status='Yes'
        ).select_related('asset', 'vulnerability')
        
        for av in asset_vulns_with_risk:
            # Calculate risk level for this vulnerability
            criticality = av.asset.get_criticality()
            highest_risk = 0
            
            for threat in av.vulnerability.threats.all():
                threat_impact_value = calculate_threat_impact_value(threat.probability, threat.impact)
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                if value_of_risk > highest_risk:
                    highest_risk = value_of_risk
            
            risk_level = calculate_risk_level(highest_risk, company_id=av.asset.company_id)
            if risk_level:
                risk_level_name = risk_level.get_name_by_language(current_language) or risk_level.get_name()
                priority_counts[risk_level_name] = priority_counts.get(risk_level_name, 0) + 1
        
        # Convert to list format for chart
        for risk_level_name, count in priority_counts.items():
            # Get color from RiskLevel model
            risk_level_obj = RiskLevel.get_by_display_name(risk_level_name)
            color = risk_level_obj.color if risk_level_obj else '#6c757d'
            
            vuln_priority_data.append({
                'priority': risk_level_name,
                'count': count,
                'color': color
            })
        
        # 5. Vulnerabilities by Asset Group
        vuln_group_data = []
        group_counts = AssetVulnerability.objects.filter(
            status='Yes'
        ).values('vulnerability__asset_group__id').annotate(
            count=Count('id')
        ).order_by('-count')
        
        for group_count in group_counts:
            # Get the asset group object to access localized name
            try:
                asset_group = AssetGroup.objects.get(id=group_count['vulnerability__asset_group__id'])
                group_name = asset_group.name_local or asset_group.name
            except AssetGroup.DoesNotExist:
                group_name = '-'
            
            vuln_group_data.append({
                'group': group_name or '-',
                'count': group_count['count'],
                'color': '#17a2b8'  # Default color
            })
        
        # 6. Treatment Status for Vulnerabilities
        vuln_treatment_data = []
        treatment_counts = {
            'treated': 0,
            'partially_treated': 0,
            'not_treated': 0
        }
        
        # Count treated vulnerabilities
        treated_vulns = RiskTreatment.objects.exclude(
            Q(treatment_type__code='Undefined') &
            Q(description='') &
            Q(responsible='') &
            Q(deadline__isnull=True) &
            Q(status__code='Undefined')
        ).values('vulnerability').distinct().count()
        
        total_vulns = AssetVulnerability.objects.filter(status='Yes').count()
        not_treated = total_vulns - treated_vulns
        
        vuln_treatment_data = [
            {
                'status': _('Treated'),
                'count': treated_vulns,
                'color': '#28a745'
            },
            {
                'status': _('Not Treated'),
                'count': not_treated,
                'color': '#dc3545'
            }
        ]
        
        return JsonResponse({
            'vulnerability_data': vulnerability_data,
            'criticality_data': list(criticality_data),
            'risk_matrix_data': risk_matrix_data,
            'all_risk_levels': all_risk_levels_data,
            'treatment_data': treatment_data,
            'translations': translations,
            # Treatment Details data
            'effectiveness_data': effectiveness_data,
            'priority_data': priority_data,
            'residual_risk_data': residual_risk_data,
            'monitoring_data': monitoring_data,
            'cost_data': cost_data,
            # Vulnerability Analytics data
            'vulnerability_types_data': vulnerability_types_data,
            'most_common_vulns': most_common_vulns,
            'vuln_status_data': vuln_status_data,
            'vuln_priority_data': vuln_priority_data,
            'vuln_group_data': vuln_group_data,
            'vuln_treatment_data': vuln_treatment_data
        })

    except Exception as e:
        logger.error(f"Error in risk_assessment_dashboard: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'An error occurred while processing the dashboard data.'}, status=500)


@user_passes_test(can_edit_risk_assessment)
@require_POST
def save_asset_vulnerabilities(request):
    asset_id = request.POST.get('asset_id')
    vulnerabilities_data = json.loads(request.POST.get('vulnerabilities', '[]'))

    # print(f"Received data - asset_id: {asset_id}, vulnerabilities: {vulnerabilities_data}")
    try:
        asset = InformationAsset.objects.get(id=asset_id)
        # logger.info(f"Raw last_modified from DB: {asset.last_modified}")
        for vuln_data in vulnerabilities_data:
            vulnerability = Vulnerability.objects.get(id=vuln_data['id'])
            asset_vulnerability, created = AssetVulnerability.objects.get_or_create(
                asset=asset,
                vulnerability=vulnerability,
                defaults={'status': vuln_data['status'], 'comment': vuln_data['comment']}
            )
            if not created:
                asset_vulnerability.status = vuln_data['status']
                asset_vulnerability.comment = vuln_data['comment']
            asset_vulnerability.modified_by = request.user
            asset_vulnerability.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        print(f"Error saving vulnerabilities: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)})

@api_view(['GET'])
@user_passes_test(has_risk_assessment_access)
def asset_vulnerabilities_data(request):
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = normalize_risk_table_length(request.GET.get('length', 25))
    search_value = request.GET.get('search[value]', '')
    show_deleted = request.GET.get('showDeleted', 'false').lower() == 'true'

    # Get filter parameters
    company_filter = request.GET.get('company', '')
    criticality_filter = request.GET.get('criticality', '')
    asset_group_filter = request.GET.get('asset_group', '')
    asset_type_filter = request.GET.get('asset_type', '')
    vulnerability_status_filter = request.GET.get('vulnerability_status', '')
    risk_level_filter = request.GET.get('risk_level', '')
    treatment_status_filter = request.GET.get('treatment_status', '')

    custom_search = request.GET.get('search', '')
    
    # Log filter parameters for debugging
    logger.debug(f"Filter parameters: company={company_filter}, criticality={criticality_filter}, "
                f"asset_group={asset_group_filter}, asset_type={asset_type_filter}, "
                f"vulnerability_status={vulnerability_status_filter}, risk_level={risk_level_filter}, "
                f"treatment_status={treatment_status_filter}")

    order_column_index = int(request.GET.get('order[0][column]', 0))
    order_direction = request.GET.get('order[0][dir]', 'asc')

    # Map frontend column index (0=checkbox, 1=asset_id, 2=criticality, 3=name, ... 8=deletion_date, 9=vulnerabilities, ...)
    # to valid InformationAsset/annotated fields only (no last_modified - it is computed in Python).
    orderable_columns = {
        0: 'asset_id',
        1: 'asset_id',
        2: 'asset_id',  # criticality: not a single DB field, fallback to asset_id
        3: 'name',
        4: 'company__name',
        5: 'group__name',
        6: 'description',
        7: 'location',
        8: 'deletion_date',
        9: 'vulnerabilities_count',
        10: 'asset_id',
        11: 'asset_id',
    }
    order_column = orderable_columns.get(order_column_index, 'asset_id')
    if order_direction == 'desc':
        order_column = f'-{order_column}'

    user_permissions = get_user_risk_assessment_permissions(request.user)

    queryset = InformationAsset.objects.filter(
        company__id__in=user_permissions['companies']
    ).select_related(
        'company', 'group', 'asset_type', 'confidentiality', 'integrity', 'availability'
    ).annotate(
        vulnerabilities_count=Count(
            Case(
                When(assetvulnerability__status='Yes', then=1),
                output_field=IntegerField()
            )
        )
    )

    # Apply filters
    if not show_deleted:
        queryset = queryset.filter(deletion_date__isnull=True, is_active=True)

    # Company filter
    if company_filter:
        # Check if the filter is a name (string) or ID (number)
        try:
            # Try to convert to integer (ID)
            company_id = int(company_filter)
            queryset = queryset.filter(company_id=company_id)
        except ValueError:
            # If it's not a number, treat it as a name
            queryset = queryset.filter(company__name=company_filter)

    # Criticality filter
    if criticality_filter:
        # Check if the filter is a name (string) or ID (number)
        try:
            # Try to convert to integer (ID)
            criticality_id = int(criticality_filter)
            queryset = queryset.filter(
                Q(confidentiality_id=criticality_id) |
                Q(integrity_id=criticality_id) |
                Q(availability_id=criticality_id)
            )
        except ValueError:
            # If it's not a number, treat it as a name
            queryset = queryset.filter(
                Q(confidentiality__name=criticality_filter) |
                Q(confidentiality__code=criticality_filter) |
                Q(integrity__name=criticality_filter) |
                Q(integrity__code=criticality_filter) |
                Q(availability__name=criticality_filter) |
                Q(availability__code=criticality_filter)
            )

    # Asset group filter
    if asset_group_filter:
        # Check if the filter is a name (string) or ID (number)
        try:
            # Try to convert to integer (ID)
            group_id = int(asset_group_filter)
            queryset = queryset.filter(group_id=group_id)
        except ValueError:
            # If it's not a number, treat it as a name (AssetGroup has name, name_local)
            queryset = queryset.filter(
                Q(group__name=asset_group_filter) |
                Q(group__name_local=asset_group_filter)
            )

    # Asset type filter
    if asset_type_filter:
        # Check if the filter is a name (string) or ID (number)
        try:
            # Try to convert to integer (ID)
            type_id = int(asset_type_filter)
            queryset = queryset.filter(asset_type_id=type_id)
        except ValueError:
            # If it's not a number, treat it as a name
            queryset = queryset.filter(
                Q(asset_type__name=asset_type_filter) |
                Q(asset_type__code=asset_type_filter)
            )



    # Search filter (combine DataTables search with custom search)
    search_terms = []
    if search_value:
        search_terms.append(search_value)
    if custom_search:
        search_terms.append(custom_search)
    
    if search_terms:
        search_query = Q()
        for term in search_terms:
                    search_query |= (
            Q(asset_id__icontains=term) |
            Q(name__icontains=term) |
            Q(company__name__icontains=term) |
            Q(group__name__icontains=term) |
            Q(group__name_local__icontains=term) |
            Q(asset_type__name__icontains=term) |
            Q(asset_type__name_local__icontains=term) |
            Q(description__icontains=term) |
            Q(location__icontains=term)
        )
        queryset = queryset.filter(search_query)

    # Apply vulnerability status filter
    if vulnerability_status_filter:
        # Map common vulnerability status names to their values
        status_mapping = {
            'vulnerable': 'Yes',
            'not_vulnerable': 'No',
            'undefined': 'Undefined',
            'yes': 'Yes',
            'no': 'No'
        }
        
        # Get the actual status value
        status_value = status_mapping.get(vulnerability_status_filter.lower(), vulnerability_status_filter)
        
        if status_value in ['Yes', 'No', 'Undefined']:
            queryset = queryset.filter(assetvulnerability__status=status_value)
            # Remove duplicates that might occur due to joins
            queryset = queryset.distinct()

    # Apply risk level filter (this requires additional processing after getting assets)
    risk_level_filter_applied = bool(risk_level_filter)
    
    # Store the risk level filter value for later processing
    risk_level_filter_value = risk_level_filter
    
    # Apply treatment status filter (this requires additional processing after getting assets)
    treatment_status_filter_applied = bool(treatment_status_filter)
    
    # Store the treatment status filter value for later processing
    treatment_status_filter_value = treatment_status_filter

    # Get available filter values for current dataset
    available_filters = {}
    if not any([company_filter, criticality_filter, asset_group_filter, asset_type_filter, 
                vulnerability_status_filter, risk_level_filter, treatment_status_filter, 
                custom_search, search_value]):
        # Only get available filters if no other filters are applied (to avoid circular dependency)
        available_filters = {
            'asset_groups': list(queryset.values_list('group__id', 'group__name').distinct()),
            'asset_types': list(queryset.values_list('asset_type__id', 'asset_type__name').distinct()),
        }

    asset_total = queryset.count()

    # ── Build Software Register queryset with same filters ─────────────────────
    sw_queryset = SoftwareRegister.objects.filter(
        Q(company__id__in=user_permissions['companies']) | Q(company__isnull=True)
    ).select_related(
        'company', 'group', 'asset_type', 'confidentiality', 'integrity', 'availability'
    ).annotate(
        vulnerabilities_count=Count(
            Case(
                When(risk_vulnerabilities__status='Yes', then=1),
                output_field=IntegerField()
            )
        )
    ).filter(is_active=True)

    if company_filter:
        try:
            sw_queryset = sw_queryset.filter(company_id=int(company_filter))
        except ValueError:
            sw_queryset = sw_queryset.filter(company__name=company_filter)

    if criticality_filter:
        try:
            crit_id = int(criticality_filter)
            sw_queryset = sw_queryset.filter(
                Q(confidentiality_id=crit_id) | Q(integrity_id=crit_id) | Q(availability_id=crit_id)
            )
        except ValueError:
            sw_queryset = sw_queryset.filter(
                Q(confidentiality__name=criticality_filter) | Q(integrity__name=criticality_filter) |
                Q(availability__name=criticality_filter)
            )

    if asset_group_filter:
        try:
            sw_queryset = sw_queryset.filter(group_id=int(asset_group_filter))
        except ValueError:
            sw_queryset = sw_queryset.filter(Q(group__name=asset_group_filter) | Q(group__name_local=asset_group_filter))

    if asset_type_filter:
        try:
            sw_queryset = sw_queryset.filter(asset_type_id=int(asset_type_filter))
        except ValueError:
            sw_queryset = sw_queryset.filter(Q(asset_type__name=asset_type_filter) | Q(asset_type__code=asset_type_filter))

    if search_terms:
        sw_search_query = Q()
        for term in search_terms:
            sw_search_query |= (
                Q(name__icontains=term) | Q(company__name__icontains=term) |
                Q(group__name__icontains=term) | Q(group__name_local__icontains=term) |
                Q(asset_type__name__icontains=term) | Q(asset_type__name_local__icontains=term) |
                Q(description__icontains=term) | Q(manufacturer__icontains=term)
            )
        sw_queryset = sw_queryset.filter(sw_search_query)

    if vulnerability_status_filter:
        status_mapping = {
            'vulnerable': 'Yes', 'not_vulnerable': 'No', 'undefined': 'Undefined', 'yes': 'Yes', 'no': 'No'
        }
        sv_status = status_mapping.get(vulnerability_status_filter.lower(), vulnerability_status_filter)
        if sv_status in ['Yes', 'No', 'Undefined']:
            sw_queryset = sw_queryset.filter(risk_vulnerabilities__status=sv_status).distinct()

    # Map asset order column to software field
    sw_order_map = {
        'asset_id': 'id', '-asset_id': '-id',
        'name': 'name', '-name': '-name',
        'company__name': 'company__name', '-company__name': '-company__name',
        'group__name': 'group__name', '-group__name': '-group__name',
        'description': 'description', '-description': '-description',
        'vulnerabilities_count': 'vulnerabilities_count', '-vulnerabilities_count': '-vulnerabilities_count',
    }
    sw_order_column = sw_order_map.get(order_column, 'id')
    sw_queryset = sw_queryset.order_by(sw_order_column)

    sw_total = sw_queryset.count()
    # ── External Media Register queryset with same filters ─────────────────────
    em_queryset = ExternalMediaRegister.objects.filter(
        Q(company__id__in=user_permissions['companies']) | Q(company__isnull=True)
    ).select_related(
        'company', 'group', 'asset_type', 'confidentiality', 'integrity', 'availability'
    ).annotate(
        vulnerabilities_count=Count(
            Case(
                When(risk_vulnerabilities__status='Yes', then=1),
                output_field=IntegerField()
            )
        )
    ).filter(is_active=True)

    if company_filter:
        try:
            em_queryset = em_queryset.filter(company_id=int(company_filter))
        except ValueError:
            em_queryset = em_queryset.filter(company__name=company_filter)

    if criticality_filter:
        try:
            crit_id = int(criticality_filter)
            em_queryset = em_queryset.filter(
                Q(confidentiality_id=crit_id) | Q(integrity_id=crit_id) | Q(availability_id=crit_id)
            )
        except ValueError:
            em_queryset = em_queryset.filter(
                Q(confidentiality__name=criticality_filter) | Q(integrity__name=criticality_filter) |
                Q(availability__name=criticality_filter)
            )

    if asset_group_filter:
        try:
            em_queryset = em_queryset.filter(group_id=int(asset_group_filter))
        except ValueError:
            em_queryset = em_queryset.filter(Q(group__name=asset_group_filter) | Q(group__name_local=asset_group_filter))

    if asset_type_filter:
        try:
            em_queryset = em_queryset.filter(asset_type_id=int(asset_type_filter))
        except ValueError:
            em_queryset = em_queryset.filter(Q(asset_type__name=asset_type_filter) | Q(asset_type__code=asset_type_filter))

    if search_terms:
        em_search_query = Q()
        for term in search_terms:
            em_search_query |= (
                Q(name__icontains=term) | Q(company__name__icontains=term) |
                Q(group__name__icontains=term) | Q(group__name_local__icontains=term) |
                Q(asset_type__name__icontains=term) | Q(asset_type__name_local__icontains=term) |
                Q(description__icontains=term) | Q(serial_number__icontains=term)
            )
        em_queryset = em_queryset.filter(em_search_query)

    if vulnerability_status_filter:
        status_mapping = {
            'vulnerable': 'Yes', 'not_vulnerable': 'No', 'undefined': 'Undefined', 'yes': 'Yes', 'no': 'No'
        }
        em_status = status_mapping.get(vulnerability_status_filter.lower(), vulnerability_status_filter)
        if em_status in ['Yes', 'No', 'Undefined']:
            em_queryset = em_queryset.filter(risk_vulnerabilities__status=em_status).distinct()

    em_order_map = {
        'asset_id': 'id', '-asset_id': '-id',
        'name': 'name', '-name': '-name',
        'company__name': 'company__name', '-company__name': '-company__name',
        'group__name': 'group__name', '-group__name': '-group__name',
        'description': 'description', '-description': '-description',
        'vulnerabilities_count': 'vulnerabilities_count', '-vulnerabilities_count': '-vulnerabilities_count',
    }
    em_order_column = em_order_map.get(order_column, 'id')
    em_queryset = em_queryset.order_by(em_order_column)

    em_total = em_queryset.count()
    total_records = asset_total + sw_total + em_total

    # Combined pagination: assets first, then software, then external media
    asset_slice_count = max(0, min(start + length, asset_total) - start)
    sw_from = max(0, start - asset_total)
    sw_slice_count = max(0, min(length - asset_slice_count, max(0, sw_total - sw_from)))
    em_from = max(0, start - asset_total - sw_total)
    em_slice_count = max(0, length - asset_slice_count - sw_slice_count)

    queryset = queryset.order_by(order_column)
    assets = queryset[start:start + asset_slice_count]
    sw_entries = sw_queryset[sw_from:sw_from + sw_slice_count] if sw_slice_count > 0 else []
    em_entries = em_queryset[em_from:em_from + em_slice_count] if em_slice_count > 0 else []
    # ───────────────────────────────────────────────────────────────────────────

    user_timezone = request.session.get('user_timezone', 'UTC')
    tz = pytz.timezone(user_timezone)
    current_language = get_language()[:2]
    data = []
    
    # Get the Undefined treatment type for reference
    undefined_treatment_type = Treatment_type.objects.filter(code='Undefined').first()
    undefined_status = Treatment_status.objects.filter(code='Undefined').first()
    
    for asset in assets:
        last_modified = AssetVulnerability.objects.filter(asset_id=asset.id).order_by('-modified_at').first()

        if last_modified:
            formatted_last_modified = last_modified.modified_at.astimezone(tz).strftime('%Y-%m-%d %H:%M:%S %Z')
        else:
            formatted_last_modified = ''

        group_name = asset.group.get_name() if asset.group else ''
        asset_type_name = asset.asset_type.get_name() if asset.asset_type else ''
        criticality = asset.get_criticality()

        # Prepare related collections
        owners = list(asset.owners.select_related('cabinet_user__user').all())
        administrators = list(asset.administrators.select_related('cabinet_user__user').all())
        software_entries = list(asset.software_entries.all())

        formatted_last_modified = format_time(last_modified.modified_at) if last_modified else ''

        # Updated query to use foreign key relationships
        treated_vulnerabilities = RiskTreatment.objects.filter(
            asset=asset
        ).exclude(
            Q(treatment_type=undefined_treatment_type) &
            Q(description='') &
            Q(responsible='') &
            Q(deadline__isnull=True) &
            Q(status=undefined_status)
        ).count()
        
        # Build override lookup for this asset (used by risk level filter and by Yes/No counts below)
        from .models import ManualRiskLevelOverride
        _override_lookup_filter = {}
        try:
            for _ov in ManualRiskLevelOverride.objects.filter(asset=asset).select_related('manual_risk_level'):
                _override_lookup_filter[(_ov.vulnerability_id, _ov.threat_id if _ov.threat_id else None)] = _ov
        except Exception:
            pass

        # Apply risk level filter
        if risk_level_filter_applied:
            asset_vulnerabilities = AssetVulnerability.objects.filter(
                asset=asset, status='Yes'
            ).select_related('vulnerability')
            highest_risk_level_id = None
            
            for av in asset_vulnerabilities:
                for threat in av.vulnerability.threats.all():
                    _ov = _override_lookup_filter.get((av.vulnerability_id, threat.id if threat else None))
                    if _ov and _ov.manual_risk_level:
                        risk_level = _ov.manual_risk_level
                    else:
                        threat_impact_value = calculate_threat_impact_value(threat.probability, threat.impact)
                        threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                        value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                        risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)
                    
                    risk_level_matches = False
                    try:
                        filter_id = int(risk_level_filter_value)
                        if risk_level and risk_level.id == filter_id:
                            risk_level_matches = True
                    except ValueError:
                        if risk_level and (
                            risk_level.get_name_by_language('uk') == risk_level_filter_value or
                            risk_level.get_name_by_language('en') == risk_level_filter_value or
                            risk_level.get_name_by_language('ru') == risk_level_filter_value or
                            risk_level.get_name() == risk_level_filter_value
                        ):
                            risk_level_matches = True
                    
                    if risk_level_matches:
                        highest_risk_level_id = risk_level.id
                        break
                if highest_risk_level_id:
                    break
            
            if not highest_risk_level_id:
                continue  # Skip this asset if it doesn't match the risk level filter
        
        # Apply treatment status filter
        if treatment_status_filter_applied:
            total_vulnerabilities = asset.vulnerabilities_count
            if treatment_status_filter_value == 'not_treated' and treated_vulnerabilities > 0:
                continue
            elif treatment_status_filter_value == 'partially_treated' and (treated_vulnerabilities == 0 or treated_vulnerabilities >= total_vulnerabilities):
                continue
            elif treatment_status_filter_value == 'fully_treated' and treated_vulnerabilities < total_vulnerabilities:
                continue

        # Calculate risk levels for this asset - separate for Yes and No statuses
        # Use same override lookup as risk level filter (manual overrides count toward overridden level)
        override_lookup = _override_lookup_filter

        risk_levels_yes = {}
        risk_levels_no = {}
        risk_levels_colors = {}
        highest_risk_level_value = 0
        highest_risk_level_name = ''

        def _risk_level_for_threat(asset, av, threat, criticality, is_yes):
            override = override_lookup.get((av.vulnerability_id, threat.id if threat else None))
            if override and override.manual_risk_level:
                return override.manual_risk_level
            if is_yes:
                probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN) / 100
            else:
                probability = Decimal('0.001')
                impact = Decimal('0.001')
            threat_impact_value = probability * impact * 100
            threat_impact_level = calculate_threat_impact_level(threat_impact_value)
            value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
            return calculate_risk_level(value_of_risk, company_id=asset.company_id)
        
        # Process Yes status vulnerabilities (INNER JOIN via select_related excludes orphaned records)
        asset_vulnerabilities_yes = AssetVulnerability.objects.filter(
            asset=asset, status='Yes'
        ).select_related('vulnerability')
        for av in asset_vulnerabilities_yes:
            threats = av.vulnerability.threats.all()
            if not threats.exists():
                continue
            
            for threat in threats:
                risk_level = _risk_level_for_threat(asset, av, threat, criticality, is_yes=True)
                
                if risk_level:
                    risk_level_name = get_risk_level_name(risk_level, current_language)
                    risk_levels_yes[risk_level_name] = risk_levels_yes.get(risk_level_name, 0) + 1
                    risk_levels_colors[risk_level_name] = risk_level.color
                    
                    if risk_level.max_value > highest_risk_level_value:
                        highest_risk_level_value = risk_level.max_value
                        highest_risk_level_name = risk_level_name
        
        # Process No status vulnerabilities (INNER JOIN via select_related excludes orphaned records)
        asset_vulnerabilities_no = AssetVulnerability.objects.filter(
            asset=asset, status='No'
        ).select_related('vulnerability')
        for av in asset_vulnerabilities_no:
            threats = av.vulnerability.threats.all()
            
            if not threats.exists():
                # If no threats, create a default risk entry with minimal values for 'No' status
                override = override_lookup.get((av.vulnerability_id, None))
                if override and override.manual_risk_level:
                    risk_level = override.manual_risk_level
                else:
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                    threat_impact_value = probability * impact * 100
                    threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                    value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                    risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)
                
                if risk_level:
                    risk_level_name = get_risk_level_name(risk_level, current_language)
                    risk_levels_no[risk_level_name] = risk_levels_no.get(risk_level_name, 0) + 1
                    risk_levels_colors[risk_level_name] = risk_level.color
                    
                    if risk_level.max_value > highest_risk_level_value:
                        highest_risk_level_value = risk_level.max_value
                        highest_risk_level_name = risk_level_name
                continue
            
            for threat in threats:
                risk_level = _risk_level_for_threat(asset, av, threat, criticality, is_yes=False)
                
                if risk_level:
                    risk_level_name = get_risk_level_name(risk_level, current_language)
                    risk_levels_no[risk_level_name] = risk_levels_no.get(risk_level_name, 0) + 1
                    risk_levels_colors[risk_level_name] = risk_level.color
                    
                    if risk_level.max_value > highest_risk_level_value:
                        highest_risk_level_value = risk_level.max_value
                        highest_risk_level_name = risk_level_name
        
        # Combine for backward compatibility
        risk_levels_count = {}
        for level_name in set(list(risk_levels_yes.keys()) + list(risk_levels_no.keys())):
            risk_levels_count[level_name] = (risk_levels_yes.get(level_name, 0) + risk_levels_no.get(level_name, 0))

        # Count total AssetVulnerability records for this asset (to determine if assessment was done)
        total_vulnerabilities_count = AssetVulnerability.objects.filter(asset=asset).count()

        data.append({
            'id': asset.id,
            'asset_id': asset.asset_id,
            'criticality': {
                'text': criticality['name'],
                'cost': criticality['cost'],
                'color': criticality['color']
            },
            'name': asset.name,
            'company': asset.company.name if asset.company else '',
            'group': f"{group_name}/{asset_type_name}" if group_name and asset_type_name else group_name or asset_type_name,
            'description': asset.description,
            'location': asset.location,
            'deletion_date': asset.deletion_date.strftime('%Y-%m-%d') if asset.deletion_date else '',
            'vulnerabilities': asset.vulnerabilities_count,
            'total_vulnerabilities_count': total_vulnerabilities_count,
            'treated_vulnerabilities': treated_vulnerabilities,
            'risk_levels': risk_levels_count,
            'risk_levels_yes': risk_levels_yes,
            'risk_levels_no': risk_levels_no,
            'risk_levels_colors': risk_levels_colors,
            'highest_risk_level': {
                'name': highest_risk_level_name,
                'value': highest_risk_level_value
            },
            'last_modified': {
                'datetime': formatted_last_modified,
                'user': AssetVulnerability.get_user_full_name(last_modified.modified_by) if last_modified else ''
            }
        })

    # ── Software Register rows ──────────────────────────────────────────────────
    for sw in sw_entries:
        # Criticality from CIA
        sw_levels = [
            (sw.confidentiality, sw.confidentiality.cost if sw.confidentiality else 0),
            (sw.integrity, sw.integrity.cost if sw.integrity else 0),
            (sw.availability, sw.availability.cost if sw.availability else 0),
        ]
        sw_max = max(sw_levels, key=lambda x: x[1])
        if sw_max[0]:
            sw_criticality = {'name': sw_max[0].get_name(), 'cost': sw_max[0].cost, 'color': sw_max[0].color}
        else:
            sw_criticality = {'name': str(_('Undefined')), 'cost': 0, 'color': '#888888'}

        sw_group_name = sw.group.get_name() if sw.group else ''
        sw_type_name = sw.asset_type.get_name() if sw.asset_type else ''

        sw_risk_levels_yes = {}
        sw_risk_levels_no = {}
        sw_risk_levels_colors = {}
        sw_highest_risk_value = 0
        sw_highest_risk_name = ''

        sw_vulns_qs = SoftwareVulnerability.objects.filter(
            software_register=sw, status__in=['Yes', 'No']
        ).select_related('vulnerability')

        for sv in sw_vulns_qs:
            threats = sv.vulnerability.threats.all()
            for threat in threats:
                if sv.status == 'Yes':
                    prob = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    imp = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                else:
                    prob = Decimal('0.001')
                    imp = Decimal('0.001')
                tiv = calculate_threat_impact_value(prob, imp)
                til = calculate_threat_impact_level(tiv)
                vor = Decimal(str(sw_criticality['cost'])) * Decimal(str(til))
                rl = calculate_risk_level(vor, company_id=sw.company_id)
                if rl:
                    rl_name = get_risk_level_name(rl, current_language)
                    sw_risk_levels_colors[rl_name] = rl.color
                    if sv.status == 'Yes':
                        sw_risk_levels_yes[rl_name] = sw_risk_levels_yes.get(rl_name, 0) + 1
                    else:
                        sw_risk_levels_no[rl_name] = sw_risk_levels_no.get(rl_name, 0) + 1
                    if rl.max_value > sw_highest_risk_value:
                        sw_highest_risk_value = rl.max_value
                        sw_highest_risk_name = rl_name

        sw_last_vuln = SoftwareVulnerability.objects.filter(software_register=sw).order_by('-modified_at').first()
        sw_total_vuln_count = SoftwareVulnerability.objects.filter(software_register=sw).count()

        data.append({
            'id': sw.id,
            'row_type': 'software',
            'asset_id': f'S{sw.id:06d}',
            'criticality': {
                'text': sw_criticality['name'],
                'cost': sw_criticality['cost'],
                'color': sw_criticality['color'],
            },
            'name': sw.name,
            'company': sw.company.name if sw.company else '',
            'group': f"{sw_group_name}/{sw_type_name}" if sw_group_name and sw_type_name else sw_group_name or sw_type_name,
            'description': sw.description,
            'location': sw.manufacturer or '',
            'deletion_date': '',
            'vulnerabilities': sw.vulnerabilities_count,
            'total_vulnerabilities_count': sw_total_vuln_count,
            'treated_vulnerabilities': 0,
            'risk_levels': {},
            'risk_levels_yes': sw_risk_levels_yes,
            'risk_levels_no': sw_risk_levels_no,
            'risk_levels_colors': sw_risk_levels_colors,
            'highest_risk_level': {'name': sw_highest_risk_name, 'value': sw_highest_risk_value},
            'last_modified': {
                'datetime': format_time(sw_last_vuln.modified_at) if sw_last_vuln else '',
                'user': SoftwareVulnerability.get_user_full_name(sw_last_vuln.modified_by) if sw_last_vuln else '',
            }
        })
    # ───────────────────────────────────────────────────────────────────────────

    # ── External Media Register rows ───────────────────────────────────────────
    for em in em_entries:
        em_levels = [
            (em.confidentiality, em.confidentiality.cost if em.confidentiality else 0),
            (em.integrity, em.integrity.cost if em.integrity else 0),
            (em.availability, em.availability.cost if em.availability else 0),
        ]
        em_max = max(em_levels, key=lambda x: x[1])
        if em_max[0]:
            em_criticality = {'name': em_max[0].get_name(), 'cost': em_max[0].cost, 'color': em_max[0].color}
        else:
            em_criticality = {'name': str(_('Undefined')), 'cost': 0, 'color': '#888888'}

        em_group_name = em.group.get_name() if em.group else ''
        em_type_name = em.asset_type.get_name() if em.asset_type else ''

        em_risk_levels_yes = {}
        em_risk_levels_no = {}
        em_risk_levels_colors = {}
        em_highest_risk_value = 0
        em_highest_risk_name = ''

        em_vulns_qs = ExternalMediaVulnerability.objects.filter(
            external_media_register=em, status__in=['Yes', 'No']
        ).select_related('vulnerability')

        for ev in em_vulns_qs:
            threats = ev.vulnerability.threats.all()
            for threat in threats:
                if ev.status == 'Yes':
                    prob = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    imp = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                else:
                    prob = Decimal('0.001')
                    imp = Decimal('0.001')
                tiv = calculate_threat_impact_value(prob, imp)
                til = calculate_threat_impact_level(tiv)
                vor = Decimal(str(em_criticality['cost'])) * Decimal(str(til))
                rl = calculate_risk_level(vor, company_id=em.company_id)
                if rl:
                    rl_name = get_risk_level_name(rl, current_language)
                    em_risk_levels_colors[rl_name] = rl.color
                    if ev.status == 'Yes':
                        em_risk_levels_yes[rl_name] = em_risk_levels_yes.get(rl_name, 0) + 1
                    else:
                        em_risk_levels_no[rl_name] = em_risk_levels_no.get(rl_name, 0) + 1
                    if rl.max_value > em_highest_risk_value:
                        em_highest_risk_value = rl.max_value
                        em_highest_risk_name = rl_name

        em_last_vuln = ExternalMediaVulnerability.objects.filter(external_media_register=em).order_by('-modified_at').first()
        em_total_vuln_count = ExternalMediaVulnerability.objects.filter(external_media_register=em).count()

        data.append({
            'id': em.id,
            'row_type': 'external_media',
            'asset_id': f'M{em.id:05d}',
            'criticality': {
                'text': em_criticality['name'],
                'cost': em_criticality['cost'],
                'color': em_criticality['color'],
            },
            'name': em.name,
            'company': em.company.name if em.company else '',
            'group': f"{em_group_name}/{em_type_name}" if em_group_name and em_type_name else em_group_name or em_type_name,
            'description': em.description,
            'location': em.serial_number or '',
            'deletion_date': '',
            'vulnerabilities': em.vulnerabilities_count,
            'total_vulnerabilities_count': em_total_vuln_count,
            'treated_vulnerabilities': 0,
            'risk_levels': {},
            'risk_levels_yes': em_risk_levels_yes,
            'risk_levels_no': em_risk_levels_no,
            'risk_levels_colors': em_risk_levels_colors,
            'highest_risk_level': {'name': em_highest_risk_name, 'value': em_highest_risk_value},
            'last_modified': {
                'datetime': format_time(em_last_vuln.modified_at) if em_last_vuln else '',
                'user': ExternalMediaVulnerability.get_user_full_name(em_last_vuln.modified_by) if em_last_vuln else '',
            }
        })
    # ───────────────────────────────────────────────────────────────────────────

    response = {
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': total_records,
        'data': data,
        'available_filters': available_filters
    }

    # Log response for debugging
    logger.debug(f"Response: {len(data)} records returned, total_records={total_records}")

    return Response(response)




@login_required
@user_passes_test(has_risk_assessment_access)
def export_risk_calculation(request):
    """Export risk calculation for selected assets to XLSX.

    Uses current methodology: Threat Impact Value = Probability (L) × Overall Impact (E) × 100 (%).
    Overall Impact (E) is the arithmetic mean of Financial/Operational/Reputational impacts
    when available via threat.calculate_overall_impact().
    """
    selected_ids = request.GET.get('ids', '').split(',') if request.GET.get('ids') else []
    assets = InformationAsset.objects.filter(asset_id__in=selected_ids)
    language = request.GET.get('language', get_language()[:2])

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = _("Risk Calculation")

    # Styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0366D6', end_color='0366D6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_border = Border(
        left=Side(style='thin', color='E1E1E1'),
        right=Side(style='thin', color='E1E1E1'),
        top=Side(style='thin', color='E1E1E1'),
        bottom=Side(style='thin', color='E1E1E1')
    )

    row_fills = [
        PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'),
        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    ]

    headers = [
        _('Asset ID'), _('Asset Name'), _('Company'), _('Criticality'),
        _('Vulnerability'), _('Threat'), _('Probability/Impact'),
        _('Threat Impact Value'), _('Threat Impact Level'),
        _('Value of Risk'), _('Risk Level'), _('Risk Mitigation Controls')
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    ws.auto_filter.ref = "A1:L1"

    # Fill data
    row_idx = 2
    for asset in assets:
        criticality = get_localized_criticality(asset, language)
        for av in asset.assetvulnerability_set.filter(status__in=['Yes', 'No']).select_related('vulnerability'):
            for threat in av.vulnerability.threats.all():
                # Probability and Impact
                if av.status == 'No':
                    probability = Decimal('0.001')
                    overall_impact = Decimal('0.001')
                else:
                    probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    try:
                        if hasattr(threat, 'calculate_overall_impact'):
                            overall_impact = Decimal(str(threat.calculate_overall_impact()))
                        else:
                            overall_impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN) / 100
                    except Exception:
                        overall_impact = Decimal('0')

                threat_impact_value = probability * overall_impact * 100  # percentage
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)

                fill = row_fills[(row_idx - 2) % 2]
                row_data = [
                    asset.asset_id,
                    asset.name,
                    asset.company.name if asset.company else '',
                    f"{criticality['name']} / {criticality['cost']}",
                    av.vulnerability.get_name(language),
                    threat.get_name(language),
                    f"{probability:.4f}/{overall_impact:.2f}",
                    f"{threat_impact_value:.4f}%",
                    threat_impact_level,
                    float(value_of_risk),
                    risk_level.get_name(language) if risk_level else _("Undefined"),
                    getattr(av.vulnerability, f'risk_mitigation_controls_{language}', av.vulnerability.risk_mitigation_controls_uk)
                ]

                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=c_idx, value=value)
                    cell.border = cell_border
                    cell.fill = fill
                    if c_idx == 4:  # Criticality column background by color
                        color = criticality['color'].lstrip('#') if criticality.get('color') else None
                        if color:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)

                row_idx += 1

    # Column widths
    column_widths = [12, 20, 20, 18, 30, 20, 18, 18, 18, 15, 20, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="risk_calculation.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(has_risk_assessment_access)
def export_asset_vulnerabilities(request):
    selected_ids = request.GET.get('ids', '').split(',')
    assets = InformationAsset.objects.filter(id__in=selected_ids)
    current_language = get_language()[:2]

    # Create a new workbook and set active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = _("Asset Vulnerabilities")
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0366D6', end_color='0366D6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='E1E1E1'),
        right=Side(style='thin', color='E1E1E1'),
        top=Side(style='thin', color='E1E1E1'),
        bottom=Side(style='thin', color='E1E1E1')
    )
    
    row_fills = [
        PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'),
        PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    ]
    
    # Status colors for vulnerability status
    status_colors = {
        'Yes': 'DC3545',    # Red (Danger)
        'No': '28A745',     # Green (Success)
        'Undefined': 'FFC107' # Yellow (Warning)
    }
    
    # Risk level colors
    risk_level_colors = {
        'Critical': 'DC3545',    # Red
        'High': 'FD7E14',        # Orange
        'Medium': 'FFC107',      # Yellow
        'Low': '20C997',         # Teal
        'Minimal': '6F42C1',     # Purple
        'Absent': '6C757D'       # Gray
    }
    
    # Define headers
    headers = [
        _('Asset ID'), _('Name'), _('Company'), _('Group/Type'), _('Description'), 
        _('Location'), _('Deletion Date'), 
        _('Critical Risks'), _('High Risks'), _('Medium Risks'), _('Low Risks'), _('Minimal Risks'), _('Absent Risks'),
        _('Highest Risk Level'), _('Vulnerability'), _('Status'), _('Comment')
    ]
    
    # Set headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:Q1"
    
    # Apply risk level colors to header columns
    for col_idx in range(8, 14):  # Risk level columns (8-13)
        risk_level_names = ['Critical', 'High', 'Medium', 'Low', 'Minimal', 'Absent']
        risk_level_index = col_idx - 8
        if risk_level_index < len(risk_level_names):
            risk_level_name = risk_level_names[risk_level_index]
            if risk_level_name in risk_level_colors:
                cell = ws.cell(row=1, column=col_idx)
                color = risk_level_colors[risk_level_name]
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    
    # Function to get localized names
    def get_localized_name(obj, field_prefix='name'):
        return getattr(obj, f'{field_prefix}_{current_language}', '') or getattr(obj, f'{field_prefix}_en', '') or ''
    
    # Fill data
    row_idx = 2
    for asset in assets:
        asset_vulnerabilities = AssetVulnerability.objects.filter(
            asset=asset
        ).select_related('vulnerability')

        group_name = get_localized_name(asset.group) if asset.group else ''
        asset_type_name = get_localized_name(asset.asset_type) if asset.asset_type else ''
        group_type = f"{group_name}/{asset_type_name}" if group_name and asset_type_name else group_name or asset_type_name
        
        # Calculate risk levels for this asset
        risk_levels_count = {}
        highest_risk_level_value = 0
        highest_risk_level_name = ''
        
        # Get criticality for this asset
        levels = [
            (asset.confidentiality, asset.confidentiality.cost if asset.confidentiality else 0),
            (asset.integrity, asset.integrity.cost if asset.integrity else 0),
            (asset.availability, asset.availability.cost if asset.availability else 0)
        ]
        max_level = max(levels, key=lambda x: x[1])
        criticality_cost = max_level[1] if max_level[0] else 0
        
        # Get all vulnerabilities for this asset
        all_vulnerabilities = Vulnerability.objects.filter(asset_type=asset.asset_type, is_active=True)

        for vulnerability in all_vulnerabilities:
            # Get asset vulnerability status
            asset_vuln = asset_vulnerabilities.filter(vulnerability=vulnerability).first()
            if asset_vuln and asset_vuln.status == 'Yes':
                # Calculate risk level for each threat of this vulnerability
                for threat in vulnerability.threats.all():
                    threat_impact_value = calculate_threat_impact_value(threat.probability, threat.impact)
                    threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                    value_of_risk = Decimal(str(criticality_cost)) * Decimal(str(threat_impact_level))
                    risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)
                    
                    if risk_level:
                        risk_level_name = get_risk_level_name(risk_level, current_language)
                        risk_levels_count[risk_level_name] = risk_levels_count.get(risk_level_name, 0) + 1
                        
                        # Track highest risk level
                        if risk_level.max_value > highest_risk_level_value:
                            highest_risk_level_value = risk_level.max_value
                            highest_risk_level_name = risk_level_name
        
        # Common asset data
        asset_data = [
            asset.asset_id,
            asset.name,
            asset.company.name if asset.company else '',
            group_type,
            asset.description,
            asset.location,
            asset.deletion_date.strftime('%Y-%m-%d') if asset.deletion_date else '',
            risk_levels_count.get('Critical', 0),
            risk_levels_count.get('High', 0),
            risk_levels_count.get('Medium', 0),
            risk_levels_count.get('Low', 0),
            risk_levels_count.get('Minimal', 0),
            risk_levels_count.get('Absent', 0),
            highest_risk_level_name or 'N/A'
        ]
        
        if asset_vulnerabilities.exists():
            for av in asset_vulnerabilities:
                fill = row_fills[(row_idx - 2) % 2]
                
                # Combine asset data with vulnerability data
                row_data = asset_data + [
                    get_localized_name(av.vulnerability, 'vulnerability'),
                    av.status,
                    av.comment
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = cell_border
                    cell.fill = fill
                    
                    # Highlight deleted assets with red text
                    if asset.deletion_date and col_idx <= 7:  # Asset data columns
                        cell.font = Font(color='DC3545')
                    
                    # Apply risk level colors to risk count columns (8-13)
                    if 8 <= col_idx <= 13 and value > 0:
                        risk_level_names = ['Critical', 'High', 'Medium', 'Low', 'Minimal', 'Absent']
                        risk_level_index = col_idx - 8
                        if risk_level_index < len(risk_level_names):
                            risk_level_name = risk_level_names[risk_level_index]
                            if risk_level_name in risk_level_colors:
                                color = risk_level_colors[risk_level_name]
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                                cell.font = Font(color='FFFFFF', bold=True)
                    
                    # Apply status colors
                    if col_idx == 16 and av.status in status_colors:  # Status column (updated index)
                        color = status_colors[av.status]
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        # White text for all status colors for better visibility
                        cell.font = Font(color='FFFFFF', bold=True)
                        
                row_idx += 1
        else:
            # If no vulnerabilities, add a single row with N/A
            fill = row_fills[(row_idx - 2) % 2]
            row_data = asset_data + ['N/A', 'N/A', 'N/A']
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = cell_border
                cell.fill = fill
                
                # Highlight deleted assets with red text
                if asset.deletion_date and col_idx <= 7:  # Asset data columns
                    cell.font = Font(color='DC3545')
                
                # Apply risk level colors to risk count columns (8-13) for rows without vulnerabilities
                if 8 <= col_idx <= 13 and value > 0:
                    risk_level_names = ['Critical', 'High', 'Medium', 'Low', 'Minimal', 'Absent']
                    risk_level_index = col_idx - 8
                    if risk_level_index < len(risk_level_names):
                        risk_level_name = risk_level_names[risk_level_index]
                        if risk_level_name in risk_level_colors:
                            color = risk_level_colors[risk_level_name]
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            cell.font = Font(color='FFFFFF', bold=True)
                    
            row_idx += 1
    
    # Adjust column widths
    column_widths = [12, 20, 20, 20, 30, 20, 15, 15, 15, 15, 15, 15, 15, 20, 30, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="asset_vulnerabilities.xlsx"'
    
    # Save the workbook to response
    wb.save(response)
    
    return response

@user_passes_test(has_risk_assessment_access)
def get_asset_vulnerabilities(request):
    asset_id = request.GET.get('asset_id')
    if not asset_id:
        return JsonResponse({'error': 'Asset ID is required'}, status=400)

    try:
        asset = get_object_or_404(InformationAsset, id=asset_id)
        vulnerabilities = Vulnerability.objects.filter(asset_type=asset.asset_type, is_active=True)
        asset_vulnerabilities = AssetVulnerability.objects.filter(asset=asset)

        current_language = get_language()[:2]

        def get_localized_field(obj, field_name):
            if not obj:
                return ''
            # Try current language, then Ukrainian, then non-localized field as a final fallback
            return (
                getattr(obj, f'{field_name}_{current_language}', '') or
                getattr(obj, f'{field_name}_uk', '') or
                getattr(obj, field_name, '')
            )

        vulnerabilities_data = [
            {
                'id': v.id,
                'scope': get_localized_field(v, 'scope'),
                'vulnerability': v.get_name(current_language),
                'risk_mitigation_controls': get_localized_field(v, 'risk_mitigation_controls'),
                'pci_dss_requirement': get_localized_field(v, 'pci_dss_requirement'),
                'iso27001_requirement': get_localized_field(v, 'iso27001_requirement'),
                'note': get_localized_field(v, 'note'),
            }
            for v in vulnerabilities
        ]

        asset_vulnerabilities_data = {
            av.vulnerability_id: {
                'id': av.id,
                'status': av.status,
                'comment': av.comment
            }
            for av in asset_vulnerabilities
        }

        criticality = asset.get_criticality()

        # Prepare related collections for asset info in modals
        owners = list(asset.owners.select_related('cabinet_user__user').all())
        administrators = list(asset.administrators.select_related('cabinet_user__user').all())
        software_entries = list(asset.software_entries.all())

        context = {
            'asset': {
                'id': asset.id,
                'name': asset.name,
                'asset_id': asset.asset_id,
                'company': asset.company.name if asset.company else None,
                'group': get_localized_field(asset.group, 'name') if asset.group else None,
                'asset_type': get_localized_field(asset.asset_type, 'name') if asset.asset_type else None,
                'description': asset.description,
                'location': asset.location,
                'registration_date': asset.registration_date.isoformat() if asset.registration_date else None,
                'deletion_date': asset.deletion_date.isoformat() if asset.deletion_date else None,
                'notes': asset.notes,
                'is_active': asset.is_active,
                'criticality': criticality,
                'cia': {
                    'confidentiality': {
                        'name': asset.confidentiality.get_name() if asset.confidentiality else None,
                        'cost': asset.confidentiality.cost if asset.confidentiality else None,
                        'color': asset.confidentiality.color if asset.confidentiality else None,
                    },
                    'integrity': {
                        'name': asset.integrity.get_name() if asset.integrity else None,
                        'cost': asset.integrity.cost if asset.integrity else None,
                        'color': asset.integrity.color if asset.integrity else None,
                    },
                    'availability': {
                        'name': asset.availability.get_name() if asset.availability else None,
                        'cost': asset.availability.cost if asset.availability else None,
                        'color': asset.availability.color if asset.availability else None,
                    },
                },
                'actualization': {
                    'actualization_date': asset.actualization_date.isoformat() if asset.actualization_date else None,
                    'actualized_by': asset.actualized_by.get_full_name() if asset.actualized_by else None,
                    'marked_no_longer_actual_at': asset.marked_no_longer_actual_at.isoformat() if asset.marked_no_longer_actual_at else None,
                    'marked_no_longer_comment': asset.marked_no_longer_comment,
                },
                'owners': [
                    {
                        'name': o.name,
                        'department': str(o.department) if o.department is not None else None,
                        'position': str(o.position) if o.position is not None else None,
                        'email': o.email,
                        'phone': o.phone,
                    }
                    for o in owners
                ],
                'administrators': [
                    {
                        'name': a.name,
                        'department': str(a.department) if a.department is not None else None,
                        'position': str(a.position) if a.position is not None else None,
                        'email': a.email,
                        'phone': a.phone,
                    }
                    for a in administrators
                ],
                'software': [
                    {
                        'id': sw.id,
                        'name': sw.name,
                        'version': getattr(sw, 'version', ''),
                    }
                    for sw in software_entries
                ],
            },
            'vulnerabilities': vulnerabilities_data,
            'asset_vulnerabilities': asset_vulnerabilities_data,
            'has_vulnerabilities': vulnerabilities.exists()
        }

        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_access)
def export_risk_details(request):
    asset_id = request.GET.get('asset_id')
    language = request.GET.get('language', 'uk')
    
    if not asset_id:
        return JsonResponse({'error': 'Asset ID is required'}, status=400)
    
    try:
        asset = InformationAsset.objects.get(id=asset_id)
        
        # Get criticality with localization based on request language
        def get_localized_criticality(asset, language):
            levels = [
                (asset.confidentiality, asset.confidentiality.cost if asset.confidentiality else 0),
                (asset.integrity, asset.integrity.cost if asset.integrity else 0),
                (asset.availability, asset.availability.cost if asset.availability else 0)
            ]
            max_level = max(levels, key=lambda x: x[1])
            if max_level[0]:
                return {
                    'name': max_level[0].get_name() if max_level[0] else '',
                    'cost': max_level[0].cost,
                    'color': max_level[0].color
                }
            return {'name': _("Undefined"), 'cost': 0, 'color': "#000000"}
        
        criticality = get_localized_criticality(asset, language)
        
        # Get acceptable risk level for this asset
        acceptable_risk = get_acceptable_risk_for_asset(asset, language)
        
        # Get asset vulnerabilities
        asset_vulnerabilities = AssetVulnerability.objects.filter(asset=asset, status__in=['Yes', 'No']).select_related('vulnerability')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = _("Risk Details")
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0366D6', end_color='0366D6', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_border = Border(
            left=Side(style='thin', color='E1E1E1'),
            right=Side(style='thin', color='E1E1E1'),
            top=Side(style='thin', color='E1E1E1'),
            bottom=Side(style='thin', color='E1E1E1')
        )
        
        row_fills = [
            PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'),
            PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ]
        
        # Define headers - match the table headers from the modal
        headers = [
            _('Asset ID'), _('Asset Name'), _('Company'), _('Criticality'), 
            _('Vulnerability'), _('Status'), _('Description'), _('Threat'), _('Probability/Impact'), 
            _('Impact Val'), _('Impact Lev'), _('Risk Val'), _('Risk Lev'), 
            _('Exceeds Acceptable'), _('Acceptable Risk Level'), _('Risk Mitigation Controls')
        ]
        
        # Set headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:P1"
        
        # Collect and prepare risk data
        risks = []
        row_idx = 2
        
        for av in asset_vulnerabilities:
            for threat in av.vulnerability.threats.all():
                # For 'No' status vulnerabilities, use minimal values for probability and impact
                if av.status == 'No':
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                else:
                    probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)
                
                # Вплив загрози = Ймовірність (L) × Загальний вплив (E) × 100 для відсоткового формату
                threat_impact_value = probability * impact * 100
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                risk_level = calculate_risk_level(value_of_risk, company_id=asset.company_id)
                
                fill = row_fills[(row_idx - 2) % 2]
                
                # Get localized text for vulnerability
                vulnerability_text = av.vulnerability.get_name(language)
                vulnerability_description = av.vulnerability.get_description(language) or av.vulnerability.description
                
                # Check if risk level exceeds acceptable risk
                exceeds_acceptable = False
                acceptable_risk_name = ''
                if acceptable_risk and risk_level:
                    if isinstance(risk_level, RiskLevel) and risk_level.max_value > acceptable_risk['max_value']:
                        exceeds_acceptable = True
                    acceptable_risk_name = acceptable_risk['name']
                
                # Prepare row data
                row_data = [
                    asset.asset_id,
                    asset.name,
                    asset.company.name if asset.company else '',
                    criticality['name'],
                    vulnerability_text,
                    av.status,
                    vulnerability_description,
                    threat.get_name(language),
                    f"{probability:.4f} / {impact:.2f}",
                    f"{threat_impact_value:.4f}%",
                    threat_impact_level,
                    float(value_of_risk),
                    get_risk_level_name(risk_level, language),
                    'YES' if exceeds_acceptable else 'NO',
                    acceptable_risk_name,
                    getattr(av.vulnerability, f'risk_mitigation_controls_{language}', av.vulnerability.risk_mitigation_controls_uk)
                ]
                
                # Add data to worksheet
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = cell_border
                    cell.fill = fill
                    
                    # Format description column with wrap text
                    if col_idx == 7:  # Description column
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Format risk mitigation controls column with wrap text
                    if col_idx == 16:  # Risk Mitigation Controls column
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Set special formatting for criticality
                    if col_idx == 4:  # Criticality column
                        if 'color' in criticality:
                            color = criticality['color'].lstrip('#')
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            # Set font color based on background brightness
                            r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
                            brightness = (r * 299 + g * 587 + b * 114) / 1000
                            cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    
                    # Set special formatting for risk level
                    if col_idx == 13 and risk_level:  # Risk Level column
                        color = risk_level.color.lstrip('#')
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        # Set font color based on background brightness
                        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        cell.font = Font(color='FFFFFF' if brightness < 128 else '000000', bold=True)
                    
                    # Set special formatting for exceeds acceptable column
                    if col_idx == 14:  # Exceeds Acceptable column
                        if value == 'YES':
                            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            cell.font = Font(color='FFFFFF', bold=True)
                        else:
                            cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                            cell.font = Font(color='000000', bold=True)
                
                row_idx += 1
        
        # Adjust column widths
        column_widths = [12, 20, 20, 15, 25, 10, 40, 20, 15, 15, 10, 10, 15, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Set row height for better readability
        for row in range(2, row_idx):
            ws.row_dimensions[row].height = 30
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="risk_details_{asset.asset_id}.xlsx"'
        
        # Save the workbook to response
        wb.save(response)
        
        return response
    except InformationAsset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_access)
def get_impact_levels(request):
    """API endpoint to get all impact levels for threat assessment.
    Uses Country-based translations (e.g. Germany DE) when available; fallback to legacy uk/ru/en."""
    raw_lang = request.GET.get('language') or (get_language() or 'en')[:2]
    # Activate language so get_country_for_current_language() returns correct Country (e.g. DE for de)
    from django.utils import translation
    translation.activate(raw_lang)

    data = {
        'financial_impacts': [
            {
                'id': impact.id,
                'name': impact.get_name(),
                'description': impact.get_description(),
                'min_value': float(impact.min_value),
                'max_value': float(impact.max_value),
                'impact_value': float(impact.impact_value),
                'color': impact.color,
                'criteria': impact.get_criteria(),
                'examples': impact.get_examples()
            }
            for impact in FinancialImpact.objects.filter(is_active=True).order_by('min_value')
        ],
        'operational_impacts': [
            {
                'id': impact.id,
                'name': impact.get_name(),
                'description': impact.get_description(),
                'min_downtime_hours': float(impact.min_downtime_hours),
                'max_downtime_hours': float(impact.max_downtime_hours),
                'impact_value': float(impact.impact_value),
                'color': impact.color,
                'criteria': impact.get_criteria(),
                'examples': impact.get_examples()
            }
            for impact in OperationalImpact.objects.filter(is_active=True).order_by('min_downtime_hours')
        ],
        'reputational_impacts': [
            {
                'id': impact.id,
                'name': impact.get_name(),
                'description': impact.get_description(),
                'impact_value': float(impact.impact_value),
                'color': impact.color,
                'criteria': impact.get_criteria(),
                'examples': impact.get_examples()
            }
            for impact in ReputationalImpact.objects.filter(is_active=True).order_by('impact_value')
        ]
    }
    
    print(f"API: Returning data with {len(data['financial_impacts'])} financial, {len(data['operational_impacts'])} operational, {len(data['reputational_impacts'])} reputational impacts")
    return JsonResponse(data)


@login_required
@user_passes_test(has_risk_assessment_config_access)
def impact_settings_summary(request):
    """View for displaying impact settings summary dashboard"""
    from django.db.models import Count, Avg
    
    # Get summary statistics
    financial_impacts = FinancialImpact.objects.count()
    operational_impacts = OperationalImpact.objects.count()
    reputational_impacts = ReputationalImpact.objects.count()
    
    threats_with_impacts = Threat.objects.filter(
        Q(financial_impact__isnull=False) |
        Q(operational_impact__isnull=False) |
        Q(reputational_impact__isnull=False)
    ).count()
    
    total_threats = Threat.objects.count()
    
    # Get impact value ranges
    financial_range = FinancialImpact.objects.aggregate(
        min_impact=Avg('impact_value'),
        max_impact=Avg('impact_value')
    )
    
    operational_range = OperationalImpact.objects.aggregate(
        min_impact=Avg('impact_value'),
        max_impact=Avg('impact_value')
    )
    
    reputational_range = ReputationalImpact.objects.aggregate(
        min_impact=Avg('impact_value'),
        max_impact=Avg('impact_value')
    )
    
    summary_data = {
        'financial_impacts': financial_impacts,
        'operational_impacts': operational_impacts,
        'reputational_impacts': reputational_impacts,
        'threats_with_impacts': threats_with_impacts,
        'total_threats': total_threats,
        'impact_coverage': f"{(threats_with_impacts / total_threats * 100):.1f}%" if total_threats > 0 else "0%",
        'financial_range': financial_range,
        'operational_range': operational_range,
        'reputational_range': reputational_range,
    }
    
    context = {
        'summary_data': summary_data,
        'title': _('Impact Assessment Settings Summary'),
    }
    
    return render(request, 'admin/app_risk/impact_settings_summary.html', context)


@login_required
@user_passes_test(has_risk_assessment_access)
def get_treatment_reference_data(request):
    """Return reference data for treatment modal dropdowns"""
    from app_cabinet.models import CabinetUser
    from app_asset.models import InformationAsset
    
    User = get_user_model()
    asset_id = request.GET.get('asset_id')
    current_language = get_language()[:2]
    
    logger.debug(f"get_treatment_reference_data called with asset_id={asset_id}, language={current_language}")

    try:
        # Defaults so we never hit UnboundLocalError
        cabinet_users = []
        asset = None
        departments_seen = {}
        positions_seen = {}

        # Get Cabinet users filtered by asset's company if asset_id is provided
        if asset_id:
            try:
                # Get the asset and its company
                asset = InformationAsset.objects.select_related('company').get(
                    Q(asset_id=asset_id) | Q(id=asset_id)
                )
                
                if asset.company_id not in _get_permitted_risk_company_ids(request.user):
                    return JsonResponse({'error': 'Permission denied for selected company'}, status=403)
                
                # Filter Cabinet users by the asset's company
                cabinet_users_qs = CabinetUser.objects.filter(
                    company=asset.company,
                    user__is_active=True
                ).select_related('user', 'department', 'position')
                
                for cu in cabinet_users_qs:
                    dept_name = cu.department.get_name(current_language) if cu.department else ''
                    pos_name = cu.position.get_name(current_language) if cu.position else ''
                    if cu.department and cu.department_id not in departments_seen:
                        departments_seen[cu.department_id] = dept_name
                    if cu.position and cu.position_id not in positions_seen:
                        positions_seen[cu.position_id] = pos_name
                    cabinet_users.append({
                        'id': cu.user.id,
                        'username': cu.user.username,
                        'full_name': f"{cu.user.first_name} {cu.user.last_name}".strip() or cu.user.username,
                        'display_name': f"{cu.user.first_name} {cu.user.last_name}".strip() + 
                                       (f" ({dept_name}/{pos_name})" if cu.department and cu.position
                                        else f" ({dept_name})" if cu.department
                                        else f" ({pos_name})" if cu.position
                                        else ""),
                        'department': dept_name,
                        'department_id': cu.department_id,
                        'position': pos_name,
                        'position_id': cu.position_id,
                        'company': cu.company.name if cu.company else ''
                    })
            except Exception:
                # If asset not found or any other error, fall back to all active users
                cabinet_users = [
                    {
                        'id': user.id,
                        'username': user.username,
                        'full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
                        'display_name': f"{user.first_name} {user.last_name}".strip() or user.username,
                        'department': '',
                        'department_id': None,
                        'position': '',
                        'position_id': None,
                        'company': ''
                    }
                    for user in User.objects.filter(is_active=True)
                ]
        else:
            # If no asset_id provided, fall back to all active users
            cabinet_users = [
                {
                    'id': user.id,
                    'username': user.username,
                    'full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
                    'display_name': f"{user.first_name} {user.last_name}".strip() or user.username,
                    'department': '',
                    'department_id': None,
                    'position': '',
                    'position_id': None,
                    'company': ''
                }
                for user in User.objects.filter(is_active=True)
            ]
            
        departments_data = [{'id': k, 'name': v} for k, v in sorted(departments_seen.items())]
        positions_data = [{'id': k, 'name': v} for k, v in sorted(positions_seen.items())]
        
        data = {
            'residual_risk_levels': [
                {
                    'id': level.id,
                    'name': level.get_name(current_language),
                    'color': level.color,
                    'min_value': float(level.min_value) if level.min_value is not None else 0,
                    'max_value': float(level.max_value) if level.max_value is not None else 0
                }
                for level in get_company_risk_levels_queryset(asset.company_id if asset else None)
            ],
            'priority_levels': [
                {
                    'id': priority.id,
                    'name': priority.get_name(current_language),
                    'color': priority.color,
                    'value': float(priority.value) if priority.value is not None else 0,
                    'description_uk': priority.get_description('uk'),
                    'description_en': priority.get_description('en'),
                    'description_ru': priority.get_description('ru')
                }
                for priority in TreatmentPriority.objects.all()
            ],
            'effectiveness_levels': [
                {
                    'id': eff.id,
                    'name': eff.get_name(current_language),
                    'color': eff.color,
                    'value': float(eff.value) if eff.value is not None else 0,
                    'description_uk': eff.get_description('uk'),
                    'description_en': eff.get_description('en'),
                    'description_ru': eff.get_description('ru')
                }
                for eff in TreatmentEffectiveness.objects.all()
            ],
            'monitoring_frequencies': [
                {
                    'id': freq.id,
                    'name': freq.get_name(current_language),
                    'days': freq.days
                }
                for freq in MonitoringFrequency.objects.all()
            ],
            'users': cabinet_users,
            'cabinet_users': cabinet_users,
            'departments': departments_data,
            'positions': positions_data,
            'available_treatments': [
                {
                    'id': treatment.id,
                    'name': f"Treatment #{treatment.id}" + (f" - {treatment.asset.name}" if treatment.asset else " - No Asset")
                }
                for treatment in RiskTreatment.objects.select_related('asset').all()
            ],
            'available_assets': [
                {
                    'id': asset.id,
                    'name': asset.name
                }
                for asset in InformationAsset.objects.all()
            ]
        }
        
        return JsonResponse(data)
    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.exception("CRITICAL ERROR in get_treatment_reference_data")
        logger.error(f"Exception type: {type(e).__name__}")
        logger.error(f"Exception message: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        # Return a minimal but valid payload so UI can still work in degraded mode
        fallback = {
            'residual_risk_levels': [],
            'priority_levels': [],
            'effectiveness_levels': [],
            'monitoring_frequencies': [],
            'users': [],
            'cabinet_users': [],
            'departments': [],
            'positions': [],
            'available_treatments': [],
            'available_assets': [],
            'error': str(e),
        }
        return JsonResponse(fallback, status=500)


# Interactive Dashboard API Views
@login_required
@user_passes_test(has_risk_assessment_access)
@log_data_access_decorator("COMPANIES_API")
def get_companies_api(request):
    """API endpoint to get companies for filter dropdown"""
    try:
        user_permissions = get_user_risk_assessment_permissions(request.user)
        companies = Company.objects.filter(id__in=user_permissions['companies'])
        
        companies_data = [
            {
                'id': company.id,
                'name': company.name
            }
            for company in companies
        ]
        
        # Log the API access
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="COMPANIES_API_ACCESS",
            details={
                'companies_count': len(companies_data),
                'user_permissions': len(user_permissions['companies'])
            },
            request_path=request.path
        )
        
        return JsonResponse(companies_data, safe=False)
    except Exception as e:
        RiskAssessmentLogger.log_error(
            user=request.user,
            error_type="COMPANIES_API_ERROR",
            error_message=str(e),
            request_path=request.path
        )
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_access)
@log_risk_action("DASHBOARD_FILTERED_DATA")
def get_dashboard_filtered_data(request):
    """API endpoint to get filtered dashboard data"""
    try:
        # Get filter parameters
        time_range = request.GET.get('time_range', '30')
        company_id = request.GET.get('company', 'all')
        risk_level = request.GET.get('risk_level', 'all')
        
        logger.info(f"Dashboard filter request: company={company_id}, time_range={time_range}, risk_level={risk_level}")
        
        # Apply filters to the existing dashboard logic
        user_permissions = get_user_risk_assessment_permissions(request.user)
        
        # Filter by company if specified
        if company_id != 'all':
            companies = [int(company_id)]
            logger.info(f"Filtering by company: {company_id}")
        else:
            companies = user_permissions['companies']
            logger.info(f"Using all user companies: {companies}")
        
        # Filter by time range
        from datetime import datetime, timedelta
        if time_range != 'all':
            days = int(time_range)
            start_date = datetime.now() - timedelta(days=days)
            # Apply date filtering to queries (implementation depends on your models)
        
        # Get current language for localization
        current_language = get_language()[:2]
        
        # Get filtered data using existing dashboard logic
        vulnerability_data = AssetVulnerability.objects.filter(
            asset__company__id__in=companies
        ).aggregate(
            vulnerable=Count(Case(When(status='Yes', then=1))),
            not_vulnerable=Count(Case(When(status='No', then=1))),
            undefined=Count(Case(When(status='Undefined', then=1)))
        )
        
        # Apply risk level filtering if needed - CriticalityLevel uses name/name_local
        default_name = 'Невизначено' if current_language == 'uk' else 'Undefined' if current_language == 'en' else 'Неопределено'

        criticality_data = InformationAsset.objects.filter(
            company__id__in=companies
        ).annotate(
            criticality_name=Case(
                When(confidentiality__cost__gt=F('integrity__cost'), then=Coalesce(F('confidentiality__name_local'), F('confidentiality__name'))),
                When(confidentiality__cost__gt=F('availability__cost'), then=Coalesce(F('confidentiality__name_local'), F('confidentiality__name'))),
                When(integrity__cost__gt=F('confidentiality__cost'), then=Coalesce(F('integrity__name_local'), F('integrity__name'))),
                When(integrity__cost__gt=F('availability__cost'), then=Coalesce(F('integrity__name_local'), F('integrity__name'))),
                When(availability__cost__gt=F('confidentiality__cost'), then=Coalesce(F('availability__name_local'), F('availability__name'))),
                When(availability__cost__gt=F('integrity__cost'), then=Coalesce(F('availability__name_local'), F('availability__name'))),
                default=Value(default_name),
                output_field=CharField()
            ),
            color=Case(
                When(confidentiality__cost__gt=F('integrity__cost'), then=F('confidentiality__color')),
                When(confidentiality__cost__gt=F('availability__cost'), then=F('confidentiality__color')),
                When(integrity__cost__gt=F('confidentiality__cost'), then=F('integrity__color')),
                When(integrity__cost__gt=F('availability__cost'), then=F('integrity__color')),
                When(availability__cost__gt=F('confidentiality__cost'), then=F('availability__color')),
                When(availability__cost__gt=F('integrity__cost'), then=F('availability__color')),
                default=Value('#808080'),
                output_field=CharField()
            )
        ).values('criticality_name', 'color').annotate(
            count=Count('id')
        ).order_by('criticality_name')
        
        # Get treatment data (active types only)
        treatment_data = []
        treatment_types = Treatment_type.objects.filter(is_active=True).order_by('name', 'code')
        
        for treatment_type in treatment_types:
            treatment_count = {
                'treatment_type': treatment_type.get_name(current_language),
                'treatment_type_code': treatment_type.code,
                'planned': RiskTreatment.objects.filter(
                    treatment_type=treatment_type,
                    status__code='Planned',
                    asset__company__id__in=companies
                ).count(),
                'in_progress': RiskTreatment.objects.filter(
                    treatment_type=treatment_type,
                    status__code='In Progress',
                    asset__company__id__in=companies
                ).count(),
                'completed': RiskTreatment.objects.filter(
                    treatment_type=treatment_type,
                    status__code='Completed',
                    asset__company__id__in=companies
                ).count()
            }
            treatment_data.append(treatment_count)
        
        # Get Treatment Details data for enhanced charts
        from .models import TreatmentEffectiveness, TreatmentPriority, MonitoringFrequency
        
        # Effectiveness data
        effectiveness_data = []
        effectiveness_levels = TreatmentEffectiveness.objects.all()
        for effectiveness in effectiveness_levels:
            count = RiskTreatment.objects.filter(
                effectiveness=effectiveness,
                asset__company__id__in=companies
            ).count()
            if count > 0:
                effectiveness_data.append({
                    'name': effectiveness.get_name(current_language),
                    'value': effectiveness.value,
                    'count': count,
                    'color': effectiveness.color
                })
        
        # Priority data
        priority_data = []
        priority_levels = TreatmentPriority.objects.all()
        for priority in priority_levels:
            count = RiskTreatment.objects.filter(
                priority=priority,
                asset__company__id__in=companies
            ).count()
            if count > 0:
                priority_data.append({
                    'name': priority.get_name(current_language),
                    'value': priority.value,
                    'count': count,
                    'color': priority.color
                })
        
        # Residual Risk data
        residual_risk_data = []
        residual_risk_levels = get_company_risk_levels_queryset(None)
        for residual_risk in residual_risk_levels:
            count = RiskTreatment.objects.filter(
                residual_risk_level=residual_risk,
                asset__company__id__in=companies
            ).count()
            if count > 0:
                residual_risk_data.append({
                    'name': residual_risk.get_name(current_language),
                    'value': residual_risk.max_value,
                    'count': count,
                    'color': residual_risk.color
                })
        
        # Monitoring Frequency data
        monitoring_data = []
        monitoring_frequencies = MonitoringFrequency.objects.all()
        for frequency in monitoring_frequencies:
            count = RiskTreatment.objects.filter(
                monitoring_frequency=frequency,
                asset__company__id__in=companies
            ).count()
            if count > 0:
                monitoring_data.append({
                    'name': frequency.get_name(current_language),
                    'count': count
                })
        
        # Cost analysis data
        cost_data = {
            'total_implementation_cost': 0,
            'total_maintenance_cost': 0,
            'treatments_with_cost': 0,
            'cost_by_priority': []
        }
        
        treatments_with_costs = RiskTreatment.objects.filter(
            asset__company__id__in=companies,
            implementation_cost__isnull=False
        ).exclude(implementation_cost=0)
        
        if treatments_with_costs.exists():
            cost_data['total_implementation_cost'] = treatments_with_costs.aggregate(
                total=Sum('implementation_cost')
            )['total'] or 0
            
            maintenance_costs = RiskTreatment.objects.filter(
                asset__company__id__in=companies,
                annual_maintenance_cost__isnull=False
            ).exclude(annual_maintenance_cost=0)
            
            if maintenance_costs.exists():
                cost_data['total_maintenance_cost'] = maintenance_costs.aggregate(
                    total=Sum('annual_maintenance_cost')
                )['total'] or 0
            
            cost_data['treatments_with_cost'] = treatments_with_costs.count()
            
            # Cost by priority
            for priority in priority_levels:
                priority_cost = treatments_with_costs.filter(priority=priority).aggregate(
                    total=Sum('implementation_cost')
                )['total'] or 0
                if priority_cost > 0:
                    cost_data['cost_by_priority'].append({
                        'priority': priority.get_name(current_language),
                        'cost': float(priority_cost),
                        'color': priority.color
                    })
        
        # Generate risk matrix data (simplified)
        risk_matrix_data = [
            {'probability': 'Low', 'impact': 'Low', 'count': 2},
            {'probability': 'Medium', 'impact': 'Medium', 'count': 5},
            {'probability': 'High', 'impact': 'High', 'count': 3}
        ]
        
        # Get translations
        status_translations = {}
        treatment_type_translations = {}
        
        for status in Treatment_status.objects.all():
            status_translations[status.code] = status.get_name(current_language)
            
        for ttype in Treatment_type.objects.all():
            treatment_type_translations[ttype.code] = ttype.get_name(current_language)
        
        translations = {
            'planned': _('Planned'),
            'in_progress': _('In Progress'),
            'completed': _('Completed'),
            'risk_treatment_progress': _('Risk Treatment Progress'),
            'status': status_translations,
            'treatment_type': treatment_type_translations
        }
        
        # Log the filtered data request
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="DASHBOARD_FILTERED_DATA_REQUEST",
            details={
                'filters': {
                    'time_range': time_range,
                    'company': company_id,
                    'risk_level': risk_level
                },
                'results': {
                    'vulnerability_count': sum(vulnerability_data.values()),
                    'criticality_categories': len(list(criticality_data)),
                    'treatment_types': len(treatment_data)
                }
            },
            request_path=request.path
        )
        
        # Calculate basic statistics for Quick Statistics component
        total_assets = InformationAsset.objects.filter(company__id__in=companies).count()
        total_vulnerabilities = AssetVulnerability.objects.filter(
            asset__company__id__in=companies,
            status='Yes'
        ).count()
        
        # Calculate high risk count
        high_risk_count = 0
        try:
            # This is a simplified calculation - you may want to use your existing logic
            high_risk_assets = InformationAsset.objects.filter(company__id__in=companies)
            for asset in high_risk_assets:
                criticality = asset.get_criticality()
                if criticality['cost'] >= 3:  # Assuming high criticality is 3 or higher
                    asset_vulnerabilities = AssetVulnerability.objects.filter(asset=asset, status='Yes')
                    if asset_vulnerabilities.exists():
                        high_risk_count += 1
        except:
            pass
        
        statistics = {
            'total_assets': total_assets,
            'total_vulnerabilities': total_vulnerabilities,
            'high_risk_count': high_risk_count
        }
        
        logger.info(f"Statistics calculated: {statistics}")

        return JsonResponse({
            'vulnerability_data': vulnerability_data,
            'criticality_data': list(criticality_data),
            'treatment_data': treatment_data,
            'risk_matrix_data': risk_matrix_data,
            'statistics': statistics,
            'translations': translations,
            'filters_applied': {
                'time_range': time_range,
                'company': company_id,
                'risk_level': risk_level
            }
        })
        
    except Exception as e:
        logger.error(f"Error in get_dashboard_filtered_data: {str(e)}", exc_info=True)
        RiskAssessmentLogger.log_error(
            user=request.user,
            error_type="DASHBOARD_FILTERED_DATA_ERROR",
            error_message=str(e),
            request_path=request.path,
            additional_context={
                'filters': {
                    'time_range': time_range,
                    'company': company_id,
                    'risk_level': risk_level
                }
            }
        )
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_access)
def get_drilldown_data(request):
    """API endpoint to get detailed drilldown data for charts"""
    try:
        chart_type = request.GET.get('chart_type')
        label = request.GET.get('label')
        current_language = get_language()[:2]
        
        user_permissions = get_user_risk_assessment_permissions(request.user)
        
        if chart_type == 'vulnerability':
            # Get assets with specific vulnerability status
            if label == 'Vulnerable':
                assets = InformationAsset.objects.filter(
                    company__id__in=user_permissions['companies'],
                    assetvulnerability__status='Yes'
                ).distinct()
            elif label == 'Not Vulnerable':
                assets = InformationAsset.objects.filter(
                    company__id__in=user_permissions['companies'],
                    assetvulnerability__status='No'
                ).distinct()
            else:  # Undefined
                assets = InformationAsset.objects.filter(
                    company__id__in=user_permissions['companies'],
                    assetvulnerability__status='Undefined'
                ).distinct()
                
        elif chart_type == 'criticality':
            # Get assets with specific criticality level - CriticalityLevel uses name/name_local
            default_name = 'Невизначено' if current_language == 'uk' else 'Undefined' if current_language == 'en' else 'Неопределено'

            assets = InformationAsset.objects.filter(
                company__id__in=user_permissions['companies']
            ).annotate(
                criticality_name=Case(
                    When(confidentiality__cost__gt=F('integrity__cost'), then=Coalesce(F('confidentiality__name_local'), F('confidentiality__name'))),
                    When(confidentiality__cost__gt=F('availability__cost'), then=Coalesce(F('confidentiality__name_local'), F('confidentiality__name'))),
                    When(integrity__cost__gt=F('confidentiality__cost'), then=Coalesce(F('integrity__name_local'), F('integrity__name'))),
                    When(integrity__cost__gt=F('availability__cost'), then=Coalesce(F('integrity__name_local'), F('integrity__name'))),
                    When(availability__cost__gt=F('confidentiality__cost'), then=Coalesce(F('availability__name_local'), F('availability__name'))),
                    When(availability__cost__gt=F('integrity__cost'), then=Coalesce(F('availability__name_local'), F('availability__name'))),
                    default=Value(default_name),
                    output_field=CharField()
                )
            ).filter(criticality_name=label)
        
        else:
            assets = InformationAsset.objects.filter(
                company__id__in=user_permissions['companies']
            )[:10]  # Limit to 10 for demo
        
        # Format asset data for response
        assets_data = []
        for asset in assets[:20]:  # Limit to 20 assets
            criticality = asset.get_criticality()
            
            # Get risk level for this asset
            risk_level = _('Medium')  # Default
            try:
                # Calculate actual risk level based on vulnerabilities
                asset_vulnerabilities = AssetVulnerability.objects.filter(asset=asset, status='Yes')
                if asset_vulnerabilities.exists():
                    # Get highest risk from vulnerabilities
                    highest_risk = 0
                    for av in asset_vulnerabilities:
                        for threat in av.vulnerability.threats.all():
                            threat_impact_value = calculate_threat_impact_value(threat.probability, threat.impact)
                            threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                            value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                            if value_of_risk > highest_risk:
                                highest_risk = value_of_risk
                    
                    risk_level_obj = get_risk_level(highest_risk)
                    if risk_level_obj:
                        risk_level = risk_level_obj.get_name_by_language(current_language) or risk_level_obj.get_name()
            except:
                pass
            
            # Get treatment status
            treatment_status = _('Not Treated')
            try:
                treatments = RiskTreatment.objects.filter(asset=asset)
                if treatments.exists():
                    completed_treatments = treatments.filter(status__code='Completed').count()
                    total_treatments = treatments.count()
                    if completed_treatments == total_treatments:
                        treatment_status = _('Fully Treated')
                    elif completed_treatments > 0:
                        treatment_status = _('Partially Treated')
                    else:
                        treatment_status = _('In Progress')
            except:
                pass
            
            assets_data.append({
                'id': asset.asset_id,
                'name': asset.name,
                'risk': risk_level,
                'status': treatment_status,
                'company': asset.company.name if asset.company else '',
                'criticality': criticality['name']
            })
        
        return JsonResponse({
            'assets': assets_data,
            'total_count': len(assets_data),
            'chart_type': chart_type,
            'label': label
        })
        
    except Exception as e:
        logger.error(f"Error in get_drilldown_data: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_access)
def get_realtime_updates(request):
    """API endpoint to check for real-time updates"""
    try:
        last_check = request.GET.get('last_check')
        
        # Get current data timestamp (simplified)
        from django.utils import timezone
        current_time = timezone.now()
        
        # Check if there are any recent changes
        # This is a simplified implementation - in production you'd check actual modification timestamps
        recent_changes = {
            'has_updates': False,
            'last_update': current_time.isoformat(),
            'changes': []
        }
        
        # Check for recent asset vulnerability changes
        if last_check:
            try:
                last_check_time = timezone.datetime.fromisoformat(last_check.replace('Z', '+00:00'))
                recent_av_changes = AssetVulnerability.objects.filter(
                    modified_at__gt=last_check_time
                ).count()
                
                if recent_av_changes > 0:
                    recent_changes['has_updates'] = True
                    recent_changes['changes'].append({
                        'type': 'vulnerability',
                        'count': recent_av_changes,
                        'message': f'{recent_av_changes} vulnerability assessments updated'
                    })
            except:
                pass
        
        # Check for recent treatment changes
        if last_check:
            try:
                last_check_time = timezone.datetime.fromisoformat(last_check.replace('Z', '+00:00'))
                recent_treatment_changes = RiskTreatment.objects.filter(
                    last_modified__gt=last_check_time
                ).count()
                
                if recent_treatment_changes > 0:
                    recent_changes['has_updates'] = True
                    recent_changes['changes'].append({
                        'type': 'treatment',
                        'count': recent_treatment_changes,
                        'message': f'{recent_treatment_changes} risk treatments updated'
                    })
            except:
                pass
        
        return JsonResponse(recent_changes)
        
    except Exception as e:
        logger.error(f"Error in get_realtime_updates: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_config_access)
def get_acceptable_risk_data(request):
    """API endpoint to get acceptable risk data for DataTables"""
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        current_language = get_language()[:2]
        
        # Get user permissions
        from .access_utils import get_user_risk_permissions
        user_permissions = get_user_risk_permissions(request.user)
        
        # Filter by user's companies
        queryset = AcceptableRisk.objects.filter(
            company__id__in=user_permissions.get('companies', [])
        ).select_related(
            'company', 'asset_group', 'asset_type', 'criticality_level', 'acceptable_risk_level'
        )
        
        # Apply search
        if search_value:
            queryset = queryset.filter(
                Q(company__name__icontains=search_value) |
                Q(asset_group__name__icontains=search_value) |
                Q(asset_group__code__icontains=search_value) |
                Q(asset_group__abbreviation__icontains=search_value) |
                Q(asset_type__name__icontains=search_value) |
                Q(asset_type__code__icontains=search_value) |
                Q(asset_type__name_en__icontains=search_value) |
                Q(asset_type__name_ru__icontains=search_value) |
                Q(criticality_level__name__icontains=search_value) |
                Q(criticality_level__name_local__icontains=search_value) |
                Q(acceptable_risk_level__name__icontains=search_value) |
                Q(acceptable_risk_level__translations__name_local__icontains=search_value)
            )
        
        total_records = queryset.count()
        filtered_records = queryset.count()
        
        # Apply pagination
        acceptable_risks = queryset[start:start + length]
        
        data = []
        for ar in acceptable_risks:
            data.append({
                'id': ar.id,
                'company': ar.company.name,
                'asset_group': ar.asset_group.name_local or ar.asset_group.name,
                'asset_type': ar.asset_type.name_local or ar.asset_type.name if ar.asset_type else '',
                'criticality_level': ar.criticality_level.get_name(),
                'criticality_cost': ar.criticality_level.cost,
                'criticality_color': ar.criticality_level.color,
                'acceptable_risk_level': ar.get_acceptable_risk_level_name(current_language),
                'acceptable_risk_color': ar.acceptable_risk_level.color,
                'created_at': ar.created_at.strftime('%Y-%m-%d %H:%M:%S') if ar.created_at else '',
                'updated_at': ar.updated_at.strftime('%Y-%m-%d %H:%M:%S') if ar.updated_at else '',
                'created_by': ar.created_by.get_full_name() if ar.created_by else '',
                'updated_by': ar.updated_by.get_full_name() if ar.updated_by else ''
            })
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        logger.error(f"Error in get_acceptable_risk_data: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
@user_passes_test(has_risk_assessment_config_access)
def save_acceptable_risk(request):
    """API endpoint to save acceptable risk settings"""
    try:
        data = json.loads(request.body)
        
        company_id = data.get('company_id')
        asset_group_id = data.get('asset_group_id')
        asset_type_id = data.get('asset_type_id')  # Can be None
        criticality_level_id = data.get('criticality_level_id')
        acceptable_risk_level_id = data.get('acceptable_risk_level_id')
        
        # Validate required fields
        if not all([company_id, asset_group_id, criticality_level_id, acceptable_risk_level_id]):
            return JsonResponse({
                'status': 'error',
                'message': 'All fields are required'
            }, status=400)

        allowed_companies = _get_permitted_risk_company_ids(request.user)
        if int(company_id) not in allowed_companies:
            return JsonResponse({
                'status': 'error',
                'message': 'Permission denied for selected company'
            }, status=403)

        allowed_risk_level_ids = set(get_company_risk_levels_queryset(company_id).values_list('id', flat=True))
        if int(acceptable_risk_level_id) not in allowed_risk_level_ids:
            return JsonResponse({
                'status': 'error',
                'message': 'Selected Risk Level is not available for the chosen company'
            }, status=400)
        
        # Check if record already exists
        acceptable_risk, created = AcceptableRisk.objects.get_or_create(
            company_id=company_id,
            asset_group_id=asset_group_id,
            asset_type_id=asset_type_id,
            criticality_level_id=criticality_level_id,
            defaults={
                'acceptable_risk_level_id': acceptable_risk_level_id,
                'created_by': request.user,
                'updated_by': request.user
            }
        )
        
        if not created:
            # Update existing record
            acceptable_risk.acceptable_risk_level_id = acceptable_risk_level_id
            acceptable_risk.updated_by = request.user
            acceptable_risk.save()
            message = 'Acceptable risk settings updated successfully'
        else:
            message = 'Acceptable risk settings created successfully'
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'id': acceptable_risk.id
        })
        
    except Exception as e:
        logger.error(f"Error in save_acceptable_risk: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(has_risk_assessment_config_access)
def delete_acceptable_risk(request, risk_id):
    """API endpoint to delete acceptable risk settings"""
    try:
        acceptable_risk = get_object_or_404(AcceptableRisk, id=risk_id)
        acceptable_risk.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Acceptable risk settings deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error in delete_acceptable_risk: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(has_risk_assessment_config_access)
def get_acceptable_risk_reference_data(request):
    """API endpoint to get reference data for acceptable risk form"""
    try:
        current_language = get_language()[:2]
        
        # Get user permissions
        from .access_utils import get_user_risk_permissions
        user_permissions = get_user_risk_permissions(request.user)
        selected_company_id = request.GET.get('company_id')
        allowed_company_ids = set(user_permissions.get('companies', []))
        if selected_company_id and int(selected_company_id) not in allowed_company_ids:
            return JsonResponse({'error': 'Permission denied for selected company'}, status=403)
        
        data = {
            'companies': [
                {
                    'id': company.id,
                    'name': company.name
                }
                for company in Company.objects.filter(id__in=user_permissions.get('companies', []))
            ],
            'asset_groups': [
                {
                    'id': group.id,
                    'name': group.get_name()
                }
                for group in AssetGroup.objects.all()
            ],
            'asset_types': [
                {
                    'id': asset_type.id,
                    'name': asset_type.get_name(),
                    'group_id': asset_type.group.id
                }
                for asset_type in AssetType.objects.all().order_by('group__name', 'name')
            ],
            'criticality_levels': [
                {
                    'id': level.id,
                    'name': level.get_name(),
                    'cost': level.cost,
                    'color': level.color
                }
                for level in CriticalityLevel.objects.all().order_by('cost')
            ],
            'risk_levels': [
                {
                    'id': level.id,
                    'name': level.get_name_by_language(current_language) or level.get_name(),
                    'color': level.color
                }
                for level in get_company_risk_levels_queryset(selected_company_id)
            ]
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f"Error in get_acceptable_risk_reference_data: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(has_risk_assessment_access)
@require_POST
def upload_treatment_attachment(request):
    """API endpoint to upload attachment for risk treatment"""
    try:
        treatment_id = request.POST.get('treatment_id')
        file = request.FILES.get('file')
        description = request.POST.get('description', '')
        
        if not treatment_id or not file:
            return JsonResponse({
                'status': 'error',
                'message': 'Treatment ID and file are required'
            }, status=400)
        
        # Validate treatment exists and user has access
        treatment = get_object_or_404(RiskTreatment, id=treatment_id)
        
        # Check file size (max 10MB)
        if file.size > 10 * 1024 * 1024:
            return JsonResponse({
                'status': 'error',
                'message': 'File size must be less than 10MB'
            }, status=400)
        
        # Check file type
        allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.txt', '.jpg', '.jpeg', '.png', '.gif']
        import os
        file_extension = os.path.splitext(file.name)[1].lower()
        if file_extension not in allowed_extensions:
            return JsonResponse({
                'status': 'error',
                'message': f'File type not allowed. Allowed types: {", ".join(allowed_extensions)}'
            }, status=400)
        
        # Create attachment
        attachment = RiskTreatmentAttachment.objects.create(
            treatment=treatment,
            file=file,
            filename=file.name,
            file_size=file.size,
            file_type=file_extension,
            uploaded_by=request.user,
            description=description
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'File uploaded successfully',
            'attachment': {
                'id': attachment.id,
                'filename': attachment.filename,
                'file_size': attachment.get_file_size_display(),
                'file_type': attachment.file_type,
                'uploaded_by': attachment.uploaded_by.get_full_name() if attachment.uploaded_by else '',
                'uploaded_at': format_time(attachment.uploaded_at),
                'description': attachment.description or '',
                'file_url': attachment.file.url if attachment.file else ''
            }
        })
        
    except Exception as e:
        logger.error(f"Error in upload_treatment_attachment: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
@user_passes_test(has_risk_assessment_access)
@require_POST
def delete_treatment_attachment(request):
    """API endpoint to delete attachment from risk treatment"""
    try:
        attachment_id = request.POST.get('attachment_id')
        
        if not attachment_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Attachment ID is required'
            }, status=400)
        
        # Get attachment and check permissions
        attachment = get_object_or_404(RiskTreatmentAttachment, id=attachment_id)
        
        # Check if user can delete (uploader or admin)
        if not (request.user.is_staff or attachment.uploaded_by == request.user):
            return JsonResponse({
                'status': 'error',
                'message': 'You do not have permission to delete this file'
            }, status=403)
        
        # Delete file and record
        attachment.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': 'File deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error in delete_treatment_attachment: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
@require_POST
@log_risk_action("BULK_UPDATE_RISK_TREATMENTS")
def bulk_update_risk_treatments(request):
    """API endpoint to bulk update treatment type and status for multiple treatments"""
    try:
        data = json.loads(request.body)
        treatment_ids = data.get('treatment_ids', [])
        treatment_type = data.get('treatment_type')
        status = data.get('status')
        
        if not treatment_ids:
            return JsonResponse({'success': False, 'message': 'No treatments selected'}, status=400)
        
        if not treatment_type and not status:
            return JsonResponse({'success': False, 'message': 'No changes specified'}, status=400)
        
        # Log the bulk update action
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="BULK_UPDATE_RISK_TREATMENTS_START",
            details={
                'treatment_ids': treatment_ids,
                'treatment_type': treatment_type,
                'status': status,
                'treatments_count': len(treatment_ids)
            },
            request_path=request.path
        )
        
        updated_treatments = []
        
        with transaction.atomic():
            for treatment_id in treatment_ids:
                treatment = RiskTreatment.objects.filter(id=treatment_id).first()
                if not treatment:
                    continue
                
                # Check if any fields have changed
                has_changes = False
                old_data = {
                    'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                    'status': treatment.status.code if treatment.status else None
                }
                
                # Update treatment type if provided
                if treatment_type:
                    treatment_type_obj = Treatment_type.objects.filter(code=treatment_type).first()
                    if treatment_type_obj and treatment.treatment_type != treatment_type_obj:
                        treatment.treatment_type = treatment_type_obj
                        has_changes = True
                
                # Update status if provided
                if status:
                    status_obj = Treatment_status.objects.filter(code=status).first()
                    if status_obj and treatment.status != status_obj:
                        treatment.status = status_obj
                        has_changes = True
                
                if has_changes:
                    treatment.last_modified = timezone.now()
                    treatment.last_modified_by = request.user
                    treatment.save()
                    
                    # Log the change
                    new_data = {
                        'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                        'status': treatment.status.code if treatment.status else None
                    }
                    
                    RiskAssessmentLogger.log_data_modification(
                        user=request.user,
                        operation="BULK_UPDATE",
                        data_type="RISK_TREATMENT",
                        data_before=old_data,
                        data_after=new_data,
                        asset_id=treatment.asset.asset_id,
                        request_path=request.path
                    )
                    
                    updated_treatments.append({
                        'id': treatment.id,
                        'last_modified': format_time(treatment.last_modified),
                        'last_modified_by': treatment.last_modified_by.get_full_name() if treatment.last_modified_by else ''
                    })
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully updated {len(updated_treatments)} treatments',
            'updated_treatments': updated_treatments
        })
        
    except Exception as e:
        logger.error(f"Error in bulk update risk treatments: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@user_passes_test(has_risk_assessment_access)
def get_risk_levels_for_editing(request):
    """API endpoint to get all available risk levels for manual editing"""
    try:
        current_language = get_language()[:2]
        asset_id = request.GET.get('asset_id')
        company_id = request.GET.get('company_id')
        allowed_company_ids = _get_permitted_risk_company_ids(request.user)

        if asset_id:
            asset = InformationAsset.objects.filter(Q(id=asset_id) | Q(asset_id=asset_id)).select_related('company').first()
            if not asset:
                return JsonResponse({'error': 'Asset not found'}, status=404)
            if asset.company_id not in allowed_company_ids:
                return JsonResponse({'error': 'Permission denied for selected company'}, status=403)
            company_id = asset.company_id
        elif company_id:
            if int(company_id) not in allowed_company_ids:
                return JsonResponse({'error': 'Permission denied for selected company'}, status=403)
        else:
            company_id = next(iter(allowed_company_ids), None)

        risk_levels = get_company_risk_levels_queryset(company_id)
        
        data = [
            {
                'id': level.id,
                'name': level.get_name(current_language),
                'color': level.color,
                'min_value': level.min_value,
                'max_value': level.max_value
            }
            for level in risk_levels
        ]
        
        return JsonResponse({'risk_levels': data})
        
    except Exception as e:
        logger.error(f"Error in get_risk_levels_for_editing: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(can_edit_risk_assessment)
@require_POST
@log_risk_action("UPDATE_RISK_LEVEL_MANUAL")
def update_risk_level_manual(request):
    """API endpoint to manually update risk level for a specific asset vulnerability threat combination"""
    try:
        data = json.loads(request.body)
        asset_id = data.get('asset_id')
        vulnerability_id = data.get('vulnerability_id')
        threat_id = data.get('threat_id')
        new_risk_level_id = data.get('risk_level_id')
        justification = data.get('justification', '')
        
        if not all([asset_id, vulnerability_id, threat_id, new_risk_level_id]):
            return JsonResponse({
                'success': False,
                'message': 'Asset ID, Vulnerability ID, Threat ID, and Risk Level ID are required'
            }, status=400)
        
        # Validate that all objects exist (asset_id can be database id or asset_id string)
        try:
            asset = InformationAsset.objects.filter(
                Q(id=asset_id) | Q(asset_id=asset_id)
            ).first()
            if not asset:
                raise InformationAsset.DoesNotExist(f'Asset not found: {asset_id}')
            vulnerability = Vulnerability.objects.get(id=vulnerability_id)
            threat = Threat.objects.get(id=threat_id)
            new_risk_level = RiskLevel.objects.get(id=new_risk_level_id)
        except (InformationAsset.DoesNotExist, Vulnerability.DoesNotExist, 
                Threat.DoesNotExist, RiskLevel.DoesNotExist) as e:
            return JsonResponse({
                'success': False,
                'message': f'One or more objects not found: {str(e)}'
            }, status=404)

        allowed_company_ids = _get_permitted_risk_company_ids(request.user)
        if asset.company_id not in allowed_company_ids:
            return JsonResponse({
                'success': False,
                'message': 'Permission denied for selected company'
            }, status=403)

        allowed_risk_level_ids = set(get_company_risk_levels_queryset(asset.company_id).values_list('id', flat=True))
        if int(new_risk_level_id) not in allowed_risk_level_ids:
            return JsonResponse({
                'success': False,
                'message': 'Selected Risk Level is not available for the asset company'
            }, status=400)
        
        # Check if AssetVulnerability exists
        try:
            asset_vulnerability = AssetVulnerability.objects.get(
                asset=asset,
                vulnerability=vulnerability
            )
        except AssetVulnerability.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Asset vulnerability relationship not found'
            }, status=404)
        
        # Log the manual risk level update
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="MANUAL_RISK_LEVEL_UPDATE",
            details={
                'asset_id': asset_id,
                'asset_name': asset.name,
                'vulnerability_id': vulnerability_id,
                'vulnerability_name': vulnerability.get_name(),
                'threat_id': threat_id,
                'threat_name': threat.get_name(),
                'old_risk_level': 'Calculated',  # Since we don't store the old calculated level
                'new_risk_level_id': new_risk_level_id,
                'new_risk_level_name': new_risk_level.get_name(),
                'justification': justification
            },
            asset_id=asset.asset_id,
            request_path=request.path
        )
        
        # Create or update a manual risk level override record
        from .models import ManualRiskLevelOverride
        
        try:
            override, created = ManualRiskLevelOverride.objects.get_or_create(
                asset=asset,
                vulnerability=vulnerability,
                threat=threat,
                defaults={
                    'manual_risk_level': new_risk_level,
                    'justification': justification,
                    'created_by': request.user,
                    'updated_by': request.user
                }
            )
            
            if not created:
                # Update existing override
                override.manual_risk_level = new_risk_level
                override.justification = justification
                override.updated_by = request.user
                override.save()
        except Exception as e:
            # If the table doesn't exist, log the warning but continue
            logger.warning(f"ManualRiskLevelOverride table not available for saving override: {e}")
            # Continue without saving the override
        
        # Get the calculated risk level for comparison
        criticality = asset.get_criticality()
        probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
        impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN) / 100
        threat_impact_value = probability * impact * 100
        threat_impact_level = calculate_threat_impact_level(threat_impact_value)
        calculated_value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
        calculated_risk_level = calculate_risk_level(calculated_value_of_risk)
        
        # Log the comparison
        RiskAssessmentLogger.log_data_modification(
            user=request.user,
            operation="MANUAL_OVERRIDE",
            data_type="RISK_LEVEL",
            data_before={
                'calculated_risk_level': calculated_risk_level.get_name() if calculated_risk_level else 'Unknown',
                'calculated_value': float(calculated_value_of_risk),
                'calculation_method': 'Automatic'
            },
            data_after={
                'manual_risk_level': new_risk_level.get_name(),
                'manual_value': new_risk_level.max_value,
                'calculation_method': 'Manual Override',
                'justification': justification
            },
            asset_id=asset.asset_id,
            request_path=request.path
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Risk level manually updated to {new_risk_level.get_name()}',
            'new_risk_level': {
                'id': new_risk_level.id,
                'name': new_risk_level.get_name(),
                'color': new_risk_level.color,
                'min_value': new_risk_level.min_value,
                'max_value': new_risk_level.max_value
            },
            'comparison': {
                'calculated_risk_level': calculated_risk_level.get_name() if calculated_risk_level else 'Unknown',
                'calculated_value': float(calculated_value_of_risk),
                'manual_risk_level': new_risk_level.get_name(),
                'manual_value': new_risk_level.max_value
            }
        })
        
    except Exception as e:
        logger.error(f"Error in update_risk_level_manual: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_POST
@log_risk_action("BULK_UPDATE_RISK_TREATMENTS")
def bulk_update_risk_treatments(request):
    """API endpoint to bulk update treatment type and status for multiple treatments"""
    try:
        data = json.loads(request.body)
        treatment_ids = data.get('treatment_ids', [])
        treatment_type = data.get('treatment_type')
        status = data.get('status')
        
        if not treatment_ids:
            return JsonResponse({'success': False, 'message': 'No treatments selected'}, status=400)
        
        if not treatment_type and not status:
            return JsonResponse({'success': False, 'message': 'No changes specified'}, status=400)
        
        # Log the bulk update action
        RiskAssessmentLogger.log_user_action(
            user=request.user,
            action="BULK_UPDATE_RISK_TREATMENTS_START",
            details={
                'treatment_ids': treatment_ids,
                'treatment_type': treatment_type,
                'status': status,
                'treatments_count': len(treatment_ids)
            },
            request_path=request.path
        )
        
        updated_treatments = []
        
        with transaction.atomic():
            for treatment_id in treatment_ids:
                treatment = RiskTreatment.objects.filter(id=treatment_id).first()
                if not treatment:
                    continue
                
                # Check if any fields have changed
                has_changes = False
                old_data = {
                    'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                    'status': treatment.status.code if treatment.status else None
                }
                
                # Update treatment type if provided
                if treatment_type:
                    treatment_type_obj = Treatment_type.objects.filter(code=treatment_type).first()
                    if treatment_type_obj and treatment.treatment_type != treatment_type_obj:
                        treatment.treatment_type = treatment_type_obj
                        has_changes = True
                
                # Update status if provided
                if status:
                    status_obj = Treatment_status.objects.filter(code=status).first()
                    if status_obj and treatment.status != status_obj:
                        treatment.status = status_obj
                        has_changes = True
                
                if has_changes:
                    treatment.last_modified = timezone.now()
                    treatment.last_modified_by = request.user
                    treatment.save()
                    
                    # Log the change
                    new_data = {
                        'treatment_type': treatment.treatment_type.code if treatment.treatment_type else None,
                        'status': treatment.status.code if treatment.status else None
                    }
                    
                    RiskAssessmentLogger.log_data_modification(
                        user=request.user,
                        operation="BULK_UPDATE",
                        data_type="RISK_TREATMENT",
                        data_before=old_data,
                        data_after=new_data,
                        asset_id=treatment.asset.asset_id,
                        request_path=request.path
                    )
                    
                    updated_treatments.append({
                        'id': treatment.id,
                        'last_modified': format_time(treatment.last_modified),
                        'last_modified_by': treatment.last_modified_by.get_full_name() if treatment.last_modified_by else ''
                    })
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully updated {len(updated_treatments)} treatments',
            'updated_treatments': updated_treatments
        })
        
    except Exception as e:
        logger.error(f"Error in bulk update risk treatments: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
@require_POST
@log_risk_action("GET_TREATMENT_HISTORY")
def get_treatment_history(request):
    """API endpoint to fetch the history of changes for a specific risk treatment"""
    try:
        data = json.loads(request.body)
        treatment_id = data.get('treatment_id')
        
        if not treatment_id:
            return JsonResponse({'success': False, 'message': 'Treatment ID is required'}, status=400)
        
        # Get the treatment and verify it exists
        treatment = RiskTreatment.objects.filter(id=treatment_id).first()
        if not treatment:
            return JsonResponse({'success': False, 'message': 'Treatment not found'}, status=404)
        
        # Get the history records for this treatment
        try:
            history_records = treatment.history.all().order_by('-changed_at')
            logger.info(f"Found {history_records.count()} history records for treatment {treatment_id}")
        except Exception as e:
            logger.error(f"Error accessing history for treatment {treatment_id}: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error accessing history: {str(e)}'}, status=500)
        
        history_data = []
        for record in history_records:
            # Get field name display safely
            try:
                field_name = record.get_field_name_display()
            except:
                field_name = record.field_name
            
            history_data.append({
                'field_name': field_name,
                'old_value': record.old_value or _('Not Set'),
                'new_value': record.new_value or _('Not Set'),
                'changed_at': record.get_formatted_timestamp(),
                'changed_by': record.get_changed_by_name(),
                'comment': record.change_reason or ''
            })
        
        return JsonResponse({
            'success': True,
            'treatment_id': treatment_id,
            'history': history_data
        })
        
    except Exception as e:
        logger.error(f"Error fetching treatment history: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ── Software Vulnerabilities ─────────────────────────────────────────────────

@login_required
@user_passes_test(has_risk_assessment_access)
def get_software_vulnerabilities(request):
    """Return vulnerability list + current statuses for a SoftwareRegister entry."""
    sw_id = request.GET.get('software_id')
    if not sw_id:
        return JsonResponse({'error': 'software_id is required'}, status=400)
    try:
        software = get_object_or_404(SoftwareRegister, id=sw_id)
        vulnerabilities = Vulnerability.objects.filter(
            asset_type=software.asset_type, is_active=True
        ) if software.asset_type else Vulnerability.objects.none()

        sw_vulns = SoftwareVulnerability.objects.filter(software_register=software)
        current_language = get_language()[:2]

        def get_localized_field(obj, field_name):
            if not obj:
                return ''
            return (
                getattr(obj, f'{field_name}_{current_language}', '') or
                getattr(obj, f'{field_name}_uk', '') or
                getattr(obj, field_name, '')
            )

        vulnerabilities_data = [
            {
                'id': v.id,
                'scope': get_localized_field(v, 'scope'),
                'vulnerability': v.get_name(current_language),
                'risk_mitigation_controls': get_localized_field(v, 'risk_mitigation_controls'),
                'pci_dss_requirement': get_localized_field(v, 'pci_dss_requirement'),
                'iso27001_requirement': get_localized_field(v, 'iso27001_requirement'),
                'note': get_localized_field(v, 'note'),
            }
            for v in vulnerabilities
        ]

        sw_vulns_data = {
            sv.vulnerability_id: {
                'id': sv.id,
                'status': sv.status,
                'comment': sv.comment,
            }
            for sv in sw_vulns
        }

        group_name = get_localized_field(software.group, 'name') if software.group else None
        asset_type_name = get_localized_field(software.asset_type, 'name') if software.asset_type else None

        context = {
            'asset': {
                'id': software.id,
                'name': software.name,
                'asset_id': f'S{software.id:06d}',
                'company': software.company.name if software.company else None,
                'group': group_name,
                'asset_type': asset_type_name,
            },
            'vulnerabilities': vulnerabilities_data,
            'asset_vulnerabilities': sw_vulns_data,
            'has_vulnerabilities': len(vulnerabilities_data) > 0,
        }
        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(has_risk_assessment_access)
@require_http_methods(['POST'])
def save_software_vulnerabilities(request):
    """Save vulnerability statuses for a SoftwareRegister entry."""
    sw_id = request.POST.get('software_id')
    vulnerabilities_data = json.loads(request.POST.get('vulnerabilities', '[]'))
    try:
        software = get_object_or_404(SoftwareRegister, id=sw_id)
        for vuln_data in vulnerabilities_data:
            vulnerability = Vulnerability.objects.get(id=vuln_data['id'])
            sv, created = SoftwareVulnerability.objects.get_or_create(
                software_register=software,
                vulnerability=vulnerability,
                defaults={'status': vuln_data['status'], 'comment': vuln_data.get('comment', '')}
            )
            if not created:
                sv.status = vuln_data['status']
                sv.comment = vuln_data.get('comment', '')
            sv.modified_by = request.user
            sv.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required
@user_passes_test(has_risk_assessment_access)
def get_software_risks(request):
    """Return risk details (vulnerability × threat × risk level) for a SoftwareRegister entry."""
    sw_id = request.GET.get('software_id')
    language = request.GET.get('language', get_language()[:2])
    if not sw_id:
        return JsonResponse({'error': 'software_id is required'}, status=400)
    try:
        software = get_object_or_404(SoftwareRegister, id=sw_id)

        # Compute criticality from CIA
        levels = [
            (software.confidentiality, software.confidentiality.cost if software.confidentiality else 0),
            (software.integrity, software.integrity.cost if software.integrity else 0),
            (software.availability, software.availability.cost if software.availability else 0),
        ]
        max_level = max(levels, key=lambda x: x[1])
        if max_level[0]:
            criticality = {'name': max_level[0].get_name(), 'cost': max_level[0].cost, 'color': max_level[0].color}
        else:
            criticality = {'name': str(_('Undefined')), 'cost': 0, 'color': '#000000'}

        sw_vulns = SoftwareVulnerability.objects.filter(
            software_register=software,
            status__in=['Yes', 'No']
        ).select_related('vulnerability')

        risks = []
        for sv in sw_vulns:
            vuln = sv.vulnerability
            threats = vuln.threats.all()
            if not threats.exists():
                if sv.status == 'No':
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                    threat_impact_value = probability * impact * 100
                    threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                    value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                    risk_level = calculate_risk_level(value_of_risk, company_id=software.company_id)
                    risks.append({
                        'asset_id': f'S{software.id:06d}',
                        'asset_name': software.name,
                        'criticality': {
                            'name': criticality['name'],
                            'cost': criticality['cost'],
                            'color': criticality['color'],
                        },
                        'vulnerability': vuln.get_name(language),
                        'vulnerability_description': (
                            vuln.get_translated_value('description', language)
                            if hasattr(vuln, 'get_translated_value')
                            else getattr(vuln, f'description_{language}', '')
                        ) or getattr(vuln, 'description', '') or '',
                        'threat': '',
                        'probability': str(probability),
                        'impact': str(impact),
                        'probability_impact': str(probability * impact),
                        'threat_impact_value': str(threat_impact_value),
                        'threat_impact_level': threat_impact_level,
                        'value_of_risk': str(value_of_risk),
                        'risk_level': {
                            'text': risk_level.get_name_by_language(language) if risk_level else '',
                            'name': risk_level.get_name_by_language(language) if risk_level else '',
                            'color': risk_level.color if risk_level else '',
                            'value': risk_level.max_value if risk_level else 0,
                        } if risk_level else None,
                        'risk_mitigation_controls': getattr(vuln, f'risk_mitigation_controls_{language}', '') or vuln.risk_mitigation_controls,
                        'status': sv.status,
                        'vulnerability_status': sv.status,
                    })
                continue

            for threat in threats:
                if sv.status == 'No':
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                else:
                    probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)

                threat_impact_value = calculate_threat_impact_value(probability, impact)
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                risk_level = calculate_risk_level(value_of_risk, company_id=software.company_id)

                risks.append({
                    'asset_id': f'S{software.id:06d}',
                    'asset_name': software.name,
                    'criticality': {
                        'name': criticality['name'],
                        'cost': criticality['cost'],
                        'color': criticality['color'],
                    },
                    'vulnerability': vuln.get_name(language),
                    'vulnerability_description': (
                        vuln.get_translated_value('description', language)
                        if hasattr(vuln, 'get_translated_value')
                        else getattr(vuln, f'description_{language}', '')
                    ) or getattr(vuln, 'description', '') or '',
                    'threat': threat.get_name(language) if hasattr(threat, 'get_name') else str(threat),
                    'probability': str(probability),
                    'impact': str(impact),
                    'probability_impact': str(probability * impact),
                    'threat_impact_value': str(threat_impact_value),
                    'threat_impact_level': threat_impact_level,
                    'value_of_risk': str(value_of_risk),
                    'risk_level': {
                        'text': risk_level.get_name_by_language(language) if risk_level else '',
                        'name': risk_level.get_name_by_language(language) if risk_level else '',
                        'color': risk_level.color if risk_level else '',
                        'value': risk_level.max_value if risk_level else 0,
                    } if risk_level else None,
                    'risk_mitigation_controls': getattr(vuln, f'risk_mitigation_controls_{language}', '') or vuln.risk_mitigation_controls,
                    'status': sv.status,
                    'vulnerability_status': sv.status,
                })

        return JsonResponse({
            'asset': {
                'asset_id': f'S{software.id:06d}',
                'name': software.name,
                'company': software.company.name if software.company else '',
            },
            'risks': risks,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ── External Media Vulnerabilities ───────────────────────────────────────────

@login_required
@user_passes_test(has_risk_assessment_access)
def get_external_media_vulnerabilities(request):
    """Return vulnerability list + current statuses for an ExternalMediaRegister entry."""
    em_id = request.GET.get('external_media_id')
    if not em_id:
        return JsonResponse({'error': 'external_media_id is required'}, status=400)
    try:
        media = get_object_or_404(ExternalMediaRegister, id=em_id)
        vulnerabilities = Vulnerability.objects.filter(
            asset_type=media.asset_type, is_active=True
        ) if media.asset_type else Vulnerability.objects.none()

        em_vulns = ExternalMediaVulnerability.objects.filter(external_media_register=media)
        current_language = get_language()[:2]

        def get_localized_field(obj, field_name):
            if not obj:
                return ''
            return (
                getattr(obj, f'{field_name}_{current_language}', '') or
                getattr(obj, f'{field_name}_uk', '') or
                getattr(obj, field_name, '')
            )

        vulnerabilities_data = [
            {
                'id': v.id,
                'scope': get_localized_field(v, 'scope'),
                'vulnerability': v.get_name(current_language),
                'risk_mitigation_controls': get_localized_field(v, 'risk_mitigation_controls'),
                'pci_dss_requirement': get_localized_field(v, 'pci_dss_requirement'),
                'iso27001_requirement': get_localized_field(v, 'iso27001_requirement'),
                'note': get_localized_field(v, 'note'),
            }
            for v in vulnerabilities
        ]

        em_vulns_data = {
            ev.vulnerability_id: {
                'id': ev.id,
                'status': ev.status,
                'comment': ev.comment,
            }
            for ev in em_vulns
        }

        group_name = get_localized_field(media.group, 'name') if media.group else None
        asset_type_name = get_localized_field(media.asset_type, 'name') if media.asset_type else None

        context = {
            'asset': {
                'id': media.id,
                'name': media.name,
                'asset_id': f'M{media.id:05d}',
                'company': media.company.name if media.company else None,
                'group': group_name,
                'asset_type': asset_type_name,
            },
            'vulnerabilities': vulnerabilities_data,
            'asset_vulnerabilities': em_vulns_data,
            'has_vulnerabilities': len(vulnerabilities_data) > 0,
        }
        return JsonResponse(context)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(has_risk_assessment_access)
@require_http_methods(['POST'])
def save_external_media_vulnerabilities(request):
    """Save vulnerability statuses for an ExternalMediaRegister entry."""
    em_id = request.POST.get('external_media_id')
    vulnerabilities_data = json.loads(request.POST.get('vulnerabilities', '[]'))
    try:
        media = get_object_or_404(ExternalMediaRegister, id=em_id)
        for vuln_data in vulnerabilities_data:
            vulnerability = Vulnerability.objects.get(id=vuln_data['id'])
            ev, created = ExternalMediaVulnerability.objects.get_or_create(
                external_media_register=media,
                vulnerability=vulnerability,
                defaults={'status': vuln_data['status'], 'comment': vuln_data.get('comment', '')}
            )
            if not created:
                ev.status = vuln_data['status']
                ev.comment = vuln_data.get('comment', '')
            ev.modified_by = request.user
            ev.save()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


@login_required
@user_passes_test(has_risk_assessment_access)
def get_external_media_risks(request):
    """Return risk details (vulnerability × threat × risk level) for an ExternalMediaRegister entry."""
    em_id = request.GET.get('external_media_id')
    language = request.GET.get('language', get_language()[:2])
    if not em_id:
        return JsonResponse({'error': 'external_media_id is required'}, status=400)
    try:
        media = get_object_or_404(ExternalMediaRegister, id=em_id)

        levels = [
            (media.confidentiality, media.confidentiality.cost if media.confidentiality else 0),
            (media.integrity, media.integrity.cost if media.integrity else 0),
            (media.availability, media.availability.cost if media.availability else 0),
        ]
        max_level = max(levels, key=lambda x: x[1])
        if max_level[0]:
            criticality = {'name': max_level[0].get_name(), 'cost': max_level[0].cost, 'color': max_level[0].color}
        else:
            criticality = {'name': str(_('Undefined')), 'cost': 0, 'color': '#000000'}

        em_vulns = ExternalMediaVulnerability.objects.filter(
            external_media_register=media,
            status__in=['Yes', 'No']
        ).select_related('vulnerability')

        risks = []
        for ev in em_vulns:
            vuln = ev.vulnerability
            threats = vuln.threats.all()
            if not threats.exists():
                if ev.status == 'No':
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                    threat_impact_value = probability * impact * 100
                    threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                    value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                    risk_level = calculate_risk_level(value_of_risk, company_id=media.company_id)
                    risks.append({
                        'asset_id': f'M{media.id:05d}',
                        'asset_name': media.name,
                        'criticality': {
                            'name': criticality['name'],
                            'cost': criticality['cost'],
                            'color': criticality['color'],
                        },
                        'vulnerability': vuln.get_name(language),
                        'vulnerability_description': (
                            vuln.get_translated_value('description', language)
                            if hasattr(vuln, 'get_translated_value')
                            else getattr(vuln, f'description_{language}', '')
                        ) or getattr(vuln, 'description', '') or '',
                        'threat': '',
                        'probability': str(probability),
                        'impact': str(impact),
                        'probability_impact': str(probability * impact),
                        'threat_impact_value': str(threat_impact_value),
                        'threat_impact_level': threat_impact_level,
                        'value_of_risk': str(value_of_risk),
                        'risk_level': {
                            'text': risk_level.get_name_by_language(language) if risk_level else '',
                            'name': risk_level.get_name_by_language(language) if risk_level else '',
                            'color': risk_level.color if risk_level else '',
                            'value': risk_level.max_value if risk_level else 0,
                        } if risk_level else None,
                        'risk_mitigation_controls': getattr(vuln, f'risk_mitigation_controls_{language}', '') or vuln.risk_mitigation_controls,
                        'status': ev.status,
                        'vulnerability_status': ev.status,
                    })
                continue

            for threat in threats:
                if ev.status == 'No':
                    probability = Decimal('0.001')
                    impact = Decimal('0.001')
                else:
                    probability = Decimal(str(threat.probability)).quantize(Decimal('0.0001'), rounding=ROUND_DOWN)
                    impact = Decimal(str(threat.impact)).quantize(Decimal('0.01'), rounding=ROUND_DOWN)

                threat_impact_value = calculate_threat_impact_value(probability, impact)
                threat_impact_level = calculate_threat_impact_level(threat_impact_value)
                value_of_risk = Decimal(str(criticality['cost'])) * Decimal(str(threat_impact_level))
                risk_level = calculate_risk_level(value_of_risk, company_id=media.company_id)

                risks.append({
                    'asset_id': f'M{media.id:05d}',
                    'asset_name': media.name,
                    'criticality': {
                        'name': criticality['name'],
                        'cost': criticality['cost'],
                        'color': criticality['color'],
                    },
                    'vulnerability': vuln.get_name(language),
                    'vulnerability_description': (
                        vuln.get_translated_value('description', language)
                        if hasattr(vuln, 'get_translated_value')
                        else getattr(vuln, f'description_{language}', '')
                    ) or getattr(vuln, 'description', '') or '',
                    'threat': threat.get_name(language) if hasattr(threat, 'get_name') else str(threat),
                    'probability': str(probability),
                    'impact': str(impact),
                    'probability_impact': str(probability * impact),
                    'threat_impact_value': str(threat_impact_value),
                    'threat_impact_level': threat_impact_level,
                    'value_of_risk': str(value_of_risk),
                    'risk_level': {
                        'text': risk_level.get_name_by_language(language) if risk_level else '',
                        'name': risk_level.get_name_by_language(language) if risk_level else '',
                        'color': risk_level.color if risk_level else '',
                        'value': risk_level.max_value if risk_level else 0,
                    } if risk_level else None,
                    'risk_mitigation_controls': getattr(vuln, f'risk_mitigation_controls_{language}', '') or vuln.risk_mitigation_controls,
                    'status': ev.status,
                    'vulnerability_status': ev.status,
                })

        return JsonResponse({
            'asset': {
                'asset_id': f'M{media.id:05d}',
                'name': media.name,
                'company': media.company.name if media.company else '',
            },
            'risks': risks,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ── Allowed Software ────────────────────────────────────────────────────────

@login_required
@user_passes_test(has_risk_assessment_config_access)
def get_allowed_software_data(request):
    """DataTable endpoint for Allowed Software list."""
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 25))
        search_value = request.GET.get('search[value]', '').strip()

        from .access_utils import get_user_risk_permissions
        user_permissions = get_user_risk_permissions(request.user)
        allowed_company_ids = user_permissions.get('companies', [])

        qs = AllowedSoftware.objects.filter(
            Q(company__id__in=allowed_company_ids) | Q(company__isnull=True)
        ).select_related('company', 'software_register', 'software_register__status', 'created_by')

        if search_value:
            qs = qs.filter(
                Q(company__name__icontains=search_value) |
                Q(software_register__name__icontains=search_value) |
                Q(software_register__manufacturer__icontains=search_value) |
                Q(notes__icontains=search_value)
            )

        total = qs.count()
        records = qs.order_by('company__name', 'software_register__name')[start:start + length]

        data = []
        for rec in records:
            sw = rec.software_register
            data.append({
                'id': rec.id,
                'company': rec.company.name if rec.company else str(_('All companies')),
                'software_id': f"S{sw.id:06d}",
                'software_name': sw.name,
                'manufacturer': sw.manufacturer or '',
                'group': sw.group.get_name() if sw.group else ('—' if not sw.category else sw.category.get_name()),
                'asset_type': sw.asset_type.get_name() if sw.asset_type else '',
                'status_name': sw.status.get_name() if sw.status else '',
                'status_color': sw.status.color if sw.status else '#aaa',
                'notes': rec.notes or '',
                'created_at': rec.created_at.strftime('%d.%m.%Y') if rec.created_at else '',
                'created_by': rec.created_by.get_full_name() or rec.created_by.username if rec.created_by else '',
            })

        return JsonResponse({
            'draw': draw,
            'recordsTotal': total,
            'recordsFiltered': total,
            'data': data,
        })
    except Exception as e:
        logger.error(f"Error in get_allowed_software_data: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(has_risk_assessment_config_access)
def get_allowed_software_reference_data(request):
    """Returns companies + software entries for the add/edit form selects."""
    try:
        from .access_utils import get_user_risk_permissions
        user_permissions = get_user_risk_permissions(request.user)
        allowed_company_ids = user_permissions.get('companies', [])
        company_id = request.GET.get('company_id')

        companies = [
            {'id': c.id, 'name': c.name}
            for c in Company.objects.filter(id__in=allowed_company_ids).order_by('name')
        ]

        sw_qs = SoftwareRegister.objects.filter(is_active=True).select_related(
            'status', 'company', 'group', 'asset_type'
        ).order_by('name')
        if company_id:
            sw_qs = sw_qs.filter(Q(company__isnull=True) | Q(company__id=company_id))

        software = [
            {
                'id': sw.id,
                'label': f"S{sw.id:06d} — {sw.name}" + (f" ({sw.company.name})" if sw.company else ''),
                'name': sw.name,
                'company_id': sw.company_id,
            }
            for sw in sw_qs
        ]

        return JsonResponse({'companies': companies, 'software': software})
    except Exception as e:
        logger.error(f"Error in get_allowed_software_reference_data: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(has_risk_assessment_config_access)
@require_POST
def save_allowed_software(request):
    """Create or update an AllowedSoftware record."""
    try:
        data = json.loads(request.body)
        record_id = data.get('id')
        company_id = data.get('company_id') or None
        software_register_id = data.get('software_register_id')
        notes = data.get('notes', '')

        if not software_register_id:
            return JsonResponse({'status': 'error', 'message': _('Software is required')}, status=400)

        from .access_utils import get_user_risk_permissions
        user_permissions = get_user_risk_permissions(request.user)
        allowed_company_ids = user_permissions.get('companies', [])

        if company_id and int(company_id) not in allowed_company_ids:
            return JsonResponse({'status': 'error', 'message': _('Permission denied for selected company')}, status=403)

        sw = get_object_or_404(SoftwareRegister, id=software_register_id)

        if record_id:
            obj = get_object_or_404(AllowedSoftware, id=record_id)
            obj.company_id = company_id
            obj.software_register = sw
            obj.notes = notes
            obj.save()
            msg = _('Allowed software updated successfully')
        else:
            if AllowedSoftware.objects.filter(company_id=company_id, software_register=sw).exists():
                return JsonResponse({'status': 'error', 'message': _('This software entry already exists for the selected company')}, status=400)
            obj = AllowedSoftware.objects.create(
                company_id=company_id,
                software_register=sw,
                notes=notes,
                created_by=request.user,
            )
            msg = _('Allowed software added successfully')

        return JsonResponse({'status': 'success', 'message': str(msg), 'id': obj.id})
    except Exception as e:
        logger.error(f"Error in save_allowed_software: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@user_passes_test(has_risk_assessment_config_access)
@require_http_methods(["GET", "POST"])
def delete_allowed_software(request, sw_id):
    """Delete an AllowedSoftware record."""
    try:
        obj = get_object_or_404(AllowedSoftware, id=sw_id)
        obj.delete()
        return JsonResponse({'status': 'success', 'message': str(_('Deleted successfully'))})
    except Exception as e:
        logger.error(f"Error in delete_allowed_software: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
