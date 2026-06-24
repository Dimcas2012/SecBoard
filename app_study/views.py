# D:\Python\MyProject\SecBoard\SecBoard\app_study\views.py
import json
import os
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.core.exceptions import PermissionDenied
from django.http import HttpResponseForbidden, Http404, FileResponse, JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.db.models import Sum, Q, Count, Case, When, IntegerField
from django.core.paginator import Paginator
import logging
from django.utils import timezone
import random
from django.contrib import messages
from .decorators import group_required, user_has_quiz_access
from  app_cabinet.models import CabinetUser
from .models import QuizAttempt, Quiz, QuizAnswer, Answer, Question, AccessQuiz, AccessPage, Page, PageManagerGuide, PageManagerGuideTranslation, QuizManagerGuide, QuizManagerGuideTranslation
from .forms import QuizForm, QuestionForm, AnswerForm
from .page_forms import PageForm
from .pagination_utils import get_study_table_page_size, STUDY_TABLE_PAGE_SIZE_OPTIONS
from django.utils.translation import gettext as _, get_language, activate as translation_activate
import hashlib
from app_doc.models import RegisterDocs, RelatedDocs
from app_conf.models import Country

# Import openpyxl for Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import io
from datetime import datetime



logger = logging.getLogger(__name__)

logger.debug("This is a debug message")
logger.info("This is an info message")
logger.warning("This is a warning message")
logger.error("This is an error message")


# Helper functions for Excel export
def set_column_widths(ws):
    """Set appropriate column widths for the worksheet"""
    column_widths = {
        'A': 5,   # № (Number)
        'B': 30,  # Full Name
        'C': 20,  # Company
        'D': 20,  # Department
        'E': 20,  # Position
        'F': 20,  # Completed
        'G': 15,  # Result
        'H': 15,  # Successful
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def apply_header_style(ws, header_row=1):
    """Apply styles to the header row"""
    # Use light blue header with dark text to match the screenshot
    header_font = Font(name='Arial', size=12, bold=True, color='000000')
    header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply style to all header cells
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Special style for the "Successful" header
    successful_cell = ws.cell(row=header_row, column=8)
    successful_cell.font = Font(name='Arial', size=12, bold=True, color='FF0000')


def apply_data_styles(ws, start_row=2, end_row=None):
    """Apply styles to data rows"""
    if not end_row:
        end_row = ws.max_row
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Apply styles to data rows - no alternating colors as per screenshot
    for row in range(start_row, end_row + 1):
        for col in range(1, 9):  # Updated to 8 columns + 1
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            # Apply alignment based on column
            if col in [1]:  # Number column
                cell.alignment = center_alignment
            elif col in [2, 3, 4, 5]:  # Text columns
                cell.alignment = left_alignment
            else:  # Numeric/date columns
                cell.alignment = center_alignment
            
            # Apply special formatting for the "Successful" column
            # This is now handled in the specific export functions


def create_title_row(ws, quiz_title, num_columns=8):  # Updated column count to 8
    """Create a title row at the top of the worksheet"""
    # Merge cells for the title
    ws.merge_cells(f'A1:{get_column_letter(num_columns)}1')
    
    # Set title content and style
    title_cell = ws['A1']
    title_cell.value = f"{quiz_title} - {_('Results')}"
    title_cell.font = Font(name='Arial', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    
    # Add border to the title row
    thick_border = Border(
        bottom=Side(style='medium')
    )
    title_cell.border = thick_border
    
    # Set row height
    ws.row_dimensions[1].height = 30
    
    # Return the next row index
    return 2


def format_excel_sheet(ws, quiz_title, has_title=True):
    """Apply all formatting to a worksheet"""
    start_row = 1
    
    if has_title:
        # Add title row and get the header row index
        start_row = create_title_row(ws, quiz_title) 
    
    # Add header row
    headers = [
        _("№"),       # New numbering column
        _("Full Name"),
        _("Company"),
        _("Department"),
        _("Position"),
        _("Completed"),
        _("Result"),
        _("Successful")
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    
    # Set column widths
    set_column_widths(ws)
    
    # Apply header styles
    apply_header_style(ws, header_row=start_row)
    
    # Return the row index where data should start
    return start_row + 1


def _sanitize_excel_sheet_name(name, used_names, max_len=31):
    """Excel worksheet title: strip invalid characters, enforce length, avoid duplicates."""
    invalid = '[]:*?/\\'
    for ch in invalid:
        name = name.replace(ch, '-')
    name = (name or '').strip() or str(_("Quiz"))
    base = name[:max_len]
    candidate = base
    n = 2
    while candidate in used_names:
        suffix = f" ({n})"
        room = max_len - len(suffix)
        candidate = (base[:room] + suffix) if room > 0 else suffix[-max_len:]
        n += 1
    used_names.add(candidate)
    return candidate


def _quiz_max_score_points(quiz):
    """Maximum achievable score (sum of all answer score fields), using prefetched relations when present."""
    total = 0
    for question in quiz.questions.all():
        for answer in question.answers.all():
            total += answer.score
    return total


def _quiz_export_user_id_to_cabinet(quiz, request_user, accessible_companies, filter_company_id):
    """
    Map user_id -> CabinetUser for users who have access to the quiz (same rules as quiz_manager).
    When filter_company_id is set, only that company is considered for company_ids.
    """
    quiz_companies = list(quiz.companies.all())
    company_ids = [c.id for c in quiz_companies if request_user.is_superuser or c in accessible_companies]
    if filter_company_id is not None:
        if filter_company_id not in company_ids:
            return {}
        company_ids = [filter_company_id]
    if not company_ids:
        return {}

    all_cabinet_users = CabinetUser.objects.filter(
        company_id__in=company_ids
    ).prefetch_related('user__groups', 'user').select_related('user', 'company')

    quiz_cabinet_user_ids = set(quiz.cabinet_users.values_list('id', flat=True))
    quiz_cabinet_group_ids = set(quiz.cabinet_groups.values_list('id', flat=True))
    quiz_has_restrictions = quiz.cabinet_groups.exists() or quiz.cabinet_users.exists()
    cabinet_group_row_ids = [g.id for g in quiz.cabinet_groups.all()]

    user_to_cabinet = {}
    for cabinet_user in all_cabinet_users:
        if not request_user.is_superuser and cabinet_user.company not in accessible_companies:
            continue
        user = cabinet_user.user
        has_access = False
        if cabinet_user.id in quiz_cabinet_user_ids:
            has_access = True
        elif quiz_cabinet_group_ids and user.groups.filter(id__in=cabinet_group_row_ids).exists():
            has_access = True
        elif not quiz_has_restrictions:
            has_access = True
        if has_access:
            uid = user.id
            if uid not in user_to_cabinet:
                user_to_cabinet[uid] = cabinet_user
    return user_to_cabinet


def _quiz_export_rows_for_users(quiz, user_id_to_cabinet, attempts_by_user_id, max_points):
    """
    Build ordered list of display rows: one per assigned user.
    Same rules as quiz_manager.html combineUsersAndResults / bestResult:
    - If the user has any passed attempt, use the latest passed (by completed_at).
    - Otherwise use the latest attempt by (completed_at or started_at) among all attempts.
    attempts_by_user_id: user_id -> list of QuizAttempt for this quiz (any completion state).
    """
    rows = []
    passing = quiz.passing_score

    def sort_key_uid(uid):
        cu = user_id_to_cabinet.get(uid)
        u = cu.user if cu else None
        if not u:
            return ('', '', uid)
        name = (u.get_full_name() or u.username or '').strip().lower()
        return (name, u.username.lower(), uid)

    def ts_attempt(a):
        if a.completed:
            return (a.completed_at or a.started_at, a.started_at)
        return (a.started_at, a.started_at)

    for uid in sorted(user_id_to_cabinet.keys(), key=sort_key_uid):
        cu = user_id_to_cabinet[uid]
        user = cu.user
        atts = attempts_by_user_id.get(uid, [])

        if not atts:
            rows.append({
                'user': user,
                'cabinet_user': cu,
                'status_key': 'not_started',
                'started_at': None,
                'completed_at': None,
                'score': None,
                'max_points': max_points,
            })
            continue

        completed = [a for a in atts if a.completed]
        passed_attempts = [a for a in completed if a.score >= passing]

        if passed_attempts:
            passed_attempts.sort(key=ts_attempt, reverse=True)
            att = passed_attempts[0]
            rows.append({
                'user': user,
                'cabinet_user': cu,
                'status_key': 'passed',
                'started_at': att.started_at,
                'completed_at': att.completed_at,
                'score': att.score,
                'max_points': max_points,
            })
            continue

        sorted_all = sorted(atts, key=ts_attempt, reverse=True)
        att = sorted_all[0]
        if att.completed:
            rows.append({
                'user': user,
                'cabinet_user': cu,
                'status_key': 'failed',
                'started_at': att.started_at,
                'completed_at': att.completed_at,
                'score': att.score,
                'max_points': max_points,
            })
        else:
            rows.append({
                'user': user,
                'cabinet_user': cu,
                'status_key': 'in_progress',
                'started_at': att.started_at,
                'completed_at': None,
                'score': None,
                'max_points': max_points,
            })
    return rows


@login_required
def export_quiz_manager_results(request):
    """Multi-sheet Excel: one row per assigned user; Passed, Failed, In progress, or Not started."""
    if not AccessQuiz.objects.filter(group__in=request.user.groups.all(), has_access=True).exists() and not request.user.is_superuser:
        messages.error(request, _("You don't have permission to access quiz management."))
        return redirect('quiz_list')

    accessible_companies = get_accessible_companies_for_user(request.user)
    accessible_company_ids = {c.id for c in accessible_companies}

    filter_company_id = None
    raw_company = request.GET.get('company', '').strip()
    if raw_company:
        try:
            cid = int(raw_company)
        except (TypeError, ValueError):
            cid = None
        if cid is not None and cid in accessible_company_ids:
            filter_company_id = cid

    status_param = (request.GET.get('status') or 'all').strip().lower()
    if status_param not in ('all', 'active', 'inactive'):
        status_param = 'all'

    if request.user.is_superuser:
        quizzes = Quiz.objects.all()
    else:
        if accessible_companies:
            quizzes = Quiz.objects.filter(companies__in=accessible_companies).distinct()
        else:
            quizzes = Quiz.objects.none()

    if filter_company_id is not None:
        quizzes = quizzes.filter(companies__id=filter_company_id).distinct()

    if status_param == 'active':
        quizzes = quizzes.filter(is_active=True)
    elif status_param == 'inactive':
        quizzes = quizzes.filter(is_active=False)

    quizzes = quizzes.prefetch_related(
        'questions__answers',
        'companies',
        'cabinet_groups',
        'cabinet_users',
    ).order_by('-updated_at')
    quiz_list = list(quizzes)
    quiz_ids = [q.id for q in quiz_list]

    filter_labels = []
    if filter_company_id is not None:
        from app_conf.models import Company
        try:
            co = Company.objects.get(pk=filter_company_id)
            filter_labels.append(f"{_('Company')}: {co.name}")
        except Company.DoesNotExist:
            filter_labels.append(f"{_('Company')} ID: {filter_company_id}")
    if status_param == 'active':
        filter_labels.append(f"{_('Status')}: {_('Active')}")
    elif status_param == 'inactive':
        filter_labels.append(f"{_('Status')}: {_('Inactive')}")
    filter_summary = '   |   '.join(filter_labels) if filter_labels else ''

    if not quiz_list:
        wb_empty = Workbook()
        ws0 = wb_empty.active
        ws0.title = 'Info'
        empty_msg = (
            _('No quizzes match the selected filters.')
            if (filter_company_id is not None or status_param != 'all')
            else _('No quizzes available for export.')
        )
        ws0['A1'] = empty_msg
        ws0['A1'].font = Font(name='Arial', size=12)
        resp_empty = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp_empty['Content-Disposition'] = (
            f'attachment; filename="quiz_manager_results_{datetime.now().strftime("%Y-%m-%d_%H%M")}.xlsx"'
        )
        wb_empty.save(resp_empty)
        return resp_empty

    attempts_by_quiz_user = {}
    if quiz_ids:
        for att in (
            QuizAttempt.objects.filter(quiz_id__in=quiz_ids)
            .select_related('user', 'quiz')
            .order_by('-started_at')
        ):
            attempts_by_quiz_user.setdefault((att.quiz_id, att.user_id), []).append(att)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    thin = Side(style='thin')
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    meta_font = Font(name='Arial', size=10, color='000000')
    meta_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    hdr_font = Font(name='Arial', size=11, bold=True, color='000000')
    hdr_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')

    used_sheet_titles = set()
    summary_ws = wb.create_sheet(_sanitize_excel_sheet_name(_("Summary"), used_sheet_titles), 0)

    summary_headers = [
        _('№'),
        _('Quiz ID'),
        _('Title'),
        _('Passing score'),
        _('Max points'),
        _('Active'),
        _('With access'),
        _('Passed'),
        _('Failed'),
        _('Not started'),
        _('In progress'),
    ]
    summary_ws.merge_cells('A1:K1')
    c = summary_ws['A1']
    c.value = (
        _('Quiz manager — results overview')
        + (f" ({filter_summary})" if filter_summary else '')
    )
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    summary_ws.row_dimensions[1].height = 36 if filter_summary else 28

    for col, h in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = grid_border

    widths_summary = [6, 10, 40, 12, 12, 10, 12, 10, 10, 12, 12]
    for i, w in enumerate(widths_summary, 1):
        summary_ws.column_dimensions[get_column_letter(i)].width = w

    summary_data_row = 3
    summary_index = 0

    for quiz in quiz_list:
        max_points = _quiz_max_score_points(quiz)
        sheet_title = _sanitize_excel_sheet_name(quiz.title, used_sheet_titles)
        ws = wb.create_sheet(sheet_title)

        ws.merge_cells('A1:L1')
        t = ws['A1']
        t.value = f"{quiz.title} — {_('Results')}"
        t.font = title_font
        t.fill = title_fill
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        active_txt = _('Yes') if quiz.is_active else _('No')
        ws.merge_cells('A2:L2')
        meta = ws['A2']
        meta_parts = [
            f"{_('Passing score')}: {quiz.passing_score}",
            f"{_('Max points')}: {max_points}",
            f"{_('Active')}: {active_txt}",
        ]
        if filter_summary:
            meta_parts.append(filter_summary)
        meta.value = '   |   '.join(meta_parts)
        meta.font = meta_font
        meta.fill = meta_fill
        meta.alignment = wrap_left
        ws.row_dimensions[2].height = 22

        headers = [
            _('№'),
            _('Full name'),
            _('Username'),
            _('Company'),
            _('Department'),
            _('Position'),
            _('Started at'),
            _('Completed at'),
            _('Score'),
            _('Max points'),
            _('Result %'),
            _('Status'),
        ]
        header_row = 4
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = grid_border

        col_widths = [5, 26, 16, 22, 18, 18, 18, 18, 10, 10, 10, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        user_id_to_cabinet = _quiz_export_user_id_to_cabinet(
            quiz, request.user, accessible_companies, filter_company_id
        )
        attempts_by_user_id = {}
        for uid in user_id_to_cabinet:
            key = (quiz.id, uid)
            if key in attempts_by_quiz_user:
                attempts_by_user_id[uid] = attempts_by_quiz_user[key]

        export_rows = _quiz_export_rows_for_users(
            quiz, user_id_to_cabinet, attempts_by_user_id, max_points
        )

        n_access = len(export_rows)
        n_passed = sum(1 for r in export_rows if r['status_key'] == 'passed')
        n_failed = sum(1 for r in export_rows if r['status_key'] == 'failed')
        n_not_started = sum(1 for r in export_rows if r['status_key'] == 'not_started')
        n_in_progress = sum(1 for r in export_rows if r['status_key'] == 'in_progress')

        summary_index += 1
        summary_ws.cell(row=summary_data_row, column=1, value=summary_index).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=2, value=quiz.id).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=3, value=quiz.title).alignment = wrap_left
        summary_ws.cell(row=summary_data_row, column=4, value=quiz.passing_score).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=5, value=max_points).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=6, value=active_txt).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=7, value=n_access).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=8, value=n_passed).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=9, value=n_failed).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=10, value=n_not_started).alignment = center_align
        summary_ws.cell(row=summary_data_row, column=11, value=n_in_progress).alignment = center_align
        for col in range(1, 12):
            summary_ws.cell(row=summary_data_row, column=col).border = grid_border
        summary_data_row += 1

        data_start = header_row + 1
        if not export_rows:
            ws.merge_cells(f'A{data_start}:L{data_start}')
            empty_cell = ws[f'A{data_start}']
            empty_cell.value = _('No users with access to this quiz for the selected scope.')
            empty_cell.font = Font(name='Arial', italic=True, color='666666')
            empty_cell.alignment = Alignment(horizontal='center', vertical='center')
            continue

        pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        pass_font = Font(name='Arial', color='006100', bold=True)
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        fail_font = Font(name='Arial', color='9C0006', bold=True)
        prog_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        prog_font = Font(name='Arial', color='9C6500', bold=True)
        neutral_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        neutral_font = Font(name='Arial', color='525252', bold=True)

        for idx, er in enumerate(export_rows, start=1):
            row = data_start + idx - 1
            user = er['user']
            cu = er['cabinet_user']
            full_name = user.get_full_name().strip()
            if not full_name:
                full_name = user.username
            company_name = ''
            dept = ''
            position = ''
            if cu:
                company_name = str(cu.company.name) if cu.company else ''
                dept = str(cu.department) if cu.department else ''
                position = str(cu.position) if cu.position else ''

            started_txt = (
                er['started_at'].strftime('%d.%m.%Y %H:%M') if er['started_at'] else ''
            )
            completed_txt = (
                er['completed_at'].strftime('%d.%m.%Y %H:%M') if er['completed_at'] else ''
            )

            sk = er['status_key']
            if sk == 'not_started':
                score_txt = '—'
                pct_txt = '—'
                status_txt = _('Not started')
                status_fill, status_font = neutral_fill, neutral_font
                row_bg = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            elif sk == 'in_progress':
                score_txt = '—'
                pct_txt = '—'
                status_txt = _('In progress')
                status_fill, status_font = prog_fill, prog_font
                row_bg = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            elif sk == 'passed':
                score_txt = er['score']
                if max_points > 0:
                    pct_txt = f"{int(round(100 * er['score'] / max_points))}%"
                else:
                    pct_txt = '—'
                status_txt = _('Passed')
                status_fill, status_font = pass_fill, pass_font
                row_bg = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            else:
                score_txt = er['score']
                if max_points > 0:
                    pct_txt = f"{int(round(100 * er['score'] / max_points))}%"
                else:
                    pct_txt = '—'
                status_txt = _('Failed')
                status_fill, status_font = fail_fill, fail_font
                row_bg = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

            values = [
                idx,
                full_name,
                user.username,
                company_name,
                dept,
                position,
                started_txt,
                completed_txt,
                score_txt,
                er['max_points'],
                pct_txt,
                status_txt,
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = grid_border
                if col == 12:
                    cell.alignment = center_align
                    cell.font = status_font
                    cell.fill = status_fill
                else:
                    cell.fill = row_bg
                    if col == 1:
                        cell.alignment = center_align
                        cell.font = Font(name='Arial', bold=True)
                    elif col in (7, 8, 9, 10, 11):
                        cell.alignment = center_align
                        cell.font = Font(name='Arial')
                    else:
                        cell.alignment = wrap_left
                        cell.font = Font(name='Arial')

    if summary_data_row > 3:
        zebra = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
        for row in range(3, summary_data_row):
            if (row - 3) % 2 == 0:
                for col in range(1, 12):
                    summary_ws.cell(row=row, column=col).fill = zebra

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fname_parts = ['quiz_manager_results']
    if filter_company_id is not None:
        fname_parts.append(f'company{filter_company_id}')
    if status_param != 'all':
        fname_parts.append(status_param)
    fname_parts.append(datetime.now().strftime('%Y-%m-%d_%H%M'))
    fname = '_'.join(fname_parts) + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


@login_required
def export_quiz_results(request, quiz_id):
    """Export results for a specific quiz as an Excel file"""
    # Import security utilities
    from .security_utils import QuizSecurityManager, QuizResultsAuditLogger, get_client_ip
    
    # Validate quiz access
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    # Check permissions using security manager
    if request.user.is_superuser:
        has_access = True
    else:
        user_groups = request.user.groups.all()
        has_access = AccessQuiz.objects.filter(
            group__in=user_groups,
            has_access_to_results=True
        ).exists()
    
    if not has_access:
        raise PermissionDenied("You do not have permission to export quiz results")
    
    allowed_companies = QuizSecurityManager.get_user_accessible_companies(request.user)
    
    # Get company filter from request (same logic as quiz_results view)
    company_filter = request.GET.get('company', None)
    if company_filter:
        try:
            selected_company_id = int(company_filter)
            # Verify user has access to this company
            if request.user.is_superuser:
                # Superuser can access any company
                from app_conf.models import Company
                try:
                    company = Company.objects.get(id=selected_company_id)
                    filtered_companies = [company]
                except Company.DoesNotExist:
                    filtered_companies = allowed_companies
                    selected_company_id = None
            elif selected_company_id in [c.id for c in allowed_companies]:
                filtered_companies = [c for c in allowed_companies if c.id == selected_company_id]
            else:
                filtered_companies = allowed_companies
                selected_company_id = None
        except (ValueError, TypeError):
            filtered_companies = allowed_companies
            selected_company_id = None
    else:
        filtered_companies = allowed_companies
        selected_company_id = None
    
    # Log export action for security monitoring
    QuizResultsAuditLogger.log_access(
        user=request.user,
        action='export_quiz_results',
        quiz_id=quiz_id,
        ip_address=get_client_ip(request),
        details={'quiz_title': quiz.title, 'company_id': selected_company_id}
    )
    
    # Check if user has access to this quiz
    if not quiz.companies.filter(id__in=[c.id for c in filtered_companies]).exists():
        return HttpResponseForbidden(_("You do not have permission to export results for this quiz."))
    
    # Create a workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = quiz.title[:31]  # Excel sheet names are limited to 31 chars
    
    # Apply formatting
    data_start_row = format_excel_sheet(ws, quiz.title)
    
    # Get quiz results data - filter by selected company
    cabinet_users = CabinetUser.objects.filter(
        company__in=filtered_companies
    ).prefetch_related('user', 'company')
    
    # Get user IDs from filtered cabinet_users
    user_ids = list(cabinet_users.values_list('user_id', flat=True))
    
    # Collect all attempts first so we can sort them by date
    all_attempts_data = []
    
    # Get all successful attempts for this quiz from filtered users only
    if user_ids:
        attempts = QuizAttempt.objects.filter(
            quiz=quiz,
            user_id__in=user_ids,
            completed=True,
            score__gte=quiz.passing_score
        ).order_by('completed_at')  # Order by ascending date (earliest first)
        
        for attempt in attempts:
            try:
                cabinet_user = CabinetUser.objects.get(user=attempt.user)
                # Double-check that cabinet_user is in filtered companies
                if cabinet_user.company.id in [c.id for c in filtered_companies]:
                    all_attempts_data.append({
                        'user': cabinet_user,
                        'attempt': attempt,
                        'completed_at': attempt.completed_at
                    })
            except CabinetUser.DoesNotExist:
                logger.warning(f"No cabinet user found for user {attempt.user.id}")
                continue
    
    # Sort all attempts by completion date (earliest first)
    all_attempts_data.sort(key=lambda x: x['completed_at'].timestamp() if x['completed_at'] else float('inf'))
    
    # Now insert the sorted data
    row = data_start_row
    for index, attempt_data in enumerate(all_attempts_data, 1):
        cabinet_user = attempt_data['user']
        attempt = attempt_data['attempt']
        
        # Add sequential number
        ws.cell(row=row, column=1, value=index)
        
        # Fill data row
        ws.cell(row=row, column=2, value=cabinet_user.user.get_full_name())
        ws.cell(row=row, column=3, value=str(cabinet_user.company.name))
        ws.cell(row=row, column=4, value=str(cabinet_user.department) if cabinet_user.department else "None")
        ws.cell(row=row, column=5, value=str(cabinet_user.position) if cabinet_user.position else "None")
        
        # Format and add the completion date
        completion_date = attempt.completed_at.strftime('%d.%m.%Y %H:%M') if attempt.completed_at else "N/A"
        ws.cell(row=row, column=6, value=completion_date)
        
        # Safely calculate the result
        try:
            total_questions = attempt.quiz.questions.count()
            if total_questions > 0:
                percentage = int((attempt.score / total_questions) * 100)
                result_text = f"{percentage}% ({attempt.score}/{total_questions})"
            else:
                result_text = f"0% (0/0)"
        except Exception as e:
            # If there's any error in the calculation, just show raw numbers
            result_text = f"{attempt.score}/{total_questions if 'total_questions' in locals() else '?'}"
            logger.error(f"Error calculating result percentage: {str(e)}")
        
        # Assign a simple string value, not a complex object
        ws.cell(row=row, column=7, value=result_text)
        
        # Successful column - "Yes" with red color
        ws.cell(row=row, column=8, value="Yes")
        
        row += 1
    
    # Apply styles to all data rows
    if row > data_start_row:  # Only if we have data
        apply_data_styles(ws, data_start_row)
        
        # Apply additional styling to match the image
        for row_idx in range(data_start_row, row):
            # Make the Successful column red
            successful_cell = ws.cell(row=row_idx, column=8)
            successful_cell.font = Font(color='FF0000', bold=True)
            
            # Format sequence number with center alignment and bold
            num_cell = ws.cell(row=row_idx, column=1)
            num_cell.font = Font(bold=True)
    
    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{quiz.title}_results_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response


@login_required
def export_all_results(request):
    """Export results for all quizzes as a multi-sheet Excel file"""
    # Check permissions using security manager
    from .security_utils import QuizSecurityManager
    allowed_companies = QuizSecurityManager.get_user_accessible_companies(request.user)
    
    # Get company filter from request (same logic as quiz_results view)
    company_filter = request.GET.get('company', None)
    if company_filter:
        try:
            selected_company_id = int(company_filter)
            # Verify user has access to this company
            if request.user.is_superuser:
                # Superuser can access any company
                from app_conf.models import Company
                try:
                    company = Company.objects.get(id=selected_company_id)
                    filtered_companies = [company]
                except Company.DoesNotExist:
                    filtered_companies = allowed_companies
                    selected_company_id = None
            elif selected_company_id in [c.id for c in allowed_companies]:
                filtered_companies = [c for c in allowed_companies if c.id == selected_company_id]
            else:
                filtered_companies = allowed_companies
                selected_company_id = None
        except (ValueError, TypeError):
            filtered_companies = allowed_companies
            selected_company_id = None
    else:
        filtered_companies = allowed_companies
        selected_company_id = None
    
    # Create a workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Get all quizzes accessible to this user - filter by selected company
    if hasattr(filtered_companies, 'values_list'):
        filtered_company_ids = list(filtered_companies.values_list('id', flat=True))
    elif hasattr(filtered_companies, '__iter__') and not isinstance(filtered_companies, str):
        filtered_company_ids = [c.id for c in filtered_companies]
    else:
        filtered_company_ids = []
    
    quizzes = Quiz.objects.filter(
        Q(companies__id__in=filtered_company_ids) |
        Q(cabinet_groups__company__id__in=filtered_company_ids) |
        Q(cabinet_users__company__id__in=filtered_company_ids)
    ).distinct()
    
    # Get all cabinet users - filter by selected company
    cabinet_users = CabinetUser.objects.filter(
        company__in=filtered_companies
    ).prefetch_related('user', 'company')
    
    # Get user IDs from filtered cabinet_users
    user_ids = list(cabinet_users.values_list('user_id', flat=True))
    
    # Create a sheet for each quiz
    for quiz in quizzes:
        # Create a new worksheet
        ws = wb.create_sheet(quiz.title[:31])  # Excel sheet names are limited to 31 chars
        
        # Apply formatting
        data_start_row = format_excel_sheet(ws, quiz.title)
        
        # Collect all attempts first so we can sort them by date
        all_attempts_data = []
        
        # Get all successful attempts for this quiz from filtered users only
        if user_ids:
            attempts = QuizAttempt.objects.filter(
                quiz=quiz,
                user_id__in=user_ids,
                completed=True,
                score__gte=quiz.passing_score
            ).order_by('completed_at')  # Order by ascending date (earliest first)
            
            for attempt in attempts:
                try:
                    cabinet_user = CabinetUser.objects.get(user=attempt.user)
                    # Double-check that cabinet_user is in filtered companies
                    if cabinet_user.company.id in [c.id for c in filtered_companies]:
                        all_attempts_data.append({
                            'user': cabinet_user,
                            'attempt': attempt,
                            'completed_at': attempt.completed_at
                        })
                except CabinetUser.DoesNotExist:
                    logger.warning(f"No cabinet user found for user {attempt.user.id}")
                    continue
        
        # Sort all attempts by completion date (earliest first)
        all_attempts_data.sort(key=lambda x: x['completed_at'].timestamp() if x['completed_at'] else float('inf'))
        
        # Now insert the sorted data
        row = data_start_row
        for index, attempt_data in enumerate(all_attempts_data, 1):
            cabinet_user = attempt_data['user']
            attempt = attempt_data['attempt']
            
            # Add sequential number
            ws.cell(row=row, column=1, value=index)
            
            # Fill data row
            ws.cell(row=row, column=2, value=cabinet_user.user.get_full_name())
            ws.cell(row=row, column=3, value=str(cabinet_user.company.name))
            ws.cell(row=row, column=4, value=str(cabinet_user.department) if cabinet_user.department else "None")
            ws.cell(row=row, column=5, value=str(cabinet_user.position) if cabinet_user.position else "None")
            
            # Format and add the completion date
            completion_date = attempt.completed_at.strftime('%d.%m.%Y %H:%M') if attempt.completed_at else "N/A"
            ws.cell(row=row, column=6, value=completion_date)
            
            # Safely calculate the result
            try:
                total_questions = attempt.quiz.questions.count()
                if total_questions > 0:
                    percentage = int((attempt.score / total_questions) * 100)
                    result_text = f"{percentage}% ({attempt.score}/{total_questions})"
                else:
                    result_text = f"0% (0/0)"
            except Exception as e:
                # If there's any error in the calculation, just show raw numbers
                result_text = f"{attempt.score}/{total_questions if 'total_questions' in locals() else '?'}"
                logger.error(f"Error calculating result percentage: {str(e)}")
            
            # Assign a simple string value, not a complex object
            ws.cell(row=row, column=7, value=result_text)
            
            # Successful column - "Yes" with red color
            ws.cell(row=row, column=8, value="Yes")
            
            row += 1
        
        # Apply styles to all data rows
        if row > data_start_row:  # Only if we have data
            apply_data_styles(ws, data_start_row)
            
            # Apply additional styling to match the image
            for row_idx in range(data_start_row, row):
                # Make the Successful column red
                successful_cell = ws.cell(row=row_idx, column=8)
                successful_cell.font = Font(color='FF0000', bold=True)
                
                # Format sequence number with center alignment and bold
                num_cell = ws.cell(row=row_idx, column=1)
                num_cell.font = Font(bold=True)
    
    # Create a summary sheet
    summary_ws = wb.create_sheet(_("Summary"), 0)
    summary_row = format_excel_sheet(summary_ws, _("Quiz Results Summary"))
    
    # Add summary data - one row per quiz with count of successful completions
    # Collect all quiz data with dates to sort them
    quiz_summary_data = []
    for quiz in quizzes:
        # Get all successful attempt completions for this quiz
        successful_attempts = QuizAttempt.objects.filter(
            quiz=quiz,
            completed=True,
            score__gte=quiz.passing_score
        ).order_by('completed_at')  # Get earliest completion first
        
        successful_count = successful_attempts.count()
        
        # Find earliest completion date
        earliest_completion = None
        if successful_attempts.exists():
            # Get the earliest completion
            earliest_completion = successful_attempts.first().completed_at
        
        if successful_count > 0:
            quiz_summary_data.append({
                'quiz': quiz,
                'successful_count': successful_count,
                'earliest_completion': earliest_completion or timezone.now()
            })
    
    # Sort quiz summary by earliest completion date
    quiz_summary_data.sort(key=lambda x: x['earliest_completion'].timestamp())
    
    # Now add the sorted summary data to the sheet
    for index, data in enumerate(quiz_summary_data, 1):
        quiz = data['quiz']
        successful_count = data['successful_count']
        
        # Add sequential number
        summary_ws.cell(row=summary_row, column=1, value=index)
        
        # Fill summary row
        summary_ws.cell(row=summary_row, column=2, value=quiz.title)
        summary_ws.cell(row=summary_row, column=3, value=str(quiz.companies.first() if quiz.companies.exists() else ""))
        summary_ws.cell(row=summary_row, column=4, value="-")  # Department 
        summary_ws.cell(row=summary_row, column=5, value="-")  # Position
        
        # Format date properly - full date with time if available
        earliest_date = data['earliest_completion'].strftime('%d.%m.%Y %H:%M') if data['earliest_completion'] else "-"
        summary_ws.cell(row=summary_row, column=6, value=earliest_date)
        
        # Result column now showing users count - ensure it's a simple string
        summary_ws.cell(row=summary_row, column=7, value=f"{successful_count} {_('users')}")
        
        # Instead of "-", we'll use "-" but with proper styling
        summary_ws.cell(row=summary_row, column=8, value="-")
        
        summary_row += 1
    
    # Apply styles to summary data
    if summary_row > 2:  # Only if we have data
        apply_data_styles(summary_ws, 2)
        
        # Apply special formatting to the summary sheet to match individual sheets
        for row in range(2, summary_row):
            # Format the result cell to be right-aligned and same color as "Yes"
            result_cell = summary_ws.cell(row=row, column=7)
            result_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format the "Successful" column in the summary
            successful_cell = summary_ws.cell(row=row, column=8)
            successful_cell.font = Font(color='000000', bold=True)  # Black color for dash
            
            # Format sequence number with center alignment and bold
            num_cell = summary_ws.cell(row=row, column=1)
            num_cell.font = Font(bold=True)
    
    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="all_quiz_results_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    
    # Save the workbook to the response
    wb.save(response)
    return response


@login_required
def quiz_list(request):
    cabinet_user = request.user.cabinet
    if not cabinet_user.company:
        messages.warning(request, _("You are not assigned to any company. Please contact the administrator."))
        return redirect(reverse('personal_cabinet') + '#user-info')

    # Get all ACTIVE quizzes and filter by access
    all_quizzes = Quiz.objects.filter(is_active=True).prefetch_related('companies', 'cabinet_groups', 'cabinet_users')
    quizzes = [quiz for quiz in all_quizzes if quiz.has_user_access(request.user)]
    quiz_data = []

    for quiz in quizzes:
        quiz_info = {
            'id': quiz.id,
            'title': quiz.title,
            'description': quiz.description,
            'youtube_video_id': quiz.youtube_video_id,
            'attempt_count': QuizAttempt.objects.filter(user=request.user, quiz=quiz).count(),
        }

        if quiz.pdf_material:
            quiz_info['pdf_url'] = quiz.pdf_material.url
            quiz_info['pdf_filename'] = quiz.pdf_filename
        else:
            quiz_info['pdf_url'] = None
            quiz_info['pdf_filename'] = None

        quiz_data.append(quiz_info)

    context = {
        'quizzes': quiz_data,
    }

    return render(request, 'app_study/quiz_list.html', context)





@login_required
def start_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)

    # Check if the user has access to this quiz
    if not quiz.has_user_access(request.user):
        return HttpResponseForbidden(_("You do not have access to this quiz."))

    attempt = QuizAttempt.objects.create(user=request.user, quiz=quiz)

    questions = list(quiz.questions.all())
    if quiz.shuffle_questions:
        random.shuffle(questions)

    for question in questions:
        answers = list(question.answers.all())
        if quiz.shuffle_answers:
            random.shuffle(answers)
        question.shuffled_answers = answers

    context = {
        'quiz': quiz,
        'questions': questions,
        'attempt': attempt,
    }

    return render(request, 'app_study/quiz_questions.html', context)


def submit_quiz(request, attempt_id):
    if request.method == 'POST':
        attempt = get_object_or_404(QuizAttempt, id=attempt_id)
        for question in attempt.quiz.questions.all():
            answer_id = request.POST.get(f'question_{question.id}')
            if answer_id:
                answer = Answer.objects.get(id=answer_id)
                QuizAnswer.objects.create(attempt=attempt, question=question, answer=answer)
                if answer.is_correct:
                    attempt.score += answer.score

        attempt.completed = True
        attempt.completed_at = timezone.now()
        attempt.save()

        return redirect('quiz_result', attempt_id=attempt.id)
    return redirect('quiz_list')


@login_required
def quiz_result(request, attempt_id):
    # Import security utilities
    from .security_utils import QuizSecurityManager, QuizResultsAuditLogger, get_client_ip
    
    # Validate access to this specific attempt
    attempt = QuizSecurityManager.validate_attempt_access(request.user, attempt_id)
    
    # Ensure attempt is completed
    if not attempt.completed:
        raise Http404("Quiz attempt not found or not completed")
    
    # Log access for security monitoring
    QuizResultsAuditLogger.log_access(
        user=request.user,
        action='view_quiz_result',
        target_user=attempt.user if attempt.user != request.user else None,
        attempt_id=attempt_id,
        quiz_id=attempt.quiz.id,
        ip_address=get_client_ip(request)
    )
    
    return render(request, 'app_study/quiz_result.html', {'attempt': attempt})


@login_required
def quiz_result_secure(request, secure_token):
    """
    Secure quiz result view using UUID token instead of predictable ID
    Prevents IDOR vulnerabilities
    NOTE: Temporarily disabled until migration is complete
    """
    # Import security utilities
    from .security_utils import QuizSecurityManager, QuizResultsAuditLogger, get_client_ip
    
    # Temporary implementation until migration is complete
    # For now, redirect to regular quiz_result with enhanced security
    logger.info(f"Secure token access attempted by {request.user.username}, redirecting to enhanced security view")
    
    # Extract attempt ID from secure token if needed (placeholder logic)
    # In production, this would use the actual secure_token field
    raise Http404("Secure token access temporarily unavailable during migration")


def quiz_history(request):
    if 'cabinet_user_id' not in request.session:
        return redirect('first_login')

    cabinet_user = CabinetUser.objects.get(id=request.session['cabinet_user_id'])
    attempts = QuizAttempt.objects.filter(user=cabinet_user).order_by('-started_at')
    return render(request, 'app_study/quiz_history.html', {'attempts': attempts})






@login_required
def quiz_results(request):
    # Import security utilities
    from .security_utils import QuizSecurityManager, QuizResultsAuditLogger, require_quiz_results_access, get_client_ip
    
    # Check if user has access to quiz results
    if request.user.is_superuser:
        has_access = True
    else:
        user_groups = request.user.groups.all()
        has_access = AccessQuiz.objects.filter(
            group__in=user_groups,
            has_access_to_results=True
        ).exists()
    
    if not has_access:
        logger.warning(
            f"User {request.user.username} attempted to access quiz results "
            "without proper permissions"
        )
        raise PermissionDenied("You do not have permission to view quiz results")
    
    # Log access for security monitoring
    QuizResultsAuditLogger.log_access(
        user=request.user,
        action='view_quiz_results',
        ip_address=get_client_ip(request)
    )
    
    # Use security manager to get filtered results
    allowed_companies = QuizSecurityManager.get_user_accessible_companies(request.user)
    
    # Get company filter from request
    selected_company_id = None
    company_filter = request.GET.get('company', None)
    if company_filter:
        try:
            selected_company_id = int(company_filter)
            # Verify user has access to this company
            if request.user.is_superuser:
                # Superuser can access any company
                from app_conf.models import Company
                try:
                    company = Company.objects.get(id=selected_company_id)
                    filtered_companies = [company]
                except Company.DoesNotExist:
                    filtered_companies = allowed_companies
                    selected_company_id = None
            elif selected_company_id in [c.id for c in allowed_companies]:
                filtered_companies = [c for c in allowed_companies if c.id == selected_company_id]
            else:
                filtered_companies = allowed_companies
                selected_company_id = None
        except (ValueError, TypeError):
            filtered_companies = allowed_companies
            selected_company_id = None
    else:
        filtered_companies = allowed_companies
    
    cabinet_users = CabinetUser.objects.filter(
        company__in=filtered_companies
    ).prefetch_related(
        'user',
        'company'
    )

    # Get accessible quizzes using database queries instead of Python loops
    # Use filtered_companies (selected company) instead of all allowed_companies
    if hasattr(filtered_companies, 'values_list'):
        # It's a queryset
        filtered_company_ids = list(filtered_companies.values_list('id', flat=True))
    elif hasattr(filtered_companies, '__iter__') and not isinstance(filtered_companies, str):
        # It's a list or other iterable
        filtered_company_ids = [c.id for c in filtered_companies]
    else:
        filtered_company_ids = []
    
    # Filter quizzes by selected company (filtered_companies)
    quizzes = Quiz.objects.filter(
        Q(companies__id__in=filtered_company_ids) |
        Q(cabinet_groups__company__id__in=filtered_company_ids) |
        Q(cabinet_users__company__id__in=filtered_company_ids)
    ).distinct().prefetch_related('companies', 'cabinet_groups', 'cabinet_users')

    # Get all user IDs from cabinet_users
    user_ids = list(cabinet_users.values_list('user_id', flat=True))
    
    # Fetch all attempts in a single query with proper prefetching (only if there are users)
    if user_ids:
        all_attempts = QuizAttempt.objects.filter(
            user_id__in=user_ids,
            completed=True
        ).select_related(
            'quiz',
            'user'
        ).prefetch_related(
            'answers__question',
            'answers__answer'
        ).order_by('completed_at')
    else:
        all_attempts = QuizAttempt.objects.none()

    # Group attempts by user_id
    attempts_by_user = {}
    for attempt in all_attempts:
        user_id = attempt.user_id
        if user_id not in attempts_by_user:
            attempts_by_user[user_id] = []
        attempts_by_user[user_id].append(attempt)

    # Collect results in a format that can be sorted by completion date
    all_user_results = []
    sorted_quiz_attempts = {}
    
    for cabinet_user in cabinet_users:
        user_id = cabinet_user.user_id
        attempts = attempts_by_user.get(user_id, [])
        
        # Calculate counts (already filtered by completed=True)
        successful_attempts = sum(1 for attempt in attempts if attempt.score >= attempt.quiz.passing_score)
        failed_attempts = sum(1 for attempt in attempts if attempt.score < attempt.quiz.passing_score)

        user_results = {
            'cabinet_user': cabinet_user,
            'attempts': [],
            'total_attempts': len(attempts),
            'successful_attempts': successful_attempts,
            'failed_attempts': failed_attempts
        }

        # Process each attempt - answers are already prefetched
        for attempt in attempts:
            # Answers are already prefetched, just access them
            answers = list(attempt.answers.all())
            
            # Store this attempt with its quiz
            if attempt.quiz_id not in sorted_quiz_attempts:
                sorted_quiz_attempts[attempt.quiz_id] = []
                
            # Only add successful attempts to the quiz-specific list
            if attempt.score >= attempt.quiz.passing_score:
                sorted_quiz_attempts[attempt.quiz_id].append({
                    'cabinet_user': cabinet_user,
                    'attempt': attempt,
                    'completed_at': attempt.completed_at or timezone.now()
                })
                
            # Add to the user's attempts list
            user_results['attempts'].append({
                'attempt': attempt,
                'answers': answers
            })

        all_user_results.append(user_results)
    
    # Sort each quiz's attempts by completion date (earliest first)
    for quiz_id in sorted_quiz_attempts:
        sorted_quiz_attempts[quiz_id].sort(
            key=lambda x: x['completed_at'].timestamp() if x['completed_at'] else float('inf')
        )
    
    # Calculate total attempts count for each quiz (for badges) - optimized
    quiz_attempts_count = {}
    if user_ids:
        # Get all quiz IDs and passing scores
        quiz_ids = list(quizzes.values_list('id', flat=True))
        quiz_passing_scores = {q.id: q.passing_score for q in quizzes}
        
        # Initialize all quizzes with zero counts
        for quiz_id in quiz_ids:
            quiz_attempts_count[quiz_id] = {'total': 0, 'successful': 0}
        
        # Count attempts from already-fetched data (much faster)
        for attempt in all_attempts:
            if attempt.quiz_id in quiz_attempts_count:
                quiz_attempts_count[attempt.quiz_id]['total'] += 1
                if attempt.score >= quiz_passing_scores.get(attempt.quiz_id, 0):
                    quiz_attempts_count[attempt.quiz_id]['successful'] += 1
    
    # Get all accessible companies for the filter dropdown
    from app_conf.models import Company
    if request.user.is_superuser:
        all_companies_for_filter = Company.objects.all().order_by('name')
    else:
        all_companies_for_filter = Company.objects.filter(id__in=[c.id for c in allowed_companies]).order_by('name')
    
    context = {
        'results': all_user_results,
        'quizzes': quizzes,
        'sorted_quiz_attempts': sorted_quiz_attempts,  # Add the sorted data for the template
        'quiz_attempts_count': quiz_attempts_count,  # Add counts for badges
        'companies': all_companies_for_filter,
        'selected_company_id': selected_company_id,
    }

    return render(request, 'app_study/quiz_results.html', context)


@login_required
@user_has_quiz_access
def protected_pdf_serve(request, quiz_id):
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        if not quiz.pdf_material:
            messages.warning(request, _("This quiz does not have a PDF attachment."))
            # Redirect back to the referring page or to the quiz list
            referer = request.META.get('HTTP_REFERER')
            if referer:
                return redirect(referer)
            return redirect('quiz_list')

        file_path = quiz.pdf_material.path
        if not os.path.exists(file_path):
            logger.error(f"PDF file not found at path: {file_path} for quiz {quiz_id}")
            messages.error(request, _("The PDF file for this quiz could not be found. Please contact the administrator."))
            # Redirect back to the referring page or to the quiz list
            referer = request.META.get('HTTP_REFERER')
            if referer:
                return redirect(referer)
            return redirect('quiz_list')

        # Відкриваємо та віддаємо файл
        response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
        return response

    except Exception as e:
        logger.error(f"Error serving PDF for quiz {quiz_id}: {str(e)}")
        messages.error(request, _("An error occurred while trying to access the PDF. Please try again later."))
        # Redirect back to the referring page or to the quiz list
        referer = request.META.get('HTTP_REFERER')
        if referer:
            return redirect(referer)
        return redirect('quiz_list')


@login_required
def get_quiz_file_content(request, quiz_id):
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)

        if not quiz.has_user_access(request.user):
            raise PermissionDenied("Access denied")

        if not quiz.pdf_material:
            return JsonResponse({
                'success': False,
                'error': _('No file attached')
            }, status=404)

        file_hash = ''
        if quiz.pdf_material:
            with quiz.pdf_material.open('rb') as f:
                file_content = f.read()
                file_hash = hashlib.sha256(file_content).hexdigest()

        return JsonResponse({
            'success': True,
            'hash': file_hash,
            'filename': quiz.pdf_filename
        })

    except Quiz.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': _('Quiz not found')
        }, status=404)
    except PermissionDenied:
        return JsonResponse({
            'success': False,
            'error': _('Access denied')
        }, status=403)
    except Exception as e:
        logger.error(f"Error getting file content: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_available_quizzes(request):
    """
    API endpoint to get available quizzes for a user/company
    Used by the user management modals to show which quizzes are accessible
    """
    try:
        user_id = request.GET.get('user_id')
        company_id = request.GET.get('company_id')
        
        if not company_id:
            return JsonResponse({
                'status': 'error',
                'message': _('Company ID is required')
            }, status=400)
        
        # Get all active quizzes that are available for the specified company
        available_quizzes = Quiz.objects.filter(
            companies__id=company_id,
            is_active=True
        ).distinct().prefetch_related('companies', 'cabinet_groups', 'cabinet_users')
        
        quiz_data = []
        
        for quiz in available_quizzes:
            quiz_info = {
                'id': quiz.id,
                'title': quiz.title,
                'description': quiz.description,
                'passing_score': quiz.passing_score,
                'youtube_video_id': quiz.youtube_video_id,
                'pdf_material': bool(quiz.pdf_material),
                'user_attempts': [],
                'is_assigned': False
            }
            
            # If user_id is provided, get their attempt history and assignment status
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                    cabinet_user = user.cabinet
                    
                    # Check if quiz is individually assigned to this user
                    quiz_info['is_assigned'] = quiz.cabinet_users.filter(id=cabinet_user.id).exists()
                    
                    attempts = QuizAttempt.objects.filter(
                        user=user,
                        quiz=quiz,
                        completed=True
                    ).order_by('-completed_at')
                    
                    quiz_info['user_attempts'] = [
                        {
                            'id': attempt.id,
                            'score': attempt.score,
                            'completed_at': attempt.completed_at.isoformat() if attempt.completed_at else None,
                            'passed': attempt.score >= quiz.passing_score
                        }
                        for attempt in attempts
                    ]
                except (User.DoesNotExist, AttributeError):
                    pass
            
            quiz_data.append(quiz_info)
        
        return JsonResponse({
            'status': 'success',
            'data': quiz_data
        })
        
    except Exception as e:
        logger.error(f"Error getting available quizzes: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': _('Error loading quizzes')
        }, status=500)


# Quiz Manager Views
def get_accessible_companies_for_user(user):
    """Get companies that user has access to based on AccessQuiz settings"""
    # Only superusers have access to all companies
    if user.is_superuser:
        from app_conf.models import Company
        all_companies = list(Company.objects.all())
        logger.info(f"User {user.username} is superuser - giving access to all companies: {[c.name for c in all_companies]}")
        return all_companies
    
    user_groups = user.groups.all()
    logger.info(f"Getting accessible companies for user: {user.username}, groups: {[g.name for g in user_groups]}")
    
    # Get all AccessQuiz entries for user's groups where has_access=True
    access_quiz_entries = AccessQuiz.objects.filter(
        group__in=user_groups,
        has_access=True
    ).prefetch_related('companies')
    
    logger.info(f"Found {access_quiz_entries.count()} AccessQuiz entries with has_access=True")
    
    # Collect all companies
    accessible_companies = set()
    for entry in access_quiz_entries:
        companies = entry.companies.all()
        accessible_companies.update(companies)
        logger.info(f"AccessQuiz entry for group {entry.group.name} has companies: {[c.name for c in companies]}")
    
    # Convert set to list to maintain consistency
    accessible_companies_list = list(accessible_companies)
    logger.info(f"Total accessible companies for user {user.username}: {[c.name for c in accessible_companies_list]}")
    return accessible_companies_list


def can_user_edit_quizzes(user):
    """Check if user can edit quizzes based on AccessQuiz settings"""
    # Superusers always have edit access
    if user.is_superuser:
        return True
    
    # Check AccessQuiz permissions for user's groups
    user_groups = user.groups.all()
    can_edit = AccessQuiz.objects.filter(
        group__in=user_groups,
        can_edit=True
    ).exists()
    
    logger.info(f"User {user.username} can_edit_quizzes check: {can_edit} (groups: {[g.name for g in user_groups]})")
    return can_edit


@login_required
def quiz_manager(request):
    """List all quizzes for management"""
    # Check AccessQuiz permissions
    if not AccessQuiz.objects.filter(group__in=request.user.groups.all(), has_access=True).exists() and not request.user.is_superuser:
        messages.error(request, _("You don't have permission to access quiz management."))
        return redirect('quiz_list')
    
    # Get accessible companies for the user
    accessible_companies = get_accessible_companies_for_user(request.user)

    selected_company_id = None
    company_filter = request.GET.get('company')
    if company_filter:
        try:
            selected_company_id = int(company_filter)
            if request.user.is_superuser:
                from app_conf.models import Company
                if not Company.objects.filter(id=selected_company_id).exists():
                    selected_company_id = None
            elif selected_company_id not in [c.id for c in accessible_companies]:
                selected_company_id = None
        except (ValueError, TypeError):
            selected_company_id = None

    from app_conf.models import Company
    if request.user.is_superuser:
        companies_for_filter = Company.objects.all().order_by('name')
    else:
        companies_for_filter = Company.objects.filter(
            id__in=[c.id for c in accessible_companies]
        ).order_by('name')
    
    # If user is superuser, show all quizzes
    if request.user.is_superuser:
        quizzes = Quiz.objects.all()
    else:
        # Filter quizzes by accessible companies
        if accessible_companies:
            quizzes = Quiz.objects.filter(companies__in=accessible_companies).distinct()
        else:
            quizzes = Quiz.objects.none()

    if selected_company_id:
        quizzes = quizzes.filter(companies__id=selected_company_id).distinct()
    
    quizzes = quizzes.prefetch_related(
        'questions', 
        'companies', 
        'cabinet_groups__group', 
        'cabinet_users__user',
        'cabinet_groups__company',
        'cabinet_users__company'
    ).order_by('-updated_at')

    per_page = get_study_table_page_size(request)
    paginator = Paginator(quizzes, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    page_quizzes = list(page_obj.object_list)
    
    # Calculate progress statistics for each quiz
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Count, Q
    
    # Prefetch all quiz attempts for all quizzes to reduce queries
    quiz_ids = [q.id for q in page_quizzes]
    all_attempts = QuizAttempt.objects.filter(
        quiz_id__in=quiz_ids,
        completed=True
    ).select_related('quiz', 'user')
    
    # Group attempts by quiz_id for faster lookup
    attempts_by_quiz = {}
    for attempt in all_attempts:
        quiz_id = attempt.quiz_id
        if quiz_id not in attempts_by_quiz:
            attempts_by_quiz[quiz_id] = {'completed': set(), 'passed': set()}
        attempts_by_quiz[quiz_id]['completed'].add(attempt.user_id)
        if attempt.score >= attempt.quiz.passing_score:
            attempts_by_quiz[quiz_id]['passed'].add(attempt.user_id)
    
    quizzes_with_progress = []
    for quiz in page_quizzes:
        # Get all users who have access to this quiz
        users_with_access = set()
        quiz_companies = list(quiz.companies.all())
        
        # Prefetch cabinet users for all companies at once
        company_ids = [c.id for c in quiz_companies if request.user.is_superuser or c in accessible_companies]
        if company_ids:
            all_cabinet_users = CabinetUser.objects.filter(
                company_id__in=company_ids
            ).prefetch_related('user__groups', 'user').select_related('user', 'company')
            
            # Prefetch quiz's cabinet_users and cabinet_groups for faster checking
            quiz_cabinet_user_ids = set(quiz.cabinet_users.values_list('id', flat=True))
            quiz_cabinet_group_ids = set(quiz.cabinet_groups.values_list('id', flat=True))
            quiz_has_restrictions = quiz.cabinet_groups.exists() or quiz.cabinet_users.exists()
            
            for cabinet_user in all_cabinet_users:
                # Only process companies accessible to the current user
                if not request.user.is_superuser and cabinet_user.company not in accessible_companies:
                    continue
                
                user = cabinet_user.user
                
                # Check if this user has access to the quiz
                has_access = False
                
                # Check direct user access
                if cabinet_user.id in quiz_cabinet_user_ids:
                    has_access = True
                # Check group access
                elif quiz_cabinet_group_ids and user.groups.filter(id__in=[g.id for g in quiz.cabinet_groups.all()]).exists():
                    has_access = True
                # Check company access (if no specific groups/users are defined)
                elif not quiz_has_restrictions:
                    has_access = True
                
                if has_access:
                    users_with_access.add(user.id)
        
        # Calculate progress statistics using pre-fetched attempts
        total_users = len(users_with_access)
        quiz_attempts = attempts_by_quiz.get(quiz.id, {'completed': set(), 'passed': set()})
        completed_users = len(users_with_access & quiz_attempts['completed'])
        passed_users = len(users_with_access & quiz_attempts['passed'])
        
        # Calculate percentage
        completion_percentage = (completed_users / total_users * 100) if total_users > 0 else 0
        pass_percentage = (passed_users / total_users * 100) if total_users > 0 else 0
        
        # Add progress data to quiz object
        quiz.progress_data = {
            'total_users': total_users,
            'completed_users': completed_users,
            'passed_users': passed_users,
            'completion_percentage': round(completion_percentage, 1),
            'pass_percentage': round(pass_percentage, 1),
        }
        
        quizzes_with_progress.append(quiz)
    
    # Default to tomorrow at 9 AM
    tomorrow = timezone.now() + timedelta(days=1)
    default_datetime = tomorrow.replace(hour=9, minute=0, second=0, microsecond=0)
    
    # Minimum datetime is now
    min_datetime = timezone.now().strftime('%Y-%m-%dT%H:%M')
    default_datetime_str = default_datetime.strftime('%Y-%m-%dT%H:%M')
    
    context = {
        'quizzes': quizzes_with_progress,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.count > 0,
        'current_page_size': per_page,
        'page_size_options': STUDY_TABLE_PAGE_SIZE_OPTIONS,
        'total_quizzes': paginator.count,
        'can_edit': can_user_edit_quizzes(request.user),
        'accessible_companies': accessible_companies,
        'companies': companies_for_filter,
        'selected_company_id': selected_company_id,
        'min_datetime': min_datetime,
        'default_datetime': default_datetime_str,
        'pagination_item_label': _('quizzes'),
        'pagination_aria_label': _('Quiz manager pagination'),
    }
    
    return render(request, 'app_study/quiz_manager.html', context)


@login_required
@require_http_methods(["POST"])
def quiz_toggle_active(request, quiz_id):
    """Toggle quiz active flag from Quiz Manager via checkbox."""
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    # Reuse same access check as quiz_manager
    if not AccessQuiz.objects.filter(group__in=request.user.groups.all(), has_access=True).exists() and not request.user.is_superuser:
        messages.error(request, _("You don't have permission to change quiz status."))
        return redirect('quiz_manager')
    quiz.is_active = not quiz.is_active
    quiz.save(update_fields=['is_active'])
    return redirect('quiz_manager')


@login_required
def quiz_create(request):
    """Create a new quiz"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to create quizzes."))
        return redirect('quiz_list')
    
    # Get accessible companies for the user to filter form choices
    accessible_companies = get_accessible_companies_for_user(request.user)
    
    if request.method == 'POST':
        form = QuizForm(request.POST, request.FILES)
        # Filter companies and groups based on accessible companies from AccessQuiz
        # Apply filtering for all users based on AccessQuiz settings
        form.filter_by_companies(accessible_companies)
        if form.is_valid():
            quiz = form.save()
            messages.success(request, _("Quiz created successfully. You can now add questions."))
            return redirect('quiz_edit', quiz_id=quiz.id)
    else:
        form = QuizForm()
        # Filter companies and groups based on accessible companies from AccessQuiz
        # Apply filtering for all users based on AccessQuiz settings
        form.filter_by_companies(accessible_companies)
    
    context = {
        'form': form,
        'title': _("Create New Quiz"),
        'action': 'create'
    }
    
    return render(request, 'app_study/quiz_form.html', context)


@login_required
def quiz_edit(request, quiz_id):
    """Edit an existing quiz with questions and answers"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to edit quizzes."))
        return redirect('quiz_list')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to edit this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        form = QuizForm(request.POST, request.FILES, instance=quiz)
        # Filter companies and groups based on accessible companies from AccessQuiz
        # Apply filtering for all users based on AccessQuiz settings
        form.filter_by_companies(accessible_companies)
        if form.is_valid():
            quiz = form.save()
            messages.success(request, _("Quiz updated successfully."))
            return redirect('quiz_edit', quiz_id=quiz.id)
    else:
        form = QuizForm(instance=quiz)
        # Filter companies and groups based on accessible companies from AccessQuiz
        # Apply filtering for all users based on AccessQuiz settings
        form.filter_by_companies(accessible_companies)
    
    # Get questions with their answers
    questions = quiz.questions.all().prefetch_related('answers').order_by('order')
    
    context = {
        'form': form,
        'quiz': quiz,
        'questions': questions,
        'title': _("Edit Quiz: {}").format(quiz.title),
        'action': 'edit'
    }
    
    return render(request, 'app_study/quiz_form.html', context)


@login_required
def quiz_delete(request, quiz_id):
    """Delete a quiz"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to delete quizzes."))
        return redirect('quiz_list')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    if request.method == 'POST':
        quiz_title = quiz.title
        quiz.delete()
        messages.success(request, _("Quiz '{}' deleted successfully.").format(quiz_title))
        return redirect('quiz_manager')
    
    context = {
        'quiz': quiz,
        'title': _("Delete Quiz: {}").format(quiz.title)
    }
    
    return render(request, 'app_study/quiz_delete.html', context)


@login_required
def question_create(request, quiz_id):
    """Create a new question for a quiz"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to create questions."))
        return redirect('quiz_list')
    
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to add questions to this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.quiz = quiz
            question.save()
            messages.success(request, _("Question created successfully."))
            return redirect('question_edit', question_id=question.id)
    else:
        # Set default order to next available number
        next_order = quiz.questions.count() + 1
        form = QuestionForm(initial={'order': next_order})
    
    context = {
        'form': form,
        'quiz': quiz,
        'title': _("Add Question to: {}").format(quiz.title),
        'action': 'create'
    }
    
    return render(request, 'app_study/question_form.html', context)


@login_required
def question_edit(request, question_id):
    """Edit a question with its answers"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to edit questions."))
        return redirect('quiz_list')
    
    question = get_object_or_404(Question, id=question_id)
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not question.quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to edit questions in this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            question = form.save()
            messages.success(request, _("Question updated successfully."))
            return redirect('question_edit', question_id=question.id)
    else:
        form = QuestionForm(instance=question)
    
    # Get answers for this question
    answers = question.answers.all()
    
    context = {
        'form': form,
        'question': question,
        'answers': answers,
        'title': _("Edit Question"),
        'action': 'edit'
    }
    
    return render(request, 'app_study/question_form.html', context)


@login_required
def question_delete(request, question_id):
    """Delete a question"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to delete questions."))
        return redirect('quiz_list')
    
    question = get_object_or_404(Question, id=question_id)
    quiz = question.quiz
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to delete questions from this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, _("Question deleted successfully."))
        return redirect('quiz_edit', quiz_id=quiz.id)
    
    context = {
        'question': question,
        'quiz': quiz,
        'title': _("Delete Question")
    }
    
    return render(request, 'app_study/question_delete.html', context)


@login_required
def answer_create(request, question_id):
    """Create a new answer for a question"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to create answers."))
        return redirect('quiz_list')
    
    question = get_object_or_404(Question, id=question_id)
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not question.quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to add answers to this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        form = AnswerForm(request.POST)
        if form.is_valid():
            answer = form.save(commit=False)
            answer.question = question
            answer.save()
            messages.success(request, _("Answer created successfully."))
            return redirect('question_edit', question_id=question.id)
    else:
        form = AnswerForm()
    
    context = {
        'form': form,
        'question': question,
        'title': _("Add Answer"),
        'action': 'create'
    }
    
    return render(request, 'app_study/answer_form.html', context)


@login_required
def answer_edit(request, answer_id):
    """Edit an answer"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to edit answers."))
        return redirect('quiz_list')
    
    answer = get_object_or_404(Answer, id=answer_id)
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not answer.question.quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to edit answers in this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        form = AnswerForm(request.POST, instance=answer)
        if form.is_valid():
            answer = form.save()
            messages.success(request, _("Answer updated successfully."))
            return redirect('question_edit', question_id=answer.question.id)
    else:
        form = AnswerForm(instance=answer)
    
    context = {
        'form': form,
        'answer': answer,
        'question': answer.question,
        'title': _("Edit Answer"),
        'action': 'edit'
    }
    
    return render(request, 'app_study/answer_form.html', context)


@login_required
def answer_delete(request, answer_id):
    """Delete an answer"""
    if not can_user_edit_quizzes(request.user):
        messages.error(request, _("You don't have permission to delete answers."))
        return redirect('quiz_list')
    
    answer = get_object_or_404(Answer, id=answer_id)
    question = answer.question
    
    # Check if user has access to this specific quiz
    accessible_companies = get_accessible_companies_for_user(request.user)
    if not request.user.is_superuser:
        if not answer.question.quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to delete answers from this quiz."))
            return redirect('quiz_manager')
    
    if request.method == 'POST':
        answer.delete()
        messages.success(request, _("Answer deleted successfully."))
        return redirect('question_edit', question_id=question.id)
    
    context = {
        'answer': answer,
        'question': question,
        'title': _("Delete Answer")
    }
    
    return render(request, 'app_study/answer_delete.html', context)


# ===== PAGE MANAGER FUNCTIONS =====

def get_accessible_companies_for_page_user(user):
    """Get companies that user has access to based on AccessPage settings"""
    # Only superusers have access to all companies
    if user.is_superuser:
        from app_conf.models import Company
        all_companies = list(Company.objects.all())
        logger.info(f"User {user.username} is superuser - giving access to all companies: {[c.name for c in all_companies]}")
        return all_companies
    
    user_groups = user.groups.all()
    logger.info(f"Getting accessible companies for page user: {user.username}, groups: {[g.name for g in user_groups]}")
    
    # Get all AccessPage entries for user's groups where has_access=True
    access_page_entries = AccessPage.objects.filter(
        group__in=user_groups,
        has_access=True
    ).prefetch_related('companies')
    
    logger.info(f"Found {access_page_entries.count()} AccessPage entries with has_access=True")
    
    # Collect all companies
    accessible_companies = set()
    for entry in access_page_entries:
        companies = entry.companies.all()
        accessible_companies.update(companies)
        logger.info(f"AccessPage entry for group {entry.group.name} has companies: {[c.name for c in companies]}")
    
    # Convert set to list to maintain consistency
    accessible_companies_list = list(accessible_companies)
    logger.info(f"Total accessible companies for page user {user.username}: {[c.name for c in accessible_companies_list]}")
    return accessible_companies_list


def can_user_edit_pages(user):
    """Check if user can edit pages based on AccessPage settings"""
    # Superusers always have edit access
    if user.is_superuser:
        return True
    
    # Check AccessPage permissions for user's groups
    user_groups = user.groups.all()
    can_edit = AccessPage.objects.filter(
        group__in=user_groups,
        can_edit=True
    ).exists()
    
    logger.info(f"User {user.username} can_edit_pages check: {can_edit} (groups: {[g.name for g in user_groups]})")
    return can_edit


def _handle_document_files(request, page):
    """Handle document file uploads and deletions for a page"""
    from .models import PageDocumentFile
    import os
    
    # Handle document deletions
    delete_document_ids = request.POST.getlist('delete_document_ids')
    if delete_document_ids:
        documents_to_delete = PageDocumentFile.objects.filter(
            id__in=delete_document_ids,
            page=page
        )
        for doc in documents_to_delete:
            # Delete the file from storage
            if doc.document_file and os.path.exists(doc.document_file.path):
                os.remove(doc.document_file.path)
            doc.delete()
    
    # Handle new document uploads
    form_index = 0
    while True:
        title_key = f'new_document_title_{form_index}'
        file_key = f'new_document_file_{form_index}'
        description_key = f'new_document_description_{form_index}'
        
        if title_key not in request.POST:
            break
            
        title = request.POST.get(title_key)
        description = request.POST.get(description_key, '')
        
        if title and file_key in request.FILES:
            document_file = request.FILES[file_key]
            
            # Create new document
            PageDocumentFile.objects.create(
                page=page,
                title=title,
                document_file=document_file,
                description=description,
                order=form_index
            )
        
        form_index += 1


@login_required
def protected_document_serve(request, document_id):
    """Serve protected document files with access control"""
    from .models import PageDocumentFile
    from django.http import HttpResponse, Http404
    from django.core.exceptions import PermissionDenied
    from django.utils.translation import gettext as _
    import os
    import mimetypes
    
    try:
        document = PageDocumentFile.objects.get(id=document_id)
    except PageDocumentFile.DoesNotExist:
        raise Http404(_("Document not found"))
    
    # Check if user has access to the page containing this document
    if not document.page.has_user_access(request.user):
        raise PermissionDenied(_("You don't have permission to access this document"))
    
    # Check if file exists
    if not document.document_file or not os.path.exists(document.document_file.path):
        raise Http404(_("Document file not found"))
    
    # Determine content type
    content_type, _unused = mimetypes.guess_type(document.document_file.path)
    if not content_type:
        content_type = 'application/octet-stream'
    
    # Read and serve the file
    with open(document.document_file.path, 'rb') as f:
        response = HttpResponse(f.read(), content_type=content_type)
        
    # Set appropriate headers
    filename = os.path.basename(document.document_file.name)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def page_manager(request):
    """List all pages for management"""
    # Check AccessPage permissions
    if not AccessPage.objects.filter(group__in=request.user.groups.all(), has_access=True).exists() and not request.user.is_superuser:
        messages.error(request, _("You don't have permission to access page management."))
        return redirect('personal_cabinet')
    
    # Get accessible companies for the user
    accessible_companies = get_accessible_companies_for_page_user(request.user)

    selected_company_id = None
    company_filter = request.GET.get('company')
    if company_filter:
        try:
            selected_company_id = int(company_filter)
            if request.user.is_superuser:
                from app_conf.models import Company
                if not Company.objects.filter(id=selected_company_id).exists():
                    selected_company_id = None
            elif selected_company_id not in [c.id for c in accessible_companies]:
                selected_company_id = None
        except (ValueError, TypeError):
            selected_company_id = None

    from app_conf.models import Company
    if request.user.is_superuser:
        companies_for_filter = Company.objects.all().order_by('name')
    else:
        companies_for_filter = Company.objects.filter(
            id__in=[c.id for c in accessible_companies]
        ).order_by('name')
    
    # If user is superuser, show all pages
    if request.user.is_superuser:
        pages = Page.objects.all()
    else:
        # Filter pages by accessible companies
        if accessible_companies:
            pages = Page.objects.filter(companies__in=accessible_companies).distinct()
        else:
            pages = Page.objects.none()

    if selected_company_id:
        pages = pages.filter(companies__id=selected_company_id).distinct()
    
    pages = pages.prefetch_related('companies', 'cabinet_groups__group', 'cabinet_users__user').order_by('-updated_at')

    per_page = get_study_table_page_size(request)
    paginator = Paginator(pages, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))
    
    context = {
        'pages': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.count > 0,
        'current_page_size': per_page,
        'page_size_options': STUDY_TABLE_PAGE_SIZE_OPTIONS,
        'total_pages': paginator.count,
        'can_edit': can_user_edit_pages(request.user),
        'accessible_companies': accessible_companies,
        'companies': companies_for_filter,
        'selected_company_id': selected_company_id,
        'pagination_item_label': _('pages'),
        'pagination_aria_label': _('Page manager pagination'),
    }
    
    return render(request, 'app_study/page_manager.html', context)


@login_required
@require_http_methods(["POST"])
def page_toggle_active(request, page_id):
    """Toggle page active flag from Page Manager via checkbox."""
    page = get_object_or_404(Page, pk=page_id)
    # Reuse same access check as page_manager
    if not AccessPage.objects.filter(group__in=request.user.groups.all(), has_access=True).exists() and not request.user.is_superuser:
        messages.error(request, _("You don't have permission to change page status."))
        return redirect('page_manager')
    page.is_active = not page.is_active
    page.save(update_fields=['is_active'])
    return redirect('page_manager')


def _has_page_manager_access(user):
    """True if user can access Page Manager (same check as page_manager view)."""
    if user.is_superuser:
        return True
    return AccessPage.objects.filter(group__in=user.groups.all(), has_access=True).exists()


@login_required
@require_http_methods(["GET"])
def page_manager_guide(request):
    """Return JSON { content: html } for the Page Manager guide (localized)."""
    if not _has_page_manager_access(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = PageManagerGuide.objects.first()
    if guide:
        if country:
            trans = PageManagerGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = PageManagerGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def page_manager_guide_translate(request):
    """API for AI translation of Page Manager guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


def _has_quiz_manager_access(user):
    """True if user can access Quiz Manager (same check as quiz_manager view)."""
    if user.is_superuser:
        return True
    return AccessQuiz.objects.filter(group__in=user.groups.all(), has_access=True).exists()


@login_required
@require_http_methods(["GET"])
def quiz_manager_guide(request):
    """Return JSON { content: html } for the Quiz Manager guide (localized)."""
    if not _has_quiz_manager_access(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = QuizManagerGuide.objects.first()
    if guide:
        if country:
            trans = QuizManagerGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = QuizManagerGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def quiz_manager_guide_translate(request):
    """API for AI translation of Quiz Manager guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
def page_create(request):
    """Create a new page"""
    if not can_user_edit_pages(request.user):
        messages.error(request, _("You don't have permission to create pages."))
        return redirect('personal_cabinet')
    
    # Get accessible companies for the user to filter form choices
    accessible_companies = get_accessible_companies_for_page_user(request.user)
    
    if request.method == 'POST':
        form = PageForm(request.POST, request.FILES)
        # Filter companies based on accessible companies from AccessPage
        form.filter_by_companies(accessible_companies)
        if form.is_valid():
            page = form.save()
            
            # Handle document files
            _handle_document_files(request, page)
            
            messages.success(request, _("Page created successfully."))
            return redirect('page_edit', page_id=page.id)
    else:
        form = PageForm()
        # Filter companies based on accessible companies from AccessPage
        form.filter_by_companies(accessible_companies)
    
    context = {
        'form': form,
        'title': _("Create New Page"),
        'action': 'create'
    }
    
    return render(request, 'app_study/page_form.html', context)


@login_required
def page_edit(request, page_id):
    """Edit an existing page"""
    if not can_user_edit_pages(request.user):
        messages.error(request, _("You don't have permission to edit pages."))
        return redirect('personal_cabinet')
    
    page = get_object_or_404(Page, id=page_id)
    
    # Check if user has access to this specific page
    accessible_companies = get_accessible_companies_for_page_user(request.user)
    if not request.user.is_superuser:
        if not page.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to edit this page."))
            return redirect('page_manager')
    
    if request.method == 'POST':
        form = PageForm(request.POST, request.FILES, instance=page)
        # Filter companies based on accessible companies from AccessPage
        form.filter_by_companies(accessible_companies)
        if form.is_valid():
            page = form.save()
            
            # Handle document files
            _handle_document_files(request, page)
            
            messages.success(request, _("Page updated successfully."))
            return redirect('page_edit', page_id=page.id)
    else:
        form = PageForm(instance=page)
        # Filter companies based on accessible companies from AccessPage
        form.filter_by_companies(accessible_companies)
    
    context = {
        'form': form,
        'page': page,
        'title': _("Edit Page: {}").format(page.title),
        'action': 'edit'
    }
    
    return render(request, 'app_study/page_form.html', context)


@login_required
def page_delete(request, page_id):
    """Delete a page"""
    if not can_user_edit_pages(request.user):
        messages.error(request, _("You don't have permission to delete pages."))
        return redirect('personal_cabinet')
    
    page = get_object_or_404(Page, id=page_id)
    
    # Check if user has access to this specific page
    accessible_companies = get_accessible_companies_for_page_user(request.user)
    if not request.user.is_superuser:
        if not page.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
            messages.error(request, _("You don't have permission to delete this page."))
            return redirect('page_manager')
    
    if request.method == 'POST':
        page_title = page.title
        page.delete()
        messages.success(request, _("Page '{}' deleted successfully.").format(page_title))
        return redirect('page_manager')
    
    context = {
        'page': page,
        'title': _("Delete Page: {}").format(page.title)
    }
    
    return render(request, 'app_study/page_delete.html', context)


@login_required
def page_view(request, slug):
    """View a page content"""
    from app_cabinet.models import CabinetUser
    
    page = get_object_or_404(Page, slug=slug)
    cabinet_user = CabinetUser.objects.get(user=request.user)

    if not cabinet_user.company:
        return HttpResponseForbidden(_("You do not have access to this page."))

    if not page.companies.filter(id=cabinet_user.company.id).exists():
        return HttpResponseForbidden(_("You do not have access to this page."))

    content = page.get_content()
    return render(request, 'app_study/page.html', {'page': page, 'content': content})


@login_required
def learning_hub(request):
    """Learning Hub - central place for tests and materials"""
    from app_cabinet.models import CabinetUser
    import html
    from django.utils.html import strip_tags
    from django.utils import translation
    
    cabinet_user = request.user.cabinet
    if not cabinet_user.company:
        messages.warning(request, _("You are not assigned to any company. Please contact the administrator."))
        return redirect(reverse('personal_cabinet') + '#user-info')

    # Get all ACTIVE quizzes and filter by access
    all_quizzes = Quiz.objects.filter(is_active=True).prefetch_related('companies', 'cabinet_groups', 'cabinet_users')
    quizzes = [quiz for quiz in all_quizzes if quiz.has_user_access(request.user)]
    quiz_data = []

    for quiz in quizzes:
        # Process description to handle HTML entities
        clean_description = html.unescape(strip_tags(quiz.description)) if quiz.description else ""
        
        quiz_info = {
            'id': quiz.id,
            'title': quiz.title,
            'description': quiz.description,  # Original for template processing
            'clean_description': clean_description,  # Clean version for display
            'youtube_video_id': quiz.youtube_video_id,
            'attempt_count': QuizAttempt.objects.filter(user=request.user, quiz=quiz).count(),
            'page': quiz.page,
        }

        if quiz.pdf_material:
            quiz_info['pdf_url'] = quiz.pdf_material.url
            quiz_info['pdf_filename'] = quiz.pdf_filename
        else:
            quiz_info['pdf_url'] = None
            quiz_info['pdf_filename'] = None

        # Check quiz status
        latest_attempt = QuizAttempt.objects.filter(
            user=request.user, 
            quiz=quiz, 
            completed=True
        ).order_by('-completed_at').first()
        
        if latest_attempt:
            quiz_info['is_completed'] = True
            quiz_info['is_passed'] = latest_attempt.score >= quiz.passing_score
            quiz_info['needs_retake'] = not quiz_info['is_passed']
        else:
            quiz_info['is_completed'] = False
            quiz_info['is_passed'] = False
            quiz_info['needs_retake'] = False

        quiz_data.append(quiz_info)

    # Sort quizzes: unpassed first, then retakes, then passed
    def quiz_sort_key(quiz):
        if not quiz['is_completed']:
            return (0, quiz['title'].lower())  # Not started - highest priority
        elif quiz['needs_retake']:
            return (1, quiz['title'].lower())  # Failed/needs retake - medium priority  
        else:
            return (2, quiz['title'].lower())  # Passed - lowest priority
    
    quiz_data.sort(key=quiz_sort_key)

    # Get all quiz attempts for history
    all_attempts = QuizAttempt.objects.filter(
        user=request.user, 
        completed=True
    ).select_related('quiz').order_by('-completed_at')
    
    # Add calculated fields for attempts
    for attempt in all_attempts:
        attempt.total_questions_count = attempt.quiz.questions.count()
        attempt.correct_answers_count = QuizAnswer.objects.filter(
            attempt=attempt,
            answer__is_correct=True
        ).count()

    # Get accessible pages for materials
    accessible_pages = []
    if cabinet_user.company:
        pages = Page.objects.filter(companies=cabinet_user.company, is_active=True).distinct().order_by('title')
        for page in pages:
            if page.has_user_access(request.user):
                accessible_pages.append(page)
    
    # Sort pages by title for consistent display
    accessible_pages.sort(key=lambda page: page.title.lower())

    # UI strings: default English; other languages via Django Translations (.po)
    current_language = get_language()
    lang_code = (current_language or 'en')[:2]
    _prev_lang = get_language() or 'en'
    translation_activate(lang_code)
    try:
        translations = {
            'passed': _('Passed'),
            'failed': _('Failed'),
            'not_started': _('New'),
            'start_test': _('Start Test'),
            'retake_test': _('Retake Test'),
            'preparation_materials': _('Preparation Materials'),
            'info_materials': _('Information and Materials'),
            'download_pdf': _('Download PDF'),
            'attempt': _('attempt'),
            'attempts': _('attempts'),
        }
    finally:
        translation_activate(_prev_lang)

    context = {
        'quizzes': quiz_data,
        'all_attempts': all_attempts,
        'accessible_pages': accessible_pages,
        'translations': translations,
    }

    return render(request, 'app_study/learning_hub.html', context)


@login_required
def get_cabinet_data_by_company(request):
    """API endpoint to get cabinet groups and users for one or more companies"""
    from app_cabinet.models import CabinetGroup, CabinetUser
    from app_conf.models import Company
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        company_ids = data.get('company_ids', [])
        
        # Support both single company_id (for backward compatibility) and multiple company_ids
        if not company_ids:
            company_id = data.get('company_id')
            if company_id:
                company_ids = [company_id]
        
        if not company_ids:
            return JsonResponse({'error': 'Company ID(s) are required'}, status=400)
        
        # Ensure company_ids is a list
        if not isinstance(company_ids, list):
            company_ids = [company_ids]
        
        # Convert to integers
        try:
            company_ids = [int(cid) for cid in company_ids]
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid company ID format'}, status=400)
        
        # Get companies
        companies = Company.objects.filter(id__in=company_ids)
        
        if not companies.exists():
            return JsonResponse({'error': 'No companies found'}, status=404)
        
        # Check if user has access to these companies for both pages and quizzes
        accessible_companies_pages = get_accessible_companies_for_page_user(request.user)
        accessible_companies_quizzes = get_accessible_companies_for_user(request.user)
        
        # Combine both access lists
        all_accessible_companies = set(accessible_companies_pages) | set(accessible_companies_quizzes)
        
        # Check access for all companies
        for company in companies:
            if company not in all_accessible_companies and not request.user.is_superuser:
                logger.warning(f"User {request.user.username} tried to access company {company.name} without permission")
                return JsonResponse({'error': f'No access to company {company.name}'}, status=403)
        
        # Get cabinet groups for all selected companies
        cabinet_groups = CabinetGroup.objects.filter(company_id__in=company_ids).order_by('name')
        groups_data = []
        seen_group_ids = set()
        for group in cabinet_groups:
            if group.id not in seen_group_ids:
                seen_group_ids.add(group.id)
                groups_data.append({
                    'id': group.id,
                    'name': group.name or f'Group {group.id}',
                    'company_id': group.company_id
                })
        
        # Get cabinet users for all selected companies (only active users)
        cabinet_users = CabinetUser.objects.filter(
            company_id__in=company_ids, user__is_active=True
        ).select_related('user', 'department', 'position').order_by('user__first_name', 'user__last_name')
        users_data = []
        seen_user_ids = set()
        departments_seen = {}
        positions_seen = {}
        for user in cabinet_users:
            if user.id not in seen_user_ids:
                seen_user_ids.add(user.id)
                full_name = f"{user.user.first_name} {user.user.last_name}".strip()
                if not full_name:
                    full_name = user.user.username
                # Use localized names for current site language
                dept = (user.department.get_name() or '') if user.department else ''
                pos = (user.position.get_name() or '') if user.position else ''
                if user.department and user.department_id not in departments_seen:
                    departments_seen[user.department_id] = dept
                if user.position and user.position_id not in positions_seen:
                    positions_seen[user.position_id] = pos
                users_data.append({
                    'id': user.id,
                    'name': full_name,
                    'company_id': user.company_id,
                    'department': dept,
                    'department_id': user.department_id,
                    'position': pos,
                    'position_id': user.position_id
                })
        departments_data = [{'id': k, 'name': v} for k, v in sorted(departments_seen.items())]
        positions_data = [{'id': k, 'name': v} for k, v in sorted(positions_seen.items())]
        
        company_names = ', '.join([c.name for c in companies])
        logger.info(f"Retrieved {len(groups_data)} groups and {len(users_data)} users for companies: {company_names}")
        
        return JsonResponse({
            'cabinet_groups': groups_data,
            'cabinet_users': users_data,
            'departments': departments_data,
            'positions': positions_data
        })
        
    except json.JSONDecodeError:
        logger.error("Invalid JSON in request body")
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error in get_cabinet_data_by_company: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def quiz_users_api(request, quiz_id):
    """API endpoint to get users who have access to a specific quiz"""
    from app_cabinet.models import CabinetUser
    from django.http import JsonResponse
    import logging
    
    logger = logging.getLogger(__name__)
    
    logger.info(f"quiz_users_api called for quiz_id: {quiz_id} by user: {request.user.username}")
    
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        # Check if user has access to this quiz
        accessible_companies = get_accessible_companies_for_user(request.user)
        if not request.user.is_superuser:
            if not quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
                logger.warning(f"User {request.user.username} tried to access quiz {quiz.id} without permission")
                return JsonResponse({'error': 'No access to this quiz'}, status=403)
        
        users_with_access = []
        
        # Get users from companies that have access to this quiz
        quiz_companies = quiz.companies.all()
        
        for company in quiz_companies:
            # Get all cabinet users from this company
            cabinet_users = CabinetUser.objects.filter(company=company).select_related('user', 'company')
            
            for cabinet_user in cabinet_users:
                user = cabinet_user.user
                
                # Check if this user has access to the quiz
                has_access = False
                access_type = None
                access_type_display = None
                
                # Check direct user access
                if quiz.cabinet_users.filter(id=cabinet_user.id).exists():
                    has_access = True
                    access_type = 'direct'
                    access_type_display = _('Direct Access')
                
                # Check group access
                elif quiz.cabinet_groups.filter(group__in=user.groups.all()).exists():
                    has_access = True
                    access_type = 'group'
                    access_type_display = _('Group Access')
                
                # Check company access (if no specific groups/users are defined)
                elif not quiz.cabinet_groups.exists() and not quiz.cabinet_users.exists():
                    has_access = True
                    access_type = 'company'
                    access_type_display = _('Company Access')
                
                if has_access:
                    full_name = f"{user.first_name} {user.last_name}".strip()
                    if not full_name:
                        full_name = user.username
                    
                    # Check if user has started or passed the quiz
                    attempts = QuizAttempt.objects.filter(quiz=quiz, user=user)
                    has_started = attempts.exists()
                    has_passed = attempts.filter(completed=True, score__gte=quiz.passing_score).exists()
                    
                    users_with_access.append({
                        'id': user.id,
                        'username': user.username,
                        'full_name': full_name,
                        'email': user.email,
                        'company': company.name,
                        'access_type': access_type,
                        'access_type_display': access_type_display,
                        'has_started': has_started,
                        'has_passed': has_passed
                    })
        
        # Remove duplicates (same user might appear multiple times)
        unique_users = []
        seen_users = set()
        for user in users_with_access:
            if user['id'] not in seen_users:
                unique_users.append(user)
                seen_users.add(user['id'])
        
        logger.info(f"Retrieved {len(unique_users)} users with access to quiz {quiz.id}")
        
        return JsonResponse({
            'users': unique_users,
            'total_count': len(unique_users)
        })
        
    except Exception as e:
        logger.error(f"Error in quiz_users_api: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def quiz_results_api(request, quiz_id):
    """API endpoint to get quiz results for a specific quiz"""
    from django.http import JsonResponse
    import logging
    
    logger = logging.getLogger(__name__)
    
    logger.info(f"quiz_results_api called for quiz_id: {quiz_id} by user: {request.user.username}")
    
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        # Check if user has access to this quiz
        accessible_companies = get_accessible_companies_for_user(request.user)
        if not request.user.is_superuser:
            if not quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
                logger.warning(f"User {request.user.username} tried to access quiz {quiz.id} results without permission")
                return JsonResponse({'error': 'No access to this quiz'}, status=403)
        
        # Get all quiz attempts for this quiz
        attempts = QuizAttempt.objects.filter(quiz=quiz).select_related('user').order_by('-started_at')
        
        results = []
        for attempt in attempts:
            # Calculate max score for this quiz
            max_score = quiz.questions.aggregate(total=Sum('answers__score'))['total'] or 0
            
            # Determine if passed
            passed = attempt.score >= quiz.passing_score if attempt.completed else False
            
            # Format user name
            full_name = f"{attempt.user.first_name} {attempt.user.last_name}".strip()
            if not full_name:
                full_name = attempt.user.username
            
            # Format date
            date_str = attempt.started_at.strftime('%d.%m.%Y %H:%M')
            if attempt.completed_at:
                date_str = attempt.completed_at.strftime('%d.%m.%Y %H:%M')
            
            results.append({
                'id': attempt.id,
                'user_id': attempt.user.id,
                'user_name': full_name,
                'user_email': attempt.user.email,
                'score': attempt.score,
                'max_score': max_score,
                'completed': attempt.completed,
                'passed': passed,
                'date': date_str,
                'started_at': attempt.started_at.isoformat(),
                'completed_at': attempt.completed_at.isoformat() if attempt.completed_at else None
            })
        
        logger.info(f"Retrieved {len(results)} results for quiz {quiz.id}")
        
        return JsonResponse({
            'results': results,
            'total_count': len(results),
            'quiz_info': {
                'id': quiz.id,
                'title': quiz.title,
                'passing_score': quiz.passing_score
            }
        })
        
    except Exception as e:
        logger.error(f"Error in quiz_results_api: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def simulate_quiz_pass_api(request, quiz_id):
    """API endpoint to simulate a positive quiz pass for a user"""
    from django.http import JsonResponse
    from django.utils import timezone
    import logging
    import json
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    logger.info(f"simulate_quiz_pass_api called for quiz_id: {quiz_id} by user: {request.user.username}")
    
    try:
        # Parse request body
        data = json.loads(request.body)
        user_id = data.get('user_id')
        
        if not user_id:
            return JsonResponse({'error': 'User ID is required'}, status=400)
        
        # Get quiz
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        # Check if requesting user has access to this quiz (admin/manager check)
        accessible_companies = get_accessible_companies_for_user(request.user)
        if not request.user.is_superuser:
            if not quiz.companies.filter(pk__in=[c.pk for c in accessible_companies]).exists():
                logger.warning(f"User {request.user.username} tried to simulate pass for quiz {quiz.id} without permission")
                return JsonResponse({'error': 'No access to this quiz'}, status=403)
        
        # Get target user
        target_user = get_object_or_404(User, id=user_id)
        
        # Check if user already has a passing attempt
        existing_passing_attempt = QuizAttempt.objects.filter(
            user=target_user,
            quiz=quiz,
            completed=True,
            score__gte=quiz.passing_score
        ).first()
        
        if existing_passing_attempt:
            return JsonResponse({
                'error': 'User already has a passing attempt for this quiz',
                'attempt_id': existing_passing_attempt.id
            }, status=400)
        
        # Calculate the score needed to pass (use passing_score or slightly above)
        # Get all questions to calculate max possible score
        questions = quiz.questions.all()
        total_questions = questions.count()
        
        if total_questions == 0:
            return JsonResponse({'error': 'Quiz has no questions'}, status=400)
        
        # Use passing score as the simulated score
        simulated_score = quiz.passing_score
        
        # Create a new quiz attempt with passing score
        attempt = QuizAttempt.objects.create(
            user=target_user,
            quiz=quiz,
            score=simulated_score,
            completed=True,
            started_at=timezone.now(),
            completed_at=timezone.now()
        )
        
        # Create simulated answers for all questions (selecting correct answers)
        for question in questions:
            correct_answer = question.answers.filter(is_correct=True).first()
            if correct_answer:
                QuizAnswer.objects.create(
                    attempt=attempt,
                    question=question,
                    answer=correct_answer,
                    is_correct=True
                )
        
        logger.info(f"Simulated quiz pass created: attempt_id={attempt.id}, user={target_user.username}, quiz={quiz.title}, score={simulated_score}")
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully simulated quiz pass for {target_user.get_full_name() or target_user.username}',
            'attempt': {
                'id': attempt.id,
                'score': attempt.score,
                'completed_at': attempt.completed_at.isoformat()
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error in simulate_quiz_pass_api: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def email_accounts_api(request):
    """API endpoint to get available email accounts"""
    try:
        from app_conf.models import MailAccount
        
        # Get active email accounts
        accounts = MailAccount.objects.filter(is_active=True).select_related('server')
        
        accounts_data = []
        for account in accounts:
            accounts_data.append({
                'id': account.id,
                'username': account.username,
                'server_name': account.server.name,
                'server_host': account.server.smtp_host,
                'server_port': account.server.smtp_port
            })
        
        return JsonResponse({
            'accounts': accounts_data,
            'total_count': len(accounts_data)
        })
        
    except Exception as e:
        logger.error(f"Error in email_accounts_api: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def quiz_default_template_api(request, quiz_id):
    """API endpoint to get/set default email template for a quiz"""
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        if request.method == 'GET':
            # Get default template
            from .models import EmailTemplate
            template = EmailTemplate.objects.filter(quiz=quiz, is_default=True).first()
            
            if template:
                return JsonResponse({
                    'template': {
                        'id': template.id,
                        'subject': template.subject,
                        'body': template.body
                    }
                })
            else:
                return JsonResponse({'template': None})
        
        elif request.method == 'POST':
            # Save default template
            import json
            data = json.loads(request.body)
            
            from .models import EmailTemplate
            
            # Create or update default template
            template, created = EmailTemplate.objects.get_or_create(
                quiz=quiz,
                is_default=True,
                defaults={
                    'subject': data.get('subject', ''),
                    'body': data.get('body', '')
                }
            )
            
            if not created:
                template.subject = data.get('subject', '')
                template.body = data.get('body', '')
                template.save()
            
            return JsonResponse({
                'success': True,
                'template_id': template.id
            })
        
    except Exception as e:
        logger.error(f"Error in quiz_default_template_api: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def send_test_email_api(request, quiz_id):
    """API endpoint to send a test email"""
    try:
        import json
        data = json.loads(request.body)
        
        quiz = get_object_or_404(Quiz, id=quiz_id)
        email_account_id = data.get('email_account')
        subject_template = data.get('subject', '')
        body_template = data.get('body', '')
        
        if not email_account_id:
            return JsonResponse({'success': False, 'error': 'Email account not specified'})
        
        # Get email account
        from app_conf.models import MailAccount
        email_account = get_object_or_404(MailAccount, id=email_account_id)
        
        # Get site settings for URL construction
        from app_conf.models import SiteSettings
        site_settings = SiteSettings.get_settings()
        base_url = site_settings.get_site_url()
        
        # Replace template variables with test data
        personalized_subject = subject_template.replace('{{ user_name }}', request.user.get_full_name() or request.user.username)
        personalized_subject = personalized_subject.replace('{{ user_email }}', request.user.email)
        personalized_subject = personalized_subject.replace('{{ quiz_title }}', quiz.title)
        personalized_subject = personalized_subject.replace('{{ quiz_url }}', f"{base_url}/en/app_study/quiz/start/{quiz.id}/")
        personalized_subject = personalized_subject.replace('{{ company_name }}', getattr(request.user.cabinet.company, 'name', 'Test Company') if hasattr(request.user, 'cabinet') else 'Test Company')
        personalized_subject = personalized_subject.replace('{{ current_date }}', timezone.now().strftime('%d.%m.%Y'))
        
        personalized_body = body_template.replace('{{ user_name }}', request.user.get_full_name() or request.user.username)
        personalized_body = personalized_body.replace('{{ user_email }}', request.user.email)
        personalized_body = personalized_body.replace('{{ quiz_title }}', quiz.title)
        personalized_body = personalized_body.replace('{{ quiz_url }}', f"{base_url}/en/app_study/quiz/start/{quiz.id}/")
        personalized_body = personalized_body.replace('{{ company_name }}', getattr(request.user.cabinet.company, 'name', 'Test Company') if hasattr(request.user, 'cabinet') else 'Test Company')
        personalized_body = personalized_body.replace('{{ current_date }}', timezone.now().strftime('%d.%m.%Y'))
        
        # Send personalized test email using direct SMTP connection
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            import ssl
            
            # Create the email message
            msg = MIMEMultipart()
            msg['From'] = email_account.username
            msg['To'] = request.user.email
            msg['Subject'] = personalized_subject
            msg.attach(MIMEText(personalized_body, 'plain'))
            
            # Connect to the server directly using smtplib
            if email_account.server.use_ssl:
                # Create SSL context without keyfile issues
                context = ssl.create_default_context()
                context.check_hostname = False
                context.verify_mode = ssl.CERT_NONE
                
                # Connect with SSL
                smtp = smtplib.SMTP_SSL(
                    host=email_account.server.smtp_host,
                    port=email_account.server.smtp_port,
                    context=context
                )
            else:
                # Connect without SSL
                smtp = smtplib.SMTP(
                    host=email_account.server.smtp_host,
                    port=email_account.server.smtp_port
                )
                
                # Use TLS if needed
                if email_account.server.use_tls:
                    smtp.starttls()
            
            # Login and send
            smtp.login(email_account.username, email_account.password)
            smtp.send_message(msg)
            smtp.quit()
            
            return JsonResponse({
                'success': True,
                'message': 'Test email sent successfully with personalized content!'
            })
            
        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"SMTP Authentication error: {str(e)}")
            if "534" in str(e) and "5.7.9" in str(e):
                return JsonResponse({
                    'success': False,
                    'error': 'Authentication error. Make sure you\'re using an app password if you have two-factor authentication enabled.'
                })
            return JsonResponse({
                'success': False,
                'error': f'Authentication failed: {str(e)}'
            })
        except Exception as email_error:
            logger.error(f"Error sending test email: {str(email_error)}")
            return JsonResponse({
                'success': False,
                'error': f'Failed to send test email: {str(email_error)}'
            })
        
    except Exception as e:
        logger.error(f"Error in send_test_email_api: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def send_reminder_emails_api(request, quiz_id):
    """API endpoint to send reminder emails to selected users"""
    try:
        import json
        data = json.loads(request.body)
        
        quiz = get_object_or_404(Quiz, id=quiz_id)
        email_account_id = data.get('email_account')
        subject_template = data.get('subject', '')
        body_template = data.get('body', '')
        user_ids = data.get('user_ids', [])
        save_template = data.get('save_template', False)
        
        if not email_account_id:
            return JsonResponse({'success': False, 'error': 'Email account not specified'})
        
        if not user_ids:
            return JsonResponse({'success': False, 'error': 'No users selected'})
        
        # Get email account
        from app_conf.models import MailAccount
        email_account = get_object_or_404(MailAccount, id=email_account_id)
        
        # Save template if requested
        if save_template:
            from .models import EmailTemplate
            template, created = EmailTemplate.objects.get_or_create(
                quiz=quiz,
                is_default=True,
                defaults={
                    'subject': subject_template,
                    'body': body_template
                }
            )
            if not created:
                template.subject = subject_template
                template.body = body_template
                template.save()
        
        # Get users
        users = User.objects.filter(id__in=user_ids)
        
        # Get site settings for URL construction
        from app_conf.models import SiteSettings
        site_settings = SiteSettings.get_settings()
        base_url = site_settings.get_site_url()
        
        sent_count = 0
        failed_count = 0
        error_messages = []
        
        for user in users:
            try:
                # Replace template variables
                personalized_subject = subject_template.replace('{{ user_name }}', user.get_full_name() or user.username)
                personalized_subject = personalized_subject.replace('{{ user_email }}', user.email)
                personalized_subject = personalized_subject.replace('{{ quiz_title }}', quiz.title)
                personalized_subject = personalized_subject.replace('{{ quiz_url }}', f"{base_url}/en/app_study/quiz/start/{quiz.id}/")
                personalized_subject = personalized_subject.replace('{{ company_name }}', getattr(user.cabinet.company, 'name', 'Unknown Company'))
                personalized_subject = personalized_subject.replace('{{ current_date }}', timezone.now().strftime('%d.%m.%Y'))
                
                personalized_body = body_template.replace('{{ user_name }}', user.get_full_name() or user.username)
                personalized_body = personalized_body.replace('{{ user_email }}', user.email)
                personalized_body = personalized_body.replace('{{ quiz_title }}', quiz.title)
                personalized_body = personalized_body.replace('{{ quiz_url }}', f"{base_url}/en/app_study/quiz/start/{quiz.id}/")
                personalized_body = personalized_body.replace('{{ company_name }}', getattr(user.cabinet.company, 'name', 'Unknown Company'))
                personalized_body = personalized_body.replace('{{ current_date }}', timezone.now().strftime('%d.%m.%Y'))
                
                # Send email using direct SMTP connection
                import smtplib
                from email.mime.text import MIMEText
                from email.mime.multipart import MIMEMultipart
                import ssl
                
                # Create the email message
                msg = MIMEMultipart()
                msg['From'] = email_account.username
                msg['To'] = user.email
                msg['Subject'] = personalized_subject
                msg.attach(MIMEText(personalized_body, 'plain'))
                
                # Connect to the server directly using smtplib
                if email_account.server.use_ssl:
                    # Create SSL context without keyfile issues
                    context = ssl.create_default_context()
                    context.check_hostname = False
                    context.verify_mode = ssl.CERT_NONE
                    
                    # Connect with SSL
                    smtp = smtplib.SMTP_SSL(
                        host=email_account.server.smtp_host,
                        port=email_account.server.smtp_port,
                        context=context
                    )
                else:
                    # Connect without SSL
                    smtp = smtplib.SMTP(
                        host=email_account.server.smtp_host,
                        port=email_account.server.smtp_port
                    )
                    
                    # Use TLS if needed
                    if email_account.server.use_tls:
                        smtp.starttls()
                
                # Login and send
                smtp.login(email_account.username, email_account.password)
                smtp.send_message(msg)
                smtp.quit()
                
                sent_count += 1
                    
            except Exception as e:
                error_msg = f"Failed to send email to {user.email}: {str(e)}"
                logger.error(error_msg)
                error_messages.append(error_msg)
                failed_count += 1
        
        # Log the immediate email sending
        from .models import ImmediateEmailLog
        email_log = ImmediateEmailLog.objects.create(
            quiz=quiz,
            sent_by=request.user,
            subject=subject_template,
            sent_count=sent_count,
            failed_count=failed_count,
            error_message='\n'.join(error_messages) if error_messages else '',
            email_account=email_account
        )
        # Add recipients to the log
        email_log.recipients.set(users)
        
        return JsonResponse({
            'success': True,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'total_count': len(user_ids)
        })
        
    except Exception as e:
        logger.error(f"Error in send_reminder_emails_api: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def scheduled_reminders_api(request, quiz_id):
    """API endpoint to get/set scheduled reminders for a quiz"""
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        if request.method == 'GET':
            # Get all scheduled reminders for this quiz
            from .models import ScheduledReminder
            reminders = ScheduledReminder.objects.filter(quiz=quiz).select_related('email_account')
            
            reminders_data = []
            for reminder in reminders:
                reminders_data.append({
                    'id': reminder.id,
                    'interval': reminder.interval,
                    'interval_display': reminder.get_interval_display(),
                    'is_active': reminder.is_active,
                    'next_send': reminder.next_send.isoformat() if reminder.next_send else None,
                    'last_sent': reminder.last_sent.isoformat() if reminder.last_sent else None,
                    'subject_template': reminder.subject_template,
                    'body_template': reminder.body_template,
                    'email_account': {
                        'id': reminder.email_account.id,
                        'username': reminder.email_account.username,
                        'server_name': reminder.email_account.server.name
                    },
                    'target_users': list(reminder.target_users.values_list('id', flat=True)),
                    'target_groups': list(reminder.target_groups.values_list('id', flat=True)),
                    'target_all_users': reminder.target_all_users,
                    'created_at': reminder.created_at.isoformat()
                })
            
            return JsonResponse({
                'reminders': reminders_data,
                'total_count': len(reminders_data)
            })
        
        elif request.method == 'POST':
            # Create or update scheduled reminder
            import json
            data = json.loads(request.body)
            
            from .models import ScheduledReminder
            from app_conf.models import MailAccount
            
            email_account = get_object_or_404(MailAccount, id=data.get('email_account'))
            
            reminder_data = {
                'quiz': quiz,
                'email_account': email_account,
                'subject_template': data.get('subject_template', ''),
                'body_template': data.get('body_template', ''),
                'interval': data.get('interval', 'week'),
                'is_active': data.get('is_active', True),
                'target_all_users': data.get('target_all_users', False)
            }
            
            # Set next_send date
            from datetime import datetime, timedelta
            from dateutil.relativedelta import relativedelta
            
            if reminder_data['interval'] == 'once':
                # For once interval, use the provided datetime
                once_datetime = data.get('once_datetime')
                if not once_datetime:
                    return JsonResponse({'success': False, 'error': 'Date and time required for once interval'})
                
                try:
                    next_send = datetime.fromisoformat(once_datetime.replace('Z', '+00:00'))
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Invalid date format'})
            else:
                # For recurring intervals, calculate from now
                base_date = datetime.now()
                if reminder_data['interval'] == 'week':
                    next_send = base_date + timedelta(weeks=1)
                elif reminder_data['interval'] == 'month':
                    next_send = base_date + relativedelta(months=1)
                elif reminder_data['interval'] == 'quarter':
                    next_send = base_date + relativedelta(months=3)
                elif reminder_data['interval'] == 'half_year':
                    next_send = base_date + relativedelta(months=6)
                else:
                    next_send = base_date + timedelta(weeks=1)
            
            reminder_data['next_send'] = next_send
            
            # Create the reminder
            reminder = ScheduledReminder.objects.create(**reminder_data)
            
            # Add target users and groups
            if data.get('target_users'):
                reminder.target_users.set(data.get('target_users'))
            
            if data.get('target_groups'):
                reminder.target_groups.set(data.get('target_groups'))
            
            return JsonResponse({
                'success': True,
                'reminder_id': reminder.id,
                'message': 'Scheduled reminder created successfully'
            })
        
        elif request.method == 'DELETE':
            # Delete scheduled reminder
            reminder_id = request.GET.get('reminder_id')
            if not reminder_id:
                return JsonResponse({'success': False, 'error': 'Reminder ID required'})
            
            from .models import ScheduledReminder
            reminder = get_object_or_404(ScheduledReminder, id=reminder_id, quiz=quiz)
            reminder.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Scheduled reminder deleted successfully'
            })
        
    except Exception as e:
        logger.error(f"Error in scheduled_reminders_api: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def reminder_logs_api(request, quiz_id):
    """API endpoint to get reminder logs for a quiz"""
    try:
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        from .models import ReminderLog, ImmediateEmailLog
        
        # Get scheduled reminder logs
        scheduled_logs = ReminderLog.objects.filter(
            scheduled_reminder__quiz=quiz
        ).select_related('scheduled_reminder').order_by('-sent_at')
        
        # Get immediate email logs
        immediate_logs = ImmediateEmailLog.objects.filter(
            quiz=quiz
        ).select_related('sent_by', 'email_account').order_by('-sent_at')
        
        logs_data = []
        
        # Add scheduled reminder logs
        for log in scheduled_logs:
            # Get recipients for this log
            recipients = []
            for recipient in log.recipients.all():
                recipients.append({
                    'id': recipient.id,
                    'username': recipient.username,
                    'full_name': recipient.get_full_name(),
                    'email': recipient.email
                })
            
            logs_data.append({
                'id': f"scheduled_{log.id}",
                'type': 'scheduled',
                'scheduled_reminder_id': log.scheduled_reminder.id,
                'interval': log.scheduled_reminder.get_interval_display(),
                'sent_at': log.sent_at.isoformat(),
                'sent_count': log.sent_count,
                'failed_count': log.failed_count,
                'error_message': log.error_message,
                'subject': log.scheduled_reminder.subject_template,
                'sent_by': 'System (Scheduled)',
                'recipients': recipients
            })
        
        # Add immediate email logs
        for log in immediate_logs:
            # Get recipients for this log
            recipients = []
            for recipient in log.recipients.all():
                recipients.append({
                    'id': recipient.id,
                    'username': recipient.username,
                    'full_name': recipient.get_full_name(),
                    'email': recipient.email
                })
            
            logs_data.append({
                'id': f"immediate_{log.id}",
                'type': 'immediate',
                'sent_at': log.sent_at.isoformat(),
                'sent_count': log.sent_count,
                'failed_count': log.failed_count,
                'error_message': log.error_message,
                'subject': log.subject,
                'sent_by': log.sent_by.get_full_name() or log.sent_by.username,
                'email_account': log.email_account.username if log.email_account else 'Unknown',
                'recipients': recipients
            })
        
        # Sort all logs by sent_at (most recent first)
        logs_data.sort(key=lambda x: x['sent_at'], reverse=True)
        
        return JsonResponse({
            'logs': logs_data,
            'total_count': len(logs_data)
        })
        
    except Exception as e:
        logger.error(f"Error in reminder_logs_api: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})
