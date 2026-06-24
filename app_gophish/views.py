# app_gophish/views.py

import json
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.translation import gettext_lazy as _, get_language
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.utils import timezone
from .pagination_utils import paginate_gophish_queryset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    GophishServer, GophishGroup, GophishTemplate, GophishLandingPage,
    GophishSendingProfile, GophishCampaign, GophishEvent, GophishSyncLog, AccessGophish,
    GophishGuide, GophishGuideTranslation
)
from .forms import (
    GophishServerForm, GophishCampaignForm, GophishTemplateForm,
    GophishLandingPageForm, GophishSendingProfileForm, GophishGroupForm,
    CampaignLaunchForm, SyncForm
)
from .api_client import gophish_manager, GophishAPIError
from .tasks import sync_gophish_data
from .sync_utils import sync_gophish_data_direct
from app_conf.models import Country

logger = logging.getLogger(__name__)


def get_user_accessible_companies_gophish(user):
    """
    Get list of companies accessible to user based on AccessGophish
    Returns None if user has access to all companies
    """
    if not user.is_authenticated:
        return []
    
    # Superusers have access to all companies
    if user.is_superuser:
        return None
    
    # Get user groups
    user_groups = user.groups.all()
    if not user_groups.exists():
        return []
    
    accessible_companies = set()
    has_unrestricted_access = False
    
    for group in user_groups:
        try:
            access = AccessGophish.objects.get(group=group, has_access=True)
            # If no companies specified, user has access to all
            if not access.companies.exists():
                has_unrestricted_access = True
                break
            # Add companies from this access configuration
            accessible_companies.update(access.companies.all())
        except AccessGophish.DoesNotExist:
            continue
    
    # If any group has unrestricted access, return None (all companies)
    if has_unrestricted_access:
        return None
    
    return list(accessible_companies) if accessible_companies else []


def check_gophish_access(user, required_permission=None):
    """
    Check if user has access to Gophish based on AccessGophish model
    """
    if not user.is_authenticated:
        return False
    
    # Superusers have full access
    if user.is_superuser:
        return True
    
    # Get user groups
    user_groups = user.groups.all()
    if not user_groups.exists():
        return False
    
    # Check if any of user's groups have Gophish access with has_access=True
    for group in user_groups:
        try:
            access = AccessGophish.objects.get(group=group)
            if access.has_access:
                # If specific permission is required, check it
                if required_permission:
                    if required_permission == 'view_campaigns' and not access.can_view_campaigns:
                        continue
                    elif required_permission == 'view_templates' and not access.can_view_templates:
                        continue
                    elif required_permission == 'view_landing_pages' and not access.can_view_landing_pages:
                        continue
                    elif required_permission == 'view_sending_profiles' and not access.can_view_sending_profiles:
                        continue
                    elif required_permission == 'view_groups' and not access.can_view_groups:
                        continue
                    elif required_permission == 'manage_servers' and not access.can_manage_servers:
                        continue
                    elif required_permission == 'sync' and not access.can_sync:
                        continue
                
                return True
        except AccessGophish.DoesNotExist:
            continue
    
    # No access if no group has has_access=True
    return False


def gophish_access_required(permission=None):
    """
    Decorator to check Gophish access
    """
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not check_gophish_access(request.user, permission):
                raise PermissionDenied(_("You don't have permission to access Gophish"))
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


@login_required
@gophish_access_required()
def dashboard(request):
    """Main dashboard for Gophish integration"""
    # Filter servers by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    
    servers_qs = GophishServer.objects.filter(is_active=True)
    if accessible_companies is not None:
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    campaigns_qs = GophishCampaign.objects.select_related('server').order_by('-created_at')
    if accessible_companies is not None:
        campaigns_qs = campaigns_qs.filter(server__company__in=accessible_companies)
    campaigns_qs = campaigns_qs[:10]
    
    context = {
        'servers': servers_qs,
        'campaigns': campaigns_qs,
        'recent_events': GophishEvent.objects.select_related('campaign').filter(campaign__server__in=servers_qs).order_by('-timestamp')[:10],
        'sync_logs': GophishSyncLog.objects.select_related('server').filter(server__in=servers_qs).order_by('-started_at')[:5],
    }
    return render(request, 'app_gophish/dashboard.html', context)


@login_required
@gophish_access_required()
@require_http_methods(["GET"])
def gophish_guide(request):
    """Return JSON { content: html } for the Gophish guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = GophishGuide.objects.first()
    if guide:
        if country:
            trans = GophishGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = GophishGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def gophish_guide_translate(request):
    """API for AI translation of Gophish guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
@gophish_access_required()
def server_list(request):
    """List all Gophish servers"""
    # Filter servers by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    
    servers = GophishServer.objects.all().order_by('name')
    if accessible_companies is not None:
        servers = servers.filter(company__in=accessible_companies)
    
    return render(request, 'app_gophish/server_list.html', {'servers': servers})


@login_required
@gophish_access_required('manage_servers')
def server_create(request):
    """Create a new Gophish server"""
    if request.method == 'POST':
        logger.info(f"POST request received for server creation")
        logger.info(f"POST data: {request.POST}")
        
        form = GophishServerForm(request.POST, user=request.user)
        logger.info(f"Form is valid: {form.is_valid()}")
        
        if form.is_valid():
            try:
                server = form.save()
                logger.info(f"Server created successfully with ID: {server.id}")
                messages.success(request, _('Gophish server created successfully'))
                return redirect('app_gophish:server_detail', server_id=server.id)
            except Exception as e:
                logger.error(f"Error creating server: {str(e)}")
                messages.error(request, _('Error creating server: {}').format(str(e)))
        else:
            # Log form errors for debugging
            logger.error(f"Form validation errors: {form.errors}")
            logger.error(f"Form non-field errors: {form.non_field_errors()}")
            for field, errors in form.errors.items():
                logger.error(f"Field '{field}' errors: {errors}")
    else:
        logger.info("GET request for server creation form")
        form = GophishServerForm(user=request.user)
    
    return render(request, 'app_gophish/server_form.html', {'form': form, 'title': _('Create Server')})


@login_required
@gophish_access_required()
def server_detail(request, server_id):
    """View server details and test connection"""
    server = get_object_or_404(GophishServer, id=server_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this server"))
    
    # Test connection
    connection_status = gophish_manager.test_server_connection(server)
    
    context = {
        'server': server,
        'connection_status': connection_status,
        'campaigns': server.campaigns.all().order_by('-created_at')[:5],
        'groups': server.groups.all().order_by('name')[:5],
        'templates': server.templates.all().order_by('name')[:5],
    }
    
    return render(request, 'app_gophish/server_detail.html', context)


@login_required
@gophish_access_required('sync')
def server_sync_direct(request, server_id):
    """Direct synchronization for a server"""
    server = get_object_or_404(GophishServer, id=server_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this server"))
    
    if request.method == 'POST':
        sync_type = request.POST.get('sync_type', 'full')
        force_update = request.POST.get('force_update') == 'on'
        
        try:
            result = sync_gophish_data_direct(server.id, sync_type, force_update)
            if result['status'] == 'completed':
                messages.success(request, _('Synchronization completed successfully. Processed: {}, Created: {}, Updated: {}').format(
                    result.get('records_processed', 0),
                    result.get('records_created', 0),
                    result.get('records_updated', 0)
                ))
            else:
                messages.error(request, _('Synchronization failed: {}').format(result.get('message', 'Unknown error')))
        except Exception as e:
            messages.error(request, _('Synchronization failed: {}').format(str(e)))
    
    return redirect('app_gophish:server_detail', server_id=server.id)


@login_required
@gophish_access_required('manage_servers')
def server_edit(request, server_id):
    """Edit a Gophish server"""
    server = get_object_or_404(GophishServer, id=server_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this server"))
    
    if request.method == 'POST':
        form = GophishServerForm(request.POST, instance=server, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, _('Server updated successfully'))
            return redirect('app_gophish:server_detail', server_id=server.id)
    else:
        form = GophishServerForm(instance=server, user=request.user)
    
    return render(request, 'app_gophish/server_form.html', {
        'form': form,
        'title': _('Edit Server'),
        'server': server
    })


@login_required
@gophish_access_required('manage_servers')
def server_delete(request, server_id):
    """Delete a Gophish server"""
    server = get_object_or_404(GophishServer, id=server_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this server"))
    
    if request.method == 'POST':
        try:
            server_name = server.name
            server.delete()
            messages.success(request, _('Server "{}" deleted successfully').format(server_name))
            return redirect('app_gophish:server_list')
        except Exception as e:
            logger.error(f"Error deleting server {server_id}: {str(e)}")
            messages.error(request, _('Error deleting server: {}').format(str(e)))
            return redirect('app_gophish:server_detail', server_id=server.id)
    
    # Check if server has associated data
    has_campaigns = server.campaigns.exists()
    has_templates = server.templates.exists()
    has_landing_pages = server.landing_pages.exists()
    has_sending_profiles = server.sending_profiles.exists()
    has_groups = server.groups.exists()
    
    context = {
        'server': server,
        'has_campaigns': has_campaigns,
        'has_templates': has_templates,
        'has_landing_pages': has_landing_pages,
        'has_sending_profiles': has_sending_profiles,
        'has_groups': has_groups,
        'has_associated_data': any([has_campaigns, has_templates, has_landing_pages, has_sending_profiles, has_groups])
    }
    
    return render(request, 'app_gophish/server_delete.html', context)


@login_required
@gophish_access_required('view_campaigns')
def campaign_list(request):
    """List all campaigns with filtering and pagination"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    campaigns = GophishCampaign.objects.select_related('server', 'template', 'landing_page', 'sending_profile').order_by('-created_at')
    if accessible_companies is not None:
        campaigns = campaigns.filter(server__company__in=accessible_companies)
    
    # Filtering
    search = request.GET.get('search', '')
    server_id = request.GET.get('server', '')
    status = request.GET.get('status', '')
    
    if search:
        campaigns = campaigns.filter(
            Q(name__icontains=search) |
            Q(template__name__icontains=search) |
            Q(landing_page__name__icontains=search)
        )
    
    if server_id:
        campaigns = campaigns.filter(server_id=server_id)
    
    if status:
        campaigns = campaigns.filter(status=status)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, campaigns)
    
    # Filter servers for context based on accessible companies
    servers_qs = GophishServer.objects.filter(is_active=True)
    if accessible_companies is not None:
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    context = {
        **pagination_context,
        'search': search,
        'server_id': server_id,
        'status': status,
        'servers': servers_qs,
        'status_choices': GophishCampaign.CAMPAIGN_STATUS_CHOICES,
    }
    
    return render(request, 'app_gophish/campaign_list.html', context)


@login_required
@gophish_access_required('view_campaigns')
def campaign_export_excel(request):
    """Export campaigns to Excel with styling and colors"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    campaigns = GophishCampaign.objects.select_related(
        'server', 'template', 'landing_page', 'sending_profile'
    ).prefetch_related('groups').order_by('-created_at')
    
    if accessible_companies is not None:
        campaigns = campaigns.filter(server__company__in=accessible_companies)
    
    # Apply same filters as campaign_list view
    search = request.GET.get('search', '')
    server_id = request.GET.get('server', '')
    status = request.GET.get('status', '')
    
    if search:
        campaigns = campaigns.filter(
            Q(name__icontains=search) |
            Q(template__name__icontains=search) |
            Q(landing_page__name__icontains=search)
        )
    
    if server_id:
        campaigns = campaigns.filter(server_id=server_id)
    
    if status:
        campaigns = campaigns.filter(status=status)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaigns"
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Status colors
    status_colors = {
        'draft': 'D3D3D3',      # Light Gray
        'running': 'FFC000',    # Orange
        'completed': '00B050',  # Green
        'paused': '92D050',     # Light Green
        'error': 'C00000',      # Red
    }
    
    # ========== SUMMARY SHEET ==========
    # Calculate statistics
    total_campaigns = campaigns.count()
    status_counts = {}
    for status_key, status_label in GophishCampaign.CAMPAIGN_STATUS_CHOICES:
        status_counts[status_label] = campaigns.filter(status=status_key).count()
    
    total_targets = sum(c.total_targets for c in campaigns)
    total_emails_sent = sum(c.emails_sent for c in campaigns)
    total_emails_opened = sum(c.emails_opened for c in campaigns)
    total_links_clicked = sum(c.links_clicked for c in campaigns)
    total_credentials_submitted = sum(c.credentials_submitted for c in campaigns)
    total_data_submitted = sum(c.data_submitted for c in campaigns)
    
    # Calculate percentages
    open_rate = (total_emails_opened / total_emails_sent * 100) if total_emails_sent > 0 else 0
    click_rate = (total_links_clicked / total_emails_sent * 100) if total_emails_sent > 0 else 0
    credential_rate = (total_credentials_submitted / total_emails_sent * 100) if total_emails_sent > 0 else 0
    
    # Write summary title
    cell = ws_summary.cell(row=1, column=1, value=str(_('Gophish Campaigns Export Summary')))
    cell.font = Font(bold=True, size=16, color='4472C4')
    ws_summary.merge_cells('A1:D1')
    
    # Export information
    row = 3
    ws_summary.cell(row=row, column=1, value=str(_('Export Date:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Total Campaigns:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=total_campaigns)
    
    # Status breakdown with colors
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Status Breakdown'))).font = Font(bold=True, size=12)
    row += 1
    
    for status_label, count in status_counts.items():
        ws_summary.cell(row=row, column=1, value=status_label).font = Font(bold=True)
        cell = ws_summary.cell(row=row, column=2, value=count)
        # Find status key for color
        for status_key, label in GophishCampaign.CAMPAIGN_STATUS_CHOICES:
            if label == status_label:
                color = status_colors.get(status_key, 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                if status_key in ['running', 'completed', 'error']:
                    cell.font = Font(bold=True, color='FFFFFF')
                break
        row += 1
    
    # Overall statistics with colors
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Overall Statistics'))).font = Font(bold=True, size=12)
    row += 1
    
    stats_data = [
        (_('Total Targets'), total_targets, '4472C4'),
        (_('Emails Sent'), total_emails_sent, '70AD47'),
        (_('Emails Opened'), total_emails_opened, 'FFC000'),
        (_('Links Clicked'), total_links_clicked, 'ED7D31'),
        (_('Credentials Submitted'), total_credentials_submitted, 'C00000'),
        (_('Data Submitted'), total_data_submitted, 'A5A5A5'),
    ]
    
    for label, value, color in stats_data:
        ws_summary.cell(row=row, column=1, value=str(label)).font = Font(bold=True)
        cell = ws_summary.cell(row=row, column=2, value=value)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        row += 1
    
    # Engagement rates
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Engagement Rates'))).font = Font(bold=True, size=12)
    row += 1
    
    rates_data = [
        (_('Open Rate'), f'{open_rate:.2f}%', '70AD47' if open_rate < 50 else 'FFC000'),
        (_('Click Rate'), f'{click_rate:.2f}%', 'FFC000' if click_rate < 30 else 'ED7D31'),
        (_('Credential Submission Rate'), f'{credential_rate:.2f}%', 'ED7D31' if credential_rate < 10 else 'C00000'),
    ]
    
    for label, value, color in rates_data:
        ws_summary.cell(row=row, column=1, value=str(label)).font = Font(bold=True)
        cell = ws_summary.cell(row=row, column=2, value=value)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        row += 1
    
    # Adjust summary sheet column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ========== CAMPAIGNS DATA SHEET ==========
    # Define headers
    headers = [
        _('Campaign Name'),
        _('Server'),
        _('Company'),
        _('Status'),
        _('Template'),
        _('Landing Page'),
        _('Sending Profile'),
        _('Target Groups'),
        _('Total Targets'),
        _('Emails Sent'),
        _('Emails Opened'),
        _('Links Clicked'),
        _('Credentials Submitted'),
        _('Data Submitted'),
        _('Launch Date'),
        _('Created Date'),
        _('Last Sync'),
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Write data
    row_idx = 2
    for campaign in campaigns:
        # Get target groups
        groups = campaign.groups.all()
        groups_text = ', '.join([group.name for group in groups]) if groups else ''
        
        # Prepare row data
        row_data = [
            campaign.name,
            campaign.server.name if campaign.server else '',
            campaign.server.company.name if campaign.server and campaign.server.company else '',
            campaign.get_status_display(),
            campaign.template.name if campaign.template else '',
            campaign.landing_page.name if campaign.landing_page else '',
            campaign.sending_profile.name if campaign.sending_profile else '',
            groups_text,
            campaign.total_targets,
            campaign.emails_sent,
            campaign.emails_opened,
            campaign.links_clicked,
            campaign.credentials_submitted,
            campaign.data_submitted,
            campaign.launch_date.strftime('%Y-%m-%d %H:%M') if campaign.launch_date else '',
            campaign.created_at.strftime('%Y-%m-%d %H:%M') if campaign.created_at else '',
            campaign.last_sync.strftime('%Y-%m-%d %H:%M') if campaign.last_sync else '',
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx in [1, 8]))
            
            # Apply status color
            if col_idx == 4:  # Status column
                status_color = status_colors.get(campaign.status, 'FFFFFF')
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF' if campaign.status in ['running', 'completed', 'error'] else '000000')
            
            # Highlight metrics columns
            if col_idx in [9, 10, 11, 12, 13, 14]:  # Numeric columns
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value and value > 0:
                    if col_idx == 13:  # Credentials Submitted - highlight in red
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                        cell.font = Font(bold=True, color='C00000')
                    else:
                        cell.fill = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
        
        row_idx += 1
    
    # Adjust column widths
    column_widths = [30, 20, 20, 15, 25, 25, 25, 30, 12, 12, 12, 12, 18, 12, 18, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Set row height for data rows
    for row in range(2, row_idx):
        ws.row_dimensions[row].height = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gophish_campaigns_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
@gophish_access_required('view_campaigns')
def events_list(request):
    """List all events with filtering and pagination"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    
    # Get base queryset
    events = GophishEvent.objects.select_related(
        'campaign', 'campaign__server', 'campaign__server__company'
    ).order_by('-timestamp')
    
    if accessible_companies is not None:
        events = events.filter(campaign__server__company__in=accessible_companies)
    
    # Filtering
    search = request.GET.get('search', '')
    campaign_id = request.GET.get('campaign', '')
    event_type = request.GET.get('event_type', '')
    server_id = request.GET.get('server', '')
    
    if search:
        events = events.filter(
            Q(target_email__icontains=search) |
            Q(target_name__icontains=search) |
            Q(campaign__name__icontains=search)
        )
    
    if campaign_id:
        events = events.filter(campaign_id=campaign_id)
    
    if event_type:
        events = events.filter(event_type=event_type)
    
    if server_id:
        events = events.filter(campaign__server_id=server_id)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, events)
    
    # Get filter options
    campaigns_qs = GophishCampaign.objects.select_related('server')
    servers_qs = GophishServer.objects.filter(is_active=True)
    
    if accessible_companies is not None:
        campaigns_qs = campaigns_qs.filter(server__company__in=accessible_companies)
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    context = {
        **pagination_context,
        'search': search,
        'campaign_id': campaign_id,
        'event_type': event_type,
        'server_id': server_id,
        'campaigns': campaigns_qs.order_by('-created_at')[:100],
        'servers': servers_qs,
        'event_type_choices': GophishEvent.EVENT_TYPE_CHOICES,
    }
    
    return render(request, 'app_gophish/events_list.html', context)


@login_required
@gophish_access_required('view_campaigns')
def events_export_excel(request):
    """Export events to Excel with styling and colors"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    
    # Get base queryset
    events = GophishEvent.objects.select_related(
        'campaign', 'campaign__server', 'campaign__server__company'
    ).order_by('-timestamp')
    
    if accessible_companies is not None:
        events = events.filter(campaign__server__company__in=accessible_companies)
    
    # Apply same filters as events_list view
    search = request.GET.get('search', '')
    campaign_id = request.GET.get('campaign', '')
    event_type = request.GET.get('event_type', '')
    server_id = request.GET.get('server', '')
    
    if search:
        events = events.filter(
            Q(target_email__icontains=search) |
            Q(target_name__icontains=search) |
            Q(campaign__name__icontains=search)
        )
    
    if campaign_id:
        events = events.filter(campaign_id=campaign_id)
    
    if event_type:
        events = events.filter(event_type=event_type)
    
    if server_id:
        events = events.filter(campaign__server_id=server_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Event type colors
    event_colors = {
        'email_sent': '4472C4',         # Blue
        'email_opened': '70AD47',       # Green
        'link_clicked': 'FFC000',       # Orange
        'credentials_submitted': 'C00000',  # Red
        'data_submitted': 'ED7D31',     # Orange-Red
        'error': 'A5A5A5',              # Gray
    }
    
    # ========== SUMMARY SHEET ==========
    # Calculate statistics
    total_events = events.count()
    event_type_counts = {}
    for event_key, event_label in GophishEvent.EVENT_TYPE_CHOICES:
        event_type_counts[event_label] = events.filter(event_type=event_key).count()
    
    # Count unique targets and campaigns
    unique_targets = events.values('target_email').distinct().count()
    unique_campaigns = events.values('campaign').distinct().count()
    
    # Write summary title
    cell = ws_summary.cell(row=1, column=1, value=str(_('Gophish Events Export Summary')))
    cell.font = Font(bold=True, size=16, color='4472C4')
    ws_summary.merge_cells('A1:D1')
    
    # Export information
    row = 3
    ws_summary.cell(row=row, column=1, value=str(_('Export Date:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Total Events:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=total_events)
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Unique Targets:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=unique_targets)
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Unique Campaigns:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=unique_campaigns)
    
    # Event type breakdown with colors
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Event Type Breakdown'))).font = Font(bold=True, size=12)
    row += 1
    
    for event_label, count in event_type_counts.items():
        ws_summary.cell(row=row, column=1, value=event_label).font = Font(bold=True)
        cell = ws_summary.cell(row=row, column=2, value=count)
        # Find event key for color
        for event_key, label in GophishEvent.EVENT_TYPE_CHOICES:
            if label == event_label:
                color = event_colors.get(event_key, 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                if event_key in ['email_sent', 'email_opened', 'link_clicked', 'credentials_submitted', 'data_submitted']:
                    cell.font = Font(bold=True, color='FFFFFF')
                break
        row += 1
    
    # Calculate event distribution percentages
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Event Distribution'))).font = Font(bold=True, size=12)
    row += 1
    
    for event_label, count in event_type_counts.items():
        if total_events > 0:
            percentage = (count / total_events * 100)
            ws_summary.cell(row=row, column=1, value=event_label).font = Font(bold=True)
            cell = ws_summary.cell(row=row, column=2, value=f'{percentage:.2f}%')
            cell.alignment = Alignment(horizontal='center')
            row += 1
    
    # Adjust summary sheet column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ========== EVENTS DATA SHEET ==========
    # Define headers
    headers = [
        _('Event Type'),
        _('Campaign'),
        _('Server'),
        _('Company'),
        _('Target Email'),
        _('Target Name'),
        _('Timestamp'),
        _('Date'),
        _('Time'),
        _('IP Address'),
        _('User Agent'),
        _('Has Details'),
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Write data
    row_idx = 2
    for event in events:
        # Prepare row data
        row_data = [
            event.get_event_type_display(),
            event.campaign.name if event.campaign else '',
            event.campaign.server.name if event.campaign and event.campaign.server else '',
            event.campaign.server.company.name if event.campaign and event.campaign.server and event.campaign.server.company else '',
            event.target_email,
            event.target_name or '',
            event.timestamp.strftime('%Y-%m-%d %H:%M:%S') if event.timestamp else '',
            event.timestamp.strftime('%Y-%m-%d') if event.timestamp else '',
            event.timestamp.strftime('%H:%M:%S') if event.timestamp else '',
            event.ip_address or '',
            event.user_agent or '',
            'Yes' if event.details else 'No',
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center')
            
            # Apply event type color
            if col_idx == 1:  # Event Type column
                event_color = event_colors.get(event.event_type, 'FFFFFF')
                cell.fill = PatternFill(start_color=event_color, end_color=event_color, fill_type='solid')
                if event.event_type in ['email_sent', 'email_opened', 'link_clicked', 'credentials_submitted', 'data_submitted']:
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.font = Font(bold=True, color='000000')
            
            # Highlight credentials submitted in red
            if event.event_type == 'credentials_submitted' and col_idx > 1:
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            # Center align specific columns
            if col_idx in [7, 8, 9, 12]:  # Timestamp, Date, Time, Has Details
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx += 1
    
    # Adjust column widths
    column_widths = [20, 25, 20, 20, 30, 25, 20, 15, 12, 15, 40, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Set row height for data rows
    for row in range(2, row_idx):
        ws.row_dimensions[row].height = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gophish_events_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
@gophish_access_required('view_campaigns')
def campaign_events_export_excel(request, campaign_id):
    """Export campaign events to Excel with styling and colors"""
    campaign = get_object_or_404(GophishCampaign, id=campaign_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and campaign.server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this campaign"))
    
    # Get events for this campaign
    events = GophishEvent.objects.filter(campaign=campaign).order_by('-timestamp')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Events"
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Event type colors
    event_colors = {
        'email_sent': '4472C4',         # Blue
        'email_opened': '70AD47',       # Green
        'link_clicked': 'FFC000',       # Orange
        'credentials_submitted': 'C00000',  # Red
        'data_submitted': 'ED7D31',     # Orange-Red
        'error': 'A5A5A5',              # Gray
    }
    
    # ========== SUMMARY SHEET ==========
    # Calculate statistics
    total_events = events.count()
    event_type_counts = {}
    for event_key, event_label in GophishEvent.EVENT_TYPE_CHOICES:
        event_type_counts[event_label] = events.filter(event_type=event_key).count()
    
    # Count unique targets
    unique_targets = events.values('target_email').distinct().count()
    
    # Write summary title
    cell = ws_summary.cell(row=1, column=1, value=str(_('Campaign Events Export Summary')))
    cell.font = Font(bold=True, size=16, color='4472C4')
    ws_summary.merge_cells('A1:D1')
    
    # Campaign information
    row = 3
    ws_summary.cell(row=row, column=1, value=str(_('Campaign:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=campaign.name)
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Server:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=campaign.server.name if campaign.server else '')
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Status:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=campaign.get_status_display())
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Export Date:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Total Events:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=total_events)
    
    row += 1
    ws_summary.cell(row=row, column=1, value=str(_('Unique Targets:'))).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=unique_targets)
    
    # Event type breakdown with colors
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Event Type Breakdown'))).font = Font(bold=True, size=12)
    row += 1
    
    for event_label, count in event_type_counts.items():
        ws_summary.cell(row=row, column=1, value=event_label).font = Font(bold=True)
        cell = ws_summary.cell(row=row, column=2, value=count)
        # Find event key for color
        for event_key, label in GophishEvent.EVENT_TYPE_CHOICES:
            if label == event_label:
                color = event_colors.get(event_key, 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                if event_key in ['email_sent', 'email_opened', 'link_clicked', 'credentials_submitted', 'data_submitted']:
                    cell.font = Font(bold=True, color='FFFFFF')
                break
        row += 1
    
    # Calculate event distribution percentages
    row += 2
    ws_summary.cell(row=row, column=1, value=str(_('Event Distribution'))).font = Font(bold=True, size=12)
    row += 1
    
    for event_label, count in event_type_counts.items():
        if total_events > 0:
            percentage = (count / total_events * 100)
            ws_summary.cell(row=row, column=1, value=event_label).font = Font(bold=True)
            cell = ws_summary.cell(row=row, column=2, value=f'{percentage:.2f}%')
            cell.alignment = Alignment(horizontal='center')
            row += 1
    
    # Adjust summary sheet column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ========== EVENTS DATA SHEET ==========
    # Define headers
    headers = [
        _('Event Type'),
        _('Target Email'),
        _('Target Name'),
        _('Timestamp'),
        _('Date'),
        _('Time'),
        _('IP Address'),
        _('User Agent'),
        _('Has Details'),
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Write data
    row_idx = 2
    for event in events:
        # Prepare row data
        row_data = [
            event.get_event_type_display(),
            event.target_email,
            event.target_name or '',
            event.timestamp.strftime('%Y-%m-%d %H:%M:%S') if event.timestamp else '',
            event.timestamp.strftime('%Y-%m-%d') if event.timestamp else '',
            event.timestamp.strftime('%H:%M:%S') if event.timestamp else '',
            event.ip_address or '',
            event.user_agent or '',
            'Yes' if event.details else 'No',
        ]
        
        # Write row
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center')
            
            # Apply event type color
            if col_idx == 1:  # Event Type column
                event_color = event_colors.get(event.event_type, 'FFFFFF')
                cell.fill = PatternFill(start_color=event_color, end_color=event_color, fill_type='solid')
                if event.event_type in ['email_sent', 'email_opened', 'link_clicked', 'credentials_submitted', 'data_submitted']:
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.font = Font(bold=True, color='000000')
            
            # Highlight credentials submitted in red
            if event.event_type == 'credentials_submitted' and col_idx > 1:
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            
            # Center align specific columns
            if col_idx in [4, 5, 6, 9]:  # Timestamp, Date, Time, Has Details
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx += 1
    
    # Adjust column widths
    column_widths = [20, 30, 25, 20, 15, 12, 15, 40, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Set row height for data rows
    for row in range(2, row_idx):
        ws.row_dimensions[row].height = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="campaign_{campaign.id}_events_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
@gophish_access_required('view_campaigns')
def campaign_create(request):
    """Create a new campaign"""
    if request.method == 'POST':
        form = GophishCampaignForm(request.POST, user=request.user)
        if form.is_valid():
            campaign = form.save()
            messages.success(request, _('Campaign created successfully'))
            return redirect('app_gophish:campaign_detail', campaign_id=campaign.id)
    else:
        form = GophishCampaignForm(user=request.user)
    
    return render(request, 'app_gophish/campaign_form.html', {
        'form': form,
        'title': _('Create Campaign')
    })


@login_required
@gophish_access_required('view_campaigns')
def campaign_detail(request, campaign_id):
    """View campaign details and results"""
    campaign = get_object_or_404(GophishCampaign, id=campaign_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and campaign.server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this campaign"))
    
    # Get statistics
    stats = {
        'total_targets': campaign.total_targets,
        'emails_sent': campaign.emails_sent,
        'emails_opened': campaign.emails_opened,
        'links_clicked': campaign.links_clicked,
        'credentials_submitted': campaign.credentials_submitted,
        'data_submitted': campaign.data_submitted,
    }
    
    context = {
        'campaign': campaign,
        'stats': stats,
    }
    
    return render(request, 'app_gophish/campaign_detail.html', context)


@login_required
@gophish_access_required('view_campaigns')
def campaign_edit(request, campaign_id):
    """Edit a campaign"""
    campaign = get_object_or_404(GophishCampaign, id=campaign_id)
    
    if request.method == 'POST':
        form = GophishCampaignForm(request.POST, instance=campaign, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, _('Campaign updated successfully'))
            return redirect('app_gophish:campaign_detail', campaign_id=campaign.id)
    else:
        form = GophishCampaignForm(instance=campaign, user=request.user)
    
    return render(request, 'app_gophish/campaign_form.html', {
        'form': form,
        'title': _('Edit Campaign'),
        'campaign': campaign
    })


@login_required
@gophish_access_required('view_campaigns')
def campaign_launch(request, campaign_id):
    """Launch a campaign"""
    campaign = get_object_or_404(GophishCampaign, id=campaign_id)
    
    if campaign.status not in ['draft']:
        messages.error(request, _('Only draft campaigns can be launched'))
        return redirect('app_gophish:campaign_detail', campaign_id=campaign.id)
    
    if request.method == 'POST':
        form = CampaignLaunchForm(campaign, request.POST)
        if form.is_valid():
            try:
                # Get API client
                client = gophish_manager.get_client(campaign.server)
                
                # Create campaign in Gophish if not exists
                if not campaign.gophish_id:
                    campaign_data = {
                        'name': campaign.name,
                        'template': {'id': campaign.template.gophish_id},
                        'page': {'id': campaign.landing_page.gophish_id},
                        'smtp': {'id': campaign.sending_profile.gophish_id},
                        'groups': [{'id': group.gophish_id} for group in campaign.groups.all()],
                    }
                    
                    if campaign.url:
                        campaign_data['url'] = campaign.url
                    
                    result = client.create_campaign(campaign_data)
                    campaign.gophish_id = result['id']
                    campaign.status = 'running'
                    campaign.save()
                
                # Launch the campaign
                launch_date = form.cleaned_data.get('launch_date')
                if launch_date:
                    campaign.launch_date = launch_date
                    campaign.save()
                
                client.launch_campaign(campaign.gophish_id)
                campaign.status = 'running'
                campaign.save()
                
                messages.success(request, _('Campaign launched successfully'))
                
            except GophishAPIError as e:
                messages.error(request, _('Failed to launch campaign: {}').format(str(e)))
                logger.error(f"Failed to launch campaign {campaign.id}: {str(e)}")
            
            return redirect('app_gophish:campaign_detail', campaign_id=campaign.id)
    else:
        form = CampaignLaunchForm(campaign)
    
    return render(request, 'app_gophish/campaign_launch.html', {
        'form': form,
        'campaign': campaign
    })


@login_required
@gophish_access_required('view_landing_pages')
def landing_page_list(request):
    """List all landing pages"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    pages = GophishLandingPage.objects.select_related('server').order_by('name')
    if accessible_companies is not None:
        pages = pages.filter(server__company__in=accessible_companies)
    
    # Filtering
    search = request.GET.get('search', '')
    server_id = request.GET.get('server', '')
    
    if search:
        pages = pages.filter(name__icontains=search)
    
    if server_id:
        pages = pages.filter(server_id=server_id)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, pages)
    
    # Filter servers for context based on accessible companies
    servers_qs = GophishServer.objects.filter(is_active=True)
    if accessible_companies is not None:
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    context = {
        **pagination_context,
        'search': search,
        'server_id': server_id,
        'servers': servers_qs,
    }
    
    return render(request, 'app_gophish/landing_page_list.html', context)


@login_required
@gophish_access_required('view_landing_pages')
def landing_page_preview(request, page_id):
    """Preview a landing page"""
    page = get_object_or_404(GophishLandingPage, id=page_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and page.server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this landing page"))
    
    context = {
        'page': page,
    }
    
    return render(request, 'app_gophish/landing_page_preview.html', context)


@login_required
@gophish_access_required('view_templates')
def email_template_list(request):
    """List all email templates"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    templates = GophishTemplate.objects.select_related('server').order_by('name')
    if accessible_companies is not None:
        templates = templates.filter(server__company__in=accessible_companies)
    
    # Filtering
    search = request.GET.get('search', '')
    server_id = request.GET.get('server', '')
    
    if search:
        templates = templates.filter(
            Q(name__icontains=search) |
            Q(subject__icontains=search)
        )
    
    if server_id:
        templates = templates.filter(server_id=server_id)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, templates)
    
    # Filter servers for context based on accessible companies
    servers_qs = GophishServer.objects.filter(is_active=True)
    if accessible_companies is not None:
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    context = {
        **pagination_context,
        'search': search,
        'server_id': server_id,
        'servers': servers_qs,
    }
    
    return render(request, 'app_gophish/email_template_list.html', context)


@login_required
@gophish_access_required('view_templates')
def email_template_detail(request, template_id):
    """View email template details"""
    template = get_object_or_404(GophishTemplate, id=template_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and template.server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this template"))
    
    # Get campaigns using this template
    campaigns = template.campaigns.select_related('server').order_by('-created_at')[:10]
    
    context = {
        'template': template,
        'campaigns': campaigns,
    }
    
    return render(request, 'app_gophish/email_template_detail.html', context)


@login_required
@gophish_access_required('view_sending_profiles')
def sending_profile_list(request):
    """List all sending profiles"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    profiles = GophishSendingProfile.objects.select_related('server').order_by('name')
    if accessible_companies is not None:
        profiles = profiles.filter(server__company__in=accessible_companies)
    
    # Filtering
    search = request.GET.get('search', '')
    server_id = request.GET.get('server', '')
    
    if search:
        profiles = profiles.filter(name__icontains=search)
    
    if server_id:
        profiles = profiles.filter(server_id=server_id)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, profiles)
    
    # Filter servers for context based on accessible companies
    servers_qs = GophishServer.objects.filter(is_active=True)
    if accessible_companies is not None:
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    context = {
        **pagination_context,
        'search': search,
        'server_id': server_id,
        'servers': servers_qs,
    }
    
    return render(request, 'app_gophish/sending_profile_list.html', context)




@login_required
@gophish_access_required('view_groups')
def group_list(request):
    """List all groups"""
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    groups = GophishGroup.objects.select_related('server').order_by('name')
    if accessible_companies is not None:
        groups = groups.filter(server__company__in=accessible_companies)
    
    # Filtering
    search = request.GET.get('search', '')
    server_id = request.GET.get('server', '')
    
    if search:
        groups = groups.filter(name__icontains=search)
    
    if server_id:
        groups = groups.filter(server_id=server_id)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, groups)
    
    # Filter servers for context based on accessible companies
    servers_qs = GophishServer.objects.filter(is_active=True)
    if accessible_companies is not None:
        servers_qs = servers_qs.filter(company__in=accessible_companies)
    
    context = {
        **pagination_context,
        'search': search,
        'server_id': server_id,
        'servers': servers_qs,
    }
    
    return render(request, 'app_gophish/group_list.html', context)




@login_required
@gophish_access_required('sync')
def sync_data(request):
    """Manual synchronization with Gophish servers"""
    if request.method == 'POST':
        form = SyncForm(request.POST, user=request.user)
        if form.is_valid():
            server = form.cleaned_data['server']
            sync_type = form.cleaned_data['sync_type']
            force_update = form.cleaned_data['force_update']
            
            # Normal sync: pull data from Gophish
            try:
                # Check if Celery is available
                from celery import current_app
                if current_app.control.inspect().stats():
                    # Celery is running, use async task
                    task = sync_gophish_data.delay(server.id, sync_type, force_update)
                    messages.success(request, _('Synchronization started. Check sync logs for progress.'))
                else:
                    # Celery not running, use direct sync
                    result = sync_gophish_data_direct(server.id, sync_type, force_update)
                    if result['status'] == 'completed':
                        messages.success(request, _('Synchronization completed successfully.'))
                    else:
                        messages.error(request, _('Synchronization failed: {}').format(result.get('message', 'Unknown error')))
            except Exception as e:
                # Fallback to direct sync
                result = sync_gophish_data_direct(server.id, sync_type, force_update)
                if result['status'] == 'completed':
                    messages.success(request, _('Synchronization completed successfully.'))
                else:
                    messages.error(request, _('Synchronization failed: {}').format(result.get('message', 'Unknown error')))
            
            return redirect('app_gophish:sync_logs')
    else:
        form = SyncForm(user=request.user)
    
    return render(request, 'app_gophish/sync_form.html', {'form': form})


@login_required
@gophish_access_required('sync')
def sync_logs(request):
    """View synchronization logs"""
    # Filter logs by accessible companies
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    logs = GophishSyncLog.objects.select_related('server').order_by('-started_at')
    if accessible_companies is not None:
        logs = logs.filter(server__company__in=accessible_companies)
    
    page_obj, pagination_context = paginate_gophish_queryset(request, logs)
    
    return render(request, 'app_gophish/sync_logs.html', pagination_context)


@login_required
@gophish_access_required('sync')
def sync_log_detail(request, log_id):
    """View detailed sync log"""
    log = get_object_or_404(GophishSyncLog, id=log_id)
    
    # Check company access
    accessible_companies = get_user_accessible_companies_gophish(request.user)
    if accessible_companies is not None and log.server.company not in accessible_companies:
        raise PermissionDenied(_("You don't have access to this sync log"))
    
    return render(request, 'app_gophish/sync_log_detail.html', {'log': log})


@login_required
@gophish_access_required('view_campaigns')
@require_http_methods(["GET"])
def campaign_events_api(request, campaign_id):
    """API endpoint for campaign events"""
    campaign = get_object_or_404(GophishCampaign, id=campaign_id)
    events = campaign.events.all().order_by('-timestamp')
    
    # Filter by event type if provided
    event_type = request.GET.get('type')
    if event_type:
        events = events.filter(event_type=event_type)
    
    # Limit results
    limit = int(request.GET.get('limit', 100))
    events = events[:limit]
    
    data = []
    for event in events:
        data.append({
            'id': event.id,
            'event_type': event.event_type,
            'target_email': event.target_email,
            'target_name': event.target_name,
            'timestamp': event.timestamp.isoformat(),
            'ip_address': event.ip_address,
            'user_agent': event.user_agent,
            'details': event.details,
        })
    
    return JsonResponse({'events': data})


@csrf_exempt
def webhook_handler(request):
    """Handle webhooks from Gophish"""
    if request.method != 'POST':
        return HttpResponse(status=405)
    
    try:
        data = json.loads(request.body)
        logger.info(f"Received Gophish webhook: {data}")
        
        # Process webhook data
        # This would need to be implemented based on your webhook configuration
        
        return JsonResponse({'status': 'success'})
    
    except json.JSONDecodeError:
        logger.error("Invalid JSON in webhook request")
        return HttpResponse(status=400)
    except Exception as e:
        logger.error(f"Error processing webhook: {str(e)}")
        return HttpResponse(status=500)


@login_required
@gophish_access_required('manage_servers')
@require_http_methods(["POST"])
def test_server_connection(request, server_id):
    """Test connection to a specific server and return detailed results"""
    server = get_object_or_404(GophishServer, id=server_id)
    
    try:
        connected = gophish_manager.test_server_connection(server)
        if connected:
            return JsonResponse({
                'connected': True,
                'message': _('Connection successful'),
                'server_name': server.name,
                'server_url': server.base_url
            })
        else:
            return JsonResponse({
                'connected': False,
                'message': _('Connection failed'),
                'server_name': server.name,
                'server_url': server.base_url
            })
    except Exception as e:
        logger.error(f"Error testing connection to server {server_id}: {str(e)}")
        return JsonResponse({
            'connected': False,
            'message': _('Connection test error: {}').format(str(e)),
            'server_name': server.name,
            'server_url': server.base_url
        })


@login_required
@gophish_access_required('manage_servers')
@require_http_methods(["POST"])
def diagnose_server(request, server_id):
    """Diagnose server connectivity and API availability"""
    server = get_object_or_404(GophishServer, id=server_id)
    
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    diagnosis = {
        'server_name': server.name,
        'server_url': server.base_url,
        'tests': []
    }
    
    # Test 1: Basic connectivity
    try:
        response = requests.get(
            f"{server.base_url}/",
            verify=False,
            timeout=10
        )
        diagnosis['tests'].append({
            'test': 'Basic Connectivity',
            'status': 'success',
            'message': f'Server reachable (HTTP {response.status_code})',
            'details': f'Response size: {len(response.content)} bytes'
        })
    except Exception as e:
        diagnosis['tests'].append({
            'test': 'Basic Connectivity',
            'status': 'error',
            'message': f'Server not reachable: {str(e)}',
            'details': 'Check if server is running and accessible'
        })
        return JsonResponse(diagnosis)
    
    # Test 2: Check if it's a Gophish server
    try:
        response = requests.get(
            f"{server.base_url}/",
            verify=False,
            timeout=10
        )
        content = response.text.lower()
        if 'gophish' in content:
            diagnosis['tests'].append({
                'test': 'Gophish Detection',
                'status': 'success',
                'message': 'Gophish server detected',
                'details': 'Server appears to be running Gophish'
            })
        else:
            diagnosis['tests'].append({
                'test': 'Gophish Detection',
                'status': 'warning',
                'message': 'Gophish not detected in response',
                'details': 'Server might not be Gophish or API might be disabled'
            })
    except Exception as e:
        diagnosis['tests'].append({
            'test': 'Gophish Detection',
            'status': 'error',
            'message': f'Error checking Gophish: {str(e)}',
            'details': 'Could not analyze server response'
        })
    
    # Test 3: API endpoints
    api_endpoints = [
        '/api/campaigns',
        '/api/groups',
        '/api/templates',
        '/api/pages',
        '/api/smtp',
        '/api/version',
        '/api/users'
    ]
    
    api_results = []
    for endpoint in api_endpoints:
        try:
            response = requests.get(
                f"{server.base_url}{endpoint}",
                verify=False,
                timeout=10
            )
            api_results.append({
                'endpoint': endpoint,
                'status_code': response.status_code,
                'accessible': response.status_code in [200, 401]  # 401 means endpoint exists but needs auth
            })
        except Exception as e:
            api_results.append({
                'endpoint': endpoint,
                'status_code': 'error',
                'accessible': False,
                'error': str(e)
            })
    
    accessible_endpoints = [r for r in api_results if r['accessible']]
    
    if accessible_endpoints:
        diagnosis['tests'].append({
            'test': 'API Endpoints',
            'status': 'success',
            'message': f'{len(accessible_endpoints)} API endpoints accessible',
            'details': f'Accessible: {[r["endpoint"] for r in accessible_endpoints]}'
        })
    else:
        diagnosis['tests'].append({
            'test': 'API Endpoints',
            'status': 'error',
            'message': 'No API endpoints accessible',
            'details': 'API might be disabled or endpoints are different'
        })
    
    # Test 4: Authentication
    if accessible_endpoints:
        try:
            # Try with the API key
            headers = {'Authorization': f'Bearer {server.api_key}'}
            response = requests.get(
                f"{server.base_url}{accessible_endpoints[0]['endpoint']}",
                headers=headers,
                verify=False,
                timeout=10
            )
            
            if response.status_code == 200:
                diagnosis['tests'].append({
                    'test': 'API Authentication',
                    'status': 'success',
                    'message': 'API key authentication successful',
                    'details': 'Server is ready for integration'
                })
            elif response.status_code == 401:
                diagnosis['tests'].append({
                    'test': 'API Authentication',
                    'status': 'error',
                    'message': 'API key authentication failed',
                    'details': 'Check API key in Gophish Settings > Account Settings'
                })
            else:
                diagnosis['tests'].append({
                    'test': 'API Authentication',
                    'status': 'warning',
                    'message': f'Unexpected response: HTTP {response.status_code}',
                    'details': 'API might be configured differently'
                })
        except Exception as e:
            diagnosis['tests'].append({
                'test': 'API Authentication',
                'status': 'error',
                'message': f'Authentication test failed: {str(e)}',
                'details': 'Could not test API key'
            })
    
    return JsonResponse(diagnosis)


@login_required
@gophish_access_required()
@require_http_methods(["POST"])
def ajax_load_server_data(request):
    """AJAX endpoint to load server-specific data"""
    server_id = request.POST.get('server_id')
    data_type = request.POST.get('data_type')
    
    if not server_id or not data_type:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    try:
        server = get_object_or_404(GophishServer, id=server_id)
        
        if data_type == 'templates':
            data = server.templates.all().values('id', 'name')
        elif data_type == 'landing_pages':
            data = server.landing_pages.all().values('id', 'name')
        elif data_type == 'sending_profiles':
            data = server.sending_profiles.all().values('id', 'name')
        elif data_type == 'groups':
            data = server.groups.all().values('id', 'name')
        elif data_type == 'connection_test':
            # Test server connection
            try:
                connected = gophish_manager.test_server_connection(server)
                return JsonResponse({'connected': connected})
            except Exception as e:
                logger.error(f"Connection test failed for server {server_id}: {str(e)}")
                return JsonResponse({'connected': False, 'error': str(e)})
        else:
            return JsonResponse({'error': 'Invalid data type'}, status=400)
        
        return JsonResponse({'data': list(data)})
    
    except Exception as e:
        logger.error(f"Error loading server data: {str(e)}")
        return JsonResponse({'error': 'Internal error'}, status=500)
