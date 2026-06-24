# SecBoard\SecBoard\app_gdpr\reports.py

from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext as _
from io import BytesIO
import json

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_gdpr_compliance_to_excel(report_data, companies=None):
    """
    Експорт GDPR Compliance звіту в Excel з підтримкою фільтрації по компаніях
    
    Args:
        report_data: дані звіту згенеровані generate_compliance_report_data
        companies: список компаній для фільтрації
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Install it with: pip install openpyxl")
    
    from .models import (
        DataSubject,
        ConsentRecord,
        DataSubjectRequest,
        DataBreachIncident,
        DataProcessingActivity,
        DPIAAssessment
    )
    
    # Створюємо workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Стилі
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Лист 1: Загальна інформація
    ws_summary = wb.create_sheet("Summary")
    
    # Заголовок звіту
    ws_summary.cell(row=1, column=1, value="GDPR Compliance Report").font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:B1')
    
    ws_summary.cell(row=2, column=1, value="Report Date:").font = Font(bold=True)
    ws_summary.cell(row=2, column=2, value=timezone.now().strftime("%Y-%m-%d %H:%M"))
    
    ws_summary.cell(row=3, column=1, value="Company Filter:").font = Font(bold=True)
    if companies:
        company_names = ", ".join([c.name for c in companies])
        ws_summary.cell(row=3, column=2, value=company_names)
    else:
        ws_summary.cell(row=3, column=2, value="All Companies")
    
    # Статистика по Data Subjects
    row = 5
    ws_summary.cell(row=row, column=1, value="Data Subjects").font = subheader_font
    ws_summary.cell(row=row, column=1).fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    summary_data = [
        ["Total Data Subjects", report_data['data_subjects']['total']],
        ["With Active Consent", report_data['data_subjects']['with_active_consent']],
        ["Anonymized", report_data['data_subjects']['anonymized']],
    ]
    
    for label, value in summary_data:
        ws_summary.cell(row=row, column=1, value=label).border = border
        ws_summary.cell(row=row, column=2, value=value).border = border
        row += 1
    
    # Статистика по Consents
    row += 1
    ws_summary.cell(row=row, column=1, value="Consents").font = subheader_font
    ws_summary.cell(row=row, column=1).fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    consent_data = [
        ["Total Consents", report_data['consents']['total']],
        ["Active", report_data['consents']['active']],
        ["Withdrawn", report_data['consents']['withdrawn']],
    ]
    
    for label, value in consent_data:
        ws_summary.cell(row=row, column=1, value=label).border = border
        ws_summary.cell(row=row, column=2, value=value).border = border
        row += 1
    
    # Статистика по DSR
    row += 1
    ws_summary.cell(row=row, column=1, value="Data Subject Requests").font = subheader_font
    ws_summary.cell(row=row, column=1).fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    dsr_data = [
        ["Total DSR", report_data['dsr']['total']],
        ["Pending", report_data['dsr']['pending']],
        ["Completed", report_data['dsr']['completed']],
        ["Overdue", report_data['dsr']['overdue']],
    ]
    
    for label, value in dsr_data:
        ws_summary.cell(row=row, column=1, value=label).border = border
        cell_value = ws_summary.cell(row=row, column=2, value=value)
        cell_value.border = border
        if label == "Overdue" and value > 0:
            cell_value.font = Font(color="FF0000", bold=True)
        row += 1
    
    # Статистика по Data Breaches
    row += 1
    ws_summary.cell(row=row, column=1, value="Data Breaches").font = subheader_font
    ws_summary.cell(row=row, column=1).fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    breach_data = [
        ["Total Breaches", report_data['breaches']['total']],
        ["Critical Severity", report_data['breaches']['by_severity']['critical']],
        ["High Severity", report_data['breaches']['by_severity']['high']],
        ["Medium Severity", report_data['breaches']['by_severity']['medium']],
        ["Low Severity", report_data['breaches']['by_severity']['low']],
        ["Reported to Authority", report_data['breaches']['reported_to_authority']],
        ["Notification Overdue", report_data['breaches']['notification_overdue']],
    ]
    
    for label, value in breach_data:
        ws_summary.cell(row=row, column=1, value=label).border = border
        cell_value = ws_summary.cell(row=row, column=2, value=value)
        cell_value.border = border
        if "Critical" in label and value > 0:
            cell_value.font = Font(color="FF0000", bold=True)
        row += 1
    
    # Статистика по Processing Activities
    row += 1
    ws_summary.cell(row=row, column=1, value="Processing Activities").font = subheader_font
    ws_summary.cell(row=row, column=1).fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    activity_data = [
        ["Total Activities", report_data['processing_activities']['total']],
        ["Active", report_data['processing_activities']['active']],
        ["With International Transfers", report_data['processing_activities']['with_international_transfers']],
    ]
    
    for label, value in activity_data:
        ws_summary.cell(row=row, column=1, value=label).border = border
        ws_summary.cell(row=row, column=2, value=value).border = border
        row += 1
    
    # Статистика по DPIA
    row += 1
    ws_summary.cell(row=row, column=1, value="DPIA Assessments").font = subheader_font
    ws_summary.cell(row=row, column=1).fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    row += 1
    
    dpia_data = [
        ["Total DPIAs", report_data['dpias']['total']],
        ["Approved", report_data['dpias']['approved']],
        ["In Review", report_data['dpias']['in_review']],
    ]
    
    for label, value in dpia_data:
        ws_summary.cell(row=row, column=1, value=label).border = border
        ws_summary.cell(row=row, column=2, value=value).border = border
        row += 1
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Лист 2: Детальні дані по Data Subjects
    ws_subjects = wb.create_sheet("Data Subjects")
    
    headers = ["ID", "Name", "Email", "Phone", "Company", "Consent Status", "Created Date", "Anonymized"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_subjects.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Фільтруємо по компаніях
    subjects_filter = {}
    if companies:
        subjects_filter['company__in'] = companies
    
    subjects = DataSubject.objects.filter(**subjects_filter).select_related('company').order_by('-created_date')[:1000]
    
    row_idx = 2
    for subject in subjects:
        ws_subjects.cell(row=row_idx, column=1, value=subject.id).border = border
        ws_subjects.cell(row=row_idx, column=2, value=f"{subject.first_name} {subject.last_name}").border = border
        ws_subjects.cell(row=row_idx, column=3, value=subject.email).border = border
        ws_subjects.cell(row=row_idx, column=4, value=subject.phone or "-").border = border
        ws_subjects.cell(row=row_idx, column=5, value=subject.company.name if subject.company else "-").border = border
        ws_subjects.cell(row=row_idx, column=6, value=subject.get_consent_status_display()).border = border
        ws_subjects.cell(row=row_idx, column=7, value=subject.created_date.strftime("%Y-%m-%d")).border = border
        ws_subjects.cell(row=row_idx, column=8, value="Yes" if subject.is_anonymized else "No").border = border
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_subjects.column_dimensions[col].width = 20
    ws_subjects.column_dimensions['B'].width = 30
    ws_subjects.column_dimensions['C'].width = 30
    
    # Лист 3: Data Subject Requests
    ws_dsr = wb.create_sheet("DSR Requests")
    
    dsr_headers = ["Request #", "Type", "Data Subject", "Company", "Status", "Request Date", "Due Date", "Completed Date"]
    for col_idx, header in enumerate(dsr_headers, start=1):
        cell = ws_dsr.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    dsr_filter = {}
    if companies:
        dsr_filter['company__in'] = companies
    
    dsr_requests = DataSubjectRequest.objects.filter(**dsr_filter).select_related(
        'data_subject', 'company'
    ).order_by('-request_date')[:1000]
    
    row_idx = 2
    for dsr in dsr_requests:
        ws_dsr.cell(row=row_idx, column=1, value=dsr.request_number).border = border
        ws_dsr.cell(row=row_idx, column=2, value=dsr.get_request_type_display()).border = border
        ws_dsr.cell(row=row_idx, column=3, value=str(dsr.data_subject) if dsr.data_subject else "-").border = border
        ws_dsr.cell(row=row_idx, column=4, value=dsr.company.name if dsr.company else "-").border = border
        
        status_cell = ws_dsr.cell(row=row_idx, column=5, value=dsr.get_status_display())
        status_cell.border = border
        if dsr.status == 'pending' and dsr.is_overdue():
            status_cell.font = Font(color="FF0000", bold=True)
        
        ws_dsr.cell(row=row_idx, column=6, value=dsr.request_date.strftime("%Y-%m-%d")).border = border
        ws_dsr.cell(row=row_idx, column=7, value=dsr.due_date.strftime("%Y-%m-%d") if dsr.due_date else "-").border = border
        ws_dsr.cell(row=row_idx, column=8, value=dsr.completion_date.strftime("%Y-%m-%d") if dsr.completion_date else "-").border = border
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_dsr.column_dimensions[col].width = 20
    ws_dsr.column_dimensions['A'].width = 25
    ws_dsr.column_dimensions['C'].width = 30
    
    # Лист 4: Data Breaches
    ws_breaches = wb.create_sheet("Data Breaches")
    
    breach_headers = ["Incident #", "Company", "Severity", "Description", "Detected Date", "Reported to Authority", "Status"]
    for col_idx, header in enumerate(breach_headers, start=1):
        cell = ws_breaches.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    breach_filter = {}
    if companies:
        breach_filter['company__in'] = companies
    
    breaches = DataBreachIncident.objects.filter(**breach_filter).select_related('company').order_by('-detected_date')[:1000]
    
    row_idx = 2
    for breach in breaches:
        ws_breaches.cell(row=row_idx, column=1, value=breach.incident_number).border = border
        ws_breaches.cell(row=row_idx, column=2, value=breach.company.name if breach.company else "-").border = border
        
        severity_cell = ws_breaches.cell(row=row_idx, column=3, value=breach.get_severity_display())
        severity_cell.border = border
        if breach.severity == 'critical':
            severity_cell.font = Font(color="FF0000", bold=True)
        elif breach.severity == 'high':
            severity_cell.font = Font(color="FF6600", bold=True)
        
        ws_breaches.cell(row=row_idx, column=4, value=breach.description[:100] + "..." if len(breach.description) > 100 else breach.description).border = border
        ws_breaches.cell(row=row_idx, column=5, value=breach.detected_date.strftime("%Y-%m-%d %H:%M")).border = border
        ws_breaches.cell(row=row_idx, column=6, value="Yes" if breach.reported_to_authority else "No").border = border
        ws_breaches.cell(row=row_idx, column=7, value=breach.get_status_display()).border = border
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_breaches.column_dimensions[col].width = 20
    ws_breaches.column_dimensions['D'].width = 50
    
    # Лист 5: Processing Activities
    ws_activities = wb.create_sheet("Processing Activities")
    
    activity_headers = ["Name", "Company", "Purpose", "Data Categories", "Legal Basis", "International Transfer", "Active"]
    for col_idx, header in enumerate(activity_headers, start=1):
        cell = ws_activities.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    activity_filter = {}
    if companies:
        activity_filter['company__in'] = companies
    
    activities = DataProcessingActivity.objects.filter(**activity_filter).select_related('company')[:1000]
    
    row_idx = 2
    for activity in activities:
        ws_activities.cell(row=row_idx, column=1, value=activity.name).border = border
        ws_activities.cell(row=row_idx, column=2, value=activity.company.name if activity.company else "-").border = border
        ws_activities.cell(row=row_idx, column=3, value=activity.purpose[:50] + "..." if len(activity.purpose) > 50 else activity.purpose).border = border
        ws_activities.cell(row=row_idx, column=4, value=activity.data_categories or "-").border = border
        ws_activities.cell(row=row_idx, column=5, value=activity.get_legal_basis_display()).border = border
        ws_activities.cell(row=row_idx, column=6, value="Yes" if activity.international_transfers else "No").border = border
        ws_activities.cell(row=row_idx, column=7, value="Yes" if activity.is_active else "No").border = border
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_activities.column_dimensions[col].width = 25
    ws_activities.column_dimensions['C'].width = 40
    
    # Зберігаємо в пам'ять
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_gdpr_compliance_to_pdf(report_data, companies=None):
    """
    Експорт GDPR Compliance звіту в PDF з підтримкою фільтрації по компаніях
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed. Install it with: pip install reportlab")
    
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Заголовок
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
    )
    
    title = Paragraph("GDPR Compliance Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Інформація про звіт
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
    )
    
    report_date = Paragraph(f"<b>Report Date:</b> {timezone.now().strftime('%Y-%m-%d %H:%M')}", info_style)
    elements.append(report_date)
    
    if companies:
        company_names = ", ".join([c.name for c in companies])
        company_filter = Paragraph(f"<b>Company Filter:</b> {company_names}", info_style)
    else:
        company_filter = Paragraph("<b>Company Filter:</b> All Companies", info_style)
    elements.append(company_filter)
    elements.append(Spacer(1, 20))
    
    heading_style = styles['Heading2']
    
    # Data Subjects Summary
    elements.append(Paragraph("Data Subjects", heading_style))
    elements.append(Spacer(1, 12))
    
    subjects_data = [
        ["Metric", "Count"],
        ["Total Data Subjects", str(report_data['data_subjects']['total'])],
        ["With Active Consent", str(report_data['data_subjects']['with_active_consent'])],
        ["Anonymized", str(report_data['data_subjects']['anonymized'])],
    ]
    
    subjects_table = Table(subjects_data, colWidths=[4*inch, 2*inch])
    subjects_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(subjects_table)
    elements.append(Spacer(1, 20))
    
    # Consents Summary
    elements.append(Paragraph("Consents", heading_style))
    elements.append(Spacer(1, 12))
    
    consents_data = [
        ["Metric", "Count"],
        ["Total Consents", str(report_data['consents']['total'])],
        ["Active", str(report_data['consents']['active'])],
        ["Withdrawn", str(report_data['consents']['withdrawn'])],
    ]
    
    consents_table = Table(consents_data, colWidths=[4*inch, 2*inch])
    consents_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(consents_table)
    elements.append(Spacer(1, 20))
    
    # DSR Summary
    elements.append(Paragraph("Data Subject Requests", heading_style))
    elements.append(Spacer(1, 12))
    
    dsr_data = [
        ["Metric", "Count"],
        ["Total DSR", str(report_data['dsr']['total'])],
        ["Pending", str(report_data['dsr']['pending'])],
        ["Completed", str(report_data['dsr']['completed'])],
        ["Overdue", str(report_data['dsr']['overdue'])],
    ]
    
    dsr_table = Table(dsr_data, colWidths=[4*inch, 2*inch])
    dsr_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(dsr_table)
    elements.append(PageBreak())
    
    # Data Breaches Summary
    elements.append(Paragraph("Data Breaches", heading_style))
    elements.append(Spacer(1, 12))
    
    breach_data = [
        ["Metric", "Count"],
        ["Total Breaches", str(report_data['breaches']['total'])],
        ["Critical Severity", str(report_data['breaches']['by_severity']['critical'])],
        ["High Severity", str(report_data['breaches']['by_severity']['high'])],
        ["Medium Severity", str(report_data['breaches']['by_severity']['medium'])],
        ["Low Severity", str(report_data['breaches']['by_severity']['low'])],
        ["Reported to Authority", str(report_data['breaches']['reported_to_authority'])],
        ["Notification Overdue", str(report_data['breaches']['notification_overdue'])],
    ]
    
    breach_table = Table(breach_data, colWidths=[4*inch, 2*inch])
    breach_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(breach_table)
    elements.append(Spacer(1, 20))
    
    # Processing Activities Summary
    elements.append(Paragraph("Processing Activities", heading_style))
    elements.append(Spacer(1, 12))
    
    activity_data = [
        ["Metric", "Count"],
        ["Total Activities", str(report_data['processing_activities']['total'])],
        ["Active", str(report_data['processing_activities']['active'])],
        ["With International Transfers", str(report_data['processing_activities']['with_international_transfers'])],
    ]
    
    activity_table = Table(activity_data, colWidths=[4*inch, 2*inch])
    activity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(activity_table)
    elements.append(Spacer(1, 20))
    
    # DPIA Summary
    elements.append(Paragraph("DPIA Assessments", heading_style))
    elements.append(Spacer(1, 12))
    
    dpia_data = [
        ["Metric", "Count"],
        ["Total DPIAs", str(report_data['dpias']['total'])],
        ["Approved", str(report_data['dpias']['approved'])],
        ["In Review", str(report_data['dpias']['in_review'])],
    ]
    
    dpia_table = Table(dpia_data, colWidths=[4*inch, 2*inch])
    dpia_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(dpia_table)
    
    # Генеруємо PDF
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

