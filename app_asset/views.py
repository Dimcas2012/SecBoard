#  SecBoard\SecBoard\app_asset\views.py
import openpyxl
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.http import require_POST, require_http_methods
from openpyxl.styles import Font
from .models import InformationAsset, AssetAdministrator, AssetGroup, CriticalityLevel, \
    AssetOwner, AssetType, AccessAssets, AssetHistory, AssetGuide, AssetGuideTranslation, InformationAssetSoftwareSelection, \
    SoftwareRegister, SoftwareRegisterFile, SoftwareRegisterHistory, SoftwareStatus, SoftwareLicenseType, \
    ExternalMediaRegister, ExternalMediaRegisterFile, ExternalMediaRegisterHistory, ExternalMediaStatus, \
    SoftwareGuide, SoftwareGuideTranslation, ExternalMediaGuide, ExternalMediaGuideTranslation
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.models import User
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
import logging
from app_conf.models import Company, Country
from django_datatables_view.base_datatable_view import BaseDatatableView
from django.utils.translation import gettext as _

import json
from collections import defaultdict
from django.utils.translation import get_language
from django.db.models import F, Q, Case, When, Value, CharField, Prefetch, Count, Sum
from datetime import datetime
from django.utils import timezone
from app_cabinet.models import CabinetUser


logger = logging.getLogger(__name__)



logger.debug("This is a debug message")
logger.info("This is an info message")
logger.warning("This is a warning message")
logger.error("This is an error message")


def _software_used_license_qty(software_id, exclude_asset_id=None):
    qs = InformationAssetSoftwareSelection.objects.filter(software_register_id=software_id)
    if exclude_asset_id:
        qs = qs.exclude(information_asset_id=exclude_asset_id)
    return qs.aggregate(total=Sum('selected_license_quantity')).get('total') or 0


def _parse_software_entries_payload(raw):
    """Support [id, ...] and [{id, version, license_quantity}, ...]."""
    result = []
    try:
        payload = json.loads(raw or '[]')
    except (TypeError, ValueError):
        return result
    if not isinstance(payload, list):
        return result
    for item in payload:
        if isinstance(item, int):
            result.append({'id': item, 'version': '', 'license_quantity': None})
            continue
        if isinstance(item, dict):
            sw_id = item.get('id')
            if not sw_id:
                continue
            qty = item.get('license_quantity')
            try:
                qty = int(qty) if qty not in (None, '') else None
            except (TypeError, ValueError):
                qty = None
            result.append({
                'id': int(sw_id),
                'version': (item.get('version') or '').strip(),
                'license_quantity': qty,
            })
    return result


def _parse_version_tags(raw_value):
    tokens = []
    for part in (raw_value or '').replace(';', ',').split(','):
        val = (part or '').strip()
        if val:
            tokens.append(val)
    return tokens


def _get_allowed_companies_for_asset_user(user):
    user_groups = user.groups.all()
    access_assets = AccessAssets.objects.filter(group__in=user_groups, has_access=True)
    companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    if companies.exists():
        return companies.order_by('name')
    if not access_assets.exists():
        return Company.objects.none()
    # Staff/superuser may see all assets in the table while M2M companies are unset
    if user.is_staff or user.is_superuser:
        return Company.objects.filter(
            informationasset__company__isnull=False,
        ).distinct().order_by('name')
    return Company.objects.none()


def _get_company_criticality_queryset(company_id):
    return CriticalityLevel.objects.filter(
        is_active=True
    ).filter(
        Q(company__isnull=True) | Q(company_id=company_id)
    ).prefetch_related('translations__country').order_by('display_order', 'cost', 'name')


def get_user_asset_permissions(user):
    access = AccessAssets.objects.filter(group__in=user.groups.all(), has_access=True)
    if access.exists():
        companies = set()
        can_edit = False
        for a in access:
            companies.update(a.companies.values_list('id', flat=True))
            can_edit |= a.can_edit
        return {
            'can_edit': can_edit,
            'show_link': True,
            'companies': list(companies)
        }
    return {
        'can_edit': False,
        'show_link': False,
        'companies': []
    }

def has_asset_permission(user, company):
    return AccessAssets.objects.filter(
        group__in=user.groups.all(),
        has_access=True
    ).exists()

def has_access_to_assets(user):
    user_groups = user.groups.all()
    has_access = AccessAssets.objects.filter(
        group__in=user_groups,
        has_access=True
    ).exists()
    logger.info(
        f"Access check for user {user}: Groups={[g.name for g in user_groups]}, "
        f"Has access={has_access}"
    )
    return has_access


def has_view_software_register(user):
    """User can view Software Register page if any of their groups has can_view_software_register."""
    if not user.is_authenticated:
        return False
    return AccessAssets.objects.filter(
        group__in=user.groups.all(),
        can_view_software_register=True
    ).exists()


def has_edit_software_register(user):
    """User can add/edit/delete Software Register entries if any of their groups has can_edit_software_register."""
    if not user.is_authenticated:
        return False
    return AccessAssets.objects.filter(
        group__in=user.groups.all(),
        can_edit_software_register=True
    ).exists()


def has_view_external_media_register(user):
    """User can view External Media Register if any of their groups has can_view_external_media_register."""
    if not user.is_authenticated:
        return False
    return AccessAssets.objects.filter(
        group__in=user.groups.all(),
        can_view_external_media_register=True
    ).exists()


def has_edit_external_media_register(user):
    """User can add/edit/delete External Media Register entries if any of their groups has can_edit_external_media_register."""
    if not user.is_authenticated:
        return False
    return AccessAssets.objects.filter(
        group__in=user.groups.all(),
        can_edit_external_media_register=True
    ).exists()


@login_required
def information_assets(request):
    user = request.user
    # print('user = ', user)
    user_groups = user.groups.all()
    # print('user_groups = ', user_groups)

    # Отримати записи AccessAssets для груп користувача з дозволом на доступ до активів
    access_assets = AccessAssets.objects.filter(group__in=user_groups, has_access=True)
    # print('access_assets = ', access_assets)

    # Перевірити, чи є хоча б один запис з дозволом на доступ
    has_access = access_assets.filter(has_access=True).exists()
    can_edit = access_assets.filter(can_edit=True).exists()
    manage_adm_own = access_assets.filter(manage_adm_own=True).exists()
    manage_types = access_assets.filter(manage_types=True).exists()
    # print('has_access = ', has_access)
    # print('can_edit = ', can_edit)

    # Отримати компанії, до яких користувач має доступ через AccessAssets
    allowed_companies = _get_allowed_companies_for_asset_user(user)
    # print('allowed_companies = ', allowed_companies)


    current_language = get_language()[:2]

    # Get groups with asset types (active only) and their translations for correct language in modal
    asset_types_prefetch = AssetType.objects.filter(is_active=True).prefetch_related('translations__country')
    groups = AssetGroup.objects.filter(is_active=True).prefetch_related(
        Prefetch('asset_types', queryset=asset_types_prefetch),
        'translations__country'
    ).order_by('display_order', 'name')

    # Precompute localized display names for groups and asset types (uses current request language)
    for group in groups:
        group.display_name = group.get_name()
        for at in group.asset_types.all():
            at.display_name = at.get_name()

    criticality_levels = CriticalityLevel.objects.filter(
        is_active=True,
        company__isnull=True,
    ).prefetch_related('translations__country').order_by('display_order', 'cost', 'name')
    
    # Precompute localized display names for criticality levels (uses current request language)
    for level in criticality_levels:
        level.display_name = level.get_name()
    
    if allowed_companies.exists():
        asset_owners = AssetOwner.objects.filter(company__in=allowed_companies).order_by('company__name')
        asset_administrators = AssetAdministrator.objects.filter(company__in=allowed_companies).order_by('company__name')
    elif user.is_staff or user.is_superuser:
        asset_owners = AssetOwner.objects.all().order_by('company__name')
        asset_administrators = AssetAdministrator.objects.all().order_by('company__name')
    else:
        asset_owners = AssetOwner.objects.none()
        asset_administrators = AssetAdministrator.objects.none()
    software_register_entries = SoftwareRegister.objects.filter(is_active=True).select_related('status', 'company', 'group', 'asset_type').order_by('name')
    software_statuses = SoftwareStatus.objects.filter(is_active=True).order_by('display_order', 'name')
    software_license_types = SoftwareLicenseType.objects.filter(is_active=True).order_by('display_order', 'name')
    _sw_asset_types_qs = AssetType.objects.filter(is_active=True).prefetch_related('translations__country')
    sw_groups = list(AssetGroup.objects.filter(is_active=True, show_in_software_register=True).prefetch_related(
        Prefetch('asset_types', queryset=_sw_asset_types_qs),
        'translations__country',
    ).order_by('display_order', 'name'))
    for _sg in sw_groups:
        _sg.display_name = _sg.get_name()
        for _sat in _sg.asset_types.all():
            _sat.display_name = _sat.get_name()
    if allowed_companies.exists():
        software_register_entries = software_register_entries.filter(Q(company__isnull=True) | Q(company__in=allowed_companies))
    else:
        software_register_entries = software_register_entries.filter(company__isnull=True)
    usage_rows = InformationAssetSoftwareSelection.objects.values('software_register_id').annotate(
        used_qty=Sum('selected_license_quantity')
    )
    usage_map = {r['software_register_id']: (r.get('used_qty') or 0) for r in usage_rows}
    software_register_entries = list(software_register_entries)
    for sw in software_register_entries:
        used_qty = usage_map.get(sw.id, 0)
        sw.used_license_quantity = used_qty
        sw.free_license_quantity = max((sw.license_quantity or 0) - used_qty, 0) if sw.license_quantity is not None else None

    context = {
        'allowed_companies': allowed_companies,
        'groups': groups,
        'criticality_levels': criticality_levels,
        'asset_owners': asset_owners,
        'asset_administrators': asset_administrators,
        'software_register_entries': software_register_entries,
        'software_statuses': software_statuses,
        'software_license_types': software_license_types,
        'sw_groups': sw_groups,
        'can_edit': can_edit,
        'manage_adm_own': manage_adm_own,
        'manage_types': manage_types,
        'current_language': current_language,
        'hasAccess': has_access
    }

    return render(request, 'app_asset/information_assets.html', context)


@login_required
@user_passes_test(has_view_software_register)
def software_register(request):
    """Реєстр дозволеного/забороненого програмного забезпечення."""
    user_groups = request.user.groups.all()
    access_assets_view = AccessAssets.objects.filter(group__in=user_groups, can_view_software_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()

    # Show global entries (company=None) and entries for user's companies (active + inactive; UI hides inactive unless toggled)
    qs = SoftwareRegister.objects.select_related(
        'status', 'group', 'asset_type',
        'confidentiality', 'integrity', 'availability',
        'company', 'actualized_by',
    ).prefetch_related('files', 'owners')
    if allowed_companies.exists():
        qs = qs.filter(Q(company__isnull=True) | Q(company__in=allowed_companies))
    else:
        qs = qs.filter(company__isnull=True)

    software_list = list(qs.order_by('status__display_order', 'display_order', 'name'))
    today_local = timezone.localdate()
    for item in software_list:
        item.license_expires_soon = False
        if item.license_valid_until:
            days_left = (item.license_valid_until - today_local).days
            item.license_expires_soon = 0 <= days_left <= 14
    software_statuses = list(SoftwareStatus.objects.filter(is_active=True).order_by('display_order', 'name'))
    software_license_types = list(SoftwareLicenseType.objects.filter(is_active=True).order_by('display_order', 'name'))
    _asset_types_qs = AssetType.objects.filter(is_active=True).prefetch_related('translations__country')
    groups = list(AssetGroup.objects.filter(is_active=True, show_in_software_register=True).prefetch_related(
        Prefetch('asset_types', queryset=_asset_types_qs),
        'translations__country',
    ).order_by('display_order', 'name'))
    for _g in groups:
        _g.display_name = _g.get_name()
        for _at in _g.asset_types.all():
            _at.display_name = _at.get_name()
    criticality_levels = list(CriticalityLevel.objects.filter(
        is_active=True,
    ).filter(
        Q(company__isnull=True) | Q(company__in=allowed_companies)
    ).prefetch_related('translations__country').order_by('display_order', 'cost', 'name'))
    for _cl in criticality_levels:
        _cl.display_name = _cl.get_name()

    can_edit = has_edit_software_register(request.user)
    manage_types = AccessAssets.objects.filter(group__in=user_groups, manage_types=True).exists()
    owner_entry_ids = set()
    if software_list:
        owner_entry_ids = set(
            qs.filter(owners__cabinet_user__user=request.user).values_list('id', flat=True).distinct()
        )
        usage_map = defaultdict(lambda: defaultdict(list))
        software_ids = [s.id for s in software_list]
        selections = InformationAssetSoftwareSelection.objects.filter(
            software_register_id__in=software_ids
        ).select_related('information_asset', 'software_register')
        for sel in selections:
            version_key = (sel.selected_version or '').strip() or '—'
            asset_label = f"{sel.information_asset.asset_id}: {sel.information_asset.name}"
            usage_map[sel.software_register_id][version_key].append(asset_label)
        for item in software_list:
            lines = []
            by_version = usage_map.get(item.id, {})
            for ver, assets in by_version.items():
                lines.append(f"{ver}: {', '.join(sorted(set(assets)))}")
            item.version_usage_tooltip = "\n".join(lines)
        license_usage_rows = InformationAssetSoftwareSelection.objects.filter(
            software_register_id__in=software_ids
        ).select_related('information_asset', 'software_register')
        license_usage_map = defaultdict(list)
        for row in license_usage_rows:
            asset_label = f"{row.information_asset.asset_id}: {row.information_asset.name}"
            qty = row.selected_license_quantity
            qty_label = qty if qty is not None else '—'
            license_usage_map[row.software_register_id].append(f"{asset_label} — {qty_label}")
        for item in software_list:
            item.license_usage_tooltip = "\n".join(sorted(license_usage_map.get(item.id, [])))
    context = {
        'software_list': software_list,
        'allowed_companies': allowed_companies,
        'software_statuses': software_statuses,
        'software_license_types': software_license_types,
        'groups': groups,
        'criticality_levels': criticality_levels,
        'can_edit': can_edit,
        'manage_types': manage_types,
        'software_owner_entry_ids': owner_entry_ids,
    }
    return render(request, 'app_asset/software_register.html', context)


def _user_can_edit_software_entry(user, entry):
    """Check if user can edit/delete this software register entry (can_edit_software_register + company scope)."""
    access_assets = AccessAssets.objects.filter(group__in=user.groups.all(), can_edit_software_register=True)
    if not access_assets.exists():
        return False
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    if entry.company is None:
        return True
    return entry.company in allowed_companies


def _user_is_software_owner(user, entry):
    if not entry.company_id:
        return False
    cabinet_user = CabinetUser.objects.filter(user=user, company=entry.company).first()
    if not cabinet_user:
        return False
    return entry.owners.filter(cabinet_user=cabinet_user).exists()


def _resolve_software_license_type_code(raw_value):
    raw_value = (raw_value or '').strip()
    if not raw_value:
        return ''
    normalized = raw_value.lower()
    if SoftwareLicenseType.objects.filter(code=normalized, is_active=True).exists():
        return normalized
    return ''


def _format_software_history_changes(changes):
    if not isinstance(changes, dict):
        return []

    field_labels = {
        'name': _('Name'),
        'status': _('Status'),
        'description': _('Description'),
        'company': _('Company'),
        'version_pattern': _('Version Pattern'),
        'manufacturer': _('Manufacturer'),
        'url': _('URL'),
        'license_type': _('License Type'),
        'license_quantity': _('License Quantity'),
        'license_valid_until': _('License Valid Until'),
        'notes': _('Notes'),
        'is_active': _('Is Active'),
        'display_order': _('Display Order'),
        'actualization_date': _('Actualization date'),
        'actualized_by': _('Actualized by'),
        'marked_no_longer_actual_at': _('Marked no longer actual at'),
        'marked_no_longer_comment': _('Marked no longer actual comment'),
    }

    status_map = {s.id: s.get_name() for s in SoftwareStatus.objects.filter(is_active=True)}
    company_map = {c.id: c.name for c in Company.objects.all()}
    license_type_map = {lt.code: lt.get_name() for lt in SoftwareLicenseType.objects.filter(is_active=True)}

    def _is_empty(v):
        return v in (None, '', [])

    def _stringify(field, value):
        if _is_empty(value):
            return '—'
        if field == 'status':
            return status_map.get(value, str(value))
        if field == 'company':
            return company_map.get(value, str(value))
        if field == 'license_type':
            return license_type_map.get(str(value), str(value))
        if field == 'is_active':
            return _('Yes') if bool(value) else _('No')
        return str(value)

    lines = []
    for field, payload in changes.items():
        if not isinstance(payload, dict) or ('old' not in payload and 'new' not in payload):
            continue
        old_v = payload.get('old')
        new_v = payload.get('new')
        if old_v == new_v:
            continue
        label = field_labels.get(field, field.replace('_', ' ').title())
        lines.append(f"{label}: {_stringify(field, old_v)} -> {_stringify(field, new_v)}")
    return lines


@login_required
def get_software_register_history(request, pk):
    """Get history records for a software register entry (JSON). View permission required."""
    try:
        entry = get_object_or_404(SoftwareRegister, pk=pk)
        if not request.user.groups.filter(accessassets__can_view_software_register=True).exists():
            return JsonResponse({'error': _('Permission denied')}, status=403)
        access_assets_view = AccessAssets.objects.filter(group__in=request.user.groups.all(), can_view_software_register=True)
        allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
        if entry.company and entry.company not in allowed_companies:
            return JsonResponse({'error': _('Permission denied')}, status=403)
        records = SoftwareRegisterHistory.objects.filter(software_register=entry).select_related('action_by').order_by('-timestamp')[:100]
        data = []
        for r in records:
            data.append({
                'timestamp': timezone.localtime(r.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                'action': str(r.get_action_display()),
                'action_by': r.action_by.get_full_name() or r.action_by.username if r.action_by else '',
                'details': r.details or '',
                'changes': r.changes,
                'changes_display': _format_software_history_changes(r.changes),
            })
        return JsonResponse({'history': data})
    except Exception:
        logger.exception('get_software_register_history error')
        return JsonResponse({'error': _('Error loading history.')}, status=500)


@login_required
@user_passes_test(has_edit_software_register)
def get_software_register_entry(request, pk):
    """Get one software register entry as JSON (for edit modal)."""
    entry = get_object_or_404(SoftwareRegister, pk=pk)
    if not _user_can_edit_software_entry(request.user, entry):
        return JsonResponse({'error': _('Permission denied')}, status=403)
    files = [
        {
            'id': f.id,
            'name': f.file.name.split('/')[-1] if f.file else '',
            'url': f.file.url if f.file else '',
            'hash': f.file_hash or '',
            'label': f.label or '',
            'uploaded_at': f.uploaded_at.isoformat() if f.uploaded_at else '',
        }
        for f in entry.files.all()
    ]
    owners = [{'id': o.id, 'name': o.name} for o in entry.owners.all()]
    used_license_quantity = _software_used_license_qty(entry.id)
    free_license_quantity = max((entry.license_quantity or 0) - used_license_quantity, 0) if entry.license_quantity is not None else None
    used_versions_set = set()
    for raw_ver in InformationAssetSoftwareSelection.objects.filter(
        software_register=entry
    ).exclude(selected_version__isnull=True).exclude(selected_version__exact='').values_list('selected_version', flat=True):
        for token in _parse_version_tags(raw_ver):
            used_versions_set.add(token)
    used_versions = sorted(used_versions_set)
    return JsonResponse({
        'id': entry.id,
        'name': entry.name,
        'status': entry.status_id,
        'description': entry.description or '',
        'company': entry.company_id or '',
        'version_pattern': entry.version_pattern or '',
        'manufacturer': entry.manufacturer or '',
        'url': entry.url or '',
        'group': entry.group_id or '',
        'asset_type': entry.asset_type_id or '',
        'confidentiality': entry.confidentiality_id or '',
        'integrity': entry.integrity_id or '',
        'availability': entry.availability_id or '',
        'license_type': entry.license_type or '',
        'license_quantity': entry.license_quantity or '',
        'license_valid_until': entry.license_valid_until.isoformat() if entry.license_valid_until else '',
        'notes': entry.notes or '',
        'is_active': entry.is_active,
        'display_order': entry.display_order,
        'files': files,
        'owners': owners,
        'used_license_quantity': used_license_quantity,
        'free_license_quantity': free_license_quantity,
        'used_versions': used_versions,
    })


@require_POST
@login_required
@user_passes_test(has_edit_software_register)
def add_software_register(request):
    """Create a new software register entry."""
    access_assets = AccessAssets.objects.filter(
        group__in=request.user.groups.all(), can_edit_software_register=True
    )
    if not access_assets.exists():
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()

    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': _('Name is required')}, status=400)
    status_id = request.POST.get('status')
    status_obj = None
    if status_id and SoftwareStatus.objects.filter(pk=status_id, is_active=True).exists():
        status_obj = SoftwareStatus.objects.get(pk=status_id)
    else:
        status_obj = SoftwareStatus.objects.filter(is_active=True).order_by('display_order').first()
    if not status_obj:
        return JsonResponse({'status': 'error', 'message': _('No active software status defined')}, status=400)
    company_id = request.POST.get('company')
    company = None
    if company_id and allowed_companies.filter(pk=company_id).exists():
        company = Company.objects.get(pk=company_id)
    elif company_id and company_id != '':
        return JsonResponse({'status': 'error', 'message': _('Invalid company')}, status=400)
    group_id = request.POST.get('group')
    sw_group = None
    if group_id and AssetGroup.objects.filter(pk=group_id, is_active=True).exists():
        sw_group = AssetGroup.objects.get(pk=group_id)
    asset_type_id = request.POST.get('asset_type')
    sw_asset_type = None
    if asset_type_id and AssetType.objects.filter(pk=asset_type_id, is_active=True).exists():
        sw_asset_type = AssetType.objects.get(pk=asset_type_id)
    confidentiality_id = request.POST.get('confidentiality')
    sw_confidentiality = CriticalityLevel.objects.filter(pk=confidentiality_id, is_active=True).first() if confidentiality_id else None
    integrity_id = request.POST.get('integrity')
    sw_integrity = CriticalityLevel.objects.filter(pk=integrity_id, is_active=True).first() if integrity_id else None
    availability_id = request.POST.get('availability')
    sw_availability = CriticalityLevel.objects.filter(pk=availability_id, is_active=True).first() if availability_id else None
    license_valid_until = request.POST.get('license_valid_until')
    from django.utils.dateparse import parse_date as parse_date_field
    license_valid_until_date = parse_date_field(license_valid_until) if license_valid_until else None
    license_quantity = request.POST.get('license_quantity')
    try:
        license_quantity_int = int(license_quantity) if license_quantity else None
    except (TypeError, ValueError):
        license_quantity_int = None

    try:
        entry = SoftwareRegister.objects.create(
            name=name,
            status=status_obj,
            description=(request.POST.get('description') or '').strip(),
            company=company,
            version_pattern=(request.POST.get('version_pattern') or '').strip(),
            manufacturer=(request.POST.get('manufacturer') or '').strip(),
            url=(request.POST.get('url') or '').strip(),
            group=sw_group,
            asset_type=sw_asset_type,
            confidentiality=sw_confidentiality,
            integrity=sw_integrity,
            availability=sw_availability,
            license_type=_resolve_software_license_type_code(request.POST.get('license_type')),
            license_quantity=license_quantity_int,
            license_valid_until=license_valid_until_date,
            notes=(request.POST.get('notes') or '').strip(),
            is_active=request.POST.get('is_active', 'true').lower() in ('true', '1', 'on'),
            display_order=int(request.POST.get('display_order') or 0),
        )
        owners_ids = []
        try:
            owners_json = request.POST.get('owners', '[]')
            owners_ids = json.loads(owners_json) if owners_json else []
        except (ValueError, TypeError):
            pass
        if owners_ids and company:
            valid_owners = AssetOwner.objects.filter(id__in=owners_ids, company=company)
            entry.owners.set(valid_owners)
        SoftwareRegisterHistory.objects.create(
            software_register=entry,
            action=SoftwareRegisterHistory.ACTION_CREATED,
            action_by=request.user,
            details=_('Created'),
        )
        return JsonResponse({'status': 'success', 'message': _('Software entry added'), 'id': entry.id})
    except Exception as e:
        logger.error(f"Error adding software register entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_software_register)
def edit_software_register(request):
    """Update a software register entry."""
    entry_id = request.POST.get('id')
    if not entry_id:
        return JsonResponse({'status': 'error', 'message': _('ID is required')}, status=400)
    entry = get_object_or_404(SoftwareRegister, pk=entry_id)
    if not _user_can_edit_software_entry(request.user, entry):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)

    access_assets = AccessAssets.objects.filter(
        group__in=request.user.groups.all(), can_edit_software_register=True
    )
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()

    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': _('Name is required')}, status=400)
    status_id = request.POST.get('status')
    if status_id and SoftwareStatus.objects.filter(pk=status_id, is_active=True).exists():
        entry.status = SoftwareStatus.objects.get(pk=status_id)
    company_id = request.POST.get('company')
    company = None
    if company_id and allowed_companies.filter(pk=company_id).exists():
        company = Company.objects.get(pk=company_id)
    elif company_id and company_id != '':
        return JsonResponse({'status': 'error', 'message': _('Invalid company')}, status=400)
    group_id = request.POST.get('group')
    sw_group = None
    if group_id and AssetGroup.objects.filter(pk=group_id, is_active=True).exists():
        sw_group = AssetGroup.objects.get(pk=group_id)
    asset_type_id = request.POST.get('asset_type')
    sw_asset_type = None
    if asset_type_id and AssetType.objects.filter(pk=asset_type_id, is_active=True).exists():
        sw_asset_type = AssetType.objects.get(pk=asset_type_id)
    confidentiality_id = request.POST.get('confidentiality')
    sw_confidentiality = CriticalityLevel.objects.filter(pk=confidentiality_id, is_active=True).first() if confidentiality_id else None
    integrity_id = request.POST.get('integrity')
    sw_integrity = CriticalityLevel.objects.filter(pk=integrity_id, is_active=True).first() if integrity_id else None
    availability_id = request.POST.get('availability')
    sw_availability = CriticalityLevel.objects.filter(pk=availability_id, is_active=True).first() if availability_id else None
    license_valid_until = request.POST.get('license_valid_until')
    from django.utils.dateparse import parse_date as parse_date_field
    license_valid_until_date = parse_date_field(license_valid_until) if license_valid_until else None
    license_quantity = request.POST.get('license_quantity')
    requested_version_pattern = (request.POST.get('version_pattern') or '').strip()
    try:
        license_quantity_int = int(license_quantity) if license_quantity else None
    except (TypeError, ValueError):
        license_quantity_int = None

    try:
        old_versions = set(_parse_version_tags(entry.version_pattern))
        new_versions = set(_parse_version_tags(requested_version_pattern))
        removed_versions = old_versions - new_versions
        if removed_versions:
            locked_rows = InformationAssetSoftwareSelection.objects.filter(
                software_register=entry
            ).select_related('information_asset')
            by_version = {}
            for row in locked_rows:
                row_versions = set(_parse_version_tags(row.selected_version or ''))
                for ver in sorted(removed_versions.intersection(row_versions)):
                    by_version.setdefault(ver, set()).add(
                        f"{row.information_asset.asset_id}: {row.information_asset.name}"
                    )
            if by_version:
                detail = "; ".join(
                    [f"{ver} -> {', '.join(sorted(assets))}" for ver, assets in sorted(by_version.items())]
                )
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot remove used Versions from Linked Software: {}').format(detail)
                }, status=400)
        used_license_quantity = _software_used_license_qty(entry.id)
        if license_quantity_int is not None and license_quantity_int < used_license_quantity:
            return JsonResponse({
                'status': 'error',
                'message': _('License Quantity cannot be lower than used quantity: {}').format(used_license_quantity)
            }, status=400)

        from django.forms.models import model_to_dict
        old_dict = model_to_dict(entry, exclude=['id'], fields=[f.name for f in entry._meta.fields if f.name != 'id'])
        entry.name = name
        entry.description = (request.POST.get('description') or '').strip()
        entry.company = company
        entry.version_pattern = requested_version_pattern
        entry.manufacturer = (request.POST.get('manufacturer') or '').strip()
        entry.url = (request.POST.get('url') or '').strip()
        entry.group = sw_group
        entry.asset_type = sw_asset_type
        entry.confidentiality = sw_confidentiality
        entry.integrity = sw_integrity
        entry.availability = sw_availability
        entry.license_type = _resolve_software_license_type_code(request.POST.get('license_type'))
        entry.license_quantity = license_quantity_int
        entry.license_valid_until = license_valid_until_date
        entry.notes = (request.POST.get('notes') or '').strip()
        entry.is_active = request.POST.get('is_active', 'true').lower() in ('true', '1', 'on')
        entry.display_order = int(request.POST.get('display_order') or 0)
        if status_id and SoftwareStatus.objects.filter(pk=status_id, is_active=True).exists():
            entry.status = SoftwareStatus.objects.get(pk=status_id)
        entry.save()
        owners_ids = []
        try:
            owners_json = request.POST.get('owners', '[]')
            owners_ids = json.loads(owners_json) if owners_json else []
        except (ValueError, TypeError):
            pass
        if entry.company:
            valid_owners = AssetOwner.objects.filter(id__in=owners_ids, company=entry.company)
            entry.owners.set(valid_owners)
        else:
            entry.owners.clear()
        new_dict = model_to_dict(entry, exclude=['id'], fields=[f.name for f in entry._meta.fields if f.name != 'id'])
        changes = {k: {'old': old_dict.get(k), 'new': new_dict.get(k)} for k in new_dict if old_dict.get(k) != new_dict.get(k)}
        SoftwareRegisterHistory.objects.create(
            software_register=entry,
            action=SoftwareRegisterHistory.ACTION_MODIFIED,
            action_by=request.user,
            details=_('Updated'),
            changes=changes if changes else None,
        )
        return JsonResponse({'status': 'success', 'message': _('Software entry updated')})
    except Exception as e:
        logger.error(f"Error editing software register entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_software_register)
def delete_software_register(request, pk):
    """Delete a software register entry."""
    entry = get_object_or_404(SoftwareRegister, pk=pk)
    if not _user_can_edit_software_entry(request.user, entry):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    try:
        entry_name = entry.name
        SoftwareRegisterHistory.objects.create(
            software_register=None,
            entry_name=entry_name,
            action=SoftwareRegisterHistory.ACTION_DELETED,
            action_by=request.user,
            details=_('Deleted'),
        )
        entry.delete()
        return JsonResponse({'status': 'success', 'message': _('Software entry deleted')})
    except Exception as e:
        logger.error(f"Error deleting software register entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_software_register)
def upload_software_register_files(request, pk):
    """Upload one or more files for a software register entry. Returns list of added files (id, name, url, hash)."""
    entry = get_object_or_404(SoftwareRegister, pk=pk)
    if not _user_can_edit_software_entry(request.user, entry):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    uploaded = request.FILES.getlist('files') or request.FILES.getlist('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': _('No files provided')}, status=400)
    added = []
    try:
        for f in uploaded:
            if not f.name:
                continue
            obj = SoftwareRegisterFile(software_register=entry, file=f)
            obj.save()
            obj.refresh_from_db()
            added.append({
                'id': obj.id,
                'name': obj.file.name.split('/')[-1] if obj.file else f.name,
                'url': obj.file.url if obj.file else '',
                'hash': obj.file_hash or '',
                'label': obj.label or '',
                'uploaded_at': obj.uploaded_at.isoformat() if obj.uploaded_at else '',
            })
        return JsonResponse({'status': 'success', 'message': _('Files uploaded'), 'files': added})
    except Exception as e:
        logger.error(f"Error uploading software register files: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_software_register)
def delete_software_register_file(request, file_pk):
    """Delete an attached file from a software register entry."""
    file_obj = get_object_or_404(SoftwareRegisterFile, pk=file_pk)
    if not _user_can_edit_software_entry(request.user, file_obj.software_register):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    try:
        file_obj.delete()
        return JsonResponse({'status': 'success', 'message': _('File deleted')})
    except Exception as e:
        logger.error(f"Error deleting software register file: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def _format_owners_export(owners_queryset):
    """Format owners for export: 'Name (dept/pos); Name2 (dept2/pos2)'."""
    def _owner_sidecar_text(val):
        if val is None:
            return ''
        if hasattr(val, 'get_name'):
            return (val.get_name() or '').strip()
        return (val if isinstance(val, str) else str(val)).strip()

    parts = []
    for o in owners_queryset:
        dept = _owner_sidecar_text(o.department)
        pos = _owner_sidecar_text(o.position)
        s = o.name or ''
        if dept or pos:
            s += f" ({dept}{' / ' if dept and pos else ''}{pos})"
        if s:
            parts.append(s)
    return '; '.join(parts)


def _get_software_register_export_queryset(user):
    """Return SoftwareRegister queryset visible to user (same as list view)."""
    access_assets_view = AccessAssets.objects.filter(group__in=user.groups.all(), can_view_software_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
    qs = SoftwareRegister.objects.filter(is_active=True).select_related(
        'status', 'group', 'asset_type',
        'confidentiality', 'integrity', 'availability',
        'company', 'actualized_by',
    ).prefetch_related('files', 'owners')
    if allowed_companies.exists():
        qs = qs.filter(Q(company__isnull=True) | Q(company__in=allowed_companies))
    else:
        qs = qs.filter(company__isnull=True)
    return qs.order_by('status__display_order', 'display_order', 'name'), allowed_companies


@login_required
@user_passes_test(has_view_software_register)
def export_software_register(request):
    """Export Software Register to Excel or CSV. ?include_files=1 adds file names and hashes; ?format=csv for CSV."""
    include_files = request.GET.get('include_files', '').lower() in ('1', 'true', 'yes')
    as_csv = request.GET.get('format', '').lower() == 'csv'
    qs, _allowed_companies = _get_software_register_export_queryset(request.user)
    entries = list(qs)

    if as_csv:
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="software_register.csv"'
        response.write('\ufeff')
        writer = csv.writer(response, delimiter=';')
        headers = [
            _('Name'), _('Status'), _('Company'), _('Owners'), _('Manufacturer'), _('URL'),
            _('Version'), _('License Type'), _('License Qty'), _('License Until'), _('Description'),
            _('Notes'), _('Actualized At'), _('Actualized By'), _('No longer actual at'),
            _('No longer actual comment'), _('Active'), _('Display Order')
        ]
        if include_files:
            headers.extend([_('File Name'), _('File Hash (SHA256)'), _('Uploaded At')])
        writer.writerow(headers)
        for entry in entries:
            row = [
                entry.name,
                entry.status.get_name() if entry.status else '',
                entry.company.name if entry.company else '',
                _format_owners_export(entry.owners.all()),
                entry.manufacturer or '',
                entry.url or '',
                entry.version_pattern or '',
                entry.license_type or '',
                entry.license_quantity or '',
                entry.license_valid_until.strftime('%Y-%m-%d') if entry.license_valid_until else '',
                entry.description or '',
                entry.notes or '',
                timezone.localtime(entry.actualization_date).strftime('%Y-%m-%d %H:%M:%S') if entry.actualization_date else '',
                (entry.actualized_by.get_full_name() or entry.actualized_by.username) if entry.actualized_by else '',
                timezone.localtime(entry.marked_no_longer_actual_at).strftime('%Y-%m-%d %H:%M:%S') if entry.marked_no_longer_actual_at else '',
                entry.marked_no_longer_comment or '',
                _('Yes') if entry.is_active else _('No'),
                entry.display_order,
            ]
            if include_files and entry.files.exists():
                for f in entry.files.all():
                    file_row = row + [
                        f.file.name.split('/')[-1] if f.file else '',
                        f.file_hash or '',
                        f.uploaded_at.strftime('%Y-%m-%d %H:%M') if f.uploaded_at else '',
                    ]
                    writer.writerow(file_row)
            else:
                if include_files:
                    row.extend(['', '', ''])
                writer.writerow(row)
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Software Register")
    headers = [
        _('Name'), _('Status'), _('Company'), _('Owners'), _('Manufacturer'), _('URL'),
        _('Version'), _('License Type'), _('License Qty'), _('License Until'), _('Description'),
        _('Notes'), _('Actualized At'), _('Actualized By'), _('No longer actual at'),
        _('No longer actual comment'), _('Active'), _('Display Order')
    ]
    if include_files:
        headers.extend([_('Files (name; hash; date)')])
    for col_num, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=h)
        ws.cell(row=1, column=col_num).font = Font(bold=True)
    for row_num, entry in enumerate(entries, 2):
        row = [
            entry.name,
            entry.status.get_name() if entry.status else '',
            entry.company.name if entry.company else '',
            _format_owners_export(entry.owners.all()),
            entry.manufacturer or '',
            entry.url or '',
            entry.version_pattern or '',
            entry.license_type or '',
            entry.license_quantity or '',
            entry.license_valid_until.strftime('%Y-%m-%d') if entry.license_valid_until else '',
            entry.description or '',
            entry.notes or '',
            timezone.localtime(entry.actualization_date).strftime('%Y-%m-%d %H:%M:%S') if entry.actualization_date else '',
            (entry.actualized_by.get_full_name() or entry.actualized_by.username) if entry.actualized_by else '',
            timezone.localtime(entry.marked_no_longer_actual_at).strftime('%Y-%m-%d %H:%M:%S') if entry.marked_no_longer_actual_at else '',
            entry.marked_no_longer_comment or '',
            _('Yes') if entry.is_active else _('No'),
            entry.display_order,
        ]
        if include_files:
            files_text = ' | '.join(
                f"{f.file.name.split('/')[-1] if f.file else ''}; {f.file_hash or ''}; {f.uploaded_at.strftime('%Y-%m-%d %H:%M') if f.uploaded_at else ''}"
                for f in entry.files.all()
            )
            row.append(files_text)
        for col_num, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=val)
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="software_register.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(has_edit_software_register)
def import_software_register(request):
    """Import Software Register from CSV or Excel. POST with file=... . Returns JSON { created, updated, errors }."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': _('POST required')}, status=405)
    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'status': 'error', 'message': _('No file provided')}, status=400)
    access_assets = AccessAssets.objects.filter(group__in=request.user.groups.all(), can_edit_software_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    status_default = SoftwareStatus.objects.filter(is_active=True).order_by('display_order').first()
    if not status_default:
        return JsonResponse({'status': 'error', 'message': _('No active software status defined')}, status=400)
    from django.utils.dateparse import parse_date as parse_date_field
    created = 0
    updated = 0
    errors = []
    name_col = _('Name')
    try:
        rows = []
        if (upload.name or '').lower().endswith('.xlsx') or (upload.name or '').lower().endswith('.xls'):
            wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else '' for c in row])
            wb.close()
        else:
            import csv
            import io
            content = upload.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content), delimiter=';')
            rows = list(reader)
        if not rows:
            return JsonResponse({'status': 'error', 'message': _('File is empty')}, status=400)
        headers = [h.strip() for h in rows[0]]
        name_idx = next((i for i, h in enumerate(headers) if h and name_col.lower() in h.lower()), 0)
        col = lambda key, default='': next((i for i, h in enumerate(headers) if h and key.lower() in h.lower()), None)
        idx = {k: col(k) for k in [
            _('Name'), _('Status'), _('Company'), _('Owners'), _('Manufacturer'), _('URL'),
            _('Version'), _('License Type'), _('License Qty'), _('License Until'), _('Description'),
            _('Notes'), _('Actualized At'), _('Actualized By'), _('No longer actual at'),
            _('No longer actual comment'), _('Active'), _('Display Order')
        ]}
        for row_num, row in enumerate(rows[1:], 2):
            if len(row) <= name_idx or not (row[name_idx] or '').strip():
                continue
            name = (row[name_idx] or '').strip()
            company = None
            if idx.get(_('Company')) is not None and row[idx[_('Company')]]:
                cname = (row[idx[_('Company')]] or '').strip()
                if cname:
                    company = allowed_companies.filter(name=cname).first()
            status = status_default
            if idx.get(_('Status')) is not None and row[idx[_('Status')]]:
                code = (row[idx[_('Status')]] or '').strip().lower()
                if SoftwareStatus.objects.filter(code=code, is_active=True).exists():
                    status = SoftwareStatus.objects.get(code=code)
            license_type = ''
            if idx.get(_('License Type')) is not None and row[idx[_('License Type')]]:
                license_type = _resolve_software_license_type_code(row[idx[_('License Type')]])
            license_valid_until = None
            if idx.get(_('License Until')) is not None and row[idx[_('License Until')]]:
                license_valid_until = parse_date_field(row[idx[_('License Until')]])
            license_quantity = None
            if idx.get(_('License Qty')) is not None and row[idx[_('License Qty')]]:
                try:
                    license_quantity = int(row[idx[_('License Qty')]])
                except (TypeError, ValueError):
                    pass
            is_active = True
            if idx.get(_('Active')) is not None and idx[_('Active')] < len(row):
                val = (row[idx[_('Active')]] or '').strip().lower()
                is_active = val not in ('no', '0', 'false', 'n')
            display_order = 0
            if idx.get(_('Display Order')) is not None and idx[_('Display Order')] < len(row) and row[idx[_('Display Order')]]:
                try:
                    display_order = int(row[idx[_('Display Order')]])
                except (TypeError, ValueError):
                    pass
            actualization_date = None
            if idx.get(_('Actualized At')) is not None and idx[_('Actualized At')] < len(row) and row[idx[_('Actualized At')]]:
                actualization_date = parse_datetime((row[idx[_('Actualized At')]] or '').strip())
            marked_no_longer_actual_at = None
            if idx.get(_('No longer actual at')) is not None and idx[_('No longer actual at')] < len(row) and row[idx[_('No longer actual at')]]:
                marked_no_longer_actual_at = parse_datetime((row[idx[_('No longer actual at')]] or '').strip())
            def v(k):
                i = idx.get(k)
                return (row[i] or '').strip() if i is not None and i < len(row) else ''
            def resolve_actualized_by(user_text):
                user_text = (user_text or '').strip()
                if not user_text:
                    return None
                user_obj = User.objects.filter(username=user_text).first()
                if user_obj:
                    return user_obj
                user_obj = User.objects.filter(email=user_text).first()
                if user_obj:
                    return user_obj
                users = User.objects.filter(first_name__isnull=False)
                for u in users:
                    full_name = (u.get_full_name() or '').strip()
                    if full_name and full_name == user_text:
                        return u
                return None

            def resolve_owners_from_cell(company_obj, owners_str):
                """Parse 'Name (dept/pos); Name2' and return list of AssetOwner for company."""
                if not owners_str or not company_obj:
                    return []
                result = []
                company_owners = list(AssetOwner.objects.filter(company=company_obj).select_related('cabinet_user__user'))
                for part in (owners_str or '').split(';'):
                    name = (part.split('(')[0] if '(' in part else part).strip()
                    if not name:
                        continue
                    for o in company_owners:
                        if (o.name or '').strip() == name:
                            result.append(o)
                            break
                return result

            try:
                existing = SoftwareRegister.objects.filter(name=name, company=company).first()
                owners_str = v(_('Owners'))
                resolved_owners = resolve_owners_from_cell(company, owners_str) if company else []
                if existing:
                    existing.status = status
                    existing.manufacturer = v(_('Manufacturer'))
                    existing.url = v(_('URL'))
                    existing.version_pattern = v(_('Version'))
                    existing.license_type = license_type
                    existing.license_quantity = license_quantity
                    existing.license_valid_until = license_valid_until
                    existing.description = v(_('Description'))
                    existing.notes = v(_('Notes'))
                    existing.actualization_date = actualization_date
                    existing.actualized_by = resolve_actualized_by(v(_('Actualized By'))) if actualization_date else None
                    existing.marked_no_longer_actual_at = marked_no_longer_actual_at
                    existing.marked_no_longer_comment = v(_('No longer actual comment'))
                    existing.is_active = is_active
                    existing.display_order = display_order
                    existing.save()
                    if company:
                        existing.owners.set(resolved_owners)
                    else:
                        existing.owners.clear()
                    updated += 1
                else:
                    new_entry = SoftwareRegister.objects.create(
                        name=name,
                        status=status,
                        company=company,
                        manufacturer=v(_('Manufacturer')),
                        url=v(_('URL')),
                        version_pattern=v(_('Version')),
                        license_type=license_type,
                        license_quantity=license_quantity,
                        license_valid_until=license_valid_until,
                        description=v(_('Description')),
                        notes=v(_('Notes')),
                        actualization_date=actualization_date,
                        actualized_by=resolve_actualized_by(v(_('Actualized By'))) if actualization_date else None,
                        marked_no_longer_actual_at=marked_no_longer_actual_at,
                        marked_no_longer_comment=v(_('No longer actual comment')),
                        is_active=is_active,
                        display_order=display_order,
                    )
                    if company and resolved_owners:
                        new_entry.owners.set(resolved_owners)
                    created += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        return JsonResponse({
            'status': 'success',
            'message': _('Import completed'),
            'created': created,
            'updated': updated,
            'errors': errors[:20],
        })
    except Exception as e:
        logger.error(f"Error importing software register: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@user_passes_test(has_edit_software_register)
def download_software_register_import_template(request):
    fmt = (request.GET.get('format') or 'csv').lower()
    headers = [
        _('Name'), _('Status'), _('Category'), _('Company'), _('Owners'), _('Manufacturer'), _('URL'),
        _('Version'), _('License Type'), _('License Qty'), _('License Until'), _('Description'),
        _('Notes'), _('Actualized At'), _('Actualized By'), _('No longer actual at'),
        _('No longer actual comment'), _('Active'), _('Display Order')
    ]
    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _("Software Import Template")
        for col_num, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=h)
            ws.cell(row=1, column=col_num).font = Font(bold=True)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="software_register_import_template.xlsx"'
        wb.save(response)
        return response

    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="software_register_import_template.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    return response


@require_POST
@login_required
@user_passes_test(has_edit_software_register)
def duplicate_software_register(request, pk):
    """Duplicate a software register entry. POST copy_files=1 to also copy attached files."""
    import os
    original = get_object_or_404(SoftwareRegister, pk=pk)
    if not _user_can_edit_software_entry(request.user, original):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    copy_files = request.POST.get('copy_files', '').lower() in ('1', 'true', 'yes')
    try:
        new_entry = SoftwareRegister(
            name=f"{original.name} ({_('copy')})",
            status=original.status,
            description=original.description,
            company=original.company,
            version_pattern=original.version_pattern,
            manufacturer=original.manufacturer,
            url=original.url,
            category=original.category,
            license_type=original.license_type,
            license_quantity=original.license_quantity,
            license_valid_until=original.license_valid_until,
            notes=original.notes,
            is_active=original.is_active,
            display_order=original.display_order,
        )
        new_entry.save()
        new_entry.owners.set(original.owners.all())
        if copy_files:
            for old_f in original.files.all():
                if not old_f.file:
                    continue
                new_f = SoftwareRegisterFile(software_register=new_entry, label=old_f.label)
                fname = os.path.basename(old_f.file.name) or 'file'
                old_f.file.open('rb')
                from django.core.files.base import ContentFile
                new_f.file.save(fname, ContentFile(old_f.file.read()), save=True)
                old_f.file.close()
        return JsonResponse({'status': 'success', 'message': _('Entry duplicated'), 'id': new_entry.id})
    except Exception as e:
        logger.error(f"Error duplicating software register: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@login_required
@require_POST
def actualize_software_register(request, entry_id):
    try:
        entry = get_object_or_404(SoftwareRegister, id=entry_id, is_active=True)
        action = request.POST.get('action', 'actualize')
        user = request.user

        access_assets_view = AccessAssets.objects.filter(
            group__in=user.groups.all(),
            can_view_software_register=True
        )
        allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
        if entry.company and entry.company not in allowed_companies:
            return JsonResponse({'status': 'error', 'message': _('You do not have access to this entry')}, status=403)

        if not _user_is_software_owner(user, entry):
            return JsonResponse(
                {'status': 'error', 'message': _('Only entry owners can actualize this record')},
                status=403
            )

        if action == 'mark_inactive':
            comment = (request.POST.get('comment') or '').strip()
            entry.actualization_date = None
            entry.actualized_by = None
            entry.marked_no_longer_actual_at = timezone.now()
            entry.marked_no_longer_comment = comment or ''
            entry.save(
                update_fields=[
                    'actualization_date',
                    'actualized_by',
                    'marked_no_longer_actual_at',
                    'marked_no_longer_comment',
                    'updated_date',
                ]
            )
            changes = {'action': 'mark_inactive', 'actualization_cleared': True}
            if comment:
                changes['comment'] = comment
            SoftwareRegisterHistory.objects.create(
                software_register=entry,
                action=SoftwareRegisterHistory.ACTION_MODIFIED,
                action_by=user,
                details=_("Entry marked as no longer actual by owner"),
                changes=changes
            )
            return JsonResponse({'status': 'success', 'message': _('Entry marked as no longer actual')})

        entry.actualization_date = timezone.now()
        entry.actualized_by = user
        entry.marked_no_longer_actual_at = None
        entry.marked_no_longer_comment = ''
        entry.save(
            update_fields=[
                'actualization_date',
                'actualized_by',
                'marked_no_longer_actual_at',
                'marked_no_longer_comment',
                'updated_date',
            ]
        )
        SoftwareRegisterHistory.objects.create(
            software_register=entry,
            action=SoftwareRegisterHistory.ACTION_MODIFIED,
            action_by=user,
            details=_("Entry actualized by owner"),
            changes={
                'actualization_date': timezone.localtime(entry.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
                'actualized_by': user.get_full_name() or user.username
            }
        )
        return JsonResponse({
            'status': 'success',
            'message': _('Entry actualized successfully'),
            'actualization_date': timezone.localtime(entry.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
            'actualized_by': user.get_full_name() or user.username
        })
    except Exception as e:
        logger.error(f"Error actualizing software entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@login_required
@require_POST
def bulk_actualize_software_register(request):
    """Bulk actualize software register entries by owner."""
    try:
        data = json.loads(request.body)
        entry_ids = data.get('entry_ids', [])

        if not entry_ids:
            return JsonResponse({
                'status': 'error',
                'message': _('No software entries selected')
            }, status=400)

        user = request.user
        entries = SoftwareRegister.objects.filter(id__in=entry_ids, is_active=True)
        if entries.count() != len(entry_ids):
            return JsonResponse({
                'status': 'error',
                'message': _('Some software entries were not found')
            }, status=404)

        access_assets_view = AccessAssets.objects.filter(
            group__in=user.groups.all(),
            can_view_software_register=True
        )
        allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()

        actualized_count = 0
        skipped_count = 0
        errors = []

        for entry in entries:
            try:
                if entry.company and entry.company not in allowed_companies:
                    skipped_count += 1
                    errors.append(f"{entry.name}: {_('No access to company')}")
                    continue

                if not _user_is_software_owner(user, entry):
                    skipped_count += 1
                    errors.append(f"{entry.name}: {_('Not an owner')}")
                    continue

                entry.actualization_date = timezone.now()
                entry.actualized_by = user
                entry.marked_no_longer_actual_at = None
                entry.marked_no_longer_comment = ''
                entry.save(
                    update_fields=[
                        'actualization_date',
                        'actualized_by',
                        'marked_no_longer_actual_at',
                        'marked_no_longer_comment',
                        'updated_date',
                    ]
                )

                SoftwareRegisterHistory.objects.create(
                    software_register=entry,
                    action=SoftwareRegisterHistory.ACTION_MODIFIED,
                    action_by=user,
                    details=_("Entry actualized by owner (bulk operation)"),
                    changes={
                        'actualization_date': timezone.localtime(entry.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
                        'actualized_by': user.get_full_name() or user.username
                    }
                )
                actualized_count += 1
            except Exception as e:
                logger.error(f"Error actualizing software entry {entry.id}: {str(e)}")
                skipped_count += 1
                errors.append(f"{entry.name}: {str(e)}")

        if actualized_count == 0:
            return JsonResponse({
                'status': 'error',
                'message': _('No software entries were actualized. You may not be an owner of the selected entries.'),
                'errors': errors[:10]
            }, status=400)

        message = _('Successfully actualized {} software entries').format(actualized_count)
        if skipped_count > 0:
            message += f". {skipped_count} {_('skipped')}"

        return JsonResponse({
            'status': 'success',
            'message': message,
            'actualized_count': actualized_count,
            'skipped_count': skipped_count,
            'errors': errors[:10] if errors else []
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': _('Invalid JSON data')
        }, status=400)
    except Exception as e:
        logger.error(f"Error in bulk actualize software register: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


# --- External Media Register views ---

def _user_can_edit_external_media_entry(user, entry):
    access_assets = AccessAssets.objects.filter(group__in=user.groups.all(), can_edit_external_media_register=True)
    if not access_assets.exists():
        return False
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    if entry.company is None:
        return True
    return entry.company in allowed_companies


def _user_is_external_media_owner(user, entry):
    if not entry.company_id:
        return False
    cabinet_user = CabinetUser.objects.filter(user=user, company=entry.company).first()
    if not cabinet_user:
        return False
    return entry.owners.filter(cabinet_user=cabinet_user).exists()


@login_required
@user_passes_test(has_view_external_media_register)
def external_media_register(request):
    user_groups = request.user.groups.all()
    access_assets_view = AccessAssets.objects.filter(group__in=user_groups, can_view_external_media_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
    qs = ExternalMediaRegister.objects.select_related(
        'status', 'company', 'actualized_by', 'group', 'asset_type',
        'confidentiality', 'integrity', 'availability',
    ).prefetch_related('files', 'owners')
    if allowed_companies.exists():
        qs = qs.filter(Q(company__isnull=True) | Q(company__in=allowed_companies))
    else:
        qs = qs.filter(company__isnull=True)
    media_list = list(qs.order_by('status__display_order', 'display_order', 'name'))
    statuses = list(ExternalMediaStatus.objects.filter(is_active=True).order_by('display_order', 'name'))
    can_edit = has_edit_external_media_register(request.user)
    manage_types = AccessAssets.objects.filter(group__in=user_groups, manage_types=True).exists()
    groups = []
    if manage_types:
        _em_asset_types_qs = AssetType.objects.filter(is_active=True).prefetch_related('translations__country')
        groups = list(AssetGroup.objects.filter(is_active=True, show_in_external_media_register=True).prefetch_related(
            Prefetch('asset_types', queryset=_em_asset_types_qs),
            'translations__country',
        ).order_by('display_order', 'name'))
        for _g in groups:
            _g.display_name = _g.get_name()
            for _at in _g.asset_types.all():
                _at.display_name = _at.get_name()
    owner_entry_ids = set()
    if media_list:
        owner_entry_ids = set(
            qs.filter(owners__cabinet_user__user=request.user).values_list('id', flat=True).distinct()
        )
    criticality_levels = list(CriticalityLevel.objects.filter(
        is_active=True,
    ).filter(
        Q(company__isnull=True) | Q(company__in=allowed_companies)
    ).prefetch_related('translations__country').order_by('display_order', 'cost', 'name'))
    for _cl in criticality_levels:
        _cl.display_name = _cl.get_name()
    context = {
        'media_list': media_list,
        'allowed_companies': allowed_companies,
        'external_media_statuses': statuses,
        'can_edit': can_edit,
        'external_media_owner_entry_ids': owner_entry_ids,
        'manage_types': manage_types,
        'groups': groups,
        'criticality_levels': criticality_levels,
    }
    return render(request, 'app_asset/external_media_register.html', context)


@login_required
def get_external_media_register_history(request, pk):
    try:
        entry = get_object_or_404(ExternalMediaRegister, pk=pk)
        if not request.user.groups.filter(accessassets__can_view_external_media_register=True).exists():
            return JsonResponse({'error': _('Permission denied')}, status=403)
        access_assets_view = AccessAssets.objects.filter(group__in=request.user.groups.all(), can_view_external_media_register=True)
        allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
        if entry.company and entry.company not in allowed_companies:
            return JsonResponse({'error': _('Permission denied')}, status=403)
        records = ExternalMediaRegisterHistory.objects.filter(external_media_register=entry).select_related('action_by').order_by('-timestamp')[:100]
        data = []
        for r in records:
            data.append({
                'timestamp': timezone.localtime(r.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                'action': str(r.get_action_display()),
                'action_by': r.action_by.get_full_name() or r.action_by.username if r.action_by else '',
                'details': r.details or '',
                'changes': r.changes,
            })
        return JsonResponse({'history': data})
    except Exception:
        logger.exception('get_external_media_register_history error')
        return JsonResponse({'error': _('Error loading history.')}, status=500)


@login_required
@user_passes_test(has_edit_external_media_register)
def get_external_media_register_entry(request, pk):
    entry = get_object_or_404(ExternalMediaRegister, pk=pk)
    if not _user_can_edit_external_media_entry(request.user, entry):
        return JsonResponse({'error': _('Permission denied')}, status=403)
    files = [
        {'id': f.id, 'name': f.file.name.split('/')[-1] if f.file else '', 'url': f.file.url if f.file else '', 'hash': f.file_hash or '', 'label': f.label or '', 'uploaded_at': f.uploaded_at.isoformat() if f.uploaded_at else ''}
        for f in entry.files.all()
    ]
    owners = [{'id': o.id, 'name': o.name} for o in entry.owners.all()]
    return JsonResponse({
        'id': entry.id, 'name': entry.name, 'status': entry.status_id, 'company': entry.company_id or '',
        'group': entry.group_id,
        'asset_type': entry.asset_type_id,
        'confidentiality': entry.confidentiality_id or '',
        'integrity': entry.integrity_id or '',
        'availability': entry.availability_id or '',
        'serial_number': entry.serial_number or '', 'description': entry.description or '', 'notes': entry.notes or '',
        'is_active': entry.is_active, 'display_order': entry.display_order, 'files': files, 'owners': owners,
    })


@require_POST
@login_required
@user_passes_test(has_edit_external_media_register)
def add_external_media_register(request):
    access_assets = AccessAssets.objects.filter(group__in=request.user.groups.all(), can_edit_external_media_register=True)
    if not access_assets.exists():
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': _('Name is required')}, status=400)
    status_id = request.POST.get('status')
    status_obj = ExternalMediaStatus.objects.filter(pk=status_id, is_active=True).first() if status_id else ExternalMediaStatus.objects.filter(is_active=True).order_by('display_order').first()
    if not status_obj:
        return JsonResponse({'status': 'error', 'message': _('No active status defined')}, status=400)
    company_id = request.POST.get('company')
    company = None
    if company_id and allowed_companies.filter(pk=company_id).exists():
        company = Company.objects.get(pk=company_id)
    group_id = request.POST.get('group')
    em_group = None
    if group_id and AssetGroup.objects.filter(pk=group_id, is_active=True).exists():
        em_group = AssetGroup.objects.get(pk=group_id)
    asset_type_id = request.POST.get('asset_type')
    em_asset_type = None
    if asset_type_id and AssetType.objects.filter(pk=asset_type_id, is_active=True).exists():
        em_asset_type = AssetType.objects.get(pk=asset_type_id)
        if em_group and em_asset_type.group_id != em_group.id:
            em_asset_type = None
        elif not em_group and em_asset_type:
            em_group = em_asset_type.group
    confidentiality_id = request.POST.get('confidentiality')
    em_confidentiality = CriticalityLevel.objects.filter(pk=confidentiality_id, is_active=True).first() if confidentiality_id else None
    integrity_id = request.POST.get('integrity')
    em_integrity = CriticalityLevel.objects.filter(pk=integrity_id, is_active=True).first() if integrity_id else None
    availability_id = request.POST.get('availability')
    em_availability = CriticalityLevel.objects.filter(pk=availability_id, is_active=True).first() if availability_id else None
    try:
        entry = ExternalMediaRegister.objects.create(
            name=name, status=status_obj, company=company,
            group=em_group,
            asset_type=em_asset_type,
            confidentiality=em_confidentiality,
            integrity=em_integrity,
            availability=em_availability,
            serial_number=(request.POST.get('serial_number') or '').strip(),
            description=(request.POST.get('description') or '').strip(),
            notes=(request.POST.get('notes') or '').strip(),
            is_active=request.POST.get('is_active', 'true').lower() in ('true', '1', 'on'),
            display_order=int(request.POST.get('display_order') or 0),
        )
        owners_ids = []
        try:
            owners_json = request.POST.get('owners', '[]')
            owners_ids = json.loads(owners_json) if owners_json else []
        except (ValueError, TypeError):
            pass
        if owners_ids and company:
            valid_owners = AssetOwner.objects.filter(id__in=owners_ids, company=company)
            entry.owners.set(valid_owners)
        ExternalMediaRegisterHistory.objects.create(external_media_register=entry, action=ExternalMediaRegisterHistory.ACTION_CREATED, action_by=request.user, details=_('Created'))
        return JsonResponse({'status': 'success', 'message': _('Entry added'), 'id': entry.id})
    except Exception as e:
        logger.error(f"Error adding external media entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_external_media_register)
def edit_external_media_register(request):
    entry_id = request.POST.get('id')
    if not entry_id:
        return JsonResponse({'status': 'error', 'message': _('ID is required')}, status=400)
    entry = get_object_or_404(ExternalMediaRegister, pk=entry_id)
    if not _user_can_edit_external_media_entry(request.user, entry):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'status': 'error', 'message': _('Name is required')}, status=400)
    access_assets = AccessAssets.objects.filter(group__in=request.user.groups.all(), can_edit_external_media_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    company_id = request.POST.get('company')
    company = None
    if company_id and allowed_companies.filter(pk=company_id).exists():
        company = Company.objects.get(pk=company_id)
    status_id = request.POST.get('status')
    if status_id and ExternalMediaStatus.objects.filter(pk=status_id, is_active=True).exists():
        entry.status = ExternalMediaStatus.objects.get(pk=status_id)
    try:
        from django.forms.models import model_to_dict
        old_dict = model_to_dict(entry, exclude=['id'], fields=[f.name for f in entry._meta.fields if f.name != 'id'])
        entry.name = name
        entry.serial_number = (request.POST.get('serial_number') or '').strip()
        entry.description = (request.POST.get('description') or '').strip()
        entry.notes = (request.POST.get('notes') or '').strip()
        entry.company = company
        group_id = request.POST.get('group')
        em_group = None
        if group_id and AssetGroup.objects.filter(pk=group_id, is_active=True).exists():
            em_group = AssetGroup.objects.get(pk=group_id)
        asset_type_id = request.POST.get('asset_type')
        em_asset_type = None
        if asset_type_id and AssetType.objects.filter(pk=asset_type_id, is_active=True).exists():
            em_asset_type = AssetType.objects.get(pk=asset_type_id)
            if em_group and em_asset_type.group_id != em_group.id:
                em_asset_type = None
            elif not em_group and em_asset_type:
                em_group = em_asset_type.group
        confidentiality_id = request.POST.get('confidentiality')
        em_confidentiality = CriticalityLevel.objects.filter(pk=confidentiality_id, is_active=True).first() if confidentiality_id else None
        integrity_id = request.POST.get('integrity')
        em_integrity = CriticalityLevel.objects.filter(pk=integrity_id, is_active=True).first() if integrity_id else None
        availability_id = request.POST.get('availability')
        em_availability = CriticalityLevel.objects.filter(pk=availability_id, is_active=True).first() if availability_id else None
        entry.group = em_group
        entry.asset_type = em_asset_type
        entry.confidentiality = em_confidentiality
        entry.integrity = em_integrity
        entry.availability = em_availability
        entry.is_active = request.POST.get('is_active', 'true').lower() in ('true', '1', 'on')
        entry.display_order = int(request.POST.get('display_order') or 0)
        entry.save()
        owners_ids = []
        try:
            owners_json = request.POST.get('owners', '[]')
            owners_ids = json.loads(owners_json) if owners_json else []
        except (ValueError, TypeError):
            pass
        if entry.company:
            valid_owners = AssetOwner.objects.filter(id__in=owners_ids, company=entry.company)
            entry.owners.set(valid_owners)
        else:
            entry.owners.clear()
        new_dict = model_to_dict(entry, exclude=['id'], fields=[f.name for f in entry._meta.fields if f.name != 'id'])
        changes = {k: {'old': old_dict.get(k), 'new': new_dict.get(k)} for k in new_dict if old_dict.get(k) != new_dict.get(k)}
        ExternalMediaRegisterHistory.objects.create(
            external_media_register=entry, action=ExternalMediaRegisterHistory.ACTION_MODIFIED,
            action_by=request.user, details=_('Updated'), changes=changes if changes else None,
        )
        return JsonResponse({'status': 'success', 'message': _('Entry updated')})
    except Exception as e:
        logger.error(f"Error editing external media entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_external_media_register)
def delete_external_media_register(request, pk):
    entry = get_object_or_404(ExternalMediaRegister, pk=pk)
    if not _user_can_edit_external_media_entry(request.user, entry):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    try:
        entry_name = entry.name
        ExternalMediaRegisterHistory.objects.create(
            external_media_register=None, entry_name=entry_name, action=ExternalMediaRegisterHistory.ACTION_DELETED,
            action_by=request.user, details=_('Deleted'),
        )
        entry.delete()
        return JsonResponse({'status': 'success', 'message': _('Entry deleted')})
    except Exception as e:
        logger.error(f"Error deleting external media entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_external_media_register)
def upload_external_media_register_files(request, pk):
    entry = get_object_or_404(ExternalMediaRegister, pk=pk)
    if not _user_can_edit_external_media_entry(request.user, entry):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    uploaded = request.FILES.getlist('files') or request.FILES.getlist('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': _('No files provided')}, status=400)
    added = []
    try:
        for f in uploaded:
            if not f.name:
                continue
            obj = ExternalMediaRegisterFile(external_media_register=entry, file=f)
            obj.save()
            obj.refresh_from_db()
            added.append({
                'id': obj.id, 'name': obj.file.name.split('/')[-1] if obj.file else f.name,
                'url': obj.file.url if obj.file else '', 'hash': obj.file_hash or '', 'label': obj.label or '',
                'uploaded_at': obj.uploaded_at.isoformat() if obj.uploaded_at else '',
            })
        return JsonResponse({'status': 'success', 'message': _('Files uploaded'), 'files': added})
    except Exception as e:
        logger.error(f"Error uploading external media files: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
@user_passes_test(has_edit_external_media_register)
def delete_external_media_register_file(request, file_pk):
    file_obj = get_object_or_404(ExternalMediaRegisterFile, pk=file_pk)
    if not _user_can_edit_external_media_entry(request.user, file_obj.external_media_register):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    try:
        file_obj.delete()
        return JsonResponse({'status': 'success', 'message': _('File deleted')})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def _get_external_media_export_queryset(user):
    access_assets_view = AccessAssets.objects.filter(group__in=user.groups.all(), can_view_external_media_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
    qs = ExternalMediaRegister.objects.filter(is_active=True).select_related(
        'status', 'company', 'actualized_by', 'group', 'asset_type',
        'confidentiality', 'integrity', 'availability',
    ).prefetch_related('files', 'owners')
    if allowed_companies.exists():
        qs = qs.filter(Q(company__isnull=True) | Q(company__in=allowed_companies))
    else:
        qs = qs.filter(company__isnull=True)
    return qs.order_by('status__display_order', 'display_order', 'name'), allowed_companies


@login_required
@user_passes_test(has_view_external_media_register)
def export_external_media_register(request):
    include_files = request.GET.get('include_files', '').lower() in ('1', 'true', 'yes')
    as_csv = request.GET.get('format', '').lower() == 'csv'
    qs, _allowed = _get_external_media_export_queryset(request.user)
    entries = list(qs)
    if as_csv:
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="external_media_register.csv"'
        response.write('\ufeff')
        writer = csv.writer(response, delimiter=';')
        headers = [
            _('Name'), _('Status'), _('Company'), _('Group'), _('Asset Type'), _('Owners'), _('Serial number'),
            _('Confidentiality'), _('Integrity'), _('Availability'),
            _('Description'), _('Notes'), _('Actualized At'), _('Actualized By'),
            _('No longer actual at'), _('No longer actual comment'), _('Active'), _('Display Order')
        ]
        if include_files:
            headers.extend([_('File Name'), _('File Hash (SHA256)'), _('Uploaded At')])
        writer.writerow(headers)
        for entry in entries:
            row = [
                entry.name, entry.status.get_name() if entry.status else '',
                entry.company.name if entry.company else '',
                entry.group.code if entry.group else '',
                entry.asset_type.code if entry.asset_type else '',
                _format_owners_export(entry.owners.all()), entry.serial_number or '',
                entry.confidentiality.get_name() if entry.confidentiality else '',
                entry.integrity.get_name() if entry.integrity else '',
                entry.availability.get_name() if entry.availability else '',
                entry.description or '', entry.notes or '',
                timezone.localtime(entry.actualization_date).strftime('%Y-%m-%d %H:%M:%S') if entry.actualization_date else '',
                (entry.actualized_by.get_full_name() or entry.actualized_by.username) if entry.actualized_by else '',
                timezone.localtime(entry.marked_no_longer_actual_at).strftime('%Y-%m-%d %H:%M:%S') if entry.marked_no_longer_actual_at else '',
                entry.marked_no_longer_comment or '',
                _('Yes') if entry.is_active else _('No'), entry.display_order
            ]
            if include_files and entry.files.exists():
                for f in entry.files.all():
                    writer.writerow(row + [f.file.name.split('/')[-1] if f.file else '', f.file_hash or '', f.uploaded_at.strftime('%Y-%m-%d %H:%M') if f.uploaded_at else ''])
            else:
                if include_files:
                    row.extend(['', '', ''])
                writer.writerow(row)
        return response
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("External Media Register")
    headers = [
        _('Name'), _('Status'), _('Company'), _('Group'), _('Asset Type'), _('Owners'), _('Serial number'),
        _('Confidentiality'), _('Integrity'), _('Availability'),
        _('Description'), _('Notes'), _('Actualized At'), _('Actualized By'),
        _('No longer actual at'), _('No longer actual comment'), _('Active'), _('Display Order')
    ]
    if include_files:
        headers.append(_('Files (name; hash; date)'))
    for col_num, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=h)
        ws.cell(row=1, column=col_num).font = Font(bold=True)
    for row_num, entry in enumerate(entries, 2):
        row = [
            entry.name, entry.status.get_name() if entry.status else '',
            entry.company.name if entry.company else '',
            entry.group.code if entry.group else '',
            entry.asset_type.code if entry.asset_type else '',
            _format_owners_export(entry.owners.all()), entry.serial_number or '',
            entry.confidentiality.get_name() if entry.confidentiality else '',
            entry.integrity.get_name() if entry.integrity else '',
            entry.availability.get_name() if entry.availability else '',
            entry.description or '', entry.notes or '',
            timezone.localtime(entry.actualization_date).strftime('%Y-%m-%d %H:%M:%S') if entry.actualization_date else '',
            (entry.actualized_by.get_full_name() or entry.actualized_by.username) if entry.actualized_by else '',
            timezone.localtime(entry.marked_no_longer_actual_at).strftime('%Y-%m-%d %H:%M:%S') if entry.marked_no_longer_actual_at else '',
            entry.marked_no_longer_comment or '',
            _('Yes') if entry.is_active else _('No'), entry.display_order
        ]
        if include_files:
            files_text = ' | '.join(f"{f.file.name.split('/')[-1] if f.file else ''}; {f.file_hash or ''}; {f.uploaded_at.strftime('%Y-%m-%d %H:%M') if f.uploaded_at else ''}" for f in entry.files.all())
            row.append(files_text)
        for col_num, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=val)
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="external_media_register.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(has_edit_external_media_register)
def import_external_media_register(request):
    """Import External Media Register from CSV or Excel. POST with file=... ."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': _('POST required')}, status=405)
    upload = request.FILES.get('file')
    if not upload:
        return JsonResponse({'status': 'error', 'message': _('No file provided')}, status=400)

    access_assets = AccessAssets.objects.filter(group__in=request.user.groups.all(), can_edit_external_media_register=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
    status_default = ExternalMediaStatus.objects.filter(is_active=True).order_by('display_order').first()
    if not status_default:
        return JsonResponse({'status': 'error', 'message': _('No active external media status defined')}, status=400)

    created = 0
    updated = 0
    errors = []
    try:
        rows = []
        if (upload.name or '').lower().endswith('.xlsx') or (upload.name or '').lower().endswith('.xls'):
            wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else '' for c in row])
            wb.close()
        else:
            import csv
            import io
            content = upload.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content), delimiter=';')
            rows = list(reader)

        if not rows:
            return JsonResponse({'status': 'error', 'message': _('File is empty')}, status=400)

        headers = [h.strip() for h in rows[0]]
        col = lambda key: next((i for i, h in enumerate(headers) if h and key.lower() in h.lower()), None)
        idx = {k: col(k) for k in [
            _('Name'), _('Status'), _('Company'), _('Group'), _('Asset Type'), _('Owners'),
            _('Confidentiality'), _('Integrity'), _('Availability'),
            _('Serial number'), _('Description'), _('Notes'),
            _('Actualized At'), _('Actualized By'), _('No longer actual at'),
            _('No longer actual comment'), _('Active'), _('Display Order')
        ]}
        name_idx = idx.get(_('Name')) if idx.get(_('Name')) is not None else 0

        def v(row, key):
            i = idx.get(key)
            return (row[i] or '').strip() if i is not None and i < len(row) else ''

        def resolve_owners_from_cell(company_obj, owners_str):
            if not owners_str or not company_obj:
                return []
            result = []
            company_owners = list(AssetOwner.objects.filter(company=company_obj).select_related('cabinet_user__user'))
            for part in (owners_str or '').split(';'):
                name = (part.split('(')[0] if '(' in part else part).strip()
                if not name:
                    continue
                for o in company_owners:
                    if (o.name or '').strip() == name:
                        result.append(o)
                        break
            return result

        def resolve_actualized_by(user_text):
            user_text = (user_text or '').strip()
            if not user_text:
                return None
            user_obj = User.objects.filter(username=user_text).first() or User.objects.filter(email=user_text).first()
            if user_obj:
                return user_obj
            for u in User.objects.all():
                if (u.get_full_name() or '').strip() == user_text:
                    return u
            return None

        for row_num, row in enumerate(rows[1:], 2):
            if len(row) <= name_idx or not (row[name_idx] or '').strip():
                continue
            try:
                name = (row[name_idx] or '').strip()
                company = None
                company_name = v(row, _('Company'))
                if company_name:
                    company = allowed_companies.filter(name=company_name).first()

                status = status_default
                status_code_or_name = v(row, _('Status')).lower()
                if status_code_or_name:
                    status = ExternalMediaStatus.objects.filter(
                        Q(code=status_code_or_name) | Q(name__iexact=status_code_or_name),
                        is_active=True
                    ).first() or status_default

                is_active = True
                active_val = v(row, _('Active')).lower()
                if active_val:
                    is_active = active_val not in ('no', '0', 'false', 'n')

                display_order = 0
                display_order_val = v(row, _('Display Order'))
                if display_order_val:
                    try:
                        display_order = int(display_order_val)
                    except (TypeError, ValueError):
                        display_order = 0

                actualization_date = parse_datetime(v(row, _('Actualized At'))) if v(row, _('Actualized At')) else None
                marked_no_longer_actual_at = parse_datetime(v(row, _('No longer actual at'))) if v(row, _('No longer actual at')) else None
                actualized_by = resolve_actualized_by(v(row, _('Actualized By'))) if actualization_date else None

                existing = ExternalMediaRegister.objects.filter(name=name, company=company).first()
                owners = resolve_owners_from_cell(company, v(row, _('Owners'))) if company else []

                g_cell = (v(row, _('Group')) or '').strip()
                t_cell = (v(row, _('Asset Type')) or '').strip()
                em_group_imp = None
                em_type_imp = None
                if g_cell:
                    em_group_imp = AssetGroup.objects.filter(
                        Q(code__iexact=g_cell) | Q(name__iexact=g_cell),
                        is_active=True,
                    ).first()
                if t_cell:
                    if em_group_imp:
                        em_type_imp = AssetType.objects.filter(
                            group=em_group_imp,
                            is_active=True,
                        ).filter(
                            Q(code__iexact=t_cell) | Q(name__iexact=t_cell),
                        ).first()
                    else:
                        em_type_imp = AssetType.objects.filter(
                            Q(code__iexact=t_cell) | Q(name__iexact=t_cell),
                            is_active=True,
                        ).first()
                        if em_type_imp:
                            em_group_imp = em_type_imp.group
                conf_cell = (v(row, _('Confidentiality')) or '').strip()
                int_cell = (v(row, _('Integrity')) or '').strip()
                avail_cell = (v(row, _('Availability')) or '').strip()
                em_conf_imp = CriticalityLevel.objects.filter(
                    Q(name__iexact=conf_cell) | Q(name_local__iexact=conf_cell),
                    is_active=True,
                ).first() if conf_cell else None
                em_int_imp = CriticalityLevel.objects.filter(
                    Q(name__iexact=int_cell) | Q(name_local__iexact=int_cell),
                    is_active=True,
                ).first() if int_cell else None
                em_avail_imp = CriticalityLevel.objects.filter(
                    Q(name__iexact=avail_cell) | Q(name_local__iexact=avail_cell),
                    is_active=True,
                ).first() if avail_cell else None

                if existing:
                    existing.status = status
                    existing.serial_number = v(row, _('Serial number'))
                    existing.description = v(row, _('Description'))
                    existing.notes = v(row, _('Notes'))
                    existing.actualization_date = actualization_date
                    existing.actualized_by = actualized_by
                    existing.marked_no_longer_actual_at = marked_no_longer_actual_at
                    existing.marked_no_longer_comment = v(row, _('No longer actual comment'))
                    existing.is_active = is_active
                    existing.display_order = display_order
                    existing.group = em_group_imp
                    existing.asset_type = em_type_imp
                    existing.confidentiality = em_conf_imp
                    existing.integrity = em_int_imp
                    existing.availability = em_avail_imp
                    existing.save()
                    if company:
                        existing.owners.set(owners)
                    else:
                        existing.owners.clear()
                    updated += 1
                else:
                    new_entry = ExternalMediaRegister.objects.create(
                        name=name,
                        status=status,
                        company=company,
                        group=em_group_imp,
                        asset_type=em_type_imp,
                        confidentiality=em_conf_imp,
                        integrity=em_int_imp,
                        availability=em_avail_imp,
                        serial_number=v(row, _('Serial number')),
                        description=v(row, _('Description')),
                        notes=v(row, _('Notes')),
                        actualization_date=actualization_date,
                        actualized_by=actualized_by,
                        marked_no_longer_actual_at=marked_no_longer_actual_at,
                        marked_no_longer_comment=v(row, _('No longer actual comment')),
                        is_active=is_active,
                        display_order=display_order,
                    )
                    if company and owners:
                        new_entry.owners.set(owners)
                    created += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        return JsonResponse({
            'status': 'success',
            'message': _('Import completed'),
            'created': created,
            'updated': updated,
            'errors': errors[:20],
        })
    except Exception as e:
        logger.error(f"Error importing external media register: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
@user_passes_test(has_edit_external_media_register)
def download_external_media_register_import_template(request):
    fmt = (request.GET.get('format') or 'csv').lower()
    headers = [
        _('Name'), _('Status'), _('Company'), _('Group'), _('Asset Type'), _('Owners'), _('Serial number'),
        _('Confidentiality'), _('Integrity'), _('Availability'),
        _('Description'), _('Notes'), _('Actualized At'), _('Actualized By'),
        _('No longer actual at'), _('No longer actual comment'), _('Active'), _('Display Order')
    ]
    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _("External Media Import Template")
        for col_num, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=h)
            ws.cell(row=1, column=col_num).font = Font(bold=True)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="external_media_register_import_template.xlsx"'
        wb.save(response)
        return response

    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="external_media_register_import_template.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    return response


@require_POST
@login_required
@user_passes_test(has_edit_external_media_register)
def duplicate_external_media_register(request, pk):
    import os
    original = get_object_or_404(ExternalMediaRegister, pk=pk)
    if not _user_can_edit_external_media_entry(request.user, original):
        return JsonResponse({'status': 'error', 'message': _('Permission denied')}, status=403)
    copy_files = request.POST.get('copy_files', '').lower() in ('1', 'true', 'yes')
    try:
        new_entry = ExternalMediaRegister(
            name=f"{original.name} ({_('copy')})", status=original.status, company=original.company,
            group=original.group,
            asset_type=original.asset_type,
            confidentiality=original.confidentiality,
            integrity=original.integrity,
            availability=original.availability,
            serial_number=original.serial_number, description=original.description, notes=original.notes,
            is_active=original.is_active, display_order=original.display_order,
        )
        new_entry.save()
        new_entry.owners.set(original.owners.all())
        if copy_files:
            for old_f in original.files.all():
                if not old_f.file:
                    continue
                new_f = ExternalMediaRegisterFile(external_media_register=new_entry, label=old_f.label)
                fname = os.path.basename(old_f.file.name) or 'file'
                old_f.file.open('rb')
                from django.core.files.base import ContentFile
                new_f.file.save(fname, ContentFile(old_f.file.read()), save=True)
                old_f.file.close()
        return JsonResponse({'status': 'success', 'message': _('Entry duplicated'), 'id': new_entry.id})
    except Exception as e:
        logger.error(f"Error duplicating external media entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@login_required
@require_POST
def actualize_external_media_register(request, entry_id):
    try:
        entry = get_object_or_404(ExternalMediaRegister, id=entry_id, is_active=True)
        action = request.POST.get('action', 'actualize')
        user = request.user

        access_assets_view = AccessAssets.objects.filter(
            group__in=user.groups.all(),
            can_view_external_media_register=True
        )
        allowed_companies = Company.objects.filter(access_assets__in=access_assets_view).distinct()
        if entry.company and entry.company not in allowed_companies:
            return JsonResponse({'status': 'error', 'message': _('You do not have access to this entry')}, status=403)

        if not _user_is_external_media_owner(user, entry):
            return JsonResponse(
                {'status': 'error', 'message': _('Only entry owners can actualize this record')},
                status=403
            )

        if action == 'mark_inactive':
            comment = (request.POST.get('comment') or '').strip()
            entry.actualization_date = None
            entry.actualized_by = None
            entry.marked_no_longer_actual_at = timezone.now()
            entry.marked_no_longer_comment = comment or ''
            entry.save(
                update_fields=[
                    'actualization_date',
                    'actualized_by',
                    'marked_no_longer_actual_at',
                    'marked_no_longer_comment',
                    'updated_date',
                ]
            )
            changes = {'action': 'mark_inactive', 'actualization_cleared': True}
            if comment:
                changes['comment'] = comment
            ExternalMediaRegisterHistory.objects.create(
                external_media_register=entry,
                action=ExternalMediaRegisterHistory.ACTION_MODIFIED,
                action_by=user,
                details=_("Entry marked as no longer actual by owner"),
                changes=changes
            )
            return JsonResponse({'status': 'success', 'message': _('Entry marked as no longer actual')})

        entry.actualization_date = timezone.now()
        entry.actualized_by = user
        entry.marked_no_longer_actual_at = None
        entry.marked_no_longer_comment = ''
        entry.save(
            update_fields=[
                'actualization_date',
                'actualized_by',
                'marked_no_longer_actual_at',
                'marked_no_longer_comment',
                'updated_date',
            ]
        )
        ExternalMediaRegisterHistory.objects.create(
            external_media_register=entry,
            action=ExternalMediaRegisterHistory.ACTION_MODIFIED,
            action_by=user,
            details=_("Entry actualized by owner"),
            changes={
                'actualization_date': timezone.localtime(entry.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
                'actualized_by': user.get_full_name() or user.username
            }
        )
        return JsonResponse({
            'status': 'success',
            'message': _('Entry actualized successfully'),
            'actualization_date': timezone.localtime(entry.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
            'actualized_by': user.get_full_name() or user.username
        })
    except Exception as e:
        logger.error(f"Error actualizing external media entry: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
@login_required
def add_asset(request):
    try:
        with transaction.atomic():
            data = request.POST

            # Parse JSON data for owners and administrators
            owners_ids = json.loads(data.get('owners', '[]'))
            administrators_ids = json.loads(data.get('administrators', '[]'))
            software_entries_payload = _parse_software_entries_payload(data.get('software_entries', '[]'))
            software_entries_ids = [x['id'] for x in software_entries_payload]

            # Parse dates
            registration_date_str = data.get('registrationDate')
            deletion_date_str = data.get('deletionDate')

            def parse_date(date_str):
                if date_str and date_str != 'NaN.NaN.NaN':
                    try:
                        day, month, year = date_str.split('.')
                        return datetime.strptime(f"{day}.{month}.{year}", "%d.%m.%Y").date()
                    except ValueError:
                        return None
                return None

            company_id = data.get('company')
            allowed_level_ids = set(_get_company_criticality_queryset(company_id).values_list('id', flat=True))
            confidentiality_id = data.get('confidentiality')
            integrity_id = data.get('integrity')
            availability_id = data.get('availability')

            for level_id in (confidentiality_id, integrity_id, availability_id):
                if not level_id or int(level_id) not in allowed_level_ids:
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Selected criticality level is not available for the chosen company.')
                    }, status=400)

            # Create the asset first to get ID
            asset = InformationAsset.objects.create(
                name=data.get('name'),
                company_id=company_id,
                description=data.get('description'),
                location=data.get('location'),
                confidentiality_id=confidentiality_id,
                integrity_id=integrity_id,
                availability_id=availability_id,
                registration_date=parse_date(registration_date_str),
                deletion_date=parse_date(deletion_date_str),
                notes=data.get('notes'),
                access_manage=data.get('access_manage') == 'on',
                is_active=data.get('is_active') == 'on',
                last_modified_by=request.user
            )

            # Handle group and asset type
            group_asset_type = data.get('group_asset_type')
            if group_asset_type:
                group_id, asset_type_id = map(int, group_asset_type.split(','))
                asset.group_id = group_id
                asset.asset_type_id = asset_type_id
                asset.save()

            # Add existing owners
            asset_owners = AssetOwner.objects.filter(id__in=owners_ids)
            asset.owners.add(*asset_owners)

            # Add existing administrators
            asset_administrators = AssetAdministrator.objects.filter(id__in=administrators_ids)
            asset.administrators.add(*asset_administrators)
            software_qs = SoftwareRegister.objects.filter(id__in=software_entries_ids, is_active=True)
            software_qs = software_qs.filter(Q(company__isnull=True) | Q(company_id=asset.company_id))
            for row in software_entries_payload:
                sw = software_qs.filter(id=row['id']).first()
                if not sw:
                    continue
                req_qty = row.get('license_quantity')
                if sw.license_valid_until and sw.license_valid_until < timezone.localdate() and req_qty not in (None, 0):
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Cannot set license quantity for {} because license validity has expired ({}).').format(
                            sw.name,
                            sw.license_valid_until.strftime('%d.%m.%Y')
                        )
                    }, status=400)
                if req_qty is None or sw.license_quantity is None:
                    continue
                free_qty = max(sw.license_quantity - _software_used_license_qty(sw.id), 0)
                if req_qty > free_qty:
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Requested license quantity for {} exceeds free licenses (free: {}).').format(sw.name, free_qty)
                    }, status=400)
            asset.software_entries.set(software_qs)
            InformationAssetSoftwareSelection.objects.filter(information_asset=asset).delete()
            selected_by_id = {x['id']: x for x in software_entries_payload}
            for sw in software_qs:
                row = selected_by_id.get(sw.id, {})
                InformationAssetSoftwareSelection.objects.create(
                    information_asset=asset,
                    software_register=sw,
                    selected_version=(row.get('version') or sw.version_pattern or '')[:100],
                    selected_license_quantity=row.get('license_quantity'),
                )

            logger.info(f"Asset created successfully with ID: {asset.id} and asset_id: {asset.asset_id}")
            logger.info(f"Added owners: {list(asset_owners.values_list('id', flat=True))}")
            logger.info(f"Added administrators: {list(asset_administrators.values_list('id', flat=True))}")

            return JsonResponse({
                'status': 'success',
                'message': _('Asset added successfully'),
                'asset_id': asset.id
            })

    except Exception as e:
        logger.error(f"Error adding asset: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@csrf_exempt
@require_POST
@user_passes_test(has_access_to_assets)
def edit_asset(request, asset_id):
    try:
        with transaction.atomic():
            if not has_asset_permission(request.user, InformationAsset.objects.get(id=asset_id).company):
                return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

            asset = get_object_or_404(InformationAsset, id=asset_id)
            
            # Store old values for history tracking
            old_confidentiality_id = asset.confidentiality_id
            old_integrity_id = asset.integrity_id
            old_availability_id = asset.availability_id
            old_group_id = asset.group_id
            old_asset_type_id = asset.asset_type_id
            old_name = asset.name
            old_description = asset.description
            old_location = asset.location
            old_access_manage = asset.access_manage
            old_is_active = asset.is_active
            old_registration_date = asset.registration_date
            old_deletion_date = asset.deletion_date
            old_notes = asset.notes
            old_owner_ids = set(asset.owners.values_list('id', flat=True))
            old_admin_ids = set(asset.administrators.values_list('id', flat=True))
            old_software_ids = set(asset.software_entries.values_list('id', flat=True))
            old_confidentiality_name = (
                asset.confidentiality.get_name() if asset.confidentiality else _("None")
            )
            old_integrity_name = (
                asset.integrity.get_name() if asset.integrity else _("None")
            )
            old_availability_name = (
                asset.availability.get_name() if asset.availability else _("None")
            )

            # Log received data for debugging
            logger.info(f"Received POST data for asset {asset_id}:")
            logger.info(f"Confidentiality: {request.POST.get('confidentiality')}")
            logger.info(f"Integrity: {request.POST.get('integrity')}")
            logger.info(f"Availability: {request.POST.get('availability')}")
            logger.info(f"All POST data: {dict(request.POST)}")

            # Get group and asset type IDs from the request
            group_id = request.POST.get('group')
            asset_type_id = request.POST.get('asset_type')

            # Validate group and asset type
            if not group_id or not asset_type_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Group and asset type are required',
                    'errors': {
                        'group': ['Group is required'] if not group_id else [],
                        'asset_type': ['Asset type is required'] if not asset_type_id else []
                    }
                }, status=400)

            try:
                group = AssetGroup.objects.get(id=group_id)
                asset_type = AssetType.objects.get(id=asset_type_id, group=group)
            except (AssetGroup.DoesNotExist, AssetType.DoesNotExist):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid group or asset type selection'
                }, status=400)

            # Update the asset with the new group and type
            asset.group = group
            asset.asset_type = asset_type

            # Update other fields
            asset.name = request.POST.get('name', asset.name)
            asset.company_id = request.POST.get('company', asset.company_id)
            asset.description = request.POST.get('description', '')
            asset.location = request.POST.get('location', '')

            # Handle criticality values with explicit logging
            confidentiality_id = request.POST.get('confidentiality')
            integrity_id = request.POST.get('integrity')
            availability_id = request.POST.get('availability')

            logger.info(f"Setting criticality values - Conf: {confidentiality_id}, Int: {integrity_id}, Avail: {availability_id}")

            allowed_level_ids = set(_get_company_criticality_queryset(asset.company_id).values_list('id', flat=True))
            for level_id in (confidentiality_id, integrity_id, availability_id):
                if not level_id or int(level_id) not in allowed_level_ids:
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Selected criticality level is not available for the chosen company.')
                    }, status=400)

            if confidentiality_id:
                asset.confidentiality_id = confidentiality_id
            if integrity_id:
                asset.integrity_id = integrity_id
            if availability_id:
                asset.availability_id = availability_id

            registration_date = request.POST.get('registrationDate')
            deletion_date = request.POST.get('deletionDate')

            asset.registration_date = parse_date(registration_date) if registration_date else None
            asset.deletion_date = parse_date(deletion_date) if deletion_date else None

            asset.notes = request.POST.get('notes', asset.notes)
            
            # Handle access_manage field
            access_manage = request.POST.get('access_manage') == 'on'
            asset.access_manage = access_manage
            asset.is_active = request.POST.get('is_active') == 'on'
            if asset.deletion_date:
                asset.is_active = False
            elif old_deletion_date and asset.deletion_date is None:
                asset.is_active = True

            # Handle owners and administrators
            owners_data = json.loads(request.POST.get('owners', '[]'))
            administrators_data = json.loads(request.POST.get('administrators', '[]'))

            # Process owners
            existing_owner_ids = []
            for owner_data in owners_data:
                if isinstance(owner_data, dict) and owner_data.get('isNew'):
                    new_owner = AssetOwner.objects.create(
                        name=owner_data['data']['name'],
                        department=owner_data['data']['department'],
                        position=owner_data['data']['position'],
                        email=owner_data['data']['email'],
                        phone=owner_data['data']['phone'],
                        company_id=asset.company_id
                    )
                    existing_owner_ids.append(new_owner.id)
                else:
                    owner_id = owner_data if isinstance(owner_data, int) else owner_data['id']
                    existing_owner_ids.append(owner_id)

            asset.owners.set(AssetOwner.objects.filter(id__in=existing_owner_ids))

            # Process administrators
            existing_admin_ids = []
            for admin_data in administrators_data:
                if isinstance(admin_data, dict) and admin_data.get('isNew'):
                    new_admin = AssetAdministrator.objects.create(
                        name=admin_data['data']['name'],
                        department=admin_data['data']['department'],
                        position=admin_data['data']['position'],
                        email=admin_data['data']['email'],
                        phone=admin_data['data']['phone'],
                        company_id=asset.company_id
                    )
                    existing_admin_ids.append(new_admin.id)
                else:
                    admin_id = admin_data if isinstance(admin_data, int) else admin_data['id']
                    existing_admin_ids.append(admin_id)

            asset.administrators.set(AssetAdministrator.objects.filter(id__in=existing_admin_ids))
            software_entries_payload = _parse_software_entries_payload(request.POST.get('software_entries', '[]'))
            software_entries_data = [x['id'] for x in software_entries_payload]
            software_qs = SoftwareRegister.objects.filter(id__in=software_entries_data, is_active=True)
            software_qs = software_qs.filter(Q(company__isnull=True) | Q(company_id=asset.company_id))
            current_qty_map = {
                s.software_register_id: (s.selected_license_quantity or 0)
                for s in InformationAssetSoftwareSelection.objects.filter(information_asset=asset)
            }
            for row in software_entries_payload:
                sw = software_qs.filter(id=row['id']).first()
                if not sw:
                    continue
                req_qty = row.get('license_quantity')
                if sw.license_valid_until and sw.license_valid_until < timezone.localdate() and req_qty not in (None, 0):
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Cannot set license quantity for {} because license validity has expired ({}).').format(
                            sw.name,
                            sw.license_valid_until.strftime('%d.%m.%Y')
                        )
                    }, status=400)
                if req_qty is None or sw.license_quantity is None:
                    continue
                used_except_current = _software_used_license_qty(sw.id, exclude_asset_id=asset.id)
                free_for_asset = max(sw.license_quantity - used_except_current, 0)
                if req_qty > free_for_asset:
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Requested license quantity for {} exceeds free licenses (free: {}).').format(sw.name, free_for_asset)
                    }, status=400)
            asset.software_entries.set(software_qs)
            InformationAssetSoftwareSelection.objects.filter(information_asset=asset).delete()
            selected_by_id = {x['id']: x for x in software_entries_payload}
            for sw in software_qs:
                row = selected_by_id.get(sw.id, {})
                InformationAssetSoftwareSelection.objects.create(
                    information_asset=asset,
                    software_register=sw,
                    selected_version=(row.get('version') or sw.version_pattern or '')[:100],
                    selected_license_quantity=row.get('license_quantity'),
                )

            asset.last_modified_by = request.user
            asset.save()
            
            # Log CIA changes manually if signals didn't catch them
            if old_confidentiality_id != asset.confidentiality_id:
                old_confidentiality = (
                    CriticalityLevel.objects.filter(id=old_confidentiality_id).first()
                    if old_confidentiality_id else None
                )
                new_confidentiality = (
                    CriticalityLevel.objects.filter(id=asset.confidentiality_id).first()
                    if asset.confidentiality_id else None
                )
                new_confidentiality_name = (
                    new_confidentiality.get_name() if new_confidentiality else _("None")
                )
                old_confidentiality_color = (
                    old_confidentiality.color if old_confidentiality else None
                )
                new_confidentiality_color = (
                    new_confidentiality.color if new_confidentiality else None
                )
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_CIA_CONFIDENTIALITY_CHANGED,
                    action_by=request.user,
                    details=_("Confidentiality changed from {} to {}").format(
                        old_confidentiality_name,
                        new_confidentiality_name
                    ),
                    changes={
                        'field': 'confidentiality',
                        'old_value': old_confidentiality_id,
                        'new_value': asset.confidentiality_id,
                        'old_name': old_confidentiality_name,
                        'new_name': new_confidentiality_name,
                        'old_color': old_confidentiality_color,
                        'new_color': new_confidentiality_color,
                    }
                )
            
            if old_integrity_id != asset.integrity_id:
                old_integrity = (
                    CriticalityLevel.objects.filter(id=old_integrity_id).first()
                    if old_integrity_id else None
                )
                new_integrity = (
                    CriticalityLevel.objects.filter(id=asset.integrity_id).first()
                    if asset.integrity_id else None
                )
                new_integrity_name = (
                    new_integrity.get_name() if new_integrity else _("None")
                )
                old_integrity_color = (
                    old_integrity.color if old_integrity else None
                )
                new_integrity_color = (
                    new_integrity.color if new_integrity else None
                )
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_CIA_INTEGRITY_CHANGED,
                    action_by=request.user,
                    details=_("Integrity changed from {} to {}").format(
                        old_integrity_name,
                        new_integrity_name
                    ),
                    changes={
                        'field': 'integrity',
                        'old_value': old_integrity_id,
                        'new_value': asset.integrity_id,
                        'old_name': old_integrity_name,
                        'new_name': new_integrity_name,
                        'old_color': old_integrity_color,
                        'new_color': new_integrity_color,
                    }
                )
            
            if old_availability_id != asset.availability_id:
                old_availability = (
                    CriticalityLevel.objects.filter(id=old_availability_id).first()
                    if old_availability_id else None
                )
                new_availability = (
                    CriticalityLevel.objects.filter(id=asset.availability_id).first()
                    if asset.availability_id else None
                )
                new_availability_name = (
                    new_availability.get_name() if new_availability else _("None")
                )
                old_availability_color = (
                    old_availability.color if old_availability else None
                )
                new_availability_color = (
                    new_availability.color if new_availability else None
                )
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_CIA_AVAILABILITY_CHANGED,
                    action_by=request.user,
                    details=_("Availability changed from {} to {}").format(
                        old_availability_name,
                        new_availability_name
                    ),
                    changes={
                        'field': 'availability',
                        'old_value': old_availability_id,
                        'new_value': asset.availability_id,
                        'old_name': old_availability_name,
                        'new_name': new_availability_name,
                        'old_color': old_availability_color,
                        'new_color': new_availability_color,
                    }
                )

            old_group = AssetGroup.objects.filter(id=old_group_id).first() if old_group_id else None
            new_group = AssetGroup.objects.filter(id=asset.group_id).first() if asset.group_id else None
            old_group_name = old_group.get_name() if old_group else _("None")
            new_group_name = new_group.get_name() if new_group else _("None")
            if old_group_id != asset.group_id:
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Group changed from {} to {}").format(
                        old_group_name,
                        new_group_name
                    ),
                    changes={
                        'field': 'group',
                        'old_value': old_group_id,
                        'new_value': asset.group_id,
                        'old_name': old_group_name,
                        'new_name': new_group_name,
                        'old_color': old_group.color if old_group else None,
                        'new_color': new_group.color if new_group else None,
                    }
                )

            old_asset_type = AssetType.objects.filter(id=old_asset_type_id).first() if old_asset_type_id else None
            new_asset_type = AssetType.objects.filter(id=asset.asset_type_id).first() if asset.asset_type_id else None
            old_asset_type_name = old_asset_type.get_name() if old_asset_type else _("None")
            new_asset_type_name = new_asset_type.get_name() if new_asset_type else _("None")
            if old_asset_type_id != asset.asset_type_id:
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Asset Type changed from {} to {}").format(
                        old_asset_type_name,
                        new_asset_type_name
                    ),
                    changes={
                        'field': 'asset_type',
                        'old_value': old_asset_type_id,
                        'new_value': asset.asset_type_id,
                        'old_name': old_asset_type_name,
                        'new_name': new_asset_type_name,
                        'old_color': old_asset_type.color if old_asset_type else None,
                        'new_color': new_asset_type.color if new_asset_type else None,
                    }
                )

            def _format_date(date_value):
                return date_value.strftime('%d-%m-%Y') if date_value else _("None")

            if old_registration_date != asset.registration_date:
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Registration Date changed from {} to {}").format(
                        _format_date(old_registration_date),
                        _format_date(asset.registration_date)
                    ),
                    changes={
                        'field': 'registration_date',
                        'old_value': _format_date(old_registration_date),
                        'new_value': _format_date(asset.registration_date),
                    }
                )

            if old_deletion_date != asset.deletion_date:
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Deletion Date changed from {} to {}").format(
                        _format_date(old_deletion_date),
                        _format_date(asset.deletion_date)
                    ),
                    changes={
                        'field': 'deletion_date',
                        'old_value': _format_date(old_deletion_date),
                        'new_value': _format_date(asset.deletion_date),
                    }
                )

            text_fields = [
                ('name', _("Name"), old_name, asset.name),
                ('description', _("Description"), old_description, asset.description),
                ('location', _("Location"), old_location, asset.location),
                ('notes', _("Notes"), old_notes, asset.notes),
            ]
            for field_key, field_label, old_value, new_value in text_fields:
                old_value = old_value or ""
                new_value = new_value or ""
                if old_value != new_value:
                    AssetHistory.objects.create(
                        asset=asset,
                        action=AssetHistory.ACTION_MODIFIED,
                        action_by=request.user,
                        details=_("{} changed from '{}' to '{}'").format(
                            field_label,
                            old_value,
                            new_value
                        ),
                        changes={
                            'field': field_key,
                            'old_value': old_value,
                            'new_value': new_value,
                        }
                    )

            if old_access_manage != asset.access_manage:
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Access Manage changed from {} to {}").format(
                        _("Yes") if old_access_manage else _("No"),
                        _("Yes") if asset.access_manage else _("No")
                    ),
                    changes={
                        'field': 'access_manage',
                        'old_value': old_access_manage,
                        'new_value': asset.access_manage,
                    }
                )

            if old_is_active != asset.is_active:
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Active changed from {} to {}").format(
                        _("Yes") if old_is_active else _("No"),
                        _("Yes") if asset.is_active else _("No")
                    ),
                    changes={
                        'field': 'is_active',
                        'old_value': old_is_active,
                        'new_value': asset.is_active,
                    }
                )

            new_owner_ids = set(asset.owners.values_list('id', flat=True))
            if old_owner_ids != new_owner_ids:
                old_owners = sorted(
                    AssetOwner.objects.filter(id__in=old_owner_ids),
                    key=lambda owner: (owner.name or "").lower()
                )
                new_owners = sorted(
                    AssetOwner.objects.filter(id__in=new_owner_ids),
                    key=lambda owner: (owner.name or "").lower()
                )
                old_owner_names = [owner.name for owner in old_owners]
                new_owner_names = [owner.name for owner in new_owners]
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_OWNERS_CHANGED,
                    action_by=request.user,
                    details=_("Asset Owners changed from {} to {}").format(
                        ", ".join(old_owner_names) if old_owner_names else _("None"),
                        ", ".join(new_owner_names) if new_owner_names else _("None")
                    ),
                    changes={
                        'field': 'owners',
                        'old_items': [{'id': owner.id, 'name': owner.name} for owner in old_owners],
                        'new_items': [{'id': owner.id, 'name': owner.name} for owner in new_owners],
                    }
                )

            new_admin_ids = set(asset.administrators.values_list('id', flat=True))
            if old_admin_ids != new_admin_ids:
                old_admins = sorted(
                    AssetAdministrator.objects.filter(id__in=old_admin_ids),
                    key=lambda admin: (admin.name or "").lower()
                )
                new_admins = sorted(
                    AssetAdministrator.objects.filter(id__in=new_admin_ids),
                    key=lambda admin: (admin.name or "").lower()
                )
                old_admin_names = [admin.name for admin in old_admins]
                new_admin_names = [admin.name for admin in new_admins]
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_ADMINISTRATORS_CHANGED,
                    action_by=request.user,
                    details=_("Asset Administrators changed from {} to {}").format(
                        ", ".join(old_admin_names) if old_admin_names else _("None"),
                        ", ".join(new_admin_names) if new_admin_names else _("None")
                    ),
                    changes={
                        'field': 'administrators',
                        'old_items': [{'id': admin.id, 'name': admin.name} for admin in old_admins],
                        'new_items': [{'id': admin.id, 'name': admin.name} for admin in new_admins],
                     }
                )

            new_software_ids = set(asset.software_entries.values_list('id', flat=True))
            if old_software_ids != new_software_ids:
                old_software = sorted(
                    SoftwareRegister.objects.filter(id__in=old_software_ids),
                    key=lambda s: (s.name or "").lower()
                )
                new_software = sorted(
                    SoftwareRegister.objects.filter(id__in=new_software_ids),
                    key=lambda s: (s.name or "").lower()
                )
                old_software_names = [s.name for s in old_software]
                new_software_names = [s.name for s in new_software]
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=request.user,
                    details=_("Linked Software changed from {} to {}").format(
                        ", ".join(old_software_names) if old_software_names else _("None"),
                        ", ".join(new_software_names) if new_software_names else _("None")
                    ),
                    changes={
                        'field': 'software_entries',
                        'old_items': [{'id': s.id, 'name': s.name} for s in old_software],
                        'new_items': [{'id': s.id, 'name': s.name} for s in new_software],
                    }
                )

            # Log final state
            logger.info(f"Asset {asset_id} updated successfully")
            logger.info(f"Final criticality values - Conf: {asset.confidentiality_id}, Int: {asset.integrity_id}, Avail: {asset.availability_id}")

            return JsonResponse({
                'status': 'success',
                'message': 'Asset updated successfully',
                'data': {
                    'id': asset.id,
                    'confidentiality': asset.confidentiality_id,
                    'integrity': asset.integrity_id,
                    'availability': asset.availability_id
                }
            })

    except Exception as e:
        logger.error(f"Error updating asset: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def delete_asset(request, asset_id):
    try:
        with transaction.atomic():
            asset = get_object_or_404(InformationAsset, id=asset_id)

            user_groups = request.user.groups.all()
            access_assets = AccessAssets.objects.filter(
                group__in=user_groups,
                companies=asset.company,
                can_edit=True
            )

            logger.info(f"Delete attempt for asset {asset_id} by user {request.user.username}")
            logger.info(f"User groups: {[g.name for g in user_groups]}")
            logger.info(f"Access records found: {access_assets.count()}")

            if not access_assets.exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('You do not have permission to delete this asset')
                }, status=403)

            if asset.deletion_date:
                return JsonResponse({
                    'status': 'error',
                    'message': _('This asset is already removed from the active register')
                }, status=400)

            asset.deletion_date = timezone.localdate()
            asset.is_active = False
            asset.last_modified_by = request.user
            asset.save()

            AssetHistory.objects.create(
                asset=asset,
                action=AssetHistory.ACTION_DELETED,
                action_by=request.user,
                details=_("Asset removed from the active register (deletion date set)"),
                changes={
                    'deletion_date': asset.deletion_date.isoformat(),
                    'soft_delete': True,
                },
            )

            logger.info(
                "Asset %s soft-deleted by user %s (deletion_date=%s)",
                asset.asset_id,
                request.user.username,
                asset.deletion_date,
            )

            return JsonResponse({
                'status': 'success',
                'message': _(
                    'Asset removed from the active list. Enable "Show Inactive Assets" to view or edit it.'
                ),
            })

    except InformationAsset.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': _('Asset not found')
        }, status=404)

    except Exception as e:
        logger.error(f"Error deleting asset {asset_id}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('Error deleting asset: {}').format(str(e))
        }, status=400)

def get_asset(request, asset_id):
    asset = get_object_or_404(InformationAsset, id=asset_id)
    asset.last_modified_by = request.user
    asset.save()
    
    # Format last modified datetime
    last_modified_datetime = timezone.localtime(asset.last_modified).strftime('%d-%m-%Y %H:%M:%S') if asset.last_modified else None
    last_modified_user = asset.last_modified_by.get_full_name() if asset.last_modified_by else None
    
    selection_map = {
        s.software_register_id: s
        for s in InformationAssetSoftwareSelection.objects.filter(information_asset=asset)
    }

    data = {
        'id': asset.id,
        'name': asset.name,
        'company': asset.company.id,
        'group': asset.group.id if asset.group else None,
        'asset_type': asset.asset_type.id if asset.asset_type else None,
        'description': asset.description,
        'location': asset.location,
        'confidentiality': asset.confidentiality.id if asset.confidentiality else None,
        'integrity': asset.integrity.id if asset.integrity else None,
        'availability': asset.availability.id if asset.availability else None,
        'registration_date': asset.registration_date.isoformat() if asset.registration_date else None,
        'deletion_date': asset.deletion_date.isoformat() if asset.deletion_date else None,
        'notes': asset.notes,
        'access_manage': asset.access_manage,
        'is_active': asset.is_active,
        'owners': [{'id': owner.id, 'name': owner.name} for owner in asset.owners.all()],
        'administrators': [{'id': admin.id, 'name': admin.name} for admin in asset.administrators.all()],
        'software_entries': [
            {
                'id': sw.id,
                'name': sw.name,
                'company': sw.company_id,
                'status': sw.status.get_name() if sw.status_id else '',
                'version': (selection_map[sw.id].selected_version if sw.id in selection_map else (sw.version_pattern or '')),
                'license_quantity': (selection_map[sw.id].selected_license_quantity if sw.id in selection_map else sw.license_quantity),
            }
            for sw in asset.software_entries.select_related('status', 'company').all()
        ],
        'last_modified': last_modified_datetime,
        'last_modified_by': last_modified_user,
    }
    return JsonResponse(data)


class AssetDatatableView(BaseDatatableView):
    model = InformationAsset
    columns = ['asset_id', 'name', 'company', 'group', 'description', 'location',
               'software', 'cia', 'registration_date',
               'deletion_date', 'notes', 'access_manage', 'actual', 'owners', 'administrators']

    order_columns = [
        'asset_id',
        'name',
        'company',  # Will be handled by get_ordering()
        'group_name',  # Using annotated field
        'description',
        'location',
        '',  # Software column - not sortable
        '',  # CIA column - not sortable
        'registration_date',
        'deletion_date',
        'notes',
        'access_manage',  # Sortable boolean field
        'actualization_date',  # Actual column
        'owners_list',  # owners (M2M) - custom sorting
        'administrators_list',  # administrators (M2M) - custom sorting
    ]

    def get_ordering(self):
        """
        Get parameters from the request and return list of list of fields
        by which we need to order.
        """
        order = []
        ordering_column = self.request.GET.get('order[0][column]', '')
        ordering_dir = self.request.GET.get('order[0][dir]', '')

        if ordering_column and ordering_dir:
            ordering_column = int(ordering_column)
            if ordering_column < len(self.order_columns):
                if self.order_columns[ordering_column] == 'company':
                    order.append(
                        '-company__name' if ordering_dir == 'desc'
                        else 'company__name'
                    )
                elif self.order_columns[ordering_column] == 'actualization_date':
                    # Sort by actualization date
                    order.append(
                        '-actualization_date' if ordering_dir == 'desc'
                        else 'actualization_date'
                    )
                elif self.order_columns[ordering_column] == 'owners_list':
                    # Sort by first owner's name via cabinet_user
                    order.append(
                        '-owners__cabinet_user__user__first_name' if ordering_dir == 'desc'
                        else 'owners__cabinet_user__user__first_name'
                    )
                elif self.order_columns[ordering_column] == 'administrators_list':
                    # Sort by first administrator's name via cabinet_user
                    order.append(
                        '-administrators__cabinet_user__user__first_name' if ordering_dir == 'desc'
                        else 'administrators__cabinet_user__user__first_name'
                    )
                elif self.order_columns[ordering_column]:
                    order.append(
                        f"-{self.order_columns[ordering_column]}"
                        if ordering_dir == 'desc'
                        else self.order_columns[ordering_column]
                    )

        if not order:
            order = ['-asset_id']  # Default ordering

        return order
    
    def ordering(self, qs):
        """Override ordering to handle Many-to-Many relationships properly"""
        ordering = self.get_ordering()
        if ordering:
            # If ordering includes owners or administrators, we need distinct() to avoid duplicates
            if any('owners__' in order or 'administrators__' in order for order in ordering):
                return qs.order_by(*ordering).distinct()
            else:
                return qs.order_by(*ordering)
        return qs
    def get_initial_queryset(self):
        current_language = get_language()[:2]
        qs = super().get_initial_queryset()

        # Get user's groups and access permissions (same logic as information_assets view)
        user_groups = self.request.user.groups.all()
        access_assets = AccessAssets.objects.filter(group__in=user_groups, has_access=True)
        allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()

        # Filter by allowed companies; staff/superuser see all if no group-based access
        if allowed_companies.exists():
            qs = qs.filter(company__in=allowed_companies)
        elif self.request.user.is_staff or self.request.user.is_superuser:
            # Admins see all assets when they have no group-based companies
            pass
        else:
            qs = qs.none()

        # Hide inactive assets by default (no deletion date + active flag); "Show Inactive Assets" shows the rest
        show_deleted = self.request.GET.get('showDeleted') == 'true'
        if not show_deleted:
            qs = qs.filter(deletion_date__isnull=True, is_active=True)

        # Annotate localized fields for display
        qs = qs.annotate(
            group_name=Case(
                When(group__isnull=True, then=Value('')),
                When(group__name_local__exact='', then=F('group__name')),
                default=F('group__name_local'),
                output_field=CharField()
            ),
            confidentiality_name=Case(
                When(confidentiality__isnull=True, then=Value('')),
                When(confidentiality__name_local__exact='', then=F('confidentiality__name')),
                default=F('confidentiality__name_local'),
                output_field=CharField()
            ),
            integrity_name=Case(
                When(integrity__isnull=True, then=Value('')),
                When(integrity__name_local__exact='', then=F('integrity__name')),
                default=F('integrity__name_local'),
                output_field=CharField()
            ),
            availability_name=Case(
                When(availability__isnull=True, then=Value('')),
                When(availability__name_local__exact='', then=F('availability__name')),
                default=F('availability__name_local'),
                output_field=CharField()
            )
        )

        return qs.select_related(
            'company',
            'group',
            'asset_type',
            'confidentiality',
            'integrity',
            'availability',
            'last_modified_by'
        ).prefetch_related(
            'owners',
            'administrators',
            'software_entries',
            Prefetch(
                'software_selections',
                queryset=InformationAssetSoftwareSelection.objects.only(
                    'information_asset_id', 'software_register_id', 'selected_version'
                ),
                to_attr='prefetched_software_selections',
            ),
        )

    def filter_queryset(self, qs):
        try:
            # Handle global search
            search = self.request.GET.get('search[value]', None)
            if search:
                qs = qs.filter(
                    Q(asset_id__icontains=search) |
                    Q(name__icontains=search) |
                    Q(company__name__icontains=search) |
                    Q(description__icontains=search) |
                    Q(location__icontains=search)
                )

            # Handle individual column filters (DataTable has checkbox at index 0)
            column_offset = 1
            for i, column in enumerate(self.columns):
                column_search = self.request.GET.get(
                    f'columns[{i + column_offset}][search][value]', ''
                )
                if not column_search:
                    continue

                if column == 'asset_id':
                    qs = qs.filter(asset_id__icontains=column_search)
                elif column == 'name':
                    qs = qs.filter(name__icontains=column_search)
                elif column == 'company':
                    qs = qs.filter(company__name__icontains=column_search)
                elif column == 'group':
                    val = column_search.strip()
                    if '/' in val:
                        g_part, t_part = val.split('/', 1)
                        g_part, t_part = g_part.strip(), t_part.strip()
                        if g_part:
                            qs = qs.filter(group_name__icontains=g_part)
                        if t_part:
                            qs = qs.filter(
                                Q(asset_type__name__icontains=t_part) |
                                Q(asset_type__name_local__icontains=t_part)
                            )
                    else:
                        qs = qs.filter(
                            Q(group_name__icontains=val) |
                            Q(asset_type__name__icontains=val) |
                            Q(asset_type__name_local__icontains=val)
                        )
                elif column == 'description':
                    qs = qs.filter(description__icontains=column_search)
                elif column == 'location':
                    qs = qs.filter(location__icontains=column_search)
                elif column == 'software':
                    qs = qs.filter(
                        Q(software_entries__name__icontains=column_search) |
                        Q(software_entries__version_pattern__icontains=column_search)
                    ).distinct()
                elif column == 'cia':
                    qs = qs.filter(
                        Q(confidentiality_name__icontains=column_search) |
                        Q(integrity_name__icontains=column_search) |
                        Q(availability_name__icontains=column_search)
                    )
                elif column == 'registration_date':
                    try:
                        search_date = parse_date(column_search)
                        if search_date:
                            qs = qs.filter(registration_date=search_date)
                    except ValueError:
                        pass
                elif column == 'deletion_date':
                    try:
                        search_date = parse_date(column_search)
                        if search_date:
                            qs = qs.filter(deletion_date=search_date)
                    except ValueError:
                        pass
                elif column == 'notes':
                    qs = qs.filter(notes__icontains=column_search)
                elif column == 'access_manage':
                    # Handle boolean search for access_manage
                    if column_search.lower() in ['true', 'yes', 'так', '1']:
                        qs = qs.filter(access_manage=True)
                    elif column_search.lower() in ['false', 'no', 'ні', '0']:
                        qs = qs.filter(access_manage=False)
                elif column == 'actual':
                    # Search by actualization date or user
                    if column_search.lower() in ['true', 'yes', 'так', '1', 'actualized', 'актуалізовано']:
                        qs = qs.filter(actualization_date__isnull=False)
                    elif column_search.lower() in ['false', 'no', 'ні', '0', 'not actualized', 'не актуалізовано']:
                        qs = qs.filter(actualization_date__isnull=True)
                    else:
                        # Try to search by actualized_by user name
                        qs = qs.filter(
                            Q(actualized_by__first_name__icontains=column_search) |
                            Q(actualized_by__last_name__icontains=column_search) |
                            Q(actualized_by__username__icontains=column_search)
                        ).distinct()
                elif column == 'owners':
                    # Search in owners' names, emails, and departments via cabinet_user
                    qs = qs.filter(
                        Q(owners__cabinet_user__user__first_name__icontains=column_search) |
                        Q(owners__cabinet_user__user__last_name__icontains=column_search) |
                        Q(owners__cabinet_user__user__email__icontains=column_search) |
                        Q(owners__cabinet_user__department__icontains=column_search)
                    ).distinct()
                elif column == 'administrators':
                    # Search in administrators' names, emails, and departments via cabinet_user
                    qs = qs.filter(
                        Q(administrators__cabinet_user__user__first_name__icontains=column_search) |
                        Q(administrators__cabinet_user__user__last_name__icontains=column_search) |
                        Q(administrators__cabinet_user__user__email__icontains=column_search) |
                        Q(administrators__cabinet_user__department__icontains=column_search)
                    ).distinct()

            return qs.distinct()

        except Exception as e:
            logger.error(f"Error in filter_queryset: {str(e)}", exc_info=True)
            raise

    def _build_software_list(self, item):
        """Return list of {name, version} using selected_version from InformationAssetSoftwareSelection."""
        selections = getattr(item, 'prefetched_software_selections', None)
        if selections is not None:
            version_map = {sel.software_register_id: sel.selected_version for sel in selections}
        else:
            version_map = {}
        return [
            {'name': sw.name, 'version': version_map.get(sw.id) or ''}
            for sw in item.software_entries.all()
        ]

    def _build_actual_dict(self, item, user_cabinet_users):
        """Build actual column dict; use getattr for new fields in case migration not applied."""
        marked_at = getattr(item, 'marked_no_longer_actual_at', None)
        marked_comment = getattr(item, 'marked_no_longer_comment', None)
        return {
            'date': timezone.localtime(item.actualization_date).strftime('%d-%m-%Y %H:%M:%S') if item.actualization_date else None,
            'user': item.actualized_by.get_full_name() if item.actualized_by else None,
            'has_actualization': bool(item.actualization_date),
            'marked_no_longer_at': timezone.localtime(marked_at).strftime('%d-%m-%Y %H:%M:%S') if marked_at else None,
            'marked_no_longer_comment': (marked_comment or '').strip() or None,
            'is_owner': item.owners.filter(cabinet_user=user_cabinet_users.get(item.company.id)).exists() if user_cabinet_users.get(item.company.id) else False
        }

    def prepare_results(self, qs):
        try:
            logger.debug(f"Preparing results for {qs.count()} items")
            current_language = get_language()[:2]
            user = self.request.user
            # Отримуємо CabinetUser для поточного користувача для перевірки власника
            user_cabinet_users = {}
            for company in Company.objects.filter(informationasset__in=qs).distinct():
                cabinet_user = CabinetUser.objects.filter(user=user, company=company).first()
                if cabinet_user:
                    user_cabinet_users[company.id] = cabinet_user
            
            data = []
            for item in qs:
                criticality = item.get_criticality()
                group_display_name = (item.group.get_name() if item.group else '')
                asset_type_name = (item.asset_type.get_name() if item.asset_type else '')
                group_type = (f"{group_display_name}/{asset_type_name}"
                              if group_display_name and asset_type_name else
                              group_display_name or asset_type_name)

                # Convert owners and admins to serializable dictionaries with safe fallbacks
                owners_data = []
                for owner in item.owners.all():
                    try:
                        owner_data = {
                            'name': owner.name if owner.cabinet_user else str(owner),
                            'department': str(owner.department) if owner.department else '',
                            'position': str(owner.position) if owner.position else '',
                            'email': owner.email or '',
                            'phone': owner.phone or ''
                        }
                        owners_data.append(owner_data)
                    except Exception as e:
                        logger.warning(f"Error processing owner data: {str(e)}")
                        owners_data.append({
                            'name': 'Unknown',
                            'department': '',
                            'position': '',
                            'email': '',
                            'phone': ''
                        })

                admins_data = []
                for admin in item.administrators.all():
                    try:
                        admin_data = {
                            'name': admin.name if admin.cabinet_user else str(admin),
                            'department': str(admin.department) if admin.department else '',
                            'position': str(admin.position) if admin.position else '',
                            'email': admin.email or '',
                            'phone': admin.phone or ''
                        }
                        admins_data.append(admin_data)
                    except Exception as e:
                        logger.warning(f"Error processing administrator data: {str(e)}")
                        admins_data.append({
                            'name': 'Unknown',
                            'department': '',
                            'position': '',
                            'email': '',
                            'phone': ''
                        })

                data.append({
                    'id': item.id,
                    'asset_id': item.asset_id,
                    'name': item.name,
                    'company': item.company.name if item.company else '',
                    'group': group_type,
                    'description': item.description,
                    'location': item.location,
                    'software': self._build_software_list(item),
                    'confidentiality': {
                        'text': item.confidentiality.get_name() if item.confidentiality else '',
                        'color': item.confidentiality.color if item.confidentiality else '#CCCCCC',
                        'description': item.confidentiality.get_description_confid() if item.confidentiality else ''
                    },
                    'integrity': {
                        'text': item.integrity.get_name() if item.integrity else '',
                        'color': item.integrity.color if item.integrity else '#CCCCCC',
                        'description': item.integrity.get_description_integ() if item.integrity else ''
                    },
                    'availability': {
                        'text': item.availability.get_name() if item.availability else '',
                        'color': item.availability.color if item.availability else '#CCCCCC',
                        'description': item.availability.get_description_avail() if item.availability else ''
                    },
                    'registration_date': (item.registration_date.strftime('%Y-%m-%d')
                                          if item.registration_date else ''),
                    'deletion_date': (item.deletion_date.strftime('%Y-%m-%d')
                                      if item.deletion_date else ''),
                    'is_deleted': bool(item.deletion_date),
                    'is_active': item.is_active,
                    'inactive_date': (
                        item.deletion_date.strftime('%Y-%m-%d') if item.deletion_date
                        else (
                            timezone.localtime(item.last_modified).strftime('%Y-%m-%d')
                            if not item.is_active else ''
                        )
                    ),
                    'notes': item.notes,
                    'access_manage': item.access_manage,
                    'actual': self._build_actual_dict(item, user_cabinet_users),
                    'owners': json.dumps(owners_data),
                    'administrators': json.dumps(admins_data),
                    'last_modified': {
                        'datetime': timezone.localtime(item.last_modified).strftime('%d-%m-%Y %H:%M:%S'),
                        'user': item.last_modified_by.get_full_name() if item.last_modified_by else ''
                    },
                    'criticality': {
                        'text': criticality['name'],
                        'cost': criticality['cost'],
                        'color': criticality['color']
                    }
                })
            logger.debug(f"Successfully prepared {len(data)} items")
            return data

        except Exception as e:
            logger.error(f"Error in prepare_results: {str(e)}", exc_info=True)
            logger.error(f"Queryset count: {qs.count() if qs else 'None'}")
            # Return empty data to prevent complete failure
            return []

    def get_context_data(self, *args, **kwargs):
        try:
            context = super().get_context_data(*args, **kwargs)
            context['can_edit'] = self.request.user.has_perm('app_asset.change_informationasset')
            return context

        except Exception as e:
            logger.error(f"Error in get_context_data: {str(e)}", exc_info=True)
            raise



def get_asset_details(request, asset_id):
    logger.info(f"Fetching details for asset with ID: {asset_id}")
    try:
        asset = get_object_or_404(InformationAsset, id=asset_id)
        owners = [{"id": owner.id, "name": owner.name} for owner in asset.owners.all()]
        administrators = [{"id": admin.id, "name": admin.name} for admin in asset.administrators.all()]

        logger.info(f"Asset details fetched successfully. Owners: {len(owners)}, Administrators: {len(administrators)}")
        return JsonResponse({
            "owners": owners,
            "administrators": administrators
        })
    except Exception as e:
        logger.error(f"Error fetching asset details: {str(e)}")
        return JsonResponse({"error": str(e)}, status=400)


@login_required
def get_asset_history(request, asset_id):
    """Отримати повну історію змін активу"""
    try:
        asset = get_object_or_404(InformationAsset, id=asset_id)
        history_records = AssetHistory.objects.filter(asset=asset).order_by('-timestamp')
        
        history_data = []
        for record in history_records:
            history_data.append({
                'id': record.id,
                'timestamp': timezone.localtime(record.timestamp).strftime('%d-%m-%Y %H:%M:%S'),
                'action': record.action,
                'action_display': record.get_action_display(),
                'action_by': record.get_action_by_name(),
                'action_by_id': record.action_by.id if record.action_by else None,
                'details': record.details,
                'changes': record.changes,
            })
        
        return JsonResponse({
            'status': 'success',
            'asset_id': asset.id,
            'asset_name': asset.name,
            'asset_asset_id': asset.asset_id,
            'history': history_data
        })
    except Exception as e:
        logger.error(f"Error fetching asset history: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@csrf_exempt
@login_required
@require_POST
def actualize_asset(request, asset_id):
    """Актуалізувати актив власником або позначити як неактуальний.
    action: 'actualize' (default) | 'mark_inactive'
    """
    try:
        asset = get_object_or_404(InformationAsset, id=asset_id)
        action = request.POST.get('action', 'actualize')  # actualize | mark_inactive

        # Перевірка, чи користувач є власником активу
        user = request.user
        logger.info(f"Actualize/mark_inactive attempt by user {user.username} (ID: {user.id}) for asset {asset_id}, action={action}")

        # Перевірка доступу до компанії
        user_groups = user.groups.all()
        access_assets = AccessAssets.objects.filter(group__in=user_groups, has_access=True)
        allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()

        if asset.company not in allowed_companies:
            logger.warning(f"User {user.username} does not have access to company {asset.company.id}")
            return JsonResponse({
                'status': 'error',
                'message': _('You do not have access to this asset')
            }, status=403)

        # Перевірка, чи користувач є власником через CabinetUser
        cabinet_user = CabinetUser.objects.filter(user=user, company=asset.company).first()

        if not cabinet_user:
            logger.warning(f"User {user.username} does not have CabinetUser for company {asset.company.id}")
            owners_list = [f"{owner.name} (cabinet_user_id: {owner.cabinet_user.id if owner.cabinet_user else 'None'})"
                          for owner in asset.owners.all()]
            logger.info(f"Asset {asset_id} owners: {owners_list}")
            return JsonResponse({
                'status': 'error',
                'message': _('You are not authorized to actualize this asset. You need to be an owner of this asset.')
            }, status=403)

        # Перевірка, чи користувач є власником
        is_owner = asset.owners.filter(cabinet_user=cabinet_user).exists()

        logger.info(f"User {user.username} is owner: {is_owner}, cabinet_user: {cabinet_user.id}")
        logger.info(f"Asset {asset_id} has {asset.owners.count()} owners")

        if not is_owner:
            owners_list = [f"{owner.name} (cabinet_user_id: {owner.cabinet_user.id if owner.cabinet_user else 'None'})"
                          for owner in asset.owners.all()]
            logger.warning(f"User {user.username} (cabinet_user_id: {cabinet_user.id}) is not an owner of asset {asset_id}")
            logger.info(f"Asset {asset_id} owners: {owners_list}")
            return JsonResponse({
                'status': 'error',
                'message': _('Only asset owners can actualize the asset')
            }, status=403)

        if action == 'mark_inactive':
            # Позначити актив як неактуальний — очистити дату актуалізації, зберегти дату та опційний коментар
            comment = (request.POST.get('comment') or '').strip()
            asset.actualization_date = None
            asset.actualized_by = None
            asset.marked_no_longer_actual_at = timezone.now()
            asset.marked_no_longer_comment = comment or ''
            asset.save()

            changes = {'action': 'mark_inactive', 'actualization_cleared': True}
            if comment:
                changes['comment'] = comment
            AssetHistory.objects.create(
                asset=asset,
                action=AssetHistory.ACTION_MODIFIED,
                action_by=user,
                details=_("Asset marked as no longer actual by owner"),
                changes=changes
            )

            return JsonResponse({
                'status': 'success',
                'message': _('Asset marked as no longer actual'),
                'actualization_date': None,
                'actualized_by': None
            })
        else:
            # Актуалізація активу (default)
            asset.actualization_date = timezone.now()
            asset.actualized_by = user
            asset.marked_no_longer_actual_at = None
            asset.marked_no_longer_comment = ''
            asset.save()

            AssetHistory.objects.create(
                asset=asset,
                action=AssetHistory.ACTION_MODIFIED,
                action_by=user,
                details=_("Asset actualized by owner"),
                changes={
                    'actualization_date': timezone.localtime(asset.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
                    'actualized_by': user.get_full_name() or user.username
                }
            )

            return JsonResponse({
                'status': 'success',
                'message': _('Asset actualized successfully'),
                'actualization_date': timezone.localtime(asset.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
                'actualized_by': user.get_full_name() or user.username
            })
        
    except Exception as e:
        logger.error(f"Error actualizing asset: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@csrf_exempt
@login_required
@require_POST
def bulk_actualize_assets(request):
    """Масово актуалізувати активи власником"""
    try:
        data = json.loads(request.body)
        asset_ids = data.get('asset_ids', [])
        
        if not asset_ids:
            return JsonResponse({
                'status': 'error',
                'message': _('No assets selected')
            }, status=400)
        
        user = request.user
        logger.info(f"Bulk actualize attempt by user {user.username} (ID: {user.id}) for {len(asset_ids)} assets")
        
        # Отримуємо активи
        assets = InformationAsset.objects.filter(id__in=asset_ids)
        
        if assets.count() != len(asset_ids):
            return JsonResponse({
                'status': 'error',
                'message': _('Some assets were not found')
            }, status=404)
        
        # Перевірка доступу до компаній
        user_groups = user.groups.all()
        access_assets = AccessAssets.objects.filter(group__in=user_groups, has_access=True)
        allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()
        
        actualized_count = 0
        skipped_count = 0
        errors = []
        
        for asset in assets:
            try:
                # Перевірка доступу до компанії
                if asset.company not in allowed_companies:
                    skipped_count += 1
                    errors.append(f"{asset.asset_id}: {_('No access to company')}")
                    continue
                
                # Перевірка, чи користувач є власником через CabinetUser
                cabinet_user = CabinetUser.objects.filter(user=user, company=asset.company).first()
                
                if not cabinet_user:
                    skipped_count += 1
                    errors.append(f"{asset.asset_id}: {_('No CabinetUser for company')}")
                    continue
                
                # Перевірка, чи користувач є власником
                is_owner = asset.owners.filter(cabinet_user=cabinet_user).exists()
                
                if not is_owner:
                    skipped_count += 1
                    errors.append(f"{asset.asset_id}: {_('Not an owner')}")
                    continue
                
                # Актуалізація активу
                asset.actualization_date = timezone.now()
                asset.actualized_by = user
                asset.save()
                
                # Логування в історію
                AssetHistory.objects.create(
                    asset=asset,
                    action=AssetHistory.ACTION_MODIFIED,
                    action_by=user,
                    details=_("Asset actualized by owner (bulk operation)"),
                    changes={
                        'actualization_date': timezone.localtime(asset.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
                        'actualized_by': user.get_full_name() or user.username
                    }
                )
                
                actualized_count += 1
                
            except Exception as e:
                logger.error(f"Error actualizing asset {asset.id}: {str(e)}")
                skipped_count += 1
                errors.append(f"{asset.asset_id}: {str(e)}")
        
        if actualized_count == 0:
            return JsonResponse({
                'status': 'error',
                'message': _('No assets were actualized. You may not be an owner of the selected assets.'),
                'errors': errors[:10]  # Перші 10 помилок
            }, status=400)
        
        message = _('Successfully actualized {} asset(s)').format(actualized_count)
        if skipped_count > 0:
            message += f". {skipped_count} {_('skipped')}"
        
        return JsonResponse({
            'status': 'success',
            'message': message,
            'actualized_count': actualized_count,
            'skipped_count': skipped_count,
            'errors': errors[:10] if errors else []
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': _('Invalid JSON data')
        }, status=400)
    except Exception as e:
        logger.error(f"Error in bulk actualize: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


def get_asset_owners(request, asset_id):
    asset = get_object_or_404(InformationAsset, id=asset_id)
    owners = asset.owners.all()
    return JsonResponse({
        'owners': [{
            'id': owner.id,
            'name': owner.name,
            'department': str(owner.department) if owner.department else '',
            'position': str(owner.position) if owner.position else '',
            'company': owner.company.name
        } for owner in owners]
    })


@require_POST
@login_required
def add_asset_owner(request):
    try:
        cabinet_user_id = request.POST.get('cabinet_user_id')
        company_id = request.POST.get('company_id')

        if not cabinet_user_id or not company_id:
            return JsonResponse({'status': 'error', 'message': _('Required fields missing')}, status=400)

        cabinet_user = CabinetUser.objects.select_related('user', 'company').get(id=cabinet_user_id)
        company = Company.objects.get(id=company_id)

        # Create owner without department/position
        owner = AssetOwner.objects.create(
            cabinet_user=cabinet_user,
            company=company
        )

        return JsonResponse({
            'status': 'success',
            'message': _('Owner added successfully'),
            'owner': {
                'id': owner.id,
                'name': owner.name,
                'department': str(cabinet_user.department) if cabinet_user.department else '',
                'position': str(cabinet_user.position) if cabinet_user.position else '',
                'email': owner.email or '',
                'phone': owner.phone or '',
                'company': {'id': company.id, 'name': company.name}
            }
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
def get_or_create_asset_owner(request):
    """Get or create AssetOwner by cabinet_user_id and company_id (for Software/External Media Register owner selection)."""
    cabinet_user_id = request.GET.get('cabinet_user_id') or request.POST.get('cabinet_user_id')
    company_id = request.GET.get('company_id') or request.POST.get('company_id')
    if not cabinet_user_id or not company_id:
        return JsonResponse({'error': _('cabinet_user_id and company_id required')}, status=400)
    cabinet_user = get_object_or_404(CabinetUser.objects.select_related('user', 'company'), id=cabinet_user_id)
    company = get_object_or_404(Company, id=company_id)
    if cabinet_user.company_id != company.id:
        return JsonResponse({'error': _('Cabinet user company mismatch')}, status=400)
    owner, created = AssetOwner.objects.get_or_create(cabinet_user=cabinet_user, company=company)
    return JsonResponse({
        'id': owner.id,
        'name': owner.name,
        'department': str(owner.department or ''),
        'position': str(owner.position or ''),
        'email': owner.email or '',
        'phone': owner.phone or '',
        'company': {'id': company.id, 'name': company.name},
    })


@require_POST
def edit_asset_owner(request):
    try:
        owner = get_object_or_404(AssetOwner, id=request.POST.get('owner_id'))
        owner.name = request.POST.get('name')
        owner.department = request.POST.get('department')
        owner.position = request.POST.get('position')
        owner.email = request.POST.get('email')
        owner.phone = request.POST.get('phone')
        owner.save()
        return JsonResponse({'status': 'success', 'message': 'Owner updated successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def remove_asset_owner(request, asset_id, owner_id):
    asset = get_object_or_404(InformationAsset, id=asset_id)
    owner = get_object_or_404(AssetOwner, id=owner_id)
    asset.owners.remove(owner)
    return JsonResponse({'status': 'success', 'message': 'Owner removed successfully'})




def get_owner(request, owner_id):
    owner = get_object_or_404(AssetOwner, id=owner_id)
    return JsonResponse({
        'id': owner.id,
        'name': owner.name,
        'department': str(owner.department) if owner.department else '',
        'position': str(owner.position) if owner.position else '',
        'email': owner.email,
        'phone': owner.phone
    })

def get_asset_administrators(request, asset_id):
    asset = get_object_or_404(InformationAsset, id=asset_id)
    administrators = asset.administrators.all()
    return JsonResponse({
        'administrators': [{
            'id': admin.id,
            'name': admin.name,
            'department': str(admin.department) if admin.department else '',
            'position': str(admin.position) if admin.position else '',
            'company': admin.company.name
        } for admin in administrators]
    })


@require_POST
@login_required
def add_asset_administrator(request):
    try:
        cabinet_user_id = request.POST.get('cabinet_user_id')
        company_id = request.POST.get('company_id')

        if not cabinet_user_id or not company_id:
            return JsonResponse({'status': 'error', 'message': _('Required fields missing')}, status=400)

        cabinet_user = CabinetUser.objects.select_related('user', 'company').get(id=cabinet_user_id)
        company = Company.objects.get(id=company_id)

        # Create admin without department/position
        admin = AssetAdministrator.objects.create(
            cabinet_user=cabinet_user,
            company=company
        )

        return JsonResponse({
            'status': 'success',
            'message': _('Administrator added successfully'),
            'administrator': {
                'id': admin.id,
                'name': admin.name,
                'department': str(cabinet_user.department) if cabinet_user.department else '',
                'position': str(cabinet_user.position) if cabinet_user.position else '',
                'email': admin.email or '',
                'phone': admin.phone or '',
                'company': {'id': company.id, 'name': company.name}
            }
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def edit_asset_administrator(request):
    try:
        admin = get_object_or_404(AssetAdministrator, id=request.POST.get('admin_id'))
        admin.name = request.POST.get('name')
        admin.department = request.POST.get('department')
        admin.position = request.POST.get('position')
        admin.email = request.POST.get('email')
        admin.phone = request.POST.get('phone')
        admin.save()
        return JsonResponse({'status': 'success', 'message': 'Administrator updated successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
def remove_asset_administrator(request, asset_id, admin_id):
    asset = get_object_or_404(InformationAsset, id=asset_id)
    admin = get_object_or_404(AssetAdministrator, id=admin_id)
    asset.administrators.remove(admin)
    return JsonResponse({'status': 'success', 'message': 'Administrator removed successfully'})

@require_POST
@user_passes_test(has_access_to_assets)
def delete_asset_owners(request):
    try:
        owner_id = request.POST.get('owner_id')
        if not owner_id:
            return JsonResponse({
                'status': 'error',
                'message': _('Owner ID is required')
            }, status=400)

        owner = get_object_or_404(AssetOwner, id=owner_id)
        owner.delete()

        return JsonResponse({
            'status': 'success',
            'message': _('Owner deleted successfully')
        })
    except Exception as e:
        logger.error(f"Error deleting owner: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@require_POST
@user_passes_test(has_access_to_assets)
def delete_asset_administrators(request):
    try:
        admin_id = request.POST.get('admin_id')
        if not admin_id:
            return JsonResponse({
                'status': 'error',
                'message': _('Administrator ID is required')
            }, status=400)

        admin = get_object_or_404(AssetAdministrator, id=admin_id)
        admin.delete()

        return JsonResponse({
            'status': 'success',
            'message': _('Administrator deleted successfully')
        })
    except Exception as e:
        logger.error(f"Error deleting administrator: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

def get_administrator(request, admin_id):
    admin = get_object_or_404(AssetAdministrator, id=admin_id)
    return JsonResponse({
        'id': admin.id,
        'name': admin.name,
        'department': str(admin.department) if admin.department else '',
        'position': str(admin.position) if admin.position else '',
        'email': admin.email,
        'phone': admin.phone
    })

def get_all_administrators(request):
    administrators = AssetAdministrator.objects.all()
    data = [{
        'id': admin.id,
        'name': admin.name,
        'department': str(admin.department) if admin.department else '',
        'position': str(admin.position) if admin.position else ''
    } for admin in administrators]
    return JsonResponse({'administrators': data})

@login_required
def get_criticality_levels(request):
    try:
        # Use language from request (e.g. Add/Edit Asset modals) so response matches current UI language
        from django.utils import translation
        lang = (request.GET.get('lang') or get_language() or '')[:2]
        if lang:
            translation.activate(lang)
        allowed_companies = _get_allowed_companies_for_asset_user(request.user)
        requested_company_id = request.GET.get('company_id')
        company_id = requested_company_id
        if company_id:
            if not allowed_companies.filter(id=company_id).exists():
                return JsonResponse({'error': _('Permission denied for selected company')}, status=403)
        else:
            company_id = allowed_companies.values_list('id', flat=True).first()

        if company_id:
            levels = _get_company_criticality_queryset(company_id)
        else:
            levels = CriticalityLevel.objects.filter(
                is_active=True,
                company__isnull=True,
            ).prefetch_related('translations__country').order_by('display_order', 'cost', 'name')
        data = []
        for level in levels:
            data.append({
                'id': level.id,
                'name': level.get_name(),  # Use get_name() to get translation for current language
                'cost': level.cost,
                'color': level.color,
                'description_confid': level.get_description_confid(),
                'description_avail': level.get_description_avail(),
                'description_integ': level.get_description_integ(),
            })
        return JsonResponse({'levels': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_company_people(request):
    try:
        company_id = request.GET.get('company_id')
        if not company_id:
            return JsonResponse({'error': 'Company ID is required'}, status=400)

        cabinet_users = CabinetUser.objects.filter(company_id=company_id).select_related('user')

        existing_owners = AssetOwner.objects.filter(company_id=company_id).values_list('cabinet_user_id', flat=True)
        existing_admins = AssetAdministrator.objects.filter(company_id=company_id).values_list('cabinet_user_id', flat=True)

        people_data = {
            'owners': [
                {
                    'id': cu.id,
                    'name': f"{cu.user.first_name} {cu.user.last_name}",
                    'department': str(cu.department) if cu.department else '',
                    'position': str(cu.position) if cu.position else '',
                    'email': cu.user.email,
                    'isOwner': cu.id in existing_owners
                } for cu in cabinet_users
            ],
            'administrators': [
                {
                    'id': cu.id,
                    'name': f"{cu.user.first_name} {cu.user.last_name}",
                    'department': str(cu.department) if cu.department else '',
                    'position': str(cu.position) if cu.position else '',
                    'email': cu.user.email,
                    'isAdmin': cu.id in existing_admins
                } for cu in cabinet_users
            ]
        }

        return JsonResponse(people_data)

    except Exception as e:
        logger.error(f"Error in get_company_people: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def check_asset_permissions(request):
    action = request.GET.get('action')
    permissions = get_user_asset_permissions(request.user)
    return JsonResponse({action: permissions.get(f'can_{action}', False)})



def get_asset_types(request):
    group_id = request.GET.get('group_id')
    current_language = get_language()[:2]

    if group_id:
        types = AssetType.objects.filter(group_id=group_id, is_active=True).order_by('display_order', 'name')
        
        # Convert to list with localized names using get_name() method
        types_list = []
        for type_obj in types:
            types_list.append({
                'id': type_obj.id,
                'name': type_obj.name,
                'name_local': type_obj.name_local,
                'code': type_obj.code,
                'color': type_obj.color,
                'localized_name': type_obj.get_name()  # Use get_name() to get translation for current language
            })

        return JsonResponse({'types': types_list})

    return JsonResponse({'types': []})


def get_localized_name(obj, field_prefix='name'):
    """
    Get localized name for an object based on current language.

    Args:
        obj: Object that has localized name fields or get_name() method
        field_prefix (str): Prefix for the localized field (default: 'name')

    Returns:
        str: Localized name or empty string if object is None
    """
    if not obj:
        return ''

    # For AssetGroup and AssetType, use get_name() method which handles translations
    if hasattr(obj, 'get_name') and callable(getattr(obj, 'get_name')):
        return obj.get_name()

    from django.utils.translation import get_language
    current_language = get_language()[:2]

    # Try to get name in current language
    localized_field = f'{field_prefix}_{current_language}'
    name = getattr(obj, localized_field, None)

    # If empty, fallback to Ukrainian
    if not name:
        name = getattr(obj, f'{field_prefix}_uk', '')

    return name or ''

def export_assets_to_excel(request):
    current_language = get_language()[:2]

    # Get user permissions
    user_groups = request.user.groups.all()
    access_assets = AccessAssets.objects.filter(group__in=user_groups, can_edit=True)
    allowed_companies = Company.objects.filter(access_assets__in=access_assets).distinct()

    # Get assets for allowed companies
    assets = InformationAsset.objects.filter(
        company__in=allowed_companies
    ).select_related(
        'company', 'group', 'asset_type',
        'confidentiality', 'integrity', 'availability',
        'actualized_by'
    ).prefetch_related(
        Prefetch('owners', queryset=AssetOwner.objects.all()),
        Prefetch('administrators', queryset=AssetAdministrator.objects.all()),
        Prefetch('history', queryset=AssetHistory.objects.all().order_by('-timestamp'))
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    # Add headers with translations
    headers = [
        _('Asset ID'), _('Name'), _('Company'), _('Group/Type'), _('Description'), _('Location'),
        _('Confidentiality'), _('Integrity'), _('Availability'), _('Registration Date'),
        _('Deletion Date'), _('Notes'), _('Access Manage'), _('Actualization Date'), _('Actualized By'),
        _('Owners'), _('Administrators'), _('Asset History')
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Add data
    for row_num, asset in enumerate(assets, 2):
        # Basic fields
        ws.cell(row=row_num, column=1, value=asset.asset_id)
        ws.cell(row=row_num, column=2, value=asset.name)
        ws.cell(row=row_num, column=3, value=asset.company.name if asset.company else '')

        # Group/Type
        group_type = ''
        if asset.group and asset.asset_type:
            group_name = get_localized_name(asset.group)
            type_name = get_localized_name(asset.asset_type)
            if group_name and type_name:
                group_type = f"{group_name}/{type_name}"
        ws.cell(row=row_num, column=4, value=group_type)

        # Other basic fields
        ws.cell(row=row_num, column=5, value=asset.description)
        ws.cell(row=row_num, column=6, value=asset.location)

        # Criticality levels
        ws.cell(row=row_num, column=7,
                value=get_localized_name(asset.confidentiality, 'critical_name') if asset.confidentiality else '')
        ws.cell(row=row_num, column=8,
                value=get_localized_name(asset.integrity, 'critical_name') if asset.integrity else '')
        ws.cell(row=row_num, column=9,
                value=get_localized_name(asset.availability, 'critical_name') if asset.availability else '')

        # Dates
        ws.cell(row=row_num, column=10,
                value=asset.registration_date.strftime('%Y-%m-%d') if asset.registration_date else '')
        ws.cell(row=row_num, column=11,
                value=asset.deletion_date.strftime('%Y-%m-%d') if asset.deletion_date else '')

        ws.cell(row=row_num, column=12, value=asset.notes)
        
        # Access Manage
        ws.cell(row=row_num, column=13, value='Yes' if asset.access_manage else 'No')

        # Actualization
        if asset.actualization_date:
            actual_date = timezone.localtime(asset.actualization_date).strftime('%d-%m-%Y %H:%M:%S')
            ws.cell(row=row_num, column=14, value=actual_date)
        else:
            ws.cell(row=row_num, column=14, value='')
        
        if asset.actualized_by:
            actualized_by = asset.actualized_by.get_full_name() or asset.actualized_by.username
            ws.cell(row=row_num, column=15, value=actualized_by)
        else:
            ws.cell(row=row_num, column=15, value='')

        # Owners
        owners_text = '; '.join([
            f"{owner.name} ({owner.department}/{owner.position})"
            for owner in asset.owners.all()
        ])
        ws.cell(row=row_num, column=16, value=owners_text)

        # Administrators
        admins_text = '; '.join([
            f"{admin.name} ({admin.department}/{admin.position})"
            for admin in asset.administrators.all()
        ])
        ws.cell(row=row_num, column=17, value=admins_text)
        
        # Asset History
        history_items = []
        for history in asset.history.all()[:50]:  # Обмежуємо до 50 останніх записів
            action_name = history.get_action_display()  # Використовуємо локалізовану назву
            timestamp = timezone.localtime(history.timestamp).strftime('%d-%m-%Y %H:%M:%S')
            action_by = history.action_by.get_full_name() if history.action_by else ''
            details = history.details or ''
            
            history_text = f"{timestamp} - {action_name}"
            if action_by:
                history_text += f" ({action_by})"
            if details:
                # Обмежуємо довжину details, щоб не перевантажувати Excel
                details_short = details[:200] + '...' if len(details) > 200 else details
                history_text += f": {details_short}"
            
            history_items.append(history_text)
        
        history_text = ' | '.join(history_items)
        ws.cell(row=row_num, column=18, value=history_text)

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value else '') for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=assets.xlsx'
    wb.save(response)
    return response


@login_required
def asset_guide(request):
    """Return JSON { content: html } for the Information Assets guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = AssetGuide.objects.first()
    if guide:
        if country:
            trans = AssetGuideTranslation.objects.filter(asset_guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = AssetGuideTranslation.objects.filter(asset_guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def asset_guide_translate(request):
    """API for AI translation of guide content (admin). Same contract as compliance translate."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
@require_http_methods(["GET"])
def software_guide(request):
    """Return JSON { content: html } for the Software Register guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = SoftwareGuide.objects.first()
    if guide:
        if country:
            trans = SoftwareGuideTranslation.objects.filter(software_guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = SoftwareGuideTranslation.objects.filter(software_guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def software_guide_translate(request):
    """API for AI translation of software guide content (admin). Same contract as compliance translate."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
@require_http_methods(["GET"])
def external_media_guide(request):
    """Return JSON { content: html } for the External Media Register guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = ExternalMediaGuide.objects.first()
    if guide:
        if country:
            trans = ExternalMediaGuideTranslation.objects.filter(external_media_guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = ExternalMediaGuideTranslation.objects.filter(external_media_guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def external_media_guide_translate(request):
    """API for AI translation of external media guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
def get_all_asset_owners(request):
    company_id = request.GET.get('company_id')
    if not company_id:
        return JsonResponse({'error': 'Company ID required'}, status=400)

    try:
        owners = AssetOwner.objects.filter(company_id=company_id)
        data = []

        for owner in owners:
            try:
                owner_data = {
                    'id': owner.id,
                    # Fallback logic for name
                    'name': (owner.cabinet_user.user.get_full_name()
                             if owner.cabinet_user and owner.cabinet_user.user
                             else owner.name if hasattr(owner, 'name')
                    else "Unknown User"),
                    # Safe conversion of department and position
                    'department': str(owner.department) if owner.department else '',
                    'position': str(owner.position) if owner.position else '',
                    # Company information with safety checks
                    'company': {
                        'id': owner.company.id,
                        'name': owner.company.name
                    } if owner.company else None,
                    # Basic fields with defaults
                    'email': (owner.cabinet_user.user.email
                              if owner.cabinet_user and owner.cabinet_user.user
                              else owner.email if hasattr(owner, 'email')
                    else ''),
                    'phone': owner.phone if hasattr(owner, 'phone') else ''
                }
                data.append(owner_data)
            except Exception as e:
                logger.warning(f"Error processing owner {owner.id}: {str(e)}")
                # Include minimal data if there's an error
                data.append({
                    'id': owner.id,
                    'name': "Unknown User",
                    'department': '',
                    'position': '',
                    'company': {'id': owner.company.id, 'name': owner.company.name} if owner.company else None,
                    'email': '',
                    'phone': ''
                })

        return JsonResponse(data, safe=False)
    except Exception as e:
        logger.error(f"Error getting owners: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_all_asset_administrators(request):
    company_id = request.GET.get('company_id')
    if not company_id:
        return JsonResponse({'error': 'Company ID required'}, status=400)

    try:
        admins = AssetAdministrator.objects.filter(company_id=company_id)
        data = []

        for admin in admins:
            try:
                admin_data = {
                    'id': admin.id,
                    # Fallback logic for name
                    'name': (admin.cabinet_user.user.get_full_name()
                            if admin.cabinet_user and admin.cabinet_user.user
                            else admin.name if hasattr(admin, 'name')
                            else "Unknown User"),
                    # Safe conversion of department and position
                    'department': str(admin.department) if admin.department else '',
                    'position': str(admin.position) if admin.position else '',
                    # Company information with safety checks
                    'company': {
                        'id': admin.company.id,
                        'name': admin.company.name
                    } if admin.company else None,
                    # Basic fields with defaults
                    'email': (admin.cabinet_user.user.email
                             if admin.cabinet_user and admin.cabinet_user.user
                             else admin.email if hasattr(admin, 'email')
                             else ''),
                    'phone': admin.phone if hasattr(admin, 'phone') else ''
                }
                data.append(admin_data)
            except Exception as e:
                logger.warning(f"Error processing administrator {admin.id}: {str(e)}")
                # Include minimal data if there's an error
                data.append({
                    'id': admin.id,
                    'name': "Unknown User",
                    'department': '',
                    'position': '',
                    'company': {'id': admin.company.id, 'name': admin.company.name} if admin.company else None,
                    'email': '',
                    'phone': ''
                })

        return JsonResponse(data, safe=False)
    except Exception as e:
        logger.error(f"Error getting administrators: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_asset_type(request):
    try:
        type_id = request.GET.get('type_id')
        asset_type = get_object_or_404(AssetType, id=type_id)
        return JsonResponse({
            'status': 'success',
            'type': {
                'id': asset_type.id,
                'name': asset_type.get_name(),
                'name_local': asset_type.name_local,
                'code': asset_type.code,
                'group_id': asset_type.group_id,
                'color': asset_type.color,
                'description': asset_type.description,
                'display_order': asset_type.display_order,
                'is_active': asset_type.is_active
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def get_asset_types_by_group(request):
    try:
        group_id = request.GET.get('group_id')
        if not group_id:
            return JsonResponse({
                'status': 'error',
                'message': 'Group ID is required'
            }, status=400)

        types = AssetType.objects.filter(group_id=group_id, is_active=True).order_by('display_order', 'name')
        
        # Convert to list with localized names using get_name() method
        types_list = []
        for type_obj in types:
            types_list.append({
                'id': type_obj.id,
                'name': type_obj.name,
                'name_local': type_obj.name_local,
                'code': type_obj.code,
                'color': type_obj.color,
                'group_id': type_obj.group_id,
                'localized_name': type_obj.get_name()  # Use get_name() to get translation for current language
            })

        return JsonResponse({
            'status': 'success',
            'types': types_list
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@require_POST
@login_required
def add_asset_type(request):
    try:
        with transaction.atomic():
            group_id = request.POST.get('group_id')
            group = get_object_or_404(AssetGroup, id=group_id)
            
            # Validate required fields
            name = request.POST.get('name')
            code = request.POST.get('code')
            
            if not name or not code:
                return JsonResponse({
                    'status': 'error',
                    'message': _('Name and Code are required')
                }, status=400)
            
            # Check if code already exists for this group
            if AssetType.objects.filter(code=code, group=group).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Code already exists for this group')
                }, status=400)

            asset_type = AssetType.objects.create(
                name=name,
                name_local=request.POST.get('name_local', ''),
                code=code,
                group=group,
                color=request.POST.get('color', '#007bff'),
                description=request.POST.get('description', ''),
                display_order=int(request.POST.get('display_order', 0)),
                is_active=request.POST.get('is_active', '') in ('on', 'true', '1'),
            )

            return JsonResponse({'status': 'success', 'id': asset_type.id})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
def edit_asset_type(request):
    try:
        with transaction.atomic():
            type_id = request.POST.get('id')
            asset_type = get_object_or_404(AssetType, id=type_id)

            # Отримуємо нову групу
            new_group_id = request.POST.get('group_id')
            if not new_group_id:
                return JsonResponse({
                    'status': 'error',
                    'message': _('Group is required')
                }, status=400)

            new_group = get_object_or_404(AssetGroup, id=new_group_id)
            
            # Validate required fields
            name = request.POST.get('name')
            code = request.POST.get('code')
            
            if not name or not code:
                return JsonResponse({
                    'status': 'error',
                    'message': _('Name and Code are required')
                }, status=400)
            
            # Check if code already exists for this group (excluding current type)
            if AssetType.objects.filter(code=code, group=new_group).exclude(id=type_id).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Code already exists for this group')
                }, status=400)

            # Оновлюємо дані типу
            asset_type.group = new_group
            asset_type.name = name
            asset_type.name_local = request.POST.get('name_local', '')
            asset_type.code = code
            asset_type.color = request.POST.get('color', '#007bff')
            asset_type.description = request.POST.get('description', '')
            asset_type.display_order = int(request.POST.get('display_order', 0))
            asset_type.is_active = request.POST.get('is_active', '') in ('on', 'true', '1')
            asset_type.save()

            return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@require_POST
@login_required
def delete_asset_type(request):
    try:
        with transaction.atomic():
            type_id = request.POST.get('type_id')
            asset_type = get_object_or_404(AssetType, id=type_id)

            # Перевірка чи тип використовується
            if InformationAsset.objects.filter(asset_type=asset_type).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete type that is in use')
                }, status=400)
            if SoftwareRegister.objects.filter(asset_type=asset_type).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete type that is used in Software Register')
                }, status=400)
            if ExternalMediaRegister.objects.filter(asset_type=asset_type).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete type that is used in External Media Register')
                }, status=400)

            asset_type.delete()
            return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# Add to your views.py
@login_required
def search_cabinet_users(request):
    company_id = request.GET.get('company_id')
    query = request.GET.get('query', '').lower()
    user_type = request.GET.get('type')

    try:
        # Get base queryset filtered by company
        users = CabinetUser.objects.select_related('user', 'company').filter(
            company_id=company_id
        ).distinct()

        # Apply search filter if query exists
        if query:
            users = users.filter(
                Q(user__first_name__icontains=query) |
                Q(user__last_name__icontains=query) |
                Q(user__email__icontains=query) |
                Q(department__icontains=query) |
                Q(position__icontains=query)
            )

        # Check existing assignments
        if user_type == 'owner':
            existing_owners = AssetOwner.objects.filter(
                company_id=company_id
            ).values_list('cabinet_user_id', flat=True)
            users = users.exclude(id__in=existing_owners)
        elif user_type == 'admin':
            existing_admins = AssetAdministrator.objects.filter(
                company_id=company_id
            ).values_list('cabinet_user_id', flat=True)
            users = users.exclude(id__in=existing_admins)

        # Order results
        users = users.order_by('user__first_name', 'user__last_name')

        data = {
            'users': [{
                'id': user.id,
                'user': {
                    'first_name': user.user.first_name,
                    'last_name': user.user.last_name,
                    'email': user.user.email,
                },
                'department': str(user.department) if user.department else '',
                'position': str(user.position) if user.position else '',
                'phone': user.phone or '',
                'company': {
                    'id': user.company.id,
                    'name': user.company.name
                } if user.company else None,
            } for user in users]
        }
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error in search_cabinet_users: {str(e)}")
        return JsonResponse({
            'error': str(e)
        }, status=400)

@require_POST
@login_required
def add_asset_group(request):
    try:
        with transaction.atomic():
            # Validate required fields
            name = request.POST.get('name')
            code = request.POST.get('code')
            abbreviation = request.POST.get('abbreviation')
            
            if not name or not code or not abbreviation:
                return JsonResponse({
                    'status': 'error',
                    'message': _('Name, Code, and Abbreviation are required')
                }, status=400)
            
            # Check if code already exists
            if AssetGroup.objects.filter(code=code).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Code already exists')
                }, status=400)
            
            # Check if abbreviation already exists
            if AssetGroup.objects.filter(abbreviation=abbreviation).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Abbreviation already exists')
                }, status=400)
            
            asset_group = AssetGroup.objects.create(
                name=name,
                name_local=request.POST.get('name_local', ''),
                code=code,
                abbreviation=abbreviation,
                color=request.POST.get('color', '#007bff'),
                description=request.POST.get('description', ''),
                display_order=int(request.POST.get('display_order', 0)),
                is_active=request.POST.get('is_active', '') in ('on', 'true', '1'),
                show_in_software_register=request.POST.get('show_in_software_register', '') in ('on', 'true', '1'),
                show_in_external_media_register=request.POST.get('show_in_external_media_register', '') in ('on', 'true', '1'),
            )
            
            return JsonResponse({
                'status': 'success',
                'id': asset_group.id,
                'message': _('Asset group added successfully')
            })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@require_POST
@login_required
def edit_asset_group(request):
    try:
        with transaction.atomic():
            group_id = request.POST.get('id')
            asset_group = get_object_or_404(AssetGroup, id=group_id)
            
            # Validate required fields
            name = request.POST.get('name')
            code = request.POST.get('code')
            abbreviation = request.POST.get('abbreviation')
            
            if not name or not code or not abbreviation:
                return JsonResponse({
                    'status': 'error',
                    'message': _('Name, Code, and Abbreviation are required')
                }, status=400)
            
            # Check if code already exists for other groups
            if AssetGroup.objects.filter(code=code).exclude(id=group_id).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Code already exists')
                }, status=400)
            
            # Check if abbreviation already exists for other groups
            if AssetGroup.objects.filter(abbreviation=abbreviation).exclude(id=group_id).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Abbreviation already exists')
                }, status=400)
            
            # Update the asset group
            asset_group.name = name
            asset_group.name_local = request.POST.get('name_local', '')
            asset_group.code = code
            asset_group.abbreviation = abbreviation
            asset_group.color = request.POST.get('color', '#007bff')
            asset_group.description = request.POST.get('description', '')
            asset_group.display_order = int(request.POST.get('display_order', 0))
            asset_group.is_active = request.POST.get('is_active', '') in ('on', 'true', '1')
            asset_group.show_in_software_register = request.POST.get('show_in_software_register', '') in ('on', 'true', '1')
            asset_group.show_in_external_media_register = request.POST.get('show_in_external_media_register', '') in ('on', 'true', '1')
            asset_group.save()
            
            return JsonResponse({
                'status': 'success',
                'message': _('Asset group updated successfully')
            })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@require_POST
@login_required
def delete_asset_group(request):
    try:
        with transaction.atomic():
            group_id = request.POST.get('group_id')
            asset_group = get_object_or_404(AssetGroup, id=group_id)
            
            # Check if there are any asset types associated with this group
            if asset_group.asset_types.exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete group that has associated asset types')
                }, status=400)
            
            # Check if there are any assets using this group
            if InformationAsset.objects.filter(group=asset_group).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete group that has associated assets')
                }, status=400)
            if SoftwareRegister.objects.filter(group=asset_group).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete group that is used in Software Register')
                }, status=400)
            if ExternalMediaRegister.objects.filter(group=asset_group).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot delete group that is used in External Media Register')
                }, status=400)
            
            asset_group.delete()
            
            return JsonResponse({
                'status': 'success',
                'message': _('Asset group deleted successfully')
            })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def get_asset_group(request):
    try:
        group_id = request.GET.get('group_id')
        asset_group = get_object_or_404(AssetGroup, id=group_id)
        return JsonResponse({
            'status': 'success',
            'group': {
                'id': asset_group.id,
                'name': asset_group.name,
                'name_local': asset_group.name_local,
                'code': asset_group.code,
                'abbreviation': asset_group.abbreviation,
                'color': asset_group.color,
                'description': asset_group.description,
                'display_order': asset_group.display_order,
                'is_active': asset_group.is_active,
                'show_in_software_register': asset_group.show_in_software_register,
                'show_in_external_media_register': asset_group.show_in_external_media_register,
            }
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def get_all_asset_groups(request):
    try:
        groups = AssetGroup.objects.filter(is_active=True).prefetch_related('translations__country').order_by('display_order', 'name')
        
        groups_list = []
        for group in groups:
            groups_list.append({
                'id': group.id,
                'name': group.name,
                'name_local': group.name_local,
                'code': group.code,
                'abbreviation': group.abbreviation,
                'description': group.description or '',
                'color': group.color,
                'display_order': group.display_order,
                'is_active': group.is_active,
                'localized_name': group.get_name(),
                'show_in_software_register': group.show_in_software_register,
                'show_in_external_media_register': group.show_in_external_media_register,
            })
        
        return JsonResponse({
            'status': 'success',
            'groups': groups_list
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)