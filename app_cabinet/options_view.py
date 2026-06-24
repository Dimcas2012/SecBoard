import json
from datetime import datetime

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, get_object_or_404
import logging
from django.utils import formats, timezone, translation
from django.contrib import messages
from django.views.decorators.http import require_POST, require_http_methods
from django.utils.translation import gettext as _, get_language
from django.shortcuts import render
from app_study.models import Page
from app_conf.models import Company
from django.db.models import Prefetch, Count
from django.core.paginator import Paginator
from .pagination_utils import (
    CABINET_TABLE_PAGE_SIZE_OPTIONS,
    get_cabinet_table_page_size,
)
from django.db.models import Q
from django.core.exceptions import ValidationError
from .models import (
    CabinetUser,
    CabinetGroup,
    Position,
    Department,
    UserActivity,
    OrgStructureGuide,
    OrgStructureGuideTranslation,
    CabinetUsersGuide,
    CabinetUsersGuideTranslation,
    CabinetGroupsGuide,
    CabinetGroupsGuideTranslation,
    PlatformRole,
    CabinetTaskReminderSchedule,
    CabinetSettings,
)
from .permissions import require_permission, has_permission
from .task_reminder_utils import estimate_next_periodic_task_run
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Task reminder schedules table: dd.mm.yyyy hh:mm (24h)
_TASK_REMINDER_LIST_DATETIME_FORMAT = 'd.m.Y H:i'


def get_user_companies(user):
    """Get companies accessible to user"""
    if user.is_superuser:
        return Company.objects.all()
    
    # Check if user is staff and has no specific access restrictions
    if user.is_staff:
        user_groups = user.groups.all()
        # Check if user has any access restrictions through AccessOptions
        from .models import AccessOptions
        access_records = AccessOptions.objects.filter(
            group__in=user_groups,
            has_access_org_structure=True
        )
        
        if not access_records.exists():
            # If staff user has no access restrictions, give access to all companies
            return Company.objects.all()
    
    # Get companies through access records
    user_groups = user.groups.all()
    from .models import AccessOptions
    access_records = AccessOptions.objects.filter(
        group__in=user_groups,
        has_access_org_structure=True
    )
    
    companies = Company.objects.none()
    for access in access_records:
        companies = companies | access.companies.all()
    
    # If no companies found through access records, try to get companies where user is a cabinet user
    if not companies.exists():
        try:
            cabinet_user = CabinetUser.objects.filter(user=user).first()
            if cabinet_user and cabinet_user.company:
                companies = Company.objects.filter(id=cabinet_user.company.id)
        except:
            pass
    
    # Last fallback: if still no companies and user is authenticated, give access to all companies
    if not companies.exists() and user.is_authenticated:
        companies = Company.objects.all()
    
    return companies.distinct()


def _cabinet_users_task_reminder_urls(request):
    """
    Resolve task-reminder API URLs. Prefer reverse(); if names are missing (stale worker,
    duplicate includes), derive from the current users list path so both /en/users/ and
    /en/app_cabinet/users/ work.

    Returns (submit_url, list_url, schedules_base_url) where schedules_base_url has no
    trailing slash (append /{pk}/ or /{pk}/update/).
    """
    from django.urls import NoReverseMatch, reverse

    try:
        submit = reverse('task_reminder_submit')
        list_url = reverse('task_reminder_schedules_list')
    except NoReverseMatch:
        base = request.path.rstrip('/')
        if base.endswith('/users'):
            submit = f'{base}/task-reminder/'
            list_url = f'{base}/task-reminder-schedules/'
        else:
            return '', '', ''
    base_sched = list_url.rstrip('/')
    return submit, list_url, base_sched


def _cabinet_users_telegram_broadcast_url(request):
    from django.urls import NoReverseMatch, reverse

    try:
        return reverse('telegram_broadcast_submit')
    except NoReverseMatch:
        base = request.path.rstrip('/')
        if base.endswith('/users'):
            return f'{base}/telegram-broadcast/'
        return ''


def _cabinet_users_email_broadcast_url(request):
    from django.urls import NoReverseMatch, reverse

    try:
        return reverse('email_broadcast_submit')
    except NoReverseMatch:
        base = request.path.rstrip('/')
        if base.endswith('/users'):
            return f'{base}/email-broadcast/'
        return ''


def _platform_roles_for_companies(accessible_companies):
    """Return active platform roles that are global (no companies) or scoped to at least one of the given companies."""
    ids_global = PlatformRole.objects.filter(is_active=True).annotate(c=Count('companies')).filter(c=0).values_list('id', flat=True)
    ids_scoped = PlatformRole.objects.filter(is_active=True, companies__in=accessible_companies).values_list('id', flat=True).distinct()
    role_ids = set(ids_global) | set(ids_scoped)
    return PlatformRole.objects.filter(id__in=role_ids).order_by('order', 'name')


@require_permission('users', 'view')
@require_http_methods(['GET'])
def get_platform_roles_by_company(request):
    """Return platform roles available for the given company (global roles + roles scoped to that company). Used in Add/Edit User."""
    company_id = request.GET.get('company_id', '').strip()
    if not company_id:
        return JsonResponse({'platform_roles': []})
    try:
        cid = int(company_id)
    except ValueError:
        return JsonResponse({'platform_roles': []})
    accessible = get_user_companies(request.user)
    if not accessible.filter(id=cid).exists():
        return JsonResponse({'platform_roles': []})
    companies = Company.objects.filter(id=cid)
    roles = _platform_roles_for_companies(companies).prefetch_related(
        'companies',
        Prefetch('groups', queryset=Group.objects.select_related('cabinet_details', 'cabinet_details__company'))
    )
    platform_roles = []
    for r in roles:
        companies_list = list(r.companies.all())
        company_names = [c.name for c in companies_list]
        company_ids = [c.id for c in companies_list]
        role_groups = []
        for g in r.groups.all():
            details = getattr(g, 'cabinet_details', None)
            if details:
                role_groups.append({
                    'id': g.id,
                    'name': details.get_name() or g.name,
                    'description': (details.get_description() or '').strip(),
                    'company_id': details.company_id,
                    'company_name': details.company.name if details.company else '',
                    'color': details.color or '#000000',
                })
            else:
                role_groups.append({
                    'id': g.id,
                    'name': g.name,
                    'description': '',
                    'company_id': None,
                    'company_name': '',
                    'color': '#000000',
                })
        platform_roles.append({
            'id': r.id,
            'name': r.name,
            'color': r.color or '#6c757d',
            'description': (r.description or '').strip(),
            'company_names': company_names,
            'company_ids': company_ids,
            'groups': role_groups,
        })
    return JsonResponse({'platform_roles': platform_roles})


logger = logging.getLogger(__name__)

logger.debug("This is a debug message")
logger.info("This is an info message")
logger.warning("This is a warning message")
logger.error("This is an error message")




def is_in_group(user, group_name):
    return user.groups.filter(name=group_name).exists()


@login_required
# In views.py
@login_required
def get_departments(request):
    try:
        lang = request.LANGUAGE_CODE[:2]
        departments = Department.objects.filter(company=request.user.cabinetuser.company)

        data = []
        for dept in departments:
            data.append({
                'id': dept.id,
                'department_name_ua': dept.get_name('ua') or dept.get_name('uk'),
                'department_name_en': dept.get_name('en'),
                'department_name_ru': dept.get_name('ru'),
                'description_ua': dept.get_description('ua') or dept.get_description('uk'),
                'description_en': dept.get_description('en'),
                'description_ru': dept.get_description('ru'),
                'level': dept.level,
                'color': dept.color,
                'parent_id': dept.parent.id if dept.parent else None,
                'name': dept.get_name(lang),  # For backward compatibility
                'description': dept.get_description(lang)  # For backward compatibility
            })

        return JsonResponse(data, safe=False)
    except Exception as e:
        logger.error(f"Error getting departments: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


def _department_serialize_for_form(department):
    """Build dict of name_<lang>, description_<lang> for each language (for Add/Edit modal)."""
    from django.conf import settings
    langs = [code for code, _ in get_department_form_languages()]
    out = {
        'id': department.id,
        'company_id': department.company.id if department.company else None,
        'level': department.level,
        'color': department.color,
        'parent_id': department.parent.id if department.parent else None,
        'parent_position_id': department.parent_position.id if department.parent_position else None,
    }
    for lang in langs:
        out[f'name_{lang}'] = department.get_name(lang) or ''
        out[f'description_{lang}'] = department.get_description(lang) or ''
    return out


@login_required
def get_department(request, pk):
    try:
        # Check if user has a cabinetuser profile
        if hasattr(request.user, 'cabinetuser'):
            department = get_object_or_404(Department, pk=pk, company=request.user.cabinetuser.company)
        else:
            department = get_object_or_404(Department, pk=pk)

        data = {
            'status': 'success',
            'department': _department_serialize_for_form(department)
        }
        return JsonResponse(data)

    except Department.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': _('Department not found')
        }, status=404)

    except Exception as e:
        logger.error(f"Error getting department {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)



@require_permission('org_structure', 'add_departments')
@require_POST
def add_department(request):
    try:
        # Name (English) is required by default
        name_en = request.POST.get('name_en', '').strip()
        required_fields = {
            'company': _('Company is required'),
            'name_en': _('Name (English) is required'),
        }
        errors = {}
        for field, error_message in required_fields.items():
            if not request.POST.get(field):
                errors[field] = [error_message]
        if errors:
            return JsonResponse({'status': 'error', 'errors': errors}, status=400)

        parent_id = request.POST.get('parent') or None
        parent_position_id = request.POST.get('parent_position') or None
        if parent_id and parent_position_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'parent': [_('You can select either Parent Department or Parent Position, but not both')]}
            }, status=400)

        if parent_position_id:
            from app_cabinet.models import Position
            try:
                parent_position = Position.objects.get(id=parent_position_id)
                if parent_position.company_id != int(request.POST['company']):
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'parent_position': [_('Parent position must be from the same company')]}
                    }, status=400)
            except Position.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'parent_position': [_('Invalid parent position')]}
                }, status=400)

        department = Department.objects.create(
            company_id=request.POST['company'],
            parent_id=parent_id,
            parent_position_id=parent_position_id,
            name=name_en[:255],
            description=request.POST.get('description_en', '').strip(),
            color=request.POST.get('color', '#2e6da4')
        )
        _save_department_translations_from_post(department, request)
        department.save()

        return JsonResponse({
            'status': 'success',
            'message': _('Department created successfully'),
            'department': {
                'id': department.id,
                'name': department.get_name(request.LANGUAGE_CODE[:2]),
                'level': department.level,
                'color': department.color
            }
        })

    except Exception as e:
        logger.error(f"Error creating department: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)





@require_permission('org_structure', 'edit_departments')
@require_POST
def edit_department(request, pk):
    try:
        department = get_object_or_404(Department, pk=pk)

        name_en = request.POST.get('name_en', '').strip()
        required_fields = {
            'company': _('Company is required'),
            'name_en': _('Name (English) is required'),
        }
        errors = {}
        for field, error_message in required_fields.items():
            if not request.POST.get(field):
                errors[field] = [error_message]
        if errors:
            return JsonResponse({'status': 'error', 'errors': errors}, status=400)

        parent_id = request.POST.get('parent') or None
        parent_position_id = request.POST.get('parent_position') or None
        if parent_id and parent_position_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'parent': [_('You can select either Parent Department or Parent Position, but not both')]}
            }, status=400)

        if parent_id:
            if int(parent_id) == department.id:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'parent': [_('Department cannot be its own parent')]}
                }, status=400)
            try:
                parent = Department.objects.get(id=parent_id)
                if parent.company_id != int(request.POST['company']):
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'parent': [_('Parent department must be from the same company')]}
                    }, status=400)
            except Department.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'parent': [_('Invalid parent department')]}
                }, status=400)

        if parent_position_id:
            from app_cabinet.models import Position
            try:
                parent_position = Position.objects.get(id=parent_position_id)
                if parent_position.company_id != int(request.POST['company']):
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'parent_position': [_('Parent position must be from the same company')]}
                    }, status=400)
            except Position.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'parent_position': [_('Invalid parent position')]}
                }, status=400)

        department.company_id = request.POST['company']
        department.parent_id = parent_id or None
        department.parent_position_id = parent_position_id
        department.color = request.POST.get('color', '#2e6da4')
        _save_department_translations_from_post(department, request)
        department.save()

        return JsonResponse({
            'status': 'success',
            'message': _('Department updated successfully'),
            'department': {
                'id': department.id,
                'name': department.get_name(request.LANGUAGE_CODE[:2]),
                'level': department.level,
                'color': department.color
            }
        })


    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@require_permission('org_structure', 'delete_departments')
@require_POST
def delete_department(request, pk):
    if not request.user.is_authenticated:
        print("User not authenticated")
        return JsonResponse({'status': 'error', 'message': 'User not authenticated'}, status=403)


    # Proceed with existing logic
    try:
        department = get_object_or_404(Department, pk=pk)
        if Department.objects.filter(parent=department).exists():
            return JsonResponse({
                'status': 'error',
                'message': 'Cannot delete department because it has child departments.'
            }, status=400)

        if department.position_set.exists():
            return JsonResponse({
                'status': 'error',
                'message': 'Cannot delete department because it has associated positions.'
            }, status=400)

        department.delete()
        return JsonResponse({'status': 'success', 'message': 'Department deleted successfully.'})
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An unexpected error occurred: {str(e)}'
        }, status=500)

# Positions CRUD
@login_required
def get_positions(request):
    try:
        lang = request.LANGUAGE_CODE[:2]
        positions = Position.objects.filter(company=request.user.cabinetuser.company)

        data = []
        for pos in positions:
            data.append({
                'id': pos.id,
                'position_name_ua': pos.get_name('ua') or pos.get_name('uk'),
                'position_name_en': pos.get_name('en'),
                'position_name_ru': pos.get_name('ru'),
                'description_ua': pos.get_description('ua') or pos.get_description('uk'),
                'description_en': pos.get_description('en'),
                'description_ru': pos.get_description('ru'),
                'department_id': pos.department_id,
                'color': pos.color,
                'name': pos.get_name(lang),  # For backward compatibility
                'description': pos.get_description(lang),  # For backward compatibility
                'users': pos.get_users_data()  # Assuming you have this method
            })
        print('get_positions data =  ',  data)
        return JsonResponse(data, safe=False)
    except Exception as e:
        logger.error(f"Error getting positions: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_position(request, pk):
    try:
        position = get_object_or_404(Position, pk=pk)
        payload = _position_serialize_for_form(position)
        payload['users_count'] = position.cabinetuser_set.filter(user__is_active=True).count()
        data = {'status': 'success', 'position': payload}
        return JsonResponse(data)

    except Position.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': _('Position not found')
        }, status=404)

    except Exception as e:
        logger.error(f"Error getting position {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@require_permission('org_structure', 'add_positions')
@require_POST
def add_position(request):
    try:
        name_en = request.POST.get('name_en', '').strip()
        required_fields = {
            'company': _('Company is required'),
            'name_en': _('Name (English) is required'),
        }
        errors = {}
        for field, error_message in required_fields.items():
            if not request.POST.get(field):
                errors[field] = [error_message]
        if errors:
            return JsonResponse({'status': 'error', 'errors': errors}, status=400)

        department_id = request.POST.get('department') or None
        parent_position_id = request.POST.get('parent_position') or None
        if not department_id and not parent_position_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'department': [_('You must select either Department or Parent Position')]}
            }, status=400)
        if department_id and parent_position_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'department': [_('You can select either Department or Parent Position, but not both')]}
            }, status=400)

        try:
            company = Company.objects.get(id=request.POST['company'])
        except Company.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Selected company does not exist')]}
            }, status=400)

        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id, company=company)
            except Department.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'department': [_('Selected department does not exist in this company')]}
                }, status=400)

        parent_position = None
        if parent_position_id:
            try:
                parent_position = Position.objects.get(id=parent_position_id)
                if parent_position.company_id != company.id:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'parent_position': [_('Parent position must be from the same company')]}
                    }, status=400)
            except Position.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'parent_position': [_('Invalid parent position')]}
                }, status=400)

        position = Position.objects.create(
            company=company,
            department=department,
            parent_position=parent_position,
            name=name_en[:255],
            description=request.POST.get('description_en', '').strip(),
            color=request.POST.get('color', '#2e6da4')
        )
        _save_position_translations_from_post(position, request)
        position.save()

        current_language = request.LANGUAGE_CODE[:2]
        return JsonResponse({
            'status': 'success',
            'message': _('Position created successfully'),
            'position': {
                'id': position.id,
                'name': position.get_name(current_language),
                'description': position.get_description(current_language),
                'department_name': department.get_name(current_language) if department else None,
                'parent_position_name': parent_position.get_name(current_language) if parent_position else None,
                'company_name': company.name,
                'color': position.color
            }
        })

    except Exception as e:
        logger.error(f"Error creating position: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while creating the position'),
            'errors': {'__all__': [str(e)]}
        }, status=500)

@require_permission('org_structure', 'edit_positions')
@require_POST
def edit_position(request, pk):
    try:
        # Get the position instance
        position = get_object_or_404(Position, pk=pk)

        name_en = request.POST.get('name_en', '').strip()
        required_fields = {
            'company': _('Company is required'),
            'name_en': _('Name (English) is required'),
        }
        errors = {}
        for field, error_message in required_fields.items():
            if not request.POST.get(field):
                errors[field] = [error_message]
        if errors:
            return JsonResponse({'status': 'error', 'errors': errors}, status=400)

        department_id = request.POST.get('department') or None
        parent_position_id = request.POST.get('parent_position') or None
        if not department_id and not parent_position_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'department': [_('You must select either Department or Parent Position')]}
            }, status=400)
        if department_id and parent_position_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'department': [_('You can select either Department or Parent Position, but not both')]}
            }, status=400)

        try:
            company = Company.objects.get(id=request.POST['company'])
        except Company.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Selected company does not exist')]}
            }, status=400)

        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id, company=company)
            except Department.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'department': [_('Selected department does not exist in this company')]}
                }, status=400)

        parent_position = None
        if parent_position_id:
            try:
                parent_position = Position.objects.get(id=parent_position_id)
                if parent_position.company_id != company.id:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'parent_position': [_('Parent position must be from the same company')]}
                    }, status=400)
                if int(parent_position_id) == position.id:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'parent_position': [_('Position cannot be its own parent')]}
                    }, status=400)
            except Position.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'parent_position': [_('Invalid parent position')]}
                }, status=400)

        if position.cabinetuser_set.exists() and department_id:
            old_has_department = position.department_id is not None
            new_has_department = department_id is not None
            if old_has_department and new_has_department and position.department_id != int(department_id):
                return JsonResponse({
                    'status': 'error',
                    'message': _('Cannot change department for position that has associated users')
                }, status=400)

        position.company = company
        position.department = department
        position.parent_position = parent_position
        position.color = request.POST.get('color', '#2e6da4')
        _save_position_translations_from_post(position, request)
        position.save()

        current_language = request.LANGUAGE_CODE[:2]
        return JsonResponse({
            'status': 'success',
            'message': _('Position updated successfully'),
            'position': {
                'id': position.id,
                'name': position.get_name(current_language),
                'description': position.get_description(current_language),
                'department_name': department.get_name(current_language) if department else None,
                'parent_position_name': parent_position.get_name(current_language) if parent_position else None,
                'company_name': company.name,
                'color': position.color
            }
        })

    except Exception as e:
        logger.error(f"Error updating position {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while updating the position'),
            'errors': {'__all__': [str(e)]}
        }, status=500)


@require_permission('org_structure', 'delete_positions')
@require_http_methods(["DELETE"])
def delete_position(request, pk):
    try:
        position = get_object_or_404(Position, pk=pk)

        # Check for active users
        if position.cabinetuser_set.filter(user__is_active=True).exists():
            return JsonResponse({
                'status': 'error',
                'message': _('Cannot delete position with active users')
            }, status=400)

        position_name = position.get_name(request.LANGUAGE_CODE[:2])
        position.delete()

        return JsonResponse({
            'status': 'success',
            'message': _('Position deleted successfully')
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)



@login_required
def get_department_positions(request, company_id, department_id):
    try:
        logger.info(f"Getting positions for company_id={company_id}, department_id={department_id}")

        # Check if user has access to this company through Chosen Companies
        accessible_companies = get_user_companies(request.user)
        if not accessible_companies.filter(id=company_id).exists():
            return JsonResponse({'error': 'Access denied'}, status=403)

        # Get positions only for the selected department
        positions = Position.objects.filter(
            company_id=company_id,
            department_id=department_id
        )

        logger.info(f"Found {positions.count()} positions")
        logger.debug(f"Positions query: {positions.query}")

        current_language = request.LANGUAGE_CODE[:2]

        data = []
        for pos in positions:
            pos_data = {
                'id': pos.id,
                'name': pos.get_name(current_language),
                'description': getattr(pos, f'description_{current_language}', ''),
                'department_id': pos.department_id,
                'parent_position_id': pos.parent_position_id
            }
            logger.debug(f"Processed position: {pos_data}")
            data.append(pos_data)

        logger.info(f"Returning {len(data)} positions")
        return JsonResponse(data, safe=False)

    except Exception as e:
        logger.error(f"Error getting positions: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
@require_http_methods(["GET"])
def org_structure_guide(request):
    """Return JSON { content: html } for the Organization Structure guide (localized)."""
    if not has_permission(request.user, 'org_structure', 'view'):
        return JsonResponse({'content': ''})
    from app_conf.models import Country
    import json
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = OrgStructureGuide.objects.first()
    if guide:
        if country:
            trans = OrgStructureGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = OrgStructureGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def org_structure_guide_translate(request):
    """API for AI translation of Organization Structure guide content (admin)."""
    from app_conf.models import Country
    import json
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def cabinet_users_guide(request):
    """Return JSON { content: html } for the Cabinet Users guide (localized)."""
    if not has_permission(request.user, 'users', 'view'):
        return JsonResponse({'content': ''})
    from app_conf.models import Country
    import json
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = CabinetUsersGuide.objects.first()
    if guide:
        if country:
            trans = CabinetUsersGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = CabinetUsersGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def cabinet_users_guide_translate(request):
    """API for AI translation of Cabinet Users guide content (admin)."""
    from app_conf.models import Country
    import json
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_department_form_languages():
    """Return list of (code, name) for Department Name/Description translation fields (same as Vulnerability)."""
    from django.conf import settings
    return list(getattr(settings, 'LANGUAGES', [('en', 'English')])) or [('en', 'English')]


def _save_department_translations_from_post(department, request):
    """Save Department name and description from request.POST (single language)."""
    name_en = request.POST.get('name_en', '').strip()
    description_en = request.POST.get('description_en', '').strip()
    department.name = name_en[:255] if name_en else (department.name or '')
    department.description = description_en or (department.description or '')


def _position_serialize_for_form(position):
    """Build dict of name_<lang>, description_<lang> for each language (for Add/Edit Position modal)."""
    langs = [code for code, _ in get_department_form_languages()]
    out = {
        'id': position.id,
        'company_id': position.company_id,
        'department_id': position.department_id,
        'parent_position_id': position.parent_position_id,
        'color': position.color,
    }
    for lang in langs:
        out[f'name_{lang}'] = position.get_name(lang) or ''
        out[f'description_{lang}'] = position.get_description(lang) or ''
    return out


def _save_position_translations_from_post(position, request):
    """Save Position name and description from request.POST (single language)."""
    name_en = request.POST.get('name_en', '').strip()
    description_en = request.POST.get('description_en', '').strip()
    position.name = name_en[:255] if name_en else (position.name or '')
    position.description = description_en or (position.description or '')


@require_permission('org_structure', 'view')
def org_structure_view(request):
    """
    View для відображення сторінки організаційної структури
    """
    try:
        context = {
            'page_title': _("Organization Structure"),
            'current_language': request.LANGUAGE_CODE[:2],
            'companies': get_user_companies(request.user).order_by('name'),
            'department_form_languages': get_department_form_languages(),
            'position_form_languages': get_department_form_languages(),
        }
        # print('org_structure_view context = ', context)
        return render(request, 'app_cabinet/org_structure.html', context)

    except Exception as e:
        logger.error(f"Error in org_structure_view: {str(e)}")
        messages.error(request, _("An error occurred while loading the organization structure"))
        return redirect('index')


@login_required
def get_company_departments(request, pk):
    try:
        # Check if user has access to this company through Chosen Companies
        accessible_companies = get_user_companies(request.user)
        company = get_object_or_404(Company, pk=pk, id__in=accessible_companies)
        departments = Department.objects.filter(company=company)
        current_language = request.LANGUAGE_CODE[:2]

        data = []
        for dept in departments:
            name = dept.get_name(current_language)
            description = dept.get_description(current_language)

            parent_position_name = None
            if dept.parent_position:
                parent_position_name = dept.parent_position.get_name(current_language)

            data.append({
                'id': dept.id,
                'department_name_ua': dept.get_name('ua') or dept.get_name('uk'),
                'department_name_ru': dept.get_name('ru'),
                'department_name_en': dept.get_name('en'),
                'description_ua': dept.get_description('ua') or dept.get_description('uk'),
                'description_ru': dept.get_description('ru'),
                'description_en': dept.get_description('en'),
                'name': name,
                'description': description,
                'parent_id': dept.parent.id if dept.parent else None,
                'parent_position_id': dept.parent_position.id if dept.parent_position else None,
                'parent_position_name': parent_position_name,
                'color': dept.color,
                'level': dept.level
            })

        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def get_company_users(request, pk):
    try:
        # Check if user has access to this company through Chosen Companies
        accessible_companies = get_user_companies(request.user)
        if not accessible_companies.filter(id=pk).exists():
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        # Get current language
        current_language = request.LANGUAGE_CODE[:2]
        
        users = CabinetUser.objects.filter(company_id=pk).select_related('user', 'position', 'department', 'company')
        user_data = []
        
        for user in users:
            department_name = user.department.get_name(current_language) if user.department else None
            position_name = user.position.get_name(current_language) if user.position else None
            
            user_data.append({
                'id': user.user.id,
                'cabinet_user_id': user.id,
                'username': user.user.username,
                'full_name': user.user.get_full_name(),
                'email': user.user.email,
                'is_active': user.user.is_active,
                'telegram_linked': bool((user.telegram_chat_id or '').strip()),
                'avatar_url': user.avatar.url if user.avatar else None,
                'position_id': user.position.id if user.position else None,
                'position_name': position_name,
                'department_id': user.department.id if user.department else None,
                'department_name': department_name,
                'company': user.company.name if user.company else None,
            })
        
        logger.info(f'get_company_users for company {pk}: {len(user_data)} users')
        return JsonResponse(user_data, safe=False)
    except Exception as e:
        logger.error(f"Error in get_company_users for company {pk}: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_permission('users', 'edit')
@require_http_methods(['POST'])
def task_reminder_submit(request):
    """Send one-time task reminder emails or create a recurring schedule (Celery Beat)."""
    from app_cabinet.task_reminder_utils import send_task_reminder_emails_for_user_ids

    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': _('Invalid JSON')}, status=400)

    company_id = data.get('company_id')
    cabinet_user_ids = data.get('cabinet_user_ids')
    schedule = (data.get('schedule') or 'once').lower()

    if not company_id:
        return JsonResponse({'success': False, 'error': _('Company is required.')}, status=400)
    if not isinstance(cabinet_user_ids, list) or not cabinet_user_ids:
        return JsonResponse({'success': False, 'error': _('Select at least one cabinet user.')}, status=400)

    try:
        company_id = int(company_id)
        cabinet_user_ids = [int(x) for x in cabinet_user_ids]
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': _('Invalid identifiers.')}, status=400)

    accessible = get_user_companies(request.user)
    if not accessible.filter(id=company_id).exists():
        return JsonResponse({'success': False, 'error': _('Access denied.')}, status=403)

    valid_ids = list(
        CabinetUser.objects.filter(
            pk__in=cabinet_user_ids,
            company_id=company_id,
        ).values_list('id', flat=True)
    )
    if len(valid_ids) != len(set(cabinet_user_ids)):
        return JsonResponse({'success': False, 'error': _('Some users do not belong to the selected company.')}, status=400)

    settings_row = CabinetSettings.objects.first()
    if not settings_row or not settings_row.mail_account:
        return JsonResponse({'success': False, 'error': _('Cabinet mail account is not configured.')}, status=400)

    if schedule == 'once':
        sent, skipped, errors = send_task_reminder_emails_for_user_ids(valid_ids)
        return JsonResponse({
            'success': True,
            'sent': sent,
            'skipped': skipped,
            'errors': errors,
            'message': _('Reminders sent.'),
        })

    if schedule not in (
        CabinetTaskReminderSchedule.FREQUENCY_DAILY,
        CabinetTaskReminderSchedule.FREQUENCY_WEEKLY,
        CabinetTaskReminderSchedule.FREQUENCY_MONTHLY,
    ):
        return JsonResponse({'success': False, 'error': _('Invalid schedule type.')}, status=400)

    send_time_raw = data.get('send_time') or '09:00'
    try:
        st = datetime.strptime(str(send_time_raw).strip(), '%H:%M').time()
    except ValueError:
        return JsonResponse({'success': False, 'error': _('Invalid send time (use HH:MM).')}, status=400)

    weekday = data.get('weekday')
    month_day = data.get('month_day')
    if schedule == CabinetTaskReminderSchedule.FREQUENCY_WEEKLY:
        try:
            weekday = int(weekday)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': _('Select a weekday.')}, status=400)
        if weekday < 0 or weekday > 6:
            return JsonResponse({'success': False, 'error': _('Weekday must be 0–6 (Mon–Sun).')}, status=400)
    else:
        weekday = None

    if schedule == CabinetTaskReminderSchedule.FREQUENCY_MONTHLY:
        try:
            month_day = int(month_day)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': _('Select a day of month.')}, status=400)
        if month_day < 1 or month_day > 31:
            return JsonResponse({'success': False, 'error': _('Day of month must be 1–31.')}, status=400)
    else:
        month_day = None

    sch = CabinetTaskReminderSchedule(
        company_id=company_id,
        frequency=schedule,
        send_time=st,
        weekday=weekday,
        month_day=month_day,
        is_active=True,
        created_by=request.user,
    )
    try:
        sch.full_clean()
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

    sch.save()
    sch.recipients.set(valid_ids)
    sch.sync_periodic_task()

    return JsonResponse({
        'success': True,
        'schedule_id': sch.id,
        'message': _('Scheduled reminders saved. Ensure Celery Beat is running.'),
    })


@login_required
@require_permission('users', 'edit')
@require_http_methods(['POST'])
def telegram_broadcast_submit(request):
    """Send a one-time Telegram message to selected cabinet users."""
    from app_cabinet.telegram_broadcast_utils import (
        MESSAGE_TYPE_CUSTOM,
        MESSAGE_TYPE_TASKS,
        send_telegram_broadcast_for_user_ids,
    )

    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': _('Invalid JSON')}, status=400)

    company_id = data.get('company_id')
    cabinet_user_ids = data.get('cabinet_user_ids')
    message_type = (data.get('message_type') or MESSAGE_TYPE_CUSTOM).strip().lower()
    custom_text = (data.get('message') or '').strip()

    if not company_id:
        return JsonResponse({'success': False, 'error': _('Company is required.')}, status=400)
    if not isinstance(cabinet_user_ids, list) or not cabinet_user_ids:
        return JsonResponse({'success': False, 'error': _('Select at least one cabinet user.')}, status=400)
    if message_type not in (MESSAGE_TYPE_CUSTOM, MESSAGE_TYPE_TASKS):
        return JsonResponse({'success': False, 'error': _('Invalid message type.')}, status=400)
    if message_type == MESSAGE_TYPE_CUSTOM and not custom_text:
        return JsonResponse({'success': False, 'error': _('Message text is required.')}, status=400)
    if message_type == MESSAGE_TYPE_CUSTOM and len(custom_text) > 4000:
        return JsonResponse({'success': False, 'error': _('Message is too long (max 4000 characters).')}, status=400)

    try:
        company_id = int(company_id)
        cabinet_user_ids = [int(x) for x in cabinet_user_ids]
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': _('Invalid identifiers.')}, status=400)

    accessible = get_user_companies(request.user)
    if not accessible.filter(id=company_id).exists():
        return JsonResponse({'success': False, 'error': _('Access denied.')}, status=403)

    valid_ids = list(
        CabinetUser.objects.filter(
            pk__in=cabinet_user_ids,
            company_id=company_id,
        ).values_list('id', flat=True)
    )
    if len(valid_ids) != len(set(cabinet_user_ids)):
        return JsonResponse(
            {'success': False, 'error': _('Some users do not belong to the selected company.')},
            status=400,
        )

    sent, skipped, errors = send_telegram_broadcast_for_user_ids(
        valid_ids,
        message_type=message_type,
        custom_text=custom_text,
    )

    if sent == 0 and errors:
        return JsonResponse({
            'success': False,
            'sent': sent,
            'skipped': skipped,
            'errors': errors[:20],
            'error': errors[0],
        }, status=400)

    return JsonResponse({
        'success': True,
        'sent': sent,
        'skipped': skipped,
        'errors': errors[:20],
        'message': _('Telegram messages sent: %(sent)s. Skipped: %(skipped)s.') % {
            'sent': sent,
            'skipped': skipped,
        },
    })


@login_required
@require_permission('users', 'edit')
@require_http_methods(['POST'])
def email_broadcast_submit(request):
    """Send a one-time email to selected cabinet users."""
    from app_cabinet.email_broadcast_utils import (
        EMAIL_BROADCAST_MAX_HTML_LENGTH,
        MESSAGE_TYPE_CUSTOM,
        MESSAGE_TYPE_TASKS,
        _custom_message_has_content,
        send_email_broadcast_for_user_ids,
    )

    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': _('Invalid JSON')}, status=400)

    company_id = data.get('company_id')
    cabinet_user_ids = data.get('cabinet_user_ids')
    message_type = (data.get('message_type') or MESSAGE_TYPE_CUSTOM).strip().lower()
    custom_subject = (data.get('subject') or '').strip()
    custom_text = (data.get('message') or '').strip()

    if not company_id:
        return JsonResponse({'success': False, 'error': _('Company is required.')}, status=400)
    if not isinstance(cabinet_user_ids, list) or not cabinet_user_ids:
        return JsonResponse({'success': False, 'error': _('Select at least one cabinet user.')}, status=400)
    if message_type not in (MESSAGE_TYPE_CUSTOM, MESSAGE_TYPE_TASKS):
        return JsonResponse({'success': False, 'error': _('Invalid message type.')}, status=400)
    if message_type == MESSAGE_TYPE_CUSTOM:
        if not custom_subject:
            return JsonResponse({'success': False, 'error': _('Subject is required.')}, status=400)
        if not _custom_message_has_content(custom_text):
            return JsonResponse({'success': False, 'error': _('Message text is required.')}, status=400)
        if len(custom_subject) > 200:
            return JsonResponse({'success': False, 'error': _('Subject is too long (max 200 characters).')}, status=400)
        if len(custom_text) > EMAIL_BROADCAST_MAX_HTML_LENGTH:
            return JsonResponse({
                'success': False,
                'error': _('Message is too long (max %(max)s characters).') % {
                    'max': EMAIL_BROADCAST_MAX_HTML_LENGTH,
                },
            }, status=400)

    try:
        company_id = int(company_id)
        cabinet_user_ids = [int(x) for x in cabinet_user_ids]
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': _('Invalid identifiers.')}, status=400)

    accessible = get_user_companies(request.user)
    if not accessible.filter(id=company_id).exists():
        return JsonResponse({'success': False, 'error': _('Access denied.')}, status=403)

    valid_ids = list(
        CabinetUser.objects.filter(
            pk__in=cabinet_user_ids,
            company_id=company_id,
        ).values_list('id', flat=True)
    )
    if len(valid_ids) != len(set(cabinet_user_ids)):
        return JsonResponse(
            {'success': False, 'error': _('Some users do not belong to the selected company.')},
            status=400,
        )

    sent, skipped, errors = send_email_broadcast_for_user_ids(
        valid_ids,
        message_type=message_type,
        custom_subject=custom_subject,
        custom_text=custom_text,
    )

    if sent == 0 and errors:
        return JsonResponse({
            'success': False,
            'sent': sent,
            'skipped': skipped,
            'errors': errors[:20],
            'error': errors[0],
        }, status=400)

    return JsonResponse({
        'success': True,
        'sent': sent,
        'skipped': skipped,
        'errors': errors[:20],
        'message': _('Email messages sent: %(sent)s. Skipped: %(skipped)s.') % {
            'sent': sent,
            'skipped': skipped,
        },
    })


@login_required
@require_permission('users', 'edit')
@require_http_methods(['GET'])
def task_reminder_schedules_list(request):
    """JSON list of task reminder schedules for companies the user can access."""
    accessible = get_user_companies(request.user)
    lang = request.LANGUAGE_CODE or get_language() or 'en'

    qs = (
        CabinetTaskReminderSchedule.objects.filter(company__in=accessible)
        .select_related('company', 'periodic_task', 'periodic_task__crontab')
        .annotate(recipient_count=Count('recipients', distinct=True))
        .order_by('-created_at')
        .prefetch_related(
            Prefetch(
                'recipients',
                queryset=CabinetUser.objects.select_related('user').order_by(
                    'user__last_name', 'user__first_name', 'user__username'
                ),
            )
        )
    )

    schedules = []
    with translation.override(lang):
        weekday_labels = (
            _('Monday'),
            _('Tuesday'),
            _('Wednesday'),
            _('Thursday'),
            _('Friday'),
            _('Saturday'),
            _('Sunday'),
        )
        for sch in qs:
            previews = []
            for cu in sch.recipients.all()[:3]:
                previews.append(cu.user.get_full_name() or cu.user.username or '')

            extra = ''
            if sch.frequency == CabinetTaskReminderSchedule.FREQUENCY_WEEKLY:
                if sch.weekday is not None and 0 <= sch.weekday <= 6:
                    extra = str(weekday_labels[sch.weekday])
            elif sch.frequency == CabinetTaskReminderSchedule.FREQUENCY_MONTHLY:
                if sch.month_day is not None:
                    extra = _('Day %(d)s of month') % {'d': sch.month_day}

            time_s = sch.send_time.strftime('%H:%M') if sch.send_time else ''
            summary_parts = [str(sch.get_frequency_display())]
            if time_s:
                summary_parts.append(time_s)
            if extra:
                summary_parts.append(extra)
            schedule_summary = ' · '.join(summary_parts)

            if sch.last_sent_at:
                last_disp = formats.date_format(
                    timezone.localtime(sch.last_sent_at),
                    _TASK_REMINDER_LIST_DATETIME_FORMAT,
                )
            else:
                last_disp = ''

            next_disp = ''
            if sch.is_active and sch.periodic_task_id:
                next_dt = estimate_next_periodic_task_run(sch.periodic_task)
                if next_dt:
                    next_disp = formats.date_format(
                        timezone.localtime(next_dt),
                        _TASK_REMINDER_LIST_DATETIME_FORMAT,
                    )

            schedules.append({
                'id': sch.id,
                'company_id': sch.company_id,
                'company_name': sch.company.name,
                'schedule_summary': schedule_summary,
                'frequency_display': str(sch.get_frequency_display()),
                'send_time': time_s,
                'schedule_detail': extra,
                'recipient_count': sch.recipient_count,
                'recipient_preview': previews,
                'last_sent_at': last_disp,
                'next_sent_display': next_disp or '—',
                'is_active': sch.is_active,
                'status_display': _('Active') if sch.is_active else _('Inactive'),
            })

    return JsonResponse({'schedules': schedules})


@login_required
@require_permission('users', 'edit')
@require_http_methods(['GET', 'DELETE'])
def task_reminder_schedule_detail(request, pk):
    """GET: JSON for editing. DELETE: remove schedule and linked PeriodicTask."""
    accessible = get_user_companies(request.user)
    sch = get_object_or_404(
        CabinetTaskReminderSchedule.objects.filter(company__in=accessible),
        pk=pk,
    )
    if request.method == 'GET':
        ids = list(sch.recipients.order_by('id').values_list('id', flat=True))
        return JsonResponse({
            'id': sch.id,
            'company_id': sch.company_id,
            'cabinet_user_ids': ids,
            'frequency': sch.frequency,
            'send_time': sch.send_time.strftime('%H:%M') if sch.send_time else '09:00',
            'weekday': sch.weekday,
            'month_day': sch.month_day,
            'is_active': sch.is_active,
        })
    sch.delete()
    return JsonResponse({'success': True, 'message': _('Scheduled reminder deleted.')})


@login_required
@require_permission('users', 'edit')
@require_http_methods(['POST'])
def task_reminder_schedule_update(request, pk):
    """Update an existing recurring task reminder schedule."""
    accessible = get_user_companies(request.user)
    sch = get_object_or_404(
        CabinetTaskReminderSchedule.objects.filter(company__in=accessible),
        pk=pk,
    )
    try:
        data = json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'success': False, 'error': _('Invalid JSON')}, status=400)

    company_id = data.get('company_id', sch.company_id)
    cabinet_user_ids = data.get('cabinet_user_ids')
    schedule = (data.get('schedule') or sch.frequency).lower()

    if not company_id:
        return JsonResponse({'success': False, 'error': _('Company is required.')}, status=400)
    if not isinstance(cabinet_user_ids, list) or not cabinet_user_ids:
        return JsonResponse({'success': False, 'error': _('Select at least one cabinet user.')}, status=400)

    try:
        company_id = int(company_id)
        cabinet_user_ids = [int(x) for x in cabinet_user_ids]
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': _('Invalid identifiers.')}, status=400)

    if not accessible.filter(id=company_id).exists():
        return JsonResponse({'success': False, 'error': _('Access denied.')}, status=403)

    valid_ids = list(
        CabinetUser.objects.filter(
            pk__in=cabinet_user_ids,
            company_id=company_id,
        ).values_list('id', flat=True)
    )
    if len(valid_ids) != len(set(cabinet_user_ids)):
        return JsonResponse({'success': False, 'error': _('Some users do not belong to the selected company.')}, status=400)

    settings_row = CabinetSettings.objects.first()
    if not settings_row or not settings_row.mail_account:
        return JsonResponse({'success': False, 'error': _('Cabinet mail account is not configured.')}, status=400)

    if schedule == 'once':
        return JsonResponse({
            'success': False,
            'error': _('Saved schedules cannot be changed to a one-time send. Use “Once” to send now, or delete this schedule.'),
        }, status=400)

    if schedule not in (
        CabinetTaskReminderSchedule.FREQUENCY_DAILY,
        CabinetTaskReminderSchedule.FREQUENCY_WEEKLY,
        CabinetTaskReminderSchedule.FREQUENCY_MONTHLY,
    ):
        return JsonResponse({'success': False, 'error': _('Invalid schedule type.')}, status=400)

    send_time_raw = data.get('send_time') or '09:00'
    try:
        st = datetime.strptime(str(send_time_raw).strip(), '%H:%M').time()
    except ValueError:
        return JsonResponse({'success': False, 'error': _('Invalid send time (use HH:MM).')}, status=400)

    weekday = data.get('weekday')
    month_day = data.get('month_day')
    if schedule == CabinetTaskReminderSchedule.FREQUENCY_WEEKLY:
        try:
            weekday = int(weekday)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': _('Select a weekday.')}, status=400)
        if weekday < 0 or weekday > 6:
            return JsonResponse({'success': False, 'error': _('Weekday must be 0–6 (Mon–Sun).')}, status=400)
    else:
        weekday = None

    if schedule == CabinetTaskReminderSchedule.FREQUENCY_MONTHLY:
        try:
            month_day = int(month_day)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': _('Select a day of month.')}, status=400)
        if month_day < 1 or month_day > 31:
            return JsonResponse({'success': False, 'error': _('Day of month must be 1–31.')}, status=400)
    else:
        month_day = None

    sch.company_id = company_id
    sch.frequency = schedule
    sch.send_time = st
    sch.weekday = weekday
    sch.month_day = month_day
    if 'is_active' in data:
        sch.is_active = bool(data['is_active'])

    try:
        sch.full_clean()
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

    sch.save()
    sch.recipients.set(valid_ids)
    sch.sync_periodic_task()

    return JsonResponse({
        'success': True,
        'message': _('Schedule updated.'),
    })


@login_required
def get_company_positions(request, pk):
    try:
        # Check if user has access to this company through Chosen Companies
        accessible_companies = get_user_companies(request.user)
        if not accessible_companies.filter(id=pk).exists():
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        # Get current language code
        current_language = request.LANGUAGE_CODE[:2]

        # Optional filters for user modals (department or parent_position)
        department_id = request.GET.get('department')
        parent_position_id = request.GET.get('parent_position')

        positions_qs = Position.objects.filter(company_id=pk)
        if department_id:
            positions_qs = positions_qs.filter(department_id=department_id)
        elif parent_position_id:
            positions_qs = positions_qs.filter(parent_position_id=parent_position_id)

        # Get positions with prefetched user data
        positions = positions_qs.prefetch_related(
            Prefetch(
                'cabinetuser_set',
                queryset=CabinetUser.objects.select_related(
                    'user'
                ).prefetch_related(
                    'user__groups',
                    'user__groups__cabinet_details'
                )
            )
        )

        data = []
        for position in positions:
            users_data = []

            # Process each cabinet user for this position
            for cabinet_user in position.cabinetuser_set.all():
                # Skip if user relationship is broken
                if not cabinet_user.user:
                    continue

                # Check if user is active based on dates
                is_active_employee = cabinet_user.is_active_employee()
                if not is_active_employee:
                    continue

                # Process groups
                cabinet_groups = []
                other_groups = []
                for group in cabinet_user.user.groups.all():
                    if hasattr(group, 'cabinet_details'):
                        cabinet_groups.append({
                            'id': group.id,
                            'name': group.cabinet_details.get_name(current_language)
                        })
                    else:
                        other_groups.append({
                            'id': group.id,
                            'name': group.name
                        })

                # Create user data dictionary
                user_data = {
                    'id': cabinet_user.id,
                    'full_name': f"{cabinet_user.user.first_name} {cabinet_user.user.last_name}".strip(),
                    'email': cabinet_user.user.email,  # Ensure email is included
                    'avatar_url': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'phone': cabinet_user.phone,
                    'start_date': cabinet_user.start_date.strftime(
                        '%d.%m.%y %H:%M') if cabinet_user.start_date else None,
                    'end_date': cabinet_user.end_date.strftime('%d.%m.%y %H:%M') if cabinet_user.end_date else None,
                    'is_active': cabinet_user.user.is_active,
                    'cabinet_groups': cabinet_groups,
                    'other_groups': other_groups,
                    'initials': ''.join(word[0].upper() for word in
                                        f"{cabinet_user.user.first_name} {cabinet_user.user.last_name}".split() if word)
                }
                users_data.append(user_data)

            name = position.get_name(current_language)
            description = position.get_description(current_language)

            position_data = {
                'id': position.id,
                'name': name,
                'description': description,
                'position_name_ua': position.get_name('ua') or position.get_name('uk'),
                'position_name_ru': position.get_name('ru'),
                'position_name_en': position.get_name('en'),
                'department_id': position.department_id,
                'parent_position_id': position.parent_position_id,
                'color': position.color,
                'users': users_data
            }
            data.append(position_data)

        return JsonResponse(data, safe=False)

    except Exception as e:
        logger.error(f"Error in get_company_positions: {str(e)}")
        return JsonResponse({
            'error': str(e)
        }, status=400)


@login_required
@login_required
def get_company_groups(request, pk):
    try:
        # Check if user has access to this company through Chosen Companies
        accessible_companies = get_user_companies(request.user)
        company = get_object_or_404(Company, pk=pk, id__in=accessible_companies)

        # Get current language
        current_language = request.LANGUAGE_CODE[:2]

        # Get all cabinet groups for the company
        cabinet_groups = CabinetGroup.objects.select_related('group').filter(
            company=company
        )

        # Get other groups (those without cabinet details)
        # Use a left outer join to ensure we don't get errors if cabinet_details doesn't exist
        other_groups = Group.objects.filter(
            Q(cabinet_details__isnull=True) | ~Q(cabinet_details__company=company)
        ).distinct()

        data = []

        # Process cabinet groups
        for cabinet_group in cabinet_groups:
            try:
                data.append({
                    'id': cabinet_group.group.id,
                    'name': cabinet_group.name,
                    'description': cabinet_group.description,
                    'color': cabinet_group.color,
                    'has_cabinet_details': True
                })
            except Exception as e:
                logger.warning(f"Error processing cabinet group {cabinet_group.id}: {str(e)}")
                continue

        # Process other groups
        for group in other_groups:
            try:
                data.append({
                    'id': group.id,
                    'name': group.name,
                    'description': '',
                    'color': '#6c757d',  # Default color for non-cabinet groups
                    'has_cabinet_details': False
                })
            except Exception as e:
                logger.warning(f"Error processing group {group.id}: {str(e)}")
                continue

        return JsonResponse(data, safe=False)

    except Company.DoesNotExist:
        logger.error(f"Company with ID {pk} not found")
        return JsonResponse({
            'status': 'error',
            'message': _('Company not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error getting groups for company {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while fetching groups')
        }, status=400)

@require_permission('groups', 'view')
def get_group(request, pk):
   try:
       # Fetch the CabinetGroup instance
       cabinet_group = get_object_or_404(CabinetGroup.objects.select_related('company', 'group'), pk=pk)

       # Prepare the response data
       data = {
           'system_group_name': cabinet_group.group.name,
           'company_id': cabinet_group.company.id if cabinet_group.company else None,
           'name': cabinet_group.name,
           'description': cabinet_group.description,
           'color': cabinet_group.color,
           'user_ids': list(cabinet_group.group.user_set.values_list('id', flat=True))
       }

       return JsonResponse({
           'status': 'success',
           'data': data
       })

   except CabinetGroup.DoesNotExist:
       logger.warning(f"Group with ID {pk} not found")
       return JsonResponse({
           'status': 'error',
           'message': _('Group not found')
       }, status=404)

   except Exception as e:
       logger.error(f"Error getting group data for ID {pk}: {str(e)}")
       return JsonResponse({
           'status': 'error',
           'message': _('An unexpected error occurred')
       }, status=500)




@require_permission('groups', 'add')
@require_POST
def create_group(request):
    try:
        # Validate company
        company_id = request.POST.get('company')
        if not company_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Company is required')]}
            }, status=400)

        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Selected company does not exist')]}
            }, status=400)

        # Validate required fields
        required_fields = {
            'group_name': _('System group name is required'),
            'name': _('Name is required'),
        }

        errors = {}
        for field, error_message in required_fields.items():
            if not request.POST.get(field):
                errors[field] = [error_message]

        if errors:
            return JsonResponse({
                'status': 'error',
                'errors': errors
            }, status=400)

        # Check if group name already exists
        if Group.objects.filter(name=request.POST['group_name']).exists():
            return JsonResponse({
                'status': 'error',
                'errors': {'group_name': [_('Group with this name already exists')]}
            }, status=400)

        # Create system group
        group = Group.objects.create(
            name=request.POST['group_name']
        )

        # Create cabinet group
        cabinet_group = CabinetGroup.objects.create(
            group=group,
            company=company,
            name=request.POST['name'],
            description=request.POST.get('description', ''),
            color=request.POST.get('color', '#000000')
        )
        
        # Add selected users to the group
        user_ids = request.POST.getlist('users')
        if user_ids:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            for user_id in user_ids:
                try:
                    user = User.objects.get(id=user_id)
                    group.user_set.add(user)
                    logger.info(f"Added user {user.username} (ID: {user_id}) to group {group.name}")
                except User.DoesNotExist:
                    logger.warning(f"User with ID {user_id} does not exist, skipping")
                except Exception as e:
                    logger.error(f"Error adding user {user_id} to group: {str(e)}")

        # Return success response
        return JsonResponse({
            'status': 'success',
            'message': _('Group created successfully'),
            'group': {
                'id': cabinet_group.id,
                'name': cabinet_group.name,
                'company_name': company.name,
                'color': cabinet_group.color
            }
        })

    except Exception as e:
        logger.error(f"Error creating group: {str(e)}")
        # Clean up system group if it was created
        if 'group' in locals():
            group.delete()

        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while creating the group'),
            'errors': {'__all__': [str(e)]}
        }, status=500)



@require_permission('groups', 'edit')
@require_POST
def edit_group(request, pk):
    try:
        # Get the cabinet group instance
        cabinet_group = get_object_or_404(CabinetGroup, pk=pk)

        # Validate company
        company_id = request.POST.get('company')
        if not company_id:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Company is required')]}
            }, status=400)

        try:
            company = Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Selected company does not exist')]}
            }, status=400)

        # Validate required fields
        required_fields = {
            'group_name': _('System group name is required'),
            'name': _('Name is required'),
        }

        errors = {}
        for field, error_message in required_fields.items():
            if not request.POST.get(field):
                errors[field] = [error_message]

        if errors:
            return JsonResponse({
                'status': 'error',
                'errors': errors
            }, status=400)

        # Check if new group name already exists for other groups
        existing_group = Group.objects.filter(
            name=request.POST['group_name']
        ).exclude(id=cabinet_group.group.id).exists()

        if existing_group:
            return JsonResponse({
                'status': 'error',
                'errors': {'group_name': [_('Group with this name already exists')]}
            }, status=400)

        # Update system group
        cabinet_group.group.name = request.POST['group_name']
        cabinet_group.group.save()

        # Update cabinet group
        cabinet_group.company = company
        cabinet_group.name = request.POST['name']
        cabinet_group.description = request.POST.get('description', '')
        cabinet_group.color = request.POST.get('color', '#000000')
        cabinet_group.save()
        
        # Update group users
        user_ids = request.POST.getlist('users')
        # Clear current users
        cabinet_group.group.user_set.clear()
        # Add selected users
        if user_ids:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            for user_id in user_ids:
                try:
                    user = User.objects.get(id=user_id)
                    cabinet_group.group.user_set.add(user)
                    logger.info(f"Added user {user.username} (ID: {user_id}) to group {cabinet_group.group.name}")
                except User.DoesNotExist:
                    logger.warning(f"User with ID {user_id} does not exist, skipping")
                except Exception as e:
                    logger.error(f"Error adding user {user_id} to group: {str(e)}")

        # Return success response
        return JsonResponse({
            'status': 'success',
            'message': _('Group updated successfully'),
            'group': {
                'id': cabinet_group.id,
                'name': cabinet_group.name,
                'company_name': company.name,
                'color': cabinet_group.color
            }
        })

    except Exception as e:
        logger.error(f"Error updating group {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while updating the group'),
            'errors': {'__all__': [str(e)]}
        }, status=500)


@require_permission('groups', 'delete')
@require_POST
def delete_group(request, pk):
   try:
       # Get the cabinet group instance
       cabinet_group = get_object_or_404(CabinetGroup.objects.select_related('group'), pk=pk)

       # Store the group name for the response message
       group_name = cabinet_group.get_name(request.LANGUAGE_CODE[:2])

       # Delete the system group (this will cascade delete the cabinet_group due to OneToOne relationship)
       cabinet_group.group.delete()

       return JsonResponse({
           'status': 'success',
           'message': _('Group "{}" was successfully deleted').format(group_name)
       })

   except CabinetGroup.DoesNotExist:
       logger.warning(f"Group with ID {pk} not found")
       return JsonResponse({
           'status': 'error',
           'message': _('Group not found')
       }, status=404)

   except Exception as e:
       logger.error(f"Error deleting group {pk}: {str(e)}")
       return JsonResponse({
           'status': 'error',
           'message': _('An error occurred while deleting the group')
       }, status=500)


# ----- Platform Roles (Manage roles) -----

@require_permission('roles', 'view')
@require_http_methods(['GET'])
def get_role_groups_by_companies(request):
    """Return access groups (CabinetGroup) that belong to the given company IDs. Used in Edit Role to show groups only for selected companies."""
    company_ids = request.GET.get('companies', '')
    if not company_ids:
        return JsonResponse({'groups': []})
    try:
        ids = [int(x.strip()) for x in company_ids.split(',') if x.strip()]
    except ValueError:
        return JsonResponse({'groups': []})
    accessible = get_user_companies(request.user)
    valid_company_ids = set(accessible.filter(id__in=ids).values_list('id', flat=True))
    if not valid_company_ids:
        return JsonResponse({'groups': []})
    cabinet_groups = CabinetGroup.objects.filter(
        company_id__in=valid_company_ids
    ).select_related('company', 'group').order_by('company__name', 'name')
    groups = []
    for cg in cabinet_groups:
        groups.append({
            'id': cg.group.id,
            'name': cg.name,
            'description': cg.description or '',
            'company_name': cg.company.name if cg.company else '',
            'color': cg.color or '#000000',
        })
    return JsonResponse({'groups': groups})


@require_permission('roles', 'view')
@require_http_methods(['GET'])
def get_role(request, pk):
    """Return single role as JSON for edit modal."""
    try:
        role = get_object_or_404(
            PlatformRole.objects.prefetch_related('groups', 'companies'),
            pk=pk
        )
        data = {
            'id': role.id,
            'name': role.name,
            'description': role.description or '',
            'slug': role.slug or '',
            'is_active': role.is_active,
            'color': role.color or '#6c757d',
            'order': role.order,
            'group_ids': list(role.groups.values_list('id', flat=True)),
            'company_ids': list(role.companies.values_list('id', flat=True)),
            'allowed_metrics_modules': role.allowed_metrics_modules or [],
        }
        return JsonResponse({'status': 'success', 'data': data})
    except PlatformRole.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': _('Role not found')}, status=404)
    except Exception as e:
        logger.error(f"Error getting role {pk}: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_permission('roles', 'add')
@require_POST
def create_role(request):
    try:
        name = (request.POST.get('name') or '').strip()
        if not name:
            return JsonResponse({
                'status': 'error',
                'errors': {'name': [_('Name is required')]}
            }, status=400)
        slug = (request.POST.get('slug') or '').strip()
        if not slug:
            from django.utils.text import slugify
            slug = slugify(name) or f'role-{PlatformRole.objects.count() + 1}'
        if PlatformRole.objects.filter(slug=slug).exists():
            return JsonResponse({
                'status': 'error',
                'errors': {'slug': [_('A role with this slug already exists')]}
            }, status=400)
        description = request.POST.get('description', '')
        is_active = request.POST.get('is_active') in ('1', 'true', 'on', 'yes')
        color = request.POST.get('color', '#6c757d') or '#6c757d'
        try:
            order = int(request.POST.get('order', 0))
        except (ValueError, TypeError):
            order = 0
        group_ids = request.POST.getlist('groups')
        allowed_raw = request.POST.get('allowed_metrics_modules', '')
        allowed_metrics_modules = []
        if allowed_raw:
            try:
                import json
                allowed_metrics_modules = json.loads(allowed_raw)
            except Exception:
                allowed_metrics_modules = [x.strip() for x in allowed_raw.split(',') if x.strip()]
        role = PlatformRole.objects.create(
            name=name,
            slug=slug,
            description=description,
            is_active=is_active,
            color=color,
            order=order,
            allowed_metrics_modules=allowed_metrics_modules,
        )
        if group_ids:
            role.groups.set(Group.objects.filter(id__in=group_ids))
        company_ids = request.POST.getlist('companies')
        if company_ids:
            accessible = get_user_companies(request.user)
            valid_ids = accessible.filter(id__in=company_ids).values_list('id', flat=True)
            role.companies.set(Company.objects.filter(id__in=valid_ids))
        return JsonResponse({
            'status': 'success',
            'message': _('Role created successfully'),
            'role': {'id': role.id, 'name': role.name, 'slug': role.slug, 'color': role.color},
        })
    except Exception as e:
        logger.error(f"Error creating role: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while creating the role'),
            'errors': {'__all__': [str(e)]},
        }, status=500)


@require_permission('roles', 'edit')
@require_POST
def edit_role(request, pk):
    try:
        role = get_object_or_404(PlatformRole, pk=pk)
        name = (request.POST.get('name') or '').strip()
        if not name:
            return JsonResponse({
                'status': 'error',
                'errors': {'name': [_('Name is required')]}
            }, status=400)
        slug = (request.POST.get('slug') or '').strip()
        if not slug:
            from django.utils.text import slugify
            slug = slugify(name) or role.slug
        if PlatformRole.objects.filter(slug=slug).exclude(pk=pk).exists():
            return JsonResponse({
                'status': 'error',
                'errors': {'slug': [_('A role with this slug already exists')]}
            }, status=400)
        description = request.POST.get('description', '')
        is_active = request.POST.get('is_active') in ('1', 'true', 'on', 'yes')
        color = request.POST.get('color', '#6c757d') or '#6c757d'
        try:
            order = int(request.POST.get('order', 0))
        except (ValueError, TypeError):
            order = role.order
        group_ids = request.POST.getlist('groups')
        allowed_raw = request.POST.get('allowed_metrics_modules', '')
        allowed_metrics_modules = role.allowed_metrics_modules or []
        if allowed_raw:
            try:
                import json
                allowed_metrics_modules = json.loads(allowed_raw)
            except Exception:
                allowed_metrics_modules = [x.strip() for x in allowed_raw.split(',') if x.strip()]
        role.name = name
        role.slug = slug
        role.description = description
        role.is_active = is_active
        role.color = color
        role.order = order
        role.allowed_metrics_modules = allowed_metrics_modules
        role.save()
        if group_ids is not None:
            role.groups.set(Group.objects.filter(id__in=group_ids))
        company_ids = request.POST.getlist('companies')
        if company_ids is not None:
            accessible = get_user_companies(request.user)
            valid_ids = accessible.filter(id__in=company_ids).values_list('id', flat=True)
            role.companies.set(Company.objects.filter(id__in=valid_ids))
        return JsonResponse({
            'status': 'success',
            'message': _('Role updated successfully'),
            'role': {'id': role.id, 'name': role.name, 'slug': role.slug, 'color': role.color},
        })
    except PlatformRole.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': _('Role not found')}, status=404)
    except Exception as e:
        logger.error(f"Error updating role {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while updating the role'),
            'errors': {'__all__': [str(e)]},
        }, status=500)


@require_permission('roles', 'delete')
@require_POST
def delete_role(request, pk):
    try:
        role = get_object_or_404(PlatformRole, pk=pk)
        name = role.name
        role.delete()
        return JsonResponse({
            'status': 'success',
            'message': _('Role "{}" was successfully deleted').format(name),
        })
    except PlatformRole.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': _('Role not found')}, status=404)
    except Exception as e:
        logger.error(f"Error deleting role {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('An error occurred while deleting the role'),
        }, status=500)


@require_permission('roles', 'view')
def roles_view(request):
    """List platform roles (Manage roles) with pagination. Roles and companies scoped by Access to Cabinet Management."""
    per_page = get_cabinet_table_page_size(request)
    search = request.GET.get('search', '')
    company_filter = request.GET.get('company')  # optional filter by company in list
    accessible_companies = get_user_companies(request.user)
    # Show roles that have no companies (global) or at least one company in user's accessible companies
    ids_global = PlatformRole.objects.annotate(c=Count('companies')).filter(c=0).values_list('id', flat=True)
    ids_scoped = PlatformRole.objects.filter(companies__in=accessible_companies).values_list('id', flat=True).distinct()
    role_ids = set(ids_global) | set(ids_scoped)
    roles = PlatformRole.objects.filter(id__in=role_ids).prefetch_related('groups', 'companies')
    if company_filter and company_filter.isdigit():
        roles = roles.filter(Q(companies__id=company_filter) | Q(id__in=ids_global))
    if search:
        roles = roles.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(slug__icontains=search)
        )
    roles = roles.order_by('order', 'name')
    paginator = Paginator(roles, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    all_groups = Group.objects.select_related('cabinet_details').order_by('name')
    companies = accessible_companies.order_by('name')
    context = {
        'page_obj': page_obj,
        'roles': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.count > 0,
        'current_page_size': per_page,
        'page_size_options': CABINET_TABLE_PAGE_SIZE_OPTIONS,
        'pagination_item_label': _('roles'),
        'all_groups': all_groups,
        'companies': companies,
        'active_filters': {'search': search, 'per_page': per_page, 'company': company_filter},
        'page_title': _('Manage Roles'),
        'current_language': request.LANGUAGE_CODE[:2],
    }
    return render(request, 'app_cabinet/roles.html', context)


@require_permission('users', 'edit')
@require_POST
def set_user_force_two_factor(request, pk):
    try:
        cabinet_user = get_object_or_404(CabinetUser, pk=pk)
        force_value = request.POST.get('force_two_factor')

        if force_value is None:
            return JsonResponse({
                'status': 'error',
                'message': _('Missing force_two_factor parameter')
            }, status=400)

        enabled = str(force_value).lower() in ['1', 'true', 'on', 'yes']
        cabinet_user.force_two_factor = enabled
        cabinet_user.save(update_fields=['force_two_factor'])

        message = _('Two-factor authentication requirement updated.')
        return JsonResponse({
            'status': 'success',
            'message': message,
            'force_two_factor': enabled
        })
    except CabinetUser.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': _('User not found')
        }, status=404)
    except Exception as e:
        logger.error(f"Error updating force two factor for user {pk}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': _('Failed to update two-factor authentication requirement')
        }, status=500)


@login_required
@require_http_methods(["GET"])
def cabinet_groups_guide(request):
    """Return JSON { content: html } for the Cabinet Groups guide (localized)."""
    if not has_permission(request.user, 'groups', 'view'):
        return JsonResponse({'content': ''})
    from app_conf.models import Country
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = CabinetGroupsGuide.objects.first()
    if guide:
        if country:
            trans = CabinetGroupsGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = CabinetGroupsGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def cabinet_groups_guide_translate(request):
    """API for AI translation of Cabinet Groups guide content (admin)."""
    from app_conf.models import Country
    import json
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_permission('groups', 'view')
def groups_view(request):
    per_page = get_cabinet_table_page_size(request)

    # Get filter parameters
    search = request.GET.get('search')
    company_id = request.GET.get('company')

    # Get current language
    current_language = request.LANGUAGE_CODE[:2]

    # Start with groups filtered by accessible companies (Chosen Companies)
    accessible_companies = get_user_companies(request.user)
    groups = CabinetGroup.objects.select_related(
        'company',
        'group'
    ).filter(company__in=accessible_companies)

    # Apply search filter if provided
    if search:
        groups = groups.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search) |
            Q(group__name__icontains=search)
        )

    # Apply company filter if provided
    if company_id:
        groups = groups.filter(company_id=company_id)

    # Order groups by name
    groups = groups.order_by('name')

    # Pagination
    paginator = Paginator(groups, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get companies based on user's access permissions (Chosen Companies)
    companies = get_user_companies(request.user).order_by('name')

    context = {
        'cabinet_groups': page_obj,
        'companies': companies,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.count > 0,
        'current_page_size': per_page,
        'page_size_options': CABINET_TABLE_PAGE_SIZE_OPTIONS,
        'pagination_item_label': _('groups'),
        'per_page': per_page,
        'active_filters': {
            'search': search,
            'company': company_id,
            'per_page': per_page,
        },
        'page_title': _("Groups Management"),
        'current_language': current_language
    }

    return render(request, 'app_cabinet/groups.html', context)


@require_permission('users', 'view')
def users_view(request):
   try:
       # Get filter parameters
       search = request.GET.get('search')
       company_id = request.GET.get('company')
       status = request.GET.get('status')
       department_id = request.GET.get('department')
       position_id = request.GET.get('position')
       cabinet_group_id = request.GET.get('cabinet_group')
       other_group_id = request.GET.get('other_group')
       employment_status = request.GET.get('employment_status')
       date_from = request.GET.get('date_from')
       date_to = request.GET.get('date_to')
       active_only_values = request.GET.getlist('active_only')

       if active_only_values:
           active_only = '1' if '1' in active_only_values else '0'
       else:
           active_only = '1'
       per_page = get_cabinet_table_page_size(request)

       # Initialize empty querysets
       departments = Department.objects.none()
       positions = Position.objects.none()
       cabinet_groups = CabinetGroup.objects.none()

       # Start with users query filtered by accessible companies (Chosen Companies)
       accessible_companies = get_user_companies(request.user)
       users = CabinetUser.objects.select_related(
           'user',
           'company',
           'department',
           'position'
       ).prefetch_related(
           Prefetch('user__groups', queryset=Group.objects.select_related('cabinet_details', 'cabinet_details__company')),
           Prefetch('roles', queryset=PlatformRole.objects.prefetch_related('groups', 'companies'))
       ).filter(company__in=accessible_companies)

       # Apply search filter
       if search:
           users = users.filter(
               Q(user__first_name__icontains=search) |
               Q(user__last_name__icontains=search) |
               Q(user__email__icontains=search) |
               Q(phone__icontains=search)
           )

       # Get companies based on user's access permissions (Chosen Companies)
       companies = get_user_companies(request.user).order_by('name')

       # Get current language code
       lang_code = request.LANGUAGE_CODE[:2]

       def get_group_label(group):
           """Return group name for sorting."""
           details = getattr(group, 'cabinet_details', None)
           if details:
               return getattr(details, 'name', '')
           return getattr(group, 'name', '')

       # Apply company filter
       if company_id:
           users = users.filter(company_id=company_id)

           # Get departments for selected company
           departments = Department.objects.filter(company_id=company_id)
           departments = departments.order_by('name')

           # Get base positions query
           positions = Position.objects.filter(company_id=company_id).order_by('name')

           # Get cabinet groups
           cabinet_groups = CabinetGroup.objects.filter(company_id=company_id).select_related('group').order_by('name')

       # Apply department filter
       if department_id:
           users = users.filter(department_id=department_id)
           if company_id:
               positions = positions.filter(department_id=department_id)

       # Apply position filter
       if position_id:
           users = users.filter(position_id=position_id)

       # Apply cabinet group filter
       if cabinet_group_id:
           users = users.filter(user__groups__cabinet_details__id=cabinet_group_id)

       # Apply other group filter
       if other_group_id:
           users = users.filter(
               user__groups__id=other_group_id,
               user__groups__cabinet_details__isnull=True
           )

       # Apply date filters
       if date_from:
           try:
               date_from = datetime.strptime(date_from, '%d.%m.%y').date()
               users = users.filter(start_date__date__gte=date_from)
           except ValueError:
               logger.warning(f"Invalid date_from format: {date_from}")

       if date_to:
           try:
               date_to = datetime.strptime(date_to, '%d.%m.%y').date()
               users = users.filter(
                   Q(end_date__isnull=True) |
                   Q(end_date__date__lte=date_to)
               )
           except ValueError:
               logger.warning(f"Invalid date_to format: {date_to}")

       # Apply status filters
       if status:
           status_filters = {
               'active': {'user__is_active': True},
               'inactive': {'user__is_active': False},
               'profile_completed': {'is_profile_completed': True},
               'staff': {'user__is_staff': True}
           }
           if status in status_filters:
               users = users.filter(**status_filters[status])

       # Apply active users filter (enabled by default)
       if active_only == '1':
           users = users.filter(user__is_active=True)
               
       # Apply employment status filters
       current_date = timezone.now().date()
       if employment_status:
           if employment_status == 'active':
               # Currently employed: start_date <= today and (end_date is null or end_date >= today)
               users = users.filter(
                   Q(start_date__date__lte=current_date) & 
                   (Q(end_date__isnull=True) | Q(end_date__date__gte=current_date))
               )
           elif employment_status == 'future':
               # Future employment: start_date > today
               users = users.filter(start_date__date__gt=current_date)
           elif employment_status == 'past':
               # Past employment: end_date < today
               users = users.filter(end_date__date__lt=current_date)
           elif employment_status == 'no_dates':
               # No dates specified: start_date and end_date are both null
               users = users.filter(start_date__isnull=True, end_date__isnull=True)

       # Get other groups (non-cabinet groups)
       other_groups = Group.objects.filter(
           cabinet_details__isnull=True
       ).order_by('name')

       # Order users by sort column and direction
       sort_field = request.GET.get('sort', 'name')
       sort_direction = request.GET.get('direction', 'asc')
       if sort_direction not in ('asc', 'desc'):
           sort_direction = 'asc'
       order_prefix = '' if sort_direction == 'asc' else '-'
       sort_mapping = {
           'name': [f'{order_prefix}user__last_name', f'{order_prefix}user__first_name'],
           'company': [f'{order_prefix}company__name', 'user__last_name', 'user__first_name'],
           'department': [f'{order_prefix}department__name', 'user__last_name', 'user__first_name'],
           'position': [f'{order_prefix}position__name', 'user__last_name', 'user__first_name'],
           'status': [f'{order_prefix}user__is_active', 'user__last_name', 'user__first_name'],
       }
       order_fields = sort_mapping.get(sort_field, sort_mapping['name'])
       users = users.order_by(*order_fields)

       # Pagination
       paginator = Paginator(users, per_page)
       page_number = request.GET.get('page')
       page_obj = paginator.get_page(page_number)

       # Calculate active period percentage
       for user in page_obj:
           if user.start_date and user.end_date:
               total_days = (user.end_date - user.start_date).days
               if total_days > 0:
                   current_date = timezone.now().date()
                   days_passed = (current_date - user.start_date.date()).days
                   user.activity_percentage = min(100, round((days_passed / total_days) * 100))
               else:
                   user.activity_percentage = 0
           else:
               user.activity_percentage = None

           # Add full name initials
           full_name = user.user.get_full_name()
           user.initials = ''.join(word[0].upper() for word in full_name.split() if word)

           # Modify color for inactive users
           if not user.user.is_active:
               user.display_color = '#6c757d'
               # Используем gettext для перевода
               user.display_name = f"{user.user.get_full_name()} ({_('No active')})"
           else:
               user.display_color = user.color or '#2e6da4'
               user.display_name = user.user.get_full_name()

           # Limit cabinet and other groups displayed in template
           # Prepare group listings with limits and totals
           user_groups = list(user.user.groups.all())
           cabinet_groups_full = sorted(
               (group for group in user_groups if getattr(group, 'cabinet_details', None)),
               key=lambda g: get_group_label(g).lower()
           )
           other_groups_full = sorted(
               (group for group in user_groups if not getattr(group, 'cabinet_details', None)),
               key=lambda g: getattr(g, 'name', '').lower()
           )

           user.display_cabinet_groups = cabinet_groups_full[:10]
           user.cabinet_groups_total = len(cabinet_groups_full)
           user.all_cabinet_groups = cabinet_groups_full

           user.display_other_groups = other_groups_full[:10]
           user.other_groups_total = len(other_groups_full)
           user.all_other_groups = other_groups_full

           # For Groups modal: groups from roles, additional cabinet groups, other groups
           role_group_ids = set()
           role_company_ids = set()
           for role in user.roles.all():
               for g in role.groups.all():
                   role_group_ids.add(g.id)
               for c in role.companies.all():
                   role_company_ids.add(c.id)
           groups_from_roles_list = [g for g in user_groups if g.id in role_group_ids]
           user.groups_from_roles = sorted(
               groups_from_roles_list,
               key=lambda g: get_group_label(g).lower()
           )
           # Additional = cabinet groups user has that are not from roles; if roles have companies, filter by those companies; else include all such cabinet groups (global roles)
           additional = []
           for g in cabinet_groups_full:
               if g.id in role_group_ids:
                   continue
               details = getattr(g, 'cabinet_details', None)
               if not details:
                   continue
               comp_id = getattr(details, 'company_id', None)
               if role_company_ids:
                   if comp_id is not None and comp_id in role_company_ids:
                       additional.append(g)
               else:
                   additional.append(g)
           user.additional_cabinet_groups = sorted(
               additional,
               key=lambda g: get_group_label(g).lower()
           )

           # Lazy import to avoid circular import (views imports options_view)
           from .views import get_tasks_count_for_cabinet_user
           user.tasks_count = get_tasks_count_for_cabinet_user(user)

       (
           task_reminder_submit_url,
           task_reminder_schedules_list_url,
           task_reminder_schedules_base_url,
       ) = _cabinet_users_task_reminder_urls(request)
       telegram_broadcast_submit_url = _cabinet_users_telegram_broadcast_url(request)
       email_broadcast_submit_url = _cabinet_users_email_broadcast_url(request)

       context = {
           'cabinet_users': page_obj,
           'page_obj': page_obj,
           'paginator': paginator,
           'is_paginated': paginator.count > 0,
           'current_page_size': per_page,
           'page_size_options': CABINET_TABLE_PAGE_SIZE_OPTIONS,
           'pagination_item_label': _('users'),
           'companies': companies,
           'departments': departments,
           'positions': positions,
           'cabinet_groups': cabinet_groups,
           'other_groups': other_groups,
           'platform_roles': _platform_roles_for_companies(accessible_companies),
           'active_filters': {
               'search': search,
               'company': company_id,
               'status': status,
               'department': department_id,
               'position': position_id,
               'cabinet_group': cabinet_group_id,
               'other_group': other_group_id,
               'employment_status': employment_status,
               'date_from': date_from,
               'date_to': date_to,
               'per_page': per_page,
               'active_only': active_only,
               'sort': sort_field,
               'direction': sort_direction,
           },
           'page_title': _("Users Management"),
           'task_reminder_submit_url': task_reminder_submit_url,
           'task_reminder_schedules_list_url': task_reminder_schedules_list_url,
           'task_reminder_schedules_base_url': task_reminder_schedules_base_url,
           'telegram_broadcast_submit_url': telegram_broadcast_submit_url,
           'email_broadcast_submit_url': email_broadcast_submit_url,
       }

       return render(request, 'app_cabinet/users.html', context)

   except Exception as e:
       logger.error(f"Error in users_view: {str(e)}", exc_info=True)
       messages.error(request, _("An error occurred while loading the users list"))
       return redirect('index')


def format_date(date_str):
    """Convert date string from dd.mm.yy to datetime object"""
    try:
        return datetime.strptime(date_str, '%d.%m.%y')
    except (ValueError, TypeError):
        return None


def set_user_active_status(user, start_date, end_date):
    """
    Sets user's is_active status based on start and end dates
    """
    now = timezone.now()

    # Default to active if no dates are set
    if not start_date and not end_date:
        user.is_active = True
        return user

    # Handle only start date
    if start_date and not end_date:
        user.is_active = start_date <= now
        return user

    # Handle only end date
    if end_date and not start_date:
        user.is_active = now <= end_date
        return user

    # Handle both dates
    user.is_active = start_date <= now <= end_date
    return user


@require_permission('users', 'add')
@require_POST
def create_user(request):
    try:
        # Check if we're creating a new user or using an existing one
        mode = request.GET.get('mode', 'new')
        
        # Basic validation based on mode
        if mode == 'new':
            # Validating new user creation
            required_fields = {
                'email': _('Email is required'),
                'first_name': _('First name is required'),
                'last_name': _('Last name is required'),
                'company': _('Company is required'),
            }
            
            errors = {}
            for field, error_message in required_fields.items():
                if not request.POST.get(field):
                    errors[field] = [error_message]
    
            if errors:
                return JsonResponse({
                    'status': 'error',
                    'errors': errors
                }, status=400)

            # Validate email format
            email = (request.POST.get('email') or '').strip().lower()
            if email:
                try:
                    from django.core.validators import validate_email
                    validate_email(email)
                except ValidationError:
                    errors.setdefault('email', []).append(_('Enter a valid email address.'))
            if errors:
                return JsonResponse({'status': 'error', 'errors': errors}, status=400)

            # Validate phone format if provided
            phone = request.POST.get('phone', '').strip()
            if phone:
                try:
                    from .security_validators import PersonalCabinetSecurityValidator
                    PersonalCabinetSecurityValidator.validate_phone_number(phone)
                except ValidationError as e:
                    msg = (e.messages[0] if getattr(e, 'messages', None) and len(e.messages) > 0 else str(e))
                    errors.setdefault('phone', []).append(msg)
            if errors:
                return JsonResponse({'status': 'error', 'errors': errors}, status=400)
    
            # Check if email already exists
            if User.objects.filter(Q(email=email) | Q(username=email)).exists():
                return JsonResponse({
                    'status': 'error',
                    'errors': {
                        'email': [_('A user with this email already exists')]
                    },
                    'message': _('A user with this email already exists in the system')
                }, status=400)
        else:
            # Validating existing user selection
            required_fields = {
                'user_id': _('User selection is required'),
                'company': _('Company is required'),
            }
            
            errors = {}
            for field, error_message in required_fields.items():
                if not request.POST.get(field):
                    errors[field] = [error_message]
    
            if errors:
                return JsonResponse({
                    'status': 'error',
                    'errors': errors
                }, status=400)
                
            # Check if user already has a cabinet user
            user_id = request.POST.get('user_id')
            if CabinetUser.objects.filter(user_id=user_id).exists():
                return JsonResponse({
                    'status': 'error',
                    'errors': {
                        'user_id': [_('This user already has a Cabinet User profile')]
                    },
                    'message': _('This user already has a Cabinet User profile')
                }, status=400)

        # Validate avatar if provided
        avatar = request.FILES.get('avatar')
        if avatar:
            # Check file size (2MB limit)
            if avatar.size > 2 * 1024 * 1024:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'avatar': [_('Image size should not exceed 2MB')]}
                }, status=400)

            # Check file type
            allowed_types = ['image/jpeg', 'image/png', 'image/jpg']
            if avatar.content_type not in allowed_types:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'avatar': [_('Please upload a JPEG or PNG image')]}
                }, status=400)

            # Check image dimensions
            from PIL import Image
            try:
                img = Image.open(avatar)
                if img.width < 200 or img.height < 200:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'avatar': [_('Image dimensions should be at least 200x200 pixels')]}
                    }, status=400)
            except Exception as e:
                logger.error(f"Error validating image: {str(e)}")
                return JsonResponse({
                    'status': 'error',
                    'errors': {'avatar': [_('Invalid image file')]}
                }, status=400)

        # Validate company
        try:
            company = Company.objects.get(id=request.POST.get('company'))
        except Company.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'errors': {'company': [_('Selected company does not exist')]}
            }, status=400)

        # Validate department if provided
        department = None
        department_id = request.POST.get('department')
        if department_id:
            try:
                department = Department.objects.get(id=department_id, company=company)
            except Department.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'department': [_('Selected department does not exist in this company')]}
                }, status=400)

        # Validate position if provided
        position = None
        position_id = request.POST.get('position')
        if position_id:
            try:
                position = Position.objects.get(id=position_id, company=company)
            except Position.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'position': [_('Selected position does not exist in this company')]}
                }, status=400)

        # Validate groups if provided
        groups_to_add = []
        group_ids = request.POST.getlist('groups[]')
        if group_ids:
            for group_id in group_ids:
                try:
                    group = Group.objects.get(id=group_id)
                    cabinet_group = CabinetGroup.objects.get(group=group, company=company)
                    groups_to_add.append(group)
                except (Group.DoesNotExist, CabinetGroup.DoesNotExist):
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'groups': [_('One or more selected groups do not exist in this company')]}
                    }, status=400)

        # Handle datetime fields
        start_datetime = None
        end_datetime = None

        start_date = request.POST.get('start_date')
        start_time = request.POST.get('start_time', '09:00')
        end_date = request.POST.get('end_date')
        end_time = request.POST.get('end_time', '18:00')

        if start_date:
            try:
                parsed_date = datetime.strptime(start_date, '%d.%m.%y')
                start_datetime = timezone.make_aware(
                    datetime.combine(
                        parsed_date.date(),
                        datetime.strptime(start_time, '%H:%M').time()
                    )
                )
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'start_date': [_('Invalid date format. Use DD.MM.YY')]}
                }, status=400)

        if end_date:
            try:
                parsed_date = datetime.strptime(end_date, '%d.%m.%y')
                end_datetime = timezone.make_aware(
                    datetime.combine(
                        parsed_date.date(),
                        datetime.strptime(end_time, '%H:%M').time()
                    )
                )
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'end_date': [_('Invalid date format. Use DD.MM.YY')]}
                }, status=400)

        if start_datetime and end_datetime and start_datetime > end_datetime:
            return JsonResponse({
                'status': 'error',
                'errors': {'end_date': [_('End date must be after start date')]}
            }, status=400)

        # Get or create user based on mode
        if mode == 'new':
            # Create new user with default active status
            user = User.objects.create_user(
                username=request.POST['email'],
                email=request.POST['email'],
                first_name=request.POST['first_name'],
                last_name=request.POST['last_name']
            )
        else:
            # Get existing user
            try:
                user = User.objects.get(id=request.POST.get('user_id'))
            except User.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'errors': {'user_id': [_('Selected user does not exist')]}
                }, status=400)

        force_two_factor = request.POST.get('force_two_factor') == 'on'

        try:
            from .security_validators import PersonalCabinetSecurityValidator
            telegram_chat_id = PersonalCabinetSecurityValidator.validate_telegram_chat_id(
                request.POST.get('telegram_chat_id', '')
            )
        except ValidationError as e:
            msg = (e.messages[0] if getattr(e, 'messages', None) and len(e.messages) > 0 else str(e))
            return JsonResponse({
                'status': 'error',
                'errors': {'telegram_chat_id': [msg]}
            }, status=400)

        # Set active status based on dates
        user = set_user_active_status(user, start_datetime, end_datetime)
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.save()

        # Create cabinet user
        cabinet_user = CabinetUser.objects.create(
            user=user,
            company=company,
            department=department,
            position=position,
            phone=request.POST.get('phone'),
            telegram_chat_id=telegram_chat_id,
            start_date=start_datetime,
            end_date=end_datetime,
            is_profile_completed=True,
            color=request.POST.get('color', '#2e6da4'),
            force_two_factor=force_two_factor
        )

        # Handle avatar upload
        if avatar:
            cabinet_user.avatar = avatar
            cabinet_user.save()

        # Add groups
        for group in groups_to_add:
            user.groups.add(group)

        # Add platform roles
        role_ids = request.POST.getlist('roles[]')
        if role_ids:
            from .models import PlatformRole
            cabinet_user.roles.set(PlatformRole.objects.filter(id__in=role_ids))
        if request.POST.get('apply_groups_from_roles') == 'on':
            for role in cabinet_user.roles.all():
                for grp in role.groups.all():
                    user.groups.add(grp)
        # Additional cabinet groups (filtered by companies of selected roles)
        additional_group_ids = request.POST.getlist('additional_cabinet_groups[]')
        if additional_group_ids and cabinet_user.roles.exists():
            role_company_ids = set(
                cabinet_user.roles.values_list('companies__id', flat=True).distinct()
            )
            role_company_ids.discard(None)
            if role_company_ids:
                for gid in additional_group_ids:
                    try:
                        cg = CabinetGroup.objects.select_related('group').get(
                            group_id=gid, company_id__in=role_company_ids
                        )
                        user.groups.add(cg.group)
                    except (CabinetGroup.DoesNotExist, ValueError):
                        pass

        # Handle quiz assignments
        assigned_quiz_ids = request.POST.getlist('assigned_quizzes[]')
        if assigned_quiz_ids:
            try:
                from app_study.models import Quiz
                
                # Add quiz assignments
                for quiz_id in assigned_quiz_ids:
                    try:
                        quiz = Quiz.objects.get(id=quiz_id)
                        quiz.cabinet_users.add(cabinet_user)
                    except Quiz.DoesNotExist:
                        logger.warning(f"Quiz with ID {quiz_id} not found")
                        
            except Exception as e:
                logger.error(f"Error handling quiz assignments: {str(e)}")

        # Get current language for response
        current_language = request.LANGUAGE_CODE[:2]

        # Prepare response with localized names
        response_data = {
            'status': 'success',
            'message': _('User created successfully'),
            'user': {
                'id': cabinet_user.id,
                'full_name': user.get_full_name(),
                'email': user.email,
                'phone': cabinet_user.phone or '',
                'telegram_chat_id': cabinet_user.telegram_chat_id or '',
                'color': cabinet_user.color,
                'avatar_url': cabinet_user.avatar.url if cabinet_user.avatar else None,
                'company': company.name,
                'department': department.get_name(current_language) if department else None,
                'position': position.get_name(current_language) if position else None,
                'start_date': start_datetime.strftime('%d.%m.%y %H:%M') if start_datetime else None,
                'end_date': end_datetime.strftime('%d.%m.%y %H:%M') if end_datetime else None,
                'is_active': user.is_active,
                'is_staff': user.is_staff,
                'groups': [
                    {
                        'id': group.id,
                        'name': group.cabinet_details.get_name(current_language) if hasattr(group,'cabinet_details') else group.name
                    } for group in user.groups.all()
                ]
            }
        }

        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error creating user: {str(e)}")
        if 'user' in locals() and mode == 'new':
            # Only delete the user if we created a new one and an error occurred
            user.delete()
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'errors': {'__all__': [_('An error occurred while creating the user')]}
        }, status=500)


@login_required
def get_user(request, pk):
    try:
        cabinet_user = CabinetUser.objects.select_related(
            'user',
            'company',
            'department',
            'position'
        ).get(pk=pk)

        # Отримуємо окремо cabinet groups та other groups
        cabinet_groups = []
        other_groups = []

        for group in cabinet_user.user.groups.all():
            try:
                if hasattr(group, 'cabinet_details'):
                    cabinet_groups.append(group.id)
                else:
                    other_groups.append(group.id)
            except:
                other_groups.append(group.id)

        data = {
            'status': 'success',
            'user': {
                'id': cabinet_user.id,
                'first_name': cabinet_user.user.first_name,
                'last_name': cabinet_user.user.last_name,
                'email': cabinet_user.user.email,
                'phone': cabinet_user.phone,
                'color': cabinet_user.color,
                'avatar_url': cabinet_user.avatar.url if cabinet_user.avatar else None,
                'company': cabinet_user.company_id,
                'department': cabinet_user.department_id,
                'position': cabinet_user.position_id,
                'start_date': cabinet_user.start_date.strftime(
                    '%Y-%m-%d %H:%M:%S') if cabinet_user.start_date else None,
                'end_date': cabinet_user.end_date.strftime('%Y-%m-%d %H:%M:%S') if cabinet_user.end_date else None,
                'is_active': cabinet_user.user.is_active,
                'is_staff': cabinet_user.user.is_staff,
                'cabinet_groups': cabinet_groups,
                'other_groups': other_groups
            }
        }

        return JsonResponse(data)
    except CabinetUser.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': _('User not found')
        }, status=404)
@login_required
def get_department_users(request, department_id):
    try:
        # Retrieve the department and ensure it exists
        department = get_object_or_404(Department, id=department_id, company=request.user.cabinetuser.company)

        # Fetch all users associated with the department
        users = CabinetUser.objects.filter(department=department).select_related('user')

        # Serialize user data
        data = [
            {
                'id': user.id,
                'full_name': f"{user.user.first_name} {user.user.last_name}",
                'avatar_url': user.avatar.url if user.avatar else None,
                'email': user.user.email,
                'position': user.position.get_name('en') or user.position.get_name() if user.position else None,
                'phone': user.phone,
                'is_active': user.user.is_active,
            }
            for user in users
        ]

        return JsonResponse(data, safe=False)

    except Exception as e:
        logger.error(f"Error fetching users for department {department_id}: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@require_permission('users', 'edit')
@require_POST
def update_user(request, pk):
    from .input_security import PersonalCabinetSecurityValidator, PersonalCabinetAuditLogger, get_client_ip
    
    try:
        cabinet_user = get_object_or_404(CabinetUser.objects.select_related('user'), pk=pk)
        user = cabinet_user.user
        ip_address = get_client_ip(request)

        logger.debug(f"Updating user {pk}")
        logger.debug(f"POST data: {request.POST}")
        logger.debug(f"FILES data: {request.FILES}")

        # Security: Validate privilege escalation attempts
        validator = PersonalCabinetSecurityValidator()
        try:
            validator.validate_privilege_escalation_attempt(
                request.user,
                target_user_id=user.id,
                company_change=request.POST.get('company'),
                department_change=request.POST.get('department'),
                position_change=request.POST.get('position')
            )
        except ValidationError as e:
            PersonalCabinetAuditLogger.log_privilege_escalation_attempt(
                request.user, f'admin_user_update_user_{pk}', ip_address, str(e)
            )
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=403)

        # Validate and sanitize input fields
        sanitized_data = {}
        try:
            sanitized_data['first_name'] = validator.validate_and_sanitize_text_field(
                request.POST.get('first_name', ''), 'first_name', 30
            )
            sanitized_data['last_name'] = validator.validate_and_sanitize_text_field(
                request.POST.get('last_name', ''), 'last_name', 30  
            )
            sanitized_data['email'] = request.POST.get('email', '').strip().lower()
            
            # Validate email format
            from django.core.validators import validate_email
            validate_email(sanitized_data['email'])
            
        except ValidationError as e:
            PersonalCabinetAuditLogger.log_malicious_input_attempt(
                request.user, 'admin_user_update', str(request.POST), ip_address
            )
            return JsonResponse({
                'status': 'error',
                'errors': {'input_validation': [str(e)]}
            }, status=400)

        # Validate required fields
        required_fields = {
            'first_name': _('First name is required'),
            'last_name': _('Last name is required'), 
            'email': _('Email is required'),
            'company': _('Company is required')
        }

        errors = {}
        for field, error_message in required_fields.items():
            value = sanitized_data.get(field) or request.POST.get(field)
            if not value:
                errors[field] = [error_message]

        if errors:
            logger.warning(f"Validation errors: {errors}")
            return JsonResponse({
                'status': 'error',
                'errors': errors
            }, status=400)

        try:
            # Handle avatar upload if provided
            avatar = request.FILES.get('avatar')
            if avatar:
                # Check file size (2MB limit)
                if avatar.size > 2 * 1024 * 1024:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'avatar': [_('Image size should not exceed 2MB')]}
                    }, status=400)

                # Check file type
                allowed_types = ['image/jpeg', 'image/png', 'image/jpg']
                if avatar.content_type not in allowed_types:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'avatar': [_('Please upload a JPEG or PNG image')]}
                    }, status=400)

                # Check image dimensions
                try:
                    from PIL import Image
                    img = Image.open(avatar)
                    if img.width < 200 or img.height < 200:
                        return JsonResponse({
                            'status': 'error',
                            'errors': {'avatar': [_('Image dimensions should be at least 200x200 pixels')]}
                        }, status=400)
                except Exception as e:
                    logger.error(f"Error validating image: {str(e)}")
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'avatar': [_('Invalid image file')]}
                    }, status=400)

                # If there's an existing avatar, delete it
                if cabinet_user.avatar:
                    cabinet_user.avatar.delete(save=False)

                # Save new avatar
                cabinet_user.avatar = avatar

            # Handle avatar removal if requested
            elif request.POST.get('remove_avatar') == 'true' and cabinet_user.avatar:
                cabinet_user.avatar.delete()
                cabinet_user.avatar = None

            # Handle datetime fields
            start_date = request.POST.get('start_date')
            start_time = request.POST.get('start_time', '09:00')
            end_date = request.POST.get('end_date')
            end_time = request.POST.get('end_time', '18:00')

            logger.debug(f"Processing dates - Start: {start_date} {start_time}, End: {end_date} {end_time}")

            start_datetime = None
            end_datetime = None

            if start_date:
                try:
                    parsed_date = datetime.strptime(start_date, '%d.%m.%y')
                    start_datetime = timezone.make_aware(
                        datetime.combine(
                            parsed_date.date(),
                            datetime.strptime(start_time, '%H:%M').time()
                        )
                    )
                except ValueError:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'start_date': [_('Invalid date format. Use DD.MM.YY')]}
                    }, status=400)

            if end_date:
                try:
                    parsed_date = datetime.strptime(end_date, '%d.%m.%y')
                    end_datetime = timezone.make_aware(
                        datetime.combine(
                            parsed_date.date(),
                            datetime.strptime(end_time, '%H:%M').time()
                        )
                    )
                except ValueError:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'end_date': [_('Invalid date format. Use DD.MM.YY')]}
                    }, status=400)

            if start_datetime and end_datetime and start_datetime > end_datetime:
                logger.warning("End date is before start date")
                return JsonResponse({
                    'status': 'error',
                    'errors': {'end_date': [_('End date must be after start date')]}
                }, status=400)

            # Update User basic info with sanitized data
            user.first_name = sanitized_data['first_name']
            user.last_name = sanitized_data['last_name']
            user.email = sanitized_data['email']
            user.username = sanitized_data['email']
            
            # Handle password change if requested
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')
            force_password_change = request.POST.get('force_password_change') == 'on'
            
            if password1 and password2:
                # Validate passwords
                if password1 != password2:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'password2': [_('Passwords do not match')]}
                    }, status=400)
                
                if len(password1) < 8:
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'password1': [_('Password must be at least 8 characters long')]}
                    }, status=400)
                
                # Check if new password is the same as current password
                if user.check_password(password1):
                    return JsonResponse({
                        'status': 'error',
                        'errors': {'password1': [_('The new password must be different from the current password')]}
                    }, status=400)
                
                # Set the new password
                user.set_password(password1)
                
                # If force_password_change is checked, we'll set a flag on the user session
                # indicating they need to change their password on next login
                if force_password_change:
                    # We'll use UserActivity to track that this user needs to change their password
                    UserActivity.objects.create(
                        user=user,
                        action='password_reset',
                        details={'force_change': True, 'reset_by': request.user.username}
                    )
                
                logger.info(f"Password changed for user {user.username}. Force password change: {force_password_change}")
            
            # Set active status based on dates
            now = timezone.now()
            if start_datetime or end_datetime:
                if start_datetime and end_datetime:
                    user.is_active = start_datetime <= now <= end_datetime
                elif start_datetime:
                    user.is_active = start_datetime <= now
                else:  # only end_datetime
                    user.is_active = now <= end_datetime
            else:
                # If no dates set, use the checkbox value
                user.is_active = request.POST.get('is_active') == 'on'

            user.is_staff = request.POST.get('is_staff') == 'on'
            user.save()

            # Update cabinet user
            cabinet_user.company_id = request.POST.get('company')
            cabinet_user.department_id = request.POST.get('department') or None
            cabinet_user.position_id = request.POST.get('position') or None
            cabinet_user.phone = request.POST.get('phone')
            try:
                from .security_validators import PersonalCabinetSecurityValidator as CabinetFieldValidator
                cabinet_user.telegram_chat_id = CabinetFieldValidator.validate_telegram_chat_id(
                    request.POST.get('telegram_chat_id', '')
                )
            except ValidationError as e:
                msg = (e.messages[0] if getattr(e, 'messages', None) and len(e.messages) > 0 else str(e))
                return JsonResponse({
                    'status': 'error',
                    'errors': {'telegram_chat_id': [msg]}
                }, status=400)
            cabinet_user.start_date = start_datetime
            cabinet_user.end_date = end_datetime
            cabinet_user.color = request.POST.get('color', '#2e6da4')
            cabinet_user.force_two_factor = request.POST.get('force_two_factor') == 'on'
            cabinet_user.save()

            logger.debug(
                f"Updated cabinet user info: company={cabinet_user.company_id}, "
                f"department={cabinet_user.department_id}, position={cabinet_user.position_id}"
            )

            # Update platform roles first (needed for apply_groups_from_roles and additional_cabinet_groups)
            role_ids = request.POST.getlist('roles[]')
            if role_ids is not None:
                from .models import PlatformRole
                cabinet_user.roles.set(PlatformRole.objects.filter(id__in=role_ids))

            # Build full group list: cabinet from form + other (preserved) + role groups + additional cabinet groups
            cabinet_group_ids = request.POST.getlist('cabinet_groups[]')
            cabinet_group_ids_in_db = list(CabinetGroup.objects.values_list('group_id', flat=True))
            other_group_ids = []
            for grp in user.groups.all():
                if grp.id not in cabinet_group_ids_in_db:
                    other_group_ids.append(str(grp.id))
            all_group_ids = list(cabinet_group_ids) + other_group_ids

            if request.POST.get('apply_groups_from_roles') == 'on':
                for role in cabinet_user.roles.all():
                    for grp in role.groups.all():
                        all_group_ids.append(str(grp.id))
            additional_group_ids = request.POST.getlist('additional_cabinet_groups[]')
            if additional_group_ids and cabinet_user.roles.exists():
                role_company_ids = set(
                    cabinet_user.roles.values_list('companies__id', flat=True).distinct()
                )
                role_company_ids.discard(None)
                for gid in additional_group_ids:
                    try:
                        if role_company_ids:
                            cg = CabinetGroup.objects.get(
                                group_id=gid, company_id__in=role_company_ids
                            )
                        else:
                            cg = CabinetGroup.objects.filter(group_id=gid).first()
                        if cg:
                            all_group_ids.append(str(cg.group_id))
                    except (CabinetGroup.DoesNotExist, ValueError):
                        pass
            all_group_ids = list(dict.fromkeys(all_group_ids))

            logger.debug(f"All group IDs to set: {all_group_ids}")

            user.groups.clear()
            for group_id in all_group_ids:
                try:
                    grp = Group.objects.get(pk=group_id)
                    user.groups.add(grp)
                except Group.DoesNotExist:
                    pass

            # Log the admin user update for audit purposes
            changed_fields = []
            if user.first_name != sanitized_data['first_name']:
                changed_fields.append('first_name')
            if user.last_name != sanitized_data['last_name']:
                changed_fields.append('last_name')
            if user.email != sanitized_data['email']:
                changed_fields.append('email')
            if password1:
                changed_fields.append('password')
            if request.POST.get('company') != str(cabinet_user.company_id):
                changed_fields.append('company')
            if request.POST.get('department') != str(cabinet_user.department_id or ''):
                changed_fields.append('department')
            if request.POST.get('position') != str(cabinet_user.position_id or ''):
                changed_fields.append('position')
            if all_group_ids:
                changed_fields.append('groups')
            
            PersonalCabinetAuditLogger.log_profile_update(
                request.user, changed_fields, ip_address, target_user=user
            )

            # Create success message
            success_message = _('User updated successfully')
            if password1 and password2:
                if force_password_change:
                    success_message = _('User updated successfully. Password has been changed and user will be required to set a new password at next login.')
                else:
                    success_message = _('User updated successfully. Password has been changed.')
            
            # Return success response with updated user data
            return JsonResponse({
                'status': 'success',
                'message': success_message,
                'user': {
                    'id': cabinet_user.id,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                    'phone': cabinet_user.phone,
                    'color': cabinet_user.color,
                    'avatar_url': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'company': cabinet_user.company_id,
                    'department': cabinet_user.department_id,
                    'position': cabinet_user.position_id,
                    'is_active': user.is_active,
                    'is_staff': user.is_staff,
                    'cabinet_groups': [g for g in all_group_ids if int(g) in cabinet_group_ids_in_db],
                    'other_groups': [g for g in all_group_ids if int(g) not in cabinet_group_ids_in_db],
                    'start_date': start_datetime.isoformat() if start_datetime else None,
                    'end_date': end_datetime.isoformat() if end_datetime else None
                }
            })

        except Exception as e:
            logger.exception("Error updating user details")
            return JsonResponse({
                'status': 'error',
                'errors': {'__all__': [str(e)]}
            }, status=400)

    except Exception as e:
        logger.exception(f"Unexpected error updating user {pk}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@require_permission('users', 'delete')
@require_POST
def delete_user(request, pk):
    try:
        user = CabinetUser.objects.get(pk=pk)
        user_name = f"{user.user.first_name} {user.user.last_name}" or user.user.email
        user.user.delete()
        return JsonResponse({'status': 'success', 'message': f'User "{user_name}" deleted successfully.'})
    except CabinetUser.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def get_system_users_not_in_cabinet(request):
    """Get Django auth users that don't have corresponding CabinetUser records"""
    try:
        # Get IDs of users who already have CabinetUser records
        cabinet_user_ids = CabinetUser.objects.all().values_list('user_id', flat=True)
        
        # Find all system users who don't have CabinetUser records
        system_users = User.objects.exclude(id__in=cabinet_user_ids).filter(is_active=True)
        
        # Apply search filter if provided
        search_query = request.GET.get('query', '')
        if search_query:
            system_users = system_users.filter(
                Q(username__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query)
            )
        
        # Format user data
        users_data = [{
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'full_name': user.get_full_name() or user.username,
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'date_joined': user.date_joined.strftime('%d.%m.%Y') if user.date_joined else None
        } for user in system_users]
        
        return JsonResponse({
            'status': 'success',
            'users': users_data
        })
        
    except Exception as e:
        logger.error(f"Error getting system users: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@require_permission('users', 'export')
def export_users_excel(request):
    """Export cabinet users to Excel with styling"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils.translation import gettext as _
    from datetime import datetime
    import io
    
    try:
        # Get filter parameters (same as in users_view function)
        search = request.GET.get('search')
        company_id = request.GET.get('company')
        status = request.GET.get('status')
        department_id = request.GET.get('department')
        position_id = request.GET.get('position')
        cabinet_group_id = request.GET.get('cabinet_group')
        other_group_id = request.GET.get('other_group')
        employment_status = request.GET.get('employment_status')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        # Start with all users query
        users = CabinetUser.objects.select_related(
            'user',
            'company',
            'department',
            'position'
        ).prefetch_related(
            Prefetch('user__groups', queryset=Group.objects.select_related('cabinet_details'))
        ).all()
        
        # Apply search filter
        if search:
            users = users.filter(
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search) |
                Q(phone__icontains=search)
            )
            
        # Apply company filter
        if company_id:
            users = users.filter(company_id=company_id)
            
        # Apply department filter
        if department_id:
            users = users.filter(department_id=department_id)
            
        # Apply position filter
        if position_id:
            users = users.filter(position_id=position_id)
            
        # Apply cabinet group filter
        if cabinet_group_id:
            users = users.filter(user__groups__cabinet_details__id=cabinet_group_id)
            
        # Apply other group filter
        if other_group_id:
            users = users.filter(
                user__groups__id=other_group_id,
                user__groups__cabinet_details__isnull=True
            )
            
        # Apply date filters
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%d.%m.%y').date()
                users = users.filter(start_date__date__gte=date_from)
            except ValueError:
                logger.warning(f"Invalid date_from format: {date_from}")
                
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%d.%m.%y').date()
                users = users.filter(
                    Q(end_date__isnull=True) |
                    Q(end_date__date__lte=date_to)
                )
            except ValueError:
                logger.warning(f"Invalid date_to format: {date_to}")
                
        # Apply status filters
        if status:
            status_filters = {
                'active': {'user__is_active': True},
                'inactive': {'user__is_active': False},
                'profile_completed': {'is_profile_completed': True},
                'staff': {'user__is_staff': True}
            }
            if status in status_filters:
                users = users.filter(**status_filters[status])
                
        # Apply employment status filters
        current_date = timezone.now().date()
        if employment_status:
            if employment_status == 'active':
                users = users.filter(
                    Q(start_date__date__lte=current_date) & 
                    (Q(end_date__isnull=True) | Q(end_date__date__gte=current_date))
                )
            elif employment_status == 'future':
                users = users.filter(start_date__date__gt=current_date)
            elif employment_status == 'past':
                users = users.filter(end_date__date__lt=current_date)
            elif employment_status == 'no_dates':
                users = users.filter(start_date__isnull=True, end_date__isnull=True)
                
        # Order users
        users = users.order_by('user__first_name', 'user__last_name')
        
        # Get current language for localization
        current_language = request.LANGUAGE_CODE[:2]
        
        # Create a workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = _("Users")[:31]  # Excel sheet names are limited to 31 chars
        
        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3F6791', end_color='3F6791', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=10)
        data_font_bold = Font(name='Arial', size=10, bold=True)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        active_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        inactive_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define headers
        headers = [
            _("Full Name"),
            _("Email"),
            _("Phone"),
            _("Company"),
            _("Department"),
            _("Position"),
            _("Status"),
            _("Cabinet Groups"),
            _("Other Groups"),
            _("Start Date"),
            _("End Date")
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # Set column widths
        column_widths = [30, 30, 15, 20, 25, 25, 15, 30, 30, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        # Write data rows
        for row_num, cabinet_user in enumerate(users, 2):
            user = cabinet_user.user
            
            # Prepare color from hex format
            user_color = cabinet_user.color or "#2e6da4"
            # Remove # from hex color if present
            if user_color.startswith('#'):
                user_color = user_color[1:]
            
            # Create a fill with the user's color (slightly transparent for readability)
            try:
                # Creating a lighter version of the color for better readability
                user_color_fill = PatternFill(
                    start_color=user_color,
                    end_color=user_color,
                    fill_type='solid'
                )
                
                # Create a lighter version for row background
                # This is a hacky approach since openpyxl doesn't support opacity
                # We'll just use the color as is but could implement a function to lighten it if needed
                light_color_fill = PatternFill(
                    start_color=user_color,
                    end_color=user_color,
                    fill_type='solid'
                )
            except:
                # Fallback to default color if invalid hex
                user_color_fill = PatternFill(start_color="2E6DA4", end_color="2E6DA4", fill_type='solid')
                light_color_fill = PatternFill(start_color="E6F0F9", end_color="E6F0F9", fill_type='solid')
            
            # Apply custom text color based on background for better contrast
            # For dark backgrounds, use white text; for light backgrounds, use dark text
            # This is a simple way to determine contrast - a more sophisticated algorithm could be used
            is_dark_color = sum(int(user_color[i:i+2], 16) for i in (0, 2, 4)) < 382
            text_color = 'FFFFFF' if is_dark_color else '000000'
            colored_font = Font(name='Arial', size=10, bold=True, color=text_color)
            
            # Full Name - apply user color as fill
            cell = ws.cell(row=row_num, column=1, value=user.get_full_name())
            cell.font = colored_font
            cell.alignment = data_alignment
            cell.border = thin_border
            cell.fill = user_color_fill
            
            # Email
            cell = ws.cell(row=row_num, column=2, value=user.email)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Phone
            cell = ws.cell(row=row_num, column=3, value=cabinet_user.phone or "-")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Company
            cell = ws.cell(row=row_num, column=4, value=cabinet_user.company.name if cabinet_user.company else "-")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Department
            if cabinet_user.department:
                dept_name = cabinet_user.department.get_name(current_language)
            else:
                dept_name = "-"
                
            cell = ws.cell(row=row_num, column=5, value=dept_name)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Position
            pos_name = cabinet_user.position.get_name(current_language) if cabinet_user.position else "-"
                
            cell = ws.cell(row=row_num, column=6, value=pos_name)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Status
            status_text = _("Active") if user.is_active else _("Inactive")
            if user.is_staff:
                status_text += ", " + _("Staff")
            if cabinet_user.is_profile_completed:
                status_text += ", " + _("Profile Completed")
                
            cell = ws.cell(row=row_num, column=7, value=status_text)
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Apply cell background based on active status
            if user.is_active:
                cell.fill = active_fill
            else:
                cell.fill = inactive_fill
                
            # Cabinet Groups
            cabinet_groups = []
            for group in user.groups.all():
                try:
                    if hasattr(group, 'cabinet_details'):
                        group_name = group.cabinet_details.get_name(current_language) or ''
                        if group_name:
                            cabinet_groups.append(group_name)
                except:
                    pass
                    
            cell = ws.cell(row=row_num, column=8, value=", ".join(cabinet_groups) or "-")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Other Groups
            other_groups = []
            for group in user.groups.all():
                if not hasattr(group, 'cabinet_details'):
                    other_groups.append(group.name)
                    
            cell = ws.cell(row=row_num, column=9, value=", ".join(other_groups) or "-")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            
            # Start Date
            start_date = ""
            if cabinet_user.start_date:
                start_date = cabinet_user.start_date.strftime('%d.%m.%Y')
                
            cell = ws.cell(row=row_num, column=10, value=start_date or "-")
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # End Date
            end_date = ""
            if cabinet_user.end_date:
                end_date = cabinet_user.end_date.strftime('%d.%m.%Y')
                
            cell = ws.cell(row=row_num, column=11, value=end_date or "-")
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=Cabinet_Users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Save workbook to response
        wb.save(response)
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting users to Excel: {str(e)}", exc_info=True)
        messages.error(request, _("An error occurred while exporting users"))
        return redirect('users')


# Company management views
@login_required
def get_companies(request):
    """Get all companies for the current user"""
    try:
        user_companies = get_user_companies(request.user)
        companies = Company.objects.filter(id__in=user_companies)
        
        companies_data = []
        for company in companies:
            companies_data.append({
                'id': company.id,
                'name': company.name,
                'group_name': company.group.name if company.group else None,
            })
        
        return JsonResponse({'companies': companies_data})
    except Exception as e:
        logger.error(f"Error getting companies: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_company(request, pk):
    """Get a specific company by ID"""
    try:
        user_companies = get_user_companies(request.user)
        company = get_object_or_404(Company, id=pk, id__in=user_companies)
        
        company_data = {
            'id': company.id,
            'name': company.name,
            'group_id': company.group.id if company.group else None,
            'group_name': company.group.name if company.group else None,
        }
        
        return JsonResponse(company_data)
    except Exception as e:
        logger.error(f"Error getting company {pk}: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_permission('org_structure', 'add_companies')
@require_POST
def add_company(request):
    """Add a new company"""
    try:
        name = request.POST.get('name', '').strip()
        
        if not name:
            return JsonResponse({'error': _('Company name is required')}, status=400)
        
        # Check if company already exists
        if Company.objects.filter(name=name).exists():
            return JsonResponse({'error': _('Company with this name already exists')}, status=400)
        
        # Create the company
        company = Company.objects.create(name=name)
        
        # Log the action
        logger.info(f"Company created: {company.name} by user {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': _('Company added successfully'),
            'company': {
                'id': company.id,
                'name': company.name,
                'group_name': company.group.name if company.group else None,
            }
        })
        
    except Exception as e:
        logger.error(f"Error adding company: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_permission('org_structure', 'edit_companies')
@require_POST
def edit_company(request, pk):
    """Edit an existing company"""
    try:
        user_companies = get_user_companies(request.user)
        company = get_object_or_404(Company, id=pk, id__in=user_companies)
        
        name = request.POST.get('name', '').strip()
        
        if not name:
            return JsonResponse({'error': _('Company name is required')}, status=400)
        
        # Check if another company with this name exists
        if Company.objects.filter(name=name).exclude(id=pk).exists():
            return JsonResponse({'error': _('Company with this name already exists')}, status=400)
        
        # Update the company
        old_name = company.name
        company.name = name
        company.save()
        
        # Update the associated group name if it exists
        if company.group:
            company.group.name = name
            company.group.save()
        
        # Log the action
        logger.info(f"Company updated: {old_name} -> {company.name} by user {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': _('Company updated successfully'),
            'company': {
                'id': company.id,
                'name': company.name,
                'group_name': company.group.name if company.group else None,
            }
        })
        
    except Exception as e:
        logger.error(f"Error editing company {pk}: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_permission('org_structure', 'delete_companies')
@require_POST
def delete_company(request, pk):
    """Delete a company"""
    try:
        user_companies = get_user_companies(request.user)
        company = get_object_or_404(Company, id=pk, id__in=user_companies)
        
        # Check if company has associated users
        if CabinetUser.objects.filter(company=company).exists():
            return JsonResponse({
                'error': _('Cannot delete company with associated users')
            }, status=400)
        
        # Check if company has associated departments
        if Department.objects.filter(company=company).exists():
            return JsonResponse({
                'error': _('Cannot delete company with associated departments')
            }, status=400)
        
        # Check if company has associated positions
        if Position.objects.filter(company=company).exists():
            return JsonResponse({
                'error': _('Cannot delete company with associated positions')
            }, status=400)
        
        # Check if company has associated cabinet groups
        if CabinetGroup.objects.filter(company=company).exists():
            return JsonResponse({
                'error': _('Cannot delete company with associated cabinet groups')
            }, status=400)
        
        company_name = company.name
        company.delete()
        
        # Log the action
        logger.info(f"Company deleted: {company_name} by user {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': _('Company deleted successfully')
        })
        
    except Exception as e:
        logger.error(f"Error deleting company {pk}: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_permission('groups', 'view')
def get_group_users(request, pk):
    """Get Django group details with list of users"""
    try:
        from django.contrib.auth.models import Group
        
        group = get_object_or_404(Group, pk=pk)
        current_language = request.LANGUAGE_CODE[:2]
        
        # Get group details
        group_data = {
            'id': group.id,
            'name': group.name,
        }
        
        # Add cabinet group details if available
        if hasattr(group, 'cabinet_details'):
            cabinet_group = group.cabinet_details
            group_data['name'] = cabinet_group.name
        
        # Get all users in this group
        users_data = []
        for user in group.user_set.all():
            try:
                cabinet_user = CabinetUser.objects.filter(user=user).first()
                department_name = None
                position_name = None
                if cabinet_user:
                    if cabinet_user.department:
                        department_name = cabinet_user.department.get_name(lang=current_language)
                    if cabinet_user.position:
                        position_name = cabinet_user.position.get_name(lang=current_language)
                user_info = {
                    'id': user.id,
                    'username': user.username,
                    'full_name': user.get_full_name(),
                    'email': user.email,
                    'is_active': user.is_active,
                    'company': cabinet_user.company.name if cabinet_user and cabinet_user.company else None,
                    'department': department_name,
                    'position': position_name,
                }
                users_data.append(user_info)
            except Exception as e:
                logger.error(f"Error processing user {user.id}: {str(e)}")
                continue
        
        group_data['users'] = users_data
        
        return JsonResponse({
            'status': 'success',
            'data': group_data
        })
        
    except Exception as e:
        logger.error(f"Error getting group users {pk}: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
@require_permission('groups', 'edit')
@require_POST
def remove_user_from_group(request, pk):
    """Remove user from group"""
    try:
        from django.contrib.auth.models import Group
        
        logger.info(f"remove_user_from_group called with pk={pk}, POST data: {request.POST.dict()}")
        
        group = get_object_or_404(Group, pk=pk)
        user_id = request.POST.get('user_id')
        
        if not user_id:
            logger.error(f"User ID not provided in request. POST data: {request.POST.dict()}")
            return JsonResponse({
                'status': 'error',
                'message': _('User ID is required')
            }, status=400)
        
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            logger.error(f"User with ID {user_id} not found")
            return JsonResponse({
                'status': 'error',
                'message': _('User not found')
            }, status=400)
        
        # Check if user is in the group
        if not user.groups.filter(id=group.id).exists():
            logger.warning(f"User {user.username} is not in group {group.name}")
            return JsonResponse({
                'status': 'error',
                'message': _('User is not in this group')
            }, status=400)
        
        # Remove user from group
        user.groups.remove(group)
        
        # Log the action
        logger.info(f"User {user.username} removed from group {group.name} by {request.user.username}")
        
        # Create user activity log
        UserActivity.objects.create(
            user=user,
            action='removed_from_group',
            details={
                'group_name': group.name,
                'removed_by': request.user.username,
                'timestamp': timezone.now().isoformat()
            }
        )
        
        return JsonResponse({
            'status': 'success',
            'message': _('User removed from group successfully')
        })
        
    except Exception as e:
        logger.error(f"Error removing user from group {pk}: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
