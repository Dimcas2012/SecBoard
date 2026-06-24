import json
import re
from datetime import datetime, date, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext_lazy as _, get_language
from django.contrib import messages
from django.db import models
from django.utils import timezone
from django.http import Http404, JsonResponse, HttpResponse, QueryDict
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST, require_GET
from .models import (
    Vendor, VendorHistory, VendorAssessment, VendorDocument, TprmOwner,
    QuestionnaireTemplate, Question, VendorQuestionnaire, QuestionResponse,
    VendorSurveyLink, TprmGuide, TprmGuideTranslation, cabinet_users_active_for_company,
)
from .forms import (
    VendorForm, VendorAssessmentForm, VendorDocumentForm, VendorFilterForm,
    QuestionnaireTemplateForm, QuestionForm, QuestionFormSet,
    VendorSurveyLinkForm
)
from .permissions import (
    tprm_access_required, check_tprm_access,
    get_user_accessible_companies_tprm, has_company_access_tprm
)
from app_conf.models import Country, Company
from .pagination_utils import paginate_tprm_queryset

_VENDOR_LIST_FILTER_KEYS = frozenset({
    'search', 'company', 'risk_level', 'status', 'nda_in_contract', 'is_active',
    'include_inactive', 'criticality_level', 'sanctions_verification_status',
    'data_access_level', 'data_access_rights',
})


def _sanitize_vendor_list_query(raw):
    """
    Rebuild vendor list query string from client; only allow VendorFilterForm field names.
    Used when returning to the vendor list after modal add/edit.
    """
    if not raw or len(raw) > 4096:
        return ''
    if any(c in raw for c in ('\r', '\n', '\x00')):
        return ''
    try:
        qd = QueryDict(raw, mutable=True)
    except Exception:
        return ''
    safe = QueryDict(mutable=True)
    for key in _VENDOR_LIST_FILTER_KEYS:
        if key not in qd:
            continue
        for val in qd.getlist(key):
            if val is not None and str(val).strip() != '':
                safe.appendlist(key, val)
    return safe.urlencode()


def _vendor_list_redirect_after_modal(request):
    url = reverse('app_tprm:vendor_list')
    q = _sanitize_vendor_list_query(request.POST.get('vendor_list_q', ''))
    if q:
        url = url + '?' + q
    return url


def _user_can_actualize_vendor(user, vendor):
    """User must be a TprmOwner (cabinet user) for the vendor's company."""
    from app_cabinet.models import CabinetUser
    if not vendor.company_id:
        return False
    if not has_company_access_tprm(user, vendor.company):
        return False
    cu = CabinetUser.objects.filter(user=user, company_id=vendor.company_id).first()
    if not cu:
        return False
    return vendor.owners.filter(cabinet_user_id=cu.pk).exists()


def _vendor_ids_user_can_actualize(user, vendors):
    """Set of vendor PKs the current user may actualize (owner for that company)."""
    from app_cabinet.models import CabinetUser
    cabinet_by_company = {cu.company_id: cu for cu in CabinetUser.objects.filter(user=user)}
    out = set()
    for v in vendors:
        if not v.company_id:
            continue
        if not has_company_access_tprm(user, v.company):
            continue
        cu = cabinet_by_company.get(v.company_id)
        if not cu:
            continue
        if any(o.cabinet_user_id == cu.pk for o in v.owners.all()):
            out.add(v.pk)
    return out


def _get_vendor_form_tprm_descriptions(form):
    """Build a dict of field_name -> { value_pk: localized_description } for VendorForm TPRM level fields."""
    if not isinstance(form, VendorForm):
        return {}
    result = {}
    for fn in [
        'risk_level', 'status', 'criticality_level',
        'sanctions_verification_status', 'data_access_level', 'data_access_rights',
    ]:
        if fn not in form.fields:
            continue
        w = form.fields[fn].widget
        result[fn] = getattr(w, 'option_titles', None) or {}
    return result


def _save_vendor_form_attachments(request, vendor):
    """Create VendorDocument rows from files posted with vendor add/edit."""
    files = request.FILES.getlist('vendor_attachments')
    if not files:
        return 0
    if not check_tprm_access(request.user, 'can_upload_documents'):
        messages.warning(
            request,
            _('Selected files were not uploaded: you do not have permission to upload vendor documents.')
        )
        return 0
    document_type = request.POST.get('attachment_document_type') or 'other'
    valid_types = {c[0] for c in VendorDocument.DOCUMENT_TYPE_CHOICES}
    if document_type not in valid_types:
        document_type = 'other'
    count = 0
    for file in files:
        VendorDocument.objects.create(
            vendor=vendor,
            document_type=document_type,
            title=file.name,
            description='',
            file=file,
            expiry_date=None,
            uploaded_by=request.user,
        )
        count += 1
    return count


def _parse_vendor_owner_ids(request):
    raw = request.POST.get('owners', '[]')
    try:
        ids = json.loads(raw) if raw else []
    except json.JSONDecodeError:
        return []
    if not isinstance(ids, list):
        return []
    result = []
    for x in ids:
        try:
            result.append(int(x))
        except (TypeError, ValueError):
            continue
    return result


def _apply_vendor_owners(request, vendor):
    """Assign TprmOwner M2M from JSON list of CabinetUser ids (active users for vendor company)."""
    ids = _parse_vendor_owner_ids(request)
    if not vendor.company_id:
        vendor.owners.clear()
        return
    allowed = set(cabinet_users_active_for_company(vendor.company_id).values_list('pk', flat=True))
    existing_cu = (
        set(vendor.owners.values_list('cabinet_user_id', flat=True))
        if vendor.pk
        else set()
    )
    tprm_instances = []
    seen_cu = set()
    for cid in ids:
        if cid in seen_cu:
            continue
        if cid not in allowed and cid not in existing_cu:
            continue
        seen_cu.add(cid)
        to, _ = TprmOwner.objects.get_or_create(
            cabinet_user_id=cid,
            company_id=vendor.company_id,
        )
        tprm_instances.append(to)
    vendor.owners.set(tprm_instances)


def _vendor_owners_initial_data(vendor):
    """Serialize current owners for the vendor form (ids are CabinetUser pks for the picker)."""
    if not vendor or not getattr(vendor, 'pk', None):
        return []
    out = []
    for o in vendor.owners.select_related('cabinet_user__user', 'cabinet_user__department', 'cabinet_user__position').all():
        cu = o.cabinet_user
        dept = cu.department.get_name() if cu.department else ''
        pos = cu.position.get_name() if cu.position else ''
        out.append({
            'id': cu.pk,
            'name': o.name,
            'department': dept,
            'position': pos,
            'email': o.email or '',
            'phone': o.phone or '',
        })
    return out


def _vendor_owner_emails_for_export(vendor):
    """Semicolon-separated User emails for TPRM owners (for Excel re-import)."""
    emails = []
    for o in vendor.owners.all():
        try:
            e = o.cabinet_user.user.email
            if e:
                emails.append(e.strip())
        except Exception:
            continue
    return '; '.join(emails)


def _parse_vendor_excel_date(val):
    """Parse a cell value to date for contract_end_date import."""
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _set_vendor_owners_from_import_emails(vendor, emails_raw):
    """Set owners M2M from semicolon/comma/newline-separated user emails (CabinetUser for vendor.company)."""
    from django.contrib.auth.models import User
    from app_cabinet.models import CabinetUser
    if not vendor.company_id:
        vendor.owners.clear()
        return
    allowed = set(cabinet_users_active_for_company(vendor.company_id).values_list('pk', flat=True))
    existing_cu = (
        set(vendor.owners.values_list('cabinet_user_id', flat=True))
        if vendor.pk
        else set()
    )
    if emails_raw is None:
        return
    text = str(emails_raw).strip()
    if not text:
        vendor.owners.clear()
        return
    parts = re.split(r'[;\n,]+', text)
    tprm_instances = []
    seen_cu = set()
    for part in parts:
        email = part.strip()
        if not email:
            continue
        user = User.objects.filter(email__iexact=email).first()
        if not user:
            continue
        cu = CabinetUser.objects.filter(user=user, company_id=vendor.company_id).first()
        if not cu:
            continue
        if cu.pk not in allowed and cu.pk not in existing_cu:
            continue
        to, _ = TprmOwner.objects.get_or_create(
            cabinet_user_id=cu.pk,
            company_id=vendor.company_id,
        )
        if cu.pk not in seen_cu:
            seen_cu.add(cu.pk)
            tprm_instances.append(to)
    vendor.owners.set(tprm_instances)


@login_required
@tprm_access_required()
def tprm_dashboard(request):
    """
    Third-Party Risk Management Dashboard
    """
    from app_conf.models import Company
    
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_tprm(request.user)
    
    # Get selected company from request
    selected_company_id = request.GET.get('company')
    selected_company = None
    
    if accessible_companies is None:
        # Access to all companies
        all_companies = Company.objects.all()
        vendors = Vendor.objects.all()
    elif accessible_companies:
        # Access to specific companies
        all_companies = Company.objects.filter(id__in=[c.id for c in accessible_companies])
        vendors = Vendor.objects.filter(
            models.Q(company__in=accessible_companies) | models.Q(company__isnull=True)
        )
    else:
        # No access
        all_companies = Company.objects.none()
        vendors = Vendor.objects.none()
    
    # Apply company filter if selected
    if selected_company_id:
        try:
            selected_company = Company.objects.get(id=selected_company_id)
            if accessible_companies is None or selected_company in accessible_companies:
                vendors = vendors.filter(company=selected_company)
        except Company.DoesNotExist:
            pass
    
    assessments = VendorAssessment.objects.filter(
        vendor__in=vendors,
        status__in=['draft', 'in_progress']
    )
    high_risk_vendors = vendors.filter(risk_level__code__in=['high', 'critical'])
    questionnaires = VendorQuestionnaire.objects.filter(vendor__in=vendors)
    
    # Risk distribution (keys expected by dashboard template)
    vendors_by_risk = {
        'low': vendors.filter(risk_level__code='low').count(),
        'medium': vendors.filter(risk_level__code='medium').count(),
        'high': vendors.filter(risk_level__code='high').count(),
        'critical': vendors.filter(risk_level__code='critical').count(),
    }
    vendors_by_status = {
        'active': vendors.filter(status__code='active').count(),
        'pending': vendors.filter(status__code='pending').count(),
        'inactive': vendors.filter(status__code='inactive').count(),
        'suspended': vendors.filter(status__code='suspended').count(),
    }
    
    context = {
        'title': _('Third-Party Risk Management'),
        'total_vendors': vendors.count(),
        'active_assessments': assessments.count(),
        'high_risk_vendors': high_risk_vendors.count(),
        'total_questionnaires': questionnaires.count(),
        'recent_vendors': vendors[:5],
        'vendors_by_risk': vendors_by_risk,
        'vendors_by_status': vendors_by_status,
        # Add permission context
        'can_edit_vendors': check_tprm_access(request.user, 'can_edit_vendors'),
        'can_delete_vendors': check_tprm_access(request.user, 'can_delete_vendors'),
        'can_conduct_assessments': check_tprm_access(request.user, 'can_conduct_assessments'),
        'can_generate_reports': check_tprm_access(request.user, 'can_generate_reports'),
        # Company filter
        'all_companies': all_companies,
        'selected_company': selected_company,
    }
    return render(request, 'app_tprm/dashboard.html', context)


@login_required
@tprm_access_required()
@require_http_methods(["GET"])
def tprm_guide(request):
    """Return JSON { content: html } for the TPRM guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = TprmGuide.objects.first()
    if guide:
        if country:
            trans = TprmGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = TprmGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def tprm_guide_translate(request):
    """API for AI translation of TPRM guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
@tprm_access_required('can_edit_vendors')
def vendor_add(request):
    """
    Add new vendor
    """
    is_modal_get = request.method == 'GET' and request.GET.get('modal') == '1'

    if request.method == 'POST':
        is_modal = request.POST.get('modal') == '1'
        form = VendorForm(request.POST, request.FILES)
        if form.is_valid():
            vendor = form.save(commit=False)
            vendor.created_by = request.user
            vendor._tprm_history_user = request.user
            vendor.save()
            _apply_vendor_owners(request, vendor)
            attached = _save_vendor_form_attachments(request, vendor)
            msg = _('Vendor added successfully!')
            messages.success(request, msg)
            if attached:
                messages.success(
                    request,
                    _('%(count)s file(s) attached to the vendor.') % {'count': attached}
                )
            if is_modal:
                return JsonResponse({
                    'success': True,
                    'redirect': _vendor_list_redirect_after_modal(request),
                    'message': str(msg),
                })
            return redirect('app_tprm:vendor_detail', pk=vendor.pk)
        if is_modal:
            tprm_descriptions = _get_vendor_form_tprm_descriptions(form)
            ctx = {
                'form': form,
                'tprm_descriptions': tprm_descriptions,
                'can_upload_documents': check_tprm_access(request.user, 'can_upload_documents'),
                'vendor_owners_initial': _vendor_owners_initial_data(None),
                'form_action_url': reverse('app_tprm:vendor_add'),
                'vendor_list_query': _sanitize_vendor_list_query(request.POST.get('vendor_list_q', '')),
            }
            html = render_to_string('app_tprm/vendor_form_modal_body.html', ctx, request=request)
            return JsonResponse({'success': False, 'html': html})
    else:
        form = VendorForm()

    # Pass localized descriptions for TPRM level fields (by current site language) for display in form
    tprm_descriptions = _get_vendor_form_tprm_descriptions(form)
    context = {
        'title': _('Add Vendor'),
        'form': form,
        'tprm_descriptions': tprm_descriptions,
        'can_upload_documents': check_tprm_access(request.user, 'can_upload_documents'),
        'vendor_owners_initial': _vendor_owners_initial_data(None),
        'form_action_url': reverse('app_tprm:vendor_add'),
        'vendor_list_query': _sanitize_vendor_list_query(request.GET.get('list_q', '')),
    }
    if is_modal_get:
        return render(request, 'app_tprm/vendor_form_modal_body.html', context)
    return render(request, 'app_tprm/vendor_form.html', context)


@login_required
@tprm_access_required('has_access_vendors')
def vendor_list(request):
    """
    List all vendors
    """
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_tprm(request.user)
    
    if accessible_companies is None:
        vendors = Vendor.objects.all()
    elif accessible_companies:
        vendors = Vendor.objects.filter(
            models.Q(company__in=accessible_companies) | models.Q(company__isnull=True)
        )
    else:
        vendors = Vendor.objects.none()
    
    filter_form = VendorFilterForm(request.GET)
    
    # Apply filters
    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        company = filter_form.cleaned_data.get('company')
        risk_level = filter_form.cleaned_data.get('risk_level')
        status = filter_form.cleaned_data.get('status')
        nda_in_contract = filter_form.cleaned_data.get('nda_in_contract')
        is_active = filter_form.cleaned_data.get('is_active')
        include_inactive = filter_form.cleaned_data.get('include_inactive')
        criticality_level = filter_form.cleaned_data.get('criticality_level')
        sanctions_verification_status = filter_form.cleaned_data.get('sanctions_verification_status')
        data_access_level = filter_form.cleaned_data.get('data_access_level')
        data_access_rights = filter_form.cleaned_data.get('data_access_rights')
        
        if search:
            vendors = vendors.filter(
                models.Q(name__icontains=search) |
                models.Q(contact_person__icontains=search) |
                models.Q(contact_email__icontains=search) |
                models.Q(contract__icontains=search) |
                models.Q(contract_validity__icontains=search)
            )
        if company:
            vendors = vendors.filter(company=company)
        if risk_level:
            vendors = vendors.filter(risk_level=risk_level)
        if status:
            vendors = vendors.filter(status=status)
        if nda_in_contract != '' and nda_in_contract is not None:
            vendors = vendors.filter(nda_in_contract=(nda_in_contract == '1'))
        if is_active != '' and is_active is not None:
            vendors = vendors.filter(is_active=(is_active == '1'))
        elif not include_inactive:
            vendors = vendors.filter(is_active=True)
        if criticality_level:
            vendors = vendors.filter(criticality_level=criticality_level)
        if sanctions_verification_status:
            vendors = vendors.filter(sanctions_verification_status=sanctions_verification_status)
        if data_access_level:
            vendors = vendors.filter(data_access_level=data_access_level)
        if data_access_rights:
            vendors = vendors.filter(data_access_rights=data_access_rights)
    
    vendors = vendors.select_related(
        'company', 'risk_level', 'status', 'criticality_level', 'sanctions_verification_status', 'data_access_level', 'data_access_rights', 'actualized_by'
    ).prefetch_related(
        models.Prefetch(
            'owners',
            queryset=TprmOwner.objects.select_related('cabinet_user__user', 'company'),
        )
    ).order_by('-created_at', '-pk')

    page_obj, pagination_context = paginate_tprm_queryset(request, vendors)
    actualize_vendor_ids = _vendor_ids_user_can_actualize(request.user, list(page_obj))

    _today = timezone.now().date()
    context = {
        **pagination_context,
        'title': _('Vendors'),
        'vendors': page_obj,
        'filter_form': filter_form,
        'can_edit_vendors': check_tprm_access(request.user, 'can_edit_vendors'),
        'can_delete_vendors': check_tprm_access(request.user, 'can_delete_vendors'),
        'actualize_vendor_ids': actualize_vendor_ids,
        'today': _today,
        'contract_warn_until': _today + timedelta(days=30),
    }
    return render(request, 'app_tprm/vendor_list.html', context)


@login_required
@tprm_access_required('has_access_vendors')
def vendor_detail(request, pk):
    """
    Vendor detail view
    """
    vendor = get_object_or_404(
        Vendor.objects.select_related(
            'company', 'risk_level', 'status', 'criticality_level', 'sanctions_verification_status',
            'data_access_level', 'data_access_rights', 'actualized_by',
        ).prefetch_related(
            models.Prefetch(
                'owners',
                queryset=TprmOwner.objects.select_related('cabinet_user__user', 'company'),
            )
        ),
        pk=pk
    )
    assessments = vendor.assessments.all()
    documents = vendor.documents.all()
    survey_links = vendor.survey_links.all()
    
    # Calculate dates for expiry checking
    from datetime import date, timedelta
    today = date.today()
    next_month = today + timedelta(days=30)
    
    context = {
        'title': vendor.name,
        'vendor': vendor,
        'assessments': assessments,
        'documents': documents,
        'survey_links': survey_links,
        'today': today,
        'next_month': next_month,
        'can_edit_vendors': check_tprm_access(request.user, 'can_edit_vendors'),
        'can_delete_vendors': check_tprm_access(request.user, 'can_delete_vendors'),
        'can_conduct_assessments': check_tprm_access(request.user, 'can_conduct_assessments'),
        'can_upload_documents': check_tprm_access(request.user, 'can_upload_documents'),
        'can_delete_documents': check_tprm_access(request.user, 'can_delete_documents'),
        'can_actualize_vendor': _user_can_actualize_vendor(request.user, vendor),
    }
    return render(request, 'app_tprm/vendor_detail.html', context)


@login_required
@tprm_access_required('can_edit_vendors')
def vendor_edit(request, pk):
    """
    Edit vendor
    """
    vendor = get_object_or_404(Vendor, pk=pk)
    is_modal_get = request.method == 'GET' and request.GET.get('modal') == '1'

    if request.method == 'POST':
        is_modal = request.POST.get('modal') == '1'
        form = VendorForm(request.POST, request.FILES, instance=vendor)
        if form.is_valid():
            obj = form.save(commit=False)
            obj._tprm_history_user = request.user
            obj.save()
            _apply_vendor_owners(request, vendor)
            attached = _save_vendor_form_attachments(request, vendor)
            msg = _('Vendor updated successfully!')
            messages.success(request, msg)
            if attached:
                messages.success(
                    request,
                    _('%(count)s file(s) attached to the vendor.') % {'count': attached}
                )
            if is_modal:
                return JsonResponse({
                    'success': True,
                    'redirect': _vendor_list_redirect_after_modal(request),
                    'message': str(msg),
                })
            return redirect('app_tprm:vendor_detail', pk=vendor.pk)
        if is_modal:
            tprm_descriptions = _get_vendor_form_tprm_descriptions(form)
            ctx = {
                'form': form,
                'vendor': vendor,
                'tprm_descriptions': tprm_descriptions,
                'can_upload_documents': check_tprm_access(request.user, 'can_upload_documents'),
                'vendor_owners_initial': _vendor_owners_initial_data(vendor),
                'form_action_url': reverse('app_tprm:vendor_edit', args=[vendor.pk]),
                'vendor_list_query': _sanitize_vendor_list_query(request.POST.get('vendor_list_q', '')),
            }
            html = render_to_string('app_tprm/vendor_form_modal_body.html', ctx, request=request)
            return JsonResponse({'success': False, 'html': html})
    else:
        form = VendorForm(instance=vendor)

    # Pass localized descriptions for TPRM level fields (by current site language) for display in form
    tprm_descriptions = _get_vendor_form_tprm_descriptions(form)
    context = {
        'title': _('Edit Vendor'),
        'form': form,
        'vendor': vendor,
        'tprm_descriptions': tprm_descriptions,
        'can_upload_documents': check_tprm_access(request.user, 'can_upload_documents'),
        'vendor_owners_initial': _vendor_owners_initial_data(vendor),
        'form_action_url': reverse('app_tprm:vendor_edit', args=[vendor.pk]),
        'vendor_list_query': _sanitize_vendor_list_query(request.GET.get('list_q', '')),
    }
    if is_modal_get:
        return render(request, 'app_tprm/vendor_form_modal_body.html', context)
    return render(request, 'app_tprm/vendor_form.html', context)


@login_required
@tprm_access_required('can_delete_vendors')
def vendor_delete(request, pk):
    """
    Delete vendor
    """
    vendor = get_object_or_404(Vendor, pk=pk)
    
    if request.method == 'POST':
        vendor_name = vendor.name
        vendor.delete()
        messages.success(request, _('Vendor "%(name)s" deleted successfully!') % {'name': vendor_name})
        return redirect('app_tprm:vendor_list')
    
    context = {
        'title': _('Delete Vendor'),
        'vendor': vendor,
    }
    return render(request, 'app_tprm/vendor_confirm_delete.html', context)


def _get_vendors_queryset_for_user(request):
    """Return Vendor queryset according to user TPRM access and optional filters (same as vendor_list)."""
    accessible_companies = get_user_accessible_companies_tprm(request.user)
    if accessible_companies is None:
        vendors = Vendor.objects.all()
    elif accessible_companies:
        vendors = Vendor.objects.filter(
            models.Q(company__in=accessible_companies) | models.Q(company__isnull=True)
        )
    else:
        vendors = Vendor.objects.none()
    filter_form = VendorFilterForm(request.GET)
    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        company = filter_form.cleaned_data.get('company')
        risk_level = filter_form.cleaned_data.get('risk_level')
        status = filter_form.cleaned_data.get('status')
        nda_in_contract = filter_form.cleaned_data.get('nda_in_contract')
        is_active = filter_form.cleaned_data.get('is_active')
        include_inactive = filter_form.cleaned_data.get('include_inactive')
        criticality_level = filter_form.cleaned_data.get('criticality_level')
        sanctions_verification_status = filter_form.cleaned_data.get('sanctions_verification_status')
        data_access_level = filter_form.cleaned_data.get('data_access_level')
        data_access_rights = filter_form.cleaned_data.get('data_access_rights')
        if search:
            vendors = vendors.filter(
                models.Q(name__icontains=search) |
                models.Q(contact_person__icontains=search) |
                models.Q(contact_email__icontains=search) |
                models.Q(contract__icontains=search) |
                models.Q(contract_validity__icontains=search)
            )
        if company:
            vendors = vendors.filter(company=company)
        if risk_level:
            vendors = vendors.filter(risk_level=risk_level)
        if status:
            vendors = vendors.filter(status=status)
        if nda_in_contract != '' and nda_in_contract is not None:
            vendors = vendors.filter(nda_in_contract=(nda_in_contract == '1'))
        if is_active != '' and is_active is not None:
            vendors = vendors.filter(is_active=(is_active == '1'))
        elif not include_inactive:
            vendors = vendors.filter(is_active=True)
        if criticality_level:
            vendors = vendors.filter(criticality_level=criticality_level)
        if sanctions_verification_status:
            vendors = vendors.filter(sanctions_verification_status=sanctions_verification_status)
        if data_access_level:
            vendors = vendors.filter(data_access_level=data_access_level)
        if data_access_rights:
            vendors = vendors.filter(data_access_rights=data_access_rights)
    return vendors.select_related(
        'company', 'risk_level', 'status', 'criticality_level', 'sanctions_verification_status', 'data_access_level', 'data_access_rights', 'actualized_by'
    )


@login_required
@tprm_access_required('has_access_vendors')
@require_POST
def actualize_vendor(request, pk):
    """Confirm vendor as actual or mark as no longer actual (owners only, like Information Assets)."""
    vendor = get_object_or_404(
        Vendor.objects.select_related('company').prefetch_related(
            models.Prefetch('owners', queryset=TprmOwner.objects.select_related('cabinet_user')),
        ),
        pk=pk,
    )
    if not _user_can_actualize_vendor(request.user, vendor):
        return JsonResponse(
            {'status': 'error', 'message': str(_('Only vendor owners can actualize this record.'))},
            status=403,
        )
    action = request.POST.get('action', 'actualize')
    user = request.user
    if action == 'mark_inactive':
        comment = (request.POST.get('comment') or '').strip()
        vendor.actualization_date = None
        vendor.actualized_by = None
        vendor.marked_no_longer_actual_at = timezone.now()
        vendor.marked_no_longer_comment = comment
        vendor._tprm_history_user = user
        vendor._tprm_history_action_override = VendorHistory.ACTION_MARKED_NOT_ACTUAL
        vendor.save()
        return JsonResponse({
            'status': 'success',
            'message': str(_('Vendor marked as no longer actual')),
            'actualization_date': None,
            'actualized_by': None,
        })
    vendor.actualization_date = timezone.now()
    vendor.actualized_by = user
    vendor.marked_no_longer_actual_at = None
    vendor.marked_no_longer_comment = ''
    vendor._tprm_history_user = user
    vendor._tprm_history_action_override = VendorHistory.ACTION_ACTUALIZED
    vendor.save()
    return JsonResponse({
        'status': 'success',
        'message': str(_('Vendor actualized successfully')),
        'actualization_date': timezone.localtime(vendor.actualization_date).strftime('%d-%m-%Y %H:%M:%S'),
        'actualized_by': user.get_full_name() or user.username,
    })


@login_required
@tprm_access_required('has_access_vendors')
@require_GET
def get_vendor_history(request, pk):
    """JSON timeline of vendor changes (for modal on vendor list)."""
    vendor = get_object_or_404(Vendor.objects.select_related('company'), pk=pk)
    accessible_companies = get_user_accessible_companies_tprm(request.user)
    if accessible_companies is not None:
        if not accessible_companies:
            return JsonResponse({'status': 'error', 'message': str(_('Forbidden'))}, status=403)
        allowed_ids = {c.pk for c in accessible_companies}
        if vendor.company_id and vendor.company_id not in allowed_ids:
            return JsonResponse({'status': 'error', 'message': str(_('Forbidden'))}, status=403)

    history_records = VendorHistory.objects.filter(vendor=vendor).order_by('-timestamp')
    history_data = []
    for record in history_records:
        history_data.append({
            'id': record.id,
            'timestamp': timezone.localtime(record.timestamp).strftime('%d-%m-%Y %H:%M:%S'),
            'action': record.action,
            'action_display': str(record.get_action_display()),
            'action_by': str(record.get_action_by_name()),
            'action_by_id': record.action_by_id,
            'details': record.details or '',
            'changes': record.changes,
        })

    return JsonResponse({
        'status': 'success',
        'vendor_id': vendor.pk,
        'vendor_name': vendor.name,
        'history': history_data,
    })


@login_required
@tprm_access_required('has_access_vendors')
@require_GET
def get_all_tprm_owners(request):
    """JSON list of active CabinetUser rows for a company (vendor owner picker; ids are CabinetUser pks)."""
    company_id = request.GET.get('company_id')
    company_name = (request.GET.get('company_name') or '').strip()
    if not company_id and not company_name:
        return JsonResponse({'error': 'Company ID required'}, status=400)
    try:
        if company_id:
            company = get_object_or_404(Company, pk=company_id)
        else:
            company = get_object_or_404(Company, name=company_name)
        if not has_company_access_tprm(request.user, company):
            return JsonResponse({'error': 'Forbidden'}, status=403)
        data = []
        for cu in cabinet_users_active_for_company(company.pk):
            u = cu.user
            data.append({
                'id': cu.pk,
                'name': u.get_full_name() or u.username or '',
                'department': cu.department.get_name() if cu.department else '',
                'position': cu.position.get_name() if cu.position else '',
                'company': {
                    'id': company.pk,
                    'name': company.name,
                },
                'email': u.email or '',
                'phone': cu.phone or '',
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@tprm_access_required('has_access_vendors')
@require_GET
def export_vendors_xlsx(request):
    """Export vendors to XLSX (filtered like vendor list). Level columns use TprmLevel.code for re-import."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    vendors = _get_vendors_queryset_for_user(request).prefetch_related(
        models.Prefetch(
            'owners',
            queryset=TprmOwner.objects.select_related('cabinet_user__user'),
        )
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Vendors")

    headers = [
        _('Name'), _('Description'), _('Contract'), _('Contract validity period'), _('Contract end date'),
        _('Website'), _('Contact Person'), _('Contact Email'), _('Contact Phone'),
        _('Risk Level'), _('Status'), _('Services Provided'),
        _('NDA in contract'), _('Criticality level'), _('Sanctions verification'),
        _('Data Access Level'), _('Data Access rights'),
        _('Company'), _('Active'), _('Owners'),
        _('Actualization date'), _('Actualized by'), _('Marked no longer actual at'),
        _('Marked no longer actual comment'),
        _('Created At'), _('Updated At'),
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_num, vendor in enumerate(vendors, 2):
        c = 1
        ws.cell(row=row_num, column=c, value=vendor.name or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.description or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.contract or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.contract_validity or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.contract_end_date.isoformat() if vendor.contract_end_date else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.website or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.contact_person or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.contact_email or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.contact_phone or ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.risk_level.code if vendor.risk_level else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.status.code if vendor.status else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.services_provided or ''); c += 1
        ws.cell(row=row_num, column=c, value='1' if vendor.nda_in_contract else '0'); c += 1
        ws.cell(row=row_num, column=c, value=vendor.criticality_level.code if vendor.criticality_level else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.sanctions_verification_status.code if vendor.sanctions_verification_status else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.data_access_level.code if vendor.data_access_level else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.data_access_rights.code if vendor.data_access_rights else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.company.name if vendor.company else ''); c += 1
        ws.cell(row=row_num, column=c, value='1' if vendor.is_active else '0'); c += 1
        ws.cell(row=row_num, column=c, value=_vendor_owner_emails_for_export(vendor)); c += 1
        ws.cell(row=row_num, column=c,
                value=timezone.localtime(vendor.actualization_date).strftime('%Y-%m-%d %H:%M') if vendor.actualization_date else ''); c += 1
        ws.cell(row=row_num, column=c,
                value=(vendor.actualized_by.get_full_name() or vendor.actualized_by.username) if vendor.actualized_by else ''); c += 1
        ws.cell(row=row_num, column=c,
                value=timezone.localtime(vendor.marked_no_longer_actual_at).strftime('%Y-%m-%d %H:%M') if vendor.marked_no_longer_actual_at else ''); c += 1
        ws.cell(row=row_num, column=c, value=vendor.marked_no_longer_comment or ''); c += 1
        ws.cell(row=row_num, column=c,
                value=timezone.localtime(vendor.created_at).strftime('%Y-%m-%d %H:%M') if vendor.created_at else ''); c += 1
        ws.cell(row=row_num, column=c,
                value=timezone.localtime(vendor.updated_at).strftime('%Y-%m-%d %H:%M') if vendor.updated_at else '')

    for column_cells in ws.columns:
        length = max(min(len(str(cell.value) if cell.value else '') + 2, 50) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=vendors_{timestamp}.xlsx'
    wb.save(response)
    return response


@login_required
@tprm_access_required('can_edit_vendors')
@require_GET
def vendor_import_template_xlsx(request):
    """Download empty XLSX template for vendor import."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(_("Vendors"))[:31]  # openpyxl sheet title max 31 chars

    headers = [
        str(_('Name')), str(_('Description')), str(_('Contract')), str(_('Contract validity period')), str(_('Contract end date')),
        str(_('Website')), str(_('Contact Person')),
        str(_('Contact Email')), str(_('Contact Phone')), str(_('Risk Level')), str(_('Status')),
        str(_('Services Provided')), str(_('NDA in contract')), str(_('Criticality level')), str(_('Sanctions verification')),
        str(_('Data Access Level')), str(_('Data Access rights')), str(_('Company')), str(_('Active')), str(_('Owners')),
        str(_('Actualization date')), str(_('Actualized by')), str(_('Marked no longer actual at')),
        str(_('Marked no longer actual comment')),
        str(_('Created At')), str(_('Updated At')),
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Example row: levels = TprmLevel code; Owners = cabinet user emails (same company), separated by ; or ,
    example = [
        str(_('Example Vendor Ltd')),
        str(_('Short description')),
        str(_('Contract summary text')),
        str(_('12 months')),
        '2026-12-31',
        'https://example.com',
        str(_('John Doe')),
        'contact@example.com',
        '+1 234 567 8900',
        'low',
        'active',
        str(_('IT services')),
        '1',
        'medium',
        'high',
        'low',
        'read_only',
        str(_('Company name or leave empty')),
        '1',
        'owner1@example.com; owner2@example.com',
        '', '', '', '',
        '', '',
    ]
    for col_num, val in enumerate(example, 1):
        ws.cell(row=2, column=col_num, value=val)

    for column_cells in ws.columns:
        length = max(min(len(str(cell.value) if cell.value is not None else '') + 2, 50) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=vendors_import_template.xlsx'
    wb.save(response)
    return response


@login_required
@tprm_access_required('can_edit_vendors')
@require_http_methods(["GET", "POST"])
def import_vendors_xlsx(request):
    """Import vendors from XLSX. GET: show form; POST: process file."""
    import openpyxl

    if request.method == 'GET':
        return render(request, 'app_tprm/vendor_import.html', {
            'title': _('Import Vendors'),
        })

    excel_file = request.FILES.get('file')
    if not excel_file:
        messages.error(request, _('No file uploaded.'))
        return redirect('app_tprm:import_vendors_xlsx')

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, _('Please upload an Excel file (.xlsx).'))
        return redirect('app_tprm:import_vendors_xlsx')

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    except Exception as e:
        messages.error(request, _('Unable to read Excel file: %(error)s') % {'error': str(e)})
        return redirect('app_tprm:import_vendors_xlsx')

    ws = wb.active
    if not ws:
        messages.error(request, _('The file has no sheet.'))
        return redirect('app_tprm:import_vendors_xlsx')

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        messages.error(request, _('The file must have a header row and at least one data row.'))
        return redirect('app_tprm:import_vendors_xlsx')

    header_row = [str(c).strip() if c is not None else '' for c in rows[0]]
    name_col = None
    for i, h in enumerate(header_row):
        if h and _('Name').lower() in h.lower():
            name_col = i
            break
    if name_col is None:
        for i, h in enumerate(header_row):
            if h and 'name' in h.lower():
                name_col = i
                break
    if name_col is None:
        messages.error(request, _('Could not find a "Name" column in the first row.'))
        return redirect('app_tprm:import_vendors_xlsx')

    def col_index(label_choices):
        for lbl in label_choices:
            for i, h in enumerate(header_row):
                if h and lbl.lower() in h.lower():
                    return i
        return None

    desc_col = col_index([_('Description'), 'description'])
    contract_col = col_index([_('Contract'), 'contract'])
    contract_validity_col = col_index([_('Contract validity period'), 'contract validity', 'contract_validity'])
    contract_end_col = col_index([_('Contract end date'), 'contract end', 'contract_end_date'])
    website_col = col_index([_('Website'), 'website'])
    contact_person_col = col_index([_('Contact Person'), 'contact person', 'contact_person'])
    contact_email_col = col_index([_('Contact Email'), 'contact email', 'contact_email', 'email'])
    contact_phone_col = col_index([_('Contact Phone'), 'contact phone', 'contact_phone', 'phone'])
    risk_col = col_index([_('Risk Level'), 'risk level', 'risk_level', 'risk'])
    status_col = col_index([_('Status'), 'status'])
    services_col = col_index([_('Services Provided'), 'services provided', 'services_provided'])
    nda_col = col_index([_('NDA in contract'), 'nda', 'nda_in_contract'])
    criticality_col = col_index([_('Criticality level'), 'criticality', 'criticality_level'])
    sanctions_col = col_index([_('Sanctions verification'), 'sanctions', 'sanctions_verification'])
    data_access_col = col_index([_('Data Access Level'), 'data access', 'data_access_level'])
    data_access_rights_col = col_index([_('Data Access rights'), 'data access rights', 'data_access_rights'])
    company_col = col_index([_('Company'), 'company'])
    active_col = col_index([_('Active'), 'active', 'is_active'])
    owners_col = col_index([_('Owners'), 'owners', 'owner emails', 'owner_emails'])
    marked_comment_col = col_index([_('Marked no longer actual comment'), 'marked_no_longer_comment', 'marked comment'])

    from .models import TprmLevel
    def level_by_code_or_name(val, level_type):
        if not val or not str(val).strip():
            return None
        v = str(val).strip().lower()
        for level in TprmLevel.objects.filter(type=level_type, is_active=True).prefetch_related('translations'):
            if level.code and level.code.lower() == v:
                return level
            if level.name and level.name.lower() == v:
                return level
            if level.get_name() and level.get_name().lower() == v:
                return level
        return None

    accessible_companies = get_user_accessible_companies_tprm(request.user)
    if accessible_companies is not None and not accessible_companies:
        messages.error(request, _('You do not have access to any company for importing vendors.'))
        return redirect('app_tprm:vendor_list')

    risk_levels = {lvl.code: lvl for lvl in TprmLevel.objects.filter(type=TprmLevel.TYPE_RISK_LEVEL, is_active=True)}
    risk_map = {}
    for lvl in risk_levels.values():
        risk_map[str(lvl.code).lower()] = lvl
        if lvl.name:
            risk_map[str(lvl.name).lower()] = lvl
        if lvl.get_name():
            risk_map[str(lvl.get_name()).lower()] = lvl
    status_levels = {lvl.code: lvl for lvl in TprmLevel.objects.filter(type=TprmLevel.TYPE_STATUS, is_active=True)}
    status_map = {}
    for lvl in status_levels.values():
        status_map[str(lvl.code).lower()] = lvl
        if lvl.name:
            status_map[str(lvl.name).lower()] = lvl
        if lvl.get_name():
            status_map[str(lvl.get_name()).lower()] = lvl

    created = 0
    updated = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        row = list(row) if row else []
        _idxs = [
            name_col, desc_col, contract_col, contract_validity_col, contract_end_col, website_col,
            contact_person_col, contact_email_col, contact_phone_col, risk_col, status_col, services_col,
            nda_col, criticality_col, sanctions_col, data_access_col, data_access_rights_col, company_col,
            active_col, owners_col, marked_comment_col,
        ]
        _max_i = max((i for i in _idxs if i is not None), default=0)
        while len(row) <= _max_i:
            row.append(None)
        name = (row[name_col] or '').strip() if name_col is not None else ''
        if not name:
            continue

        company = None
        if company_col is not None and row[company_col]:
            company_val = str(row[company_col]).strip()
            if company_val:
                if accessible_companies is None:
                    company = Company.objects.filter(name=company_val).first() or \
                             Company.objects.filter(id=company_val).first()
                else:
                    company = next((c for c in accessible_companies if c.name == company_val or str(c.id) == company_val), None)
                    if company is None:
                        company = Company.objects.filter(id__in=[c.id for c in accessible_companies], name=company_val).first()
                if company is None and company_val.isdigit():
                    company = Company.objects.filter(id=int(company_val)).first()
                    if company and accessible_companies is not None and company not in accessible_companies:
                        company = None

        risk_level = risk_levels.get('medium') or list(risk_levels.values())[0] if risk_levels else None
        if risk_col is not None and row[risk_col]:
            rv = str(row[risk_col]).strip().lower()
            risk_level = risk_map.get(rv) or risk_level
        status = status_levels.get('pending') or list(status_levels.values())[0] if status_levels else None
        if status_col is not None and row[status_col]:
            sv = str(row[status_col]).strip().lower()
            status = status_map.get(sv) or status

        description = row[desc_col] if desc_col is not None else ''
        description = str(description).strip() if description else ''
        contract = ''
        if contract_col is not None and row[contract_col] is not None:
            contract = str(row[contract_col]).strip()
        contract_validity = ''
        if contract_validity_col is not None and row[contract_validity_col] is not None:
            contract_validity = str(row[contract_validity_col]).strip()
        contract_end_date = None
        if contract_end_col is not None:
            contract_end_date = _parse_vendor_excel_date(row[contract_end_col])
        website = row[website_col] if website_col is not None else ''
        website = str(website).strip() if website else ''
        contact_person = row[contact_person_col] if contact_person_col is not None else ''
        contact_person = str(contact_person).strip() if contact_person else ''
        contact_email = row[contact_email_col] if contact_email_col is not None else ''
        contact_email = str(contact_email).strip() if contact_email else ''
        contact_phone = row[contact_phone_col] if contact_phone_col is not None else ''
        contact_phone = str(contact_phone).strip() if contact_phone else ''
        services_provided = row[services_col] if services_col is not None else ''
        services_provided = str(services_provided).strip() if services_provided else ''
        nda_val = row[nda_col] if nda_col is not None else None
        nda_in_contract = (
            str(nda_val).strip().lower() in ('1', 'yes', 'true', 'так')
            if nda_val is not None and str(nda_val).strip() != ''
            else False
        )
        criticality_level = level_by_code_or_name(row[criticality_col] if criticality_col is not None else None, TprmLevel.TYPE_CRITICALITY)
        sanctions_verification_status = level_by_code_or_name(row[sanctions_col] if sanctions_col is not None else None, TprmLevel.TYPE_SANCTIONS)
        data_access_level = level_by_code_or_name(row[data_access_col] if data_access_col is not None else None, TprmLevel.TYPE_DATA_ACCESS)
        data_access_rights = level_by_code_or_name(row[data_access_rights_col] if data_access_rights_col is not None else None, TprmLevel.TYPE_DATA_ACCESS_RIGHTS)

        active_val = row[active_col] if active_col is not None else None
        if active_val is not None and str(active_val).strip() != '':
            is_active_import = str(active_val).strip().lower() in ('1', 'yes', 'true', 'так', 'active')
        else:
            is_active_import = None

        existing = Vendor.objects.filter(name=name).first()
        if existing:
            existing.description = description or existing.description
            if contract_col is not None:
                existing.contract = contract or existing.contract
            if contract_validity_col is not None:
                existing.contract_validity = contract_validity or existing.contract_validity
            if contract_end_col is not None:
                raw_end = row[contract_end_col]
                if raw_end is None or (isinstance(raw_end, str) and not str(raw_end).strip()):
                    existing.contract_end_date = None
                elif contract_end_date is not None:
                    existing.contract_end_date = contract_end_date
            existing.website = website or existing.website
            existing.contact_person = contact_person or existing.contact_person
            existing.contact_email = contact_email or existing.contact_email
            existing.contact_phone = contact_phone or existing.contact_phone
            existing.services_provided = services_provided or existing.services_provided
            if nda_col is not None:
                if nda_val is not None and str(nda_val).strip() != '':
                    existing.nda_in_contract = str(nda_val).strip().lower() in ('1', 'yes', 'true', 'так')
            if risk_level is not None:
                existing.risk_level = risk_level
            if status is not None:
                existing.status = status
            if criticality_level is not None:
                existing.criticality_level = criticality_level
            if sanctions_verification_status is not None:
                existing.sanctions_verification_status = sanctions_verification_status
            if data_access_level is not None:
                existing.data_access_level = data_access_level
            if data_access_rights is not None:
                existing.data_access_rights = data_access_rights
            if company is not None:
                existing.company = company
            if is_active_import is not None:
                existing.is_active = is_active_import
            if marked_comment_col is not None:
                mc = row[marked_comment_col]
                existing.marked_no_longer_comment = str(mc).strip() if mc is not None else ''
            existing._tprm_skip_history = True
            existing.save()
            if owners_col is not None:
                _set_vendor_owners_from_import_emails(existing, row[owners_col])
            updated += 1
        else:
            try:
                v_new = Vendor(
                    name=name,
                    description=description,
                    contract=contract,
                    contract_validity=contract_validity,
                    contract_end_date=contract_end_date,
                    website=website or '',
                    contact_person=contact_person,
                    contact_email=contact_email,
                    contact_phone=contact_phone,
                    risk_level=risk_level,
                    status=status,
                    services_provided=services_provided,
                    nda_in_contract=nda_in_contract if nda_col is not None else False,
                    criticality_level=criticality_level,
                    sanctions_verification_status=sanctions_verification_status,
                    data_access_level=data_access_level,
                    data_access_rights=data_access_rights,
                    company=company,
                    is_active=is_active_import if is_active_import is not None else True,
                    created_by=request.user,
                )
                v_new._tprm_skip_history = True
                v_new.save()
                if owners_col is not None:
                    v_new._tprm_skip_history = True
                    _set_vendor_owners_from_import_emails(v_new, row[owners_col])
                if marked_comment_col is not None:
                    mc = row[marked_comment_col]
                    if mc is not None and str(mc).strip():
                        v_new.marked_no_longer_comment = str(mc).strip()
                        v_new._tprm_skip_history = True
                        v_new.save(update_fields=['marked_no_longer_comment'])
                created += 1
            except Exception as e:
                errors.append(_('Row %(row)s "%(name)s": %(err)s') % {'row': row_idx, 'name': name, 'err': str(e)})

    wb.close()
    if created or updated:
        messages.success(
            request,
            _('Import completed: %(created)s created, %(updated)s updated.') % {'created': created, 'updated': updated}
        )
    if errors:
        for err in errors[:10]:
            messages.warning(request, err)
        if len(errors) > 10:
            messages.warning(request, _('… and %(count)s more errors.') % {'count': len(errors) - 10})
    if not created and not updated and not errors:
        messages.info(request, _('No rows were imported. Check that the file has a "Name" column and data rows.'))
    return redirect('app_tprm:vendor_list')


@login_required
@tprm_access_required('can_conduct_assessments')
def assessment_add(request):
    """
    Create new assessment
    """
    if request.method == 'POST':
        form = VendorAssessmentForm(request.POST)
        if form.is_valid():
            assessment = form.save(commit=False)
            assessment.assessed_by = request.user
            assessment.save()
            messages.success(request, _('Assessment created successfully!'))
            return redirect('app_tprm:dashboard')
    else:
        form = VendorAssessmentForm()
    
    context = {
        'title': _('New Assessment'),
        'form': form,
    }
    return render(request, 'app_tprm/assessment_form.html', context)


@login_required
@tprm_access_required('can_generate_reports')
def reports(request):
    """
    TPRM Reports
    """
    # Filter by accessible companies
    accessible_companies = get_user_accessible_companies_tprm(request.user)
    
    if accessible_companies is None:
        vendors = Vendor.objects.all()
    elif accessible_companies:
        vendors = Vendor.objects.filter(
            models.Q(company__in=accessible_companies) | models.Q(company__isnull=True)
        )
    else:
        vendors = Vendor.objects.none()
    
    vendors = vendors
    assessments = VendorAssessment.objects.all()
    
    # Calculate statistics
    risk_distribution = {
        'low': vendors.filter(risk_level__code='low').count(),
        'medium': vendors.filter(risk_level__code='medium').count(),
        'high': vendors.filter(risk_level__code='high').count(),
        'critical': vendors.filter(risk_level__code='critical').count(),
    }
    status_distribution = {
        'active': vendors.filter(status__code='active').count(),
        'inactive': vendors.filter(status__code='inactive').count(),
        'pending': vendors.filter(status__code='pending').count(),
        'suspended': vendors.filter(status__code='suspended').count(),
    }
    
    context = {
        'title': _('TPRM Reports'),
        'vendors': vendors,
        'assessments': assessments,
        'risk_distribution': risk_distribution,
        'status_distribution': status_distribution,
    }
    return render(request, 'app_tprm/reports.html', context)


@login_required
@tprm_access_required('has_access_questionnaires')
def questionnaire_list(request):
    """List of questionnaire templates"""
    templates = QuestionnaireTemplate.objects.filter(is_active=True)
    vendors = Vendor.objects.all()
    
    context = {
        'title': _('Questionnaire Templates'),
        'templates': templates,
        'vendors': vendors,
    }
    return render(request, 'app_tprm/questionnaire_list.html', context)


@login_required
@tprm_access_required('can_complete_questionnaires')
def questionnaire_start(request, vendor_pk, template_pk):
    """Start a new questionnaire for a vendor"""
    vendor = get_object_or_404(Vendor, pk=vendor_pk)
    template = get_object_or_404(QuestionnaireTemplate, pk=template_pk)
    
    # Check if questionnaire already exists
    questionnaire, created = VendorQuestionnaire.objects.get_or_create(
        vendor=vendor,
        template=template,
        status__in=['not_started', 'in_progress'],
        defaults={
            'status': 'in_progress',
            'started_date': timezone.now(),
            'completed_by': request.user,
        }
    )
    
    if created:
        # Create responses for all questions
        for question in template.questions.all():
            QuestionResponse.objects.create(
                questionnaire=questionnaire,
                question=question
            )
        messages.success(request, _('Questionnaire started successfully!'))
    
    return redirect('app_tprm:questionnaire_fill', pk=questionnaire.pk)


@login_required
@tprm_access_required('can_complete_questionnaires')
def questionnaire_fill(request, pk):
    """Fill out a questionnaire"""
    questionnaire = get_object_or_404(VendorQuestionnaire, pk=pk)
    responses = questionnaire.responses.all().select_related('question')
    
    if request.method == 'POST':
        # Process responses
        for response in responses:
            question = response.question
            field_name = f'question_{question.pk}'
            
            if question.question_type == 'yes_no':
                value = request.POST.get(field_name)
                response.response_bool = value == 'yes' if value else None
            elif question.question_type == 'scale':
                value = request.POST.get(field_name)
                response.response_scale = int(value) if value else None
            elif question.question_type == 'text':
                response.response_text = request.POST.get(field_name, '')
            elif question.question_type == 'multiple_choice':
                response.response_choice = request.POST.get(field_name, '')
            
            response.answered_at = timezone.now()
            response.auto_score()
        
        # Update questionnaire status
        if 'complete' in request.POST:
            questionnaire.status = 'completed'
            questionnaire.completed_date = timezone.now()
            questionnaire.completed_by = request.user
            questionnaire.calculate_score()
            messages.success(request, _('Questionnaire completed successfully!'))
            return redirect('app_tprm:questionnaire_view', pk=questionnaire.pk)
        else:
            questionnaire.save()
            messages.success(request, _('Progress saved!'))
    
    context = {
        'title': f"{questionnaire.vendor.name} - {questionnaire.template.name}",
        'questionnaire': questionnaire,
        'responses': responses,
    }
    return render(request, 'app_tprm/questionnaire_fill.html', context)


@login_required
@tprm_access_required('has_access_questionnaires')
def questionnaire_view(request, pk):
    """View completed questionnaire"""
    questionnaire = get_object_or_404(VendorQuestionnaire, pk=pk)
    responses = questionnaire.responses.all().select_related('question')
    
    context = {
        'title': f"{questionnaire.vendor.name} - {questionnaire.template.name}",
        'questionnaire': questionnaire,
        'responses': responses,
    }
    return render(request, 'app_tprm/questionnaire_view.html', context)


@login_required
@tprm_access_required('has_access_templates')
def template_list(request):
    """List all questionnaire templates"""
    templates = QuestionnaireTemplate.objects.all().annotate(
        question_count=models.Count('questions'),
        conditional_count=models.Count('questions', filter=models.Q(questions__parent_question__isnull=False))
    )
    
    context = {
        'title': _('Questionnaire Templates'),
        'templates': templates,
        'can_edit_templates': check_tprm_access(request.user, 'can_edit_templates'),
    }
    return render(request, 'app_tprm/template_list.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def template_add(request):
    """Create new questionnaire template"""
    if request.method == 'POST':
        form = QuestionnaireTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.created_by = request.user
            template.save()
            messages.success(request, _('Template created successfully!'))
            return redirect('app_tprm:template_edit', pk=template.pk)
    else:
        form = QuestionnaireTemplateForm()
    
    context = {
        'title': _('Create Questionnaire Template'),
        'form': form,
    }
    return render(request, 'app_tprm/template_form.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def template_edit(request, pk):
    """Edit questionnaire template and its questions"""
    template = get_object_or_404(QuestionnaireTemplate, pk=pk)
    
    if request.method == 'POST':
        form = QuestionnaireTemplateForm(request.POST, instance=template)
        formset = QuestionFormSet(request.POST, instance=template)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, _('Template and questions updated successfully!'))
            return redirect('app_tprm:template_edit', pk=template.pk)
    else:
        form = QuestionnaireTemplateForm(instance=template)
        formset = QuestionFormSet(instance=template)
    
    # Calculate statistics
    questions = template.questions.all()
    root_questions = questions.filter(parent_question__isnull=True)
    conditional_questions = questions.filter(parent_question__isnull=False)
    
    context = {
        'title': _('Edit Template'),
        'template': template,
        'form': form,
        'formset': formset,
        'root_questions_count': root_questions.count(),
        'conditional_questions_count': conditional_questions.count(),
        'max_score': template.get_total_score(),
    }
    return render(request, 'app_tprm/template_edit.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def template_delete(request, pk):
    """Delete questionnaire template"""
    template = get_object_or_404(QuestionnaireTemplate, pk=pk)
    
    if request.method == 'POST':
        template_name = template.name
        template.delete()
        messages.success(request, _('Template "%(name)s" deleted successfully!') % {'name': template_name})
        return redirect('app_tprm:template_list')
    
    context = {
        'title': _('Delete Template'),
        'template': template,
    }
    return render(request, 'app_tprm/template_confirm_delete.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def template_duplicate(request, pk):
    """Duplicate a questionnaire template"""
    original = get_object_or_404(QuestionnaireTemplate, pk=pk)
    
    # Create copy of template
    new_template = QuestionnaireTemplate.objects.create(
        name=f"{original.name} (Copy)",
        description=original.description,
        category=original.category,
        is_active=False,  # Inactive by default
        created_by=request.user
    )
    
    # Create question_mapping for parent relationships
    question_mapping = {}
    
    # Copy all questions
    for question in original.questions.all():
        new_question = Question.objects.create(
            template=new_template,
            question_text=question.question_text,
            question_type=question.question_type,
            choices=question.choices,
            weight=question.weight,
            correct_answer=question.correct_answer,
            order=question.order,
            is_required=question.is_required,
            help_text=question.help_text,
            show_if_answer=question.show_if_answer,
        )
        question_mapping[question.pk] = new_question
    
    # Update parent relationships
    for old_id, new_question in question_mapping.items():
        old_question = Question.objects.get(pk=old_id)
        if old_question.parent_question:
            new_question.parent_question = question_mapping[old_question.parent_question.pk]
            new_question.save()
    
    messages.success(request, _('Template duplicated successfully!'))
    return redirect('app_tprm:template_edit', pk=new_template.pk)


@login_required
@tprm_access_required('can_upload_documents')
def vendor_document_add(request, vendor_pk):
    """Add one or multiple documents to vendor"""
    vendor = get_object_or_404(Vendor, pk=vendor_pk)
    
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        document_type = request.POST.get('document_type')
        title = request.POST.get('title', '')
        description = request.POST.get('description', '')
        expiry_date = request.POST.get('expiry_date') or None
        
        if not files:
            messages.error(request, _('Please select at least one file to upload.'))
            return redirect('app_tprm:vendor_detail', pk=vendor.pk)
        
        if not document_type:
            messages.error(request, _('Please select document type.'))
            return redirect('app_tprm:vendor_detail', pk=vendor.pk)
        
        uploaded_count = 0
        for file in files:
            # Generate title from filename if not provided
            doc_title = title or file.name
            # If multiple files and title provided, append filename
            if len(files) > 1 and title:
                doc_title = f"{title} - {file.name}"
            
            document = VendorDocument.objects.create(
                vendor=vendor,
                document_type=document_type,
                title=doc_title,
                description=description,
                file=file,
                expiry_date=expiry_date,
                uploaded_by=request.user
            )
            uploaded_count += 1
        
        if uploaded_count == 1:
            messages.success(request, _('Document uploaded successfully!'))
        else:
            messages.success(request, _('%(count)s documents uploaded successfully!') % {'count': uploaded_count})
        
        return redirect('app_tprm:vendor_detail', pk=vendor.pk)
    
    context = {
        'title': _('Add Documents'),
        'vendor': vendor,
        'form': VendorDocumentForm(initial={'vendor': vendor}),
    }
    return render(request, 'app_tprm/vendor_document_add.html', context)


@login_required
@tprm_access_required('can_delete_documents')
def vendor_document_delete(request, pk):
    """Delete a single document"""
    document = get_object_or_404(VendorDocument, pk=pk)
    vendor = document.vendor
    
    if request.method == 'POST':
        document_title = document.title
        document.file.delete()  # Delete file from storage
        document.delete()
        messages.success(request, _('Document "%(title)s" deleted successfully!') % {'title': document_title})
        return redirect('app_tprm:vendor_detail', pk=vendor.pk)
    
    context = {
        'title': _('Delete Document'),
        'document': document,
        'vendor': vendor,
    }
    return render(request, 'app_tprm/vendor_document_confirm_delete.html', context)


@login_required
@tprm_access_required('can_delete_documents')
def vendor_document_bulk_delete(request, vendor_pk):
    """Delete multiple documents at once"""
    vendor = get_object_or_404(Vendor, pk=vendor_pk)
    
    if request.method == 'POST':
        document_ids = request.POST.getlist('document_ids')
        
        if not document_ids:
            messages.warning(request, _('No documents selected for deletion.'))
            return redirect('app_tprm:vendor_detail', pk=vendor.pk)
        
        documents = VendorDocument.objects.filter(pk__in=document_ids, vendor=vendor)
        deleted_count = 0
        
        for document in documents:
            document.file.delete()  # Delete file from storage
            document.delete()
            deleted_count += 1
        
        if deleted_count == 1:
            messages.success(request, _('Document deleted successfully!'))
        else:
            messages.success(request, _('%(count)s documents deleted successfully!') % {'count': deleted_count})
        
        return redirect('app_tprm:vendor_detail', pk=vendor.pk)
    
    messages.error(request, _('Invalid request method.'))
    return redirect('app_tprm:vendor_detail', pk=vendor.pk)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_list(request):
    """List all survey links"""
    links = VendorSurveyLink.objects.all().select_related('vendor', 'questionnaire', 'template', 'created_by')
    
    # Filter by vendor if provided
    vendor_id = request.GET.get('vendor')
    if vendor_id:
        links = links.filter(vendor_id=vendor_id)
    
    # Filter by status if provided
    status = request.GET.get('status')
    if status:
        links = links.filter(status=status)

    search = request.GET.get('search', '').strip()
    if search:
        links = links.filter(
            models.Q(token__icontains=search) |
            models.Q(vendor__name__icontains=search) |
            models.Q(status__icontains=search)
        )

    links = links.order_by('-created_at', '-pk')
    page_obj, pagination_context = paginate_tprm_queryset(request, links)

    context = {
        **pagination_context,
        'title': _('External Survey Links'),
        'links': page_obj,
        'search': search,
        'can_edit_vendors': check_tprm_access(request.user, 'can_edit_vendors'),
    }
    return render(request, 'app_tprm/survey_link_list.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_create(request):
    """Create new external survey link"""
    initial = {}
    
    # Pre-fill vendor if provided in query string
    vendor_id = request.GET.get('vendor')
    if vendor_id:
        try:
            vendor = Vendor.objects.get(pk=vendor_id)
            initial['vendor'] = vendor
        except Vendor.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = VendorSurveyLinkForm(request.POST)
        if form.is_valid():
            link = form.save(commit=False)
            link.created_by = request.user
            link.save()
            
            # Generate URL
            link_url = link.get_absolute_url(request)
            
            messages.success(request, _('Survey link created successfully!'))
            messages.info(request, _('Link URL: %(url)s') % {'url': link_url})
            return redirect('app_tprm:survey_link_detail', pk=link.pk)
    else:
        form = VendorSurveyLinkForm(initial=initial)
    
    context = {
        'title': _('Create External Survey Link'),
        'form': form,
    }
    return render(request, 'app_tprm/survey_link_form.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_detail(request, pk):
    """View survey link details"""
    link = get_object_or_404(VendorSurveyLink, pk=pk)
    link_url = link.get_absolute_url(request)
    
    context = {
        'title': _('Survey Link Details'),
        'link': link,
        'link_url': link_url,
        'is_valid': link.is_valid(),
    }
    return render(request, 'app_tprm/survey_link_detail.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_edit(request, pk):
    """Edit survey link"""
    link = get_object_or_404(VendorSurveyLink, pk=pk)
    
    if request.method == 'POST':
        form = VendorSurveyLinkForm(request.POST, instance=link)
        if form.is_valid():
            form.save()
            messages.success(request, _('Survey link updated successfully!'))
            return redirect('app_tprm:survey_link_detail', pk=link.pk)
    else:
        form = VendorSurveyLinkForm(instance=link)
    
    context = {
        'title': _('Edit Survey Link'),
        'form': form,
        'link': link,
    }
    return render(request, 'app_tprm/survey_link_form.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_revoke(request, pk):
    """Revoke a survey link"""
    link = get_object_or_404(VendorSurveyLink, pk=pk)
    
    if request.method == 'POST':
        link.status = 'revoked'
        link.save()
        messages.success(request, _('Survey link revoked successfully!'))
        return redirect('app_tprm:survey_link_detail', pk=link.pk)
    
    context = {
        'title': _('Revoke Survey Link'),
        'link': link,
    }
    return render(request, 'app_tprm/survey_link_confirm_revoke.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_delete(request, pk):
    """Delete a survey link"""
    link = get_object_or_404(VendorSurveyLink, pk=pk)
    
    if request.method == 'POST':
        link.delete()
        messages.success(request, _('Survey link deleted successfully!'))
        return redirect('app_tprm:survey_link_list')
    
    context = {
        'title': _('Delete Survey Link'),
        'link': link,
    }
    return render(request, 'app_tprm/survey_link_confirm_delete.html', context)


@login_required
@tprm_access_required('can_edit_templates')
def survey_link_get_questionnaires(request, vendor_id):
    """AJAX endpoint to get questionnaires for a vendor"""
    from django.http import JsonResponse
    
    try:
        vendor = Vendor.objects.get(pk=vendor_id)
        questionnaires = VendorQuestionnaire.objects.filter(vendor=vendor).select_related('template')
        
        data = [{
            'id': q.pk,
            'text': f"{q.template.name} ({q.get_status_display})"
        } for q in questionnaires]
        
        return JsonResponse({'questionnaires': data})
    except Vendor.DoesNotExist:
        return JsonResponse({'questionnaires': []}, status=404)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def survey_link_access(request, token):
    """Public access to questionnaire via external link (no authentication required)"""
    try:
        link = VendorSurveyLink.objects.get(token=token)
    except VendorSurveyLink.DoesNotExist:
        raise Http404(_('Survey link not found'))
    
    # Check if link is valid
    if not link.is_valid():
        context = {
            'title': _('Link Expired'),
            'link': link,
            'error': _('This survey link is no longer valid. It may have expired, been used, or revoked.'),
        }
        return render(request, 'app_tprm/survey_link_expired.html', context, status=403)
    
    # Get or create questionnaire
    questionnaire = link.get_questionnaire_or_create()
    if not questionnaire:
        raise Http404(_('Questionnaire not found'))
    
    # Record access only on GET requests (first visit)
    # For POST requests, we'll record after successful submission
    if request.method == 'GET':
        ip_address = request.META.get('REMOTE_ADDR')
        link.record_access(ip_address)
        
        # If questionnaire is not started, mark it as in progress
        if questionnaire.status == 'not_started':
            questionnaire.status = 'in_progress'
            questionnaire.started_date = timezone.now()
            questionnaire.save()
    
    # Handle form submission
    if request.method == 'POST':
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"POST request received for survey link {link.token}")
        logger.info(f"POST data keys: {list(request.POST.keys())}")
        
        try:
            responses = questionnaire.responses.all().select_related('question')
            logger.info(f"Found {responses.count()} responses to process")
            
            # Process responses
            processed_count = 0
            for response in responses:
                question = response.question
                field_name = f'question_{question.pk}'
                
                try:
                    if question.question_type == 'yes_no':
                        value = request.POST.get(field_name)
                        response.response_bool = value == 'yes' if value else None
                        logger.debug(f"Question {question.pk} (yes_no): value={value}, response_bool={response.response_bool}")
                    elif question.question_type == 'scale':
                        value = request.POST.get(field_name)
                        if value:
                            try:
                                response.response_scale = int(value)
                            except (ValueError, TypeError):
                                response.response_scale = None
                        else:
                            response.response_scale = None
                        logger.debug(f"Question {question.pk} (scale): value={value}, response_scale={response.response_scale}")
                    elif question.question_type == 'text':
                        response.response_text = request.POST.get(field_name, '')
                        logger.debug(f"Question {question.pk} (text): response_text length={len(response.response_text)}")
                    elif question.question_type == 'multiple_choice':
                        response.response_choice = request.POST.get(field_name, '')
                        logger.debug(f"Question {question.pk} (multiple_choice): response_choice={response.response_choice}")
                    
                    response.answered_at = timezone.now()
                    response.auto_score()
                    response.save()
                    processed_count += 1
                except Exception as e:
                    logger.error(f"Error processing response for question {question.pk}: {e}", exc_info=True)
                    continue
            
            logger.info(f"Processed {processed_count} responses successfully")
            
            # Update questionnaire status
            # Check action parameter (from button value) or fallback to button name
            action = request.POST.get('action', '')
            if action == 'complete' or 'complete' in request.POST:
                logger.info("Completing questionnaire")
                questionnaire.status = 'completed'
                questionnaire.completed_date = timezone.now()
                questionnaire.calculate_score()
                questionnaire.save()
                messages.success(request, _('Questionnaire completed successfully! Thank you for your response.'))
                # Redirect to completed page
                return render(request, 'app_tprm/survey_link_completed.html', {
                    'title': _('Questionnaire Completed'),
                    'link': link,
                    'questionnaire': questionnaire,
                })
            elif action == 'save' or 'save' in request.POST:
                logger.info("Saving progress")
                questionnaire.save()
                messages.success(request, _('Progress saved!'))
                # Redirect to avoid resubmission
                from django.urls import reverse
                return redirect(reverse('app_tprm:survey_link_access', kwargs={'token': link.token}))
            else:
                logger.warning(f"POST request but action not recognized. POST keys: {list(request.POST.keys())}, action='{action}'")
                # If we got here, it means form was submitted but action wasn't clear
                # Default to saving progress
                logger.info("No clear action, defaulting to save progress")
                questionnaire.save()
                messages.success(request, _('Progress saved!'))
                from django.urls import reverse
                return redirect(reverse('app_tprm:survey_link_access', kwargs={'token': link.token}))
        except Exception as e:
            logger.error(f"Error processing form submission: {e}", exc_info=True)
            messages.error(request, _('An error occurred while saving your responses. Please try again.'))
    
    # Get responses for display
    responses = questionnaire.responses.all().select_related('question').order_by('question__order')
    
    context = {
        'title': f"{questionnaire.vendor.name} - {questionnaire.template.name}",
        'link': link,
        'questionnaire': questionnaire,
        'responses': responses,
        'is_external': True,  # Flag to indicate external access
    }
    return render(request, 'app_tprm/survey_link_fill.html', context)
