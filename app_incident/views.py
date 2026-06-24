#  SecBoard\SecBoard\app_incident\incidents_views.py
import json
import logging
import os
import io
from datetime import datetime

import pytz
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.utils.timezone import make_aware
from django.utils.translation import get_language, gettext as _
from django.views.decorators.csrf import csrf_exempt, csrf_protect, ensure_csrf_cookie
from django.views.decorators.http import require_http_methods, require_POST
from .models import Incident, Currentstate, Classification, Incidenttype, AccessIncidents, IncidentFile, IncidentRegisterGuide, IncidentRegisterGuideTranslation
from app_conf.models import Company, MailAccount, Country
from django.core.serializers import serialize
from django.db import DatabaseError
from django.core.paginator import Paginator
from django.utils.dateparse import parse_datetime
from django.db.models import F, Q
from django.shortcuts import render, get_object_or_404
from django.db import transaction
from django.utils import timezone
from django.contrib import messages
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .forms import IncidentForm, IncidentFileForm, IncidentFilterForm
from .pagination_utils import get_incident_table_page_size, INCIDENT_TABLE_PAGE_SIZE_OPTIONS
from functools import wraps

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# For email functionality
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

from django.contrib.contenttypes.models import ContentType
from django.utils import formats

from app_cabinet.models import CabinetUser
from django.contrib.auth.models import User

logger = logging.getLogger(__name__)


logger.debug("This is a debug message")
logger.info("This is an info message")
logger.warning("This is a warning message")
logger.error("This is an error message")


def check_user_incident_access(user, company=None):
    """
    Check if user has access to incidents, optionally for a specific company.
    Returns a dictionary with has_access, can_edit, can_add, can_delete, and can_mail flags.
    """
    result = {
        'has_access': False, 
        'can_edit': False,
        'can_add': False,
        'can_delete': False,
        'can_mail': False
    }
    
    # Superuser has all permissions
    if user.is_superuser:
        result['has_access'] = True
        result['can_edit'] = True
        result['can_add'] = True
        result['can_delete'] = True
        result['can_mail'] = True
        return result
    
    # Get user groups
    user_groups = user.groups.all()
    
    # Check access based on AccessIncidents model
    for group in user_groups:
        try:
            access = AccessIncidents.objects.get(group=group)
            if access.has_access:
                # If checking access for a specific company
                if company:
                    if company in access.companies.all():
                        result['has_access'] = True
                        if access.can_edit:
                            result['can_edit'] = True
                        if access.can_add:
                            result['can_add'] = True
                        if access.can_delete:
                            result['can_delete'] = True
                        if access.can_mail:
                            result['can_mail'] = True
                # If checking general access
                else:
                    result['has_access'] = True
                    if access.can_edit:
                        result['can_edit'] = True
                    if access.can_add:
                        result['can_add'] = True
                    if access.can_delete:
                        result['can_delete'] = True
                    if access.can_mail:
                        result['can_mail'] = True
        except AccessIncidents.DoesNotExist:
            continue
    
    return result


def get_user_accessible_companies(user):
    """
    Get QuerySet of companies user has access to based on AccessIncidents
    """
    # Superuser has access to all companies
    if user.is_superuser:
        return Company.objects.all()
    
    # Get user groups
    user_groups = user.groups.all()
    
    # Get accessible companies
    company_ids = set()
    for group in user_groups:
        try:
            access = AccessIncidents.objects.get(group=group, has_access=True)
            for company in access.companies.all():
                company_ids.add(company.id)
        except AccessIncidents.DoesNotExist:
            continue
    
    # Return a QuerySet of accessible companies
    return Company.objects.filter(id__in=company_ids)


def user_can_edit_incident(view_func):
    """
    Decorator to check if user has edit permissions for incidents.
    If a specific incident is involved, checks access for that incident's company.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # If editing a specific incident, check company-specific access
        incident_id = kwargs.get('incident_id')
        if incident_id:
            try:
                incident = Incident.objects.get(id=incident_id)
                access = check_user_incident_access(request.user, incident.company)
                if not access['can_edit']:
                    messages.error(request, _("You don't have permission to edit this incident."))
                    return redirect('incident_detail', incident_id=incident_id)
            except Incident.DoesNotExist:
                messages.error(request, _("Incident not found."))
                return redirect('incident_register')
        else:
            # Check general edit access for adding new incidents
            access = check_user_incident_access(request.user)
            if not access['can_edit']:
                messages.error(request, _("You don't have permission to add or edit incidents."))
                return redirect('incident_register')
        
        return view_func(request, *args, **kwargs)
    
    return _wrapped_view


def user_can_add_incident(view_func):
    """
    Decorator to check if user has add permissions for incidents.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Check general add access
        access = check_user_incident_access(request.user)
        if not access['can_add']:
            messages.error(request, _("You don't have permission to add new incidents."))
            return redirect('incident_register')
        
        return view_func(request, *args, **kwargs)
    
    return _wrapped_view


def user_can_delete_incident(view_func):
    """
    Decorator to check if user has delete permissions for incidents.
    If a specific incident is involved, checks access for that incident's company.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # If deleting a specific incident, check company-specific access
        incident_id = kwargs.get('incident_id')
        file_id = kwargs.get('file_id')
        
        if incident_id:
            try:
                incident = Incident.objects.get(id=incident_id)
                access = check_user_incident_access(request.user, incident.company)
                if not access['can_delete']:
                    messages.error(request, _("You don't have permission to delete this incident."))
                    return redirect('incident_detail', incident_id=incident_id)
            except Incident.DoesNotExist:
                messages.error(request, _("Incident not found."))
                return redirect('incident_register')
        elif file_id:
            try:
                file = IncidentFile.objects.get(id=file_id)
                access = check_user_incident_access(request.user, file.incident.company)
                if not access['can_delete']:
                    messages.error(request, _("You don't have permission to delete files for this incident."))
                    return redirect('incident_detail', incident_id=file.incident.id)
            except IncidentFile.DoesNotExist:
                messages.error(request, _("File not found."))
                return redirect('incident_register')
        else:
            # Check general delete access
            access = check_user_incident_access(request.user)
            if not access['can_delete']:
                messages.error(request, _("You don't have permission to delete incidents."))
                return redirect('incident_register')
        
        return view_func(request, *args, **kwargs)
    
    return _wrapped_view


def user_can_mail_incident(view_func):
    """
    Decorator to check if user has email permissions for incidents.
    """
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Check general mail access - for specific incident, we'll check in the view
        access = check_user_incident_access(request.user)
        if not access['can_mail']:
            return JsonResponse({
                'success': False, 
                'message': _("You don't have permission to send emails for incidents.")
            })
        
        return view_func(request, *args, **kwargs)
    
    return _wrapped_view


@login_required
def incident_register(request):
    """
    View for displaying the incident register page with a paginated table of incidents.
    """
    # Initialize filters
    filter_form = IncidentFilterForm(request.GET, user=request.user)
    incidents_list = Incident.objects.all().select_related('company', 'classification', 'incident_type', 'current_state')
    
    # Check access permissions
    access = check_user_incident_access(request.user)
    accessible_companies = get_user_accessible_companies(request.user)
    
    # Filter incidents by accessible companies if not superuser
    if not request.user.is_superuser and accessible_companies:
        incidents_list = incidents_list.filter(company__in=accessible_companies)
    elif not access['has_access']:
        # If no access at all, show no incidents
        incidents_list = Incident.objects.none()
    
    # Apply filters if form is valid
    if filter_form.is_valid():
        # Filter by company
        company = filter_form.cleaned_data.get('company')
        if company:
            incidents_list = incidents_list.filter(company=company)
        
        # Filter by classification
        classification = filter_form.cleaned_data.get('classification')
        if classification:
            incidents_list = incidents_list.filter(classification=classification)
        
        # Filter by incident type
        incident_type = filter_form.cleaned_data.get('incident_type')
        if incident_type:
            incidents_list = incidents_list.filter(incident_type=incident_type)
        
        # Filter by current state
        current_state = filter_form.cleaned_data.get('current_state')
        if current_state:
            incidents_list = incidents_list.filter(current_state=current_state)
        
        # Filter by date range
        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            incidents_list = incidents_list.filter(occurrence_datetime__date__gte=date_from)
        
        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            incidents_list = incidents_list.filter(occurrence_datetime__date__lte=date_to)
    
    # Get sort parameters
    sort_by = request.GET.get('sort', 'id')  # Default sort by ID
    sort_dir = request.GET.get('dir', 'desc')  # Default sort direction is descending (newest first)
    
    # Validate and apply sorting
    valid_sort_fields = {
        'id': 'id',
        '-id': '-id',
        'datetime': 'occurrence_datetime',
        '-datetime': '-occurrence_datetime'
    }
    
    # Determine sort field and direction
    if sort_by in ['id', 'datetime']:
        sort_field = valid_sort_fields[sort_by]
        if sort_dir == 'desc':
            sort_field = f"-{sort_field}"
    else:
        # Default to ID descending (latest first)
        sort_field = '-id'
    
    # Apply sorting
    incidents_list = incidents_list.order_by(sort_field)
    
    # Pagination
    page_size = get_incident_table_page_size(request)
    paginator = Paginator(incidents_list, page_size)
    page_obj = paginator.get_page(request.GET.get('page'))
    incidents = page_obj

    context = {
        'incidents': incidents,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.count > 0,
        'current_page_size': page_size,
        'page_size_options': INCIDENT_TABLE_PAGE_SIZE_OPTIONS,
        'can_edit': access['can_edit'],
        'can_add': access['can_add'],
        'can_delete': access['can_delete'],
        'can_mail': access['can_mail'],
        'accessible_companies': accessible_companies,
        'filter_form': filter_form,
        'page_size': page_size,
        'sort_by': sort_by,
        'sort_dir': sort_dir,
        'pagination_item_label': _('incidents'),
    }
    
    return render(request, 'app_incident/incident_register.html', context)


@login_required
@require_http_methods(["GET"])
def incident_register_guide(request):
    """Return JSON { content: html } for the Incident Register guide (localized)."""
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = IncidentRegisterGuide.objects.first()
    if guide:
        if country:
            trans = IncidentRegisterGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = IncidentRegisterGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def incident_register_guide_translate(request):
    """API for AI translation of Incident Register guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        err = str(e)
        if 'No support for the provided language' in err:
            return JsonResponse({
                'success': True,
                'translated_text': text,
                'target_language': target,
                'country_name': country.name,
                'warning': 'Language not supported, returned original',
            })
        return JsonResponse({'success': False, 'error': err}, status=500)


@login_required
def incident_detail(request, incident_id):
    """
    View for displaying details of a specific incident.
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    # Check access permissions
    access = check_user_incident_access(request.user, incident.company)
    
    if not access['has_access']:
        messages.error(request, _("You don't have permission to view this incident."))
        return redirect('incident_register')
    
    context = {
        'incident': incident,
        'can_edit': access['can_edit'],
        'can_delete': access['can_delete'],
        'can_mail': access['can_mail'],
    }
    
    return render(request, 'app_incident/incident_detail.html', context)


@login_required
@user_can_add_incident
def incident_add(request):
    """
    View for adding a new incident.
    """
    # Ensure media directory exists
    media_root = settings.MEDIA_ROOT
    upload_dir = os.path.join(media_root, 'incident_reports')
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir, exist_ok=True)
        logger.info(f"Created incident reports directory: {upload_dir}")
    
    # Get accessible companies
    accessible_companies = get_user_accessible_companies(request.user)
    
    if request.method == 'POST':
        # Create a form instance with POST data and files
        form = IncidentForm(request.POST, request.FILES, available_companies=accessible_companies)
        
        if form.is_valid():
            try:
                # Use transaction.atomic to ensure database consistency
                with transaction.atomic():
                    # Create incident but don't save to DB yet
                    incident = form.save(commit=False)
                    
                    # Add fields not in the form
                    incident.registered_by = request.user.get_full_name() or request.user.username
                    
                    # Set default dates if not provided
                    if not incident.reported_datetime:
                        incident.reported_datetime = timezone.now()
                    if not incident.registered_datetime:
                        incident.registered_datetime = timezone.now()
                    
                    # Save to DB
                    incident.save()
                    
                    # Process additional files (support multiple files)
                    files = request.FILES.getlist('additional_files')
                    logger.debug(f"Processing {len(files)} additional files")
                    
                    for uploaded_file in files:
                        # Create each file record
                        incident_file = IncidentFile(
                            incident=incident,
                            file=uploaded_file,
                            filename=uploaded_file.name
                        )
                        incident_file.save()
                        logger.debug(f"Saved file: {uploaded_file.name}")
                
                messages.success(request, _("Incident added successfully."))
                return redirect('incident_detail', incident_id=incident.id)
                
            except Exception as e:
                logger.error(f"Error adding incident: {str(e)}")
                messages.error(request, _("An error occurred while adding the incident."))
        else:
            logger.warning(f"Form validation errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # GET request - create a blank form
        form = IncidentForm(available_companies=accessible_companies)
        
        # Set initial values
        initial_data = {
            'registered_by': request.user.get_full_name() or request.user.username,
        }
        for field, value in initial_data.items():
            if field in form.fields:
                form.fields[field].initial = value
    
    # Prepare form context
    context = {
        'form': form,
        'companies': accessible_companies,
        'classifications': Classification.objects.filter(is_active=True).order_by('name'),
        'incident_types': Incidenttype.objects.filter(is_active=True).order_by('name'),
        'current_states': Currentstate.objects.filter(is_active=True).order_by('name'),
    }
    
    return render(request, 'app_incident/incident_form.html', context)


@login_required
@user_can_edit_incident
def incident_edit(request, incident_id):
    """
    View for editing an existing incident.
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    # Get accessible companies
    accessible_companies = get_user_accessible_companies(request.user)
    
    if request.method == 'POST':
        # Create a form instance with POST data, files, and the instance to edit
        form = IncidentForm(request.POST, request.FILES, instance=incident, available_companies=accessible_companies)
        
        if form.is_valid():
            try:
                # Use transaction.atomic to ensure database consistency
                with transaction.atomic():
                    # Save form but don't commit yet
                    incident = form.save(commit=False)
                    
                    # Set the registered_by field if it's empty
                    if not incident.registered_by:
                        incident.registered_by = request.user.get_full_name() or request.user.username
                    
                    # Set default dates if not provided
                    if not incident.reported_datetime:
                        incident.reported_datetime = timezone.now()
                    if not incident.registered_datetime:
                        incident.registered_datetime = timezone.now()
                    
                    # Save the changes
                    incident.save()
                    
                    # Process additional files (support multiple files)
                    files = request.FILES.getlist('additional_files')
                    logger.debug(f"Processing {len(files)} additional files for edit")
                    
                    for uploaded_file in files:
                        # Create each file record
                        incident_file = IncidentFile(
                            incident=incident,
                            file=uploaded_file,
                            filename=uploaded_file.name
                        )
                        incident_file.save()
                        logger.debug(f"Saved file: {uploaded_file.name}")
                
                messages.success(request, _("Incident updated successfully."))
                return redirect('incident_detail', incident_id=incident.id)
                
            except Exception as e:
                logger.error(f"Error updating incident: {str(e)}")
                messages.error(request, _("An error occurred while updating the incident."))
        else:
            logger.warning(f"Form validation errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # For GET requests, initialize the form with the incident instance
        form = IncidentForm(instance=incident, available_companies=accessible_companies)
    
    # Prepare form context
    context = {
        'form': form,
        'incident': incident,
        'companies': accessible_companies,
        'classifications': Classification.objects.filter(is_active=True).order_by('name'),
        'incident_types': Incidenttype.objects.filter(is_active=True).order_by('name'),
        'current_states': Currentstate.objects.filter(is_active=True).order_by('name'),
        'edit_mode': True
    }
    
    return render(request, 'app_incident/incident_form.html', context)


@login_required
@user_can_delete_incident
def incident_delete(request, incident_id):
    """
    View for deleting an existing incident.
    """
    logger.info(f"Delete view accessed for incident #{incident_id} by user {request.user.username}")
    incident = get_object_or_404(Incident, id=incident_id)
    
    # Handle the actual deletion
    if request.method == 'POST':
        try:
            # Save incident details for the success message
            incident_id_str = str(incident.id)
            company_name = incident.company.name
            
            logger.info(f"About to delete incident #{incident_id_str} for {company_name}")
            
            # Delete the incident
            incident.delete()
            
            logger.info(f"Successfully deleted incident #{incident_id_str}")
            
            messages.success(request, _("Incident #%(id)s for %(company)s was deleted successfully.") % 
                             {'id': incident_id_str, 'company': company_name})
            return redirect('incident_register')
            
        except Exception as e:
            logger.error(f"Error deleting incident #{incident_id}: {str(e)}")
            logger.exception("Detailed exception information:")
            messages.error(request, _("An error occurred while deleting the incident."))
            return redirect('incident_detail', incident_id=incident.id)
    
    # If it's a GET request, show the confirmation page
    logger.info(f"Rendering delete confirmation page for incident #{incident_id}")
    context = {
        'incident': incident,
    }
    
    return render(request, 'app_incident/incident_delete.html', context)


@login_required
@user_can_delete_incident
def file_delete(request, file_id):
    """
    View for deleting a specific file.
    """
    file = get_object_or_404(IncidentFile, id=file_id)
    incident_id = file.incident.id
    
    # Handle the actual deletion
    if request.method == 'POST':
        try:
            # Save file details for the success message
            file_id_str = str(file.id)
            company_name = file.incident.company.name
            
            logger.info(f"About to delete file #{file_id_str} for {company_name}")
            
            # Delete the file
            file.delete()
            
            logger.info(f"Successfully deleted file #{file_id_str}")
            
            messages.success(request, _("File #%(id)s for %(company)s was deleted successfully.") % 
                             {'id': file_id_str, 'company': company_name})
            return redirect('incident_detail', incident_id=incident_id)
            
        except Exception as e:
            logger.error(f"Error deleting file #{file_id}: {str(e)}")
            logger.exception("Detailed exception information:")
            messages.error(request, _("An error occurred while deleting the file."))
            return redirect('incident_detail', incident_id=incident_id)
    
    # If it's a GET request, show the confirmation page
    context = {
        'file': file,
    }
    
    return render(request, 'app_incident/file_delete.html', context)


@login_required
@require_POST
def delete_file_ajax(request):
    """
    AJAX endpoint for deleting a file
    """
    try:
        logger.debug(f"Request content type: {request.content_type}")
        logger.debug(f"Request body: {request.body[:100]}...")
        
        # Handle both JSON and form data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            file_id = data.get('file_id')
        else:
            file_id = request.POST.get('file_id')
        
        logger.debug(f"Parsed file_id: {file_id}")
        
        if not file_id:
            logger.warning("File ID is missing in request")
            return JsonResponse({'success': False, 'error': _('File ID is required')})
        
        # Get the file
        try:
            file = IncidentFile.objects.get(id=file_id)
        except IncidentFile.DoesNotExist:
            logger.warning(f"File with ID {file_id} not found")
            return JsonResponse({'success': False, 'error': _('File not found')})
        
        # Check permissions
        access = check_user_incident_access(request.user, file.incident.company)
        if not access['can_delete']:
            logger.warning(f"User {request.user.username} does not have permission to delete file {file_id}")
            return JsonResponse({'success': False, 'error': _('You do not have permission to delete this file')})
        
        # Store incident ID for logging
        incident_id = file.incident.id
        filename = file.filename
        
        # Get the file path to delete the actual file
        file_path = file.file.path if file.file else None
        
        # Delete the file from database
        file.delete()
        
        # Attempt to delete the physical file if it exists
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
                logger.info(f"Physical file deleted: {file_path}")
            except OSError as e:
                logger.warning(f"Could not delete physical file {file_path}: {str(e)}")
        
        # Log the deletion
        logger.info(f"File '{filename}' (ID: {file_id}) for incident #{incident_id} was deleted by {request.user.username}")
        
        return JsonResponse({'success': True})
        
    except json.JSONDecodeError:
        logger.error("Invalid JSON in request")
        return JsonResponse({'success': False, 'error': _('Invalid JSON')})
    except Exception as e:
        logger.error(f"Error in delete_file_ajax: {str(e)}")
        logger.exception("Exception details:")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def delete_main_incident_file(request):
    """
    AJAX endpoint for deleting the main incident file
    """
    try:
        logger.debug(f"Request content type: {request.content_type}")
        logger.debug(f"Request body: {request.body[:100]}...")
        
        # Handle both JSON and form data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            incident_id = data.get('incident_id')
        else:
            incident_id = request.POST.get('incident_id')
        
        logger.debug(f"Parsed incident_id: {incident_id}")
        
        if not incident_id:
            logger.warning("Incident ID is missing in request")
            return JsonResponse({'success': False, 'error': _('Incident ID is required')})
        
        # Get the incident
        try:
            # Convert to integer in case it's passed as a string
            incident_id = int(incident_id)
            incident = Incident.objects.get(id=incident_id)
            logger.debug(f"Found incident: {incident}")
        except (ValueError, TypeError):
            logger.warning(f"Invalid incident ID format: {incident_id}")
            return JsonResponse({'success': False, 'error': _('Invalid incident ID format')})
        except Incident.DoesNotExist:
            logger.warning(f"Incident with ID {incident_id} not found")
            return JsonResponse({'success': False, 'error': _('Incident not found')})
        
        # Check permissions
        access = check_user_incident_access(request.user, incident.company)
        if not access['can_delete']:
            logger.warning(f"User {request.user.username} does not have permission to delete main file for incident {incident_id}")
            return JsonResponse({'success': False, 'error': _('You do not have permission to delete this file')})
        
        # Store filename for logging
        filename = incident.file_incident.name if incident.file_incident else "Unknown"
        
        # Get the file path to delete the actual file
        file_path = incident.file_incident.path if incident.file_incident else None
        
        # Clear the file field
        incident.file_incident = None
        incident.save()
        
        # Attempt to delete the physical file if it exists
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
                logger.info(f"Physical main file deleted: {file_path}")
            except OSError as e:
                logger.warning(f"Could not delete physical file {file_path}: {str(e)}")
        
        # Log the deletion
        logger.info(f"Main file '{filename}' for incident #{incident_id} was deleted by {request.user.username}")
        
        return JsonResponse({'success': True})
        
    except json.JSONDecodeError:
        logger.error("Invalid JSON in request")
        return JsonResponse({'success': False, 'error': _('Invalid JSON')})
    except Exception as e:
        logger.error(f"Error in delete_main_incident_file: {str(e)}")
        logger.exception("Exception details:")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def export_incidents_excel(request):
    """
    Export incidents to Excel with enhanced styling
    """
    # Get filtered incidents using the same logic as incident_register
    filter_form = IncidentFilterForm(request.GET, user=request.user)
    incidents_list = Incident.objects.all().select_related('company', 'classification', 'incident_type', 'current_state')
    
    # Check access permissions
    access = check_user_incident_access(request.user)
    accessible_companies = get_user_accessible_companies(request.user)
    
    # Filter incidents by accessible companies if not superuser
    if not request.user.is_superuser and accessible_companies:
        incidents_list = incidents_list.filter(company__in=accessible_companies)
    elif not access['has_access']:
        # If no access at all, show no incidents
        incidents_list = Incident.objects.none()
    
    # Apply filters if form is valid (same as in incident_register)
    if filter_form.is_valid():
        # Filter by company
        company = filter_form.cleaned_data.get('company')
        if company:
            incidents_list = incidents_list.filter(company=company)
        
        # Filter by classification
        classification = filter_form.cleaned_data.get('classification')
        if classification:
            incidents_list = incidents_list.filter(classification=classification)
        
        # Filter by incident type
        incident_type = filter_form.cleaned_data.get('incident_type')
        if incident_type:
            incidents_list = incidents_list.filter(incident_type=incident_type)
        
        # Filter by current state
        current_state = filter_form.cleaned_data.get('current_state')
        if current_state:
            incidents_list = incidents_list.filter(current_state=current_state)
        
        # Filter by date range
        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            incidents_list = incidents_list.filter(occurrence_datetime__date__gte=date_from)
        
        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            incidents_list = incidents_list.filter(occurrence_datetime__date__lte=date_to)
    
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = _("Incidents")
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    section_font = Font(name='Arial', size=12, bold=True)
    section_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    section_alignment = Alignment(horizontal='left', vertical='center')
    
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    date_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers - more comprehensive list
    headers = [
        _("ID"),
        _("Company"),
        _("Date & Time of Occurrence"),
        _("Place"),
        _("Classification"),
        _("Incident Type"),
        _("Current State"),
        _("Reported By"),
        _("Reported Date & Time"),
        _("Registered By"),
        _("Registered Date & Time"),
        _("Description"),
        _("Features and Signs"),
        _("Impact"),
        _("Responsible for Resolution"),
        _("Measures Taken"),
        _("Additional Measures"),
        _("Reports and Records"),
        _("Comment"),
        _("Has Files"),
        _("Created At"),
        _("Updated At")
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set language for localized content
    language_code = get_language()
    
    # Write data rows
    for row_num, incident in enumerate(incidents_list, 2):
        # ID
        cell = ws.cell(row=row_num, column=1, value=incident.id)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Company
        cell = ws.cell(row=row_num, column=2, value=incident.company.name)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Date & Time of Occurrence
        if incident.occurrence_datetime:
            # Remove timezone information from datetime
            naive_datetime = incident.occurrence_datetime.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=3, value=naive_datetime)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=3, value="-")
        cell.font = data_font
        cell.alignment = date_alignment
        cell.border = thin_border
        
        # Place
        cell = ws.cell(row=row_num, column=4, value=incident.place)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Classification
        if incident.classification:
            classification_text = incident.classification.get_name()
            cell = ws.cell(row=row_num, column=5, value=classification_text)
            # Try to convert hex color to RGB for Excel
            try:
                hex_color = incident.classification.color.lstrip('#')
                rgb_color = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
                fill = PatternFill(start_color=incident.classification.color.lstrip('#'), 
                                  end_color=incident.classification.color.lstrip('#'), 
                                  fill_type='solid')
                cell.fill = fill
                # Add white font if background is dark
                luminance = (0.299 * rgb_color[0] + 0.587 * rgb_color[1] + 0.114 * rgb_color[2]) / 255
                if luminance < 0.5:
                    cell.font = Font(name='Arial', size=11, color='FFFFFF')
            except:
                # Fallback if color processing fails
                pass
        else:
            cell = ws.cell(row=row_num, column=5, value="-")
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Incident Type
        if incident.incident_type:
            incident_type_text = incident.incident_type.get_name()
            cell = ws.cell(row=row_num, column=6, value=incident_type_text)
            # Try to convert hex color to RGB for Excel
            try:
                hex_color = incident.incident_type.color.lstrip('#')
                rgb_color = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
                fill = PatternFill(start_color=incident.incident_type.color.lstrip('#'), 
                                  end_color=incident.incident_type.color.lstrip('#'), 
                                  fill_type='solid')
                cell.fill = fill
                # Add white font if background is dark
                luminance = (0.299 * rgb_color[0] + 0.587 * rgb_color[1] + 0.114 * rgb_color[2]) / 255
                if luminance < 0.5:
                    cell.font = Font(name='Arial', size=11, color='FFFFFF')
            except:
                # Fallback if color processing fails
                pass
        else:
            cell = ws.cell(row=row_num, column=6, value="-")
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Current State
        if incident.current_state:
            current_state_text = incident.current_state.get_name()
            cell = ws.cell(row=row_num, column=7, value=current_state_text)
            # Try to convert hex color to RGB for Excel
            try:
                hex_color = incident.current_state.color.lstrip('#')
                rgb_color = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
                fill = PatternFill(start_color=incident.current_state.color.lstrip('#'), 
                                  end_color=incident.current_state.color.lstrip('#'), 
                                  fill_type='solid')
                cell.fill = fill
                # Add white font if background is dark
                luminance = (0.299 * rgb_color[0] + 0.587 * rgb_color[1] + 0.114 * rgb_color[2]) / 255
                if luminance < 0.5:
                    cell.font = Font(name='Arial', size=11, color='FFFFFF')
            except:
                # Fallback if color processing fails
                pass
        else:
            cell = ws.cell(row=row_num, column=7, value="-")
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Reported By
        cell = ws.cell(row=row_num, column=8, value=incident.reported_by)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Reported Date & Time
        if incident.reported_datetime:
            naive_datetime = incident.reported_datetime.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=9, value=naive_datetime)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=9, value="-")
        cell.font = data_font
        cell.alignment = date_alignment
        cell.border = thin_border
        
        # Registered By
        cell = ws.cell(row=row_num, column=10, value=incident.registered_by)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Registered Date & Time
        if incident.registered_datetime:
            naive_datetime = incident.registered_datetime.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=11, value=naive_datetime)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=11, value="-")
        cell.font = data_font
        cell.alignment = date_alignment
        cell.border = thin_border
        
        # Description
        cell = ws.cell(row=row_num, column=12, value=incident.description)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Features and Signs
        cell = ws.cell(row=row_num, column=13, value=incident.features)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Impact
        cell = ws.cell(row=row_num, column=14, value=incident.impact)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Responsible for Resolution
        cell = ws.cell(row=row_num, column=15, value=incident.responsible)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Measures Taken
        cell = ws.cell(row=row_num, column=16, value=incident.measures_taken)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Additional Measures
        cell = ws.cell(row=row_num, column=17, value=incident.additional_measures)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Reports and Records
        cell = ws.cell(row=row_num, column=18, value=incident.reports_and_records)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Comment
        cell = ws.cell(row=row_num, column=19, value=incident.comment)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Has Files
        has_files = "✓" if incident.files.exists() or incident.file_incident else "✗"
        cell = ws.cell(row=row_num, column=20, value=has_files)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        
        # Created At
        if incident.created_at:
            naive_datetime = incident.created_at.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=21, value=naive_datetime)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=21, value="-")
        cell.font = data_font
        cell.alignment = date_alignment
        cell.border = thin_border
        
        # Updated At
        if incident.updated_at:
            naive_datetime = incident.updated_at.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=22, value=naive_datetime)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=22, value="-")
        cell.font = data_font
        cell.alignment = date_alignment
        cell.border = thin_border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 40)  # Cap width at 40
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Create a second sheet for file information
    if incidents_list.exists():
        # Get all incident IDs
        incident_ids = incidents_list.values_list('id', flat=True)
        
        # Fetch all files for these incidents
        all_files = IncidentFile.objects.filter(incident_id__in=incident_ids).select_related('incident')
        
        # If there are files or any incidents have a main file, create the files sheet
        if all_files.exists() or Incident.objects.filter(id__in=incident_ids, file_incident__isnull=False).exists():
            ws_files = wb.create_sheet(title=_("Files"))
            
            # Define file headers
            file_headers = [
                _("Incident ID"),
                _("Company"),
                _("File Type"),
                _("Filename"),
                _("Uploaded At"),
                _("File Path")
            ]
            
            # Write file headers
            for col_num, header in enumerate(file_headers, 1):
                cell = ws_files.cell(row=1, column=col_num, value=str(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            row_num = 2
            
            # Add main incident files
            for incident in incidents_list.filter(file_incident__isnull=False):
                # Incident ID
                cell = ws_files.cell(row=row_num, column=1, value=incident.id)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                
                # Company
                cell = ws_files.cell(row=row_num, column=2, value=incident.company.name)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # File Type
                cell = ws_files.cell(row=row_num, column=3, value=_("Main Incident File"))
                cell.font = Font(name='Arial', size=11, bold=True)
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Filename
                filename = incident.file_incident.name.split('/')[-1] if incident.file_incident else "-"
                cell = ws_files.cell(row=row_num, column=4, value=filename)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Uploaded At
                if incident.created_at:
                    naive_datetime = incident.created_at.replace(tzinfo=None)
                    cell = ws_files.cell(row=row_num, column=5, value=naive_datetime)
                    cell.number_format = 'DD.MM.YYYY HH:MM'
                else:
                    cell = ws_files.cell(row=row_num, column=5, value="-")
                cell.font = data_font
                cell.alignment = date_alignment
                cell.border = thin_border
                
                # File Path
                cell = ws_files.cell(row=row_num, column=6, value=incident.file_incident.name if incident.file_incident else "-")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                row_num += 1
            
            # Add additional files
            for file in all_files:
                # Incident ID
                cell = ws_files.cell(row=row_num, column=1, value=file.incident.id)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                
                # Company
                cell = ws_files.cell(row=row_num, column=2, value=file.incident.company.name)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # File Type
                cell = ws_files.cell(row=row_num, column=3, value=_("Additional File"))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Filename
                cell = ws_files.cell(row=row_num, column=4, value=file.filename)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Uploaded At
                if file.uploaded_at:
                    naive_datetime = file.uploaded_at.replace(tzinfo=None)
                    cell = ws_files.cell(row=row_num, column=5, value=naive_datetime)
                    cell.number_format = 'DD.MM.YYYY HH:MM'
                else:
                    cell = ws_files.cell(row=row_num, column=5, value="-")
                cell.font = data_font
                cell.alignment = date_alignment
                cell.border = thin_border
                
                # File Path
                cell = ws_files.cell(row=row_num, column=6, value=file.file.name if file.file else "-")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                row_num += 1
            
            # Auto-adjust column widths for files sheet
            for column in ws_files.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws_files.column_dimensions[column_letter].width = min(adjusted_width, 40)  # Cap width at 40
            
            # Set row height for header
            ws_files.row_dimensions[1].height = 30
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Incidents_{}.xlsx'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
def get_incident_email_template(request):
    """Get email template with incident details for the specified incident."""
    incident_id = request.GET.get('incident_id')
    if not incident_id:
        return JsonResponse({'success': False, 'message': _('Incident ID is required')})
    
    try:
        incident = Incident.objects.get(id=incident_id)
        
        # Check if user has permission to view this incident
        access = check_user_incident_access(request.user, incident.company)
        if not access['has_access']:
            return JsonResponse({'success': False, 'message': _('You do not have permission to view this incident')})
        
        # Check if user has permission to send emails
        if not access['can_mail']:
            return JsonResponse({'success': False, 'message': _('You do not have permission to send emails for this incident')})
        
        # Prepare email subject
        subject = _("Incident Report") + f" #{incident.id}: {incident.place}"
        
        # Prepare incident details for email body
        classification = incident.classification.get_name() if incident.classification else ""
        
        incident_type = incident.incident_type.get_name() if incident.incident_type else ""
        
        current_state = incident.current_state.get_name() if incident.current_state else ""
        
        # Format the date string according to the locale
        date_format = "d E Y H:i"  # Default format for Ukrainian and Russian
        if request.LANGUAGE_CODE == 'en':
            date_format = "M d, Y H:i"  # English format
            
        def format_datetime(dt):
            if not dt:
                return _('N/A')
            try:
                return formats.date_format(dt, date_format)
            except:
                return dt.strftime('%d.%m.%Y %H:%M')
        
        occurrence_date = format_datetime(incident.occurrence_datetime)
        reported_date = format_datetime(incident.reported_datetime)
        registered_date = format_datetime(incident.registered_datetime)
        
        # Format the email message with incident details using HTML for better presentation
        # Creating an HTML email instead of plain text for better formatting
        message_html = f"""<!DOCTYPE html>
<html lang="{request.LANGUAGE_CODE}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{_('Incident Report')}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 800px;
            margin: 0 auto;
            background-color: #f8f9fa;
        }}
        .container {{
            background-color: #fff;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            overflow: hidden;
            margin: 20px auto;
        }}
        .header {{
            background-color: #234a87;
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .content {{
            padding: 20px;
        }}
        .greeting {{
            margin-bottom: 20px;
            font-size: 16px;
            color: #333;
        }}
        .section {{
            margin-bottom: 25px;
            padding: 15px;
            background-color: #f9f9f9;
            border-left: 5px solid #3498db;
            border-radius: 4px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .section-title {{
            font-weight: bold;
            font-size: 18px;
            margin-bottom: 12px;
            color: #2c3e50;
            text-transform: uppercase;
            border-bottom: 2px solid #3498db;
            padding-bottom: 6px;
        }}
        .field {{
            margin-bottom: 10px;
            display: flex;
            flex-wrap: wrap;
        }}
        .field-name {{
            font-weight: bold;
            min-width: 180px;
            color: #555;
            padding-right: 10px;
        }}
        .field-value {{
            flex: 1;
            min-width: 60%;
        }}
        .highlight {{
            font-weight: bold;
            color: #c0392b;
        }}
        .footer {{
            margin-top: 30px;
            padding: 15px;
            border-top: 1px solid #eee;
            font-size: 14px;
            color: #777;
            background-color: #f9f9f9;
            text-align: center;
        }}
        .classification {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 3px;
            color: white;
            font-weight: bold;
            background-color: {incident.classification.color if incident.classification else '#6c757d'};
        }}
        .incident-type {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 3px;
            color: white;
            font-weight: bold;
            background-color: {incident.incident_type.color if incident.incident_type else '#6c757d'};
        }}
        .current-state {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 3px;
            color: white;
            font-weight: bold;
            background-color: {incident.current_state.color if incident.current_state else '#6c757d'};
        }}
        .note {{
            background-color: #fcf8e3;
            border-left: 5px solid #f0ad4e;
            padding: 15px;
            margin-top: 20px;
            border-radius: 4px;
        }}
        /* Responsive styles */
        @media screen and (max-width: 600px) {{
            .field-name, .field-value {{
                width: 100%;
            }}
            .field {{
                flex-direction: column;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{_('Incident Report')} #{incident.id}</h1>
        </div>
        <div class="content">
            <div class="greeting">
                <p><strong>{_('Dear Recipient')},</strong></p>
                <p>{_('Please find below details of the incident report')}:</p>
            </div>
            
            <div class="section">
                <div class="section-title">{_('INCIDENT INFORMATION')}</div>
                <div class="field">
                    <div class="field-name">{_('ID')}:</div>
                    <div class="field-value"><strong>{incident.id}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Company')}:</div>
                    <div class="field-value"><strong>{incident.company.name}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Date & Time')}:</div>
                    <div class="field-value"><strong>{occurrence_date}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Place')}:</div>
                    <div class="field-value"><strong>{incident.place}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Classification')}:</div>
                    <div class="field-value"><span class="classification">{classification}</span></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Incident Type')}:</div>
                    <div class="field-value"><span class="incident-type">{incident_type}</span></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Current State')}:</div>
                    <div class="field-value"><span class="current-state">{current_state}</span></div>
                </div>
            </div>
"""

        # Add description section
        message_html += f"""
            <div class="section">
                <div class="section-title">{_('DESCRIPTION')}</div>
                <p>{incident.description}</p>
            </div>
"""

        # Add features if available
        if incident.features:
            message_html += f"""
            <div class="section">
                <div class="section-title">{_('FEATURES AND SIGNS')}</div>
                <p>{incident.features}</p>
            </div>
"""

        # Add impact if available
        if incident.impact:
            message_html += f"""
            <div class="section">
                <div class="section-title">{_('IMPACT')}</div>
                <p class="highlight">{incident.impact}</p>
            </div>
"""

        # Add measures taken if available
        if incident.measures_taken:
            message_html += f"""
            <div class="section">
                <div class="section-title">{_('MEASURES TAKEN')}</div>
                <p>{incident.measures_taken}</p>
            </div>
"""

        # Add additional measures if available
        if incident.additional_measures:
            message_html += f"""
            <div class="section">
                <div class="section-title">{_('ADDITIONAL MEASURES')}</div>
                <p>{incident.additional_measures}</p>
            </div>
"""

        # Add reported by information
        message_html += f"""
            <div class="section">
                <div class="section-title">{_('ADDITIONAL INFORMATION')}</div>
                <div class="field">
                    <div class="field-name">{_('Reported By')}:</div>
                    <div class="field-value"><strong>{incident.reported_by}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Reported Date')}:</div>
                    <div class="field-value"><strong>{reported_date}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Responsible for Resolution')}:</div>
                    <div class="field-value"><strong>{incident.responsible}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Registered By')}:</div>
                    <div class="field-value"><strong>{incident.registered_by if incident.registered_by else _('N/A')}</strong></div>
                </div>
                <div class="field">
                    <div class="field-name">{_('Registered Date')}:</div>
                    <div class="field-value"><strong>{registered_date}</strong></div>
                </div>
            </div>
"""

        # Add note about attachments if there are files
        has_files = bool(incident.file_incident) or incident.files.exists()
        if has_files:
            message_html += f"""
            <div class="note">
                <strong>{_('NOTE')}:</strong> {_('This email includes incident report files as attachments')}.
            </div>
"""

        # Add footer
        message_html += f"""
        </div>
        <div class="footer">
            <p>
                <strong>{_('Best regards')},</strong><br>
                {request.user.get_full_name() or request.user.username}
            </p>
        </div>
    </div>
</body>
</html>
"""

        # Also create a plain text version for email clients that don't support HTML
        message_plain = f"""{_('Dear Recipient')},

{_('Please find below details of the incident report')}:

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('INCIDENT INFORMATION')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{_('ID')}: {incident.id}
{_('Company')}: {incident.company.name}
{_('Date & Time')}: {occurrence_date}
{_('Place')}: {incident.place}
{_('Classification')}: {classification}
{_('Incident Type')}: {incident_type}
{_('Current State')}: {current_state}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('DESCRIPTION')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{incident.description}
"""
        # Add features if available
        if incident.features:
            message_plain += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('FEATURES AND SIGNS')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{incident.features}
"""

        # Add impact if available
        if incident.impact:
            message_plain += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('IMPACT')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{incident.impact}
"""

        # Add measures taken if available
        if incident.measures_taken:
            message_plain += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('MEASURES TAKEN')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{incident.measures_taken}
"""

        # Add additional measures if available
        if incident.additional_measures:
            message_plain += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('ADDITIONAL MEASURES')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{incident.additional_measures}
"""

        # Add reported by information
        message_plain += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('ADDITIONAL INFORMATION')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{_('Reported By')}: {incident.reported_by}
{_('Reported Date')}: {reported_date}
{_('Responsible for Resolution')}: {incident.responsible}
{_('Registered By')}: {incident.registered_by if incident.registered_by else _('N/A')}
{_('Registered Date')}: {registered_date}
"""

        # Add note about attachments if there are files
        if has_files:
            message_plain += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{_('NOTE')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{_('This email includes incident report files as attachments')}.
"""

        message_plain += f"""
{_('Best regards')},
{request.user.get_full_name() or request.user.username}
"""

        return JsonResponse({
            'success': True, 
            'subject': subject,
            'message': message_plain,
            'message_html': message_html,
            'has_files': has_files
        })
    
    except Incident.DoesNotExist:
        return JsonResponse({'success': False, 'message': _('Incident not found')})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_can_mail_incident
@require_POST
def send_incident_email(request):
    """Send email about incident to specified recipients."""
    incident_id = request.POST.get('incident_id')
    recipients = request.POST.get('recipients', '')
    subject = request.POST.get('subject', '')
    message = request.POST.get('message', '')
    message_html = request.POST.get('message_html', '')
    include_files = request.POST.get('include_files') == 'on'
    
    logger.info(f"Email request received - Incident ID: {incident_id}, Recipients: {recipients}")
    
    # Validate required fields
    if not incident_id or not recipients or not subject or not message:
        logger.warning(f"Missing required fields - ID: {bool(incident_id)}, Recipients: {bool(recipients)}, Subject: {bool(subject)}, Message: {bool(message)}")
        return JsonResponse({'success': False, 'message': _('All fields are required')})
    
    # Split recipients and validate email format
    recipient_list = [email.strip() for email in recipients.split(',') if email.strip()]
    if not recipient_list:
        logger.warning("No valid recipient emails found")
        return JsonResponse({'success': False, 'message': _('At least one valid recipient email is required')})
    
    try:
        # Get the incident
        incident = Incident.objects.get(id=incident_id)
        logger.info(f"Found incident #{incident.id} for company {incident.company.name}")
        
        # Check if user has permission to access this incident
        access = check_user_incident_access(request.user, incident.company)
        logger.info(f"User {request.user.username} access check: {access}")
        
        if not access['has_access']:
            logger.warning(f"User {request.user.username} does not have permission to access incident #{incident_id}")
            return JsonResponse({'success': False, 'message': _('You do not have permission to access this incident')})
        
        # Get active mail account
        mail_account = MailAccount.objects.filter(is_active=True).first()
        if not mail_account:
            logger.error("No active mail account found in the system")
            return JsonResponse({'success': False, 'message': _('No active mail account found')})
        
        logger.info(f"Using mail account: {mail_account.username} with server {mail_account.server.smtp_host}:{mail_account.server.smtp_port}")
        
        # Create the email message
        msg = MIMEMultipart('alternative')
        msg['From'] = mail_account.username
        msg['To'] = ', '.join(recipient_list)
        msg['Subject'] = subject
        
        # Add text body
        msg.attach(MIMEText(message, 'plain'))
        
        # Add HTML body if available
        if message_html:
            msg.attach(MIMEText(message_html, 'html'))
        
        attachment_count = 0
        # Add attachments if requested
        if include_files:
            # Add main incident file if it exists
            if incident.file_incident:
                try:
                    file_path = incident.file_incident.path
                    logger.info(f"Attaching main file: {file_path}")
                    with open(file_path, 'rb') as file:
                        file_data = file.read()
                        file_name = os.path.basename(file_path)
                        part = MIMEApplication(file_data, Name=file_name)
                        part['Content-Disposition'] = f'attachment; filename="{file_name}"'
                        msg.attach(part)
                        attachment_count += 1
                except Exception as e:
                    logger.error(f"Error attaching main file: {str(e)}")
            
            # Add additional files if they exist
            for incident_file in incident.files.all():
                try:
                    file_path = incident_file.file.path
                    logger.info(f"Attaching additional file: {file_path}")
                    with open(file_path, 'rb') as file:
                        file_data = file.read()
                        file_name = incident_file.filename
                        part = MIMEApplication(file_data, Name=file_name)
                        part['Content-Disposition'] = f'attachment; filename="{file_name}"'
                        msg.attach(part)
                        attachment_count += 1
                except Exception as e:
                    logger.error(f"Error attaching file {incident_file.id}: {str(e)}")
        
        logger.info(f"Email prepared with {attachment_count} attachments. Attempting to send...")
        
        # Send the email
        try:
            # Connect to the server directly using smtplib
            if mail_account.server.use_ssl:
                # Create SSL context without keyfile issues
                context = ssl.create_default_context()
                context.check_hostname = False
                context.verify_mode = ssl.CERT_NONE
                
                # Connect with SSL
                logger.info(f"Connecting with SSL to {mail_account.server.smtp_host}:{mail_account.server.smtp_port}")
                smtp = smtplib.SMTP_SSL(
                    host=mail_account.server.smtp_host,
                    port=mail_account.server.smtp_port,
                    context=context
                )
                logger.info("Connected to mail server using SSL")
            else:
                # Connect without SSL
                logger.info(f"Connecting without SSL to {mail_account.server.smtp_host}:{mail_account.server.smtp_port}")
                smtp = smtplib.SMTP(
                    host=mail_account.server.smtp_host,
                    port=mail_account.server.smtp_port
                )
                
                # Use TLS if needed
                if mail_account.server.use_tls:
                    logger.info("Starting TLS")
                    smtp.starttls()
                    logger.info("Connected to mail server using TLS")
            
            # Login and send
            logger.info(f"Logging in with username: {mail_account.username}")
            smtp.login(mail_account.username, mail_account.password)
            
            logger.info(f"Sending message to {recipient_list}")
            smtp.send_message(msg)
            smtp.quit()
            
            # Log the email sending
            logger.info(f"Email sent successfully for incident #{incident_id} to {recipients}")
            
            return JsonResponse({
                'success': True,
                'message': _('Email sent successfully')
            })
            
        except smtplib.SMTPAuthenticationError as e:
            error_msg = _('Authentication failed. Please check your mail account credentials.')
            logger.error(f"SMTP authentication error: {str(e)}")
            return JsonResponse({'success': False, 'message': error_msg})
        
        except Exception as e:
            error_msg = _('Failed to send email: {}').format(str(e))
            logger.error(f"Error sending email: {str(e)}")
            return JsonResponse({'success': False, 'message': error_msg})
    
    except Incident.DoesNotExist:
        logger.error(f"Incident not found: {incident_id}")
        return JsonResponse({'success': False, 'message': _('Incident not found')})
    
    except Exception as e:
        logger.error(f"Error in send_incident_email: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def test_mail_account(request):
    """Test if there's an active mail account available."""
    from django.http import JsonResponse
    
    try:
        mail_accounts = MailAccount.objects.all()
        active_accounts = MailAccount.objects.filter(is_active=True)
        
        result = {
            'success': True,
            'total_accounts': mail_accounts.count(),
            'active_accounts': active_accounts.count(),
            'details': []
        }
        
        for account in mail_accounts:
            result['details'].append({
                'id': account.id,
                'username': account.username,
                'is_active': account.is_active,
                'server_host': account.server.smtp_host if account.server else 'No server',
                'server_port': account.server.smtp_port if account.server else None,
                'use_ssl': account.server.use_ssl if account.server else None,
                'use_tls': account.server.use_tls if account.server else None
            })
        
        return JsonResponse(result)
    except Exception as e:
        logger.error(f"Error testing mail account: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def get_company_users(request):
    """Get all users belonging to the same company as the specified incident."""
    incident_id = request.GET.get('incident_id')
    if not incident_id:
        return JsonResponse({'success': False, 'message': _('Incident ID is required')})
    
    try:
        # Get the incident
        incident = Incident.objects.get(id=incident_id)
        
        # Check if user has permission to view this incident
        access = check_user_incident_access(request.user, incident.company)
        if not access['has_access']:
            return JsonResponse({'success': False, 'message': _('You do not have permission to view this incident')})
        
        # Get the company from the incident
        company = incident.company
        
        # Get all cabinet users from this company
        cabinet_users = CabinetUser.objects.filter(
            company=company,
            user__is_active=True
        ).select_related('user', 'department', 'position')
        
        # Get current language code
        language_code = request.LANGUAGE_CODE[:2]
        
        # Prepare the response data
        users_data = []
        for cabinet_user in cabinet_users:
            user = cabinet_user.user
            
            # Get department name based on language
            department_name = cabinet_user.department.get_name(language_code) if cabinet_user.department else None
            
            # Get position name based on language
            position_name = cabinet_user.position.get_name(language_code) if cabinet_user.position else None
            
            users_data.append({
                'id': user.id,
                'email': user.email,
                'full_name': f"{user.first_name} {user.last_name}".strip() or None,
                'position': position_name,
                'department': department_name
            })
        
        return JsonResponse({
            'success': True,
            'users': users_data
        })
    
    except Incident.DoesNotExist:
        return JsonResponse({'success': False, 'message': _('Incident not found')})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


 