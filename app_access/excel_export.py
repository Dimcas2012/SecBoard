"""
Excel export functionality for Access Requests with openpyxl styling and colors
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
from django.utils.translation import get_language, gettext as _
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Prefetch
from io import BytesIO
import datetime

from .models import (
    AccessRequest,
    AccessRequestApprover,
    AccessObjectIS,
    ObjectRoles,
    ObjectAccessRights,
    AccessObjectFunction,
    ObjectRoleFunctions,
)
from . import matrix_view
from app_conf.models import Company
from app_asset.models import InformationAsset


def _hex_fill(color):
    """Normalize color to 6-char hex for openpyxl PatternFill (no leading #)."""
    if not color:
        return "FFFFFF"
    s = (color or "").strip().lstrip("#")
    if len(s) == 3:
        s = "".join(c * 2 for c in s)
    return (s + "000000")[:6].upper()


@login_required
def export_access_requests_excel(request):
    """
    Export access requests to Excel with enhanced styling and colors
    Supports both selected records and all records export
    """
    # Get export parameters
    export_type = request.GET.get('export_type', 'all')  # 'selected' or 'all'
    selected_ids = request.GET.getlist('selected_ids[]')
    current_language = get_language()[:2]
    
    # Base query - same as in user_access_request view
    requests_query = AccessRequest.objects.filter(
        Q(requested_by=request.user) |
        Q(system__owners__cabinet_user__user=request.user) |
        Q(system__administrators__cabinet_user__user=request.user) |
        Q(request_approvers__cabinet_user__user=request.user)
    ).distinct().select_related(
        'company',
        'system',
        'requested_by',
        'requested_for',
        'cancelled_by'
    ).prefetch_related(
        'request_approvers__cabinet_user',
        'access_records',
        'access_records__roles',
        'access_records__access_object',
        'access_records__access_right',
        'third_party_users',
        'attachments'
    )
    
    # Apply filters from request
    if request.GET.get('company'):
        requests_query = requests_query.filter(company_id=request.GET.get('company'))
    
    if request.GET.get('system'):
        requests_query = requests_query.filter(system_id=request.GET.get('system'))
    
    if request.GET.get('status'):
        requests_query = requests_query.filter(status=request.GET.get('status'))
    
    if request.GET.get('admin_status'):
        requests_query = requests_query.filter(admin_status=request.GET.get('admin_status'))
    
    if request.GET.get('request_type'):
        requests_query = requests_query.filter(request_type=request.GET.get('request_type'))
    
    if request.GET.get('search'):
        search_query = request.GET.get('search')
        requests_query = requests_query.filter(
            Q(id__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(requested_for__username__icontains=search_query) |
            Q(justification__icontains=search_query)
        )
    
    # Filter by selected IDs if export_type is 'selected'
    if export_type == 'selected' and selected_ids:
        requests_query = requests_query.filter(id__in=selected_ids)
    
    # Order by creation date (newest first)
    requests = requests_query.order_by('-created_at')
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(_("Access Requests"))
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
    title_font = Font(name='Arial', bold=True, size=14)
    normal_font = Font(name='Arial', size=10)
    
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    # Status colors
    status_colors = {
        'pending': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'approved': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'rejected': PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),
        'cancelled': PatternFill(start_color="D6D8DB", end_color="D6D8DB", fill_type="solid"),
    }
    
    # Admin Status colors  
    # Color scheme: Yellow=Pending, Green=Granted, Red=Denied, Light Blue=In Progress
    admin_status_colors = {
        'pending': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),      # Yellow - Pending
        'granted': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),      # Green - Access Granted
        'denied': PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),       # Red - Access Denied
        'in_progress': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),  # Light Blue - In Progress
    }
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Alternating row colors for better readability
    even_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # Light gray
    
    # Column headers
    headers = [
        str(_("ID")),
        str(_("Request Type")),
        str(_("Company")),
        str(_("System")),
        str(_("Grant access records")),
        str(_("Revoke access records")),
        str(_("Requested For")),
        str(_("Requested By")),
        str(_("Status")),
        str(_("Admin Status")),
        str(_("Start Date")),
        str(_("End Date")),
        str(_("Created At")),
        str(_("Justification")),
        str(_("Notes")),
        str(_("Approvers")),
        str(_("Attachments"))
    ]
    
    # Add title
    export_type_text = str(_("Selected Records")) if export_type == 'selected' else str(_("All Records"))
    title = f"{str(_('Access Requests Export'))} - {export_type_text} - {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A1:Q1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    title_cell.fill = title_fill
    
    # Add records count
    count_text = f"{str(_('Total Records'))}: {requests.count()}"
    ws.merge_cells('A2:Q2')
    count_cell = ws.cell(row=2, column=1, value=count_text)
    count_cell.font = Font(name='Arial', bold=True, size=11)
    count_cell.alignment = left_alignment
    
    # Headers
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = [8, 12, 20, 25, 30, 30, 20, 20, 12, 15, 18, 18, 18, 40, 40, 30, 15]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Write data
    row_num = header_row + 1
    for idx, req in enumerate(requests):
        # Determine if this is an even row for alternating colors
        is_even_row = idx % 2 == 1
        row_fill = even_row_fill if is_even_row else PatternFill()
        
        # ID
        cell = ws.cell(row=row_num, column=1, value=req.id)
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Request Type
        request_type_display = dict(AccessRequest.REQUEST_TYPE_CHOICES).get(req.request_type, req.request_type)
        cell = ws.cell(row=row_num, column=2, value=str(request_type_display))
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Company
        cell = ws.cell(row=row_num, column=3, value=req.company.name if req.company else '-')
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # System
        cell = ws.cell(row=row_num, column=4, value=req.system.name if req.system else '-')
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Grant/Revoke access records summary (with per-record revoked marking for grants)
        grant_lines_info = []  # list of tuples (text, is_revoked)
        grant_records_text = '-'
        revoke_records_text = '-'
        try:
            if hasattr(req, 'access_records'):
                lines = []
                for record in req.access_records.all():
                    # Object name
                    obj_name = '-'
                    if getattr(record, 'access_object', None):
                        if current_language == 'uk':
                            obj_name = record.access_object.get_name() or record.access_object.name or '-'
                        elif current_language == 'ru':
                            obj_name = record.access_object.get_name() or record.access_object.name or '-'
                        else:
                            obj_name = record.access_object.get_name() or record.access_object.name or '-'

                    # Roles list
                    role_names = []
                    for role in record.roles.all():
                        if current_language == 'uk':
                            role_names.append(role.get_name() or role.name or '')
                        elif current_language == 'ru':
                            role_names.append(role.get_name() or role.name or '')
                        else:
                            role_names.append(role.get_name() or role.name or '')
                    roles_str = ', '.join([rn for rn in role_names if rn])

                    line = f"{obj_name}{': ' + roles_str if roles_str else ''}"
                    lines.append(line)

                    # Determine if this grant record was revoked (Admin Status granted)
                    is_revoked_for_grant = False
                    if req.request_type == 'grant':
                        try:
                            from .models import AccessRequestSequence as ARS
                            seq = ARS.objects.filter(access_record=record, grant_request=req).order_by('order_number').first()
                            if seq and seq.sequence_status == 'revoked' and getattr(seq, 'revoke_request', None) and getattr(seq.revoke_request, 'admin_status', None) == 'granted':
                                is_revoked_for_grant = True
                            elif not seq:
                                # Fallback by sequence_id prefix
                                seq2 = ARS.objects.filter(
                                    sequence_id__startswith=f"{record.id}.{req.id}.",
                                    sequence_status='revoked',
                                    revoke_request__admin_status='granted'
                                ).order_by('-revoked_at').first()
                                if seq2:
                                    is_revoked_for_grant = True
                        except Exception:
                            pass

                    grant_lines_info.append((line, is_revoked_for_grant))

                if req.request_type == 'grant':
                    grant_records_text = '\n'.join(lines) if lines else '-'
                elif req.request_type == 'revoke':
                    revoke_records_text = '\n'.join(lines) if lines else '-'
        except Exception:
            pass

        # Write Grant access records (col 5)
        cell = ws.cell(row=row_num, column=5)
        # Use Unicode combining long stroke overlay to visually strike revoked lines across Excel versions
        def strike_overlay(text: str) -> str:
            try:
                return ''.join(ch + '\u0336' for ch in text)
            except Exception:
                return text
        if grant_lines_info:
            rendered_lines = []
            for txt, is_rev in grant_lines_info:
                rendered_lines.append(strike_overlay(txt) if is_rev else txt)
            cell.value = '\n'.join(rendered_lines)
        else:
            cell.value = grant_records_text
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border

        # Write Revoke access records (col 6)
        cell = ws.cell(row=row_num, column=6, value=revoke_records_text)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border

        # Requested For
        requested_for_name = req.requested_for.get_full_name() or req.requested_for.username if req.requested_for else '-'
        cell = ws.cell(row=row_num, column=7, value=requested_for_name)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Requested By
        requested_by_name = req.requested_by.get_full_name() or req.requested_by.username if req.requested_by else '-'
        cell = ws.cell(row=row_num, column=8, value=requested_by_name)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Status
        status_display = dict(AccessRequest.STATUS_CHOICES).get(req.status, req.status)
        cell = ws.cell(row=row_num, column=9, value=str(status_display))
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = center_alignment
        cell.fill = status_colors.get(req.status, PatternFill())
        cell.border = thin_border
        
        # Admin Status
        admin_status_display = dict(AccessRequest.ADMIN_STATUS_CHOICES).get(req.admin_status, req.admin_status)
        cell = ws.cell(row=row_num, column=10, value=str(admin_status_display))
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = center_alignment
        cell.fill = admin_status_colors.get(req.admin_status, PatternFill())
        cell.border = thin_border
        
        # Start Date
        if req.start_date:
            start_date_naive = req.start_date.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=11, value=start_date_naive)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=11, value='-')
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # End Date
        if req.end_date:
            end_date_naive = req.end_date.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=12, value=end_date_naive)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=12, value='-')
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Created At
        if req.created_at:
            created_at_naive = req.created_at.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=13, value=created_at_naive)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=13, value='-')
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Justification
        cell = ws.cell(row=row_num, column=14, value=req.justification or '-')
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Notes
        cell = ws.cell(row=row_num, column=15, value=req.notes or '-')
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Approvers
        approvers_info = []
        if req.request_approvers.exists():
            for approver in req.request_approvers.all():
                status_text = dict(AccessRequestApprover.APPROVING_STATUS_CHOICES).get(approver.current_status, approver.current_status)
                approver_name = approver.cabinet_user.user.get_full_name() or approver.cabinet_user.user.username
                approvers_info.append(f"{approver_name} ({status_text})")
        
        approvers_text = '\n'.join(approvers_info) if approvers_info else '-'
        cell = ws.cell(row=row_num, column=16, value=approvers_text)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Attachments
        attachments_count = req.attachments.count()
        attachments_text = f"{attachments_count} {str(_('files'))}" if attachments_count > 0 else '-'
        cell = ws.cell(row=row_num, column=17, value=attachments_text)
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        row_num += 1
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    export_type_suffix = 'selected' if export_type == 'selected' else 'all'
    filename = f"access_requests_{export_type_suffix}_{timestamp}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


@login_required
def export_access_matrix_excel(request):
    """
    Export access matrix (default or object) to Excel with styles and colors.
    GET params: system_id (required), object_id (optional), environment (optional).
    """
    system_id = request.GET.get('system_id')
    if not system_id:
        return HttpResponse(_('Missing system_id'), status=400)
    try:
        system_id = int(system_id)
    except (TypeError, ValueError):
        return HttpResponse(_('Invalid system_id'), status=400)
    object_id = request.GET.get('object_id')
    if object_id:
        try:
            object_id = int(object_id)
        except (TypeError, ValueError):
            object_id = None
    environment = request.GET.get('environment') or None

    result = matrix_view.get_matrix_data_for_export(
        request, system_id, object_id=object_id, environment=environment
    )
    if result[0] is None:
        return JsonResponse(result[1], status=result[2])

    data, mappings = result
    wb = openpyxl.Workbook()
    ws = wb.active
    first_col_header = _('Object Roles & Functions') if data.get('is_object_matrix') else _('Roles & Functions')
    ws.title = _('Access Matrix')[:31]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    role_font = Font(name='Arial', bold=True, size=10)
    role_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    normal_font = Font(name='Arial', size=9)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    system_mapping_fill = PatternFill(start_color='FD7E14', end_color='FD7E14', fill_type='solid')
    checked_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

    access_rights = sorted(data['access_rights'], key=lambda r: r.get('order', 0))
    num_rights = len(access_rights)
    roles = data.get('roles') or []

    row = 1
    title = _('Default Access Rights Matrix')
    if data.get('is_object_matrix') and data.get('object_name'):
        title = _('Object Access Rights Matrix') + ' - ' + data['object_name']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_rights + 1)
    ws.cell(row=1, column=1, value=title).font = Font(name='Arial', bold=True, size=12)
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    row = 2

    cell0 = ws.cell(row=row, column=1, value=first_col_header)
    cell0.font = header_font
    cell0.fill = header_fill
    cell0.alignment = left_align
    cell0.border = thin_border
    for c, right in enumerate(access_rights, 2):
        cell = ws.cell(row=row, column=c, value=right['name'])
        cell.font = header_font
        right_fill = PatternFill(start_color=_hex_fill(right.get('color')), end_color=_hex_fill(right.get('color')), fill_type='solid')
        cell.fill = right_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    mappings_set = {(m['role_id'], m['function_id'], m['right_id']) for m in (mappings or [])}
    system_mappings = {(sm['role_id'], sm['function_id'], sm['right_id']) for sm in (data.get('system_mappings') or [])}

    right_ids = [r['id'] for r in access_rights]

    for role in roles:
        role_id = role['id']
        role_color = _hex_fill(role.get('color'))
        role_fill_cell = PatternFill(start_color=role_color, end_color=role_color, fill_type='solid')
        cell0 = ws.cell(row=row, column=1, value=role['name'])
        cell0.font = role_font
        cell0.fill = role_fill_cell
        cell0.alignment = left_align
        cell0.border = thin_border
        for col_idx in range(2, num_rights + 2):
            ws.cell(row=row, column=col_idx, value='').fill = role_fill
            ws.cell(row=row, column=col_idx).border = thin_border
        row += 1

        funcs = sorted(role['columns']['functions'], key=lambda x: x.get('order', 0))
        subs = role['columns']['subfunctions']
        subsubs = role['columns']['subsubfunctions']
        for func in funcs:
            _write_function_row(ws, row, 1, func, role_id, right_ids, mappings_set, system_mappings,
                               data.get('is_object_matrix'), thin_border, normal_font, left_align,
                               center_align, checked_fill, system_mapping_fill, _hex_fill)
            row += 1
            for sub in sorted([s for s in subs if s.get('parent_id') == func['id']], key=lambda x: x.get('order', 0)):
                _write_function_row(ws, row, 2, sub, role_id, right_ids, mappings_set, system_mappings,
                                   data.get('is_object_matrix'), thin_border, normal_font, left_align,
                                   center_align, checked_fill, system_mapping_fill, _hex_fill)
                row += 1
                for subsub in sorted([s for s in subsubs if s.get('parent_id') == sub['id']], key=lambda x: x.get('order', 0)):
                    _write_function_row(ws, row, 3, subsub, role_id, right_ids, mappings_set, system_mappings,
                                       data.get('is_object_matrix'), thin_border, normal_font, left_align,
                                       center_align, checked_fill, system_mapping_fill, _hex_fill)
                    row += 1

    ws.column_dimensions['A'].width = 40
    for i in range(2, num_rights + 2):
        ws.column_dimensions[get_column_letter(i)].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"access_matrix_{system_id}"
    if object_id:
        filename += f"_object_{object_id}"
    filename += f"_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_objects_excel(request):
    """
    Export access objects (Object Management) to Excel for the selected system and environment.
    GET params: system_id (required), environment (required).
    """
    system_id = request.GET.get('system_id')
    environment = request.GET.get('environment')
    if not system_id:
        return HttpResponse(_('Missing system_id'), status=400)
    if not environment:
        return HttpResponse(_('Missing environment'), status=400)
    try:
        system_id = int(system_id)
    except (TypeError, ValueError):
        return HttpResponse(_('Invalid system_id'), status=400)

    user_companies = matrix_view.get_user_companies_for_config_is(request.user)
    try:
        system = InformationAsset.objects.get(id=system_id)
    except InformationAsset.DoesNotExist:
        return HttpResponse(_('Information system not found'), status=404)
    if user_companies and not user_companies.filter(id=system.company_id).exists():
        return HttpResponse(_('Access denied to this information system'), status=403)

    objects_list = AccessObjectIS.objects.filter(
        asset_id=system_id,
        environment=environment
    ).prefetch_related(
        Prefetch(
            'object_roles',
            queryset=ObjectRoles.objects.filter(is_active=True).select_related('role').prefetch_related(
                Prefetch('role_functions', queryset=ObjectRoleFunctions.objects.filter(is_active=True).select_related('function'))
            ),
        ),
        Prefetch(
            'object_access_rights',
            queryset=ObjectAccessRights.objects.filter(is_active=True).select_related('access_right'),
        ),
        Prefetch(
            'object_functions',
            queryset=AccessObjectFunction.objects.filter(is_active=True).select_related('function'),
        ),
    ).order_by('tree_id', 'lft')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    normal_font = Font(name='Arial', size=9)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(_("Objects"))[:31]

    headers = [
        _('Name'),
        _('Code'),
        _('Description'),
        _('Parent'),
        _('Level'),
        _('Is Active'),
        _('Order'),
        _('Color'),
        _('Object Roles'),
        _('Object Access Rights'),
        _('Object Functions'),
        _('Object Roles & Functions'),
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align
        cell.border = thin_border

    for row_idx, obj in enumerate(objects_list, 2):
        level = obj.get_ancestors().count() + 1
        parent_name = (obj.parent.name or obj.parent.name_local or '') if obj.parent else ''
        ws.cell(row=row_idx, column=1, value=obj.name or obj.name_local or '')
        ws.cell(row=row_idx, column=2, value=obj.code or '')
        ws.cell(row=row_idx, column=3, value=obj.description or '')
        ws.cell(row=row_idx, column=4, value=parent_name)
        ws.cell(row=row_idx, column=5, value=level)
        ws.cell(row=row_idx, column=6, value=_('Yes') if obj.is_active else _('No'))
        ws.cell(row=row_idx, column=7, value=obj.order or 0)
        ws.cell(row=row_idx, column=8, value=obj.color or '')

        obj_roles_str = ', '.join(
            (or_.role.name or or_.role.name_local or '')
            for or_ in obj.object_roles.all()
        )
        obj_rights_str = ', '.join(
            (oar.access_right.name or oar.access_right.name_local or '')
            for oar in obj.object_access_rights.all()
        )
        obj_funcs_str = ', '.join(
            (aof.function.name or aof.function.name_local or '')
            for aof in obj.object_functions.all()
        )
        roles_functions_parts = []
        for or_ in obj.object_roles.all():
            role_name = or_.role.name or or_.role.name_local or ''
            func_names = [
                (rf.function.name or rf.function.name_local or '')
                for rf in or_.role_functions.all()
            ]
            if func_names:
                roles_functions_parts.append(f"{role_name}: {', '.join(func_names)}")
            elif role_name:
                roles_functions_parts.append(role_name)
        obj_roles_functions_str = '; '.join(roles_functions_parts)

        ws.cell(row=row_idx, column=9, value=obj_roles_str)
        ws.cell(row=row_idx, column=10, value=obj_rights_str)
        ws.cell(row=row_idx, column=11, value=obj_funcs_str)
        ws.cell(row=row_idx, column=12, value=obj_roles_functions_str)

        for col in range(1, 13):
            c = ws.cell(row=row_idx, column=col)
            c.font = normal_font
            c.border = thin_border
            c.alignment = left_align

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 35
    for col in range(5, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 40
    ws.column_dimensions['L'].width = 50

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"objects_{system_id}_{environment}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _write_function_row(ws, row, level, func, role_id, right_ids, mappings_set, system_mappings,
                        is_object_matrix, thin_border, normal_font, left_align, center_align,
                        checked_fill, system_mapping_fill, hex_fill_fn):
    indent = '    ' * level
    name = (indent + (func.get('name') or ''))
    func_color = hex_fill_fn(func.get('color'))
    fill_first = PatternFill(start_color=func_color, end_color=func_color, fill_type='solid')
    cell0 = ws.cell(row=row, column=1, value=name)
    cell0.font = normal_font
    cell0.fill = fill_first
    cell0.alignment = left_align
    cell0.border = thin_border
    func_id = func['id']
    for col_idx, right_id in enumerate(right_ids, 2):
        cell = ws.cell(row=row, column=col_idx)
        cell.border = thin_border
        is_mapped = (role_id, func_id, right_id) in mappings_set
        is_system = is_object_matrix and ((role_id, func_id, right_id) in system_mappings)
        cell.value = '✓' if is_mapped else ''
        cell.alignment = center_align
        if is_mapped:
            cell.fill = system_mapping_fill if is_system else checked_fill
        cell.font = normal_font