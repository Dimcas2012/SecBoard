from app_access.models import AccessRequest, AccessRequestAttachment, AccessRequestApprover, AccessRequestAdminStatusHistory, EmailNotificationHistory, ThirdPartyUser, ThirdPartyOrganization, SystemAccessStatusHistory, AccessJustificationTemplate
from .matrix_view import (has_access_manage_ar_permission, can_add_manage_ar, 
                         can_edit_manage_ar, can_delete_manage_ar, 
                         get_user_companies_for_manage_ar)
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.http import JsonResponse
from django.utils.translation import get_language, activate, gettext as _
from django.utils import translation
from app_conf.models import Company, Country
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
import json
from django.views.decorators.http import require_http_methods, require_POST
from .models import SystemAccess, AccessRoles, AccessApprover, AccessStatus, UserAccessRequestGuide, UserAccessRequestGuideTranslation
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.db.models import Q, Prefetch
from django.db import transaction
from django.conf import settings
from django.db import models
import logging
from app_asset.models import InformationAsset
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
import os
from django.core.files.storage import default_storage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .email_utils import send_access_request_notification, send_access_request_status_notification
from .access_is_view import check_and_update_expired_records
from .pagination_utils import (
    ACCESS_TABLE_DEFAULT_PAGE_SIZE,
    ACCESS_TABLE_PAGE_SIZE_OPTIONS,
    get_access_table_page_size,
)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
import datetime

logger = logging.getLogger(__name__)
logger.setLevel(logging.WARNING)


def _serialize_access_justification_templates(company_id=None):
    templates_query = AccessJustificationTemplate.objects.filter(is_active=True)
    if company_id:
        templates_query = templates_query.filter(
            Q(company_id=company_id) | Q(company__isnull=True)
        )
    else:
        templates_query = templates_query.filter(company__isnull=True)

    return [
        {
            'id': tpl.id,
            'name': tpl.get_name(),
            'content': tpl.get_content(),
        }
        for tpl in templates_query
        .prefetch_related('translations__country')
        .order_by('sort_order', 'name')
    ]


def parse_client_datetime(value):
    """
    Parse datetime strings from HTML datetime-local or ISO payloads.
    - Supports values without seconds (YYYY-MM-DDTHH:MM)
    - Makes naive datetimes timezone-aware in current timezone
    """
    if not value:
        return None

    value_str = str(value).strip()
    if not value_str:
        return None

    if len(value_str) == 16 and value_str.count(':') == 1:
        value_str += ':00'

    dt = parse_datetime(value_str)
    if not dt:
        return None

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def is_cabinet_employee_access_request(access_request):
    """Grant/revoke for an internal employee; requested_for is not the submitter placeholder."""
    return bool(
        access_request.requested_for_id
        and access_request.requested_by_id
        and access_request.requested_for_id != access_request.requested_by_id
    )


def is_third_party_access_request(access_request):
    """
    True when the grant/revoke targets third-party users, not the cabinet user in requested_for.
    Cabinet requests may have stale third_party_users M2M or legacy fields; those must not
    override requested_for on the list page.
    """
    if is_cabinet_employee_access_request(access_request):
        return False
    if access_request.third_party_users.exists():
        return True
    has_legacy_third_party = bool(
        (access_request.third_party_first_name or '').strip()
        or (access_request.third_party_last_name or '').strip()
        or (access_request.third_party_email or '').strip()
    )
    if not has_legacy_third_party:
        return False
    return access_request.requested_for_id == access_request.requested_by_id


def get_effective_requested_for_summary(access_request):
    """Human-readable beneficiary shown on list/admin (aligned with user-access-request UI)."""
    if is_third_party_access_request(access_request):
        if access_request.third_party_users.exists():
            parts = []
            for tp in access_request.third_party_users.all():
                label = tp.full_name or f'{tp.first_name} {tp.last_name}'.strip() or tp.email
                org = tp.organization_name or (tp.organization.name if tp.organization else '')
                if org:
                    label = f'{label} ({org})'
                parts.append(label)
            return '; '.join(parts)
        name = f'{access_request.third_party_first_name or ""} {access_request.third_party_last_name or ""}'.strip()
        org = (access_request.third_party_organization or '').strip()
        if name and org:
            return f'{name} ({org})'
        return name or access_request.third_party_email or str(_('Third party'))
    if (
        access_request.requested_for_users_data
        and (access_request.requested_for_count or 1) > 1
    ):
        names = [
            u.get('full_name') or u.get('email') or ''
            for u in access_request.requested_for_users_data
            if isinstance(u, dict)
        ]
        return '; '.join(n for n in names if n)
    if access_request.requested_for:
        user = access_request.requested_for
        return user.get_full_name() or user.email or user.username
    return ''


def _access_request_display_context(req):
    """Fields for user_access_request template (dict rows, not model instances)."""
    is_tp = is_third_party_access_request(req)
    return {
        'show_third_party_display': is_tp,
        'is_third_party_request': is_tp,
        'third_party_users_list': list(req.third_party_users.all()) if is_tp else [],
        'requested_for_users_data': (
            req.requested_for_users_data or []
            if (req.requested_for_count or 1) > 1 and req.requested_for_users_data
            else []
        ),
        'effective_requested_for_summary': get_effective_requested_for_summary(req),
    }


@login_required
@ensure_csrf_cookie
def user_access_request(request):
    """Відображення сторінки запиту доступу"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        messages.error(request, _("Access denied - you are not authorized to submit access requests. Please contact your administrator to be added to the request list for at least one system."))
        return redirect('index')
    
    # Автоматично перевіряємо та оновлюємо записи з минулими датами
    expired_count = check_and_update_expired_records()
    if expired_count > 0 and settings.DEBUG is True and False:
        # Suppressed noisy info log even in DEBUG
        logger.info(f"Automatically set {expired_count} expired records to inactive before loading user access request page")
    
    current_language = get_language()[:2]

    # Базовий запит - показувати запити де користувач є: Requested By, Owner, Administrator, або Approving Person
    requests_query = AccessRequest.objects.filter(
        Q(requested_by=request.user) |  # Запити, які подав користувач
        Q(system__owners__cabinet_user__user=request.user) |  # Запити для систем, де користувач є власником
        Q(system__administrators__cabinet_user__user=request.user) |  # Запити для систем, де користувач є адміністратором
        Q(request_approvers__cabinet_user__user=request.user)  # Запити, які користувач має затверждувати
    ).distinct().select_related(
        'company',
        'system',
        'requested_by',
        'requested_for',
        'requested_for__cabinet',
        'requested_for__cabinet__department',
        'requested_for__cabinet__position'
    ).prefetch_related(
        'request_approvers__cabinet_user',
        'request_approvers__cabinet_user__user',
        'request_approvers__cabinet_user__department',
        'request_approvers__cabinet_user__position',
        'request_approvers__status_history',
        'system__owners__cabinet_user',
        'system__owners__cabinet_user__user',
        'system__owners__cabinet_user__department',
        'system__owners__cabinet_user__position',
        'system__administrators__cabinet_user',
        'system__administrators__cabinet_user__user',
        'system__administrators__cabinet_user__department',
        'system__administrators__cabinet_user__position',
        'attachments',
        Prefetch('admin_status_history', 
                 queryset=AccessRequestAdminStatusHistory.objects.select_related('changed_by').order_by('-changed_at')),
        Prefetch('email_notifications',
                 queryset=EmailNotificationHistory.objects.select_related('triggered_by', 'mail_account').order_by('-created_at')),
        'third_party_users',
        'access_records',
        'access_records__roles',
        'access_records__access_object'
    )
    
    # Застосовуємо фільтри
    if request.GET.get('company'):
        requests_query = requests_query.filter(company_id=request.GET.get('company'))
    
    if request.GET.get('system'):
        requests_query = requests_query.filter(system_id=request.GET.get('system'))
    
    if request.GET.get('object'):
        requests_query = requests_query.filter(access_records__access_object_id=request.GET.get('object'))
    
    if request.GET.get('status'):
        requests_query = requests_query.filter(status=request.GET.get('status'))
    
    if request.GET.get('environment'):
        requests_query = requests_query.filter(environment=request.GET.get('environment'))
    
    if request.GET.get('role'):
        role_id = request.GET.get('role')
        requests_query = requests_query.filter(access_records__roles__id=role_id)
    
    if request.GET.get('search'):
        search_query = request.GET.get('search')
        requests_query = requests_query.filter(
            Q(id__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(requested_for__username__icontains=search_query) |
            Q(justification__icontains=search_query) |
            Q(requirements__icontains=search_query) |
            Q(notes__icontains=search_query) |
            Q(third_party_first_name__icontains=search_query) |
            Q(third_party_last_name__icontains=search_query) |
            Q(third_party_email__icontains=search_query) |
            Q(third_party_organization__icontains=search_query) |
            Q(third_party_users__first_name__icontains=search_query) |
            Q(third_party_users__last_name__icontains=search_query) |
            Q(third_party_users__email__icontains=search_query) |
            Q(third_party_users__organization__icontains=search_query)
        )
    
    if request.GET.get('period'):
        now = timezone.now()
        if request.GET.get('period') == 'active':
            requests_query = requests_query.filter(
                Q(start_date__lte=now) & 
                (Q(end_date__gt=now) | Q(end_date__isnull=True))
            )
        elif request.GET.get('period') == 'expired':
            requests_query = requests_query.filter(end_date__lt=now)
        elif request.GET.get('period') == 'future':
            requests_query = requests_query.filter(start_date__gt=now)
    
    # Фільтрація за admin_status
    if request.GET.get('admin_status'):
        requests_query = requests_query.filter(admin_status=request.GET.get('admin_status'))
    
    # Фільтрація за request_type
    if request.GET.get('request_type'):
        requests_query = requests_query.filter(request_type=request.GET.get('request_type'))
    
    # Фільтрація за датою створення
    if request.GET.get('date'):
        try:
            date = timezone.datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()
            requests_query = requests_query.filter(created_at__date=date)
        except ValueError:
            logger.warning(f"Invalid date format: {request.GET.get('date')}")
    
    # Сортуємо результати
    requests = requests_query.order_by('-created_at')

    # Пагінація (за замовчуванням 25 записів на сторінку)
    page_size = get_access_table_page_size(request)
    page = request.GET.get('page', 1)
    paginator = Paginator(requests, page_size)
    try:
        requests_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        requests_page = paginator.page(1)

    # Відображення коду мови для моделі
    lang_code = 'ua' if current_language == 'uk' else current_language

    # Підготовка даних для відображення
    formatted_requests = []
    for req in requests_page:
        try:
            # Backfill request approvers if missing
            if not req.request_approvers.exists():
                # Вибираємо джерело затверджувачів: з першого access_record або з системних
                source_approvers = []
                try:
                    source_record = req.access_records.first()
                    if source_record:
                        source_approvers = list(source_record.approvers.all().order_by('order'))
                except Exception:
                    source_approvers = []

                if not source_approvers:
                    try:
                        source_approvers = list(req.system.approving_persons.all().order_by('order'))
                    except Exception:
                        source_approvers = []

                if source_approvers:
                    from .models import AccessRequestApprover, AccessApprover, ApprovingPerson
                    created_any = False
                    for sa in source_approvers:
                        # Підтримка двох типів джерел: AccessApprover або ApprovingPerson
                        if isinstance(sa, AccessApprover):
                            cabinet_user = sa.cabinet_user
                            order = sa.order
                        else:  # ApprovingPerson
                            cabinet_user = sa.cabinet_user
                            order = sa.order

                        try:
                            AccessRequestApprover.objects.create(
                                access_request=req,
                                access_approver=sa if isinstance(sa, AccessApprover) else None,
                                cabinet_user=cabinet_user,
                                order=order,
                                current_status='approved' if req.status == 'approved' else 'pending'
                            )
                            created_any = True
                        except Exception:
                            # Ignore duplicates or creation issues; continue best-effort
                            pass
                    if created_any:
                        # Refresh relation for subsequent usage
                        req.refresh_from_db()
        except Exception:
            # Never break page due to backfill issues
            pass
        # Визначаємо чи може користувач адмініструвати цей запит
        user_can_admin = False
        if hasattr(request.user, 'cabinet') and request.user.cabinet:
            user_can_admin = req.system.administrators.filter(cabinet_user=request.user.cabinet).exists()
        
        # Отримуємо структуровані дані про ролі та об'єкти з усіх записів доступу
        access_records_data = []
        roles = []
        if req.access_records.exists():
            # Підготуємо можливий оригінальний grant-запит для revoke
            original_request = None
            if req.request_type == 'revoke':
                if req.notes:
                    import re
                    match = re.search(r'request #(\d+)', req.notes)
                    if match:
                        try:
                            original_request = AccessRequest.objects.get(id=int(match.group(1)))
                        except AccessRequest.DoesNotExist:
                            original_request = None
                # Fallback for granular revoke: original grant ID from revoked_grant_access_record_ids (format A.B.C = access_record.grant_request.order)
                if original_request is None:
                    revoked_ids = getattr(req, 'revoked_grant_access_record_ids', None) or []
                    if revoked_ids:
                        try:
                            first_id = str(revoked_ids[0])
                            parts = first_id.split('.')
                            if len(parts) >= 2 and parts[1].isdigit():
                                grant_request_id = int(parts[1])
                                original_request = AccessRequest.objects.filter(
                                    id=grant_request_id,
                                    request_type='grant',
                                    status='approved',
                                    admin_status='granted'
                                ).first()
                        except (IndexError, ValueError, TypeError):
                            pass

            requested_role_by_record = {}
            if getattr(req, 'requested_access_record_roles', None) and isinstance(req.requested_access_record_roles, list):
                for item in req.requested_access_record_roles:
                    if isinstance(item, dict) and 'access_record_id' in item and 'role_id' in item:
                        requested_role_by_record[item['access_record_id']] = item['role_id']
            # For Revoke: show the requested role from the original grant (the role being revoked)
            if req.request_type == 'revoke' and original_request is not None and getattr(original_request, 'requested_access_record_roles', None) and isinstance(original_request.requested_access_record_roles, list):
                for item in original_request.requested_access_record_roles:
                    if isinstance(item, dict) and 'access_record_id' in item and 'role_id' in item:
                        requested_role_by_record[item['access_record_id']] = item['role_id']

            # Збираємо структуровані дані з усіх записів доступу
            for access_record in req.access_records.all():
                record_roles = list(access_record.roles.all())
                roles.extend(record_roles)
                requested_role = None
                rid = requested_role_by_record.get(access_record.id)
                if rid:
                    try:
                        requested_role = AccessRoles.objects.get(id=rid)
                    except AccessRoles.DoesNotExist:
                        pass
                
                                # Перевіряємо чи є цей Access Record revoked
                is_revoked = False
                revoke_info = None
                grant_access_record_id = None

                # (moved) Grant revocation check based on full ID happens after we resolve grant_access_record_id below
                
                # Визначаємо Grant Access Record ID для відображення в списку "My Access Requests"
                try:
                    from .models import AccessRequestSequence
                    sequence_record = None
                    grant_access_record_id_full = None
                    if req.request_type == 'revoke':
                        # Спершу пробуємо зв'язок за revoke_request
                        sequence_record = AccessRequestSequence.objects.filter(
                            access_record=access_record,
                            revoke_request=req
                        ).order_by('order_number').first()
                        # Якщо не знайдено, пробуємо за оригінальним grant-запитом
                        if not sequence_record and original_request is not None:
                            sequence_record = AccessRequestSequence.objects.filter(
                                access_record=access_record,
                                grant_request=original_request
                            ).order_by('order_number').first()
                        # Додатковий fallback: шукаємо серед відкликаних послідовностей із суфіксом .<revoke_id>
                        if not sequence_record:
                            sequence_record = AccessRequestSequence.objects.filter(
                                access_record=access_record,
                                sequence_status='revoked',
                                sequence_id__endswith=f".{req.id}"
                            ).order_by('order_number').first()
                        # Додатковий fallback: використовуємо список IDs із самого Revoke запиту
                        # Знаходимо A.B.C для конкретного access_record за префіксом A = access_record.id
                        if not sequence_record:
                            try:
                                revoked_ids = getattr(req, 'revoked_grant_access_record_ids', None) or []
                                matched_id = None
                                for rid in revoked_ids:
                                    rid_str = str(rid)
                                    if rid_str.startswith(f"{access_record.id}."):
                                        matched_id = rid_str
                                        break
                                if matched_id:
                                    # Синтезуємо псевдо sequence_record-дані для відображення іконки "i"
                                    grant_access_record_id = '.'.join(matched_id.split('.')[:3])
                                    grant_access_record_id_full = f"{grant_access_record_id}.{req.id}"
                            except Exception:
                                pass
                    elif req.request_type == 'grant':
                        sequence_record = AccessRequestSequence.objects.filter(
                            access_record=access_record,
                            grant_request=req
                        ).order_by('order_number').first()
                        
                        # Debug logging for sequence lookup
                        sequence_count = AccessRequestSequence.objects.filter(
                            access_record=access_record,
                            grant_request=req
                        ).count()
                        #
                        
                        # Якщо з якоїсь причини прямий зв'язок не знайдено (після повторних циклів grant/revoke),
                        # шукаємо по префіксу A.B.* для цього grant запиту
                        if not sequence_record:
                            from .models import AccessRequestSequence as ARS
                            sequence_record = ARS.objects.filter(
                                sequence_id__startswith=f"{access_record.id}.{req.id}.",
                                grant_request=req
                            ).order_by('order_number').first()
                            
                            if sequence_record:
                                logger.debug(f"Found sequence by prefix search: {sequence_record.sequence_id}")
                            else:
                                # Auto-create missing AccessRequestSequence for approved grant requests
                                if req.status == 'approved' and req.admin_status == 'granted':
                                    try:
                                        from .models import AccessRequestSequence
                                        from django.db import models
                                        # Знаходимо максимальний порядковий номер для цього grant request
                                        max_order = AccessRequestSequence.objects.filter(
                                            grant_request=req
                                        ).aggregate(models.Max('order_number'))['order_number__max'] or 0
                                        order_number = max_order + 1
                                        
                                        sequence_record = AccessRequestSequence.objects.create(
                                            grant_request=req,
                                            access_record=access_record,
                                            order_number=order_number,
                                            sequence_status='active'
                                        )
                                        logger.debug(
                                            f"Auto-created missing AccessRequestSequence {access_record.id}.{req.id}.{order_number} for Grant Request {req.id}, Access Record {access_record.id}"
                                        )
                                    except Exception as e:
                                        logger.error(
                                            f"Failed to auto-create AccessRequestSequence for Grant Request {req.id}, Access Record {access_record.id}: {str(e)}"
                                        )

                    if sequence_record:
                        # Для відображення нормалізуємо до перших трьох частин
                        seq_id_str = str(sequence_record.sequence_id)
                        grant_access_record_id = '.'.join(seq_id_str.split('.')[:3])
                        # Повний ID для відображення: додаємо D-частину (0 якщо активний, або ID Revoke)
                        if sequence_record.sequence_status == 'revoked' and sequence_record.revoke_request:
                            grant_access_record_id_full = f"{grant_access_record_id}.{sequence_record.revoke_request.id}"
                            # Іконку "revoked" виставляємо лише для рядків типу Grant (не для Revoke)
                            if req.request_type == 'grant' and sequence_record.revoke_request.admin_status == 'granted':
                                is_revoked = True
                                revoke_request = sequence_record.revoke_request
                                revoked_for_correct = None
                                if revoke_request.third_party_email:
                                    revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                elif revoke_request.requested_for:
                                    revoked_for_correct = revoke_request.requested_for.get_full_name()
                                revoke_info = {
                                    'revoked_at': getattr(sequence_record, 'revoked_at', None) or revoke_request.created_at,
                                    'revoked_by': revoke_request.requested_by.get_full_name() if revoke_request.requested_by else 'System',
                                    'revoked_for': revoked_for_correct,
                                    'revoke_request_id': revoke_request.id,
                                    'change_reason': 'Access revoked for this grant sequence'
                                }
                        else:
                            grant_access_record_id_full = f"{grant_access_record_id}.0"
                except Exception as e:
                    logger.error(f"Failed to resolve Grant Access Record ID for request {req.id}, record {access_record.id}: {str(e)}")

                # Перевірка по послідовності для рядків Grant: базуємось на повному Grant Access Record ID (A.B.C.D)
                if not is_revoked and req.request_type == 'grant' and grant_access_record_id:
                    try:
                        from .models import AccessRequestSequence as ARS
                        seq_rev = ARS.objects.filter(
                            sequence_id__startswith=f"{grant_access_record_id}.",
                            sequence_status='revoked',
                            revoke_request__isnull=False,
                            revoke_request__admin_status='granted'
                        ).order_by('-revoked_at').first()
                        if seq_rev:
                            revoke_request = seq_rev.revoke_request
                            
                            # Перевіряємо, чи revoke стосується того ж користувача, що і grant
                            grant_user_email = None
                            revoke_user_email = None
                            
                            if req.third_party_email:
                                grant_user_email = req.third_party_email
                            elif req.requested_for:
                                grant_user_email = req.requested_for.email
                            
                            if revoke_request.third_party_email:
                                revoke_user_email = revoke_request.third_party_email
                            elif revoke_request.requested_for:
                                revoke_user_email = revoke_request.requested_for.email
                            
                            # Тільки позначаємо як revoked, якщо це стосується того ж користувача
                            if grant_user_email and revoke_user_email and grant_user_email == revoke_user_email:
                                is_revoked = True
                                revoked_for_correct = None
                                if revoke_request.third_party_email:
                                    revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                elif revoke_request.requested_for:
                                    revoked_for_correct = revoke_request.requested_for.get_full_name()
                                revoke_info = {
                                    'revoked_at': getattr(seq_rev, 'revoked_at', None) or revoke_request.created_at,
                                    'revoked_by': revoke_request.requested_by.get_full_name() if revoke_request.requested_by else 'System',
                                    'revoked_for': revoked_for_correct,
                                    'revoke_request_id': revoke_request.id,
                                    'change_reason': 'Access revoked for this grant sequence'
                                }
                        else:
                            # Фолбек: якщо немає revoke_request (дані старих записів), але послідовність має статус revoked
                            seq_rev2 = ARS.objects.filter(
                                sequence_id__startswith=f"{grant_access_record_id}.",
                                sequence_status='revoked'
                            ).order_by('-revoked_at').first()
                            if seq_rev2:
                                is_revoked = True
                                revoke_info = {
                                    'revoked_at': getattr(seq_rev2, 'revoked_at', None),
                                    'revoked_by': 'System',
                                    'revoked_for': None,
                                    'revoke_request_id': getattr(seq_rev2.revoke_request, 'id', None),
                                    'change_reason': 'Access revoked (no linked request)'
                                }
                            else:
                                # Додатковий фолбек: шукаємо Revoke запити, які містять цей Grant Access Record ID у списку
                                existing_rr = AccessRequest.objects.filter(
                                    request_type='revoke',
                                    admin_status='granted',
                                    revoked_grant_access_record_ids__contains=[grant_access_record_id]
                                ).order_by('-created_at').first()
                                if existing_rr:
                                    revoke_request = existing_rr
                                    
                                    # Перевіряємо, чи revoke стосується того ж користувача, що і grant
                                    grant_user_email = None
                                    revoke_user_email = None
                                    
                                    if req.third_party_email:
                                        grant_user_email = req.third_party_email
                                    elif req.requested_for:
                                        grant_user_email = req.requested_for.email
                                    
                                    if revoke_request.third_party_email:
                                        revoke_user_email = revoke_request.third_party_email
                                    elif revoke_request.requested_for:
                                        revoke_user_email = revoke_request.requested_for.email
                                    
                                    # Тільки позначаємо як revoked, якщо це стосується того ж користувача
                                    if grant_user_email and revoke_user_email and grant_user_email == revoke_user_email:
                                        is_revoked = True
                                        revoked_for_correct = None
                                        if revoke_request.third_party_email:
                                            revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                        elif revoke_request.requested_for:
                                            revoked_for_correct = revoke_request.requested_for.get_full_name()
                                        revoke_info = {
                                            'revoked_at': revoke_request.created_at,
                                            'revoked_by': revoke_request.requested_by.get_full_name() if revoke_request.requested_by else 'System',
                                            'revoked_for': revoked_for_correct,
                                            'revoke_request_id': revoke_request.id,
                                            'change_reason': 'Access revoked for this grant sequence'
                                        }
                    except Exception as e:
                        logger.error(f"Sequence revoke check by full ID failed for request {req.id}, record {access_record.id}: {str(e)}")
                
                # Debug logging removed

                access_records_data.append({
                    'id': access_record.id,
                    'object_id': access_record.access_object.id if access_record.access_object else None,
                    'object_name': access_record.access_object.get_name(current_language) if access_record.access_object else _('No Object'),
                    'object_color': access_record.access_object.color if access_record.access_object else '#6c757d',
                    'roles': record_roles,
                    'requested_role': requested_role,
                    'environment': access_record.environment,
                    'start_date': access_record.start_date,
                    'end_date': access_record.end_date,
                    'is_revoked': is_revoked,
                    'revoke_info': revoke_info,
                    'grant_access_record_id': grant_access_record_id,
                    'grant_access_record_id_full': grant_access_record_id_full or (f"{grant_access_record_id}.0" if grant_access_record_id else None)
                })
        # Fallback logic removed - all records should use access_records now
            
        # Для revoke запитів отримуємо інформацію про оригінальний запит
        original_request_id = None
        original_request = None
        if req.request_type == 'revoke' and req.notes:
            # Витягуємо ID оригінального запиту з notes
            import re
            match = re.search(r'request #(\d+)', req.notes)
            if match:
                original_request_id = int(match.group(1))
                try:
                    original_request = AccessRequest.objects.get(id=original_request_id)
                except AccessRequest.DoesNotExist:
                    original_request = None
            
        # Обробляємо revocation timing для revoke запитів
        revocation_timing = None
        is_immediate_revocation = False
        if req.request_type == 'revoke':
            # Перевіряємо чи це негайне скасування
            if req.notes and 'Revocation timing: Immediate' in req.notes:
                is_immediate_revocation = True
                revocation_timing = _('Immediately')
            elif req.start_date:
                # Запланована дата скасування
                if req.end_date:
                    revocation_timing = f"{timezone.localtime(req.start_date).strftime('%d.%m.%Y %H:%M')} - {timezone.localtime(req.end_date).strftime('%d.%m.%Y %H:%M')}"
                else:
                    revocation_timing = f"{timezone.localtime(req.start_date).strftime('%d.%m.%Y %H:%M')} - {_('Permanent')}"
            else:
                revocation_timing = _('Immediately')
                is_immediate_revocation = True

        # Отримуємо існуючі granted ролі користувача для цього Object та Environment
        existing_granted_roles = []
        # TODO: Fix existing granted roles logic for ManyToMany access_records
        # Temporarily disabled to prevent errors

        # Sort access records for Grant: keep revoked at the bottom
        if req.request_type == 'grant':
            try:
                access_records_data.sort(key=lambda item: item.get('is_revoked', False))
            except Exception:
                pass

        # Побудова короткого тексту історії погоджень для тултіпа статусу запиту
        approval_history_text = None
        try:
            events = []
            # 1) Події з поточних request_approvers
            if req.request_approvers.exists():
                for ra in req.request_approvers.all().order_by('order'):
                    history_qs = getattr(ra, 'status_history', None)
                    if history_qs is not None:
                        for h in history_qs.all().order_by('changed_at'):
                            approver_name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                            changer_name = h.changed_by.get_full_name() if h.changed_by else ''
                            ts = timezone.localtime(h.changed_at).strftime('%d.%m.%Y %H:%M')
                            events.append((h.changed_at, ra.order, f"Lvl {ra.order}: {approver_name} → {h.new_status.title()} ({ts} by {changer_name})"))
                    else:
                        approver_name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                        events.append((None, ra.order, f"Lvl {ra.order}: {approver_name} - {ra.current_status.title()}"))

            # 2) Snapshot-історія, прив'язана безпосередньо до запиту (на випадок заміни/видалення approver'ів)
            try:
                from .models import AccessRequestApproverStatusHistory as ARASH
                snapshot_qs = ARASH.objects.filter(access_request=req).order_by('changed_at')
                for h in snapshot_qs:
                    approver_name = h.approver_name or (h.approver_cabinet_user.user.get_full_name() if h.approver_cabinet_user and getattr(h.approver_cabinet_user, 'user', None) else '')
                    changer_name = h.changed_by.get_full_name() if h.changed_by else ''
                    ts = timezone.localtime(h.changed_at).strftime('%d.%m.%Y %H:%M')
                    order_num = getattr(h, 'order_at_change', None) or 0
                    events.append((h.changed_at, order_num, f"Lvl {order_num}: {approver_name} → {h.new_status.title()} ({ts} by {changer_name})"))
            except Exception:
                pass

            if events:
                events.sort(key=lambda x: (x[0] is None, x[0] or 0, x[1] or 0))
                approval_history_text = " | ".join([e[2] for e in events])
        except Exception:
            approval_history_text = None

        # Форматуємо дані
        formatted_requests.append({
            'id': req.id,
            'request_type': req.request_type,
            'company_id': req.company_id,
            'company_name': req.company.name,
            'system_id': req.system_id,
            'system_name': req.system.name,
            'object_id': req.access_records.first().access_object.id if req.access_records.exists() and req.access_records.first().access_object else None,
            'object_name': req.access_records.first().access_object.get_name(current_language) if req.access_records.exists() and req.access_records.first().access_object else _('No Object'),
            'environment': req.environment,
            'requested_by': req.requested_by,
            'requested_for': req.requested_for,
            'approvers': req.request_approvers.all() if hasattr(req, 'request_approvers') else [],
            'owners': req.system.owners.all() if hasattr(req.system, 'owners') else [],
            'administrators': req.system.administrators.all() if hasattr(req.system, 'administrators') else [],
            'system_approvers': req.system.approving_persons.all() if hasattr(req.system, 'approving_persons') else [],
            'roles': roles,
            'access_records_data': access_records_data,
            'access_records_count': len(access_records_data),
            'justification': req.justification,
            'requirements': req.requirements,
            'notes': req.notes,
            'attachments': req.attachments,
            'created_at': req.created_at,
            'record_start_date': req.start_date,
            'record_end_date': req.end_date,
            'record_progress': calculate_progress(req.start_date, req.end_date),
            'status': req.status,
            'admin_status': req.admin_status,
            'approval_history_text': approval_history_text,
            'admin_status_history': req.admin_status_history,
            'can_be_cancelled': req.can_be_cancelled(request.user),
            'can_be_edited': req.can_be_edited(request.user),
            'is_cancelled': req.is_cancelled,
            'cancelled_at': req.cancelled_at,
            'cancelled_by': req.cancelled_by,
            'cancellation_reason': req.cancellation_reason,
            'email_notifications': req.email_notifications.all().order_by('-created_at')[:5],  # Останні 5 email повідомлень
            # Додаємо поля третьої сторони
            'third_party_first_name': req.third_party_first_name,
            'third_party_last_name': req.third_party_last_name,
            'third_party_email': req.third_party_email,
            'third_party_phone': req.third_party_phone,
            'third_party_organization': req.third_party_organization,
            'third_party_description': req.third_party_description,
            # Новий зв'язок з ThirdPartyUser
            'third_party_users': req.third_party_users.all() if hasattr(req, 'third_party_users') else [],
            'third_party_users_count': req.third_party_users.count() if hasattr(req, 'third_party_users') else 0,
            # Додаємо інформацію про кількість записів доступу
            'access_records_count': req.access_records.count() if hasattr(req, 'access_records') else 0,
            # Додаємо інформацію про оригінальний запит для revoke запитів
            'original_request_id': original_request_id,
            'original_request': original_request,
            # Додаємо інформацію про revocation timing
            'revocation_timing': revocation_timing,
            'is_immediate_revocation': is_immediate_revocation,
            # Права користувача для цього запиту
            'user_can_admin': user_can_admin,
            # Додаємо інформацію про існуючі granted ролі
            'existing_granted_roles': existing_granted_roles,
            **_access_request_display_context(req),

        })

    # Отримуємо список компаній та систем для фільтрів, доступних користувачу
    user_groups = request.user.groups.all()
    
    # Отримуємо тільки компанії, в яких є системи доступні користувачу
    companies = Company.objects.filter(
        informationasset__access_records__is_active=True
    ).filter(
        Q(informationasset__access_records__request_users=request.user) |
        Q(informationasset__access_records__request_groups__in=user_groups)
    ).distinct().order_by('name')
    
    systems = []
    objects = []
    if request.GET.get('company'):
        # Отримуємо тільки системи доступні користувачу в вибраній компанії
        systems = InformationAsset.objects.filter(
            company_id=request.GET.get('company'),
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().order_by('name')
    
    if request.GET.get('system'):
        from .models import AccessObjectIS
        # Отримуємо тільки об'єкти доступні користувачу в вибраній системі
        objects = AccessObjectIS.objects.filter(
            asset_id=request.GET.get('system'),
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().order_by('order', 'name')
    
    # Отримуємо ID ролей з запитів доступу користувача
    role_ids = AccessRequest.objects.filter(
        requested_by=request.user
    ).values_list('access_records__roles__id', flat=True).distinct()
    
    # Отримуємо ролі за ID
    roles = AccessRoles.objects.filter(
        id__in=role_ids, is_active=True
    ).order_by('order', 'name', 'code')
    
    # Отримуємо список організацій для dropdown (JSON-serializable)
    third_party_organizations = list(
        ThirdPartyOrganization.objects.filter(is_active=True)
        .order_by('name')
        .values('id', 'name')
    )
    selected_company_id = request.GET.get('company')
    access_justification_templates = _serialize_access_justification_templates(
        company_id=selected_company_id
    )
    
    return render(request, 'app_access/user_access_request.html', {
        'requests': formatted_requests,
        'companies': companies,
        'systems': systems,
        'objects': objects,
        'roles': roles,
        'current_language': current_language,
        'third_party_organizations': third_party_organizations,
        'access_justification_templates': access_justification_templates,
        # Пагінація
        'paginator': paginator,
        'page_obj': requests_page,
        'is_paginated': paginator.count > 0,
        'current_page_size': page_size,
        'page_size_options': ACCESS_TABLE_PAGE_SIZE_OPTIONS,
    })


@login_required
@require_http_methods(["GET"])
def user_access_request_guide(request):
    """Return JSON { content: html } for the User Access Request guide (localized)."""
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = UserAccessRequestGuide.objects.first()
    if guide:
        if country:
            trans = UserAccessRequestGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = UserAccessRequestGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def user_access_request_guide_translate(request):
    """API for AI translation of User Access Request guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_access_records(request, system_id, object_id, user_id):
    """Отримання доступних записів доступу для користувача"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({
            'success': False,
            'error': _("Access denied - you are not authorized to view access records.")
        }, status=403)
    """Отримання доступних записів доступу для користувача"""
    try:
        # Автоматично перевіряємо та оновлюємо записи з минулими датами
        expired_count = check_and_update_expired_records()
        if expired_count > 0:
            logger.info(f"Automatically set {expired_count} expired records to inactive before loading available records")
        
        current_language = get_language()[:2]
        user_groups = request.user.groups.all()
        
        # Отримуємо environment з GET параметрів
        environment = request.GET.get('environment')
        
        # Базова фільтрація записів
        records = SystemAccess.objects.filter(
            asset_id=system_id,
            access_object_id=object_id,
            is_active=True
        ).filter(
            Q(end_date__gt=timezone.now()) |
            Q(end_date__isnull=True)
        ).filter(
            # Тільки ті записи, до яких поточний користувач може запитувати доступ
            Q(request_users=request.user) |
            Q(request_groups__in=user_groups)
        )
        
        # Фільтруємо по environment, якщо параметр передано
        if environment:
            logger.info(f"Filtering records by environment: {environment}")
            records = records.filter(environment=environment)
        else:
            logger.info("Loading all access records without environment filter")
        
        # Перевіряємо чи це третя сторона
        if user_id == 'third_parties':
            # Для третьої сторони показуємо записи з увімкненим third_parties
            records = records.filter(third_parties=True)
        else:
            # Для звичайних користувачів фільтруємо за доступом користувача
            # Конвертуємо user_id в int, якщо це не 'third_parties'
            try:
                user_id_int = int(user_id)
                records = records.filter(
                    Q(access_users__id=user_id_int) |
                    Q(access_groups__user__id=user_id_int)
                )
            except (ValueError, TypeError):
                logger.error(f"Invalid user_id format: {user_id}")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid user ID format'
                }, status=400)
        
        records = records.select_related(
            'asset',
            'access_object'
        ).prefetch_related(
            'roles',
            'asset__owners__cabinet_user__user',
            'asset__owners__cabinet_user__department',
            'asset__owners__cabinet_user__position',
            'asset__administrators__cabinet_user__user',
            'asset__administrators__cabinet_user__department',
            'asset__administrators__cabinet_user__position',
            'approvers__cabinet_user__user',
            'approvers__cabinet_user__department',
            'approvers__cabinet_user__position'
        ).distinct()

        processed_ids = set()
        formatted_records = []

        for record in records:
            if record.id in processed_ids:
                continue

            processed_ids.add(record.id)

            # Форматуємо власників
            owners = []
            for owner in record.asset.owners.all():
                cabinet_user = owner.cabinet_user
                owners.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None
                })

            # Форматуємо адміністраторів
            administrators = []
            for admin in record.asset.administrators.all():
                cabinet_user = admin.cabinet_user
                administrators.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None
                })

            # Форматуємо затверджувачів
            approvers = []
            for approver in record.approvers.all().order_by('order'):
                cabinet_user = approver.cabinet_user
                approvers.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None,
                    'order': approver.order
                })

            formatted_records.append({
                'id': record.id,
                'asset_id': system_id,  # Додаємо system_id як asset_id
                'system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо system_name
                'information_system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо information_system_name
                'object_id': record.access_object.id if record.access_object else None,  # Додаємо object_id
                'object_name': record.access_object.get_name(current_language) if record.access_object else 'Unknown Object',  # Додаємо object_name
                'object_color': record.access_object.color if record.access_object and hasattr(record.access_object, 'color') else '#6c757d',  # Додаємо object_color
                'environment': record.environment,  # Додаємо environment
                'roles': [{
                    'id': role.id,
                    'name': role.get_name(current_language),
                    'color': role.color or '#6c757d',
                    'description': role.get_description(current_language) if hasattr(role, 'get_description') else None
                } for role in record.roles.all()],
                'start_date': record.start_date.isoformat() if record.start_date else None,
                'end_date': record.end_date.isoformat() if record.end_date else None,
                'owners': owners,
                'administrators': administrators,
                'approvers': approvers
            })

        return JsonResponse({
            'status': 'success',
            'access_records': formatted_records
        })
    except Exception as e:
        logger.error(f"Error getting access records: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def get_access_records_by_system(request, system_id, user_id):
    """Отримання доступних записів доступу для вибраної системи та користувача (без фільтрації по об'єкту)"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({
            'success': False,
            'error': _("Access denied - you are not authorized to view access records.")
        }, status=403)
    
    try:
        # Автоматично перевіряємо та оновлюємо записи з минулими датами
        expired_count = check_and_update_expired_records()
        if expired_count > 0:
            logger.info(f"Automatically set {expired_count} expired records to inactive before loading available records")
        
        # Отримуємо мову з параметрів запиту або використовуємо поточну мову
        current_language = request.GET.get('language', get_language())[:2]
        user_groups = request.user.groups.all()
        
        # Базова фільтрація записів по системі (без об'єкта)
        records = SystemAccess.objects.filter(
            asset_id=system_id,
            is_active=True
        ).filter(
            Q(end_date__gt=timezone.now()) |
            Q(end_date__isnull=True)
        ).filter(
            # Тільки ті записи, до яких поточний користувач може запитувати доступ
            Q(request_users=request.user) |
            Q(request_groups__in=user_groups)
        )
        
        # Перевіряємо чи це третя сторона
        if user_id == 'third_parties':
            # Для третьої сторони показуємо записи з увімкненим third_parties
            records = records.filter(third_parties=True)
        else:
            # Для звичайних користувачів фільтруємо за доступом користувача
            # Конвертуємо user_id в int, якщо це не 'third_parties'
            try:
                user_id_int = int(user_id)
                records = records.filter(
                    Q(access_users__id=user_id_int) |
                    Q(access_groups__user__id=user_id_int)
                )
            except (ValueError, TypeError):
                logger.error(f"Invalid user_id format: {user_id}")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid user ID format'
                }, status=400)
        
        records = records.select_related(
            'asset',
            'access_object'
        ).prefetch_related(
            'roles',
            'asset__owners__cabinet_user__user',
            'asset__owners__cabinet_user__department',
            'asset__owners__cabinet_user__position',
            'asset__administrators__cabinet_user__user',
            'asset__administrators__cabinet_user__department',
            'asset__administrators__cabinet_user__position',
            'approvers__cabinet_user__user',
            'approvers__cabinet_user__department',
            'approvers__cabinet_user__position'
        ).distinct()

        processed_ids = set()
        formatted_records = []

        for record in records:
            if record.id in processed_ids:
                continue

            processed_ids.add(record.id)

            # Форматуємо власників
            owners = []
            for owner in record.asset.owners.all():
                cabinet_user = owner.cabinet_user
                owners.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None
                })

            # Форматуємо адміністраторів
            administrators = []
            for admin in record.asset.administrators.all():
                cabinet_user = admin.cabinet_user
                administrators.append({
                    'name': cabinet_user.user.get_full_name(),
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': admin.cabinet_user.position.get_name(current_language) if admin.cabinet_user.position else None
                })

            # Форматуємо затверджувачів
            approvers = []
            for approver in record.approvers.all().order_by('order'):
                cabinet_user = approver.cabinet_user
                approvers.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None,
                    'order': approver.order
                })

            formatted_records.append({
                'id': record.id,
                'asset_id': system_id,  # Додаємо system_id як asset_id
                'system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо system_name
                'information_system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо information_system_name
                'object_id': record.access_object.id if record.access_object else None,  # Додаємо object_id
                'object_name': record.access_object.get_name(current_language) if record.access_object else 'Unknown Object',  # Додаємо object_name
                'object_color': record.access_object.color if record.access_object and hasattr(record.access_object, 'color') else '#6c757d',  # Додаємо object_color
                'environment': record.environment,  # Додаємо environment
                'roles': [{
                    'id': role.id,
                    'name': role.get_name(current_language),
                    'color': role.color or '#6c757d',
                    'description': role.get_description(current_language) if hasattr(role, 'get_description') else None
                } for role in record.roles.all()],
                'start_date': record.start_date.isoformat() if record.start_date else None,
                'end_date': record.end_date.isoformat() if record.end_date else None,
                'owners': owners,
                'administrators': administrators,
                'approvers': approvers
            })

        return JsonResponse({
            'status': 'success',
            'access_records': formatted_records
        })
    except Exception as e:
        logger.error(f"Error getting access records by system: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


def handle_revoke_request(request, data, attachments):
    """Обробка Revoke запитів"""
    try:
        from django.utils import translation
        
        # Перевіряємо чи це новий granular revoke запит чи старий формат
        selected_access_records = data.get('selected_access_records')
        #
        
        if selected_access_records:
            # Новий формат - granular revoke з вибраними access records
            logger.debug("Using granular revoke request handler")
            return handle_granular_revoke_request(request, data, attachments)
        
        # Старий формат - відкликання цілого запиту
        # Отримуємо ID оригінального запиту для відкликання
        original_request_id = data.get('original_request_id')
        if not original_request_id:
            return JsonResponse({
                'success': False,
                'message': _('Original request ID or selected access records are required for revoke requests')
            }, status=400)
        
        # Перевіряємо що оригінальний запит існує і належить користувачу
        try:
            original_request = AccessRequest.objects.get(
                id=original_request_id,
                requested_by=request.user,
                request_type='grant',
                status='approved',
                admin_status='granted'
            )
        except AccessRequest.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': _('Original request not found or not accessible')
            }, status=404)
        
        # Перевіряємо чи немає вже активного revoke запиту для цього grant запиту
        # Використовуємо більш точну перевірку по ключових полях
        existing_revoke = AccessRequest.objects.filter(
            request_type='revoke',
            requested_by=request.user,
            requested_for_id=original_request.requested_for_id,
            access_record_id=original_request.access_record_id,
            system_id=original_request.system_id,
            environment=original_request.environment,
            status__in=['pending', 'approved'],
            notes__contains=f'request #{original_request_id}'
        ).first()
        
        if existing_revoke:
            # Формуємо детальне повідомлення про існуючий запит
            status_text = _('pending') if existing_revoke.status == 'pending' else _('approved')
            return JsonResponse({
                'success': False,
                'message': _('A revoke request for this access is already {status} (Request #{request_id}). Only one revoke request per grant request is allowed.').format(
                    status=status_text,
                    request_id=existing_revoke.id
                )
            }, status=400)
        
        # Обробляємо дані про час скасування
        revoke_immediately = data.get('revoke_immediately', 'true').lower() == 'true'
        revocation_start_date = None
        revocation_end_date = None
        
        if not revoke_immediately:
            # Якщо не негайно, обробляємо дати скасування
            if data.get('revocation_start_date'):
                revocation_start_date = parse_client_datetime(data.get('revocation_start_date'))
            
            if data.get('revocation_end_date'):
                revocation_end_date = parse_client_datetime(data.get('revocation_end_date'))
        
        # Формуємо додаткові примітки про час скасування
        timing_notes = ""
        if revoke_immediately:
            timing_notes = "Revocation timing: Immediate upon approval. "
        else:
            timing_notes = "Revocation timing: Scheduled. "
            if revocation_start_date:
                timing_notes += f"Start: {revocation_start_date}. "
            else:
                timing_notes += "Start: Immediate upon approval. "
            
            if revocation_end_date:
                timing_notes += f"End: {revocation_end_date}. "
            else:
                timing_notes += "End: Permanent revocation. "
        
        # Встановлюємо дати для revoke запиту
        from django.utils import timezone
        
        # Для revoke запитів start_date завжди має бути встановлена
        if revoke_immediately:
            # Для негайного скасування використовуємо поточну дату як start_date
            request_start_date = timezone.now()
            request_end_date = None  # Негайне скасування не має кінцевої дати
        else:
            # Для запланованого скасування
            request_start_date = revocation_start_date or timezone.now()
            request_end_date = revocation_end_date
        
        # Створюємо Revoke запит (без записів у notes про "request #...")
        revoke_request = AccessRequest.objects.create(
            request_type='revoke',
            company_id=data.get('company_id') or original_request.company_id,
            system_id=data.get('system_id') or original_request.system_id,
            environment=data.get('environment') or original_request.environment,
            requested_by=request.user,
            requested_for_id=original_request.requested_for_id,
            access_record_id=data.get('access_record_id') or original_request.access_record_id,
            justification=data.get('justification', ''),
            requirements=data.get('requirements', ''),
            notes=data.get('notes', ''),
            # Використовуємо поля start_date та end_date для зберігання часу скасування
            start_date=request_start_date,
            end_date=request_end_date,
            # Копіюємо дані третьої сторони якщо є
            third_party_first_name=original_request.third_party_first_name,
            third_party_last_name=original_request.third_party_last_name,
            third_party_email=original_request.third_party_email,
            third_party_phone=original_request.third_party_phone,
            third_party_organization=original_request.third_party_organization,
            third_party_description=original_request.third_party_description,
            third_party_count=original_request.third_party_count,
            third_party_users_data=original_request.third_party_users_data
        )
        
        # Копіюємо зв'язки з ThirdPartyUser якщо є
        if original_request.third_party_users.exists():
            revoke_request.third_party_users.set(original_request.third_party_users.all())
        
        # Обробка файлів
        if attachments:
            max_file_size = 10 * 1024 * 1024  # 10MB
            allowed_types = [
                'application/pdf',
                'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'text/plain',
                'image/jpeg',
                'image/png',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel'
            ]
            
            for attachment in attachments:
                # Перевіряємо розмір файлу
                if attachment.size > max_file_size:
                    raise ValueError(f"File {attachment.name} exceeds maximum size of 10MB")
                
                # Перевіряємо тип файлу
                if attachment.content_type not in allowed_types:
                    raise ValueError(f"File type {attachment.content_type} is not allowed")
                
                # Зберігаємо файл
                AccessRequestAttachment.objects.create(
                    access_request=revoke_request,
                    file=attachment,
                    original_filename=attachment.name,
                    file_size=attachment.size,
                    content_type=attachment.content_type,
                    uploaded_by=request.user
                )
        
        # Створюємо AccessRequestApprover записи базуючись на оригінальному запиті
        if original_request.access_record:
            for approver in original_request.access_record.approvers.all():
                AccessRequestApprover.objects.create(
                    access_request=revoke_request,
                    access_approver=approver,
                    cabinet_user=approver.cabinet_user,
                    order=approver.order,
                    current_status='pending'
                )
        
            logger.debug(f"Created revoke request {revoke_request.id} for original request {original_request_id} by user {request.user.username}")
        
        # Відправляємо email повідомлення
        try:
            send_access_request_notification(revoke_request, recipients_type='all')
            logger.debug(f"Email notifications sent for revoke request {revoke_request.id}")
        except Exception as e:
            logger.error(f"Failed to send email notifications for revoke request {revoke_request.id}: {e}")
        
        return JsonResponse({
            'success': True,
            'message': _('Revoke access request submitted successfully'),
            'request_id': revoke_request.id,
            'original_request_id': original_request_id
        })
        
    except Exception as e:
        logger.error(f"Error handling revoke request: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


def handle_granular_revoke_request(request, data, attachments):
    """Обробка granular revoke запитів з вибраними access records"""
    try:
        import json
        from django.utils import timezone
        
        #
        
        # Парсимо вибрані записи доступу
        selected_records_json = data.get('selected_access_records')
        #
        
        if isinstance(selected_records_json, str):
            selected_records = json.loads(selected_records_json)
        else:
            selected_records = selected_records_json
        
        if not selected_records or len(selected_records) == 0:
            return JsonResponse({
                'success': False,
                'message': _('At least one access record must be selected for revocation')
            }, status=400)
        
        # Отримуємо дані форми
        company_id = data.get('company_id')
        system_id = data.get('system_id')
        environment = data.get('environment')
        access_justification = data.get('access_justification', '')
        additional_requirements = data.get('additional_requirements', '')
        notes = data.get('notes', '')
        revocation_timing = data.get('revocation_timing', 'immediately')
        
        if not all([company_id, system_id, environment, access_justification]):
            return JsonResponse({
                'success': False,
                'message': _('Company, System, Environment, and Access Justification are required')
            }, status=400)
        
        # Обробляємо дати
        start_date = timezone.now()
        end_date = None
        
        if revocation_timing == 'scheduled':
            if data.get('start_date'):
                start_date = parse_client_datetime(data.get('start_date')) or timezone.now()
            if data.get('end_date'):
                end_date = parse_client_datetime(data.get('end_date'))
        
        # Групуємо записи по оригінальним запитам
        records_by_request = {}
        for record_info in selected_records:
            request_id = record_info['request_id']
            record_id = record_info['record_id']
            
            if request_id not in records_by_request:
                records_by_request[request_id] = []
            records_by_request[request_id].append(record_id)
        
        created_revoke_requests = []
        
        with transaction.atomic():
            for original_request_id, record_ids in records_by_request.items():
                # Перевіряємо оригінальний запит
                try:
                    original_request = AccessRequest.objects.get(
                        id=original_request_id,
                        status='approved',
                        admin_status='granted',
                        request_type='grant'
                    )
                except AccessRequest.DoesNotExist:
                    logger.warning(f"Original request {original_request_id} not found or not accessible")
                    continue
                
                # Отримуємо access records для цього запиту
                selected_access_records = original_request.access_records.filter(
                    id__in=record_ids
                )
                
                if not selected_access_records.exists():
                    logger.warning(f"No valid access records found for request {original_request_id}")
                    continue
                
                    logger.debug(f"Found {selected_access_records.count()} access records for request {original_request_id}")
                    logger.debug(f"Record IDs: {list(selected_access_records.values_list('id', flat=True))}")
                
                # Формуємо notes з інформацією про вибрані записи
                records_info = []
                revoked_grant_access_record_ids = []
                
                for record_info in selected_records:
                    if record_info['request_id'] != int(original_request_id):
                        continue  # Skip records not belonging to this request
                        
                    record_id = record_info['record_id']
                    record = selected_access_records.filter(id=record_id).first()
                    if not record:
                        continue
                        
                    object_name = record.access_object.get_name('en') if record.access_object else 'Default Object'
                    roles = [role.get_name('en') or role.name or '' for role in record.roles.all()]
                    records_info.append(f"{object_name}: {', '.join(roles)}")
                    
                    # Використовуємо конкретний Grant Access Record ID, який передав фронтенд
                    frontend_grant_id = record_info.get('grant_access_record_id')
                    if frontend_grant_id:
                        revoked_grant_access_record_ids.append(frontend_grant_id)
                        logger.debug(f"Using frontend-provided Grant Access Record ID: {frontend_grant_id}")
                    else:
                        # Fallback: знаходимо конкретні Grant Access Record ID для цього запису
                        from .models import AccessRequestSequence
                        sequences = AccessRequestSequence.objects.filter(
                            access_record=record,
                            grant_request=original_request,
                            sequence_status='active'
                        )
                        # Initialize to avoid UnboundLocalError when no sequences found
                        grant_access_record_id = None
                        for sequence in sequences:
                            grant_access_record_id = f"{sequence.access_record.id}.{sequence.grant_request.id}.{sequence.order_number}"
                            revoked_grant_access_record_ids.append(grant_access_record_id)
                        if grant_access_record_id:
                            logger.debug(f"Fallback: found Grant Access Record ID: {grant_access_record_id}")
                
                detailed_notes = notes or ''
                
                # Створюємо revoke запит
                revoke_request = AccessRequest.objects.create(
                    request_type='revoke',
                    company_id=company_id,
                    system_id=system_id,
                    environment=environment,
                    requested_by=request.user,
                    requested_for=original_request.requested_for,
                    start_date=start_date,
                    end_date=end_date,
                    justification=access_justification,
                    requirements=additional_requirements,
                    notes=detailed_notes,
                    # Зберігаємо конкретні Grant Access Record ID, які скасовуються
                    revoked_grant_access_record_ids=revoked_grant_access_record_ids,
                    # Копіюємо дані третьої сторони якщо є
                    third_party_first_name=original_request.third_party_first_name,
                    third_party_last_name=original_request.third_party_last_name,
                    third_party_email=original_request.third_party_email,
                    third_party_phone=original_request.third_party_phone,
                    third_party_organization=original_request.third_party_organization,
                    third_party_description=original_request.third_party_description,
                    third_party_count=original_request.third_party_count,
                    third_party_users_data=original_request.third_party_users_data
                )
                
                # Прив'язуємо тільки вибрані access records
                revoke_request.access_records.set(selected_access_records)
                
                # Копіюємо зв'язки з ThirdPartyUser якщо є
                if original_request.third_party_users.exists():
                    revoke_request.third_party_users.set(original_request.third_party_users.all())
                
                # Створюємо AccessRequestApprover записи (уникаємо дублювання)
                created_approvers = set()  # Відстежуємо вже створених approvers
                
                for selected_record in selected_access_records:
                    for approver in selected_record.approvers.all():
                        # Створюємо унікальний ключ для approver
                        approver_key = (approver.cabinet_user.id, approver.order)
                        
                        if approver_key not in created_approvers:
                            AccessRequestApprover.objects.get_or_create(
                                access_request=revoke_request,
                                cabinet_user=approver.cabinet_user,
                                defaults={
                                    'access_approver': approver,
                                    'order': approver.order,
                                    'current_status': 'pending'
                                }
                            )
                            created_approvers.add(approver_key)
                            logger.info(f"Created approver for user {approver.cabinet_user.id} with order {approver.order}")
                
                # Обробка файлів
                if attachments:
                    max_file_size = 10 * 1024 * 1024  # 10MB
                    allowed_types = [
                        'application/pdf',
                        'application/msword',
                        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                        'text/plain',
                        'image/jpeg',
                        'image/png',
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        'application/vnd.ms-excel'
                    ]
                    
                    for attachment in attachments:
                        # Перевіряємо розмір файлу
                        if attachment.size > max_file_size:
                            raise ValueError(f"File {attachment.name} exceeds maximum size of 10MB")
                        
                        # Перевіряємо тип файлу
                        if attachment.content_type not in allowed_types:
                            raise ValueError(f"File type {attachment.content_type} is not allowed")
                        
                        # Зберігаємо файл
                        AccessRequestAttachment.objects.create(
                            access_request=revoke_request,
                            file=attachment,
                            original_filename=attachment.name,
                            file_size=attachment.size,
                            content_type=attachment.content_type,
                            uploaded_by=request.user
                        )
                
                created_revoke_requests.append(revoke_request)
                logger.info(f"Created granular revoke request {revoke_request.id} for {len(selected_access_records)} access records from original request {original_request_id}")
        
        if not created_revoke_requests:
            return JsonResponse({
                'success': False,
                'message': _('No valid revoke requests could be created')
            }, status=400)
        
        # Відправляємо повідомлення для першого створеного запиту (можна розширити для всіх)
        try:
            from .email_utils import send_access_request_notification
            for revoke_request in created_revoke_requests:
                send_access_request_notification(revoke_request, 'submitted')
        except Exception as e:
            logger.error(f"Error sending notification for revoke request: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': _('Revoke request(s) submitted successfully'),
            'created_requests': len(created_revoke_requests),
            'request_ids': [req.id for req in created_revoke_requests]
        })
        
    except Exception as e:
        logger.error(f"Error handling granular revoke request: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def process_approved_grant_request(grant_request):
    """
    Обробка схваленого grant запиту - додавання користувачів до access_users
    """
    try:
        from django.utils import timezone
        
        logger.info(f"Processing approved grant request {grant_request.id} for user {grant_request.requested_for}")
        
        # Отримуємо всі Access Records, пов'язані з цим grant запитом
        access_records = grant_request.access_records.all()
        
        if not access_records.exists():
            logger.warning(f"No access records found for grant request {grant_request.id}")
            return False
        
        updated_records = []
        target_user = grant_request.requested_for
        
        for record in access_records:
            # Перевіряємо, чи користувач вже має доступ до цього запису
            if record.access_users.filter(id=target_user.id).exists():
                logger.info(f"User {target_user} already has access to record {record.id}, skipping")
                continue
            
            # Додаємо користувача до access_users
            record.access_users.add(target_user)
            record.modified_at = timezone.now()
            record.modified_by = grant_request.requested_by
            record.save()
            
            # Створюємо запис в історії змін
            try:
                from .models import SystemAccessStatusHistory
                SystemAccessStatusHistory.objects.create(
                    access_record=record,
                    old_status=record.status,
                    new_status=record.status,  # Статус не змінюється, тільки додається користувач
                    changed_by=grant_request.requested_by,
                    change_reason=f"Access granted to user {target_user.get_full_name()} via request #{grant_request.id}"
                )
                logger.info(f"Created access history record for access record {record.id}")
            except Exception as e:
                logger.error(f"Failed to create access history for access record {record.id}: {str(e)}")
            
            # Створюємо запис AccessRequestSequence для відстеження послідовності
            try:
                from .models import AccessRequestSequence
                # Знаходимо максимальний порядковий номер для цього grant request
                max_order = AccessRequestSequence.objects.filter(
                    grant_request=grant_request
                ).aggregate(models.Max('order_number'))['order_number__max'] or 0
                order_number = max_order + 1
                
                AccessRequestSequence.objects.create(
                    grant_request=grant_request,
                    access_record=record,
                    order_number=order_number,
                    sequence_status='active'
                )
                logger.info(f"Created access sequence record {record.id}.{grant_request.id}.{order_number} for grant request {grant_request.id} and access record {record.id}")
            except Exception as e:
                logger.error(f"Failed to create access sequence for grant request {grant_request.id} and access record {record.id}: {str(e)}")
            
            updated_records.append({
                'record_id': record.id,
                'system': record.asset.name,
                'environment': record.environment,
                'granted_at': timezone.now(),
                'granted_by': grant_request.requested_by.get_full_name(),
                'granted_to': target_user.get_full_name(),
                'grant_request_id': grant_request.id
            })
            
            logger.info(f"Added user {target_user} to access record {record.id}")
        
        # Логуємо результат
        logger.info(f"Successfully processed grant request {grant_request.id} for user {target_user}. Updated {len(updated_records)} access records")
        
        return True
        
    except Exception as e:
        logger.error(f"Error processing approved grant request {grant_request.id}: {str(e)}", exc_info=True)
        return False


def process_approved_revoke_request(revoke_request):
    """
    Обробка схваленого revoke запиту - позначення Access Records як revoked
    ТІЛЬКИ для конкретного користувача (requested_for)
    """
    try:
        from django.utils import timezone
        
        # Визначаємо користувача для логування
        if revoke_request.third_party_email:
            user_info = f"Third Party user {revoke_request.third_party_email}"
        else:
            user_info = f"Cabinet user {revoke_request.requested_for}"
        
        logger.info(f"Processing approved revoke request {revoke_request.id} for {user_info}")
        
        # Отримуємо всі Access Records, пов'язані з цим revoke запитом
        access_records = revoke_request.access_records.all()
        
        if not access_records.exists():
            logger.warning(f"No access records found for revoke request {revoke_request.id}")
            return False
        
        # Знаходимо або створюємо статус "Revoked" для кожної системи
        revoked_statuses = {}
        updated_records = []
        
        # Отримуємо користувача, для якого скасовується доступ
        target_user = revoke_request.requested_for
        target_user_email = None
        
        # Для Third Party користувачів використовуємо email
        if revoke_request.third_party_email:
            target_user_email = revoke_request.third_party_email
            logger.info(f"Processing revoke for Third Party user: {target_user_email}")
        elif target_user:
            target_user_email = target_user.email
            logger.info(f"Processing revoke for Cabinet user: {target_user_email}")
        
        # Отримуємо список конкретних Grant Access Record ID, які скасовуються
        revoked_grant_access_record_ids = revoke_request.revoked_grant_access_record_ids or []
        logger.info(f"Revoking specific Grant Access Record IDs: {revoked_grant_access_record_ids}")
        
        for record in access_records:
            system = record.asset
            
            # ВАЖЛИВО: Перевіряємо, чи цей користувач дійсно має доступ до цього запису
            if target_user and not record.access_users.filter(id=target_user.id).exists():
                logger.warning(f"User {target_user} doesn't have access to record {record.id}, skipping revocation")
                continue
            
            # Знаходимо всі Grant запити для цього користувача з такими ж ролями та об'єктом
            # Це потрібно для того, щоб скасувати доступ тільки в тих записах, які були надані цьому користувачу
            if target_user:
                # Для Cabinet користувачів
                grant_requests = AccessRequest.objects.filter(
                    request_type='grant',
                    requested_for=target_user,
                    system=system,
                    environment=record.environment,
                    admin_status='granted',
                    access_records=record
                )
            else:
                # Для Third Party користувачів
                grant_requests = AccessRequest.objects.filter(
                    request_type='grant',
                    third_party_email=target_user_email,
                    system=system,
                    environment=record.environment,
                    admin_status='granted',
                    access_records=record
                )
            
            if not grant_requests.exists():
                logger.warning(f"No matching grant requests found for user {target_user_email} and record {record.id}, skipping")
                continue
            
            # Перевіряємо, чи є конкретні Grant Access Record ID для скасування
            if revoked_grant_access_record_ids:
                # Знаходимо конкретні AccessRequestSequence записи для скасування
                from .models import AccessRequestSequence
                sequences_to_revoke = AccessRequestSequence.objects.filter(
                    access_record=record,
                    grant_request__in=grant_requests,
                    sequence_status='active'
                )

                # Підтримка формату з четвертою частиною A.B.C.D
                normalized_ids = set()
                for rid in revoked_grant_access_record_ids:
                    parts = str(rid).split('.')
                    key = '.'.join(parts[:3]) if len(parts) >= 3 else rid
                    normalized_ids.add(key)

                # Фільтруємо тільки ті записи, які мають відповідні перші три частини Grant Access Record ID
                sequences_to_revoke = [
                    seq for seq in sequences_to_revoke 
                    if f"{seq.access_record.id}.{seq.grant_request.id}.{seq.order_number}" in normalized_ids
                ]
                
                if not sequences_to_revoke:
                    logger.warning(f"No matching Grant Access Record IDs found for record {record.id}, skipping")
                    continue
                
                logger.info(f"Found {len(sequences_to_revoke)} sequences to revoke for record {record.id}")
            else:
                # Якщо не вказано конкретні ID, скасовуємо всі активні записи
                from .models import AccessRequestSequence
                sequences_to_revoke = AccessRequestSequence.objects.filter(
                    access_record=record,
                    grant_request__in=grant_requests,
                    sequence_status='active'
                )
                logger.info(f"No specific Grant Access Record IDs provided, revoking all {sequences_to_revoke.count()} active sequences for record {record.id}")
            
            # Отримуємо або створюємо статус "Revoked" для цієї системи та environment
            if (system.id, record.environment) not in revoked_statuses:
                revoked_status, created = AccessStatus.objects.get_or_create(
                    system=system,
                    environment=record.environment,
                    name='Revoked',
                    defaults={
                        'description': 'Access revoked by request',
                        'color': '#dc3545',
                        'order': 999
                    }
                )
                revoked_statuses[(system.id, record.environment)] = revoked_status
                
                if created:
                    logger.info(f"Created new 'Revoked' status for system {system.name} in {record.environment}")
            
            # Скасовуємо конкретні sequences (завжди, незалежно від створення статусу)
            for sequence in sequences_to_revoke:
                sequence.revoke_sequence(revoke_request)
                # Оновлюємо sequence_id до розширеного формату A.B.C.D, де D = revoke_request.id
                try:
                    parts = str(sequence.sequence_id).split('.')
                    if len(parts) >= 3:
                        sequence.sequence_id = f"{parts[0]}.{parts[1]}.{parts[2]}.{revoke_request.id}"
                        sequence.save()
                except Exception as e:
                    logger.warning(f"Failed to extend sequence_id for {sequence.sequence_id}: {str(e)}")
                logger.info(f"Revoked sequence {sequence.sequence_id} for record {record.id}")
            
            # Видаляємо користувача з access_users цього запису (тільки для Cabinet користувачів)
            if target_user:
                record.access_users.remove(target_user)
            
            # Оновлюємо статус запису ТІЛЬКИ якщо більше немає користувачів з доступом
            old_status = record.status
            if not record.access_users.exists():
                # Якщо більше немає користувачів з доступом, позначаємо запис як Revoked
                record.status = revoked_statuses[(system.id, record.environment)]
                record.modified_at = timezone.now()
                record.modified_by = revoke_request.requested_by
                record.save()
            else:
                # Якщо ще є користувачі з доступом, залишаємо статус без змін
                logger.info(f"Access record {record.id} still has {record.access_users.count()} users with access, status unchanged")
            
            # Створюємо запис в історії змін статусу Access Record
            try:
                from .models import SystemAccessStatusHistory
                # Формуємо change_reason для різних типів користувачів
                if target_user:
                    change_reason = f"Access revoked for user {target_user.get_full_name()} via request #{revoke_request.id}"
                else:
                    change_reason = f"Access revoked for Third Party user {target_user_email} via request #{revoke_request.id}"
                
                # Додаємо інформацію про конкретні Grant Access Record ID
                if revoked_grant_access_record_ids:
                    revoked_ids_str = ", ".join(revoked_grant_access_record_ids)
                    change_reason += f" (Grant Access Record IDs: {revoked_ids_str})"
                
                SystemAccessStatusHistory.objects.create(
                    access_record=record,
                    old_status=old_status,
                    new_status=record.status,
                    changed_by=revoke_request.requested_by,
                    change_reason=change_reason,
                    revoke_request=revoke_request
                )
                logger.info(f"Created status history record for access record {record.id}")
            except Exception as e:
                logger.error(f"Failed to create status history for access record {record.id}: {str(e)}")
            
            # ВАЖЛИВО: не відкликати будь-яку "active" послідовність за замовчуванням,
            # якщо були вказані конкретні Grant Access Record IDs. Вищe ми вже відкликали
            # лише ті послідовності, які співпали з revoked_grant_access_record_ids або всі, якщо список порожній.
            
            new_status_name = (record.status.name if record.status else None) or 'Unknown'
            updated_records.append({
                'record_id': record.id,
                'old_status': (old_status.name if old_status else None) or 'None',
                'new_status': new_status_name,
                'system': system.name,
                'environment': record.environment,
                'revoked_at': timezone.now(),
                'revoked_by': revoke_request.requested_by.get_full_name(),
                'revoked_for': target_user.get_full_name() if target_user else target_user_email,
                'revoke_request_id': revoke_request.id,
                'remaining_users': record.access_users.count()
            })
            
            if record.status and (record.status.name or '') == 'Revoked':
                logger.info(f"Updated access record {record.id} status from '{old_status}' to 'Revoked' for user {target_user_email} (no more users with access)")
            else:
                logger.info(f"Removed user {target_user_email} from access record {record.id}, status unchanged ({record.access_users.count()} users still have access)")
        
        # Логуємо результат
        logger.info(f"Successfully processed revoke request {revoke_request.id} for user {target_user_email}. Updated {len(updated_records)} access records")
        
        return True
        
    except Exception as e:
        logger.error(f"Error processing approved revoke request {revoke_request.id}: {str(e)}", exc_info=True)
        return False


@login_required
@ensure_csrf_cookie
def user_access_request(request):
    """Відображення сторінки запиту доступу"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        messages.error(request, _("Access denied - you are not authorized to submit access requests. Please contact your administrator to be added to the request list for at least one system."))
        return redirect('index')
    
    # Автоматично перевіряємо та оновлюємо записи з минулими датами
    expired_count = check_and_update_expired_records()
    if expired_count > 0 and settings.DEBUG is True and False:
        # Suppressed noisy info log even in DEBUG
        logger.info(f"Automatically set {expired_count} expired records to inactive before loading user access request page")
    
    current_language = get_language()[:2]

    # Базовий запит - показувати запити де користувач є: Requested By, Owner, Administrator, або Approving Person
    requests_query = AccessRequest.objects.filter(
        Q(requested_by=request.user) |  # Запити, які подав користувач
        Q(system__owners__cabinet_user__user=request.user) |  # Запити для систем, де користувач є власником
        Q(system__administrators__cabinet_user__user=request.user) |  # Запити для систем, де користувач є адміністратором
        Q(request_approvers__cabinet_user__user=request.user)  # Запити, які користувач має затверждувати
    ).distinct().select_related(
        'company',
        'system',
        'requested_by',
        'requested_for',
        'requested_for__cabinet',
        'requested_for__cabinet__department',
        'requested_for__cabinet__position'
    ).prefetch_related(
        'request_approvers__cabinet_user',
        'request_approvers__cabinet_user__user',
        'request_approvers__cabinet_user__department',
        'request_approvers__cabinet_user__position',
        'request_approvers__status_history',
        'system__owners__cabinet_user',
        'system__owners__cabinet_user__user',
        'system__owners__cabinet_user__department',
        'system__owners__cabinet_user__position',
        'system__administrators__cabinet_user',
        'system__administrators__cabinet_user__user',
        'system__administrators__cabinet_user__department',
        'system__administrators__cabinet_user__position',
        'attachments',
        Prefetch('admin_status_history', 
                 queryset=AccessRequestAdminStatusHistory.objects.select_related('changed_by').order_by('-changed_at')),
        Prefetch('email_notifications',
                 queryset=EmailNotificationHistory.objects.select_related('triggered_by', 'mail_account').order_by('-created_at')),
        'third_party_users',
        'access_records',
        'access_records__roles',
        'access_records__access_object'
    )
    
    # Застосовуємо фільтри
    if request.GET.get('company'):
        requests_query = requests_query.filter(company_id=request.GET.get('company'))
    
    if request.GET.get('system'):
        requests_query = requests_query.filter(system_id=request.GET.get('system'))
    
    if request.GET.get('object'):
        requests_query = requests_query.filter(access_records__access_object_id=request.GET.get('object'))
    
    if request.GET.get('status'):
        requests_query = requests_query.filter(status=request.GET.get('status'))
    
    if request.GET.get('environment'):
        requests_query = requests_query.filter(environment=request.GET.get('environment'))
    
    if request.GET.get('role'):
        role_id = request.GET.get('role')
        requests_query = requests_query.filter(access_records__roles__id=role_id)
    
    if request.GET.get('search'):
        search_query = request.GET.get('search')
        requests_query = requests_query.filter(
            Q(id__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(requested_for__username__icontains=search_query) |
            Q(justification__icontains=search_query) |
            Q(requirements__icontains=search_query) |
            Q(notes__icontains=search_query) |
            Q(third_party_first_name__icontains=search_query) |
            Q(third_party_last_name__icontains=search_query) |
            Q(third_party_email__icontains=search_query) |
            Q(third_party_organization__icontains=search_query) |
            Q(third_party_users__first_name__icontains=search_query) |
            Q(third_party_users__last_name__icontains=search_query) |
            Q(third_party_users__email__icontains=search_query) |
            Q(third_party_users__organization__icontains=search_query)
        )
    
    if request.GET.get('period'):
        now = timezone.now()
        if request.GET.get('period') == 'active':
            requests_query = requests_query.filter(
                Q(start_date__lte=now) & 
                (Q(end_date__gt=now) | Q(end_date__isnull=True))
            )
        elif request.GET.get('period') == 'expired':
            requests_query = requests_query.filter(end_date__lt=now)
        elif request.GET.get('period') == 'future':
            requests_query = requests_query.filter(start_date__gt=now)
    
    # Фільтрація за admin_status
    if request.GET.get('admin_status'):
        requests_query = requests_query.filter(admin_status=request.GET.get('admin_status'))
    
    # Фільтрація за request_type
    if request.GET.get('request_type'):
        requests_query = requests_query.filter(request_type=request.GET.get('request_type'))
    
    # Фільтрація за датою створення
    if request.GET.get('date'):
        try:
            date = timezone.datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()
            requests_query = requests_query.filter(created_at__date=date)
        except ValueError:
            logger.warning(f"Invalid date format: {request.GET.get('date')}")
    
    # Сортуємо результати
    requests = requests_query.order_by('-created_at')

    # Пагінація (за замовчуванням 25 записів на сторінку)
    page_size = get_access_table_page_size(request)
    page = request.GET.get('page', 1)
    paginator = Paginator(requests, page_size)
    try:
        requests_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        requests_page = paginator.page(1)

    # Відображення коду мови для моделі
    lang_code = 'ua' if current_language == 'uk' else current_language

    # Підготовка даних для відображення
    formatted_requests = []
    for req in requests_page:
        try:
            # Backfill request approvers if missing
            if not req.request_approvers.exists():
                # Вибираємо джерело затверджувачів: з першого access_record або з системних
                source_approvers = []
                try:
                    source_record = req.access_records.first()
                    if source_record:
                        source_approvers = list(source_record.approvers.all().order_by('order'))
                except Exception:
                    source_approvers = []

                if not source_approvers:
                    try:
                        source_approvers = list(req.system.approving_persons.all().order_by('order'))
                    except Exception:
                        source_approvers = []

                if source_approvers:
                    from .models import AccessRequestApprover, AccessApprover, ApprovingPerson
                    created_any = False
                    for sa in source_approvers:
                        # Підтримка двох типів джерел: AccessApprover або ApprovingPerson
                        if isinstance(sa, AccessApprover):
                            cabinet_user = sa.cabinet_user
                            order = sa.order
                        else:  # ApprovingPerson
                            cabinet_user = sa.cabinet_user
                            order = sa.order

                        try:
                            AccessRequestApprover.objects.create(
                                access_request=req,
                                access_approver=sa if isinstance(sa, AccessApprover) else None,
                                cabinet_user=cabinet_user,
                                order=order,
                                current_status='approved' if req.status == 'approved' else 'pending'
                            )
                            created_any = True
                        except Exception:
                            # Ignore duplicates or creation issues; continue best-effort
                            pass
                    if created_any:
                        # Refresh relation for subsequent usage
                        req.refresh_from_db()
        except Exception:
            # Never break page due to backfill issues
            pass
        # Визначаємо чи може користувач адмініструвати цей запит
        user_can_admin = False
        if hasattr(request.user, 'cabinet') and request.user.cabinet:
            user_can_admin = req.system.administrators.filter(cabinet_user=request.user.cabinet).exists()
        
        # Отримуємо структуровані дані про ролі та об'єкти з усіх записів доступу
        access_records_data = []
        roles = []
        if req.access_records.exists():
            # Підготуємо можливий оригінальний grant-запит для revoke
            original_request = None
            if req.request_type == 'revoke':
                if req.notes:
                    import re
                    match = re.search(r'request #(\d+)', req.notes)
                    if match:
                        try:
                            original_request = AccessRequest.objects.get(id=int(match.group(1)))
                        except AccessRequest.DoesNotExist:
                            original_request = None
                # Fallback for granular revoke: original grant ID from revoked_grant_access_record_ids (format A.B.C = access_record.grant_request.order)
                if original_request is None:
                    revoked_ids = getattr(req, 'revoked_grant_access_record_ids', None) or []
                    if revoked_ids:
                        try:
                            first_id = str(revoked_ids[0])
                            parts = first_id.split('.')
                            if len(parts) >= 2 and parts[1].isdigit():
                                grant_request_id = int(parts[1])
                                original_request = AccessRequest.objects.filter(
                                    id=grant_request_id,
                                    request_type='grant',
                                    status='approved',
                                    admin_status='granted'
                                ).first()
                        except (IndexError, ValueError, TypeError):
                            pass

            requested_role_by_record = {}
            if getattr(req, 'requested_access_record_roles', None) and isinstance(req.requested_access_record_roles, list):
                for item in req.requested_access_record_roles:
                    if isinstance(item, dict) and 'access_record_id' in item and 'role_id' in item:
                        requested_role_by_record[item['access_record_id']] = item['role_id']
            # For Revoke: show the requested role from the original grant (the role being revoked)
            if req.request_type == 'revoke' and original_request is not None and getattr(original_request, 'requested_access_record_roles', None) and isinstance(original_request.requested_access_record_roles, list):
                for item in original_request.requested_access_record_roles:
                    if isinstance(item, dict) and 'access_record_id' in item and 'role_id' in item:
                        requested_role_by_record[item['access_record_id']] = item['role_id']

            # Збираємо структуровані дані з усіх записів доступу
            for access_record in req.access_records.all():
                record_roles = list(access_record.roles.all())
                roles.extend(record_roles)
                requested_role = None
                rid = requested_role_by_record.get(access_record.id)
                if rid:
                    try:
                        requested_role = AccessRoles.objects.get(id=rid)
                    except AccessRoles.DoesNotExist:
                        pass
                
                                # Перевіряємо чи є цей Access Record revoked
                is_revoked = False
                revoke_info = None
                grant_access_record_id = None

                # (moved) Grant revocation check based on full ID happens after we resolve grant_access_record_id below
                
                # Визначаємо Grant Access Record ID для відображення в списку "My Access Requests"
                try:
                    from .models import AccessRequestSequence
                    sequence_record = None
                    grant_access_record_id_full = None
                    if req.request_type == 'revoke':
                        # Спершу пробуємо зв'язок за revoke_request
                        sequence_record = AccessRequestSequence.objects.filter(
                            access_record=access_record,
                            revoke_request=req
                        ).order_by('order_number').first()
                        # Якщо не знайдено, пробуємо за оригінальним grant-запитом
                        if not sequence_record and original_request is not None:
                            sequence_record = AccessRequestSequence.objects.filter(
                                access_record=access_record,
                                grant_request=original_request
                            ).order_by('order_number').first()
                        # Додатковий fallback: шукаємо серед відкликаних послідовностей із суфіксом .<revoke_id>
                        if not sequence_record:
                            sequence_record = AccessRequestSequence.objects.filter(
                                access_record=access_record,
                                sequence_status='revoked',
                                sequence_id__endswith=f".{req.id}"
                            ).order_by('order_number').first()
                        # Додатковий fallback: використовуємо список IDs із самого Revoke запиту
                        # Знаходимо A.B.C для конкретного access_record за префіксом A = access_record.id
                        if not sequence_record:
                            try:
                                revoked_ids = getattr(req, 'revoked_grant_access_record_ids', None) or []
                                matched_id = None
                                for rid in revoked_ids:
                                    rid_str = str(rid)
                                    if rid_str.startswith(f"{access_record.id}."):
                                        matched_id = rid_str
                                        break
                                if matched_id:
                                    # Синтезуємо псевдо sequence_record-дані для відображення іконки "i"
                                    grant_access_record_id = '.'.join(matched_id.split('.')[:3])
                                    grant_access_record_id_full = f"{grant_access_record_id}.{req.id}"
                            except Exception:
                                pass
                    elif req.request_type == 'grant':
                        sequence_record = AccessRequestSequence.objects.filter(
                            access_record=access_record,
                            grant_request=req
                        ).order_by('order_number').first()
                        
                        # Debug logging for sequence lookup
                        sequence_count = AccessRequestSequence.objects.filter(
                            access_record=access_record,
                            grant_request=req
                        ).count()
                        #
                        
                        # Якщо з якоїсь причини прямий зв'язок не знайдено (після повторних циклів grant/revoke),
                        # шукаємо по префіксу A.B.* для цього grant запиту
                        if not sequence_record:
                            from .models import AccessRequestSequence as ARS
                            sequence_record = ARS.objects.filter(
                                sequence_id__startswith=f"{access_record.id}.{req.id}.",
                                grant_request=req
                            ).order_by('order_number').first()
                            
                            if sequence_record:
                                logger.debug(f"Found sequence by prefix search: {sequence_record.sequence_id}")
                            else:
                                # Auto-create missing AccessRequestSequence for approved grant requests
                                if req.status == 'approved' and req.admin_status == 'granted':
                                    try:
                                        from .models import AccessRequestSequence
                                        from django.db import models
                                        # Знаходимо максимальний порядковий номер для цього grant request
                                        max_order = AccessRequestSequence.objects.filter(
                                            grant_request=req
                                        ).aggregate(models.Max('order_number'))['order_number__max'] or 0
                                        order_number = max_order + 1
                                        
                                        sequence_record = AccessRequestSequence.objects.create(
                                            grant_request=req,
                                            access_record=access_record,
                                            order_number=order_number,
                                            sequence_status='active'
                                        )
                                        logger.debug(
                                            f"Auto-created missing AccessRequestSequence {access_record.id}.{req.id}.{order_number} for Grant Request {req.id}, Access Record {access_record.id}"
                                        )
                                    except Exception as e:
                                        logger.error(
                                            f"Failed to auto-create AccessRequestSequence for Grant Request {req.id}, Access Record {access_record.id}: {str(e)}"
                                        )

                    if sequence_record:
                        # Для відображення нормалізуємо до перших трьох частин
                        seq_id_str = str(sequence_record.sequence_id)
                        grant_access_record_id = '.'.join(seq_id_str.split('.')[:3])
                        # Повний ID для відображення: додаємо D-частину (0 якщо активний, або ID Revoke)
                        if sequence_record.sequence_status == 'revoked' and sequence_record.revoke_request:
                            grant_access_record_id_full = f"{grant_access_record_id}.{sequence_record.revoke_request.id}"
                            # Іконку "revoked" виставляємо лише для рядків типу Grant (не для Revoke)
                            if req.request_type == 'grant' and sequence_record.revoke_request.admin_status == 'granted':
                                is_revoked = True
                                revoke_request = sequence_record.revoke_request
                                revoked_for_correct = None
                                if revoke_request.third_party_email:
                                    revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                elif revoke_request.requested_for:
                                    revoked_for_correct = revoke_request.requested_for.get_full_name()
                                revoke_info = {
                                    'revoked_at': getattr(sequence_record, 'revoked_at', None) or revoke_request.created_at,
                                    'revoked_by': revoke_request.requested_by.get_full_name() if revoke_request.requested_by else 'System',
                                    'revoked_for': revoked_for_correct,
                                    'revoke_request_id': revoke_request.id,
                                    'change_reason': 'Access revoked for this grant sequence'
                                }
                        else:
                            grant_access_record_id_full = f"{grant_access_record_id}.0"
                except Exception as e:
                    logger.error(f"Failed to resolve Grant Access Record ID for request {req.id}, record {access_record.id}: {str(e)}")

                # Перевірка по послідовності для рядків Grant: базуємось на повному Grant Access Record ID (A.B.C.D)
                if not is_revoked and req.request_type == 'grant' and grant_access_record_id:
                    try:
                        from .models import AccessRequestSequence as ARS
                        seq_rev = ARS.objects.filter(
                            sequence_id__startswith=f"{grant_access_record_id}.",
                            sequence_status='revoked',
                            revoke_request__isnull=False,
                            revoke_request__admin_status='granted'
                        ).order_by('-revoked_at').first()
                        if seq_rev:
                            revoke_request = seq_rev.revoke_request
                            
                            # Перевіряємо, чи revoke стосується того ж користувача, що і grant
                            grant_user_email = None
                            revoke_user_email = None
                            
                            if req.third_party_email:
                                grant_user_email = req.third_party_email
                            elif req.requested_for:
                                grant_user_email = req.requested_for.email
                            
                            if revoke_request.third_party_email:
                                revoke_user_email = revoke_request.third_party_email
                            elif revoke_request.requested_for:
                                revoke_user_email = revoke_request.requested_for.email
                            
                            # Тільки позначаємо як revoked, якщо це стосується того ж користувача
                            if grant_user_email and revoke_user_email and grant_user_email == revoke_user_email:
                                is_revoked = True
                                revoked_for_correct = None
                                if revoke_request.third_party_email:
                                    revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                elif revoke_request.requested_for:
                                    revoked_for_correct = revoke_request.requested_for.get_full_name()
                                revoke_info = {
                                    'revoked_at': getattr(seq_rev, 'revoked_at', None) or revoke_request.created_at,
                                    'revoked_by': revoke_request.requested_by.get_full_name() if revoke_request.requested_by else 'System',
                                    'revoked_for': revoked_for_correct,
                                    'revoke_request_id': revoke_request.id,
                                    'change_reason': 'Access revoked for this grant sequence'
                                }
                        else:
                            # Фолбек: якщо немає revoke_request (дані старих записів), але послідовність має статус revoked
                            seq_rev2 = ARS.objects.filter(
                                sequence_id__startswith=f"{grant_access_record_id}.",
                                sequence_status='revoked'
                            ).order_by('-revoked_at').first()
                            if seq_rev2:
                                is_revoked = True
                                revoke_info = {
                                    'revoked_at': getattr(seq_rev2, 'revoked_at', None),
                                    'revoked_by': 'System',
                                    'revoked_for': None,
                                    'revoke_request_id': getattr(seq_rev2.revoke_request, 'id', None),
                                    'change_reason': 'Access revoked (no linked request)'
                                }
                            else:
                                # Додатковий фолбек: шукаємо Revoke запити, які містять цей Grant Access Record ID у списку
                                existing_rr = AccessRequest.objects.filter(
                                    request_type='revoke',
                                    admin_status='granted',
                                    revoked_grant_access_record_ids__contains=[grant_access_record_id]
                                ).order_by('-created_at').first()
                                if existing_rr:
                                    revoke_request = existing_rr
                                    
                                    # Перевіряємо, чи revoke стосується того ж користувача, що і grant
                                    grant_user_email = None
                                    revoke_user_email = None
                                    
                                    if req.third_party_email:
                                        grant_user_email = req.third_party_email
                                    elif req.requested_for:
                                        grant_user_email = req.requested_for.email
                                    
                                    if revoke_request.third_party_email:
                                        revoke_user_email = revoke_request.third_party_email
                                    elif revoke_request.requested_for:
                                        revoke_user_email = revoke_request.requested_for.email
                                    
                                    # Тільки позначаємо як revoked, якщо це стосується того ж користувача
                                    if grant_user_email and revoke_user_email and grant_user_email == revoke_user_email:
                                        is_revoked = True
                                        revoked_for_correct = None
                                        if revoke_request.third_party_email:
                                            revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                        elif revoke_request.requested_for:
                                            revoked_for_correct = revoke_request.requested_for.get_full_name()
                                        revoke_info = {
                                            'revoked_at': revoke_request.created_at,
                                            'revoked_by': revoke_request.requested_by.get_full_name() if revoke_request.requested_by else 'System',
                                            'revoked_for': revoked_for_correct,
                                            'revoke_request_id': revoke_request.id,
                                            'change_reason': 'Access revoked for this grant sequence'
                                        }
                    except Exception as e:
                        logger.error(f"Sequence revoke check by full ID failed for request {req.id}, record {access_record.id}: {str(e)}")
                
                # Debug logging removed

                access_records_data.append({
                    'id': access_record.id,
                    'object_id': access_record.access_object.id if access_record.access_object else None,
                    'object_name': access_record.access_object.get_name(current_language) if access_record.access_object else _('No Object'),
                    'object_color': access_record.access_object.color if access_record.access_object else '#6c757d',
                    'roles': record_roles,
                    'requested_role': requested_role,
                    'environment': access_record.environment,
                    'start_date': access_record.start_date,
                    'end_date': access_record.end_date,
                    'is_revoked': is_revoked,
                    'revoke_info': revoke_info,
                    'grant_access_record_id': grant_access_record_id,
                    'grant_access_record_id_full': grant_access_record_id_full or (f"{grant_access_record_id}.0" if grant_access_record_id else None)
                })
        # Fallback logic removed - all records should use access_records now
            
        # Для revoke запитів отримуємо інформацію про оригінальний запит
        original_request_id = None
        original_request = None
        if req.request_type == 'revoke' and req.notes:
            # Витягуємо ID оригінального запиту з notes
            import re
            match = re.search(r'request #(\d+)', req.notes)
            if match:
                original_request_id = int(match.group(1))
                try:
                    original_request = AccessRequest.objects.get(id=original_request_id)
                except AccessRequest.DoesNotExist:
                    original_request = None
            
        # Обробляємо revocation timing для revoke запитів
        revocation_timing = None
        is_immediate_revocation = False
        if req.request_type == 'revoke':
            # Перевіряємо чи це негайне скасування
            if req.notes and 'Revocation timing: Immediate' in req.notes:
                is_immediate_revocation = True
                revocation_timing = _('Immediately')
            elif req.start_date:
                # Запланована дата скасування
                if req.end_date:
                    revocation_timing = f"{timezone.localtime(req.start_date).strftime('%d.%m.%Y %H:%M')} - {timezone.localtime(req.end_date).strftime('%d.%m.%Y %H:%M')}"
                else:
                    revocation_timing = f"{timezone.localtime(req.start_date).strftime('%d.%m.%Y %H:%M')} - {_('Permanent')}"
            else:
                revocation_timing = _('Immediately')
                is_immediate_revocation = True

        # Отримуємо існуючі granted ролі користувача для цього Object та Environment
        existing_granted_roles = []
        # TODO: Fix existing granted roles logic for ManyToMany access_records
        # Temporarily disabled to prevent errors

        # Sort access records for Grant: keep revoked at the bottom
        if req.request_type == 'grant':
            try:
                access_records_data.sort(key=lambda item: item.get('is_revoked', False))
            except Exception:
                pass

        # Побудова короткого тексту історії погоджень для тултіпа статусу запиту
        approval_history_text = None
        try:
            events = []
            # 1) Події з поточних request_approvers
            if req.request_approvers.exists():
                for ra in req.request_approvers.all().order_by('order'):
                    history_qs = getattr(ra, 'status_history', None)
                    if history_qs is not None:
                        for h in history_qs.all().order_by('changed_at'):
                            approver_name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                            changer_name = h.changed_by.get_full_name() if h.changed_by else ''
                            ts = timezone.localtime(h.changed_at).strftime('%d.%m.%Y %H:%M')
                            events.append((h.changed_at, ra.order, f"Lvl {ra.order}: {approver_name} → {h.new_status.title()} ({ts} by {changer_name})"))
                    else:
                        approver_name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                        events.append((None, ra.order, f"Lvl {ra.order}: {approver_name} - {ra.current_status.title()}"))

            # 2) Snapshot-історія, прив'язана безпосередньо до запиту (на випадок заміни/видалення approver'ів)
            try:
                from .models import AccessRequestApproverStatusHistory as ARASH
                snapshot_qs = ARASH.objects.filter(access_request=req).order_by('changed_at')
                for h in snapshot_qs:
                    approver_name = h.approver_name or (h.approver_cabinet_user.user.get_full_name() if h.approver_cabinet_user and getattr(h.approver_cabinet_user, 'user', None) else '')
                    changer_name = h.changed_by.get_full_name() if h.changed_by else ''
                    ts = timezone.localtime(h.changed_at).strftime('%d.%m.%Y %H:%M')
                    order_num = getattr(h, 'order_at_change', None) or 0
                    events.append((h.changed_at, order_num, f"Lvl {order_num}: {approver_name} → {h.new_status.title()} ({ts} by {changer_name})"))
            except Exception:
                pass

            if events:
                events.sort(key=lambda x: (x[0] is None, x[0] or 0, x[1] or 0))
                approval_history_text = " | ".join([e[2] for e in events])
        except Exception:
            approval_history_text = None

        # Форматуємо дані
        formatted_requests.append({
            'id': req.id,
            'request_type': req.request_type,
            'company_id': req.company_id,
            'company_name': req.company.name,
            'system_id': req.system_id,
            'system_name': req.system.name,
            'object_id': req.access_records.first().access_object.id if req.access_records.exists() and req.access_records.first().access_object else None,
            'object_name': req.access_records.first().access_object.get_name(current_language) if req.access_records.exists() and req.access_records.first().access_object else _('No Object'),
            'environment': req.environment,
            'requested_by': req.requested_by,
            'requested_for': req.requested_for,
            'approvers': req.request_approvers.all() if hasattr(req, 'request_approvers') else [],
            'owners': req.system.owners.all() if hasattr(req.system, 'owners') else [],
            'administrators': req.system.administrators.all() if hasattr(req.system, 'administrators') else [],
            'system_approvers': req.system.approving_persons.all() if hasattr(req.system, 'approving_persons') else [],
            'roles': roles,
            'access_records_data': access_records_data,
            'access_records_count': len(access_records_data),
            'justification': req.justification,
            'requirements': req.requirements,
            'notes': req.notes,
            'attachments': req.attachments,
            'created_at': req.created_at,
            'record_start_date': req.start_date,
            'record_end_date': req.end_date,
            'record_progress': calculate_progress(req.start_date, req.end_date),
            'status': req.status,
            'admin_status': req.admin_status,
            'approval_history_text': approval_history_text,
            'admin_status_history': req.admin_status_history,
            'can_be_cancelled': req.can_be_cancelled(request.user),
            'can_be_edited': req.can_be_edited(request.user),
            'is_cancelled': req.is_cancelled,
            'cancelled_at': req.cancelled_at,
            'cancelled_by': req.cancelled_by,
            'cancellation_reason': req.cancellation_reason,
            'email_notifications': req.email_notifications.all().order_by('-created_at')[:5],  # Останні 5 email повідомлень
            # Додаємо поля третьої сторони
            'third_party_first_name': req.third_party_first_name,
            'third_party_last_name': req.third_party_last_name,
            'third_party_email': req.third_party_email,
            'third_party_phone': req.third_party_phone,
            'third_party_organization': req.third_party_organization,
            'third_party_description': req.third_party_description,
            # Новий зв'язок з ThirdPartyUser
            'third_party_users': req.third_party_users.all() if hasattr(req, 'third_party_users') else [],
            'third_party_users_count': req.third_party_users.count() if hasattr(req, 'third_party_users') else 0,
            # Додаємо інформацію про кількість записів доступу
            'access_records_count': req.access_records.count() if hasattr(req, 'access_records') else 0,
            # Додаємо інформацію про оригінальний запит для revoke запитів
            'original_request_id': original_request_id,
            'original_request': original_request,
            # Додаємо інформацію про revocation timing
            'revocation_timing': revocation_timing,
            'is_immediate_revocation': is_immediate_revocation,
            # Права користувача для цього запиту
            'user_can_admin': user_can_admin,
            # Додаємо інформацію про існуючі granted ролі
            'existing_granted_roles': existing_granted_roles,
            **_access_request_display_context(req),

        })

    # Отримуємо список компаній та систем для фільтрів, доступних користувачу
    user_groups = request.user.groups.all()
    
    # Отримуємо тільки компанії, в яких є системи доступні користувачу
    companies = Company.objects.filter(
        informationasset__access_records__is_active=True
    ).filter(
        Q(informationasset__access_records__request_users=request.user) |
        Q(informationasset__access_records__request_groups__in=user_groups)
    ).distinct().order_by('name')
    
    systems = []
    objects = []
    if request.GET.get('company'):
        # Отримуємо тільки системи доступні користувачу в вибраній компанії
        systems = InformationAsset.objects.filter(
            company_id=request.GET.get('company'),
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().order_by('name')
    
    if request.GET.get('system'):
        from .models import AccessObjectIS
        # Отримуємо тільки об'єкти доступні користувачу в вибраній системі
        objects = AccessObjectIS.objects.filter(
            asset_id=request.GET.get('system'),
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().order_by('order', 'name')
    
    # Отримуємо ID ролей з запитів доступу користувача
    role_ids = AccessRequest.objects.filter(
        requested_by=request.user
    ).values_list('access_records__roles__id', flat=True).distinct()
    
    # Отримуємо ролі за ID
    roles = AccessRoles.objects.filter(
        id__in=role_ids, is_active=True
    ).order_by('order', 'name', 'code')
    
    # Отримуємо список організацій для dropdown (JSON-serializable)
    third_party_organizations = list(
        ThirdPartyOrganization.objects.filter(is_active=True)
        .order_by('name')
        .values('id', 'name')
    )
    selected_company_id = request.GET.get('company')
    access_justification_templates = _serialize_access_justification_templates(
        company_id=selected_company_id
    )
    
    return render(request, 'app_access/user_access_request.html', {
        'requests': formatted_requests,
        'companies': companies,
        'systems': systems,
        'objects': objects,
        'roles': roles,
        'current_language': current_language,
        'third_party_organizations': third_party_organizations,
        'access_justification_templates': access_justification_templates,
        # Пагінація
        'paginator': paginator,
        'page_obj': requests_page,
        'is_paginated': paginator.count > 0,
        'current_page_size': page_size,
        'page_size_options': ACCESS_TABLE_PAGE_SIZE_OPTIONS,
    })


@login_required
@require_http_methods(["GET"])
def user_access_request_guide(request):
    """Return JSON { content: html } for the User Access Request guide (localized)."""
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({'content': ''})
    lang = (get_language() or 'en')[:2]
    lang_to_code = {'uk': 'UA', 'en': 'GB', 'ru': 'RU'}
    code = lang_to_code.get(lang, 'GB')
    try:
        country = Country.objects.filter(code=code).first()
    except Exception:
        country = None
    content = ''
    guide = UserAccessRequestGuide.objects.first()
    if guide:
        if country:
            trans = UserAccessRequestGuideTranslation.objects.filter(guide=guide, country=country).first()
            if trans and trans.content:
                content = trans.content
        if not content and guide.base_content:
            content = guide.base_content
        if not content:
            trans = UserAccessRequestGuideTranslation.objects.filter(guide=guide).select_related('country').first()
            if trans and trans.content:
                content = trans.content
    return JsonResponse({'content': content})


@login_required
@require_POST
def user_access_request_guide_translate(request):
    """API for AI translation of User Access Request guide content (admin)."""
    try:
        data = json.loads(request.body)
        text = (data.get('text') or '').strip()
        country_id = data.get('country_id')
        if not text:
            return JsonResponse({'error': 'Text is required'}, status=400)
        if not country_id:
            return JsonResponse({'error': 'Country ID is required'}, status=400)
        country = Country.objects.get(id=country_id)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
    lang_map = {
        'ua': 'uk', 'gb': 'en', 'us': 'en', 'uk': 'en', 'ru': 'ru',
        'kz': 'kk', 'by': 'be', 'md': 'ro', 'ge': 'ka', 'am': 'hy', 'az': 'az',
        'ch': 'de', 'at': 'de', 'be': 'nl', 'dk': 'da', 'no': 'no', 'se': 'sv',
        'fi': 'fi', 'ee': 'et', 'lv': 'lv', 'lt': 'lt', 'cz': 'cs', 'sk': 'sk',
        'hu': 'hu', 'ro': 'ro', 'bg': 'bg', 'pl': 'pl', 'fr': 'fr', 'es': 'es',
        'it': 'it', 'cn': 'zh-cn', 'jp': 'ja', 'kr': 'ko', 'tr': 'tr',
    }
    target = lang_map.get(country.code.lower(), country.code.lower())
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source='auto', target=target)
        translated = translator.translate(text)
        return JsonResponse({
            'success': True,
            'translated_text': translated,
            'target_language': target,
            'country_name': country.name,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_access_records(request, system_id, object_id, user_id):
    """Отримання доступних записів доступу для користувача"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({
            'success': False,
            'error': _("Access denied - you are not authorized to view access records.")
        }, status=403)
    """Отримання доступних записів доступу для користувача"""
    try:
        # Автоматично перевіряємо та оновлюємо записи з минулими датами
        expired_count = check_and_update_expired_records()
        if expired_count > 0:
            logger.info(f"Automatically set {expired_count} expired records to inactive before loading available records")
        
        current_language = get_language()[:2]
        user_groups = request.user.groups.all()
        
        # Отримуємо environment з GET параметрів
        environment = request.GET.get('environment')
        
        # Базова фільтрація записів
        records = SystemAccess.objects.filter(
            asset_id=system_id,
            access_object_id=object_id,
            is_active=True
        ).filter(
            Q(end_date__gt=timezone.now()) |
            Q(end_date__isnull=True)
        ).filter(
            # Тільки ті записи, до яких поточний користувач може запитувати доступ
            Q(request_users=request.user) |
            Q(request_groups__in=user_groups)
        )
        
        # Фільтруємо по environment, якщо параметр передано
        if environment:
            logger.info(f"Filtering records by environment: {environment}")
            records = records.filter(environment=environment)
        else:
            logger.info("Loading all access records without environment filter")
        
        # Перевіряємо чи це третя сторона
        if user_id == 'third_parties':
            # Для третьої сторони показуємо записи з увімкненим third_parties
            records = records.filter(third_parties=True)
        else:
            # Для звичайних користувачів фільтруємо за доступом користувача
            # Конвертуємо user_id в int, якщо це не 'third_parties'
            try:
                user_id_int = int(user_id)
                records = records.filter(
                    Q(access_users__id=user_id_int) |
                    Q(access_groups__user__id=user_id_int)
                )
            except (ValueError, TypeError):
                logger.error(f"Invalid user_id format: {user_id}")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid user ID format'
                }, status=400)
        
        records = records.select_related(
            'asset',
            'access_object'
        ).prefetch_related(
            'roles',
            'asset__owners__cabinet_user__user',
            'asset__owners__cabinet_user__department',
            'asset__owners__cabinet_user__position',
            'asset__administrators__cabinet_user__user',
            'asset__administrators__cabinet_user__department',
            'asset__administrators__cabinet_user__position',
            'approvers__cabinet_user__user',
            'approvers__cabinet_user__department',
            'approvers__cabinet_user__position'
        ).distinct()

        processed_ids = set()
        formatted_records = []

        for record in records:
            if record.id in processed_ids:
                continue

            processed_ids.add(record.id)

            # Форматуємо власників
            owners = []
            for owner in record.asset.owners.all():
                cabinet_user = owner.cabinet_user
                owners.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None
                })

            # Форматуємо адміністраторів
            administrators = []
            for admin in record.asset.administrators.all():
                cabinet_user = admin.cabinet_user
                administrators.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None
                })

            # Форматуємо затверджувачів
            approvers = []
            for approver in record.approvers.all().order_by('order'):
                cabinet_user = approver.cabinet_user
                approvers.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None,
                    'order': approver.order
                })

            formatted_records.append({
                'id': record.id,
                'asset_id': system_id,  # Додаємо system_id як asset_id
                'system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо system_name
                'information_system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо information_system_name
                'object_id': record.access_object.id if record.access_object else None,  # Додаємо object_id
                'object_name': record.access_object.get_name(current_language) if record.access_object else 'Unknown Object',  # Додаємо object_name
                'object_color': record.access_object.color if record.access_object and hasattr(record.access_object, 'color') else '#6c757d',  # Додаємо object_color
                'environment': record.environment,  # Додаємо environment
                'roles': [{
                    'id': role.id,
                    'name': role.get_name(current_language),
                    'color': role.color or '#6c757d',
                    'description': role.get_description(current_language) if hasattr(role, 'get_description') else None
                } for role in record.roles.all()],
                'start_date': record.start_date.isoformat() if record.start_date else None,
                'end_date': record.end_date.isoformat() if record.end_date else None,
                'owners': owners,
                'administrators': administrators,
                'approvers': approvers
            })

        return JsonResponse({
            'status': 'success',
            'access_records': formatted_records
        })
    except Exception as e:
        logger.error(f"Error getting access records: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def get_access_records_by_system(request, system_id, user_id):
    """Отримання доступних записів доступу для вибраної системи та користувача (без фільтрації по об'єкту)"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({
            'success': False,
            'error': _("Access denied - you are not authorized to view access records.")
        }, status=403)
    
    try:
        # Автоматично перевіряємо та оновлюємо записи з минулими датами
        expired_count = check_and_update_expired_records()
        if expired_count > 0:
            logger.info(f"Automatically set {expired_count} expired records to inactive before loading available records")
        
        # Отримуємо мову з параметрів запиту або використовуємо поточну мову
        current_language = request.GET.get('language', get_language())[:2]
        user_groups = request.user.groups.all()
        
        # Базова фільтрація записів по системі (без об'єкта)
        records = SystemAccess.objects.filter(
            asset_id=system_id,
            is_active=True
        ).filter(
            Q(end_date__gt=timezone.now()) |
            Q(end_date__isnull=True)
        ).filter(
            # Тільки ті записи, до яких поточний користувач може запитувати доступ
            Q(request_users=request.user) |
            Q(request_groups__in=user_groups)
        )
        
        # Перевіряємо чи це третя сторона
        if user_id == 'third_parties':
            # Для третьої сторони показуємо записи з увімкненим third_parties
            records = records.filter(third_parties=True)
        else:
            # Для звичайних користувачів фільтруємо за доступом користувача
            # Конвертуємо user_id в int, якщо це не 'third_parties'
            try:
                user_id_int = int(user_id)
                records = records.filter(
                    Q(access_users__id=user_id_int) |
                    Q(access_groups__user__id=user_id_int)
                )
            except (ValueError, TypeError):
                logger.error(f"Invalid user_id format: {user_id}")
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid user ID format'
                }, status=400)
        
        records = records.select_related(
            'asset',
            'access_object'
        ).prefetch_related(
            'roles',
            'asset__owners__cabinet_user__user',
            'asset__owners__cabinet_user__department',
            'asset__owners__cabinet_user__position',
            'asset__administrators__cabinet_user__user',
            'asset__administrators__cabinet_user__department',
            'asset__administrators__cabinet_user__position',
            'approvers__cabinet_user__user',
            'approvers__cabinet_user__department',
            'approvers__cabinet_user__position'
        ).distinct()

        processed_ids = set()
        formatted_records = []

        for record in records:
            if record.id in processed_ids:
                continue

            processed_ids.add(record.id)

            # Форматуємо власників
            owners = []
            for owner in record.asset.owners.all():
                cabinet_user = owner.cabinet_user
                owners.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None
                })

            # Форматуємо адміністраторів
            administrators = []
            for admin in record.asset.administrators.all():
                cabinet_user = admin.cabinet_user
                administrators.append({
                    'name': cabinet_user.user.get_full_name(),
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': admin.cabinet_user.position.get_name(current_language) if admin.cabinet_user.position else None
                })

            # Форматуємо затверджувачів
            approvers = []
            for approver in record.approvers.all().order_by('order'):
                cabinet_user = approver.cabinet_user
                approvers.append({
                    'full_name': cabinet_user.user.get_full_name(),
                    'name': cabinet_user.user.get_full_name(),
                    'username': cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': cabinet_user.department.get_name(
                        current_language) if cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user.position else None,
                    'order': approver.order
                })

            formatted_records.append({
                'id': record.id,
                'asset_id': system_id,  # Додаємо system_id як asset_id
                'system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо system_name
                'information_system_name': record.asset.name if record.asset else 'Unknown System',  # Додаємо information_system_name
                'object_id': record.access_object.id if record.access_object else None,  # Додаємо object_id
                'object_name': record.access_object.get_name(current_language) if record.access_object else 'Unknown Object',  # Додаємо object_name
                'object_color': record.access_object.color if record.access_object and hasattr(record.access_object, 'color') else '#6c757d',  # Додаємо object_color
                'environment': record.environment,  # Додаємо environment
                'roles': [{
                    'id': role.id,
                    'name': role.get_name(current_language),
                    'color': role.color or '#6c757d',
                    'description': role.get_description(current_language) if hasattr(role, 'get_description') else None
                } for role in record.roles.all()],
                'start_date': record.start_date.isoformat() if record.start_date else None,
                'end_date': record.end_date.isoformat() if record.end_date else None,
                'owners': owners,
                'administrators': administrators,
                'approvers': approvers
            })

        return JsonResponse({
            'status': 'success',
            'access_records': formatted_records
        })
    except Exception as e:
        logger.error(f"Error getting access records by system: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


def handle_revoke_request(request, data, attachments):
    """Обробка Revoke запитів"""
    try:
        from django.utils import translation
        
        # Перевіряємо чи це новий granular revoke запит чи старий формат
        selected_access_records = data.get('selected_access_records')
        #
        
        if selected_access_records:
            # Новий формат - granular revoke з вибраними access records
            logger.debug("Using granular revoke request handler")
            return handle_granular_revoke_request(request, data, attachments)
        
        # Старий формат - відкликання цілого запиту
        # Отримуємо ID оригінального запиту для відкликання
        original_request_id = data.get('original_request_id')
        if not original_request_id:
            return JsonResponse({
                'success': False,
                'message': _('Original request ID or selected access records are required for revoke requests')
            }, status=400)
        
        # Перевіряємо що оригінальний запит існує і належить користувачу
        try:
            original_request = AccessRequest.objects.get(
                id=original_request_id,
                requested_by=request.user,
                request_type='grant',
                status='approved',
                admin_status='granted'
            )
        except AccessRequest.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': _('Original request not found or not accessible')
            }, status=404)
        
        # Перевіряємо чи немає вже активного revoke запиту для цього grant запиту
        # Використовуємо більш точну перевірку по ключових полях
        existing_revoke = AccessRequest.objects.filter(
            request_type='revoke',
            requested_by=request.user,
            requested_for_id=original_request.requested_for_id,
            access_record_id=original_request.access_record_id,
            system_id=original_request.system_id,
            environment=original_request.environment,
            status__in=['pending', 'approved'],
            notes__contains=f'request #{original_request_id}'
        ).first()
        
        if existing_revoke:
            # Формуємо детальне повідомлення про існуючий запит
            status_text = _('pending') if existing_revoke.status == 'pending' else _('approved')
            return JsonResponse({
                'success': False,
                'message': _('A revoke request for this access is already {status} (Request #{request_id}). Only one revoke request per grant request is allowed.').format(
                    status=status_text,
                    request_id=existing_revoke.id
                )
            }, status=400)
        
        # Обробляємо дані про час скасування
        revoke_immediately = data.get('revoke_immediately', 'true').lower() == 'true'
        revocation_start_date = None
        revocation_end_date = None
        
        if not revoke_immediately:
            # Якщо не негайно, обробляємо дати скасування
            if data.get('revocation_start_date'):
                revocation_start_date = parse_client_datetime(data.get('revocation_start_date'))
            
            if data.get('revocation_end_date'):
                revocation_end_date = parse_client_datetime(data.get('revocation_end_date'))
        
        # Формуємо додаткові примітки про час скасування
        timing_notes = ""
        if revoke_immediately:
            timing_notes = "Revocation timing: Immediate upon approval. "
        else:
            timing_notes = "Revocation timing: Scheduled. "
            if revocation_start_date:
                timing_notes += f"Start: {revocation_start_date}. "
            else:
                timing_notes += "Start: Immediate upon approval. "
            
            if revocation_end_date:
                timing_notes += f"End: {revocation_end_date}. "
            else:
                timing_notes += "End: Permanent revocation. "
        
        # Встановлюємо дати для revoke запиту
        from django.utils import timezone
        
        # Для revoke запитів start_date завжди має бути встановлена
        if revoke_immediately:
            # Для негайного скасування використовуємо поточну дату як start_date
            request_start_date = timezone.now()
            request_end_date = None  # Негайне скасування не має кінцевої дати
        else:
            # Для запланованого скасування
            request_start_date = revocation_start_date or timezone.now()
            request_end_date = revocation_end_date
        
        # Створюємо Revoke запит (без записів у notes про "request #...")
        revoke_request = AccessRequest.objects.create(
            request_type='revoke',
            company_id=data.get('company_id') or original_request.company_id,
            system_id=data.get('system_id') or original_request.system_id,
            environment=data.get('environment') or original_request.environment,
            requested_by=request.user,
            requested_for_id=original_request.requested_for_id,
            access_record_id=data.get('access_record_id') or original_request.access_record_id,
            justification=data.get('justification', ''),
            requirements=data.get('requirements', ''),
            notes=data.get('notes', ''),
            # Використовуємо поля start_date та end_date для зберігання часу скасування
            start_date=request_start_date,
            end_date=request_end_date,
            # Копіюємо дані третьої сторони якщо є
            third_party_first_name=original_request.third_party_first_name,
            third_party_last_name=original_request.third_party_last_name,
            third_party_email=original_request.third_party_email,
            third_party_phone=original_request.third_party_phone,
            third_party_organization=original_request.third_party_organization,
            third_party_description=original_request.third_party_description,
            third_party_count=original_request.third_party_count,
            third_party_users_data=original_request.third_party_users_data
        )
        
        # Копіюємо зв'язки з ThirdPartyUser якщо є
        if original_request.third_party_users.exists():
            revoke_request.third_party_users.set(original_request.third_party_users.all())
        
        # Обробка файлів
        if attachments:
            max_file_size = 10 * 1024 * 1024  # 10MB
            allowed_types = [
                'application/pdf',
                'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'text/plain',
                'image/jpeg',
                'image/png',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel'
            ]
            
            for attachment in attachments:
                # Перевіряємо розмір файлу
                if attachment.size > max_file_size:
                    raise ValueError(f"File {attachment.name} exceeds maximum size of 10MB")
                
                # Перевіряємо тип файлу
                if attachment.content_type not in allowed_types:
                    raise ValueError(f"File type {attachment.content_type} is not allowed")
                
                # Зберігаємо файл
                AccessRequestAttachment.objects.create(
                    access_request=revoke_request,
                    file=attachment,
                    original_filename=attachment.name,
                    file_size=attachment.size,
                    content_type=attachment.content_type,
                    uploaded_by=request.user
                )
        
        # Створюємо AccessRequestApprover записи базуючись на оригінальному запиті
        if original_request.access_record:
            for approver in original_request.access_record.approvers.all():
                AccessRequestApprover.objects.create(
                    access_request=revoke_request,
                    access_approver=approver,
                    cabinet_user=approver.cabinet_user,
                    order=approver.order,
                    current_status='pending'
                )
        
            logger.debug(f"Created revoke request {revoke_request.id} for original request {original_request_id} by user {request.user.username}")
        
        # Відправляємо email повідомлення
        try:
            send_access_request_notification(revoke_request, recipients_type='all')
            logger.debug(f"Email notifications sent for revoke request {revoke_request.id}")
        except Exception as e:
            logger.error(f"Failed to send email notifications for revoke request {revoke_request.id}: {e}")
        
        return JsonResponse({
            'success': True,
            'message': _('Revoke access request submitted successfully'),
            'request_id': revoke_request.id,
            'original_request_id': original_request_id
        })
        
    except Exception as e:
        logger.error(f"Error handling revoke request: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


def handle_granular_revoke_request(request, data, attachments):
    """Обробка granular revoke запитів з вибраними access records"""
    try:
        import json
        from django.utils import timezone
        
        #
        
        # Парсимо вибрані записи доступу
        selected_records_json = data.get('selected_access_records')
        #
        
        if isinstance(selected_records_json, str):
            selected_records = json.loads(selected_records_json)
        else:
            selected_records = selected_records_json
        
        if not selected_records or len(selected_records) == 0:
            return JsonResponse({
                'success': False,
                'message': _('At least one access record must be selected for revocation')
            }, status=400)
        
        # Отримуємо дані форми
        company_id = data.get('company_id')
        system_id = data.get('system_id')
        environment = data.get('environment')
        access_justification = data.get('access_justification', '')
        additional_requirements = data.get('additional_requirements', '')
        notes = data.get('notes', '')
        revocation_timing = data.get('revocation_timing', 'immediately')
        
        if not all([company_id, system_id, environment, access_justification]):
            return JsonResponse({
                'success': False,
                'message': _('Company, System, Environment, and Access Justification are required')
            }, status=400)
        
        # Обробляємо дати
        start_date = timezone.now()
        end_date = None
        
        if revocation_timing == 'scheduled':
            if data.get('start_date'):
                start_date = parse_client_datetime(data.get('start_date')) or timezone.now()
            if data.get('end_date'):
                end_date = parse_client_datetime(data.get('end_date'))
        
        # Групуємо записи по оригінальним запитам
        records_by_request = {}
        for record_info in selected_records:
            request_id = record_info['request_id']
            record_id = record_info['record_id']
            
            if request_id not in records_by_request:
                records_by_request[request_id] = []
            records_by_request[request_id].append(record_id)
        
        created_revoke_requests = []
        
        with transaction.atomic():
            for original_request_id, record_ids in records_by_request.items():
                # Перевіряємо оригінальний запит
                try:
                    original_request = AccessRequest.objects.get(
                        id=original_request_id,
                        status='approved',
                        admin_status='granted',
                        request_type='grant'
                    )
                except AccessRequest.DoesNotExist:
                    logger.warning(f"Original request {original_request_id} not found or not accessible")
                    continue
                
                # Отримуємо access records для цього запиту
                selected_access_records = original_request.access_records.filter(
                    id__in=record_ids
                )
                
                if not selected_access_records.exists():
                    logger.warning(f"No valid access records found for request {original_request_id}")
                    continue
                
                    logger.debug(f"Found {selected_access_records.count()} access records for request {original_request_id}")
                    logger.debug(f"Record IDs: {list(selected_access_records.values_list('id', flat=True))}")
                
                # Формуємо notes з інформацією про вибрані записи
                records_info = []
                revoked_grant_access_record_ids = []
                
                for record_info in selected_records:
                    if record_info['request_id'] != int(original_request_id):
                        continue  # Skip records not belonging to this request
                        
                    record_id = record_info['record_id']
                    record = selected_access_records.filter(id=record_id).first()
                    if not record:
                        continue
                        
                    object_name = record.access_object.get_name('en') if record.access_object else 'Default Object'
                    roles = [role.get_name('en') or role.name or '' for role in record.roles.all()]
                    records_info.append(f"{object_name}: {', '.join(roles)}")
                    
                    # Використовуємо конкретний Grant Access Record ID, який передав фронтенд
                    frontend_grant_id = record_info.get('grant_access_record_id')
                    if frontend_grant_id:
                        revoked_grant_access_record_ids.append(frontend_grant_id)
                        logger.debug(f"Using frontend-provided Grant Access Record ID: {frontend_grant_id}")
                    else:
                        # Fallback: знаходимо конкретні Grant Access Record ID для цього запису
                        from .models import AccessRequestSequence
                        sequences = AccessRequestSequence.objects.filter(
                            access_record=record,
                            grant_request=original_request,
                            sequence_status='active'
                        )
                        # Initialize to avoid UnboundLocalError when no sequences found
                        grant_access_record_id = None
                        for sequence in sequences:
                            grant_access_record_id = f"{sequence.access_record.id}.{sequence.grant_request.id}.{sequence.order_number}"
                            revoked_grant_access_record_ids.append(grant_access_record_id)
                        if grant_access_record_id:
                            logger.debug(f"Fallback: found Grant Access Record ID: {grant_access_record_id}")
                
                detailed_notes = notes or ''
                
                # Створюємо revoke запит
                revoke_request = AccessRequest.objects.create(
                    request_type='revoke',
                    company_id=company_id,
                    system_id=system_id,
                    environment=environment,
                    requested_by=request.user,
                    requested_for=original_request.requested_for,
                    start_date=start_date,
                    end_date=end_date,
                    justification=access_justification,
                    requirements=additional_requirements,
                    notes=detailed_notes,
                    # Зберігаємо конкретні Grant Access Record ID, які скасовуються
                    revoked_grant_access_record_ids=revoked_grant_access_record_ids,
                    # Копіюємо дані третьої сторони якщо є
                    third_party_first_name=original_request.third_party_first_name,
                    third_party_last_name=original_request.third_party_last_name,
                    third_party_email=original_request.third_party_email,
                    third_party_phone=original_request.third_party_phone,
                    third_party_organization=original_request.third_party_organization,
                    third_party_description=original_request.third_party_description,
                    third_party_count=original_request.third_party_count,
                    third_party_users_data=original_request.third_party_users_data
                )
                
                # Прив'язуємо тільки вибрані access records
                revoke_request.access_records.set(selected_access_records)
                
                # Копіюємо зв'язки з ThirdPartyUser якщо є
                if original_request.third_party_users.exists():
                    revoke_request.third_party_users.set(original_request.third_party_users.all())
                
                # Створюємо AccessRequestApprover записи (уникаємо дублювання)
                created_approvers = set()  # Відстежуємо вже створених approvers
                
                for selected_record in selected_access_records:
                    for approver in selected_record.approvers.all():
                        # Створюємо унікальний ключ для approver
                        approver_key = (approver.cabinet_user.id, approver.order)
                        
                        if approver_key not in created_approvers:
                            AccessRequestApprover.objects.get_or_create(
                                access_request=revoke_request,
                                cabinet_user=approver.cabinet_user,
                                defaults={
                                    'access_approver': approver,
                                    'order': approver.order,
                                    'current_status': 'pending'
                                }
                            )
                            created_approvers.add(approver_key)
                            logger.info(f"Created approver for user {approver.cabinet_user.id} with order {approver.order}")
                
                # Обробка файлів
                if attachments:
                    max_file_size = 10 * 1024 * 1024  # 10MB
                    allowed_types = [
                        'application/pdf',
                        'application/msword',
                        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                        'text/plain',
                        'image/jpeg',
                        'image/png',
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        'application/vnd.ms-excel'
                    ]
                    
                    for attachment in attachments:
                        # Перевіряємо розмір файлу
                        if attachment.size > max_file_size:
                            raise ValueError(f"File {attachment.name} exceeds maximum size of 10MB")
                        
                        # Перевіряємо тип файлу
                        if attachment.content_type not in allowed_types:
                            raise ValueError(f"File type {attachment.content_type} is not allowed")
                        
                        # Зберігаємо файл
                        AccessRequestAttachment.objects.create(
                            access_request=revoke_request,
                            file=attachment,
                            original_filename=attachment.name,
                            file_size=attachment.size,
                            content_type=attachment.content_type,
                            uploaded_by=request.user
                        )
                
                created_revoke_requests.append(revoke_request)
                logger.info(f"Created granular revoke request {revoke_request.id} for {len(selected_access_records)} access records from original request {original_request_id}")
        
        if not created_revoke_requests:
            return JsonResponse({
                'success': False,
                'message': _('No valid revoke requests could be created')
            }, status=400)
        
        # Відправляємо повідомлення для першого створеного запиту (можна розширити для всіх)
        try:
            from .email_utils import send_access_request_notification
            for revoke_request in created_revoke_requests:
                send_access_request_notification(revoke_request, 'submitted')
        except Exception as e:
            logger.error(f"Error sending notification for revoke request: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': _('Revoke request(s) submitted successfully'),
            'created_requests': len(created_revoke_requests),
            'request_ids': [req.id for req in created_revoke_requests]
        })
        
    except Exception as e:
        logger.error(f"Error handling granular revoke request: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def process_approved_grant_request(grant_request):
    """
    Обробка схваленого grant запиту - додавання користувачів до access_users
    """
    try:
        from django.utils import timezone
        
        logger.info(f"Processing approved grant request {grant_request.id} for user {grant_request.requested_for}")
        
        # Отримуємо всі Access Records, пов'язані з цим grant запитом
        access_records = grant_request.access_records.all()
        
        if not access_records.exists():
            logger.warning(f"No access records found for grant request {grant_request.id}")
            return False
        
        updated_records = []
        target_user = grant_request.requested_for
        
        for record in access_records:
            # Перевіряємо, чи користувач вже має доступ до цього запису
            if record.access_users.filter(id=target_user.id).exists():
                logger.info(f"User {target_user} already has access to record {record.id}, skipping")
                continue
            
            # Додаємо користувача до access_users
            record.access_users.add(target_user)
            record.modified_at = timezone.now()
            record.modified_by = grant_request.requested_by
            record.save()
            
            # Створюємо запис в історії змін
            try:
                from .models import SystemAccessStatusHistory
                SystemAccessStatusHistory.objects.create(
                    access_record=record,
                    old_status=record.status,
                    new_status=record.status,  # Статус не змінюється, тільки додається користувач
                    changed_by=grant_request.requested_by,
                    change_reason=f"Access granted to user {target_user.get_full_name()} via request #{grant_request.id}"
                )
                logger.info(f"Created access history record for access record {record.id}")
            except Exception as e:
                logger.error(f"Failed to create access history for access record {record.id}: {str(e)}")
            
            # Створюємо запис AccessRequestSequence для відстеження послідовності
            try:
                from .models import AccessRequestSequence
                # Знаходимо максимальний порядковий номер для цього grant request
                max_order = AccessRequestSequence.objects.filter(
                    grant_request=grant_request
                ).aggregate(models.Max('order_number'))['order_number__max'] or 0
                order_number = max_order + 1
                
                AccessRequestSequence.objects.create(
                    grant_request=grant_request,
                    access_record=record,
                    order_number=order_number,
                    sequence_status='active'
                )
                logger.info(f"Created access sequence record {record.id}.{grant_request.id}.{order_number} for grant request {grant_request.id} and access record {record.id}")
            except Exception as e:
                logger.error(f"Failed to create access sequence for grant request {grant_request.id} and access record {record.id}: {str(e)}")
            
            updated_records.append({
                'record_id': record.id,
                'system': record.asset.name,
                'environment': record.environment,
                'granted_at': timezone.now(),
                'granted_by': grant_request.requested_by.get_full_name(),
                'granted_to': target_user.get_full_name(),
                'grant_request_id': grant_request.id
            })
            
            logger.info(f"Added user {target_user} to access record {record.id}")
        
        # Логуємо результат
        logger.info(f"Successfully processed grant request {grant_request.id} for user {target_user}. Updated {len(updated_records)} access records")
        
        return True
        
    except Exception as e:
        logger.error(f"Error processing approved grant request {grant_request.id}: {str(e)}", exc_info=True)
        return False


def process_approved_revoke_request(revoke_request):
    """
    Обробка схваленого revoke запиту - позначення Access Records як revoked
    ТІЛЬКИ для конкретного користувача (requested_for)
    """
    try:
        from django.utils import timezone
        
        # Визначаємо користувача для логування
        if revoke_request.third_party_email:
            user_info = f"Third Party user {revoke_request.third_party_email}"
        else:
            user_info = f"Cabinet user {revoke_request.requested_for}"
        
        logger.info(f"Processing approved revoke request {revoke_request.id} for {user_info}")
        
        # Отримуємо всі Access Records, пов'язані з цим revoke запитом
        access_records = revoke_request.access_records.all()
        
        if not access_records.exists():
            logger.warning(f"No access records found for revoke request {revoke_request.id}")
            return False
        
        # Знаходимо або створюємо статус "Revoked" для кожної системи
        revoked_statuses = {}
        updated_records = []
        
        # Отримуємо користувача, для якого скасовується доступ
        target_user = revoke_request.requested_for
        target_user_email = None
        
        # Для Third Party користувачів використовуємо email
        if revoke_request.third_party_email:
            target_user_email = revoke_request.third_party_email
            logger.info(f"Processing revoke for Third Party user: {target_user_email}")
        elif target_user:
            target_user_email = target_user.email
            logger.info(f"Processing revoke for Cabinet user: {target_user_email}")
        
        # Отримуємо список конкретних Grant Access Record ID, які скасовуються
        revoked_grant_access_record_ids = revoke_request.revoked_grant_access_record_ids or []
        logger.info(f"Revoking specific Grant Access Record IDs: {revoked_grant_access_record_ids}")
        
        for record in access_records:
            system = record.asset
            
            # ВАЖЛИВО: Перевіряємо, чи цей користувач дійсно має доступ до цього запису
            if target_user and not record.access_users.filter(id=target_user.id).exists():
                logger.warning(f"User {target_user} doesn't have access to record {record.id}, skipping revocation")
                continue
            
            # Знаходимо всі Grant запити для цього користувача з такими ж ролями та об'єктом
            # Це потрібно для того, щоб скасувати доступ тільки в тих записах, які були надані цьому користувачу
            if target_user:
                # Для Cabinet користувачів
                grant_requests = AccessRequest.objects.filter(
                    request_type='grant',
                    requested_for=target_user,
                    system=system,
                    environment=record.environment,
                    admin_status='granted',
                    access_records=record
                )
            else:
                # Для Third Party користувачів
                grant_requests = AccessRequest.objects.filter(
                    request_type='grant',
                    third_party_email=target_user_email,
                    system=system,
                    environment=record.environment,
                    admin_status='granted',
                    access_records=record
                )
            
            if not grant_requests.exists():
                logger.warning(f"No matching grant requests found for user {target_user_email} and record {record.id}, skipping")
                continue
            
            # Перевіряємо, чи є конкретні Grant Access Record ID для скасування
            if revoked_grant_access_record_ids:
                # Знаходимо конкретні AccessRequestSequence записи для скасування
                from .models import AccessRequestSequence
                sequences_to_revoke = AccessRequestSequence.objects.filter(
                    access_record=record,
                    grant_request__in=grant_requests,
                    sequence_status='active'
                )

                # Підтримка формату з четвертою частиною A.B.C.D
                normalized_ids = set()
                for rid in revoked_grant_access_record_ids:
                    parts = str(rid).split('.')
                    key = '.'.join(parts[:3]) if len(parts) >= 3 else rid
                    normalized_ids.add(key)

                # Фільтруємо тільки ті записи, які мають відповідні перші три частини Grant Access Record ID
                sequences_to_revoke = [
                    seq for seq in sequences_to_revoke 
                    if f"{seq.access_record.id}.{seq.grant_request.id}.{seq.order_number}" in normalized_ids
                ]
                
                if not sequences_to_revoke:
                    logger.warning(f"No matching Grant Access Record IDs found for record {record.id}, skipping")
                    continue
                
                logger.info(f"Found {len(sequences_to_revoke)} sequences to revoke for record {record.id}")
            else:
                # Якщо не вказано конкретні ID, скасовуємо всі активні записи
                from .models import AccessRequestSequence
                sequences_to_revoke = AccessRequestSequence.objects.filter(
                    access_record=record,
                    grant_request__in=grant_requests,
                    sequence_status='active'
                )
                logger.info(f"No specific Grant Access Record IDs provided, revoking all {sequences_to_revoke.count()} active sequences for record {record.id}")
            
            # Отримуємо або створюємо статус "Revoked" для цієї системи та environment
            if (system.id, record.environment) not in revoked_statuses:
                revoked_status, created = AccessStatus.objects.get_or_create(
                    system=system,
                    environment=record.environment,
                    name='Revoked',
                    defaults={
                        'description': 'Access revoked by request',
                        'color': '#dc3545',
                        'order': 999
                    }
                )
                revoked_statuses[(system.id, record.environment)] = revoked_status
                
                if created:
                    logger.info(f"Created new 'Revoked' status for system {system.name} in {record.environment}")
            
            # Скасовуємо конкретні sequences (завжди, незалежно від створення статусу)
            for sequence in sequences_to_revoke:
                sequence.revoke_sequence(revoke_request)
                # Оновлюємо sequence_id до розширеного формату A.B.C.D, де D = revoke_request.id
                try:
                    parts = str(sequence.sequence_id).split('.')
                    if len(parts) >= 3:
                        sequence.sequence_id = f"{parts[0]}.{parts[1]}.{parts[2]}.{revoke_request.id}"
                        sequence.save()
                except Exception as e:
                    logger.warning(f"Failed to extend sequence_id for {sequence.sequence_id}: {str(e)}")
                logger.info(f"Revoked sequence {sequence.sequence_id} for record {record.id}")
            
            # Видаляємо користувача з access_users цього запису (тільки для Cabinet користувачів)
            if target_user:
                record.access_users.remove(target_user)
            
            # Оновлюємо статус запису ТІЛЬКИ якщо більше немає користувачів з доступом
            old_status = record.status
            if not record.access_users.exists():
                # Якщо більше немає користувачів з доступом, позначаємо запис як Revoked
                record.status = revoked_statuses[(system.id, record.environment)]
                record.modified_at = timezone.now()
                record.modified_by = revoke_request.requested_by
                record.save()
            else:
                # Якщо ще є користувачі з доступом, залишаємо статус без змін
                logger.info(f"Access record {record.id} still has {record.access_users.count()} users with access, status unchanged")
            
            # Створюємо запис в історії змін статусу Access Record
            try:
                from .models import SystemAccessStatusHistory
                # Формуємо change_reason для різних типів користувачів
                if target_user:
                    change_reason = f"Access revoked for user {target_user.get_full_name()} via request #{revoke_request.id}"
                else:
                    change_reason = f"Access revoked for Third Party user {target_user_email} via request #{revoke_request.id}"
                
                # Додаємо інформацію про конкретні Grant Access Record ID
                if revoked_grant_access_record_ids:
                    revoked_ids_str = ", ".join(revoked_grant_access_record_ids)
                    change_reason += f" (Grant Access Record IDs: {revoked_ids_str})"
                
                SystemAccessStatusHistory.objects.create(
                    access_record=record,
                    old_status=old_status,
                    new_status=record.status,
                    changed_by=revoke_request.requested_by,
                    change_reason=change_reason,
                    revoke_request=revoke_request
                )
                logger.info(f"Created status history record for access record {record.id}")
            except Exception as e:
                logger.error(f"Failed to create status history for access record {record.id}: {str(e)}")
            
            # ВАЖЛИВО: не відкликати будь-яку "active" послідовність за замовчуванням,
            # якщо були вказані конкретні Grant Access Record IDs. Вищe ми вже відкликали
            # лише ті послідовності, які співпали з revoked_grant_access_record_ids або всі, якщо список порожній.
            
            new_status_name = (record.status.name if record.status else None) or 'Unknown'
            updated_records.append({
                'record_id': record.id,
                'old_status': (old_status.name if old_status else None) or 'None',
                'new_status': new_status_name,
                'system': system.name,
                'environment': record.environment,
                'revoked_at': timezone.now(),
                'revoked_by': revoke_request.requested_by.get_full_name(),
                'revoked_for': target_user.get_full_name() if target_user else target_user_email,
                'revoke_request_id': revoke_request.id,
                'remaining_users': record.access_users.count()
            })
            
            if record.status and (record.status.name or '') == 'Revoked':
                logger.info(f"Updated access record {record.id} status from '{old_status}' to 'Revoked' for user {target_user_email} (no more users with access)")
            else:
                logger.info(f"Removed user {target_user_email} from access record {record.id}, status unchanged ({record.access_users.count()} users still have access)")
        
        # Логуємо результат
        logger.info(f"Successfully processed revoke request {revoke_request.id} for user {target_user_email}. Updated {len(updated_records)} access records")
        
        return True
        
    except Exception as e:
        logger.error(f"Error processing approved revoke request {revoke_request.id}: {str(e)}", exc_info=True)
        return False


@login_required
@require_http_methods(['POST'])
def submit_access_request(request):
    """Обробка подання запиту на доступ"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({
            'success': False,
            'error': _("Access denied - you are not authorized to submit access requests.")
        }, status=403)
    
    try:
        from django.utils import translation
        # Перевіряємо, чи це multipart form (з файлами) чи JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData з файлами
            data = {}
            for key in request.POST.keys():
                values = request.POST.getlist(key)
                if len(values) == 1:
                    data[key] = values[0]
                else:
                    data[key] = values
            attachments = request.FILES.getlist('attachments')
        else:
            # JSON без файлів
            data = json.loads(request.body)
            attachments = []
            
        try:
            logger.debug(f"Received access request data: {data}")
        except UnicodeEncodeError:
            logger.debug("Received access request data with special characters")

        # Отримуємо тип запиту (grant або revoke)
        request_type = data.get('request_type', 'grant')
        if request_type not in ['grant', 'revoke']:
            return JsonResponse({
                'success': False,
                'message': _('Invalid request type')
            }, status=400)

        # Спеціальна обробка для Revoke запитів
        if request_type == 'revoke':
            return handle_revoke_request(request, data, attachments)

        # Валідація доступу користувача до Company, Information System та Object
        user_groups = request.user.groups.all()
        
        # 1. Перевіряємо доступ до Company
        company_id = data.get('company_id')
        if not company_id:
            return JsonResponse({
                'success': False,
                'message': _('Company is required')
            }, status=400)
        
        accessible_companies = Company.objects.filter(
            informationasset__access_records__is_active=True
        ).filter(
            Q(informationasset__access_records__request_users=request.user) |
            Q(informationasset__access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_companies.filter(id=company_id).exists():
            logger.warning(f"User {request.user.username} attempted to access company {company_id} without permission")
            return JsonResponse({
                'success': False,
                'message': _('You do not have permission to access this company')
            }, status=403)
        
        # 2. Перевіряємо доступ до Information System
        system_id = data.get('system_id')
        if not system_id:
            return JsonResponse({
                'success': False,
                'message': _('Information System is required')
            }, status=400)
        
        accessible_systems = InformationAsset.objects.filter(
            company_id=company_id,
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_systems.filter(id=system_id).exists():
            logger.warning(f"User {request.user.username} attempted to access system {system_id} without permission")
            return JsonResponse({
                'success': False,
                'message': _('You do not have permission to access this information system')
            }, status=403)
        
        # 3. Перевіряємо доступ до Object (якщо вказано)
        object_id = data.get('object_id')
        if object_id:
            from .models import AccessObjectIS
            accessible_objects = AccessObjectIS.objects.filter(
                asset_id=system_id,
                access_records__is_active=True
            ).filter(
                Q(access_records__request_users=request.user) |
                Q(access_records__request_groups__in=user_groups)
            ).distinct()
            
            if not accessible_objects.filter(id=object_id).exists():
                logger.warning(f"User {request.user.username} attempted to access object {object_id} without permission")
                return JsonResponse({
                    'success': False,
                    'message': _('You do not have permission to access this object')
                }, status=403)
        
        # Перевіряємо чи це третя сторона або отримуємо список користувачів
        user_id = data.get('user_id')
        access_users_raw = data.get('access_users', [])
        
        # Debug logging
        logger.debug(f"Received user data - user_id: {user_id}, access_users_raw: {access_users_raw}")
        logger.debug(f"All data keys: {list(data.keys())}")
        
        # Парсимо access_users якщо це JSON string
        access_users = []
        if access_users_raw:
            if isinstance(access_users_raw, str):
                try:
                    access_users = json.loads(access_users_raw)
                except json.JSONDecodeError:
                    logger.error(f"Failed to parse access_users JSON: {access_users_raw}")
                    return JsonResponse({
                        'success': False,
                        'message': _('Invalid user data format')
                    }, status=400)
            elif isinstance(access_users_raw, list):
                access_users = access_users_raw
        
        # Підтримка старого формату (одного користувача) та нового (кілька користувачів)
        if access_users:
            # Новий формат - кілька користувачів (пріоритет)
            logger.debug(f"Using access_users format: {access_users}")
            is_third_party = False
            # Конвертуємо user IDs в integers
            try:
                requested_users = [int(user_id) for user_id in access_users]
                logger.debug(f"Converted requested_users: {requested_users}")
            except (ValueError, TypeError) as e:
                logger.error(f"Invalid user ID format in access_users: {access_users}")
                return JsonResponse({
                    'success': False,
                    'message': _('Invalid user ID format')
                }, status=400)
        elif user_id:
            # Старий формат - один користувач
            logger.debug(f"Using user_id format: {user_id}")
            if user_id == 'third_parties':
                is_third_party = True
                requested_users = []
            else:
                is_third_party = False
                try:
                    requested_users = [int(user_id)]
                    logger.debug(f"Converted requested_users: {requested_users}")
                except (ValueError, TypeError):
                    logger.error(f"Invalid user_id format: {user_id}")
                    return JsonResponse({
                        'success': False,
                        'message': _('Invalid user ID format')
                    }, status=400)
        else:
            logger.error("No user data found in request")
            return JsonResponse({
                'success': False,
                'message': _('User or users are required')
            }, status=400)
        
        # Для старої сумісності
        if not requested_users and not is_third_party:
            return JsonResponse({
                'success': False,
                'message': _('At least one user must be selected')
            }, status=400)
        
        # 4. Перевіряємо доступ до Access Record
        access_record_ids = data.get('access_record_ids')
        if not access_record_ids:
            access_record_id = data.get('access_record_id')
            if access_record_id:
                access_record_ids = [access_record_id]
            else:
                return JsonResponse({
                    'success': False,
                    'message': _('At least one access record must be selected')
                }, status=400)
        if isinstance(access_record_ids, str):
            # If sent as comma-separated string
            access_record_ids = [rid.strip() for rid in access_record_ids.split(',') if rid.strip()]
        try:
            access_record_ids = [int(rid) for rid in access_record_ids]
        except Exception:
            return JsonResponse({
                'success': False,
                'message': _('Invalid access record IDs')
            }, status=400)
        accessible_records = SystemAccess.objects.filter(id__in=access_record_ids, is_active=True)
        if accessible_records.count() != len(access_record_ids):
            return JsonResponse({
                'success': False,
                'message': _('One or more selected access records are not available')
            }, status=400)
        
        # Для третьої сторони додатково перевіряємо поле third_parties=True
        if is_third_party:
            accessible_records = accessible_records.filter(third_parties=True)
            logger.debug(f"Filtering third party records, found: {accessible_records.count()}")
        
        if not accessible_records.exists():
            error_msg = f"User {request.user.username} attempted to access record {access_record_id} without permission"
            if is_third_party:
                error_msg += " (third party access not allowed for this record)"
            logger.warning(error_msg)
            return JsonResponse({
                'success': False,
                'message': _('You do not have permission to access this access record')
            }, status=403)
        
        # 5. Перевіряємо доступ до користувачів, для яких запитується доступ
        
        if not is_third_party and accessible_records and requested_users:
            # Користувач повинен мати можливість надати доступ вибраним користувачам через access record
            access_record = accessible_records.first()
            for req_user_id in requested_users:
                if not (access_record.access_users.filter(id=req_user_id).exists() or 
                        access_record.access_groups.filter(user__id=req_user_id).exists()):
                    logger.warning(f"User {request.user.username} attempted to request access for user {req_user_id} who is not in access record")
                    return JsonResponse({
                        'success': False,
                        'message': _('One or more selected users are not eligible for this access record')
                    }, status=403)
        
        # 6. Перевіряємо environment
        environment = data.get('environment')
        if not environment:
            return JsonResponse({
                'success': False,
                'message': _('Environment is required')
            }, status=400)
        
        if environment not in ['production', 'test', 'development']:
            return JsonResponse({
                'success': False,
                'message': _('Invalid environment value')
            }, status=400)

        # Parse datetime fields
        from django.utils.dateparse import parse_datetime
        from datetime import datetime
        
        # Parse start_date
        start_date_str = data.get('start_date')
        if start_date_str:
            # If it's in format '2025-06-01T19:48', add seconds
            if len(start_date_str) == 16 and start_date_str.count(':') == 1:
                start_date_str += ':00'
            
            # Parse the datetime
            start_date = parse_datetime(start_date_str)
            if not start_date:
                return JsonResponse({
                    'success': False,
                    'message': _('Invalid start date format')
                }, status=400)
            
            # Make datetime timezone-aware if it's naive
            if start_date.tzinfo is None:
                start_date = timezone.make_aware(start_date)
            
            # Validate that start_date is not in the past for Grant requests
            if request_type == 'grant':
                now = timezone.now()
                if start_date < now:
                    return JsonResponse({
                        'success': False,
                        'message': _('Start date cannot be in the past')
                    }, status=400)
                
                # Validate that start_date is within the access record period
                if accessible_records:
                    access_record = accessible_records.first()
                    if access_record.start_date and start_date < access_record.start_date:
                        return JsonResponse({
                            'success': False,
                            'message': _('Start date cannot be earlier than the access record start date')
                        }, status=400)
                    
                    if access_record.end_date and start_date > access_record.end_date:
                        return JsonResponse({
                            'success': False,
                            'message': _('Start date cannot be later than the access record end date')
                        }, status=400)
        else:
            return JsonResponse({
                'success': False,
                'message': _('Start date is required')
            }, status=400)
        
        # Parse end_date (optional)
        end_date_str = data.get('end_date')
        end_date = None
        if end_date_str and end_date_str.strip():
            # If it's in format '2025-06-01T19:48', add seconds
            if len(end_date_str) == 16 and end_date_str.count(':') == 1:
                end_date_str += ':00'
            
            # Parse the datetime
            end_date = parse_datetime(end_date_str)
            if not end_date:
                return JsonResponse({
                    'success': False,
                    'message': _('Invalid end date format')
                }, status=400)
            
            # Make datetime timezone-aware if it's naive
            if end_date.tzinfo is None:
                end_date = timezone.make_aware(end_date)
            
            # Validate that end_date is in the future for Grant requests
            if request_type == 'grant':
                now = timezone.now()
                if end_date <= now:
                    return JsonResponse({
                        'success': False,
                        'message': _('End date must be in the future')
                    }, status=400)
                
                # Validate that end_date is within the access record period
                if accessible_records:
                    access_record = accessible_records.first()
                    if access_record.end_date and end_date > access_record.end_date:
                        return JsonResponse({
                            'success': False,
                            'message': _('End date cannot be later than the access record end date')
                        }, status=400)

        # Parse third party users data before duplicate checking
        third_party_users_data = []
        if is_third_party:
            third_party_users_json = data.get('third_party_users_json', '[]')
            logger.debug(f"Third party users JSON: {third_party_users_json}")
            if third_party_users_json:
                try:
                    third_party_users_data = json.loads(third_party_users_json)
                    logger.debug(f"Parsed third party users data: {third_party_users_data}")
                    logger.debug(f"Number of third party users to check: {len(third_party_users_data)}")
                except json.JSONDecodeError:
                    third_party_users_data = []

        # Parse requested Object Role per access record (for grant requests from Grant form)
        requested_access_record_roles = []
        if request_type == 'grant':
            arr_json = data.get('access_record_roles', '[]')
            try:
                requested_access_record_roles = json.loads(arr_json) if isinstance(arr_json, str) else (arr_json or [])
            except json.JSONDecodeError:
                requested_access_record_roles = []
            valid_arr = []
            for item in requested_access_record_roles:
                if not isinstance(item, dict) or 'access_record_id' not in item or 'role_id' not in item:
                    continue
                ar_id = item.get('access_record_id')
                role_id = item.get('role_id')
                if ar_id not in access_record_ids:
                    continue
                rec = accessible_records.filter(id=ar_id).first()
                if not rec or not rec.roles.filter(id=role_id).exists():
                    continue
                valid_arr.append({'access_record_id': int(ar_id), 'role_id': int(role_id)})
            requested_access_record_roles = valid_arr

        # CHECK FOR DUPLICATE REQUESTS BEFORE CREATING NEW ONES
        # Prevent duplicate requests for users who already have approved and granted access
        if request_type == 'grant':
            if is_third_party:
                # Check for duplicate third party requests
                # Враховуємо запити з статусом 'pending' та 'approved'
                # ВАЖЛИВО: Перевіряємо дублікати для конкретних третьосторонніх користувачів, а не для всіх запитів користувача
                duplicate_found = False
                existing_request = None
                
                # Збираємо всі дублікати для детального відображення
                all_duplicate_details = []
                
                # Перевіряємо кожного третьостороннього користувача окремо
                logger.debug(f"Starting duplicate check for {len(third_party_users_data)} third party users")
                for i, tp_user_data in enumerate(third_party_users_data):
                    logger.debug(f"Checking third party user {i+1}/{len(third_party_users_data)}: {tp_user_data}")
                    if 'id' in tp_user_data and tp_user_data['id']:
                        # Це існуючий ThirdPartyUser
                        try:
                            third_party_user = ThirdPartyUser.objects.get(id=tp_user_data['id'])
                            # Знаходимо всі існуючі запити для цього third party user
                            # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                            potential_existing_requests = AccessRequest.objects.filter(
                                third_party_users=third_party_user,
                                system_id=system_id,
                                environment=environment,
                                request_type='grant',
                                status__in=['pending', 'approved']
                            ).exclude(admin_status='denied')
                            
                            # Збираємо всі дублікати для детального відображення
                            for existing_request in potential_existing_requests:
                                existing_access_record_ids = set(existing_request.access_records.values_list('id', flat=True))
                                new_access_record_ids = set(access_record_ids)
                                
                                # Знаходимо перетин access records
                                overlapping_records = existing_access_record_ids & new_access_record_ids
                                if overlapping_records:
                                    # Перевіряємо чи є успішний revoke запит для цих overlapping records
                                    # ВАЖЛИВО: Перевіряємо тільки revoke запити, які були створені ПІСЛЯ поточного grant запиту
                                    revoke_requests = AccessRequest.objects.filter(
                                        third_party_users=third_party_user,
                                        system_id=system_id,
                                        environment=environment,
                                        request_type='revoke',
                                        admin_status='granted',  # Успішно відкликано
                                        access_records__id__in=overlapping_records,
                                        created_at__gt=existing_request.created_at  # Revoke запит створений після grant запиту
                                    ).distinct()
                                    
                                    logger.info(f"Checking for revoke requests for user {third_party_user.id} with overlapping records {overlapping_records} created after {existing_request.created_at}: found {revoke_requests.count()} revokes")
                                    
                                    if revoke_requests.exists():
                                        # Є успішний revoke запит, створений після grant запиту
                                        latest_revoke = revoke_requests.order_by('-created_at').first()
                                        logger.info(f"Found successful revoke request #{latest_revoke.id} created at {latest_revoke.created_at} for third party user {third_party_user.id}")
                                        
                                        # Додатково перевіряємо, чи не було створено новий grant запит після revoke
                                        # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                                        newer_grant_requests = AccessRequest.objects.filter(
                                            third_party_users=third_party_user,
                                            system_id=system_id,
                                            environment=environment,
                                            request_type='grant',
                                            status__in=['pending', 'approved'],
                                            access_records__id__in=overlapping_records,
                                            created_at__gt=latest_revoke.created_at  # Grant запит створений після revoke
                                        ).exclude(admin_status='denied').distinct()
                                        
                                        if newer_grant_requests.exists():
                                            logger.info(f"Found newer grant request(s) after revoke for user {third_party_user.id}, treating as active duplicate")
                                            # Є новіший grant запит після revoke - це дублікат
                                        else:
                                            logger.info(f"No newer grant requests found after revoke, skipping this existing request due to revoke")
                                            continue  # Пропускаємо цей існуючий запит, оскільки він був відкликаний і немає новіших grant запитів
                                    
                                    # Збираємо інформацію про кожен дублікат access record
                                    for overlapping_record_id in overlapping_records:
                                        overlapping_record = existing_request.access_records.filter(id=overlapping_record_id).first()
                                        if overlapping_record:
                                            all_duplicate_details.append({
                                                'request': existing_request,
                                                'access_record': overlapping_record,
                                                'overlapping_record_id': overlapping_record_id,
                                                'third_party_user': third_party_user
                                            })
                        except ThirdPartyUser.DoesNotExist:
                            continue
                    else:
                        # Це новий користувач - перевіряємо по email
                        email = tp_user_data.get('email')
                        if email:
                            # Знаходимо всі існуючі запити для цього email
                            # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                            potential_existing_requests = AccessRequest.objects.filter(
                                Q(
                                    Q(third_party_email=email) | 
                                    Q(third_party_users__email=email)
                                ) & 
                                Q(
                                    Q(third_party_users__isnull=False) |
                                    Q(third_party_first_name__isnull=False)
                                ),
                                system_id=system_id,
                                environment=environment,
                                request_type='grant',
                                status__in=['pending', 'approved']
                            ).exclude(admin_status='denied')
                            
                            # Збираємо всі дублікати для детального відображення
                            for existing_request in potential_existing_requests:
                                existing_access_record_ids = set(existing_request.access_records.values_list('id', flat=True))
                                new_access_record_ids = set(access_record_ids)
                                
                                # Знаходимо перетин access records
                                overlapping_records = existing_access_record_ids & new_access_record_ids
                                if overlapping_records:
                                    # Перевіряємо чи є успішний revoke запит для цих overlapping records
                                    # ВАЖЛИВО: Перевіряємо тільки revoke запити, які були створені ПІСЛЯ поточного grant запиту
                                    revoke_requests = AccessRequest.objects.filter(
                                        Q(
                                            Q(third_party_email=email) | 
                                            Q(third_party_users__email=email)
                                        ),
                                        system_id=system_id,
                                        environment=environment,
                                        request_type='revoke',
                                        admin_status='granted',  # Успішно відкликано
                                        access_records__id__in=overlapping_records,
                                        created_at__gt=existing_request.created_at  # Revoke запит створений після grant запиту
                                    ).distinct()
                                    
                                    logger.info(f"Checking for revoke requests for email {email} with overlapping records {overlapping_records} created after {existing_request.created_at}: found {revoke_requests.count()} revokes")
                                    
                                    if revoke_requests.exists():
                                        # Є успішний revoke запит, створений після grant запиту
                                        latest_revoke = revoke_requests.order_by('-created_at').first()
                                        logger.info(f"Found successful revoke request #{latest_revoke.id} created at {latest_revoke.created_at} for third party email {email}")
                                        
                                        # Додатково перевіряємо, чи не було створено новий grant запит після revoke
                                        # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                                        newer_grant_requests = AccessRequest.objects.filter(
                                            Q(
                                                Q(third_party_email=email) | 
                                                Q(third_party_users__email=email)
                                            ),
                                            system_id=system_id,
                                            environment=environment,
                                            request_type='grant',
                                            status__in=['pending', 'approved'],
                                            access_records__id__in=overlapping_records,
                                            created_at__gt=latest_revoke.created_at  # Grant запит створений після revoke
                                        ).exclude(admin_status='denied').distinct()
                                        
                                        if newer_grant_requests.exists():
                                            logger.info(f"Found newer grant request(s) after revoke for email {email}, treating as active duplicate")
                                            # Є новіший grant запит після revoke - це дублікат
                                        else:
                                            logger.info(f"No newer grant requests found after revoke, skipping this existing request due to revoke")
                                            continue  # Пропускаємо цей існуючий запит, оскільки він був відкликаний і немає новіших grant запитів
                                    
                                    # Збираємо інформацію про кожен дублікат access record
                                    for overlapping_record_id in overlapping_records:
                                        overlapping_record = existing_request.access_records.filter(id=overlapping_record_id).first()
                                        if overlapping_record:
                                            all_duplicate_details.append({
                                                'request': existing_request,
                                                'access_record': overlapping_record,
                                                'overlapping_record_id': overlapping_record_id,
                                                'email': email
                                            })
                
                duplicate_found = len(all_duplicate_details) > 0
                
                if duplicate_found:
                    message = _('Access request cannot be created. There is already a pending or approved third party access request for this system and environment.')
                    
                    # Формуємо детальну інформацію про всі дублікати
                    all_duplicate_details_formatted = []
                    
                    if all_duplicate_details:
                        # Отримуємо поточну мову інтерфейсу
                        current_language = translation.get_language()
                        
                        for dup_info in all_duplicate_details:
                            existing_request = dup_info['request']
                            overlapping_record = dup_info['access_record']
                            
                            # Get object name
                            object_name = _('No Object')
                            if overlapping_record and overlapping_record.access_object:
                                if current_language == 'uk':
                                    object_name = overlapping_record.access_object.get_name('uk')
                                elif current_language == 'ru':
                                    object_name = overlapping_record.access_object.get_name('ru')
                                else:
                                    object_name = overlapping_record.access_object.get_name('en')
                        
                            requested_by = existing_request.requested_by.get_full_name() or existing_request.requested_by.username
                            
                            # Third party info - handle both old and new formats
                            if 'third_party_user' in dup_info:
                                # New format - use ThirdPartyUser
                                tp_user = dup_info['third_party_user']
                                third_party_info = f"{tp_user.first_name} {tp_user.last_name}".strip()
                                if tp_user.email:
                                    third_party_info += f" ({tp_user.email})"
                            elif existing_request.third_party_users.exists():
                                # New format - use ThirdPartyUser from request
                                tp_user = existing_request.third_party_users.first()
                                third_party_info = f"{tp_user.first_name} {tp_user.last_name}".strip()
                                if tp_user.email:
                                    third_party_info += f" ({tp_user.email})"
                            else:
                                # Old format - use direct fields
                                third_party_info = f"{existing_request.third_party_first_name} {existing_request.third_party_last_name}".strip()
                                if existing_request.third_party_email:
                                    third_party_info += f" ({existing_request.third_party_email})"
                            
                            # Get roles for this specific access record
                            roles = []
                            if overlapping_record:
                                for role in overlapping_record.roles.all():
                                    if current_language == 'uk':
                                        role_name = role.get_name() or role.name or ''
                                    elif current_language == 'ru':
                                        role_name = role.get_name() or role.name or ''
                                    else:
                                        role_name = role.get_name('en') or role.name or ''
                                    roles.append(role_name)
                            roles_str = ", ".join(roles) if roles else _("No roles")
                            
                            # Get attachments count
                            attachments_count = existing_request.attachments.count() if hasattr(existing_request, 'attachments') else 0
                            attachments_str = f"{attachments_count} files" if attachments_count > 0 else _("No attachments")
                            
                            # Get approving persons
                            approvers = []
                            if hasattr(existing_request, 'request_approvers'):
                                for approver in existing_request.request_approvers.all()[:3]:
                                    approver_name = approver.cabinet_user.user.get_full_name() or approver.cabinet_user.user.username
                                    approvers.append(f"{approver_name} ({approver.current_status})")
                            approvers_str = "; ".join(approvers) if approvers else _("No approvers")
                            if existing_request.request_approvers.count() > 3:
                                approvers_str += f" +{existing_request.request_approvers.count() - 3} more"
                            
                            # Format dates
                            created_date = existing_request.created_at.strftime('%d.%m.%Y %H:%M')
                            period_str = ""
                            # Беремо період з конкретного AccessRecord
                            if overlapping_record:
                                if overlapping_record.start_date and overlapping_record.end_date:
                                    period_str = f"{overlapping_record.start_date.strftime('%d.%m.%Y %H:%M')} - {overlapping_record.end_date.strftime('%d.%m.%Y %H:%M')}"
                                elif overlapping_record.start_date:
                                    period_str = f"{overlapping_record.start_date.strftime('%d.%m.%Y %H:%M')} - Indefinite"
                            # Fallback: якщо немає AccessRecord, перевіряємо AccessRequest
                            elif existing_request.start_date and existing_request.end_date:
                                period_str = f"{existing_request.start_date.strftime('%d.%m.%Y')} - {existing_request.end_date.strftime('%d.%m.%Y')}"
                            
                            # Додаємо інформацію про цей дублікат
                            duplicate_detail = {
                                'request_id': existing_request.id,
                                'object_name': object_name,
                                'requested_by': requested_by,
                                'third_party': third_party_info,
                                'roles': roles_str,
                                'period': period_str or 'Not specified',
                                'attachments': attachments_str,
                                'approvers': approvers_str,
                                'status': existing_request.status,
                                'admin_status': existing_request.admin_status,
                                'created_date': created_date,
                                'access_record_id': overlapping_record.id
                            }
                            all_duplicate_details_formatted.append(duplicate_detail)
                    
                    # Формуємо відповідь з усіма дублікатами
                    duplicate_details = {}
                    if all_duplicate_details_formatted:
                        # Використовуємо перший дублікат для основної інформації (для зворотної сумісності)
                        duplicate_details = all_duplicate_details_formatted[0].copy()
                        # Додаємо список всіх дублікатів
                        duplicate_details['all_duplicates'] = all_duplicate_details_formatted
                    
                    return JsonResponse({
                        'success': False,
                        'message': message.replace('\n', '<br>'),
                        'message_type': 'duplicate_request',
                        'duplicate_details': duplicate_details
                    }, status=400)
            
            elif requested_users:
                # Check for duplicate regular user requests
                # Враховуємо запити з статусом 'pending' та 'approved'
                # Отримуємо поточну мову інтерфейсу
                current_language = translation.get_language()
                duplicate_users = []
                for req_user_id in requested_users:
                    # Знаходимо всі існуючі запити для цього користувача в тій же системі та середовищі
                    # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                    potential_existing_requests = AccessRequest.objects.filter(
                        requested_for_id=req_user_id,
                        system_id=system_id,
                        environment=environment,
                        request_type='grant',
                        status__in=['pending', 'approved']
                    ).exclude(admin_status='denied')
                    
                    # Збираємо всі дублікати для детального відображення
                    duplicate_info = []
                    for existing_request in potential_existing_requests:
                        existing_access_record_ids = set(existing_request.access_records.values_list('id', flat=True))
                        new_access_record_ids = set(access_record_ids)
                        
                        # Знаходимо перетин access records
                        overlapping_records = existing_access_record_ids & new_access_record_ids
                        if overlapping_records:
                            # Перевіряємо чи є успішний revoke запит для цих overlapping records
                            # ВАЖЛИВО: Перевіряємо тільки revoke запити, які були створені ПІСЛЯ поточного grant запиту
                            revoke_requests = AccessRequest.objects.filter(
                                requested_for_id=req_user_id,
                                system_id=system_id,
                                environment=environment,
                                request_type='revoke',
                                admin_status='granted',  # Успішно відкликано
                                access_records__id__in=overlapping_records,
                                created_at__gt=existing_request.created_at  # Revoke запит створений після grant запиту
                            ).distinct()
                            
                            logger.info(f"Checking for revoke requests for cabinet user {req_user_id} with overlapping records {overlapping_records} created after {existing_request.created_at}: found {revoke_requests.count()} revokes")
                            
                            if revoke_requests.exists():
                                # Є успішний revoke запит, створений після grant запиту
                                latest_revoke = revoke_requests.order_by('-created_at').first()
                                logger.info(f"Found successful revoke request #{latest_revoke.id} created at {latest_revoke.created_at} for cabinet user {req_user_id}")
                                
                                # Додатково перевіряємо, чи не було створено новий grant запит після revoke
                                # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                                newer_grant_requests = AccessRequest.objects.filter(
                                    requested_for_id=req_user_id,
                                    system_id=system_id,
                                    environment=environment,
                                    request_type='grant',
                                    status__in=['pending', 'approved'],
                                    access_records__id__in=overlapping_records,
                                    created_at__gt=latest_revoke.created_at  # Grant запит створений після revoke
                                ).exclude(admin_status='denied').distinct()
                                
                                if newer_grant_requests.exists():
                                    logger.info(f"Found newer grant request(s) after revoke for cabinet user {req_user_id}, treating as active duplicate")
                                    # Є новіший grant запит після revoke - це дублікат
                                else:
                                    logger.info(f"No newer grant requests found after revoke, skipping this existing request due to revoke")
                                    continue  # Пропускаємо цей існуючий запит, оскільки він був відкликаний і немає новіших grant запитів
                            
                            # Збираємо інформацію про кожен дублікат access record
                            for overlapping_record_id in overlapping_records:
                                overlapping_record = existing_request.access_records.filter(id=overlapping_record_id).first()
                                if overlapping_record:
                                    duplicate_info.append({
                                        'request': existing_request,
                                        'access_record': overlapping_record,
                                        'overlapping_record_id': overlapping_record_id
                                    })
                    
                    existing_requests = len(duplicate_info) > 0
                    
                    if existing_requests:
                        try:
                            from django.contrib.auth.models import User as AuthUser
                            existing_user = AuthUser.objects.get(id=req_user_id)
                            
                            user_info = existing_user.get_full_name() or existing_user.username
                            
                            # Формуємо детальну інформацію про всі дублікати
                            all_duplicate_details = []
                            for dup_info in duplicate_info:
                                existing_request = dup_info['request']
                                overlapping_record = dup_info['access_record']
                                
                                # Get object name
                                object_name = _('No Object')
                                if overlapping_record and overlapping_record.access_object:
                                    if current_language == 'uk':
                                        object_name = overlapping_record.access_object.get_name('uk')
                                    elif current_language == 'ru':
                                        object_name = overlapping_record.access_object.get_name('ru')
                                    else:
                                        object_name = overlapping_record.access_object.get_name('en')
                                
                                # Get roles for this specific access record
                                roles = []
                                if overlapping_record:
                                    for role in overlapping_record.roles.all():
                                        if current_language == 'uk':
                                            role_name = role.get_name() or role.name or ''
                                        elif current_language == 'ru':
                                            role_name = role.get_name() or role.name or ''
                                        else:
                                            role_name = role.get_name('en') or role.name or ''
                                        roles.append(role_name)
                                roles_str = ", ".join(roles) if roles else _("No roles")
                                
                                requested_by = existing_request.requested_by.get_full_name() or existing_request.requested_by.username
                                requested_for = existing_request.requested_for.get_full_name() or existing_request.requested_for.username
                                
                                # Get attachments count
                                attachments_count = existing_request.attachments.count() if hasattr(existing_request, 'attachments') else 0
                                attachments_str = f"{attachments_count} files" if attachments_count > 0 else _("No attachments")
                                
                                # Get approving persons
                                approvers = []
                                if hasattr(existing_request, 'request_approvers'):
                                    for approver in existing_request.request_approvers.all()[:3]:  # Show first 3 approvers
                                        approver_name = approver.cabinet_user.user.get_full_name() or approver.cabinet_user.user.username
                                        approvers.append(f"{approver_name} ({approver.current_status})")
                                approvers_str = "; ".join(approvers) if approvers else _("No approvers")
                                if existing_request.request_approvers.count() > 3:
                                    approvers_str += f" +{existing_request.request_approvers.count() - 3} more"
                                
                                # Format dates
                                created_date = existing_request.created_at.strftime('%d.%m.%Y %H:%M')
                                period_str = ""
                                # Беремо період з конкретного AccessRecord
                                if overlapping_record:
                                    if overlapping_record.start_date and overlapping_record.end_date:
                                        period_str = f"{overlapping_record.start_date.strftime('%d.%m.%Y %H:%M')} - {overlapping_record.end_date.strftime('%d.%m.%Y %H:%M')}"
                                    elif overlapping_record.start_date:
                                        period_str = f"{overlapping_record.start_date.strftime('%d.%m.%Y %H:%M')} - Indefinite"
                                # Fallback: якщо немає AccessRecord, перевіряємо AccessRequest
                                elif existing_request.start_date and existing_request.end_date:
                                    period_str = f"{existing_request.start_date.strftime('%d.%m.%Y')} - {existing_request.end_date.strftime('%d.%m.%Y')}"
                                
                                # Додаємо інформацію про цей дублікат
                                duplicate_detail = {
                                    'request_id': existing_request.id,
                                    'object_name': object_name,
                                    'requested_by': requested_by,
                                    'requested_for': requested_for,
                                    'roles': roles_str,
                                    'period': period_str or 'Not specified',
                                    'attachments': attachments_str,
                                    'approvers': approvers_str,
                                    'status': existing_request.status,
                                    'admin_status': existing_request.admin_status,
                                    'created_date': created_date,
                                    'access_record_id': overlapping_record.id
                                }
                                all_duplicate_details.append(duplicate_detail)
                            
                            # Додаємо користувача до списку дублікатів тільки один раз
                            if user_info not in duplicate_users:
                                duplicate_users.append(user_info)
                        except AuthUser.DoesNotExist:
                            duplicate_users.append(f"User ID {req_user_id}")
                
                if duplicate_users:
                    users_list = "\n\n".join(duplicate_users)
                    message = _('Access request cannot be created. The following users already have pending or approved access requests for this system and environment:\n\n{}').format(users_list)
                    
                    # Використовуємо зібрані дані про всі дублікати
                    duplicate_details = {}
                    if all_duplicate_details:
                        # Використовуємо перший дублікат для основної інформації (для зворотної сумісності)
                        duplicate_details = all_duplicate_details[0].copy()  # Робимо копію
                        # Додаємо список всіх дублікатів
                        duplicate_details['all_duplicates'] = all_duplicate_details
                    
                    return JsonResponse({
                        'success': False,
                        'message': message.replace('\n', '<br>'),
                        'message_type': 'duplicate_request',
                        'duplicate_details': duplicate_details
                    }, status=400)

        with transaction.atomic():
            created_requests = []
            
            # Using access_record_ids for all users
            
            if is_third_party:
                # Для третьої сторони створюємо окремий запит для кожного користувача
                requested_for_id = request.user.id
                
                # Спочатку створюємо або отримуємо ThirdPartyUser з унікальним email
                created_third_party_users = []
                
                # Обробляємо данні про третіх користувачів
                # third_party_users_data is already parsed above for duplicate checking
                if third_party_users_data:
                    for tp_user_data in third_party_users_data:
                        # Перевіряємо, чи це існуючий користувач
                        existing_user_id = tp_user_data.get('id')  # Changed from 'existing_user_id' to 'id'
                        if existing_user_id:
                            try:
                                # Використовуємо існуючого користувача
                                third_party_user = ThirdPartyUser.objects.get(
                                    id=existing_user_id,
                                    is_active=True
                                )
                                created_third_party_users.append(third_party_user)
                                logger.info(f"Using existing ThirdPartyUser: {third_party_user.email}")
                                continue
                                
                            except ThirdPartyUser.DoesNotExist:
                                logger.error(f"ThirdPartyUser with id {existing_user_id} not found")
                                continue
                        
                        # Якщо це не існуючий користувач, створюємо новий
                        # Перевіряємо чи є email та чи він не порожній
                        email = tp_user_data.get('email', '').strip()
                        if email:
                            # Намагаємося створити або отримати користувача
                            try:
                                # Обробляємо організацію - або знаходимо, або створюємо
                                organization_obj = None
                                organization_name = tp_user_data.get('organization', '').strip()
                                if organization_name:
                                    organization_obj, org_created = ThirdPartyOrganization.objects.get_or_create(
                                        name=organization_name,
                                        defaults={
                                            'created_by': request.user,
                                            'is_active': True
                                        }
                                    )
                                    if org_created:
                                        logger.info(f"Created new ThirdPartyOrganization: {organization_name}")
                                
                                # Спочатку перевіряємо чи існує користувач з таким email
                                try:
                                    third_party_user, created = ThirdPartyUser.objects.get_or_create(
                                        email=email,
                                        defaults={
                                            'first_name': tp_user_data.get('first_name', '').strip(),
                                            'last_name': tp_user_data.get('last_name', '').strip(),
                                            'phone': tp_user_data.get('phone', '').strip(),
                                            'organization': organization_obj,
                                            'organization_name': organization_name,  # Legacy field
                                            'description': tp_user_data.get('description', '').strip(),
                                            'created_by': request.user,
                                            'is_active': True
                                        }
                                    )
                                except Exception as e:
                                    logger.error(f"Error in get_or_create for ThirdPartyUser: {str(e)}")
                                    logger.error(f"Error type: {type(e)}")
                                    logger.error(f"User data: {tp_user_data}")
                                    raise e
                                
                                # Якщо користувач вже існував, оновлюємо його данні якщо потрібно
                                if not created:
                                    try:
                                        updated = False
                                        new_first_name = tp_user_data.get('first_name', '').strip()
                                        new_last_name = tp_user_data.get('last_name', '').strip()
                                        new_phone = tp_user_data.get('phone', '').strip()
                                        new_organization_name = tp_user_data.get('organization', '').strip()
                                        new_description = tp_user_data.get('description', '').strip()
                                        
                                        if new_first_name and third_party_user.first_name != new_first_name:
                                            third_party_user.first_name = new_first_name
                                            updated = True
                                        if new_last_name and third_party_user.last_name != new_last_name:
                                            third_party_user.last_name = new_last_name
                                            updated = True
                                        if new_phone and third_party_user.phone != new_phone:
                                            third_party_user.phone = new_phone
                                            updated = True
                                            
                                        # Оновлюємо організацію
                                        if new_organization_name and (
                                            not third_party_user.organization or 
                                            third_party_user.organization.name != new_organization_name
                                        ):
                                            # Знаходимо або створюємо організацію
                                            new_org_obj, org_created = ThirdPartyOrganization.objects.get_or_create(
                                                name=new_organization_name,
                                                defaults={
                                                    'created_by': request.user,
                                                    'is_active': True
                                                }
                                            )
                                            third_party_user.organization = new_org_obj
                                            third_party_user.organization_name = new_organization_name
                                            updated = True
                                        
                                        if new_description and third_party_user.description != new_description:
                                            third_party_user.description = new_description
                                            updated = True
                                        
                                        if updated:
                                            third_party_user.save()
                                            logger.info(f"Updated ThirdPartyUser {third_party_user.email} information")
                                    except Exception as e:
                                        logger.error(f"Error updating ThirdPartyUser: {str(e)}")
                                        logger.error(f"Error type: {type(e)}")
                                        raise e
                                else:
                                    logger.info(f"Created new ThirdPartyUser: {third_party_user.email}")
                                
                                created_third_party_users.append(third_party_user)
                                
                            except Exception as e:
                                logger.error(f"Error creating/updating ThirdPartyUser with email {email}: {str(e)}")
                                logger.error(f"Error type: {type(e)}")
                                logger.error(f"Error details: {repr(e)}")
                                # Продовжуємо обробку інших користувачів
                                continue
                
                # Створюємо окремий AccessRequest для кожного ThirdPartyUser
                for third_party_user in created_third_party_users:
                    # Знаходимо дані користувача для цього ThirdPartyUser
                    user_data = None
                    for tp_user_data in third_party_users_data:
                        if tp_user_data.get('id') == third_party_user.id or tp_user_data.get('email') == third_party_user.email:
                            user_data = tp_user_data
                            break

                    # Створюємо AccessRequest для цього користувача
                    access_request = AccessRequest.objects.create(
                        request_type=request_type,
                        company_id=company_id,
                        system_id=system_id,
                        environment=data['environment'],
                        requested_by=request.user,
                        requested_for_id=requested_for_id,
                        start_date=start_date,
                        end_date=end_date,
                        justification=data['justification'],
                        requirements=data.get('requirements', ''),
                        notes=data.get('notes', ''),
                        # Поля третьої сторони (для зворотної сумісності)
                        third_party_first_name=third_party_user.first_name,
                        third_party_last_name=third_party_user.last_name,
                        third_party_email=third_party_user.email,
                        third_party_phone=third_party_user.phone or '',
                        third_party_organization=third_party_user.organization_name or (third_party_user.organization.name if third_party_user.organization else ''),
                        third_party_description=third_party_user.description or '',
                        # Поля для кількох третіх сторін
                        third_party_count=1,  # One user per request
                        third_party_users_data=[user_data] if user_data else None
                    )
                    access_request.access_records.set(accessible_records)
                    if requested_access_record_roles:
                        access_request.requested_access_record_roles = requested_access_record_roles
                        access_request.save(update_fields=['requested_access_record_roles'])
                    # Пов'язуємо ThirdPartyUser з AccessRequest
                    access_request.third_party_users.add(third_party_user)
                    logger.info(f"Created AccessRequest {access_request.id} for ThirdPartyUser {third_party_user.email}")
                    created_requests.append(access_request)
            else:
                # Окремий AccessRequest на кожного cabinet-користувача
                for req_user_id in requested_users:
                    access_request = AccessRequest.objects.create(
                        request_type=request_type,
                        company_id=company_id,
                        system_id=system_id,
                        environment=data['environment'],
                        requested_by=request.user,
                        requested_for_id=req_user_id,
                        start_date=start_date,
                        end_date=end_date,
                        justification=data['justification'],
                        requirements=data.get('requirements', ''),
                        notes=data.get('notes', ''),
                        requested_for_count=1,
                        requested_for_users_data=None,
                    )
                    access_request.access_records.set(accessible_records)
                    if requested_access_record_roles:
                        access_request.requested_access_record_roles = requested_access_record_roles
                        access_request.save(update_fields=['requested_access_record_roles'])
                    created_requests.append(access_request)

            # Обробка завантажених файлів для всіх створених запитів
            if attachments:
                max_file_size = 10 * 1024 * 1024  # 10MB
                allowed_types = [
                    'application/pdf',
                    'application/msword',
                    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    'text/plain',
                    'image/jpeg',
                    'image/png',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/vnd.ms-excel'
                ]
                
                for attachment in attachments:
                    # Перевіряємо розмір файлу
                    if attachment.size > max_file_size:
                        raise ValueError(f"File {attachment.name} exceeds maximum size of 10MB")
                    
                    # Перевіряємо тип файлу
                    if attachment.content_type not in allowed_types:
                        raise ValueError(f"File type {attachment.content_type} is not allowed")
                
                # Зберігаємо файли для всіх створених запитів
                for access_request in created_requests:
                    for attachment in attachments:
                        AccessRequestAttachment.objects.create(
                            access_request=access_request,
                            file=attachment,
                            original_filename=attachment.name,
                            file_size=attachment.size,
                            content_type=attachment.content_type,
                            uploaded_by=request.user
                        )

            # Створюємо AccessRequestApprover записи для всіх запитів
            # базуючись на approver'ах з access_record
            source_access_record = accessible_records.first()
            logger.debug(f"Using source access record: {source_access_record.id} for approvers")
            
            for access_request in created_requests:
                for approver in source_access_record.approvers.all():
                    AccessRequestApprover.objects.create(
                        access_request=access_request,
                        access_approver=approver,
                        cabinet_user=approver.cabinet_user,
                        order=approver.order,
                        current_status='pending'
                    )

            logger.info(f"Created {len(created_requests)} access request(s) by user {request.user.username}")

            # Відправляємо email повідомлення для всіх запитів
            for access_request in created_requests:
                try:
                    send_access_request_notification(access_request, recipients_type='all')
                    logger.info(f"Email notifications sent for access request {access_request.id}")
                except Exception as e:
                    logger.error(f"Failed to send email notifications for access request {access_request.id}: {e}")
                    # Не зупиняємо процес через помилку email, тільки логуємо

            # Повертаємо інформацію про створені запити
            request_ids = [req.id for req in created_requests]
            return JsonResponse({
                'success': True,
                'message': _('Access request(s) submitted successfully'),
                'request_ids': request_ids,
                'request_count': len(created_requests)
            })
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': _('Invalid request data format')
        }, status=400)
    except KeyError as e:
        logger.error(f"Missing required field: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': _('Missing required field: {}').format(str(e))
        }, status=400)
    except Exception as e:
        logger.error(f"Error submitting access request: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def get_request_details(request, request_id):
    """Отримання деталей запиту доступу"""
    try:
        current_language = get_language()[:2]
        lang_code = 'ua' if current_language == 'uk' else current_language
        
        access_request = get_object_or_404(AccessRequest, id=request_id)
        
        # Отримуємо ролі
        roles = []
        if access_request.access_record:
            for role in access_request.access_record.roles.all():
                roles.append({
                    'id': role.id,
                    'name': role.get_name() or role.name or '',
                    'color': role.color
                })
        
        # Отримуємо власників
        owners = []
        if access_request.system:
            for owner in access_request.system.owners.all():
                cabinet_user = owner.cabinet_user
                # Отримуємо назву відділу та посади як рядки
                department_name = ''
                position_name = ''
                
                if hasattr(cabinet_user, 'department') and cabinet_user.department:
                    department_name = cabinet_user.department.get_name(current_language)
                
                if hasattr(cabinet_user, 'position') and cabinet_user.position:
                    position_name = cabinet_user.position.get_name(current_language)
                
                owners.append({
                    'name': cabinet_user.user.get_full_name() or cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': department_name,
                    'position': position_name
                })
        
        # Отримуємо адміністраторів
        administrators = []
        if access_request.system:
            for admin in access_request.system.administrators.all():
                cabinet_user = admin.cabinet_user
                # Отримуємо назву відділу та посади як рядки
                department_name = ''
                position_name = ''
                
                if hasattr(cabinet_user, 'department') and cabinet_user.department:
                    department_name = cabinet_user.department.get_name(current_language)
                
                if hasattr(cabinet_user, 'position') and cabinet_user.position:
                    position_name = cabinet_user.position.get_name(current_language)
                
                administrators.append({
                    'name': cabinet_user.user.get_full_name() or cabinet_user.user.username,
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': department_name,
                    'position': position_name
                })
        
        # Отримуємо схвалювачів для цього конкретного запиту
        approvers = []
        for request_approver in access_request.request_approvers.all().order_by('order'):
                cabinet_user = request_approver.cabinet_user
                # Отримуємо назву відділу та посади як рядки
                department_name = ''
                position_name = ''
                
                if hasattr(cabinet_user, 'department') and cabinet_user.department:
                    department_name = cabinet_user.department.get_name(current_language)
                
                if hasattr(cabinet_user, 'position') and cabinet_user.position:
                    position_name = cabinet_user.position.get_name(current_language)
                
                # Отримуємо історію статусу
                status_history = []
                for history in request_approver.get_status_history():
                    status_history.append({
                        'old_status': history.old_status,
                        'new_status': history.new_status,
                        'changed_at': history.changed_at.isoformat(),
                        'changed_by_name': history.changed_by.get_full_name() if history.changed_by else '',
                        'comment': history.comment
                    })

                approvers.append({
                    'id': request_approver.id,  # ID для AccessRequestApprover
                    'name': cabinet_user.user.get_full_name() or cabinet_user.user.username,
                    'email': cabinet_user.user.email if cabinet_user.user.email else '',
                    'avatar': cabinet_user.avatar.url if cabinet_user.avatar else None,
                    'color': cabinet_user.color,
                    'department': department_name,
                    'position': position_name,
                    'order': request_approver.order,
                    'current_status': request_approver.current_status,
                    'status_history': status_history
                })
        
        # Форматуємо дані для відповіді
        request_details = {
            'id': access_request.id,
            'company_name': access_request.company.name if access_request.company else '',
            'system_name': access_request.system.name if access_request.system else '',
            'object_name': access_request.access_record.access_object.get_name(current_language) if access_request.access_record and access_request.access_record.access_object else _('No Object'),
            'environment': access_request.environment,
            'status': access_request.status,
            'admin_status': access_request.admin_status,
            'requested_by': {
                'first_name': access_request.requested_by.first_name,
                'last_name': access_request.requested_by.last_name,
                'email': access_request.requested_by.email,
            },
            'requested_for': {
                'first_name': access_request.requested_for.first_name,
                'last_name': access_request.requested_for.last_name,
                'email': access_request.requested_for.email,
                'name': access_request.requested_for.get_full_name() or access_request.requested_for.username,
                'department': getattr(access_request.requested_for.cabinet.department, 'name', '') if hasattr(access_request.requested_for, 'cabinet') and access_request.requested_for.cabinet and hasattr(access_request.requested_for.cabinet, 'department') else '',
                'position': getattr(access_request.requested_for.cabinet.position, 'name', '') if hasattr(access_request.requested_for, 'cabinet') and access_request.requested_for.cabinet and hasattr(access_request.requested_for.cabinet, 'position') else '',
                'color': getattr(access_request.requested_for.cabinet, 'color', '#cccccc'),
                'avatar': access_request.requested_for.cabinet.avatar.url if hasattr(access_request.requested_for, 'cabinet') and access_request.requested_for.cabinet and access_request.requested_for.cabinet.avatar else None,
            },
            'effective_requested_for': get_effective_requested_for_summary(access_request),
            'is_third_party_request': is_third_party_access_request(access_request),
            'third_party_first_name': access_request.third_party_first_name,
            'third_party_last_name': access_request.third_party_last_name,
            'third_party_email': access_request.third_party_email,
            'roles': roles,
            'record_start_date': access_request.start_date,
            'record_end_date': access_request.end_date,
            'justification': access_request.justification,
            'requirements': access_request.requirements,
            'notes': access_request.notes,
            'created_at': access_request.created_at,
            'owners': owners,
            'administrators': administrators,
            'approvers': approvers
        }
        
        return JsonResponse({
            'success': True,
            'request': request_details
        })
    except Exception as e:
        logger.error(f"Error getting request details: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def get_available_systems(request, company_id):
    """Отримання доступних систем для компанії"""
    try:
        user_groups = request.user.groups.all()
        logger.debug(f"Checking systems for user groups: {[g.name for g in user_groups]}")

        # Отримуємо басейни, де користувач або його групи є в Request Users and Groups
        systems = InformationAsset.objects.filter(
            company_id=company_id,
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__is_active=True
        ).filter(
            # Перевіряємо дату закінчення
            Q(access_records__end_date__gt=timezone.now()) |
            Q(access_records__end_date__isnull=True)
        ).filter(
            # Перевіряємо користувача та групи
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().values(
            'id',
            'name'
        ).order_by('name')

        # Додаємо детальне логування
        for system in systems:
            logger.debug(f"Found system: {system['name']}")

        logger.info(f"Found {len(systems)} available systems for user {request.user} in company {company_id}")
        logger.debug(f"User groups: {list(request.user.groups.values_list('name', flat=True))}")

        # Форматуємо відповідь
        formatted_systems = [{
            'id': system['id'],
            'name': system['name']
        } for system in systems]

        return JsonResponse({
            'status': 'success',
            'systems': formatted_systems
        })
    except Exception as e:
        logger.error(f"Error getting available systems: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_available_objects(request, system_id):
    """Отримання доступних об'єктів для системи на основі прав користувача"""
    try:
        from .models import AccessObjectIS
        
        user_groups = request.user.groups.all()
        logger.debug(f"Checking objects for user groups: {[g.name for g in user_groups]}")
        
        # Отримуємо об'єкти для системи, де користувач або його групи є в Request Users and Groups
        objects = AccessObjectIS.objects.filter(
            asset_id=system_id,
            access_records__is_active=True
        ).filter(
            # Перевіряємо дату закінчення записів доступу
            Q(access_records__end_date__gt=timezone.now()) |
            Q(access_records__end_date__isnull=True)
        ).filter(
            # Перевіряємо користувача та групи в записах доступу
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().order_by('order', 'name')

        logger.debug(f"Found {objects.count()} available objects for user {request.user} in system {system_id}")

        # Форматуємо відповідь
        formatted_objects = []
        for obj in objects:
            formatted_objects.append({
                'id': obj.id,
                'name': obj.get_name(get_language()[:2])
            })

        return JsonResponse({
            'status': 'success',
            'objects': formatted_objects
        })
    except Exception as e:
        logger.error(f"Error getting available objects: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

@login_required
def get_available_users(request, system_id, object_id):
    """Отримання користувачів для вибраної системи та об'єкта, до яких поточний користувач має права запитувати доступ"""
    try:
        user_groups = request.user.groups.all()
        
        # Отримуємо активні записи доступу для системи та об'єкта, 
        # до яких поточний користувач має права запитувати доступ
        active_records = SystemAccess.objects.filter(
            asset_id=system_id,
            access_object_id=object_id,
            is_active=True
        ).filter(
            Q(end_date__gt=timezone.now()) |
            Q(end_date__isnull=True)
        ).filter(
            # Тільки ті записи, до яких поточний користувач може запитувати доступ
            Q(request_users=request.user) |
            Q(request_groups__in=user_groups)
        )

        logger.debug(f"Found {active_records.count()} active records available to user {request.user} for system {system_id} and object {object_id}")

        # Отримуємо користувачів з Access Users та з груп в Access Groups цих записів
        users = User.objects.filter(
            Q(system_access_granted__in=active_records) |  # Користувачі напряму
            Q(groups__system_access_granted__in=active_records)  # Користувачі через групи
        ).distinct().select_related(
            'cabinet',
            'cabinet__department',
            'cabinet__position'
        ).order_by(
            'first_name',
            'last_name'
        )

        logger.debug(f"Found {users.count()} users for system {system_id} and object {object_id}")

        formatted_users = []
        for user in users:
            user_data = {
                'id': user.id,
                'name': f"{user.first_name} {user.last_name}".strip(),
                'avatar': user.cabinet.avatar.url if hasattr(user, 'cabinet') and user.cabinet.avatar else None,
                'color': user.cabinet.color if hasattr(user, 'cabinet') else '#000000'
            }

            if hasattr(user, 'cabinet'):
                if user.cabinet.department:
                    user_data['department'] = user.cabinet.department.get_name(get_language()[:2])
                if user.cabinet.position:
                    user_data['position'] = user.cabinet.position.get_name(get_language()[:2])

            formatted_users.append(user_data)

        logger.debug(f"Formatted users: {formatted_users}")

        return JsonResponse({
            'status': 'success',
            'users': formatted_users
        })
    except Exception as e:
        logger.error(f"Error getting available users: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def get_available_users_by_system(request, system_id):
    """Отримання користувачів для вибраної системи, до яких поточний користувач має права запитувати доступ"""
    try:
        user_groups = request.user.groups.all()
        
        # Отримуємо активні записи доступу для системи, 
        # до яких поточний користувач має права запитувати доступ
        active_records = SystemAccess.objects.filter(
            asset_id=system_id,
            is_active=True
        ).filter(
            Q(end_date__gt=timezone.now()) |
            Q(end_date__isnull=True)
        ).filter(
            # Тільки ті записи, до яких поточний користувач може запитувати доступ
            Q(request_users=request.user) |
            Q(request_groups__in=user_groups)
        )
        
        # Фільтрація за environment, якщо вказано
        environment = request.GET.get('environment')
        if environment:
            active_records = active_records.filter(environment=environment)

        logger.debug(f"Found {active_records.count()} active records available to user {request.user} for system {system_id}")

        # Отримуємо користувачів з Access Users та з груп в Access Groups цих записів
        users = User.objects.filter(
            Q(system_access_granted__in=active_records) |  # Користувачі напряму
            Q(groups__system_access_granted__in=active_records)  # Користувачі через групи
        ).distinct().select_related(
            'cabinet',
            'cabinet__department',
            'cabinet__position'
        ).order_by(
            'first_name',
            'last_name'
        )

        logger.debug(f"Found {users.count()} users for system {system_id}")

        formatted_users = []
        for user in users:
            user_data = {
                'id': user.id,
                'name': f"{user.first_name} {user.last_name}".strip(),
                'avatar': user.cabinet.avatar.url if hasattr(user, 'cabinet') and user.cabinet.avatar else None,
                'color': user.cabinet.color if hasattr(user, 'cabinet') else '#000000'
            }

            if hasattr(user, 'cabinet'):
                if user.cabinet.department:
                    user_data['department'] = user.cabinet.department.get_name(get_language()[:2])
                if user.cabinet.position:
                    user_data['position'] = user.cabinet.position.get_name(get_language()[:2])

            formatted_users.append(user_data)

        logger.debug(f"Formatted users: {formatted_users}")

        return JsonResponse({
            'status': 'success',
            'users': formatted_users
        })
    except Exception as e:
        logger.error(f"Error getting available users by system: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required
def admin_access_requests(request):
    """Об'єднана сторінка для обробки заявок: погодження та адміністрування"""
    # Перевіряємо права доступу через AccessISAM
    user = request.user
    if not user.is_authenticated:
        return redirect('login')
    
    # Перевіряємо дозволи через AccessISAM
    if not has_access_manage_ar_permission(user):
        messages.error(request, _('You do not have permission to access Manage Access Requests page'))
        return redirect('user_access_request')
    
    # Отримуємо дозволи користувача для передачі в контекст
    user_can_add = can_add_manage_ar(user)
    user_can_edit = can_edit_manage_ar(user)  
    user_can_delete = can_delete_manage_ar(user)
    
    # Отримуємо компанії користувача
    user_companies = get_user_companies_for_manage_ar(user)
    
    current_language = get_language()[:2]

    # Отримуємо всі запити з фільтрацією та сортуванням
    requests = AccessRequest.objects.select_related(
        'company',
        'system',
        'requested_for',
        'requested_for__cabinet',
        'requested_for__cabinet__department',
        'requested_for__cabinet__position',
        'requested_by',
        'requested_by__cabinet',
        'requested_by__cabinet__department',
        'requested_by__cabinet__position'
    ).prefetch_related(
        'system__owners__cabinet_user__user',
        'system__administrators__cabinet_user__user',
        'system__approving_persons__cabinet_user__user',
        'access_records__roles',
        'attachments',
        'request_approvers__cabinet_user__user',
        'request_approvers__cabinet_user__department',
        'request_approvers__cabinet_user__position',
        Prefetch('admin_status_history', 
                 queryset=AccessRequestAdminStatusHistory.objects.select_related('changed_by').order_by('-changed_at'))
    ).order_by('-created_at')

    # Фільтрація за статусом
    status_filter = request.GET.get('status')
    if status_filter and status_filter.lower() != 'undefined':
        requests = requests.filter(status=status_filter)

    # Фільтрація за пошуком
    search_query = request.GET.get('search')
    if search_query and search_query.lower() != 'undefined':
        requests = requests.filter(
            Q(id__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(requested_for__first_name__icontains=search_query) |
            Q(requested_for__last_name__icontains=search_query) |
            Q(requested_by__first_name__icontains=search_query) |
            Q(requested_by__last_name__icontains=search_query) |
            Q(access_records__roles__name__icontains=search_query) |
            Q(access_records__roles__name_local__icontains=search_query) |
            Q(third_party_first_name__icontains=search_query) |
            Q(third_party_last_name__icontains=search_query) |
            Q(third_party_email__icontains=search_query) |
            Q(third_party_organization__icontains=search_query)
        ).distinct()

    # Фільтрація за компанією
    company_filter = request.GET.get('company')
    if company_filter and company_filter.lower() != 'undefined':
        requests = requests.filter(company_id=company_filter)

    # Фільтрація за системою
    system_filter = request.GET.get('system')
    if system_filter and system_filter.lower() != 'undefined':
        requests = requests.filter(system_id=system_filter)

    # Фільтрація за датою
    date_filter = request.GET.get('date')
    if date_filter and date_filter.lower() != 'undefined':
        try:
            date = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
            requests = requests.filter(created_at__date=date)
        except ValueError:
            logger.warning(f"Invalid date format: {date_filter}")

    # Фільтрація за Requested By
    requested_by_filter = request.GET.get('requested_by')
    if requested_by_filter and requested_by_filter.lower() != 'undefined':
        requests = requests.filter(requested_by=requested_by_filter)

    # Фільтрація за Requested For
    requested_for_filter = request.GET.get('requested_for')
    if requested_for_filter and requested_for_filter.lower() != 'undefined':
        requests = requests.filter(requested_for=requested_for_filter)

    # Фільтрація за власником
    owner_filter = request.GET.get('owner')
    if owner_filter and owner_filter.lower() != 'undefined':
        requests = requests.filter(system__owners__cabinet_user__user=owner_filter)

    # Фільтрація за адміністратором
    administrator_filter = request.GET.get('administrator')
    if administrator_filter and administrator_filter.lower() != 'undefined':
        requests = requests.filter(system__administrators__cabinet_user__user=administrator_filter)

    # Застосовуємо сортування
    requests = requests.order_by('-created_at')

    # Отримуємо системи з поточних запитів
    systems_in_requests = requests.values_list('system', flat=True).distinct()

    # Фільтрація за Record Period
    record_period = request.GET.get('record_period')
    if record_period and record_period.lower() != 'undefined':
        if record_period == 'active':
            requests = requests.filter(
                Q(access_records__end_date__gt=timezone.now()) | Q(access_records__end_date__isnull=True),
                access_records__start_date__lte=timezone.now()
            )
        elif record_period == 'expired':
            requests = requests.filter(
                access_records__end_date__lte=timezone.now()
            )
        elif record_period == 'future':
            requests = requests.filter(
                access_records__start_date__gt=timezone.now()
            )

    # Фільтрація за User Period
    user_period = request.GET.get('user_period')
    if user_period and user_period.lower() != 'undefined':
        if user_period == 'active':
            requests = requests.filter(
                Q(end_date__gt=timezone.now()) | Q(end_date__isnull=True),
                start_date__lte=timezone.now()
            )
        elif user_period == 'expired':
            requests = requests.filter(
                end_date__lte=timezone.now()
            )
        elif user_period == 'future':
            requests = requests.filter(
                start_date__gt=timezone.now()
            )

    # Отримуємо ролі з поточних запитів
    roles = AccessRoles.objects.filter(
        system_accesses__in=requests.values_list('access_records', flat=True).distinct(),
        is_active=True
    ).distinct().order_by('order', 'name', 'code')

    # Отримуємо окремі списки користувачів
    requested_by_users = User.objects.filter(
        access_requests_by__isnull=False
    ).distinct().order_by('first_name', 'last_name')

    requested_for_users = User.objects.filter(
        access_requests_for__isnull=False
    ).distinct().order_by('first_name', 'last_name')

    # Отримуємо списки власників та адміністраторів
    owners = User.objects.filter(
        cabinet__assetowner__owned_assets__in=systems_in_requests
    ).distinct().order_by('first_name', 'last_name')

    administrators = User.objects.filter(
        cabinet__assetadministrator__administered_assets__in=systems_in_requests
    ).distinct().order_by('first_name', 'last_name')

    # Фільтрація за роллю
    role_filter = request.GET.get('role')
    if role_filter and role_filter.lower() != 'undefined':
        requests = requests.filter(access_records__roles=role_filter)

    # Фільтрація за Environment
    environment_filter = request.GET.get('environment')
    if environment_filter and environment_filter.lower() != 'undefined':
        requests = requests.filter(environment=environment_filter)

    # Фільтрація за Approving Status
    approving_status_filter = request.GET.get('approving_status')
    if approving_status_filter and approving_status_filter.lower() != 'undefined':
        requests = requests.filter(request_approvers__current_status=approving_status_filter).distinct()

    # Фільтрація за Admin Status
    admin_status_filter = request.GET.get('admin_status')
    if admin_status_filter and admin_status_filter.lower() != 'undefined':
        requests = requests.filter(admin_status=admin_status_filter)

    # Отримуємо унікальні значення environment з поточних запитів
    environments_in_requests = requests.values_list('environment', flat=True).distinct()
    
    # Фільтруємо ENVIRONMENT_CHOICES, залишаючи тільки ті, що є в запитах
    filtered_environment_choices = [
        choice for choice in AccessRequest.ENVIRONMENT_CHOICES 
        if choice[0] in environments_in_requests
    ]

    # Пагінація (за замовчуванням 25 записів на сторінку)
    page = request.GET.get('page', 1)
    page_size = get_access_table_page_size(request)
    paginator = Paginator(requests, page_size)
    try:
        requests_page = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        requests_page = paginator.page(1)

    # Визначаємо права поточного користувача
    user_can_edit_status = False
    user_is_owner = False
    user_is_administrator = False
    user_is_approver = False

    if hasattr(user, 'cabinet') and user.cabinet:
        try:
            user_is_owner = user.cabinet.assetowner_set.exists()
            user_is_administrator = user.cabinet.assetadministrator_set.exists()
            user_is_approver = user.cabinet.accessapprover_set.exists()
        except Exception as e:
            logger.warning(f"Error checking detailed user permissions for {user.username}: {e}")

    # Форматуємо дані для шаблону
    formatted_requests = []
    
    for req in requests_page:
        # Визначаємо чи може користувач адмініструвати цей запит
        user_can_admin = False
        if hasattr(user, 'cabinet') and user.cabinet:
            user_can_admin = req.system.administrators.filter(cabinet_user=user.cabinet).exists()
        
        # Для revoke запитів отримуємо інформацію про оригінальний запит
        original_request_id = None
        original_request = None
        if req.request_type == 'revoke' and req.notes:
            # Витягуємо ID оригінального запиту з notes
            import re
            match = re.search(r'request #(\d+)', req.notes)
            if match:
                original_request_id = int(match.group(1))
                try:
                    original_request = AccessRequest.objects.get(id=original_request_id)
                except AccessRequest.DoesNotExist:
                    original_request = None
        
        # Обробляємо revocation timing для revoke запитів
        revocation_timing = None
        is_immediate_revocation = False
        if req.request_type == 'revoke':
            # Перевіряємо чи це негайне скасування
            if req.notes and 'Revocation timing: Immediate' in req.notes:
                is_immediate_revocation = True
                revocation_timing = _('Immediately')
            elif req.start_date:
                # Запланована дата скасування
                if req.end_date:
                    revocation_timing = f"{req.start_date.strftime('%d.%m.%Y %H:%M')} - {req.end_date.strftime('%d.%m.%Y %H:%M')}"
                else:
                    revocation_timing = f"{req.start_date.strftime('%d.%m.%Y %H:%M')} - {_('Permanent')}"
            else:
                revocation_timing = _('Immediately')
                is_immediate_revocation = True
        
        # Отримуємо існуючі granted ролі користувача для цього Object та Environment
        existing_granted_roles = []
        if req.access_records.exists():
            try:
                # Для кабінетних користувачів
                if req.requested_for:
                    # Спочатку знаходимо всі схвалені та надані доступи користувача для цієї системи
                    existing_requests = AccessRequest.objects.filter(
                        requested_for=req.requested_for,
                        system=req.system,
                        status='approved',
                        admin_status='granted',
                        request_type='grant'  # Тільки grant запити
                    ).exclude(id=req.id).prefetch_related('access_records__roles')
                    
                    # Збираємо всі granted ролі користувача в цій системі
                    for existing_req in existing_requests:
                        for access_record in existing_req.access_records.all():
                            for role in access_record.roles.all():
                                role_data = {
                                    'id': role.id,
                                    'color': role.color or '#6c757d',
                                    'request_id': existing_req.id,
                                    'start_date': existing_req.start_date,
                                    'end_date': existing_req.end_date,
                                    'environment': existing_req.environment
                                }
                                
                                # Додаємо назву ролі залежно від мови
                                if current_language == 'uk':
                                    role_data['name'] = role.get_name() or role.name or ''
                                elif current_language == 'ru':
                                    role_data['name'] = role.get_name() or role.name or ''
                                else:
                                    role_data['name'] = role.get_name('en') or role.name or ''
                                
                                existing_granted_roles.append(role_data)
                    
                # Для третьосторонніх користувачів
                elif req.third_party_first_name or req.third_party_last_name:
                    # Знаходимо всі схвалені та надані доступи третьосторонніх користувачів з такими ж іменем, email та організацією
                    filter_conditions = {
                        'system': req.system,
                        'status': 'approved',
                        'admin_status': 'granted',
                        'request_type': 'grant'
                    }
                    
                    # Додаємо умови для третьосторонніх користувачів
                    if req.third_party_first_name:
                        filter_conditions['third_party_first_name'] = req.third_party_first_name
                    if req.third_party_last_name:
                        filter_conditions['third_party_last_name'] = req.third_party_last_name
                    if req.third_party_email:
                        filter_conditions['third_party_email'] = req.third_party_email
                    
                    existing_requests = AccessRequest.objects.filter(**filter_conditions).exclude(id=req.id).prefetch_related('access_records__roles')
                    
                    # Збираємо всі ролі з цих запитів
                    for existing_req in existing_requests:
                        for access_record in existing_req.access_records.all():
                            for role in access_record.roles.all():
                                role_data = {
                                    'id': role.id,
                                    'color': role.color or '#6c757d',
                                    'request_id': existing_req.id,
                                    'start_date': existing_req.start_date,
                                    'end_date': existing_req.end_date,
                                    'environment': existing_req.environment
                                }
                                
                                # Додаємо назву ролі залежно від мови
                                if current_language == 'uk':
                                    role_data['name'] = role.get_name() or role.name or ''
                                elif current_language == 'ru':
                                    role_data['name'] = role.get_name() or role.name or ''
                                else:
                                    role_data['name'] = role.get_name('en') or role.name or ''
                                
                                existing_granted_roles.append(role_data)
                else:
                    existing_requests = AccessRequest.objects.none()
                

                    
            except Exception as e:
                logger.warning(f"Error getting existing granted roles for request {req.id}: {e}")
        
        # Додаємо інформацію про оригінальний grant запит для кожного access record
        access_records_with_original_grant = []
        for access_record in req.access_records.all():
            # Знаходимо оригінальний grant запит для цього access record
            original_grant_request = None
            # Шукаємо grant запит, який створив цей access record (перший за часом)
            original_grant_request = AccessRequest.objects.filter(
                access_records=access_record,
                request_type='grant',
                status='approved',
                admin_status='granted'
            ).order_by('created_at').first()
            
            # Removed verbose debug logs
            
            # Знаходимо revoke запит для цього access record через AccessRequestSequence
            revoke_request = None
            if access_record.status and (access_record.status.name or '') == 'Revoked':
                try:
                    from .models import AccessRequestSequence
                    sequence_record = AccessRequestSequence.objects.filter(
                        access_record=access_record,
                        sequence_status='revoked',
                        revoke_request__isnull=False
                    ).order_by('-revoked_at').first()
                    
                    if sequence_record:
                        revoke_request = sequence_record.revoke_request
                    else:
                        # Fallback до SystemAccessStatusHistory якщо запис послідовності не знайдено
                        revoke_history_entry = SystemAccessStatusHistory.objects.filter(
                            access_record=access_record,
                            new_status__name='Revoked',
                            revoke_request__isnull=False
                        ).order_by('-changed_at').first()
                        
                        if revoke_history_entry:
                            revoke_request = revoke_history_entry.revoke_request
                            logger.debug(f"DEBUG: Admin View - Found revoke request #{revoke_request.id} for record {access_record.id} via SystemAccessStatusHistory (fallback)")
                except Exception as e:
                    logger.error(f"Error finding revoke request for record {access_record.id}: {str(e)}")
            
            access_records_with_original_grant.append({
                'access_record': access_record,
                'original_grant_request': original_grant_request,
                'revoke_request': revoke_request
            })
        
        # Побудова короткого тексту історії погоджень для тултіпа статусу запиту (як у My Access Requests)
        approval_history_text = None
        try:
            events = []
            # 1) Поточні request approvers зі своєю історією
            if req.request_approvers.exists():
                for ra in req.request_approvers.all().order_by('order'):
                    history_qs = getattr(ra, 'status_history', None)
                    if history_qs is not None:
                        for h in history_qs.all().order_by('changed_at'):
                            approver_name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                            changer_name = h.changed_by.get_full_name() if h.changed_by else ''
                            ts = timezone.localtime(h.changed_at).strftime('%d.%m.%Y %H:%M')
                            events.append((h.changed_at, ra.order, f"Lvl {ra.order}: {approver_name} → {h.new_status.title()} ({ts} by {changer_name})"))
                    else:
                        approver_name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                        events.append((None, ra.order, f"Lvl {ra.order}: {approver_name} - {ra.current_status.title()}"))

            # 2) Snapshot-історія, прив'язана до запиту (на випадок заміни/видалення погоджувачів)
            try:
                from .models import AccessRequestApproverStatusHistory as ARASH
                snapshot_qs = ARASH.objects.filter(access_request=req).order_by('changed_at')
                for h in snapshot_qs:
                    approver_name = h.approver_name or (h.approver_cabinet_user.user.get_full_name() if h.approver_cabinet_user and getattr(h.approver_cabinet_user, 'user', None) else '')
                    changer_name = h.changed_by.get_full_name() if h.changed_by else ''
                    ts = timezone.localtime(h.changed_at).strftime('%d.%m.%Y %H:%M')
                    order_num = getattr(h, 'order_at_change', None) or 0
                    events.append((h.changed_at, order_num, f"Lvl {order_num}: {approver_name} → {h.new_status.title()} ({ts} by {changer_name})"))
            except Exception:
                pass

            if events:
                events.sort(key=lambda x: (x[0] is None, x[0] or 0, x[1] or 0))
                approval_history_text = " | ".join([e[2] for e in events])
            else:
                # Фолбек: якщо немає історії, покажемо поточний стан кожного погоджувача
                if req.request_approvers.exists():
                    lines = []
                    for ra in req.request_approvers.all().order_by('order'):
                        name = ra.cabinet_user.user.get_full_name() or ra.cabinet_user.user.username
                        lines.append(f"Lvl {ra.order}: {name} - {ra.current_status.title()}")
                    if lines:
                        approval_history_text = " | ".join(lines)
        except Exception:
            approval_history_text = None

        formatted_requests.append({
            'id': req.id,
            'company': req.company,
            'system': req.system,
            'access_records': req.access_records,
            'access_records_with_original_grant': access_records_with_original_grant,
            'requested_by': req.requested_by,
            'requested_for': req.requested_for,
            'request_type': req.request_type,
            'status': req.status,
            'admin_status': req.admin_status,
            'environment': req.environment,
            'start_date': req.start_date,
            'end_date': req.end_date,
            'justification': req.justification,
            'requirements': req.requirements,
            'notes': req.notes,
            'created_at': req.created_at,
            'attachments': req.attachments,
            'request_approvers': req.request_approvers,
            'admin_status_history': req.admin_status_history.all(),
            'approval_history_text': approval_history_text,
            # Поля для перевірки закінчення періоду
            'is_period_expired': req.is_period_expired,
            'can_be_approved': req.can_be_approved,
            # Поля третьої сторони
            'third_party_first_name': req.third_party_first_name,
            'third_party_last_name': req.third_party_last_name,
            'third_party_email': req.third_party_email,
            'third_party_phone': req.third_party_phone,
            'third_party_organization': req.third_party_organization,
            'third_party_description': req.third_party_description,
            # Права користувача для цього запиту
            'user_can_admin': user_can_admin,
            # Додаємо інформацію про оригінальний запит для revoke запитів
            'original_request_id': original_request_id,
            'original_request': original_request,
            # Додаємо інформацію про revocation timing
            'revocation_timing': revocation_timing,
            'is_immediate_revocation': is_immediate_revocation,
            # Додаємо інформацію про існуючі granted ролі
            'existing_granted_roles': existing_granted_roles,
        })

    context = {
        'requests': formatted_requests,
        'companies': Company.objects.all(),
        'systems': InformationAsset.objects.filter(company_id=company_filter, access_manage=True, deletion_date__isnull=True) if company_filter else [],
        'requested_by_users': requested_by_users,
        'requested_for_users': requested_for_users,
        'owners': owners,
        'administrators': administrators,
        'roles': roles,
        'environment_choices': filtered_environment_choices,
        'admin_status_choices': AccessRequest.ADMIN_STATUS_CHOICES,
        'current_language': current_language,
        'user_can_edit_status': user_can_edit_status,
        'user_is_owner': user_is_owner,
        'user_is_administrator': user_is_administrator,
        'user_is_approver': user_is_approver,
        'current_user': user,
        # Access permissions from AccessISAM
        'can_add_manage_ar': user_can_add,
        'can_edit_manage_ar': user_can_edit,
        'can_delete_manage_ar': user_can_delete,
        'user_companies': user_companies,
        # Пагінація
        'paginator': paginator,
        'page_obj': requests_page,
        'is_paginated': paginator.count > 0,
        'current_page_size': page_size,
        'page_size_options': ACCESS_TABLE_PAGE_SIZE_OPTIONS,
    }
    
    return render(request, 'app_access/manage_access_requests.html', context)


@login_required
def revoke_access_form(request):
    """Відображення окремої форми для скасування доступу"""
    # Перевіряємо чи користувач може подавати запити доступу
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        messages.error(request, _("Access denied - you are not authorized to submit access requests. Please contact your administrator to be added to the request list for at least one system."))
        return redirect('index')
    
    # Автоматично перевіряємо та оновлюємо записи з минулими датами
    expired_count = check_and_update_expired_records()
    if expired_count > 0 and settings.DEBUG is True and False:
        # Suppressed noisy info log even in DEBUG
        logger.info(f"Automatically set {expired_count} expired records to inactive before loading revoke access form")
    
    current_language = get_language()[:2]
    
    user_groups = request.user.groups.all()
    
    # Отримуємо тільки компанії, в яких є системи доступні користувачу
    companies = Company.objects.filter(
        informationasset__access_manage=True, # Only include assets marked for access management
        informationasset__deletion_date__isnull=True, # Only include active assets
        informationasset__access_records__is_active=True
    ).filter(
        Q(informationasset__access_records__request_users=request.user) |
        Q(informationasset__access_records__request_groups__in=user_groups)
    ).distinct().order_by('name')
    
    return render(request, 'app_access/revoke_access_form.html', {
        'companies': companies,
        'current_language': current_language,
        'access_justification_templates': _serialize_access_justification_templates(),
    })


@login_required
@require_http_methods(["GET"])
def get_access_justification_templates(request):
    company_id = request.GET.get("company_id")
    try:
        templates = _serialize_access_justification_templates(company_id=company_id)
        return JsonResponse({
            "status": "success",
            "templates": templates,
        })
    except Exception as exc:
        logger.exception("Failed to load access justification templates: %s", exc)
        return JsonResponse({
            "status": "error",
            "message": _("Error loading access justification templates"),
        }, status=500)


@login_required
def get_approved_access_requests(request, company_id, system_id, environment, user_type, user_id):
    """Отримання схвалених запитів доступу для скасування"""
    try:
        # Визначаємо поточну мову більш точно
        current_language = getattr(request, 'LANGUAGE_CODE', get_language())
        if current_language:
            current_language = current_language[:2]
        else:
            current_language = 'en'  # fallback
        
        # Принудово активуємо мову для цього запиту
        if current_language in ['en', 'uk', 'ru']:
            activate(current_language)
        
        # Suppressed noisy debug prints
        
        user_groups = request.user.groups.all()
        
        # Отримуємо додаткові параметри для третіх сторін
        organization_id = request.GET.get('organization_id', 'all')
        third_party_user_id = request.GET.get('third_party_user_id', 'all')
        
        # Базова перевірка доступу до компанії та системи
        accessible_companies = Company.objects.filter(
            informationasset__access_manage=True, # Only include assets marked for access management
            informationasset__deletion_date__isnull=True, # Only include active assets
            informationasset__access_records__is_active=True
        ).filter(
            Q(informationasset__access_records__request_users=request.user) |
            Q(informationasset__access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_companies.filter(id=company_id).exists():
            return JsonResponse({
                'status': 'error',
                'message': _('You do not have permission to access this company')
            }, status=403)
        
        accessible_systems = InformationAsset.objects.filter(
            company_id=company_id,
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_systems.filter(id=system_id).exists():
            return JsonResponse({
                'status': 'error',
                'message': _('You do not have permission to access this system')
            }, status=403)
        
        # Фільтруємо схвалені запити доступу
        approved_requests_query = AccessRequest.objects.filter(
            company_id=company_id,
            system_id=system_id,
            environment=environment,
            status='approved',
            admin_status='granted',
            request_type='grant'
        )
        
        # Фільтруємо за типом користувача
        if user_type == 'third_parties':
            # Для третіх сторін показуємо запити з third_party_users
            approved_requests_query = approved_requests_query.filter(
                third_party_users__isnull=False
            ).distinct()
            
            # Додаткові фільтри для третіх сторін
            if organization_id != 'all':
                try:
                    org_id = int(organization_id)
                    approved_requests_query = approved_requests_query.filter(
                        third_party_users__organization_id=org_id
                    )
                except (ValueError, TypeError):
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Invalid organization ID format')
                    }, status=400)
            
            if third_party_user_id != 'all':
                try:
                    tp_user_id = int(third_party_user_id)
                    approved_requests_query = approved_requests_query.filter(
                        third_party_users__id=tp_user_id
                    )
                except (ValueError, TypeError):
                    return JsonResponse({
                        'status': 'error',
                        'message': _('Invalid third party user ID format')
                    }, status=400)
            
        elif user_type == 'cabinet_users' and user_id != 'all':
            # Для конкретного кабінетного користувача
            try:
                user_id_int = int(user_id)
                approved_requests_query = approved_requests_query.filter(
                    requested_for_id=user_id_int,
                    third_party_users__isnull=True  # Виключаємо записи третіх сторін
                )
            except (ValueError, TypeError):
                return JsonResponse({
                    'status': 'error',
                    'message': _('Invalid user ID format')
                }, status=400)
        elif user_type == 'cabinet_users' and user_id == 'all':
            # Для всіх кабінетних користувачів
            approved_requests_query = approved_requests_query.filter(
                requested_for__isnull=False,
                third_party_users__isnull=True
            )
        
        # Перевіряємо чи немає активних запитів на скасування
        approved_requests_query = approved_requests_query.exclude(
            id__in=AccessRequest.objects.filter(
                request_type='revoke',
                status__in=['pending', 'approved'],
                notes__regex=r'request #\d+'
            ).values_list('notes', flat=True)
        )
        
        approved_requests = approved_requests_query.select_related(
            'company',
            'system',
            'requested_by',
            'requested_for',
            'requested_for__cabinet',
            'requested_for__cabinet__department',
            'requested_for__cabinet__position'
        ).prefetch_related(
            'access_records',
            'access_records__roles',
            'access_records__access_object',
            'third_party_users'
        )
        
        formatted_requests = []
        for req in approved_requests:
            # Форматуємо інформацію про користувача
            if is_third_party_access_request(req):
                user_info = []
                for tp_user in req.third_party_users.all():
                    user_info.append({
                        'type': 'third_party',
                        'name': f"{tp_user.first_name} {tp_user.last_name}",
                        'email': tp_user.email,
                        'organization': tp_user.organization.name if tp_user.organization else tp_user.organization_name
                    })
                if not user_info and (req.third_party_first_name or req.third_party_last_name):
                    user_info.append({
                        'type': 'third_party',
                        'name': f"{req.third_party_first_name} {req.third_party_last_name}".strip(),
                        'email': req.third_party_email,
                        'organization': req.third_party_organization or '',
                    })
            else:
                cabinet_user = req.requested_for.cabinet if hasattr(req.requested_for, 'cabinet') else None
                user_info = [{
                    'type': 'cabinet',
                    'name': req.requested_for.get_full_name(),
                    'email': req.requested_for.email,
                    'department': cabinet_user.department.get_name(current_language) if cabinet_user and cabinet_user.department else None,
                    'position': cabinet_user.position.get_name(current_language) if cabinet_user and cabinet_user.position else None
                }]
            
            # Requested Object Role per record (from grant request)
            requested_role_by_record = {}
            if getattr(req, 'requested_access_record_roles', None) and isinstance(req.requested_access_record_roles, list):
                for item in req.requested_access_record_roles:
                    if isinstance(item, dict) and 'access_record_id' in item and 'role_id' in item:
                        requested_role_by_record[item['access_record_id']] = item['role_id']

            # Форматуємо записи доступу
            access_records_info = []
            for record in req.access_records.all():
                roles_info = []
                for role in record.roles.all():
                    # Suppressed noisy debug prints
                    
                    # Використовуємо метод get_name для отримання локалізованої назви
                    role_name = role.get_name(current_language)
                    
                    # Остаточний fallback якщо назва порожня
                    if not role_name:
                        role_name = f"Role #{role.id}"
                    
                    # Suppressed noisy debug prints
                    
                    roles_info.append({
                        'id': role.id,
                        'name': role_name,
                        'color': role.color or '#6c757d',
                        'name': role.get_name() or role.name or ''
                    })
                
                # Визначаємо оригінальний grant-запит для цього запису (для відображення довідкової інформації)
                original_grant_request = AccessRequest.objects.filter(
                    access_records=record,
                    request_type='grant',
                    status='approved',
                    admin_status='granted'
                ).order_by('created_at').first()
                
                # Обчислюємо статус відкликання ТІЛЬКИ для поточного Grant Access Record (послідовність цього запиту)
                from .models import AccessRequestSequence
                revoke_request = None
                grant_access_record_id = None
                order_number = None
                is_revoked = False

                try:
                    sequence_record = AccessRequestSequence.objects.filter(
                        access_record=record,
                        grant_request=req
                    ).order_by('order_number').first()
                    
                    if sequence_record:
                        # Для відображення нормалізуємо ID до перших трьох частин
                        grant_access_record_id = '.'.join(str(sequence_record.sequence_id).split('.')[:3])
                        order_number = sequence_record.order_number
                        # Якщо саме цей Grant Access Record вже відкликано (і адміністратором підтверджено), показуємо відповідний revoke_request
                        if (
                            sequence_record.sequence_status == 'revoked'
                            and sequence_record.revoke_request
                            and sequence_record.revoke_request.admin_status == 'granted'
                        ):
                            is_revoked = True
                            revoke_request = sequence_record.revoke_request
                            logger.debug(f"DEBUG: Sequence {grant_access_record_id} is revoked by request #{revoke_request.id}")
                    else:
                        # Додаткова перевірка за sequence_id (на випадок розсинхронізації)
                        from .models import AccessRequestSequence as ARS
                        revoked_seq_by_id = ARS.objects.filter(
                            sequence_id__startswith=f"{record.id}.{req.id}.",
                            sequence_status='revoked',
                            revoke_request__isnull=False,
                            revoke_request__admin_status='granted'
                        ).order_by('-revoked_at').first()
                        if revoked_seq_by_id:
                            is_revoked = True
                            revoke_request = revoked_seq_by_id.revoke_request
                            grant_access_record_id = '.'.join(str(revoked_seq_by_id.sequence_id).split('.')[:3])
                            logger.debug(f"DEBUG: Sequence {grant_access_record_id} resolved revoked via direct lookup, request #{revoke_request.id}")
                    
                    if not grant_access_record_id:
                        # Фолбек ID якщо запис послідовності не знайдено
                        grant_access_record_id = f"{record.id}.{req.id}.1"
                        order_number = 1

                    # Враховуємо лише відклики, де Admin Status = Access Revoked (granted)
                    if not is_revoked and grant_access_record_id:
                        existing_rr = AccessRequest.objects.filter(
                            request_type='revoke',
                            admin_status='granted',
                            access_records=record,
                            revoked_grant_access_record_ids__contains=[grant_access_record_id]
                        ).order_by('-created_at').first()
                        if existing_rr:
                            is_revoked = True
                            revoke_request = existing_rr

                    # Видаляємо глобальний фолбек за статусом запису.
                    # is_revoked визначається виключно за AccessRequestSequence для конкретного grant-запиту
                except Exception as e:
                    logger.error(f"Error resolving sequence/revoke info for record {record.id}, request {req.id}: {str(e)}")
                
                # Додаємо повний ID A.B.C.D для відображення
                full_id = None
                try:
                    if grant_access_record_id and revoke_request and is_revoked:
                        full_id = f"{grant_access_record_id}.{revoke_request.id}"
                    elif grant_access_record_id:
                        full_id = f"{grant_access_record_id}.0"
                except Exception:
                    full_id = grant_access_record_id

                requested_role_info = None
                rid = requested_role_by_record.get(record.id)
                if rid:
                    try:
                        req_role = AccessRoles.objects.get(id=rid)
                        requested_role_info = {
                            'id': req_role.id,
                            'name': req_role.get_name(current_language) or req_role.name or f"Role #{req_role.id}",
                            'color': req_role.color or '#6c757d'
                        }
                    except AccessRoles.DoesNotExist:
                        pass

                access_records_info.append({
                    'id': record.id,
                    'object_name': record.access_object.get_name(current_language) if record.access_object else None,
                    'object_color': (record.access_object.color if getattr(record, 'access_object', None) and getattr(record.access_object, 'color', None) else '#495057'),
                    'roles': roles_info,
                    'requested_role': requested_role_info,
                    'start_date': record.start_date.isoformat() if record.start_date else None,
                    'end_date': record.end_date.isoformat() if record.end_date else None,
                    'status': (record.status.name if record.status else None),
                    'is_revoked': is_revoked,
                    'grant_access_record_id': grant_access_record_id,
                    'grant_access_record_id_full': full_id,
                    'order_number': order_number,
                    'original_grant_request': {
                        'id': original_grant_request.id if original_grant_request else None,
                        'created_at': original_grant_request.created_at.isoformat() if original_grant_request else None,
                        'requested_by': original_grant_request.requested_by.get_full_name() if original_grant_request and original_grant_request.requested_by else None
                    } if original_grant_request else None,
                    'revoke_request': {
                        'id': revoke_request.id if revoke_request else None,
                        'created_at': revoke_request.created_at.isoformat() if revoke_request else None,
                        'requested_by': revoke_request.requested_by.get_full_name() if revoke_request and revoke_request.requested_by else None
                    } if revoke_request else None
                })
            
            formatted_requests.append({
                'id': req.id,
                'company_name': req.company.name,
                'system_name': req.system.name,
                'environment': req.environment,
                'environment_display': str(dict(AccessRequest.ENVIRONMENT_CHOICES).get(req.environment, req.environment)),
                'users': user_info,
                'access_records': access_records_info,
                'created_at': req.created_at.isoformat(),
                'start_date': req.start_date.isoformat() if req.start_date else None,
                'end_date': req.end_date.isoformat() if req.end_date else None,
                'justification': req.justification
            })
        
        return JsonResponse({
            'status': 'success',
            'requests': formatted_requests
        })
        
    except Exception as e:
        logger.error(f"Error getting approved access requests: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
def get_third_party_organizations(request, company_id, system_id, environment):
    """Отримання організацій третіх сторін для вибраної компанії, системи та середовища"""
    try:
        current_language = get_language()[:2]
        user_groups = request.user.groups.all()
        
        # Перевіряємо доступ до компанії та системи
        accessible_companies = Company.objects.filter(
            informationasset__access_records__is_active=True
        ).filter(
            Q(informationasset__access_records__request_users=request.user) |
            Q(informationasset__access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_companies.filter(id=company_id).exists():
            return JsonResponse({
                'status': 'error',
                'message': _('You do not have permission to access this company')
            }, status=403)
        
        accessible_systems = InformationAsset.objects.filter(
            company_id=company_id,
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_systems.filter(id=system_id).exists():
            return JsonResponse({
                'status': 'error',
                'message': _('You do not have permission to access this system')
            }, status=403)
        
        # Отримуємо організації з approved запитів третіх сторін
        # Спочатку знаходимо approved запити третіх сторін
        approved_tp_requests = AccessRequest.objects.filter(
            company_id=company_id,
            system_id=system_id,
            environment=environment,
            status='approved',
            admin_status='granted',
            request_type='grant',
            third_party_users__isnull=False
        ).distinct()
        
        # Отримуємо ID організацій з цих запитів
        org_ids = set()
        for request_obj in approved_tp_requests:
            for tp_user in request_obj.third_party_users.all():
                if tp_user.organization_id:
                    org_ids.add(tp_user.organization_id)
        
        # Отримуємо організації за ID
        organizations_query = ThirdPartyOrganization.objects.filter(
            id__in=org_ids,
            is_active=True
        ).order_by('name')
        
        organizations = []
        for org in organizations_query:
            organizations.append({
                'id': org.id,
                'name': org.name,
                'description': org.description or ''
            })
        
        return JsonResponse({
            'status': 'success',
            'organizations': organizations
        })
        
    except Exception as e:
        logger.error(f"Error getting third party organizations: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
def get_third_party_users_by_org(request, company_id, system_id, environment, organization_id):
    """Отримання користувачів третіх сторін за організацією"""
    try:
        current_language = get_language()[:2]
        user_groups = request.user.groups.all()
        
        # Перевіряємо доступ до компанії та системи
        accessible_companies = Company.objects.filter(
            informationasset__access_records__is_active=True
        ).filter(
            Q(informationasset__access_records__request_users=request.user) |
            Q(informationasset__access_records__request_groups__in=user_groups)
        ).distinct()
        
        if not accessible_companies.filter(id=company_id).exists():
            return JsonResponse({
                'status': 'error',
                'message': _('You do not have permission to access this company')
            }, status=403)
        
        # Спочатку знаходимо approved запити третіх сторін
        approved_tp_requests = AccessRequest.objects.filter(
            company_id=company_id,
            system_id=system_id,
            environment=environment,
            status='approved',
            admin_status='granted',
            request_type='grant',
            third_party_users__isnull=False
        ).distinct()
        
        # Отримуємо ID користувачів з цих запитів
        user_ids = set()
        for request_obj in approved_tp_requests:
            for tp_user in request_obj.third_party_users.all():
                user_ids.add(tp_user.id)
        
        # Базовий запит для користувачів третіх сторін
        users_query = ThirdPartyUser.objects.filter(
            id__in=user_ids,
            is_active=True
        )
        
        # Фільтруємо по організації
        if organization_id != 'all':
            try:
                org_id = int(organization_id)
                users_query = users_query.filter(organization_id=org_id)
            except (ValueError, TypeError):
                return JsonResponse({
                    'status': 'error',
                    'message': _('Invalid organization ID format')
                }, status=400)
        
        users_query = users_query.distinct().order_by('first_name', 'last_name')
        
        users = []
        for user in users_query:
            users.append({
                'id': user.id,
                'name': f"{user.first_name} {user.last_name}",
                'email': user.email,
                'organization_name': user.organization.name if user.organization else user.organization_name,
                'phone': user.phone or ''
            })
        
        return JsonResponse({
            'status': 'success',
            'users': users
        })
        
    except Exception as e:
        logger.error(f"Error getting third party users by organization: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
def get_user_current_access(request, system_id, object_id, user_id):
    """Отримання поточних доступів користувача для скасування"""
    try:
        from .models import AccessObjectIS
        
        # Отримуємо environment з GET параметрів
        environment = request.GET.get('environment')
        
        # Отримуємо користувача та об'єкт
        user = get_object_or_404(User, id=user_id)
        access_object = get_object_or_404(AccessObjectIS, id=object_id)
        system = get_object_or_404(InformationAsset, id=system_id)
        current_language = get_language()[:2]
        
        # Знаходимо поточні активні запити доступу для користувача
        current_requests = AccessRequest.objects.filter(
            requested_for=user,
            system=system,
            access_record__access_object=access_object,
            status='approved'  # Тільки схвалені запити можна скасувати
        )
        
        # Фільтруємо по environment, якщо параметр передано
        if environment:
            logger.info(f"Filtering current access by environment: {environment}")
            current_requests = current_requests.filter(environment=environment)
        else:
            logger.info("Loading all current access without environment filter")
            
        current_requests = current_requests.select_related(
            'access_record',
            'access_record__access_object',
            'system'
        ).prefetch_related(
            'access_record__roles',
            'system__owners__cabinet_user__user',
            'system__administrators__cabinet_user__user',
            'request_approvers__cabinet_user__user'
        )
        
        current_access = []
        for req in current_requests:
            access_record = req.access_record
            
            # Формуємо дані про поточний доступ
            access_data = {
                'id': access_record.id,
                'request_id': req.id,  # ID запиту, який надав цей доступ
                'environment': req.environment,
                'start_date': access_record.start_date.isoformat() if access_record.start_date else None,
                'end_date': access_record.end_date.isoformat() if access_record.end_date else None,
                'asset_id': system.id,
                'object_id': access_object.id,
                'roles': [],
                'owners': [],
                'administrators': [],
                'approvers': []
            }
            
            # Додаємо ролі
            for role in access_record.roles.all():
                role_data = {
                    'id': role.id,
                    'color': role.color or '#6c757d'
                }
                
                if current_language == 'uk':
                    role_data['name'] = role.get_name() or role.name or ''
                elif current_language == 'ru':
                    role_data['name'] = role.get_name() or role.name or ''
                else:
                    role_data['name'] = role.get_name('en') or role.name or ''
                    
                access_data['roles'].append(role_data)
            
            # Додаємо власників з системи (не з access_record)
            for owner in system.owners.all():
                access_data['owners'].append({
                    'name': owner.cabinet_user.user.get_full_name() or owner.cabinet_user.user.username,
                    'email': owner.cabinet_user.user.email,
                    'department': (owner.cabinet_user.department.get_name(current_language) if owner.cabinet_user.department else ''),
                    'position': (owner.cabinet_user.position.get_name(current_language) if owner.cabinet_user.position else '')
                })
            
            # Додаємо адміністраторів з системи (не з access_record)
            for admin in system.administrators.all():
                access_data['administrators'].append({
                    'name': admin.cabinet_user.user.get_full_name() or admin.cabinet_user.user.username,
                    'email': admin.cabinet_user.user.email,
                    'department': (admin.cabinet_user.department.get_name(current_language) if admin.cabinet_user.department else ''),
                    'position': (admin.cabinet_user.position.get_name(current_language) if admin.cabinet_user.position else '')
                })
            
            # Додаємо approvers з поточного запиту
            for approver in req.request_approvers.all().order_by('order'):
                approver_data = {
                    'order': approver.order,
                    'name': approver.cabinet_user.user.get_full_name() or approver.cabinet_user.user.username,
                    'email': approver.cabinet_user.user.email,
                    'current_status': approver.current_status,
                    'department': (approver.cabinet_user.department.get_name(current_language) if approver.cabinet_user.department else ''),
                    'position': (approver.cabinet_user.position.get_name(current_language) if approver.cabinet_user.position else ''),
                    'status_history': []
                }
                
                # Додаємо історію статусів
                for history in approver.get_status_history():
                    approver_data['status_history'].append({
                        'old_status': history.old_status,
                        'new_status': history.new_status,
                        'changed_at': history.changed_at.isoformat(),
                        'changed_by_name': history.changed_by.get_full_name() if history.changed_by else '',
                        'comment': history.comment
                    })
                
                access_data['approvers'].append(approver_data)
            
            current_access.append(access_data)
        
        return JsonResponse({
            'status': 'success',
            'current_access': current_access,
            'message': f'Found {len(current_access)} current access records for user'
        })
        
    except Exception as e:
        logger.error(f"Error getting user current access: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required
def cancel_access_request(request, request_id):
    """Скасування заявки користувачем"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': _('Only POST method allowed')
        }, status=405)

    try:
        access_request = get_object_or_404(AccessRequest, id=request_id)
        
        # Перевіряємо, чи може користувач скасувати заявку
        if not access_request.can_be_cancelled(request.user):
            return JsonResponse({
                'success': False,
                'message': _('Request cannot be cancelled')
            }, status=400)
        
        # Отримуємо причину скасування з запиту
        reason = request.POST.get('reason', '')
        
        # Скасовуємо заявку
        access_request.cancel_request(request.user, reason)
        
        return JsonResponse({
            'success': True,
            'message': _('Request cancelled successfully')
        })
        
    except Exception as e:
        logger.error(f"Error cancelling request {request_id}: {e}")
        return JsonResponse({
            'success': False,
            'message': _('Error cancelling request')
        }, status=500)


def _validate_grant_access_record_permissions(request, company_id, system_id, object_id, access_record_ids):
    """Shared validation for grant submit/edit: company, system, object, access records."""
    user_groups = request.user.groups.all()

    if not company_id:
        return None, _('Company is required')
    accessible_companies = Company.objects.filter(
        informationasset__access_records__is_active=True
    ).filter(
        Q(informationasset__access_records__request_users=request.user) |
        Q(informationasset__access_records__request_groups__in=user_groups)
    ).distinct()
    if not accessible_companies.filter(id=company_id).exists():
        return None, _('You do not have permission to access this company')

    if not system_id:
        return None, _('Information System is required')
    accessible_systems = InformationAsset.objects.filter(
        company_id=company_id,
        access_manage=True,
        deletion_date__isnull=True,
        access_records__is_active=True
    ).filter(
        Q(access_records__request_users=request.user) |
        Q(access_records__request_groups__in=user_groups)
    ).distinct()
    if not accessible_systems.filter(id=system_id).exists():
        return None, _('You do not have permission to access this information system')

    if object_id:
        from .models import AccessObjectIS
        accessible_objects = AccessObjectIS.objects.filter(
            asset_id=system_id,
            access_records__is_active=True
        ).filter(
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct()
        if not accessible_objects.filter(id=object_id).exists():
            return None, _('You do not have permission to access this object')

    if not access_record_ids:
        return None, _('At least one access record must be selected')
    if isinstance(access_record_ids, str):
        access_record_ids = [rid.strip() for rid in access_record_ids.split(',') if rid.strip()]
    try:
        access_record_ids = [int(rid) for rid in access_record_ids]
    except (TypeError, ValueError):
        return None, _('Invalid access record IDs')

    accessible_records = SystemAccess.objects.filter(id__in=access_record_ids, is_active=True)
    if accessible_records.count() != len(access_record_ids):
        return None, _('One or more selected access records are not available')

    return accessible_records, None


def _parse_grant_request_post_data(request):
    """Parse multipart or JSON body for grant submit/edit."""
    if request.content_type and 'multipart/form-data' in request.content_type:
        data = {}
        for key in request.POST.keys():
            values = request.POST.getlist(key)
            data[key] = values[0] if len(values) == 1 else values
        attachments = request.FILES.getlist('attachments')
    else:
        data = json.loads(request.body)
        attachments = []
    return data, attachments


def _sync_request_approvers_from_access_record(access_request, source_access_record):
    """Rebuild pending approvers for a request from the source access record."""
    access_request.request_approvers.all().delete()
    for approver in source_access_record.approvers.all():
        AccessRequestApprover.objects.create(
            access_request=access_request,
            access_approver=approver,
            cabinet_user=approver.cabinet_user,
            order=approver.order,
            current_status='pending'
        )


@login_required
@require_http_methods(['GET'])
def get_access_request_edit_data(request, request_id):
    """Дані grant-заявки для форми редагування (тільки Requested By, до першого погодження)."""
    access_request = get_object_or_404(AccessRequest, id=request_id)
    if not access_request.can_be_edited(request.user):
        return JsonResponse({
            'success': False,
            'can_be_edited': False,
            'message': _('This request cannot be edited')
        }, status=403)

    requested_role_by_record = {}
    if access_request.requested_access_record_roles and isinstance(access_request.requested_access_record_roles, list):
        for item in access_request.requested_access_record_roles:
            if isinstance(item, dict) and 'access_record_id' in item and 'role_id' in item:
                requested_role_by_record[item['access_record_id']] = item['role_id']

    access_record_roles = []
    for record in access_request.access_records.all():
        role_id = requested_role_by_record.get(record.id)
        if not role_id:
            first_role = record.roles.filter(is_active=True).first() or record.roles.first()
            if first_role:
                role_id = first_role.id
        if role_id:
            access_record_roles.append({
                'access_record_id': record.id,
                'role_id': role_id,
            })

    attachments = []
    for attachment in access_request.attachments.all():
        attachments.append({
            'id': attachment.id,
            'original_filename': attachment.original_filename,
            'file_size': attachment.file_size,
            'url': attachment.file.url if attachment.file else '',
            'icon_class': attachment.get_file_icon_class(),
            'file_size_display': attachment.get_file_size_display(),
        })

    is_third_party = is_third_party_access_request(access_request)
    third_party_user_ids = list(access_request.third_party_users.values_list('id', flat=True))

    return JsonResponse({
        'success': True,
        'can_be_edited': True,
        'request': {
            'id': access_request.id,
            'company_id': access_request.company_id,
            'company_name': access_request.company.name if access_request.company else '',
            'system_id': access_request.system_id,
            'system_name': access_request.system.name if access_request.system else '',
            'environment': access_request.environment,
            'start_date': access_request.start_date.isoformat() if access_request.start_date else None,
            'end_date': access_request.end_date.isoformat() if access_request.end_date else None,
            'justification': access_request.justification,
            'requirements': access_request.requirements or '',
            'notes': access_request.notes or '',
            'access_record_ids': list(access_request.access_records.values_list('id', flat=True)),
            'access_record_roles': access_record_roles,
            'is_third_party': is_third_party,
            'requested_for_id': access_request.requested_for_id,
            'third_party_user_ids': third_party_user_ids,
            'attachments': attachments,
        }
    })


@login_required
@require_http_methods(['POST'])
def edit_access_request(request, request_id):
    """Редагування grant-заявки Requested By до першого погодження."""
    from .matrix_view import can_submit_access_requests
    if not can_submit_access_requests(request.user):
        return JsonResponse({
            'success': False,
            'message': _("Access denied - you are not authorized to submit access requests.")
        }, status=403)

    access_request = get_object_or_404(AccessRequest, id=request_id)
    if not access_request.can_be_edited(request.user):
        return JsonResponse({
            'success': False,
            'message': _('This request cannot be edited')
        }, status=403)

    try:
        data, attachments = _parse_grant_request_post_data(request)

        company_id = access_request.company_id
        system_id = access_request.system_id

        environment = data.get('environment')
        if not environment or environment not in ['production', 'test', 'development']:
            return JsonResponse({
                'success': False,
                'message': _('Invalid environment value')
            }, status=400)

        access_record_ids = data.get('access_record_ids')
        if not access_record_ids:
            access_record_id = data.get('access_record_id')
            access_record_ids = [access_record_id] if access_record_id else []
        accessible_records, perm_error = _validate_grant_access_record_permissions(
            request, company_id, system_id, data.get('object_id'), access_record_ids
        )
        if perm_error:
            return JsonResponse({'success': False, 'message': perm_error}, status=403 if 'permission' in perm_error.lower() else 400)

        if is_third_party_access_request(access_request):
            accessible_records = accessible_records.filter(third_parties=True)
            if not accessible_records.exists():
                return JsonResponse({
                    'success': False,
                    'message': _('You do not have permission to access this access record')
                }, status=403)

        start_date = parse_client_datetime(data.get('start_date'))
        if not start_date:
            return JsonResponse({
                'success': False,
                'message': _('Invalid start date format')
            }, status=400)

        now = timezone.now()
        original_start = access_request.start_date
        start_changed = (
            not original_start
            or abs((start_date - original_start).total_seconds()) > 1
        )
        if start_date < now and start_changed:
            return JsonResponse({
                'success': False,
                'message': _('Start date cannot be in the past')
            }, status=400)

        end_date = parse_client_datetime(data.get('end_date'))
        original_end = access_request.end_date
        end_changed = True
        if end_date and original_end:
            end_changed = abs((end_date - original_end).total_seconds()) > 1
        elif not end_date and not original_end:
            end_changed = False
        if end_date and end_date <= now and end_changed:
            return JsonResponse({
                'success': False,
                'message': _('End date must be in the future')
            }, status=400)

        if end_date and end_date < start_date:
            return JsonResponse({
                'success': False,
                'message': _('End date cannot be earlier than start date')
            }, status=400)

        justification = (data.get('justification') or '').strip()
        if not justification:
            return JsonResponse({
                'success': False,
                'message': _('Justification is required')
            }, status=400)

        requested_access_record_roles = []
        arr_json = data.get('access_record_roles', '[]')
        try:
            requested_access_record_roles = json.loads(arr_json) if isinstance(arr_json, str) else (arr_json or [])
        except json.JSONDecodeError:
            requested_access_record_roles = []
        valid_arr = []
        record_id_set = set(accessible_records.values_list('id', flat=True))
        for item in requested_access_record_roles:
            if not isinstance(item, dict) or 'access_record_id' not in item or 'role_id' not in item:
                continue
            ar_id = int(item['access_record_id'])
            role_id = int(item['role_id'])
            if ar_id not in record_id_set:
                continue
            rec = accessible_records.filter(id=ar_id).first()
            if not rec or not rec.roles.filter(id=role_id).exists():
                continue
            valid_arr.append({'access_record_id': ar_id, 'role_id': role_id})
        requested_access_record_roles = valid_arr

        deleted_attachment_ids = []
        deleted_raw = data.get('deleted_attachment_ids', '[]')
        try:
            deleted_attachment_ids = json.loads(deleted_raw) if isinstance(deleted_raw, str) else (deleted_raw or [])
            deleted_attachment_ids = [int(i) for i in deleted_attachment_ids]
        except (TypeError, ValueError, json.JSONDecodeError):
            deleted_attachment_ids = []

        with transaction.atomic():
            access_request.environment = environment
            access_request.start_date = start_date
            access_request.end_date = end_date
            access_request.justification = justification
            access_request.requirements = data.get('requirements', '') or ''
            access_request.notes = data.get('notes', '') or ''
            access_request.requested_access_record_roles = requested_access_record_roles or None
            access_request.save()

            records_changed = set(access_request.access_records.values_list('id', flat=True)) != record_id_set
            access_request.access_records.set(accessible_records)
            if records_changed:
                source_access_record = accessible_records.first()
                if source_access_record:
                    _sync_request_approvers_from_access_record(access_request, source_access_record)

            if deleted_attachment_ids:
                for attachment in access_request.attachments.filter(id__in=deleted_attachment_ids):
                    if attachment.file:
                        attachment.file.delete(save=False)
                    attachment.delete()

            if attachments:
                max_file_size = 10 * 1024 * 1024
                allowed_types = [
                    'application/pdf',
                    'application/msword',
                    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                    'text/plain',
                    'image/jpeg',
                    'image/png',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    'application/vnd.ms-excel'
                ]
                for attachment in attachments:
                    if attachment.size > max_file_size:
                        raise ValueError(f"File {attachment.name} exceeds maximum size of 10MB")
                    if attachment.content_type not in allowed_types:
                        raise ValueError(f"File type {attachment.content_type} is not allowed")
                    AccessRequestAttachment.objects.create(
                        access_request=access_request,
                        file=attachment,
                        original_filename=attachment.name,
                        file_size=attachment.size,
                        content_type=attachment.content_type,
                        uploaded_by=request.user
                    )

        logger.info(f"Access request {access_request.id} edited by user {request.user.username}")
        return JsonResponse({
            'success': True,
            'message': _('Access request updated successfully'),
            'request_id': access_request.id,
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': _('Invalid request data format')
        }, status=400)
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
    except Exception as e:
        logger.error(f"Error editing access request {request_id}: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': _('Error updating access request')
        }, status=500)


@login_required
def get_approver_history(request, approver_id):
    """Отримання історії змін статусу approver"""
    try:
        try:
            approver = AccessRequestApprover.objects.get(id=approver_id)
        except AccessRequestApprover.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': _('Approver not found')
            }, status=404)
        
        # Перевіряємо права доступу: 
        # 1. Адміністратор може бачити все
        # 2. Самий approver може бачити свою історію
        # 3. Користувач може бачити історію запитів, пов'язаних з ним (requested_by, request_users, access_users)
        can_view = (
            request.user.is_staff or  # Адміністратор
            (hasattr(request.user, 'cabinet') and 
             request.user.cabinet == approver.cabinet_user) or  # Самий approver
            # Користувач, який зробив запит
            approver.access_request.requested_by == request.user or
            approver.access_request.requested_for == request.user
        )
        
        if not can_view:
            return JsonResponse({
                'success': False,
                'message': _('Permission denied')
            }, status=403)
        
        # Формуємо історію на основі бази даних
        history = []
        
        # Додаємо базовий запис про призначення approver'а
        history.append({
            'status': 'assigned',
            'timestamp': approver.created_at.isoformat(),
            'user': approver.access_request.requested_by.get_full_name(),
            'comment': f'Assigned to approval level {approver.order}'
        })
        
        # Отримуємо історію змін статусу з бази даних
        status_history = approver.get_status_history()
        
        for history_record in status_history:
            history.append({
                'status': history_record.new_status,
                'old_status': history_record.old_status,
                'timestamp': history_record.changed_at.isoformat(),
                'user': history_record.changed_by.get_full_name(),
                'comment': history_record.comment or f'Status changed from {history_record.old_status} to {history_record.new_status}'
            })
        
        # Якщо історії немає, але є поточний статус, додаємо його
        if not status_history and approver.current_status != 'pending':
            history.append({
                'status': approver.current_status,
                'old_status': 'pending',
                'timestamp': approver.status_changed_at.isoformat() if approver.status_changed_at else '',
                'user': approver.status_changed_by.get_full_name() if approver.status_changed_by else 'System',
                'comment': approver.status_comment or f'Status: {approver.current_status}'
            })
        
        return JsonResponse({
            'success': True,
            'history': history
        })
    except Exception as e:
        logger.error(f"Error getting approver history: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

@login_required
def get_company_systems_for_filter(request, company_id):
    """Отримання списку систем для фільтрації"""
    try:
        logger.info(f"Getting systems for company ID: {company_id}")
        systems = InformationAsset.objects.filter(
            company_id=company_id,
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True,  # Only include active assets
            access_records__isnull=False
        ).distinct().values('id', 'name').order_by('name')
        
        logger.info(f"Found {systems.count()} systems")
        
        # Перевіряємо, чи є системи
        if not systems.exists():
            # Якщо немає систем з access_records, спробуємо отримати всі системи компанії
            systems = InformationAsset.objects.filter(
                company_id=company_id,
                access_manage=True,  # Only include assets marked for access management
                deletion_date__isnull=True  # Only include active assets
            ).distinct().values('id', 'name').order_by('name')
            logger.info(f"Fallback: Found {systems.count()} systems without access_records filter")

        return JsonResponse({
            'status': 'success',
            'data': {
                'systems': list(systems)
            }
        })
    except Exception as e:
        logger.error(f"Error getting company systems: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'error': str(e)
        }, status=500)


@login_required
def get_system_objects_for_filter(request, system_id):
    """Отримання об'єктів для фільтрації на основі існуючих запитів користувача"""
    try:
        from .models import AccessObjectIS
        
        logger.info(f"Getting objects for system ID: {system_id}")
        
        # Отримуємо об'єкти з існуючих запитів користувача для цієї системи
        # Базовий запит - показувати об'єкти, до яких користувач має доступ
        user_groups = request.user.groups.all()
        
        objects = AccessObjectIS.objects.filter(
            asset_id=system_id,
            access_records__is_active=True
        ).filter(
            # Перевіряємо дату закінчення записів доступу
            Q(access_records__end_date__gt=timezone.now()) |
            Q(access_records__end_date__isnull=True)
        ).filter(
            # Перевіряємо користувача та групи в записах доступу
            Q(access_records__request_users=request.user) |
            Q(access_records__request_groups__in=user_groups)
        ).distinct().order_by('order', 'name')
        
        logger.info(f"Found {objects.count()} objects from existing requests")
        
        # Форматуємо відповідь
        objects_data = []
        for obj in objects:
            objects_data.append({
                'id': obj.id,
                'name': obj.get_name() or obj.name or '',
                'name': obj.get_name() or obj.name or '',
                'object_name_en': obj.get_name('en') or obj.object_name_en,
                'description': obj.get_description() or obj.description_ua
            })

        return JsonResponse({
            'status': 'success',
            'data': objects_data
        })
    except Exception as e:
        logger.error(f"Error getting system objects for filter: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'error': str(e)
        }, status=500)

def calculate_progress(start_date, end_date):
    """Розрахунок прогресу для періоду"""
    now = timezone.now()
    
    # Якщо немає кінцевої дати, вважаємо що період безстроковий
    if not end_date:
        if start_date > now:
            return {
                'percentage': 0,
                'is_active': False,
                'is_expired': False,
                'is_future': True
            }
        else:
            return {
                'percentage': 100,
                'is_active': True,
                'is_expired': False,
                'is_future': False
            }
    
    # Якщо період ще не почався
    if start_date > now:
        return {
            'percentage': 0,
            'is_active': False,
            'is_expired': False,
            'is_future': True
        }
    
    # Якщо період вже закінчився
    if end_date < now:
        return {
            'percentage': 100,
            'is_active': False,
            'is_expired': True,
            'is_future': False
        }
    
    # Розрахунок прогресу для активного періоду
    total_duration = (end_date - start_date).total_seconds()
    elapsed_duration = (now - start_date).total_seconds()
    percentage = min(100, max(0, (elapsed_duration / total_duration) * 100))
    
    return {
        'percentage': round(percentage, 1),
        'is_active': True,
        'is_expired': False,
        'is_future': False
    }

@login_required
@require_http_methods(["POST"])
def set_approver_status(request, approver_id):
    """Встановлення статусу approver'а"""
    try:
        import json
        
        approver = get_object_or_404(AccessRequestApprover, id=approver_id)
        
        # Перевіряємо права доступу:
        # 1. Адміністратор може змінювати все
        # 2. Самий approver може змінювати свій статус
        # 3. Перевіряємо додаткові права (owner, administrator системи)
        can_edit = (
            request.user.is_staff or  # Адміністратор
            (hasattr(request.user, 'cabinet') and 
             request.user.cabinet == approver.cabinet_user) or  # Самий approver
            # Owner системи
            approver.access_request.system.owners.filter(cabinet_user__user=request.user).exists() or
            # Administrator системи
            approver.access_request.system.administrators.filter(cabinet_user__user=request.user).exists()
        )
        
        if not can_edit:
            return JsonResponse({
                'success': False,
                'message': _('Permission denied: You cannot change this approver status')
            }, status=403)
        
        # Отримуємо дані з POST запиту
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, TypeError):
            data = request.POST
        
        new_status = data.get('status')
        comment = data.get('comment', '')
        
        if new_status not in ['pending', 'approved', 'rejected', 'cancelled']:
            return JsonResponse({
                'success': False,
                'message': _('Invalid status. Allowed values: pending, approved, rejected, cancelled')
            }, status=400)
        
        # Перевіряємо чи може цей approver встановлювати статус (логіка level)
        if new_status in ['approved', 'rejected'] and not approver.can_approve():
            # Знаходимо які рівні ще не затверджені
            pending_levels = []
            previous_levels = AccessRequestApprover.objects.filter(
                access_request=approver.access_request,
                order__lt=approver.order
            ).order_by('order')
            
            for prev_approver in previous_levels:
                if prev_approver.current_status != 'approved':
                    pending_levels.append(f"Level {prev_approver.order} ({prev_approver.cabinet_user.user.get_full_name()})")
            
            return JsonResponse({
                'success': False,
                'message': _('Cannot approve: Previous levels must be completed first. Pending: {}').format(', '.join(pending_levels))
            }, status=400)
        
        # Встановлюємо статус
        try:
            old_status = approver.current_status
            approver.set_status(new_status, request.user, comment)
            
            # Відправляємо email повідомлення про зміну статусу запитувачу
            try:
                # Тільки якщо статус кардинально змінився (не pending -> pending)
                if old_status != new_status and new_status in ['approved', 'rejected']:
                    approver_user = approver.cabinet_user.user
                    send_access_request_status_notification(
                        approver.access_request,
                        old_status,
                        new_status,
                        request.user,
                        comment,
                        status_change_context={
                            'status_type': 'approver',
                            'changed_approver_order': approver.order,
                            'changed_approver_email': approver_user.email if approver_user else '',
                        },
                    )
                    logger.info(f"Status update notification sent for access request {approver.access_request.id}")
            except Exception as e:
                logger.error(f"Failed to send status update notification: {e}")
                # Не зупиняємо процес через помилку email
            
            # Log status change
            if old_status != new_status:
                logger.info(f"Approver status changed for request {approver.access_request.id}: {old_status} -> {new_status}")
            
            return JsonResponse({
                'success': True,
                'message': _('Approver status updated successfully'),
                'new_status': new_status,
                'approver_name': approver.cabinet_user.user.get_full_name(),
                'level': approver.order
            })
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
        
    except Exception as e:
        logger.error(f"Error setting approver status: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': _('Internal server error')
        }, status=500)



@login_required
@require_http_methods(["POST"])
def set_admin_status(request, request_id):
    """Встановлення адміністративного статусу заявки"""
    try:
        import json
        
        access_request = get_object_or_404(AccessRequest, id=request_id)
        
        # Перевіряємо права доступу: тільки адміністратори системи можуть змінювати статус
        can_edit = (
            hasattr(request.user, 'cabinet') and 
            access_request.system.administrators.filter(cabinet_user=request.user.cabinet).exists()
        )
        
        if not can_edit:
            return JsonResponse({
                'success': False,
                'message': _('Permission denied: You are not an administrator of this system')
            }, status=403)
        
        # Перевіряємо що заявка погоджена
        if access_request.status != 'approved':
            return JsonResponse({
                'success': False,
                'message': _('Only approved requests can have administrative status changed')
            }, status=400)
        
        # Перевіряємо чи не намагаються змінити статус "Access Denied"
        if access_request.admin_status == 'denied':
            return JsonResponse({
                'success': False,
                'message': _('Access Denied status cannot be changed once set')
            }, status=400)
        
        # Отримуємо дані з POST запиту
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, TypeError):
            data = request.POST
        
        new_admin_status = data.get('admin_status')
        comment = data.get('comment', '')
        
        # Перевіряємо чи не закінчився період доступу (тільки для надання доступу)
        if access_request.is_period_expired and new_admin_status in ['granted', 'in_progress']:
            return JsonResponse({
                'success': False,
                'message': _('Cannot grant access: the access period has expired')
            }, status=400)
        
        # Перевіряємо допустимі статуси залежно від типу запиту
        if access_request.request_type == 'revoke':
            # Перевіряємо чи вже встановлено статус 'granted' (Access Revoked) - він не може бути змінений
            if access_request.admin_status == 'granted' and new_admin_status != 'granted':
                return JsonResponse({
                    'success': False,
                    'message': _('Access Revoked status cannot be changed once set. The revocation process is irreversible.')
                }, status=400)
            
            # Для Revoke запитів дозволені: granted (Access Revoked), pending, in_progress, denied (Reject Request)
            if new_admin_status not in ['granted', 'pending', 'in_progress', 'denied']:
                return JsonResponse({
                    'success': False,
                    'message': _('Invalid admin status for revoke request. Allowed values: granted, pending, in_progress, denied')
                }, status=400)
        else:
            # Для Grant запитів дозволені всі статуси
            if new_admin_status not in ['pending', 'granted', 'denied', 'in_progress']:
                return JsonResponse({
                    'success': False,
                    'message': _('Invalid admin status. Allowed values: pending, granted, denied, in_progress')
                }, status=400)
        
        # Зберігаємо попередній статус для логування
        old_admin_status = getattr(access_request, 'admin_status', 'pending')
        
        # Створюємо запис в історії тільки якщо статус змінився
        if old_admin_status != new_admin_status:
            AccessRequestAdminStatusHistory.objects.create(
                access_request=access_request,
                old_status=old_admin_status,
                new_status=new_admin_status,
                comment=comment,
                changed_by=request.user
            )
            
            # Відправляємо email повідомлення про зміну адміністративного статусу
            try:
                send_access_request_status_notification(
                    access_request,
                    old_admin_status,
                    new_admin_status,
                    request.user,
                    comment,
                    status_type='admin'
                )
                logger.info(f"Admin status notification sent for access request {access_request.id}")
            except Exception as e:
                logger.error(f"Failed to send admin status notification: {e}")
                # Не зупиняємо процес через помилку email
            
            # Log admin status change
            logger.info(f"Admin status changed for request {access_request.id}: {old_admin_status} -> {new_admin_status}")
        
        # Встановлюємо новий статус
        access_request.admin_status = new_admin_status
        access_request.admin_status_changed_by = request.user
        access_request.admin_status_changed_at = timezone.now()
        access_request.admin_status_comment = comment
        access_request.save()
        
        # Оновлюємо статуси схвалювачів при зміні admin_status на 'granted'
        if new_admin_status == 'granted' and old_admin_status != 'granted':
            # Автоматично встановлюємо статус 'approved' для всіх схвалювачів
            approvers_updated = 0
            for approver in access_request.request_approvers.all():
                if approver.current_status == 'pending':
                    # Створюємо запис в історії статусу схвалювача
                    try:
                        from .models import AccessRequestApproverStatusHistory
                        AccessRequestApproverStatusHistory.objects.create(
                            approver=approver,
                            old_status=approver.current_status,
                            new_status='approved',
                            comment=f'Automatically approved due to admin status change to granted by {request.user.get_full_name()}',
                            changed_by=request.user
                        )
                    except Exception as e:
                        logger.error(f"Failed to create approver status history for approver {approver.id}: {str(e)}")
                    
                    # Оновлюємо статус схвалювача
                    approver.current_status = 'approved'
                    approver.status_comment = f'Automatically approved due to admin grant by {request.user.get_full_name()}'
                    approver.status_changed_at = timezone.now()
                    approver.status_changed_by = request.user
                    approver.save()
                    approvers_updated += 1
                    
                    logger.info(f"Updated approver {approver.cabinet_user.user.get_full_name()} status to 'approved' for request {access_request.id}")
            
            if approvers_updated > 0:
                logger.info(f"Automatically updated {approvers_updated} approver statuses to 'approved' for request {access_request.id}")
        
        # Обробляємо логіку для схвалених Grant запитів
        if access_request.request_type == 'grant' and new_admin_status == 'granted':
            # Процесуємо схвалений grant запит - додаємо користувачів до access_users
            success = process_approved_grant_request(access_request)
            if success:
                logger.info(f"Successfully processed approved grant request {access_request.id}")
            else:
                logger.error(f"Failed to process approved grant request {access_request.id}")
        
        # Обробляємо логіку для схвалених Revoke запитів
        if access_request.request_type == 'revoke' and new_admin_status == 'granted':
            # Процесуємо схвалений revoke запит - позначаємо Access Records як revoked
            success = process_approved_revoke_request(access_request)
            if success:
                logger.info(f"Successfully processed approved revoke request {access_request.id}")
            else:
                logger.error(f"Failed to process approved revoke request {access_request.id}")
        
        # Також оновлюємо статуси схвалювачів для revoke запитів при зміні admin_status на 'granted'
        if access_request.request_type == 'revoke' and new_admin_status == 'granted' and old_admin_status != 'granted':
            # Автоматично встановлюємо статус 'approved' для всіх схвалювачів revoke запиту
            approvers_updated = 0
            for approver in access_request.request_approvers.all():
                if approver.current_status == 'pending':
                    # Створюємо запис в історії статусу схвалювача
                    try:
                        from .models import AccessRequestApproverStatusHistory
                        AccessRequestApproverStatusHistory.objects.create(
                            approver=approver,
                            old_status=approver.current_status,
                            new_status='approved',
                            comment=f'Automatically approved due to revoke admin status change to granted by {request.user.get_full_name()}',
                            changed_by=request.user
                        )
                    except Exception as e:
                        logger.error(f"Failed to create approver status history for revoke approver {approver.id}: {str(e)}")
                    
                    # Оновлюємо статус схвалювача
                    approver.current_status = 'approved'
                    approver.status_comment = f'Automatically approved due to revoke admin grant by {request.user.get_full_name()}'
                    approver.status_changed_at = timezone.now()
                    approver.status_changed_by = request.user
                    approver.save()
                    approvers_updated += 1
                    
                    logger.info(f"Updated revoke approver {approver.cabinet_user.user.get_full_name()} status to 'approved' for request {access_request.id}")
            
            if approvers_updated > 0:
                logger.info(f"Automatically updated {approvers_updated} revoke approver statuses to 'approved' for request {access_request.id}")
        
        # Обробляємо логіку для Revoke запитів з статусом "denied"
        if access_request.request_type == 'revoke' and new_admin_status == 'denied':
            # Знаходимо оригінальний Grant запит
            if access_request.notes:
                import re
                match = re.search(r'request #(\d+)', access_request.notes)
                if match:
                    original_request_id = int(match.group(1))
                    try:
                        original_request = AccessRequest.objects.get(
                            id=original_request_id,
                            request_type='grant'
                        )
                        
                        # ЗМІНЕНО: НЕ змінюємо admin_status оригінального запиту на 'denied'
                        # Замість цього тільки логуємо відхилення revoke запиту
                        if original_request.admin_status == 'granted':
                            logger.info(f"Revoke request {access_request.id} was denied by admin {request.user}. Original grant request {original_request_id} admin_status remains 'granted'.")
                            logger.info(f"Access Records remain active with their current status, only revoke request was denied.")
                        
                    except AccessRequest.DoesNotExist:
                        logger.warning(f"Original request {original_request_id} not found for revoke request {access_request.id}")
        
        # Логуємо зміну
        logger.info(f"Admin status changed for request {request_id}: {old_admin_status} -> {new_admin_status} by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': _('Administrative status updated successfully'),
            'new_status': new_admin_status,
            'request_id': request_id
        })
        
    except Exception as e:
        logger.error(f"Error setting admin status: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': _('Internal server error')
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_third_party_users(request):
    """Отримання списку існуючих Third Party користувачів"""
    try:
        # Отримуємо всіх активних третьосторонніх користувачів
        users = ThirdPartyUser.objects.filter(is_active=True).select_related('organization').order_by('first_name', 'last_name')
        
        users_data = []
        for user in users:
            users_data.append({
                'id': user.id,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'phone': user.phone or '',
                'organization_name': user.organization.name if user.organization else user.organization_name,
                'organization_id': user.organization.id if user.organization else None,
                'description': user.description or '',
                'full_name': user.full_name,
                'created_at': user.created_at.isoformat() if user.created_at else None
            })
        
        return JsonResponse({
            'success': True,
            'users': users_data
        })
        
    except Exception as e:
        logger.error(f"Error getting third party users: {e}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
@csrf_protect
def create_third_party_organization(request):
    """Створення нової організації третьої сторони"""
    try:
        data = json.loads(request.body)
        
        organization_name = data.get('name', '').strip()
        if not organization_name:
            return JsonResponse({
                'success': False,
                'message': _('Organization name is required')
            }, status=400)
        
        # Перевіряємо чи вже існує така організація
        if ThirdPartyOrganization.objects.filter(name__iexact=organization_name).exists():
            return JsonResponse({
                'success': False,
                'message': _('Organization with this name already exists')
            }, status=400)
        
        # Створюємо нову організацію
        organization = ThirdPartyOrganization.objects.create(
            name=organization_name,
            description=data.get('description', ''),
            contact_email=data.get('contact_email', ''),
            contact_phone=data.get('contact_phone', ''),
            website=data.get('website', ''),
            address=data.get('address', ''),
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'organization': {
                'id': organization.id,
                'name': organization.name,
                'description': organization.description,
                'contact_email': organization.contact_email,
                'contact_phone': organization.contact_phone,
                'website': organization.website,
                'address': organization.address
            },
            'message': _('Organization created successfully')
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': _('Invalid JSON data')
        }, status=400)
    except Exception as e:
        logger.error(f"Error creating third party organization: {e}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def _serialize_third_party_user(user):
    return {
        'id': user.id,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'phone': user.phone or '',
        'organization_name': user.organization.name if user.organization else user.organization_name,
        'organization_id': user.organization.id if user.organization else None,
        'description': user.description or '',
        'full_name': user.full_name,
        'created_at': user.created_at.isoformat() if user.created_at else None,
    }


def _create_or_update_third_party_user(user_data, created_by):
    """
    Create or update a ThirdPartyUser record.
    Returns (user, created) tuple.
    """
    email = (user_data.get('email') or '').strip()
    if not email:
        raise ValueError(_('Email is required'))

    organization_obj = None
    organization_name = (user_data.get('organization') or '').strip()
    if organization_name:
        organization_obj, _ = ThirdPartyOrganization.objects.get_or_create(
            name=organization_name,
            defaults={
                'created_by': created_by,
                'is_active': True,
            }
        )

    third_party_user, created = ThirdPartyUser.objects.get_or_create(
        email=email,
        defaults={
            'first_name': (user_data.get('first_name') or '').strip(),
            'last_name': (user_data.get('last_name') or '').strip(),
            'phone': (user_data.get('phone') or '').strip(),
            'organization': organization_obj,
            'organization_name': organization_name,
            'description': (user_data.get('description') or '').strip(),
            'created_by': created_by,
            'is_active': True,
        }
    )

    if not created:
        updated = False
        field_updates = {
            'first_name': (user_data.get('first_name') or '').strip(),
            'last_name': (user_data.get('last_name') or '').strip(),
            'phone': (user_data.get('phone') or '').strip(),
            'description': (user_data.get('description') or '').strip(),
        }
        for field, value in field_updates.items():
            if value and getattr(third_party_user, field) != value:
                setattr(third_party_user, field, value)
                updated = True

        if organization_name and (
            not third_party_user.organization
            or third_party_user.organization.name != organization_name
        ):
            new_org_obj, _ = ThirdPartyOrganization.objects.get_or_create(
                name=organization_name,
                defaults={
                    'created_by': created_by,
                    'is_active': True,
                }
            )
            third_party_user.organization = new_org_obj
            third_party_user.organization_name = organization_name
            updated = True

        if updated:
            third_party_user.save()

    return third_party_user, created


@login_required
@require_http_methods(["POST"])
@csrf_protect
def create_third_party_user(request):
    """Створення нового Third Party користувача"""
    try:
        data = json.loads(request.body)

        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        email = data.get('email', '').strip()

        if not first_name:
            return JsonResponse({
                'success': False,
                'message': _('First name is required')
            }, status=400)

        if not last_name:
            return JsonResponse({
                'success': False,
                'message': _('Last name is required')
            }, status=400)

        if not email:
            return JsonResponse({
                'success': False,
                'message': _('Email is required')
            }, status=400)

        third_party_user, created = _create_or_update_third_party_user(data, request.user)

        if created:
            logger.info(f"Created new ThirdPartyUser via API: {third_party_user.email}")
            message = _('Third party user created successfully')
        else:
            logger.info(f"Returned existing ThirdPartyUser via API: {third_party_user.email}")
            message = _('Third party user already exists and was updated')

        return JsonResponse({
            'success': True,
            'created': created,
            'user': _serialize_third_party_user(third_party_user),
            'message': message,
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': _('Invalid JSON data')
        }, status=400)
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
    except Exception as e:
        logger.error(f"Error creating third party user: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def check_duplicate_request(request):
    """Перевірка на duplicate запити при виборі access record"""
    try:
        import json
        from django.utils import translation
        
        # Отримуємо дані з POST запиту
        # Спочатку перевіряємо, чи це FormData
        if request.content_type and 'multipart/form-data' in request.content_type:
            data = request.POST
        else:
            try:
                data = json.loads(request.body)
            except (json.JSONDecodeError, TypeError, Exception):
                # Fallback до request.POST якщо не можемо прочитати body
                data = request.POST
        
        # Перевіряємо тип перевірки
        check_type = data.get('check_type', 'grant_duplicate')
        
        if check_type == 'revoke_duplicate':
            # Перевірка дублікатів revoke запитів
            return check_revoke_duplicates(request, data)
        
        # Стандартна перевірка дублікатів grant запитів
        system_id = data.get('system_id')
        access_record_id = data.get('access_record_id')
        environment = data.get('environment')
        user_ids = data.get('user_ids', [])
        is_third_party = data.get('is_third_party', False)
        
        logger.info(f"Checking duplicates for system_id={system_id}, access_record_id={access_record_id}, environment={environment}, users={user_ids}, is_third_party={is_third_party}")
        
        # Валідація обов'язкових полів
        if not all([system_id, access_record_id, environment]):
            return JsonResponse({
                'success': False,
                'message': _('Missing required fields')
            }, status=400)
        
        # Перевірка дублікатів
        duplicates_found = []
        
        if is_third_party:
            # Для third party перевіряємо кожного обраного користувача
            # Отримуємо поточну мову інтерфейсу
            try:
                current_language = translation.get_language()
            except Exception as lang_error:
                logger.error(f"Error getting current language: {lang_error}")
                current_language = 'en'  # fallback
            
            for user_id in user_ids:
                # Перевіряємо чи user_id це ID існуючого ThirdPartyUser чи email нового користувача
                existing_requests = None
                third_party_user = None
                
                logger.info(f"Processing user_id: {user_id} (type: {type(user_id)})")
                
                if isinstance(user_id, int) or (isinstance(user_id, str) and user_id.isdigit()):
                    # Це ID існуючого ThirdPartyUser
                    try:
                        third_party_user = ThirdPartyUser.objects.get(id=int(user_id))
                        # Шукаємо існуючі запити для цього ThirdPartyUser
                        # ВАЖЛИВО: Перевіряємо тільки запити які мають зв'язок з ThirdPartyUser
                        # Знаходимо всі існуючі запити для цього third party user
                        # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                        potential_existing_requests = AccessRequest.objects.filter(
                            third_party_users=third_party_user,
                            system_id=system_id,
                            environment=environment,
                            request_type='grant',
                            status__in=['pending', 'approved']
                        ).exclude(admin_status='denied')
                        
                        logger.info(f"Found {potential_existing_requests.count()} potential existing requests for third party user {user_id} ({third_party_user.first_name} {third_party_user.last_name})")
                        for req in potential_existing_requests:
                            logger.info(f"  - Request #{req.id}: status={req.status}, admin_status={req.admin_status}, access_records={[ar.id for ar in req.access_records.all()]}")
                        
                        # Перевіряємо чи є запит з перетином access records
                        existing_requests = None
                        for potential_request in potential_existing_requests:
                            existing_access_record_ids = set(potential_request.access_records.values_list('id', flat=True))
                            new_access_record_ids = set([int(access_record_id)])
                            
                            if existing_access_record_ids & new_access_record_ids:
                                # Додаткова перевірка: чи є успішний revoke запит для цього grant запиту
                                # Перевіряємо чи є revoke запит з перетином access records
                                overlapping_access_record_ids = existing_access_record_ids & new_access_record_ids
                                logger.info(f"Found overlap between existing request #{potential_request.id} and new request: overlapping_records={overlapping_access_record_ids}")
                                
                                revoke_requests = AccessRequest.objects.filter(
                                    third_party_users=third_party_user,
                                    system_id=system_id,
                                    environment=environment,
                                    request_type='revoke',
                                    admin_status='granted',  # Успішно відкликано
                                    access_records__id__in=overlapping_access_record_ids
                                ).distinct()
                                
                                logger.info(f"Searching for revoke requests with criteria: user={third_party_user.id}, system={system_id}, env={environment}, type=revoke, admin_status=granted, access_records__id__in={overlapping_access_record_ids}")
                                logger.info(f"Found {revoke_requests.count()} revoke requests:")
                                for revoke_req in revoke_requests:
                                    logger.info(f"  - Revoke request #{revoke_req.id}: status={revoke_req.status}, admin_status={revoke_req.admin_status}, access_records={[ar.id for ar in revoke_req.access_records.all()]}")
                                
                                if revoke_requests.exists():
                                    logger.info(f"Found successful revoke request for third party user {user_id} with overlapping access records {overlapping_access_record_ids}, allowing duplicate")
                                    continue  # Пропускаємо цей grant запит, оскільки він був відкликаний
                                
                                logger.info(f"No revoke requests found, blocking duplicate for request #{potential_request.id}")
                                existing_requests = potential_request
                                break
                    except ThirdPartyUser.DoesNotExist:
                        continue
                else:
                    # Це email нового користувача - шукаємо по старих полях або по ThirdPartyUser з таким email
                    # ВАЖЛИВО: Додаємо додаткову перевірку, щоб шукати тільки third party запити
                    # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                    potential_existing_requests = AccessRequest.objects.filter(
                        Q(
                            Q(third_party_email=user_id) | 
                            Q(third_party_users__email=user_id)
                        ) & 
                        Q(
                            Q(third_party_users__isnull=False) |
                            Q(third_party_first_name__isnull=False)
                        ),
                        system_id=system_id,
                        environment=environment,
                        request_type='grant',
                        status__in=['pending', 'approved']
                    ).exclude(admin_status='denied')
                    
                    # Перевіряємо чи є запит з перетином access records
                    existing_requests = None
                    for potential_request in potential_existing_requests:
                        existing_access_record_ids = set(potential_request.access_records.values_list('id', flat=True))
                        new_access_record_ids = set([int(access_record_id)])
                        
                        if existing_access_record_ids & new_access_record_ids:
                            # Додаткова перевірка: чи є успішний revoke запит для цього grant запиту
                            # Перевіряємо чи є revoke запит з перетином access records
                            overlapping_access_record_ids = existing_access_record_ids & new_access_record_ids
                            
                            revoke_requests = AccessRequest.objects.filter(
                                Q(
                                    Q(third_party_email=user_id) | 
                                    Q(third_party_users__email=user_id)
                                ),
                                system_id=system_id,
                                environment=environment,
                                request_type='revoke',
                                admin_status='granted',  # Успішно відкликано
                                access_records__id__in=overlapping_access_record_ids
                            ).distinct()
                            
                            if revoke_requests.exists():
                                logger.info(f"Found successful revoke request for third party email {user_id} with overlapping access records {overlapping_access_record_ids}, allowing duplicate")
                                continue  # Пропускаємо цей grant запит, оскільки він був відкликаний
                            
                            existing_requests = potential_request
                            break
            
                if existing_requests:
                    existing_request = existing_requests
                    
                    # Додаткова перевірка: переконуємося, що це дійсно third party request
                    is_actually_third_party = (
                        existing_request.third_party_users.exists() or
                        existing_request.third_party_first_name or
                        existing_request.third_party_email
                    )
                    
                    if not is_actually_third_party:
                        # Це не third party request, пропускаємо
                        continue
                    
                    # Отримуємо детальну інформацію для third party
                    object_name = _('No Object')
                    if existing_request.access_records.first() and existing_request.access_records.first().access_object:
                        if current_language == 'uk':
                            object_name = existing_request.access_records.first().access_object.get_name('uk')
                        elif current_language == 'ru':
                            object_name = existing_request.access_records.first().access_object.get_name('ru')
                        else:
                            object_name = existing_request.access_records.first().access_object.get_name('en')
                    
                    requested_by = existing_request.requested_by.get_full_name() or existing_request.requested_by.username
                        
                    # Визначаємо інформацію про третьосторонього користувача
                    if third_party_user:
                        third_party_info = f"{third_party_user.first_name} {third_party_user.last_name}".strip()
                        if third_party_user.email:
                            third_party_info += f" ({third_party_user.email})"
                    elif existing_request.third_party_users.exists():
                        # Беремо першого користувача з нового зв'язку
                        tp_user = existing_request.third_party_users.first()
                        third_party_info = f"{tp_user.first_name} {tp_user.last_name}".strip()
                        if tp_user.email:
                            third_party_info += f" ({tp_user.email})"
                    else:
                        # Fallback до старих полів
                        third_party_info = f"{existing_request.third_party_first_name} {existing_request.third_party_last_name}".strip()
                        if existing_request.third_party_email:
                            third_party_info += f" ({existing_request.third_party_email})"
                    
                    # Get roles with current language
                    roles = []
                    if existing_request.access_records.first():
                        for role in existing_request.access_records.first().roles.all():
                            if current_language == 'uk':
                                role_name = role.get_name() or role.name or ''
                            elif current_language == 'ru':
                                role_name = role.get_name() or role.name or ''
                            else:
                                role_name = role.get_name('en') or role.name or ''
                            roles.append(role_name)
                    roles_str = ", ".join(roles) if roles else _("No roles")
                    
                    # Get attachments count
                    attachments_count = existing_request.attachments.count() if hasattr(existing_request, 'attachments') else 0
                    attachments_str = f"{attachments_count} files" if attachments_count > 0 else _("No attachments")
                    
                    # Get approving persons
                    approvers = []
                    if hasattr(existing_request, 'request_approvers'):
                        for approver in existing_request.request_approvers.all()[:3]:
                            approver_name = approver.cabinet_user.user.get_full_name() or approver.cabinet_user.user.username
                            approvers.append(f"{approver_name} ({approver.current_status})")
                    approvers_str = "; ".join(approvers) if approvers else _("No approvers")
                    if existing_request.request_approvers.count() > 3:
                        approvers_str += f" +{existing_request.request_approvers.count() - 3} more"
                    
                    # Format dates
                    created_date = existing_request.created_at.strftime('%d.%m.%Y %H:%M')
                    period_str = ""
                    # Беремо період з AccessRecord, а не з AccessRequest
                    if existing_request.access_records.first():
                        access_record = existing_request.access_records.first()
                        if access_record.start_date and access_record.end_date:
                            period_str = f"{access_record.start_date.strftime('%d.%m.%Y %H:%M')} - {access_record.end_date.strftime('%d.%m.%Y %H:%M')}"
                        elif access_record.start_date:
                            period_str = f"{access_record.start_date.strftime('%d.%m.%Y %H:%M')} - Indefinite"
                    # Fallback: якщо немає AccessRecord, перевіряємо AccessRequest
                    elif existing_request.start_date and existing_request.end_date:
                        period_str = f"{existing_request.start_date.strftime('%d.%m.%Y')} - {existing_request.end_date.strftime('%d.%m.%Y')}"
                    
                    duplicates_found.append({
                        'type': 'third_party',
                        'user_info': third_party_info,
                        'request_id': existing_request.id,
                        'created_date': created_date,
                        'details': {
                            'object_name': object_name,
                            'requested_by': requested_by,
                            'third_party': third_party_info,
                            'roles': roles_str,
                            'period': period_str or 'Not specified',
                            'attachments': attachments_str,
                            'approvers': approvers_str,
                            'status': existing_request.status,
                            'admin_status': existing_request.admin_status
                        }
                    })
        
        elif user_ids:
            # Для cabinet users перевіряємо кожного користувача
            # Отримуємо поточну мову інтерфейсу
            current_language = translation.get_language()
            
            for user_id in user_ids:
                # Знаходимо всі існуючі запити для цього користувача в тій же системі та середовищі
                # ІГНОРУЄМО запити з admin_status='denied' оскільки доступ не був наданий
                potential_existing_requests = AccessRequest.objects.filter(
                    requested_for_id=user_id,
                    system_id=system_id,
                    environment=environment,
                    request_type='grant',
                    status__in=['pending', 'approved']
                ).exclude(admin_status='denied')
                
                logger.info(f"Found {potential_existing_requests.count()} potential existing requests for cabinet user {user_id}")
                
                # Перевіряємо чи є запит з перетином access records
                existing_request = None
                for potential_request in potential_existing_requests:
                    existing_access_record_ids = set(potential_request.access_records.values_list('id', flat=True))
                    new_access_record_ids = set([int(access_record_id)])  # Конвертуємо в int для порівняння
                    
                    # Перевіряємо чи є перетин access records
                    if existing_access_record_ids & new_access_record_ids:
                        # Додаткова перевірка: чи є успішний revoke запит для цього grant запиту
                        # Перевіряємо чи є revoke запит з перетином access records
                        overlapping_access_record_ids = existing_access_record_ids & new_access_record_ids
                        
                        revoke_requests = AccessRequest.objects.filter(
                            requested_for_id=user_id,
                            system_id=system_id,
                            environment=environment,
                            request_type='revoke',
                            admin_status='granted',  # Успішно відкликано
                            access_records__id__in=overlapping_access_record_ids
                        ).distinct()
                        
                        if revoke_requests.exists():
                            logger.info(f"Found successful revoke request for cabinet user {user_id} with overlapping access records {overlapping_access_record_ids}, allowing duplicate")
                            continue  # Пропускаємо цей grant запит, оскільки він був відкликаний
                        
                        existing_request = potential_request
                        break
                
                if existing_request:
                    try:
                        from django.contrib.auth.models import User as AuthUser
                        user = AuthUser.objects.get(id=user_id)
                        user_info = user.get_full_name() or user.username
                        
                        # Отримуємо детальну інформацію для cabinet user
                        object_name = _('No Object')
                        if existing_request.access_records.first() and existing_request.access_records.first().access_object:
                            if current_language == 'uk':
                                object_name = existing_request.access_records.first().access_object.get_name('uk')
                            elif current_language == 'ru':
                                object_name = existing_request.access_records.first().access_object.get_name('ru')
                            else:
                                object_name = existing_request.access_records.first().access_object.get_name('en')
                        
                        requested_by = existing_request.requested_by.get_full_name() or existing_request.requested_by.username
                        requested_for = existing_request.requested_for.get_full_name() or existing_request.requested_for.username
                        
                        # Get roles with current language
                        roles = []
                        if existing_request.access_records.first():
                            for role in existing_request.access_records.first().roles.all():
                                if current_language == 'uk':
                                    role_name = role.get_name() or role.name or ''
                                elif current_language == 'ru':
                                    role_name = role.get_name() or role.name or ''
                                else:
                                    role_name = role.get_name('en') or role.name or ''
                                roles.append(role_name)
                        roles_str = ", ".join(roles) if roles else _("No roles")
                        
                        # Get attachments count
                        attachments_count = existing_request.attachments.count() if hasattr(existing_request, 'attachments') else 0
                        attachments_str = f"{attachments_count} files" if attachments_count > 0 else _("No attachments")
                        
                        # Get approving persons
                        approvers = []
                        if hasattr(existing_request, 'request_approvers'):
                            for approver in existing_request.request_approvers.all()[:3]:
                                approver_name = approver.cabinet_user.user.get_full_name() or approver.cabinet_user.user.username
                                approvers.append(f"{approver_name} ({approver.current_status})")
                        approvers_str = "; ".join(approvers) if approvers else _("No approvers")
                        if existing_request.request_approvers.count() > 3:
                            approvers_str += f" +{existing_request.request_approvers.count() - 3} more"
                        
                        # Format dates
                        created_date = existing_request.created_at.strftime('%d.%m.%Y %H:%M')
                        period_str = ""
                        # Беремо період з AccessRecord, а не з AccessRequest
                        if existing_request.access_records.first():
                            access_record = existing_request.access_records.first()
                            if access_record.start_date and access_record.end_date:
                                period_str = f"{access_record.start_date.strftime('%d.%m.%Y %H:%M')} - {access_record.end_date.strftime('%d.%m.%Y %H:%M')}"
                            elif access_record.start_date:
                                period_str = f"{access_record.start_date.strftime('%d.%m.%Y %H:%M')} - Indefinite"
                        # Fallback: якщо немає AccessRecord, перевіряємо AccessRequest
                        elif existing_request.start_date and existing_request.end_date:
                            period_str = f"{existing_request.start_date.strftime('%d.%m.%Y')} - {existing_request.end_date.strftime('%d.%m.%Y')}"
                        
                        duplicates_found.append({
                            'type': 'cabinet_user',
                            'user_info': user_info,
                            'request_id': existing_request.id,
                            'created_date': created_date,
                            'details': {
                                'object_name': object_name,
                                'requested_by': requested_by,
                                'requested_for': requested_for,
                                'roles': roles_str,
                                'period': period_str or 'Not specified',
                                'attachments': attachments_str,
                                'approvers': approvers_str,
                                'status': existing_request.status,
                                'admin_status': existing_request.admin_status
                            }
                        })
                    except AuthUser.DoesNotExist:
                        duplicates_found.append({
                            'type': 'cabinet_user',
                            'user_info': f'User ID {user_id}',
                            'request_id': existing_request.id,
                            'created_date': existing_request.created_at.strftime('%d.%m.%Y %H:%M'),
                            'details': None
                        })
        
        # Повертаємо результат
        return JsonResponse({
            'success': True,
            'has_duplicates': len(duplicates_found) > 0,
            'duplicates': duplicates_found,
            'message': _('Duplicate check completed')
        })
        
    except Exception as e:
        logger.error(f"Error checking duplicate requests: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': _('Internal server error')
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_user_active_requests(request, system_id):
    """
    Отримання активних запитів користувача для відкликання
    """
    try:
        logger.info(f"get_user_active_requests called: user={request.user.id}, system_id={system_id}")
        
        # Отримуємо параметри фільтрації
        filter_user_id = request.GET.get('user_id')
        filter_environment = request.GET.get('environment')
        filter_object_id = request.GET.get('object_id')
        
        logger.info(f"Filters: user_id={filter_user_id}, environment={filter_environment}, object_id={filter_object_id}")
        
        # Спочатку перевіримо загальну кількість запитів користувача
        total_requests = AccessRequest.objects.filter(requested_by=request.user).count()
        logger.info(f"Total user requests: {total_requests}")
        
        # Базовий запит для активних запитів користувача (approved і granted)
        user_requests = AccessRequest.objects.filter(
            requested_by=request.user,
            system_id=system_id,
            request_type='grant',
            status='approved',
            admin_status='granted'
        )
        
        # Застосовуємо фільтри
        if filter_user_id and filter_user_id != 'all':
            if filter_user_id == 'third_party':
                # Фільтр для третьосторонніх користувачів
                user_requests = user_requests.filter(
                    Q(third_party_users__isnull=False) |
                    Q(third_party_first_name__isnull=False)
                )
            else:
                # Фільтр для конкретного користувача кабінету
                user_requests = user_requests.filter(requested_for_id=filter_user_id)
        
        if filter_environment and filter_environment != 'all':
            user_requests = user_requests.filter(environment=filter_environment)
            
        if filter_object_id and filter_object_id != 'all':
            user_requests = user_requests.filter(access_record__access_object_id=filter_object_id)
        
        user_requests = user_requests.select_related(
            'access_record',
            'access_record__access_object',
            'system',
            'requested_for'
        ).prefetch_related(
            'access_record__roles',
            'third_party_users'
        ).order_by('-created_at')
        
        logger.info(f"Found {user_requests.count()} user requests")
        
        # Отримуємо поточну мову інтерфейсу
        current_language = translation.get_language()
        
        requests_data = []
        for req in user_requests:
            object_id = None
            object_name = _('No Object')
            if req.access_record and req.access_record.access_object:
                object_id = req.access_record.access_object.id
                object_name = req.access_record.access_object.get_name(current_language[:2]) or str(req.access_record.access_object)

            # Отримуємо ролі з урахуванням мови
            roles = []
            if req.access_record:
                for role in req.access_record.roles.all():
                    if current_language == 'uk':
                        role_name = role.get_name() or role.name or ''
                    elif current_language == 'ru':
                        role_name = role.get_name() or role.name or ''
                    else:
                        role_name = role.get_name('en') or role.name or ''
                    
                    roles.append({
                        'id': role.id,
                        'name': role_name,
                        'color': role.color
                    })
            
            # Форматуємо дати
            start_date = None
            end_date = None
            if req.access_record:
                if req.access_record.start_date:
                    start_date = req.access_record.start_date.isoformat()
                if req.access_record.end_date:
                    end_date = req.access_record.end_date.isoformat()
            elif req.start_date:
                start_date = req.start_date.isoformat()
                if req.end_date:
                    end_date = req.end_date.isoformat()
            
            # Отримуємо інформацію про користувача (для кого запит)
            requested_for_info = None
            if req.requested_for:
                requested_for_info = {
                    'id': req.requested_for.id,
                    'name': req.requested_for.get_full_name() or req.requested_for.username,
                    'username': req.requested_for.username
                }
            
            # Отримуємо інформацію про третьосторонніх користувачів (лише для справжніх third-party запитів)
            third_party_info = []
            if is_third_party_access_request(req):
                if req.third_party_users.exists():
                    for tp_user in req.third_party_users.all():
                        organization_name = ""
                        if tp_user.organization:
                            if hasattr(tp_user.organization, 'name'):
                                organization_name = tp_user.organization.name
                            else:
                                organization_name = str(tp_user.organization)

                        third_party_info.append({
                            'id': tp_user.id,
                            'name': f"{tp_user.first_name} {tp_user.last_name}".strip(),
                            'email': tp_user.email,
                            'organization': organization_name
                        })
                elif req.third_party_first_name:
                    third_party_info.append({
                        'name': f"{req.third_party_first_name} {req.third_party_last_name}".strip(),
                        'email': req.third_party_email,
                        'organization': req.third_party_organization
                    })
            
            requests_data.append({
                'id': req.id,
                'access_record_id': req.access_record.id if req.access_record else None,
                'access_record': {
                    'id': req.access_record.id if req.access_record else None,
                    'access_object': {
                        'id': object_id,
                        'name': object_name
                    }
                },
                'object_id': object_id,
                'object_name': object_name,
                'environment': req.environment,
                'roles': roles,
                'start_date': start_date,
                'end_date': end_date,
                'justification': req.justification,
                'status': req.status,
                'admin_status': req.admin_status,
                'created_at': req.created_at.isoformat(),
                'requested_for': requested_for_info,
                'third_party_users': third_party_info,
                'is_third_party_request': is_third_party_access_request(req),
            })
        
        logger.info(f"Returning {len(requests_data)} requests")
        
        return JsonResponse({
            'success': True,
            'requests': requests_data,
            'count': len(requests_data),
            'message': _('Active requests loaded successfully')
        })
        
    except Exception as e:
        logger.error(f"Error loading user active requests: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': _('Error loading active requests'),
            'error': str(e)
        }, status=500)


def check_revoke_duplicates(request, data):
    """
    Перевірка на дублікати revoke запитів для конкретного оригінального запиту
    """
    try:
        from django.utils import translation
        from django.http import JsonResponse
        from django.utils.translation import gettext as _
        import logging
        from app_access.models import AccessRequest
        
        logger = logging.getLogger(__name__)
        
        # Додаємо debug логування
        logger.info(f"check_revoke_duplicates called with data: {data}")
        
        original_request_id = data.get('original_request_id')
        logger.info(f"original_request_id: {original_request_id}")
        
        if not original_request_id:
            return JsonResponse({
                'success': False,
                'message': _('Original request ID is required')
            }, status=400)
        
        # Знаходимо оригінальний запит
        try:
            original_request = AccessRequest.objects.get(
                id=original_request_id,
                requested_by=request.user,
                request_type='grant'
            )
        except AccessRequest.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': _('Original request not found or not accessible')
            }, status=404)
        
        # Шукаємо існуючі revoke запити для цього оригінального запиту
        existing_revoke_requests = AccessRequest.objects.filter(
            request_type='revoke',
            requested_by=request.user,
            requested_for_id=original_request.requested_for_id,
            access_record_id=original_request.access_record_id,
            system_id=original_request.system_id,
            environment=original_request.environment,
            status__in=['pending', 'approved'],
            notes__contains=f'request #{original_request_id}'
        )
        
        duplicates_found = []
        
        if existing_revoke_requests.exists():
            current_language = translation.get_language()
            
            for revoke_request in existing_revoke_requests:
                # Отримуємо детальну інформацію про існуючий revoke запит
                object_name = _('No Object')
                if revoke_request.access_record and revoke_request.access_record.access_object:
                    if current_language == 'uk':
                        object_name = revoke_request.access_record.access_object.get_name('uk')
                    elif current_language == 'ru':
                        object_name = revoke_request.access_record.access_object.get_name('ru')
                    else:
                        object_name = revoke_request.access_record.access_object.get_name('en')
                
                requested_by = revoke_request.requested_by.get_full_name() or revoke_request.requested_by.username
                requested_for = revoke_request.requested_for.get_full_name() or revoke_request.requested_for.username
                
                # Форматуємо дату створення
                created_date = revoke_request.created_at.strftime('%d.%m.%Y %H:%M')
                
                duplicates_found.append({
                    'type': 'revoke_duplicate',
                    'request_id': revoke_request.id,
                    'original_request_id': original_request_id,
                    'status': revoke_request.status,
                    'created_date': created_date,
                    'details': {
                        'object_name': object_name,
                        'requested_by': requested_by,
                        'requested_for': requested_for,
                        'environment': revoke_request.environment,
                        'status': revoke_request.status,
                        'admin_status': revoke_request.admin_status,
                        'justification': revoke_request.justification or _('No justification provided'),
                        'revoke_immediately': 'Immediate' if not revoke_request.start_date else 'Scheduled',
                        'revoke_start_date': revoke_request.start_date.strftime('%d.%m.%Y') if revoke_request.start_date else None,
                        'revoke_end_date': revoke_request.end_date.strftime('%d.%m.%Y') if revoke_request.end_date else None
                    }
                })
        
        return JsonResponse({
            'success': True,
            'has_duplicates': len(duplicates_found) > 0,
            'duplicates_found': duplicates_found,
            'message': _('Revoke duplicate check completed')
        })
        
    except Exception as e:
        logger.error(f"Error checking revoke duplicates: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': _('Internal server error')
        }, status=500)


@login_required
def export_access_requests_excel(request):
    """
    Export access requests to Excel with enhanced styling and colors
    Supports both selected records and all records export
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from io import BytesIO
    import datetime
    
    # Get export parameters
    export_type = request.GET.get('export_type', 'all')  # 'selected' or 'all'
    selected_ids = request.GET.getlist('selected_ids[]')
    current_language = get_language()[:2]
    
    # Base query - same as in user_access_request view
    requests_query = AccessRequest.objects.filter(
        Q(requested_by=request.user) |
        Q(system__owners__cabinet_user__user=request.user) |
        Q(system__administrators__cabinet_user__user=request.user) |
        Q(request_approvers__cabinet_user__user=request.user)
    ).distinct().select_related(
        'company',
        'system',
        'requested_by',
        'requested_for',
        'requested_for__cabinet',
        'requested_for__cabinet__department',
        'requested_for__cabinet__position',
        'access_record',
        'access_record__status',
        'access_record__access_right',
        'access_record__access_object',
        'cancelled_by',
        'admin_status_changed_by'
    ).prefetch_related(
        'request_approvers__cabinet_user',
        'request_approvers__cabinet_user__user',
        'request_approvers__cabinet_user__department',
        'request_approvers__cabinet_user__position',
        'access_record__roles',
        'third_party_users',
        'attachments'
    )
    
    # Apply filters from request (same as in user_access_request view)
    if request.GET.get('company'):
        requests_query = requests_query.filter(company_id=request.GET.get('company'))
    
    if request.GET.get('system'):
        requests_query = requests_query.filter(system_id=request.GET.get('system'))
    
    if request.GET.get('object'):
        requests_query = requests_query.filter(access_records__access_object_id=request.GET.get('object'))
    
    if request.GET.get('status'):
        requests_query = requests_query.filter(status=request.GET.get('status'))
    
    if request.GET.get('environment'):
        requests_query = requests_query.filter(environment=request.GET.get('environment'))
    
    if request.GET.get('role'):
        role_id = request.GET.get('role')
        requests_query = requests_query.filter(access_records__roles__id=role_id)
    
    if request.GET.get('admin_status'):
        requests_query = requests_query.filter(admin_status=request.GET.get('admin_status'))
    
    if request.GET.get('request_type'):
        requests_query = requests_query.filter(request_type=request.GET.get('request_type'))
    
    if request.GET.get('search'):
        search_query = request.GET.get('search')
        requests_query = requests_query.filter(
            Q(id__icontains=search_query) |
            Q(company__name__icontains=search_query) |
            Q(system__name__icontains=search_query) |
            Q(requested_for__username__icontains=search_query) |
            Q(justification__icontains=search_query) |
            Q(third_party_first_name__icontains=search_query) |
            Q(third_party_last_name__icontains=search_query) |
            Q(third_party_email__icontains=search_query)
        )
    
    # Filter by selected IDs if export_type is 'selected'
    if export_type == 'selected' and selected_ids:
        requests_query = requests_query.filter(id__in=selected_ids)
    
    # Order by creation date (newest first)
    requests = requests_query.order_by('-created_at')
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Access Requests")
    
    # Define styles
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    # Fonts
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=12)
    title_font = Font(name='Arial', bold=True, size=14)
    normal_font = Font(name='Arial', size=10)
    id_font = Font(name='Arial', bold=True, size=10)
    date_font = Font(name='Arial', size=10)
    small_font = Font(name='Arial', size=9)
    
    # Fills
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    odd_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Status colors
    status_colors = {
        'pending': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'approved': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'rejected': PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),
        'cancelled': PatternFill(start_color="D6D8DB", end_color="D6D8DB", fill_type="solid"),
    }
    
    # Admin status colors
    admin_status_colors = {
        'pending': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'granted': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'denied': PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),
        'in_progress': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
    }
    
    # Request type colors
    request_type_colors = {
        'grant': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'revoke': PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid"),
    }
    
    # Alignments
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # Define column headers
    headers = [
        str(_("ID")),
        str(_("Request Type")),
        str(_("Company")),
        str(_("System")),
        str(_("Object")),
        str(_("Environment")),
        str(_("Requested For")),
        str(_("Requested By")),
        str(_("Access Right")),
        str(_("Roles")),
        str(_("Status")),
        str(_("Admin Status")),
        str(_("Start Date")),
        str(_("End Date")),
        str(_("Created At")),
        str(_("Justification")),
        str(_("Requirements")),
        str(_("Notes")),
        str(_("Admin Comment")),
        str(_("Third Party")),
        str(_("Approvers")),
        str(_("Attachments"))
    ]
    
    # Add title row with metadata
    export_type_text = _("Selected Records") if export_type == 'selected' else _("All Records")
    title = f"{_('Access Requests Export')} - {export_type_text} - {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.merge_cells('A1:V1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    title_cell.fill = title_fill
    
    # Add filter info if applied
    filter_row = 2
    applied_filters = []
    
    for key, value in request.GET.items():
        if key not in ['export_type', 'selected_ids[]'] and value:
            if key == 'company':
                try:
                    company = Company.objects.get(id=value)
                    applied_filters.append(f"{_('Company')}: {company.name}")
                except Company.DoesNotExist:
                    pass
            elif key == 'system':
                try:
                    system = InformationAsset.objects.get(id=value)
                    applied_filters.append(f"{_('System')}: {system.name}")
                except InformationAsset.DoesNotExist:
                    pass
            elif key == 'status':
                status_display = dict(AccessRequest.STATUS_CHOICES).get(value, value)
                applied_filters.append(f"{_('Status')}: {status_display}")
            elif key == 'admin_status':
                admin_status_display = dict(AccessRequest.ADMIN_STATUS_CHOICES).get(value, value)
                applied_filters.append(f"{_('Admin Status')}: {admin_status_display}")
            elif key == 'request_type':
                request_type_display = dict(AccessRequest.REQUEST_TYPE_CHOICES).get(value, value)
                applied_filters.append(f"{_('Request Type')}: {request_type_display}")
            elif key == 'environment':
                env_display = str(dict(AccessRequest.ENVIRONMENT_CHOICES).get(value, value))
                applied_filters.append(f"{_('Environment')}: {env_display}")
            elif key == 'search':
                applied_filters.append(f"{_('Search')}: {value}")
    
    if applied_filters:
        filter_text = f"{_('Applied Filters')}: {', '.join(applied_filters)}"
        ws.merge_cells(f'A{filter_row}:V{filter_row}')
        filter_cell = ws.cell(row=filter_row, column=1, value=filter_text)
        filter_cell.font = small_font
        filter_cell.alignment = left_alignment
        filter_row += 1
    
    # Add total records count
    count_text = f"{_('Total Records')}: {requests.count()}"
    ws.merge_cells(f'A{filter_row}:V{filter_row}')
    count_cell = ws.cell(row=filter_row, column=1, value=count_text)
    count_cell.font = Font(name='Arial', bold=True, size=11)
    count_cell.alignment = left_alignment
    
    # Start headers at appropriate row
    header_row = filter_row + 1
    
    # Write headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thick_bottom_border
    
    # Set column widths
    column_widths = [8, 12, 20, 25, 20, 12, 20, 20, 15, 25, 12, 15, 18, 18, 18, 40, 40, 40, 40, 25, 30, 15]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Write data to worksheet
    row_num = header_row + 1
    for idx, req in enumerate(requests):
        # Determine row fill (alternating colors)
        row_fill = odd_row_fill if idx % 2 == 0 else even_row_fill
        
        # ID
        cell = ws.cell(row=row_num, column=1, value=req.id)
        cell.font = id_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Request Type
        request_type_display = dict(AccessRequest.REQUEST_TYPE_CHOICES).get(req.request_type, req.request_type)
        cell = ws.cell(row=row_num, column=2, value=str(request_type_display))
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = center_alignment
        cell.fill = request_type_colors.get(req.request_type, row_fill)
        cell.border = thin_border
        
        # Company
        cell = ws.cell(row=row_num, column=3, value=req.company.name if req.company else '-')
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # System
        cell = ws.cell(row=row_num, column=4, value=req.system.name if req.system else '-')
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Object
        object_name = '-'
        if req.access_record and req.access_record.access_object:
            object_name = req.access_record.access_object.get_name(current_language) or req.access_record.access_object.get_name('ua')
        
        cell = ws.cell(row=row_num, column=5, value=object_name)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Environment
        env_display = str(dict(AccessRequest.ENVIRONMENT_CHOICES).get(req.environment, req.environment))
        cell = ws.cell(row=row_num, column=6, value=str(env_display))
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Requested For
        requested_for_name = '-'
        if req.requested_for:
            if hasattr(req.requested_for, 'cabinet') and req.requested_for.cabinet:
                requested_for_name = f"{req.requested_for.cabinet.full_name} ({req.requested_for.username})"
            else:
                requested_for_name = f"{req.requested_for.get_full_name()} ({req.requested_for.username})"
        
        cell = ws.cell(row=row_num, column=7, value=requested_for_name)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Requested By
        requested_by_name = '-'
        if req.requested_by:
            if hasattr(req.requested_by, 'cabinet') and req.requested_by.cabinet:
                requested_by_name = f"{req.requested_by.cabinet.full_name} ({req.requested_by.username})"
            else:
                requested_by_name = f"{req.requested_by.get_full_name()} ({req.requested_by.username})"
        
        cell = ws.cell(row=row_num, column=8, value=requested_by_name)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Access Right
        access_right_name = '-'
        if req.access_record and req.access_record.access_right:
            access_right_name = req.access_record.access_right.get_name(current_language) or req.access_record.access_right.get_name('ua')
        
        cell = ws.cell(row=row_num, column=9, value=access_right_name)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Roles
        roles_text = '-'
        if req.access_record:
            roles = req.access_record.roles.all()
            if roles:
                role_names = []
                for role in roles:
                    if current_language == 'uk':
                        role_names.append(role.get_name() or role.name or '')
                    elif current_language == 'ru':
                        role_names.append(role.get_name() or role.name or '')
                    else:
                        role_names.append(role.get_name() or role.name or '')
                roles_text = ', '.join(role_names)
        
        cell = ws.cell(row=row_num, column=10, value=roles_text)
        cell.font = normal_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Status
        status_display = dict(AccessRequest.STATUS_CHOICES).get(req.status, req.status)
        cell = ws.cell(row=row_num, column=11, value=str(status_display))
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = center_alignment
        cell.fill = status_colors.get(req.status, row_fill)
        cell.border = thin_border
        
        # Admin Status
        admin_status_display = dict(AccessRequest.ADMIN_STATUS_CHOICES).get(req.admin_status, req.admin_status)
        cell = ws.cell(row=row_num, column=12, value=str(admin_status_display))
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = center_alignment
        cell.fill = admin_status_colors.get(req.admin_status, row_fill)
        cell.border = thin_border
        
        # Start Date
        if req.start_date:
            start_date_naive = req.start_date.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=13, value=start_date_naive)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=13, value='-')
        cell.font = date_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # End Date
        if req.end_date:
            end_date_naive = req.end_date.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=14, value=end_date_naive)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=14, value='-')
        cell.font = date_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Created At
        if req.created_at:
            created_at_naive = req.created_at.replace(tzinfo=None)
            cell = ws.cell(row=row_num, column=15, value=created_at_naive)
            cell.number_format = 'DD.MM.YYYY HH:MM'
        else:
            cell = ws.cell(row=row_num, column=15, value='-')
        cell.font = date_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Justification
        cell = ws.cell(row=row_num, column=16, value=req.justification or '-')
        cell.font = small_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Requirements
        cell = ws.cell(row=row_num, column=17, value=req.requirements or '-')
        cell.font = small_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Notes
        cell = ws.cell(row=row_num, column=18, value=req.notes or '-')
        cell.font = small_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Admin Comment
        cell = ws.cell(row=row_num, column=19, value=req.admin_comment or '-')
        cell.font = small_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Third Party Information
        third_party_info = []
        if req.third_party_first_name or req.third_party_last_name:
            third_party_info.append(f"{req.third_party_first_name} {req.third_party_last_name}".strip())
        if req.third_party_email:
            third_party_info.append(req.third_party_email)
        if req.third_party_organization:
            third_party_info.append(req.third_party_organization)
        
        # Add third party users
        if req.third_party_users.exists():
            for tp_user in req.third_party_users.all():
                third_party_info.append(f"{tp_user.full_name} ({tp_user.email})")
        
        third_party_text = '\n'.join(third_party_info) if third_party_info else '-'
        cell = ws.cell(row=row_num, column=20, value=third_party_text)
        cell.font = small_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Approvers
        approvers_info = []
        if req.request_approvers.exists():
            for approver in req.request_approvers.all():
                status_text = dict(AccessRequestApprover.APPROVING_STATUS_CHOICES).get(approver.current_status, approver.current_status)
                approvers_info.append(f"{approver.cabinet_user.full_name} ({status_text})")
        
        approvers_text = '\n'.join(approvers_info) if approvers_info else '-'
        cell = ws.cell(row=row_num, column=21, value=approvers_text)
        cell.font = small_font
        cell.alignment = left_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        # Attachments
        attachments_count = req.attachments.count()
        attachments_text = f"{attachments_count} {_('files')}" if attachments_count > 0 else '-'
        cell = ws.cell(row=row_num, column=22, value=attachments_text)
        cell.font = normal_font
        cell.alignment = center_alignment
        cell.fill = row_fill
        cell.border = thin_border
        
        row_num += 1
    
    # Auto-fit row heights for multiline content
    for row in ws.iter_rows(min_row=header_row + 1, max_row=row_num - 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                lines = len(cell.value.split('\n'))
                ws.row_dimensions[cell.row].height = max(15, lines * 12)
    
    # Create the HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    export_type_suffix = 'selected' if export_type == 'selected' else 'all'
    filename = f"access_requests_{export_type_suffix}_{timestamp}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def get_access_records_with_sequence(request_id):
    """
    Отримує Access Records з урахуванням послідовності grant/revoke
    """
    try:
        from django.db.models import Q
        from django.utils import timezone
        
        # Отримуємо основний запит
        main_request = AccessRequest.objects.select_related(
            'requested_for', 'requested_by', 'company', 'system'
        ).prefetch_related(
            'access_records__roles',
            'access_records__access_object'
        ).get(id=request_id)
        
        # Визначаємо користувача, для якого надається доступ
        target_user = None
        if main_request.third_party_first_name or main_request.third_party_last_name:
            # Третя сторона
            target_user = {
                'type': 'third_party',
                'name': f"{main_request.third_party_first_name} {main_request.third_party_last_name}".strip(),
                'email': main_request.third_party_email,
                'organization': main_request.third_party_organization,
                'phone': main_request.third_party_phone
            }
        else:
            # Кабінетний користувач
            cabinet_user = main_request.requested_for.cabinet if hasattr(main_request.requested_for, 'cabinet') else None
            target_user = {
                'type': 'cabinet',
                'name': main_request.requested_for.get_full_name(),
                'email': main_request.requested_for.email,
                'department': cabinet_user.department.get_name() if cabinet_user and cabinet_user.department else None,
                'position': cabinet_user.position.get_name() if cabinet_user and cabinet_user.position else None,
                'user_id': main_request.requested_for.id
            }
        
        # Логуємо інформацію про основного користувача
        logger.debug(f"Main request {main_request.id}: target_user={target_user.get('name', 'Unknown')} "
                   f"(email: {target_user.get('email', 'Unknown')}, type: {target_user.get('type', 'Unknown')})")
        
        # Отримуємо всі запити для цього користувача в цій системі
        # Використовуємо email як унікальний ідентифікатор для всіх користувачів
        user_requests = AccessRequest.objects.filter(
            system=main_request.system,
            status__in=['approved', 'pending'],
            admin_status__in=['granted', 'pending']
        ).order_by('created_at')
        
        # Фільтруємо запити за email користувача
        if target_user['type'] == 'cabinet':
            user_requests = user_requests.filter(requested_for__email=target_user['email'])
        else:
            user_requests = user_requests.filter(third_party_email=target_user['email'])
        
        # Логуємо кількість знайдених запитів
        logger.debug(f"Found {user_requests.count()} requests for user {target_user.get('email', 'Unknown')} "
                   f"in system {main_request.system.id}")
        
        # Логуємо деталі кожного запиту
        for req in user_requests:
            req_user_email = None
            if req.third_party_email:
                req_user_email = req.third_party_email
                req_user_name = f"{req.third_party_first_name} {req.third_party_last_name}".strip()
            else:
                req_user_email = req.requested_for.email if req.requested_for else None
                req_user_name = req.requested_for.get_full_name() if req.requested_for else "Unknown"
            
            logger.debug(f"Request {req.id}: {req_user_name} ({req_user_email}) - {req.request_type}")
        
        # Збираємо послідовність подій
        access_sequence = []
        
        for req in user_requests:
            # Логуємо інформацію про кожен запит
            req_user_email = None
            if req.third_party_email:
                req_user_email = req.third_party_email
                req_user_name = f"{req.third_party_first_name} {req.third_party_last_name}".strip()
            else:
                req_user_email = req.requested_for.email if req.requested_for else None
                req_user_name = req.requested_for.get_full_name() if req.requested_for else "Unknown"
            
            logger.debug(f"Processing request {req.id}: user={req_user_name} (email: {req_user_email})")
            
            # Перевіряємо, чи запит належить правильному користувачу
            if req_user_email != target_user.get('email'):
                logger.warning(f"Request {req.id} belongs to user {req_user_email} but we're processing for {target_user.get('email')} - skipping")
                continue
            
            for access_record in req.access_records.all():
                for role in access_record.roles.all():
                    # Визначаємо стан для конкретного запиту (req)
                    # Спочатку встановлюємо базовий стан на основі типу запиту
                    if req.status == 'approved' and req.admin_status == 'granted':
                        if req.request_type == 'grant':
                            role_state = 'granted'
                        elif req.request_type == 'revoke':
                            role_state = 'revoked'
                    elif req.status == 'approved' and req.admin_status == 'denied':
                        role_state = 'denied'
                    else:
                        role_state = 'pending'
                    
                    # Тепер перевіряємо, чи є пізніші операції, які змінюють стан
                    # Знаходимо всі запити для цього користувача з цією роллю
                    # Використовуємо email як унікальний ідентифікатор
                    user_role_requests = AccessRequest.objects.filter(
                        system=main_request.system,
                        access_records__roles=role,
                        status__in=['approved', 'pending'],
                        admin_status__in=['granted', 'pending']
                    ).order_by('created_at')
                    
                    # Фільтруємо запити за email користувача
                    if target_user['type'] == 'cabinet':
                        user_role_requests = user_role_requests.filter(requested_for__email=target_user['email'])
                    else:
                        user_role_requests = user_role_requests.filter(third_party_email=target_user['email'])
                    
                    # Логуємо кількість знайдених запитів
                    logger.debug(f"Found {user_role_requests.count()} requests for user {target_user.get('email', 'Unknown')} "
                               f"with role {role.id} in system {main_request.system.id}")
                    
                    # Знаходимо останній grant та revoke для цього користувача
                    last_grant_time = None
                    last_revoke_time = None
                    last_revoke_request = None
                    
                    # Логуємо деталі кожного запиту для цієї ролі
                    for role_req in user_role_requests:
                        req_user_email = None
                        if role_req.third_party_email:
                            req_user_email = role_req.third_party_email
                            req_user_name = f"{role_req.third_party_first_name} {role_req.third_party_last_name}".strip()
                        else:
                            req_user_email = role_req.requested_for.email if role_req.requested_for else None
                            req_user_name = role_req.requested_for.get_full_name() if role_req.requested_for else "Unknown"
                        
                        logger.debug(f"Role request {role_req.id}: {req_user_name} ({req_user_email}) - {role_req.request_type}")
                        
                        if role_req.status == 'approved' and role_req.admin_status == 'granted':
                            if role_req.request_type == 'grant':
                                last_grant_time = role_req.created_at
                                logger.debug(f"  -> Grant at {last_grant_time}")
                            elif role_req.request_type == 'revoke':
                                last_revoke_time = role_req.created_at
                                last_revoke_request = role_req
                                logger.debug(f"  -> Revoke at {last_revoke_time}")
                    
                    # Визначаємо фінальний стан на основі хронології
                    # Але тільки якщо це не конкретний запит, який ми обробляємо
                    # Для конкретного запиту стан повинен відповідати його типу
                    if req.request_type == 'grant' and req.status == 'approved' and req.admin_status == 'granted':
                        # Це grant запит - стан повинен бути granted
                        role_state = 'granted'
                    elif req.request_type == 'revoke' and req.status == 'approved' and req.admin_status == 'granted':
                        # Це revoke запит - стан повинен бути revoked
                        role_state = 'revoked'
                    else:
                        # Для інших випадків використовуємо хронологічну логіку
                        if last_grant_time and last_revoke_time:
                            if last_grant_time > last_revoke_time:
                                # Останній grant після revoke - доступ відновлений
                                role_state = 'granted'
                            else:
                                # Останній revoke після grant - доступ скасований
                                role_state = 'revoked'
                        elif last_grant_time and not last_revoke_time:
                            # Є тільки grant - доступ наданий
                            role_state = 'granted'
                        elif last_revoke_time and not last_grant_time:
                            # Є тільки revoke - доступ скасований
                            role_state = 'revoked'
                    
                    # Логуємо для діагностики
                    logger.debug(f"Request {req.id}, Role {role.id}, User {target_user.get('name', 'Unknown')} "
                               f"(email: {target_user.get('email', 'Unknown')}): "
                               f"last_grant={last_grant_time}, last_revoke={last_revoke_time}, final_state={role_state}")
                    
                    # Додаткова перевірка - чи запит дійсно належить цьому користувачу
                    req_user_email = None
                    if req.third_party_email:
                        req_user_email = req.third_party_email
                    else:
                        req_user_email = req.requested_for.email if req.requested_for else None
                    
                    if req_user_email != target_user.get('email'):
                        logger.warning(f"SKIPPING: Request {req.id} belongs to {req_user_email} but processing for {target_user.get('email')}")
                        continue
                    
                    # Логуємо, що запит пройшов перевірку
                    logger.debug(f"PROCESSING: Request {req.id} belongs to {req_user_email} - matches target user {target_user.get('email')}")
                    
                    # Логуємо фінальний стан для цього запиту
                    logger.debug(f"FINAL STATE: Request {req.id}, Role {role.id} -> {role_state}")
                    
                    access_sequence.append({
                        'request_id': req.id,
                        'request_type': req.request_type,
                        'role_id': role.id,
                        'role_name': role.get_name(),
                        'role_color': role.color or '#6c757d',
                        'access_record_id': access_record.id,
                        'object_name': access_record.access_object.get_name() if access_record.access_object else None,
                        'environment': req.environment,
                        'state': role_state,
                        'created_at': req.created_at,
                        'start_date': req.start_date,
                        'end_date': req.end_date,
                        'justification': req.justification,
                        'requested_by': req.requested_by.get_full_name(),
                        'revoked_at': last_revoke_time,
                        'revoked_by': last_revoke_request.requested_by.get_full_name() if last_revoke_request else None
                    })
        
        return {
            'target_user': target_user,
            'access_sequence': access_sequence,
            'main_request': main_request
        }
        
    except Exception as e:
        logger.error(f"Error getting access records with sequence: {str(e)}", exc_info=True)
        return None


@login_required
def get_access_sequence_view(request, request_id):
    """
    API endpoint для отримання послідовності Access Records
    """
    try:
        sequence_data = get_access_records_with_sequence(request_id)
        
        if sequence_data is None:
            return JsonResponse({
                'status': 'error',
                'message': _('Failed to get access sequence data')
            }, status=400)
        
        return JsonResponse({
            'status': 'success',
            'data': sequence_data
        })
        
    except Exception as e:
        logger.error(f"Error in get_access_sequence_view: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)





