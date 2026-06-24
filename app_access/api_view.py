import json
import requests
import base64
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.utils import timezone
from django.db import transaction, IntegrityError
from .models import ApiCredential, ApiUser, ApiUserRole, ApiUserMerchant, ApiUserRoleMapping, ApiUserStatus, ApiUserPermissionHistory, ApiSyncStatus, ApiUserLoginHistory, ApiUserMerchantLink, ScheduledSync
from django.conf import settings
from django.core.paginator import Paginator
import logging
import math
import time
from django.views.decorators.http import require_POST
from django.db import models
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from redis import Redis
import uuid
from datetime import datetime
from io import BytesIO
from django.views.decorators.cache import never_cache
from django.utils.decorators import decorator_from_middleware_with_args
from django.middleware.locale import LocaleMiddleware
from django.views.decorators.http import require_GET, require_http_methods
from django.db import models
from django.utils import timezone

logger = logging.getLogger(__name__)
SYNC_HTTP_TIMEOUT = (10, 20)
# Sub-steps per user while syncing (roles, status, permission history, login history, merchants)
USER_SYNC_PHASES = 5

def get_user_access_request_statuses(api_user, credentials):
    """
    Get access request statuses for an API user based on their email and role mappings
    Returns a list of access request statuses for each role mapping
    
    Matching logic:
    1. Company from credentials matches Access Request company
    2. Information System from credentials matches Access Request system  
    3. Environment from credentials matches Access Request environment
    4. Merchant from API matches Object from Access Request
    5. Role from API matches Role from Access Request
    6. Email matches between API user and Access Request user
    
    Filters:
    - Access Type: Grant and Revoke (to show complete access history)
    - Request Status: Approved, Pending only
    - Admin Status: pending, in_progress, granted, denied only
    """
    from app_access.models import AccessRequest, ThirdPartyUser
    from app_conf.models import Company
    from app_asset.models import InformationAsset
    from django.contrib.auth.models import User
    
    statuses = []
    
    try:
        # Find Django User by email (for requested_for field)
        django_user = None
        try:
            django_user = User.objects.get(email=api_user.email)
        except User.DoesNotExist:
            pass
        
        # Find Third Party User by email
        third_party_user = None
        try:
            third_party_user = ThirdPartyUser.objects.get(email=api_user.email)
        except ThirdPartyUser.DoesNotExist:
            pass
        
        # Note: We continue processing even if no Django/ThirdParty user found
        # to show role mappings without corresponding Access Requests
        
        # Get company, information system, and environment from credentials
        companies_from_credentials = set()
        systems_from_credentials = set()
        environments_from_credentials = set()
        
        for credential in credentials:
            if credential.company:
                companies_from_credentials.add(credential.company)
            if credential.information_system:
                systems_from_credentials.add(credential.information_system)
            if credential.environment:
                environments_from_credentials.add(credential.environment)
        
        # For each role mapping, find corresponding access requests with proper matching
        for role_mapping in api_user.role_mappings.all():
            merchant_name = role_mapping.merchant.name  # This is Object name
            role_name = role_mapping.role.name
            
            # Build base query for Access Requests (Grant and Revoke)
            base_query = AccessRequest.objects.filter(
                # Filter by Access Type - Grant and Revoke requests to show complete history
                request_type__in=['grant', 'revoke']
            ).filter(
                # Filter by Request Status - only Approved and Pending
                status__in=['approved', 'pending']
            ).filter(
                # Filter by Admin Status - include all relevant statuses
                admin_status__in=['pending', 'in_progress', 'granted', 'denied']
            )
            
            # Apply credential filters if available
            if companies_from_credentials:
                base_query = base_query.filter(company__in=companies_from_credentials)
            if systems_from_credentials:
                base_query = base_query.filter(system__in=systems_from_credentials)
            if environments_from_credentials:
                base_query = base_query.filter(environment__in=environments_from_credentials)
            
            # Find matching requests for this user
            matching_requests = []
            
            # Search for Django user requests (only if Django user exists)
            if django_user:
                django_requests = base_query.filter(requested_for=django_user)
                matching_requests.extend(list(django_requests))
            
            # Search for Third Party user requests (only if Third Party user exists)
            if third_party_user:
                third_party_requests = base_query.filter(third_party_users=third_party_user)
                matching_requests.extend(list(third_party_requests))
            
            # Always search by email in third party fields (legacy support)
            email_requests = base_query.filter(third_party_email=api_user.email)
            matching_requests.extend(list(email_requests))
            
            # Remove duplicates
            unique_requests = {}
            for request_obj in matching_requests:
                if request_obj.id not in unique_requests:
                    unique_requests[request_obj.id] = request_obj
            
            # Filter requests that match Object (Merchant) and Role
            # Collect ALL matching requests, not just the first one
            matched_requests = []
            for access_request in unique_requests.values():
                # Check if this request matches our Merchant (Object) and Role
                # First, collect all access records for this request
                access_records = []
                if access_request.access_records.exists():
                    access_records = list(access_request.access_records.all())
                elif access_request.access_record:
                    # Fallback to single access_record for backward compatibility
                    access_records = [access_request.access_record]
                
                for access_record in access_records:
                    # Check Object match (Merchant from API = Object from Access Request)
                    object_match = False
                    if access_record.access_object:
                        # Check for EXACT match between merchant name and object name
                        object_name_ua = access_record.access_object.get_name() or ""
                        object_name_en = access_record.access_object.get_name('en') or ""
                        object_name_ru = access_record.access_object.get_name('ru') or ""
                        
                        if (merchant_name.lower() == object_name_ua.lower() or
                            merchant_name.lower() == object_name_en.lower() or
                            merchant_name.lower() == object_name_ru.lower()):
                            object_match = True
                    else:
                        # If no specific object, consider it a system-level request
                        # EXACT match against company or system name
                        company_name = access_request.company.name if access_request.company else ""
                        system_name = access_request.system.name if access_request.system else ""
                        
                        if (merchant_name.lower() == company_name.lower() or
                            merchant_name.lower() == system_name.lower()):
                            object_match = True
                    
                    # Check Role match - EXACT match
                    role_match = False
                    if access_record.roles.exists():
                        for role in access_record.roles.all():
                            role_name_ua = role.get_name() or role.name or ""
                            role_name_en = role.accessrole_name_en or ""
                            role_name_ru = role.get_name() or role.name or ""
                            
                            if (role_name.lower() == role_name_ua.lower() or
                                role_name.lower() == role_name_en.lower() or
                                role_name.lower() == role_name_ru.lower()):
                                role_match = True
                                break
                    
                    # If both Object and Role match, add this combination
                    if object_match and role_match:
                        matched_requests.append({
                            'request': access_request,
                            'access_record': access_record
                        })
                        break  # Only need one match per access_request
            
            # Group matched requests by access record to avoid duplicates
            if matched_requests:
                # Group requests by access record ID
                grouped_requests = {}
                for match in matched_requests:
                    access_record = match['access_record']
                    access_record_id = access_record.id if access_record else None
                    
                    if access_record_id not in grouped_requests:
                        grouped_requests[access_record_id] = []
                    grouped_requests[access_record_id].append(match)
                
                # Process each group to show consolidated status
                for access_record_id, matches in grouped_requests.items():
                    # Sort requests by creation date (newest first) and prioritize revoke requests
                    matches.sort(key=lambda x: (x['request'].created_at, x['request'].request_type == 'revoke'), reverse=True)
                    
                    # Use the most recent request for basic info, but consolidate status
                    primary_match = matches[0]
                    matched_request = primary_match['request']
                    access_record = primary_match['access_record']
                    
                    # Analyze the chronological sequence of all requests to determine final status
                    # Sort all requests by creation date (newest first) to get chronological order
                    all_requests = [match['request'] for match in matches]
                    all_requests.sort(key=lambda x: x.created_at, reverse=True)
                    
                    # Find the most recent granted request to determine final status
                    most_recent_granted_grant = None
                    most_recent_granted_revoke = None
                    
                    for request in all_requests:
                        if request.admin_status == 'granted':
                            if request.request_type == 'grant' and not most_recent_granted_grant:
                                most_recent_granted_grant = request
                            elif request.request_type == 'revoke' and not most_recent_granted_revoke:
                                most_recent_granted_revoke = request
                    
                    # Keep the old logic for backward compatibility
                    revoke_request = most_recent_granted_revoke
                    grant_request = most_recent_granted_grant
                    
                    # Use grant request for basic info if available, otherwise use primary
                    if grant_request:
                        matched_request = grant_request
                    
                    # Determine the user type for this request
                    user_type = "Unknown"
                    if django_user and matched_request.requested_for == django_user:
                        user_type = "Cabinet User"
                    elif third_party_user and third_party_user in matched_request.third_party_users.all():
                        user_type = "Third Party User"
                    elif matched_request.third_party_email == api_user.email:
                        user_type = "Third Party (Legacy)"
                    
                    # Get object name for display
                    object_display = ""
                    if access_record and access_record.access_object:
                        object_display = (access_record.access_object.get_name() or 
                                        access_record.access_object.get_name('en') or 
                                        access_record.access_object.get_name('ua') or 
                                        "Unknown Object")
                    
                    # Get requested_by information
                    requested_by_name = ""
                    if matched_request.requested_by:
                        requested_by_name = f"{matched_request.requested_by.first_name or ''} {matched_request.requested_by.last_name or ''}".strip()
                        if not requested_by_name:
                            requested_by_name = matched_request.requested_by.username
                    
                    # Get access type (request_type)
                    access_type = matched_request.request_type or ''
                    
                    # Get approving persons
                    approving_persons = []
                    if access_record:
                        # Get approving persons from the access record
                        for approver in access_record.approvers.all().order_by('order'):
                            approver_name = f"{approver.cabinet_user.user.first_name or ''} {approver.cabinet_user.user.last_name or ''}".strip()
                            if not approver_name:
                                approver_name = approver.cabinet_user.user.username
                            approving_persons.append({
                                'name': approver_name,
                                'order': approver.order,
                                'status': approver.current_status
                            })
                    
                    # Check if this access record is revoked
                    is_revoked = False
                    revoke_info = None
                    access_record_status = None
                    
                    if access_record:
                        # Primary check: access record status is "Revoked"
                        if hasattr(access_record, 'status') and access_record.status:
                            if access_record.status and (access_record.status.name or '') == 'Revoked':
                                is_revoked = True
                                access_record_status = 'Revoked'
                                # Get revoke information from status history if available
                                if hasattr(access_record, 'status_history'):
                                    latest_revoke_history = access_record.status_history.filter(
                                        new_status__name='Revoked'
                                    ).order_by('-changed_at').first()
                                    
                                    if latest_revoke_history:
                                        # Визначаємо, для кого був скасований доступ
                                        revoked_for = None
                                        if latest_revoke_history.revoke_request:
                                            if latest_revoke_history.revoke_request.requested_for:
                                                revoked_for = latest_revoke_history.revoke_request.requested_for.get_full_name()
                                            elif latest_revoke_history.revoke_request.third_party_email:
                                                revoked_for = f"{latest_revoke_history.revoke_request.third_party_first_name} {latest_revoke_history.revoke_request.third_party_last_name}".strip()
                                        else:
                                            # Спробуємо витягти з change_reason
                                            import re
                                            user_match = re.search(r'user\s+([^,\s]+(?:\s+[^,\s]+)*)', latest_revoke_history.change_reason or '')
                                            if user_match:
                                                revoked_for = user_match.group(1)
                                        
                                        # Правильно визначаємо "Revoked for" з revoke запиту
                                        revoked_for_correct = None
                                        if latest_revoke_history.revoke_request:
                                            revoke_request = latest_revoke_history.revoke_request
                                            # Пріоритет: third_party_* поля з revoke запиту
                                            if revoke_request.third_party_email:
                                                revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                            elif revoke_request.requested_for:
                                                revoked_for_correct = revoke_request.requested_for.get_full_name()
                                        else:
                                            revoked_for_correct = revoked_for or 'Unknown'
                                        
                                        revoke_info = {
                                            'revoked_at': latest_revoke_history.changed_at,
                                            'revoked_by': latest_revoke_history.changed_by.get_full_name() if latest_revoke_history.changed_by else 'Unknown',
                                            'revoked_for': revoked_for_correct,
                                            'change_reason': latest_revoke_history.change_reason or 'No reason provided'
                                        }
                            else:
                                access_record_status = (access_record.status.name if access_record.status else None) or str(access_record.status)
                        
                        # Secondary check: access record is marked as inactive
                        if not is_revoked and hasattr(access_record, 'is_active') and not access_record.is_active:
                            is_revoked = True
                            access_record_status = 'Inactive/Revoked'
                            # Try to get revoke information from status history
                            if hasattr(access_record, 'status_history'):
                                latest_history = access_record.status_history.order_by('-changed_at').first()
                                if latest_history:
                                    # Визначаємо, для кого був скасований доступ
                                    revoked_for = None
                                    if latest_history.revoke_request:
                                        if latest_history.revoke_request.requested_for:
                                            revoked_for = latest_history.revoke_request.requested_for.get_full_name()
                                        elif latest_history.revoke_request.third_party_email:
                                            revoked_for = f"{latest_history.revoke_request.third_party_first_name} {latest_history.revoke_request.third_party_last_name}".strip()
                                    else:
                                        # Спробуємо витягти з change_reason
                                        import re
                                        user_match = re.search(r'user\s+([^,\s]+(?:\s+[^,\s]+)*)', latest_history.change_reason or '')
                                        if user_match:
                                            revoked_for = user_match.group(1)
                                    
                                    # Правильно визначаємо "Revoked for" з revoke запиту
                                    revoked_for_correct = None
                                    if latest_history.revoke_request:
                                        revoke_request = latest_history.revoke_request
                                        # Пріоритет: third_party_* поля з revoke запиту
                                        if revoke_request.third_party_email:
                                            revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                        elif revoke_request.requested_for:
                                            revoked_for_correct = revoke_request.requested_for.get_full_name()
                                    else:
                                        revoked_for_correct = revoked_for or 'Unknown'
                                    
                                    revoke_info = {
                                        'revoked_at': latest_history.changed_at,
                                        'revoked_by': latest_history.changed_by.get_full_name() if latest_history.changed_by else 'Unknown',
                                        'revoked_for': revoked_for_correct,
                                        'change_reason': latest_history.change_reason or 'Access record deactivated'
                                    }
                        
                        # Tertiary check: if grant request is approved but user is not in access_users
                        if (not is_revoked and 
                            matched_request.request_type == 'grant' and 
                              matched_request.admin_status == 'granted' and 
                              matched_request.requested_for):
                            # Check if user is still in access_users
                            if not access_record.access_users.filter(id=matched_request.requested_for.id).exists():
                                # Check if user was in access_users but removed (revoked)
                                if hasattr(access_record, 'status_history'):
                                    user_revoke_history = access_record.status_history.filter(
                                        change_reason__icontains=f"user {matched_request.requested_for.get_full_name()}"
                                    ).order_by('-changed_at').first()
                                    
                                    if user_revoke_history:
                                        is_revoked = True
                                        access_record_status = 'User Access Revoked'
                                        # Визначаємо, для кого був скасований доступ
                                        revoked_for = matched_request.requested_for.get_full_name() if matched_request.requested_for else 'Unknown'
                                        
                                        # Правильно визначаємо "Revoked for" з revoke запиту
                                        revoked_for_correct = None
                                        if user_revoke_history.revoke_request:
                                            revoke_request = user_revoke_history.revoke_request
                                            # Пріоритет: third_party_* поля з revoke запиту
                                            if revoke_request.third_party_email:
                                                revoked_for_correct = f"{revoke_request.third_party_first_name} {revoke_request.third_party_last_name}".strip()
                                            elif revoke_request.requested_for:
                                                revoked_for_correct = revoke_request.requested_for.get_full_name()
                                        else:
                                            revoked_for_correct = revoked_for
                                        
                                        revoke_info = {
                                            'revoked_at': user_revoke_history.changed_at,
                                            'revoked_by': user_revoke_history.changed_by.get_full_name() if user_revoke_history.changed_by else 'Unknown',
                                            'revoked_for': revoked_for_correct,
                                            'change_reason': user_revoke_history.change_reason or 'User access revoked'
                                        }
                    
                    # Determine final access and admin status based on consolidated requests
                    final_access_status = access_record_status
                    final_admin_status = matched_request.admin_status
                    
                    # Determine final status based on chronological sequence
                    # Priority: Check if there's a more recent grant after the most recent revoke
                    if (most_recent_granted_grant and most_recent_granted_revoke and 
                        most_recent_granted_grant.created_at > most_recent_granted_revoke.created_at):
                        # Grant is more recent than revoke - access is re-granted
                        final_access_status = 'Granted'
                        final_admin_status = 'granted'
                    elif revoke_request and revoke_request.admin_status == 'granted':
                        # Most recent revoke is granted and no newer grant - access is revoked
                        final_access_status = 'Revoked'
                        final_admin_status = 'Access Revoked'
                    elif revoke_request and revoke_request.admin_status in ['pending', 'in_progress']:
                        # Revoke is pending
                        final_access_status = 'Revoke Pending'
                        final_admin_status = revoke_request.admin_status
                    elif is_revoked:
                        # For access records revoked through other means (not via revoke request)
                        final_access_status = 'Revoked'
                        if matched_request.admin_status == 'granted':
                            final_admin_status = 'Access Revoked'
                    
                    statuses.append({
                        'merchant': merchant_name,
                        'role': role_name,
                        'request_status': matched_request.status,
                        'admin_status': final_admin_status,
                        'request_id': matched_request.id,
                        'company': matched_request.company.name if matched_request.company else '',
                        'system': matched_request.system.name if matched_request.system else '',
                        'environment': matched_request.environment or '',
                        'object': object_display,
                        'object_name': object_display,
                        'created_at': matched_request.created_at,
                        'modified_at': matched_request.modified_at,
                        'start_date': matched_request.start_date,
                        'end_date': matched_request.end_date,
                        'user_type': user_type,
                        'requested_by': requested_by_name,
                        'request_type': matched_request.request_type,
                        'access_type': access_type,
                        'justification': matched_request.justification or '',
                        'requirements': matched_request.requirements or '',
                        'notes': matched_request.notes or '',
                        'admin_comment': matched_request.admin_comment or '',
                        'third_party_first_name': matched_request.third_party_first_name or '',
                        'third_party_last_name': matched_request.third_party_last_name or '',
                        'third_party_email': matched_request.third_party_email or '',
                        'third_party_phone': matched_request.third_party_phone or '',
                        'third_party_organization': matched_request.third_party_organization or '',
                        'approving_persons': approving_persons,
                        # New fields for revoked access tracking
                        'access_record_id': access_record.id if access_record else None,
                        'access_record_status': final_access_status,
                        'is_revoked': (final_access_status in ['Revoked', 'Revoke Pending'] and 
                                     not (most_recent_granted_grant and most_recent_granted_revoke and 
                                          most_recent_granted_grant.created_at > most_recent_granted_revoke.created_at)),
                        'revoke_info': revoke_info,
                    })
            else:
                # No matching request found - show role mapping without request details
                statuses.append({
                    'merchant': merchant_name,
                    'role': role_name,
                    'request_status': None,
                    'admin_status': None,
                    'request_id': None,
                    'company': '',
                    'system': '',
                    'environment': '',
                    'object': '',
                    'object_name': '',
                    'created_at': None,
                    'modified_at': None,
                    'start_date': None,
                    'end_date': None,
                    'user_type': None,
                    'requested_by': '',
                    'request_type': '',
                    'access_type': '',
                    'justification': '',
                    'requirements': '',
                    'notes': '',
                    'admin_comment': '',
                    'third_party_first_name': '',
                    'third_party_last_name': '',
                    'third_party_email': '',
                    'third_party_phone': '',
                    'third_party_organization': '',
                    'approving_persons': [],
                })
    
    except Exception as e:
        logger.error(f"Error getting access request statuses for user {api_user.email}: {str(e)}")
    
    return statuses

@login_required
def api_request_page(request):
    """
    Render the API request page
    """
    # Check if user has permission to access API functionality
    from .matrix_view import has_access_api_permission, get_user_companies_for_api
    if not has_access_api_permission(request.user):
        messages.error(request, _("Access denied - you do not have permission to access API functionality. Please contact your administrator to grant you access rights."))
        return redirect('index')
    
    # Filter data based on user's allowed companies
    user_companies = get_user_companies_for_api(request.user)
    
    # Get user's saved credentials - filter by allowed companies if restrictions exist
    credentials = ApiCredential.objects.filter(user=request.user)
    if user_companies.exists():
        credentials = credentials.filter(company__in=user_companies)
    credentials = credentials.order_by('-is_default', 'name')
    default_credential = credentials.filter(is_default=True).first()
    
    # Get query parameters
    page_number = request.GET.get('page', 1)
    show_all = request.GET.get('show_all', False)
    search_query = request.GET.get('q', '').strip()
    
    # Get column filters
    filters = {}
    filter_params = {}
    for key, value in request.GET.items():
        if key.startswith('filter_') and value:
            field_name = key[7:]  # Remove 'filter_' prefix
            filters[field_name] = value
            filter_params[key] = value
    
    # Get synced users with pagination and include status and role mappings
    users = ApiUser.objects.select_related('status_info').prefetch_related(
        'role_mappings',
        'role_mappings__merchant',
        'role_mappings__role',
        'permission_history',
        'login_history',
        'merchant_links',
    ).all().order_by('-updated_at')
    
    # Apply search filter if provided
    if search_query:
        users = users.filter(
            # Search in multiple fields - case insensitive
            models.Q(email__icontains=search_query) |
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(hash__icontains=search_query) |
            models.Q(phone__icontains=search_query) |
            models.Q(user_id__icontains=search_query)
        )
    
    # Apply column filters
    if 'user_id' in filters:
        users = users.filter(user_id=filters['user_id'])
    if 'email' in filters:
        users = users.filter(email__icontains=filters['email'])
    if 'name' in filters:
        name_filter = filters['name']
        users = users.filter(
            models.Q(first_name__icontains=name_filter) | 
            models.Q(last_name__icontains=name_filter)
        )
    if 'phone' in filters:
        users = users.filter(phone__icontains=filters['phone'])
    if 'status' in filters:
        users = users.filter(status_info__status=filters['status'])
    
    # Date range filters for last_login
    if 'last_login_from' in filters and filters['last_login_from']:
        users = users.filter(last_login__date__gte=filters['last_login_from'])
    if 'last_login_to' in filters and filters['last_login_to']:
        users = users.filter(last_login__date__lte=filters['last_login_to'])
    
    # Date range filters for updated_at
    if 'updated_from' in filters and filters['updated_from']:
        users = users.filter(updated_at__date__gte=filters['updated_from'])
    if 'updated_to' in filters and filters['updated_to']:
        users = users.filter(updated_at__date__lte=filters['updated_to'])
    
    if 'merchant_name' in filters:
        merchant_filter = filters['merchant_name']
        users = users.filter(merchant_links__merchant_name__icontains=merchant_filter)
    if 'role_name' in filters:
        role_filter = filters['role_name']
        users = users.filter(role_mappings__role__name__icontains=role_filter)
    if 'merchant' in filters:
        merchant_filter = filters['merchant']
        users = users.filter(role_mappings__merchant__name__icontains=merchant_filter)
    
    # Apply distinct to remove duplicates that may occur due to joining with related models
    if any(key in filters for key in ['merchant_name', 'role_name', 'merchant']):
        users = users.distinct()
    
    # Get unique values for filter dropdowns
    unique_statuses = ApiUserStatus.objects.values_list('status', flat=True).distinct()
    
    # Get unique merchants and roles for dropdowns
    unique_merchants = ApiUserMerchant.objects.values_list('name', flat=True).distinct()
    unique_roles = ApiUserRole.objects.values_list('name', flat=True).distinct()

    # Build merchant -> roles mapping for "All Merchants" modal
    merchant_role_map = {}
    merchant_user_map = {}
    merchant_user_roles_map = {}
    merchant_role_rows = ApiUserRoleMapping.objects.select_related('merchant', 'role').values_list(
        'merchant__name',
        'role__name'
    )
    for merchant_name, role_name in merchant_role_rows:
        if not merchant_name or not role_name:
            continue
        merchant_role_map.setdefault(merchant_name, set()).add(role_name)

    merchant_user_role_rows = ApiUserRoleMapping.objects.select_related('merchant', 'role', 'user').values_list(
        'merchant__name',
        'user__email',
        'role__name'
    )
    for merchant_name, user_email, role_name in merchant_user_role_rows:
        if not merchant_name or not user_email or not role_name:
            continue
        merchant_user_roles_map.setdefault(merchant_name, {})
        merchant_user_roles_map[merchant_name].setdefault(user_email, set()).add(role_name)

    merchant_user_rows = ApiUserMerchantLink.objects.select_related('user').values_list(
        'merchant_name',
        'user__email'
    )
    for merchant_name, user_email in merchant_user_rows:
        if not merchant_name or not user_email:
            continue
        merchant_user_map.setdefault(merchant_name, set()).add(user_email)

    merchant_role_map = {
        merchant_name: sorted(role_names)
        for merchant_name, role_names in merchant_role_map.items()
    }
    merchant_user_map = {
        merchant_name: sorted(user_emails)
        for merchant_name, user_emails in merchant_user_map.items()
    }
    merchant_user_roles_map = {
        merchant_name: {
            user_email: sorted(role_names)
            for user_email, role_names in user_roles.items()
        }
        for merchant_name, user_roles in merchant_user_roles_map.items()
    }
    
    # Get total users count
    total_users = users.count()
    
    # Determine page size - use a large value if show_all is true
    page_size = 1000 if show_all else 20
    
    # Paginate the results
    paginator = Paginator(users, page_size)  # Show 20 users per page by default or 1000 if show_all
    page_obj = paginator.get_page(page_number)
    
    # Get companies and systems for dropdowns - filter by user permissions
    from app_conf.models import Company
    from app_asset.models import InformationAsset
    
    if user_companies.exists():
        companies = user_companies.order_by('name')
        information_systems = InformationAsset.objects.filter(
            company__in=user_companies,
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True  # Only include active assets
        ).order_by('name')
    else:
        companies = Company.objects.all().order_by('name')
        information_systems = InformationAsset.objects.filter(
            access_manage=True,  # Only include assets marked for access management
            deletion_date__isnull=True  # Only include active assets
        ).order_by('name')
    
    # Enrich users with access request statuses
    for user in page_obj:
        user.access_request_statuses = get_user_access_request_statuses(user, credentials)
    
    # Import permission functions for template
    from .matrix_view import can_add_access_api, can_edit_access_api, can_delete_access_api
    
    context = {
        'title': 'API Request',
        'result': None,
        'credentials': credentials,
        'default_credential': default_credential,
        'page_obj': page_obj,
        'total_users': total_users,
        'show_all': show_all,
        'search_query': search_query,
        'filtered_count': users.count() if search_query or filters else None,
        'filters': filters,
        'filter_params': filter_params,
        'unique_statuses': unique_statuses,
        'unique_merchants': unique_merchants,
        'unique_roles': unique_roles,
        'merchant_role_map': merchant_role_map,
        'merchant_user_map': merchant_user_map,
        'merchant_user_roles_map': merchant_user_roles_map,
        'companies': companies,
        'information_systems': information_systems,
        # Add API permissions
        'can_add_access_api': can_add_access_api(request.user),
        'can_edit_access_api': can_edit_access_api(request.user),
        'can_delete_access_api': can_delete_access_api(request.user),
    }
    return render(request, 'app_access/api_request.html', context)

@login_required
@csrf_exempt
def save_api_credential(request):
    """
    Save API credential for the user
    """
    # Check permissions for add/edit API credentials
    from .matrix_view import can_add_access_api, can_edit_access_api
    
    credential_id = request.POST.get('credential_id', '')
    if credential_id:
        # Editing existing credential
        if not can_edit_access_api(request.user):
            messages.error(request, _("Access denied - you do not have permission to edit API credentials."))
            return redirect('api_request_page')
    else:
        # Adding new credential
        if not can_add_access_api(request.user):
            messages.error(request, _("Access denied - you do not have permission to add API credentials."))
            return redirect('api_request_page')
    if request.method == 'POST':
        name = request.POST.get('name', '')
        url = request.POST.get('url', '')
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        company_id = request.POST.get('company', '') or None
        information_system_id = request.POST.get('information_system', '') or None
        environment = request.POST.get('environment', 'test')
        is_default = request.POST.get('is_default') == 'on'
        
        if not name or not url or not email:
            messages.error(request, _("Name, URL, and Email are required"))
            return redirect('api_request_page')
        
        # Check if user has access to selected company
        from .matrix_view import get_user_companies_for_api
        user_companies = get_user_companies_for_api(request.user)
        if company_id and user_companies.exists():
            from app_conf.models import Company
            try:
                company = Company.objects.get(id=company_id)
                if company not in user_companies:
                    messages.error(request, _("Access denied - you do not have permission to create credentials for this company."))
                    return redirect('api_request_page')
            except Company.DoesNotExist:
                messages.error(request, _("Selected company does not exist."))
                return redirect('api_request_page')
            
        # Additional validation for new credentials
        if not credential_id and not password:
            messages.error(request, _("Password is required for new credentials"))
            return redirect('api_request_page')
        
        # Update existing or create new credential
        if credential_id:
            try:
                credential = get_object_or_404(ApiCredential, id=credential_id, user=request.user)
                # Additional company access check for existing credential
                if credential.company and user_companies.exists() and credential.company not in user_companies:
                    messages.error(request, _("Access denied - you do not have permission to edit credentials for this company."))
                    return redirect('api_request_page')
                credential.name = name
                credential.url = url
                credential.email = email
                # Only update password if provided
                if password:
                    credential.password = password
                credential.company_id = company_id
                credential.information_system_id = information_system_id
                credential.environment = environment
                credential.is_default = is_default
                credential.save()
                messages.success(request, _("Credential updated successfully"))
            except Exception as e:
                messages.error(request, f"{_('Error updating credential')}: {str(e)}")
        else:
            try:
                ApiCredential.objects.create(
                    user=request.user,
                    name=name,
                    url=url,
                    email=email,
                    password=password,
                    company_id=company_id,
                    information_system_id=information_system_id,
                    environment=environment,
                    is_default=is_default
                )
                messages.success(request, _("Credential saved successfully"))
            except Exception as e:
                messages.error(request, f"{_('Error saving credential')}: {str(e)}")
    
    return redirect('api_request_page')

@login_required
@csrf_exempt
def delete_api_credential(request):
    """
    Delete API credential
    """
    # Check permission to delete API credentials
    from .matrix_view import can_delete_access_api
    if not can_delete_access_api(request.user):
        messages.error(request, _("Access denied - you do not have permission to delete API credentials."))
        return redirect('api_request_page')
    if request.method == 'POST':
        credential_id = request.POST.get('credential_id')
        if not credential_id:
            messages.error(request, _("Credential ID is required"))
            return redirect('api_request_page')
        
        credential = get_object_or_404(ApiCredential, id=credential_id, user=request.user)
        
        # Check if user has access to credential's company
        from .matrix_view import get_user_companies_for_api
        user_companies = get_user_companies_for_api(request.user)
        if credential.company and user_companies.exists() and credential.company not in user_companies:
            messages.error(request, _("Access denied - you do not have permission to delete credentials for this company."))
            return redirect('api_request_page')
        
        try:
            credential.delete()
            messages.success(request, _("Credential deleted successfully"))
        except Exception as e:
            messages.error(request, f"{_('Error deleting credential')}: {str(e)}")
    
    return redirect('api_request_page')

@login_required
def get_api_credential(request, credential_id):
    """
    Get API credential details
    """
    # Check permission to edit API credentials (needed to get credential details)
    from .matrix_view import can_edit_access_api, get_user_companies_for_api
    if not can_edit_access_api(request.user):
        return JsonResponse({'error': _('Access denied - you do not have permission to view API credentials.')}, status=403)
    
    credential = get_object_or_404(ApiCredential, id=credential_id, user=request.user)
    
    # Check if user has access to credential's company
    user_companies = get_user_companies_for_api(request.user)
    if credential.company and user_companies.exists() and credential.company not in user_companies:
        return JsonResponse({'error': _('Access denied - you do not have permission to view credentials for this company.')}, status=403)
    
    data = {
        'id': credential.id,
        'name': credential.name,
        'url': credential.url,
        'email': credential.email,
        'password': credential.password,
        'company': credential.company_id,
        'information_system': credential.information_system_id,
        'environment': credential.environment,
        'is_default': credential.is_default
    }
    
    return JsonResponse(data)

@login_required
@require_http_methods(["GET"])
def get_user_access_statuses(request, user_id):
    """
    Get access request statuses for a specific API user via AJAX
    """
    try:
        api_user = get_object_or_404(ApiUser, id=user_id)
        credentials = ApiCredential.objects.filter(user=request.user)
        
        # Get fresh status data
        statuses = get_user_access_request_statuses(api_user, credentials)
        
        # Format the data for JSON response
        formatted_statuses = []
        for status in statuses:
            formatted_status = {
                'merchant': status['merchant'] or '',
                'role': status['role'] or '',
                'request_status': status['request_status'],
                'admin_status': status['admin_status'],
                'user_type': status['user_type'],
                'company': status['company'] or '',
                'system': status['system'] or '',
                'environment': status['environment'] or '',
                'object_name': status['object_name'] or '',
                'request_id': status['request_id'],
                'created_at': status['created_at'].strftime('%d-%m-%Y %H:%M') if status['created_at'] else None,
                'modified_at': status['modified_at'].strftime('%d-%m-%Y %H:%M') if status['modified_at'] else None,
                'start_date': status['start_date'].strftime('%d-%m-%Y %H:%M') if status['start_date'] else None,
                'end_date': status['end_date'].strftime('%d-%m-%Y %H:%M') if status['end_date'] else None,
                'requested_by': status['requested_by'] or '',
                'request_type': status['request_type'] or '',
                'access_type': status['access_type'] or '',
                'justification': status['justification'] or '',
                'requirements': status['requirements'] or '',
                'notes': status['notes'] or '',
                'admin_comment': status['admin_comment'] or '',
                'third_party_first_name': status['third_party_first_name'] or '',
                'third_party_last_name': status['third_party_last_name'] or '',
                'third_party_email': status['third_party_email'] or '',
                'third_party_phone': status['third_party_phone'] or '',
                'third_party_organization': status['third_party_organization'] or '',
                'approving_persons': status['approving_persons'] or [],
            }
            formatted_statuses.append(formatted_status)
        
        return JsonResponse({
            'success': True,
            'user_email': api_user.email,
            'statuses': formatted_statuses,
            'updated_at': timezone.now().strftime('%d-%m-%Y %H:%M:%S')
        })
        
    except ApiUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'})
    except Exception as e:
        logger.error(f"Error getting access statuses for user {user_id}: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_http_methods(["GET"])
def get_access_record_history(request, access_record_id):
    """
    Get chronological history of all grant and revoke requests for a specific access record
    Returns a list of requests sorted by date to show complete approval/revocation timeline
    """
    from app_access.models import AccessRequest, SystemAccess
    from django.utils.translation import get_language
    
    try:
        # Get the access record
        try:
            access_record = SystemAccess.objects.get(id=access_record_id)
        except SystemAccess.DoesNotExist:
            return JsonResponse({'error': 'Access record not found'}, status=404)
        
        # Get current language for localization
        current_language = get_language() or 'en'
        
        # Find all related access requests for this access record
        related_requests = AccessRequest.objects.filter(
            access_records=access_record
        ).order_by('created_at')
        
        # Also check for legacy single access_record relationship
        legacy_requests = AccessRequest.objects.filter(
            access_record=access_record
        ).order_by('created_at')
        
        # Combine and deduplicate
        all_requests = list(related_requests) + list(legacy_requests)
        unique_requests = {}
        for req in all_requests:
            if req.id not in unique_requests:
                unique_requests[req.id] = req
        
        # Sort by creation date
        sorted_requests = sorted(unique_requests.values(), key=lambda x: x.created_at)
        
        history = []
        for req in sorted_requests:
            # Get user type
            user_type = "Unknown"
            requested_by_name = ""
            
            if req.requested_for:
                user_type = "Cabinet User"
                requested_by_name = f"{req.requested_for.first_name or ''} {req.requested_for.last_name or ''}".strip()
                if not requested_by_name:
                    requested_by_name = req.requested_for.username
            elif req.third_party_users.exists():
                user_type = "Third Party User"
                third_party = req.third_party_users.first()
                requested_by_name = f"{third_party.first_name or ''} {third_party.last_name or ''}".strip()
                if not requested_by_name:
                    requested_by_name = third_party.email
            elif req.third_party_email:
                user_type = "Third Party (Legacy)"
                requested_by_name = f"{req.third_party_first_name or ''} {req.third_party_last_name or ''}".strip()
                if not requested_by_name:
                    requested_by_name = req.third_party_email
            
            # Get requesting user
            requesting_user_name = ""
            if req.requested_by:
                requesting_user_name = f"{req.requested_by.first_name or ''} {req.requested_by.last_name or ''}".strip()
                if not requesting_user_name:
                    requesting_user_name = req.requested_by.username
            
            # Get object name
            object_name = ""
            if access_record.access_object:
                object_name = access_record.access_object.get_name(current_language) or access_record.access_object.get_name('ua')
            
            # Get roles
            roles = []
            for role in access_record.roles.all():
                if current_language == 'uk':
                    role_name = role.get_name() or role.name or ''
                elif current_language == 'ru':
                    role_name = role.get_name() or role.name or ''
                else:
                    role_name = role.accessrole_name_en
                
                if not role_name:
                    role_name = role.get_name() or role.name or ''
                
                roles.append({
                    'id': role.id,
                    'name': role_name,
                    'color': role.color or '#6c757d'
                })
            
            history.append({
                'request_id': req.id,
                'request_type': req.request_type,
                'request_status': req.status,
                'admin_status': req.admin_status,
                'user_type': user_type,
                'requested_by': requesting_user_name,
                'requested_for': requested_by_name,
                'company': req.company.name if req.company else '',
                'system': req.system.name if req.system else '',
                'environment': req.environment,
                'object_name': object_name,
                'roles': roles,
                'justification': req.justification or '',
                'requirements': req.requirements or '',
                'notes': req.notes or '',
                'admin_comment': req.admin_comment or '',
                'created_at': req.created_at.strftime('%d-%m-%Y %H:%M') if req.created_at else '',
                'modified_at': req.modified_at.strftime('%d-%m-%Y %H:%M') if req.modified_at else '',
                'start_date': req.start_date.strftime('%d-%m-%Y %H:%M') if req.start_date else '',
                'end_date': req.end_date.strftime('%d-%m-%Y %H:%M') if req.end_date else '',
                'third_party_email': req.third_party_email or '',
                'third_party_first_name': req.third_party_first_name or '',
                'third_party_last_name': req.third_party_last_name or '',
                'third_party_phone': req.third_party_phone or '',
                'third_party_organization': req.third_party_organization or '',
            })
        
        return JsonResponse({
            'success': True,
            'access_record': {
                'id': access_record.id,
                'object_name': object_name,
                'system': access_record.asset.name if access_record.asset else '',
                'environment': access_record.environment,
                'roles': roles,
            },
            'history': history
        })
        
    except Exception as e:
        logger.error(f"Error in get_access_record_history: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


def get_all_user_requests_by_merchant_role(api_user, credentials, merchant, role):
    """
    Get ALL individual requests (not consolidated) for a specific user, merchant, and role
    This returns each grant/revoke request separately for chronological history
    """
    from app_access.models import AccessRequest, ThirdPartyUser
    from django.contrib.auth.models import User
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Find Django User by email (for requested_for field)
    django_user = None
    try:
        django_user = User.objects.get(email=api_user.email)
    except User.DoesNotExist:
        pass
    
    # Find Third Party User by email
    third_party_user = None
    try:
        third_party_user = ThirdPartyUser.objects.get(email=api_user.email)
    except ThirdPartyUser.DoesNotExist:
        pass
    
    # Get company, information system, and environment from credentials
    companies_from_credentials = set()
    systems_from_credentials = set()
    environments_from_credentials = set()
    
    for credential in credentials:
        if credential.company:
            companies_from_credentials.add(credential.company)
        if credential.information_system:
            systems_from_credentials.add(credential.information_system)
        if credential.environment:
            environments_from_credentials.add(credential.environment)
    
    # Build base query for Access Requests (Grant and Revoke)
    base_query = AccessRequest.objects.filter(
        request_type__in=['grant', 'revoke']
    ).filter(
        status__in=['approved', 'pending']
    ).filter(
        admin_status__in=['pending', 'in_progress', 'granted', 'denied']
    )
    
    # Apply credential filters if available
    if companies_from_credentials:
        base_query = base_query.filter(company__in=companies_from_credentials)
    if systems_from_credentials:
        base_query = base_query.filter(system__in=systems_from_credentials)
    if environments_from_credentials:
        base_query = base_query.filter(environment__in=environments_from_credentials)
    
    # Find matching requests for this user
    matching_requests = []
    
    # Search for Django user requests (only if Django user exists)
    if django_user:
        django_requests = base_query.filter(requested_for=django_user)
        matching_requests.extend(list(django_requests))
    
    # Search for Third Party user requests (only if Third Party user exists)
    if third_party_user:
        third_party_requests = base_query.filter(third_party_users=third_party_user)
        matching_requests.extend(list(third_party_requests))
    
    # Always search by email in third party fields (legacy support)
    email_requests = base_query.filter(third_party_email=api_user.email)
    matching_requests.extend(list(email_requests))
    
    # Remove duplicates
    unique_requests = {}
    for request_obj in matching_requests:
        if request_obj.id not in unique_requests:
            unique_requests[request_obj.id] = request_obj
    
    # Filter requests that match Object (Merchant) and Role
    filtered_requests = []
    logger.info(f"Filtering {len(unique_requests)} unique requests for merchant: {merchant}, role: {role}")
    
    for access_request in unique_requests.values():
        # Check if this request matches our Merchant (Object) and Role
        access_records = []
        if access_request.access_records.exists():
            access_records = list(access_request.access_records.all())
        elif hasattr(access_request, 'access_record') and access_request.access_record:
            access_records = [access_request.access_record]
        
        for access_record in access_records:
            # Check Object match (Merchant from API = Object from Access Request)
            object_match = False
            if access_record.access_object:
                object_name_ua = access_record.access_object.get_name() or ""
                object_name_en = access_record.access_object.get_name('en') or ""
                object_name_ru = access_record.access_object.get_name('ru') or ""
                
                if (merchant.lower() == object_name_ua.lower() or
                    merchant.lower() == object_name_en.lower() or
                    merchant.lower() == object_name_ru.lower()):
                    object_match = True
            else:
                # If no specific object, check against company or system name
                company_name = access_request.company.name if access_request.company else ""
                system_name = access_request.system.name if access_request.system else ""
                
                if (merchant.lower() == company_name.lower() or
                    merchant.lower() == system_name.lower()):
                    object_match = True
            
            # Check Role match
            role_match = False
            if access_record.roles.exists():
                for role_obj in access_record.roles.all():
                    role_name_ua = role_obj.get_name() or role_obj.name or ""
                    role_name_en = role_obj.accessrole_name_en or ""
                    role_name_ru = role_obj.get_name() or role_obj.name or ""
                    
                    if (role.lower() == role_name_ua.lower() or
                        role.lower() == role_name_en.lower() or
                        role.lower() == role_name_ru.lower()):
                        role_match = True
                        break
            
            # If both Object and Role match, add this request
            if object_match and role_match:
                filtered_requests.append(access_request)
                break  # Only need one match per access_request
    
    return filtered_requests


@login_required
@require_http_methods(["GET"])
def get_merchant_role_history(request, user_id, merchant, role):
    """
    Get chronological history of all requests for a specific user, merchant, and role combination
    This is used when access_record_id is not available
    """
    from app_access.models import ApiUser
    from django.utils.translation import get_language
    from urllib.parse import unquote
    from django.contrib.auth.models import User
    
    try:
        # Decode URL-encoded parameters
        merchant = unquote(merchant)
        role = unquote(role)
        
        # Get the API user
        try:
            api_user = ApiUser.objects.get(id=user_id)
        except ApiUser.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)
        
        # Get current language for localization
        current_language = get_language() or 'en'
        
        # Get credentials for the requesting user
        credentials = ApiCredential.objects.filter(user=request.user)
        
        # Try to get the actual merchant name from the API user's role mappings
        merchant_display_name = merchant  # fallback
        try:
            # The merchant parameter might be the merchant name or ID, so we need to try both
            role_mapping = None
            
            # First, try to find by merchant name
            role_mapping = api_user.role_mappings.filter(
                merchant__name=merchant,
                role__name=role
            ).first()
            
            # If not found, try to find by merchant ID (in case merchant param is numeric)
            if not role_mapping and merchant.isdigit():
                role_mapping = api_user.role_mappings.filter(
                    merchant__id=int(merchant),
                    role__name=role
                ).first()
            
            # If still not found, try a broader search for any role mapping with this user and role
            if not role_mapping:
                role_mapping = api_user.role_mappings.filter(role__name=role).first()
            
            if role_mapping and role_mapping.merchant:
                # Get the merchant name from the ApiUserMerchant model
                raw_merchant_name = role_mapping.merchant.name
                
                # If the merchant name is just a number, try to make it more descriptive
                if raw_merchant_name and raw_merchant_name.isdigit():
                    # Try to find a more descriptive name from access records or other sources
                    merchant_display_name = f"Merchant {raw_merchant_name}"
                else:
                    merchant_display_name = raw_merchant_name
                    
                logger.info(f"Resolved merchant '{merchant}' to display name: '{merchant_display_name}' (raw: '{raw_merchant_name}', ID: {role_mapping.merchant.id})")
        except Exception as e:
            logger.debug(f"Could not resolve merchant display name: {e}")
        
        # Get ALL individual requests for this user and filter by merchant and role
        # Instead of using the consolidated statuses, get raw requests
        matching_requests = get_all_user_requests_by_merchant_role(api_user, credentials, merchant, role)
        
        logger.info(f"Found {len(matching_requests)} matching requests for user {user_id}, merchant: {merchant}, role: {role}")
        
        # Try to get a better merchant display name from the access records in the matching requests
        if merchant_display_name == merchant or (merchant_display_name and merchant_display_name.startswith("Merchant ")):
            for request_obj in matching_requests:
                access_records = []
                if request_obj.access_records.exists():
                    access_records = list(request_obj.access_records.all())
                elif hasattr(request_obj, 'access_record') and request_obj.access_record:
                    access_records = [request_obj.access_record]
                
                for access_record in access_records:
                    if access_record.access_object:
                        # Get localized object name
                        object_name = (access_record.access_object.get_name() or 
                                     access_record.access_object.get_name('en') or 
                                     access_record.access_object.get_name('ua'))
                        if object_name and object_name != merchant:
                            merchant_display_name = object_name
                            logger.info(f"Enhanced merchant display name from access record: '{merchant_display_name}'")
                            break
                if merchant_display_name != merchant and not merchant_display_name.startswith("Merchant "):
                    break
        
        if not matching_requests:
            return JsonResponse({
                'success': True,
                'access_record': {
                    'id': 'unknown',
                    'object_name': merchant_display_name,
                    'system': '',
                    'environment': '',
                    'roles': [{'name': role, 'color': '#6c757d'}],
                },
                'history': []
            })
        
        # Get the first matching request to extract common information
        first_request = matching_requests[0]
        
        # Create history entries from all matching requests
        history = []
        for access_request in matching_requests:
            # Determine user type
            user_type = "Unknown"
            requested_for_name = api_user.email
            
            # Check if this is a Django user request
            try:
                django_user = User.objects.get(email=api_user.email)
                if access_request.requested_for == django_user:
                    user_type = "Cabinet User"
                    requested_for_name = f"{django_user.first_name or ''} {django_user.last_name or ''}".strip() or django_user.username
            except User.DoesNotExist:
                pass
            
            # Check if this is a Third Party user request
            if access_request.third_party_email == api_user.email or access_request.third_party_users.filter(email=api_user.email).exists():
                user_type = "Third Party User"
                if access_request.third_party_first_name or access_request.third_party_last_name:
                    requested_for_name = f"{access_request.third_party_first_name or ''} {access_request.third_party_last_name or ''}".strip()
                elif access_request.third_party_users.filter(email=api_user.email).exists():
                    third_party_user = access_request.third_party_users.filter(email=api_user.email).first()
                    requested_for_name = f"{third_party_user.first_name or ''} {third_party_user.last_name or ''}".strip() or third_party_user.email
            
            # Get requested_by information
            requested_by_name = ""
            if access_request.requested_by:
                requested_by_name = f"{access_request.requested_by.first_name or ''} {access_request.requested_by.last_name or ''}".strip()
                if not requested_by_name:
                    requested_by_name = access_request.requested_by.username
            
            # Get object name - try multiple sources for better name resolution
            object_name = merchant  # fallback
            access_records = []
            if access_request.access_records.exists():
                access_records = list(access_request.access_records.all())
            elif hasattr(access_request, 'access_record') and access_request.access_record:
                access_records = [access_request.access_record]
            
            if access_records:
                access_record = access_records[0]
                if access_record.access_object:
                    # Get localized object name
                    object_name = (access_record.access_object.object_name_en or 
                                 access_record.access_object.get_name() or access_record.access_object.name or 
                                 access_record.access_object.get_name() or access_record.access_object.name or 
                                 merchant)
                else:
                    # If no access object, try to get a friendly name from the system or company
                    if access_request.system:
                        # For system-level access, use system name
                        object_name = access_request.system.name
                    elif access_request.company:
                        # For company-level access, use company name
                        object_name = access_request.company.name
            
            # If we still have the merchant code, use the display name or make it more readable
            if object_name == merchant:
                # First try to use the merchant display name we resolved earlier
                object_name = merchant_display_name
                # If still the same and has underscores, make it more readable
                if object_name == merchant and '_' in merchant:
                    object_name = merchant.replace('_', ' ')
            
            # Get roles - use provided role as fallback
            roles = [{'name': role, 'color': '#6c757d'}]
            if access_records:
                roles = []
                for access_record in access_records:
                    for role_obj in access_record.roles.all():
                        role_name = role_obj.get_name() or role_obj.name or ''
                        roles.append({'name': role_name, 'color': '#6c757d'})
                if not roles:  # Fallback if no roles found
                    roles = [{'name': role, 'color': '#6c757d'}]
            
            history.append({
                'request_id': access_request.id,
                'request_type': access_request.request_type or 'grant',
                'request_status': access_request.status,
                'admin_status': access_request.admin_status,
                'user_type': user_type,
                'requested_by': requested_by_name,
                'requested_for': requested_for_name,
                'company': access_request.company.name if access_request.company else '',
                'system': access_request.system.name if access_request.system else '',
                'environment': access_request.environment or '',
                'object_name': object_name,
                'roles': roles,
                'justification': access_request.justification or '',
                'requirements': access_request.requirements or '',
                'notes': access_request.notes or '',
                'admin_comment': access_request.admin_comment or '',
                'created_at': access_request.created_at.strftime('%d-%m-%Y %H:%M') if access_request.created_at else '',
                'modified_at': access_request.modified_at.strftime('%d-%m-%Y %H:%M') if access_request.modified_at else '',
                'start_date': access_request.start_date.strftime('%d-%m-%Y %H:%M') if access_request.start_date else '',
                'end_date': access_request.end_date.strftime('%d-%m-%Y %H:%M') if access_request.end_date else '',
                'third_party_email': access_request.third_party_email or '',
                'third_party_first_name': access_request.third_party_first_name or '',
                'third_party_last_name': access_request.third_party_last_name or '',
                'third_party_phone': access_request.third_party_phone or '',
                'third_party_organization': access_request.third_party_organization or '',
            })
        
        # Sort history by creation date (latest first - reverse chronological)
        history.sort(key=lambda x: x['created_at'] or '0000-00-00 00:00', reverse=True)
        
        # Get access record ID safely
        access_record_id = 'unknown'
        if hasattr(first_request, 'access_record') and first_request.access_record:
            access_record_id = first_request.access_record.id
        elif first_request.access_records.exists():
            access_record_id = first_request.access_records.first().id
        
        return JsonResponse({
            'success': True,
            'access_record': {
                'id': access_record_id,
                'object_name': merchant_display_name,
                'system': first_request.system.name if first_request.system else '',
                'environment': first_request.environment or '',
                'roles': [{'name': role, 'color': '#6c757d'}],
            },
            'history': history
        })
        
    except Exception as e:
        logger.error(f"Error in get_merchant_role_history: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_access_requests_by_record(request, access_record_id):
    """
    Get all access requests for a specific access record
    """
    try:
        from app_access.models import SystemAccess, AccessRequest, AccessRequestSequence
        
        # Get the access record
        access_record = get_object_or_404(SystemAccess, id=access_record_id)
        
        # Pull all requests that include this record via M2M
        access_requests = AccessRequest.objects.filter(
            access_records=access_record
        ).select_related('requested_for', 'requested_by', 'company', 'system').order_by('-created_at')

        from app_access.models import AccessRoles
        current_language = getattr(request, 'LANGUAGE_CODE', 'en')[:2] if getattr(request, 'LANGUAGE_CODE', None) else 'en'

        # Map requests to Grant Access Record sequences (A.B.C), attach D when revoked
        requests_data = []
        for request_obj in access_requests:
            grant_ids = []
            if request_obj.request_type == 'grant':
                seqs = AccessRequestSequence.objects.filter(access_record=access_record, grant_request=request_obj).order_by('order_number')
                for s in seqs:
                    prefix = '.'.join(str(s.sequence_id).split('.')[:3])
                    if s.sequence_status == 'revoked' and s.revoke_request:
                        grant_ids.append(f"{prefix}.{s.revoke_request.id}")
                    else:
                        grant_ids.append(f"{prefix}.0")
            elif request_obj.request_type == 'revoke':
                seqs = AccessRequestSequence.objects.filter(access_record=access_record, revoke_request=request_obj).order_by('order_number')
                for s in seqs:
                    prefix = '.'.join(str(s.sequence_id).split('.')[:3])
                    d_part = s.revoke_request.id if s.revoke_request else 0
                    grant_ids.append(f"{prefix}.{d_part}")

            # Requested Object Role for this access record (grant: from request; revoke: from original grant)
            requested_role_info = None
            role_id = None
            if request_obj.request_type == 'grant' and getattr(request_obj, 'requested_access_record_roles', None) and isinstance(request_obj.requested_access_record_roles, list):
                for item in request_obj.requested_access_record_roles:
                    if isinstance(item, dict) and item.get('access_record_id') == access_record.id and 'role_id' in item:
                        role_id = item['role_id']
                        break
            elif request_obj.request_type == 'revoke':
                original_request = None
                import re
                if request_obj.notes and re.search(r'request #(\d+)', request_obj.notes):
                    match = re.search(r'request #(\d+)', request_obj.notes)
                    if match:
                        try:
                            original_request = AccessRequest.objects.filter(
                                id=int(match.group(1)), request_type='grant', status='approved', admin_status='granted'
                            ).first()
                        except (ValueError, TypeError):
                            pass
                if original_request is None and getattr(request_obj, 'revoked_grant_access_record_ids', None):
                    ids_list = request_obj.revoked_grant_access_record_ids or []
                    if ids_list:
                        try:
                            parts = str(ids_list[0]).split('.')
                            if len(parts) >= 2 and parts[1].isdigit():
                                orig = AccessRequest.objects.filter(
                                    id=int(parts[1]), request_type='grant', status='approved', admin_status='granted'
                                ).first()
                                if orig:
                                    original_request = orig
                        except (IndexError, ValueError, TypeError):
                            pass
                if original_request and getattr(original_request, 'requested_access_record_roles', None) and isinstance(original_request.requested_access_record_roles, list):
                    for item in original_request.requested_access_record_roles:
                        if isinstance(item, dict) and item.get('access_record_id') == access_record.id and 'role_id' in item:
                            role_id = item['role_id']
                            break
            if role_id:
                try:
                    role = AccessRoles.objects.get(id=role_id)
                    requested_role_info = {
                        'id': role.id,
                        'name': role.get_name(current_language) or role.name or '',
                        'color': role.color or '#6c757d',
                    }
                except AccessRoles.DoesNotExist:
                    pass

            requests_data.append({
                'id': request_obj.id,
                'request_type': request_obj.request_type,
                'status': request_obj.status,
                'admin_status': request_obj.admin_status,
                'requested_for_name': request_obj.requested_for.get_full_name() if request_obj.requested_for else '',
                'requested_for_email': request_obj.requested_for.email if request_obj.requested_for else '',
                'requested_by_name': request_obj.requested_by.get_full_name() if request_obj.requested_by else '',
                'requested_by_email': request_obj.requested_by.email if request_obj.requested_by else '',
                'created_at': request_obj.created_at.isoformat(),
                'company_name': request_obj.company.name if request_obj.company else '',
                'system_name': request_obj.system.name if request_obj.system else '',
                'environment': request_obj.environment,
                'justification': request_obj.justification,
                'requirements': request_obj.requirements,
                'notes': request_obj.notes,
                'admin_comment': request_obj.admin_comment,
                'admin_status_comment': request_obj.admin_status_comment,
                'start_date': request_obj.start_date.isoformat() if request_obj.start_date else None,
                'end_date': request_obj.end_date.isoformat() if request_obj.end_date else None,
                'grant_access_record_ids': grant_ids,
                'requested_role': requested_role_info,
            })
        
        return JsonResponse({
            'status': 'success',
            'data': {
                'access_record': {
                    'id': access_record.id,
                    'company': access_record.asset.company.name if access_record.asset.company else '',
                    'system': access_record.asset.name if access_record.asset else '',
                    'environment': access_record.environment,
                    'object': access_record.access_object.get_name() if access_record.access_object else '',
                },
                'requests': requests_data,
                # Total unique sequences (A.B.C) involved across all requests for this record
                'total_count': len({ '.'.join(gid.split('.')[:3]) for r in requests_data for gid in r.get('grant_access_record_ids', []) }) or len(requests_data)
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting access requests for record {access_record_id}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })


@login_required
def test_base64_decode(request):
    """
    Test function to decode the Base64 content directly
    """
    # The Base64 encoded content from the response
    base64_content = "eyJjb3VudCI6MTEyLCJwYWdlcyI6OCwiZGF0YSI6W3siaWQiOjE1NywiZW1haWwiOiJuLnN1bXNrYXlhQGNyZWRpdDcudWEiLCJoYXNoIjoiNTA1MzkxMTU2Mjk3N2VkOWYwMzdiIiwiZmlyc3ROYW1lIjpudWxsLCJsYXN0TmFtZSI6bnVsbCwicGhvbmUiOiIzODA2Njg2MDI3NjMiLCJsYXN0X2xvZ2luIjoiMjAyNC0wMS0yNVQwOTozNToyNiJ9LHsiaWQiOjE5NiwiZW1haWwiOiJuaWtpdGEua2xpbW92aWNoQGF2ZW50dXMud29yayIsImhhc2giOiI5ODk2NjUwODYzZTM0YTRiYTJmNzEiLCJmaXJzdE5hbWUiOm51bGwsImxhc3ROYW1lIjpudWxsLCJwaG9uZSI6IjM4MDAwMDAwMDAwMCIsImxhc3RfbG9naW4iOiIyMDI1LTAzLTE0VDExOjExOjI2In0seyJpZCI6MTI5LCJlbWFpbCI6Iml2YW5AcGF5dGVjaC5jb20udWEiLCJoYXNoIjoiMzY1MzQ2MTVhZjExZjA1MTZjNC44MzMwNDc4OCIsImZpcnN0TmFtZSI6bnVsbCwibGFzdE5hbWUiOm51bGwsInBob25lIjoiMzgwNTA0MjIyMjkxIiwibGFzdF9sb2dpbiI6IjIwMjItMDYtMDlUMTQ6MjQ6MzYifSx7ImlkIjoyMTAsImVtYWlsIjoiZWxlbmEuc3Rhcm9rYWRvbXNrYXlhQHBheXRlY2guY29tLnVhIiwiaGFzaCI6IjI1Njg1MTUxNjVkZjVkZjBhN2Q0ZCIsImZpcnN0TmFtZSI6bnVsbCwibGFzdE5hbWUiOm51bGwsInBob25lIjoiMzgwOTczNzkxNDg5IiwibGFzdF9sb2dpbiI6IjIwMjUtMDMtMTRUMTM6MjY6MDgifSx7ImlkIjoxNTQsImVtYWlsIjoiYW5kcmV5QHNlbGZp"
    
    try:
        # Add padding if needed
        padding_needed = len(base64_content) % 4
        if padding_needed:
            base64_content += '=' * (4 - padding_needed)
        
        # Decode Base64
        decoded_bytes = base64.b64decode(base64_content)
        
        # Try UTF-8 decoding first
        try:
            decoded_content = decoded_bytes.decode('utf-8')
            content_type = 'text/plain'
        except UnicodeDecodeError:
            # If UTF-8 decoding fails, return the raw bytes
            decoded_content = str(decoded_bytes)
            content_type = 'text/plain'
        
        # Return the decoded content
        return HttpResponse(decoded_content, content_type=content_type)
    
    except Exception as e:
        return HttpResponse(f"Error decoding Base64: {str(e)}", content_type='text/plain')

@login_required
@csrf_exempt
def api_request_answer(request):
    """
    Handle the API request to the Paytech login endpoint
    """
    if request.method == 'POST':
        try:
            # Get email and password from the form
            email = request.POST.get('email')
            password = request.POST.get('password')
            
            # Check if using saved credential
            credential_id = request.POST.get('credential_id')
            if credential_id:
                credential = get_object_or_404(ApiCredential, id=credential_id, user=request.user)
                email = credential.email
                password = credential.password
                api_url = credential.url
            else:
                api_url = 'https://bn2-crss.paytech.com.ua/api/login'
            
            # Prepare the payload
            payload = json.dumps({
                "email": email,
                "password": password
            })
            
            # Set headers
            headers = {
                'Content-Type': 'application/json'
            }
            
            # Make the API request
            response = requests.post(
                api_url,
                headers=headers,
                data=payload,
                timeout=SYNC_HTTP_TIMEOUT
            )
            
            # Get the response data
            response_data = response.json()
            
            # Extract token if available
            token = response_data.get('token', '')
            
            # Store token in session if using a saved credential
            if credential_id and token:
                if not hasattr(request, 'session'):
                    request.session = {}
                if 'api_tokens' not in request.session:
                    request.session['api_tokens'] = {}
                request.session['api_tokens'][credential_id] = token
                request.session.modified = True
            
            # Prepare context for rendering
            context = {
                'title': 'API Request',
                'result': {
                    'status_code': response.status_code,
                    'data': response_data,
                    'token': token,
                    'pretty_json': json.dumps(response_data, indent=4)
                },
                'credentials': ApiCredential.objects.filter(user=request.user).order_by('-is_default', 'name'),
                'default_credential': ApiCredential.objects.filter(user=request.user, is_default=True).first()
            }
            
            return render(request, 'app_access/api_request.html', context)
            
        except Exception as e:
            # Handle errors
            context = {
                'title': 'API Request',
                'result': {
                    'error': str(e)
                },
                'credentials': ApiCredential.objects.filter(user=request.user).order_by('-is_default', 'name'),
                'default_credential': ApiCredential.objects.filter(user=request.user, is_default=True).first()
            }
            return render(request, 'app_access/api_request.html', context)
    
    # If not POST, redirect to the request page
    return api_request_page(request)

@login_required
def refresh_token(request, credential_id):
    """
    Refresh the token for a credential by making a login request
    Returns the new token in JSON format
    """
    try:
        # Get the credential
        credential = get_object_or_404(ApiCredential, id=credential_id, user=request.user)
        
        # Prepare the login request
        api_url = credential.url
        if not api_url.endswith('/'):
            if '/login' not in api_url:
                api_url = f"{api_url}/login"
        
        # Prepare the payload
        payload = json.dumps({
            "email": credential.email,
            "password": credential.password
        })
        
        # Set headers
        headers = {
            'Content-Type': 'application/json'
        }
        
        # Make the API request
        response = requests.post(
            api_url,
            headers=headers,
            data=payload,
            timeout=SYNC_HTTP_TIMEOUT
        )
        
        # Check if the request was successful
        if response.status_code == 200:
            # Get the response data
            response_data = response.json()
            
            # Extract token if available
            token = response_data.get('token', '')
            
            # Store token in session
            if token:
                if not hasattr(request, 'session'):
                    request.session = {}
                if 'api_tokens' not in request.session:
                    request.session['api_tokens'] = {}
                request.session['api_tokens'][str(credential_id)] = token
                request.session.modified = True
                
                return JsonResponse({'success': True, 'token': token})
        
        return JsonResponse({'success': False, 'error': 'Failed to obtain token'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@csrf_exempt
def sync_api_users(request):
    """
    Synchronize users from the API with the database
    """
    if request.method == 'POST':
        # Skip noisy request logging for frequent status probes
        is_status_check_probe = (
            request.content_type
            and 'application/json' in request.content_type
            and b'"status_check"' in request.body
            and b'true' in request.body
        )
        if not is_status_check_probe:
            logger.info(f"Request content type: {request.content_type}")
            logger.info(f"Request headers: {dict(request.headers)}")

        # Handle JSON content type for AJAX requests
        if request.content_type and 'application/json' in request.content_type:
            try:
                data = json.loads(request.body)
                credential_id = data.get('credential_id')
                stop_sync = data.get('stop_sync')
                status_check = data.get('status_check')
                if not data.get('status_check'):
                    logger.info(f"Parsed JSON data: {data}")
            except Exception as e:
                logger.error(f"Error parsing JSON request: {str(e)}")
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
        else:
            # Handle form data for regular form submissions
            credential_id = request.POST.get('credential_id')
            stop_sync = request.POST.get('stop_sync') == 'true'
            status_check = request.POST.get('status_check') == 'true'
        
        # Add debugging log only for non-polling actions
        if not status_check:
            logger.info(f"Sync request initiated with credential_id={credential_id}, stop_sync={stop_sync}, status_check={status_check}")
        
        if not credential_id:
            messages.error(request, _("No API credential selected"))
            return JsonResponse({'success': False, 'error': _("No API credential selected")}, status=400) if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('api_request_page')
        
        try:
            # Get credential
            credential = get_object_or_404(ApiCredential, id=credential_id, user=request.user)
            
            # Handle stop sync request
            if stop_sync:
                try:
                    sync_status = ApiSyncStatus.objects.filter(
                        credential=credential,
                        status='running'
                    ).latest('started_at')
                    sync_status.stop()
                    return JsonResponse({
                        'success': True,
                        'message': _("Synchronization process stopped successfully.")
                    })
                except ApiSyncStatus.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': _("No running sync process found.")
                    })
            
            # Handle status check request
            if status_check:
                try:
                    sync_status = ApiSyncStatus.objects.filter(
                        credential=credential,
                        status='running'
                    ).latest('started_at')
                    cs = int(sync_status.completed_steps or 0)
                    ts = int(sync_status.total_steps or 0)
                    pct = min(100, int((cs / ts) * 100)) if ts > 0 else int(sync_status.percent_complete or 0)
                    return JsonResponse({
                        'status': sync_status.status,
                        'percent': pct,
                        'message': sync_status.current_step,
                        'completed_steps': cs,
                        'total_steps': ts,
                        'extra_data': sync_status.extra_data or {}
                    })
                except ApiSyncStatus.DoesNotExist:
                    # Check if sync was recently completed
                    try:
                        recent_sync = ApiSyncStatus.objects.filter(
                            credential=credential
                        ).exclude(status='running').latest('completed_at')
                        
                        # If completed within the last minute, return its status
                        if recent_sync.completed_at and recent_sync.completed_at > timezone.now() - timezone.timedelta(minutes=1):
                            return JsonResponse({
                                'status': recent_sync.status,
                                'percent': 100 if recent_sync.status == 'completed' else recent_sync.percent_complete,
                                'message': recent_sync.current_step or (_("Completed") if recent_sync.status == 'completed' else 
                                           _("Error: {}").format(recent_sync.error_message) if recent_sync.status == 'error' else 
                                           _("Stopped")),
                                'completed_steps': recent_sync.completed_steps,
                                'total_steps': recent_sync.total_steps,
                                'extra_data': recent_sync.extra_data or {}
                            })
                    except ApiSyncStatus.DoesNotExist:
                        pass
                        
                    # No recent sync found
                    return JsonResponse({
                        'status': 'not_running',
                        'percent': 0,
                        'message': _("No sync process is currently running.")
                    })
            
            # Create new sync status
            # Generate unique ID for this sync operation
            sync_unique_id = f"SYN-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:8]}"
            is_scheduled = False  # This is a manual sync
            
            # If this is a scheduled sync, update the is_scheduled flag
            scheduled_sync_id = request.POST.get('scheduled_sync_id') or (data.get('scheduled_sync_id') if 'data' in locals() else None)
            if scheduled_sync_id:
                is_scheduled = True
                sync_unique_id = f"SCH-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:8]}"
            
            sync_status = ApiSyncStatus.objects.create(
                credential=credential,
                status='running',
                current_step='Initializing...',
                completed_steps=0,
                total_steps=0,
                unique_id=sync_unique_id,
                is_scheduled=is_scheduled,
                extra_data={'terminal_lines': []}
            )

            terminal_lines = []

            def append_terminal_line(message, persist=True):
                """Persist rolling terminal lines for UI live-view."""
                timestamp = timezone.now().strftime('%H:%M:%S')
                terminal_lines.append(f"[{timestamp}] {message}")
                del terminal_lines[:-500]

                if not persist:
                    return

                current_extra = sync_status.extra_data if isinstance(sync_status.extra_data, dict) else {}
                current_extra['terminal_lines'] = terminal_lines.copy()
                sync_status.extra_data = current_extra
                sync_status.save(update_fields=['extra_data'])

            def update_sync_progress(message, completed_steps, total_steps, extra_data=None):
                """Update progress while preserving terminal output in extra_data."""
                payload = dict(extra_data or {})
                payload['terminal_lines'] = terminal_lines.copy()
                sync_status.update_progress(message, completed_steps, total_steps, payload)
            
            # First, get a fresh token by authenticating
            api_login_url = credential.url
            if '/login' not in api_login_url:
                api_login_url = f"{api_login_url}/login" if not api_login_url.endswith('/') else f"{api_login_url}login"
            
            # Update sync status
            append_terminal_line(_('Sync started'))
            append_terminal_line(_('Authenticating with API...'))
            update_sync_progress('Authenticating...', 1, 5)
            
            # Prepare login payload
            login_payload = json.dumps({
                "email": credential.email,
                "password": credential.password
            })
            
            # Set headers for login
            login_headers = {
                'Content-Type': 'application/json'
            }
            
            # Make the login API request
            logger.info(f"Making login request to {api_login_url}")
            login_response = requests.post(
                api_login_url,
                headers=login_headers,
                data=login_payload,
                timeout=SYNC_HTTP_TIMEOUT
            )
            
            # Log the response status
            logger.info(f"Login response status: {login_response.status_code}")
            
            # Check if login was successful
            if login_response.status_code != 200:
                # Log the error response
                try:
                    error_content = login_response.json()
                    logger.error(f"Authentication failed with error: {error_content}")
                    
                    # Extract detailed error message if available
                    detailed_error = ""
                    if isinstance(error_content, dict):
                        if 'message' in error_content:
                            detailed_error = f": {error_content['message']}"
                        elif 'respMessage' in error_content:
                            detailed_error = f": {error_content['respMessage']}"
                    
                    sync_status.error(_("Authentication failed. Status code: {}{}").format(login_response.status_code, detailed_error))
                    messages.error(request, _("Authentication failed. Status code: {}{}").format(login_response.status_code, detailed_error))
                except:
                    error_text = login_response.text[:200] if login_response.text else "No response content"
                    logger.error(f"Authentication failed with status code: {login_response.status_code}, response: {error_text}")
                    sync_status.error(_("Authentication failed. Status code: {}. Response: {}").format(login_response.status_code, error_text))
                    messages.error(request, _("Authentication failed. Status code: {}. Response: {}").format(login_response.status_code, error_text))
                
                append_terminal_line(_('Authentication failed'))
                return redirect('api_request_page')
            
            # Update sync status
            append_terminal_line(_('Authentication successful. Fetching users...'))
            update_sync_progress('Fetching users...', 2, 5)
            
            # Get token from response
            try:
                login_data = login_response.json()
                token = login_data.get('token')
                
                if not token:
                    sync_status.error(_("No token received from authentication response"))
                    messages.error(request, _("No token received from authentication response"))
                    return redirect('api_request_page')
                
                # Store token in session
                if not hasattr(request, 'session'):
                    request.session = {}
                if 'api_tokens' not in request.session:
                    request.session['api_tokens'] = {}
                request.session['api_tokens'][credential_id] = token
                request.session.modified = True
                
                messages.success(request, _("Successfully authenticated and obtained token"))
                append_terminal_line(_('Token received successfully'))
            except Exception as e:
                sync_status.error(_("Error parsing authentication response: {}").format(str(e)))
                messages.error(request, _("Error parsing authentication response: {}").format(str(e)))
                append_terminal_line(_("Error parsing authentication response: {}").format(str(e)))
                return redirect('api_request_page')
            
            # Initialize counters for all entities
            user_count = 0
            user_count_created = 0
            user_count_updated = 0
            user_count_existing = 0
            user_count_removed = 0
            role_count_created = 0
            merchant_count_created = 0
            role_mapping_count_created = 0
            status_count_created = 0
            status_count_updated = 0
            history_entries_created = 0
            login_history_entries_created = 0  # Initialize login history counter
            merchant_links_count_created = 0
            
            # Now proceed with user synchronization
            try:
                # Extract the base URL from the login endpoint
                base_url = credential.url.rstrip('/')
                if '/login' in base_url:
                    base_url = base_url.split('/login')[0]
                
                # Make sure the URL doesn't end with a slash
                if base_url.endswith('/'):
                    base_url = base_url[:-1]
                
                # Build API URL based on API format
                api_url = base_url
                if not base_url.endswith('/api') and '/api/' not in base_url:
                    api_url = f"{base_url}/api"
                    
                # Build final URL
                api_base_url = f"{api_url}/security/users"
                
                # Log what we're doing
                logger.info(f"Base URL: {base_url}")
                logger.info(f"API Base URL: {api_base_url}")
                
                # Update sync status
                append_terminal_line(_('Processing users...'))
                update_sync_progress('Processing users...', 3, 5)
                
                # Define a helper function to decode Base64 responses
                def decode_response(response):
                    """
                    Attempts to decode the response content, handling Base64 encoded responses
                    Returns tuple (success, data, error_message)
                    """
                    try:
                        raw_content = response.text.strip()
                        
                        # If empty response
                        if not raw_content:
                            return False, None, "Empty response"
                        
                        # Check if content appears to be Base64 encoded
                        if all(c in 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=' for c in raw_content[:100]):
                            try:
                                # Add padding if needed
                                padding_needed = len(raw_content) % 4
                                if padding_needed:
                                    padded_content = raw_content + '=' * (4 - padding_needed)
                                else:
                                    padded_content = raw_content
                                # Decode Base64
                                decoded_bytes = base64.b64decode(padded_content)
                                decoded_content = decoded_bytes.decode('utf-8')
                                # Parse JSON
                                data = json.loads(decoded_content)
                                logger.debug(f"Successfully decoded Base64 response: {data}")
                                return True, data, None
                            except Exception as decode_err:
                                logger.warning(f"Error decoding Base64 response: {str(decode_err)}")
                                logger.debug(f"Raw content: {raw_content[:200]}")
                                # Try direct JSON parsing as fallback
                                try:
                                    data = response.json()
                                    return True, data, None
                                except:
                                    return False, None, f"Failed to decode Base64 and JSON: {str(decode_err)}"
                        else:
                            # Try direct JSON parsing
                            try:
                                data = response.json()
                                return True, data, None
                            except json.JSONDecodeError as json_err:
                                return False, None, f"Invalid JSON response: {str(json_err)}"
                    except Exception as e:
                        return False, None, f"Error processing response: {str(e)}"
                
                # Set headers with authorization token
                headers = {
                    'Authorization': f'Bearer {token}',
                    'Content-Type': 'application/json',
                    'Accept-Encoding': 'identity'
                }
                
                # Fetch all users from the API
                users_url = api_base_url
                params = {
                    'count': '1000',  # Try to get as many users as possible in one request
                    'page': '1'
                }
                
                logger.info(f"Fetching users from: {users_url} with params {params}")
                users_response = requests.get(users_url, headers=headers, params=params, timeout=SYNC_HTTP_TIMEOUT)
                
                # Log raw response for debugging
                logger.info(f"Users API response status: {users_response.status_code}")
                try:
                    response_text = users_response.text[:500]  # Limit to first 500 chars
                    logger.debug(f"Response content (truncated): {response_text}")
                except Exception as text_err:
                    logger.error(f"Error getting response text: {str(text_err)}")
                
                if users_response.status_code != 200:
                    # Check for token expiration
                    if users_response.status_code == 422:
                        try:
                            # Use our helper function to decode the response
                            success, resp_data, error_message = decode_response(users_response)
                            
                            if not success:
                                logger.error(f"Failed to decode 422 response: {error_message}")
                                # Try to refresh token as a last resort
                                if credential_id:
                                    logger.info("Attempting token refresh due to decode failure")
                                    new_token = refresh_token(request, credential_id)
                                    if new_token:
                                        token = new_token
                                        headers['Authorization'] = f'Bearer {token}'
                                        logger.info("Token refreshed, retrying the request")
                                        # Try with a smaller batch size
                                        params['count'] = '50'
                                        users_response = requests.get(users_url, headers=headers, params=params, timeout=SYNC_HTTP_TIMEOUT)
                                        if users_response.status_code == 200:
                                            logger.info("Request succeeded after token refresh and smaller batch size")
                                        else:
                                            logger.error(f"Request still failed: {users_response.status_code}")
                                            messages.error(request, f"Failed to fetch users. API returned status code: {users_response.status_code}")
                                            return redirect('api_request_page')
                                    else:
                                        logger.error("Failed to refresh token")
                                        messages.error(request, _("Failed to refresh token after decode failure"))
                                        return redirect('api_request_page')
                                else:
                                    messages.error(request, f"Error processing API response: {error_message}")
                                    return redirect('api_request_page')
                            else:
                                # Check for specific API errors
                                if resp_data.get('respMessage') and 'Expected a value less than 100' in resp_data.get('respMessage'):
                                    # Adjust the count parameter and retry
                                    params['count'] = '50'  # Try with smaller batch size
                                    logger.info(f"API limiting batch size. Retrying with smaller count: {params['count']}")
                                    users_response = requests.get(users_url, headers=headers, params=params, timeout=SYNC_HTTP_TIMEOUT)
                                    if users_response.status_code == 200:
                                        logger.info("Request succeeded with smaller batch size")
                                        # Use decode_response instead of json() method directly
                                        success, decoded_users_data, error = decode_response(users_response)
                                        if success:
                                            users_data = decoded_users_data
                                        else:
                                            logger.error(f"Failed to decode users response after batch size reduction: {error}")
                                            messages.error(request, f"Failed to decode API response: {error}")
                                            return redirect('api_request_page')
                                    else:
                                        logger.error(f"Request still failed with smaller batch size: {users_response.status_code}")
                                        messages.error(request, f"Failed to fetch users with smaller batch size. API returned status code: {users_response.status_code}")
                                        return redirect('api_request_page')
                                # Check for token expiration
                                elif resp_data.get('respCode') == 2401 and resp_data.get('respMessage') == 'Token expired':
                                    # Refresh the token and try again
                                    logger.info("Token expired, attempting to refresh")
                                    if credential_id:
                                        new_token = refresh_token(request, credential_id)
                                        if new_token:
                                            token = new_token
                                            headers['Authorization'] = f'Bearer {token}'
                                            logger.info("Token refreshed, retrying the request")
                                            users_response = requests.get(users_url, headers=headers, params=params, timeout=SYNC_HTTP_TIMEOUT)
                                            if users_response.status_code == 200:
                                                logger.info("Request succeeded after token refresh")
                                            else:
                                                logger.error(f"Request still failed after token refresh: {users_response.status_code}")
                                                messages.error(request, f"Failed to fetch users even after token refresh. API returned status code: {users_response.status_code}")
                                                return redirect('api_request_page')
                                        else:
                                            logger.error("Failed to refresh token")
                                            messages.error(request, _("Failed to refresh expired token"))
                                            return redirect('api_request_page')
                                    else:
                                        logger.error("No credential ID provided for token refresh")
                                        messages.error(request, _("No credential ID provided for token refresh"))
                                        return redirect('api_request_page')
                                else:
                                    # General 422 error not related to token expiration
                                    logger.error(f"API returned 422 error: {resp_data}")
                                    error_message = resp_data.get('respMessage', 'Unknown error')
                                    messages.error(request, f"API error: {error_message}")
                                    return redirect('api_request_page')
                        except Exception as e:
                            logger.error(f"Error handling token expiration: {str(e)}", exc_info=True)
                            messages.error(request, f"Error handling possible token expiration: {str(e)}")
                            return redirect('api_request_page')
                
                # If still not 200 after potential token refresh, fail
                if users_response.status_code != 200:
                    messages.error(request, f"Failed to fetch users. API returned status code: {users_response.status_code}")
                    return redirect('api_request_page')
                
                # Parse users response using the helper function instead of direct json()
                success, decoded_data, error = decode_response(users_response)
                if success:
                    users_data = decoded_data
                    
                    if not users_data or 'data' not in users_data:
                        messages.error(request, "Failed to get valid user data from API")
                        return redirect('api_request_page')
                    
                    # Detailed logging of the raw pagination data to diagnose issues
                    logger.info(f"Raw pagination data: " + json.dumps({
                        'current_page': users_data.get('current_page'),
                        'last_page': users_data.get('last_page'), 
                        'per_page': users_data.get('per_page'),
                        'total': users_data.get('total'),
                        'meta': users_data.get('meta'),
                        'links': users_data.get('links')
                    }, default=str))
                    
                    # Get info about total pages - check multiple possible pagination formats
                    # Standard Laravel pagination
                    current_page = int(users_data.get('current_page', 1))
                    total_pages = int(users_data.get('last_page', 1))
                    per_page = int(users_data.get('per_page', params.get('count', 50)))
                    total_users_count = int(users_data.get('total', len(users_data['data'])))
                    
                    # Alternative pagination formats
                    if not total_pages or total_pages == 1:
                        # Check if pagination is in meta object (common in APIs)
                        meta = users_data.get('meta', {})
                        if meta:
                            current_page = int(meta.get('current_page', current_page))
                            total_pages = int(meta.get('last_page', meta.get('total_pages', total_pages)))
                            per_page = int(meta.get('per_page', per_page))
                            total_users_count = int(meta.get('total', total_users_count))
                    
                    # If we can't determine total_pages but have total count and per_page
                    if total_pages == 1 and total_users_count > per_page:
                        total_pages = math.ceil(total_users_count / per_page)
                        logger.info(f"Calculated total_pages={total_pages} based on total={total_users_count} and per_page={per_page}")
                    
                    logger.info(f"Pagination info: current_page={current_page}, total_pages={total_pages}, per_page={per_page}, total_users={total_users_count}")
                    
                    # If API doesn't provide clear pagination information, use an alternative approach
                    if not users_data.get('last_page') and not users_data.get('meta', {}).get('last_page'):
                        # This API might not return pagination metadata - use a different approach
                        logger.info("API doesn't provide standard pagination info, will try exhaustive page fetching")
                        
                        # Store users from first page
                        all_users = users_data['data']
                        
                        # Safety limit to prevent infinite loops
                        max_pages = 100
                        current_page = 1
                        
                        # Keep fetching pages until we get an empty response or hit max pages
                        while current_page < max_pages:
                            current_page += 1
                            params['page'] = str(current_page)  # Keep as string for URL param
                            logger.info(f"Fetching page {current_page} using exhaustive pagination approach")
                            
                            # Add delay to avoid rate limiting
                            time.sleep(0.5)
                            
                            try:
                                page_response = requests.get(users_url, headers=headers, params=params, timeout=SYNC_HTTP_TIMEOUT)
                                
                                if page_response.status_code == 200:
                                    page_success, page_data, page_error = decode_response(page_response)
                                    
                                    if page_success and 'data' in page_data and isinstance(page_data['data'], list):
                                        # Check if we got any data
                                        if len(page_data['data']) > 0:
                                            logger.info(f"Successfully fetched page {current_page} with {len(page_data['data'])} users")
                                            all_users.extend(page_data['data'])
                                        else:
                                            logger.info(f"Page {current_page} returned empty data array, stopping pagination")
                                            break
                                    else:
                                        logger.error(f"Failed to decode page {current_page} response: {page_error}")
                                        break
                                else:
                                    logger.error(f"Failed to fetch page {current_page}, status code: {page_response.status_code}")
                                    break
                            except Exception as page_err:
                                logger.error(f"Error fetching page {current_page}: {str(page_err)}")
                                break
                        
                        # Update metadata about pagination
                        total_pages = current_page - 1 if current_page > 1 else 1
                        logger.info(f"Exhaustive pagination completed. Fetched {total_pages} pages with {len(all_users)} total users")
                    else:
                        # Standard pagination approach
                        all_users = users_data['data']
                        
                        # If there are more pages, fetch them
                        if total_pages > 1:
                            logger.info(f"Found {total_pages} pages of users, fetching all pages")
                            
                            for page in range(2, total_pages + 1):
                                params['page'] = str(page)  # Keep as string for URL param
                                logger.info(f"Fetching page {page} of {total_pages} from {users_url}")
                                
                                # Add delay to avoid rate limiting
                                time.sleep(0.5)
                                
                                page_response = requests.get(users_url, headers=headers, params=params, timeout=SYNC_HTTP_TIMEOUT)
                                
                                if page_response.status_code == 200:
                                    # Use decode_response helper
                                    page_success, page_data, page_error = decode_response(page_response)
                                    
                                    if page_success and 'data' in page_data and isinstance(page_data['data'], list):
                                        # Check if we actually got data
                                        if len(page_data['data']) > 0:
                                            logger.info(f"Successfully fetched page {page} with {len(page_data['data'])} users")
                                            all_users.extend(page_data['data'])
                                        else:
                                            logger.warning(f"Page {page} returned 0 users, may have reached end of data")
                                            break  # No more data, exit loop
                                    else:
                                        logger.error(f"Failed to decode page {page} response: {page_error}")
                                        messages.warning(request, f"Failed to fetch page {page} of users: {page_error}")
                                else:
                                    logger.error(f"Failed to fetch page {page}, status code: {page_response.status_code}")
                                    messages.warning(request, f"Failed to fetch page {page} of users. API returned status code: {page_response.status_code}")
                                    
                                    # Try to handle common errors
                                    if page_response.status_code == 401 or page_response.status_code == 403:
                                        logger.info("Token may have expired during pagination, attempting to refresh")
                                        if credential_id:
                                            new_token = refresh_token(request, credential_id)
                                            if new_token:
                                                token = new_token
                                                headers['Authorization'] = f'Bearer {token}'
                                                logger.info("Token refreshed, retrying the current page")
                                                # Don't increment page counter, retry the same page
                                                page -= 1
                                                continue
                                
                                    # If serious error, might be best to stop rather than continue with partial data
                                    if page_response.status_code >= 500:
                                        logger.error("Server error during pagination, stopping fetch to avoid hammering server")
                                        break
                    
                    # Check for duplicate users in the API response
                    total_users = len(all_users)
                    unique_user_ids = len(set(user.get('id') for user in all_users if user.get('id')))
                    unique_hashes = len(set(user.get('hash') for user in all_users if user.get('hash')))

                    logger.info(f"Total users: {total_users}, Unique user_ids: {unique_user_ids}, Unique hashes: {unique_hashes}")
                    if total_users > unique_user_ids or total_users > unique_hashes:
                        logger.warning(f"Found {total_users - unique_user_ids} duplicate user_ids and {total_users - unique_hashes} duplicate hashes in API response")
                    
                    # Use merged list from all pages for the rest of sync (roles, DB, stale removal)
                    users_data['data'] = all_users
                    logger.info(f"Found {len(all_users)} users to process after pagination merge")

                # Define helper functions for fetching user data
                def fetch_user_roles(user_hash):
                    try:
                        roles_url = f"{api_base_url}/{user_hash}/roles"
                        fetch_headers = {
                            'Authorization': f'Bearer {token}',
                            'Content-Type': 'application/json',
                            'Accept-Encoding': 'identity'
                        }
                        
                        logger.info(f"Fetching roles for user {user_hash} from {roles_url}")
                        append_terminal_line(f"INFO Fetching roles for user {user_hash}")
                        response = requests.get(roles_url, headers=fetch_headers, timeout=SYNC_HTTP_TIMEOUT)
                        
                        if response.status_code == 200:
                            # Use decode_response helper to handle potential Base64 encoding
                            success, data, error = decode_response(response)
                            if success:
                                return data
                            else:
                                logger.error(f"Failed to decode roles response: {error}")
                                return None
                        else:
                            logger.warning(f"Roles API returned {response.status_code} for user {user_hash}")
                            return None
                    except Exception as e:
                        logger.error(f"Error fetching roles for user {user_hash}: {str(e)}")
                        return None
                
                def fetch_user_status(user_hash, api_user):
                    try:
                        status_url = f"{api_base_url}/{user_hash}/status"
                        fetch_headers = {
                            'Authorization': f'Bearer {token}',
                            'Content-Type': 'application/json',
                            'Accept-Encoding': 'identity'
                        }
                        
                        logger.info(f"Fetching status for user {user_hash} from {status_url}")
                        append_terminal_line(f"INFO Fetching status for user {user_hash}")
                        response = requests.get(status_url, headers=fetch_headers, timeout=SYNC_HTTP_TIMEOUT)
                        
                        # Check response status code
                        if response.status_code == 200:
                            # Use decode_response helper to handle potential Base64 encoding
                            success, status_data, error = decode_response(response)
                            
                            if success:
                                # Create or update ApiUserStatus
                                raw_status = status_data.get('status', 'unknown')
                                mapped_status = 'unknown'
                                
                                if raw_status.lower() in ['active', 'активний']:
                                    mapped_status = 'active'
                                elif raw_status.lower() in ['blocked', 'заблокований']:
                                    mapped_status = 'blocked'
                                elif raw_status.lower() in ['temporary unavailable', 'тимчасово недоступний']:
                                    mapped_status = 'temporary_unavailable'
                                
                                # Create or get ApiUserStatus
                                status_obj, created = ApiUserStatus.objects.update_or_create(
                                    user_id=api_user.id,
                                    defaults={
                                        'status': mapped_status,
                                        'raw_status': raw_status,
                                        'sync': sync_status
                                    }
                                )
                                
                                # Add detailed logging about status
                                if created:
                                    logger.info(f"Created new status entry for user {api_user.email}: {mapped_status}")
                                else:
                                    logger.debug(f"Updated existing status for user {api_user.email} to: {mapped_status}")
                                
                                # Return created flag so caller can update counters
                                return status_obj, created
                            else:
                                logger.warning(f"Failed to decode status data: {error}")
                                # Create a default status object for invalid response
                                status_obj, created = ApiUserStatus.objects.update_or_create(
                                    user_id=api_user.id,
                                    defaults={
                                        'status': 'unknown',
                                        'raw_status': f'Failed to decode: {error}',
                                        'sync': sync_status
                                    }
                                )
                                return status_obj, created
                        else:
                            logger.warning(f"Status API returned {response.status_code} for user {user_hash}")
                            # Don't display message for each error to avoid cluttering the UI
                            return None, False
                    except Exception as e:
                        logger.error(f"Error fetching status for user {user_hash}: {str(e)}")
                        return None, False
                
                def fetch_user_permission_history(user_hash, api_user):
                    try:
                        # Get recent permission history (last 3 months)
                        three_months_ago = (timezone.now() - timezone.timedelta(days=90)).strftime('%Y-%m-%d %H:%M:%S')
                        today = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        history_url = f"{api_base_url}/{user_hash}/permissions-history"
                        params = {
                            'page': '1',
                            'count': '50',  # Get up to 50 entries
                            'filter[from]': three_months_ago,
                            'filter[to]': today
                        }
                        
                        # Build URL with parameters
                        full_url = f"{history_url}?{'&'.join([f'{k}={v}' for k, v in params.items()])}"
                        
                        headers = {
                            'Authorization': f'Bearer {token}',
                            'Content-Type': 'application/json',
                            'Accept-Encoding': 'identity'  # Disable compression for easier debugging
                        }
                        
                        logger.info(f"Fetching permission history for user {user_hash} from {full_url}")
                        append_terminal_line(f"INFO Fetching permission history for user {user_hash}")
                        response = requests.get(full_url, headers=headers, timeout=SYNC_HTTP_TIMEOUT)
                        
                        history_created = 0
                        if response.status_code == 200:
                            # Use decode_response helper to handle potential Base64 encoding
                            success, history_data, error = decode_response(response)
                            
                            if success:
                                if 'data' in history_data and isinstance(history_data['data'], list):
                                    for entry in history_data['data']:
                                        if 'time' in entry:
                                            try:
                                                # Parse timestamp
                                                entry_time = timezone.datetime.fromisoformat(entry['time'])
                                                
                                                # Check if entry already exists
                                                existing_entry = ApiUserPermissionHistory.objects.filter(
                                                    user=api_user,
                                                    time=entry_time
                                                ).first()
                                                
                                                if not existing_entry:
                                                    # Create new history entry
                                                    history_entry = ApiUserPermissionHistory(
                                                        user=api_user,
                                                        time=entry_time,
                                                        added_permissions=entry.get('addedPermissions', {}),
                                                        removed_permissions=entry.get('removedPermissions', {}),
                                                        raw_data=entry,
                                                        sync=sync_status
                                                    )
                                                    history_entry.save()
                                                    history_created += 1
                                                    logger.debug(f"Created permission history entry for user {user_hash} at {entry_time}")
                                            except Exception as e:
                                                logger.error(f"Error saving permission history entry: {str(e)}", exc_info=True)
                                    else:
                                        logger.warning(f"No permission history data found in response for user {user_hash}")
                                else:
                                    logger.warning(f"No permission history data found in response for user {user_hash}")
                            else:
                                logger.warning(f"Failed to decode permission history: {error}")
                        else:
                            logger.warning(f"Permission history API returned {response.status_code} for user {user_hash}")
                        
                        return history_created
                    except Exception as e:
                        logger.error(f"Error fetching permission history for user {user_hash}: {str(e)}", exc_info=True)
                        return 0
                
                def fetch_user_login_history(user_hash, api_user):
                    """
                    Fetch user login history from API
                    """
                    try:
                        # Get recent login history (last 3 months)
                        three_months_ago = (timezone.now() - timezone.timedelta(days=90)).strftime('%Y-%m-%d %H:%M:%S')
                        today = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                        
                        history_url = f"{api_base_url}/{user_hash}/login-history"
                        params = {
                            'page': '1',
                            'count': '50',  # Get up to 50 entries
                            'filter[from]': three_months_ago,
                            'filter[to]': today
                        }
                        
                        # Build URL with parameters
                        full_url = f"{history_url}?{'&'.join([f'{k}={v}' for k, v in params.items()])}"
                        
                        headers = {
                            'Authorization': f'Bearer {token}',
                            'Content-Type': 'application/json',
                            'Accept-Encoding': 'identity'
                        }
                        
                        logger.info(f"Fetching login history for user {user_hash} from {full_url}")
                        append_terminal_line(f"INFO Fetching login history for user {user_hash}")
                        response = requests.get(full_url, headers=headers, timeout=SYNC_HTTP_TIMEOUT)
                        
                        history_created = 0
                        if response.status_code == 200:
                            # Use decode_response helper to handle potential Base64 encoding
                            success, history_data, error = decode_response(response)
                            
                            if success:
                                if 'data' in history_data and isinstance(history_data['data'], list):
                                    for entry in history_data['data']:
                                        if 'time' in entry and 'ip' in entry:
                                            try:
                                                # Parse timestamp
                                                entry_time = timezone.datetime.fromisoformat(entry['time'].replace('Z', '+00:00'))
                                                
                                                # Check if entry already exists
                                                existing_entry = ApiUserLoginHistory.objects.filter(
                                                    user=api_user,
                                                    time=entry_time,
                                                    ip=entry['ip']
                                                ).first()
                                                
                                                if not existing_entry:
                                                    # Create new history entry
                                                    history_entry = ApiUserLoginHistory(
                                                        user=api_user,
                                                        time=entry_time,
                                                        ip=entry['ip'],
                                                        sync=sync_status
                                                    )
                                                    history_entry.save()
                                                    history_created += 1
                                                    logger.debug(f"Created login history entry for user {user_hash} at {entry_time}")
                                            except Exception as e:
                                                logger.error(f"Error saving login history entry: {str(e)}", exc_info=True)
                                else:
                                    logger.warning(f"No login history data found in response for user {user_hash}")
                            else:
                                logger.warning(f"Failed to decode login history: {error}")
                        else:
                            logger.warning(f"Login history API returned {response.status_code} for user {user_hash}")
                        
                        return history_created
                    except Exception as e:
                        logger.error(f"Error fetching login history for user {user_hash}: {str(e)}", exc_info=True)
                        return 0
                
                def fetch_user_merchants(user_hash, api_user):
                    """
                    Fetch user merchant links from API
                    """
                    try:
                        merchants_url = f"{api_base_url}/{user_hash}/merchants"
                        
                        headers = {
                            'Authorization': f'Bearer {token}',
                            'Content-Type': 'application/json',
                            'Accept-Encoding': 'identity'
                        }
                        
                        logger.info(f"Fetching merchant links for user {user_hash} from {merchants_url}")
                        append_terminal_line(f"INFO Fetching merchant links for user {user_hash}")
                        response = requests.get(merchants_url, headers=headers, timeout=SYNC_HTTP_TIMEOUT)
                        
                        if response.status_code == 200:
                            # Use decode_response helper to handle potential Base64 encoding
                            success, merchants_data, error = decode_response(response)
                            
                            if success and isinstance(merchants_data, list):
                                # Save merchant links - first delete existing ones
                                api_user.merchant_links.all().delete()
                                
                                merchants_created = 0
                                for merchant_name in merchants_data:
                                    if merchant_name:
                                        merchant_link = ApiUserMerchantLink(
                                            user=api_user,
                                            merchant_name=merchant_name,
                                            sync=sync_status
                                        )
                                        merchant_link.save()
                                        merchants_created += 1
                                        logger.debug(f"Created merchant link '{merchant_name}' for user {user_hash}")
                                
                                append_terminal_line(f"INFO Created {merchants_created} merchant links for user {api_user.email}")
                                return merchants_created
                            else:
                                if error:
                                    logger.warning(f"Failed to decode merchant links: {error}")
                                else:
                                    logger.warning(f"Merchant data not in expected format for user {user_hash}")
                                return 0
                        else:
                            logger.warning(f"Merchant API returned {response.status_code} for user {user_hash}")
                            return 0
                    except Exception as e:
                        logger.error(f"Error fetching merchant links for user {user_hash}: {str(e)}", exc_info=True)
                        return 0
                
                # Process each user
                logger.info(f"Starting to process {len(users_data['data'])} users")
                
                # Track processed user IDs to avoid duplicates
                processed_user_ids = set()
                processed_hashes = set()
                total_users = len(users_data['data'])
                processed_count = 0
                total_sync_steps = total_users * USER_SYNC_PHASES
                
                # Total micro-steps for progress (users × phases per user)
                sync_status.total_steps = total_sync_steps
                sync_status.save(update_fields=['total_steps'])
                
                for user_data in users_data['data']:
                    # Check if sync was stopped
                    sync_status.refresh_from_db()
                    if sync_status.status == 'stopped':
                        logger.info("Sync process was stopped by user")
                        return JsonResponse({
                            'success': True,
                            'message': _("Synchronization process stopped by user.")
                        })
                    
                    # Update progress based on processed users
                    processed_count += 1
                    
                    # Fine-grained step: between users / before sub-requests for this user
                    completed_micro = (processed_count - 1) * USER_SYNC_PHASES
                    progress_percent = min(100, int((completed_micro / total_sync_steps) * 100)) if total_sync_steps else 0
                    
                    # Update sync status with progress
                    update_sync_progress(
                        _('Processing users...'),
                        completed_micro,
                        total_sync_steps,
                        {
                            'total_api_users': total_users,
                            'processed_users': processed_count,
                            'existing_users': ApiUser.objects.count(),
                            'created_users': user_count_created,
                            'updated_users': user_count_updated,
                            'percent': progress_percent
                        }
                    )
                    
                    try:
                        # Extract user data
                        user_id = user_data.get('id')
                        email = user_data.get('email')
                        hash_value = user_data.get('hash')
                        first_name = user_data.get('firstName')
                        last_name = user_data.get('lastName')
                        phone = user_data.get('phone')
                        last_login = user_data.get('last_login')

                        append_terminal_line(
                            _("Processing user {current}/{total}: {email} (id={uid}, hash={h})").format(
                                current=processed_count,
                                total=total_users,
                                email=email or '-',
                                uid=user_id or '-',
                                h=(hash_value[:12] + '…') if hash_value and len(hash_value) > 12 else (hash_value or '-'),
                            )
                        )
                        
                        # Skip if missing essential data
                        if not user_id or not hash_value or not email:
                            logger.warning(f"Skipping user with incomplete data: id={user_id}, hash={hash_value}, email={email}")
                            sync_status.touch_step(
                                _('Skipping user {cur}/{tot} (incomplete data)').format(
                                    cur=processed_count, tot=total_users
                                ),
                                processed_count * USER_SYNC_PHASES,
                                total_sync_steps,
                            )
                            continue
                        
                        # Skip if we've already processed this user in this batch
                        if user_id in processed_user_ids or hash_value in processed_hashes:
                            logger.debug(f"Skipping duplicate user in batch: id={user_id}, hash={hash_value}, email={email}")
                            sync_status.touch_step(
                                _('Skipping duplicate user {cur}/{tot}').format(
                                    cur=processed_count, tot=total_users
                                ),
                                processed_count * USER_SYNC_PHASES,
                                total_sync_steps,
                            )
                            continue
                        
                        # Add to processed sets
                        processed_user_ids.add(user_id)
                        processed_hashes.add(hash_value)
                        
                        api_user = None
                        # Check if user already exists with this id and hash
                        logger.debug(f"Checking if user with ID {user_id} exists")
                        append_terminal_line(f"DEBUG Checking if user with ID {user_id} exists")
                        try:
                            existing_user = ApiUser.objects.filter(user_id=user_id).first()
                            
                            if existing_user:
                                logger.debug(f"User {email} (ID: {user_id}) exists, checking for changes")
                                append_terminal_line(f"DEBUG User {email} (ID: {user_id}) exists, checking for changes")
                                # Update existing user
                                was_modified = False
                                if existing_user.hash != hash_value:
                                    logger.debug(f"User {email} hash changed from {existing_user.hash} to {hash_value}")
                                    existing_user.hash = hash_value
                                    was_modified = True
                                if existing_user.email != email:
                                    logger.debug(f"User {user_id} email changed from {existing_user.email} to {email}")
                                    existing_user.email = email
                                    was_modified = True
                                if existing_user.first_name != first_name:
                                    existing_user.first_name = first_name
                                    was_modified = True
                                if existing_user.last_name != last_name:
                                    existing_user.last_name = last_name
                                    was_modified = True
                                if existing_user.phone != phone:
                                    existing_user.phone = phone
                                    was_modified = True
                                
                                # Process last_login field if it exists in the API response
                                if last_login and (not existing_user.last_login or existing_user.last_login.isoformat() != last_login):
                                    try:
                                        # Parse ISO format datetime
                                        last_login_dt = timezone.datetime.fromisoformat(last_login.replace('Z', '+00:00'))
                                        existing_user.last_login = last_login_dt
                                        was_modified = True
                                        logger.debug(f"Updated last_login for user {email} to {last_login}")
                                        append_terminal_line(f"DEBUG Updated last_login for user {email} to {last_login}")
                                    except Exception as dt_err:
                                        logger.error(f"Error parsing last_login date {last_login} for user {email}: {str(dt_err)}")
                                
                                if was_modified:
                                    # Set the sync reference
                                    existing_user.sync = sync_status
                                    existing_user.api_credential = credential
                                    existing_user.save()
                                    user_count_updated += 1
                                    logger.info(f"Updated API user: {email}")
                                    append_terminal_line(f"INFO Updated API user: {email}")
                                else:
                                    # Still update the sync reference even if no other fields changed
                                    existing_user.sync = sync_status
                                    existing_user.api_credential = credential
                                    existing_user.save(update_fields=['sync', 'api_credential'])
                                    logger.debug(f"No data changes needed for API user: {email}, but updated sync reference")
                                
                                api_user = existing_user
                                created = False
                            else:
                                logger.info(f"Creating new user {email} (ID: {user_id})")
                                append_terminal_line(f"INFO Creating new user {email} (ID: {user_id})")
                                # Create new user
                                try:
                                    # Parse last_login if available
                                    last_login_dt = None
                                    if last_login:
                                        try:
                                            last_login_dt = timezone.datetime.fromisoformat(last_login.replace('Z', '+00:00'))
                                        except Exception as dt_err:
                                            logger.error(f"Error parsing last_login date {last_login} for new user {email}: {str(dt_err)}")
                                    
                                    api_user = ApiUser.objects.create(
                                        user_id=user_id,
                                        hash=hash_value,
                                        email=email,
                                        first_name=first_name,
                                        last_name=last_name,
                                        phone=phone,
                                        last_login=last_login_dt,
                                        sync=sync_status,
                                        api_credential=credential,
                                    )
                                    user_count_created += 1
                                    logger.info(f"Created new API user: {email}")
                                    append_terminal_line(f"INFO Created new API user: {email}")
                                    created = True
                                except Exception as create_error:
                                    # Check if the exception is due to a unique constraint violation
                                    error_msg = str(create_error)
                                    if 'duplicate key value violates unique constraint' in error_msg or 'UNIQUE constraint failed' in error_msg:
                                        constraint_field = None
                                        if 'user_id' in error_msg:
                                            constraint_field = 'user_id'
                                        elif 'hash' in error_msg:
                                            constraint_field = 'hash'
                                        
                                        logger.warning(f"User creation failed due to duplicate {constraint_field or 'unknown field'} constraint: user_id={user_id}, hash={hash_value}, email={email}")
                                        
                                        try:
                                            # Try to get the user by hash if that's what's duplicated
                                            api_user = ApiUser.objects.filter(hash=hash_value).first()
                                            if not api_user:
                                                # If not found by hash, try by user_id
                                                api_user = ApiUser.objects.filter(user_id=user_id).first()
                                            
                                            if api_user:
                                                logger.info(f"Found existing user: {api_user.email} (ID: {api_user.user_id})")
                                                api_user.sync = sync_status
                                                api_user.api_credential = credential
                                                api_user.save(update_fields=['sync', 'api_credential'])
                                                created = False
                                            else:
                                                # If we can't find the user by any method, skip this user
                                                logger.error(f"Could not find or create user with ID {user_id}, hash {hash_value}")
                                                sync_status.touch_step(
                                                    _('User {cur}/{tot}: cannot resolve duplicate').format(
                                                        cur=processed_count, tot=total_users
                                                    ),
                                                    processed_count * USER_SYNC_PHASES,
                                                    total_sync_steps,
                                                )
                                                continue
                                        except Exception as fetch_error:
                                            logger.error(f"Error fetching existing user after unique constraint violation: {str(fetch_error)}")
                                            sync_status.touch_step(
                                                _('User {cur}/{tot}: error loading user after duplicate').format(
                                                    cur=processed_count, tot=total_users
                                                ),
                                                processed_count * USER_SYNC_PHASES,
                                                total_sync_steps,
                                            )
                                            continue
                                    else:
                                        # If it's not a unique constraint violation, re-raise the exception
                                        logger.error(f"Error creating new user {email}: {str(create_error)}")
                                        sync_status.touch_step(
                                            _('User {cur}/{tot}: create failed').format(
                                                cur=processed_count, tot=total_users
                                            ),
                                            processed_count * USER_SYNC_PHASES,
                                            total_sync_steps,
                                        )
                                        continue
                        except Exception as e:
                            logger.error(f"Error processing user {user_data.get('email', 'unknown')}: {str(e)}", exc_info=True)
                            messages.error(request, f"Error processing user {user_data.get('email', 'unknown')}")
                        
                        if api_user is None:
                            sync_status.touch_step(
                                _('User {cur}/{tot}: skipped — no local user record').format(
                                    cur=processed_count, tot=total_users
                                ),
                                processed_count * USER_SYNC_PHASES,
                                total_sync_steps,
                            )
                            continue
                        
                        sync_status.touch_step(
                            _('User {cur}/{tot}: {email} — {phase}').format(
                                cur=processed_count,
                                tot=total_users,
                                email=email or '-',
                                phase=_('roles'),
                            ),
                            (processed_count - 1) * USER_SYNC_PHASES + 1,
                            total_sync_steps,
                        )
                        
                        # Get user roles - wrapped in try/except
                        try:
                            roles_data = fetch_user_roles(hash_value)
                            
                            if roles_data:
                                # Delete existing role mappings for this user to avoid duplicates
                                ApiUserRoleMapping.objects.filter(user=api_user).delete()
                                
                                # Process merchant and role data
                                for merchant_data in roles_data:
                                    merchant_name = merchant_data.get('name', '')
                                    if not merchant_name:
                                        continue
                                    
                                    # Create or get merchant
                                    merchant, merchant_created = ApiUserMerchant.objects.get_or_create(
                                        name=merchant_name,
                                        defaults={'sync': sync_status}
                                    )
                                    
                                    if merchant_created:
                                        merchant_count_created += 1
                                    
                                    # Process roles for this merchant
                                    roles_list = merchant_data.get('roles', [])
                                    for role_info in roles_list:
                                        role_id = role_info.get('roleId')
                                        role_name = role_info.get('roleName', '')
                                        
                                        if not role_id or not role_name:
                                            continue
                                        
                                        # Create or get role
                                        role, role_created = ApiUserRole.objects.update_or_create(
                                            role_id=role_id,
                                            defaults={'name': role_name, 'sync': sync_status}
                                        )
                                        
                                        if role_created:
                                            role_count_created += 1
                                        
                                        # Create the role mapping
                                        role_mapping, created = ApiUserRoleMapping.objects.get_or_create(
                                            user=api_user,
                                            merchant=merchant,
                                            role=role,
                                            defaults={'sync': sync_status}
                                        )
                                        
                                        # Update sync reference even if the mapping already existed
                                        if not created:
                                            role_mapping.sync = sync_status
                                            role_mapping.save(update_fields=['sync'])
                                        
                                        if created:
                                            role_mapping_count_created += 1
                                            logger.debug(f"Created role mapping for user {api_user.email}: {merchant.name} -> {role.name}")
                                            append_terminal_line(f"DEBUG Created role mapping for user {api_user.email}: {merchant.name} -> {role.name}")
                            else:
                                logger.info(f"No roles data found for user {api_user.email}")
                                append_terminal_line(f"INFO No roles data found for user {api_user.email}")
                        except Exception as roles_error:
                            logger.error(f"Error processing roles for user {api_user.email}: {str(roles_error)}")
                        
                        sync_status.touch_step(
                            _('User {cur}/{tot}: {email} — {phase}').format(
                                cur=processed_count,
                                tot=total_users,
                                email=api_user.email,
                                phase=_('status'),
                            ),
                            (processed_count - 1) * USER_SYNC_PHASES + 2,
                            total_sync_steps,
                        )
                        
                        # Get user status - wrapped in try/except
                        try:
                            status_obj, created = fetch_user_status(hash_value, api_user)
                            if status_obj:
                                if created:
                                    status_count_created += 1
                                else:
                                    status_count_updated += 1
                                logger.debug(f"Updated status for user {api_user.email}: {status_obj.status}")
                                append_terminal_line(f"DEBUG Updated status for user {api_user.email}: {status_obj.status}")
                            else:
                                logger.info(f"No valid status data found for user {api_user.email}")
                                append_terminal_line(f"INFO No valid status data found for user {api_user.email}")
                        except Exception as status_error:
                            logger.error(f"Error processing status for user {api_user.email}: {str(status_error)}")
                        
                        sync_status.touch_step(
                            _('User {cur}/{tot}: {email} — {phase}').format(
                                cur=processed_count,
                                tot=total_users,
                                email=api_user.email,
                                phase=_('permission history'),
                            ),
                            (processed_count - 1) * USER_SYNC_PHASES + 3,
                            total_sync_steps,
                        )
                        
                        # Get user permission history - wrapped in try/except
                        try:
                            history_count = fetch_user_permission_history(hash_value, api_user)
                            if history_count > 0:
                                logger.info(f"Created {history_count} permission history entries for user {api_user.email}")
                                append_terminal_line(f"INFO Created {history_count} permission history entries for user {api_user.email}")
                            history_entries_created += history_count
                        except Exception as history_error:
                            logger.error(f"Error processing permission history for user {api_user.email}: {str(history_error)}")
                        
                        sync_status.touch_step(
                            _('User {cur}/{tot}: {email} — {phase}').format(
                                cur=processed_count,
                                tot=total_users,
                                email=api_user.email,
                                phase=_('login history'),
                            ),
                            (processed_count - 1) * USER_SYNC_PHASES + 4,
                            total_sync_steps,
                        )
                        
                        # Get user login history - wrapped in try/except
                        try:
                            login_history_count = fetch_user_login_history(hash_value, api_user)
                            if login_history_count > 0:
                                logger.info(f"Created {login_history_count} login history entries for user {api_user.email}")
                                append_terminal_line(f"INFO Created {login_history_count} login history entries for user {api_user.email}")
                            login_history_entries_created += login_history_count
                        except Exception as login_history_error:
                            logger.error(f"Error processing login history for user {api_user.email}: {str(login_history_error)}")
                        
                        sync_status.touch_step(
                            _('User {cur}/{tot}: {email} — {phase}').format(
                                cur=processed_count,
                                tot=total_users,
                                email=api_user.email,
                                phase=_('merchant links'),
                            ),
                            (processed_count - 1) * USER_SYNC_PHASES + 5,
                            total_sync_steps,
                        )
                        
                        # Get user merchant links - wrapped in try/except
                        try:
                            merchant_links_created = fetch_user_merchants(hash_value, api_user)
                            if merchant_links_created > 0:
                                logger.info(f"Created {merchant_links_created} merchant links for user {api_user.email}")
                            merchant_links_count_created += merchant_links_created
                        except Exception as merchant_links_error:
                            logger.error(f"Error processing merchant links for user {api_user.email}: {str(merchant_links_error)}")
                        
                    except Exception as e:
                        logger.error(f"Error processing user {user_data.get('email', 'unknown')}: {str(e)}", exc_info=True)
                        messages.error(request, f"Error processing user {user_data.get('email', 'unknown')}")
                
                # Users removed from the external API: delete local rows tied to this credential only
                api_seen_ids = set()
                for ud in users_data['data']:
                    uid = ud.get('id')
                    if uid is not None:
                        try:
                            api_seen_ids.add(int(uid))
                        except (TypeError, ValueError):
                            continue
                stale_users = ApiUser.objects.filter(api_credential=credential).exclude(user_id__in=api_seen_ids)
                user_count_removed = stale_users.count()
                if user_count_removed:
                    msg_rm = _('Removing {n} users no longer present in the API for this credential.').format(
                        n=user_count_removed
                    )
                    append_terminal_line(msg_rm)
                    logger.info(
                        'Removing %s ApiUser rows absent from API response (credential_id=%s)',
                        user_count_removed,
                        credential.id,
                    )
                    stale_users.delete()

                append_terminal_line(_('Finalizing synchronization...'))
                payload_fin = dict(sync_status.extra_data) if isinstance(sync_status.extra_data, dict) else {}
                payload_fin['terminal_lines'] = terminal_lines.copy()
                sync_status.update_progress(
                    _('Finalizing synchronization...'),
                    total_sync_steps,
                    total_sync_steps,
                    payload_fin,
                )
                
                user_count = total_users
                
                # Success message with summary
                summary_message = (
                    f"Sync completed successfully. {user_count} users processed, "
                    f"{user_count_created} users created, {user_count_updated} users updated, {user_count_existing} users existing, "
                    f"{user_count_removed} users removed (no longer in API), "
                    f"{role_count_created} roles created, {merchant_count_created} merchants created, "
                    f"{role_mapping_count_created} role mappings created, {status_count_created} status entries created, "
                    f"{status_count_updated} status entries updated, {history_entries_created} history entries created, "
                    f"{login_history_entries_created} login history entries created, "
                    f"{merchant_links_count_created} merchant links created."
                )
                
                # Mark sync as completed
                append_terminal_line(_('Synchronization completed successfully'))
                sync_status.complete()
                
                messages.success(request, summary_message)
                # For AJAX requests, return JSON response
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': summary_message,
                        'status': 'completed'
                    })
                # For regular form submissions, redirect to the API request page
                return redirect('api_request_page')
                
            except Exception as e:
                if isinstance(e, (requests.exceptions.Timeout, requests.exceptions.ConnectionError, requests.exceptions.RequestException)):
                    error_message = _("Synchronization was stopped on the API server side or the API server is unavailable.")
                else:
                    error_message = str(e)

                append_terminal_line(_("Sync failed: {}").format(error_message))
                sync_status.error(error_message)
                logger.error(f"Error during sync: {str(e)}", exc_info=True)
                messages.error(request, f"{_('Error during sync')}: {error_message}")
                # For AJAX requests, return JSON response
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': error_message,
                        'status': 'error'
                    })
                # For regular form submissions, redirect to the API request page
                return redirect('api_request_page')
                
        except Exception as e:
            logger.error(f"Error in sync_api_users: {str(e)}", exc_info=True)
            messages.error(request, f"Error in sync_api_users: {str(e)}")
            # For AJAX requests, return JSON response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': str(e),
                    'status': 'error'
                })
            # For regular form submissions, redirect to the API request page
            return redirect('api_request_page')
    
    return redirect('api_request_page')

@login_required
def api_synced_users(request):
    """
    Display list of users synced from the API
    """
    users = ApiUser.objects.all().order_by('-updated_at')
    
    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(users, 20)  # Show 20 users per page
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': 'Synced API Users',
        'page_obj': page_obj,
        'total_users': users.count(),
    }
    
    return render(request, 'app_access/api_synced_users.html', context)

@login_required
@require_POST
def clear_sync_data(request):
    """
    Clear all synchronized API user data
    """
    try:
        with transaction.atomic():
            # Delete all data in reverse order of dependencies
            ApiUserPermissionHistory.objects.all().delete()
            ApiUserStatus.objects.all().delete()
            ApiUserRoleMapping.objects.all().delete()
            ApiUserRole.objects.all().delete()
            ApiUserMerchant.objects.all().delete()
            ApiUser.objects.all().delete()

            logger.info("Successfully cleared all synchronized API user data")
            return JsonResponse({
                'success': True,
                'message': _("All synchronized data has been cleared successfully.")
            })
    except Exception as e:
        logger.error(f"Error clearing synchronized data: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def export_api_users_xlsx(request):
    """
    Export API users to XLSX (Excel) file
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from io import BytesIO
    import datetime
    
    # Get query parameters
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)
    show_all = request.GET.get('show_all', 'false').lower() == 'true'
    
    # Get column filters
    filters = {}
    filter_params = {}
    for key, value in request.GET.items():
        if key.startswith('filter_') and value:
            field_name = key[7:]  # Remove 'filter_' prefix
            filters[field_name] = value
            filter_params[key] = value
    
    # Get synced users with related models
    users = ApiUser.objects.select_related('status_info').prefetch_related(
        'role_mappings',
        'role_mappings__merchant',
        'role_mappings__role',
        'permission_history',
        'login_history',
        'merchant_links',
    ).all().order_by('-updated_at')
    
    # Apply search filter if provided
    if search_query:
        users = users.filter(
            # Search in multiple fields - case insensitive
            models.Q(email__icontains=search_query) |
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(hash__icontains=search_query) |
            models.Q(phone__icontains=search_query) |
            models.Q(user_id__icontains=search_query)
        )
    
    # Apply column filters
    if 'user_id' in filters:
        users = users.filter(user_id=filters['user_id'])
    if 'email' in filters:
        users = users.filter(email__icontains=filters['email'])
    if 'name' in filters:
        name_filter = filters['name']
        users = users.filter(
            models.Q(first_name__icontains=name_filter) | 
            models.Q(last_name__icontains=name_filter)
        )
    if 'phone' in filters:
        users = users.filter(phone__icontains=filters['phone'])
    if 'status' in filters:
        users = users.filter(status_info__status=filters['status'])
    
    # Date range filters for last_login
    if 'last_login_from' in filters and filters['last_login_from']:
        users = users.filter(last_login__date__gte=filters['last_login_from'])
    if 'last_login_to' in filters and filters['last_login_to']:
        users = users.filter(last_login__date__lte=filters['last_login_to'])
    
    # Date range filters for updated_at
    if 'updated_from' in filters and filters['updated_from']:
        users = users.filter(updated_at__date__gte=filters['updated_from'])
    if 'updated_to' in filters and filters['updated_to']:
        users = users.filter(updated_at__date__lte=filters['updated_to'])
    
    if 'merchant_name' in filters:
        merchant_filter = filters['merchant_name']
        users = users.filter(merchant_links__merchant_name__icontains=merchant_filter)
    if 'role_name' in filters:
        role_filter = filters['role_name']
        users = users.filter(role_mappings__role__name__icontains=role_filter)
    if 'merchant' in filters:
        merchant_filter = filters['merchant']
        users = users.filter(role_mappings__merchant__name__icontains=merchant_filter)
    
    # Apply distinct to remove duplicates that may occur due to joining with related models
    if any(key in filters for key in ['merchant_name', 'role_name', 'merchant']):
        users = users.distinct()
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Users"
    
    # Define styles
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    # Fonts
    header_font = Font(name='Arial', bold=True, color="FFFFFF", size=11)
    normal_font = Font(name='Arial', size=10)
    id_font = Font(name='Arial', bold=True, size=10)
    date_font = Font(name='Arial', italic=True, size=10)
    
    # Fills
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    odd_row_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    status_active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    status_blocked_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Alignments
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # Define column headers
    headers = [
        str(_("User ID")), 
        str(_("Email")), 
        str(_("First Name")), 
        str(_("Last Name")), 
        str(_("Phone")),
        str(_("Status")),
        str(_("Last Login")), 
        str(_("Created At")),
        str(_("Updated At")), 
        str(_("Third parties")), 
        str(_("Roles"))
    ]
    
    # Add title row with metadata
    title = f"API Users Export - {datetime.datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws.merge_cells('A1:K1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='Arial', bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    # Add filter info if applied
    filter_row = 2
    if search_query or filters:
        filter_text = f"Filters applied: {search_query if search_query else ''}"
        if filters:
            filter_text += ", " + ", ".join([f"{k}: {v}" for k, v in filters.items()])
        ws.merge_cells(f'A{filter_row}:K{filter_row}')
        filter_cell = ws.cell(row=filter_row, column=1, value=filter_text)
        filter_cell.font = Font(name='Arial', italic=True, size=10)
        filter_cell.alignment = left_alignment
        filter_row += 1
    
    # Add total records count
    count_text = f"Total records: {users.count()}"
    ws.merge_cells(f'A{filter_row}:K{filter_row}')
    count_cell = ws.cell(row=filter_row, column=1, value=count_text)
    count_cell.font = Font(name='Arial', bold=True, size=10)
    count_cell.alignment = left_alignment
    
    # Start headers at row 4
    header_row = filter_row + 1
    
    # Write headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thick_bottom_border
    
    # Set column widths
    column_widths = [12, 35, 20, 20, 15, 15, 20, 20, 20, 40, 40]
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width
    
    # Write data to worksheet
    row_num = header_row + 1
    for idx, user in enumerate(users):
        # Get status text
        status_text = getattr(user.status_info, 'status', 'Unknown') if hasattr(user, 'status_info') else 'Unknown'
        
        # Format datetime objects in dd-mm-yyyy format
        last_login = user.last_login.strftime('%d-%m-%Y %H:%M:%S') if user.last_login else 'N/A'
        created_at = user.created_at.strftime('%d-%m-%Y %H:%M:%S') if user.created_at else 'N/A'
        updated_at = user.updated_at.strftime('%d-%m-%Y %H:%M:%S') if user.updated_at else 'N/A'
        
        # Get merchants and roles as comma-separated strings
        merchants = ", ".join(link.merchant_name for link in user.merchant_links.all())
        roles = ", ".join(set(mapping.role.name for mapping in user.role_mappings.all() if mapping.role))
        
        # Row data
        row_data = [
            user.user_id,
            user.email,
            user.first_name or '',
            user.last_name or '',
            user.phone or '',
            status_text,
            last_login,
            created_at,
            updated_at,
            merchants,
            roles
        ]
        
        # Set row fill based on even/odd
        row_fill = odd_row_fill if idx % 2 == 0 else even_row_fill
        
        # Write data row
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = thin_border
            cell.fill = row_fill
            
            # Apply specific formatting based on column type
            if col_num == 1:  # User ID
                cell.font = id_font
                cell.alignment = center_alignment
            elif col_num == 2:  # Email
                cell.font = normal_font
                cell.alignment = left_alignment
            elif col_num in [3, 4]:  # Names
                cell.font = normal_font
                cell.alignment = left_alignment
            elif col_num == 5:  # Phone
                cell.font = normal_font
                cell.alignment = center_alignment
            elif col_num == 6:  # Status
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = center_alignment
                # Apply conditional formatting based on status
                if status_text.lower() in ['active', 'активний']:
                    cell.fill = status_active_fill
                elif status_text.lower() in ['blocked', 'заблоковано']:
                    cell.fill = status_blocked_fill
            elif col_num in [7, 8, 9]:  # Dates
                cell.font = date_font
                cell.alignment = center_alignment
            else:  # Merchants and Roles
                cell.font = normal_font
                cell.alignment = left_alignment
        
        row_num += 1
    
    # Apply auto filter to header row
    ws.auto_filter.ref = f"A{header_row}:K{row_num-1}"
    
    # Freeze panes to keep header visible when scrolling
    ws.freeze_panes = f'A{header_row+1}'
    
    # Create a file-like buffer to receive XLSX data
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate the response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Set filename with timestamp
    timestamp = datetime.datetime.now().strftime('%d%m%Y_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=api_users_{timestamp}.xlsx'
    
    return response

@login_required
def export_all_merchants_xlsx(request):
    """
    Export merchants with roles and API users to XLSX.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merchants"

    headers = [
        str(_("Merchant")),
        str(_("Roles")),
        str(_("API Users Count")),
        str(_("API Users")),
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='medium')
        )

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 60

    merchant_names = list(
        ApiUserMerchant.objects.order_by('name').values_list('name', flat=True).distinct()
    )

    role_map = {}
    for merchant_name, role_name in ApiUserRoleMapping.objects.select_related('merchant', 'role').values_list(
        'merchant__name',
        'role__name'
    ):
        if merchant_name and role_name:
            role_map.setdefault(merchant_name, set()).add(role_name)

    user_map = {}
    for merchant_name, user_email in ApiUserMerchantLink.objects.select_related('user').values_list(
        'merchant_name',
        'user__email'
    ):
        if merchant_name and user_email:
            user_map.setdefault(merchant_name, set()).add(user_email)

    merchant_user_roles_map = {}
    for merchant_name, user_email, role_name in ApiUserRoleMapping.objects.select_related('merchant', 'user', 'role').values_list(
        'merchant__name',
        'user__email',
        'role__name'
    ):
        if merchant_name and user_email and role_name:
            merchant_user_roles_map.setdefault(merchant_name, {})
            merchant_user_roles_map[merchant_name].setdefault(user_email, set()).add(role_name)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 2
    for merchant_name in merchant_names:
        roles = sorted(role_map.get(merchant_name, set()))
        users = sorted(user_map.get(merchant_name, set()))
        users_with_roles = []
        merchant_user_roles = merchant_user_roles_map.get(merchant_name, {})
        for user_email in users:
            user_roles = sorted(merchant_user_roles.get(user_email, set()))
            if user_roles:
                users_with_roles.append(f"{user_email} ({', '.join(user_roles)})")
            else:
                users_with_roles.append(user_email)

        row_data = [
            merchant_name,
            ", ".join(roles),
            len(users),
            "\n".join(users_with_roles),
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num, value=value)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border

        row += 1

    ws.auto_filter.ref = f"A1:D{max(row - 1, 1)}"
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=merchants_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx'
    )
    return response

@login_required
@csrf_exempt
def scheduled_syncs(request):
    """
    View to list, create, update, or delete scheduled API syncs.
    """
    # Check if user has permission to access API functionality
    from .matrix_view import has_access_api_permission, get_user_companies_for_api
    if not has_access_api_permission(request.user):
        messages.error(request, _("Access denied - you do not have permission to access API functionality."))
        return redirect('index')
    from app_access.models import ApiCredential, ScheduledSync
    
    # Get credentials for this user filtered by allowed companies
    user_companies = get_user_companies_for_api(request.user)
    credentials = ApiCredential.objects.filter(user=request.user)
    if user_companies.exists():
        credentials = credentials.filter(company__in=user_companies)
    credentials = credentials.order_by('name')
    
    if request.method == 'GET':
        # List scheduled syncs for this user filtered by allowed companies
        schedules = ScheduledSync.objects.filter(
            credential__user=request.user,
        ).select_related('credential')
        if user_companies.exists():
            schedules = schedules.filter(credential__company__in=user_companies)
        schedules = schedules.order_by('-scheduled_time')
        
        # Get API permissions for template
        from .matrix_view import can_add_access_api, can_edit_access_api, can_delete_access_api
        
        return render(request, 'app_access/scheduled_syncs.html', {
            'schedules': schedules,
            'credentials': credentials,
            'can_add_access_api': can_add_access_api(request.user),
            'can_edit_access_api': can_edit_access_api(request.user),
            'can_delete_access_api': can_delete_access_api(request.user),
        })
    
    elif request.method == 'POST':
        # Handle AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                data = json.loads(request.body)
                action = data.get('action')
                
                # Handle different actions
                if action == 'create':
                    from .matrix_view import can_add_access_api
                    if not can_add_access_api(request.user):
                        return JsonResponse({'success': False, 'error': _("Access denied - you do not have permission to create scheduled syncs.")})
                    return create_scheduled_sync(request, data)
                elif action == 'update':
                    from .matrix_view import can_edit_access_api
                    if not can_edit_access_api(request.user):
                        return JsonResponse({'success': False, 'error': _("Access denied - you do not have permission to edit scheduled syncs.")})
                    return update_scheduled_sync(request, data)
                elif action == 'delete':
                    from .matrix_view import can_delete_access_api
                    if not can_delete_access_api(request.user):
                        return JsonResponse({'success': False, 'error': _("Access denied - you do not have permission to delete scheduled syncs.")})
                    return delete_scheduled_sync(request, data)
                else:
                    return JsonResponse({
                        'success': False,
                        'error': _("Invalid action specified.")
                    })
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'error': _("Invalid JSON in request body.")
                })
        
        # Handle form submission
        else:
            # Check permission to add scheduled syncs
            from .matrix_view import can_add_access_api
            if not can_add_access_api(request.user):
                messages.error(request, _("Access denied - you do not have permission to create scheduled syncs."))
                return redirect('scheduled_syncs')
            
            # Process form data
            try:
                name = request.POST.get('name')
                credential_id = request.POST.get('credential_id')
                frequency = request.POST.get('frequency')
                scheduled_date = request.POST.get('scheduled_date')
                scheduled_time = request.POST.get('scheduled_time')
                
                # Validate input
                if not name or not credential_id or not frequency or not scheduled_date or not scheduled_time:
                    messages.error(request, _("All fields are required."))
                    return redirect('scheduled_syncs')
                
                # Get credential
                try:
                    credential = ApiCredential.objects.get(id=credential_id, user=request.user)
                except ApiCredential.DoesNotExist:
                    messages.error(request, _("Selected credential does not exist."))
                    return redirect('scheduled_syncs')
                
                # Parse datetime
                try:
                    scheduled_datetime = timezone.datetime.strptime(
                        f"{scheduled_date} {scheduled_time}",
                        "%Y-%m-%d %H:%M"
                    )
                    # Make timezone-aware
                    scheduled_datetime = timezone.make_aware(scheduled_datetime)
                except ValueError:
                    messages.error(request, _("Invalid date or time format."))
                    return redirect('scheduled_syncs')
                
                # Create scheduled sync
                scheduled_sync = ScheduledSync.objects.create(
                    name=name,
                    credential=credential,
                    frequency=frequency,
                    scheduled_time=scheduled_datetime,
                    is_active=True,
                    created_by=request.user
                )
                
                # Calculate next run time
                if frequency != 'once':
                    scheduled_sync.calculate_next_run()
                
                # Create periodic task for recurring syncs
                if frequency != 'once':
                    scheduled_sync.update_or_create_task()
                
                messages.success(request, _("Scheduled sync created successfully."))
                return redirect('scheduled_syncs')
                
            except Exception as e:
                messages.error(request, _("Error creating scheduled sync: ") + str(e))
                return redirect('scheduled_syncs')
    
    # Method not allowed
    return JsonResponse({
        'success': False,
        'error': _("Method not allowed.")
    }, status=405)

@login_required
@csrf_exempt
def convert_to_once(request):
    """Convert a recurring scheduled sync (daily, weekly, monthly) to run once immediately."""
    from app_access.models import ScheduledSync
    from app_access.tasks import sync_api_users_task
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': _("Method not allowed.")
        }, status=405)
    
    try:
        # Get sync_id from request data
        if request.headers.get('Content-Type') == 'application/json':
            data = json.loads(request.body)
            sync_id = data.get('id')
        else:
            sync_id = request.POST.get('id')
        
        if not sync_id:
            return JsonResponse({
                'success': False,
                'error': _("Scheduled sync ID is required.")
            })
        
        # Get scheduled sync
        try:
            scheduled_sync = ScheduledSync.objects.get(id=sync_id, credential__user=request.user)
        except ScheduledSync.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': _("Scheduled sync not found.")
            })
        
        # Check if this is a recurring sync
        if scheduled_sync.frequency == 'once':
            return JsonResponse({
                'success': False,
                'error': _("This sync is already scheduled to run once.")
            })
        
        # Store original values before modifying
        original_frequency = scheduled_sync.frequency
        original_next_run = scheduled_sync.next_run
        
        # Delete associated periodic task if exists
        if scheduled_sync.periodic_task:
            scheduled_sync.periodic_task.delete()
            scheduled_sync.periodic_task = None
        
        # Change frequency to once
        scheduled_sync.frequency = 'once'
        scheduled_sync.next_run = None
        scheduled_sync.save()
        
        # Run the task immediately
        task = sync_api_users_task.apply_async(
            kwargs={
                'scheduled_sync_id': scheduled_sync.id,
                'credential_id': scheduled_sync.credential.id
            }
        )
        
        # Update task ID
        scheduled_sync.celery_task_id = task.id
        scheduled_sync.save(update_fields=['celery_task_id'])
        
        return JsonResponse({
            'success': True,
            'message': _("Schedule converted to run once and task started."),
            'original_frequency': original_frequency,
            'original_next_run': original_next_run.strftime("%Y-%m-%d %H:%M") if original_next_run else None,
            'task_id': task.id
        })
    
    except Exception as e:
        logger.error(f"Error converting scheduled sync: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': _("Error converting scheduled sync: ") + str(e)
        })

def create_scheduled_sync(request, data):
    """
    Create a new scheduled sync from AJAX request.
    """
    from app_access.models import ApiCredential, ScheduledSync
    from app_access.tasks import sync_api_users_task
    from django_celery_beat.models import PeriodicTask
    import json
    
    try:
        # Extract data
        name = data.get('name')
        credential_id = data.get('credential_id')
        frequency = data.get('frequency')
        hourly_cycle = data.get('hourly_cycle', 1)  # Default to 1 hour if not provided
        scheduled_date = data.get('scheduled_date')
        scheduled_time = data.get('scheduled_time')
        
        # Validate input
        if not name or not credential_id or not frequency or not scheduled_date or not scheduled_time:
            return JsonResponse({
                'success': False,
                'error': _("All fields are required.")
            })
        
        # Validate hourly_cycle if frequency is hourly
        if frequency == 'hourly':
            try:
                hourly_cycle = int(hourly_cycle)
                if hourly_cycle not in [1, 2, 3, 4, 6, 8, 12]:
                    return JsonResponse({
                        'success': False,
                        'error': _("Invalid hourly cycle value.")
                    })
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'error': _("Invalid hourly cycle value.")
                })
        
        # Get credential
        try:
            credential = ApiCredential.objects.get(id=credential_id, user=request.user)
        except ApiCredential.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': _("Selected credential does not exist.")
            })
        
        # Parse datetime
        try:
            scheduled_datetime = timezone.datetime.strptime(
                f"{scheduled_date} {scheduled_time}",
                "%Y-%m-%d %H:%M"
            )
            # Make timezone-aware
            scheduled_datetime = timezone.make_aware(scheduled_datetime)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': _("Invalid date or time format.")
            })
        
        # Check if datetime is in the future
        if scheduled_datetime < timezone.now():
            return JsonResponse({
                'success': False,
                'error': _("Scheduled time must be in the future.")
            })
        
        # Create scheduled sync
        scheduled_sync = ScheduledSync.objects.create(
            name=name,
            credential=credential,
            frequency=frequency,
            hourly_cycle=hourly_cycle if frequency == 'hourly' else 1,
            scheduled_time=scheduled_datetime,
            is_active=True,
            created_by=request.user
        )
        
        # Calculate next run time
        if frequency != 'once':
            scheduled_sync.calculate_next_run()
        
        # Create periodic task for recurring syncs
        if frequency != 'once':
            scheduled_sync.update_or_create_task()
        
        return JsonResponse({
            'success': True,
            'message': _("Scheduled sync created successfully."),
            'id': scheduled_sync.id,
            'name': scheduled_sync.name,
            'frequency': scheduled_sync.get_frequency_display(),
            'scheduled_time': scheduled_sync.scheduled_time.strftime("%Y-%m-%d %H:%M"),
            'next_run': scheduled_sync.next_run.strftime("%Y-%m-%d %H:%M") if scheduled_sync.next_run else None,
        })
        
    except Exception as e:
        logger.error(f"Error creating scheduled sync: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': _("Error creating scheduled sync: ") + str(e)
        })

def update_scheduled_sync(request, data):
    """
    Update an existing scheduled sync from AJAX request.
    """
    from app_access.models import ApiCredential, ScheduledSync
    from app_access.tasks import sync_api_users_task
    
    try:
        # Extract data
        sync_id = data.get('id')
        name = data.get('name')
        credential_id = data.get('credential_id')
        frequency = data.get('frequency')
        hourly_cycle = data.get('hourly_cycle', 1)  # Default to 1 hour if not provided
        scheduled_date = data.get('scheduled_date')
        scheduled_time = data.get('scheduled_time')
        is_active = data.get('is_active', True)
        
        # Validate input
        if not sync_id or not name or not credential_id or not frequency or not scheduled_date or not scheduled_time:
            return JsonResponse({
                'success': False,
                'error': _("All fields are required.")
            })
        
        # Validate hourly_cycle if frequency is hourly
        if frequency == 'hourly':
            try:
                hourly_cycle = int(hourly_cycle)
                if hourly_cycle not in [1, 2, 3, 4, 6, 8, 12]:
                    return JsonResponse({
                        'success': False,
                        'error': _("Invalid hourly cycle value.")
                    })
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'error': _("Invalid hourly cycle value.")
                })
        
        # Get scheduled sync
        try:
            scheduled_sync = ScheduledSync.objects.get(id=sync_id, credential__user=request.user)
        except ScheduledSync.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': _("Scheduled sync not found.")
            })
        
        # Get credential
        try:
            credential = ApiCredential.objects.get(id=credential_id, user=request.user)
        except ApiCredential.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': _("Selected credential does not exist.")
            })
        
        # Parse datetime
        try:
            scheduled_datetime = timezone.datetime.strptime(
                f"{scheduled_date} {scheduled_time}",
                "%Y-%m-%d %H:%M"
            )
            # Make timezone-aware
            scheduled_datetime = timezone.make_aware(scheduled_datetime)
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': _("Invalid date or time format.")
            })
        
        # Update scheduled sync
        scheduled_sync.name = name
        scheduled_sync.credential = credential
        scheduled_sync.frequency = frequency
        scheduled_sync.hourly_cycle = hourly_cycle if frequency == 'hourly' else 1
        scheduled_sync.scheduled_time = scheduled_datetime
        scheduled_sync.is_active = is_active
        scheduled_sync.save()
        
        # Calculate next run time
        if frequency != 'once':
            scheduled_sync.calculate_next_run()
        
        # Update periodic task
        scheduled_sync.update_or_create_task()
        
        return JsonResponse({
            'success': True,
            'message': _("Scheduled sync updated successfully."),
            'id': scheduled_sync.id,
            'name': scheduled_sync.name,
            'frequency': scheduled_sync.get_frequency_display(),
            'scheduled_time': scheduled_sync.scheduled_time.strftime("%Y-%m-%d %H:%M"),
            'next_run': scheduled_sync.next_run.strftime("%Y-%m-%d %H:%M") if scheduled_sync.next_run else None,
        })
        
    except Exception as e:
        logger.error(f"Error updating scheduled sync: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': _("Error updating scheduled sync: ") + str(e)
        })

def delete_scheduled_sync(request, data):
    """
    Delete a scheduled sync.
    """
    try:
        if 'id' not in data:
            return JsonResponse({
                'success': False,
                'error': _("Missing scheduled sync ID")
            })
        
        scheduled_sync_id = data['id']
        
        # Get the scheduled sync
        scheduled_sync = ScheduledSync.objects.get(id=scheduled_sync_id)
        
        # Check if the user has permission to delete the scheduled sync
        if not request.user.is_superuser and scheduled_sync.created_by != request.user:
            return JsonResponse({
                'success': False,
                'error': _("You don't have permission to delete this scheduled sync")
            })
        
        # Delete the scheduled sync
        scheduled_sync.delete()
        
        return JsonResponse({
            'success': True,
            'message': _("Scheduled sync deleted successfully.")
        })
        
    except Exception as e:
        logger.error(f"Error deleting scheduled sync: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': _("Error deleting scheduled sync: ") + str(e)
        })

@login_required
@never_cache
def check_api_status(request):
    """
    Check the status of API-related services (Celery, Redis, Beat).
    This function bypasses Django's translation system to avoid encoding issues.
    """
    try:
        # Check Redis status
        redis_client = Redis(host=settings.REDIS_HOST,
                            port=settings.REDIS_PORT,
                            db=settings.REDIS_DB)
        redis_running = redis_client.ping()
    except Exception as e:
        logger.error(f"Redis connection error: {e}")
        redis_running = False

    # Check Celery status
    try:
        from celery.app.control import Control
        from SecBoard.celery import app
        
        control = Control(app=app)
        active_workers = control.inspect().active()
        celery_running = active_workers is not None and len(active_workers) > 0
    except Exception as e:
        logger.error(f"Celery check error: {e}")
        celery_running = False

    # Check Beat status
    try:
        beat_running = False
        
        # Method 1: Check celerybeat-schedule files' last modified time
        import os
        import time
        
        beat_files = ['celerybeat-schedule.dat', 'celerybeat-schedule.dir', 'celerybeat-schedule.bak']
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        for filename in beat_files:
            file_path = os.path.join(base_dir, filename)
            if os.path.exists(file_path):
                # Check if file was modified in the last 5 minutes
                mod_time = os.path.getmtime(file_path)
                if (time.time() - mod_time) < 300:  # 300 seconds = 5 minutes
                    beat_running = True
                    break
        
        # Method 2: If file check didn't confirm, check for tasks
        if not beat_running:
            from django_celery_beat.models import PeriodicTask
            from django.utils import timezone
            from datetime import timedelta
            
            # Check if any task ran recently
            latest_task = PeriodicTask.objects.filter(
                enabled=True,
                last_run_at__isnull=False
            ).order_by('-last_run_at').first()

            if latest_task:
                # Check if any task ran in the last 5 minutes
                beat_running = latest_task.last_run_at > timezone.now() - timedelta(minutes=5)
            else:
                # Check for scheduled tasks
                has_scheduled_tasks = PeriodicTask.objects.filter(enabled=True).exists()
                
                # If we have scheduled tasks but none have run, beat might be running but no task was due yet
                beat_running = has_scheduled_tasks
                
        # Method 3: Try to ping the scheduler via Redis
        if not beat_running and redis_running:
            try:
                # Check if Beat has added a heartbeat key to Redis
                beat_heartbeat = redis_client.get('celery-beat-heartbeat')
                if beat_heartbeat:
                    last_heartbeat = float(beat_heartbeat)
                    # Check if heartbeat is within last minute
                    beat_running = (time.time() - last_heartbeat) < 60
            except:
                # Ignore errors in this fallback method
                pass
    except Exception as e:
        logger.error(f"Beat check error: {e}")
        beat_running = False

    # Create direct HttpResponse with JSON content to bypass translation system
    from django.http import HttpResponse
    import json
    
    response_data = {
        'celery_status': celery_running,
        'redis_status': redis_running,
        'beat_status': beat_running,
        'timestamp': timezone.now().isoformat()
    }
    
    return HttpResponse(
        json.dumps(response_data),
        content_type='application/json'
    )

class BypassLocaleMiddleware(LocaleMiddleware):
    def process_request(self, request):
        # Do nothing, effectively bypassing the locale middleware
        pass
    
    def process_response(self, request, response):
        # Return the response unmodified
        return response

bypass_locale = decorator_from_middleware_with_args(BypassLocaleMiddleware)

@require_GET
@login_required
@bypass_locale()
def direct_service_status(request):
    """
    A direct service status check that bypasses Django's translation system completely.
    This is a fallback for the check_api_status endpoint to avoid Unicode decoding issues.
    """
    try:
        # Check Redis status
        redis_client = Redis(host=settings.REDIS_HOST,
                            port=settings.REDIS_PORT,
                            db=settings.REDIS_DB)
        redis_running = redis_client.ping()
    except Exception as e:
        logger.error(f"Redis connection error: {e}")
        redis_running = False

    # Check Celery status
    try:
        from celery.app.control import Control
        from SecBoard.celery import app
        
        control = Control(app=app)
        active_workers = control.inspect().active()
        celery_running = active_workers is not None and len(active_workers) > 0
    except Exception as e:
        logger.error(f"Celery check error: {e}")
        celery_running = False

    # Check Beat status
    try:
        beat_running = False
        
        # Method 1: Check celerybeat-schedule files' last modified time
        import os
        import time
        
        beat_files = ['celerybeat-schedule.dat', 'celerybeat-schedule.dir', 'celerybeat-schedule.bak']
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        
        for filename in beat_files:
            file_path = os.path.join(base_dir, filename)
            if os.path.exists(file_path):
                # Check if file was modified in the last 5 minutes
                mod_time = os.path.getmtime(file_path)
                if (time.time() - mod_time) < 300:  # 300 seconds = 5 minutes
                    beat_running = True
                    break
        
        # Method 2: If file check didn't confirm, check for tasks
        if not beat_running:
            from django_celery_beat.models import PeriodicTask
            from django.utils import timezone
            from datetime import timedelta
            
            # Check if any task ran recently
            latest_task = PeriodicTask.objects.filter(
                enabled=True,
                last_run_at__isnull=False
            ).order_by('-last_run_at').first()

            if latest_task:
                # Check if any task ran in the last 5 minutes
                beat_running = latest_task.last_run_at > timezone.now() - timedelta(minutes=5)
            else:
                # Check for scheduled tasks
                has_scheduled_tasks = PeriodicTask.objects.filter(enabled=True).exists()
                
                # If we have scheduled tasks but none have run, beat might be running but no task was due yet
                beat_running = has_scheduled_tasks
                
        # Method 3: Try to ping the scheduler via Redis
        if not beat_running and redis_running:
            try:
                # Check if Beat has added a heartbeat key to Redis
                beat_heartbeat = redis_client.get('celery-beat-heartbeat')
                if beat_heartbeat:
                    last_heartbeat = float(beat_heartbeat)
                    # Check if heartbeat is within last minute
                    beat_running = (time.time() - last_heartbeat) < 60
            except:
                # Ignore errors in this fallback method
                pass
    except Exception as e:
        logger.error(f"Beat check error: {e}")
        beat_running = False

    # Create direct HttpResponse with JSON content to bypass translation system completely
    from django.http import HttpResponse
    import json
    
    response_data = {
        'celery_status': celery_running,
        'redis_status': redis_running,
        'beat_status': beat_running,
        'timestamp': timezone.now().isoformat()
    }
    
    # Return raw JSON without going through Django's translation system
    return HttpResponse(
        json.dumps(response_data),
        content_type='application/json'
    )
